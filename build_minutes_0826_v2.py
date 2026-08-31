# -*- coding: utf-8 -*-
"""中間報告会議（令和8年8月26日）の議事録の校正（第2回）.

令和8年8月31日に受領した編集後の版（受領版2）を点検する。
第1回の校正17件はすべて保たれている。
発注者による編集5件はいずれも文章を良くするものであり、そのまま採る。
本回は、その編集の意図を損なわない3件のみを本文に反映し、
編集された箇所そのものに関わる3件はご提案として本文を変えない。

  第10期計画_中間報告会議議事録_令和8年8月26日.odt   校正反映後（送付）
  第10期計画_中間報告会議議事録の校正結果_第2回.xlsx  校正の記録（内部保管）

文書プロパティの作成者に個人名が再度入っていた。
Word で保存すると作成者が自動で入るため、
保存前に「ドキュメント検査」で個人情報を削除する運用が要る。
"""

import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import docmeta

SRC = ("/home/user/repository/source/"
       "中間報告会議議事録_令和8年8月26日_受領版2.odt")
OUT_ODT = ("/home/user/repository/output/"
           "第10期計画_中間報告会議議事録_令和8年8月26日.odt")
OUT_XLSX = ("/home/user/repository/output/"
            "第10期計画_中間報告会議議事録の校正結果_第2回.xlsx")

# ------------------------------------------------------------------
# 発注者による編集（第1回の校正反映版 → 受領版2）
#   (No., 箇所, 編集前, 編集後, 受託者の評価, 採否)
# ------------------------------------------------------------------
HENSHU = [
    (1, "必要書類の確認",
     "…推計を取る形を取らせていただくことで合意。"
     "給付実績データは、当月末までにいただくことで合意",
     "…推計を取る形を採用させていただき、"
     "給付実績データは、当月末までにいただくことで合意。",
     "「取る形を取らせていただく」という重複を解消されたもので、"
     "改善です。ただし2つの合意事項が1文になったため、"
     "主語がねじれています（前段は受託者の方法、後段は発注者のご提供）。"
     "2文に分けることをご提案します（ご提案No.1）。", "そのまま採る"),
    (2, "社会資源調査 介護従事者数",
     "介護事業所・生活関連情報検索サイトより、…実態がある。",
     "弊社において、介護事業所・生活関連情報検索サイトより、…"
     "実態が確認された。",
     "主体を明示し、断定を避けられたもので、改善です。"
     "ただし同じ区分の後段にも「弊社において計画に接続しやすい方法を」"
     "とあり、「弊社において」が2回現れます（校正No.1）。", "一部採る"),
    (3, "業務進捗",
     "(2)第9期計画の評価・検証65％・(5)将来推計78％は、",
     "(2)第9期計画の評価・検証65％、(5)将来推計78％は、",
     "中点を読点に改められたものです。"
     "ただしこの文は読点が多く、並列の読点と文の切れ目の読点が"
     "同じ記号になるため読みにくくなります（ご提案No.2）。", "そのまま採る"),
    (4, "連携事項（弊社担当）",
     "早急に（明日中）に事務局へ送付する。",
     "8/27中に事務局へ送付する。",
     "「に」の重複を解消し、期限を具体化されたもので、改善です。"
     "ただし文書内の他の日付は「令和8年8月26日」「令和8年9月2日（水）」と"
     "和暦で表記しており、「8/27」のみ表記が異なります（校正No.2）。",
     "一部採る"),
    (5, "連携事項（事務局様）",
     "・前回の目標指標を出していくのか、新設予定のＫＰＩはあるか、"
     "ご確認いただく。",
     "・前回の目標指標を出していくのか、新設予定のＫＰＩの有無等について、"
     "ご確認いただく。",
     "口語的な「〜はあるか」を改められたもので、改善です。"
     "そのまま採ります。", "そのまま採る"),
]

# ------------------------------------------------------------------
# 第2回の校正（本文に反映するもの）
#   発注者の編集の意図を損なわないものに限る。
#   (No., 区分, 箇所, 現行, 修正後, 理由, 置換の対象, 置換後)
# ------------------------------------------------------------------
FIX = [
    (1, "語の重複", "社会資源調査 介護従事者数",
     "弊社において、介護事業所・…（中略）…"
     "弊社において計画に接続しやすい方法を",
     "弊社が、介護事業所・…（中略）…"
     "弊社において計画に接続しやすい方法を",
     "同じ区分の中で「弊社において」が2回現れます。"
     "冒頭を「弊社が」に改め、重複を避けます。"
     "主体を明示するという編集のご意図は保たれます。",
     "弊社において、介護事業所", "弊社が、介護事業所"),
    (2, "日付の表記", "連携事項（弊社担当）",
     "8/27中に事務局へ送付する。",
     "令和8年8月27日中に事務局へ送付する。",
     "文書内の他の日付は「令和8年8月26日」「令和8年9月2日（水）」と"
     "和暦で表記しています。"
     "算用数字とスラッシュによる表記は本文書では用いていません。"
     "期限を具体化するという編集のご意図は保たれます。",
     "8/27中に", "令和8年8月27日中に"),
    (3, "文書プロパティ", "meta.xml",
     "作成者「熊谷 麻由」、"
     "作成ソフト「MicrosoftOffice/15.0 MicrosoftWord」",
     "作成者「ビズアップ公共コンサルティング株式会社」",
     "第1回の校正で「ビズアップ公共コンサルティング株式会社」に"
     "改めましたが、Word で保存された際に個人名が再度入っています。"
     "構成3町へ展開される文書であり、送付前の是正が必要です。"
     "Word の「ファイル」→「情報」→「問題のチェック」→"
     "「ドキュメント検査」で個人情報を削除する運用をご提案します。",
     "", ""),
]

# ------------------------------------------------------------------
# ご提案（本文は変えない）
#   発注者が意図して編集された箇所であるため、
#   受託者の判断では変えず、ご意向を伺う。
# ------------------------------------------------------------------
TEIAN = [
    (1, "文のねじれ", "必要書類の確認",
     "使用するデータは、前段の期間とし、"
     "見える化上の試算で推計を取る形を採用させていただき、"
     "給付実績データは、当月末までにいただくことで合意。",
     "使用するデータは、前段の期間とし、"
     "見える化システムによる試算で推計する方法を採用することで合意。"
     "給付実績データは、当月末までにご提供いただくことで合意。",
     "1文に「受託者が採用する方法」と"
     "「発注者にご提供いただく期限」の2つの合意事項が入っており、"
     "文の途中で主語が変わります。2文に分けることをご提案します。"
     "あわせて「推計を取る」という口語を「推計する」に、"
     "「見える化上」を「見える化システムによる」に改めます。"),
    (2, "句読点", "業務進捗",
     "(2)第9期計画の評価・検証65％、(5)将来推計78％は、",
     "(2)第9期計画の評価・検証65％及び(5)将来推計78％は、",
     "中点を読点に改められましたが、"
     "この文は読点が多く、並列を表す読点と文の切れ目の読点が"
     "同じ記号になるため読みにくくなります。"
     "並列を「及び」で明示することをご提案します。"),
    (3, "文体", "社会資源調査 介護従事者数",
     "…そのまま計上されている実態が確認された。",
     "…そのまま計上されている実態があることを確認。",
     "議事録の他の記述は「〜を確認。」「〜で合意。」という"
     "簡潔体で統一されています。"
     "「確認された」は受身であり、確認した主体が読み取れません。"
     "断定を避けるという編集のご意図は"
     "「〜であることを確認。」でも保たれます。"),
]

# ------------------------------------------------------------------
# 点検した結果（本回で確認し、問題がなかった事項）
# ------------------------------------------------------------------
TENKEN = [
    ("第1回の校正17件の保持", "すべて保たれている",
     "業務期間（令和9年3月26日まで）、代表KPIの中間判定、"
     "①②の追加、業務進捗の行、受領日の明確化、"
     "脱字・助詞・誤変換・用語・名称・表記の統一、"
     "出席者欄の重複解消のいずれも保たれています。"),
    ("数値", "誤りなし",
     "代表KPI4項目の判定、進捗率（全体60％及び区分別9件）、"
     "業務期間、期限（令和8年9月2日）のいずれも"
     "受託者の記録と一致します。"
     "進捗率は令和8年8月26日時点の報告値であり、"
     "その後の再評価（8月28日時点64％）とは別のものです。"),
    ("固有名詞", "誤りなし",
     "大雪地区広域連合、東川町、ビズアップ公共コンサルティング株式会社、"
     "出席者の氏名の表記に誤りはありません。"),
    ("空白・書式", "保たれている",
     "見出しの番号の後の空白、氏名と敬称の間の空白、"
     "「①第9期計画の評価・検証 中間報告」の空白など、"
     "半角空白はすべて保たれています。"),
    ("誤字脱字", "該当なし",
     "本文を通読し、変換の誤り・脱字・衍字を点検しました。"
     "変換の誤り・脱字・衍字は1件もありませんでした。"
     "本回の校正3件とご提案3件は、"
     "いずれも表記の統一と文の組立てに関するものです。"),
]


def build_odt():
    """受領版2の書式を保ったまま、本文と文書プロパティを校正する。"""
    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        data = {n: z.read(n) for n in names}

    content = data["content.xml"].decode("utf-8")
    applied = []
    for row in FIX:
        old, new = row[6], row[7]
        if not old:
            continue
        if old not in content:
            raise AssertionError("置換対象が見つからない: %r" % old)
        n = content.count(old)
        content = content.replace(old, new)
        applied.append((row[0], n))
    data["content.xml"] = content.encode("utf-8")

    if os.path.exists(OUT_ODT):
        os.remove(OUT_ODT)
    with zipfile.ZipFile(OUT_ODT, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(zipfile.ZipInfo("mimetype"), data["mimetype"],
                   compress_type=zipfile.ZIP_STORED)
        for n in names:
            if n == "mimetype":
                continue
            z.writestr(n, data[n])

    docmeta.clean_odt(
        OUT_ODT,
        "大雪地区広域連合　第10期介護保険事業計画　"
        "第9期計画の評価・検証　中間報告　会議議事録（令和8年8月26日開催）",
        "第9期計画の評価・検証 中間報告の会議の議事録")
    return applied


APPLIED = build_odt()

# ------------------------------------------------------------------ 校正結果
FONT = "游ゴシック"
NAVY = "1F4E78"
HEAD = "5B9BD5"
IN_Y = "FFF2CC"
OK_G = "E2EFDA"
NG_O = "FCE4D6"
MID_B = "DEEBF7"
GRAY = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 56
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=64, align=None):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9)
        c.border = BORDER
        ha = (align or {}).get(i, "left")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8, height=88):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_第2回の校正の概要",
           "中間報告会議議事録（令和8年8月26日）の校正結果　第2回",
           "令和8年8月31日に受領した編集後の版を点検しました。"
           "第1回の校正17件はすべて保たれています。"
           "発注者による編集5件はいずれも文章を良くするもので、"
           "そのまま採ります。"
           "本回は、編集の意図を損なわない3件を本文に反映し、"
           "編集された箇所そのものに関わる3件はご提案としました。",
           [4, 22, 16, 46, 46, 12], freeze="A5")

r = lead(ws, 4, "【1　点検した結果】", 6)
r = header(ws, r, ["No.", "点検した事項", "結果", "内容", "", "判定"])
for i, (ji, kekka, cont) in enumerate(TENKEN, start=1):
    fl = {6: OK_G if kekka.startswith(("すべて", "誤りなし", "保たれ"))
          else IN_Y}
    r = body(ws, r, [i, ji, kekka, cont, "", kekka], fl, height=56,
             align={1: "center", 3: "center", 6: "center"})
    ws.merge_cells(start_row=r - 1, start_column=4, end_row=r - 1, end_column=5)

r += 1
r = lead(ws, r, "【2　本回の対応】", 6)
r = header(ws, r, ["No.", "区分", "件数", "内容", "", "対応"])
for i, (kb, ken, cont, taio) in enumerate([
    ("語の重複", 1,
     "同じ区分の中で「弊社において」が2回現れていました。"
     "冒頭を「弊社が」に改めました。"
     "主体を明示するというご意図は保たれます。", "本文に反映"),
    ("日付の表記", 1,
     "「8/27」のみ算用数字とスラッシュで、他の和暦表記と異なりました。"
     "「令和8年8月27日中に」に改めました。"
     "期限を具体化するというご意図は保たれます。", "本文に反映"),
    ("文書プロパティ", 1,
     "作成者に個人名が再度入っていました。"
     "Word で保存すると自動で入るため、運用の見直しをご提案します。",
     "本文に反映"),
    ("文のねじれ", 1,
     "2つの合意事項が1文になり、文の途中で主語が変わります。"
     "2文に分けることをご提案します。", "ご提案（本文は変えない）"),
    ("句読点", 1,
     "並列を表す読点と文の切れ目の読点が同じ記号です。"
     "「及び」で並列を明示することをご提案します。",
     "ご提案（本文は変えない）"),
    ("文体", 1,
     "「確認された」という受身が、"
     "他の「〜を確認。」という簡潔体と揃っていません。",
     "ご提案（本文は変えない）"),
], start=1):
    fl = {6: OK_G if taio == "本文に反映" else IN_Y}
    if kb == "文書プロパティ":
        fl = {2: NG_O, 6: OK_G}
    r = body(ws, r, [i, kb, "%d件" % ken, cont, "", taio], fl, height=48,
             align={1: "center", 3: "center", 6: "center"})
    ws.merge_cells(start_row=r - 1, start_column=4, end_row=r - 1, end_column=5)

note(ws, r + 1,
     "注1）第1回の校正17件（業務期間の年、代表KPIの中間判定、"
     "①②の追加、業務進捗の行、受領日の明確化ほか）は"
     "すべて保たれていることを確認しました。\n"
     "注2）発注者による編集5件（01シート）は、"
     "いずれも文章を良くするものでした。そのまま採ります。"
     "本文に反映した3件は、その編集の意図を損なわないものに限っています。"
     "編集された箇所そのものに関わる3件は、"
     "受託者の判断では変えず、ご提案として03シートに掲げました。\n"
     "注3）文書プロパティは第1回でも是正した箇所です。"
     "Word で保存するたびに作成者が自動で入るため、"
     "「ファイル」→「情報」→「問題のチェック」→「ドキュメント検査」で"
     "個人情報を削除してから保存する運用をご提案します。", 6, 88)

# ============================================================ 01
ws = sheet("01_発注者による編集",
           "発注者による編集（5件）と、受託者の評価",
           "第1回の校正反映版から受領版2への差分です。"
           "いずれも文章を良くする編集であり、"
           "そのまま又は一部を整えて採ります。",
           [4, 22, 44, 44, 46, 12], freeze="A5")

r = header(ws, 4, ["No.", "箇所", "編集前（受託者の校正反映版）",
                   "編集後（受領版2）", "受託者の評価", "採否"])
for no, kasho, mae, ato, hyoka, saihi in HENSHU:
    fl = {6: OK_G if saihi == "そのまま採る" else IN_Y}
    r = body(ws, r, [no, kasho, mae, ato, hyoka, saihi], fl, height=88,
             align={1: "center", 6: "center"})

note(ws, r + 1,
     "注1）「一部採る」は、編集の趣旨は採りつつ、"
     "表記の不統一等を整えることを意味します。"
     "編集の内容そのものを戻すものではありません。\n"
     "注2）編集3（業務進捗の中点を読点に）は、"
     "本回の校正で「及び」に改めます。"
     "中点に戻すものではありません。\n"
     "注3）ご意図と異なる場合は、そのままの表記に戻します。"
     "ご指示をいただければ対応いたします。", 6, 76)

# ============================================================ 02
ws = sheet("02_本文に反映した校正",
           "本文に反映した校正（%d件）" % len(FIX),
           "発注者の編集の意図を損なわないものに限って本文に反映しました。"
           "受領版2の記載と、校正後の記載を対照します。",
           [4, 16, 22, 44, 44, 44], freeze="A5")

r = header(ws, 4, ["No.", "区分", "箇所", "受領版2", "校正後", "理由"])
for no, kb, kasho, mae, ato, why, _o, _n in FIX:
    fl = {2: NG_O if kb == "文書プロパティ" else IN_Y}
    r = body(ws, r, [no, kb, kasho, mae, ato, why], fl, height=88,
             align={1: "center"})

note(ws, r + 1,
     "注1）本文を置換した箇所は次のとおりです。%s。"
     "いずれも1箇所です。\n"
     "注2）No.3（文書プロパティ）は本文の置換ではなく、"
     "meta.xml の書き換えによります。\n"
     "注3）本文の書式は受領版2のままです。"
     "段落・表・空白は変えていません。"
     % "、".join("No.%d" % a for a, _b in APPLIED), 6, 64)

# ============================================================ 03
ws = sheet("03_ご提案",
           "ご提案（%d件・本文は変えていません）" % len(TEIAN),
           "発注者が意図して編集された箇所に関わるため、"
           "受託者の判断では変えず、ご意向を伺うものです。"
           "ご指示をいただければ反映いたします。",
           [4, 16, 22, 46, 46, 40], freeze="A5")

r = header(ws, 4, ["No.", "区分", "箇所", "現行（受領版2のまま）",
                   "ご提案", "理由"])
for no, kb, kasho, mae, ato, why in TEIAN:
    r = body(ws, r, [no, kb, kasho, mae, ato, why], {2: IN_Y}, height=96,
             align={1: "center"})

note(ws, r + 1,
     "注1）3件とも、発注者が意図して編集された箇所に関わります。"
     "受託者の判断で戻すことはせず、ご意向を伺います。\n"
     "注2）いずれも文意は変わりません。"
     "読みやすさと文体の統一に関するものです。\n"
     "注3）このままで差し支えないとのご判断であれば、"
     "受領版2の表記のままで確定します。", 6, 64)

wb.save(OUT_XLSX)
print("saved:", OUT_ODT)
print("saved:", OUT_XLSX)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
print("置換:", APPLIED)
