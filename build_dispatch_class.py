# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　成果品の送付区分表.

発注者へ送付する資料と、受託者が作業記録として保管する資料とを分ける。
あわせて、送付前に留意する表現を整理する。

シート構成
  00_この表について
  01_送付区分
  02_留意する表現
  03_送付前の点検手順
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_dispatch import DISPATCH, SOFU, JOKEN, NAIBU, GAI   # noqa: E402

ODIR = "/home/user/repository/output"
OUT = os.path.join(ODIR, "第10期計画_成果品の送付区分表.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FILL = {SOFU: OK_G, JOKEN: IN_Y, NAIBU: NG_O, GAI: GRAY}

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 44
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=28):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=HEAD, end_color=HEAD, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=26, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and i in fills:
            c.fill = PatternFill(start_color=fills[i], end_color=fills[i],
                                 fill_type="solid")
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=90):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_この表について", "成果品の送付区分表",
           "発注者へ送付する資料と、受託者が作業記録として保管する資料とを分ける。"
           "区分の基準と、送付前に留意する表現を示す。"
           "令和8年8月時点。成果品を追加した場合は本表に区分を追加する。",
           [14, 18, 62, 46])

r = lead(ws, 4, "【1　区分の考え方】", 4)
r = header(ws, r, ["区分", "取扱い", "基準", "根拠"])
for a in [
    ("送付", "そのまま送付する",
     "事実・出典・集計値で構成されており、受託者の判断を含まないか、"
     "含んでもその旨と根拠がシート内で完結しているもの。",
     "仕様書４（10）の納品物、及び月次の提出物"),
    ("条件付き", "お諮りする内容である旨を明示して送付する",
     "受託者の判断・推奨・試算・仮置き・留保を含むもの。"
     "送付にあたっては、確定値ではないこと、"
     "ご判断をお願いする事項であることを送り状又は表紙に明記する。",
     "計画策定の協議に必要な資料であるため、送付しないという選択はとらない"),
    ("内部保管", "納品時又は求めに応じて提出する",
     "受託者の作業過程の記録（自己点検、校正、修正指示、進行者の手元資料、"
     "旧版）。成果品の内容ではなく、成果品を作る過程を記した資料。",
     "仕様書４（10）「その他業務で作成した資料 一式」に含まれるため、"
     "廃棄はしない。日常の送付資料には含めない"),
    ("対象外", "本業務の成果品ではない",
     "他計画を含む様式部品。", "別スコープ"),
]:
    r = body(ws, r, list(a), {1: FILL.get(a[0], GRAY)}, height=76,
             align={1: "center", 2: "center"})

r += 1
r = lead(ws, r, "【2　自己点検の記録の取扱い】", 4)
r = note(ws, r,
         "受託者は、作成した整理を自ら批判的に検証する自己点検を行っており、"
         "その記録が2件（第9期評価の自己点検記録、"
         "主張の根拠水準の棚卸しと再レビュー）あります。\n"
         "これらは受託者自身の作業過程の記録であり、"
         "日常の送付資料としては用いません（内部保管）。"
         "ただし仕様書４（10）「その他業務で作成した資料 一式」に該当するため、"
         "廃棄せず保管し、納品時又は求めに応じて提出します。\n"
         "点検により判明した不足資料18件・確認事項56件は、"
         "そのまま送付できる形にして「必要事項の一覧」及び「確認依頼書」に"
         "移してあります。したがって、発注者にお伝えすべき内容は"
         "送付区分「送付」の資料に反映済みです。", 4, height=130)

r += 1
r = lead(ws, r, "【3　計画素案の版の管理】", 4)
r = note(ws, r,
         "計画素案は稿番号（第○稿）では管理しません。"
         "用途と基準日で管理します。\n"
         "　現在の版　第10期介護保険事業計画_協議用素案_令和8年8月.docx\n"
         "稿番号は受託者内部の作業単位であり、"
         "発注者にとっては版の新旧が分かりにくく、"
         "また未成熟な段階のものが流通する原因にもなります。"
         "今後は「協議用素案（令和8年8月時点）」のように、"
         "何のための版で、いつ時点の資料かを名称に含めます。\n"
         "作業単位としての稿番号は、内部保管の修正指示書にのみ残します。", 4,
         height=118)

# ============================================================ 01
ws = sheet("01_送付区分", "成果品ごとの送付区分",
           "ファイルごとに区分と理由を示す。"
           "「実体」欄は、本表の作成時点で output に存在するかどうかである。",
           [6, 13, 54, 60, 8])

r = header(ws, 4, ["No.", "区分", "ファイル名", "理由", "実体"])
cnt = {}
for i, fn in enumerate(sorted(DISPATCH), start=1):
    kb, why = DISPATCH[fn]
    cnt[kb] = cnt.get(kb, 0) + 1
    ex = "○" if os.path.exists(os.path.join(ODIR, fn)) else "―"
    r = body(ws, r, [i, kb, fn, why, ex], {2: FILL.get(kb, GRAY)},
             height=40, align={1: "center", 2: "center", 5: "center"})

r += 1
r = lead(ws, r, "【区分別の件数】", 5)
r = header(ws, r, ["区分", "件数", "取扱い", "", ""])
for kb, torii in [
    (SOFU, "そのまま送付する"),
    (JOKEN, "お諮りする内容である旨を明示して送付する"),
    (NAIBU, "納品時又は求めに応じて提出する"),
    (GAI, "本業務の成果品ではない"),
]:
    r = body(ws, r, [kb, cnt.get(kb, 0), torii, "", ""],
             {1: FILL.get(kb, GRAY)}, align={1: "center", 2: "center"})
r = body(ws, r, ["計", sum(cnt.values()), "", "", ""],
         {1: MID_B}, align={1: "center", 2: "center"}, bold=True)

r += 1
r = note(ws, r,
         "注1）「対象外」の5件は、他計画（障がい福祉計画・こども計画）を含む"
         "様式部品であり、本業務の成果品一覧には参考として掲げているものです。\n"
         "注2）区分は受託者の整理です。"
         "発注者において別の取扱いをご希望の場合はお申し付けください。"
         "特に「内部保管」としたものについて、"
         "ご確認をご希望の場合はそのまま提出します。", 5, height=76)

# ============================================================ 02
ws = sheet("02_留意する表現", "送付資料で用いない表現",
           "発注者へ送付する資料において、"
           "誤解を招くため、又は根拠の水準を超えるため用いない表現を整理する。"
           "自動点検（check_external.py）の対象としている。",
           [6, 22, 30, 46, 34])

r = lead(ws, 4, "【1　根拠の水準を超える表現】", 5)
r = header(ws, r, ["No.", "用いない表現", "何が問題か", "代わりに用いる表現", "備考"])
for a in [
    (1, "全国トップ級", "順位の確認をしていない。1時点の値である",
     "令和6年度分の指標では全国平均を上回る",
     "分子・分母の実数による安定性の検証をしていない"),
    (2, "〜に由来する", "相関を因果として述べている",
     "〜と同時に動いている／〜との関係を確認していない",
     "要因の特定には別の検証を要する"),
    (3, "〜と整合する（根拠として）",
     "別の資料と数字が近いことを、正しさの根拠にしている",
     "〜と近い値である（ただし算定方法が異なる）",
     "算定方法が異なる統計どうしの一致は根拠にならない"),
    (4, "1件も〜ない", "確認した範囲を超えて断定している",
     "確認した範囲では〜を確認できなかった",
     "確認した範囲を必ず明記する"),
    (5, "有意差がないため関係がない",
     "検出力の不足と、関係がないこととを混同している",
     "差を確認できなかった（n＝○○）",
     "標本規模を併記する"),
]:
    r = body(ws, r, list(a), {2: NG_O}, height=42, align={1: "center"})

r += 1
r = lead(ws, r, "【2　受託者内部の用語】", 5)
r = header(ws, r, ["No.", "用いない表現", "何が問題か", "代わりに用いる表現", "備考"])
for a in [
    (6, "レッドチームレビュー",
     "受託者内部の品質管理の手法名であり、発注者にとって意味が伝わらない",
     "自己点検記録",
     "資料名も「第9期評価の自己点検記録」に改めた"),
    (7, "（指摘が）的中した",
     "内部の当否の話であり、成果品の内容ではない",
     "指摘のとおりであることが確認された",
     "訂正した事実は残し、当否の評価は書かない"),
    (8, "自らの整理を否定する立場",
     "否定を目的としているように読める",
     "自らの整理を批判的に検証する立場", ""),
    (9, "第○稿",
     "受託者内部の作業単位であり、版の新旧が発注者に伝わらない",
     "協議用素案（令和8年8月時点）等、用途と基準日による表記",
     "内部保管の修正指示書にのみ残す"),
] :
    r = body(ws, r, list(a), {2: NG_O}, height=42, align={1: "center"})

r += 1
r = lead(ws, r, "【3　自己評価にあたる表現】", 5)
r = header(ws, r, ["No.", "用いない表現", "何が問題か", "代わりに用いる表現", "備考"])
for a in [
    (10, "所定の到達点に達した／万全である",
     "成果の評価は発注者が行うものである",
     "仕様書５の令和8年8月に定める内容を実施した",
     "実施した内容を書き、達否は書かない"),
    (11, "本報告の最大の特徴は",
     "自己評価である",
     "本報告では〜を扱っている", ""),
    (12, "明らかになった",
     "1つの資料で断定している場合に用いると根拠の水準を超える",
     "〜を確認した／〜と考えられる",
     "確認できた範囲を併記する"),
    (13, "（強調のための）記号の行頭付加",
     "様式が統一されず、重要度の基準も示されない",
     "表の列を分ける／「推奨」等の語で示す",
     "評価尺度としての記号（★★★）は差し支えない"),
]:
    r = body(ws, r, list(a), {2: NG_O}, height=42, align={1: "center"})

r += 1
r = note(ws, r,
         "注）上表の表現を「用いない」とすることは、"
         "都合の悪い事実を書かないという意味ではありません。"
         "訂正した事実、確認できていない事実、留保は、そのまま記載します。"
         "改めるのは、根拠の水準を超えた断定と、"
         "受託者内部の事情に属する記述の2つです。", 5, height=62)

# ============================================================ 03
ws = sheet("03_送付前の点検手順", "送付前の点検手順",
           "資料を発注者へ送付する前に行う点検の手順。"
           "機械的な点検と目視の点検を分ける。",
           [6, 26, 56, 44])

r = header(ws, 4, ["No.", "手順", "内容", "判定"])
for a in [
    (1, "送付区分の確認",
     "01シートで、送付しようとする資料の区分を確認する。"
     "「内部保管」のものが含まれていないか確認する。",
     "内部保管が含まれていないこと"),
    (2, "自動点検の実行",
     "python3 check_external.py -v を実行し、"
     "02シートの表現が残っていないか確認する。",
     "送付対象のファイルで検出0件であること"),
    (3, "版の表記の確認",
     "計画素案に稿番号が残っていないか、"
     "表紙の基準日が最新かを確認する。",
     "「協議用素案（令和8年8月時点）」と表紙・ファイル名が一致すること"),
    (4, "個人情報の確認",
     "個票データ、担当者名、電話番号、"
     "メールアドレスが含まれていないか確認する。",
     "集計値のみであること"),
    (5, "出典の確認",
     "民間の情報サイト等の二次情報が出典になっていないか確認する。",
     "原典（国・北海道・見える化システム・3町提供資料・実施済み調査）"
     "のみであること"),
    (6, "留保の確認",
     "「条件付き」の資料について、"
     "確定値でないこと・ご判断をお願いする事項であることが"
     "資料の冒頭に書かれているか確認する。",
     "00シート又は表紙に記載があること"),
    (7, "送り状の作成",
     "「条件付き」を含む場合、"
     "何をご判断いただきたいかを送り状に列記する。",
     "ご判断事項が本文を読まずに分かること"),
]:
    r = body(ws, r, list(a), {}, height=52, align={1: "center"})

r += 1
r = note(ws, r,
         "注）手順2の自動点検は、表現の見落としを減らすためのものであり、"
         "これを通ることが内容の妥当性を意味するものではありません。"
         "内容の点検は、根拠対照表（中間報告）"
         "及び各資料の00シートの留保の記載により行います。", 4, height=56)

wb.save(OUT)

n = {}
for kb, _ in DISPATCH.values():
    n[kb] = n.get(kb, 0) + 1
print("saved:", os.path.basename(OUT))
print("　".join("%s %d" % (k, v) for k, v in n.items()), "／計", len(DISPATCH))
miss = [f for f in sorted(DISPATCH) if not os.path.exists(os.path.join(ODIR, f))]
print("実体なし:", miss if miss else "なし")
