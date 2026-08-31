# -*- coding: utf-8 -*-
"""基金条例の確認.

令和8年8月28日に受領した2つの条例の確認結果。
確認事項No.52①（基金の名称）に対する回答であり、
あわせて第10期の保険料算定に関わる論点を整理した。

シート構成
  00_確認の結果    2条例の概要と、この資料により確定すること
  01_基金条例の逐条 条文ごとの内容と、計画への意味
  02_積立ての上限   第5条第3項による上限と、令和6年度末の残高
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_kikin_jorei as J
import data_kessan_r6 as K

FONT = "游ゴシック"
NAVY = "1F4E78"
HEAD = "5B9BD5"
IN_Y = "FFF2CC"
OK_G = "E2EFDA"
NG_O = "FCE4D6"
MID_B = "DEEBF7"
GRAY = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 56
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=48, align=None, numfmt="#,##0",
         bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.border = BORDER
        ha = (align or {}).get(i, "left" if isinstance(v, str) else "right")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if numfmt and not isinstance(v, str):
            c.number_format = numfmt
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8, height=88):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_確認の結果",
           "基金条例の確認",
           "令和8年8月28日に受領した2つの条例の確認結果です。"
           "確認事項No.52①（基金の名称）にご回答をいただいたものとして"
           "扱い、あわせて第10期の保険料算定に関わる論点を整理しました。",
           [4, 32, 34, 44, 30, 12], freeze="A5")

r = lead(ws, 4, "【1　受領した条例】", 6)
r = header(ws, r, ["No.", "条例", "制定・改正", "内容", "第10期計画との関係",
                   "判定"])
for no, nm, seitei, cont, kankei, hh in [
    (1, J.JOREI["名称"],
     "%s\n改正 %s" % (J.JOREI["制定"], "／".join(J.JOREI["改正"])),
     "広域連合が設置する基金の管理に関する条例です。"
     "国民健康保険事業財政調整基金と介護保険事業財政調整基金の"
     "2基金を別表に掲げています。",
     "第5条第3項（積立ての上限）と別表（設置目的・処分の要件）が"
     "第10期の保険料算定に直接関わります。", "確認済"),
    (2, J.RINJI["名称"], J.RINJI["制定"],
     "平成21年度の介護報酬改定に伴う保険料の急激な上昇を"
     "抑制するための臨時の基金です。",
     "%s" % J.RINJI["第10期計画との関係"], "対象外"),
]:
    fl = {6: OK_G if hh == "確認済" else GRAY}
    r = body(ws, r, [no, nm, seitei, cont, kankei, hh], fl, height=80,
             align={1: "center", 6: "center"}, numfmt=None)

r += 1
r = lead(ws, r, "【2　この資料により確定すること】", 6)
r = header(ws, r, ["No.", "確定すること", "内容", "これまでの状態", "反映先",
                   "重要度"])
for no, kk, cont, mae, saki, lv in [
    (1, "基金の正式名称",
     "条例上の名称は「%s」です。"
     "令和6年度決算書の財産に関する調書も同じ名称です。"
     "年報の様式4が用いる「介護給付費準備基金」は"
     "国の様式上の呼称であり、条例上の名称ではありません。"
     "計画本文には条例上の名称を用います。" % J.KIKIN_NAME,
     "決算書と年報で名称が異なり、"
     "計画本文にどちらを用いるか確定できなかった"
     "（確認事項No.52①）",
     "素案 第6章第6節\n確認事項No.52", "高"),
    (2, "基金の処分（取崩し）の要件",
     "別表により、①介護保険事業に要する経費に充てる財源に"
     "不足を生じた場合、②介護保険事業の円滑な運営に必要な場合において"
     "予算で定めるとき、の2つに限られます。"
     "いずれも予算の定めを要します（第8条）。",
     "取崩しの要件を計画本文に記載できなかった",
     "素案 第6章第6節", "高"),
    (3, "積立ての上限",
     "第5条第3項により、毎年度の決算において生じた剰余金を"
     "「当該計画期間における介護保険事業計画で算定する"
     "介護保険料必要収納額の100分の10に相当する額まで」"
     "積み立てるものとされています。"
     "計画期間ごとに上限が定まります。",
     "上限の定めがあることを把握していなかった",
     "素案 第6章第6節\n将来推計 第3段階", "最高"),
    (4, "令和6年度に剰余金を積み立てなかった理由の手がかり",
     "令和6年度は実質収支142,873,961円が生じましたが、"
     "積立ては232,197円（運用益相当）にとどまりました。"
     "第9期の上限（205,196,846円）を令和6年度末の残高"
     "231,689,018円が上回っていることと整合的です（02シート）。",
     "剰余金を積み立てず繰越金とした理由が分からなかった",
     "素案 第6章第6節", "高"),
    (5, "処遇改善臨時特例基金は現存しないこと",
     "平成24年3月31日限りで失効しています（附則第2項）。"
     "第10期計画には関係しません。",
     "―", "―", "低"),
]:
    fl = {6: NG_O if lv == "最高" else (IN_Y if lv == "高" else None)}
    r = body(ws, r, [no, kk, cont, mae, saki, lv], fl, height=88,
             align={1: "center", 6: "center"}, numfmt=None)

note(ws, r + 1,
     "注1）確認事項No.52①（基金の名称）は、"
     "条例の受領により解決しました。"
     "計画本文には条例上の名称「%s」を用い、"
     "年報の様式4が「介護給付費準備基金」と呼んでいることを"
     "注記します。\n"
     "注2）02シートの積立ての上限は、"
     "第10期の保険料算定における取崩額の設定に直結します。"
     "ご確認をお願いする事項があります。\n"
     "注3）本確認には個人が特定される情報は含まれていません。"
     % J.KIKIN_NAME, 6, 76)

# ============================================================ 01
ws = sheet("01_基金条例の逐条",
           "%s（%s）の逐条" % (J.JOREI["名称"], J.JOREI["制定"]),
           "条文ごとに内容の要旨と、第10期計画への意味を整理します。"
           "介護保険事業財政調整基金に関わる部分を中心とします。",
           [12, 22, 56, 50], freeze="A5")

r = header(ws, 4, ["条", "見出し", "内容の要旨", "第10期計画への意味"])
for jo, midashi, cont, imi in J.JOBUN:
    fl = {1: NG_O} if jo == "第5条第3項" else (
        {1: IN_Y} if jo in ("第8条",) else {1: MID_B})
    r = body(ws, r, [jo, midashi, cont, imi], fl, height=56,
             align={1: "center"}, numfmt=None)

r += 1
r = lead(ws, r, "【別表（第2条・第8条関係）　介護保険事業財政調整基金】", 4)
r = header(ws, r, ["区分", "内容", "", ""])
r = body(ws, r, ["基金の名称", J.BEPPYO["基金の名称"], "", ""],
         {1: MID_B}, height=24, numfmt=None)
ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)
r = body(ws, r, ["基金の設置目的", J.BEPPYO["基金の設置目的"], "", ""],
         {1: MID_B}, height=34, numfmt=None)
ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)
r = body(ws, r, ["基金の処分", "\n".join(J.BEPPYO["基金の処分"]), "", ""],
         {1: NG_O}, height=44, numfmt=None)
ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

r += 1
r = lead(ws, r, "【計画本文に記載する事項】", 4)
r = header(ws, r, ["No.", "記載する事項", "根拠", ""])
for i, (kisai, konkyo) in enumerate([
    ("基金の名称は「介護保険事業財政調整基金」であること",
     "基金条例 別表"),
    ("基金の目的は、介護保険事業の適正な運営及び"
     "長期にわたる財政の健全性を維持することであること",
     "基金条例 別表"),
    ("取崩しは、①財源不足が生じた場合、"
     "②円滑な運営に必要な場合において予算で定めるとき、"
     "の2つに限られること",
     "基金条例 第8条・別表"),
    ("積立ては、計画期間の保険料収納必要額の100分の10を上限とすること",
     "基金条例 第5条第3項"),
    ("積立て・取崩しのいずれも予算の定めを要すること",
     "基金条例 第3条・第8条"),
], start=1):
    r = body(ws, r, [i, kisai, konkyo, ""], {2: MID_B}, height=40,
             align={1: "center"}, numfmt=None)

note(ws, r + 1,
     "注1）第5条第2項は国民健康保険事業財政調整基金に関する規定です"
     "（北海道国保事業費納付金の1年当たり平均額の100分の15）。"
     "本業務の対象外のため掲げていません。\n"
     "注2）第7条の繰替運用は、歳計現金への一時的な繰替えであり、"
     "第8条の処分（取崩し）とは別の制度です。"
     "保険料算定における取崩額には含みません。\n"
     "注3）条例は平成23年6月20日及び令和4年6月13日に改正されています。"
     "第5条第3項が現行の内容となった時期は、"
     "受領した条例からは判別できません。", 4, 76)

# ============================================================ 02
ws = sheet("02_積立ての上限",
           "条例第5条第3項による積立ての上限と、令和6年度末の残高",
           "第10期の保険料算定における取崩額の設定に直結します。"
           "ご確認をお願いする事項があります。",
           [34, 24, 24, 24, 52], freeze="B5")

r = lead(ws, 4, "【1　第9期の上限と、令和6年度末の残高】", 5)
r = header(ws, r, ["区分", "金額", "", "", "算出の根拠・所見"])
for nm, kingaku, sk in [
    ("第9期の保険料収納必要額（3か年計）", J.J_DAI9,
     "第9期計画の保険料算定によります。"
     "保険料の所得段階と低所得者軽減の検証で用いている値です。"),
    ("条例第5条第3項の上限（100分の10）", J.JOGEN_DAI9,
     "上記の100分の10です。円未満は切り捨てています。"),
    ("令和6年度末の基金残高", J.ZANDAKA_R6,
     "令和6年度決算書の財産に関する調書によります。"),
    ("差（残高－上限）", J.chouka(),
     "令和6年度末の残高が第9期の上限を上回っています。"),
]:
    fl = {1: MID_B}
    if "差" in nm:
        fl = {1: NG_O, 2: NG_O}
    elif "上限" in nm:
        fl = {1: IN_Y}
    r = body(ws, r, [nm, kingaku, "", "", sk], fl, height=40)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

r += 1
r = lead(ws, r, "【2　令和6年度の積立ての状況】", 5)
r = header(ws, r, ["区分", "金額", "", "", "所見"])
for nm, kingaku, sk in [
    ("令和6年度の実質収支額", K.JISSHITSU["5 実質収支額"],
     "翌年度へ繰越すべき財源は0円のため、"
     "歳入歳出差引額と同額です。"),
    ("うち基金へ繰り入れた額", K.JISSHITSU["6 地方自治法第233条の2の"
                                          "規定による基金繰入額"],
     "実質収支額を基金へ繰り入れていません。"
     "全額が翌年度への繰越金となります。"),
    ("令和6年度の基金への積立額", K.KIKIN["決算年度中増減高"],
     "歳入の財産運用収入232,197円と同額であり、"
     "条例第4条による運用益の編入と解されます。"
     "剰余金からの積立て（第5条第3項）は行われていません。"),
    ("令和6年度の基金の取崩額", K.KIKIN["うち支消"],
     "第9期の初年度は取り崩していません。"),
]:
    fl = {1: MID_B}
    r = body(ws, r, [nm, kingaku, "", "", sk], fl, height=40)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

r += 1
r = lead(ws, r, "【3　受託者の考えと、ご確認をお願いする事項】", 5)
r = header(ws, r, ["No.", "事項", "受託者の考え", "", "ご確認をお願いする内容"])
for i, (ji, kang, kakunin) in enumerate([
    ("令和6年度に剰余金を積み立てなかった理由",
     "第9期の上限205,196,846円を令和6年度末の残高231,689,018円が"
     "既に上回っているため、剰余金142,873,961円を積み立てず"
     "繰越金としたものと解されます。"
     "積立ては運用益232,197円のみでした。",
     "この理解で相違ないかご確認ください。"
     "他の理由がある場合はご教示ください。"),
    ("条例第5条第3項の「100分の10」の読み方",
     "「毎年度の決算において生じた剰余金を…"
     "100分の10に相当する額まで積み立てる」という文言は、"
     "各年度の積立額の上限ではなく、"
     "積立ての結果として基金が到達しうる残高の上限を"
     "定めたものと解しています。",
     "この解釈で相違ないかご確認ください。"
     "運用上の取扱いをご教示ください。"),
    ("第9期の保険料収納必要額の確認",
     "第9期計画の保険料算定による3か年計2,051,968,467円を"
     "用いています。",
     "この値で相違ないかご確認ください。"
     "条例にいう「介護保険事業計画で算定する"
     "介護保険料必要収納額」がこの値を指すかを含みます。"),
    ("第10期の上限と取崩額の設定",
     "第10期計画で算定する保険料収納必要額の100分の10が"
     "次期の上限となります。"
     "現在の残高が新たな上限を上回る場合、"
     "超過分の取崩しを保険料の抑制に充てる考え方があります。",
     "第10期の保険料算定において、"
     "基金をどこまで取り崩す方針とするか。"
     "受託者の案は、条例上の上限に収まる水準まで"
     "取り崩して保険料の上昇を抑制することです。"),
], start=1):
    r = body(ws, r, [i, ji, kang, "", kakunin], {2: MID_B, 5: IN_Y},
             height=80, align={1: "center"}, numfmt=None)
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)

note(ws, r + 1,
     "注1）第9期の保険料収納必要額は第9期計画の算定によるものであり、"
     "決算の実績とは異なります。"
     "条例の文言も「介護保険事業計画で算定する」としています。\n"
     "注2）第10期の上限は、第10期計画の保険料算定が確定してから"
     "定まります。将来推計 第3段階の結果を待って算定します。\n"
     "注3）取崩しには、別表の要件の充足と予算の定めの双方を要します"
     "（第8条）。計画に取崩しの方針を記載する場合は、"
     "予算措置を前提とする旨を併記します。\n"
     "注4）本シートの論点は、条例を受領したことにより"
     "検討が可能となったものです。第10期の保険料の水準に影響します。",
     5, 88)

OUT = ("/home/user/repository/output/"
       "第10期計画_基金条例の確認.xlsx")
wb.save(OUT)
print("saved:", OUT)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
