"""小野町 第10期介護保険事業計画 サービス見込量・保険料の算定シートを生成する。

委託仕様書4(2)は次のとおり定める。

  「見える化」システムを活用したサービス見込量を推計及び保険料の算出を行うこと。
  （「見える化」システムへの入力作業を含む。）……算定にあたっては、新たに介護施設を
  整備する場合や制度改正により条件が変わる場合等、様々なケースで勘案する必要が
  あるため、想定されるケースごとに保険料の算定を行うものとする。

引継ぎ資料の試算テンプレートは標準／基金抑制／基金活用の3ケースであり、これは
基金の取扱いによる区分にすぎず、仕様書が例示する「施設整備」「制度改正」の
ケースを含まない。本ブックはケース定義から組み直す。

あわせて、本町固有の事情としてケース軸に「認定者数の前提」を置く。令和8年3月末の
認定者数（見える化 792人）は20期の系列で唯一の大幅減であり、確認中である。
確定値が792人か、従前の趨勢に沿う900人台かで、給付費は2割近く動く。前提を
1つに固定して算定すると、確認結果しだいで全面やり直しになる。

数値は未受領のため空欄とし、算式のみを実装する。認定者数と実績が確定した時点で
入力欄に値を入れれば、給付費と保険料まで一気に通る状態にしてある。

出力: 04_算定・見込量/小野町_第10期_見込量保険料算定シート_YYYYMMDD.xlsx
"""

import pathlib

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent
OUT = ROOT / "小野町_引継ぎ_整理済" / "04_算定・見込量"
ASOF = "20260813"

IN_FILL = PatternFill("solid", fgColor="FFF2CC")      # 入力欄
CALC_FILL = PatternFill("solid", fgColor="EAF1FB")    # 算式
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
KEY_FILL = PatternFill("solid", fgColor="FCE4E4")     # 最終結果
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YEARS = ["令和9年度", "令和10年度", "令和11年度"]
GRADES = ["要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"]

# 介護給付のサービス区分（見込量の単位つき）
SVC_KAIGO = [
    ("居宅", "訪問介護", "回"),
    ("居宅", "訪問入浴介護", "回"),
    ("居宅", "訪問看護", "回"),
    ("居宅", "訪問リハビリテーション", "回"),
    ("居宅", "居宅療養管理指導", "人"),
    ("居宅", "通所介護", "回"),
    ("居宅", "通所リハビリテーション", "回"),
    ("居宅", "短期入所生活介護", "日"),
    ("居宅", "短期入所療養介護", "日"),
    ("居宅", "福祉用具貸与", "人"),
    ("居宅", "特定福祉用具販売", "人"),
    ("居宅", "住宅改修", "人"),
    ("居宅", "特定施設入居者生活介護", "人"),
    ("居宅", "居宅介護支援", "人"),
    ("地域密着型", "定期巡回・随時対応型訪問介護看護", "人"),
    ("地域密着型", "夜間対応型訪問介護", "人"),
    ("地域密着型", "地域密着型通所介護", "回"),
    ("地域密着型", "認知症対応型通所介護", "回"),
    ("地域密着型", "小規模多機能型居宅介護", "人"),
    ("地域密着型", "認知症対応型共同生活介護", "人"),
    ("地域密着型", "地域密着型特定施設入居者生活介護", "人"),
    ("地域密着型", "地域密着型介護老人福祉施設入所者生活介護", "人"),
    ("地域密着型", "看護小規模多機能型居宅介護", "人"),
    ("施設", "介護老人福祉施設", "人"),
    ("施設", "介護老人保健施設", "人"),
    ("施設", "介護医療院", "人"),
]

# 予防給付のサービス区分
SVC_YOBO = [
    ("介護予防", "介護予防訪問入浴介護", "回"),
    ("介護予防", "介護予防訪問看護", "回"),
    ("介護予防", "介護予防訪問リハビリテーション", "回"),
    ("介護予防", "介護予防居宅療養管理指導", "人"),
    ("介護予防", "介護予防通所リハビリテーション", "人"),
    ("介護予防", "介護予防短期入所生活介護", "日"),
    ("介護予防", "介護予防短期入所療養介護", "日"),
    ("介護予防", "介護予防福祉用具貸与", "人"),
    ("介護予防", "特定介護予防福祉用具販売", "人"),
    ("介護予防", "介護予防住宅改修", "人"),
    ("介護予防", "介護予防特定施設入居者生活介護", "人"),
    ("介護予防", "介護予防支援", "人"),
    ("地域密着型予防", "介護予防認知症対応型通所介護", "回"),
    ("地域密着型予防", "介護予防小規模多機能型居宅介護", "人"),
    ("地域密着型予防", "介護予防認知症対応型共同生活介護", "人"),
]


def style_header(ws, row=1):
    for c in ws[row]:
        if c.value is not None:
            c.font = Font(bold=True, size=9)
            c.fill = HEAD_FILL
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def body_style(ws, min_row=2, wrap_cols=()):
    """罫線と折返しを一括で当てる。太字・文字サイズの個別指定は保持する。"""
    for row in ws.iter_rows(min_row=min_row):
        for i, c in enumerate(row):
            if c.value is not None:
                c.border = BORDER
            c.font = Font(size=c.font.size or 9, bold=bool(c.font.bold))
            c.alignment = Alignment(wrap_text=(i in wrap_cols), vertical="top")


# ---------------------------------------------------------------- 00_使い方

def sheet_howto(wb):
    ws = wb.create_sheet("00_使い方")
    rows = [
        ["小野町 第10期介護保険事業計画　サービス見込量・保険料 算定シート"],
        [f"作成日：{ASOF[:4]}-{ASOF[4:6]}-{ASOF[6:]}"],
        [],
        ["■ 本シートの位置づけ"],
        ["委託仕様書4(2)が定める「サービス見込量の推計」及び「想定されるケースごとの保険料の算定」を行うための算定シートです。"],
        ["数値が未受領のため入力欄は空ですが、算式は実装済みです。値を入れれば給付費・保険料まで自動で通ります。"],
        [],
        ["■ セルの色"],
        ["黄色", "入力欄。町からの受領データ、見える化システムの出力、政策判断による設定値を入れます"],
        ["水色", "算式。触らないでください"],
        ["赤色", "算定結果（保険料基準額など）"],
        [],
        ["■ 算定の流れ"],
        ["1", "02_人口・被保険者", "見える化システムの人口推計から、令和9〜11年度の被保険者数を入力"],
        ["2", "03_認定者推計", "認定率を入力すると、要介護度別の認定者数が算出されます"],
        ["3", "04_受給率", "認定者に対する受給者の比率を入力"],
        ["4", "05_見込量_介護", "サービス別の利用者数・回数・単価を入力すると給付費が算出されます"],
        ["5", "06_見込量_予防", "同上（予防給付）"],
        ["6", "07_地域支援事業", "介護予防・日常生活支援総合事業、包括的支援事業、任意事業の事業費を入力"],
        ["7", "08_標準給付費", "05〜07が集計され、標準給付費見込額が算出されます"],
        ["8", "09_保険料算定", "負担割合・調整交付金・基金取崩・収納率・所得段階別人数を入力すると保険料基準額が出ます"],
        ["9", "10_ケース比較", "01で定義した各ケースの結果を並べます"],
        [],
        ["■ ケースの考え方（仕様書4(2)）"],
        ["仕様書は「新たに介護施設を整備する場合や制度改正により条件が変わる場合等」を例示しています。"],
        ["引継ぎ時点のテンプレートは基金の取扱いによる3ケースのみで、仕様書の例示に対応していませんでした。"],
        ["本シートでは、政策要因（施設整備・制度改正・基金）に加え、本町固有の事情として「認定者数の前提」を軸に加えています。"],
        ["令和8年3月末の認定者数は20期の系列で唯一の大幅減（▲152人）であり、確認中です。"],
        ["確定値が792人か900人台かで給付費は2割近く動くため、前提を1つに固定して算定すると、確認結果しだいで全面やり直しになります。"],
        [],
        ["■ 未受領のデータ"],
        ["本シートを完成させるには、進捗管理表 06_資料受領管理 の次の項目が必要です。"],
        ["", "令和8年3月末の認定者数の確認（最優先）"],
        ["", "令和3〜8年度 サービス別給付費・利用者数"],
        ["", "介護給付費準備基金の残高と取崩方針"],
        ["", "所得段階別第1号被保険者数"],
        ["", "地域支援事業費の実績と事業別内訳"],
        ["", "令和6・7年度の決算（保険料収納率の実績を含む）"],
        [],
        ["■ 第9期の算定根拠（参考・第9期計画本文より）"],
        ["標準給付費見込額（3年計）", "3,566,486千円"],
        ["準備基金取崩額", "35,000千円"],
        ["保険料基準額（月額）", "6,600円"],
        ["見える化システムによる必要保険料", "6,131円"],
        ["※ 第10期はこれを基礎に算定します。第9期の見込みと実績の対比は成果品①第7章7を参照。"],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13)
    for cell in ("A4", "A8", "A13", "A25", "A32", "A40"):
        ws[cell].font = Font(bold=True, size=10)
    ws["A9"].fill = IN_FILL
    ws["A10"].fill = CALC_FILL
    ws["A11"].fill = KEY_FILL
    widths(ws, [22, 26, 86])
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


# ---------------------------------------------------------------- 01_ケース定義

def sheet_cases(wb):
    ws = wb.create_sheet("01_ケース定義")
    ws.append(["ケース", "名称", "認定者数の前提", "施設整備", "制度改正", "基金取崩",
               "設定の趣旨", "仕様書との対応"])
    rows = [
        ("C1", "標準（低位）",
         "令和8年3月末792人を実勢とみる。以後は人口推計に認定率を乗じる",
         "新規整備なし（第9期方針を継承）", "織り込まない",
         "第9期と同水準",
         "見える化システムの出力をそのまま前提とする、最も機械的なケース。"
         "確認の結果792人が正しければこれが基準になる",
         "仕様書4(2)の基本ケース"),
        ("C2", "標準（高位）",
         "令和7年7月末953人の水準を実勢とみる。792人は一時的な減少とみて補正する",
         "新規整備なし", "織り込まない", "第9期と同水準",
         "**認定者数の減少が更新認定の運用等による一時的なものであった場合。**"
         "20期の系列では令和2年度以外に減少年がなく、こちらが趨勢に沿う",
         "同上（本町固有のリスクへの対応）"),
        ("C3", "施設整備あり",
         "C1と同じ",
         "**計画期間中に施設又は居住系サービスを整備する**",
         "織り込まない", "第9期と同水準",
         "仕様書が明示的に例示するケース。整備の内容・定員・開設時期を入力すると、"
         "在宅から施設への移行と給付費の増を反映する",
         "**仕様書4(2)「新たに介護施設を整備する場合」**"),
        ("C4", "制度改正の影響",
         "C1と同じ", "新規整備なし",
         "**利用者負担、報酬改定、支給限度額、第2号被保険者の負担率等の改正を織り込む**",
         "第9期と同水準",
         "仕様書が明示的に例示するケース。国の制度改正の内容が示された時点で"
         "パラメータを入れる。示されるまでは感度分析として幅で持つ",
         "**仕様書4(2)「制度改正により条件が変わる場合」**"),
        ("C5", "基金活用",
         "C1と同じ", "新規整備なし", "織り込まない",
         "**準備基金を取り崩し、保険料の上昇を抑制する**",
         "繰越金182.0百万円・基金純増88.1百万円という財政余剰の還元。"
         "取崩額を変えて保険料への影響を示す",
         "仕様書4(2)（引継ぎテンプレートの3ケースはこの軸のみ）"),
        ("C6", "保険料抑制（据置）",
         "C1と同じ", "新規整備なし", "織り込まない",
         "**保険料基準額を第9期（6,600円）に据え置くために必要な取崩額を逆算**",
         "保険料を先に決め、必要な基金取崩額を求める逆算ケース。"
         "協議会に「据え置くとしたら基金がいくら必要か」を示すために使う",
         "同上"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws)
    body_style(ws, wrap_cols=(2, 3, 4, 5, 6, 7))
    widths(ws, [7, 16, 34, 30, 34, 24, 46, 34])
    ws.freeze_panes = "A2"
    ws.append([])
    ws.append(["※ C1とC2は認定者数の前提が異なるだけで、他の条件は同じです。"
               "**この2ケースの差が、認定者数の確認が済むまでの不確実性の幅を表します。**"])
    ws.append(["※ C3〜C6はC1を基準にした差分です。C2を基準にした組合せも必要に応じて算定します。"])
    ws.append(["※ 第10期の第1号被保険者負担割合・調整交付金の考え方は国の基本指針で示されます。"
               "示されるまでは第9期の値（23％）を仮置きします。"])
    return ws


# ---------------------------------------------------------------- 02_人口・被保険者

def sheet_population(wb):
    ws = wb.create_sheet("02_人口・被保険者")
    ws.append(["区分", "令和7年度（実績）", "令和8年度（見込）"] + [f"{y}（推計）" for y in YEARS] + ["出典・備考"])
    items = [
        "総人口",
        "第1号被保険者数（65歳以上）",
        "　前期高齢者（65〜74歳）",
        "　後期高齢者（75歳以上）",
        "　　うち85歳以上",
        "第2号被保険者数（40〜64歳）",
        "高齢化率（％）",
    ]
    for it in items:
        ws.append([it, None, None, None, None, None,
                   "見える化システム（社人研準拠）／A系指標"])
    # 高齢化率 = 第1号 ÷ 総人口
    for col in range(2, 7):
        L = get_column_letter(col)
        ws[f"{L}8"] = f"=IF({L}2=0,\"\",{L}3/{L}2*100)"
    style_header(ws)
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=2, max_col=6):
        for c in row:
            c.fill = IN_FILL
            c.number_format = "#,##0"
    for c in ws["B8":"F8"][0]:
        c.fill = CALC_FILL
        c.number_format = "0.0"
    body_style(ws, wrap_cols=(6,))
    widths(ws, [26, 15, 15, 15, 15, 15, 34])
    ws.freeze_panes = "B2"
    ws.append([])
    ws.append(["※ 第1号被保険者数は前期＋後期と一致すること。85歳以上は後期の内数。"])
    ws.append(["※ 令和7年度・令和8年度は実績及び見込み。令和9〜11年度が計画期間。"])
    return ws


# ---------------------------------------------------------------- 03_認定者推計

def sheet_nintei(wb):
    ws = wb.create_sheet("03_認定者推計")
    ws.append(["要介護度", "令和8年3月末（実績）", "構成比",
               "令和9年度", "令和10年度", "令和11年度", "備考"])
    start = 2
    for g in GRADES:
        ws.append([g, None, None, None, None, None, ""])
    r_total = start + len(GRADES)
    ws.append(["合計", f"=SUM(B{start}:B{r_total - 1})", "",
               f"=SUM(D{start}:D{r_total - 1})",
               f"=SUM(E{start}:E{r_total - 1})",
               f"=SUM(F{start}:F{r_total - 1})", ""])
    # 構成比
    for i in range(len(GRADES)):
        r = start + i
        ws[f"C{r}"] = f"=IF($B${r_total}=0,\"\",B{r}/$B${r_total})"
    ws[f"C{r_total}"] = f"=IF($B${r_total}=0,\"\",SUM(C{start}:C{r_total - 1}))"

    ws.append([])
    r = r_total + 2
    ws.cell(r, 1, "【推計の前提】")
    ws.cell(r, 1).font = Font(bold=True, size=10)
    rows = [
        ("認定率の設定方法", "", "", "", "", "",
         "見える化システムの認定率（第1号被保険者に対する割合）を年齢階層別に設定し、"
         "02の被保険者数に乗じる"),
        ("65〜74歳の認定率（％）", None, "", None, None, None, "入力欄"),
        ("75〜84歳の認定率（％）", None, "", None, None, None, "入力欄"),
        ("85歳以上の認定率（％）", None, "", None, None, None, "入力欄"),
        ("第2号被保険者の認定者数", None, "", None, None, None, "入力欄（人数で設定）"),
    ]
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    for rr in range(start, r_total):
        ws[f"B{rr}"].fill = IN_FILL
        for col in "DEF":
            ws[f"{col}{rr}"].fill = IN_FILL
        ws[f"C{rr}"].fill = CALC_FILL
        ws[f"C{rr}"].number_format = "0.0%"
    for col in "BDEF":
        ws[f"{col}{r_total}"].fill = CALC_FILL
    ws[f"C{r_total}"].fill = CALC_FILL
    ws[f"C{r_total}"].number_format = "0.0%"
    for rr in range(r + 2, r + 6):
        for col in ("B", "D", "E", "F"):
            ws[f"{col}{rr}"].fill = IN_FILL
    body_style(ws, wrap_cols=(6,))
    widths(ws, [24, 17, 10, 14, 14, 14, 52])
    ws.freeze_panes = "B2"

    ws.append([])
    ws.append(["※ 認定者数の推計は、年齢階層別の認定率に被保険者数を乗じる方法（自然体推計）を基本とします。"])
    ws.append(["※ **令和8年3月末の認定者数は確認中です。**20期の系列で唯一の大幅減（▲152人、うち77.0％が85歳以上）であり、"
               "確定するまではケースC1（792人）とC2（953人水準）の両方で算定します。"])
    ws.append(["※ 介護予防・日常生活支援総合事業の対象者（事業対象者）は、認定者とは別に07で扱います。"])
    return ws


# ---------------------------------------------------------------- 04_受給率

def sheet_jukyu(wb):
    ws = wb.create_sheet("04_受給率")
    ws.append(["区分", "令和7年度（実績）", "令和9年度", "令和10年度", "令和11年度", "設定の考え方"])
    items = [
        ("認定者に対する受給者の比率（全体）", "見える化 D17系。**本町は認定率が高い一方で受給率が低く、"
                                "在宅の認定者の42.1％が未利用（令和7年度調査）**"),
        ("　居宅サービス受給率", "認定者のうち居宅サービスを利用する割合"),
        ("　地域密着型サービス受給率", "同上"),
        ("　施設サービス受給率", "同上"),
        ("　サービス未利用の割合", "**令和7年度調査で在宅の認定者の42.1％。"
                        "利用に移行した場合の給付費増がリスク**"),
    ]
    for name, note_txt in items:
        ws.append([name, None, None, None, None, note_txt])
    style_header(ws)
    for row in ws.iter_rows(min_row=2, max_row=6, min_col=2, max_col=5):
        for c in row:
            c.fill = IN_FILL
            c.number_format = "0.0%"
    body_style(ws, wrap_cols=(0, 5))
    widths(ws, [30, 16, 14, 14, 14, 60])
    ws.freeze_panes = "B2"
    ws.append([])
    ws.append(["※ 未利用者が利用に移行すると給付費が急増します。"
               "**未利用の理由は需要側が92.9％（状態ではない46.8％、本人希望なし29.2％、家族が介護16.9％）、"
               "供給制約は0.6％**であり、供給を増やしても利用は増えない構造です。"])
    ws.append(["※ したがって受給率は実績の延長を基本とし、施策により押し上げる分は"
               "レスパイト・利用意向への働きかけの効果として別に見込みます。"])
    return ws


# ---------------------------------------------------------------- 05・06_見込量

def sheet_mikomi(wb, name, services, label):
    ws = wb.create_sheet(name)
    header = ["区分", "サービス", "単位"]
    for y in YEARS:
        header += [f"{y} 利用者数", f"{y} 量", f"{y} 給付費(千円)"]
    header += ["単価(円/単位)", "備考"]
    ws.append(header)
    start = 2
    for cat, svc, unit in services:
        ws.append([cat, svc, unit] + [None] * 9 + [None, ""])
    end = start + len(services) - 1
    # 給付費 = 量 × 単価 ÷ 1000（単位が「人」の場合は利用者数×単価）
    for i in range(len(services)):
        r = start + i
        for k, base in enumerate(("E", "H", "K")):        # 量の列
            gcol = ("F", "I", "L")[k]                     # 給付費の列
            ws[f"{gcol}{r}"] = f"=IF(OR({base}{r}=\"\",$M{r}=\"\"),\"\",{base}{r}*$M{r}/1000)"
    ws.append(["合計", "", ""] +
              [None, None, f"=SUM(F{start}:F{end})",
               None, None, f"=SUM(I{start}:I{end})",
               None, None, f"=SUM(L{start}:L{end})", None, ""])
    style_header(ws)
    for r in range(start, end + 1):
        for col in ("D", "E", "G", "H", "J", "K", "M"):
            ws[f"{col}{r}"].fill = IN_FILL
            ws[f"{col}{r}"].number_format = "#,##0"
        for col in ("F", "I", "L"):
            ws[f"{col}{r}"].fill = CALC_FILL
            ws[f"{col}{r}"].number_format = "#,##0"
    tot = end + 1
    for col in ("F", "I", "L"):
        ws[f"{col}{tot}"].fill = CALC_FILL
        ws[f"{col}{tot}"].number_format = "#,##0"
        ws[f"{col}{tot}"].font = Font(bold=True, size=9)
    body_style(ws, wrap_cols=(13,))
    widths(ws, [11, 30, 6] + [11, 11, 13] * 3 + [13, 30])
    ws.freeze_panes = "D2"
    ws.append([])
    ws.append([f"※ {label}。「量」は単位欄の単位（回・日・人）による月あたりの量に12を乗じた年間量を入れます。"])
    ws.append(["※ 単価は直近実績の給付費 ÷ 量で求め、報酬改定がある場合は改定率を乗じます（ケースC4）。"])
    ws.append(["※ 単位が「人」のサービスは、量の欄に延べ人数（月あたり人数×12）を入れます。"])
    return ws


# ---------------------------------------------------------------- 07_地域支援事業

def sheet_chiiki(wb):
    ws = wb.create_sheet("07_地域支援事業")
    ws.append(["区分", "事業", "令和7年度（実績）"] + YEARS + ["備考"])
    items = [
        ("総合事業", "介護予防・生活支援サービス事業（訪問型）", ""),
        ("総合事業", "介護予防・生活支援サービス事業（通所型）", ""),
        ("総合事業", "その他の生活支援サービス", ""),
        ("総合事業", "介護予防ケアマネジメント", ""),
        ("総合事業", "一般介護予防事業",
         "**令和5年度で1.0百万円と低水準。通いの場を第10期の最重点とするため増額を見込む**"),
        ("包括的支援事業", "地域包括支援センターの運営", ""),
        ("包括的支援事業", "在宅医療・介護連携推進事業", ""),
        ("包括的支援事業", "生活支援体制整備事業", ""),
        ("包括的支援事業", "認知症総合支援事業", ""),
        ("包括的支援事業", "地域ケア会議推進事業", ""),
        ("任意事業", "介護給付等費用適正化事業", ""),
        ("任意事業", "家族介護支援事業", ""),
        ("任意事業", "その他の任意事業", ""),
    ]
    start = 2
    for cat, ev, note_txt in items:
        ws.append([cat, ev, None, None, None, None, note_txt])
    end = start + len(items) - 1
    ws.append(["合計", "",
               f"=SUM(C{start}:C{end})", f"=SUM(D{start}:D{end})",
               f"=SUM(E{start}:E{end})", f"=SUM(F{start}:F{end})", ""])
    style_header(ws)
    for r in range(start, end + 1):
        for col in ("C", "D", "E", "F"):
            ws[f"{col}{r}"].fill = IN_FILL
            ws[f"{col}{r}"].number_format = "#,##0"
    for col in ("C", "D", "E", "F"):
        ws[f"{col}{end + 1}"].fill = CALC_FILL
        ws[f"{col}{end + 1}"].number_format = "#,##0"
        ws[f"{col}{end + 1}"].font = Font(bold=True, size=9)
    body_style(ws, wrap_cols=(6,))
    widths(ws, [16, 36, 15, 14, 14, 14, 52])
    ws.freeze_panes = "C2"
    ws.append([])
    ws.append(["※ 単位：千円。"])
    ws.append(["※ 総合事業には上限額（前年度実績×高齢者人口の伸び率）が設定されます。上限を超える場合は町と協議します。"])
    ws.append(["※ **見える化システムには令和3年度以降の総合事業の値がありません。**町の事業実績が必要です。"])
    return ws


# ---------------------------------------------------------------- 08_標準給付費

def sheet_hyojun(wb):
    ws = wb.create_sheet("08_標準給付費")
    ws.append(["項目"] + YEARS + ["3年計", "算式・備考"])
    rows = [
        ("介護給付費（05_見込量_介護の合計）", "svc_kaigo", "05_見込量_介護 の合計行"),
        ("予防給付費（06_見込量_予防の合計）", "svc_yobo", "06_見込量_予防 の合計行"),
        ("総給付費", "sum2", "介護給付費＋予防給付費"),
        ("特定入所者介護サービス費等給付額", "input", "補足給付。実績の伸びで見込む"),
        ("高額介護サービス費等給付額", "input", "同上"),
        ("高額医療合算介護サービス費等給付額", "input", "同上"),
        ("算定対象審査支払手数料", "input", "国保連への手数料"),
        ("標準給付費見込額", "std", "総給付費＋補足給付＋高額＋高額合算＋手数料"),
        ("地域支援事業費（07の合計）", "chiiki", "07_地域支援事業 の合計行"),
        ("総費用額", "total", "標準給付費見込額＋地域支援事業費"),
    ]
    start = 2
    for label, kind, note_txt in rows:
        ws.append([label, None, None, None, None, note_txt])
    # 行番号
    R = {label: start + i for i, (label, _, _) in enumerate(rows)}
    for ci, col in enumerate(("B", "C", "D")):
        y = ci  # 0-based 年度
        gcol = ("F", "I", "L")[y]
        ws[f"{col}{R['介護給付費（05_見込量_介護の合計）']}"] = \
            f"='05_見込量_介護'!{gcol}{2 + len(SVC_KAIGO)}"
        ws[f"{col}{R['予防給付費（06_見込量_予防の合計）']}"] = \
            f"='06_見込量_予防'!{gcol}{2 + len(SVC_YOBO)}"
        ws[f"{col}{R['総給付費']}"] = \
            f"={col}{R['介護給付費（05_見込量_介護の合計）']}+{col}{R['予防給付費（06_見込量_予防の合計）']}"
        ws[f"{col}{R['標準給付費見込額']}"] = (
            f"={col}{R['総給付費']}+{col}{R['特定入所者介護サービス費等給付額']}"
            f"+{col}{R['高額介護サービス費等給付額']}+{col}{R['高額医療合算介護サービス費等給付額']}"
            f"+{col}{R['算定対象審査支払手数料']}")
        ws[f"{col}{R['地域支援事業費（07の合計）']}"] = \
            f"='07_地域支援事業'!{('D', 'E', 'F')[y]}{2 + 13}"
        ws[f"{col}{R['総費用額']}"] = \
            f"={col}{R['標準給付費見込額']}+{col}{R['地域支援事業費（07の合計）']}"
    for label in R:
        r = R[label]
        ws[f"E{r}"] = f"=SUM(B{r}:D{r})"
    style_header(ws)
    calc_rows = {R['介護給付費（05_見込量_介護の合計）'], R['予防給付費（06_見込量_予防の合計）'],
                 R['総給付費'], R['標準給付費見込額'], R['地域支援事業費（07の合計）'], R['総費用額']}
    for label, r in R.items():
        for col in ("B", "C", "D", "E"):
            ws[f"{col}{r}"].number_format = "#,##0"
            ws[f"{col}{r}"].fill = CALC_FILL if (r in calc_rows or col == "E") else IN_FILL
    for r in (R['標準給付費見込額'], R['総費用額']):
        for col in ("A", "B", "C", "D", "E"):
            ws[f"{col}{r}"].font = Font(bold=True, size=9)
    body_style(ws, wrap_cols=(5,))
    widths(ws, [38, 16, 16, 16, 17, 46])
    ws.freeze_panes = "B2"
    ws.append([])
    ws.append(["※ 単位：千円。"])
    ws.append(["※ 第9期の標準給付費見込額（3年計）は3,566,486千円でした。第10期はこれを基礎に、"
               "認定者数の推計と実績の伸びで算定します。"])
    return ws, R


# ---------------------------------------------------------------- 09_保険料算定

def sheet_hokenryo(wb, hyojun_rows):
    ws = wb.create_sheet("09_保険料算定")
    ws.append(["#", "項目", "値", "単位", "算式・出典"])
    R_STD = hyojun_rows['標準給付費見込額']
    R_CHI = hyojun_rows['地域支援事業費（07の合計）']
    rows = [
        (1, "標準給付費見込額（3年計）", f"='08_標準給付費'!E{R_STD}", "千円", "08_標準給付費 より", "calc"),
        (2, "地域支援事業費（3年計）", f"='08_標準給付費'!E{R_CHI}", "千円", "08_標準給付費 より", "calc"),
        (3, "総費用額（3年計）", "=C2+C3", "千円", "①＋②", "calc"),
        (4, "第1号被保険者負担割合", None, "％",
         "国の基本指針による。第9期は23％。**第10期の割合は国の決定後に確定**", "in"),
        (5, "第1号被保険者負担分相当額", "=C4*C5/100", "千円", "③×④", "calc"),
        (6, "調整交付金相当額", "=C4*C8/100", "千円", "③×⑦", "calc"),
        (7, "調整交付金相当割合", None, "％", "国の基本指針による。標準は5％", "in"),
        (8, "調整交付金見込額", None, "千円",
         "**後期高齢者加入割合・所得段階別加入割合により算定。見える化システムで確認**", "in"),
        (9, "財政安定化基金拠出金見込額", None, "千円", "県の設定による。第9期は0円", "in"),
        (10, "財政安定化基金償還金", None, "千円", "借入がある場合。第9期は0円", "in"),
        (11, "準備基金取崩額", None, "千円",
         "**政策判断。第9期は35,000千円。ケースC5・C6で変える**", "in"),
        (12, "市町村特別給付費等", None, "千円", "町単独事業を保険料で賄う場合", "in"),
        (13, "保険料収納必要額", "=C6+C7-C9+C10+C11-C12+C13", "千円",
         "⑤＋⑥－⑧＋⑨＋⑩－⑪＋⑫", "calc"),
        (14, "予定保険料収納率", None, "％", "直近3年の実績による。**令和6・7年度決算が必要**", "in"),
        (15, "保険料賦課総額", "=IF(C15=0,\"\",C14/C15*100)", "千円", "⑬÷⑭", "calc"),
        (16, "所得段階別加入割合補正後被保険者数（3年計）", "='09_保険料算定'!C36", "人",
         "下表の合計。**所得段階別第1号被保険者数が必要**", "calc"),
        (17, "保険料基準額（年額）", "=IF(C17=0,\"\",C16*1000/C17)", "円", "⑮÷⑯", "key"),
        (18, "保険料基準額（月額）", "=IF(C18=\"\",\"\",C18/12)", "円", "⑰÷12", "key"),
        (19, "第9期の保険料基準額（月額）", 6600, "円", "第9期計画本文", "ref"),
        (20, "第9期との差", "=IF(C19=\"\",\"\",C19-C20)", "円", "⑱－⑲", "key"),
    ]
    for no, name, val, unit, src, kind in rows:
        ws.append([no, name, val, unit, src])
    style_header(ws)
    for i, (no, name, val, unit, src, kind) in enumerate(rows):
        r = 2 + i
        c = ws[f"C{r}"]
        c.number_format = "#,##0" if unit in ("千円", "人", "円") else "0.00"
        c.fill = {"in": IN_FILL, "calc": CALC_FILL, "key": KEY_FILL,
                  "ref": PatternFill("solid", fgColor="EEEEEE")}[kind]
        if kind == "key":
            ws[f"B{r}"].font = Font(bold=True, size=9)
            c.font = Font(bold=True, size=10)
    body_style(ws, wrap_cols=(4,))
    widths(ws, [5, 42, 18, 8, 60])
    ws.freeze_panes = "A2"

    # 所得段階別
    ws.append([])
    r0 = ws.max_row + 1
    ws.cell(r0, 1, "【所得段階別 加入割合補正後被保険者数】")
    ws.cell(r0, 1).font = Font(bold=True, size=10)
    ws.append(["段階", "対象", "保険料率", "被保険者数（3年計）", "補正後人数"])
    style_header(ws, ws.max_row)
    stages = [
        ("第1段階", "生活保護受給者、老齢福祉年金受給者、住民税非課税世帯かつ年金収入等80万円以下", 0.285),
        ("第2段階", "住民税非課税世帯かつ年金収入等80万円超120万円以下", 0.485),
        ("第3段階", "住民税非課税世帯かつ年金収入等120万円超", 0.685),
        ("第4段階", "本人非課税かつ世帯課税、年金収入等80万円以下", 0.90),
        ("第5段階", "本人非課税かつ世帯課税、年金収入等80万円超", 1.00),
        ("第6段階", "本人課税、合計所得金額120万円未満", 1.20),
        ("第7段階", "本人課税、合計所得金額120万円以上210万円未満", 1.30),
        ("第8段階", "本人課税、合計所得金額210万円以上320万円未満", 1.50),
        ("第9段階", "本人課税、合計所得金額320万円以上420万円未満", 1.70),
        ("第10段階", "本人課税、合計所得金額420万円以上520万円未満", 1.90),
        ("第11段階", "本人課税、合計所得金額520万円以上620万円未満", 2.10),
        ("第12段階", "本人課税、合計所得金額620万円以上720万円未満", 2.30),
        ("第13段階", "本人課税、合計所得金額720万円以上", 2.40),
    ]
    s0 = ws.max_row + 1
    for name, target, rate in stages:
        ws.append([name, target, rate, None, None])
    s1 = ws.max_row
    for r in range(s0, s1 + 1):
        ws[f"E{r}"] = f"=IF(D{r}=\"\",\"\",C{r}*D{r})"
        ws[f"C{r}"].fill = IN_FILL
        ws[f"D{r}"].fill = IN_FILL
        ws[f"E{r}"].fill = CALC_FILL
        ws[f"C{r}"].number_format = "0.000"
        ws[f"D{r}"].number_format = "#,##0"
        ws[f"E{r}"].number_format = "#,##0.0"
    ws.append(["合計", "", "", f"=SUM(D{s0}:D{s1})", f"=SUM(E{s0}:E{s1})"])
    tot = ws.max_row
    for col in ("D", "E"):
        ws[f"{col}{tot}"].fill = CALC_FILL
        ws[f"{col}{tot}"].font = Font(bold=True, size=9)
        ws[f"{col}{tot}"].number_format = "#,##0.0"
    ws[f"C{2 + 15}"] = f"=E{tot}"     # ⑯ を合計にリンク
    body_style(ws, min_row=r0, wrap_cols=(1,))
    ws.append([])
    ws.append(["※ 保険料率は第9期の13段階を仮置きしています。第10期の段階設定・乗率は国の基本指針を受けて確定します。"])
    ws.append(["※ 補正後人数＝保険料率×被保険者数。3年分の合計を入れます。"])
    ws.append(["※ **所得段階別第1号被保険者数は町からの受領が必要です。**"])
    return ws


# ---------------------------------------------------------------- 10_ケース比較

def sheet_compare(wb):
    ws = wb.create_sheet("10_ケース比較")
    ws.append(["ケース", "認定者数の前提", "標準給付費（3年計・千円）", "地域支援事業費（千円）",
               "総費用額（千円）", "基金取崩額（千円）", "保険料基準額（月額・円）",
               "第9期との差（円）", "備考"])
    for c in ["C1 標準（低位）", "C2 標準（高位）", "C3 施設整備あり",
              "C4 制度改正の影響", "C5 基金活用", "C6 保険料据置（逆算）"]:
        ws.append([c, None, None, None, None, None, None, None, ""])
    style_header(ws)
    for r in range(2, 8):
        for col in "BCDEFGH":
            ws[f"{col}{r}"].fill = IN_FILL
            ws[f"{col}{r}"].number_format = "#,##0"
    body_style(ws, wrap_cols=(1, 8))
    widths(ws, [20, 30, 20, 19, 18, 18, 20, 17, 40])
    ws.freeze_panes = "B2"
    ws.append([])
    ws.append(["※ 各ケースは09_保険料算定の入力値を変えて算定し、結果を本表に転記します。"
               "**ブックを複製してケースごとに保存する運用とします。**"])
    ws.append(["※ **C1とC2の差が、認定者数の確認が済むまでの不確実性の幅です。**"
               "この幅を協議会に示したうえで、確認結果を待って1本に絞ります。"])
    ws.append(["※ 協議会に諮る際は、保険料の水準だけでなく、"
               "繰越金182.0百万円・基金純増88.1百万円という財政余剰の還元方針を併せて示します。"])
    return ws


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_howto(wb)
    sheet_cases(wb)
    sheet_population(wb)
    sheet_nintei(wb)
    sheet_jukyu(wb)
    sheet_mikomi(wb, "05_見込量_介護", SVC_KAIGO, "介護給付（要介護1〜5）")
    sheet_mikomi(wb, "06_見込量_予防", SVC_YOBO, "予防給付（要支援1・2）")
    sheet_chiiki(wb)
    _, hyojun_rows = sheet_hyojun(wb)
    sheet_hokenryo(wb, hyojun_rows)
    sheet_compare(wb)
    path = OUT / f"小野町_第10期_見込量保険料算定シート_{ASOF}.xlsx"
    wb.save(path)
    print("出力:", path)
    print("  シート:", wb.sheetnames)


if __name__ == "__main__":
    main()
