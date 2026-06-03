# -*- coding: utf-8 -*-
"""
福祉計画 ワードコラム部品 管理表ジェネレータ
- 差し込み印刷で可変できるフィールド設計
- 共通項目と各計画での使用例を整理
- 計画ごとのブックを作成（高齢者/障がい/こども/共通/マスター）
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# 出力先: 環境変数 OUT_DIR で上書き可。未指定ならスクリプトと同階層の output/
OUT_DIR = os.environ.get(
    "OUT_DIR",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "output"),
)
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# 配色（画像のカラー目安に準拠）
# ============================================================
COLORS = {
    "こども": "1F77B4",   # 青
    "高齢者": "2CA02C",   # 緑
    "障がい": "FF7F0E",   # オレンジ
    "共通":   "7F4FBF",   # 紫
    "header": "1F3864",
    "subhead":"2E75B6",
    "band":   "DDEBF7",
    "alt":    "F7FAFC",
    "note":   "FFF3F3",
    "white":  "FFFFFF",
}

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================
# 部品マスタ（画像の内容を構造化）
# 共通項目: 部品ID / 分類 / 部品名 / アイコン / 主な用途 / 使用例 / 推奨レイアウト / 使用頻度 / 配色 / 対象計画 / 差込タイトル / 差込本文 / 備考
# ============================================================
COMPONENTS = [
    # --- 基本コラム部品（汎用） ---
    {"id":"BC-01","cat":"基本","name":"ポイント","icon":"電球","use":"計画の要点や重要なポイントを簡潔にまとめる",
     "ex":"基本方針、重点事項、留意点 等","layout":"枠＋アイコン上","freq":5,"color":"共通","target":"全計画",
     "title_ex":"〇〇計画推進のポイント","body_ex":"本計画では◯◯を重点に推進します。…","note":"KPI整理にも活用"},
    {"id":"BC-02","cat":"基本","name":"コラム","icon":"吹き出し","use":"本文の補足説明や背景情報、豆知識などを掲載",
     "ex":"制度説明、背景解説、考え方 等","layout":"帯＋アイコン左","freq":5,"color":"共通","target":"全計画",
     "title_ex":"〇〇とは？","body_ex":"〇〇は◯年に制定された制度で…","note":"最も汎用的に使用"},
    {"id":"BC-03","cat":"基本","name":"事例紹介","icon":"家","use":"地域や他自治体の取り組み事例を紹介",
     "ex":"活動事例、先進事例、参考事例 等","layout":"左ライン＋アイコン","freq":4,"color":"共通","target":"全計画",
     "title_ex":"〇〇市の取り組み事例","body_ex":"◯◯市では地域包括ケアの一環として…","note":"出典明記が必要"},
    {"id":"BC-04","cat":"基本","name":"解説","icon":"書類","use":"制度や用語の解説、専門的内容をわかりやすく説明",
     "ex":"制度解説、用語解説、法令解説 等","layout":"角丸＋影付き","freq":4,"color":"共通","target":"全計画",
     "title_ex":"用語解説：〇〇","body_ex":"〇〇とは、△△を指します。…","note":"巻末用語集と連携"},
    {"id":"BC-05","cat":"基本","name":"データの見方","icon":"グラフ","use":"図表・グラフの読み方や注意点を説明する補足",
     "ex":"図表の補足、数値の見方 等","layout":"二重枠","freq":5,"color":"共通","target":"全計画",
     "title_ex":"このグラフの見方","body_ex":"縦軸は◯◯、横軸は◯◯を示しています。…","note":"図表直下に配置"},
    {"id":"BC-06","cat":"基本","name":"注意・留意点","icon":"！","use":"計画推進にあたり注意すべき事項や留意点を示す",
     "ex":"留意事項、リスク、注意喚起 等","layout":"背景色ベタ","freq":4,"color":"共通","target":"全計画",
     "title_ex":"推進にあたっての留意点","body_ex":"〇〇の実施にあたっては…にご留意ください。","note":"視認性を高める"},

    # --- 高齢者計画向け ---
    {"id":"SR-01","cat":"高齢者","name":"認知症サポート","icon":"カップ","use":"認知症の理解や支援、地域での見守り・支え合いに関するコラム",
     "ex":"認知症施策の解説 等","layout":"帯＋アイコン左","freq":5,"color":"高齢者","target":"高齢者計画",
     "title_ex":"認知症の方を地域で支える","body_ex":"認知症サポーターは…として活動しています。","note":"オレンジリング啓発連動"},
    {"id":"SR-02","cat":"高齢者","name":"介護予防のヒント","icon":"歩く人","use":"介護予防やフレイル予防に役立つポイントを紹介",
     "ex":"予防の取り組み紹介 等","layout":"枠＋アイコン上","freq":5,"color":"高齢者","target":"高齢者計画",
     "title_ex":"今日からできるフレイル予防","body_ex":"運動・栄養・社会参加の3本柱で…","note":"イラスト併用推奨"},
    {"id":"SR-03","cat":"高齢者","name":"在宅生活を支える","icon":"家","use":"在宅で安心して暮らすためのサービスや支援に関するコラム",
     "ex":"在宅支援の紹介 等","layout":"左ライン＋アイコン","freq":4,"color":"高齢者","target":"高齢者計画",
     "title_ex":"在宅生活を支えるサービス","body_ex":"訪問介護・訪問看護など…","note":"サービス一覧と連携"},

    # --- 障がい計画向け ---
    {"id":"DS-01","cat":"障がい","name":"合理的配慮とは","icon":"ハート手","use":"合理的配慮の考え方や具体例をわかりやすく解説",
     "ex":"制度・考え方の解説 等","layout":"角丸＋影付き","freq":5,"color":"障がい","target":"障がい計画",
     "title_ex":"合理的配慮を理解する","body_ex":"合理的配慮とは、障害のある人が…","note":"差別解消法との関係明記"},
    {"id":"DS-02","cat":"障がい","name":"コミュニケーション支援","icon":"吹き出し","use":"意思疎通支援やコミュニケーション方法の工夫を紹介",
     "ex":"支援方法の紹介 等","layout":"帯＋アイコン左","freq":4,"color":"障がい","target":"障がい計画",
     "title_ex":"伝わるコミュニケーションの工夫","body_ex":"手話、要約筆記、絵カードなど…","note":"絵記号併記推奨"},
    {"id":"DS-03","cat":"障がい","name":"地域での共生","icon":"人々","use":"誰もが安心して暮らせる共生社会づくりに向けた取り組みを紹介",
     "ex":"共生の取り組み紹介 等","layout":"左ライン＋アイコン","freq":4,"color":"障がい","target":"障がい計画",
     "title_ex":"地域共生社会の実現に向けて","body_ex":"住民同士の支え合いと…","note":"地域包括との接続"},

    # --- こども計画向け ---
    {"id":"CH-01","cat":"こども","name":"子どもの声","icon":"こども顔","use":"アンケートや意見交換等で寄せられた子どもの声を掲載",
     "ex":"調査結果、意見の紹介 等","layout":"枠＋アイコン上","freq":5,"color":"こども","target":"こども計画",
     "title_ex":"子どもたちの声","body_ex":"「もっと安心して遊べる場所がほしい」…","note":"原文ニュアンス保持"},
    {"id":"CH-02","cat":"こども","name":"子育てワンポイント","icon":"親子","use":"子育てに役立つ情報やちょっとしたヒントを紹介",
     "ex":"子育て支援情報 等","layout":"帯＋アイコン左","freq":4,"color":"こども","target":"こども計画",
     "title_ex":"子育てワンポイントアドバイス","body_ex":"乳幼児期の関わり方のコツは…","note":"親しみやすい文体"},
    {"id":"CH-03","cat":"こども","name":"成長を支える地域","icon":"家と子","use":"地域で子どもの成長を見守り支える取り組みを紹介",
     "ex":"地域活動の紹介 等","layout":"左ライン＋アイコン","freq":4,"color":"こども","target":"こども計画",
     "title_ex":"地域で育てる子どもの未来","body_ex":"放課後子ども教室や見守り活動…","note":"地域組織と連携"},
]

# レイアウトバリエーション
LAYOUTS = [
    {"name":"帯＋アイコン左","desc":"標準的に最も使うパターン","best_for":"コラム、子育てワンポイント等"},
    {"name":"枠＋アイコン上","desc":"枠で囲み、アイコンを上に配置","best_for":"ポイント、介護予防のヒント等"},
    {"name":"左ライン＋アイコン","desc":"縦ラインでスッキリ見せる","best_for":"事例紹介、地域取組等"},
    {"name":"角丸＋影付き","desc":"やわらかく強調したいときに","best_for":"解説、合理的配慮等"},
    {"name":"二重枠","desc":"重要な補足を強調したいときに","best_for":"データの見方"},
    {"name":"背景色ベタ","desc":"注意喚起を目立たせたいときに","best_for":"注意・留意点"},
]

# デザイン統一ポイント
DESIGN_RULES = [
    "アイコンは同じ線幅・同じスタイルで統一",
    "余白（内側8〜12mm程度）を確保して読みやすく",
    "色は淡いトーンで、本文より主張させない",
    "1ページにコラムは最大2個程度が基本",
    "フォントは本文と同じ（游ゴシック等）を使用",
    "印刷したときに見やすい濃さで設計",
]

RECOMMEND_SIZE = [
    ("アイコンサイズ","10〜12mm"),
    ("コラム内の本文","9〜10.5pt"),
    ("タイトル（帯部分）","10〜11pt"),
    ("余白（内側）","上下左右 8〜12mm"),
    ("角丸半径","2〜4mm"),
]

# 差込フィールド定義
MERGE_FIELDS = [
    ("部品ID","システム識別子（例: BC-01）","必須"),
    ("分類","基本／高齢者／障がい／こども／共通","必須"),
    ("部品名","ポイント、コラム など","必須"),
    ("アイコン","アイコン名（記号）","推奨"),
    ("タイトル","コラム見出し（差込で可変）","必須"),
    ("本文","コラム本文（差込で可変）","必須"),
    ("出典","出典・参考文献等","任意"),
    ("レイアウト","推奨レイアウトパターン","必須"),
    ("配色","計画別カラー","必須"),
    ("対象計画","使用する計画名","必須"),
    ("使用頻度","★1〜5","参考"),
    ("備考","運用上の留意点","任意"),
]


# ============================================================
# 共通スタイル関数
# ============================================================
def style_title(cell, text, fill=COLORS["header"], font_color="FFFFFF", size=14):
    cell.value = text
    cell.font = Font(name="游ゴシック", size=size, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)

def style_subhead(cell, text, fill=COLORS["subhead"]):
    cell.value = text
    cell.font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal="center")

def style_header_row(ws, row, headers, fill=COLORS["subhead"]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER

def style_data_cell(cell, alt=False):
    cell.font = Font(name="游ゴシック", size=10)
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
    cell.border = BORDER
    if alt:
        cell.fill = PatternFill("solid", fgColor=COLORS["alt"])

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 各シート作成関数
# ============================================================
def add_cover_sheet(wb, plan_name, scope_desc):
    ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = "表紙・凡例"
    set_col_widths(ws, [22, 60, 18, 18])
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:D1")
    style_title(ws["A1"], f"福祉計画コラム部品 管理表  ／  {plan_name}")

    ws["A3"] = "計画名"
    ws["B3"] = plan_name
    ws["A4"] = "対象範囲"
    ws["B4"] = scope_desc
    ws["A5"] = "更新日"
    ws["B5"] = "=TODAY()"
    ws["B5"].number_format = "yyyy/mm/dd"
    ws["A6"] = "差込印刷対応"
    ws["B6"] = "可（Word差し込み印刷ウィザードで本ブックの『差込データ』シートを指定）"

    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(name="游ゴシック", size=10, bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=COLORS["band"])
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).font = Font(name="游ゴシック", size=10)
        ws.cell(row=r, column=2).border = BORDER

    # 凡例
    ws.merge_cells("A8:D8")
    style_subhead(ws["A8"], "シート構成")
    sheets_info = [
        ("差込データ","Wordへ差し込むデータ本体（1行=1コラム）"),
        ("共通項目定義","差込フィールドの一覧と必須／任意区分"),
        ("部品マスタ","部品の用途・使用例・推奨レイアウト一覧"),
        ("レイアウト一覧","6パターンのレイアウトと使い分け"),
        ("デザインルール","配色・余白・フォントの統一ルール"),
        ("使い分けガイド","部品名／主な用途／使用頻度（★）"),
    ]
    style_header_row(ws, 9, ["シート名","内容"])
    for i, (n, d) in enumerate(sheets_info, 10):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=d)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        for col in range(1, 5):
            style_data_cell(ws.cell(row=i, column=col), alt=(i % 2 == 0))

    # 配色凡例
    start = 10 + len(sheets_info) + 2
    ws.merge_cells(start_row=start-1, start_column=1, end_row=start-1, end_column=4)
    style_subhead(ws.cell(row=start-1, column=1), "計画別カラー目安")
    palette = [("こども計画","信頼・安心感・未来・つながり","こども"),
               ("高齢者計画","安心・健康・いきがい・やさしさ","高齢者"),
               ("障がい計画","支援・共生・理解・温かさ","障がい"),
               ("共通","協働・連携・包括・全体調和","共通")]
    style_header_row(ws, start, ["計画","イメージ","カラー","HEX"])
    for i, (n, img, key) in enumerate(palette, start+1):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=img)
        c = ws.cell(row=i, column=3, value=key)
        c.fill = PatternFill("solid", fgColor=COLORS[key])
        c.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=4, value="#"+COLORS[key])
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = BORDER
            if col != 3:
                style_data_cell(ws.cell(row=i, column=col), alt=False)

    ws.sheet_view.showGridLines = False
    return ws


def add_merge_data_sheet(wb, components):
    ws = wb.create_sheet("差込データ")
    headers = ["部品ID","分類","部品名","アイコン","タイトル","本文","出典","レイアウト","配色","対象計画","使用頻度","備考"]
    set_col_widths(ws, [10, 8, 16, 10, 28, 50, 16, 18, 10, 14, 10, 28])
    ws.row_dimensions[1].height = 28
    style_header_row(ws, 1, headers)

    for i, c in enumerate(components, 2):
        row = [c["id"], c["cat"], c["name"], c["icon"], c["title_ex"], c["body_ex"], "",
               c["layout"], c["color"], c["target"], "★"*c["freq"], c["note"]]
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            style_data_cell(cell, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 42
        # 配色セルを塗る
        color_cell = ws.cell(row=i, column=9)
        if c["color"] in COLORS:
            color_cell.fill = PatternFill("solid", fgColor=COLORS[c["color"]])
            color_cell.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
            color_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # オートフィルタ
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(components)+1}"
    return ws


def add_field_def_sheet(wb):
    ws = wb.create_sheet("共通項目定義")
    set_col_widths(ws, [18, 50, 12])
    ws.merge_cells("A1:C1")
    style_title(ws["A1"], "差込フィールド定義（共通項目）")
    style_header_row(ws, 3, ["フィールド名","説明","区分"])
    for i, (n, d, k) in enumerate(MERGE_FIELDS, 4):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=d)
        ws.cell(row=i, column=3, value=k)
        for col in range(1, 4):
            style_data_cell(ws.cell(row=i, column=col), alt=(i % 2 == 0))
        # 区分セルの色
        kc = ws.cell(row=i, column=3)
        kc.alignment = Alignment(horizontal="center", vertical="center")
        if k == "必須":
            kc.fill = PatternFill("solid", fgColor="C00000")
            kc.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
        elif k == "推奨":
            kc.fill = PatternFill("solid", fgColor="ED7D31")
            kc.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.sheet_view.showGridLines = False
    return ws


def add_master_sheet(wb, components):
    ws = wb.create_sheet("部品マスタ")
    headers = ["部品ID","分類","部品名","主な用途","使用例","推奨レイアウト","使用頻度","対象計画"]
    set_col_widths(ws, [10, 8, 16, 38, 28, 18, 10, 14])
    ws.row_dimensions[1].height = 28
    style_header_row(ws, 1, headers)
    for i, c in enumerate(components, 2):
        row = [c["id"], c["cat"], c["name"], c["use"], c["ex"], c["layout"], "★"*c["freq"], c["target"]]
        for col, v in enumerate(row, 1):
            style_data_cell(ws.cell(row=i, column=col, value=v), alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(components)+1}"
    ws.sheet_view.showGridLines = False
    return ws


def add_layout_sheet(wb):
    ws = wb.create_sheet("レイアウト一覧")
    set_col_widths(ws, [22, 40, 36])
    ws.merge_cells("A1:C1")
    style_title(ws["A1"], "レイアウトバリエーション（6パターン）")
    style_header_row(ws, 3, ["パターン名","特徴","適した部品"])
    for i, l in enumerate(LAYOUTS, 4):
        ws.cell(row=i, column=1, value=l["name"])
        ws.cell(row=i, column=2, value=l["desc"])
        ws.cell(row=i, column=3, value=l["best_for"])
        for col in range(1, 4):
            style_data_cell(ws.cell(row=i, column=col), alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 28
    ws.sheet_view.showGridLines = False
    return ws


def add_design_sheet(wb):
    ws = wb.create_sheet("デザインルール")
    set_col_widths(ws, [40, 30])
    ws.merge_cells("A1:B1")
    style_title(ws["A1"], "デザイン統一ルール ／ 推奨サイズ")
    ws.merge_cells("A3:B3")
    style_subhead(ws["A3"], "デザイン統一のポイント")
    for i, r in enumerate(DESIGN_RULES, 4):
        ws.cell(row=i, column=1, value=f"✓ {r}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        style_data_cell(ws.cell(row=i, column=1), alt=(i % 2 == 0))

    base = 4 + len(DESIGN_RULES) + 1
    ws.merge_cells(start_row=base, start_column=1, end_row=base, end_column=2)
    style_subhead(ws.cell(row=base, column=1), "推奨サイズ（Wordでの目安）")
    style_header_row(ws, base+1, ["項目","目安"])
    for i, (k, v) in enumerate(RECOMMEND_SIZE, base+2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        for col in range(1, 3):
            style_data_cell(ws.cell(row=i, column=col), alt=(i % 2 == 0))
    ws.sheet_view.showGridLines = False
    return ws


def add_guide_sheet(wb, components):
    ws = wb.create_sheet("使い分けガイド")
    set_col_widths(ws, [16, 40, 16])
    ws.merge_cells("A1:C1")
    style_title(ws["A1"], "部品の使い分けガイド")
    style_header_row(ws, 3, ["部品名","主な用途","使用頻度"])
    for i, c in enumerate(components, 4):
        ws.cell(row=i, column=1, value=c["name"])
        ws.cell(row=i, column=2, value=c["use"])
        ws.cell(row=i, column=3, value="★"*c["freq"])
        for col in range(1, 4):
            style_data_cell(ws.cell(row=i, column=col), alt=(i % 2 == 0))
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 32
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# ブック生成
# ============================================================
def build_book(filename, plan_name, scope_desc, components):
    wb = Workbook()
    add_cover_sheet(wb, plan_name, scope_desc)
    add_merge_data_sheet(wb, components)
    add_field_def_sheet(wb)
    add_master_sheet(wb, components)
    add_layout_sheet(wb)
    add_design_sheet(wb)
    add_guide_sheet(wb, components)
    path = os.path.join(OUT_DIR, filename)
    wb.save(path)
    print(f"  ✓ 作成: {path}  （{len(components)}部品）")
    return path


def build_master_book():
    """全データ + 計画別シート分岐を含むマスターブック"""
    wb = Workbook()
    add_cover_sheet(wb, "全計画マスター", "高齢者・障がい・こども・共通の全部品を一括管理")
    add_merge_data_sheet(wb, COMPONENTS)
    add_field_def_sheet(wb)
    add_master_sheet(wb, COMPONENTS)
    add_layout_sheet(wb)
    add_design_sheet(wb)
    add_guide_sheet(wb, COMPONENTS)

    # 計画別の差込データシートを追加（マスターから絞り込み）
    for plan_key, plan_name in [("高齢者","高齢者計画 差込"),
                                  ("障がい","障がい計画 差込"),
                                  ("こども","こども計画 差込")]:
        ws = wb.create_sheet(plan_name)
        headers = ["部品ID","分類","部品名","アイコン","タイトル","本文","レイアウト","配色","使用頻度"]
        set_col_widths(ws, [10, 8, 16, 10, 28, 50, 18, 10, 10])
        ws.row_dimensions[1].height = 28
        style_header_row(ws, 1, headers)
        target = [c for c in COMPONENTS if c["cat"] == plan_key or c["cat"] == "基本"]
        for i, c in enumerate(target, 2):
            row = [c["id"], c["cat"], c["name"], c["icon"], c["title_ex"], c["body_ex"],
                   c["layout"], c["color"], "★"*c["freq"]]
            for col, v in enumerate(row, 1):
                style_data_cell(ws.cell(row=i, column=col, value=v), alt=(i % 2 == 0))
            ws.row_dimensions[i].height = 42
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(target)+1}"
        ws.sheet_view.showGridLines = False

    path = os.path.join(OUT_DIR, "00_全計画マスター管理表.xlsx")
    wb.save(path)
    print(f"  ✓ 作成: {path}  （マスター: {len(COMPONENTS)}部品）")
    return path


if __name__ == "__main__":
    print("【1】マスター管理表を作成")
    build_master_book()

    print("\n【2】計画別ブックを作成")
    # 共通＝基本コラム部品のみ
    common = [c for c in COMPONENTS if c["cat"] == "基本"]
    senior = [c for c in COMPONENTS if c["cat"] in ("基本","高齢者")]
    disab  = [c for c in COMPONENTS if c["cat"] in ("基本","障がい")]
    child  = [c for c in COMPONENTS if c["cat"] in ("基本","こども")]

    build_book("01_共通_基本コラム部品.xlsx", "共通（基本コラム部品）",
               "全計画で共通利用する6種類の基本部品", common)
    build_book("02_高齢者介護保険事業計画.xlsx", "高齢者介護保険事業計画",
               "基本6部品＋高齢者計画向け3部品", senior)
    build_book("03_障がい福祉計画.xlsx", "障がい福祉計画",
               "基本6部品＋障がい計画向け3部品", disab)
    build_book("04_こども計画.xlsx", "こども計画",
               "基本6部品＋こども計画向け3部品", child)

    print("\n完了。出力先: " + OUT_DIR)
