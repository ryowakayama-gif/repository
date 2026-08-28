# -*- coding: utf-8 -*-
"""令和6年度決算書（介護保険特別会計）の受領点検.

令和8年8月28日に受領した「令和6年度 大雪地区広域連合
各会計歳入歳出決算書」のうち、介護保険特別会計の部の点検結果。

本受領により、年報の様式4により確定した令和6年度の決算額が、
決算書という別の出典によって裏づけられた。
あわせて介護給付費準備基金の前年度末（第8期末）残高が判明し、
第9期における基金の推移が確定した。

シート構成
  00_受領資料の概要   決算書の構成と、この資料により確定すること
  01_歳入歳出決算     款項別の予算現額・調定額・収入済額・支出済額
  02_基金と実質収支   財産に関する調書、実質収支に関する調書
  03_年報との対照     年報の様式4との対照22項目と、相違3件
  04_成果品への反映   更新する箇所と、なお必要な資料
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_kessan_r6 as K
import data_nenpo as N

FONT = "游ゴシック"
NAVY = "1F4E78"
HEAD = "5B9BD5"
IN_Y = "FFF2CC"
OK_G = "E2EFDA"
NG_O = "FCE4D6"
MID_B = "DEEBF7"
GRAY = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 56
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=22, align=None, numfmt="#,##0",
         bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.border = BORDER
        ha = (align or {}).get(i, "left" if isinstance(v, str) else "right")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if numfmt and not isinstance(v, str):
            c.number_format = numfmt
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8, height=88):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_受領資料の概要",
           "令和6年度決算書（介護保険特別会計）の受領点検",
           "令和8年8月28日に受領した「令和6年度 大雪地区広域連合 "
           "各会計歳入歳出決算書」（令和7年12月22日提出）のうち、"
           "介護保険特別会計の部の点検結果です。"
           "年報の様式4により確定していた決算額が、"
           "決算書という別の出典によって裏づけられました。",
           [4, 24, 32, 44, 34, 12], freeze="A5")

r = lead(ws, 4, "【1　受領した資料の構成】", 6)
r = header(ws, r, ["No.", "会計", "本業務との関係", "収録内容",
                   "点検の対象", "判定"])
for no, kaikei, kankei, cont, taisho, hh in [
    (1, "一般会計", "間接",
     "構成3町からの分担金及び負担金、広域連合の運営費",
     "介護保険特別会計への繰入金の原資となるが、"
     "国民健康保険・後期高齢者医療の分を含むため区分できない。"
     "本点検では扱わない。", "対象外"),
    (2, "介護保険特別会計", "直接",
     "歳入歳出決算書（款項別）、事項別明細書、"
     "実質収支に関する調書、財産に関する調書",
     "本点検の対象です。款項別の決算額と基金の残高を確定しました。",
     "点検済"),
    (3, "国民健康保険特別会計", "なし", "―",
     "本業務の対象外です。", "対象外"),
    (4, "後期高齢者医療特別会計", "なし", "―",
     "本業務の対象外です。", "対象外"),
]:
    fl = {6: OK_G if hh == "点検済" else GRAY}
    r = body(ws, r, [no, kaikei, kankei, cont, taisho, hh], fl, height=56,
             align={1: "center", 3: "center", 6: "center"}, numfmt=None)

r += 1
r = lead(ws, r, "【2　この資料により確定すること】", 6)
r = header(ws, r, ["No.", "確定すること", "値", "これまでの状態", "反映先",
                   "重要度"])
for no, kk, atai, mae, saki, lv in [
    (1, "介護給付費準備基金の第9期開始時点の残高",
     "令和5年度末（第8期末）231,456,821円。"
     "令和6年度中の積立232,197円・取崩0円により、"
     "令和6年度末231,689,018円",
     "令和6年度末の残高のみが分かっており、"
     "第9期に取り崩したかどうかを確定できなかった",
     "素案 第6章第6節\n将来推計 第3段階", "最高"),
    (2, "令和6年度の決算額の裏づけ",
     "款項別の決算額が年報の様式4と22項目すべて一致",
     "年報の様式4のみによっていた。"
     "国への報告値と決算値が一致するかを確認できていなかった",
     "素案 第6章第3節\n年報月報の受領点検", "高"),
    (3, "歳出の款項の構成",
     "7款16項。地域支援事業費は4項"
     "（介護予防・生活支援サービス事業費、包括的支援事業・任意事業費、"
     "保険者機能強化推進事業、保険者努力支援事業）",
     "年報の様式4の科目名によっていた。"
     "決算の科目との対応が確認できていなかった",
     "素案 第6章第3節\n資料提供依頼 05シート", "高"),
    (4, "不用額と予算執行の状況",
     "歳出の不用額128,956,046円（予算現額の3.79％）。"
     "うち予備費56,104,000円は全額不用",
     "把握していなかった",
     "素案 第6章第3節", "中"),
    (5, "基金の正式名称",
     "「介護保険事業財政調整基金」。"
     "年報は「介護給付費準備基金」と呼んでいる",
     "年報の名称のみによっていた",
     "素案 第6章第6節\n確認事項", "中"),
]:
    fl = {6: NG_O if lv == "最高" else (IN_Y if lv == "高" else None)}
    r = body(ws, r, [no, kk, atai, mae, saki, lv], fl, height=68,
             align={1: "center", 6: "center"}, numfmt=None)

note(ws, r + 1,
     "注1）決算書はスキャンしたPDF（85ページ）です。"
     "画像を目視で確認して値を確定しました。"
     "款の合計と項の合計、歳入合計・歳出合計、"
     "実質収支に関する調書との整合をすべて確認しています。\n"
     "注2）決算書は議会に提出された確定値です。"
     "年報は国への報告値であり、両者が一致することの確認は"
     "第三者検証の観点から意味があります。\n"
     "注3）本点検には個人が特定される情報は含まれていません。"
     "決算書に記載のある職名・氏名は成果品に収録していません。",
     6, 88)

# ============================================================ 01
ws = sheet("01_歳入歳出決算",
           "介護保険特別会計の款項別の決算額（令和6年度）",
           "決算書の歳入歳出決算書によります。"
           "款の合計と項の合計、歳入合計・歳出合計の整合を確認済みです。",
           [40, 18, 18, 18, 16, 16, 12], freeze="B5")

r = lead(ws, 4, "【1　歳入】", 7)
r = header(ws, r, ["款・項", "予算現額", "調定額", "収入済額",
                   "不納欠損額", "収入未済額", "構成比"])
ZEN = K.SAINYU["歳入合計"][2]
for k, v in K.SAINYU.items():
    kan = "／" not in k
    fl = {1: MID_B} if kan else None
    if k == "歳入合計":
        fl = {i: MID_B for i in range(1, 8)}
    nm = k if kan else "　　" + k.split("／")[1]
    r = body(ws, r, [nm, v[0], v[1], v[2], v[3], v[4],
                     v[2] / ZEN if ZEN else ""], fl,
             bold=(k == "歳入合計"))
    ws.cell(row=r - 1, column=7).number_format = "0.0%"

r += 1
r = lead(ws, r, "【2　歳出】", 7)
r = header(ws, r, ["款・項", "予算現額", "支出済額", "翌年度繰越額",
                   "不用額", "執行率", "構成比"])
ZES = K.SAISHUTSU["歳出合計"][1]
for k, v in K.SAISHUTSU.items():
    kan = "／" not in k
    fl = {1: MID_B} if kan else None
    if k == "歳出合計":
        fl = {i: MID_B for i in range(1, 8)}
    nm = k if kan else "　　" + k.split("／")[1]
    r = body(ws, r, [nm, v[0], v[1], v[2], v[3],
                     v[1] / v[0] if v[0] else "",
                     v[1] / ZES if ZES else ""], fl,
             bold=(k == "歳出合計"))
    ws.cell(row=r - 1, column=6).number_format = "0.0%"
    ws.cell(row=r - 1, column=7).number_format = "0.0%"

r += 1
r = lead(ws, r, "【3　所見】", 7)
r = header(ws, r, ["No.", "事項", "内容", "", "", "", ""])
for i, (ji, cont) in enumerate([
    ("歳入の構成",
     "支払基金交付金（第2号被保険者の負担分）が24.4％で最も大きく、"
     "保険料（第1号被保険者）は19.1％です。"
     "国庫支出金24.9％、道支出金13.4％、繰入金14.2％が続きます。"),
    ("基金繰入金は0円",
     "予算では5,946,000円を計上していましたが、"
     "実際の取崩しは0円でした。"
     "第9期の初年度は基金を用いずに収支が成立しています。"),
    ("保険給付費の執行率",
     "保険給付費の執行率は97.9％で、不用額64,342,012円が生じています。"
     "うち特定入所者介護等サービス費が8,502,554円（執行率89.9％）と"
     "最も大きく、次いで高額医療合算介護等サービス費が"
     "2,486,171円（執行率78.9％）です。"),
    ("予備費と公債費は全額不用",
     "予備費56,104,000円・公債費247,000円はいずれも執行がありません。"
     "不用額128,956,046円のうち予備費が43.5％を占めます。"),
    ("一般介護予防事業費の項がない",
     "地域支援事業費は4項ですが、"
     "「一般介護予防事業費」の項がありません。"
     "年報の様式4でも同項は0円です。"
     "総合事業の実施状況に関する調査では"
     "介護予防普及啓発事業を3町とも実施しているため、"
     "科目の立て方の確認を要します（03シート）。"),
], start=1):
    r = body(ws, r, [i, ji, cont, "", "", "", ""], {2: MID_B}, height=56,
             align={1: "center"})
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=7)

note(ws, r + 1,
     "注1）執行率は支出済額÷予算現額です。"
     "構成比は歳入合計（収入済額）又は歳出合計（支出済額）に対する割合です。\n"
     "注2）歳入の調定額3,416,485,402円と収入済額3,414,476,915円の差"
     "2,008,487円は、保険料の不納欠損額646,812円と"
     "収入未済額1,361,675円の合計です。\n"
     "注3）款の合計と項の合計、歳入合計・歳出合計の整合を"
     "すべて確認しています。", 7, 76)

# ============================================================ 02
ws = sheet("02_基金と実質収支",
           "介護給付費準備基金と実質収支（令和6年度）",
           "決算書の財産に関する調書及び実質収支に関する調書によります。"
           "第10期の保険料算定における取崩額の設定に直結します。",
           [36, 22, 22, 22, 46], freeze="B5")

r = lead(ws, 4, "【1　実質収支に関する調書】", 5)
r = header(ws, r, ["区分", "金額", "", "", "所見"])
for k, sk in [
    ("1 歳入総額", "決算書の歳入合計（収入済額）と一致します。"),
    ("2 歳出総額", "決算書の歳出合計（支出済額）と一致します。"),
    ("3 歳入歳出差引額", "―"),
    ("4 翌年度へ繰越すべき財源　計",
     "継続費・繰越明許費・事故繰越しのいずれも0円です。"),
    ("5 実質収支額",
     "繰越すべき財源が0円のため、歳入歳出差引額と同額です。"),
    ("6 地方自治法第233条の2の規定による基金繰入額",
     "実質収支額を基金へ繰り入れていません。"
     "全額が翌年度への繰越金となります。"),
]:
    fl = {1: MID_B} if k.startswith("5") else None
    r = body(ws, r, [k, K.JISSHITSU[k], "", "", sk], fl, height=30)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

r += 1
r = lead(ws, r, "【2　財産に関する調書（基金）】", 5)
r = header(ws, r, ["区分", "金額", "", "", "所見"])
for k, sk in [
    ("名称",
     "年報の様式4は「介護給付費準備基金」と呼んでいます。"
     "残高は一致するため実質的に同一のものと解されますが、"
     "計画本文に用いる名称の確認を要します。"),
    ("前年度末残高",
     "令和5年度末＝第8期の最終年度末の残高です。"
     "第9期の開始時点の基金残高にあたります。"),
    ("決算年度中増減高", "―"),
    ("うち積立", "予算233,000円に対し232,197円。"
     "歳入の財産運用収入232,197円と同額であり、利子の積立てと解されます。"),
    ("うち支消", "第9期の初年度は基金を取り崩していません。"),
    ("決算年度末現在高",
     "年報の様式4の「介護給付費準備基金保有額」231,689,018円と"
     "一致します。"),
]:
    v = K.KIKIN[k]
    fl = {1: NG_O} if k == "決算年度末現在高" else (
        {1: MID_B} if k in ("名称", "前年度末残高") else None)
    r = body(ws, r, [k, v, "", "", sk], fl, height=34,
             numfmt=None if isinstance(v, str) else "#,##0")
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

r += 1
r = lead(ws, r, "【3　第9期における基金の推移】", 5)
r = header(ws, r, ["年度", "年度当初の残高", "積立", "取崩", "年度末の残高"])
for nendo, hajime, tsumi, tori, owari, memo in [
    ("令和5年度末（第8期末）", "―", "―", "―", 231456821, ""),
    ("令和6年度（第9期1年目）", 231456821, 232197, 0, 231689018, ""),
    ("令和7年度（第9期2年目）", 231689018, "未確定", "未確定", "未確定", ""),
    ("令和8年度（第9期3年目）", "未確定", "未確定", "未確定", "未確定", ""),
]:
    fl = {5: OK_G if isinstance(owari, int) else NG_O}
    r = body(ws, r, [nendo, hajime, tsumi, tori, owari], fl, height=26)

note(ws, r + 1,
     "注1）令和7年度以降が未確定であるのは、"
     "令和7年度の年報の様式4が未入力であり、"
     "令和7年度の決算書も未受領のためです。\n"
     "注2）第10期の保険料算定における取崩額の設定には、"
     "第9期末（令和9年3月末）の基金残高の見込みが必要です。"
     "令和7年度の決算が確定すれば、"
     "令和8年度の見込みと合わせて推計できます。\n"
     "注3）令和6年度の実質収支142,873,961円は基金へ繰り入れず"
     "繰越金としています。"
     "令和7年度も同様であれば、"
     "第9期末の基金残高は2億3千万円台にとどまる見込みです。"
     "ただし推測であり、令和7年度の決算による確認を要します。", 5, 96)

# ============================================================ 03
ws = sheet("03_年報との対照",
           "介護保険事業状況報告（年報）の様式4との対照",
           "決算書と年報は別の出典です。"
           "両者が一致することの確認は第三者検証の観点から意味があります。",
           [44, 44, 22, 22, 14, 40], freeze="A5")

r = lead(ws, 4, "【1　歳入の対照】", 6)
r = header(ws, r, ["決算書の科目", "年報の様式4の科目", "決算書",
                   "年報", "判定", "備考"])
for a, b, c, d in K.TAIOU_SAINYU:
    ok = "一致" if c == d else "不一致"
    r = body(ws, r, [a, b, c, d, ok, ""],
             {5: OK_G if ok == "一致" else NG_O}, height=26,
             align={5: "center"})

r += 1
r = lead(ws, r, "【2　歳出の対照】", 6)
r = header(ws, r, ["決算書の科目", "年報の様式4の科目", "決算書",
                   "年報", "判定", "備考"])
for a, b, c, d in K.TAIOU_SAISHUTSU:
    ok = "一致" if c == d else "不一致"
    r = body(ws, r, [a, b, c, d, ok, ""],
             {5: OK_G if ok == "一致" else NG_O}, height=26,
             align={5: "center"})

r += 1
n_ok = sum(1 for a, b, c, d in K.TAIOU_SAINYU + K.TAIOU_SAISHUTSU if c == d)
n_all = len(K.TAIOU_SAINYU) + len(K.TAIOU_SAISHUTSU)
r = lead(ws, r, "【3　対照の結果】", 6)
r = header(ws, r, ["区分", "件数", "内容", "", "", ""])
for kb, ct, cont in [
    ("一致", "%d件" % n_ok,
     "歳入10項目・歳出12項目のすべてで、"
     "決算書と年報の様式4の値が一致しました。"
     "年報の科目が細分されている箇所（支払基金交付金、一般会計繰入金、"
     "介護サービス等諸費）も、合算すると決算書の項と一致します。"),
    ("不一致", "%d件" % (n_all - n_ok),
     "ありません。"),
    ("金額以外の相違", "3件",
     "保険料の不納欠損額と収入未済額の区分（1,500円）、"
     "基金の名称、一般介護予防事業費の科目の3件です（下記4）。"),
]:
    r = body(ws, r, [kb, ct, cont, "", "", ""],
             {1: MID_B, 2: OK_G if kb == "一致" else None}, height=56,
             align={2: "center"})
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=6)

r += 1
r = lead(ws, r, "【4　金額以外の相違（3件）】", 6)
r = header(ws, r, ["No.", "事項", "相違の内容", "", "受託者の考え", ""])
for i, (ji, soi, kang) in enumerate(K.SOI, start=1):
    r = body(ws, r, [i, ji, soi, "", kang, ""], {2: IN_Y}, height=64,
             align={1: "center"})
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)
    ws.merge_cells(start_row=r - 1, start_column=5, end_row=r - 1, end_column=6)

note(ws, r + 1,
     "注1）決算書は議会に提出された確定値、"
     "年報は国へ報告した値です。"
     "両者が22項目すべてで一致したことは、"
     "計画本文に用いる決算額の信頼性を高めます。\n"
     "注2）相違3件はいずれも金額の総額に影響しません。"
     "保険料の不納欠損額と収入未済額は区分が1,500円異なりますが、"
     "調定額・収納額は一致します。\n"
     "注3）計画本文に用いる基金の名称については、"
     "確認事項としてご確認をお願いします。", 6, 76)

# ============================================================ 04
ws = sheet("04_成果品への反映",
           "成果品への反映箇所と、なお必要な資料",
           "本受領により更新する箇所と、"
           "第10期の保険料算定のためになお必要な資料を整理します。",
           [4, 30, 30, 44, 28, 12], freeze="A5")

r = lead(ws, 4, "【1　更新する箇所】", 6)
r = header(ws, r, ["No.", "成果品", "箇所", "更新の内容", "根拠", "時期"])
for no, seika, kasho, cont, konkyo, jiki in [
    (1, "第10期介護保険事業計画_協議用素案",
     "第6章第6節（介護給付費準備基金）",
     "［要確認］としていた基金の残高を、"
     "令和5年度末231,456,821円・令和6年度末231,689,018円に確定します。"
     "第9期の初年度は取崩しがないことを記載します。",
     "本受領点検 02シート", "R8.9"),
    (2, "第10期介護保険事業計画_協議用素案",
     "第6章第3節（介護保険財政）",
     "款項別の決算額により、第9期の財政の状況を記載します。"
     "不用額と執行率を含めます。",
     "本受領点検 01シート", "R8.9"),
    (3, "第10期計画_年報月報の受領点検",
     "02シート（保険料と決算）",
     "決算書との対照22項目がすべて一致した旨を追記します。"
     "基金の前年度末残高を加えます。",
     "本受領点検 03シート", "R8.9"),
    (4, "第10期計画_将来推計_第3段階",
     "基金の取崩額の設定",
     "第9期末の基金残高の見込みを、"
     "令和6年度末231,689,018円を起点として推計します。"
     "令和7年度の決算が確定するまでは暫定とします。",
     "本受領点検 02シート3", "R8.10"),
    (5, "第10期計画_必要事項の一覧",
     "資料No.12（基金）",
     "令和6年度分は受領済みとし、"
     "令和7年度分の決算書又は年報の様式4に絞ります。",
     "本受領点検 04シート2", "R8.8"),
    (6, "第10期計画_資料提供依頼_第9期の施策事業実績",
     "05シート（決算の状況）・06シート（提供が難しい場合）",
     "令和6年度分は受領済みとします。"
     "「決算額は代替できない」としていた記載を、"
     "令和6年度は受領済み・令和7年度は未受領に改めます。",
     "本受領点検 03シート", "R8.9"),
]:
    r = body(ws, r, [no, seika, kasho, cont, konkyo, jiki], {4: OK_G},
             height=68, align={1: "center", 6: "center"}, numfmt=None)

r += 1
r = lead(ws, r, "【2　なお必要な資料】", 6)
r = header(ws, r, ["No.", "資料", "なぜ必要か", "代替の可否", "優先", ""])
for no, nm, naze, alt, pr in [
    (1, "令和7年度の決算書（介護保険特別会計）"
     "又は年報の様式4",
     "第9期2年目の決算額と、令和7年度末の基金残高。"
     "第10期の保険料算定における取崩額の設定に必要です。"
     "決算の調製は例年12月に議会へ提出されているため、"
     "令和7年度分は令和8年12月頃と見込まれます。",
     "代替できません。"
     "令和6年度の値からの推計は暫定にとどまります。", "最高"),
    (2, "令和8年度の予算書（介護保険特別会計）",
     "第9期最終年度の見込み。"
     "決算は本業務の期間中に確定しないため、"
     "予算額により見込みを立てます。",
     "代替できません。ご提供をお願いします。", "高"),
    (3, "介護給付費準備基金条例",
     "基金の正式名称と、積立て・処分の要件の確認。"
     "決算書は「介護保険事業財政調整基金」、"
     "年報は「介護給付費準備基金」としています。",
     "代替できません。条例をご提供いただくか、"
     "計画本文に用いる名称をご指示ください。", "中"),
    (4, "包括的支援事業・任意事業の事業別の内訳",
     "決算書の「包括的支援事業・任意事業費」75,886,638円は"
     "1つの項であり、6事業＋任意事業の内訳が分かりません。",
     "代替できません。"
     "資料提供依頼No.13の03シートのご記入が必要です。", "最高"),
    (5, "一般介護予防事業費の執行の状況",
     "決算書に「一般介護予防事業費」の項がありません。"
     "介護予防普及啓発事業を3町とも実施しているため、"
     "どの科目で執行されているかの確認が必要です。",
     "代替できません。ご確認をお願いします。", "高"),
    (6, "19施策ごとの事業量・対象実人数・決算額",
     "決算書は款項レベルであり、事業別の内訳はありません。",
     "代替できません。"
     "資料提供依頼No.13の02シートのご記入が必要です。", "最高"),
]:
    fl = {5: NG_O if pr == "最高" else (IN_Y if pr == "高" else None)}
    r = body(ws, r, [no, nm, naze, alt, pr, ""], fl, height=68,
             align={1: "center", 5: "center"}, numfmt=None)
    ws.merge_cells(start_row=r - 1, start_column=5, end_row=r - 1, end_column=6)

note(ws, r + 1,
     "注1）本受領により、資料提供依頼No.13の05シート（決算の状況）の"
     "令和6年度分は完全に埋まりました。"
     "款だけでなく項のレベルまで確定しています。\n"
     "注2）令和7年度の決算書は、例年の調製の時期からすると"
     "令和8年12月頃と見込まれます。"
     "本業務の第10期の保険料算定は、"
     "令和6年度の実績と令和7・8年度の見込みにより行い、"
     "令和7年度の決算が確定した時点で更新する方針とすることを"
     "ご提案します。\n"
     "注3）02シート（19施策の実績）と"
     "03シートの包括的支援事業・任意事業は手つかずのままです。"
     "決算書は款項レベルであり、これらの代替にはなりません。",
     6, 96)

OUT = ("/home/user/repository/output/"
       "第10期計画_令和6年度決算書の受領点検.xlsx")
wb.save(OUT)
print("saved:", OUT)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
