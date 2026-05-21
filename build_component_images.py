# -*- coding: utf-8 -*-
"""
基本コラム部品6種の見本画像をPillowで生成し、Excelに貼り付ける。
- ポイント / コラム / 事例紹介 / 解説 / データの見方 / 注意・留意点
画像は元画像のデザイン（淡色帯＋アイコン＋枠＋本文）を踏襲。
"""

import os
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR  = "/home/user/repository/output"
IMG_DIR  = os.path.join(OUT_DIR, "images_basic")
os.makedirs(IMG_DIR, exist_ok=True)

FONT_REG = "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf"
FONT_BOLD = "/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf"

# 画像サイズ
W, H = 520, 320
PAD = 20

# 6部品定義（id, name, accent色, 淡色, アイコン文字, アイコン色, 用途, 使用例）
PARTS = [
    {"id":"BC-01","name":"ポイント","accent":"#F2A93B","light":"#FFF8E8","icon":"要","stroke":"#F2A93B",
     "use":"計画の要点や重要なポイントを簡潔にまとめる欄です。","ex":"基本方針、重点事項、留意点 等"},
    {"id":"BC-02","name":"コラム","accent":"#3B82F6","light":"#E8F0FE","icon":"文","stroke":"#3B82F6",
     "use":"本文の補足説明や背景情報、豆知識などを掲載する欄です。","ex":"制度説明、背景解説、考え方 等"},
    {"id":"BC-03","name":"事例紹介","accent":"#2CA02C","light":"#E9F6EA","icon":"例","stroke":"#2CA02C",
     "use":"地域や他自治体の取り組み事例を紹介する欄です。","ex":"活動事例、先進事例、参考事例 等"},
    {"id":"BC-04","name":"解説","accent":"#7F4FBF","light":"#EFE9F8","icon":"解","stroke":"#7F4FBF",
     "use":"制度や用語の解説、専門的内容をわかりやすく説明する欄です。","ex":"制度解説、用語解説、法令解説 等"},
    {"id":"BC-05","name":"データの見方","accent":"#1F77B4","light":"#E5EEF7","icon":"図","stroke":"#1F77B4",
     "use":"図表・グラフの読み方や注意点を説明する補足です。","ex":"図表の補足、数値の見方 等"},
    {"id":"BC-06","name":"注意・留意点","accent":"#C0392B","light":"#FDECEA","icon":"注","stroke":"#C0392B",
     "use":"計画推進にあたり注意すべき事項や留意点を示す欄です。","ex":"留意事項、リスク、注意喚起 等"},
]


def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def rounded_rect(draw, xy, radius, fill=None, outline=None, width=2):
    draw.rounded_rectangle(xy, radius=radius, fill=fill, outline=outline, width=width)


def draw_icon_circle(draw, cx, cy, r, fill, icon_text, font, icon_color="white"):
    draw.ellipse([cx-r, cy-r, cx+r, cy+r], fill=fill, outline=None)
    bbox = draw.textbbox((0,0), icon_text, font=font)
    tw = bbox[2]-bbox[0]; th = bbox[3]-bbox[1]
    draw.text((cx-tw/2 - bbox[0], cy-th/2 - bbox[1]), icon_text, fill=icon_color, font=font)


def render_part(p):
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)

    f_title = ImageFont.truetype(FONT_BOLD, 28)
    f_body  = ImageFont.truetype(FONT_REG, 18)
    f_small = ImageFont.truetype(FONT_REG, 15)
    f_icon  = ImageFont.truetype(FONT_BOLD, 30)
    f_label = ImageFont.truetype(FONT_BOLD, 14)

    # 外枠（角丸＋アクセント色の細枠）
    rounded_rect(d, [4, 4, W-4, H-4], radius=14, fill=hex_to_rgb(p["light"]),
                 outline=hex_to_rgb(p["accent"]), width=2)

    # 上部ヘッダー帯
    rounded_rect(d, [4, 4, W-4, 78], radius=14, fill=hex_to_rgb(p["accent"]), outline=None, width=0)
    # 帯下半分の角を直角に見せるための補正矩形
    d.rectangle([4, 50, W-4, 78], fill=hex_to_rgb(p["accent"]))

    # アイコン円（白円にアクセント文字）
    draw_icon_circle(d, 50, 41, 26, "white", p["icon"], f_icon, icon_color=hex_to_rgb(p["accent"]))

    # タイトル（白）
    d.text((92, 22), p["name"], fill="white", font=f_title)

    # IDラベル（右上）
    d.text((W-100, 28), p["id"], fill="white", font=f_label)

    # 本文ボックス（白）
    body_y = 100
    rounded_rect(d, [PAD, body_y, W-PAD, H-PAD-50], radius=10,
                 fill="white", outline=hex_to_rgb(p["accent"]), width=1)

    # 本文テキスト（自動改行）
    def wrap(text, font, max_w):
        lines = []
        cur = ""
        for ch in text:
            test = cur + ch
            w = d.textlength(test, font=font)
            if w > max_w and cur:
                lines.append(cur)
                cur = ch
            else:
                cur = test
        if cur: lines.append(cur)
        return lines

    lines = wrap(p["use"], f_body, W - PAD*2 - 24)
    ty = body_y + 18
    for ln in lines:
        d.text((PAD+14, ty), ln, fill=(40,40,40), font=f_body)
        ty += 28

    # 使用例ラベル
    eg_y = H - PAD - 38
    d.text((PAD, eg_y), "使用例：", fill=hex_to_rgb(p["accent"]), font=f_label)
    d.text((PAD+62, eg_y), p["ex"], fill=(80,80,80), font=f_small)

    path = os.path.join(IMG_DIR, f'{p["id"]}_{p["name"]}.png')
    img.save(path, "PNG")
    return path


# ============================================================
# 画像生成
# ============================================================
print("【1】基本コラム部品の見本画像を生成")
image_paths = []
for p in PARTS:
    path = render_part(p)
    image_paths.append((p, path))
    print(f"  ✓ {p['id']} {p['name']}: {path}")


# ============================================================
# Excelへ貼り付け（新規シート『部品画像一覧』を追加）
# ============================================================
print("\n【2】共通_基本コラム部品.xlsx に画像シートを追加")

src = os.path.join(OUT_DIR, "01_共通_基本コラム部品.xlsx")
wb = load_workbook(src)

# 既存に同名シートがあれば削除
if "部品画像一覧" in wb.sheetnames:
    del wb["部品画像一覧"]

ws = wb.create_sheet("部品画像一覧", index=1)  # 表紙の次に配置

# 見出し
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "基本コラム部品（共通）　画像見本一覧"
c.font = Font(name="游ゴシック", size=14, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[1].height = 32

headers = ["部品ID","部品名","画像見本","主な用途","使用例"]
widths  = [10, 16, 78, 36, 28]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
ws.row_dimensions[3].height = 28

# データ行 + 画像
ROW_HEIGHT = 180  # 各行の高さ（pt）
IMG_SCALE = 0.45  # 画像縮尺
for idx, (p, path) in enumerate(image_paths, start=4):
    ws.row_dimensions[idx].height = ROW_HEIGHT
    ws.cell(row=idx, column=1, value=p["id"])
    ws.cell(row=idx, column=2, value=p["name"])
    ws.cell(row=idx, column=3, value="")  # 画像セル
    ws.cell(row=idx, column=4, value=p["use"])
    ws.cell(row=idx, column=5, value=p["ex"])
    for col in range(1, 6):
        cell = ws.cell(row=idx, column=col)
        cell.font = Font(name="游ゴシック", size=10)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
        cell.border = BORDER
        if idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F7FAFC")
    ws.cell(row=idx, column=2).font = Font(name="游ゴシック", size=11, bold=True)
    ws.cell(row=idx, column=2).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.cell(row=idx, column=1).alignment = Alignment(vertical="center", horizontal="center")

    # 画像挿入
    xlimg = XLImage(path)
    xlimg.width  = int(W * IMG_SCALE * 1.5)   # 表示幅
    xlimg.height = int(H * IMG_SCALE * 1.5)   # 表示高さ
    xlimg.anchor = f"C{idx}"
    ws.add_image(xlimg)

ws.sheet_view.showGridLines = False

out = src  # 上書き保存
wb.save(out)
print(f"  ✓ 保存: {out}")

# マスターブックにも同じシートを追加
print("\n【3】全計画マスター管理表にも画像シートを追加")
master_src = os.path.join(OUT_DIR, "00_全計画マスター管理表.xlsx")
wbm = load_workbook(master_src)
if "基本部品画像一覧" in wbm.sheetnames:
    del wbm["基本部品画像一覧"]
wsm = wbm.create_sheet("基本部品画像一覧", index=1)

wsm.merge_cells("A1:E1")
c = wsm["A1"]
c.value = "基本コラム部品（共通）　画像見本一覧"
c.font = Font(name="游ゴシック", size=14, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
wsm.row_dimensions[1].height = 32

for i, h in enumerate(headers, 1):
    cell = wsm.cell(row=3, column=i, value=h)
    cell.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER
    wsm.column_dimensions[get_column_letter(i)].width = widths[i-1]
wsm.row_dimensions[3].height = 28

for idx, (p, path) in enumerate(image_paths, start=4):
    wsm.row_dimensions[idx].height = ROW_HEIGHT
    wsm.cell(row=idx, column=1, value=p["id"])
    wsm.cell(row=idx, column=2, value=p["name"])
    wsm.cell(row=idx, column=3, value="")
    wsm.cell(row=idx, column=4, value=p["use"])
    wsm.cell(row=idx, column=5, value=p["ex"])
    for col in range(1, 6):
        cell = wsm.cell(row=idx, column=col)
        cell.font = Font(name="游ゴシック", size=10)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
        cell.border = BORDER
        if idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F7FAFC")
    wsm.cell(row=idx, column=2).font = Font(name="游ゴシック", size=11, bold=True)
    wsm.cell(row=idx, column=2).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    wsm.cell(row=idx, column=1).alignment = Alignment(vertical="center", horizontal="center")

    xlimg = XLImage(path)
    xlimg.width  = int(W * IMG_SCALE * 1.5)
    xlimg.height = int(H * IMG_SCALE * 1.5)
    xlimg.anchor = f"C{idx}"
    wsm.add_image(xlimg)

wsm.sheet_view.showGridLines = False
wbm.save(master_src)
print(f"  ✓ 保存: {master_src}")

print("\n完了")
