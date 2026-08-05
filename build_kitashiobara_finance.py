# -*- coding: utf-8 -*-
"""
北塩原村 第8期障がい福祉計画・第4期障がい児福祉計画
財源構成案 管理ブック ジェネレータ

出力: output/北塩原村_財源構成案.xlsx

令和8年7月7日打合せの決定事項
「国・県・村の財政負担（自立支援給付等の内訳）の推移を計画に盛り込み可視化する」
に対応する。従来の計画がサービス利用件数のみの追跡にとどまっていた点を改める。

設計方針
- 給付実績（受領データ整理版）は確定値としてハードコード。
- 負担割合を掛けた財源構成はすべてExcel数式。負担割合を変えれば全年度が再計算される。
- 地域生活支援事業・村単独事業は村資料待ちのため入力欄のみ用意する。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from kitashiobara_common import (
    BENEFIT_YEARS, COLORS, FONT, FUNDING_RULES, JIDO_KYUFU, JIDO_TOTAL,
    KAIGO_KYUFU, KAIGO_TOTAL, OUT_DIR,
    add_sheet, ensure_out_dir, set_col_widths, style_data_cell,
    style_header_row, style_note, style_status, style_title, write_row,
)

OUT_FILE = f"{OUT_DIR}/北塩原村_財源構成案.xlsx"

YEN = "#,##0"
PCT1 = "0.0%"
CNT = "#,##0"

# 負担割合を置くセル（01_負担区分マスタ）。全シートがここを参照する。
RATE_SHEET = "01_負担区分マスタ"
RATE_KUNI = f"'{RATE_SHEET}'!$C$6"   # 国 1/2
RATE_KEN = f"'{RATE_SHEET}'!$D$6"    # 県 1/4
RATE_MURA = f"'{RATE_SHEET}'!$E$6"   # 村 1/4

DATA_START = 6

# 00_概要が参照するセル位置。各シート側で assert して整合を保証する。
FUNDING_R7_ROW = DATA_START + len(BENEFIT_YEARS) - 1   # 04_財源構成試算 令和7年度
CHIIKI_ITEMS = 10                                      # 06_地域生活支援事業 の事業数
CHIIKI_TOTAL_ROW = DATA_START + CHIIKI_ITEMS           # 06 合計行
TANDOKU_ITEMS = 5                                      # 07_村単独事業 の事業数
TANDOKU_TOTAL_ROW = DATA_START + TANDOKU_ITEMS         # 07 合計行
REV_R8_ROW = 7                          # 令和8年度報酬改定の改定率入力行
REV_R9_ROW = 8                          # 令和9年度報酬改定の改定率入力行
REV_TOTAL_ROW = 9                       # 適用倍率（合計）行
PROJ_SVC_FIRST = 13                     # サービス別積上げ表の先頭行
PROJ_SVC_COUNT = 16                     # 積上げ表のサービス数
PROJ_SVC_TOTAL_ROW = PROJ_SVC_FIRST + PROJ_SVC_COUNT   # 積上げ合計行


# ============================================================
# 00_概要
# ============================================================
def sheet_overview(wb):
    ws = wb.create_sheet("00_概要")
    set_col_widths(ws, [34, 20, 20, 14, 46])
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:E1")
    style_title(ws["A1"], "北塩原村 財源構成案 管理ブック（国・県・村の負担内訳）",
                fill=COLORS["障がい"])
    ws.merge_cells("A2:E2")
    style_note(ws["A2"],
               "第8期北塩原村障がい福祉計画・第4期北塩原村障がい児福祉計画／"
               "令和8年7月7日打合せ「国・県・村の財政負担（自立支援給付等の内訳）の推移を計画に盛り込み可視化する」に対応。")
    ws.row_dimensions[2].height = 28

    r = 4
    style_header_row(ws, r, ["確認項目", "現在値（自動計算）", "参照先", "状態", "備考"])
    r += 1
    fr = FUNDING_R7_ROW
    rows = [
        ("令和7年度 自立支援給付費 総額", f"='04_財源構成試算'!D{fr}", "04_財源構成試算", "確定",
         "介護給付費＋障害児給付費。受領データ（令和2〜7年度）による"),
        ("うち 国庫負担（1/2）", f"='04_財源構成試算'!E{fr}", "04_財源構成試算", "暫定",
         "法定負担割合による試算。村の決算額との突合が必要"),
        ("うち 県負担（1/4）", f"='04_財源構成試算'!F{fr}", "04_財源構成試算", "暫定", "同上"),
        ("うち 村負担（1/4）", f"='04_財源構成試算'!G{fr}", "04_財源構成試算", "暫定", "同上"),
        ("令和2年度からの増加率（総額）", f"='04_財源構成試算'!I{fr}", "04_財源構成試算", "確定",
         "令和2年度 36,748千円 → 令和7年度 55,513千円"),
        ("地域生活支援事業費（令和7年度）", f"='06_地域生活支援事業'!C{CHIIKI_TOTAL_ROW}",
         "06_地域生活支援事業", "村資料待ち",
         "補助基準額の範囲内補助のため、村負担・超過負担の把握には村決算が必要"),
        ("村単独事業費（令和7年度）", f"='07_村単独事業'!D{TANDOKU_TOTAL_ROW}",
         "07_村単独事業", "村資料待ち",
         "重度心身障害者医療費助成、人工透析患者通院交通費助成等"),
        ("令和11年度 給付費推計（積上げ）", f"='08_R9R11推計'!H{PROJ_SVC_TOTAL_ROW}", "08_R9R11推計", "村資料待ち",
         "サービス見込量ブックの個別積上げ×単価。見込量と単価の入力後に確定する"),
    ]
    for i, row in enumerate(rows):
        write_row(ws, r, list(row), alt=(i % 2 == 1),
                  aligns=["left", "right", "center", "center", "left"],
                  numfmts=[None, YEN, None, None, None])
        style_status(ws.cell(row=r, column=4))
        r += 1
    ws.cell(row=9, column=2).number_format = PCT1  # 増加率

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "本ブックの使い方と留意点", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "1. 02・03シートの給付実績は受領データ（令和2〜7年度）の確定値です。原則として変更しません。",
        "2. 財源構成（04・05）は、01_負担区分マスタの負担割合を掛けた法定ベースの試算です。国庫負担基準額を超えた分は村負担となるため、実際の決算額とは一致しません。",
        "3. 村の歳入歳出決算（障害者自立支援給付費負担金・障害児施設給付費負担金等）を受領したら、09_村確認事項の該当行を更新し、試算値と決算額の差を確認してください。",
        "4. 地域生活支援事業（06）は国1/2以内・県1/4以内の統合補助金です。補助基準額を超える部分は全額村負担となるため、事業費・交付額・村負担を分けて把握する必要があります。",
        "5. 令和9〜11年度の給付費は、伸び率ではなく「北塩原村_サービス見込量.xlsx」の個別積上げ×単価で算定します。08シートに見込量と単価を入力してください。",
        "6. 計画書への掲載は、金額の羅列ではなく「総額の推移」「財源別構成比の推移」「サービス別の構成」の3点に絞ると読みやすくなります。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 26
        r += 1
    return ws


# ============================================================
# 01_負担区分マスタ
# ============================================================
def sheet_rules(wb):
    ws = add_sheet(
        wb, RATE_SHEET, "財源の負担区分",
        "障害者総合支援法・児童福祉法に基づく法定負担割合。04・05・08シートはこの表の負担割合（C6:E6）を参照して計算します。"
        "実際の交付額・村負担は村の決算で確認してください。",
        [40, 12, 12, 12, 12, 14, 52])
    style_header_row(ws, 5, ["費目", "国", "県", "村", "計", "性格", "根拠・留意点"])

    r = DATA_START
    for i, (name, kuni, ken, mura, seikaku, note) in enumerate(FUNDING_RULES):
        if ken is None:
            vals = [name, "－", "事業ごと", "残額", "", seikaku, note]
            numfmts = [None] * 7
        else:
            vals = [name, kuni, ken, mura, f"=SUM(B{r}:D{r})", seikaku, note]
            numfmts = [None, PCT1, PCT1, PCT1, PCT1, None, None]
        write_row(ws, r, vals, alt=(i % 2 == 1),
                  aligns=["left", "center", "center", "center", "center", "center", "left"],
                  numfmts=numfmts)
        ws.row_dimensions[r].height = 32
        r += 1

    # 参照用の負担割合セルを強調（先頭行＝自立支援給付）
    for col in range(2, 5):
        c = ws.cell(row=DATA_START, column=col)
        c.fill = PatternFill("solid", fgColor=COLORS["input"])
        c.font = Font(name=FONT, size=10, bold=True)

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "※ 黄色セル（B6:D6）が04・05・08シートの計算に使われる負担割合です。制度改正等で割合が変わった場合はここを直すと全シートに反映されます。")
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1), "計画書に記載する際の整理", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "・自立支援給付（介護給付費・訓練等給付費・計画相談支援給付費）と障害児通所給付費は義務的経費であり、利用があれば国・県の負担が伴います。村の裁量で増減させるものではありません。",
        "・これに対し地域生活支援事業は裁量的経費（統合補助金）で、国1/2以内・県1/4以内の「以内補助」です。補助基準額を超えた分は村の一般財源で賄うことになります。",
        "・したがって「実施の有無」だけで施策を評価せず、事業費・国県補助額・村負担額・超過負担の有無を分けて評価する必要があります（計画書 現行計画評価の章）。",
        "・村単独事業（重度心身障害者医療費助成、在宅重度障害者対策事業、人工透析患者通院交通費助成等）は、生活継続を支える基盤であるため、対象者数と村負担額を継続して把握します。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# ============================================================
# 02・03 給付実績
# ============================================================
def _benefit_sheet(wb, title, heading, note, data, totals, label):
    ncol = 2 + len(BENEFIT_YEARS) + 2
    ws = add_sheet(wb, title, heading, note,
                   [32, 10] + [14] * len(BENEFIT_YEARS) + [12, 14])
    hdr = ["サービス", "区分"] + [w for w, _ in BENEFIT_YEARS] + ["令和2年度比", "備考"]
    style_header_row(ws, 5, hdr)

    aligns = ["left", "center"] + ["right"] * len(BENEFIT_YEARS) + ["right", "left"]
    numfmts = [None, None] + [YEN] * len(BENEFIT_YEARS) + [PCT1, None]

    r = DATA_START
    amount_rows = []
    for i, (svc, by_year) in enumerate(data.items()):
        vals = [by_year.get(y, (None, None))[1] for _, y in BENEFIT_YEARS]
        first, last = vals[0], vals[-1]
        ratio = (f"=IF(OR(C{r}=\"\",C{r}=0,{get_column_letter(2 + len(BENEFIT_YEARS))}{r}=\"\"),\"\","
                 f"{get_column_letter(2 + len(BENEFIT_YEARS))}{r}/C{r}-1)")
        note_txt = ""
        if svc == "就労継続支援（B型）":
            note_txt = "金額最大。日中活動・就労支援の見込量根拠"
        elif svc == "短期入所":
            note_txt = "令和6年度から実績あり。拠点の緊急受入と接続"
        elif svc == "放課後等デイサービス":
            note_txt = "令和7年度に件数・金額とも大幅増"
        elif svc == "保育所等訪問支援":
            note_txt = "令和7年度に初めて実績が発生"
        write_row(ws, r, [svc, "金額"] + vals + [ratio, note_txt],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts)
        amount_rows.append(r)
        r += 1

    # 合計行（受領原票の合計と一致することを検証済み）
    sums = [f"=SUM({get_column_letter(3 + j)}{amount_rows[0]}:{get_column_letter(3 + j)}{amount_rows[-1]})"
            for j in range(len(BENEFIT_YEARS))]
    lastcol = get_column_letter(2 + len(BENEFIT_YEARS))
    write_row(ws, r, [f"{label} 合計", "金額"] + sums +
              [f"={lastcol}{r}/C{r}-1", "内訳の合計＝原票の年度合計と一致"],
              aligns=aligns, numfmts=numfmts, fills=[COLORS["band"]] * ncol)
    for col in range(1, ncol + 1):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=10, bold=True)
    total_row = r
    r += 2

    # 件数表
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    style_title(ws.cell(row=r, column=1), "参考：件数", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, hdr)
    r += 1
    cnt_first = r
    for i, (svc, by_year) in enumerate(data.items()):
        vals = [by_year.get(y, (None, None))[0] for _, y in BENEFIT_YEARS]
        write_row(ws, r, [svc, "件数"] + vals + ["", ""],
                  alt=(i % 2 == 1), aligns=aligns,
                  numfmts=[None, None] + [CNT] * len(BENEFIT_YEARS) + [None, None])
        r += 1
    sums_c = [f"=SUM({get_column_letter(3 + j)}{cnt_first}:{get_column_letter(3 + j)}{r - 1})"
              for j in range(len(BENEFIT_YEARS))]
    write_row(ws, r, [f"{label} 合計", "件数"] + sums_c + ["", ""],
              aligns=aligns, numfmts=[None, None] + [CNT] * len(BENEFIT_YEARS) + [None, None],
              fills=[COLORS["band"]] * ncol)
    for col in range(1, ncol + 1):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=10, bold=True)
    return ws, total_row


def sheet_kaigo(wb):
    return _benefit_sheet(
        wb, "02_給付実績_介護給付費", "介護給付費・訓練等給付費 実績（サービス別）",
        "受領データ「【北塩原村】障がいサービス給付実績.xlsx」の整理版による確定値（令和2〜7年度）。"
        "空欄は当該年度に実績がないことを示します。原則として変更しません。",
        KAIGO_KYUFU, KAIGO_TOTAL, "介護給付費")


def sheet_jido(wb):
    return _benefit_sheet(
        wb, "03_給付実績_障害児給付費", "障害児通所給付費・障害児相談支援給付費 実績（サービス別）",
        "受領データの整理版による確定値（令和2〜7年度）。障害児給付費は令和4年度をピークに減少し、"
        "令和7年度は放課後等デイサービスが中心となっています。",
        JIDO_KYUFU, JIDO_TOTAL, "障害児給付費")


# ============================================================
# 04_財源構成試算
# ============================================================
def sheet_funding(wb, kaigo_total_row, jido_total_row):
    ws = add_sheet(
        wb, "04_財源構成試算", "財源構成試算（国・県・村の負担内訳の推移）",
        "自立支援給付費・障害児給付費に法定負担割合を掛けた試算です。国庫負担基準額を超えた分は村負担となるため、"
        "実際の決算額とは一致しません。村の決算受領後に突合してください。単位：円。",
        [12, 16, 16, 16, 16, 16, 16, 14, 14, 30])
    style_header_row(ws, 5, ["年度", "介護給付費", "障害児給付費", "自立支援給付計",
                             "国庫負担", "県負担", "村負担", "前年度比", "令和2年度比", "備考"])

    aligns = ["center"] + ["right"] * 8 + ["left"]
    numfmts = [None] + [YEN] * 6 + [PCT1, PCT1, None]

    notes = {
        2: "基準年度",
        3: "共同生活援助・就労継続支援B型の増",
        6: "件数が605件へ増加。特定障害者特別給付費を計上",
        7: "直近年度。伸びは鈍化",
    }

    r = DATA_START
    first_row = r
    for i, (wareki, y) in enumerate(BENEFIT_YEARS):
        col = get_column_letter(3 + i)  # 02・03シートの当該年度列
        vals = [
            wareki,
            f"='02_給付実績_介護給付費'!{col}{kaigo_total_row}",
            f"='03_給付実績_障害児給付費'!{col}{jido_total_row}",
            f"=B{r}+C{r}",
            f"=ROUND(D{r}*{RATE_KUNI},0)",
            f"=ROUND(D{r}*{RATE_KEN},0)",
            f"=D{r}-E{r}-F{r}",
            ("" if i == 0 else f"=D{r}/D{r - 1}-1"),
            ("" if i == 0 else f"=D{r}/D${first_row}-1"),
            notes.get(y, ""),
        ]
        write_row(ws, r, vals, alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts,
                  fills=[None, COLORS["calc"], COLORS["calc"], COLORS["calc"],
                         COLORS["calc"], COLORS["calc"], COLORS["calc"], None, None, None])
        if y == 7:
            assert r == FUNDING_R7_ROW, f"04_財源構成試算 令和7年度の行位置が想定と異なります: {r}"
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    style_title(ws.cell(row=r, column=1), "財源別構成比（計画書掲載用）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年度", "自立支援給付計", "国庫負担", "県負担", "村負担",
                             "国 構成比", "県 構成比", "村 構成比", "", ""])
    r += 1
    for i, (wareki, _) in enumerate(BENEFIT_YEARS):
        src = first_row + i
        write_row(ws, r, [wareki, f"=D{src}", f"=E{src}", f"=F{src}", f"=G{src}",
                          f"=E{src}/D{src}", f"=F{src}/D{src}", f"=G{src}/D{src}", "", ""],
                  alt=(i % 2 == 1),
                  aligns=["center"] + ["right"] * 7 + ["left", "left"],
                  numfmts=[None, YEN, YEN, YEN, YEN, PCT1, PCT1, PCT1, None, None])
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    style_note(ws.cell(row=r, column=1),
               "読み取り：介護給付費は令和2年度36,002千円から令和7年度53,877千円へ約49.6％増加しています。"
               "一方、障害児給付費は令和4年度4,030千円をピークに令和7年度1,636千円まで減少しています。"
               "村負担（1/4相当）は総額の増加に比例して増えるため、計画書では総額の推移とあわせて村負担の推移を示してください。")
    ws.row_dimensions[r].height = 34
    return ws


# ============================================================
# 05_サービス別財源構成
# ============================================================
def sheet_by_service(wb):
    ws = add_sheet(
        wb, "05_サービス別財源構成", "令和7年度 サービス別の財源構成",
        "直近年度（令和7年度）のサービス別給付費に法定負担割合を掛けたものです。"
        "計画書では上位サービスに絞って掲載し、見込量の根拠と結び付けて説明します。単位：円。",
        [32, 16, 12, 16, 16, 16, 40])
    style_header_row(ws, 5, ["サービス", "令和7年度 給付費", "構成比",
                             "国庫負担", "県負担", "村負担", "計画上の位置付け"])

    positions = {
        "就労継続支援（B型）": "村内に事業所なし。喜多方市・会津若松市等の圏域事業所を利用。令和2年度比2.08倍",
        "生活介護": "重度者の日中活動。特別支援学校卒業予定者の進路と接続",
        "共同生活援助": "居住系の中心。地域移行・親亡き後支援と接続",
        "施設入所支援": "4人で横ばい。成果目標「施設入所者の地域生活移行」の対象",
        "計画相談支援": "圏域の相談支援事業所が実施。基幹相談支援センター広域設置と接続",
        "居宅介護": "訪問可能地区・冬季提供体制の制約を確認",
        "自立訓練（生活訓練）": "利用の増減が大きい。個別ケース確認が必要",
        "短期入所": "令和6年度から実績。地域生活支援拠点の緊急受入機能と一体で検討",
        "生活介護（基準該当）": "基準該当事業所の利用。少額",
        "共同生活援助（特定障害者特別給付費）": "家賃補助分",
        "施設入所支援（特定障害者特別給付費）": "食費・光熱水費の補足給付分",
    }

    r7 = [(svc, d[7][1]) for svc, d in KAIGO_KYUFU.items() if 7 in d]
    r7 += [(svc, d[7][1]) for svc, d in JIDO_KYUFU.items() if 7 in d]
    r7.sort(key=lambda x: -x[1])
    total = sum(v for _, v in r7)
    assert total == KAIGO_TOTAL[7][1] + JIDO_TOTAL[7][1], "令和7年度の内訳合計が総額と一致しません"

    aligns = ["left", "right", "right", "right", "right", "right", "left"]
    numfmts = [None, YEN, PCT1, YEN, YEN, YEN, None]
    r = DATA_START
    first = r
    for i, (svc, amt) in enumerate(r7):
        write_row(ws, r, [svc, amt, f"=B{r}/$B${first + len(r7)}",
                          f"=ROUND(B{r}*{RATE_KUNI},0)",
                          f"=ROUND(B{r}*{RATE_KEN},0)",
                          f"=B{r}-D{r}-E{r}",
                          positions.get(svc, "障がい児支援。教育・保育・相談支援ニーズと接続")],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts)
        r += 1
    write_row(ws, r, ["合計", f"=SUM(B{first}:B{r - 1})", "=B{0}/B{0}".format(r),
                      f"=SUM(D{first}:D{r - 1})", f"=SUM(E{first}:E{r - 1})",
                      f"=SUM(F{first}:F{r - 1})", "介護給付費53,877,076円＋障害児給付費1,635,703円"],
              aligns=aligns, numfmts=numfmts, fills=[COLORS["band"]] * 7)
    for col in range(1, 8):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=10, bold=True)
    return ws


# ============================================================
# 06_地域生活支援事業
# ============================================================
def sheet_chiiki(wb):
    ws = add_sheet(
        wb, "06_地域生活支援事業", "地域生活支援事業の財源構成（入力欄）",
        "地域生活支援事業は国1/2以内・県1/4以内の統合補助金です。補助基準額を超えた分は全額村負担となるため、"
        "事業費・交付額・村負担を分けて把握する必要があります。黄色セルに村の決算値を入力してください。単位：円。",
        [30, 16, 16, 16, 16, 16, 14, 36])
    style_header_row(ws, 5, ["事業", "令和6年度\n事業費", "令和7年度\n事業費",
                             "令和8年度\n事業費（見込）", "国補助額", "県補助額",
                             "村負担額", "確認事項"])

    jigyo = [
        ("障がい者相談支援事業", "村住民課・地域生活支援センターいなわしろの2か所。基本相談部分を委託"),
        ("成年後見制度利用支援事業", "現行計画見込5人に対し令和5年度実績1人。乖離要因を確認"),
        ("意思疎通支援事業", "実績0人。手話通訳者派遣の需要の有無を確認"),
        ("日常生活用具給付等事業", "令和5年度実績69件。排泄管理支援用具が大半"),
        ("移動支援事業", "令和3〜8年度の全期間で見込量0人・実績0人。制度未周知／代替手段／ニーズ不在の切り分けが必要"),
        ("日中一時支援事業", "令和5年度実績2人（増加傾向）。家族レスパイト・障がい児と接続"),
        ("地域活動支援センター機能強化事業", "令和5年度実績3人が見込量2人を上回っている。次期見込量に実績を反映"),
        ("訪問入浴サービス事業", "実績0人"),
        ("理解促進研修・啓発事業", "実績0件。実施方法と財源を確認"),
        ("自発的活動支援事業", "実績0件。障がい者団体・家族会の活動支援と接続"),
    ]
    assert len(jigyo) == CHIIKI_ITEMS, "06_地域生活支援事業の事業数が定数と一致しません"
    aligns = ["left"] + ["right"] * 6 + ["left"]
    numfmts = [None] + [YEN] * 6 + [None]
    r = DATA_START
    first = r
    for i, (name, note) in enumerate(jigyo):
        write_row(ws, r, [name, None, None, None, None, None,
                          f"=IF(COUNT(C{r},E{r},F{r})=0,\"\",C{r}-E{r}-F{r})", note],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts,
                  fills=[None, COLORS["input"], COLORS["input"], COLORS["input"],
                         COLORS["input"], COLORS["input"], COLORS["calc"], None])
        ws.row_dimensions[r].height = 30
        r += 1
    write_row(ws, r, ["合計"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{r - 1})"
                                 for c in range(2, 8)] +
              ["村負担＝事業費－国補助－県補助（補助基準額超過分を含む）"],
              aligns=aligns, numfmts=numfmts, fills=[COLORS["band"]] * 8)
    for col in range(1, 9):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=10, bold=True)
    assert r == CHIIKI_TOTAL_ROW, f"06_地域生活支援事業 合計行が想定と異なります: {r}"
    total_row = r

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    style_note(ws.cell(row=r, column=1),
               "留意：移動支援事業が令和3年度から令和8年度まで一貫して0人であることは、"
               "地区分散・冬季移動が中心的な地域課題である本村の実態と整合しません。"
               "実績ゼロをニーズ不在と即断せず、制度未周知・家族送迎等での代替・ニーズ自体の不在のいずれに該当するかを、"
               "アンケート結果と個別ケースで確認したうえで次期見込量を確定してください（令和8年7月7日打合せの決定事項）。")
    ws.row_dimensions[r].height = 44
    return ws, total_row


# ============================================================
# 07_村単独事業
# ============================================================
def sheet_tandoku(wb):
    ws = add_sheet(
        wb, "07_村単独事業", "村単独事業の財源構成（入力欄）",
        "村が単独で実施する助成事業です。一部に県補助金があります。障がいのある方の生活継続を支える基盤であるため、"
        "対象者数と村負担額を継続して把握します。黄色セルに村の決算値を入力してください。単位：円。",
        [30, 14, 16, 16, 16, 16, 40])
    style_header_row(ws, 5, ["事業", "対象者数", "令和6年度\n事業費", "令和7年度\n事業費",
                             "県補助額", "村負担額", "内容・確認事項"])
    jigyo = [
        ("重度心身障害者医療費助成", "重度障がい者の医療費自己負担分を助成。対象者数と一人当たり額を確認"),
        ("在宅重度障害者対策事業", "在宅の重度障がい者への助成。対象要件と実績を確認"),
        ("人工透析患者通院交通費助成", "通院交通費の助成。地区分散・冬季移動の課題と直結する事業"),
        ("障がい者団体等への補助", "障がい者団体・家族会の活動支援。自発的活動支援事業との関係を確認"),
        ("その他（　　　　　）", "村単独で実施している事業があれば追記"),
    ]
    assert len(jigyo) == TANDOKU_ITEMS, "07_村単独事業の事業数が定数と一致しません"
    aligns = ["left", "right", "right", "right", "right", "right", "left"]
    numfmts = [None, CNT, YEN, YEN, YEN, YEN, None]
    r = DATA_START
    first = r
    for i, (name, note) in enumerate(jigyo):
        write_row(ws, r, [name, None, None, None, None,
                          f"=IF(COUNT(D{r},E{r})=0,\"\",D{r}-E{r})", note],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts,
                  fills=[None, COLORS["input"], COLORS["input"], COLORS["input"],
                         COLORS["input"], COLORS["calc"], None])
        ws.row_dimensions[r].height = 28
        r += 1
    write_row(ws, r, ["合計"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{r - 1})"
                                 for c in range(2, 7)] + ["村負担＝事業費－県補助"],
              aligns=aligns, numfmts=numfmts, fills=[COLORS["band"]] * 7)
    for col in range(1, 8):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=10, bold=True)
    assert r == TANDOKU_TOTAL_ROW, f"07_村単独事業 合計行が想定と異なります: {r}"
    return ws, r


# ============================================================
# 08_R9R11推計
# ============================================================
# 推計方式の見直しにより、伸び率シナリオを主たる推計から外した。
# 給付費は「サービス見込量ブックの個別積上げ × 単価」で算定する。
# 伸び率は、未特定の潜在需要（第3層）を補正する場合にのみ用いる。


def sheet_projection(wb):
    ws = add_sheet(
        wb, "08_R9R11推計", "次期計画期間（令和9〜11年度）の給付費・財源構成 推計",
        "給付費は「サービス見込量ブック（北塩原村_サービス見込量.xlsx）の個別積上げ × 単価」で算定します。"
        "伸び率による推計は、未特定の潜在需要（第3層）を補正する場合に限って用います。"
        "本村では利用者1人の増減が生活介護で16.7%、居宅介護・施設入所支援で25%を占め、"
        "年率2〜8%の伸び率を大きく上回るためです。単位：円。",
        [26, 12, 14, 16, 16, 16, 16, 16, 34])

    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_title(ws.cell(row=r, column=1),
                "1. 報酬改定の前提（単価の補正率）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "改定率", "施行", "内容", "", "", "", "", ""])
    r += 1
    assert r == REV_R8_ROW, f"報酬改定行が想定と異なります: {r}"
    rev_rows = [
        ("令和8年度報酬改定", 1.000, "令和8年6月1日",
         "①福祉・介護職員等処遇改善加算の拡充（対象を障害福祉従事者へ拡大。月1.0万円＝3.3％の賃上げ、"
         "生産性向上・協働化に取り組む事業者は月0.3万円＝1.0％を上乗せ。計画相談支援・障害児相談支援・"
         "地域相談支援に新設）②就労継続支援B型の基本報酬区分の基準見直し "
         "③新規指定事業所に限る応急的な報酬単価（B型 所定単位数の984/1000、"
         "共同生活援助（介護サービス包括型・日中サービス支援型）972/1000 ほか児童発達支援・放課後等デイサービス）"),
        ("令和9年度報酬改定", 1.000, "令和9年4月（予定）",
         "次期計画期間の初年度に当たる。令和8年度改定の応急的単価が「令和9年度報酬改定までの間」と"
         "されていることから実施が見込まれる。内容確定後に改定率を入力する"),
    ]
    for i, (kubun, rate, shikou, naiyou) in enumerate(rev_rows):
        write_row(ws, r, [kubun, rate, shikou, naiyou, "", "", "", "", ""],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "center", "left", "left", "left", "left", "left", "left"],
                  numfmts=[None, "0.000", None, None, None, None, None, None, None],
                  fills=[None, COLORS["input"], None, None, None, None, None, None, None])
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 58
        r += 1
    write_row(ws, r, ["適用倍率（合計）", f"=B{REV_R8_ROW}*B{REV_R9_ROW}", "―",
                      "下表の令和11年度年間給付費に乗じます。改定率が確定するまでは1.000（改定なし）として計算します。"
                      "離島・中山間地域（特別地域加算の対象地域）の事業所や重度障がい者に係る基本報酬には"
                      "従前単価が適用される配慮措置があるため、圏域事業所の該当状況を確認してください。",
                      "", "", "", "", ""],
              aligns=["left", "right", "center", "left", "left", "left", "left", "left", "left"],
              numfmts=[None, "0.000", None, None, None, None, None, None, None],
              fills=[COLORS["band"], COLORS["calc"]] + [COLORS["band"]] * 7)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    ws.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True)
    ws.row_dimensions[r].height = 44
    assert r == REV_TOTAL_ROW, f"適用倍率行が想定と異なります: {r}"
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_title(ws.cell(row=r, column=1),
                "2. サービス別の積上げ（見込量×単価）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["サービス", "単位", "月額単価",
                             "令和8年度\n見込量", "令和9年度\n見込量",
                             "令和10年度\n見込量", "令和11年度\n見込量",
                             "令和11年度\n年間給付費", "備考"])
    r += 1
    assert r == PROJ_SVC_FIRST, f"積上げ表の先頭行が想定と異なります: {r}"

    svc_rows = [
        ("生活介護", "人日／月", "令和7年度 給付費13,209,810円。特別支援学校卒業者の主な移行先"),
        ("就労継続支援（B型）", "人日／月", "令和7年度 給付費19,707,121円。令和2年度比2.08倍の要因分析が必要"),
        ("共同生活援助", "人／月", "介護保険に相当サービスがなく、65歳以降も継続する"),
        ("施設入所支援", "人／月", "介護保険に相当サービスなし。成果目標（１）と整合させる"),
        ("居宅介護", "時間／月", "介護保険の訪問介護に相当。65歳到達者は個別に判定する"),
        ("計画相談支援", "人／月", "請求は計画作成・モニタリング時のみ。実利用者数と請求件数は異なる"),
        ("短期入所", "人日／月", "令和6年度2件・令和7年度3件の実績。令和9年度からの計上を検討"),
        ("自立訓練（生活訓練）", "人日／月", "令和7年度 給付費382,238円"),
        ("就労移行支援", "人日／月", "一般就労移行の成果目標と連動"),
        ("就労継続支援（A型）", "人日／月", "現利用者への継続提供"),
        ("就労定着支援", "人／月", "一般就労移行者数と連動"),
        ("就労選択支援", "人／月", "令和7年10月開始。圏域事業所の有無を確認"),
        ("児童発達支援", "人日／月", "令和7年度は請求1件。就学に伴う移行の可能性（要確認）"),
        ("放課後等デイサービス", "人日／月", "令和7年度は請求27件。就学に伴う流入の可能性（要確認）"),
        ("保育所等訪問支援", "人／月", "令和7年度に初めて実績"),
        ("障がい児相談支援", "人／月", "令和7年度 給付費330,460円"),
    ]
    assert len(svc_rows) == PROJ_SVC_COUNT, "積上げ表のサービス数が定数と一致しません"
    first = r
    for i, (svc, unit, note) in enumerate(svc_rows):
        write_row(ws, r, [svc, unit, None, None, None, None, None,
                          f"=IFERROR(ROUND(C{r}*G{r}*12*$B${REV_TOTAL_ROW},0),\"\")", note],
                  alt=(i % 2 == 1),
                  aligns=["left", "center", "right", "right", "right", "right", "right",
                          "right", "left"],
                  numfmts=[None, None, YEN, "#,##0.0", "#,##0.0", "#,##0.0", "#,##0.0",
                           YEN, None],
                  fills=[None, None] + [COLORS["input"]] * 5 + [COLORS["calc"], None])
        ws.row_dimensions[r].height = 26
        r += 1
    write_row(ws, r, ["合計", "―", "―", "―", "―", "―", "―",
                      f"=SUM(H{first}:H{r-1})", "積上げによる令和11年度の年間給付費"],
              aligns=["left", "center", "center", "center", "center", "center", "center",
                      "right", "left"],
              numfmts=[None] * 7 + [YEN, None],
              fills=[COLORS["band"]] * 9)
    for c in range(1, 10):
        ws.cell(row=r, column=c).font = Font(name=FONT, size=10, bold=True)
    assert r == PROJ_SVC_TOTAL_ROW, f"積上げ合計行が想定と異なります: {r}"
    total_row = r
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_note(ws.cell(row=r, column=1),
               "見込量は「北塩原村_サービス見込量.xlsx」の03・04シートの『月間サービス量』を転記します。"
               "単価は直近の実績単価（令和7年度の給付費÷利用量）を基本とし、"
               "報酬改定による補正は上表1の適用倍率で一括して行います（単価欄には改定前の実績単価を入力してください）。"
               "外部リンクにするとファイル移動時に壊れるため、転記は手作業としています。")
    ws.row_dimensions[r].height = 34
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_title(ws.cell(row=r, column=1), "3. 財源構成", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年度", "", "", "自立支援給付計", "国庫負担", "県負担", "村負担",
                             "対令和7年度", "区分・備考"])
    r += 1
    fin_first = r
    fin_rows = [
        ("令和7年度", f"='04_財源構成試算'!D{FUNDING_R7_ROW}", "実績", "受領データによる確定値"),
        ("令和8年度", None, "見込", "村の最新見込を受領して入力する"),
        ("令和9年度", None, "推計", "積上げ結果を入力する（次期計画1年目）"),
        ("令和10年度", None, "推計", "積上げ結果を入力する（次期計画2年目）"),
        ("令和11年度", f"=H{total_row}", "推計", "上表の積上げ合計を自動参照（目標年度）"),
    ]
    for i, (wareki, formula, kubun, note) in enumerate(fin_rows):
        is_calc = formula is not None
        write_row(ws, r, [wareki, "", "", formula,
                          f"=IFERROR(ROUND(D{r}*{RATE_KUNI},0),\"\")",
                          f"=IFERROR(ROUND(D{r}*{RATE_KEN},0),\"\")",
                          f"=IFERROR(D{r}-E{r}-F{r},\"\")",
                          f"=IFERROR(D{r}/D${fin_first}-1,\"\")",
                          f"{kubun}／{note}"],
                  alt=(i % 2 == 1),
                  aligns=["center", "left", "left", "right", "right", "right", "right",
                          "right", "left"],
                  numfmts=[None, None, None, YEN, YEN, YEN, YEN, PCT1, None],
                  fills=[None, None, None,
                         COLORS["calc"] if is_calc else COLORS["input"]] +
                        [COLORS["calc"]] * 4 + [None])
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_title(ws.cell(row=r, column=1),
                "4. 参考：伸び率による推計（第3層の補正にのみ用いる）",
                fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年度", "介護給付費", "前年度比", "障害児給付費", "前年度比",
                             "", "", "", "備考"])
    r += 1
    hist = [
        ("令和2年度", 36002274, None, 746064, None, "基準年度"),
        ("令和3年度", 45991635, 0.277, 3254621, 3.362, "共同生活援助・就労継続支援B型の増"),
        ("令和4年度", 44234021, -0.038, 4030423, 0.238, "障害児給付費のピーク"),
        ("令和5年度", 44419974, 0.004, 3577832, -0.112, "ほぼ横ばい"),
        ("令和6年度", 52698808, 0.186, 1641275, -0.541, "介護は大幅増。障害児の減は就学移行の可能性"),
        ("令和7年度", 53877076, 0.022, 1635703, -0.003, "介護の伸びは鈍化"),
    ]
    for i, (wareki, k, kg, j, jg, note) in enumerate(hist):
        write_row(ws, r, [wareki, k, kg, j, jg, "", "", "", note],
                  alt=(i % 2 == 1),
                  aligns=["center", "right", "right", "right", "right", "left", "left",
                          "left", "left"],
                  numfmts=[None, YEN, "+0.0%;-0.0%", YEN, "+0.0%;-0.0%", None, None, None, None])
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_note(ws.cell(row=r, column=1),
               "介護給付費の前年度比は＋27.7％から−3.8％まで振れており、令和2〜7年度のCAGRは＋8.4％、"
               "直近1年は＋2.2％です。いずれも単一値として将来4年に当てはめる根拠は弱く、"
               "本表は伸び率の不安定さを示す参考資料として置いています。"
               "障害児給付費の減少は、令和6・7年度に児童発達支援が減り放課後等デイサービスが増えていることから、"
               "対象児童の就学に伴う移行である可能性があります（受給者単位の確認が必要）。"
               "ニーズの減少と断定しないでください。"
               "前回計画課題整理でも「1人の利用開始・終了で実績が大きく動くため、機械的な伸び率推計を避ける」ことを"
               "方針としています。")
    ws.row_dimensions[r].height = 62
    return ws


# ============================================================
# 09_村確認事項
# ============================================================
def sheet_confirm(wb):
    ws = add_sheet(
        wb, "09_村確認事項", "村への確認・依頼事項（財源関係）",
        "受領・確認が済んだものは「状態」を更新し、該当シートに入力してください。",
        [30, 52, 26, 12, 32])
    style_header_row(ws, 5, ["確認事項", "確認したい内容", "使途・反映先", "優先度", "状態"])
    rows = [
        ("障がい福祉関係の歳入歳出決算",
         "障害者自立支援給付費負担金（国・県）、障害児施設給付費負担金、地域生活支援事業費補助金の"
         "令和6・7年度交付額と、村の一般財源負担額。",
         "04_財源構成試算との突合", "高", "村資料待ち"),
        ("地域生活支援事業の事業費・交付額",
         "事業別の事業費、補助基準額、国・県補助額、村負担額。補助基準額を超過している事業の有無。",
         "06_地域生活支援事業", "高", "村資料待ち"),
        ("国庫負担基準額の超過の有無",
         "訪問系サービス等で国庫負担基準額を超過し、村の超過負担が生じているかどうか。",
         "04_財源構成試算の注記", "高", "村資料待ち"),
        ("令和8年度の給付費見込",
         "現行計画最終年度の給付費見込額。次期推計の起点となる。",
         "08_R9R11推計", "高", "村資料待ち"),
        ("村単独事業の実績",
         "重度心身障害者医療費助成、在宅重度障害者対策事業、人工透析患者通院交通費助成等の"
         "対象者数・事業費・県補助額。",
         "07_村単独事業", "中", "村資料待ち"),
        ("就労継続支援B型の増加要因",
         "令和2年度9,458,880円から令和7年度19,707,121円へ約2.08倍に増加した要因。"
         "利用者数の増、利用日数の増、報酬改定、重度化のいずれによるものか。",
         "05_サービス別財源構成／08_R9R11推計", "高", "村資料待ち"),
        ("障害児給付費の減少要因",
         "令和4年度4,030,423円から令和6・7年度1,600千円台へ減少した要因。"
         "対象児童の就学・転出・サービス終了等の個別事情を確認する。",
         "03_給付実績_障害児給付費", "高", "村資料待ち"),
        ("短期入所の利用状況",
         "令和6年度2件・令和7年度3件の実績の内容（利用者、利用目的、受入先）。"
         "地域生活支援拠点の緊急受入機能との関係。",
         "05_サービス別財源構成／08_R9R11推計", "中", "村資料待ち"),
        ("マイナス計上の確認",
         "受領原票で一部にマイナス値が見られる。返戻・過誤調整によるものかを確認する。",
         "02・03給付実績", "中", "村資料待ち"),
        ("令和7年度実績の確定状況",
         "令和7年度は12か月分の値が入っているが、確定実績か、見込・調整後データを含むか。",
         "全シート", "中", "村資料待ち"),
    ]
    r = DATA_START
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "left", "center", "center"])
        style_status(ws.cell(row=r, column=5))
        ws.row_dimensions[r].height = 42
        r += 1
    return ws


# ============================================================
# 自己検証
# ============================================================
def verify():
    """サービス別内訳の合計が受領原票の年度合計と一致することを確認する。"""
    for label, data, totals in (("介護給付費", KAIGO_KYUFU, KAIGO_TOTAL),
                                ("障害児給付費", JIDO_KYUFU, JIDO_TOTAL)):
        for _, y in BENEFIT_YEARS:
            amt = sum(d[y][1] for d in data.values() if y in d)
            cnt = sum(d[y][0] for d in data.values() if y in d)
            exp_cnt, exp_amt = totals[y]
            assert amt == exp_amt, f"{label} 令和{y}年度 金額: 内訳{amt:,} ≠ 原票{exp_amt:,}"
            assert cnt == exp_cnt, f"{label} 令和{y}年度 件数: 内訳{cnt:,} ≠ 原票{exp_cnt:,}"
    print("  自己検証: 給付実績のサービス別内訳＝原票の年度合計（2区分×6年度、件数・金額とも一致）")


def main():
    verify()
    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb)
    sheet_rules(wb)
    _, kaigo_total_row = sheet_kaigo(wb)
    _, jido_total_row = sheet_jido(wb)
    sheet_funding(wb, kaigo_total_row, jido_total_row)
    sheet_by_service(wb)
    sheet_chiiki(wb)
    sheet_tandoku(wb)
    sheet_projection(wb)
    sheet_confirm(wb)
    wb.save(OUT_FILE)
    print(f"作成: {OUT_FILE}")


if __name__ == "__main__":
    main()
