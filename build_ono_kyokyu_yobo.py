"""入所（利用）定員と介護予防・総合事業の見える化データを整理ブックにまとめる。

入力:
  07_介護保険_見える化整理/入所（利用）定員/          D25〜D30（6ファイル）
  07_介護保険_見える化整理/介護予防・総合事業/        F1〜F14・F28〜F40（27ファイル）
出力:
  07_介護保険_見える化整理/小野町_第10期介護保険事業計画_供給体制・介護予防整理.xlsx
"""

import pathlib
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = pathlib.Path(__file__).parent
BASE = ROOT / "小野町_引継ぎ_整理済" / "07_介護保険_見える化整理"
TEIIN = BASE / "入所（利用）定員"
YOBO = BASE / "介護予防・総合事業"
NINCHI = BASE / "認知症施策"


def nendo(p):
    """見出しを年度表記に正規化する。"""
    p = str(p).replace("\n", "").replace("/", "").strip()
    m = re.match(r"^平成(\d+)年度$", p)
    if m:
        return f"H{m.group(1)}"
    if p == "令和元年度":
        return "R1"
    m = re.match(r"^令和(\d+)年度$", p)
    if m:
        return f"R{m.group(1)}"
    # 「令和3年3月」のような年度末表記は前年度を指す
    m = re.match(r"^令和(\d+)年3月$", p)
    if m:
        n = int(m.group(1))
        return f"R{n - 1}" if n > 1 else "H30"
    m = re.match(r"^平成(\d+)年3月$", p)
    if m:
        return f"H{int(m.group(1)) - 1}"
    # 「R2」「H29」のような略記、「令和6年4月」のような時点表記はそのまま扱う
    if re.match(r"^(H|R)\d{1,2}$", p) or p in ("R元", "R元年"):
        return "R1" if p in ("R元", "R元年") else p
    if re.match(r"^(令和|平成)\d+年\d+月$", p):
        return p
    return None


def read(path):
    """表形式シートから小野町の行を読み、{項目: {年度: 値}} と年度一覧を返す。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if "表形式" in s]
    if not sheets:
        return {}, []
    ws = wb[sheets[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for r in rows[:6]:
        if r and sum(1 for x in r if x and nendo(x)) >= 2:
            hdr = [nendo(x) if x else None for x in r]
            break
    if hdr is None:
        return {}, []
    out = {}
    for r in rows:
        if r and r[1] == "小野町":
            out[str(r[2])] = {
                hdr[i]: r[i] for i in range(min(len(r), len(hdr)))
                if hdr[i] and isinstance(r[i], (int, float))
            }
    return out, [p for p in hdr if p]


def main():
    wb = openpyxl.Workbook()
    head = PatternFill("solid", fgColor="DDEBF7")
    hl = PatternFill("solid", fgColor="FFF2CC")
    warn = PatternFill("solid", fgColor="FCE4E4")

    # ------------------------------------------------ 00_委員会要点
    ws = wb.active
    ws.title = "00_委員会要点"
    ws.append(["項目", "確認結果", "第10期計画への接続", "委員会での確認ポイント"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = head
    for r in [
        ("入所定員の推移",
         "施設サービス定員は平成29年度まで50人で推移し、令和元年度83人・令和2年度112人へ拡大"
         "（地域密着型介護老人福祉施設58人の整備）。居住系（認知症対応型共同生活介護）は"
         "令和元年度まで35人、令和2年度53人、令和4年度71人、令和6年度98人と段階的に増加",
         "第10期の施設整備の方針と、保険料の施設整備ケースの前提となる",
         "第9期の「新規施設整備は原則行わない」方針と、認知症対応型共同生活介護の定員が"
         "令和4年度71人から令和6年度98人へ増えている事実の関係を整理する"),
        ("供給の充足度",
         "要支援・要介護者1人あたり定員は、施設0.132（令和7年度）、居住系0.116、通所系0.226。"
         "見える化の全国比較では施設の偏差値43.36（やや少ない）、居住系57.40（多い）",
         "居住系に厚く施設に薄い供給構造。第10期のサービス見込量に反映する",
         "認知症対応型共同生活介護への依存度が高いことの評価"),
        ("通いの場",
         "週1回以上の通いの場は平成27年度に3か所・56人（参加率1.7％）があったのみで、"
         "平成28年度以降は0か所・0人が続く。月1回以上も平成28年度の11か所・214人（6.5％）を"
         "ピークに、令和元年度1か所・40人（1.2％）、令和2年度3か所・42人（1.2％）へ減少",
         "第9期の介護予防施策の評価に直結する。第10期の重点施策の設計を要する",
         "国が重点施策とする通いの場が本町でほぼ機能していない要因と、"
         "代替する介護予防の取組の有無を確認する"),
        ("通いの場の全国比較",
         "参考として新地町の週1回以上参加率は平成29年度17.8％、令和2年度15.3％。本町は0.0％",
         "近隣町村との差の要因分析が必要",
         "運営主体・活動場所・活動内容の別（F7〜F14）は本町のデータが未登録である"),
        ("総合事業サービス",
         "F28〜F40の期間は令和元年度・令和2年度の2年度分のみ。値が入っているのは"
         "旧介護予防訪問介護相当（令和元年度10件）、訪問型サービスB（408件・1,051件）、"
         "旧介護予防通所介護相当（39件）、通所型サービスB（693件・564件）の4指標のみ。"
         "訪問型A・C・D、通所型A・C、見守り、配食、介護予防ケアマネジメントは値なし",
         "見える化システムでは第9期評価に足りない。町の事業実績で代替する必要がある",
         "町の総合事業の事業別実績（令和3〜7年度）の提供を依頼する"),
        ("介護予防事業費との整合",
         "一般介護予防事業費は令和5年度で1.0百万円と低水準（特別会計経理状況）。"
         "通いの場が実質ゼロであることと整合する",
         "第10期は通いの場の立ち上げを含む介護予防の体制整備が論点となる",
         "地域支援事業費に占める一般介護予防事業の位置づけを見直すか"),
        ("認定者数の検証",
         "D28・D29の分母から逆算した要支援・要介護者数は、令和6年度933人・令和7年度848人。"
         "令和7年3月末の認定者935人とは一致するが、令和8年3月末の783人とは65人の差がある",
         "認定者数の減少は別系列でも確認できるが、減少幅は系列により異なる",
         "町の認定実績（要介護度別・年度末時点）で確定させる"),
    ]:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", (16, 66, 44, 46)):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------ 01_入所定員
    ws2 = wb.create_sheet("01_入所定員")
    files = [("D25", "D25_定員（施設サービス別）_時系列 (1).xlsx"),
             ("D26", "D26_定員（居住系サービス別）_時系列 (1).xlsx"),
             ("D27", "D27_定員（通所系サービス別）_時系列 (1).xlsx"),
             ("D28", "D28_要支援・要介護者1人あたり定員（施設サービス別）_時系列 (1).xlsx"),
             ("D29", "D29_要支援・要介護者1人あたり定員（居住系サービス別）_時系列 (1).xlsx"),
             ("D30", "D30_要支援・要介護者1人あたり定員（通所系サービス別）_時系列 (1).xlsx")]
    store, per = {}, []
    for sid, fn in files:
        d, p = read(TEIIN / fn)
        store[sid] = d
        if len(p) > len(per):
            per = p
    ws2.append(["指標ID", "項目"] + per)
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = head
    for sid, _ in files:
        for lab, vals in store[sid].items():
            ws2.append([sid, lab] + [vals.get(p) for p in per])
            if "合計" in lab:
                for c in ws2[ws2.max_row]:
                    c.fill = hl
    # 逆算した要支援・要介護者数
    ws2.append([])
    ws2.append(["計算", "定員合計（施設）÷1人あたり定員合計（施設）＝要支援・要介護者数"] +
               [round(a / b) if (a := store["D25"].get("定員合計（施設サービス）", {}).get(p))
                and (b := store["D28"].get("要支援・要介護者1人あたり定員合計（施設サービス）", {}).get(p))
                else None for p in per])
    ws2.append(["計算", "定員合計（居住系）÷1人あたり定員合計（居住系）＝要支援・要介護者数"] +
               [round(a / b) if (a := store["D26"].get("定員合計（居住系サービス）", {}).get(p))
                and (b := store["D29"].get("要支援・要介護者1人あたり定員合計（居住系サービス）", {}).get(p))
                else None for p in per])
    for row in ws2.iter_rows(min_row=ws2.max_row - 1, max_row=ws2.max_row):
        for c in row:
            c.fill = warn
    ws2.column_dimensions["A"].width = 9
    ws2.column_dimensions["B"].width = 56
    ws2.freeze_panes = "C2"

    # ------------------------------------------------ 02_通いの場
    ws3 = wb.create_sheet("02_通いの場")
    kayoi = [("F1", "F1_週１回以上の通いの場の参加率_時系列 (1).xlsx"),
             ("F2", "F2_週1回以上の通いの場の参加者数_時系列 (1).xlsx"),
             ("F3", "F3_週1回以上の通いの場の箇所数_時系列.xlsx"),
             ("F4", "F4_月１回以上の通いの場の参加率_時系列.xlsx"),
             ("F5", "F5_月１回以上の通いの場の参加者数_時系列.xlsx"),
             ("F6", "F6_月１回以上の通いの場の箇所数_時系列.xlsx")]
    kper = []
    kstore = {}
    for sid, fn in kayoi:
        d, p = read(YOBO / fn)
        kstore[sid] = d
        if len(p) > len(kper):
            kper = p
    ws3.append(["指標ID", "項目"] + kper)
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = head
    for sid, _ in kayoi:
        for lab, vals in kstore[sid].items():
            ws3.append([sid, lab] + [vals.get(p) for p in kper])
    ws3.append([])
    ws3.append(["備考", "週1回以上は平成27年度のみ実績があり、平成28年度以降は0が続く"])
    ws3.append(["備考", "月1回以上は平成28年度の214人・11か所をピークに、令和元年度以降は40人台・1〜3か所"])
    ws3.append(["備考", "F7〜F14（運営主体別・活動場所別・活動内容別／2020年地域別）は本町の値が未登録"])
    ws3.column_dimensions["A"].width = 9
    ws3.column_dimensions["B"].width = 56
    ws3.freeze_panes = "C2"

    # ------------------------------------------------ 03_総合事業サービス
    ws4 = wb.create_sheet("03_総合事業サービス")
    ws4.append(["指標ID", "指標名", "項目", "令和元年度", "令和2年度", "状態"])
    for c in ws4[1]:
        c.font = Font(bold=True)
        c.fill = head
    for f in sorted(YOBO.glob("F[23][0-9]*.xlsx"),
                    key=lambda x: int(re.search(r"F(\d+)_", x.name).group(1))):
        sid = re.match(r"^(F\d+)_", f.name).group(1)
        if int(sid[1:]) < 28:
            continue
        d, p = read(f)
        name = f.name.split("_", 1)[1].replace("_時系列.xlsx", "").replace("_時系列 (1).xlsx", "")
        any_val = False
        for lab, vals in d.items():
            v = [vals.get(x) for x in p]
            if any(x is not None for x in v):
                any_val = True
            ws4.append([sid, name[:44], lab[:40]] + v[:2] +
                       ["値あり" if any(x is not None for x in v) else "値なし"])
        if not any_val:
            for row in ws4.iter_rows(min_row=ws4.max_row - len(d) + 1, max_row=ws4.max_row):
                for c in row:
                    c.fill = warn
    ws4.append([])
    ws4.append(["備考", "期間は令和元年度・令和2年度の2年度分のみ。令和3年度以降は未登録"])
    ws4.append(["備考", "【65歳以上人口1万対】の率はすべて未算出（分母の65歳以上人口が未登録）"])
    for col, w in zip("ABCDEF", (9, 46, 42, 13, 13, 10)):
        ws4.column_dimensions[col].width = w
    ws4.freeze_panes = "A2"

    # ------------------------------------------------ 04_認知症施策
    ws45 = wb.create_sheet("04_認知症施策")
    ws45.append(["指標ID", "指標名", "項目", "前回", "直近", "備考"])
    for c in ws45[1]:
        c.font = Font(bold=True)
        c.fill = head
    for f in sorted(NINCHI.glob("J*.xlsx")):
        sid = re.match(r"^(J\d+)_", f.name).group(1)
        d, p = read(f)
        name = f.name.split("_", 1)[1].replace("_時系列.xlsx", "")
        for lab, vals in d.items():
            v = [vals.get(x) for x in p]
            ws45.append([sid, name[:40], lab[:30]] + (v + [None, None])[:2] +
                        [f"{p[0]}→{p[-1]}" if p else ""])
    ws45.append([])
    for t in ["公表されているのはJ16・J17・J18の3指標のみ。研修系（J1〜J15）は本町の値が未公表",
              "初期集中支援チームの訪問実績は令和2年度・令和5年度とも0件（県内59団体中30団体が0件）",
              "認知症地域支援推進員は1人（令和3年3月）から4人（令和6年3月）へ増員。県内中央値3人",
              "認知症カフェは1か所で横ばい。県内中央値1か所"]:
        ws45.append(["備考", t])
    for col, w in zip("ABCDEF", (9, 42, 32, 12, 12, 24)):
        ws45.column_dimensions[col].width = w

    # ------------------------------------------------ 90_取込確認
    ws5 = wb.create_sheet("90_取込確認")
    ws5.append(["区分", "指標ID", "ファイル", "自治体", "小野町の値"])
    for c in ws5[1]:
        c.font = Font(bold=True)
        c.fill = head
    for folder, kubun in ((TEIIN, "入所定員"), (YOBO, "介護予防・総合事業"), (NINCHI, "認知症施策")):
        for f in sorted(folder.glob("*.xlsx")):
            m = re.match(r"^([A-Z]\d+)_", f.name)
            if not m:
                continue
            d, p = read(f)
            n = sum(1 for vals in d.values() for x in vals.values() if x is not None)
            ws5.append([kubun, m.group(1), f.name, "小野町", f"{n}件" if n else "値なし"])
            if not n:
                for c in ws5[ws5.max_row]:
                    c.fill = warn
    for col, w in zip("ABCDE", (18, 9, 62, 10, 12)):
        ws5.column_dimensions[col].width = w

    path = BASE / "小野町_第10期介護保険事業計画_供給体制・介護予防整理.xlsx"
    wb.save(path)
    print("出力:", path)
    print(f"  定員 期間 {per[0]}〜{per[-1]} / 通いの場 期間 {kper[0]}〜{kper[-1]}")


if __name__ == "__main__":
    main()
