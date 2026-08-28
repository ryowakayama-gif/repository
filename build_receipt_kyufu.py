# -*- coding: utf-8 -*-
"""給付実績データ（資料提供依頼No.23）の受領点検.

令和8年8月28日に12ファイルを受領した。
  【大雪】【東川町】【美瑛町】【東神楽町】の各R6・R7・R8

内容を点検し、
  ① 個人情報の取扱い
  ② 代表KPI H15（給付費の計画乖離率）の算定
  ③ 見える化システムの値との差
  ④ 第2段階（サービス見込量）への反映
を整理する。

シート構成
  00_この表について（個人情報の取扱いを含む）
  01_給付費の実績と計画との乖離
  02_サービス別・要介護度別の実績
  03_成果品への反映箇所
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_kyufu_jisseki import KYUFU, KYUFU_TOTAL     # noqa: E402

ODIR = "/home/user/repository/output"
OUT = os.path.join(ODIR, "第10期計画_給付実績データの受領点検.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 第9期計画の計画値（妥当性検証報告書による）
PLAN_R6_HYOJUN = 3127780210      # 標準給付費見込額（令和6年度）
PLAN_R6_SOU = 2924324000         # 総給付費（令和6年度）
N1 = 9090                        # 第1号被保険者数（令和8年3月末）
MIERUKA_R7 = 26885               # 見える化の第1号1人当たり給付月額（R7）

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 52
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=30):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=24, align=None, bold=False,
         numfmt=None):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if numfmt and i in numfmt:
            c.number_format = numfmt[i]
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=84):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_この表について", "給付実績データの受領点検",
           "令和8年8月25日の進捗確認の打合せでご依頼した給付実績データ"
           "（資料提供依頼No.23）を、令和8年8月28日に12ファイル受領しました。"
           "内容を点検した結果を示します。",
           [16, 24, 54, 42])

r = lead(ws, 4, "【1　受領したファイル】", 4)
r = header(ws, r, ["ファイル", "規模", "収録内容", "点検の結果"])
for a in [
    ("【大雪】R6.給付費データ集計", "9シート\n57,142行",
     "給付費（個票）、償還払い、介護度別給付費、"
     "介護度別利用者数（延べ利用者数・延べ利用回数）、"
     "1月あたり利用者、1人1月あたり給付費、1人1月あたり利用回数。",
     "令和6年度の通年の実績です。"
     "34サービス×7要介護度の集計が得られました。"),
    ("【大雪】R7.給付費データ集計", "9シート\n59,576行", "同上。",
     "令和7年度の通年の実績です。"),
    ("【大雪】R8.給付費データ集計", "9シート", "同上。",
     "令和8年度は年度途中です。"
     "給付費の総額は令和6・7年度の約36％であり、"
     "4〜7か月分と考えられます。対象期間のご確認をお願いします。"),
    ("【東川町】R6・R7・R8.給付費データ", "各9シート",
     "同上。集計シートは東川町分（証記載保険者番号14589で抽出）。",
     "町別の集計が得られました。"
     "ただし第1シート（給付費）は大雪地区広域連合の全件であり、"
     "町に絞り込まれていません（下記3）。"),
    ("【美瑛町】R6・R7・R8.給付費データ", "各9シート",
     "同上。集計シートは美瑛町分（同14597で抽出）。",
     "同上。なお令和7年度のファイルは、"
     "利用回数のシート名が「（東神楽）」となっています。"
     "収録されている値は東神楽町のファイルと一致せず美瑛町の値であるため、"
     "シート名の誤りとして扱いました（下記4）。"),
    ("【東神楽町】R6・R7・R8.給付費データ", "各9シート",
     "同上。集計シートは東神楽町分（同14530で抽出）。", "同上。"),
]:
    r = body(ws, r, list(a), {}, height=64, align={2: "center"})

r += 1
r = lead(ws, r, "【2　個人情報の取扱いについて（ご確認をお願いします）】", 4)
r = note(ws, r,
         "受領した12ファイルの第1シート（給付費）には、次の個人情報が"
         "含まれています。\n"
         "　被保険者番号、氏名（カナ・漢字）、性別、生年月日、"
         "住民コード、世帯番号、行政区コード、住所コード、郵便番号\n"
         "行数は令和6年度57,142行、令和7年度59,576行です。\n"
         "\n"
         "受託者は、本業務の成果品には集計値のみを収録し、"
         "個票データを収録しない運用としています。"
         "本ファイルについても、集計シート"
         "（介護度別給付費、1月あたり利用者、1人1月あたり給付費、"
         "1人1月あたり利用回数）の値のみを用い、"
         "第1シートの個票は集計にも用いていません。\n"
         "\n"
         "つきましては、次の2点についてご確認をお願いします。\n"
         "① 個票を含むファイルの送付が意図されたものか。"
         "集計シートのみで足りる場合は、次回以降は集計シートのみを"
         "ご送付いただければ、個人情報を授受せずに作業できます。\n"
         "② 受託者における本ファイルの保管期間と廃棄の方法。"
         "ご指示がない場合は、業務終了後に消去し、"
         "消去の記録を残す取扱いといたします。", 4, height=230)

r += 1
r = lead(ws, r, "【3　町別ファイルについて】", 4)
r = note(ws, r,
         "町別（東川町・美瑛町・東神楽町）のファイルは、"
         "集計シートは当該町分に絞られていますが、"
         "第1シート（給付費）は大雪地区広域連合の全件が収録されています"
         "（行数・内容とも【大雪】のファイルと一致します）。\n"
         "町へ提供される資料である場合、"
         "他町の被保険者の個人情報が含まれることになります。"
         "意図されたものかご確認をお願いします。", 4, height=88)

r += 1
r = lead(ws, r, "【4　令和7年度の集計が一致しません（ご確認をお願いします）】", 4)
r = note(ws, r,
         "3町の給付費総額の合計と、大雪地区広域連合の給付費総額とを"
         "突き合わせた結果は次のとおりです。\n"
         "　令和6年度　3町計 3,084,937,834円＝大雪 3,084,937,834円（差0）\n"
         "　令和7年度　3町計 3,213,054,558円－大雪 3,083,474,291円"
         "＝＋129,580,267円（＋4.20％）\n"
         "　令和8年度　3町計 1,110,075,123円＝大雪 1,110,075,123円（差0）\n"
         "\n"
         "令和6年度と令和8年度は完全に一致する一方、"
         "令和7年度のみ3町の合計が大雪地区広域連合を1億2,958万円上回ります。"
         "受託者の側では原因を特定できません。\n"
         "【美瑛町】令和7年度のファイルは、"
         "利用回数のシート名が他年度と異なる（「（東神楽）」となっている）ため、"
         "他年度とは異なる手順で作成された可能性があります。\n"
         "\n"
         "令和7年度について、【大雪】と【美瑛町】のいずれかのファイルの"
         "再作成をご検討いただけますでしょうか。"
         "解消するまでは、令和7年度の町別の値は用いず、"
         "大雪地区広域連合の合計のみを用います。", 4, height=180)

r += 1
r = lead(ws, r, "【4　この資料により進むこと】", 4)
r = header(ws, r, ["項目", "内容", "", ""])
for a, b in [
    ("代表KPI H15（給付費の計画乖離率）",
     "令和6年度の実績が確定したため、第9期計画の計画値との乖離を"
     "算定できるようになりました（01シート）。"
     "分母を標準給付費見込額とするか総給付費とするかで結果が変わるため、"
     "ご決定をお願いします。"),
    ("将来推計 第2段階（サービス見込量）",
     "34サービス×7要介護度の月平均利用者数・1人1月あたり利用回数・"
     "1人1月あたり給付費が得られました。"
     "これまで見える化システムのD系列により推定していた部分を、"
     "実績により置き換えられます（02シート）。"),
    ("将来推計 第3段階（給付費・保険料）",
     "サービス別・要介護度別の単価が実績から得られるため、"
     "給付費の算定に用いられます。"),
    ("見える化システムの値との差",
     "打合せでご指摘のあったとおり、"
     "見える化は年度途中で締めた値を収録しています。"
     "本データにより通年の実績が確定しました（01シート）。"),
]:
    r = body(ws, r, [a, b, "", ""], {1: MID_B}, height=54)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

# ============================================================ 01
ws = sheet("01_給付費の実績と乖離", "給付費の実績と第9期計画との乖離",
           "受領した給付実績により、令和6年度・令和7年度の通年の給付費が"
           "確定しました。第9期計画の計画値との乖離を算定します。",
           [22, 20, 20, 20, 16, 46])

r = lead(ws, 4, "【1　給付費の総額（要介護度別）】", 6)
r = header(ws, r, ["年度", "要支援1・2", "要介護1・2", "要介護3〜5",
                   "総計", "所見"])
for y, lab in (("R6", "令和6年度"), ("R7", "令和7年度"),
               ("R8", "令和8年度（途中）")):
    v = KYUFU_TOTAL[("大雪", y)]
    ys = (v[0] or 0) + (v[1] or 0)
    k12 = (v[2] or 0) + (v[3] or 0)
    k35 = (v[4] or 0) + (v[5] or 0) + (v[6] or 0)
    obs = ""
    if y == "R7":
        obs = ("総額は令和6年度とほぼ同額ですが、"
               "要介護1・2が減り要介護4・5が増えています。")
    if y == "R8":
        obs = ("年度途中のため通年ではありません。"
               "令和7年度の36.0％です。対象期間のご確認をお願いします。")
    r = body(ws, r, [lab, ys, k12, k35, v[7], obs], {}, height=32,
             align={2: "right", 3: "right", 4: "right", 5: "right"},
             numfmt={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0"},
             bold=(y == "R7"))

r += 1
r = lead(ws, r, "【2　第9期計画との乖離（代表KPI H15）】", 6)
r = header(ws, r, ["比較の対象", "計画値", "実績（令和6年度）",
                   "差", "乖離率", "評価"])
a6 = KYUFU_TOTAL[("大雪", "R6")][7]
for nm, pv in (("標準給付費見込額", PLAN_R6_HYOJUN), ("総給付費", PLAN_R6_SOU)):
    d = a6 - pv
    rate = d / pv * 100
    ok = abs(rate) <= 5
    r = body(ws, r, [nm, pv, a6, d, "%+.2f％" % rate,
                     "目標±5％以内を満たします" if ok
                     else "目標±5％を超えます"],
             {5: OK_G if ok else NG_O, 6: OK_G if ok else NG_O},
             height=26,
             align={2: "right", 3: "right", 4: "right", 5: "right"},
             numfmt={2: "#,##0", 3: "#,##0", 4: "#,##0"})

r = note(ws, r,
         "第9期計画は標準給付費見込額と総給付費の両方を掲載しており、"
         "どちらを分母とするかで結果が変わります。\n"
         "　標準給付費見込額を分母とすると ▲1.37％（目標±5％以内）\n"
         "　総給付費を分母とすると ＋5.49％（目標±5％を超える）\n"
         "代表KPI H15の定義として、いずれを用いるかのご決定をお願いします。"
         "受託者は、計画期間を通じた財政運営の指標であることから、"
         "標準給付費見込額を分母とすることを案といたします。", 6, height=104)

r += 1
r = lead(ws, r, "【3　見える化システムの値との差】", 6)
r = header(ws, r, ["項目", "本データ（実績）", "見える化", "差", "", "説明"])
for y, lab in (("R6", "令和6年度"), ("R7", "令和7年度")):
    t = KYUFU_TOTAL[("大雪", y)][7]
    per = round(t / 12 / N1)
    r = body(ws, r, ["第1号1人当たり給付月額\n（%s）" % lab, per,
                     MIERUKA_R7 if y == "R7" else "―",
                     per - MIERUKA_R7 if y == "R7" else "―", "",
                     "見える化は年度途中で締めた値を収録しているため、"
                     "実績より小さくなります。"
                     "本データにより通年の実績が確定しました。"
                     if y == "R7" else ""],
             {}, height=34,
             align={2: "right", 3: "right", 4: "right"},
             numfmt={2: "#,##0", 3: "#,##0", 4: "#,##0"})

r = note(ws, r,
         "注1）本データの給付費を第1号被保険者数9,090人（令和8年3月末）で"
         "除した参考値です。"
         "見える化の給付月額は第2号被保険者を含むか、"
         "対象期間が異なる可能性があります。"
         "定義の対応関係は、年報の受領後に確認します。\n"
         "注2）計画本文の記述を本データに置き換えるかは、"
         "定義の確認後に判断します。"
         "現時点では見える化の値のままとしています。", 6, height=76)

r += 1
r = lead(ws, r, "【4　町別の給付費】", 6)
r = header(ws, r, ["年度", "東川町", "美瑛町", "東神楽町",
                   "3町計と大雪の差", "所見"])
N1T = {"東川町": 2654, "美瑛町": 3646, "東神楽町": 2858}
for y, lab in (("R6", "令和6年度"), ("R7", "令和7年度"),
               ("R8", "令和8年度（途中）")):
    d = KYUFU_TOTAL[("大雪", y)][7]
    v = {t: KYUFU_TOTAL[(t, y)][7] for t in N1T}
    gap = sum(v.values()) - d
    obs = ("一致します。" if gap == 0 else
           "3町計が大雪を%s円（%+.2f％）上回ります。"
           "原因を特定できていません（00シート4）。"
           % (format(gap, ","), gap / d * 100))
    r = body(ws, r, [lab, v["東川町"], v["美瑛町"], v["東神楽町"], gap, obs],
             {5: OK_G if gap == 0 else NG_O}, height=38,
             align={2: "right", 3: "right", 4: "right", 5: "right"},
             numfmt={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0"})

r += 1
r = header(ws, r, ["年度", "東川町", "美瑛町", "東神楽町", "", "参考"])
for y, lab in (("R6", "令和6年度"), ("R8", "令和8年度（途中）")):
    v = {t: round(KYUFU_TOTAL[(t, y)][7] / N1T[t] / 12) for t in N1T}
    r = body(ws, r, ["第1号1人当たり\n給付月額（%s）" % lab,
                     v["東川町"], v["美瑛町"], v["東神楽町"], "",
                     "第1号被保険者数（東川2,654人・美瑛3,646人・"
                     "東神楽2,858人）で除した参考値です。"
                     if y == "R6" else ""],
             {}, height=34,
             align={2: "right", 3: "right", 4: "right"},
             numfmt={2: "#,##0", 3: "#,##0", 4: "#,##0"})

r = note(ws, r,
         "注1）令和7年度は3町計と大雪地区広域連合の合計が一致しないため、"
         "1人当たりの値を算定していません。\n"
         "注2）月平均利用者数についても、3町計が大雪地区広域連合を"
         "100〜140人上回ります。"
         "これは町ごとに月平均を四捨五入しているためであり、"
         "34サービス×7区分の丸め差の累積として説明できます。"
         "給付費の令和7年度の差（4.20％）とは性質が異なります。", 6, height=76)

# ============================================================ 02
ws = sheet("02_サービス別の実績", "サービス別・要介護度別の実績（令和7年度）",
           "34サービスについて、月平均利用者数・1人1月あたり利用回（日）数・"
           "1人1月あたり給付費が得られました。"
           "将来推計 第2段階（サービス見込量）の基礎になります。",
           [30, 10, 10, 10, 10, 10, 10, 10, 12])

CARE = ["要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"]
for label in ("月平均利用者数", "1人1月あたり利用回数", "1人1月あたり給付費"):
    r = lead(ws, ws.max_row + 2 if ws.max_row > 3 else 4,
             "【%s（令和7年度・大雪地区広域連合）】" % label, 9)
    r = header(ws, r, ["サービス"] + CARE + ["計"])
    rows = KYUFU[("大雪", "R7")].get(label, [])
    for row in rows:
        vals = [row[0]] + list(row[1:8])
        tot = sum(x for x in row[1:8] if isinstance(x, (int, float)))
        r = body(ws, r, vals + [tot if label == "月平均利用者数" else ""],
                 {}, height=16,
                 align={i: "right" for i in range(2, 10)},
                 numfmt={i: "#,##0" for i in range(2, 10)})

r = note(ws, ws.max_row + 2,
         "注1）出典は大雪地区広域連合の給付費データ集計（令和7年度）です。"
         "受領ファイルの集計シートの値をそのまま収録しています。\n"
         "注2）「1人1月あたり給付費」は円、"
         "「1人1月あたり利用回数」は回又は日です。"
         "サービスにより単位が異なります。\n"
         "注3）予防給付のサービスは要支援1・2の列に、"
         "介護給付のサービスは要介護1〜5の列に値が入ります。\n"
         "注4）令和6年度・令和8年度及び東神楽町分も同じ形式で"
         "収録しています（data_kyufu_jisseki.py）。", 9, height=88)

# ============================================================ 03
ws = sheet("03_成果品への反映", "成果品への反映箇所",
           "受領した給付実績により更新する箇所と、その時期を示します。",
           [6, 32, 44, 28, 14])

r = header(ws, 4, ["No.", "成果品・箇所", "更新の内容", "前提", "時期"])
for a in [
    (1, "将来推計 第2段階（サービス見込量）",
     "34サービス×7要介護度の月平均利用者数・1人1月あたり利用回数を"
     "実績に置き換える。これまで見える化D系列から推定していた部分が"
     "実績によるものになる。",
     "決定1（人口推計の基礎）・決定2（認定率のシナリオ）", "令和8年9月"),
    (2, "将来推計 第3段階（給付費・保険料）",
     "サービス別・要介護度別の1人1月あたり給付費を単価として用いる。",
     "上記に加え資料No.9（決算・基金）、国の政令改正の告示",
     "令和8年12月"),
    (3, "代表KPI H15（給付費の計画乖離率）",
     "令和6年度の実績3,084,937,834円により乖離率を算定した。"
     "分母の定義のご決定を要する。",
     "分母の定義のご決定", "ご決定後"),
    (4, "計画素案 第2章第2節（給付の状況）",
     "見える化の値（年度途中で締めた値）を、"
     "通年の実績に置き換えるかを検討する。定義の対応関係の確認を要する。",
     "年報（資料No.21）の受領", "令和8年9月"),
    (5, "妥当性検証報告書 13_受給率の内訳と推移",
     "要介護度別の給付費の推移を実績で更新する。"
     "令和6→7年度で要介護1・2が減り要介護4・5が増えた点を扱う。",
     "―", "令和8年9月"),
    (6, "計画素案 第6章第2節（サービス見込量）",
     "実績に基づく見込量に差し替える。", "上記1の完了", "令和8年10月"),
    (7, "必要事項の一覧・業務工程管理表",
     "資料依頼No.23を「受領（東川町・美瑛町分は未受領）」に改める。",
     "―", "反映済み"),
]:
    f = {5: OK_G if a[4] == "反映済み" else IN_Y}
    r = body(ws, r, list(a), f, height=52, align={1: "center", 5: "center"})

r += 1
r = lead(ws, r, "【引き続きご提供をお願いする資料】", 5)
r = header(ws, r, ["No.", "資料", "理由", "", "時期"])
for a in [
    ("23-2", "東川町・美瑛町の給付実績データ（令和6・7年度）",
     "3町の比較に必要です。東神楽町分のみでは町間の差を扱えません。",
     "", "令和8年9月"),
    (21, "介護保険事業状況報告　年報（令和6・7年度）",
     "本データの給付費と見える化の値の定義の対応関係を確認するために"
     "必要です。", "", "令和8年9月"),
    (22, "同　月報（令和8年4〜7月分）",
     "令和8年度分の対象期間を確定するために必要です。", "", "令和8年9月"),
    (13, "令和6〜8年度の施策・事業実績",
     "第9期のプロセス評価に必要です。最優先で未受領です。", "", "受領次第"),
]:
    r = body(ws, r, list(a), {}, height=40, align={1: "center", 5: "center"})
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)

wb.save(OUT)
print("saved:", os.path.basename(OUT), "sheets=%d" % len(wb.sheetnames))
for k in sorted(KYUFU_TOTAL):
    print("  %-8s %-3s 総給付費 %s円" % (k[0], k[1],
                                    format(KYUFU_TOTAL[k][7], ",")))
print("  H15 対 標準給付費見込額 %+.2f％／対 総給付費 %+.2f％"
      % ((a6 - PLAN_R6_HYOJUN) / PLAN_R6_HYOJUN * 100,
         (a6 - PLAN_R6_SOU) / PLAN_R6_SOU * 100))
