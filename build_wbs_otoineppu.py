# -*- coding: utf-8 -*-
"""
音威子府村高齢者福祉計画・第10期介護保険事業計画 策定業務委託
WBS（作業分解構成図）ジェネレータ

仕様書（音威子府村 提供）に基づき、以下のシートを持つ管理ブックを生成する。
  00_業務概要 / 01_WBS / 02_スケジュール(ガント) / 03_マイルストーン
  04_成果物一覧 / 05_推計作業手順 / 06_必要資料リスト / 07_役割分担
  08_リスク・課題管理 / 09_スコープ管理 / 10_工数集計
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = "/home/user/repository/output"
os.makedirs(OUT_DIR, exist_ok=True)

FILENAME = "WBS_音威子府村_高齢者福祉計画・第10期介護保険事業計画.xlsx"
PLAN_NAME = "音威子府村高齢者福祉計画・第10期介護保険事業計画"
CLIENT = "音威子府村"

# 契約締結日は未確定のため、令和8年9月1日で仮置き（確定後に CONTRACT_START を修正）
CONTRACT_START = date(2026, 9, 1)
CONTRACT_END = date(2027, 3, 31)

FONT = "游ゴシック"

COLORS = {
    "header":  "1F3864",
    "subhead": "2E75B6",
    "band":    "DDEBF7",
    "alt":     "F7FAFC",
    "note":    "FFF3F3",
    "white":   "FFFFFF",
    "L1":      "203864",
    "L2":      "8EA9DB",
    "ms":      "C00000",
    # 大分類カラー（ガント／集計で共通利用）
    "C1": "7F4FBF",   # 業務管理・打合せ
    "C2": "2CA02C",   # 現状の評価・分析
    "C3": "1F77B4",   # 人口・認定者数推計
    "C4": "ED7D31",   # サービス見込量・保険料
    "C5": "C00000",   # 計画素案の作成
    "C6": "1F3864",   # 成果品の作成・納品
}

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ============================================================
# 共通スタイル関数（build_excel.py の作法に準拠）
# ============================================================
def style_title(cell, text, fill=COLORS["header"], font_color="FFFFFF", size=14):
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)


def style_subhead(cell, text, fill=COLORS["subhead"]):
    cell.value = text
    cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)


def style_header_row(ws, row, headers, fill=COLORS["subhead"]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER


def style_data_cell(cell, alt=False, center=False):
    cell.font = Font(name=FONT, size=10)
    cell.alignment = Alignment(vertical="center",
                               horizontal="center" if center else "left",
                               wrap_text=True, indent=0 if center else 1)
    cell.border = BORDER
    if alt:
        cell.fill = PatternFill("solid", fgColor=COLORS["alt"])


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fill_table(ws, start_row, rows, widths, center_cols=()):
    """単純な明細テーブルを敷き込む"""
    for r, values in enumerate(rows, start_row):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in center_cols))
    return start_row + len(rows) - 1


# ============================================================
# WBS データ
#   (WBS No, レベル, 作業項目, 作業内容・実施方法, 成果物, 主担当,
#    村の対応, 工数(人日), 開始, 完了, 先行作業, 仕様書該当, 備考)
#   レベル1・2は集計行（工数・日程は子タスクから自動算出）
# ============================================================
D = date

TASKS = [
 # ---------------- 1. 業務管理・打合せ ----------------
 ("1",     1, "業務管理・打合せ", "", "", "", "", None, None, None, "", "2 委託期間／4 打合せ", ""),
 ("1.1",   2, "業務着手準備", "", "", "", "", None, None, None, "", "4 打合せ", ""),
 ("1.1.1", 3, "契約締結・業務実施体制の確定",
  "契約書・仕様書の内容確認、業務従事者（計画策定担当者）の確定、連絡体制・緊急連絡先の共有",
  "業務従事者届／連絡体制表", "受託者", "契約手続", 0.5, D(2026,9,1), D(2026,9,4),
  "－", "2 委託期間", "打合せには計画策定担当者本人が出席（仕様書4）"),
 ("1.1.2", 3, "業務実施計画書・詳細工程表の作成",
  "本WBSに基づく実施計画書・工程表を作成し、村と協議のうえ確定。以降の変更は都度改訂",
  "業務実施計画書／詳細工程表（本WBS）", "受託者", "内容確認・承認", 1.0, D(2026,9,1), D(2026,9,18),
  "1.1.1", "4 打合せ", "第1回打合せで協議し確定。スケジュールは以降も適宜協議（仕様書4）"),
 ("1.1.3", 3, "【第1回打合せ】キックオフ",
  "業務実施計画・工程の協議、貸与資料の確認、計画の方向性・論点の共有、庁内体制の確認",
  "第1回打合せ資料／議事録", "受託者", "日程調整・出席・資料準備", 1.0, D(2026,9,10), D(2026,9,10),
  "1.1.1", "4 打合せ（3回程度）", "★マイルストーン／打合せ1回目"),
 ("1.1.4", 3, "個人情報・機密情報の取扱手続",
  "個人情報保護に関する誓約書の提出、データ授受方法（受渡・保管・返却）の取決め、アクセス権限設定",
  "誓約書／情報取扱ルール", "受託者", "条例に基づく確認", 0.5, D(2026,9,1), D(2026,9,11),
  "1.1.1", "7 留意事項（個人情報）", "音威子府村個人情報保護条例を遵守"),

 ("1.2",   2, "進行管理", "", "", "", "", None, None, None, "", "4 打合せ／5 その他", ""),
 ("1.2.1", 3, "定例連絡・進捗報告",
  "月1回程度のメール・電話・Web会議による進捗報告と論点確認（対面打合せ3回を補完）",
  "進捗報告メモ（月次）", "受託者", "内容確認・指示", 2.0, D(2026,9,1), D(2027,3,31),
  "1.1.1", "4 打合せ", "対面3回の制約をWeb会議等で補完"),
 ("1.2.2", 3, "【第2回打合せ】現状分析・骨子報告",
  "現状分析・現行計画評価の結果報告、課題と施策の方向の協議、計画骨子（章立て）案の提示・確定",
  "第2回打合せ資料／議事録", "受託者", "日程調整・出席・庁内調整", 1.0, D(2026,12,10), D(2026,12,10),
  "2.4.3, 5.2.1", "4 打合せ（3回程度）", "★マイルストーン／打合せ2回目"),
 ("1.2.3", 3, "【第3回打合せ】計画素案・保険料報告",
  "サービス見込量・保険料算定結果の報告と協議、計画素案の提示、パブリックコメント等の進め方確認",
  "第3回打合せ資料／議事録", "受託者", "日程調整・出席・庁内調整", 1.0, D(2027,2,10), D(2027,2,10),
  "4.3.4, 5.3.10", "4 打合せ（3回程度）", "★マイルストーン／打合せ3回目"),
 ("1.2.4", 3, "国の指示・基本指針変更への対応協議",
  "国の基本指針・通知・Q&A・介護報酬改定情報を継続的に収集し、仕様変更が必要な場合は村と協議",
  "国通知等の整理メモ／協議記録", "受託者", "協議・決定", 0.5, D(2026,9,1), D(2027,3,31),
  "1.1.1", "5 その他", "特別の事情がない限り国の指示を遵守"),

 ("1.3",   2, "資料収集・整理", "", "", "", "", None, None, None, "", "3(1) 現状の評価・分析", ""),
 ("1.3.1", 3, "村提供資料の受領・確認",
  "貸与資料リスト（本ブック『06_必要資料リスト』）を提示し、資料を受領。不足・欠測の確認と追加依頼",
  "貸与資料リスト（受領チェック済）", "受託者", "資料抽出・提供", 0.5, D(2026,9,8), D(2026,9,30),
  "1.1.3", "7 留意事項（連携・協議）", "契約後2週間以内にリスト提示"),
 ("1.3.2", 3, "公表統計・システムデータの収集",
  "国勢調査、住民基本台帳人口、社人研将来推計人口、介護保険事業状況報告、地域包括ケア「見える化」システム等",
  "基礎統計データセット", "受託者", "－", 1.0, D(2026,9,8), D(2026,10,15),
  "1.1.3", "3(1) 現状の評価・分析", "出典・取得日を明記して管理"),
 ("1.3.3", 3, "関連計画・関連資料の収集",
  "村総合計画、地域福祉計画、健康増進計画、北海道の関連計画等を収集し、記載内容・目標値を整理",
  "関連計画整理表", "受託者", "計画提供", 0.5, D(2026,9,8), D(2026,10,15),
  "1.3.1", "7 留意事項（総合計画等との整合）", "整合性確認は5.1.3で実施"),

 # ---------------- 2. 現状の評価・分析 ----------------
 ("2",     1, "現状の評価・分析", "", "", "", "", None, None, None, "", "3(1) 現状の評価・分析", ""),
 ("2.1",   2, "基礎的資料による現状把握・分析", "", "", "", "", None, None, None, "", "3(1)① 統計資料等", ""),
 ("2.1.1", 3, "人口・世帯の動向分析",
  "総人口・年齢3区分別人口・高齢化率・前期／後期高齢者、高齢者単身世帯・高齢者のみ世帯の推移を分析",
  "人口・世帯分析図表", "受託者", "住基データ提供", 1.5, D(2026,10,1), D(2026,10,23),
  "1.3.2", "3(1)① 統計資料等", "直近10年の経年変化で整理"),
 ("2.1.2", 3, "地域資源・社会基盤の整理",
  "村内・近隣の介護サービス事業所、医療機関、通所・訪問系資源、住民主体の活動、生活交通の状況を整理",
  "地域資源マップ／一覧表", "受託者", "事業所情報提供", 1.0, D(2026,10,1), D(2026,10,23),
  "1.3.1", "3(1)① 統計資料等", "圏域外（近隣市町村）利用の実態も整理"),
 ("2.1.3", 3, "北海道・全国・類似団体との比較分析",
  "高齢化率、認定率、受給率、1人当たり給付費、保険料水準を北海道平均・全国・類似規模団体と比較",
  "比較分析図表", "受託者", "－", 0.5, D(2026,10,13), D(2026,10,30),
  "2.1.1", "3(1)① 統計資料等", "「見える化」システムを活用"),

 ("2.2",   2, "高齢者福祉サービス実績の分析", "", "", "", "", None, None, None, "", "3(1)② 高齢者福祉サービス実績", ""),
 ("2.2.1", 3, "高齢者福祉施策（村単独事業）の実績分析",
  "配食・除雪・緊急通報・移送・生きがい活動等、村単独の高齢者福祉サービスの利用実績・事業費を分析",
  "高齢者福祉サービス実績分析表", "受託者", "事業実績提供", 1.0, D(2026,10,1), D(2026,10,31),
  "1.3.1", "3(1)② 高齢者福祉サービス実績", "第9期計画の掲載事業と対応づけ"),
 ("2.2.2", 3, "介護予防・日常生活支援総合事業の実績分析",
  "訪問型・通所型サービス、一般介護予防事業（通いの場・住民主体の活動）の実施状況と参加者数を分析",
  "総合事業実績分析表", "受託者", "事業実績提供", 1.0, D(2026,10,1), D(2026,10,31),
  "1.3.1", "3(1)② 高齢者福祉サービス実績", "介護予防の到達点を評価"),
 ("2.2.3", 3, "包括的支援事業・任意事業の実績分析",
  "地域包括支援センター運営、在宅医療・介護連携、認知症施策、生活支援体制整備、権利擁護の実績を分析",
  "地域支援事業実績分析表", "受託者", "事業実績提供", 0.5, D(2026,10,13), D(2026,10,31),
  "1.3.1", "3(1)② 高齢者福祉サービス実績", "地域包括ケアの推進状況を評価"),

 ("2.3",   2, "介護保険サービス実績の分析", "", "", "", "", None, None, None, "", "3(1)③ 介護保険サービス実績", ""),
 ("2.3.1", 3, "要介護認定者数・認定率の推移分析",
  "要支援1〜要介護5別・性別・年齢階級別の認定者数と認定率の推移、認定の重度化傾向を分析",
  "認定状況分析図表", "受託者", "認定データ提供", 1.5, D(2026,10,1), D(2026,10,31),
  "1.3.1", "3(1)③ 介護保険サービス実績", "3.2 認定者数推計の基礎データ"),
 ("2.3.2", 3, "サービス種類別給付実績の分析",
  "居宅・地域密着型・施設サービスの種類別に、受給者数・利用回数（日数）・給付費の推移を分析",
  "給付実績分析図表", "受託者", "給付実績提供", 1.5, D(2026,10,6), D(2026,11,13),
  "1.3.1", "3(1)③ 介護保険サービス実績", "4.1 見込量推計の基礎データ"),
 ("2.3.3", 3, "受給率・利用率・1人当たり給付費の分析",
  "認定者に対する受給率、要介護度別の利用状況、1人当たり給付費（月額）を算出し、傾向と特徴を整理",
  "利用状況分析表", "受託者", "－", 1.0, D(2026,10,20), D(2026,11,13),
  "2.3.2", "3(1)③ 介護保険サービス実績", "小規模ゆえの年変動に留意"),
 ("2.3.4", 3, "介護保険財政・保険料収納状況の分析",
  "介護保険特別会計の決算推移、第9期の給付費実績と計画対比、介護給付費準備基金残高、収納率を分析",
  "介護保険財政分析表", "受託者", "決算・基金資料提供", 1.0, D(2026,10,20), D(2026,11,20),
  "1.3.1", "3(1)③ 介護保険サービス実績", "4.3 保険料算定の前提となる"),

 ("2.4",   2, "現行計画（第9期）の点検・評価", "", "", "", "", None, None, None, "", "3(1)④ 現行計画の点検・評価", ""),
 ("2.4.1", 3, "施策・事業別の進捗評価",
  "第9期計画の施策体系に沿って、事業ごとの取組状況・目標値と実績の対比・達成度（3段階等）を評価",
  "第9期計画 進捗評価シート", "受託者", "各事業の実績・自己評価", 1.5, D(2026,11,2), D(2026,11,20),
  "2.2.3", "3(1)④ 現行計画の点検・評価", "村担当課への照会により評価を確定"),
 ("2.4.2", 3, "介護予防・日常生活圏域ニーズ調査結果の整理・分析",
  "今年度（令和8年度）村が実施した調査の集計データを整理し、リスク該当者割合・生活実態・意向を分析",
  "ニーズ調査結果分析資料", "受託者", "調査集計データ提供", 1.5, D(2026,10,13), D(2026,11,13),
  "1.3.1", "3(3) 見込量（ニーズ調査結果の反映）", "調査の実施・集計自体は業務範囲外（村実施分を反映）"),
 ("2.4.3", 3, "現状分析・評価の総括資料作成",
  "2.1〜2.4の分析結果を統合し、第2回打合せ用の中間報告資料として取りまとめ",
  "現状分析・評価 中間報告資料", "受託者", "内容確認", 1.0, D(2026,11,16), D(2026,11,27),
  "2.4.1, 2.4.2", "3(1) 現状の評価・分析", "★中間成果物／第2回打合せ資料"),

 # ---------------- 3. 人口推計・要支援要介護認定者数の推計 ----------------
 ("3",     1, "人口推計・要支援要介護認定者数の推計", "", "", "", "", None, None, None, "", "3(2) 人口推計等", ""),
 ("3.1",   2, "人口推計", "", "", "", "", None, None, None, "", "3(2) 人口推計等", ""),
 ("3.1.1", 3, "推計手法の設定・基準人口の確定",
  "コーホート要因法を基本とし、基準人口（住民基本台帳／国勢調査）、基準日、推計単位を村と確認",
  "推計条件設定書", "受託者", "基準人口の確認", 0.5, D(2026,10,13), D(2026,10,23),
  "2.1.1", "3(2) 人口推計等", "社人研の仮定値を基本に直近実績で補正"),
 ("3.1.2", 3, "将来人口推計（令和9年度から10年間程度）",
  "生残率・純移動率・出生仮定を設定し、男女5歳階級別に令和9年度〜令和18年度の人口を推計",
  "将来人口推計表・グラフ", "受託者", "－", 1.5, D(2026,10,26), D(2026,11,13),
  "3.1.1", "3(2) 令和9年から10年間程度の将来推計", "65-74歳／75歳以上・高齢化率を算出"),
 ("3.1.3", 3, "推計結果の検証・妥当性確認",
  "社人研推計・過去トレンドとの比較検証、村の開発・転入出動向を踏まえた妥当性の確認と補正",
  "推計検証資料", "受託者", "地域事情の情報提供", 0.5, D(2026,11,9), D(2026,11,20),
  "3.1.2", "3(2) 人口推計等", "国の基本指針が中長期（令和22年度等）を求める場合は対象年次を追加"),

 ("3.2",   2, "要支援・要介護認定者数の推計", "", "", "", "", None, None, None, "", "3(2) 認定者数の推計", ""),
 ("3.2.1", 3, "性・年齢階級別認定率の設定",
  "直近実績から第1号・第2号被保険者の性・年齢階級別認定率を算出し、将来の認定率仮定を設定",
  "認定率設定表", "受託者", "認定データの確認", 0.5, D(2026,11,9), D(2026,11,20),
  "2.3.1", "3(2) 認定者数の推計", "国の基本指針の考え方に従い設定"),
 ("3.2.2", 3, "各年度の要支援・要介護認定者数の推計",
  "将来人口×認定率により、要支援1〜要介護5別の各年度認定者数を推計（第10期3か年＋中長期）",
  "認定者数推計表・グラフ", "受託者", "－", 1.5, D(2026,11,16), D(2026,12,4),
  "3.1.2, 3.2.1", "3(2) 国の基本指針に従い各年度を推計", "4.1 見込量推計の前提"),
 ("3.2.3", 3, "推計結果の検証・村協議",
  "認定調査・審査会の運用状況、施設整備の予定等を踏まえて推計値を検証し、村担当課と確認",
  "推計検証・協議記録", "受託者", "実務状況の確認・確定", 0.5, D(2026,12,1), D(2026,12,9),
  "3.2.2", "3(2) 認定者数の推計", "確定値を4章以降で使用"),

 ("3.3",   2, "推計結果の資料化", "", "", "", "", None, None, None, "", "3(2) 人口推計等", ""),
 ("3.3.1", 3, "人口・認定者数推計結果の資料化",
  "推計手法・仮定条件・結果を計画掲載用の図表に整理し、根拠データを併せて保存",
  "人口・認定者数推計 結果資料", "受託者", "内容確認", 1.0, D(2026,12,7), D(2026,12,11),
  "3.1.3, 3.2.3", "3(2) 人口推計等", "★中間成果物／計画第2章に反映"),

 # ---------------- 4. 介護サービス見込量・保険料の推計 ----------------
 ("4",     1, "介護サービス見込量・保険料の推計", "", "", "", "", None, None, None, "", "3(3) 見込量・保険料", ""),
 ("4.1",   2, "介護サービス見込量の推計", "", "", "", "", None, None, None, "", "3(3) サービス見込量", ""),
 ("4.1.1", 3, "推計ツールの設定・実績データ投入",
  "地域包括ケア「見える化」システム等の推計ツールに、人口・認定者数・給付実績データを投入し初期設定",
  "推計ツール設定データ", "受託者", "システム利用環境の確認", 1.0, D(2026,12,1), D(2026,12,11),
  "3.2.3, 2.3.2", "3(3) 推計ツール等を用いて", "国の最新版ツールの公開時期に留意"),
 ("4.1.2", 3, "サービス種類別見込量の推計（令和9〜11年度）",
  "居宅・地域密着型・施設サービスの種類ごとに、各年度の受給者数・利用回数（日数）の見込量を推計",
  "サービス見込量推計表", "受託者", "－", 2.0, D(2026,12,7), D(2026,12,25),
  "4.1.1", "3(3) 令和11年度までの各年度・種類ごと", "第10期＝令和9〜11年度の3か年"),
 ("4.1.3", 3, "ニーズ調査結果・施策方針の反映",
  "ニーズ調査結果および村の施策方針（介護予防強化、在宅支援等）を見込量に反映し調整",
  "見込量調整記録", "受託者", "施策方針の提示", 1.0, D(2026,12,14), D(2027,1,8),
  "2.4.2, 4.1.2", "3(3) ニーズ調査結果の反映", "国の指示に適宜対応（仕様書3(3)）"),
 ("4.1.4", 3, "見込量の妥当性検証・村協議",
  "過去トレンド・北海道平均との比較、圏域外利用や事業所の供給力を踏まえた実現可能性を検証",
  "見込量検証資料", "受託者", "供給力の確認・確定", 0.5, D(2027,1,6), D(2027,1,15),
  "4.1.3", "3(3) サービス見込量", "★見込量の確定（マイルストーン）"),

 ("4.2",   2, "給付費・地域支援事業費の見込", "", "", "", "", None, None, None, "", "3(3) 見込量・保険料", ""),
 ("4.2.1", 3, "介護報酬改定等の国の情報の反映",
  "介護報酬改定率、区分支給限度額、負担割合、公費負担割合等の国の決定事項を確認し前提に反映",
  "算定前提条件整理メモ", "受託者", "国通知の共有", 0.5, D(2026,12,21), D(2027,1,15),
  "－", "5 その他（国の指示の遵守）", "改定率確定が遅れる場合は暫定値で算定→確定後に再算定"),
 ("4.2.2", 3, "標準給付費・地域支援事業費の算出",
  "見込量×単価により標準給付費を算出し、特定入所者介護サービス費・高額介護サービス費等を加算。地域支援事業費も算出",
  "給付費見込算出表", "受託者", "地域支援事業の方針提示", 1.0, D(2027,1,6), D(2027,1,22),
  "4.1.4, 4.2.1", "3(3) 見込量・保険料", "4.3 保険料算定の入力値"),

 ("4.3",   2, "第10期介護保険料の算定", "", "", "", "", None, None, None, "", "3(3) 第10期の介護保険料を算定", ""),
 ("4.3.1", 3, "保険料算定の前提条件整理",
  "第1号被保険者負担割合、調整交付金交付割合、予定収納率、所得段階区分・乗率、被保険者数見込を設定",
  "保険料算定前提条件表", "受託者", "所得段階別被保険者数の提供", 0.5, D(2026,12,14), D(2027,1,8),
  "2.3.4", "3(3) 保険料の算定", "村独自の多段階設定の要否を協議"),
 ("4.3.2", 3, "財政シミュレーション（複数パターン）",
  "介護給付費準備基金の取崩し額を変えた複数パターンで保険料水準を試算し、財政運営上の影響を整理",
  "財政シミュレーション表（複数案）", "受託者", "基金活用方針の検討", 1.5, D(2027,1,12), D(2027,1,29),
  "4.2.2, 4.3.1", "3(3) 保険料の算定", "3案程度を提示し村が方針決定"),
 ("4.3.3", 3, "保険料基準額・所得段階別保険料の算定",
  "保険料収納必要額から基準額（月額）を算出し、所得段階別の保険料額を設定。第9期との比較も整理",
  "第10期介護保険料算定書", "受託者", "方針決定", 1.0, D(2027,1,25), D(2027,2,5),
  "4.3.2", "3(3) 第10期の介護保険料を算定", "★保険料算定（マイルストーン）"),
 ("4.3.4", 3, "算定結果の説明資料作成・村協議",
  "算定過程・前提条件・保険料水準の根拠を説明資料に整理し、第3回打合せで報告・協議",
  "保険料説明資料", "受託者", "内容確認・庁内説明", 1.0, D(2027,2,1), D(2027,2,10),
  "4.3.3", "3(3) 保険料の算定", "議会・住民説明での活用を想定"),

 ("4.4",   2, "国の指示等に伴う再算定対応", "", "", "", "", None, None, None, "", "5 その他", ""),
 ("4.4.1", 3, "国の確定情報に基づく見込量・保険料の再算定",
  "基本指針の告示・報酬改定率の確定等を受け、必要な範囲で見込量・給付費・保険料を再算定し反映",
  "再算定結果・修正版算定書", "受託者", "協議・確認", 1.0, D(2027,2,1), D(2027,3,13),
  "4.3.3", "5 その他（国の指示により仕様変更）", "変更が必要な場合は村と協議して決定"),

 # ---------------- 5. 計画素案の作成 ----------------
 ("5",     1, "計画素案の作成", "", "", "", "", None, None, None, "", "3(4) 計画素案の作成", ""),
 ("5.1",   2, "課題及び施策の方向の整理", "", "", "", "", None, None, None, "", "3(4)① 課題及び施策の方向", ""),
 ("5.1.1", 3, "現状分析・推計結果からの課題抽出",
  "現状分析・第9期評価・ニーズ調査・推計結果を横断的に整理し、村の高齢者施策の課題を抽出・体系化",
  "課題整理シート", "受託者", "課題認識の共有", 1.0, D(2026,11,24), D(2026,12,4),
  "2.4.3", "3(4)① 課題及び施策の方向", "第2回打合せで協議"),
 ("5.1.2", 3, "基本理念・基本目標・施策の方向の整理",
  "課題に対応する基本理念、基本目標、施策の柱を整理し、施策体系図として提案",
  "基本理念・施策体系図（案）", "受託者", "方針決定", 1.0, D(2026,11,30), D(2026,12,7),
  "5.1.1", "3(4)① 課題及び施策の方向", "第9期からの継承・見直し点を明示"),
 ("5.1.3", 3, "総合計画等 関連計画との整合性確認",
  "村総合計画・地域福祉計画・健康増進計画等と理念・目標・施策の整合を確認し、記述を調整",
  "関連計画整合性チェック表", "受託者", "関連計画の方針確認", 0.5, D(2026,11,30), D(2026,12,9),
  "1.3.3, 5.1.2", "7 留意事項（総合計画等との整合）", "整合しない箇所は協議のうえ調整"),

 ("5.2",   2, "計画骨子の提案", "", "", "", "", None, None, None, "", "3(4)② 計画骨子の提案", ""),
 ("5.2.1", 3, "章立て（目次構成）案の作成",
  "国の策定指針の記載事項を踏まえた章立て・節構成、ページ配分、掲載図表の案を作成",
  "計画骨子（目次構成）案", "受託者", "内容確認", 0.5, D(2026,12,2), D(2026,12,9),
  "5.1.2", "3(4)② 計画骨子の提案", "★中間成果物"),
 ("5.2.2", 3, "国の策定指針 記載事項チェックリスト作成",
  "基本指針で定める記載事項（任意記載事項を含む）を一覧化し、計画本文との対応を管理",
  "記載事項チェックリスト", "受託者", "－", 0.5, D(2026,11,16), D(2026,11,27),
  "－", "3(4)※ 国の策定指針に従い記載", "記載漏れ防止のため全編で運用"),
 ("5.2.3", 3, "骨子案の提示・確定（第2回打合せ）",
  "骨子案を提示し、構成・記載レベル・分量を協議のうえ確定",
  "確定版 計画骨子", "受託者", "出席・決定", 0.5, D(2026,12,10), D(2026,12,10),
  "5.2.1, 5.2.2", "3(4)② 計画骨子の提案", "★マイルストーン"),

 ("5.3",   2, "計画素案の執筆", "", "", "", "", None, None, None, "", "3(4)③ 計画素案の作成", ""),
 ("5.3.1", 3, "第1章 計画の策定にあたって",
  "計画策定の趣旨、法的位置づけ、計画期間（令和9〜11年度）、関連計画との関係、策定体制・経過を執筆",
  "第1章 原稿", "受託者", "策定体制情報の提供", 0.5, D(2026,12,14), D(2026,12,25),
  "5.2.3", "3(4)③ 計画素案の作成", ""),
 ("5.3.2", 3, "第2章 高齢者を取り巻く現状",
  "人口・世帯、高齢化の状況、認定者数、給付実績、ニーズ調査結果、将来推計を図表とともに執筆",
  "第2章 原稿", "受託者", "内容確認", 2.0, D(2026,12,14), D(2027,1,15),
  "2.4.3, 3.3.1", "3(4)③ 計画素案の作成", "現状分析・推計結果を反映"),
 ("5.3.3", 3, "第3章 第9期計画の評価と課題",
  "第9期計画の進捗評価結果、達成状況、残された課題を整理して執筆",
  "第3章 原稿", "受託者", "評価結果の確認", 1.0, D(2026,12,14), D(2027,1,8),
  "2.4.1, 5.1.1", "3(1)④／3(4)③", ""),
 ("5.3.4", 3, "第4章 計画の基本的な考え方",
  "基本理念、基本目標、施策体系図、日常生活圏域の設定、地域包括ケアシステムの考え方を執筆",
  "第4章 原稿", "受託者", "方針確認", 1.0, D(2026,12,21), D(2027,1,15),
  "5.1.2", "3(4)③ 計画素案の作成", ""),
 ("5.3.5", 3, "第5章 施策の展開",
  "介護予防・健康づくり、地域包括ケア、在宅医療・介護連携、認知症施策、生活支援体制整備、権利擁護、"
  "住まい、災害・感染症対策、介護人材確保等の施策を執筆",
  "第5章 原稿", "受託者", "各事業の方針・実施内容確認", 3.0, D(2027,1,5), D(2027,2,5),
  "5.1.2, 5.2.3", "3(4)③ 計画素案の作成", "村単独事業と地域支援事業を体系的に記載"),
 ("5.3.6", 3, "第6章 介護保険事業の見込みと保険料",
  "サービス見込量、標準給付費、地域支援事業費、第10期介護保険料（所得段階別）を執筆",
  "第6章 原稿", "受託者", "保険料方針の決定", 1.5, D(2027,1,25), D(2027,2,9),
  "4.1.4, 4.3.3", "3(3)／3(4)③", "保険料は村の方針決定後に確定"),
 ("5.3.7", 3, "第7章 計画の推進体制",
  "推進体制、関係機関との連携、PDCAによる進行管理、評価指標・目標値の設定方法を執筆",
  "第7章 原稿", "受託者", "推進体制の確認", 0.5, D(2027,1,19), D(2027,1,29),
  "5.2.3", "3(4)③ 計画素案の作成", "第10期の進捗管理方法を明記"),
 ("5.3.8", 3, "資料編の作成",
  "策定経過、ニーズ調査結果概要、用語解説、関係法令抜粋、統計資料等を資料編として編集",
  "資料編 原稿", "受託者", "策定経過情報の提供", 1.0, D(2027,1,19), D(2027,2,5),
  "5.3.2", "3(4)③ 計画素案の作成", ""),
 ("5.3.9", 3, "図表作成・レイアウト・校正",
  "図表・グラフの統一デザイン適用、体裁調整、数値・図表番号・用語の整合確認、通読校正",
  "レイアウト済 素案データ", "受託者", "－", 2.0, D(2027,1,26), D(2027,2,9),
  "5.3.5, 5.3.6", "3(4)③ 計画素案の作成", "既存のコラム部品ライブラリを活用可"),
 ("5.3.10", 3, "計画素案の提出・協議（第3回打合せ）",
  "計画素案一式を提出し、内容・表現・分量を協議。修正方針を確認",
  "計画素案（第1版）", "受託者", "出席・査読", 0.5, D(2027,2,10), D(2027,2,10),
  "5.3.9", "3(4)③ 計画素案の作成", "★マイルストーン／中間成果物"),

 ("5.4",   2, "素案の修正・反映", "", "", "", "", None, None, None, "", "3(4)③／7 留意事項", ""),
 ("5.4.1", 3, "村意見の反映・修正",
  "第3回打合せおよび庁内査読で示された意見を反映し、素案を修正",
  "計画素案（修正版）", "受託者", "意見の取りまとめ", 1.5, D(2027,2,12), D(2027,3,5),
  "5.3.10", "3(4)③ 計画素案の作成", ""),
 ("5.4.2", 3, "パブリックコメント等の意見反映",
  "村が実施するパブリックコメント・住民説明等で寄せられた意見への対応案を作成し、計画に反映",
  "意見対応表／反映版原稿", "受託者", "パブコメの実施・意見提供", 0.5, D(2027,2,12), D(2027,3,13),
  "5.4.1", "7 留意事項（協議）", "パブコメの実施主体・日程は村と協議（仕様書に明記なし）"),

 # ---------------- 6. 成果品の作成・納品 ----------------
 ("6",     1, "成果品の作成・納品", "", "", "", "", None, None, None, "", "6 成果品", ""),
 ("6.1",   2, "成果品の作成", "", "", "", "", None, None, None, "", "6 成果品", ""),
 ("6.1.1", 3, "最終原稿の確定・最終校正",
  "全編の通し校正（誤字脱字、数値・図表の整合、出典表記、用語統一）を行い最終原稿を確定",
  "最終原稿", "受託者", "最終確認", 1.0, D(2027,3,8), D(2027,3,19),
  "5.4.2", "6 成果品", ""),
 ("6.1.2", 3, "成果品データの作成（CD-R収録）",
  "計画本編・資料編の編集可能データ（Word等）およびPDFを作成し、CD-Rに収録・ラベル作成",
  "計画（CDデータ）1部", "受託者", "検収", 0.5, D(2027,3,17), D(2027,3,26),
  "6.1.1", "6 成果品（CDデータ 1部）", "著作権は音威子府村に帰属"),
 ("6.1.3", 3, "業務データの整理・収録",
  "推計データ、集計表、図表の元データ、収集統計、算定シート等を整理し磁気媒体（CD-R）に収録",
  "業務データ収録媒体（CDデータ）1部", "受託者", "検収", 0.5, D(2027,3,17), D(2027,3,26),
  "6.1.1", "6 成果品（業務データ CDデータ 1部）", "村での更新作業を想定した構成とする"),

 ("6.2",   2, "納品・完了手続", "", "", "", "", None, None, None, "", "2 委託期間／6 成果品", ""),
 ("6.2.1", 3, "納品・検査対応",
  "成果品の納品、業務完了届の提出、検査対応、指摘事項の修正",
  "業務完了届／検査調書", "受託者", "検査・完了確認", 0.5, D(2027,3,29), D(2027,3,31),
  "6.1.2, 6.1.3", "2 委託期間（令和9年3月31日）", "★マイルストーン／履行期限厳守"),
 ("6.2.2", 3, "貸与資料・個人情報の返却／消去",
  "貸与資料の返却、業務上取得した個人情報を含むデータの消去・廃棄と証明書の提出",
  "資料返却書／データ消去証明", "受託者", "受領確認", 0.3, D(2027,3,29), D(2027,3,31),
  "6.2.1", "7 留意事項（個人情報）", "音威子府村個人情報保護条例を遵守"),
]

# 大分類名（WBS No の先頭桁 → 名称・カラー）
CAT_NAMES = {
    "1": "1. 業務管理・打合せ",
    "2": "2. 現状の評価・分析",
    "3": "3. 人口推計・認定者数推計",
    "4": "4. サービス見込量・保険料の推計",
    "5": "5. 計画素案の作成",
    "6": "6. 成果品の作成・納品",
}
CAT_COLORS = {k: COLORS["C" + k] for k in CAT_NAMES}


def cat_of(wbs):
    return wbs.split(".")[0]


def build_row_model():
    """TASKS を描画用モデルに変換し、集計行の子ブロック範囲を確定する"""
    rows = []
    for t in TASKS:
        wbs, lv, name, work, deliv, owner, village, eff, st, en, pred, spec, note = t
        rows.append(dict(wbs=wbs, lv=lv, name=name, work=work, deliv=deliv, owner=owner,
                         village=village, eff=eff, st=st, en=en, pred=pred, spec=spec,
                         note=note, cat=cat_of(wbs)))
    # 親名の解決
    by_wbs = {r["wbs"]: r for r in rows}
    for r in rows:
        parts = r["wbs"].split(".")
        r["l1name"] = by_wbs[parts[0]]["name"]
        r["l2name"] = by_wbs[".".join(parts[:2])]["name"] if len(parts) >= 2 else ""
        r["l3name"] = r["name"] if r["lv"] == 3 else ""
    # 集計行の子ブロック（インデックス範囲）と日程
    for i, r in enumerate(rows):
        if r["lv"] == 3:
            r["block"] = None
            continue
        j = i + 1
        while j < len(rows) and rows[j]["lv"] > r["lv"]:
            j += 1
        r["block"] = (i + 1, j - 1) if j - 1 >= i + 1 else None
        kids = [rows[k] for k in range(i + 1, j) if rows[k]["lv"] == 3]
        if kids:
            r["st"] = min(k["st"] for k in kids)
            r["en"] = max(k["en"] for k in kids)
    return rows


ROWS = build_row_model()
WBS_HEADER_ROW = 4           # 01_WBS のヘッダ行
WBS_FIRST_DATA_ROW = 5       # 01_WBS の明細開始行
WBS_LAST_DATA_ROW = WBS_FIRST_DATA_ROW + len(ROWS) - 1
WBS_TOTAL_ROW = WBS_LAST_DATA_ROW + 1


# ============================================================
# 00_業務概要
# ============================================================
def add_overview_sheet(wb):
    ws = wb.active
    ws.title = "00_業務概要"
    set_col_widths(ws, [22, 46, 30, 26, 22])
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:E1")
    style_title(ws["A1"], f"{PLAN_NAME}　策定業務委託　WBS（作業分解構成図）")

    ws.merge_cells("A2:E2")
    ws["A2"] = "本ブックは委託仕様書に基づき作成した業務管理用WBSです。契約締結日・打合せ日程等が確定した時点で日程を更新してください。"
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    r = 4
    ws.merge_cells(f"A{r}:E{r}")
    style_subhead(ws.cell(row=r, column=1), "1. 業務の基本情報")
    r += 1
    info = [
        ("業務名", f"{PLAN_NAME}　策定業務委託"),
        ("発注者", CLIENT),
        ("業務目的", "計画原案の作成等の支援（計画策定に必要となるサービス見込量等の数値の推計を含む）"),
        ("委託期間", "契約締結日 〜 令和9年3月31日"),
        ("計画期間", "第10期＝令和9年度〜令和11年度（3か年）"),
        ("人口推計期間", "令和9年から10年間程度（令和9年度〜令和18年度）"),
        ("打合せ回数", "3回程度（計画策定担当者本人が出席）"),
        ("成果品", "計画（CDデータ）1部／業務データ収録の磁気媒体（CDデータ）1部"),
        ("著作権", "成果品の著作権はすべて音威子府村に帰属"),
        ("WBS作成日", "=TODAY()"),
    ]
    for label, value in info:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name=FONT, size=10, bold=True)
        c1.fill = PatternFill("solid", fgColor=COLORS["band"])
        c1.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        c1.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c2 = ws.cell(row=r, column=2, value=value)
        style_data_cell(c2)
        if label == "WBS作成日":
            c2.number_format = "yyyy/mm/dd"
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    style_subhead(ws.cell(row=r, column=1), "2. 業務量サマリ（『01_WBS』シートと連動）")
    r += 1
    style_header_row(ws, r, ["区分", "値", "算出方法", "備考", ""])
    r += 1
    q = "'01_WBS'!"
    summary = [
        ("作業項目数（レベル3）",
         f"=COUNTIF({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3)",
         "レベル3の行数", "実作業単位の総数"),
        ("総工数（人日）",
         f"=SUMIF({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3,{q}$K${WBS_FIRST_DATA_ROW}:$K${WBS_LAST_DATA_ROW})",
         "レベル3の工数合計", "1人日＝7.5時間想定"),
        ("総工数（人月換算）",
         f"=ROUND(SUMIF({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3,{q}$K${WBS_FIRST_DATA_ROW}:$K${WBS_LAST_DATA_ROW})/20,1)",
         "総工数 ÷ 20人日", "1人月＝20人日換算"),
        ("完了件数",
         f"=COUNTIFS({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3,{q}$P${WBS_FIRST_DATA_ROW}:$P${WBS_LAST_DATA_ROW},\"完了\")",
         "進捗欄が「完了」の件数", "進捗欄の入力に応じて自動更新"),
        ("進捗率",
         f"=IFERROR(COUNTIFS({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3,{q}$P${WBS_FIRST_DATA_ROW}:$P${WBS_LAST_DATA_ROW},\"完了\")/COUNTIF({q}$C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3),0)",
         "完了件数 ÷ 作業項目数", "0.0%表示"),
    ]
    for name, formula, how, note in summary:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=formula)
        ws.cell(row=r, column=3, value=how)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=4, value=note)
        for col in range(1, 6):
            style_data_cell(ws.cell(row=r, column=col), alt=(r % 2 == 0))
        vc = ws.cell(row=r, column=2)
        vc.alignment = Alignment(vertical="center", horizontal="center")
        vc.font = Font(name=FONT, size=10, bold=True)
        vc.number_format = "0.0%" if name == "進捗率" else "#,##0.0"
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    style_subhead(ws.cell(row=r, column=1), "3. 前提条件（確定次第、各シートの日程・数値を更新）")
    r += 1
    assumptions = [
        ("契約締結日", "令和8年9月1日と仮定して日程を設定（未確定）", "確定後にスクリプトの CONTRACT_START を修正"),
        ("打合せ日程", "第1回=R8.9中旬／第2回=R8.12上旬／第3回=R9.2中旬 と仮置き", "村と協議のうえ確定"),
        ("ニーズ調査", "令和8年度に村が実施済みの集計データを受領して反映（調査実施は業務範囲外）", "仕様書3(3)による"),
        ("国の基本指針", "第10期の基本指針・介護報酬改定率は策定作業と並行して確定する前提", "確定遅延時は暫定値で算定→再算定（WBS 4.4.1）"),
        ("成果品", "CDデータのみ（印刷・製本は仕様書に記載なし）", "必要な場合は別途協議"),
        ("パブリックコメント", "実施主体・日程は仕様書に記載なし。村実施を想定し意見反映のみ計上", "実施支援が必要な場合は協議"),
        ("再委託", "予定なし。必要が生じた場合は事前に村の承諾を得る", "仕様書7 留意事項"),
        ("工数単位", "1人日＝7.5時間、1人月＝20人日で換算", "見積・要員計画の基礎"),
    ]
    style_header_row(ws, r, ["項目", "前提の置き方", "備考", "", ""])
    r += 1
    for a, b, c in assumptions:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(row=r, column=3, value=c)
        for col in range(1, 6):
            style_data_cell(ws.cell(row=r, column=col), alt=(r % 2 == 0))
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    style_subhead(ws.cell(row=r, column=1), "4. シート構成")
    r += 1
    sheets = [
        ("01_WBS", "作業分解構成図の本体（レベル1〜3・工数・日程・担当・仕様書対応）"),
        ("02_スケジュール", "半月単位のガントチャート（令和8年9月〜令和9年3月）"),
        ("03_マイルストーン", "打合せ・提出物等の節目一覧"),
        ("04_成果物一覧", "中間成果物と最終成果品の一覧"),
        ("05_推計作業手順", "人口・認定者数・見込量・保険料の推計手順とデータ・ツール"),
        ("06_必要資料リスト", "村から貸与を受ける資料・公表統計の一覧（依頼票）"),
        ("07_役割分担", "受託者／村担当課／関係機関の役割分担（RACI）"),
        ("08_リスク管理", "想定リスクと対応方針"),
        ("09_スコープ管理", "業務範囲に含む／含まない／協議事項の整理"),
        ("10_工数集計", "大分類・中分類別の工数集計"),
    ]
    style_header_row(ws, r, ["シート名", "内容", "", "", ""])
    r += 1
    for n, d in sheets:
        ws.cell(row=r, column=1, value=n)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2, value=d)
        for col in range(1, 6):
            style_data_cell(ws.cell(row=r, column=col), alt=(r % 2 == 0))
        r += 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    return ws


# ============================================================
# 01_WBS
# ============================================================
WBS_HEADERS = ["No.", "WBS No.", "レベル", "大分類", "中分類", "作業項目",
               "作業内容・実施方法", "成果物・アウトプット", "主担当",
               "村（発注者）の対応", "工数\n(人日)", "開始予定", "完了予定",
               "先行作業", "仕様書 該当箇所", "進捗", "備考・留意点"]
WBS_WIDTHS = [5, 9, 6, 24, 26, 30, 54, 26, 9, 22, 8, 12, 12, 14, 26, 9, 34]


def add_wbs_sheet(wb):
    ws = wb.create_sheet("01_WBS")
    set_col_widths(ws, WBS_WIDTHS)
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:Q1")
    style_title(ws["A1"], f"WBS（作業分解構成図）　／　{PLAN_NAME}　策定業務委託")

    ws.merge_cells("A2:Q2")
    ws["A2"] = ("委託期間：契約締結日〜令和9年3月31日　／　打合せ：3回程度　／　"
                "★印はマイルストーン　／　レベル1・2行の工数・日程は配下のレベル3から自動集計")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("A3:Q3")
    ws["A3"] = ("凡例：レベル1＝大分類（濃紺）／レベル2＝中分類（水色）／レベル3＝実作業　　"
                "進捗欄は「未着手・着手・完了・保留・対象外」から選択")
    ws["A3"].font = Font(name=FONT, size=9, color="595959")
    ws["A3"].alignment = Alignment(vertical="center", indent=1)

    ws.row_dimensions[WBS_HEADER_ROW].height = 34
    style_header_row(ws, WBS_HEADER_ROW, WBS_HEADERS, fill=COLORS["header"])

    seq = 0
    for i, r in enumerate(ROWS):
        row = WBS_FIRST_DATA_ROW + i
        lv = r["lv"]
        if lv == 3:
            seq += 1
            no = seq
        else:
            no = ""

        # 工数：レベル3は実数、集計行はSUMIF
        if lv == 3:
            eff = r["eff"]
        elif r["block"]:
            a = WBS_FIRST_DATA_ROW + r["block"][0]
            b = WBS_FIRST_DATA_ROW + r["block"][1]
            eff = f"=SUMIF($C${a}:$C${b},3,$K${a}:$K${b})"
        else:
            eff = None

        values = [no, r["wbs"], lv, r["l1name"], r["l2name"], r["l3name"],
                  r["work"], r["deliv"], r["owner"], r["village"], eff,
                  r["st"], r["en"], r["pred"], r["spec"],
                  "未着手" if lv == 3 else "", r["note"]]

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            style_data_cell(cell, alt=(lv == 3 and row % 2 == 0),
                            center=(col in (1, 2, 3, 9, 11, 12, 13, 16)))
            if col in (12, 13):
                cell.number_format = "yyyy/mm/dd"
            if col == 11:
                cell.number_format = "#,##0.0"

        # 階層別の書式
        if lv == 1:
            for col in range(1, len(WBS_HEADERS) + 1):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=CAT_COLORS[r["cat"]])
                c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
            ws.row_dimensions[row].height = 24
        elif lv == 2:
            for col in range(1, len(WBS_HEADERS) + 1):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=COLORS["L2"])
                c.font = Font(name=FONT, size=10, bold=True, color="1F3864")
            ws.row_dimensions[row].height = 20
        else:
            ws.row_dimensions[row].height = 46
            # マイルストーン行を強調
            if r["note"].startswith("★") or "★" in r["note"]:
                mc = ws.cell(row=row, column=6)
                mc.font = Font(name=FONT, size=10, bold=True, color=COLORS["ms"])

    # 合計行
    ws.merge_cells(start_row=WBS_TOTAL_ROW, start_column=1, end_row=WBS_TOTAL_ROW, end_column=10)
    tc = ws.cell(row=WBS_TOTAL_ROW, column=1, value="合計（レベル3の作業項目）")
    tc.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    tc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    total = ws.cell(row=WBS_TOTAL_ROW, column=11,
                    value=f"=SUMIF($C${WBS_FIRST_DATA_ROW}:$C${WBS_LAST_DATA_ROW},3,"
                          f"$K${WBS_FIRST_DATA_ROW}:$K${WBS_LAST_DATA_ROW})")
    total.number_format = "#,##0.0"
    total.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    total.alignment = Alignment(vertical="center", horizontal="center")
    ws.merge_cells(start_row=WBS_TOTAL_ROW, start_column=12, end_row=WBS_TOTAL_ROW, end_column=17)
    nc = ws.cell(row=WBS_TOTAL_ROW, column=12, value="人日（1人日＝7.5時間／1人月＝20人日換算）")
    nc.font = Font(name=FONT, size=10, color="FFFFFF")
    nc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col in range(1, len(WBS_HEADERS) + 1):
        c = ws.cell(row=WBS_TOTAL_ROW, column=col)
        c.fill = PatternFill("solid", fgColor=COLORS["header"])
        c.border = BORDER
    ws.row_dimensions[WBS_TOTAL_ROW].height = 24

    # 進捗欄の入力規則
    dv = DataValidation(type="list", formula1='"未着手,着手,完了,保留,対象外"', allow_blank=True)
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力値が不正です"
    ws.add_data_validation(dv)
    dv.add(f"P{WBS_FIRST_DATA_ROW}:P{WBS_LAST_DATA_ROW}")

    ws.freeze_panes = f"D{WBS_FIRST_DATA_ROW}"
    ws.auto_filter.ref = f"A{WBS_HEADER_ROW}:Q{WBS_LAST_DATA_ROW}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = f"{WBS_HEADER_ROW}:{WBS_HEADER_ROW}"
    return ws


# ============================================================
# 02_スケジュール（ガントチャート）
# ============================================================
PERIODS = [
    ("令和8年度", "9月", "前半", D(2026, 9, 1),  D(2026, 9, 15)),
    ("令和8年度", "9月", "後半", D(2026, 9, 16), D(2026, 9, 30)),
    ("令和8年度", "10月", "前半", D(2026, 10, 1),  D(2026, 10, 15)),
    ("令和8年度", "10月", "後半", D(2026, 10, 16), D(2026, 10, 31)),
    ("令和8年度", "11月", "前半", D(2026, 11, 1),  D(2026, 11, 15)),
    ("令和8年度", "11月", "後半", D(2026, 11, 16), D(2026, 11, 30)),
    ("令和8年度", "12月", "前半", D(2026, 12, 1),  D(2026, 12, 15)),
    ("令和8年度", "12月", "後半", D(2026, 12, 16), D(2026, 12, 31)),
    ("令和8年度", "1月", "前半", D(2027, 1, 1),  D(2027, 1, 15)),
    ("令和8年度", "1月", "後半", D(2027, 1, 16), D(2027, 1, 31)),
    ("令和8年度", "2月", "前半", D(2027, 2, 1),  D(2027, 2, 15)),
    ("令和8年度", "2月", "後半", D(2027, 2, 16), D(2027, 2, 28)),
    ("令和8年度", "3月", "前半", D(2027, 3, 1),  D(2027, 3, 15)),
    ("令和8年度", "3月", "後半", D(2027, 3, 16), D(2027, 3, 31)),
]
GANTT_FIRST_COL = 7   # G列から期間欄


def add_schedule_sheet(wb):
    ws = wb.create_sheet("02_スケジュール")
    widths = [9, 42, 9, 8, 12, 12] + [6] * len(PERIODS)
    set_col_widths(ws, widths)
    last_col = GANTT_FIRST_COL + len(PERIODS) - 1
    last_letter = get_column_letter(last_col)

    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{last_letter}1")
    style_title(ws["A1"], f"実施スケジュール（ガントチャート）　／　{PLAN_NAME}　策定業務委託")

    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"] = ("期間：契約締結日（令和8年9月1日と仮定）〜令和9年3月31日　半月単位　／　"
                "■＝作業期間　◆＝打合せ・提出等のマイルストーン")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    # ヘッダ（3行目＝月、4行目＝前半／後半）
    head1, head2 = 3, 4
    for c, label in enumerate(["WBS No.", "作業項目", "主担当", "工数\n(人日)", "開始予定", "完了予定"], 1):
        ws.merge_cells(start_row=head1, start_column=c, end_row=head2, end_column=c)
        cell = ws.cell(row=head1, column=c, value=label)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER
        ws.cell(row=head2, column=c).border = BORDER

    col = GANTT_FIRST_COL
    i = 0
    while i < len(PERIODS):
        month = PERIODS[i][1]
        span = sum(1 for p in PERIODS if p[1] == month)
        ws.merge_cells(start_row=head1, start_column=col, end_row=head1, end_column=col + span - 1)
        mc = ws.cell(row=head1, column=col, value=month)
        mc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        mc.fill = PatternFill("solid", fgColor=COLORS["header"])
        mc.alignment = Alignment(vertical="center", horizontal="center")
        for k in range(span):
            ws.cell(row=head1, column=col + k).border = BORDER
            hc = ws.cell(row=head2, column=col + k, value=PERIODS[i + k][2])
            hc.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor=COLORS["subhead"])
            hc.alignment = Alignment(vertical="center", horizontal="center")
            hc.border = BORDER
        col += span
        i += span
    ws.row_dimensions[head1].height = 20
    ws.row_dimensions[head2].height = 18

    row = head2 + 1
    for r in ROWS:
        if r["lv"] == 1:
            label = r["name"]
        elif r["lv"] == 2:
            label = "　" + r["name"]
        else:
            label = "　　" + r["name"]

        ws.cell(row=row, column=1, value=r["wbs"])
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=r["owner"])
        ws.cell(row=row, column=4, value=r["eff"] if r["lv"] == 3 else None)
        ws.cell(row=row, column=5, value=r["st"])
        ws.cell(row=row, column=6, value=r["en"])
        for c in range(1, 7):
            style_data_cell(ws.cell(row=row, column=c),
                            alt=(r["lv"] == 3 and row % 2 == 0),
                            center=(c in (1, 3, 4, 5, 6)))
        ws.cell(row=row, column=4).number_format = "#,##0.0"
        ws.cell(row=row, column=5).number_format = "m/d"
        ws.cell(row=row, column=6).number_format = "m/d"

        is_ms = (r["lv"] == 3 and r["st"] == r["en"])
        bar_color = CAT_COLORS[r["cat"]]
        for k, (_, _, _, ps, pe) in enumerate(PERIODS):
            cell = ws.cell(row=row, column=GANTT_FIRST_COL + k)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", horizontal="center")
            if r["st"] and r["en"] and r["st"] <= pe and r["en"] >= ps:
                if is_ms:
                    cell.value = "◆"
                    cell.font = Font(name=FONT, size=10, bold=True, color=COLORS["ms"])
                else:
                    cell.fill = PatternFill("solid", fgColor=bar_color)
            elif r["lv"] == 3 and row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt"])

        if r["lv"] == 1:
            for c in range(1, 7):
                cc = ws.cell(row=row, column=c)
                cc.fill = PatternFill("solid", fgColor=CAT_COLORS[r["cat"]])
                cc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            ws.row_dimensions[row].height = 22
        elif r["lv"] == 2:
            for c in range(1, 7):
                cc = ws.cell(row=row, column=c)
                cc.fill = PatternFill("solid", fgColor=COLORS["L2"])
                cc.font = Font(name=FONT, size=10, bold=True, color="1F3864")
            ws.row_dimensions[row].height = 18
        else:
            ws.row_dimensions[row].height = 17
        row += 1

    # 凡例
    row += 1
    ws.cell(row=row, column=1, value="凡例")
    ws.cell(row=row, column=1).font = Font(name=FONT, size=10, bold=True)
    for k, (key, name) in enumerate(CAT_NAMES.items()):
        cell = ws.cell(row=row + 1 + k, column=2, value=name)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = Alignment(vertical="center", indent=1)
        sw = ws.cell(row=row + 1 + k, column=1)
        sw.fill = PatternFill("solid", fgColor=CAT_COLORS[key])
        sw.border = BORDER
    ms_row = row + 1 + len(CAT_NAMES)
    ws.cell(row=ms_row, column=1, value="◆").font = Font(name=FONT, size=10, bold=True, color=COLORS["ms"])
    ws.cell(row=ms_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=ms_row, column=2, value="マイルストーン（打合せ・提出・納品）").font = Font(name=FONT, size=9)
    ws.cell(row=ms_row, column=2).alignment = Alignment(vertical="center", indent=1)

    ws.freeze_panes = ws.cell(row=head2 + 1, column=GANTT_FIRST_COL).coordinate
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 03_マイルストーン
# ============================================================
MILESTONES = [
    ("M-01", "契約締結・業務着手", "令和8年9月上旬", "1.1.1", "受託者・村",
     "契約書・仕様書の確認、業務従事者の確定", "－"),
    ("M-02", "【第1回打合せ】キックオフ", "令和8年9月中旬", "1.1.3", "受託者・村",
     "業務実施計画・工程の協議、論点共有、貸与資料の確認", "第1回打合せ資料／議事録"),
    ("M-03", "貸与資料一式の受領完了", "令和8年9月末", "1.3.1", "村",
     "分析着手に必要な実績・統計データが揃うこと", "貸与資料リスト（受領済）"),
    ("M-04", "現状分析・第9期評価の完了", "令和8年11月下旬", "2.4.3", "受託者",
     "現状分析・実績分析・現行計画評価の取りまとめ", "現状分析・評価 中間報告資料"),
    ("M-05", "【第2回打合せ】現状分析・骨子報告", "令和8年12月上旬", "1.2.2／5.2.3", "受託者・村",
     "課題と施策の方向の協議、計画骨子の確定", "第2回打合せ資料／確定版計画骨子"),
    ("M-06", "人口・要支援要介護認定者数推計の確定", "令和8年12月中旬", "3.3.1", "受託者・村",
     "推計手法・仮定条件・結果の村確認", "人口・認定者数推計 結果資料"),
    ("M-07", "介護サービス見込量の確定", "令和9年1月中旬", "4.1.4", "受託者・村",
     "令和9〜11年度のサービス種類別見込量の確定", "サービス見込量推計表"),
    ("M-08", "第10期介護保険料（第1次案）の算定", "令和9年2月上旬", "4.3.3", "受託者・村",
     "基金取崩し方針の決定、保険料基準額の算定", "第10期介護保険料算定書"),
    ("M-09", "【第3回打合せ】計画素案・保険料報告", "令和9年2月中旬", "1.2.3", "受託者・村",
     "計画素案および保険料算定結果の報告・協議", "第3回打合せ資料／議事録"),
    ("M-10", "計画素案の提出", "令和9年2月中旬", "5.3.10", "受託者",
     "国の策定指針の記載事項を満たした素案の提出", "計画素案（第1版）"),
    ("M-11", "住民意見等の反映完了", "令和9年3月中旬", "5.4.2", "受託者・村",
     "パブリックコメント等の意見への対応（村実施分）", "意見対応表／反映版原稿"),
    ("M-12", "成果品の納品・検査", "令和9年3月31日", "6.2.1", "受託者・村",
     "履行期限。計画（CD）1部・業務データ（CD）1部の納品", "計画CD／業務データCD／業務完了届"),
]


def add_milestone_sheet(wb):
    ws = wb.create_sheet("03_マイルストーン")
    set_col_widths(ws, [9, 34, 18, 14, 16, 46, 34])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "マイルストーン・会議体一覧")
    ws.merge_cells("A2:G2")
    ws["A2"] = "打合せは仕様書「4 打合せ」に基づき3回程度。受託者側は計画策定担当者本人が出席する。日程は村と協議のうえ確定。"
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["No.", "マイルストーン", "予定時期", "WBS No.", "関係者",
                             "達成基準・協議内容", "提出物・成果物"], fill=COLORS["header"])
    ws.row_dimensions[4].height = 26
    r = 5
    for m in MILESTONES:
        for c, v in enumerate(m, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (1, 3, 4, 5)))
        if "打合せ" in m[1]:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=COLORS["note"])
            ws.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True, color=COLORS["ms"])
        ws.row_dimensions[r].height = 32
        r += 1

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 04_成果物一覧
# ============================================================
DELIVERABLES = [
    ("最終成果品", "高齢者福祉計画・第10期介護保険事業計画", "CD-R（Word/Excel等の編集可能データ＋PDF）", "1部",
     "令和9年3月31日", "6.1.2", "仕様書「6 成果品」／著作権は音威子府村に帰属"),
    ("最終成果品", "業務データ収録の磁気媒体", "CD-R（推計データ・集計表・図表元データ等）", "1部",
     "令和9年3月31日", "6.1.3", "仕様書「6 成果品」／村での更新作業を想定した構成"),
    ("中間成果物", "業務実施計画書・詳細工程表（本WBS）", "電子データ", "1式",
     "令和8年9月中旬", "1.1.2", "第1回打合せで協議・確定"),
    ("中間成果物", "貸与資料リスト（依頼票）", "電子データ", "1式",
     "令和8年9月上旬", "1.3.1", "契約後2週間以内に提示"),
    ("中間成果物", "現状分析・評価 中間報告資料", "電子データ（PowerPoint/PDF等）", "1式",
     "令和8年11月下旬", "2.4.3", "第2回打合せ資料"),
    ("中間成果物", "人口・要支援要介護認定者数 推計結果資料", "電子データ（Excel＋図表）", "1式",
     "令和8年12月中旬", "3.3.1", "推計条件・仮定値を明記"),
    ("中間成果物", "介護サービス見込量 推計表", "電子データ（Excel）", "1式",
     "令和9年1月中旬", "4.1.2", "サービス種類別・年度別（令和9〜11年度）"),
    ("中間成果物", "第10期介護保険料 算定書・財政シミュレーション", "電子データ（Excel）", "1式",
     "令和9年2月上旬", "4.3.2／4.3.3", "複数パターンの比較を含む"),
    ("中間成果物", "計画骨子（目次構成）案", "電子データ", "1式",
     "令和8年12月上旬", "5.2.1", "国の策定指針の記載事項チェックリストを添付"),
    ("中間成果物", "計画素案（第1版）", "電子データ（Word/PDF）", "1式",
     "令和9年2月中旬", "5.3.10", "第3回打合せで提出・協議"),
    ("中間成果物", "計画素案（修正版）", "電子データ（Word/PDF）", "1式",
     "令和9年3月上旬", "5.4.1", "村意見を反映"),
    ("中間成果物", "打合せ議事録", "電子データ", "3回分",
     "各打合せ後1週間以内", "1.1.3／1.2.2／1.2.3", "決定事項・宿題事項を明記"),
    ("管理資料", "進捗報告メモ", "電子データ（メール等）", "月1回程度",
     "毎月", "1.2.1", "対面打合せ3回を補完"),
    ("管理資料", "個人情報取扱いに関する誓約書・データ消去証明", "書面／電子データ", "各1式",
     "着手時／完了時", "1.1.4／6.2.2", "音威子府村個人情報保護条例を遵守"),
]


def add_deliverable_sheet(wb):
    ws = wb.create_sheet("04_成果物一覧")
    set_col_widths(ws, [12, 40, 38, 12, 18, 16, 44])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "成果物・提出物一覧")
    ws.merge_cells("A2:G2")
    ws["A2"] = "仕様書「6 成果品」に定める最終成果品のほか、業務進行上必要となる中間成果物・管理資料を整理したもの。"
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["区分", "名称", "形式", "部数・数量", "提出時期",
                             "WBS No.", "備考・根拠"], fill=COLORS["header"])
    ws.row_dimensions[4].height = 26
    r = 5
    for d in DELIVERABLES:
        for c, v in enumerate(d, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (1, 4, 5, 6)))
        if d[0] == "最終成果品":
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=COLORS["band"])
            ws.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{r-1}"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 05_推計作業手順
# ============================================================
ESTIMATION_STEPS = [
    # (工程, 手順No, 手順, 使用データ・ツール, 留意点, WBS No.)
    ("Ⅰ. 人口推計", "1", "基準人口の設定",
     "住民基本台帳人口（各年10月1日現在、男女・5歳階級別）／国勢調査",
     "住基と国勢調査の差異を確認し、どちらを基準とするか村と合意する", "3.1.1"),
    ("Ⅰ. 人口推計", "2", "生残率・純移動率の仮定値設定",
     "社人研「日本の地域別将来推計人口」の仮定値／直近の人口動態実績",
     "小規模自治体は年変動が大きいため、複数年平均で平滑化する", "3.1.1"),
    ("Ⅰ. 人口推計", "3", "出生仮定の設定（子ども女性比・0-4歳性比）",
     "人口動態統計（出生数）／住基",
     "出生数が極少の年があるため、実績の単純外挿は避ける", "3.1.1"),
    ("Ⅰ. 人口推計", "4", "コーホート要因法による将来人口推計",
     "推計シート（Excel）／社人研仮定値",
     "令和9年度〜令和18年度（10年間程度）を男女5歳階級別に推計", "3.1.2"),
    ("Ⅰ. 人口推計", "5", "高齢者人口・高齢化率の算出",
     "推計結果",
     "65〜74歳／75歳以上、高齢化率、後期高齢者割合を年度別に整理", "3.1.2"),
    ("Ⅰ. 人口推計", "6", "推計結果の検証",
     "社人研推計値／過去10年の実績トレンド",
     "乖離が大きい場合は仮定値を再設定し、根拠を記録に残す", "3.1.3"),

    ("Ⅱ. 認定者数推計", "1", "認定者数実績の整理",
     "介護保険事業状況報告／村の認定データ（要支援1〜要介護5・性・年齢階級別）",
     "直近5年分。年度末（3月末）現在で統一する", "3.2.1"),
    ("Ⅱ. 認定者数推計", "2", "性・年齢階級別認定率の算出",
     "認定者数実績 ÷ 該当年齢階級人口",
     "第1号（65歳以上）・第2号（40〜64歳）を区分して算出", "3.2.1"),
    ("Ⅱ. 認定者数推計", "3", "将来の認定率仮定の設定",
     "国の基本指針／直近実績／北海道平均",
     "直近実績固定を基本とし、傾向が明確な場合のみトレンドを反映", "3.2.1"),
    ("Ⅱ. 認定者数推計", "4", "各年度の認定者数の推計",
     "将来人口 × 認定率",
     "令和9〜11年度（第10期）に加え、中長期年次の要否を国の指針で確認", "3.2.2"),
    ("Ⅱ. 認定者数推計", "5", "推計結果の検証・確定",
     "認定調査・審査会の運用状況／施設整備予定",
     "村担当課の実務感覚と照合し、必要に応じて補正する", "3.2.3"),

    ("Ⅲ. サービス見込量推計", "1", "推計ツールへのデータ投入",
     "地域包括ケア「見える化」システム等の推計ツール",
     "国の最新版ツールの公開時期を確認。公開前は暫定シートで先行作業", "4.1.1"),
    ("Ⅲ. サービス見込量推計", "2", "サービス種類別の受給率・利用量の設定",
     "給付実績（国保連データ・事業状況報告）",
     "小規模ゆえ1人の増減で率が大きく振れる。実人数ベースでも確認する", "4.1.2"),
    ("Ⅲ. サービス見込量推計", "3", "施設・居住系サービスの整備方針の反映",
     "村・北海道の整備方針／既存施設の定員",
     "圏域外（近隣市町村）の施設利用分を明示的に見込む", "4.1.2"),
    ("Ⅲ. サービス見込量推計", "4", "総合事業・地域支援事業の見込量設定",
     "総合事業実績／地域支援事業実施計画",
     "上限額（事業費の枠）との整合を確認する", "4.1.2"),
    ("Ⅲ. サービス見込量推計", "5", "ニーズ調査結果・施策方針の反映",
     "介護予防・日常生活圏域ニーズ調査（令和8年度村実施）",
     "介護予防強化等の施策効果をどこまで見込むか村と合意する", "4.1.3"),
    ("Ⅲ. サービス見込量推計", "6", "見込量の妥当性検証",
     "過去トレンド／北海道平均／事業所の供給力",
     "供給力を超える見込みは実現可能性の観点から再検討する", "4.1.4"),

    ("Ⅳ. 給付費・保険料算定", "1", "標準給付費見込額の算出",
     "見込量 × 単位数単価 × 介護報酬改定率",
     "改定率未確定の段階は暫定値（前提を明記）で算定する", "4.2.2"),
    ("Ⅳ. 給付費・保険料算定", "2", "特定入所者介護サービス費等の加算",
     "補足給付・高額介護サービス費・審査支払手数料の実績",
     "実績ベースで見込み、制度改正の影響を確認する", "4.2.2"),
    ("Ⅳ. 給付費・保険料算定", "3", "地域支援事業費の見込",
     "総合事業・包括的支援事業・任意事業の事業費",
     "上限額の算定式に沿って確認する", "4.2.2"),
    ("Ⅳ. 給付費・保険料算定", "4", "算定前提条件の設定",
     "第1号被保険者負担割合／調整交付金交付割合／予定収納率／所得段階別被保険者数",
     "国の基本指針で示される割合を反映。確定前は暫定値とする", "4.3.1"),
    ("Ⅳ. 給付費・保険料算定", "5", "介護給付費準備基金の取崩し方針の検討",
     "基金残高推移／第9期の収支実績",
     "取崩し額を変えた複数パターン（3案程度）を提示し村が方針決定", "4.3.2"),
    ("Ⅳ. 給付費・保険料算定", "6", "保険料収納必要額・基準額の算出",
     "標準給付費＋地域支援事業費 − 各種公費・調整交付金 − 基金取崩し",
     "第9期基準額との比較・増減要因を必ず整理する", "4.3.3"),
    ("Ⅳ. 給付費・保険料算定", "7", "所得段階別保険料額の設定",
     "国の標準段階・乗率／村独自段階の要否",
     "多段階化の要否は村の被保険者構成を踏まえて協議する", "4.3.3"),
    ("Ⅳ. 給付費・保険料算定", "8", "算定結果の説明資料化",
     "算定書・比較表・グラフ",
     "議会・住民説明に耐える根拠の明示（前提条件を必ず併記）", "4.3.4"),
    ("Ⅳ. 給付費・保険料算定", "9", "国の確定情報に基づく再算定",
     "基本指針の告示／介護報酬改定率の確定値",
     "確定が年度末にずれ込む場合、再算定の要否を村と協議（仕様書5）", "4.4.1"),
]


def add_estimation_sheet(wb):
    ws = wb.create_sheet("05_推計作業手順")
    set_col_widths(ws, [20, 7, 38, 44, 48, 12])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:F1")
    style_title(ws["A1"], "推計作業の手順・使用データ（仕様書 3(2)(3) の詳細展開）")
    ws.merge_cells("A2:F2")
    ws["A2"] = ("仕様書に定める「人口推計・要支援要介護認定者数の推計」および「介護サービス見込量、保険料の推計」の"
                "作業手順を分解したもの。国の基本指針・推計ツールの公開状況により手順は適宜見直す。")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["工程", "No.", "手順", "使用データ・ツール", "留意点", "WBS No."],
                     fill=COLORS["header"])
    ws.row_dimensions[4].height = 26

    phase_colors = {"Ⅰ": "C3", "Ⅱ": "C3", "Ⅲ": "C4", "Ⅳ": "C4"}
    r = 5
    prev_phase = None
    for phase, no, step, data, note, wbs in ESTIMATION_STEPS:
        for c, v in enumerate([phase, no, step, data, note, wbs], 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (2, 6)))
        pc = ws.cell(row=r, column=1)
        if phase == prev_phase:
            pc.value = ""
        else:
            key = phase_colors[phase.split(".")[0]]
            pc.fill = PatternFill("solid", fgColor=COLORS[key])
            pc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
            prev_phase = phase
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 06_必要資料リスト
# ============================================================
MATERIALS = [
    # (区分, 資料名, 対象期間・粒度, 入手先, 個人情報, 提出期限目安, 用途 WBS)
    ("計画", "第9期高齢者福祉計画・介護保険事業計画（本編・資料編）", "現行計画一式", "村", "なし", "令和8年9月中旬", "2.4.1／5.2.1"),
    ("計画", "第9期計画の進捗管理・評価資料（事業別実績、目標値と実績）", "令和6〜8年度", "村", "なし", "令和8年9月末", "2.4.1"),
    ("計画", "村総合計画、地域福祉計画、健康増進計画等の関連計画", "現行計画一式", "村", "なし", "令和8年9月末", "1.3.3／5.1.3"),
    ("人口", "住民基本台帳人口（男女・5歳階級別、各年10月1日現在）", "直近10年", "村", "なし（集計値）", "令和8年9月末", "2.1.1／3.1.1"),
    ("人口", "人口動態（出生・死亡・転入・転出）", "直近10年", "村", "なし（集計値）", "令和8年9月末", "3.1.1"),
    ("人口", "世帯数、高齢者単身世帯・高齢者のみ世帯の状況", "直近5年", "村", "なし（集計値）", "令和8年9月末", "2.1.1"),
    ("人口", "国勢調査結果（人口・世帯・就業等）", "直近2回", "総務省（公表）", "なし", "－", "2.1.1"),
    ("人口", "日本の地域別将来推計人口", "最新版", "社人研（公表）", "なし", "－", "3.1.1／3.1.3"),
    ("介護", "要介護認定者数（要支援1〜要介護5、性・年齢階級別、年度末現在）", "直近5年", "村", "なし（集計値）", "令和8年9月末", "2.3.1／3.2.1"),
    ("介護", "介護給付費実績（サービス種類別・月次／年次、受給者数・回数・給付費）", "直近5年", "村（国保連データ）", "なし（集計値）", "令和8年9月末", "2.3.2／4.1.1"),
    ("介護", "介護保険事業状況報告（月報・年報）", "直近5年", "村／厚労省（公表）", "なし", "令和8年9月末", "2.3.2"),
    ("介護", "介護保険特別会計 決算・予算（歳入歳出）", "直近5年", "村", "なし", "令和8年10月上旬", "2.3.4／4.3.1"),
    ("介護", "介護給付費準備基金の残高推移・積立取崩実績", "直近5年", "村", "なし", "令和8年10月上旬", "2.3.4／4.3.2"),
    ("介護", "第9期保険料の所得段階別被保険者数・保険料額・収納率", "令和6〜8年度", "村", "なし（集計値）", "令和8年10月上旬", "2.3.4／4.3.1"),
    ("介護", "介護保険事業所・施設一覧（村内・近隣市町村、定員・稼働状況）", "現況", "村／北海道", "なし", "令和8年10月上旬", "2.1.2／4.1.2"),
    ("福祉", "高齢者福祉サービス（村単独事業）の利用実績・事業費", "直近5年", "村", "なし（集計値）", "令和8年9月末", "2.2.1"),
    ("福祉", "介護予防・日常生活支援総合事業の実施状況・事業費", "直近5年", "村", "なし（集計値）", "令和8年9月末", "2.2.2"),
    ("福祉", "包括的支援事業・任意事業の実施状況（地域包括支援センター運営状況）", "直近5年", "村", "なし（集計値）", "令和8年9月末", "2.2.3"),
    ("福祉", "在宅医療・介護連携、認知症施策、生活支援体制整備の取組状況", "現況", "村", "なし", "令和8年10月上旬", "2.2.3／5.3.5"),
    ("調査", "介護予防・日常生活圏域ニーズ調査 集計データ（令和8年度実施分）", "令和8年度", "村", "個人が特定されない集計値で受領", "令和8年10月中旬", "2.4.2／4.1.3"),
    ("調査", "同 調査票・実施概要（対象者数・回収数・回収率）", "令和8年度", "村", "なし", "令和8年10月中旬", "2.4.2"),
    ("体制", "村の組織図・策定体制、庁内会議・協議会等の設置状況", "現況", "村", "なし", "令和8年9月中旬", "5.3.1／5.3.7"),
    ("体制", "計画策定・議決に関する想定スケジュール（議会日程、パブコメ日程）", "令和8〜9年度", "村", "なし", "令和8年9月中旬", "1.1.2／5.4.2"),
    ("国", "国の基本指針・関係通知・Q&A（第10期関係）", "随時", "厚生労働省（公表）", "なし", "－", "1.2.4／5.2.2"),
    ("国", "地域包括ケア「見える化」システム データ・推計ツール", "最新版", "厚生労働省", "なし", "－", "2.1.3／4.1.1"),
]


def add_material_sheet(wb):
    ws = wb.create_sheet("06_必要資料リスト")
    set_col_widths(ws, [8, 52, 18, 22, 26, 18, 16])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "必要資料リスト（貸与依頼票）")
    ws.merge_cells("A2:G2")
    ws["A2"] = ("契約後2週間以内に村へ提示し、資料の授受状況を管理する。個人情報を含む資料は原則として"
                "個人が特定されない集計値で受領し、音威子府村個人情報保護条例に基づき取り扱う。")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["区分", "資料名", "対象期間・粒度", "入手先",
                             "個人情報の取扱い", "提出期限目安", "用途（WBS No.）"],
                     fill=COLORS["header"])
    ws.row_dimensions[4].height = 26
    r = 5
    for m in MATERIALS:
        for c, v in enumerate(m, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (1, 3, 6, 7)))
        ws.row_dimensions[r].height = 28
        r += 1

    # 受領チェック欄
    ws.cell(row=4, column=8, value="受領状況")
    hc = ws.cell(row=4, column=8)
    hc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    hc.fill = PatternFill("solid", fgColor=COLORS["header"])
    hc.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    hc.border = BORDER
    ws.column_dimensions["H"].width = 12
    dv = DataValidation(type="list", formula1='"未依頼,依頼済,受領済,一部受領,対象外"', allow_blank=True)
    ws.add_data_validation(dv)
    for rr in range(5, r):
        cell = ws.cell(row=rr, column=8, value="未依頼")
        style_data_cell(cell, alt=(rr % 2 == 0), center=True)
    dv.add(f"H5:H{r-1}")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:H{r-1}"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 07_役割分担（RACI）
# ============================================================
RACI = [
    # (WBS, 作業, 受託者, 村担当課, 村庁内関係課, 国・北海道/関係機関, 備考)
    ("1.1.2", "業務実施計画書・工程表の作成", "R", "A", "I", "－", "村の承認をもって確定"),
    ("1.1.3", "第1回打合せ（キックオフ）", "R", "A", "C", "－", "計画策定担当者本人が出席"),
    ("1.1.4", "個人情報の取扱手続", "R", "A", "I", "－", "音威子府村個人情報保護条例に基づく"),
    ("1.3.1", "貸与資料の提供", "C", "R/A", "C", "－", "資料抽出は村、リスト作成は受託者"),
    ("1.3.2", "公表統計・システムデータの収集", "R/A", "I", "－", "C", "見える化システム等"),
    ("2.1", "現状把握・分析", "R/A", "C", "I", "－", "村は地域事情の情報提供"),
    ("2.2", "高齢者福祉サービス実績の分析", "R", "A", "C", "－", "事業実績の確定は村"),
    ("2.3", "介護保険サービス実績の分析", "R", "A", "C", "－", "給付データの提供は村"),
    ("2.4.1", "第9期計画の進捗評価", "R", "A", "C", "－", "各事業の自己評価は村が実施"),
    ("2.4.2", "ニーズ調査結果の分析", "R", "A", "I", "－", "調査の実施・集計は村（業務範囲外）"),
    ("3.1", "人口推計", "R/A", "C", "I", "C", "社人研仮定値を参照"),
    ("3.2", "要支援要介護認定者数の推計", "R", "A", "I", "C", "国の基本指針に従う"),
    ("4.1", "介護サービス見込量の推計", "R", "A", "C", "C", "施策方針の決定は村"),
    ("4.2", "給付費・地域支援事業費の見込", "R", "A", "C", "I", "国の改定情報を反映"),
    ("4.3.2", "財政シミュレーション（基金取崩し）", "R", "A", "C", "I", "取崩し方針の決定は村"),
    ("4.3.3", "第10期介護保険料の算定", "R", "A", "C", "I", "保険料水準の決定は村"),
    ("4.4.1", "国の確定情報に基づく再算定", "R", "A", "I", "C", "仕様変更が必要な場合は協議（仕様書5）"),
    ("5.1", "課題整理・施策の方向の提案", "R", "A", "C", "－", "方針決定は村"),
    ("5.2", "計画骨子の提案", "R", "A", "C", "－", "国の策定指針に従う"),
    ("5.3", "計画素案の執筆", "R", "A", "C", "I", "各施策の内容確認は関係課"),
    ("5.4.2", "パブリックコメント等の意見反映", "R", "A", "C", "－", "パブコメの実施は村"),
    ("6.1", "成果品データの作成", "R/A", "C", "－", "－", "著作権は村に帰属"),
    ("6.2.1", "納品・検査", "R", "A", "I", "－", "履行期限：令和9年3月31日"),
    ("－", "計画の議決・公表", "C", "R/A", "C", "I", "業務範囲外（村が実施）"),
    ("－", "介護保険条例の改正（保険料）", "C", "R/A", "C", "I", "業務範囲外（村が実施。算定根拠は受託者が提供）"),
]

RACI_LEGEND = [
    ("R", "Responsible：実行責任者（実際に作業を行う）", "70AD47"),
    ("A", "Accountable：説明責任者（成果を承認・決定する）", "C00000"),
    ("C", "Consulted：協議先（事前に意見を求める）", "ED7D31"),
    ("I", "Informed：報告先（結果を共有する）", "7F7F7F"),
]


def add_raci_sheet(wb):
    ws = wb.create_sheet("07_役割分担")
    set_col_widths(ws, [10, 42, 12, 14, 16, 20, 44])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "役割分担表（RACI）")
    ws.merge_cells("A2:G2")
    ws["A2"] = ("仕様書「7 留意事項」に基づき、村担当課と緊密な連携を図りながら業務を遂行する。"
                "決定を要する事項（施策方針・保険料水準・基金取崩し等）は村が決定し、受託者は根拠と選択肢を提示する。")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["WBS No.", "作業", "受託者", "村担当課",
                             "村庁内関係課", "国・北海道／関係機関", "備考"], fill=COLORS["header"])
    ws.row_dimensions[4].height = 30

    raci_fill = {l[0]: l[2] for l in RACI_LEGEND}
    r = 5
    for row_data in RACI:
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (1, 3, 4, 5, 6)))
            if c in (3, 4, 5, 6) and v not in ("－", ""):
                key = v.split("/")[0]
                if key in raci_fill:
                    cell.fill = PatternFill("solid", fgColor=raci_fill[key])
                    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_subhead(ws.cell(row=r, column=1), "凡例")
    r += 1
    for mark, desc, color in RACI_LEGEND:
        mc = ws.cell(row=r, column=1, value=mark)
        mc.fill = PatternFill("solid", fgColor=color)
        mc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        dc = ws.cell(row=r, column=2, value=desc)
        style_data_cell(dc)
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 08_リスク・課題管理
# ============================================================
RISKS = [
    ("R-01", "スケジュール", "国の基本指針の告示や介護報酬改定率の決定が年度末にずれ込み、保険料が確定できない",
     "計画素案の第6章が確定せず、納期に影響", "高", "大",
     "暫定値で算定し、確定後に差し替えられるよう算定シートをパラメータ化。仕様書5に基づき村と協議（WBS 4.4.1）",
     "受託者", "毎月／国通知の公表時"),
    ("R-02", "スケジュール", "村からの貸与資料の提供が遅れ、分析着手が後ろ倒しになる",
     "現状分析・推計工程が圧迫される", "中", "大",
     "契約後2週間以内に必要資料リストを提示し、提出期限目安を明示。受領状況を『06_必要資料リスト』で管理",
     "受託者・村", "毎月"),
    ("R-03", "品質", "小規模自治体のため給付実績・認定者数の年変動が大きく、推計値が不安定になる",
     "見込量・保険料が過大／過小になる", "高", "大",
     "複数年平均による平滑化、実人数ベースの併用、複数シナリオ（3案程度）の提示により幅を示す",
     "受託者", "推計工程（R8.10〜R9.1）"),
    ("R-04", "品質", "村内のサービス事業所が限られ、近隣市町村の資源に依存している",
     "見込量が供給力と乖離する", "中", "中",
     "圏域外利用の実績を明示的に分解して推計。近隣市町村・北海道の整備方針を確認して反映",
     "受託者・村", "WBS 4.1.2〜4.1.4"),
    ("R-05", "品質", "ニーズ調査の回収数が少なく、統計的な信頼性が限られる",
     "調査結果に基づく施策立案の根拠が弱くなる", "中", "中",
     "経年比較・北海道平均との比較で補完し、自由記述等の定性情報を併用。限界を計画本文に明記",
     "受託者", "WBS 2.4.2"),
    ("R-06", "コミュニケーション", "対面打合せが3回程度に限られ、認識齟齬が生じる",
     "手戻り・修正工数の増大", "中", "中",
     "月1回程度のメール・Web会議による中間確認を実施（WBS 1.2.1）。議事録で決定事項を明文化",
     "受託者", "毎月"),
    ("R-07", "スケジュール", "議会日程・パブリックコメント日程が前倒しとなる",
     "素案完成が間に合わない", "中", "大",
     "第1回打合せで議会・パブコメ日程を確認し工程に反映。素案は令和9年1月中旬完成を目標に前倒し管理",
     "受託者・村", "第1回打合せ／毎月"),
    ("R-08", "情報管理", "個人情報を含むデータの授受・保管におけるリスク",
     "条例違反・情報漏えい", "低", "大",
     "個人が特定されない集計値での受領を原則とし、やむを得ない場合は暗号化・持出制限・アクセス権限管理を実施。"
     "完了時にデータ消去証明を提出（WBS 1.1.4／6.2.2）",
     "受託者", "常時"),
    ("R-09", "体制", "業務従事者の不稼働（傷病等）により進捗が停滞する",
     "納期遅延", "低", "大",
     "社内の代替要員・レビュー体制を確保。再委託が必要となる場合は事前に村の承諾を得る（仕様書7）",
     "受託者", "常時"),
    ("R-10", "スケジュール", "冬季の積雪・交通事情により現地打合せが実施できない",
     "打合せの延期・工程遅延", "中", "小",
     "Web会議の併用を第1回打合せで合意しておく。現地訪問は必要最小限に集約",
     "受託者・村", "R8.12〜R9.3"),
    ("R-11", "スコープ", "国の指針改定により推計対象年次や記載事項が追加される",
     "作業量の増加", "中", "中",
     "仕様書5に基づき村と協議して対応方針を決定。原則として国の指示を遵守",
     "受託者・村", "国通知の公表時"),
    ("R-12", "品質", "関連計画（総合計画等）との整合が取れていない記述が生じる",
     "計画間の齟齬", "低", "中",
     "関連計画整合性チェック表により章ごとに確認（WBS 5.1.3）。齟齬は村と協議のうえ調整",
     "受託者", "WBS 5.3 執筆時"),
]


def add_risk_sheet(wb):
    ws = wb.create_sheet("08_リスク管理")
    set_col_widths(ws, [8, 16, 46, 30, 10, 10, 56, 14, 18])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:I1")
    style_title(ws["A1"], "リスク・課題管理表")
    ws.merge_cells("A2:I2")
    ws["A2"] = "業務遂行上想定されるリスクと対応方針。発生時は速やかに村と協議のうえ対応する（仕様書「7 留意事項」）。"
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["No.", "区分", "リスク内容", "想定される影響", "発生可能性",
                             "影響度", "対応方針", "主担当", "監視タイミング"], fill=COLORS["header"])
    ws.row_dimensions[4].height = 30

    level_color = {"高": "C00000", "中": "ED7D31", "低": "70AD47",
                   "大": "C00000", "小": "70AD47"}
    r = 5
    for risk in RISKS:
        for c, v in enumerate(risk, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c in (1, 2, 5, 6, 8, 9)))
            if c in (5, 6) and v in level_color:
                cell.fill = PatternFill("solid", fgColor=level_color[v])
                cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 46
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{r-1}"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 09_スコープ管理
# ============================================================
SCOPE = [
    ("含む", "現状の評価・分析", "統計資料による現状把握・分析、高齢者福祉サービス実績分析、介護保険サービス実績分析、現行計画の点検・評価",
     "仕様書 3(1)", "WBS 2"),
    ("含む", "人口推計", "令和9年から10年間程度の将来人口推計", "仕様書 3(2)", "WBS 3.1"),
    ("含む", "要支援要介護認定者数の推計", "国の基本指針に従った各年度の認定者数推計", "仕様書 3(2)", "WBS 3.2"),
    ("含む", "介護サービス見込量の推計", "推計ツール等による令和11年度までの各年度・サービス種類ごとの見込量推計",
     "仕様書 3(3)", "WBS 4.1"),
    ("含む", "介護保険料の算定", "計画期間内の見込量に基づく第10期介護保険料の算定", "仕様書 3(3)", "WBS 4.3"),
    ("含む", "ニーズ調査結果の反映", "令和8年度に村が実施した介護予防・日常生活圏域ニーズ調査結果の計画への反映",
     "仕様書 3(3)", "WBS 2.4.2／4.1.3"),
    ("含む", "計画素案の作成", "課題・施策の方向の整理、計画骨子の提案、計画素案の作成（国の策定指針に従った記載）",
     "仕様書 3(4)", "WBS 5"),
    ("含む", "打合せ", "3回程度。受託者側は計画策定担当者本人が出席", "仕様書 4", "WBS 1.1.3／1.2.2／1.2.3"),
    ("含む", "成果品の作成・納品", "計画（CDデータ）1部、業務データ収録の磁気媒体（CDデータ）1部",
     "仕様書 6", "WBS 6"),
    ("含む", "関連計画との整合性確保", "村総合計画等の関連計画との整合を図る", "仕様書 7", "WBS 5.1.3"),

    ("含まない", "ニーズ調査の実施・集計", "調査票設計、対象者抽出、発送・回収、データ入力・集計は業務範囲外（令和8年度に村が実施済み）",
     "仕様書 3(3)「今年度実施した」との記載", "結果データの提供を受けて反映"),
    ("含まない", "印刷・製本", "成果品はCDデータのみ。冊子の印刷・製本は仕様書に記載なし",
     "仕様書 6 成果品", "必要な場合は別途協議"),
    ("含まない", "計画の議決・公表手続", "議会議決、計画の公表・ホームページ掲載等の行政手続", "－", "村が実施"),
    ("含まない", "介護保険条例の改正案作成", "保険料改定に伴う条例改正案の作成・法制執務", "－", "村が実施（算定根拠は受託者が提供）"),

    ("要協議", "計画策定委員会等の設置・運営", "策定委員会・協議体を設置する場合の会議運営、委員名簿作成、会議録作成",
     "仕様書に記載なし", "設置の要否・受託者の関与範囲を第1回打合せで確認"),
    ("要協議", "パブリックコメントの実施", "実施要領作成、意見募集の運営。WBSでは意見反映のみを計上",
     "仕様書に記載なし", "実施主体・日程・支援範囲を協議（WBS 5.4.2）"),
    ("要協議", "住民説明会・議会説明", "説明会資料の作成・説明員としての出席", "仕様書に記載なし", "必要な場合は別途協議"),
    ("要協議", "中長期推計（令和22年度等）の対象年次", "国の基本指針が中長期の推計を求める場合の追加推計",
     "仕様書 3(2)「10年間程度」", "国の指針確定後に対象年次を協議（WBS 3.1.3）"),
    ("要協議", "国の指示に伴う仕様変更", "国の指示により本仕様を変更する必要が生じた場合の対応",
     "仕様書 5 その他", "委託者・受託者の協議により決定。原則として国の指示を遵守"),
    ("要協議", "再委託", "業務の全部または主要な部分の第三者への再委託", "仕様書 7 留意事項",
     "予定なし。必要が生じた場合は事前に村の承諾を得る"),
]


def add_scope_sheet(wb):
    ws = wb.create_sheet("09_スコープ管理")
    set_col_widths(ws, [12, 34, 62, 32, 44])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")
    style_title(ws["A1"], "スコープ管理表（業務範囲の整理）")
    ws.merge_cells("A2:E2")
    ws["A2"] = ("仕様書の記載から業務範囲を整理したもの。「要協議」の項目は第1回打合せで取扱いを確認する"
                "（仕様書「7 留意事項」：仕様書に定めのない事項や疑義が生じた場合は都度協議のうえ決定）。")
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    style_header_row(ws, 4, ["区分", "項目", "内容", "仕様書の根拠", "備考・対応"], fill=COLORS["header"])
    ws.row_dimensions[4].height = 26

    kind_color = {"含む": "70AD47", "含まない": "7F7F7F", "要協議": "ED7D31"}
    r = 5
    for s in SCOPE:
        for c, v in enumerate(s, 1):
            cell = ws.cell(row=r, column=c, value=v)
            style_data_cell(cell, alt=(r % 2 == 0), center=(c == 1))
        kc = ws.cell(row=r, column=1)
        kc.fill = PatternFill("solid", fgColor=kind_color[s[0]])
        kc.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 34
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{r-1}"
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# 10_工数集計
# ============================================================
def add_effort_sheet(wb):
    ws = wb.create_sheet("10_工数集計")
    set_col_widths(ws, [10, 34, 30, 12, 12, 14, 40])
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "工数集計（『01_WBS』シートと連動）")
    ws.merge_cells("A2:G2")
    ws["A2"] = "工数はレベル3の作業項目に付した見積値の合計。1人日＝7.5時間、1人月＝20人日で換算。契約・体制の確定後に見直す。"
    ws["A2"].font = Font(name=FONT, size=9, color="595959")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    q = "'01_WBS'!"
    a, b = WBS_FIRST_DATA_ROW, WBS_LAST_DATA_ROW
    total_formula = f"SUMIF({q}$C${a}:$C${b},3,{q}$K${a}:$K${b})"

    # --- 大分類別 ---
    r = 4
    ws.merge_cells(f"A{r}:G{r}")
    style_subhead(ws.cell(row=r, column=1), "1. 大分類別工数")
    r += 1
    style_header_row(ws, r, ["区分", "大分類", "対象WBS", "工数(人日)", "構成比",
                             "人月換算", "主な内容"], fill=COLORS["header"])
    r += 1
    first_cat_row = r
    cat_desc = {
        "1": "業務実施計画、打合せ3回、進行管理、資料収集",
        "2": "現状把握、高齢者福祉・介護保険サービス実績分析、第9期計画の点検評価",
        "3": "コーホート要因法による人口推計、要支援要介護認定者数の推計",
        "4": "サービス種類別見込量、標準給付費、第10期介護保険料の算定",
        "5": "課題整理、計画骨子、計画素案の執筆・修正",
        "6": "最終校正、CDデータ作成、納品・検査対応",
    }
    for key, name in CAT_NAMES.items():
        # 該当する大分類のレベル3行だけを合計（WBS No. の先頭桁で判定）
        formula = (f'=SUMPRODUCT(({q}$C${a}:$C${b}=3)*'
                   f'(LEFT({q}$B${a}:$B${b},1)="{key}")*{q}$K${a}:$K${b})')
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=f"{key}.1〜")
        ws.cell(row=r, column=4, value=formula)
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/{total_formula},0)")
        ws.cell(row=r, column=6, value=f"=ROUND(D{r}/20,2)")
        ws.cell(row=r, column=7, value=cat_desc[key])
        for c in range(1, 8):
            style_data_cell(ws.cell(row=r, column=c), alt=(r % 2 == 0), center=(c in (1, 3, 4, 5, 6)))
        c1 = ws.cell(row=r, column=1)
        c1.fill = PatternFill("solid", fgColor=CAT_COLORS[key])
        c1.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.cell(row=r, column=4).number_format = "#,##0.0"
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        ws.row_dimensions[r].height = 24
        r += 1
    last_cat_row = r - 1
    # 大分類合計
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    tc = ws.cell(row=r, column=1, value="合計")
    tc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    ws.cell(row=r, column=4, value=f"=SUM(D{first_cat_row}:D{last_cat_row})")
    ws.cell(row=r, column=5, value=f"=SUM(E{first_cat_row}:E{last_cat_row})")
    ws.cell(row=r, column=6, value=f"=SUM(F{first_cat_row}:F{last_cat_row})")
    ws.cell(row=r, column=7, value=f"=IF(ROUND(D{r}-{total_formula},4)=0,\"01_WBS の合計と一致\",\"要確認：01_WBSと不一致\")")
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center", horizontal="right", indent=1)
    ws.cell(row=r, column=7).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.cell(row=r, column=4).number_format = "#,##0.0"
    ws.cell(row=r, column=5).number_format = "0.0%"
    ws.cell(row=r, column=6).number_format = "#,##0.00"
    ws.row_dimensions[r].height = 24
    r += 2

    # --- 中分類別 ---
    ws.merge_cells(f"A{r}:G{r}")
    style_subhead(ws.cell(row=r, column=1), "2. 中分類別工数")
    r += 1
    style_header_row(ws, r, ["WBS No.", "大分類", "中分類", "工数(人日)", "構成比",
                             "開始予定", "完了予定"], fill=COLORS["header"])
    r += 1
    for i, m in enumerate(ROWS):
        if m["lv"] != 2:
            continue
        wbs_row = WBS_FIRST_DATA_ROW + i
        ws.cell(row=r, column=1, value=m["wbs"])
        ws.cell(row=r, column=2, value=m["l1name"])
        ws.cell(row=r, column=3, value=m["name"])
        ws.cell(row=r, column=4, value=f"={q}$K${wbs_row}")
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/{total_formula},0)")
        ws.cell(row=r, column=6, value=m["st"])
        ws.cell(row=r, column=7, value=m["en"])
        for c in range(1, 8):
            style_data_cell(ws.cell(row=r, column=c), alt=(r % 2 == 0), center=(c in (1, 4, 5, 6, 7)))
        ws.cell(row=r, column=4).number_format = "#,##0.0"
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6).number_format = "yyyy/mm/dd"
        ws.cell(row=r, column=7).number_format = "yyyy/mm/dd"
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    note = ws.cell(row=r, column=1,
                   value="※ 工数は仕様書の記載事項から想定した見積値であり、契約金額・要員計画の確定後に実績に合わせて更新すること。")
    note.font = Font(name=FONT, size=9, color="595959")
    note.alignment = Alignment(vertical="center", indent=1)

    ws.sheet_view.showGridLines = False
    return ws


# ============================================================
# Markdown 版（git 上でレビューするための同内容の要約）
# ============================================================
def build_markdown():
    def jp(d):
        y = d.year - 2018  # 2019年=令和1年
        return f"R{y}.{d.month}.{d.day}"

    L = []
    L.append(f"# {PLAN_NAME}　策定業務委託　WBS")
    L.append("")
    L.append(f"- **発注者**：{CLIENT}")
    L.append("- **委託期間**：契約締結日 〜 令和9年3月31日")
    L.append("- **計画期間**：第10期＝令和9年度〜令和11年度（3か年）")
    L.append("- **人口推計**：令和9年から10年間程度（令和9年度〜令和18年度）")
    L.append("- **打合せ**：3回程度（受託者側は計画策定担当者本人が出席）")
    L.append("- **成果品**：計画（CDデータ）1部／業務データ収録の磁気媒体（CDデータ）1部")
    n3 = sum(1 for r in ROWS if r["lv"] == 3)
    eff = sum(r["eff"] for r in ROWS if r["lv"] == 3)
    L.append(f"- **規模**：作業項目（レベル3）{n3}件／想定工数 {eff:.1f}人日（約{eff/20:.1f}人月）")
    L.append("")
    L.append(f"詳細は Excel 版 `output/{FILENAME}` を参照（全11シート）。")
    L.append("")
    L.append("> **注意**：契約締結日が未確定のため、令和8年9月1日と仮定して日程を設定している。")
    L.append("> 確定後は `build_wbs_otoineppu.py` の `CONTRACT_START` を修正して再生成すること。")
    L.append("")

    L.append("## 1. 大分類別の工数")
    L.append("")
    L.append("| 区分 | 大分類 | 工数(人日) | 構成比 |")
    L.append("|---|---|---:|---:|")
    for key, name in CAT_NAMES.items():
        e = sum(r["eff"] for r in ROWS if r["lv"] == 3 and r["cat"] == key)
        L.append(f"| {key} | {name[3:]} | {e:.1f} | {e/eff*100:.1f}% |")
    L.append(f"| | **合計** | **{eff:.1f}** | **100.0%** |")
    L.append("")

    L.append("## 2. WBS")
    L.append("")
    L.append("| WBS No. | 作業項目 | 成果物 | 工数 | 開始 | 完了 | 仕様書 |")
    L.append("|---|---|---|---:|---|---|---|")
    for r in ROWS:
        if r["lv"] == 1:
            L.append(f"| **{r['wbs']}** | **{r['name']}** | | **{sum(x['eff'] for x in ROWS if x['lv']==3 and x['cat']==r['cat']):.1f}** | {jp(r['st'])} | {jp(r['en'])} | {r['spec']} |")
        elif r["lv"] == 2:
            kids = [x for x in ROWS if x["lv"] == 3 and x["wbs"].startswith(r["wbs"] + ".")]
            L.append(f"| {r['wbs']} | *{r['name']}* | | {sum(x['eff'] for x in kids):.1f} | {jp(r['st'])} | {jp(r['en'])} | |")
        else:
            L.append(f"| {r['wbs']} | {r['name']} | {r['deliv']} | {r['eff']:.1f} | {jp(r['st'])} | {jp(r['en'])} | {r['spec']} |")
    L.append("")

    L.append("## 3. マイルストーン")
    L.append("")
    L.append("| No. | マイルストーン | 予定時期 | WBS No. | 提出物 |")
    L.append("|---|---|---|---|---|")
    for mm in MILESTONES:
        L.append(f"| {mm[0]} | {mm[1]} | {mm[2]} | {mm[3]} | {mm[6]} |")
    L.append("")

    L.append("## 4. スコープ（仕様書からの整理）")
    L.append("")
    for kind in ("含む", "含まない", "要協議"):
        L.append(f"### {kind}")
        L.append("")
        for sc in SCOPE:
            if sc[0] == kind:
                L.append(f"- **{sc[1]}**：{sc[2]}　（根拠：{sc[3]}／{sc[4]}）")
        L.append("")

    L.append("## 5. 主なリスク")
    L.append("")
    L.append("| No. | 区分 | リスク | 可能性 | 影響度 | 対応方針 |")
    L.append("|---|---|---|---|---|---|")
    for rk in RISKS:
        L.append(f"| {rk[0]} | {rk[1]} | {rk[2]} | {rk[4]} | {rk[5]} | {rk[6]} |")
    L.append("")

    path = os.path.join(OUT_DIR, FILENAME.replace(".xlsx", ".md"))
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print(f"  ✓ 作成: {path}")
    return path


# ============================================================
# ブック生成
# ============================================================
def build():
    wb = Workbook()
    add_overview_sheet(wb)
    add_wbs_sheet(wb)
    add_schedule_sheet(wb)
    add_milestone_sheet(wb)
    add_deliverable_sheet(wb)
    add_estimation_sheet(wb)
    add_material_sheet(wb)
    add_raci_sheet(wb)
    add_risk_sheet(wb)
    add_scope_sheet(wb)
    add_effort_sheet(wb)

    path = os.path.join(OUT_DIR, FILENAME)
    wb.save(path)
    n3 = sum(1 for r in ROWS if r["lv"] == 3)
    eff = sum(r["eff"] for r in ROWS if r["lv"] == 3)
    print(f"  ✓ 作成: {path}")
    print(f"    シート数: {len(wb.sheetnames)}  /  作業項目（レベル3）: {n3}件  /  総工数: {eff:.1f}人日")
    return path


if __name__ == "__main__":
    print(f"【{PLAN_NAME}】WBS を作成します")
    build()
    build_markdown()
    print("完了。出力先: " + OUT_DIR)
