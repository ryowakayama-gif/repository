# -*- coding: utf-8 -*-
"""
北海道・東北 案件サマリー ジェネレータ

営業会議ブック（週次シート）の最新回 20260904㉜（第32回・2026/9/4時点）から、
北海道／東北エリアの
  - 今年度（今期）の案件残数・見込売上
  - 来期の案件数・売上計画
を抽出・集計してブック化する。

金額は原本の単価×掛率で算定（原本と同じ ROUND 計算）。単位は千円。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/home/user/repository/output"
os.makedirs(OUT_DIR, exist_ok=True)
OUT_PATH = os.path.join(OUT_DIR, "北海道・東北_案件サマリー_20260904.xlsx")

SRC_BOOK = "営業会議ブック"
SRC_SHEET = "20260904㉜"        # 第32回・2026/9/4 時点（最新回）
PREV_SHEET = "20260828㉛"       # 前回（先週比の算定元）

FONT = "游ゴシック"
COLORS = {
    "header":  "1F3864",
    "subhead": "2E75B6",
    "band":    "DDEBF7",
    "alt":     "F7FAFC",
    "hokkaido":"2CA02C",
    "tohoku":  "1F77B4",
    "total":   "FFF2CC",
    "note":    "FFF3F3",
}
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0;(#,##0);-'
PCT = '0.0%'

# ============================================================
# 原本データ（出所セルはコメント参照）
# ============================================================

# 【現在の案件残】 SRC_SHEET!D32:K38   E列=北海道 F列=東北 J列=単価 K列=掛率
ZANSU = [
    # テーマ, 北海道件数, 東北件数, 関東件数, 西日本件数, 単価(千円), 掛率
    ("公会計",     0, 4, 5, 3, 1000, 0.7),
    ("会計支援",   0, 0, 0, 0, None, None),
    ("経営戦略",   0, 0, 0, 0, 3000, 0.8),
    ("経営改善",   0, 0, 2, 1, 4000, 0.7),
    ("公共施設M",  3, 3, 1, 2, 2000, 0.5),
    ("福祉計画",   1, 2, 1, 3, 2000, 0.5),
    ("その他",     0, 0, 0, 0, None, None),
]
ZANSU_GENPON = (31, 32800)   # 原本 I39 / L39（検証用）

# 【補正対象】 SRC_SHEET!D42:K48
HOSEI = [
    ("公会計",     0, 3, 0, 0, 1000, 0.7),
    ("会計支援",   0, 0, 0, 0, None, None),
    ("経営戦略",   0, 0, 4, 1, 3000, 0.8),
    ("経営改善",   0, 0, 3, 0, 4000, 0.7),
    ("公共施設M",  0, 2, 2, 0, 2000, 0.5),
    ("福祉計画",   1, 0, 0, 1, 2000, 0.5),
    ("その他",     0, 0, 1, 0, None, None),
]
HOSEI_GENPON = (18, 28500)   # 原本 I49 / L49（検証用）

# 来期テーマ別計画 SRC_SHEET!F123:P131  G/H=北海道 I/J=東北 O/P=全社
RAIKI = [
    # テーマ, 北件数, 北金額, 東件数, 東金額, 全社件数, 全社金額
    ("公会計",               182, 121100, 102, 70000, 416, 269600),
    ("会計支援",              35,  25400,  18, 11000,  70,  56000),
    ("公営企業",              14,  33500,  16, 37500,  66, 188200),
    ("公共施設マネジメント",   14,  26400,  12, 21000,  54, 110100),
    ("総合計画",               5,  10500,   0,     0,  11,  19500),
    ("福祉計画",               6,  12000,   8, 18500,  34,  79900),
    ("社会生活計画",           7,  15000,   6, 13000,  26,  61600),
    ("その他計画",            15,  25500,  10, 19000,  33,  64500),
    ("その他",                 2,    600,   0,     0,   2,    600),
]

# 今期 確保済実績 SRC_SHEET!G7:J7（随契）／G9:J9（落札）
JISSEKI = [
    ("北海道", 168748, 37614),
    ("東北",   110969, 42890),
    ("関東",    97200, 99921),
    ("西日本",      0, 52072),
]

# 見積提示済件数 SRC_SHEET!H172:I172（北海道合計）／H186:I186（東北合計）
MITSUMORI = {"北海道": (1, 44), "東北": (2, 57), "関東": (2, 29), "西日本": (0, 4)}

# 【指名停止の影響】 SRC_SHEET!C20:K26  D列=北海道 E列=東北
SHIMEI = [
    # テーマ, 北海道, 東北, 関東, 西日本, 単価, 掛率
    ("公会計",     0, 3, 5,  6, 1000, 0.7),
    ("会計支援",   0, 0, 0,  0, None, None),
    ("経営戦略",   0, 0, 8,  8, 3000, 0.8),
    ("経営改善",   0, 0, 6,  4, 4000, 0.7),
    ("公共施設M",  0, 2, 5,  4, 2000, 0.5),
    ("福祉計画",   0, 7, 9, 12, 2000, 0.5),
    ("その他",     0, 2, 0,  0, 3000, 0.5),
]
SHIMEI_GENPON = (81, 118200)   # 原本 H27 / K27（検証用）

# 来期要員数 SRC_SHEET!G133（北海道）／I133（東北）
YOIN = {"北海道": 12, "東北": 6}


# ============================================================
# 書式ヘルパ
# ============================================================
def title(ws, cell, text, size=14):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=size, bold=True, color=COLORS["header"])


def note(ws, cell, text, size=9, color="595959"):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=size, color=color)


def head(ws, row, col, text, width=None, fill="subhead"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLORS[fill])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def body(ws, row, col, value, fmt=None, bold=False, fill=None, align="right", src=False):
    """src=True … 原本からの転記値（青字）／それ以外は計算式（黒字）"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=10, bold=bold, color="0000FF" if src else "000000")
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = PatternFill("solid", fgColor=COLORS[fill])
    return c


def label(ws, row, col, text, bold=False, fill=None):
    return body(ws, row, col, text, bold=bold, fill=fill, align="left")


# ============================================================
# Sheet 1: サマリー
# ============================================================
def build_summary(ws, refs):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    for col in "CDEFGHI":
        ws.column_dimensions[col].width = 14

    title(ws, "B2", "北海道・東北　案件サマリー（今期案件残／来期計画）")
    note(ws, "B3", f"出所：{SRC_BOOK} シート「{SRC_SHEET}」（第32回・2026/9/4 時点）　単位：件数＝件、金額＝千円")

    # --- 表1 今期の案件残 ---
    r = 5
    title(ws, f"B{r}", "① 今年度（今期）の案件残", size=12)
    note(ws, f"E{r}", "金額＝件数×単価×掛率（原本の算定式と同一）")
    r += 1
    head(ws, r, 2, "エリア")
    head(ws, r, 3, "案件残\n件数")
    head(ws, r, 4, "案件残\n金額")
    head(ws, r, 5, "補正対象\n件数")
    head(ws, r, 6, "補正対象\n金額")
    head(ws, r, 7, "合計\n件数", fill="header")
    head(ws, r, 8, "合計\n金額", fill="header")
    head(ws, r, 9, "全社比\n(金額)")
    ws.row_dimensions[r].height = 30

    z, h = refs["zansu"], refs["hosei"]
    first = r + 1
    for i, (area, key) in enumerate([("北海道", "hk"), ("東北", "tk")]):
        rr = first + i
        label(ws, rr, 2, area, bold=True)
        body(ws, rr, 3, f"={z['sheet']}!{z[key+'_cnt']}", NUM)
        body(ws, rr, 4, f"={z['sheet']}!{z[key+'_amt']}", NUM)
        body(ws, rr, 5, f"={h['sheet']}!{h[key+'_cnt']}", NUM)
        body(ws, rr, 6, f"={h['sheet']}!{h[key+'_amt']}", NUM)
        body(ws, rr, 7, f"=C{rr}+E{rr}", NUM, bold=True, fill="band")
        body(ws, rr, 8, f"=D{rr}+F{rr}", NUM, bold=True, fill="band")
        body(ws, rr, 9, f"=IFERROR(H{rr}/$H${first+3},0)", PCT)

    rr = first + 2
    label(ws, rr, 2, "2エリア計", bold=True, fill="total")
    for col in range(3, 9):
        L = get_column_letter(col)
        body(ws, rr, col, f"=SUM({L}{first}:{L}{first+1})", NUM, bold=True, fill="total")
    body(ws, rr, 9, f"=IFERROR(H{rr}/$H${first+3},0)", PCT, bold=True, fill="total")

    rr = first + 3
    label(ws, rr, 2, "（参考）全社")
    body(ws, rr, 3, f"={z['sheet']}!{z['all_cnt']}", NUM)
    body(ws, rr, 4, f"={z['sheet']}!{z['all_amt']}", NUM)
    body(ws, rr, 5, f"={h['sheet']}!{h['all_cnt']}", NUM)
    body(ws, rr, 6, f"={h['sheet']}!{h['all_amt']}", NUM)
    body(ws, rr, 7, f"=C{rr}+E{rr}", NUM, bold=True)
    body(ws, rr, 8, f"=D{rr}+F{rr}", NUM, bold=True)
    body(ws, rr, 9, 1.0, PCT)

    # --- 表2 来期 ---
    r = first + 6
    title(ws, f"B{r}", "② 来期の案件計画", size=12)
    note(ws, f"E{r}", "原本の来期テーマ別計画（件数・金額とも原本の入力値）")
    r += 1
    head(ws, r, 2, "エリア")
    head(ws, r, 3, "件数")
    head(ws, r, 4, "金額")
    head(ws, r, 5, "平均単価\n(千円/件)")
    head(ws, r, 6, "全社比\n(件数)")
    head(ws, r, 7, "全社比\n(金額)")
    head(ws, r, 8, "要員数\n(名)")
    head(ws, r, 9, "1名あたり\n売上")
    ws.row_dimensions[r].height = 30

    k = refs["raiki"]
    f2 = r + 1
    for i, (area, key) in enumerate([("北海道", "hk"), ("東北", "tk")]):
        rr = f2 + i
        label(ws, rr, 2, area, bold=True)
        body(ws, rr, 3, f"={k['sheet']}!{k[key+'_cnt']}", NUM, bold=True, fill="band")
        body(ws, rr, 4, f"={k['sheet']}!{k[key+'_amt']}", NUM, bold=True, fill="band")
        body(ws, rr, 5, f"=IFERROR(D{rr}/C{rr},0)", NUM)
        body(ws, rr, 6, f"=IFERROR(C{rr}/$C${f2+3},0)", PCT)
        body(ws, rr, 7, f"=IFERROR(D{rr}/$D${f2+3},0)", PCT)
        body(ws, rr, 8, YOIN[area], NUM, src=True)
        body(ws, rr, 9, f"=IFERROR(D{rr}/H{rr},0)", NUM)

    rr = f2 + 2
    label(ws, rr, 2, "2エリア計", bold=True, fill="total")
    for col in (3, 4, 8):
        L = get_column_letter(col)
        body(ws, rr, col, f"=SUM({L}{f2}:{L}{f2+1})", NUM, bold=True, fill="total")
    body(ws, rr, 5, f"=IFERROR(D{rr}/C{rr},0)", NUM, bold=True, fill="total")
    body(ws, rr, 6, f"=IFERROR(C{rr}/$C${f2+3},0)", PCT, bold=True, fill="total")
    body(ws, rr, 7, f"=IFERROR(D{rr}/$D${f2+3},0)", PCT, bold=True, fill="total")
    body(ws, rr, 9, f"=IFERROR(D{rr}/H{rr},0)", NUM, bold=True, fill="total")

    rr = f2 + 3
    label(ws, rr, 2, "（参考）全社")
    body(ws, rr, 3, f"={k['sheet']}!{k['all_cnt']}", NUM)
    body(ws, rr, 4, f"={k['sheet']}!{k['all_amt']}", NUM)
    body(ws, rr, 5, f"=IFERROR(D{rr}/C{rr},0)", NUM)
    body(ws, rr, 6, 1.0, PCT)
    body(ws, rr, 7, 1.0, PCT)
    body(ws, rr, 8, None)
    body(ws, rr, 9, None)

    # --- 表3 参考：今期確保済 ---
    r = f2 + 6
    title(ws, f"B{r}", "③（参考）今期の確保済実績とエリア別見積状況", size=12)
    r += 1
    head(ws, r, 2, "エリア")
    head(ws, r, 3, "随契\n金額")
    head(ws, r, 4, "落札\n金額")
    head(ws, r, 5, "確保済\n計")
    head(ws, r, 6, "見積提示\n今期分")
    head(ws, r, 7, "見積提示\n来期分")
    head(ws, r, 8, "指名停止\n影響件数")
    head(ws, r, 9, "指名停止\n影響額")
    ws.row_dimensions[r].height = 30

    f3 = r + 1
    sh = refs["shimei"]
    area_keys = {"北海道": "hk", "東北": "tk", "関東": "kt", "西日本": "ns"}
    for i, (area, zui, raku) in enumerate(JISSEKI):
        rr = f3 + i
        emph = area in ("北海道", "東北")
        fill = "band" if emph else None
        k = area_keys[area]
        label(ws, rr, 2, area, bold=emph, fill=fill)
        body(ws, rr, 3, zui, NUM, fill=fill, src=True)
        body(ws, rr, 4, raku, NUM, fill=fill, src=True)
        body(ws, rr, 5, f"=C{rr}+D{rr}", NUM, bold=True, fill=fill)
        body(ws, rr, 6, MITSUMORI[area][0], NUM, fill=fill, src=True)
        body(ws, rr, 7, MITSUMORI[area][1], NUM, fill=fill, src=True)
        body(ws, rr, 8, f"={sh['sheet']}!{sh[k + '_cnt']}", NUM, fill=fill)
        body(ws, rr, 9, f"={sh['sheet']}!{sh[k + '_amt']}", NUM, fill=fill)

    rr = f3 + len(JISSEKI)
    label(ws, rr, 2, "全社計", bold=True, fill="total")
    for col in range(3, 10):
        L = get_column_letter(col)
        body(ws, rr, col, f"=SUM({L}{f3}:{L}{rr-1})", NUM, bold=True, fill="total")

    r = rr + 2
    note(ws, f"B{r}", "※ 随契・落札はエリア別の金額のみ原本に記載（件数のエリア別内訳は原本に無し）。全社は随契353件、落札161件。")
    note(ws, f"B{r+3}", "※ 全社の見積提示済は今期分5件・来期分134件、指名停止の影響額は118,200千円で、いずれも原本の集計と一致。")
    note(ws, f"B{r+1}", "※ 指名停止の影響は【指名停止の影響】表の件数に各テーマの単価×掛率を乗じた参考値。北海道は影響無し。")
    note(ws, f"B{r+2}", "※ 東北の当期実損は原本コメントで「5件：△8,000千円（うち3件公会計・須賀川市/関連組合）」と記載。上表はテーマ別集計値。")


# ============================================================
# Sheet 2: 今期案件（案件残・補正対象・指名停止の影響）の明細
# ============================================================
# 共通レイアウト:
#   B=テーマ C=単価 D=掛率 E/F=北海道 G/H=東北 I/J=2エリア計
#   K/L=(参考)関東 M/N=(参考)西日本 O/P=全社
BLOCKS = [
    ("【現在の案件残】", "今期中に決着する未確定案件", ZANSU, ZANSU_GENPON, "z"),
    ("【補正対象】", "補正予算での発注が見込まれる案件（原本上は別枠管理）", HOSEI, HOSEI_GENPON, "h"),
    ("【指名停止の影響】", "指名停止により影響が及びうる案件（参考・案件残とは別集計）", SHIMEI, SHIMEI_GENPON, "s"),
]


def build_zansu(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 20
    for col in "CDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 10.5

    title(ws, "B2", "① 今年度（今期）案件　テーマ別明細")
    note(ws, "B3", f"出所：{SRC_BOOK}「{SRC_SHEET}」【現在の案件残】D32:K38 ／【補正対象】D42:K48 ／【指名停止の影響】C20:K26")
    note(ws, "B4", "単位：金額＝千円　／　青字＝原本からの転記値、黒字＝本表の計算式")

    refs = {"sheet": ws.title}
    row = 6
    for block_name, desc, data, genpon, tag in BLOCKS:
        title(ws, f"B{row}", block_name, size=12)
        note(ws, f"E{row}", desc)
        row += 1
        for col, text in [
            (2, "テーマ"), (3, "単価\n(千円)"), (4, "掛率"),
            (5, "北海道\n件数"), (6, "北海道\n金額"),
            (7, "東北\n件数"), (8, "東北\n金額"),
            (9, "2エリア\n件数"), (10, "2エリア\n金額"),
            (11, "(参考)\n関東件数"), (12, "(参考)\n関東金額"),
            (13, "(参考)\n西日本件数"), (14, "(参考)\n西日本金額"),
            (15, "全社\n件数"), (16, "全社\n金額"),
        ]:
            head(ws, row, col, text, fill="header" if col in (5, 6, 7, 8) else "subhead")
        ws.row_dimensions[row].height = 30

        first = row + 1
        for i, (theme, hk, tk, kt, ns, unit, rate) in enumerate(data):
            rr = first + i
            fill = "alt" if i % 2 else None
            label(ws, rr, 2, theme, fill=fill)
            body(ws, rr, 3, unit, NUM, fill=fill, src=True)
            body(ws, rr, 4, rate, "0.0", fill=fill, src=True)
            for cnt_col, cnt in ((5, hk), (7, tk), (11, kt), (13, ns)):
                L = get_column_letter(cnt_col)
                body(ws, rr, cnt_col, cnt, NUM, fill=fill, src=True)
                body(ws, rr, cnt_col + 1, f"=ROUND({L}{rr}*$C{rr}*$D{rr},0)", NUM, fill=fill)
            body(ws, rr, 9, f"=E{rr}+G{rr}", NUM, fill=fill)
            body(ws, rr, 10, f"=F{rr}+H{rr}", NUM, fill=fill)
            body(ws, rr, 15, f"=E{rr}+G{rr}+K{rr}+M{rr}", NUM, fill=fill)
            body(ws, rr, 16, f"=F{rr}+H{rr}+L{rr}+N{rr}", NUM, fill=fill)

        last = first + len(data) - 1
        rr = last + 1
        label(ws, rr, 2, "計", bold=True, fill="total")
        body(ws, rr, 3, None, fill="total")
        body(ws, rr, 4, None, fill="total")
        for col in range(5, 17):
            L = get_column_letter(col)
            body(ws, rr, col, f"=SUM({L}{first}:{L}{last})", NUM, bold=True, fill="total")

        for key, cell in [("hk_cnt", f"$E${rr}"), ("hk_amt", f"$F${rr}"),
                          ("tk_cnt", f"$G${rr}"), ("tk_amt", f"$H${rr}"),
                          ("kt_cnt", f"$K${rr}"), ("kt_amt", f"$L${rr}"),
                          ("ns_cnt", f"$M${rr}"), ("ns_amt", f"$N${rr}"),
                          ("all_cnt", f"$O${rr}"), ("all_amt", f"$P${rr}")]:
            refs[tag + key] = cell

        # 原本との照合（本表の全社計が原本の集計と一致することの検証）
        gr = rr + 1
        label(ws, gr, 2, "原本の全社計")
        body(ws, gr, 15, genpon[0], NUM, src=True)
        body(ws, gr, 16, genpon[1], NUM, src=True)
        dr = gr + 1
        label(ws, dr, 2, "差異（0であること）", bold=True)
        body(ws, dr, 15, f"=O{rr}-O{gr}", NUM, bold=True)
        body(ws, dr, 16, f"=P{rr}-P{gr}", NUM, bold=True)
        row = dr + 2

    note(ws, f"B{row}", "※ 単価／掛率が空欄の行は原本に設定が無いもの（会計支援は全ブロック、その他は案件残・補正対象ブロック）。件数のみ計上し金額は0。")
    note(ws, f"B{row+1}", "※ 金額は原本と同じ ROUND(件数×単価×掛率, 0)。原本はエリア別金額を持たず全社計のみを算定しているため、")
    note(ws, f"B{row+2}", "　 本表では同じ単価・掛率をエリア別件数に適用して展開した。各ブロックの「差異」行が0であることで原本との整合を確認できる。")
    note(ws, f"B{row+3}", "※ 掛率は原本で設定された受注確度。したがって金額は「確度加味後の見込」であり、満額ベースではない。")
    note(ws, f"B{row+4}", f"※ 先週（{PREV_SHEET}）比：北海道・東北とも案件残・補正対象に増減なし。全社の減（35→31件）は関東によるもの。")
    note(ws, f"B{row+5}", "※ 指名停止の影響は案件残・補正対象とは別集計であり、①②の合計には含めていない。")
    return refs


# ============================================================
# Sheet 3: 来期案件の明細
# ============================================================
def build_raiki(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 13

    title(ws, "B2", "② 来期案件計画　テーマ別明細")
    note(ws, "B3", f"出所：{SRC_BOOK}「{SRC_SHEET}」F123:P132（来期テーマ別 件数・金額）")
    note(ws, "B4", "単位：金額＝千円　／　青字＝原本からの転記値、黒字＝本表の計算式")

    row = 5
    head(ws, row, 2, "テーマ")
    head(ws, row, 3, "北海道\n件数", fill="header")
    head(ws, row, 4, "北海道\n金額", fill="header")
    head(ws, row, 5, "東北\n件数", fill="header")
    head(ws, row, 6, "東北\n金額", fill="header")
    head(ws, row, 7, "2エリア\n件数")
    head(ws, row, 8, "2エリア\n金額")
    head(ws, row, 9, "全社\n件数")
    head(ws, row, 10, "全社\n金額")
    head(ws, row, 11, "全社比\n(金額)")
    ws.row_dimensions[row].height = 30

    first = row + 1
    for i, (theme, hc, ha, tc, ta, ac, aa) in enumerate(RAIKI):
        rr = first + i
        fill = "alt" if i % 2 else None
        label(ws, rr, 2, theme, fill=fill)
        body(ws, rr, 3, hc, NUM, fill=fill, src=True)
        body(ws, rr, 4, ha, NUM, fill=fill, src=True)
        body(ws, rr, 5, tc, NUM, fill=fill, src=True)
        body(ws, rr, 6, ta, NUM, fill=fill, src=True)
        body(ws, rr, 7, f"=C{rr}+E{rr}", NUM, fill=fill)
        body(ws, rr, 8, f"=D{rr}+F{rr}", NUM, fill=fill)
        body(ws, rr, 9, ac, NUM, fill=fill, src=True)
        body(ws, rr, 10, aa, NUM, fill=fill, src=True)
        body(ws, rr, 11, f"=IFERROR(H{rr}/J{rr},0)", PCT, fill=fill)

    last = first + len(RAIKI) - 1
    rr = last + 1
    label(ws, rr, 2, "合計", bold=True, fill="total")
    for col in range(3, 11):
        L = get_column_letter(col)
        body(ws, rr, col, f"=SUM({L}{first}:{L}{last})", NUM, bold=True, fill="total")
    body(ws, rr, 11, f"=IFERROR(H{rr}/J{rr},0)", PCT, bold=True, fill="total")

    # 平均単価行
    rr2 = rr + 1
    label(ws, rr2, 2, "平均単価(千円/件)", bold=True)
    body(ws, rr2, 3, None)
    body(ws, rr2, 4, f"=IFERROR(D{rr}/C{rr},0)", NUM, bold=True)
    body(ws, rr2, 5, None)
    body(ws, rr2, 6, f"=IFERROR(F{rr}/E{rr},0)", NUM, bold=True)
    body(ws, rr2, 7, None)
    body(ws, rr2, 8, f"=IFERROR(H{rr}/G{rr},0)", NUM, bold=True)
    body(ws, rr2, 9, None)
    body(ws, rr2, 10, f"=IFERROR(J{rr}/I{rr},0)", NUM, bold=True)
    body(ws, rr2, 11, None)

    r = rr2 + 2
    note(ws, f"B{r}", "※ 件数・金額とも原本の入力値をそのまま採用（原本合計：北海道 280件/270,000千円、東北 172件/190,000千円）。")
    note(ws, f"B{r+1}", "※ 全社の来期計画は 712件／850,000千円。今期見込売上 700,000千円に対し +150,000千円（+21.4%）の計画。")
    note(ws, f"B{r+2}", "※ 来期要員計画は北海道12名・東北6名（原本 G133／I133）。1名あたり売上は北海道22,500千円・東北31,667千円。")

    return {"sheet": ws.title,
            "hk_cnt": f"$C${rr}", "hk_amt": f"$D${rr}",
            "tk_cnt": f"$E${rr}", "tk_amt": f"$F${rr}",
            "all_cnt": f"$I${rr}", "all_amt": f"$J${rr}"}


# ============================================================
# Sheet 4: 前提・注記
# ============================================================
def build_notes(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96

    title(ws, "B2", "前提・出所・注記")

    rows = [
        ("元データ", f"{SRC_BOOK}（週次営業会議シート）"),
        ("基準シート", f"「{SRC_SHEET}」＝第32回・2026年9月4日時点（ブック内で最新の回）"),
        ("比較シート", f"「{PREV_SHEET}」＝第31回・2026年8月28日時点（先週比の算定に使用）"),
        ("単位", "件数＝件、金額＝千円"),
        ("今期案件残の定義", "【現在の案件残】＝入札・提案の結果が未確定で今期中に決着する案件。"
                             "【補正対象】＝補正予算での発注が見込まれる案件で、原本上は別枠管理。"),
        ("金額の算定", "原本はエリア別の金額を持たず全社計のみを ROUND(件数×単価×掛率,0) で算定している。"
                       "本サマリーは同じ単価・掛率をエリア別件数に適用して展開した。"
                       "4エリア合計は原本の全社計（案件残32,800千円／補正対象28,500千円）と一致することを確認済み。"),
        ("掛率の意味", "原本で設定された受注確度。公会計0.7／経営戦略0.8／経営改善0.7／公共施設M0.5／福祉計画0.5。"
                       "したがって案件残金額は「確度加味後の見込売上」であり、満額ベースではない。"),
        ("来期計画", "原本 F123:P132 のテーマ別 件数・金額をそのまま採用。原本上は「※イメージ」と付記された計画値。"),
        ("今期確保済実績", "随契は原本 G7:J7、落札は G9:J9（いずれもエリア別金額）。件数のエリア別内訳は原本に無し。"),
        ("見積提示済件数", "エリア別明細ブロック（北海道 H172:I172／東北 H186:I186）の合計行。今期分・来期分の区分は原本どおり。"),
        ("指名停止の影響", "【指名停止の影響】C20:K26 のエリア別件数に各テーマの単価×掛率を適用した参考値。"
                           "北海道は原本コメントで「無し」。東北の当期実損は原本コメントで「5件：△8,000千円」と記載されており、"
                           "本表の14件／14,100千円は影響が及びうる案件の総量を指す。"),
        ("留意点", "原本は週次で上書き更新されるため、本サマリーは2026/9/4時点のスナップショット。"),
    ]
    r = 4
    for k, v in rows:
        label(ws, r, 2, k, bold=True, fill="band")
        c = body(ws, r, 3, v, align="left")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 15 * (1 + len(v) // 55)
        r += 1


# ============================================================
def page_setup(ws, freeze=None):
    """横向き・幅1ページに収める印刷設定。"""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    if freeze:
        ws.freeze_panes = freeze


def main():
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "サマリー"
    ws_z = wb.create_sheet("今期_案件残明細")
    ws_r = wb.create_sheet("来期_案件明細")
    ws_n = wb.create_sheet("前提・注記")

    zrefs = build_zansu(ws_z)
    rrefs = build_raiki(ws_r)

    def block(tag):
        d = {"sheet": ws_z.title}
        for key in ("hk_cnt", "hk_amt", "tk_cnt", "tk_amt",
                    "kt_cnt", "kt_amt", "ns_cnt", "ns_amt", "all_cnt", "all_amt"):
            d[key] = zrefs[tag + key]
        return d

    build_summary(ws_sum, {"zansu": block("z"), "hosei": block("h"),
                           "shimei": block("s"), "raiki": rrefs})
    build_notes(ws_n)

    page_setup(ws_sum)
    page_setup(ws_z, freeze="C1")
    page_setup(ws_r, freeze="C1")
    page_setup(ws_n)

    wb.active = 0
    wb.save(OUT_PATH)
    print(f"saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
