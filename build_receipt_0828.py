# -*- coding: utf-8 -*-
"""令和8年8月28日受領資料の点検結果.

発注者より9件の資料を受領した。内容を点検し、
  ① どの資料提供依頼・確認事項が解消したか
  ② 既存の成果品のどこを更新するか
  ③ 新たに判明した事実
を整理する。

シート構成
  00_受領資料の一覧と点検の結果
  01_交付金評価　3町別・3か年
  02_要介護認定の申請と認定の状況
  03_成果品への反映箇所
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_kofukin import KOF                       # noqa: E402
import data_nintei_shinsei as NS                   # noqa: E402

ODIR = "/home/user/repository/output"
OUT = os.path.join(ODIR, "第10期計画_令和8年8月28日受領資料の点検結果.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TOWNS = ("東川町", "美瑛町", "東神楽町")
YEARS = ("R6", "R7", "R8")
YLAB = {"R6": "令和6年度", "R7": "令和7年度", "R8": "令和8年度"}

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 50
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=30):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=26, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=84):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


def grp(t, y, kind):
    """指標群別の小計。kind は '体制'（ⅰ）'活動'（ⅱ）'成果'（Ⅳ）。"""
    d = KOF[t][y]
    if kind == "成果":
        return d.get("Ⅳ合計", 0)
    suf = "（ⅰ）計" if kind == "体制" else "（ⅱ）計"
    sui = sum(d.get(r + suf, 0) for r in "ⅠⅡⅢ")
    sien = sum(d.get(r + suf + "_2", 0) for r in "ⅠⅡⅢ")
    return sui, sien


# ============================================================ 00
ws = sheet("00_受領資料の一覧", "令和8年8月28日受領資料の点検結果",
           "発注者より9件の資料を受領しました。"
           "内容を点検し、解消した資料提供依頼・確認事項、"
           "既存の成果品の更新箇所、新たに判明した事実を整理します。",
           [5, 34, 12, 40, 40])

r = header(ws, 4, ["No.", "受領資料", "形式", "内容", "点検の結果"])
ITEMS = [
    (1, "保険者機能強化推進交付金等（市町村）に係る全国集計結果"
        "　令和6年度", "ZIP 15件",
     "全国1,570余保険者の評価結果、指標群別都道府県平均、"
     "評価指標の定義（PDF）、上位50位の集計表。",
     "資料No.5（評価調書）の代替として使えます。"
     "構成3町の項目別得点が判明しました。"),
    (2, "同　令和7年度", "ZIP 9件", "同上。",
     "同上。3か年の推移が得られました。"),
    (3, "同　令和8年度", "ZIP 9件", "同上（参考資料はPDF）。", "同上。"),
    (4, "要介護認定 集計帳票　令和6年度分", "PDF 14頁",
     "申請件数（被保険者区分別・年齢階級別）、認定件数"
     "（年齢階級別・二次判定別・申請区分別）、"
     "重軽度変更率、状態区分変化率、認定所要日数、特定疾病別。",
     "資料No.19（区分変更の申請件数と変更前後の要介護度）が解消します。"
     "代表KPI H03の算定にも用いられます。"),
    (5, "同　令和7年度分", "PDF 14頁", "同上。", "同上。"),
    (6, "同　令和8年度分（令和8年8月7日時点）", "PDF 14頁",
     "同上。ただし令和8年4月1日から8月7日までの年度途中の集計。",
     "年度途中のため通年の値ではありません。"
     "年度間の比較には用いません。"),
    (7, "在宅医療を行っている医療機関リスト", "Excel 1シート",
     "北海道全域。医療機関名、住所、往診の可否、"
     "訪問診療、看取り等の対応状況。",
     "在宅医療・介護連携（基本目標1）の社会資源の把握に用います。"
     "区域内の医療機関を抽出して整理します。"),
    (8, "在宅訪問対応薬局リスト", "Excel 1シート",
     "北海道全域。薬局名、住所、無菌製剤処理、一包化、"
     "麻薬の取扱い等の対応状況。", "同上。"),
    (9, "介護予防・日常生活支援総合事業等（従前相当サービス）"
        "の指標値（平成27年〜令和元年）", "Excel 5シート",
     "全国の保険者別。短期集中リハビリテーション実施加算等の"
     "ストラクチャー指標、認定者1万人あたりの指標値。",
     "平成27年から令和元年の5か年であり、第9期（令和6〜8年度）の"
     "評価には直接用いられません。"
     "長期の推移の参考として整理します。"),
]
for a in ITEMS:
    f = OK_G if "解消" in a[4] or "使えます" in a[4] else (
        IN_Y if "参考" in a[4] or "ではありません" in a[4] else MID_B)
    r = body(ws, r, list(a), {5: f}, height=62, align={1: "center", 3: "center"})

r += 1
r = lead(ws, r, "【解消する資料提供依頼・確認事項】", 5)
r = header(ws, r, ["No.", "依頼・確認事項", "状態", "受領資料", "残る留保"])
for a in [
    ("依頼5", "保険者機能強化推進交付金等の評価調書（令和4〜7年度分）",
     "代替により解消",
     "全国集計結果（令和6〜8年度）",
     "評価調書そのものではないため、"
     "0点の項目が「未実施」か「要件未充足」かの別は依然として分かりません。"
     "ただし、記録・報告の不備によるものではないことは確認できました"
     "（保険者の回答が国の集計に反映されているため）。"),
    ("依頼19", "区分変更の申請件数と変更前後の要介護度（令和元〜令和7年度）",
     "一部解消",
     "要介護認定 集計帳票 3-7（前回二次判定別・二次判定別）",
     "令和6・7年度分のみ。令和元〜5年度分は未受領です。"),
    ("確認No.38", "交付金評価の最新調査",
     "解消", "全国集計結果（令和8年度＝令和8年度評価指標）",
     "―"),
]:
    r = body(ws, r, list(a), {3: OK_G}, height=74, align={1: "center", 3: "center"})

r += 1
r = note(ws, r,
         "注1）本点検は受領資料の内容の確認であり、"
         "計画本文への反映は03シートに掲げる箇所について順次行います。\n"
         "注2）医療機関リスト・薬局リストは北海道全域のデータです。"
         "区域内の抽出と、介護サービス情報公表システムの事業所一覧との"
         "突合は別途行います。\n"
         "注3）令和8年8月26日の打合せでご依頼した年報・月報"
         "（資料No.21・22）及び給付実績データ（同No.23）は、"
         "本受領には含まれていません。", 5, height=84)

# ============================================================ 01
ws = sheet("01_交付金評価", "保険者機能強化推進交付金等の評価　構成3町別・3か年",
           "交付金は市町村に交付されるため、評価は構成3町ごとに行われています。"
           "公表資料の大雪地区広域連合の行に得点はなく、"
           "3町それぞれの行に得点が記載されています。"
           "これまで「見える化」システムにより保険者単位で把握していた値を、"
           "町別・年度別に分解できるようになりました。",
           [10, 8, 10, 10, 10, 10, 10, 10, 10, 44])

r = lead(ws, 4, "【1　合計得点と全国順位】", 10)
r = header(ws, r, ["町", "年度", "推進\n合計", "支援\n合計", "推進・支援\n合計",
                   "全国\n順位", "", "", "", "所見"])
for t in TOWNS:
    for y in YEARS:
        d = KOF[t][y]
        r = body(ws, r, [t if y == "R6" else "", YLAB[y],
                         d.get("推進合計"), d.get("支援合計"),
                         d.get("推進・支援合計"), d.get("今年度順位"),
                         "", "", "",
                         "" if y != "R8" else
                         "3か年で合計得点は%s、順位は%s。"
                         % ("ほぼ横ばい" if abs(
                             (KOF[t]["R8"].get("推進・支援合計", 0)
                              - KOF[t]["R6"].get("推進・支援合計", 0))) <= 5
                            else "変動", "低下")],
                 {1: MID_B if y == "R6" else None} if y == "R6" else {},
                 height=20,
                 align={2: "center", 3: "right", 4: "right", 5: "right",
                        6: "right"},
                 bold=(y == "R8"))
    r += 0

r += 1
r = lead(ws, r, "【2　指標群別の得点】", 10)
r = header(ws, r, ["町", "年度", "推進\n体制取組", "推進\n活動", "推進\n成果",
                   "支援\n体制取組", "支援\n活動", "支援\n成果", "", "所見"])
for t in TOWNS:
    for y in YEARS:
        ti, si = grp(t, y, "体制")
        ta, sa = grp(t, y, "活動")
        se = KOF[t][y].get("Ⅳ合計", "")
        obs = ""
        if t == "美瑛町" and y == "R6":
            obs = "推進の活動指標群が3目標とも0点"
        if t == "東神楽町" and y == "R8":
            obs = "支援の活動指標群が3町で最も高い"
        r = body(ws, r, [t if y == "R6" else "", YLAB[y], ti, ta, se,
                         si, sa, se, "", obs],
                 {4: NG_O if ta == 0 else None, 7: NG_O if sa == 0 else None},
                 height=20,
                 align={2: "center", 3: "right", 4: "right", 5: "right",
                        6: "right", 7: "right", 8: "right"})

r += 1
r = lead(ws, r, "【3　この資料から分かったこと】", 10)
for a in [
    ("① 評価は3町別に行われている",
     "保険者は大雪地区広域連合ですが、交付金は市町村に交付されるため、"
     "評価結果は3町それぞれに記録されています。"
     "公表資料の全1,564列のうち約1,456列（93％）は3町で同一の値であり、"
     "これは保険者としての共通の取組と考えられます。"
     "残る約108列が町ごとに異なります。"),
    ("② 成果指標群は3町とも同一",
     "目標Ⅳ（成果指標群）は3町とも同じ得点です"
     "（令和6年度60点、令和7・8年度55点）。"
     "要介護認定率等の成果は保険者単位で算定されるためと考えられます。"
     "計画素案の代表KPI H01（見える化W144）の扱いと矛盾しません。"),
    ("③ 活動指標群に町間の差がある",
     "推進の活動指標群は、令和6年度で東川町12点・美瑛町0点・東神楽町12点、"
     "令和8年度で東川町28点・美瑛町16点・東神楽町25点です。"
     "支援の活動指標群も東神楽町が3町で最も高くなっています。"
     "第10期の3町との役割分担（資料編 資料2）を検討する際の材料になります。"),
    ("④ 合計得点はほぼ横ばい、全国順位は低下",
     "3町とも合計得点は3か年でほぼ横ばいですが、全国順位は低下しています"
     "（東川町1,476→1,625位、美瑛町1,492→1,639位、"
     "東神楽町1,243→1,377位）。"
     "他の保険者の得点が上昇しているためと考えられます。"),
    ("⑤ 「記録・報告の不備」ではないことが確認できた",
     "中間報告では、活動指標群の得点が低い理由を"
     "「未実施・要件未充足・記録又は報告の不備のいずれか」として"
     "未確認としていました。"
     "国の集計に保険者の回答が反映されている以上、"
     "報告そのものが行われていないという説明は成り立ちません。"
     "残るのは「未実施」か「要件未充足」かの別です。"),
]:
    r = body(ws, r, [a[0], "", "", "", "", "", "", "", "", a[1]],
             {1: MID_B}, height=64)
    ws.merge_cells(start_row=r - 1, start_column=1, end_row=r - 1, end_column=2)
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=9)

r += 1
r = note(ws, r,
         "出典　厚生労働省「保険者機能強化推進交付金等（市町村分）に係る"
         "全国集計結果」令和6年度・令和7年度・令和8年度。\n"
         "注1）各年度の交付金は当該年度の評価指標により算定されます。\n"
         "注2）令和5年度評価指標の合計得点（東川町863点・美瑛町868点・"
         "東神楽町868点）は指標の体系が異なるため、"
         "令和6年度以降の得点と直接比較できません。", 10, height=64)

# ============================================================ 02
ws = sheet("02_要介護認定の状況", "要介護認定の申請と認定の状況",
           "認定支援ネットワークの集計帳票により、"
           "申請件数・認定件数・所要日数を年度別に把握しました。"
           "令和8年度分は令和8年4月1日から8月7日までの集計であり、"
           "通年の値ではありません。",
           [22, 14, 14, 16, 8, 60])

r = lead(ws, 4, "【1　申請件数（年齢階級別）】", 6)
r = header(ws, r, ["区分", "令和6年度", "令和7年度",
                   "令和8年度\n（8月7日まで）", "", "所見"])
LAB = ["合計", "65歳未満", "65〜69歳", "70〜74歳", "75〜79歳", "80〜84歳",
       "85〜89歳", "90〜94歳", "95〜99歳", "100歳以上", "65歳以上（再掲）",
       "65〜74歳（再掲）", "75歳以上（再掲）"]
for i, lab in enumerate(LAB):
    v = [NS.AGE_APP[y][i] if i < len(NS.AGE_APP[y]) else "" for y in YEARS]
    obs = ""
    if lab == "合計":
        obs = ("令和6年度1,427件、令和7年度1,527件で100件（7.0％）増加。"
               "令和8年度は年度途中のため比較できません。")
    if lab == "75歳以上（再掲）":
        obs = ("申請の87.4〜88.0％が75歳以上です。"
               "代表KPI H03（75歳以上新規認定率）の分母の把握に用います。")
    r = body(ws, r, [lab] + v + ["", obs],
             {1: MID_B if lab == "合計" else None}, height=18,
             align={2: "right", 3: "right", 4: "right"},
             bold=(lab == "合計"))

r += 1
r = lead(ws, r, "【2　認定件数（申請区分別）】", 6)
r = header(ws, r, ["区分", "令和6年度", "令和7年度",
                   "令和8年度\n（8月7日まで）", "", "所見"])
for k in ("新規", "更新"):
    v = [NS.NINTEI[y][k][-1] for y in YEARS]
    r = body(ws, r, [k + "　認定件数"] + v + ["", ""], {}, height=18,
             align={2: "right", 3: "right", 4: "right"})
r = body(ws, r, ["新規のうち要支援1・2"] +
         [NS.NINTEI[y]["新規"][1] + NS.NINTEI[y]["新規"][2] for y in YEARS] +
         ["", "新規認定の3割強が要支援です。"
          "介護予防・生活支援の対象規模を示します。"], {}, height=30,
         align={2: "right", 3: "right", 4: "right"})
r = body(ws, r, ["新規のうち要介護3以上"] +
         [sum(NS.NINTEI[y]["新規"][5:8]) for y in YEARS] +
         ["", "新規時点で中重度の者。重度化防止の起点の把握に用います。"],
         {}, height=30, align={2: "right", 3: "right", 4: "right"})

r += 1
r = lead(ws, r, "【3　認定に要する日数】", 6)
r = header(ws, r, ["区分", "令和6年度", "令和7年度",
                   "令和8年度\n（8月7日まで）", "", "所見"])
for k in ("新規", "更新", "区分変更"):
    r = body(ws, r, ["申請日から認定調査まで　" + k] +
             ["%.1f日" % NS.DAYS_SHINSA[y][k] for y in YEARS] + ["", ""],
             {}, height=18, align={2: "right", 3: "right", 4: "right"})
for k in ("新規", "更新", "区分変更"):
    v = [NS.DAYS_HANTEI[y][k] for y in YEARS]
    obs = ""
    if k == "新規":
        obs = ("介護保険法第27条第11項は、申請から30日以内の処分を定めています。"
               "3年度とも平均が30日を超えています。"
               "令和8年度（年度途中）は更に長くなっています。")
    r = body(ws, r, ["申請日から二次判定まで　" + k] +
             ["%.1f日" % x for x in v] + ["", obs],
             {2: NG_O, 3: NG_O, 4: NG_O}, height=34 if obs else 18,
             align={2: "right", 3: "right", 4: "right"})

r += 1
r = note(ws, r,
         "注1）申請日から主治医意見書入手までの日数は、"
         "3年度とも入手日が未入力（集計対象外）であるため算定できません。\n"
         "注2）認定に要する日数は、保険者機能強化推進交付金の"
         "評価指標にも関係します。"
         "本件を第9期の評価及び第10期の課題として扱うかは、"
         "発注者のご判断をお願いします（確認事項として追加します）。\n"
         "注3）令和8年度分は年度途中の集計です。"
         "申請件数206件は、令和7年度の月平均127件に対して低い水準ですが、"
         "データの入力の遅れによる可能性があり、"
         "年度間の比較には用いません。", 6, height=84)

# ============================================================ 03
ws = sheet("03_成果品への反映", "成果品への反映箇所",
           "受領資料により更新する箇所と、その時期を示します。",
           [6, 34, 40, 30, 14])

r = header(ws, 4, ["No.", "成果品・箇所", "更新の内容", "根拠", "時期"])
for a in [
    (1, "第9期計画の評価・検証 中間報告　第3節2（交付金評価）",
     "1時点（令和5年度評価指標）の記述を、3か年（令和6〜8年度評価指標）の"
     "推移に改める。あわせて3町別の得点を示す。",
     "全国集計結果 令和6〜8年度", "令和8年9月"),
    (2, "同　第3節2（活動指標群の理由）",
     "「未実施・要件未充足・記録又は報告の不備のいずれか」から、"
     "「記録・報告の不備によるものではない。未実施か要件未充足かの別は"
     "評価調書の確認による」に改める。",
     "同上", "令和8年9月"),
    (3, "妥当性検証報告書 23_交付金評価の分析",
     "3か年・3町別の得点表を追加する。"
     "指標群別の推移を図示する。", "同上", "令和8年9月"),
    (4, "計画素案 第3章第4節（交付金評価）",
     "同上。3町の役割分担の議論に接続する。", "同上", "令和8年9月"),
    (5, "計画素案 資料編 資料2（3町との役割分担・共通指標）",
     "活動指標群の町間の差を、共通指標の設定の根拠として示す。",
     "同上", "令和8年10月"),
    (6, "代表KPI H03（75歳以上新規認定率）",
     "分母・分子の算定方法を確定する。"
     "認定帳票1-2（年齢階級別申請件数）と3-1（年齢階級別認定件数）により"
     "算定できるかを検証する。",
     "要介護認定 集計帳票 令和6・7年度分", "令和8年9月"),
    (7, "必要事項の一覧 01_資料の提供",
     "資料No.5（評価調書）を「代替により解消」、"
     "資料No.19（区分変更）を「一部解消」に改める。",
     "本点検", "反映済み"),
    (8, "計画素案 第2章（在宅医療・介護連携）",
     "区域内の在宅医療を行う医療機関数・在宅訪問対応薬局数を"
     "社会資源として記載する。",
     "医療機関リスト・薬局リスト", "令和8年9月"),
    (9, "確認事項（新規）",
     "認定に要する日数（申請から二次判定まで平均45〜47日）を"
     "第9期の評価及び第10期の課題として扱うか。",
     "要介護認定 集計帳票 4-1", "ご確認をお願いします"),
]:
    f = {5: OK_G if a[4] == "反映済み" else (
        NG_O if "ご確認" in a[4] else IN_Y)}
    r = body(ws, r, list(a), f, height=54, align={1: "center", 5: "center"})

r += 1
r = note(ws, r,
         "注）本シートの更新は、令和8年9月の作業として行います。"
         "資料No.13（施策・事業実績）の受領によるプロセス評価と"
         "並行して進めます。", 5, height=40)

wb.save(OUT)
print("saved:", os.path.basename(OUT), "sheets=%d" % len(wb.sheetnames))
for t in TOWNS:
    print("  %-5s %s" % (t, " → ".join(
        "%s %s点(%s位)" % (YLAB[y], KOF[t][y].get("推進・支援合計"),
                          KOF[t][y].get("今年度順位")) for y in YEARS)))
print("  認定申請 %s" % " → ".join(
    "%s %d件" % (YLAB[y], NS.AGE_APP[y][0]) for y in YEARS))
print("  申請から二次判定まで（新規） %s" % " → ".join(
    "%s %.1f日" % (YLAB[y], NS.DAYS_HANTEI[y]["新規"]) for y in YEARS))
