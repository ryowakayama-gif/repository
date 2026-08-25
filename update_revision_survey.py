# -*- coding: utf-8 -*-
"""
環境基本計画 営業先リスト 改定時期調査の反映（2026-08-25 第2回）

対象:
  02_営業先マスタで終期が未確認だった先のうち、義務団体（道県・中核市）と
  主要市を優先に、計画期間を調査した結果を反映する。

方針:
  - 公表情報で終期を特定できたものだけを「確認済」にする
  - 策定年・計画名のみ判明したものは「一部確認」とし、終期欄は空欄のまま
  - 推測で終期を埋めない

使い方:
  python3 update_revision_survey.py            # 日本語ファイル名 + cp932 CSV
  python3 update_revision_survey.py --ascii    # 半角英数ファイル名
"""

import copy as _copy
import os
import shutil
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import naming

OUT_DIR = "/home/user/repository/output"
CSV_DIR = os.path.join(OUT_DIR, "csv")
BOOK_STEM_JP = "環境基本計画_営業先リスト_北海道東北"
SRC = os.path.join(OUT_DIR, f"{BOOK_STEM_JP}.xlsx")
SURVEY_SHEET = "08_改定時期調査_20260825b"

C_TITLE, C_SUBHEAD, C_BAND, C_ALT, C_NOTE, C_WHITE = "1F3864", "2E75B6", "DDEBF7", "F7FAFC", "FFF3F3", "FFFFFF"
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name="Yu Gothic", size=14, bold=True, color=C_WHITE)
F_HEAD = Font(name="Yu Gothic", size=10, bold=True, color=C_WHITE)
F_BODY = Font(name="Yu Gothic", size=10)
F_BOLD = Font(name="Yu Gothic", size=10, bold=True)
F_NOTE = Font(name="Yu Gothic", size=9, color="7F3F3F")
AL_WRAP = Alignment(vertical="top", wrap_text=True)
AL_CTR = Alignment(vertical="center", horizontal="center", wrap_text=True)

ASCII = naming.ascii_mode()
ENC = naming.csv_encoding()

# ============================================================
# 調査結果
#   (計画名, 開始, 終期, R9判定, 確認状況, 出典・確認メモ, 接触優先度, 次アクション)
# ============================================================
FINDINGS = {
    # --- 終期を特定できたもの ---
    "青森県|青森県": (
        "第6次青森県環境計画（青森県環境総合プラン）", "令和6年度(2024)", "令和10年度(2028)",
        "非該当", "確認済",
        "青森県 第6次青森県環境計画（令和6年3月）。計画期間は令和6〜10年度の5年間",
        "中（R10満了・R9夏に予算要求期）",
        "令和10年度満了のため改定業務は令和10年度。県への接触は令和9年度の夏〜秋。並行して県の市町村支援メニューを確認する"),
    "宮城県|宮城県": (
        "宮城県環境基本計画（第4期）", "令和3年度(2021)", "令和12年度(2030)",
        "非該当", "確認済",
        "宮城県公式サイト 環境基本計画（第4期）。計画期間は令和3年度〜令和12年度",
        "中（県ルート）",
        "令和12年度満了。期間中は県の市町村支援・共同策定の枠組みの有無を確認する"),
    "山形県|山形県": (
        "第4次山形県環境計画", "令和3年度(2021)", "令和12年度(2030)",
        "非該当", "確認済",
        "山形県公式サイト 第4次山形県環境計画（令和3年3月策定・10年間）。従前リストの『第3次』は平成23年度〜令和2年度で終了済のため計画名を訂正",
        "中（県ルート）",
        "令和12年度満了。期間中は県の市町村支援・共同策定の枠組みの有無を確認する"),
    "福島県|福島市": (
        "第3次福島市環境基本計画", "令和3年度(2021)", "令和7年度(2025)",
        "満了済（第4次の期間を要確認）", "確認済",
        "福島市公式サイト 環境基本計画（令和3年2月策定）。計画期間は令和3年度〜令和7年度の5年間。5年周期のため次の満了も早く来る",
        "優先（義務団体・満了済）",
        "令和7年度末で満了済。第4次の策定状況を早急に確認する。策定済みなら次期（令和12年度頃）に向けた関係構築、未策定なら補正・当初での策定提案"),
    "岩手県|一関市": (
        "一関市環境基本計画", "平成29年度(2017)", "令和8年度(2026)",
        "◇今年度が改定年度（R8満了）", "確認済",
        "一関市 生活環境課 環境基本計画。計画期間は平成29年度から令和8年度。なお市の地球温暖化対策地域推進計画は令和5年度〜令和12年度",
        "最優先（今年度満了）",
        "令和8年度末で満了＝次期計画は令和9年度開始。改定業務は今年度に実施されるため、既発注か自前策定かを直ちに確認する。発注済みなら範囲外の支援に絞る"),
    "福島県|会津若松市": (
        "会津若松市第3期環境基本計画", "令和6年度(2024)", "令和12年度(2030)",
        "非該当", "確認済",
        "会津若松市 第3期環境基本計画（令和6年3月策定・7年間）",
        "低（期間先）", "令和12年度満了。次期改定は令和11年度頃から予算化を確認する"),
    "北海道|釧路市": (
        "第2次釧路市環境基本計画【改定版】", "令和3年度(2021)", "令和12年度(2030)",
        "非該当（令和6年3月に一部改定）", "確認済",
        "釧路市 第2次環境基本計画【改定版】。計画期間は令和3年度〜令和12年度。国・北海道の温暖化計画改定を受け、低炭素社会の形成の目標を令和6年3月に改定",
        "中（期間中改定の実例）",
        "計画期間中だが低炭素分野を一部改定した実績がある。国の2035年度目標を踏まえた再度の部分改定の余地を確認する"),
    "北海道|千歳市": (
        "第3次千歳市環境基本計画", "令和3年度(2021)", "令和12年度(2030)",
        "非該当", "確認済",
        "千歳市 第3次環境基本計画（令和3年度〜令和12年度）",
        "低（期間先）", "令和12年度満了。次期改定は令和11年度頃から予算化を確認する"),
    "北海道|石狩市": (
        "石狩市環境基本計画（第3次）", "令和5年度(2023)", "令和24年度(2042)頃（20年計画）",
        "非該当", "確認済（20年計画）",
        "石狩市 環境基本計画。令和5年3月策定・計画期間20年。第2次は令和2年度で終了",
        "低（20年計画）",
        "20年計画のため改定需要は当面ない。中間見直しの規定の有無だけ確認する"),
    "宮城県|石巻市": (
        "石巻市環境基本計画（第2次）", "令和8年度(2026)", "令和17年度(2035)",
        "非該当（令和8年3月策定）", "確認済",
        "石巻市 環境基本計画（案）令和8年。令和8年度から10年間、目標年度は令和17年度",
        "低（策定直後）",
        "策定直後のため新規受託は難しい。進捗管理・アンケート二次分析など周辺支援の余地を確認する"),
    "北海道|江別市": (
        "えべつアジェンダ21―江別市環境管理計画―", "平成7年度(1995)", "計画期間30年（終期到来・後継未確認）",
        "◇満了期・後継要確認", "確認済（30年計画・後期推進計画は令和5年度終了）",
        "江別市 環境管理計画。平成7年度に計画期間30年で策定。後期推進計画は平成26年度から令和5年度で終了済。後継計画の状況は未確認",
        "優先（満了期・後継未確認）",
        "30年計画の終期が到来し、後期推進計画も令和5年度で終了している。後継計画の有無と策定予定を早急に確認する"),
    # --- 計画名・策定年のみ判明（終期は未確認のまま） ---
    "北海道|函館市": (
        "函館市環境基本計画［第3次計画］", "令和2年度(2020)", "",
        "要確認", "一部確認（策定年のみ）",
        "函館市／EPO北海道。令和2年3月策定。複数回検索したが終期はHTML上で確認できず、計画PDFの確認が必要。総合計画は2017〜2026年",
        "最優先（義務団体・終期要確認）",
        "計画PDFの表紙・第1章で終期を最優先確認する。8年計画なら令和9年度満了で該当、10年なら令和11年度満了"),
    "北海道|帯広市": (
        "第三期帯広市環境基本計画", "令和2年度(2020)", "",
        "要確認", "一部確認（策定年のみ）",
        "帯広市 第三期環境基本計画（令和2年3月策定）。第二期は平成22年度〜平成31年度。終期はPDF確認が必要",
        "中（終期要確認）", "計画PDFで終期を確認する。10年計画なら令和11年度満了"),
    "北海道|北見市": (
        "第2次北見市環境基本計画（改定版）", "", "",
        "要確認", "一部確認（計画掲載確認）",
        "北見市 第2次環境基本計画（改定版）の掲載を確認。計画期間はHTML上で読めず、PDF確認が必要",
        "中（終期要確認）", "計画PDFで計画期間を確認する"),
    "北海道|北広島市": (
        "第3次北広島市環境基本計画", "令和3年度(2021)", "",
        "要確認", "一部確認（策定年・一部改定確認）",
        "北広島市 第3次環境基本計画。第2次が令和2年度で10年の期間を終了。令和5年2月のゼロカーボンシティ宣言を受け、令和5年4月に地球環境分野を一部改定",
        "中（終期要確認・期間中改定の実例）",
        "計画PDFで終期を確認する。地球環境分野の一部改定実績があるため、国の2035年度目標を踏まえた再度の部分改定の余地もある"),
    "宮城県|大崎市": (
        "第2次大崎市環境基本計画", "令和2年度(2020)", "",
        "要確認", "一部確認（策定年のみ）",
        "大崎市 第2次環境基本計画（令和2年3月策定・10年後を見据えた計画）。アクションプランも別途あり。終期はPDF確認が必要",
        "中（終期要確認）", "計画PDFで終期を確認する。10年計画なら令和11年度満了"),
    "山形県|鶴岡市": (
        "第2次鶴岡市環境基本計画", "", "",
        "要確認", "一部確認（計画掲載確認）",
        "鶴岡市 第2次環境基本計画の掲載を確認。計画期間はHTML上で読めず、PDF確認が必要",
        "中（終期要確認）", "計画PDFで計画期間を確認する"),
    "北海道|北海道": (
        "北海道環境基本計画［第3次計画］", "令和3年度(2021)", "",
        "要確認", "一部確認（策定年のみ）",
        "北海道 環境生活部環境政策課。令和3年3月策定。複数回検索したが終期はHTML上で確認できず、計画本文の確認が必要",
        "優先（義務団体・道ルート）",
        "終期の確認と併せて、道の市町村支援・共同策定の枠組みを確認する。営業ルートの選択に直結する"),
    "秋田県|秋田県": (
        "第3次秋田県環境基本計画", "令和3年度(2021)", "",
        "要確認", "一部確認（策定年のみ）",
        "秋田県 環境基本計画。第3次を令和3年3月31日に策定。終期は未確認。なお第4次秋田県循環型社会形成推進基本計画は令和3年度〜令和7年度",
        "優先（義務団体・県ルート）", "計画本文で終期を確認する。10年計画なら令和12年度満了"),
    "福島県|福島県": (
        "福島県環境基本計画（第5次）", "令和4年度(2022)", "",
        "要確認", "一部確認（次数・前計画の終了年のみ）",
        "福島県 環境基本計画（第5次）。第4次が令和3年度で終了し、令和3年11月の環境審議会答申を経て第5次を策定。終期は未確認",
        "優先（義務団体・県ルート）", "計画本文で計画期間を確認する"),
    # --- 環境基本計画そのものが確認できないもの ---
    "青森県|青森市": (
        "環境基本計画は確認できず（関連：地球温暖化対策実行計画）", "", "",
        "要確認（環境基本計画の有無から）", "公式検索確認",
        "青森市サイト・検索では環境基本計画を確認できず。地球温暖化対策実行計画（区域施策編 令和7年3月改定／事務事業編 第4期）は確認できる。中核市のため区域施策編は策定義務",
        "最優先（義務団体・計画の有無から確認）",
        "環境基本計画の有無そのものを確認する。未策定なら新規策定提案の対象。策定済みなら計画名が異なる可能性があるため計画一覧から探す"),
    "青森県|十和田市": (
        "環境基本計画は確認できず（関連：環境保全率先行動計画・地球温暖化対策実行計画）", "", "",
        "要確認（環境基本計画の有無から）", "公式検索確認",
        "十和田市サイト・検索では環境基本計画を確認できず。環境保全率先行動計画『とわだエコ・オフィスプラン』第5次が令和4年4月から令和8年度、地球温暖化対策実行計画（区域施策編）は確認できる",
        "中（関連計画がR8満了）",
        "環境保全率先行動計画が令和8年度で満了するため、その改定を入口に環境基本計画の有無を確認する"),
    "北海道|小樽市": (
        "", "", "",
        "要確認", "公式検索確認",
        "小樽市サイト・検索では環境基本計画の計画期間を確認できず。『小樽市の環境』（年次報告）と緑の基本計画は確認できる",
        "中（未確認継続）", "公式の計画一覧・生活環境部のページで環境基本計画の有無とPDFを継続確認する"),
}

# 01_R9満了候補 から取り除き、更新後の内容で置き直す先
RESOLVED_IN_01 = {"青森県", "宮城県", "山形県", "福島市", "北海道", "秋田県", "福島県", "函館市", "青森市"}

# 01_R9満了候補 へ入れ直す行（区分, 都道府県, 自治体, 義務区分, 計画名, 計画期間, 確認状況, なぜ狙うか, 接触時期, 持っていくもの, 出典）
ROWS_EXPIRED = [
    ["◇今年度満了", "岩手県", "一関市", "努力義務", "一関市環境基本計画", "平成29年度〜令和8年度（10年間）", "確認済",
     "令和8年度末で満了＝次期計画は令和9年度開始。改定業務は今年度に実施されるはずで、R9満了先より1年早い。今年度に動いている数少ない案件",
     "直ちに接触。既発注か自前策定か、支援できる範囲が残っているかを確認する",
     "①既発注・自前策定・支援範囲の確認 ②発注済みなら範囲外の二次分析・素案レビュー・財源対応表 ③未発注なら補正での短期策定支援",
     "一関市 生活環境課 環境基本計画ページ。市の地球温暖化対策地域推進計画は令和5年度〜令和12年度"],
    ["◇満了期", "北海道", "江別市", "努力義務", "えべつアジェンダ21―江別市環境管理計画―", "平成7年度〜（計画期間30年）／後期推進計画は平成26年度〜令和5年度で終了", "確認済",
     "平成7年度に30年計画として策定され終期が到来。後期推進計画も令和5年度で終了しており、後継計画の状況が確認できていない。満了したまま次期が立ち上がっていない可能性がある",
     "後継計画の有無を早急に確認する。未策定なら令和9年度当初での策定提案",
     "①後継計画の策定状況の確認 ②30年計画から現行の制度環境への組み替え提案 ③国の第六次環境基本計画・2035年度目標との差分表",
     "江別市 環境管理計画・環境管理計画後期推進計画のページ"],
    ["◆別扱い", "福島県", "福島市", "義務（中核市）", "第3次福島市環境基本計画", "令和3年度〜令和7年度（5年間・満了済）", "確認済",
     "令和7年度末で満了済。中核市で区域施策編の策定義務があり、5年周期のため次の満了も早く来る。第4次の策定状況が未確認",
     "第4次の策定状況を早急に確認する",
     "①第4次の策定状況の確認 ②発注済みなら範囲外の二次分析・レビュー ③未策定なら満了後の空白のリスク整理",
     "福島市 環境基本計画（令和3年2月策定）ページ"],
]

ROWS_NOT_APPLICABLE = [
    ["－非該当", "青森県", "青森県", "義務（県）", "第6次青森県環境計画（青森県環境総合プラン）", "令和6年度〜令和10年度（5年間）", "確認済",
     "令和10年度満了。県計画としては短い5年周期のため、次の改定が比較的早く来る",
     "令和9年度（2027年）夏〜秋に予算要求期の接触", "県の市町村支援メニューの確認は期間中でも行う", "青森県 第6次青森県環境計画"],
    ["－非該当", "宮城県", "宮城県", "義務（県）", "宮城県環境基本計画（第4期）", "令和3年度〜令和12年度", "確認済",
     "令和12年度満了", "令和11年度（2029年）夏〜秋", "県の市町村支援メニューの確認は期間中でも行う", "宮城県公式サイト 環境基本計画（第4期）"],
    ["－非該当", "山形県", "山形県", "義務（県）", "第4次山形県環境計画", "令和3年度〜令和12年度（10年間）", "確認済",
     "令和12年度満了。従前リストの『第3次』は平成23年度〜令和2年度で終了済のため計画名を訂正した",
     "令和11年度（2029年）夏〜秋", "県の市町村支援メニューの確認は期間中でも行う", "山形県 第4次山形県環境計画（令和3年3月策定）"],
    ["－非該当", "福島県", "会津若松市", "努力義務", "会津若松市第3期環境基本計画", "令和6年度〜令和12年度（7年間）", "確認済",
     "令和12年度満了", "令和11年度（2029年）夏〜秋", "—", "会津若松市 第3期環境基本計画（令和6年3月策定）"],
    ["－非該当", "北海道", "釧路市", "努力義務", "第2次釧路市環境基本計画【改定版】", "令和3年度〜令和12年度", "確認済",
     "令和12年度満了。ただし国・北海道の温暖化計画改定を受け、令和6年3月に低炭素社会の形成の目標を一部改定した実績がある",
     "令和11年度（2029年）夏〜秋。ただし部分改定の余地は随時",
     "国の2035年度60%・2040年度73%を踏まえた再度の部分改定の提案", "釧路市 第2次環境基本計画【改定版】"],
    ["－非該当", "北海道", "千歳市", "努力義務", "第3次千歳市環境基本計画", "令和3年度〜令和12年度", "確認済",
     "令和12年度満了", "令和11年度（2029年）夏〜秋", "—", "千歳市 第3次環境基本計画"],
    ["－非該当", "北海道", "石狩市", "努力義務", "石狩市環境基本計画（第3次）", "令和5年度〜（計画期間20年）", "確認済",
     "20年計画のため改定需要は当面ない。中間見直しの規定の有無だけ確認する", "—", "—", "石狩市 環境基本計画（令和5年3月策定）"],
    ["－非該当", "宮城県", "石巻市", "努力義務", "石巻市環境基本計画（第2次）", "令和8年度〜令和17年度（10年間）", "確認済",
     "令和8年3月策定で直後のため新規受託は難しい。進捗管理・アンケート二次分析など周辺支援の余地", "—", "—", "石巻市 環境基本計画（令和8年3月）"],
]

ROWS_STILL_UNKNOWN = [
    ["？要確認", "北海道", "北海道", "義務（道）", "北海道環境基本計画［第3次計画］", "令和3年度〜（終期未確認）", "一部確認",
     "令和3年3月策定。複数回検索したが終期はHTML上で確認できず。道は市町村支援の枠組みを持つ可能性があり、営業ルートとして最優先で当たる価値がある",
     "終期確認と併せて道の市町村支援メニューを確認", "—", "北海道 環境生活部環境政策課"],
    ["？要確認", "秋田県", "秋田県", "義務（県）", "第3次秋田県環境基本計画", "令和3年度〜（終期未確認）", "一部確認",
     "令和3年3月31日策定。10年計画なら令和12年度満了だが未確認", "終期を確認", "—", "秋田県 環境基本計画"],
    ["？要確認", "福島県", "福島県", "義務（県）", "福島県環境基本計画（第5次）", "令和4年度〜（終期未確認）", "一部確認",
     "第4次が令和3年度で終了し、令和3年11月の環境審議会答申を経て第5次を策定。終期は未確認", "終期を確認", "—", "福島県 環境基本計画（第5次）"],
    ["？要確認", "北海道", "函館市", "義務（中核市）", "函館市環境基本計画［第3次計画］", "令和2年度〜（終期未確認）", "一部確認",
     "令和2年3月策定。複数回検索したが終期はHTML上で確認できず。8年計画なら令和9年度満了で該当するため、確認優先度が最も高い",
     "計画PDFで終期を最優先確認", "—", "函館市／EPO北海道"],
    ["？要確認", "青森県", "青森市", "義務（中核市）", "環境基本計画は確認できず（地球温暖化対策実行計画はあり）", "（計画の有無から要確認）", "公式検索確認",
     "検索では環境基本計画を確認できず。区域施策編（令和7年3月改定）・事務事業編（第4期）はある。中核市のため区域施策編は策定義務",
     "環境基本計画の有無そのものを確認。未策定なら新規策定提案の対象", "—", "青森市サイト"],
    ["？要確認", "北海道", "帯広市", "努力義務", "第三期帯広市環境基本計画", "令和2年度〜（終期未確認）", "一部確認",
     "令和2年3月策定。第二期は平成22年度〜平成31年度", "計画PDFで終期を確認", "—", "帯広市 第三期環境基本計画"],
    ["？要確認", "北海道", "北広島市", "努力義務", "第3次北広島市環境基本計画", "令和3年度〜（終期未確認）", "一部確認",
     "第2次が令和2年度で10年の期間を終了。令和5年2月のゼロカーボンシティ宣言を受け令和5年4月に地球環境分野を一部改定",
     "計画PDFで終期を確認", "—", "北広島市 第3次環境基本計画"],
    ["？要確認", "北海道", "北見市", "努力義務", "第2次北見市環境基本計画（改定版）", "（終期未確認）", "一部確認",
     "改定版の掲載を確認。計画期間はHTML上で読めず", "計画PDFで計画期間を確認", "—", "北見市 第2次環境基本計画（改定版）"],
    ["？要確認", "宮城県", "大崎市", "努力義務", "第2次大崎市環境基本計画", "令和2年度〜（終期未確認）", "一部確認",
     "令和2年3月策定・10年後を見据えた計画。アクションプランも別途あり", "計画PDFで終期を確認", "—", "大崎市 第2次環境基本計画"],
    ["？要確認", "山形県", "鶴岡市", "努力義務", "第2次鶴岡市環境基本計画", "（終期未確認）", "一部確認",
     "第2次の掲載を確認。計画期間はHTML上で読めず", "計画PDFで計画期間を確認", "—", "鶴岡市 第2次環境基本計画"],
    ["？要確認", "青森県", "十和田市", "努力義務", "環境基本計画は確認できず（環境保全率先行動計画・温対実行計画はあり）", "（計画の有無から要確認）", "公式検索確認",
     "環境保全率先行動計画『とわだエコ・オフィスプラン』第5次が令和4年4月から令和8年度で満了。これを入口に環境基本計画の有無を確認する",
     "率先行動計画の改定を入口に接触", "—", "十和田市サイト"],
    ["？要確認", "北海道", "小樽市", "努力義務", "（環境基本計画の期間を確認できず）", "（未確認）", "公式検索確認",
     "『小樽市の環境』（年次報告）と緑の基本計画は確認できるが、環境基本計画の期間は確認できず", "計画一覧・生活環境部ページで継続確認", "—", "小樽市サイト"],
]


# ============================================================
# 共通処理
# ============================================================
def copy_style(dst, src):
    dst.font = _copy.copy(src.font)
    dst.fill = _copy.copy(src.fill)
    dst.border = _copy.copy(src.border)
    dst.alignment = _copy.copy(src.alignment)


def find_row(ws, col, needle):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and needle in str(v):
            return r
    return None


def classify(row, ncol):
    a = row[0]
    rest = [v for v in row[1:ncol] if v not in (None, "")]
    if a in (None, "") and not rest:
        return "blank"
    if isinstance(a, str) and a.startswith("■") and not rest:
        return "band"
    if not rest:
        return "note"
    return "data"


def render_sheet(wb, name, index, title, headers, widths, rows):
    """values の並びから、既定書式でシートを作り直す。"""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index)
    ncol = len(headers)
    ws.cell(row=1, column=1, value=title).font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=C_TITLE)
    ws.row_dimensions[1].height = 26
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = F_HEAD, PatternFill("solid", fgColor=C_SUBHEAD), AL_CTR, BORDER
    ws.row_dimensions[2].height = 32
    r, last_data = 3, 2
    for row in rows:
        kind = classify(row, ncol)
        if kind == "blank":
            r += 1
            continue
        if kind == "band":
            ws.cell(row=r, column=1, value=row[0]).font = F_BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_BAND)
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            continue
        if kind == "note":
            ws.cell(row=r, column=1, value=row[0]).font = F_NOTE
            ws.cell(row=r, column=1).alignment = AL_WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_NOTE)
            r += 1
            continue
        fill = PatternFill("solid", fgColor=C_ALT) if (r % 2 == 1) else None
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else None)
            cell.font, cell.alignment, cell.border = F_BODY, AL_WRAP, BORDER
            if fill:
                cell.fill = fill
        last_data = r
        r += 1
    for c, w in enumerate(widths, start=1):
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{last_data}"
    ws.sheet_view.zoomScale = 90
    return ws


# ============================================================
# 1. 02_営業先マスタ への反映
# ============================================================
def update_master(wb):
    ws = wb["02_営業先マスタ"]
    applied = 0
    for r in range(3, ws.max_row + 1):
        pref = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if not name:
            continue
        f = FINDINGS.get(f"{pref}|{name}")
        if not f:
            continue
        keikaku, start, end, hantei, joukyou, shutten, rank, action = f
        ws.cell(row=r, column=7, value=keikaku or None)
        ws.cell(row=r, column=8, value=start or None)
        ws.cell(row=r, column=9, value=end or None)
        ws.cell(row=r, column=10, value=hantei)
        ws.cell(row=r, column=11, value=joukyou)
        ws.cell(row=r, column=12, value=shutten)
        ws.cell(row=r, column=13, value=rank)
        ws.cell(row=r, column=14, value=action)
        applied += 1
    print(f"  02_営業先マスタ: {applied}件を更新")
    if applied != len(FINDINGS):
        raise SystemExit(f"エラー: 反映件数が一致しません（{applied} != {len(FINDINGS)}）。自治体名の指定を確認してください。")
    return applied


# ============================================================
# 2. 01_R9満了候補 の作り直し
# ============================================================
def update_01(wb):
    ws = wb["01_R9満了候補"]
    if find_row(ws, 3, "一関市") is not None:
        print("  01_R9満了候補: 反映済みのためスキップ")
        return 0
    ncol = ws.max_column
    widths = [ws.column_dimensions[get_column_letter(c)].width for c in range(1, ncol + 1)]
    title = ws.cell(row=1, column=1).value
    headers = [ws.cell(row=2, column=c).value for c in range(1, ncol + 1)]
    rows = [list(r) for r in ws.iter_rows(min_row=3, values_only=True)]

    # 解決済みの行を全ブロックから除去
    kept = [r for r in rows if not (classify(r, ncol) == "data" and str(r[2]) in RESOLVED_IN_01)]
    removed = len(rows) - len(kept)

    out, inserted_expired = [], False
    for row in kept:
        out.append(row)
        if classify(row, ncol) == "band" and "期間満了済・改定進行中" in str(row[0]):
            out.extend([list(x) for x in ROWS_EXPIRED])
            inserted_expired = True
    if not inserted_expired:
        raise SystemExit("エラー: 『■ 期間満了済・改定進行中』の帯が見つかりません。")

    # 非該当・要確認は末尾のブロックとして置き直す
    out.append(["■ 2026-08-25 追加調査：確認済で令和9年度満了に該当しない先"])
    out.extend([list(x) for x in ROWS_NOT_APPLICABLE])
    out.append(["■ 2026-08-25 追加調査：終期が確認できず、令和9年度満了の可能性を排除できない先"])
    out.extend([list(x) for x in ROWS_STILL_UNKNOWN])

    # データ行を1件も持たなくなった見出し帯を除去する
    pruned = []
    for i, row in enumerate(out):
        if classify(row, ncol) == "band":
            has_data = False
            for nxt in out[i + 1:]:
                k = classify(nxt, ncol)
                if k == "band":
                    break
                if k == "data":
                    has_data = True
                    break
            if not has_data:
                continue
        pruned.append(row)
    out = pruned

    idx = wb.sheetnames.index("01_R9満了候補")
    new_ws = render_sheet(wb, "01_R9満了候補", idx, title, headers, widths, out)
    before = sum(1 for r in rows if classify(r, ncol) == "data")
    after = sum(1 for r in new_ws.iter_rows(min_row=3, values_only=True) if classify(list(r), ncol) == "data")
    added = len(ROWS_EXPIRED) + len(ROWS_NOT_APPLICABLE) + len(ROWS_STILL_UNKNOWN)
    if after != before - removed + added:
        raise SystemExit(f"エラー: 行数が合いません（{before} - {removed} + {added} != {after}）")
    print(f"  01_R9満了候補: データ行 {before} -> {after}（除去{removed} / 追加{added}）")
    return added


# ============================================================
# 3. 08_改定時期調査 シートの追加
# ============================================================
SURVEY_HEADERS = ["区分", "都道府県", "自治体", "義務区分", "調査結果", "計画名", "計画期間", "R9満了判定", "確認状況", "根拠・確認メモ", "次確認"]
SURVEY_WIDTHS = [16, 10, 12, 12, 14, 34, 30, 24, 20, 62, 46]


def survey_rows():
    gimu = {"北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
            "札幌市", "仙台市", "旭川市", "函館市", "青森市", "八戸市", "盛岡市", "秋田市", "山形市", "福島市", "郡山市", "いわき市"}
    rows = [["■ 終期を特定できた先（確認済）"]]
    order_conf = ["青森県|青森県", "宮城県|宮城県", "山形県|山形県", "福島県|福島市", "岩手県|一関市",
                  "福島県|会津若松市", "北海道|釧路市", "北海道|千歳市", "北海道|石狩市", "宮城県|石巻市", "北海道|江別市"]
    order_part = ["北海道|北海道", "秋田県|秋田県", "福島県|福島県", "北海道|函館市", "北海道|帯広市",
                  "北海道|北広島市", "北海道|北見市", "宮城県|大崎市", "山形県|鶴岡市"]
    order_none = ["青森県|青森市", "青森県|十和田市", "北海道|小樽市"]
    for key in order_conf:
        pref, name = key.split("|")
        k, s, e, h, j, src, rank, act = FINDINGS[key]
        rows.append(["終期確定", pref, name, "義務" if name in gimu else "努力義務", "確認済", k,
                     f"{s}〜{e}" if e else s, h, j, src, act])
    rows.append(["■ 計画名・策定年のみ判明（終期は未確認のまま。推測で埋めない）"])
    for key in order_part:
        pref, name = key.split("|")
        k, s, e, h, j, src, rank, act = FINDINGS[key]
        rows.append(["一部確認", pref, name, "義務" if name in gimu else "努力義務", "一部確認", k,
                     s if s else "（未確認）", h, j, src, act])
    rows.append(["■ 環境基本計画そのものを確認できなかった先（未策定の可能性を含む）"])
    for key in order_none:
        pref, name = key.split("|")
        k, s, e, h, j, src, rank, act = FINDINGS[key]
        rows.append(["計画未確認", pref, name, "義務" if name in gimu else "努力義務", "公式検索確認",
                     k or "（確認できず）", "（未確認）", h, j, src, act])
    rows.append(["調査方法: 自治体サイトへの直接アクセスができない環境のため、公開情報の検索で確認できた範囲に限る。"
                 "計画期間がHTML上に出ておらずPDF本文にしか記載がない団体が多く、そこは『一部確認』として終期欄を空欄のままにしている。"])
    rows.append(["この調査で令和9年度末満了の新規該当先は出なかった。一方、令和8年度末満了（一関市＝今年度が改定年度）と、"
                 "満了済・満了期（福島市・江別市）という、R9満了より先に動く先が3件見つかった。"])
    return rows


def add_survey_sheet(wb):
    idx = len(wb.sheetnames)
    render_sheet(wb, SURVEY_SHEET, idx,
                 "⑧ 改定時期調査（2026-08-25 第2回）｜ 義務団体・主要市の未確認先を優先に確認",
                 SURVEY_HEADERS, SURVEY_WIDTHS, survey_rows())


# ============================================================
# 4. 00_概要 / 05_更新メモ
# ============================================================
def append_styled_row(ws, values, template_row):
    ncol = ws.max_column
    r = ws.max_row + 1
    for c in range(1, ncol + 1):
        copy_style(ws.cell(row=r, column=c), ws.cell(row=template_row, column=c))
        ws.cell(row=r, column=c).value = values[c - 1] if c - 1 < len(values) else None
    return r


def update_notes(wb):
    ws0 = wb["00_概要"]
    if find_row(ws0, 2, SURVEY_SHEET) is None:
        t = find_row(ws0, 1, "構成") or 3
        append_styled_row(ws0, ["構成", SURVEY_SHEET,
                                "義務団体・主要市の未確認先を優先に計画期間を調査した結果。終期を特定できた11件、計画名・策定年のみ9件、計画そのものを確認できず3件",
                                "終期を特定できたものだけを確認済にし、推測では埋めていない"], t)
        append_styled_row(ws0, ["更新", "2026-08-25 第2回の更新",
                                "R9年度末満了の新規該当先は出なかった。代わりに令和8年度末満了（一関市＝今年度が改定年度）と満了済・満了期（福島市・江別市）の3件を発見",
                                f"{SURVEY_SHEET} を参照"], t)
    ws5 = wb["05_更新メモ"]
    if find_row(ws5, 1, "2026-08-25 改定時期調査 第2回") is None:
        t = ws5.max_row
        append_styled_row(ws5, [
            "2026-08-25 改定時期調査 第2回",
            "義務団体9件・主要市14件の未確認先",
            "計画期間を調査。終期を特定できた11件を確認済へ、計画名・策定年のみの9件を一部確認へ、環境基本計画を確認できなかった3件を計画未確認へ更新",
            "一関市はR8年度末満了で今年度が改定年度＝直ちに接触。福島市（中核市）はR3〜R7で満了済、江別市は30年計画の終期到来で後継未確認。いずれもR9満了先より先に動く",
            f"{SURVEY_SHEET}",
        ], t)


# ============================================================
# 5. 出力
# ============================================================
def export_csv(wb):
    os.makedirs(CSV_DIR, exist_ok=True)
    for i, ws in enumerate(wb.worksheets):
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        naming.write_csv(os.path.join(CSV_DIR, f"{naming.sheet_slug(ws.title, i)}.csv"), rows, encoding=ENC, quiet=True)


def main():
    if not os.path.exists(SRC):
        print(f"エラー: 入力ブックがありません -> {SRC}")
        print(f"       （現在のモード … {naming.mode_label(ASCII)}）")
        sys.exit(1)
    wb = load_workbook(SRC)
    update_master(wb)
    update_01(wb)
    add_survey_sheet(wb)
    update_notes(wb)

    out_path = os.path.join(OUT_DIR, f"{naming.book_stem(BOOK_STEM_JP, ASCII)}.xlsx")
    wb.save(out_path)
    print(f"保存: {out_path}")
    if not ASCII:
        ascii_path = os.path.join(OUT_DIR, f"{naming.book_stem(BOOK_STEM_JP, True)}.xlsx")
        shutil.copyfile(out_path, ascii_path)
        print(f"保存: {ascii_path}（提出用・半角英数名）")

    values = [v for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for v in row if isinstance(v, str)]
    bad = naming.check_encodable(values, ENC)
    print(f"\n{ENC} エンコード検査: " + ("変換不能なし" if not bad else
          "、".join(f"{ch!r}×{n}→{naming.CP932_SUBSTITUTIONS.get(ch,'?')!r}" for ch, n in bad.items())))
    export_csv(wb)
    print(f"提出用CSV: {CSV_DIR} に {len(wb.worksheets)}件")
    print(f"モード … {naming.mode_label(ASCII)}")


if __name__ == "__main__":
    main()
