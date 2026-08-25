# -*- coding: utf-8 -*-
"""
09_PDF確認待ち一覧 シートの追加（2026-08-25 第3回）

背景:
  02_営業先マスタの未確認先のうち、公開情報の検索で確認できるものは
  第2回（08_改定時期調査_20260825b）で確認しきった。残っているのは
  「計画期間がPDF本文にしか書かれておらず、検索の要約に出てこない」先で、
  検索経路では原理的に確定できない。

  加えて本作業環境は、自治体サイト・府省サイトを含む全ホストへの
  アウトバウンド接続が組織のegressポリシーで遮断されている
  （CONNECTに対しゲートウェイが403を返す。Wikipediaも同様）。
  そのためPDFを取得できず、PDF→Word→抽出の手順を回せない。

このシートの役割:
  制限のない環境でPDFを取得すれば1コマンドで片付くよう、対象・掲載ページ・
  手順を1枚にまとめる。取得後は
    python3 fetch_plan_pdfs.py --out plan_pdfs
    python3 batch_extract_plan_periods.py plan_pdfs
  で計画期間の一覧CSVが出る。
"""

import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import naming
from fetch_plan_pdfs import TARGETS

OUT_DIR = "/home/user/repository/output"
CSV_DIR = os.path.join(OUT_DIR, "csv")
BOOK_STEM_JP = "環境基本計画_営業先リスト_北海道東北"
SRC = os.path.join(OUT_DIR, f"{BOOK_STEM_JP}.xlsx")
SHEET = "09_PDF確認待ち一覧"

C_TITLE, C_SUBHEAD, C_BAND, C_ALT, C_NOTE, C_WHITE = "1F3864", "2E75B6", "DDEBF7", "F7FAFC", "FFF3F3", "FFFFFF"
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name="Yu Gothic", size=14, bold=True, color=C_WHITE)
F_HEAD = Font(name="Yu Gothic", size=10, bold=True, color=C_WHITE)
F_BODY = Font(name="Yu Gothic", size=10)
F_BOLD = Font(name="Yu Gothic", size=10, bold=True)
F_NOTE = Font(name="Yu Gothic", size=9, color="7F3F3F")
AL_WRAP = Alignment(vertical="top", wrap_text=True)
AL_CTR = Alignment(vertical="center", horizontal="center", wrap_text=True)

HEADERS = ["優先度", "都道府県", "自治体", "義務区分", "現在の状況", "計画掲載ページ", "PDFで見る場所", "確定したら入れる列"]
WIDTHS = [10, 10, 12, 12, 52, 62, 40, 34]

PREF = {
    "函館市": ("北海道", "義務（中核市）"), "青森市": ("青森県", "義務（中核市）"),
    "北海道": ("北海道", "義務（道）"), "秋田県": ("秋田県", "義務（県）"), "福島県": ("福島県", "義務（県）"),
    "帯広市": ("北海道", "努力義務"), "北広島市": ("北海道", "努力義務"), "北見市": ("北海道", "努力義務"),
    "大崎市": ("宮城県", "努力義務"), "鶴岡市": ("山形県", "努力義務"), "小樽市": ("北海道", "努力義務"),
    "十和田市": ("青森県", "努力義務"), "上士幌町": ("北海道", "努力義務"), "新得町": ("北海道", "努力義務"),
    "池田町": ("北海道", "努力義務"), "厚岸町": ("北海道", "努力義務"), "鶴居村": ("北海道", "努力義務"),
    "弟子屈町": ("北海道", "努力義務"),
}
WHERE = "表紙／第1章『計画の期間』。多くは「計画期間は、○年度から○年度までの○年間」の1行"
COLS = "02_営業先マスタ H列(開始)・I列(終期)・J列(判定)・K列(確認状況)"


def rows():
    out = [["■ 対象18件（優先度順）。掲載ページを開き、PDFの計画期間の1行だけ確認する"]]
    for name, pri, note, url in TARGETS:
        pref, gimu = PREF.get(name, ("", ""))
        out.append([pri, pref, name, gimu, note, url, WHERE, COLS])
    out.append(["■ 手順（制限のない環境で実行する）"])
    out.append(["手順1", "", "", "", "計画ページからPDFを一括取得する", "python3 fetch_plan_pdfs.py --out plan_pdfs", "—", "—"])
    out.append(["手順2", "", "", "", "PDF→Word変換のうえ計画期間を一括抽出する", "python3 batch_extract_plan_periods.py plan_pdfs", "—", "計画期間抽出結果.csv が出る"])
    out.append(["手順3", "", "", "", "1件だけ確認する場合", "python3 extract_plan_period.py <PDFパス>", "—", "—"])
    out.append(["手順4", "", "", "", "抽出結果を02_営業先マスタへ転記し、終期がR9(2027)の先を01_R9満了候補へ移す", "—", "—", COLS])
    out.append(["■ 補足"])
    out.append(["注意", "", "", "", "自動取得に失敗する団体は、掲載ページを開いてPDFを手で保存し、ファイル名の先頭を自治体名にして同じフォルダへ置けば手順2で処理できる",
                "例: 函館市_環境基本計画.pdf", "—", "—"])
    out.append(["注意", "", "", "", "画像のみのスキャンPDFはOCRが必要。抽出ツールはその旨を表示して中断する", "—", "—", "—"])
    out.append(["注意", "", "", "", "抽出はスコア付きで候補を出す。スコアが低い場合は根拠行を目視で確認してから確定させる", "—", "—", "—"])
    out.append(["本シートを作った理由: 第2回調査（08シート）で公開情報の検索から確認できる先は確認しきった。"
                "残る18件は計画期間がPDF本文にしかなく、検索の要約に出てこないため検索経路では確定できない。"
                "加えて本作業環境は全ホストへのアウトバウンド接続が組織のegressポリシーで遮断されており（CONNECTに403）、PDFを取得できない。"])
    return out


def render(wb):
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, len(wb.sheetnames))
    ncol = len(HEADERS)
    ws.cell(row=1, column=1, value="⑨ PDF確認待ち一覧 ｜ 検索では確定できない18件。PDFの計画期間1行で片付く").font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=C_TITLE)
    ws.row_dimensions[1].height = 26
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = F_HEAD, PatternFill("solid", fgColor=C_SUBHEAD), AL_CTR, BORDER
    ws.row_dimensions[2].height = 30
    r, last = 3, 2
    for row in rows():
        if len(row) == 1:
            is_band = str(row[0]).startswith("■")
            ws.cell(row=r, column=1, value=row[0]).font = F_BOLD if is_band else F_NOTE
            ws.cell(row=r, column=1).alignment = AL_WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            color = C_BAND if is_band else C_NOTE
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)
                if is_band:
                    ws.cell(row=r, column=c).border = BORDER
            r += 1
            continue
        fill = PatternFill("solid", fgColor=C_ALT) if (r % 2 == 1) else None
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else None)
            cell.font, cell.alignment, cell.border = F_BODY, AL_WRAP, BORDER
            if fill:
                cell.fill = fill
        last = r
        r += 1
    for c, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{last}"
    ws.sheet_view.zoomScale = 90


def find_row(ws, col, needle):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and needle in str(v):
            return r
    return None


def main():
    wb = load_workbook(SRC)
    render(wb)

    ws5 = wb["05_更新メモ"]
    if find_row(ws5, 1, "2026-08-25 PDF確認待ち") is None:
        import copy as _copy
        t = ws5.max_row
        r = ws5.max_row + 1
        vals = ["2026-08-25 PDF確認待ち", "未確認先18件",
                "検索で確定できる先は第2回で確認しきったため、残りをPDF確認待ちとして09シートに整理。取得スクリプトと一括抽出ツールを用意",
                "制限のない環境でPDFを取得すれば1コマンドで確定できる。函館市・青森市（いずれも中核市）が最優先",
                "09_PDF確認待ち一覧 ／ fetch_plan_pdfs.py ／ batch_extract_plan_periods.py"]
        for c in range(1, ws5.max_column + 1):
            src = ws5.cell(row=t, column=c)
            cell = ws5.cell(row=r, column=c)
            cell.font = _copy.copy(src.font); cell.fill = _copy.copy(src.fill)
            cell.border = _copy.copy(src.border); cell.alignment = _copy.copy(src.alignment)
            cell.value = vals[c - 1] if c - 1 < len(vals) else None

    out_path = os.path.join(OUT_DIR, f"{naming.book_stem(BOOK_STEM_JP, naming.ascii_mode())}.xlsx")
    wb.save(out_path)
    print(f"保存: {out_path}")
    if not naming.ascii_mode():
        shutil.copyfile(out_path, os.path.join(OUT_DIR, f"{naming.book_stem(BOOK_STEM_JP, True)}.xlsx"))
        print("保存: 提出用・半角英数名も更新")

    os.makedirs(CSV_DIR, exist_ok=True)
    for i, ws in enumerate(wb.worksheets):
        naming.write_csv(os.path.join(CSV_DIR, f"{naming.sheet_slug(ws.title, i)}.csv"),
                         [list(x) for x in ws.iter_rows(values_only=True)], quiet=True)
    print(f"提出用CSV: {len(wb.worksheets)}件")


if __name__ == "__main__":
    main()
