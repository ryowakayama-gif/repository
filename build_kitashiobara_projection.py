# -*- coding: utf-8 -*-
"""
北塩原村 第8期障がい福祉計画・第4期障がい児福祉計画
将来推計 管理ブック ジェネレータ

出力: output/北塩原村_将来推計.xlsx

仕様書4-Ⅱ(3)①及び4-Ⅱ(5)③が必須記載事項として求める
「障がい者手帳所持者数（身体・療育・精神）の将来推計」に対応する。

設計方針
- 実績（確定値）はハードコード、推計はすべてExcel数式で持たせる。
  村からR6〜R8実績を受領して入力欄に入れれば、推計値が自動で再計算される。
- トレンド延長法の対象期間は COUNT() で可変にしてあるため、
  実績年が5か年から8か年に増えれば自動的に8か年回帰へ切り替わる。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from kitashiobara_common import (
    AGE_BANDS_OTHER, AGE_BANDS_PHYSICAL,
    COLORS, OUT_DIR, FONT, POPULATION, TEGATA, TEGATA_YEARS,
    add_sheet, ensure_out_dir, set_col_widths, style_data_cell,
    style_header_row, style_note, style_status, style_title, write_row,
)

OUT_FILE = f"{OUT_DIR}/北塩原村_将来推計.xlsx"

INT_FMT = "#,##0"
PCT_FMT = "0.0%"
RATE_FMT = "0.000%"

# 人口推計シートの行位置（データは6行目から）
POP_START = 6
POP_ROW = {rec[1]: POP_START + i for i, rec in enumerate(POPULATION)}  # 西暦 -> 行番号
POP_R6 = POP_ROW[2024]
POP_R11 = POP_ROW[2029]

# 手帳実績シートの行位置
TEG_START = 6
TEG_ROW = {name: TEG_START + i for i, name in enumerate(TEGATA)}
TEG_YEAR_HDR = 5  # 西暦を並べたヘッダ行
TEG_FIRST_COL = 3  # C列 = 平成31年
TEG_LAST_COL = TEG_FIRST_COL + len(TEGATA_YEARS) - 1  # J列 = 令和8年

# 手帳所持率の推移（現行計画本編の年齢3区分別人口により算定）
RATE_HISTORY = [
    ("身体障害者手帳", (0.04594, 0.04486, 0.04867, 0.04813, 0.04666), "横ばい（4.49〜4.87％）"),
    ("療育手帳", (0.00474, 0.00482, 0.00532, 0.00473, 0.00491), "横ばい"),
    ("精神障害者保健福祉手帳", (0.00875, 0.01038, 0.01027, 0.01026, 0.01228),
     "明確な上昇（平成31年比＋40％）"),
    ("合計", (0.05942, 0.06007, 0.06426, 0.06312, 0.06386), "上昇（5.94％→6.39％）"),
]

# 令和11年の推計値（方法A・B・D・中位）
METHOD_D = [
    ("身体障害者手帳", 104, 102, 111, 103,
     "実数は減少しているが所持率は横ばい。方法Dは所持率のわずかな上昇を拾って高めに出る"),
    ("療育手帳", 10, 11, 11, 11, "いずれの方法でも10〜11人。差に意味はない"),
    ("精神障害者保健福祉手帳", 35, 27, 35, 31,
     "方法Dが方法Aと一致。方法Bの27人は所持率一定の仮定による過小推計"),
    ("合計", 149, 140, 157, 145, "方法Dは中位より12人多い"),
]



# ============================================================
# 00_概要
# ============================================================
def sheet_overview(wb):
    ws = wb.create_sheet("00_概要")
    set_col_widths(ws, [30, 22, 20, 16, 46])
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:E1")
    style_title(ws["A1"], "北塩原村 将来推計 管理ブック（人口・障がい者手帳所持者数）",
                fill=COLORS["障がい"])
    ws.merge_cells("A2:E2")
    style_note(ws["A2"],
               "第8期北塩原村障がい福祉計画・第4期北塩原村障がい児福祉計画（令和9〜11年度）／"
               "業務仕様書4-Ⅱ(3)①・4-Ⅱ(5)③「障がい者手帳所持者数の将来推計」に対応。")
    ws.row_dimensions[2].height = 28

    r = 4
    style_header_row(ws, r, ["確認項目", "現在値（自動計算）", "参照先", "状態", "備考"])
    r += 1
    rows = [
        ("令和11年度 総人口",
         f"='01_人口推計'!D{POP_R11}", "01_人口推計", "確定",
         "こども・子育て計画のコーホート変化率法による村公式推計"),
        ("令和11年度 高齢化率",
         f"='01_人口推計'!H{POP_R11}", "01_人口推計", "確定", "老年人口÷総人口"),
        ("令和9年 手帳所持者数（中位推計）",
         "='03_手帳将来推計'!C16", "03_手帳将来推計", "暫定",
         "平成31〜令和5年の5か年実績に基づく暫定値。R6〜R8実績の入力で自動更新／次期計画1年目"),
        ("令和10年 手帳所持者数（中位推計）",
         "='03_手帳将来推計'!D16", "03_手帳将来推計", "暫定", "次期計画2年目"),
        ("令和11年 手帳所持者数（中位推計）",
         "='03_手帳将来推計'!E16", "03_手帳将来推計", "暫定", "次期計画目標年度"),
        ("推計の基準年",
         "='03_手帳将来推計'!C4", "03_手帳将来推計", "暫定",
         "所持率法の基準年。R8実績受領後は2026に変更する"),
        ("トレンド延長法の対象年数",
         f"=COUNT('02_手帳実績'!{get_column_letter(TEG_FIRST_COL)}{TEG_ROW['身体障害者手帳']}:"
         f"{get_column_letter(TEG_LAST_COL)}{TEG_ROW['身体障害者手帳']})",
         "02_手帳実績", "暫定", "実績を入力すると自動的に対象年数が増える"),
    ]
    for i, row in enumerate(rows):
        write_row(ws, r, list(row), alt=(i % 2 == 1),
                  aligns=["left", "center", "center", "center", "left"],
                  numfmts=[None, INT_FMT, None, None, None])
        style_status(ws.cell(row=r, column=4))
        r += 1
    # 高齢化率行はパーセント表示
    ws.cell(row=6, column=2).number_format = PCT_FMT

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "本ブックの使い方", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "1. 黄色のセル（入力欄）に村から受領した令和6〜8年度の実績を入力します。他のセルは数式のため触りません。",
        "2. 手帳所持者数の実績を入力すると、トレンド延長法の回帰対象年数が自動的に増え、推計値が更新されます。",
        "3. 所持率法の基準年（03_手帳将来推計 C4）は、最新実績が揃った年度に変更します。基準年の総人口は01_人口推計から自動参照します。",
        "4. 令和2〜4年及び令和7〜10年の人口実績が判明した場合は、01_人口推計の該当行に入力します（補間推計を実績で上書き）。",
        "5. 推計結果は計画書 第2章に転記します。単一値ではなく方法A・方法Bの幅を併記してください。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 18
        r += 1
    return ws


# ============================================================
# 01_人口推計
# ============================================================
def sheet_population(wb):
    ws = add_sheet(
        wb, "01_人口推計", "人口推計",
        "住民基本台帳（各年4月1日現在）。令和11年は北塩原村こども・子育て計画のコーホート変化率法による村公式推計。"
        "令和7〜10年は令和6年実績と令和11年公式推計の線形補間（数式）。空欄は村資料待ちの入力欄。",
        [12, 8, 12, 12, 12, 14, 12, 11, 42])
    style_header_row(ws, 5, ["年度", "西暦", "区分", "総人口", "年少人口\n(0〜14歳)",
                             "生産年齢人口\n(15〜64歳)", "老年人口\n(65歳〜)",
                             "高齢化率", "出典・備考"])

    aligns = ["center", "center", "center", "right", "right", "right", "right", "center", "left"]
    numfmts = [None, None, None, INT_FMT, INT_FMT, INT_FMT, INT_FMT, PCT_FMT, None]

    for i, (wareki, seireki, kubun, total, young, work, old, note) in enumerate(POPULATION):
        row = POP_START + i
        is_interp = (kubun == "補間推計")
        # 補間推計は数式、実績・公式推計は値（未判明は空欄＝入力欄）
        if is_interp:
            step = seireki - 2024  # 令和6年からの経過年数（1〜4）
            vals = [
                f"=ROUND($D${POP_R6}+($D${POP_R11}-$D${POP_R6})*{step}/5,0)",
                f"=ROUND($E${POP_R6}+($E${POP_R11}-$E${POP_R6})*{step}/5,0)",
                f"=ROUND($F${POP_R6}+($F${POP_R11}-$F${POP_R6})*{step}/5,0)",
                f"=ROUND($G${POP_R6}+($G${POP_R11}-$G${POP_R6})*{step}/5,0)",
            ]
        else:
            vals = [total, young, work, old]

        fills = [None, None, None] + [
            (COLORS["calc"] if is_interp else (COLORS["input"] if v is None else None))
            for v in ([total, young, work, old] if not is_interp else [1, 1, 1, 1])
        ] + [COLORS["calc"], None]

        write_row(ws, row, [wareki, seireki, kubun] + vals +
                  [f"=IFERROR(G{row}/D{row},\"\")", note],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts, fills=fills)

    # 検算行
    r = POP_START + len(POPULATION) + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    style_note(ws.cell(row=r, column=1),
               "検算：年少人口＋生産年齢人口＋老年人口＝総人口（令和11年 111＋1,023＋1,051＝2,185）。"
               "補間推計は四捨五入のため、合計が総人口と1人程度ずれる場合があります。")
    return ws


# ============================================================
# 02_手帳実績
# ============================================================
def sheet_tegata_actual(wb):
    ws = add_sheet(
        wb, "02_手帳実績", "障がい者手帳所持者数 実績",
        "各年4月1日現在。平成31年〜令和5年は現行計画本編（第4次障がい者計画・第7期障がい福祉計画）による。"
        "令和6〜8年は村資料待ちの入力欄（黄色）。入力すると03_手帳将来推計の回帰対象年数が自動的に増えます。",
        [22, 10] + [11] * len(TEGATA_YEARS) + [34])

    hdr = ["区分", "単位"] + [f"{w}\n({s})" for w, s in TEGATA_YEARS] + ["傾向"]
    style_header_row(ws, 5, hdr)
    # 数式が参照する西暦の行（非表示にせず4行目に置く）
    ws.cell(row=4, column=1, value="西暦（数式参照用）").font = Font(name=FONT, size=9, color="808080")
    for j, (_, s) in enumerate(TEGATA_YEARS):
        c = ws.cell(row=4, column=TEG_FIRST_COL + j, value=s)
        c.font = Font(name=FONT, size=9, color="808080")
        c.alignment = Alignment(horizontal="center")

    trends = {
        "身体障害者手帳": "緩やかな減少傾向（高齢化に伴う死亡等の影響を含む）",
        "療育手帳": "横ばい",
        "精神障害者保健福祉手帳": "緩やかな増加傾向",
    }
    ncol = len(hdr)
    aligns = ["left", "center"] + ["right"] * len(TEGATA_YEARS) + ["left"]
    numfmts = [None, None] + [INT_FMT] * len(TEGATA_YEARS) + [None]

    for i, (name, vals) in enumerate(TEGATA.items()):
        row = TEG_ROW[name]
        fills = [None, None] + [(None if v is not None else COLORS["input"]) for v in vals] + [None]
        write_row(ws, row, [name, "人"] + list(vals) + [trends[name]],
                  alt=(i % 2 == 1), aligns=aligns, numfmts=numfmts, fills=fills)

    # 合計行
    total_row = TEG_START + len(TEGATA)
    sums = [f"=IF(COUNT({get_column_letter(TEG_FIRST_COL + j)}{TEG_START}:"
            f"{get_column_letter(TEG_FIRST_COL + j)}{TEG_START + len(TEGATA) - 1})=0,\"\","
            f"SUM({get_column_letter(TEG_FIRST_COL + j)}{TEG_START}:"
            f"{get_column_letter(TEG_FIRST_COL + j)}{TEG_START + len(TEGATA) - 1}))"
            for j in range(len(TEGATA_YEARS))]
    write_row(ws, total_row, ["合計", "人"] + sums + ["令和5年156人（人口の約6.4％）"],
              aligns=aligns, numfmts=numfmts,
              fills=[COLORS["band"]] * ncol)
    for col in range(1, ncol + 1):
        ws.cell(row=total_row, column=col).font = Font(name=FONT, size=10, bold=True)

    # 人口・所持率
    r = total_row + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    style_title(ws.cell(row=r, column=1), "参考：総人口と手帳所持率", fill=COLORS["subhead"], size=11)
    r += 1
    pop_cells = []
    for _, s in TEGATA_YEARS:
        pr = POP_ROW.get(s)
        pop_cells.append(f"='01_人口推計'!D{pr}" if pr else None)
    write_row(ws, r, ["総人口", "人"] + pop_cells + ["01_人口推計から自動参照"],
              aligns=aligns, numfmts=numfmts,
              fills=[None, None] + [COLORS["calc"]] * len(TEGATA_YEARS) + [None])
    pop_row = r
    r += 1
    rate_cells = [
        f"=IFERROR(IF(OR({get_column_letter(TEG_FIRST_COL + j)}{total_row}=\"\","
        f"{get_column_letter(TEG_FIRST_COL + j)}{pop_row}=\"\"),\"\","
        f"{get_column_letter(TEG_FIRST_COL + j)}{total_row}/"
        f"{get_column_letter(TEG_FIRST_COL + j)}{pop_row}),\"\")"
        for j in range(len(TEGATA_YEARS))
    ]
    write_row(ws, r, ["手帳所持率（合計）", "％"] + rate_cells + ["合計手帳所持者数÷総人口"],
              alt=True, aligns=aligns,
              numfmts=[None, None] + [RATE_FMT] * len(TEGATA_YEARS) + [None],
              fills=[None, None] + [COLORS["calc"]] * len(TEGATA_YEARS) + [None])

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    style_note(ws.cell(row=r, column=1),
               "留意：仕様書に記載されたアンケート調査対象者数（身体約130人・精神約40人・療育約20人＝計約190人）は、"
               "本表の令和5年時点156人と差があります。年数経過による自然増か、集計基準・対象年齢・"
               "サービス利用者の重複計上の違いによるものかを村に確認し、確認結果を05_村確認事項に記録してください。")
    ws.row_dimensions[r].height = 32
    return ws


# ============================================================
# 03_手帳将来推計
# ============================================================
def sheet_tegata_projection(wb):
    ws = add_sheet(
        wb, "03_手帳将来推計", "障がい者手帳所持者数 将来推計",
        "性質の異なる2つの方法を併用し、幅を持たせて提示します。方法A・方法Bの差が推計の不確実性の幅です。"
        "推計の時点は実績と同じ「各年4月1日現在」です（実績が4月1日現在のため、年度末で予測すると1年ずれます）。"
        "計画書には単一値ではなく幅を併記してください。不確実性の定量評価は04_推計の不確実性を参照。",
        [24, 18, 14, 14, 14, 14, 46])

    fc = get_column_letter(TEG_FIRST_COL)
    lc = get_column_letter(TEG_LAST_COL)

    # --- 基準年設定 ---
    ws.cell(row=4, column=1, value="所持率法の基準年（西暦）").font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=4, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c = ws.cell(row=4, column=3, value=2023)
    style_data_cell(c, align="center", fill=COLORS["input"])
    c.font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=4, column=4, value="基準年総人口")
    style_data_cell(ws.cell(row=4, column=4), align="center")
    c = ws.cell(row=4, column=5,
                value=f"=INDEX('01_人口推計'!$D${POP_START}:$D${POP_START + len(POPULATION) - 1},"
                      f"MATCH($C$4,'01_人口推計'!$B${POP_START}:$B${POP_START + len(POPULATION) - 1},0))")
    style_data_cell(c, align="right", numfmt=INT_FMT, fill=COLORS["calc"])
    ws.merge_cells("F4:G4")
    style_note(ws["F4"], "R6〜R8実績の受領後は基準年を2026（令和8年）に変更してください。")

    # --- 推計表 ---
    style_header_row(ws, 6, ["区分", "手法", "令和9年\n(2027.4.1)", "令和10年\n(2028.4.1)",
                             "令和11年\n(2029.4.1)", "基準年所持率", "考え方・特性"])

    aligns = ["left", "left", "right", "right", "right", "right", "left"]
    numfmts = [None, None, INT_FMT, INT_FMT, INT_FMT, RATE_FMT, None]
    row = 7
    method_rows = {}  # (区分, 手法) -> 行番号

    for name in TEGATA:
        trow = TEG_ROW[name]
        cnt = f"COUNT('02_手帳実績'!${fc}${trow}:${lc}${trow})"
        known_y = f"OFFSET('02_手帳実績'!${fc}${trow},0,0,1,{cnt})"
        known_x = f"OFFSET('02_手帳実績'!${fc}${TEG_YEAR_HDR - 1},0,0,1,{cnt})"
        rate = (f"=IFERROR(INDEX('02_手帳実績'!${fc}${trow}:${lc}${trow},"
                f"MATCH($C$4,'02_手帳実績'!${fc}${TEG_YEAR_HDR - 1}:${lc}${TEG_YEAR_HDR - 1},0))/$E$4,\"\")")

        # 方法A トレンド延長法
        vals_a = [f"=IFERROR(ROUND(FORECAST.LINEAR({y},{known_y},{known_x}),0),\"\")"
                  for y in (2027, 2028, 2029)]
        write_row(ws, row, [name, "方法A　トレンド延長法"] + vals_a + ["",
                  "実績の推移を直線回帰して延長。直近の動き（精神の増加・身体の減少）をそのまま反映するが、単年の変動に引きずられやすい。"],
                  aligns=aligns, numfmts=numfmts,
                  fills=[None, None, COLORS["calc"], COLORS["calc"], COLORS["calc"], None, None])
        method_rows[(name, "A")] = row
        row += 1

        # 方法B 所持率法
        vals_b = [f"=IFERROR(ROUND($F{row}*'01_人口推計'!$D${POP_ROW[y]},0),\"\")"
                  for y in (2027, 2028, 2029)]
        write_row(ws, row, ["", "方法B　所持率法"] + vals_b + [rate,
                  "基準年の人口に対する所持率が一定と仮定し、村公式の将来人口推計に乗じる。人口減少の影響を反映するが、所持率自体の変化は捉えない。"],
                  alt=True, aligns=aligns, numfmts=numfmts,
                  fills=[None, None, COLORS["calc"], COLORS["calc"], COLORS["calc"], COLORS["calc"], None])
        method_rows[(name, "B")] = row
        row += 1

        # 中位
        ra, rb = method_rows[(name, "A")], method_rows[(name, "B")]
        vals_m = [f"=IFERROR(ROUND(AVERAGE({col}{ra},{col}{rb}),0),\"\")" for col in ("C", "D", "E")]
        write_row(ws, row, ["", "中位（A・Bの単純平均）"] + vals_m + ["", "前提の異なる2手法の平均であり、それ自体に理論的裏付けはない。計画書ではA〜Bの幅を主に示し、中位は参考値として扱う。"],
                  aligns=aligns, numfmts=numfmts,
                  fills=[None, COLORS["band"], COLORS["band"], COLORS["band"], COLORS["band"], None, None])
        for col in range(2, 6):
            ws.cell(row=row, column=col).font = Font(name=FONT, size=10, bold=True)
        method_rows[(name, "M")] = row
        row += 1

    # 合計（中位）— 概要シートが D13:F13 を参照する
    mid_rows = [method_rows[(n, "M")] for n in TEGATA]
    sums = [f"=IFERROR(SUM({col}{mid_rows[0]},{col}{mid_rows[1]},{col}{mid_rows[2]}),\"\")"
            for col in ("C", "D", "E")]
    write_row(ws, row, ["合計", "中位推計"] + sums + ["", "次期計画での暫定基礎値。令和5年実績156人との比較で読む。"],
              aligns=aligns, numfmts=numfmts, fills=[COLORS["障がい"]] * 5 + [None, None])
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    total_mid_row = row

    # 概要シートは '03_手帳将来推計'!D13〜F13 を参照している。ずれた場合に備えて検証。
    assert total_mid_row == 16, f"合計行が想定と異なります: {total_mid_row}"

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    style_title(ws.cell(row=row, column=1), "読み取りのポイント・留意点", fill=COLORS["subhead"], size=11)
    row += 1
    for txt in [
        "・中位推計では、手帳所持者数の合計は令和5年の156人から令和11年度末には145人程度まで緩やかに減少する見通しです。",
        "・方法Aと方法Bの差（身体障害者手帳は令和11年度末で102〜104人）が推計の幅です。単一の数値として計画に確定表記することは避けてください。",
        "・老年人口は令和3年の1,006人をピークに減少に転じている一方、身体障害者手帳所持者数は平成31年126人から令和5年114人へ減少しています。"
        "手帳更新時期・制度改正・転出入の影響である可能性があるため、令和6〜8年実績で継続的な傾向か一時的な変動かを確認します。",
        "・精神障害者保健福祉手帳は、人口減少が続く中でも所持者数・所持率とも増加傾向にあります。"
        "相談・受診機会の広がりを踏まえると次期計画期間も同様の傾向が続く可能性があり、相談支援体制の強化と整合させて記載します。",
        "・方法Bは基準年の所持率を将来にわたり一定と置くため、精神のように所持率自体が動いている区分では方法Aとの差が大きくなります。"
        "この差は誤差ではなく、想定の違いとして計画本文に注記してください。",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        style_note(ws.cell(row=row, column=1), txt)
        ws.row_dimensions[row].height = 26
        row += 1

    # 参考：現時点の静的計算値（Excelを開かずに検証できるようにする）
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    style_title(ws.cell(row=row, column=1),
                "参考：平成31〜令和5年の5か年実績に基づく計算値（静的・検証用）",
                fill=COLORS["subhead"], size=11)
    row += 1
    style_header_row(ws, row, ["区分", "手法", "令和9年", "令和10年", "令和11年", "", ""])
    row += 1
    static = [
        ("身体障害者手帳", "方法A", 108, 106, 104), ("", "方法B", 106, 104, 102), ("", "中位", 107, 105, 103),
        ("療育手帳", "方法A", 11, 11, 10), ("", "方法B", 11, 11, 11), ("", "中位", 11, 11, 11),
        ("精神障害者保健福祉手帳", "方法A", 33, 34, 35), ("", "方法B", 28, 27, 27), ("", "中位", 31, 31, 31),
        ("合計", "中位", 149, 147, 145),
    ]
    for i, rec in enumerate(static):
        write_row(ws, row, list(rec) + ["", ""], alt=(i % 2 == 1),
                  aligns=["left", "left", "right", "right", "right", "left", "left"],
                  numfmts=[None, None, INT_FMT, INT_FMT, INT_FMT, None, None])
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    style_note(ws.cell(row=row, column=1),
               "※ 上表は計画素案第9稿の記載値（合計 令和11年度末139人）ではなく、"
               "村公式の将来人口推計（こども・子育て計画）を用いた再計算値です。"
               "第9稿は簡易人口推計を用いていたため差が生じています。計画書には本表の値を用いてください。")
    ws.row_dimensions[row].height = 30
    return ws



# ============================================================
# 04_推計の不確実性（レッドチーム検証結果）
# ============================================================
# n=5 の単回帰から3〜6年先を外挿する際の予測区間、及び所持率法の
# 分母の選び方による感度を定量化する。値はビルド時に verify() で再計算し、
# 表示値と一致することを確認している。
UNCERTAINTY_A = [
    # 区分, 傾き, r2, 残差s, R9点推計, R9下限, R9上限, R11点推計, R11下限, R11上限
    ("身体障害者手帳", -2.30, 0.453, 4.62, 108, 76, 141, 104, 63, 144),
    ("療育手帳", -0.30, 0.321, 0.80, 11, 5, 17, 10, 3, 17),
    ("精神障害者保健福祉手帳", 1.00, 0.500, 1.83, 33, 20, 46, 35, 19, 51),
]

UNCERTAINTY_B = [
    # 区分, 総人口ベース, 老年人口ベース, 生産年齢人口ベース, 方法A
    ("身体障害者手帳", 102, 120, 94, 104),
    ("療育手帳", 11, 13, 10, 10),
    ("精神障害者保健福祉手帳", 27, 32, 25, 35),
]


def sheet_uncertainty(wb):
    ws = add_sheet(
        wb, "04_推計の不確実性", "推計の不確実性（レッドチーム検証）",
        "本ブックの推計値をそのまま計画書の確定値として用いてよいかを検証したものです。"
        "結論として、現時点の推計は「方向性の目安」であり、令和6〜8年度実績の受領前に確定値として"
        "記載することは避けてください。",
        [24, 12, 12, 12, 16, 20, 40])

    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1),
                "1. 方法A（トレンド延長法）の統計的信頼性", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "傾き\n(人/年)", "決定係数\nr²", "残差\n標準誤差",
                             "令和9年\n点推計", "令和9年\n95%予測区間", "評価"])
    r += 1
    evals = {
        "身体障害者手帳": "r²が0.45で、実績の変動の半分以上を直線では説明できない。予測区間が実測値の幅を大きく超える。",
        "療育手帳": "母数12〜14人に対し傾き−0.3人/年。実質的にノイズであり、10人と11人を区別する意味はない。",
        "精神障害者保健福祉手帳": "増加傾向自体は一貫しているが、予測区間は20〜46人と広く、点推計を目標値の根拠にはできない。",
    }
    for i, (name, slope, r2, se, p9, l9, u9, p11, l11, u11) in enumerate(UNCERTAINTY_A):
        write_row(ws, r, [name, slope, r2, se, f"{p9}人", f"{l9}〜{u9}人", evals[name]],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "center", "left"],
                  numfmts=[None, "+0.00;-0.00", "0.000", "0.00", None, None, None])
        ws.row_dimensions[r].height = 34
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "令和11年（目標年度）ではさらに広がり、身体障害者手帳の95%予測区間は63〜144人、"
               "精神障害者保健福祉手帳は19〜51人、療育手帳は3〜17人となります。"
               "標本が5か年しかなく、そこから3〜6年先を外挿しているためです。"
               "点推計（104人・35人・10人）は中心値にすぎず、統計的な確からしさを主張できる数値ではありません。")
    ws.row_dimensions[r].height = 44
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1),
                "2. 方法B（所持率法）の分母による感度（令和11年）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "総人口\nベース", "老年人口\nベース", "生産年齢\n人口ベース",
                             "方法A\n（参考）", "最大−最小", "考え方"])
    r += 1
    notes_b = {
        "身体障害者手帳": "所持者は高齢層に集中するため、総人口ではなく老年人口を分母にすると120人となり、18人増える。仕様書の約130人に近づく。",
        "療育手帳": "分母によらず11〜13人。母数が小さく差も小さいが、そもそも推計の意味が薄い。",
        "精神障害者保健福祉手帳": "所持者は生産年齢層が中心とみられ、分母の選択で25〜32人と振れる。方法Aの35人とも乖離する。",
    }
    for i, (name, tot, old, work, a) in enumerate(UNCERTAINTY_B):
        write_row(ws, r, [name, tot, old, work, a, max(tot, old, work) - min(tot, old, work),
                          notes_b[name]],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "right", "left"],
                  numfmts=[None, INT_FMT, INT_FMT, INT_FMT, INT_FMT, INT_FMT, None])
        ws.row_dimensions[r].height = 34
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "本ブックの方法Bは総人口を分母としています。北塩原村は総人口が減る一方で老年人口は増えるため、"
               "高齢層に偏る身体障害者手帳では総人口ベースの所持率法が推計を実態と逆方向に働かせます。"
               "ただしこれは所持率法そのものの欠陥ではなく、分母の取り方の問題です。"
               "年齢階級別の所持率を用いれば所持率法は有効に使えるため、"
               "村から年齢階級別の手帳所持者数を受領した時点で05_年齢階級別推計（方法C）へ切り替えます。")
    ws.row_dimensions[r].height = 34
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1),
                "3. 仕様書の調査対象者数との突合（最重要）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "令和5年\n実績", "仕様書\n（約）", "差", "増減率",
                             "本ブックの推計方向", "評価"])
    r += 1
    spec_rows = [
        ("身体障害者手帳", 114, 130, "減少（令和11年 102〜104人）"),
        ("療育手帳", 12, 20, "横ばい〜減少（10〜11人）"),
        ("精神障害者保健福祉手帳", 30, 40, "増加（27〜35人）"),
    ]
    for i, (name, r5, spec, direction) in enumerate(spec_rows):
        write_row(ws, r, [name, r5, spec, spec - r5, (spec - r5) / r5, direction,
                          "仕様書の数値が令和8年時点の実数であれば、推計の方向と矛盾する"],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "left", "left"],
                  numfmts=[None, INT_FMT, INT_FMT, "+0;-0", "+0%", None, None])
        ws.row_dimensions[r].height = 30
        r += 1
    write_row(ws, r, ["合計", 156, 190, 34, 34 / 156, "減少（145人）",
                      "同期間の総人口は約5%減。手帳所持者だけが22%増える説明が必要"],
              aligns=["left", "right", "right", "right", "right", "left", "left"],
              numfmts=[None, INT_FMT, INT_FMT, "+0;-0", "+0%", None, None],
              fills=[COLORS["note"]] * 7)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "これは確認事項ではなく、推計を確定できない前提条件です。仕様書の約190人が"
               "（a）発送数に余裕を見た概数なのか、（b）集計基準・対象年齢が異なるのか、"
               "（c）実際に令和6〜8年で増加しているのか、のいずれであるかによって、"
               "推計の方向（減少か増加か）そのものが変わります。"
               "村への確認が済むまで、手帳所持者数の将来推計を計画書の確定値として記載しないでください。")
    ws.row_dimensions[r].height = 48
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1),
                "4. 手帳所持率の推移と方法D（所持率トレンド法）", fill=COLORS["subhead"], size=11)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "現行計画本編（令和6年3月）の受領により、平成31年から令和5年までの年齢3区分別人口が"
               "判明したため、各年の手帳所持率を算定できるようになりました。"
               "方法Bは令和5年の所持率が将来も一定と仮定していますが、実際には所持率が上昇しています。")
    ws.row_dimensions[r].height = 30
    r += 1
    style_header_row(ws, r, ["区分", "平成31年", "令和2年", "令和3年", "令和4年", "令和5年", "傾向"])
    r += 1
    for i, (name, rates, trend) in enumerate(RATE_HISTORY):
        write_row(ws, r, [name] + list(rates) + [trend], alt=(i % 2 == 1),
                  aligns=["left"] + ["right"] * 5 + ["left"],
                  numfmts=[None] + [RATE_FMT] * 5 + [None])
        r += 1
    r += 1
    style_header_row(ws, r, ["区分", "方法A\nトレンド延長", "方法B\n所持率一定",
                             "方法D\n所持率トレンド", "中位\n（A・B平均）", "", "評価"])
    r += 1
    for i, (name, a, b, d, m, note) in enumerate(METHOD_D):
        write_row(ws, r, [name, a, b, d, m, "", note], alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "left", "left"],
                  numfmts=[None, INT_FMT, INT_FMT, INT_FMT, INT_FMT, None, None])
        ws.row_dimensions[r].height = 30
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "精神障害者保健福祉手帳では、所持率トレンドを延長した方法Dが35人となり、"
               "実数のトレンドを延長した方法Aの35人と一致します。"
               "所持率一定を仮定する方法Bの27人は、所持率が上昇している実態を反映できていません。"
               "一方で身体障害者手帳は、方法Dが111人と方法A（104人）・方法B（102人）を上回ります。"
               "母数が小さく所持率の年次変動が大きいため、方法Dも単独では確定値になりません。"
               "村から年齢階級別の手帳所持者数を受領し、05_年齢階級別推計（方法C）へ移行することが本筋です。")
    ws.row_dimensions[r].height = 56
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_title(ws.cell(row=r, column=1),
                "5. 過去稿との数値の異同", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["版", "方法Bの人口基礎", "令和11年\n総人口", "身体", "療育", "精神", "合計（中位）"])
    r += 1
    hist = [
        ("計画素案 第4〜9稿", "令和5年から75人/年の定数減（線形）", 1993, "93人", "10人", "25人", "139人"),
        ("将来推計ワークブック(20260707)", "こども・子育て計画の公式推計", 2185, "102人", "11人", "27人", "145人"),
        ("本ブック", "こども・子育て計画の公式推計", 2185, "102人", "11人", "27人", "145人"),
    ]
    for i, rec in enumerate(hist):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "right", "right", "right", "right", "right"],
                  numfmts=[None, None, INT_FMT, None, None, None, None])
        ws.row_dimensions[r].height = 28
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "計画素案第4〜9稿の方法Bは、本文では「人口減少率をそのまま延長した簡易推計」と説明していますが、"
               "実際の計算は率ではなく定数人数減（平成31年→令和5年の300人減÷4年＝75人/年）です。"
               "記載と計算が一致していません。また、第8稿で村公式推計を「第2章2-2の基礎として明記した」と"
               "整理していますが、2-2の数値は第7稿から変更されておらず、実際には反映されていません。"
               "本ブックは公式推計に統一しているため、第4〜9稿及び前回計画評価表（20260703、身体104→99等）の"
               "記載とは一致しません。計画書への転記時は本ブックの値に揃えてください。")
    ws.row_dimensions[r].height = 62
    return ws



# ============================================================
# 05_年齢階級別推計（方法C）
# ============================================================
# 総人口を分母とする所持率法（方法B）は、年齢構成の変化を捉えられない。
# 北塩原村は総人口が減る一方で老年人口が増えるため、身体障害者手帳の
# ように高齢層に偏る区分では推計が逆方向に働く。
# 年齢階級別の所持率を用いれば所持率法は有効に使えるため、村から
# 年齢階級別の手帳所持者数を受領した時点で本シートを正式版とする。
def sheet_age_band(wb):
    ws = add_sheet(
        wb, "05_年齢階級別推計", "年齢階級別所持率法（方法C・村資料受領後に使用）",
        "手帳の所持は年齢構成に強く左右されます。総人口を分母とする方法Bは簡便法であり、"
        "年齢階級別人口と年齢階級別所持率が揃った時点で本シートの方法Cへ切り替えてください。"
        "算式は「年齢区分別人口 × 年齢区分別手帳所持率」です。",
        [20, 14, 14, 14, 14, 14, 16, 40])

    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    style_title(ws.cell(row=r, column=1), "1. 身体障害者手帳（5区分）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年齢区分", "令和8年\n人口", "令和8年\n所持者数", "所持率",
                             "令和11年\n人口", "令和11年\n推計", "", "備考"])
    r += 1
    notes5 = {
        "0〜17歳": "障がい児支援の対象年齢。就学状況とあわせて確認する",
        "18〜39歳": "就労・日中活動の中心層",
        "40〜64歳": "特定疾病により介護保険第2号被保険者となる可能性がある層",
        "65〜74歳": "介護保険優先原則の対象。前期高齢者",
        "75歳以上": "後期高齢者。人口が増えるため所持者数も増える可能性が高い",
    }
    first = r
    for i, band in enumerate(AGE_BANDS_PHYSICAL):
        write_row(ws, r, [band, None, None,
                          f"=IFERROR(C{r}/B{r},\"\")", None,
                          f"=IFERROR(ROUND(D{r}*E{r},0),\"\")", "", notes5[band]],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "right", "left", "left"],
                  numfmts=[None, INT_FMT, INT_FMT, RATE_FMT, INT_FMT, INT_FMT, None, None],
                  fills=[None, COLORS["input"], COLORS["input"], COLORS["calc"],
                         COLORS["input"], COLORS["calc"], None, None])
        ws.row_dimensions[r].height = 26
        r += 1
    write_row(ws, r, ["合計", f"=SUM(B{first}:B{r-1})", f"=SUM(C{first}:C{r-1})",
                      f"=IFERROR(C{r}/B{r},\"\")", f"=SUM(E{first}:E{r-1})",
                      f"=SUM(F{first}:F{r-1})", "",
                      "方法A・方法B（03_手帳将来推計）と突合する"],
              aligns=["left", "right", "right", "right", "right", "right", "left", "left"],
              numfmts=[None, INT_FMT, INT_FMT, RATE_FMT, INT_FMT, INT_FMT, None, None],
              fills=[COLORS["band"]] * 8)
    for c in range(1, 7):
        ws.cell(row=r, column=c).font = Font(name=FONT, size=10, bold=True)
    r += 2

    for title, bands, note in (
        ("2. 療育手帳（3区分・コーホート移行）", AGE_BANDS_OTHER,
         "18歳未満から18〜64歳、64歳から65歳以上への移動をコーホートとして追跡します。"
         "母数が小さいため、推計値は1人単位で丸めてください。"),
        ("3. 精神障害者保健福祉手帳（3区分）", AGE_BANDS_OTHER,
         "人口比だけでなく、過去3〜5年の交付数・新規交付・更新・返還・死亡・転出、"
         "及び自立支援医療受給者数を併用して補正します。"),
    ):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_title(ws.cell(row=r, column=1), title, fill=COLORS["subhead"], size=11)
        r += 1
        style_header_row(ws, r, ["年齢区分", "令和8年\n人口", "令和8年\n所持者数", "所持率",
                                 "令和11年\n人口", "令和11年\n推計", "", "備考"])
        r += 1
        first = r
        for i, band in enumerate(bands):
            write_row(ws, r, [band, None, None,
                              f"=IFERROR(C{r}/B{r},\"\")", None,
                              f"=IFERROR(ROUND(D{r}*E{r},0),\"\")", "", ""],
                      alt=(i % 2 == 1),
                      aligns=["left", "right", "right", "right", "right", "right", "left", "left"],
                      numfmts=[None, INT_FMT, INT_FMT, RATE_FMT, INT_FMT, INT_FMT, None, None],
                      fills=[None, COLORS["input"], COLORS["input"], COLORS["calc"],
                             COLORS["input"], COLORS["calc"], None, None])
            r += 1
        write_row(ws, r, ["合計", f"=SUM(B{first}:B{r-1})", f"=SUM(C{first}:C{r-1})",
                          f"=IFERROR(C{r}/B{r},\"\")", f"=SUM(E{first}:E{r-1})",
                          f"=SUM(F{first}:F{r-1})", "", ""],
                  aligns=["left", "right", "right", "right", "right", "right", "left", "left"],
                  numfmts=[None, INT_FMT, INT_FMT, RATE_FMT, INT_FMT, INT_FMT, None, None],
                  fills=[COLORS["band"]] * 8)
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = Font(name=FONT, size=10, bold=True)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_note(ws.cell(row=r, column=1), note)
        ws.row_dimensions[r].height = 30
        r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    style_title(ws.cell(row=r, column=1), "4. 手帳推計とサービス見込量の関係",
                fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "・手帳所持者数の推計と、障害福祉サービス見込量の推計は分離します。手帳所持者数が増減しても、"
        "サービス利用者数が同じ割合で増減するとは限りません。",
        "・サービス見込量は「北塩原村_サービス見込量.xlsx」の個別積上げによります。"
        "手帳推計はサービス見込量の直接の根拠ではなく、計画の基礎データ及び相談支援体制・"
        "権利擁護等の施策規模を考える材料として用います。",
        "・北塩原村は母数が小さいため、推計値は1人単位で丸めます。小数点以下の統計値をそのまま計画値にしないでください。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# ============================================================
# 06_推計方法
# ============================================================
def sheet_method(wb):
    ws = add_sheet(
        wb, "06_推計方法", "推計方法と前提",
        "計画書 第2章及びアンケート調査報告書に転記するための、推計方法の説明です。",
        [22, 62, 58])
    style_header_row(ws, 5, ["方法", "考え方", "特性・留意点"])
    rows = [
        ("方法A　トレンド延長法",
         "手帳所持者数（実数）の推移を直線回帰し、令和9〜11年度まで延長する。回帰の対象期間は実績が入力されている年数（現在は平成31〜令和5年の5か年）に自動的に追随する。",
         "直近の実数の動きをそのまま反映するが、単年の変動に引きずられやすい。北塩原村のように母数が小さい場合、1〜2人の増減が傾きに影響する。"),
        ("方法B　所持率法",
         "基準年（現在は令和5年）の人口に対する手帳所持率が今後も一定と仮定し、村公式の将来人口推計（こども・子育て計画のコーホート変化率法）に乗じる。",
         "人口減少の影響を反映するが、所持率自体の変化（精神の増加傾向等）は捉えない。人口推計が村公式値であるため、村の他計画と整合が取りやすい。"),
        ("中位推計",
         "方法Aと方法Bの単純平均。",
         "前提の異なる2手法の平均であり、それ自体に理論的裏付けはない。計画書ではA〜Bの幅を主に示し、中位は参考値とする。"),
        ("方法C　年齢階級別所持率法\n（村資料受領後の正式版）",
         "年齢区分別人口 × 年齢区分別手帳所持率。身体障害者手帳は0〜17歳／18〜39歳／40〜64歳／65〜74歳／75歳以上の5区分、"
         "療育手帳・精神障害者保健福祉手帳は18歳未満／18〜64歳／65歳以上の3区分。",
         "年齢構成の変化を織り込めるため、総人口を分母とする方法Bの欠点を解消する。"
         "療育手帳はコーホート移行（18歳未満→18〜64歳→65歳以上）を反映する。"
         "精神障害者保健福祉手帳は人口比だけでなく交付数・新規交付・更新・返還等の実績を併用する。"
         "村から年齢階級別の手帳所持者数を受領した時点で、本方法を正式版とする（05_年齢階級別推計）。"),
    ]
    r = 6
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 66
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_title(ws.cell(row=r, column=1), "人口推計の前提", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "・令和11年の人口は、北塩原村こども・子育て計画（令和7年3月策定）が示すコーホート変化率法による村公式推計を用いています（総人口2,185人、年少111人、生産年齢1,023人、老年1,051人、高齢化率48.1％）。",
        "・令和7〜10年は、令和6年実績（2,394人）と令和11年公式推計（2,185人）の線形補間です。村独自の人口ビジョンや社人研推計がある場合は、そちらに置き換えて再計算してください。",
        "・計画素案第9稿では、平成31〜令和5年の人口減少率をそのまま延長した簡易推計を方法Bに用いていました。本ブックでは村公式推計に統一しています。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# ============================================================
# 05_村確認事項
# ============================================================
def sheet_confirm(wb):
    ws = add_sheet(
        wb, "07_村確認事項", "村への確認・依頼事項（将来推計関係）",
        "受領・確認が済んだものは「状態」を更新し、反映先シートに入力してください。",
        [28, 52, 26, 12, 34])
    style_header_row(ws, 5, ["確認事項", "確認したい内容", "使途・反映先", "優先度", "状態・回答"])
    rows = [
        ("令和6〜8年度の手帳所持者数",
         "身体・療育・精神の各年4月1日現在の所持者数。",
         "02_手帳実績／03_手帳将来推計", "高", "村資料待ち"),
        ("年齢階級別の手帳所持者数",
         "身体は0〜17歳／18〜39歳／40〜64歳／65〜74歳／75歳以上、療育・精神は18歳未満／18〜64歳／"
         "65歳以上の区分。方法C（年齢階級別所持率法）へ切り替えるために必須。",
         "05_年齢階級別推計", "最優先", "村資料待ち"),
        ("年齢階級別の人口",
         "令和8年及び令和11年の年齢階級別人口。上記と同じ区分。"
         "こども・子育て計画の推計に年齢階級別の内訳があればそれを用いる。",
         "05_年齢階級別推計", "最優先", "村資料待ち"),
        ("令和2〜4年、令和7〜8年の人口",
         "総人口・年少人口・生産年齢人口・老年人口（各年4月1日現在）。",
         "01_人口推計", "高", "村資料待ち"),
        ("調査対象者数と手帳所持者数の差異",
         "仕様書の約190人（身体約130・精神約40・療育約20）と、現行計画の156人（令和5年）の差の理由。"
         "集計基準・対象年齢・サービス利用者の重複計上の違いを確認する。",
         "02_手帳実績／アンケート発送設計", "高", "村資料待ち"),
        ("村独自の人口ビジョン・将来推計",
         "総合振興計画の人口ビジョン、社人研推計等がある場合はその数値。",
         "01_人口推計／06_推計方法", "中", "村資料待ち"),
        ("身体障害者手帳の減少要因",
         "老年人口が横ばい〜微減である一方、身体障害者手帳所持者数が減少している要因。"
         "手帳更新時期・制度改正・転出入・死亡等の影響を確認する。",
         "03_手帳将来推計／04_推計の不確実性", "中", "村資料待ち"),
        ("精神障害者保健福祉手帳の増加要因",
         "所持者数・所持率とも増加傾向にある背景（受診機会の広がり、相談体制、新規申請の動向等）。",
         "03_手帳将来推計／相談支援施策", "中", "村資料待ち"),
        ("障がい児の就学状況",
         "特別支援学校・特別支援学級の在籍者数（令和8年4月1日現在）、卒業予定年度。",
         "計画書 第2章／障がい児福祉計画", "中", "村資料待ち"),
    ]
    r = 6
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "left", "center", "center"])
        style_status(ws.cell(row=r, column=5))
        ws.row_dimensions[r].height = 42
        r += 1
    return ws


# ============================================================
# 自己検証
# ============================================================
# シートに埋め込んだ静的参照値（03_手帳将来推計の末尾表）。
# Excel数式（FORECAST.LINEAR／所持率法／平均）と同じ計算をPythonで再現し、
# 数式の組み立てを誤っていないことをビルド時に確認する。
STATIC_EXPECTED = {
    "身体障害者手帳":         {"A": (108, 106, 104), "B": (106, 104, 102), "M": (107, 105, 103)},
    "療育手帳":               {"A": (11, 11, 10),    "B": (11, 11, 11),    "M": (11, 11, 11)},
    "精神障害者保健福祉手帳": {"A": (33, 34, 35),    "B": (28, 27, 27),    "M": (31, 31, 31)},
}
STATIC_TOTAL_MID = (149, 147, 145)

TARGET_YEARS = (2027, 2028, 2029)


def _forecast_linear(x, ys, xs):
    """Excel の FORECAST.LINEAR と同じ単回帰予測。"""
    n = len(ys)
    mx = sum(xs) / n
    my = sum(ys) / n
    sxy = sum((xi - mx) * (yi - my) for xi, yi in zip(xs, ys))
    sxx = sum((xi - mx) ** 2 for xi in xs)
    slope = sxy / sxx
    return my + slope * (x - mx)


def _round_half_up(v):
    """Excel の ROUND（四捨五入）。Python の round は銀行丸めのため使わない。"""
    from decimal import Decimal, ROUND_HALF_UP
    return int(Decimal(str(v)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _interpolated_population(year):
    """01_人口推計の線形補間数式と同じ計算。"""
    base = dict((rec[1], rec[3]) for rec in POPULATION)
    r6, r11 = base[2024], base[2029]
    if year in (2024, 2029):
        return base[year]
    step = year - 2024
    return _round_half_up(r6 + (r11 - r6) * step / 5)


def verify():
    base_pop = dict((rec[1], rec[3]) for rec in POPULATION)[2023]  # 令和5年 2,443人
    totals = [0, 0, 0]
    for name, actual in TEGATA.items():
        xs = [s for (_, s), v in zip(TEGATA_YEARS, actual) if v is not None]
        ys = [v for v in actual if v is not None]

        got_a = tuple(_round_half_up(_forecast_linear(y, ys, xs)) for y in TARGET_YEARS)
        rate = ys[xs.index(2023)] / base_pop
        got_b = tuple(_round_half_up(rate * _interpolated_population(y)) for y in TARGET_YEARS)
        got_m = tuple(_round_half_up((a + b) / 2) for a, b in zip(got_a, got_b))

        exp = STATIC_EXPECTED[name]
        assert got_a == exp["A"], f"{name} 方法A: 計算値{got_a} ≠ 表示値{exp['A']}"
        assert got_b == exp["B"], f"{name} 方法B: 計算値{got_b} ≠ 表示値{exp['B']}"
        assert got_m == exp["M"], f"{name} 中位: 計算値{got_m} ≠ 表示値{exp['M']}"
        totals = [t + m for t, m in zip(totals, got_m)]

    assert tuple(totals) == STATIC_TOTAL_MID, f"合計（中位）: 計算値{tuple(totals)} ≠ 表示値{STATIC_TOTAL_MID}"

    # 人口の補間と内訳合計の整合
    assert _interpolated_population(2027) == 2269, "令和9年 総人口の補間が想定と異なります"
    assert _interpolated_population(2029) == 2185, "令和11年 総人口が公式推計と一致しません"
    print("  自己検証: 手帳推計3区分×3手法・合計・人口補間 いずれも一致")
    _verify_uncertainty()


def _verify_uncertainty():
    """04_推計の不確実性に表示した統計値を再計算して突合する。"""
    import math
    xs = [s for (_, s) in TEGATA_YEARS[:5]]
    xbar = sum(xs) / len(xs)
    sxx = sum((x - xbar) ** 2 for x in xs)
    t975 = 3.182  # t(0.975, df=3)

    pop = dict((rec[1], (rec[3], rec[5], rec[6])) for rec in POPULATION)  # 総/生産年齢/老年
    base_tot, base_work, base_old = pop[2023]
    proj_tot, proj_work, proj_old = pop[2029]

    for name, slope_x, r2_x, se_x, p9, l9, u9, p11, l11, u11 in UNCERTAINTY_A:
        ys = [v for v in TEGATA[name] if v is not None]
        ybar = sum(ys) / len(ys)
        sxy = sum((x - xbar) * (y - ybar) for x, y in zip(xs, ys))
        syy = sum((y - ybar) ** 2 for y in ys)
        slope = sxy / sxx
        r2 = sxy ** 2 / (sxx * syy)
        sse = syy - sxy ** 2 / sxx
        s_res = math.sqrt(sse / 3)
        assert abs(slope - slope_x) < 0.005, f"{name} 傾き {slope:.3f} ≠ 表示 {slope_x}"
        assert abs(r2 - r2_x) < 0.001, f"{name} r² {r2:.4f} ≠ 表示 {r2_x}"
        assert abs(s_res - se_x) < 0.005, f"{name} 残差s {s_res:.3f} ≠ 表示 {se_x}"
        for x0, pt, lo, hi in ((2027, p9, l9, u9), (2029, p11, l11, u11)):
            pred = ybar + slope * (x0 - xbar)
            pi = t975 * s_res * math.sqrt(1 + 1 / len(ys) + (x0 - xbar) ** 2 / sxx)
            assert abs(_round_half_up(pred) - pt) <= 1, f"{name} {x0} 点推計 {pred:.1f} ≠ 表示 {pt}"
            assert abs(_round_half_up(pred - pi) - lo) <= 1, f"{name} {x0} 下限 {pred-pi:.1f} ≠ 表示 {lo}"
            assert abs(_round_half_up(pred + pi) - hi) <= 1, f"{name} {x0} 上限 {pred+pi:.1f} ≠ 表示 {hi}"

    for name, tot_x, old_x, work_x, a_x in UNCERTAINTY_B:
        r5 = TEGATA[name][4]
        got = (_round_half_up(r5 / base_tot * proj_tot),
               _round_half_up(r5 / base_old * proj_old),
               _round_half_up(r5 / base_work * proj_work))
        assert got == (tot_x, old_x, work_x), f"{name} 分母感度 {got} ≠ 表示 ({tot_x}, {old_x}, {work_x})"

    # 過去稿の方法B（75人/年の定数減）が第9稿の記載値を再現することの確認
    for name, expected in (("身体障害者手帳", 93), ("療育手帳", 10)):
        r5 = TEGATA[name][4]
        assert _round_half_up(r5 / base_tot * (base_tot - 75 * 6)) == expected, \
            f"過去稿の方法B（{name}）が第9稿記載値 {expected}人 を再現しません"
    _verify_rates()
    print("  自己検証: 予測区間・分母感度・過去稿方法Bの再現 いずれも一致")


def _verify_rates():
    """手帳所持率の推移と方法Dの推計値を再計算して表示値と突合する。"""
    pop = dict((rec[1], rec[3]) for rec in POPULATION)
    years = [2019, 2020, 2021, 2022, 2023]
    xbar = sum(years) / len(years)
    sxx = sum((x - xbar) ** 2 for x in years)

    series = {name: [v for v in vals[:5]] for name, vals in TEGATA.items()}
    series["合計"] = [sum(series[k][i] for k in TEGATA) for i in range(5)]

    d_calc = {}
    for name, shown, _ in RATE_HISTORY:
        rates = [series[name][i] / pop[y] for i, y in enumerate(years)]
        for got, exp in zip(rates, shown):
            assert abs(got - exp) < 5e-5, f"{name} 所持率 {got:.5f} ≠ 表示 {exp}"
        ybar = sum(rates) / len(rates)
        sxy = sum((x - xbar) * (r - ybar) for x, r in zip(years, rates))
        d_calc[name] = _round_half_up((ybar + sxy / sxx * (2029 - xbar)) * pop[2029])

    for name, a, b, d, m, _ in METHOD_D:
        assert d_calc[name] == d, f"{name} 方法D {d_calc[name]} ≠ 表示 {d}"


def main():
    verify()
    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb)
    sheet_population(wb)
    sheet_tegata_actual(wb)
    sheet_tegata_projection(wb)
    sheet_uncertainty(wb)
    sheet_age_band(wb)
    sheet_method(wb)
    sheet_confirm(wb)
    wb.save(OUT_FILE)
    print(f"作成: {OUT_FILE}")


if __name__ == "__main__":
    main()
