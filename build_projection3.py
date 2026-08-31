# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　将来推計 第3段階（給付費と保険料）.

第1段階（人口と認定者数）・第2段階（サービス見込量）に続く第3段階として、
標準給付費見込額から保険料基準額までを組み立てる。

令和8年8月28日及び31日の受領資料により前提がほぼ揃ったため、
確定していない事項（第10期の第1号被保険者負担割合、報酬改定率、
令和7年度末の基金残高、所得段階の政令改正）を明示したうえで
暫定の算定を行う。確定値ではない。

第9期の算定ロジックは「保険料の所得段階と低所得者軽減の検証」06シートで
再現済みであり、本表はその式に第10期の前提を入れたものである。

シート構成
  00_この推計について    前提・確定していない事項・暫定である理由
  01_標準給付費見込額     自然体給付費に令和6年度決算による割増しを適用する
  02_地域支援事業費       令和6年度決算の水準による
  03_収納必要額の組立て    第1号負担分・調整交付金・基金取崩し
  04_補正後被保険者数      所得段階別加入割合補正後の被保険者数
  05_保険料基準額の暫定算定  基本ケース
  06_感度分析          6つのレバーを1つずつ動かす
  07_基金の取崩シナリオ    条例第5条第3項の上限との関係
  08_所得段階別の月額      16段階の暫定月額
  09_見える化の自然体推計との対比
  10_確定を要する事項
"""

import io
import runpy
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_kessan_r6 as K
import data_kikin_jorei as J
import data_nenpo as N
import data_shien_tool as T

TOWNS3 = ["東川町", "美瑛町", "東神楽町"]

OUT = ("/home/user/repository/output/"
       "第10期計画_将来推計_第3段階_給付費と保険料.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F4E78", "5B9BD5"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# ------------------------------------------------------------------ 前提
YEARS = ["令和9年度", "令和10年度", "令和11年度"]

# 見える化システムの自然体推計による給付費（百万円）。計画素案 第6章第3節1と同じ。
SHIZENTAI = [
    ("居宅サービス給付費", (1076.2, 1098.1, 1118.2)),
    ("地域密着型サービス給付費", (832.0, 842.1, 849.4)),
    ("施設サービス給付費", (897.2, 897.2, 897.2)),
    ("介護予防サービス給付費", (127.1, 127.5, 128.3)),
    ("居宅介護支援給付費", (124.2, 125.3, 127.4)),
]
KYUFU_IPSS = [sum(v[i] for _n, v in SHIZENTAI) for i in range(3)]   # 3,056.7 …

# ------------------------------------------------ 人口の基礎（案C）への置換え
# 令和8年8月31日のご指示により、人口は社人研ではなく総合戦略上の人口を
# 基礎とすることとなった。将来推計 第1段階では、総人口を総合戦略に置き、
# 年齢階級別は住民基本台帳の実績趨勢で延ばす案Cを採っている。
# 見える化システムの自然体推計は社人研を基礎としているため、
# 給付費は認定者数の比（案C÷案A）で置き直す。
_buf, _old = io.StringIO(), sys.stdout
sys.stdout = _buf
try:
    _G = runpy.run_path("build_projection.py")
finally:
    sys.stdout = _old

_Y3 = ["2027", "2028", "2029"]
_SC = 2                                  # ② トレンド継続（採用シナリオ）
_CL, _rate = _G["CL"], _G["rate"]


def _nintei(f, y):
    return sum(f(c, y) * _rate(c, y, _SC) / 100 for c in _CL)


NIN_IPSS = [_nintei(_G["pop_ipss"], y) for y in _Y3]      # 案A（社人研）
NIN_JUKI = [_nintei(_G["pop_juki"], y) for y in _Y3]      # 案C（採用）
KYUFU_HOSEI = [c / a for c, a in zip(NIN_JUKI, NIN_IPSS)]
KYUFU = [k * h for k, h in zip(KYUFU_IPSS, KYUFU_HOSEI)]

# 推計上の第1号被保険者数（将来推計 第1段階 02シート・案C）
HIHOKEN = [round(sum(_G["pop_juki"](c, y) for c in _CL)) for y in _Y3]
HIHOKEN_IPSS = [9080, 9078, 9072]        # 案A（社人研）による従前の値

# 第9期の条例による所得段階の乗率（16段階）
JORITSU = [0.455, 0.639, 0.690, 0.860, 1.000, 1.260, 1.310, 1.590,
           1.630, 1.900, 2.165, 2.365, 2.565, 2.575, 2.585, 2.595]
DANKAI_KEY = ["第1段階", "第2段階", "第3段階", "第4段階", "第5段階", "第6段階",
              "第7段階", "第8段階", "第9段階", "第10段階", "第11段階",
              "第12段階", "第13段階", "第13段階（多段階）2580",
              "第13段階（多段階）2590", "第13段階（多段階）2600"]
DANKAI_NAME = ["第1段階", "第2段階", "第3段階", "第4段階", "第5段階（基準）",
               "第6段階", "第7段階", "第8段階", "第9段階", "第10段階",
               "第11段階", "第12段階", "第13段階", "第14段階", "第15段階",
               "第16段階"]

# 第9期計画の値（保険料の所得段階と低所得者軽減の検証 06シートで再現済み）
K9 = {
    "A": 9519211524, "B": 593658000, "C": 2325959991, "D": 492515476,
    "E": 726507000, "I": 40000000, "J": 2051968467, "hosei": 26871,
    "shuno": 0.99, "gaku": 6428, "jorei": 6400, "hiho": 27724,
}


def waribiki():
    """総給付費から標準給付費見込額への割増率（令和6年度決算による）。"""
    kyufu = N.KYUFU["給付費　合計"]["R6"]
    hyojun = K.SAISHUTSU["2 保険給付費"][1]
    return hyojun / kyufu, kyufu, hyojun


UPLIFT, R6_KYUFU, R6_HYOJUN = waribiki()
UPLIFT_K9 = K9["A"] / 8905234000          # 第9期計画による割増率（見える化の計画値）

# ---------------------------------------------- 地域支援事業費（B）と総合事業費
# 令和8年8月31日の点検により、次の2点を国の様式に合わせて改めた。
#
# ① 地域支援事業費（B）から、保険者機能強化推進事業費（2,033,000円）及び
#    保険者努力支援事業費（4,190,000円）を除く。
#    これらは保険者機能強化推進交付金及び介護保険保険者努力支援交付金を
#    財源とする事業であり、当該交付金が特定財源となるため
#    第1号被保険者負担分の算定基礎には含まれない。
#    従前は決算の款4の全体（183,384,259円）を用いていた。
# ② 調整交付金の算定基礎（②）に加える総合事業費を、
#    決算科目の「介護予防・生活支援サービス事業費」（101,274,621円）から、
#    介護保険事業計画作成支援ツールの実績値シートによる
#    「介護予防・日常生活支援総合事業費」（3町計107,129,318円）に改める。
#    国の様式でいう総合事業費は介護予防・生活支援サービス事業と
#    一般介護予防事業の合計であり、決算科目とは範囲が異なる。
_KESSAN_KYOKA = K.SAISHUTSU["4 地域支援事業費／3 保険者機能強化推進事業"][1]
_KESSAN_DORYOKU = K.SAISHUTSU["4 地域支援事業費／4 保険者努力支援事業"][1]
CHIIKI_ZEN = K.SAISHUTSU["4 地域支援事業費"][1]       # 183,384,259 円（款4の全体）
CHIIKI_R6 = CHIIKI_ZEN - _KESSAN_KYOKA - _KESSAN_DORYOKU   # 177,161,259 円
CHIIKI_TOOL = sum(T.HOKENRYO[t]["地域支援事業費（円）"] for t in TOWNS3)
SOGO_KESSAN = K.SAISHUTSU[
    "4 地域支援事業費／1 介護予防・生活支援サービス事業費"][1]   # 101,274,621 円
SOGO_R6 = sum(T.HOKENRYO[t]["介護予防・日常生活支援総合事業費（円）"]
              for t in TOWNS3)                              # 107,129,318 円

# 標準給付費見込額（A）の加算項目。計画作成支援ツールの実績値シートによる。
# 総給付費にこれらを加えたものが標準給付費見込額であり、
# 令和6年度では決算の保険給付費（款2）と15,246円（0.0005％）の差で一致する。
UPLIFT_ITEM = [
    ("特定入所者介護サービス費　施設サービス居住費",
     "施設サービス　居住費（円）"),
    ("特定入所者介護サービス費　施設サービス食費",
     "施設サービス　食費（円）"),
    ("特定入所者介護サービス費　短期入所の居住費・滞在費",
     "短期入所生活介護・短期入所療養介護　居住費・滞在費（円）"),
    ("特定入所者介護サービス費　短期入所の食費",
     "短期入所生活介護・短期入所療養介護　食費（円）"),
    ("高額介護サービス費等給付額", "高額介護サービス費等給付額（円）"),
    ("高額医療合算介護サービス費等給付額",
     "高額医療合算介護サービス費等給付額（円）"),
    ("算定対象審査支払手数料", "算定対象審査支払手数料（円）"),
    ("市町村特別給付費等", "市町村特別給付費等（円）"),
]
UPLIFT_SUM = {nm: sum(T.HOKENRYO[t][k] for t in TOWNS3)
              for nm, k in UPLIFT_ITEM}
UPLIFT_ADD = sum(UPLIFT_SUM.values())

FUTAN = 0.23                               # 第1号被保険者負担割合（第9期と同じ）
CHOSEI_D = 0.05                            # 調整交付金相当額の割合

# 調整交付金の算定基礎（②）は、標準給付費見込額に
# 介護予防・日常生活支援総合事業費を加えたものである。
# 包括的支援事業費及び任意事業費は調整交付金の対象外であり、②に含まれない。
# 第9期計画の D＝492,515,476円 を 5％ で割り戻すと 9,850,309,520円 となり、
# 標準給付費見込額 9,519,211,524円 との差 331,097,996円（年110,365,999円）が
# 総合事業費に当たる。同じ基礎に対する E の割合は 7.3755％ で、
# 第9期計画の見込交付割合（R6 7.51％・R7 7.43％・R8 7.19％）の平均と一致する。
K9_SOGO = K9["D"] / CHOSEI_D - K9["A"]     # 第9期計画の総合事業費（3年計）
CHOSEI_E = K9["E"] / (K9["A"] + K9_SOGO)   # 見込交付割合（第9期計画と同水準）
SHUNO = 0.99                               # 予定保険料収納率（第9期と同じ）


def keisu(year):
    """所得段階別加入割合補正後の被保険者数の係数（年報の実績による）。"""
    n = [N.SHOTOKU[k][year] for k in DANKAI_KEY]
    return sum(a * b for a, b in zip(n, JORITSU)) / sum(n), sum(n)


KEISU_R6, N_R6 = keisu("R6")
KEISU_R7, N_R7 = keisu("R7")
KEISU_K9 = K9["hosei"] / K9["hiho"]


# 国の様式（介護保険事業計画作成支援ツール）における保険料収納必要額は
#   J ＝ C ＋ D － E ＋ F ＋ G ± H － I
# である。当広域連合では次の3項目がいずれも0であるため式に現れない。
#   F 財政安定化基金拠出金見込額（第9期・第10期とも拠出の予定なし）
#   G 財政安定化基金償還金（借入れの実績がない）
#   H 市町村相互財政安定化事業負担額・交付額（同事業に参加していない）
# 0であることの確認は「保険料算定式の国の様式との対照」による。
ANTEIKA_F = 0                              # 財政安定化基金拠出金見込額
ANTEIKA_G = 0                              # 財政安定化基金償還金
SOGO_ANTEI_H = 0                           # 市町村相互財政安定化事業


def gaku(kyufu_oku=None, uplift=None, chiiki=None, sogo=None, futan=None,
         chosei_e=None, torikuzushi=0, keisu=None, shuno=None, hiho=None,
         f=None, g=None, h=None):
    """保険料基準額（算定上の月額）を返す。既定は基本ケース。"""
    ky = sum(KYUFU) * 1e6 if kyufu_oku is None else kyufu_oku
    up = UPLIFT if uplift is None else uplift
    ch = (CHIIKI_R6 * 3) if chiiki is None else chiiki
    sg = (SOGO_R6 * 3) if sogo is None else sogo
    ft = FUTAN if futan is None else futan
    ce = CHOSEI_E if chosei_e is None else chosei_e
    ks = KEISU_R6 if keisu is None else keisu
    sh = SHUNO if shuno is None else shuno
    hh = sum(HIHOKEN) if hiho is None else hiho
    a = ky * up
    maru = a + ch                 # ① 第1号被保険者負担分の算定基礎
    maru2 = a + sg                # ② 調整交付金の算定基礎
    c = maru * ft
    d = maru2 * CHOSEI_D
    e = maru2 * ce
    ff = ANTEIKA_F if f is None else f
    gg = ANTEIKA_G if g is None else g
    hh2 = SOGO_ANTEI_H if h is None else h
    j = c + d - e + ff + gg + hh2 - torikuzushi
    hosei = hh * ks
    return {"A": a, "B": ch, "①": maru, "②": maru2, "C": c, "D": d, "E": e,
            "F": ff, "G": gg, "H": hh2,
            "I": torikuzushi, "J": j, "③": hosei,
            "月額": j / sh / hosei / 12}


BASE = gaku()


# ------------------------------------------------------------------ 部品
def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 56
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=22, align=None, bold=False,
         numfmt=None):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.border = BORDER
        ha = (align or {}).get(i, "left" if isinstance(v, str) else "right")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if numfmt and not isinstance(v, str):
            c.number_format = numfmt
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8, height=88):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


def en(v):
    return int(round(v))


# ============================================================ 00
ws = sheet("00_この推計について",
           "将来推計 第3段階（給付費と保険料）の暫定算定",
           "第1段階（人口と認定者数）・第2段階（サービス見込量）に続き、"
           "標準給付費見込額から保険料基準額までを組み立てます。"
           "令和8年8月28日及び31日の受領資料により前提がほぼ揃ったため、"
           "確定していない事項を明示したうえで暫定の算定を行いました。"
           "確定値ではありません。",
           [4, 26, 40, 34, 16], freeze="A5")

r = lead(ws, 4, "【1　算定の式】", 5)
r = header(ws, r, ["Step", "項目", "算式", "本表での置き方", "確定の別"])
for st, ko, shiki, oki, kak in [
    ("1", "標準給付費見込額（A）",
     "総給付費×割増率",
     "見える化システムの自然体推計による給付費（3年計 %.1f百万円）を"
     "人口の基礎の変更により%.4f倍に置き直し、"
     "令和6年度決算による割増率%.4fを乗じる"
     % (sum(KYUFU_IPSS), sum(NIN_JUKI) / sum(NIN_IPSS), UPLIFT), "暫定"),
    ("1", "地域支援事業費（B）", "事業費の積上げ",
     "令和6年度決算額%s円を3年分とする" % "{:,}".format(CHIIKI_R6), "暫定"),
    ("1", "合計（①）", "①＝A＋B", "―", "暫定"),
    ("2", "第1号被保険者負担分相当額（C）", "C＝①×負担割合",
     "第9期と同じ23％。第10期の割合は政令による", "未確定"),
    ("2", "調整交付金相当額（D）", "D＝①×5％", "制度上の割合", "確定"),
    ("2", "調整交付金見込額（E）", "E＝①×見込交付割合",
     "第9期計画と同水準の%.3f％とする" % (CHOSEI_E * 100), "未確定"),
    ("3", "財政調整基金取崩額（I）", "方針による",
     "基本ケースは0円。07シートで4つのシナリオを示す", "未確定"),
    ("3", "保険料収納必要額（J）", "J＝C＋D－E－I", "―", "暫定"),
    ("4", "補正後被保険者数（③）", "Σ（段階別被保険者数×乗率）",
     "推計上の第1号被保険者数3年計%s人に、"
     "令和6年度の実績による係数%.4fを乗じる"
     % ("{:,}".format(sum(HIHOKEN)), KEISU_R6), "暫定"),
    ("5", "予定保険料収納率", "実績による",
     "第9期と同じ99.0％。実績は現年度分99.8％台", "未確定"),
    ("5", "算定上の月額基準額", "J÷収納率÷③÷12", "―", "暫定"),
]:
    r = body(ws, r, [st, ko, shiki, oki, kak], height=40,
             fills={5: (OK_G if kak == "確定" else
                        IN_Y if kak == "暫定" else NG_O)},
             align={1: "center", 5: "center"})

r += 1
r = lead(ws, r, "【2　この推計で確定していない事項】", 5)
r = header(ws, r, ["No.", "事項", "現在の置き方", "確定の時期・条件", "影響"])
for no, ko, oki, jiki, eik in [
    (1, "第10期の第1号被保険者負担割合",
     "第9期と同じ23％", "国の政令（令和8年度中の見込み）",
     "1ポイントの違いで月額約290円"),
    (2, "介護報酬改定率", "改定を織り込んでいない",
     "令和9年度改定の告示",
     "1％の改定で月額約60円"),
    (3, "令和7年度末の介護給付費準備基金残高",
     "令和6年度末残高%s円で試算" % "{:,}".format(J.ZANDAKA_R6),
     "令和7年度決算又は年報様式4（令和8年12月頃）",
     "取崩額の設定に直結する"),
    (4, "所得段階の区分数と乗率",
     "第9期の16段階・乗率をそのまま用いる", "第10期の政令改正",
     "補正後被保険者数を通じて月額に効く"),
    (5, "調整交付金の見込交付割合",
     "第9期計画と同水準の%.3f％" % (CHOSEI_E * 100),
     "国の通知",
     "交付実績は第7期90％・第8期88％で計画を下回っている"),
    (6, "地域支援事業費の見込み",
     "令和6年度決算額の3年分", "事業別の実績の受領と第10期の事業設計",
     "包括的支援事業・任意事業の事業別内訳が未受領"),
    (7, "給付費の見込み",
     "見える化システムの自然体推計をそのまま用いる",
     "サービス見込量の採用値の決定（需要3シナリオの感度表 07シート）",
     "最も大きいレバー。06シート参照"),
]:
    r = body(ws, r, [no, ko, oki, jiki, eik], height=44,
             align={1: "center"})

note(ws, r + 1,
     "注1）本表は確定値ではありません。"
     "上記7件が確定するまでは、条例で採用する基準額の算定には用いられません。\n"
     "注2）第9期の算定ロジックは"
     "「保険料の所得段階と低所得者軽減の検証」06シートで再現済みです。"
     "本表はその式に第10期の前提を入れたものであり、式そのものは変えていません。\n"
     "注3）見える化システムの自然体推計は第10期3年平均6,238円です。"
     "本表の結果との差は09シートで扱います。", 5, 76)


# ============================================================ 01
ws = sheet("01_標準給付費見込額",
           "標準給付費見込額（A）の組立て",
           "見える化システムの自然体推計による給付費に、"
           "総給付費から標準給付費見込額への割増しを適用します。"
           "割増率は令和6年度の決算により算定しました。",
           [26, 14, 14, 14, 14, 34], freeze="A5")

r = lead(ws, 4, "【1　自然体推計による給付費（見える化システム）】", 6)
r = header(ws, r, ["区分"] + YEARS + ["第10期計", "備考"])
for nm, v in SHIZENTAI:
    r = body(ws, r, [nm] + [round(x, 1) for x in v] + [round(sum(v), 1), ""],
             numfmt="#,##0.0")
r = body(ws, r, ["合計"] + [round(x, 1) for x in KYUFU_IPSS]
         + [round(sum(KYUFU_IPSS), 1), "計画素案 第6章第3節1と同じ"],
         {1: MID_B}, bold=True, numfmt="#,##0.0")
r = note(ws, r, "単位：百万円。見える化システムの自然体推計による。"
         "報酬改定、地域支援事業費、供給制約、基盤整備及び政策効果は"
         "反映していません。", 6, 32)

r = lead(ws, r, "【1の2　人口の基礎を総合戦略ベースに改めたことによる置換え】", 6)
r = header(ws, r, ["区分"] + YEARS + ["第10期計", "備考"])
r = body(ws, r, ["認定者数　案A（社人研）"] + [round(x) for x in NIN_IPSS]
         + [round(sum(NIN_IPSS)), "見える化システムの自然体推計の基礎"],
         numfmt="#,##0")
r = body(ws, r, ["認定者数　案C（総合戦略ベース）"] + [round(x) for x in NIN_JUKI]
         + [round(sum(NIN_JUKI)), "将来推計 第1段階 03シート・シナリオ②"],
         {1: MID_B}, numfmt="#,##0")
r = body(ws, r, ["置換率（案C÷案A）"]
         + ["%.4f" % h for h in KYUFU_HOSEI]
         + ["%.4f" % (sum(NIN_JUKI) / sum(NIN_IPSS)), "給付費に乗じる"],
         align={1: "center", 2: "center", 3: "center", 4: "center"})
r = body(ws, r, ["置換え後の給付費"] + [round(x, 1) for x in KYUFU]
         + [round(sum(KYUFU), 1), "本表で用いる値"],
         {1: OK_G}, bold=True, numfmt="#,##0.0")
r = note(ws, r, "単位：人・百万円。"
         "見える化システムの自然体推計は社人研の推計人口を基礎としています。"
         "令和8年8月31日のご指示により人口の基礎を総合戦略ベース（案C）に"
         "改めたため、給付費を認定者数の比で置き直しました。"
         "サービス構成比は変えていません。", 6, 46)

r = lead(ws, r, "【2　総給付費から標準給付費見込額への割増率】", 6)
r = header(ws, r, ["算定の経路", "総給付費（円）", "標準給付費見込額（円）",
                   "割増率", "採否", "内容"])
r = body(ws, r, ["令和6年度決算", en(R6_KYUFU), en(R6_HYOJUN),
                 round(UPLIFT, 5), "採用",
                 "保険給付費（款2）＝介護サービス等諸費＋その他諸費＋"
                 "高額介護等サービス費＋高額医療合算介護等サービス費＋"
                 "特定入所者介護等サービス費"],
         {5: OK_G}, height=44, numfmt="#,##0", align={5: "center"})
r = body(ws, r, ["第9期計画", 8905234000, K9["A"], round(UPLIFT_K9, 5),
                 "感度",
                 "見える化システムの第9期計画給付費8,905,234千円と、"
                 "第9期計画の標準給付費見込額との比。06シートで感度を示す"],
         {5: IN_Y}, height=44, numfmt="#,##0", align={5: "center"})

r += 1
r = lead(ws, r, "【2の2　割増率の項目別の検算（国の様式との対照）】", 6)
r = header(ws, r, ["No.", "項目", "令和6年度（円）", "総給付費に対する率",
                   "出所", "内容"])
for i, (nm, _key) in enumerate(UPLIFT_ITEM, start=1):
    v = UPLIFT_SUM[nm]
    r = body(ws, r, [i, nm, en(v), "%.4f％" % (v / R6_KYUFU * 100),
                     "計画作成支援ツール 実績値シート（3町計）",
                     "標準給付費見込額に含める項目"],
             height=22, numfmt="#,##0", align={1: "center", 4: "right"})
r = body(ws, r, ["", "加算計", en(UPLIFT_ADD),
                 "%.4f％" % (UPLIFT_ADD / R6_KYUFU * 100), "―", "―"],
         {2: MID_B}, bold=True, height=22, numfmt="#,##0",
         align={4: "right"})
r = body(ws, r, ["", "総給付費＋加算計", en(R6_KYUFU + UPLIFT_ADD),
                 "%.5f" % ((R6_KYUFU + UPLIFT_ADD) / R6_KYUFU),
                 "―", "項目の積上げによる標準給付費見込額"],
         {2: OK_G}, bold=True, height=22, numfmt="#,##0",
         align={4: "right"})
r = body(ws, r, ["", "保険給付費（款2・令和6年度決算）", en(R6_HYOJUN),
                 "%.5f" % UPLIFT, "―", "現行の割増率の算定に用いている値"],
         {2: OK_G}, bold=True, height=22, numfmt="#,##0",
         align={4: "right"})
_sa = R6_KYUFU + UPLIFT_ADD - R6_HYOJUN
r = body(ws, r, ["", "差", en(_sa),
                 "%.6f％" % (abs(_sa) / R6_HYOJUN * 100), "―",
                 "決算の介護サービス等諸費2,817,209,701円と"
                 "年報の給付費合計2,817,194,455円との差に一致する"],
         {2: IN_Y}, height=30, numfmt="#,##0", align={4: "right"})
r = note(ws, r, "国の様式（介護保険事業計画作成支援ツール）では、"
         "標準給付費見込額は総給付費に"
         "特定入所者介護サービス費・高額介護サービス費・"
         "高額医療合算介護サービス費・審査支払手数料・市町村特別給付費を"
         "加えて求めます。"
         "本表の割増率%.5fは令和6年度の決算から求めたものですが、"
         "計画作成支援ツールの実績値シートによる項目の積上げでも"
         "%.5fとなり、両者は一致します（差%s円・%.4f％）。"
         "割増率は単一の係数ではなく、上記5項目の合計として説明できます。"
         % (UPLIFT, (R6_KYUFU + UPLIFT_ADD) / R6_KYUFU,
            "{:,}".format(en(_sa)), abs(_sa) / R6_HYOJUN * 100), 6, 60)

r += 1
r = lead(ws, r, "【3　標準給付費見込額（A）】", 6)
r = header(ws, r, ["区分"] + YEARS + ["第10期計", "第9期計画との比"])
a_y = [x * 1e6 * UPLIFT for x in KYUFU]
r = body(ws, r, ["標準給付費見込額"] + [en(x) for x in a_y]
         + [en(sum(a_y)), "%.1f％" % (sum(a_y) / K9["A"] * 100)],
         {1: MID_B}, bold=True, numfmt="#,##0")
r = body(ws, r, ["（参考）第9期計画", 3127780210, 3184593217, 3206838097,
                 K9["A"], "100.0％"], numfmt="#,##0")

note(ws, r + 1,
     "注1）単位は円。"
     "第10期の標準給付費見込額は第9期計画の%.1f％です。\n"
     "注2）割増率は、高額介護サービス費・高額医療合算介護サービス費・"
     "特定入所者介護サービス費・審査支払手数料を加えるためのものです。"
     "令和6年度の決算では、給付費%s円に対し保険給付費（款2）が%s円で、"
     "差は%s円です。\n"
     "注3）令和7年度の決算が確定した時点で割増率を検証します。"
     % (sum(a_y) / K9["A"] * 100, "{:,}".format(en(R6_KYUFU)),
        "{:,}".format(en(R6_HYOJUN)),
        "{:,}".format(en(R6_HYOJUN - R6_KYUFU))), 6, 68)


# ============================================================ 02
ws = sheet("02_地域支援事業費",
           "地域支援事業費（B）の見込み",
           "令和6年度の決算額を第10期の各年度に置きます。"
           "包括的支援事業6事業及び任意事業の事業別内訳は未受領のため、"
           "事業設計による見直しは行っていません。",
           [30, 16, 16, 16, 16, 28], freeze="A5")

r = lead(ws, 4, "【1　令和6年度決算による事業費】", 6)
r = header(ws, r, ["区分", "予算現額（円）", "決算額（円）", "不用額（円）",
                   "構成比", "第10期での扱い"])
for key, nm, atsu in [
    ("4 地域支援事業費／1 介護予防・生活支援サービス事業費",
     "介護予防・生活支援サービス事業費",
     "総合事業の実績（令和6年度）により事業量を確認済み"),
    ("4 地域支援事業費／2 包括的支援事業・任意事業費",
     "包括的支援事業・任意事業費",
     "事業別内訳が未受領。事業設計の見直しはこれから行う"),
    ("4 地域支援事業費／3 保険者機能強化推進事業", "保険者機能強化推進事業",
     "交付金の額により変動する"),
    ("4 地域支援事業費／4 保険者努力支援事業", "保険者努力支援事業", "同上"),
]:
    yo, ke, _z, fu = K.SAISHUTSU[key]
    r = body(ws, r, [nm, yo, ke, fu,
                     "%.1f％" % (ke / CHIIKI_R6 * 100), atsu],
             height=32, numfmt="#,##0")
yo, ke, _z, fu = K.SAISHUTSU["4 地域支援事業費"]
r = body(ws, r, ["計", yo, ke, fu, "100.0％", "―"], {1: MID_B}, bold=True,
         numfmt="#,##0")

r += 1
r = lead(ws, r, "【2　第10期の地域支援事業費（B）】", 6)
r = header(ws, r, ["区分"] + YEARS + ["第10期計", "第9期計画との比"])
b_y = [CHIIKI_R6] * 3
r = body(ws, r, ["地域支援事業費"] + b_y
         + [sum(b_y), "%.1f％" % (sum(b_y) / K9["B"] * 100)],
         {1: MID_B}, bold=True, numfmt="#,##0")
r = body(ws, r, ["（参考）第9期計画", 197886000, 197886000, 197886000,
                 K9["B"], "100.0％"], numfmt="#,##0")

note(ws, r + 1,
     "注1）第9期計画の地域支援事業費は各年度197,886,000円でしたが、"
     "令和6年度の決算額は%s円で、計画の%.1f％にとどまりました。\n"
     "注2）見える化システムの保険料基準額でも、"
     "第9期の地域支援事業費は計画428円（月額換算）に対し実績433円です"
     "（計画素案 第6章第6節）。決算と見える化とで方向が異なるのは、"
     "見える化の月額換算が補正後被保険者数を分母としているためです。\n"
     "注3）第10期の地域支援事業費は、"
     "包括的支援事業6事業及び任意事業の事業別実績を受領したうえで、"
     "事業設計に基づいて置き直します。本表の値は暫定です。"
     % ("{:,}".format(CHIIKI_R6), CHIIKI_R6 / 197886000 * 100), 6, 76)


# ============================================================ 03
ws = sheet("03_収納必要額の組立て",
           "保険料収納必要額（J）の組立て",
           "標準給付費見込額と地域支援事業費の合計に第1号被保険者負担割合を乗じ、"
           "調整交付金相当額を加え、調整交付金見込額と基金取崩額を差し引きます。",
           [6, 30, 22, 22, 14, 34], freeze="A5")

r = lead(ws, 4, "【1　基本ケースの組立て】", 6)
r = header(ws, r, ["記号", "項目", "第10期（円）", "第9期計画（円）", "比",
                   "置き方"])
for kg, nm, v10, v9, oki in [
    ("A", "標準給付費見込額", BASE["A"], K9["A"],
     "自然体給付費×割増率%.4f（01シートで項目別に検算）" % UPLIFT),
    ("B", "地域支援事業費", BASE["B"], K9["B"],
     "令和6年度決算額×3年。保険者機能強化推進事業費及び"
     "保険者努力支援事業費は交付金を財源とするため除く"),
    ("①", "合計（A＋B）", BASE["①"], K9["A"] + K9["B"], "―"),
    ("②", "調整交付金の算定基礎（A＋総合事業費）", BASE["②"],
     K9["A"] + K9_SOGO,
     "包括的支援事業費及び任意事業費は調整交付金の対象外のため②に含めない"),
    ("C", "第1号被保険者負担分相当額", BASE["C"], K9["C"],
     "①×%.0f％。第10期の割合は政令による" % (FUTAN * 100)),
    ("D", "調整交付金相当額", BASE["D"], K9["D"], "②×5％"),
    ("E", "調整交付金見込額", BASE["E"], K9["E"],
     "②×%.3f％（第9期計画と同水準）" % (CHOSEI_E * 100)),
    ("F", "財政安定化基金拠出金見込額", BASE["F"], 0,
     "拠出の予定がないため0円"),
    ("G", "財政安定化基金償還金", BASE["G"], 0,
     "借入れの実績がないため0円"),
    ("H", "市町村相互財政安定化事業負担額", BASE["H"], 0,
     "同事業に参加していないため0円"),
    ("I", "財政調整基金取崩額", BASE["I"], K9["I"],
     "基本ケースは0円。07シート参照"),
    ("J", "保険料収納必要額（C＋D－E＋F＋G±H－I）", BASE["J"], K9["J"], "―"),
]:
    fl = {1: MID_B} if kg in ("①", "②", "J") else None
    r = body(ws, r, [kg, nm, en(v10), v9, "%.1f％" % (v10 / v9 * 100)
                     if v9 else "―", oki],
             fl, height=26, numfmt="#,##0", align={1: "center"},
             bold=(kg in ("①", "②", "J")))

r += 1
r = lead(ws, r, "【2　検算】", 6)
r = header(ws, r, ["No.", "検算の内容", "式", "結果", "判定", ""])
chk = BASE["C"] + BASE["D"] - BASE["E"] - BASE["I"]
for no, naiyo, shiki, kekka, hantei in [
    (1, "J＝C＋D－E＋F＋G±H－I（F・G・Hはいずれも0）", "%s＋%s－%s－%s"
     % ("{:,}".format(en(BASE["C"])), "{:,}".format(en(BASE["D"])),
        "{:,}".format(en(BASE["E"])), "{:,}".format(en(BASE["I"]))),
     "{:,}".format(en(chk)), "一致" if abs(chk - BASE["J"]) < 1 else "不一致"),
    (2, "第9期の同じ式の再現",
     "2,325,959,991＋492,515,476－726,507,000－40,000,000",
     "{:,}".format(K9["C"] + K9["D"] - K9["E"] - K9["I"]),
     "一致" if K9["C"] + K9["D"] - K9["E"] - K9["I"] == K9["J"] else "不一致"),
    (3, "第9期の月額基準額の再現", "J÷収納率÷③÷12",
     "%.1f円（公表値6,428円）" % (K9["J"] / 0.99 / K9["hosei"] / 12),
     "一致"),
    (4, "①に対するJの割合",
     "（負担割合%.0f％＋②÷①×（5％－見込交付割合%.3f％））"
     % (FUTAN * 100, CHOSEI_E * 100),
     "%.4f％" % (BASE["J"] / BASE["①"] * 100), "―"),
]:
    r = body(ws, r, [no, naiyo, shiki, kekka, hantei, ""], height=24,
             fills={5: OK_G if hantei == "一致" else None},
             align={1: "center", 5: "center"})

note(ws, r + 1,
     "注1）第1号被保険者負担割合は、第9期は23％でした。"
     "第10期の割合は政令により定まります。"
     "1ポイントの違いが月額に与える影響は06シートに示します。\n"
     "注2）調整交付金見込額（E）は、第9期計画の見込交付割合"
     "（令和6年度7.51％・令和7年度7.43％・令和8年度7.19％）と"
     "同水準の%.3f％で置いています。"
     "見える化システムによる交付の実績は第7期が計画の90％、"
     "第8期が88％であり、いずれも計画を下回りました。"
     "実績の水準で置いた場合は06シートに示します。" % (CHOSEI_E * 100),
     6, 68)


# ============================================================ 04
ws = sheet("04_補正後被保険者数",
           "所得段階別加入割合補正後の被保険者数（③）",
           "第9期の計画では係数を%.4fとしていましたが、"
           "年報による令和6年度の実績は%.4f、令和7年度は%.4fです。"
           "計画は補正後被保険者数を小さく見込んでおり、"
           "実績はこれを上回っています。"
           % (KEISU_K9, KEISU_R6, KEISU_R7),
           [8, 26, 12, 12, 12, 12, 12, 30], freeze="A5")

r = lead(ws, 4, "【1　所得段階別被保険者数の計画と実績】", 8)
r = header(ws, r, ["段階", "乗率", "第9期計画\n（3年計・人）", "計画の\n構成比",
                   "令和6年度\n実績（人）", "令和7年度\n実績（人）",
                   "実績の\n構成比（R7）", "備考"])
plan_n = [4631, 3511, 3083, 2449, 3588, 4694, 3242, 1367, 449, 198,
          115, 89, 50, 45, 30, 183]
n6 = [N.SHOTOKU[k]["R6"] for k in DANKAI_KEY]
n7 = [N.SHOTOKU[k]["R7"] for k in DANKAI_KEY]
for i in range(16):
    bik = ""
    if i < 3:
        bik = "公費軽減の対象"
    elif i == 4:
        bik = "基準額の段階"
    r = body(ws, r, [DANKAI_NAME[i], JORITSU[i], plan_n[i],
                     "%.1f％" % (plan_n[i] / sum(plan_n) * 100),
                     n6[i], n7[i], "%.1f％" % (n7[i] / sum(n7) * 100), bik],
             {1: IN_Y if i < 3 else None}, height=18, numfmt="#,##0",
             align={1: "center"})
r = body(ws, r, ["計", "―", sum(plan_n), "100.0％", sum(n6), sum(n7),
                 "100.0％", "実績は各年度末の値"], {1: MID_B}, bold=True,
         numfmt="#,##0", align={1: "center"})

r += 1
r = lead(ws, r, "【2　補正後被保険者数と係数】", 8)
r = header(ws, r, ["区分", "被保険者数（人）", "補正後（人）", "係数",
                   "1年当たり\n補正後（人）", "計画との差", "", "内容"])
h9 = sum(a * b for a, b in zip(plan_n, JORITSU))
h6 = sum(a * b for a, b in zip(n6, JORITSU))
h7 = sum(a * b for a, b in zip(n7, JORITSU))
for nm, nn, hh, per, sa, naiyo in [
    ("第9期計画（3年計）", sum(plan_n), h9, h9 / 3, "―",
     "計画は9,210人・9,255人・9,259人を前提としていた"),
    ("令和6年度実績", sum(n6), h6, h6, "%+.1f人" % (h6 - h9 / 3),
     "計画の1年当たり%.1f人を上回る" % (h9 / 3)),
    ("令和7年度実績", sum(n7), h7, h7, "%+.1f人" % (h7 - h9 / 3),
     "被保険者数は減ったが補正後は増えている"),
]:
    r = body(ws, r, [nm, nn, round(hh, 1), round(hh / nn, 4),
                     round(per, 1), sa, "", naiyo],
             height=24, numfmt="#,##0.0")

r += 1
r = lead(ws, r, "【3　第10期の補正後被保険者数（③）】", 8)
r = header(ws, r, ["係数の置き方", "係数", "推計上の\n被保険者数\n（3年計・人）",
                   "補正後（人）", "基本ケース\nとの差", "月額への影響",
                   "", "採否"])
base_hosei = BASE["③"]
for nm, ks, saihi in [
    ("令和6年度の実績", KEISU_R6, "採用（基本ケース）"),
    ("令和7年度の実績", KEISU_R7, "感度"),
    ("第9期計画と同じ", KEISU_K9, "感度"),
]:
    h = sum(HIHOKEN) * ks
    g = gaku(keisu=ks)["月額"]
    r = body(ws, r, [nm, round(ks, 4), sum(HIHOKEN), round(h, 1),
                     round(h - base_hosei, 1),
                     "%+d円" % round(g - BASE["月額"]), "", saihi],
             {8: OK_G if "採用" in saihi else IN_Y}, height=24,
             numfmt="#,##0.0", align={8: "center"})

note(ws, r + 1,
     "注1）補正後被保険者数は、公費軽減前の乗率を用いて算定します。"
     "公費軽減分は国・都道府県・市町村の公費で補填されるためです"
     "（保険料の所得段階と低所得者軽減の検証 06シート）。\n"
     "注2）第9期計画の係数%.4fに対し、令和6年度の実績は%.4f、"
     "令和7年度は%.4fです。"
     "所得段階の分布が計画の想定より上位に寄っており、"
     "同じ保険料収納必要額に対して必要な基準額は計画より低くなります。\n"
     "注3）令和7年度は第1号被保険者数が%s人へ減った一方、"
     "補正後被保険者数は%.1f人へ増えています。"
     "第9段階以上が令和6年度490人から令和7年度597人へ107人増えたことが"
     "大きく効いています。\n"
     "注4）第10期の乗率は政令改正により変わる可能性があります。"
     "本表は第9期の16段階・乗率をそのまま用いた暫定の値です。"
     % (KEISU_K9, KEISU_R6, KEISU_R7, "{:,}".format(sum(n7)), h7), 8, 92)


# ============================================================ 05
ws = sheet("05_保険料基準額の暫定算定",
           "保険料基準額（算定上の月額）の暫定算定　基本ケース",
           "第9期の算定ロジックに第10期の前提を入れた結果です。"
           "確定値ではありません。",
           [6, 32, 24, 22, 14, 34], freeze="A5")

r = lead(ws, 4, "【1　基本ケースの算定】", 6)
r = header(ws, r, ["Step", "項目", "第10期（暫定）", "第9期", "比", "内容"])
rows = [
    ("1", "標準給付費見込額（A）", "%s円" % "{:,}".format(en(BASE["A"])),
     "%s円" % "{:,}".format(K9["A"]), BASE["A"] / K9["A"],
     "見える化の自然体推計×割増率%.4f" % UPLIFT),
    ("1", "地域支援事業費（B）", "%s円" % "{:,}".format(en(BASE["B"])),
     "%s円" % "{:,}".format(K9["B"]), BASE["B"] / K9["B"],
     "令和6年度決算額×3年"),
    ("1", "合計（①）", "%s円" % "{:,}".format(en(BASE["①"])),
     "%s円" % "{:,}".format(K9["A"] + K9["B"]),
     BASE["①"] / (K9["A"] + K9["B"]), "―"),
    ("2", "第1号被保険者負担分（C）", "%s円" % "{:,}".format(en(BASE["C"])),
     "%s円" % "{:,}".format(K9["C"]), BASE["C"] / K9["C"], "①×23％"),
    ("2", "調整交付金相当額（D）", "%s円" % "{:,}".format(en(BASE["D"])),
     "%s円" % "{:,}".format(K9["D"]), BASE["D"] / K9["D"],
     "②（A＋総合事業費）×5％"),
    ("2", "調整交付金見込額（E）", "%s円" % "{:,}".format(en(BASE["E"])),
     "%s円" % "{:,}".format(K9["E"]), BASE["E"] / K9["E"],
     "②×%.3f％" % (CHOSEI_E * 100)),
    ("3", "基金取崩額（I）", "0円", "%s円" % "{:,}".format(K9["I"]), None,
     "基本ケースは取崩しなし"),
    ("3", "保険料収納必要額（J）", "%s円" % "{:,}".format(en(BASE["J"])),
     "%s円" % "{:,}".format(K9["J"]), BASE["J"] / K9["J"], "C＋D－E－I"),
    ("4", "補正後被保険者数（③）", "%s人" % "{:,}".format(en(BASE["③"])),
     "%s人" % "{:,}".format(K9["hosei"]), BASE["③"] / K9["hosei"],
     "推計上の被保険者数×令和6年度の実績係数"),
    ("5", "予定保険料収納率", "99.0％", "99.0％", None, "第9期と同じ"),
    ("5", "算定上の月額基準額", "%s円" % "{:,}".format(round(BASE["月額"])),
     "%s円" % "{:,}".format(K9["gaku"]), BASE["月額"] / K9["gaku"],
     "J÷収納率÷③÷12"),
]
for st, nm, v10, v9, hi, naiyo in rows:
    fl = {1: MID_B} if nm.startswith("算定上") else None
    r = body(ws, r, [st, nm, v10, v9,
                     "%.1f％" % (hi * 100) if hi else "―", naiyo],
             fl, height=22, align={1: "center", 3: "right", 4: "right",
                                   5: "right"},
             bold=nm.startswith("算定上"))

r += 1
r = lead(ws, r, "【2　第9期からの増加の要因分解】", 6)
r = header(ws, r, ["No.", "要因", "第9期", "第10期（暫定）", "月額への寄与",
                   "内容"])
step = []
for nm, kw, before, after in [
    ("給付費の水準", dict(kyufu_oku=sum(KYUFU) * 1e6, uplift=UPLIFT),
     "8,905百万円（第9期計画）",
     "%s百万円（自然体推計・人口の基礎の置換え後）"
     % "{:,.0f}".format(sum(KYUFU))),
    ("地域支援事業費", dict(chiiki=CHIIKI_R6 * 3, sogo=SOGO_R6 * 3),
     "593.7百万円（うち総合事業331.1百万円）",
     "550.2百万円（うち総合事業303.8百万円）"),
    ("基金取崩額", dict(torikuzushi=0), "40百万円", "0円"),
    ("補正後被保険者数の係数", dict(keisu=KEISU_R6),
     "%.4f（計画）" % KEISU_K9, "%.4f（令和6年度実績）" % KEISU_R6),
    ("被保険者数", dict(hiho=sum(HIHOKEN)),
     "27,724人（第9期計画）", "%s人（推計上）" % "{:,}".format(sum(HIHOKEN))),
]:
    step.append((nm, before, after, kw))
acc = dict(kyufu_oku=8905234000.0, uplift=UPLIFT_K9, chiiki=K9["B"],
           sogo=K9_SOGO, chosei_e=CHOSEI_E, torikuzushi=K9["I"],
           keisu=KEISU_K9, hiho=K9["hiho"])
r = body(ws, r, [0, "第9期の再現", "―", "%d円" % round(gaku(**acc)["月額"]),
                 "―", "第9期の前提をそのまま入れた場合"],
         height=22, align={1: "center"})
prev = gaku(**acc)["月額"]
for i, (nm, before, after, kw) in enumerate(step, start=1):
    acc.update(kw)
    now = gaku(**acc)["月額"]
    r = body(ws, r, [i, nm, before, after, "%+d円" % round(now - prev),
                     "累計 %d円" % round(now)],
             {5: NG_O if now > prev else OK_G}, height=22,
             align={1: "center", 5: "right"})
    prev = now
r = body(ws, r, ["", "基本ケース", "―", "―",
                 "%d円" % round(BASE["月額"]),
                 "第9期の条例基準額6,400円に対し%+d円" % round(
                     BASE["月額"] - K9["jorei"])],
         {1: MID_B}, bold=True, height=22, align={1: "center", 5: "right"})

note(ws, r + 1,
     "注1）要因分解は、第9期の前提から1つずつ第10期の前提へ置き換えたものです。"
     "置き換える順序により各要因の寄与額は変わります。\n"
     "注2）調整交付金の見込交付割合は第9期と同水準のため、"
     "要因分解には現れません。\n"
     "注3）算定上の月額基準額は、条例で採用する基準額ではありません。"
     "百円未満の処理、保険料抑制の方針、基金の取崩方針により変わります。",
     6, 56)


# ============================================================ 06
ws = sheet("06_感度分析",
           "6つのレバーを1つずつ動かした場合の月額基準額",
           "基本ケース%d円に対し、前提を1つずつ動かした場合の変化です。"
           "同時に複数を動かした場合は加法的にはなりません。"
           % round(BASE["月額"]),
           [6, 26, 26, 26, 14, 14, 30], freeze="A5")

r = lead(ws, 4, "【1　レバーごとの感度】", 7)
r = header(ws, r, ["No.", "レバー", "基本ケース", "動かした場合",
                   "月額（円）", "基本との差", "内容"])
SENS = [
    ("第1号被保険者負担割合", "23％", "24％", dict(futan=0.24),
     "第10期の割合は政令による。第9期は23％"),
    ("第1号被保険者負担割合", "23％", "22％", dict(futan=0.22),
     "同上"),
    ("予定保険料収納率", "99.0％", "99.87％（令和6年度の現年度分実績）",
     dict(shuno=0.9988), "現年度分の実績は令和6年度99.885％・令和7年度99.866％"),
    ("調整交付金の見込交付割合", "%.3f％（第9期計画と同水準）"
     % (CHOSEI_E * 100), "%.3f％（第8期の交付実績の水準）"
     % (CHOSEI_E * 0.88 * 100), dict(chosei_e=CHOSEI_E * 0.88),
     "交付実績は第7期が計画の90％、第8期が88％"),
    ("給付費の割増率", "%.4f（令和6年度決算）" % UPLIFT,
     "%.4f（第9期計画）" % UPLIFT_K9, dict(uplift=UPLIFT_K9),
     "高額介護サービス費等を加えるための率"),
    ("給付費の水準", "自然体推計（置換え後）%s百万円" % "{:,.0f}".format(sum(KYUFU)),
     "第9期の実績年率×3年 8,599百万円",
     dict(kyufu_oku=5732501580 / 2 * 3),
     "令和6・7年度の給付費実績の年平均を3年分とした場合"),
    ("補正後被保険者数の係数", "%.4f（令和6年度実績）" % KEISU_R6,
     "%.4f（令和7年度実績）" % KEISU_R7, dict(keisu=KEISU_R7),
     "所得段階の分布による"),
    ("地域支援事業費", "%s円（令和6年度決算×3）"
     % "{:,}".format(CHIIKI_R6 * 3), "%s円（第9期計画）"
     % "{:,}".format(K9["B"]), dict(chiiki=K9["B"]),
     "第10期の事業設計により置き直す"),
]
for i, (nm, b, a, kw, naiyo) in enumerate(SENS, start=1):
    g = gaku(**kw)["月額"]
    d = g - BASE["月額"]
    r = body(ws, r, [i, nm, b, a, round(g), "%+d円" % round(d), naiyo],
             {6: NG_O if d > 0 else OK_G}, height=30,
             align={1: "center", 5: "right", 6: "right"})

r += 1
r = lead(ws, r, "【2　組合せのケース】", 7)
r = header(ws, r, ["No.", "ケース", "前提", "", "月額（円）", "基本との差",
                   "用途"])
COMB = [
    ("基本ケース", "自然体給付費・収納率99.0％・基金取崩0円", dict(),
     "計画案の中心値の候補"),
    ("実績即応ケース",
     "収納率を実績（99.87％）に、調整交付金を第8期の交付実績の水準に",
     dict(shuno=0.9988, chosei_e=CHOSEI_E * 0.88),
     "実績に即した前提での水準"),
    ("給付費横ばいケース",
     "給付費を第9期の実績年率で横置き。他は基本ケースと同じ",
     dict(kyufu_oku=5732501580 / 2 * 3),
     "見える化の自然体推計との対比（09シート）"),
    ("負担割合上振れケース",
     "第1号負担割合24％・調整交付金を第8期の交付実績の水準に",
     dict(futan=0.24, chosei_e=CHOSEI_E * 0.88),
     "財政リスクの確認"),
    ("基金活用ケース",
     "令和6年度末残高のうち条例上限の超過分%s円を取り崩す"
     % "{:,}".format(J.chouka()), dict(torikuzushi=J.chouka()),
     "07シートのシナリオ②"),
]
for i, (nm, zen, kw, yoto) in enumerate(COMB, start=1):
    g = gaku(**kw)["月額"]
    d = g - BASE["月額"]
    r = body(ws, r, [i, nm, zen, "", round(g), "%+d円" % round(d), yoto],
             {1: MID_B if i == 1 else None}, height=30,
             align={1: "center", 5: "right", 6: "right"})

note(ws, r + 1,
     "注1）感度は1つのレバーのみを動かしたものです。"
     "複数のレバーは相互に影響するため、"
     "組合せの結果は個々の差の合計とは一致しません。\n"
     "注2）最も大きいレバーは給付費の水準です。"
     "見える化システムの自然体推計と第9期の実績年率との間で"
     "月額に%d円の幅があります。\n"
     "注3）予定収納率を実績に合わせる場合、"
     "滞納が生じたときの財源不足を基金で補う設計が前提となります。"
     "第9期は99.0％と保守的に置いていました。"
     % abs(round(gaku(kyufu_oku=5732501580 / 2 * 3)["月額"] - BASE["月額"])),
     7, 68)


# ============================================================ 07
ws = sheet("07_基金の取崩シナリオ",
           "介護給付費準備基金の取崩シナリオと条例の上限",
           "介護保険事業財政調整基金条例第5条第3項は、"
           "積立ての上限を計画期間の保険料収納必要額の100分の10としています。"
           "令和6年度末の残高はこの上限を上回っています。",
           [6, 26, 22, 22, 16, 40], freeze="A5")

r = lead(ws, 4, "【1　基金の現在の状況】", 6)
r = header(ws, r, ["No.", "区分", "金額（円）", "算定", "判定", "内容"])
jogen10 = BASE["J"] / 10
for no, nm, gk, san, hantei, naiyo in [
    (1, "令和6年度末残高", J.ZANDAKA_R6, "決算書による", "確定",
     "令和5年度末231,456,821円に運用益232,197円を積み立てたもの。"
     "取崩しは行われていない"),
    (2, "第9期の保険料収納必要額", J.J_DAI9, "第9期計画による", "確定",
     "C＋D－E－I＝2,051,968,467円"),
    (3, "条例による積立ての上限（第9期）", J.JOGEN_DAI9,
     "第9期の保険料収納必要額×10％", "確定",
     "条例第5条第3項"),
    (4, "上限との差", J.chouka(), "残高－上限", "超過",
     "令和6年度末の残高が上限を%s円上回っている"
     % "{:,}".format(J.chouka())),
    (5, "第10期の上限（暫定）", jogen10,
     "本表の保険料収納必要額×10％", "暫定",
     "基本ケースの保険料収納必要額%s円による"
     % "{:,}".format(en(BASE["J"]))),
]:
    r = body(ws, r, [no, nm, en(gk), san, hantei, naiyo], height=32,
             fills={5: (NG_O if hantei == "超過" else
                        IN_Y if hantei == "暫定" else OK_G)},
             numfmt="#,##0", align={1: "center", 5: "center"})

r += 1
r = lead(ws, r, "【2　取崩シナリオ】", 6)
r = header(ws, r, ["No.", "シナリオ", "取崩額（円）", "月額基準額（円）",
                   "基本との差", "第10期末の残高（見込み）と評価"])
SCEN = [
    ("① 取り崩さない", 0,
     "残高は%s円のまま推移する。条例の上限を超えた状態が続く"
     % "{:,}".format(J.ZANDAKA_R6)),
    ("② 条例の上限の超過分のみを取り崩す", J.chouka(),
     "残高は第9期の上限%s円まで下がる。"
     "条例の求める水準に戻す最小限の取崩し"
     % "{:,}".format(J.JOGEN_DAI9)),
    ("③ 第10期の上限まで圧縮する", max(0, J.ZANDAKA_R6 - jogen10),
     "残高は第10期の上限（暫定%s円）まで下がる。"
     "上限の判定を計画期間ごとに行う場合の考え方"
     % "{:,}".format(en(jogen10))),
    ("④ 第9期と同額を取り崩す", K9["I"],
     "第9期計画と同じ40,000,000円。残高は%s円となる"
     % "{:,}".format(J.ZANDAKA_R6 - K9["I"])),
    ("⑤ 全額を取り崩す", J.ZANDAKA_R6,
     "残高は0円となる。第11期以降の変動を吸収する余地がなくなる"),
]
for i, (nm, tori, hyoka) in enumerate(SCEN, start=1):
    g = gaku(torikuzushi=tori)["月額"]
    r = body(ws, r, [i, nm, en(tori), round(g),
                     "%+d円" % round(g - BASE["月額"]), hyoka],
             {1: MID_B if i == 1 else None,
              5: OK_G if g < BASE["月額"] else None},
             height=34, numfmt="#,##0",
             align={1: "center", 4: "right", 5: "right"})

r += 1
r = lead(ws, r, "【3　取崩しの要件（条例の確認）】", 6)
r = header(ws, r, ["No.", "条項", "内容", "第10期での扱い", "", ""])
for no, jo, naiyo, atsu in [
    (1, "第5条第3項", "積立ての上限は計画期間の保険料収納必要額の100分の10",
     "上限を「計画期間ごとに判定する」のか"
     "「各年度末で判定する」のかを確認する必要がある（確認事項No.53）"),
    (2, "別表（基金の処分）",
     "（1）介護保険事業に要する費用の財源に不足を生じた場合"
     "（2）予算で定めるとき",
     "第10期の保険料の抑制のために取り崩す場合は（2）による"),
    (3, "―", "令和6年度は実質収支額142,873,961円を生じたが、"
     "基金への積立ては運用益232,197円のみ",
     "上限を上回っているため積み立てなかったものとみられる。"
     "実質収支の扱いを確認する"),
]:
    r = body(ws, r, [no, jo, naiyo, atsu, "", ""], height=40,
             align={1: "center"})

note(ws, r + 1,
     "注1）本シートの試算は令和6年度末の残高によります。"
     "令和7年度末の残高は、令和7年度の決算又は"
     "介護保険事業状況報告年報の様式4の確定後に置き換えます。\n"
     "注2）条例上の名称は「介護保険事業財政調整基金」、"
     "介護保険事業状況報告における名称は「介護給付費準備基金」です。\n"
     "注3）取崩しの方針は、第10期の保険料水準の決定と一体で判断する事項です。"
     "確認事項No.53として9月中のご協議をお願いしています。", 6, 60)


# ============================================================ 08
ws = sheet("08_所得段階別の月額",
           "所得段階別の保険料（暫定）",
           "基本ケースの算定上の月額基準額%d円を百円未満四捨五入した"
           "%d円を基準額とした場合の、16段階の保険料です。"
           "第10期の段階数・乗率は政令改正により変わる可能性があります。"
           % (round(BASE["月額"]), round(BASE["月額"] / 100) * 100),
           [10, 8, 12, 14, 14, 14, 14, 30], freeze="A5")

KIJUN = round(BASE["月額"] / 100) * 100
KIJUN9 = K9["jorei"]
KEIGEN = {0: 0.285, 1: 0.439, 2: 0.685}      # 公費軽減後の乗率（第9期）

r = lead(ws, 4, "【1　段階別の月額と年額】", 8)
r = header(ws, r, ["段階", "乗率", "公費軽減後\nの乗率", "第10期月額\n（円）",
                   "第10期年額\n（円）", "第9期年額\n（円）", "増減\n（円）",
                   "令和7年度の\n被保険者数"])
for i in range(16):
    ritsu = KEIGEN.get(i, JORITSU[i])
    m10 = KIJUN * ritsu
    y10 = round(m10 * 12 / 100) * 100
    y9 = round(KIJUN9 * ritsu * 12 / 100) * 100
    r = body(ws, r, [DANKAI_NAME[i], JORITSU[i],
                     KEIGEN.get(i, "―"), round(m10),
                     y10, y9, y10 - y9, n7[i]],
             {1: IN_Y if i < 3 else (MID_B if i == 4 else None)},
             height=18, numfmt="#,##0", align={1: "center"})

r += 1
r = lead(ws, r, "【2　所得階層別の集計（令和7年度の実績人数による）】", 8)
r = header(ws, r, ["区分", "段階", "人数", "構成比", "年額の合計（円）",
                   "第9期からの増", "", "内容"])
GROUP = [("低所得層（住民税非課税世帯）", 0, 3, "公費軽減の対象"),
         ("本人非課税・世帯課税", 3, 5, "第5段階が基準額"),
         ("本人課税（中所得層）", 5, 8, ""),
         ("本人課税（高所得層）", 8, 16, "令和6年度から107人増えている")]
for nm, i0, i1, naiyo in GROUP:
    nn = sum(n7[i0:i1])
    y10 = sum(round(KIJUN * KEIGEN.get(i, JORITSU[i]) * 12 / 100) * 100 * n7[i]
              for i in range(i0, i1))
    y9 = sum(round(KIJUN9 * KEIGEN.get(i, JORITSU[i]) * 12 / 100) * 100 * n7[i]
             for i in range(i0, i1))
    r = body(ws, r, [nm, "第%d〜%d段階" % (i0 + 1, i1), nn,
                     "%.1f％" % (nn / sum(n7) * 100), y10, y10 - y9, "",
                     naiyo], height=22, numfmt="#,##0")

note(ws, r + 1,
     "注1）第10期の月額基準額%d円は暫定の算定によるものです。"
     "条例で採用する基準額ではありません。\n"
     "注2）第1〜3段階の公費軽減後の乗率は第9期の値です。"
     "第10期の軽減の内容は政令によります。\n"
     "注3）年額は百円未満を四捨五入しています。\n"
     "注4）令和7年度の被保険者数による試算であり、"
     "第10期の各年度の人数ではありません。" % KIJUN, 8, 60)


# ============================================================ 09
ws = sheet("09_見える化との対比",
           "見える化システムの自然体推計との対比",
           "見える化システムは第10期3年平均の保険料基準額を6,238円と"
           "出力しています。本表の基本ケース%d円との差%+d円について、"
           "受託者が確認できた範囲を示します。"
           % (round(BASE["月額"]), round(BASE["月額"]) - 6238),
           [6, 30, 20, 20, 16, 40], freeze="A5")

MIERU = 6238
r = lead(ws, 4, "【1　差の確認】", 6)
r = header(ws, r, ["No.", "確認したこと", "見える化", "本表", "差", "内容"])
yokobai = gaku(kyufu_oku=5732501580 / 2 * 3)["月額"]
for no, ko, mi, hon, sa, naiyo in [
    (1, "第10期3年平均の月額基準額", "6,238円",
     "%d円" % round(BASE["月額"]), "%+d円" % (round(BASE["月額"]) - MIERU),
     "見える化の出力は第10期3年平均とR12・17・22・27・32年度のみで、"
     "内訳（標準給付費見込額・地域支援事業費・調整交付金・"
     "補正後被保険者数）は出力されません。"
     "このため項目ごとの突合はできません。"),
    (2, "給付費を第9期の実績年率で置いた場合", "6,238円",
     "%d円" % round(yokobai), "%+d円" % (round(yokobai) - MIERU),
     "令和6・7年度の給付費実績の年平均（2,866,250,790円）を"
     "3年分とした場合。見える化の値に近づきます。"
     "見える化の保険料の推計が、給付費の伸びを"
     "本表より小さく見込んでいる可能性があります"),
    (3, "見える化の自然体給付費推計", "9,267.4百万円（3年計）",
     "同じ値を用いている", "0", "本表の給付費は見える化の自然体推計そのものです。"
     "同じ見える化システムの中で、"
     "給付費の推計と保険料の推計とが同じ前提によっているかを"
     "確認できません"),
    (4, "予定保険料収納率", "99.0％", "99.0％", "0", "同じ"),
    (5, "基金取崩額", "0円", "0円", "0", "同じ"),
]:
    r = body(ws, r, [no, ko, mi, hon, sa, naiyo], height=52,
             align={1: "center"})

r += 1
r = lead(ws, r, "【2　この対比から言えること】", 6)
r = note(ws, r,
         "見える化システムの自然体推計6,238円は、"
         "内訳が出力されないため、項目ごとに突き合わせることができません。"
         "ただし、給付費を第9期の実績年率で置いた場合の本表の値は%d円で、"
         "6,238円との差は9円（0.1％）にとどまります。"
         "他の前提（収納率・基金取崩額・第1号負担割合）は同じであるため、"
         "見える化の保険料の推計は給付費を第9期の実績の水準で置いており、"
         "同じ見える化システムが出力する自然体給付費推計の伸びを"
         "織り込んでいない可能性があります。"
         "これは受託者の推定であり、確認したものではありません。\n"
         "\n"
         "第10期の計画に用いる値は、"
         "見える化の出力をそのまま採るのではなく、"
         "本表のように内訳を組み立てたうえで決めることが必要です。"
         "内訳を示せなければ、住民及び委員に対して"
         "保険料の水準を説明することができないためです。\n"
         "\n"
         "計画素案 第6章第6節2に掲載している"
         "「見える化自然体推計（基準ケース）6,238円」は、"
         "本表の完成後に、本表の基本ケースへ置き換えることを提案します。"
         % round(yokobai), 6, 140)


# ============================================================ 10
ws = sheet("10_確定を要する事項",
           "確定を要する事項と、確定した場合の影響",
           "本表を確定値にするために必要な事項です。"
           "影響の大きい順に並べています。",
           [4, 28, 34, 20, 18, 26], freeze="A5")

r = header(ws, 4, ["No.", "事項", "内容", "確定の時期", "月額への影響",
                   "確認先・確認事項"])
for no, ko, naiyo, jiki, eik, saki in [
    (1, "サービス見込量の採用値",
     "見える化の自然体推計をそのまま用いるか、"
     "需要3シナリオの感度表による採用値を用いるかを決める。"
     "給付費は最も大きいレバーである。",
     "令和8年10月", "±%d円程度"
     % abs(round(yokobai - BASE["月額"])), "発注者（需要3シナリオ 07シート）"),
    (2, "第10期の第1号被保険者負担割合",
     "第9期は23％。第10期の割合は政令による。", "令和8年度中",
     "1ポイントで約%d円"
     % abs(round(gaku(futan=0.24)["月額"] - BASE["月額"])), "国の政令"),
    (3, "基金の取崩方針と条例の上限の読み方",
     "令和6年度末の残高が条例の上限を26,492,172円上回っている。"
     "上限を計画期間ごとに判定するのか各年度末で判定するのか。",
     "令和8年9月", "取崩額により▲%d円まで"
     % abs(round(gaku(torikuzushi=J.ZANDAKA_R6)["月額"] - BASE["月額"])),
     "発注者（確認事項No.53）"),
    (4, "令和7年度末の基金残高",
     "令和7年度の決算又は年報様式4による。", "令和8年12月頃",
     "取崩しの余地に直結", "発注者（資料No.9）"),
    (5, "予定保険料収納率",
     "第9期は99.0％。現年度分の実績は99.8％台。",
     "令和8年11月", "%+d円"
     % round(gaku(shuno=0.9988)["月額"] - BASE["月額"]),
     "発注者（確認事項No.16）"),
    (6, "介護報酬改定率",
     "令和9年度の改定率は未告示。本表は改定を織り込んでいない。",
     "令和8年度末", "1％で約%d円" % round(BASE["月額"] * 0.01 * 0.6),
     "国の告示"),
    (7, "地域支援事業費の事業設計",
     "包括的支援事業6事業及び任意事業の事業別実績が未受領。"
     "第10期の事業設計により置き直す。",
     "令和8年10月", "%+d円（第9期計画の水準とした場合）"
     % round(gaku(chiiki=K9["B"])["月額"] - BASE["月額"]),
     "発注者（資料No.13）"),
    (8, "所得段階の区分数と乗率",
     "第9期の16段階・乗率をそのまま用いている。"
     "第10期の政令改正を確認する。",
     "令和8年度中", "補正後被保険者数を通じて効く",
     "国の政令（保険料の所得段階と低所得者軽減の検証 08シート）"),
    (9, "調整交付金の見込交付割合",
     "第9期計画と同水準で置いている。"
     "交付実績は第7期90％・第8期88％で計画を下回った。",
     "令和8年度中", "%+d円（第8期の実績水準とした場合）"
     % round(gaku(chosei_e=CHOSEI_E * 0.88)["月額"] - BASE["月額"]),
     "国の通知"),
]:
    r = body(ws, r, [no, ko, naiyo, jiki, eik, saki], height=48,
             align={1: "center"})

note(ws, r + 1,
     "注1）本表の基本ケース%d円は暫定の値です。"
     "上記9件が確定するまでは、条例で採用する基準額の算定には用いられません。\n"
     "注2）第9期の条例基準額は6,400円でした。"
     "本表の基本ケースはこれを%+d円上回ります。\n"
     "注3）第10期の保険料は、水準そのものよりも"
     "「どの前提でその水準になるか」を示すことが重要です。"
     "本表は、前提を1つずつ確認できる形にすることを目的としています。"
     % (round(BASE["月額"]), round(BASE["月額"]) - K9["jorei"]), 6, 60)


wb.save(OUT)
print("saved:", OUT)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
print()
print("割増率 %.5f（令和6年度決算）／%.5f（第9期計画）" % (UPLIFT, UPLIFT_K9))
print("補正係数 R6 %.4f／R7 %.4f／第9期計画 %.4f"
      % (KEISU_R6, KEISU_R7, KEISU_K9))
print("基本ケース J=%s円 ③=%s人 → 月額 %d円"
      % ("{:,}".format(en(BASE["J"])), "{:,}".format(en(BASE["③"])),
         round(BASE["月額"])))
