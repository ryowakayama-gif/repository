# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画
3町の社会資源一覧と、既存の把握との突合.

令和8年8月20日に発注者から3町の社会資源一覧（74事業所）を受領した。
事業所ごとの従業員総数・常勤・非常勤が3町分そろっている。

本表は、
  ① 受領した一覧の内容を整理する
  ② 訪問介護13事業所について、公表画面による把握と突合する
  ③ 従業員数の把握の状況を全サービスで整理する
  ④ 指定事業所一覧（北海道）との網羅性を確認する
  ⑤ 計画本文及び成果品への反映箇所を示す

シート構成
  00_この表について
  01_受領した一覧（74事業所）
  02_訪問介護13事業所の突合
  03_サービス別の従業員数
  04_指定事業所一覧との網羅性
  05_反映する箇所
"""

import re
from collections import Counter, OrderedDict

import data_hokkaido_shitei as H
import data_shakai_shigen as S
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = ("/home/user/repository/output/"
       "第10期計画_3町の社会資源一覧との突合.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TOWNS = ["東川町", "美瑛町", "東神楽町"]

# 訪問介護13事業所の対応（社会資源一覧の名称 → 公表画面の名称）
PAIR = [
    ("美瑛町", "美瑛町ホームヘルプサービスセンター",
     "美瑛町ホームヘルプサービスセンター"),
    ("美瑛町", "シルバーハウス訪問介護支援事業所",
     "シルバーハウス訪問介護事業所"),
    ("東川町", "指定訪問介護（指定介護予防訪問介護）事業所縁結び",
     "指定訪問介護（指定介護予防訪問介護）事業所 縁結び"),
    ("東川町", "指定訪問介護事業所　恩送り", "指定訪問介護事業所 恩送り"),
    ("東川町", "東川町社協訪問介護事業所", "東川町社協訪問介護事業所"),
    ("東川町", "訪問介護事業所　桜華", "訪問介護事業所 桜華"),
    ("東川町", "ヘルパーステーションフラワー", "ヘルパーステーションフラワー"),
    ("東川町", "訪問介護ステーションゆう", "訪問介護ステーションゆう"),
    ("東神楽町", "指定訪問介護事業所ひばり", "指定訪問介護事業所ひばり"),
    ("東神楽町", "医療法人回生会　花時計訪問介護事業所",
     "花時計訪問介護事業所"),
    ("東神楽町", "指定（介護予防）訪問介護ケンセイシャレバレッジ",
     "指定（介護予防）訪問介護ケンセイシャレバレッジ"),
    ("東神楽町", "東神楽町ホームヘルプサービスセンター",
     "東神楽町ホームヘルプサービスセンター"),
    ("東神楽町", "訪問介護事業所　ほがらか", "訪問介護事業所 ほがらか"),
]

KOHYO = {k["事業所名"]: k for k in H.KOHYO if k["サービス"] == "訪問介護"}
IDX = {}
for _t in TOWNS:
    for _r in S.SHIGEN[_t]:
        IDX[(_t, _r[1])] = _r


def f2(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


wb = Workbook()


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 52
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=20, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=None):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height or max(30, 13 * (len(text) // 80 + 1))
    return row + 1


# 集計
N_ALL = sum(len(S.SHIGEN[t]) for t in TOWNS)
SUM_T = {t: sum(f2(r[4]) or 0 for r in S.SHIGEN[t]) for t in TOWNS}
SV = Counter()
SVN = Counter()
for t in TOWNS:
    for r in S.SHIGEN[t]:
        SV[r[3]] += f2(r[4]) or 0
        SVN[r[3]] += 1

# ============================================================ 00
ws = sheet("00_この表について",
           "3町の社会資源一覧との突合",
           "令和8年8月20日に発注者から3町の社会資源一覧を受領した。"
           "3町合計74事業所について、従業員総数・常勤・非常勤が"
           "そろっている。"
           "これまで訪問介護のみ公表画面により把握していた従業員数が、"
           "全サービスについて得られたことになる。",
           [26, 76], freeze="A5")

r = lead(ws, 4, "【受領の内容と、確認できたこと】", 2)
r = header(ws, r, ["項目", "内容"])
_t1 = sum(f2(IDX[(a, b)][4]) or 0 for a, b, _c in PAIR)
_t2 = sum(f2(KOHYO[c]["総従業者数"]) or 0 for _a, _b, c in PAIR)
_t3 = sum(f2(KOHYO[c]["訪問介護員等_実人数"]) or 0 for _a, _b, c in PAIR)
for k, v in [
    ("① 74事業所の従業員数が\n　　そろった",
     "東川町%d・美瑛町%d・東神楽町%dの計%d事業所について、"
     "従業員総数・常勤職員数・非常勤職員数が得られた。"
     % (len(S.SHIGEN["東川町"]), len(S.SHIGEN["美瑛町"]),
        len(S.SHIGEN["東神楽町"]), N_ALL) +
     "従業員総数の合計は東川町%.0f人・美瑛町%.0f人・東神楽町%.0f人である。"
     % (SUM_T["東川町"], SUM_T["美瑛町"], SUM_T["東神楽町"]) +
     "これまで介護人材実態調査（27事業所）と"
     "公表画面（訪問介護13事業所）でしか把握できていなかった従業員数が、"
     "全サービスについて3町分そろったことになる。"),
    ("② 訪問介護13事業所は\n　　13件とも照合できた",
     "公表画面により把握していた訪問介護13事業所は、"
     "本一覧の13事業所とすべて対応した。"
     "従業員総数は本一覧%.1f人に対し公表画面%.1f人で、"
     "差は%+.1f人（%.1f％）である。"
     % (_t1, _t2, _t1 - _t2, (_t1 / _t2 - 1) * 100) +
     "3事業所（シルバーハウス・桜華・花時計）は完全に一致した。"
     "独立した2つの出所が事業所単位で対応したことになる。"),
    ("③ 差は「数える範囲」に\n　　よるものである",
     "本一覧は事業所の従業員総数（全職種）であり、"
     "公表画面の「訪問介護員等」は職種を限った実人数である。"
     "本一覧の値が公表画面を下回る事業所はなく、"
     "13事業所すべてで本一覧が同数か上回る。"
     "差の中央値は＋1人である。"
     "管理者・サービス提供責任者・事務職員の数え方の違いとみられる。"),
    ("④ 記入上の注記がある",
     "花時計訪問介護事業所について、"
     "本一覧に「事務員を0.5カウントの常勤に登録」との注記がある。"
     "総数16.5人・常勤13.5人はこれによる。"
     "公表画面の総従業者数16.5人と一致しており、"
     "同じ数え方が公表画面にも及んでいる。"),
    ("⑤ 介護人材実態調査の\n　　空白が埋まる",
     "介護人材実態調査は27事業所からの回答で、"
     "訪問系は13事業所のうち3事業所にとどまっていた。"
     "本一覧により、回答のなかった事業所を含めて"
     "従業員数を把握できる。"
     "ただし本一覧は従業員数のみであり、"
     "資格・年代・勤続年数・採用退職は含まない。"),
    ("⑥ 事業所調査の照会の\n　　一部が解消する",
     "「事業所調査の照会票と確定値管理表」の確定値表No.1・No.4"
     "（介護職員の総数）は、"
     "重複計上13人の扱いが未確定であった。"
     "本一覧は事業所ごとの従業員数であるため、"
     "重複を避けた総数を別に算定できる。"
     "ただし本一覧は「従業員」であり調査の「介護職員」とは範囲が異なるため、"
     "調査の数値をそのまま置き換えることはできない。"),
    ("⑦ 参考資料が付いている",
     "受領ファイルには、北海道の介護保険事業所一覧（全道）の"
     "サービス別シート10件（訪問介護1,704件ほか）が"
     "参考資料として付されている。"
     "これは既に受領済みの資料と同じものであり、"
     "区域内124件は分析に用いている。"),
]:
    r = body(ws, r, [k, v], height=100)

note(ws, r + 1,
     "注1）本一覧は令和8年8月20日に受領した。"
     "基準日の記載がないため、いつ時点の従業員数かは確認を要する［要確認］。"
     "注2）本一覧には担当者名・電話番号が含まれていないため、"
     "そのままデータ化した。"
     "注3）事業所名にはふりがな（ルビ）が付されているが、"
     "データ化に当たり本文のみを取り出している。", 2)

# ============================================================ 01
ws = sheet("01_受領した一覧",
           "受領した社会資源一覧（74事業所）",
           "3町分をそのまま収める。"
           "「調査への回答」欄は、実施済み調査のいずれに回答したかである。",
           [8, 5, 34, 26, 26, 9, 9, 9, 24], freeze="C5")

r = header(ws, 4, ["町", "No.", "事業所名", "法人等の名称・母体",
                   "介護サービスの種類", "従業員\n総数", "うち\n常勤",
                   "うち\n非常勤", "調査への回答"])
for t in TOWNS:
    for rr in S.SHIGEN[t]:
        no, nm, ho, sv, a, b, c, ch = rr
        r = body(ws, r, [t, no, nm, ho, sv, a, b, c,
                         "／".join(ch) if ch else ""],
                 {6: MID_B} if sv == "訪問介護" else {}, height=18,
                 align={2: "center", 6: "right", 7: "right", 8: "right"})
r = body(ws, r, ["3町計", N_ALL, "", "", "",
                 round(sum(SUM_T.values()), 1), "", "", ""],
         {1: GRAY}, height=20, bold=True,
         align={2: "center", 6: "right"})

note(ws, r + 1,
     "注1）網掛けは訪問介護である（13事業所）。"
     "注2）花時計訪問介護事業所の常勤欄には"
     "「※事務員を0.5カウントの常勤に登録」との注記が入っている。"
     "注3）介護老人保健施設回生苑（老人保健施設・入所）は"
     "従業員数の記入がない。"
     "同一施設の短期入所療養介護（73人）に含まれているとみられる［要確認］。"
     "注4）同一の施設が複数のサービスで計上されている場合がある"
     "（例：東神楽特別養護老人ホームアゼリアハイツは"
     "短期入所生活介護・介護老人福祉施設・地域密着型で3行）。"
     "従業員数を単純に合計すると重複する。", 9)

# ============================================================ 02
ws = sheet("02_訪問介護13事業所の突合",
           "訪問介護13事業所　社会資源一覧と公表画面の突合",
           "公表画面（介護サービス情報公表システム）により把握していた"
           "13事業所と、本一覧の13事業所を突き合わせる。"
           "独立した2つの出所である。",
           [8, 32, 9, 9, 9, 11, 11, 9, 28], freeze="C5")

r = header(ws, 4, ["町", "事業所名", "社会資源\n総数", "うち\n常勤",
                   "うち\n非常勤", "公表画面\n総従業者数",
                   "公表画面\n訪問介護員等", "差\n（社－公表）", "見方"])
_diff = []
for town, snm, knm in PAIR:
    s = IDX[(town, snm)]
    k = KOHYO[knm]
    a, b = f2(s[4]), f2(k["総従業者数"])
    d = a - b
    _diff.append(d)
    v = ("完全に一致する" if d == 0 else
         ("差は1人。管理者又は事務職員の数え方による" if d == 1 else
          "差が2人以上ある。数える範囲の確認を要する"))
    r = body(ws, r, [town, snm, s[4], s[5] if not isinstance(s[5], str)
                     else s[5][:6], s[6], k["総従業者数"],
                     k["訪問介護員等_実人数"], d, v],
             {8: (OK_G if d == 0 else (IN_Y if d <= 1 else NG_O))},
             height=20,
             align={3: "right", 4: "right", 5: "right", 6: "right",
                    7: "right", 8: "right"})
r = body(ws, r, ["3町計", "", round(_t1, 1), "", "", round(_t2, 1),
                 round(_t3, 1), round(_t1 - _t2, 1), "―"],
         {1: GRAY}, height=20, bold=True,
         align={3: "right", 6: "right", 7: "right", 8: "right"})

r = note(ws, r,
         "注1）13事業所すべてが対応した。"
         "完全に一致するのは%d事業所、差が1人は%d事業所、"
         "2人以上は%d事業所である。"
         % (sum(1 for d in _diff if d == 0),
            sum(1 for d in _diff if d == 1),
            sum(1 for d in _diff if d >= 2)) +
         "本一覧が公表画面を下回る事業所はない。"
         "注2）差が2人以上あるのは"
         "東神楽町ホームヘルプサービスセンター（＋4）、"
         "美瑛町ホームヘルプサービスセンター（＋3）、"
         "縁結び（＋3）、東川町社協（＋2）、"
         "ケンセイシャレバレッジ（＋2）である。"
         "いずれも社会福祉協議会又は規模の大きい事業所であり、"
         "事務職員を含めるかどうかの違いとみられる［要確認］。"
         "注3）公表画面の「訪問介護員等」は職種を限った実人数であり、"
         "本一覧の従業員総数とは定義が異なる。"
         "計画本文の「訪問介護員等181人」は公表画面によるものであり、"
         "本一覧の受領によって変わらない。", 9)

r += 1
r = lead(ws, r, "【3つの出所の対照】", 9)
r = header(ws, r, ["出所", "対象", "値", "", "把握できる事業所",
                   "", "含む職種", "", ""])
for a in [
    ("介護人材実態調査（第10期分）", "回答のあった訪問系3事業所",
     "介護職員42人", "", "3／13事業所", "",
     "介護職員（資格・年代・勤続年数・採用退職を含む）", "", ""),
    ("介護サービス情報公表システム", "訪問介護13事業所",
     "訪問介護員等の実人数181人／総従業者数%.1f人" % _t2, "",
     "13／13事業所", "",
     "訪問介護員等（サービス提供責任者を含む）", "", ""),
    ("3町の社会資源一覧（本一覧）", "訪問介護13事業所",
     "従業員総数%.1f人" % _t1, "", "13／13事業所", "",
     "全職種（管理者・事務職員を含むとみられる）", "", ""),
]:
    r = body(ws, r, list(a), {}, height=32)
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=4)
    ws.merge_cells(start_row=r - 1, start_column=5, end_row=r - 1, end_column=6)
    ws.merge_cells(start_row=r - 1, start_column=7, end_row=r - 1, end_column=9)

note(ws, r + 1,
     "注4）3つの出所は目的も定義も異なるため、"
     "どれか1つが正しいというものではない。"
     "計画本文では、供給力を示す数値として"
     "公表画面の訪問介護員等181人を用い、"
     "本一覧は事業所の規模を示す参考として扱う。", 9)

# ============================================================ 03
ws = sheet("03_サービス別の従業員数",
           "サービス別の従業員数（社会資源一覧による）",
           "本一覧をサービス別に集計する。"
           "同一施設が複数のサービスで計上されている場合があるため、"
           "合計は延べ人数である。",
           [34, 9, 11, 11, 11, 11, 34], freeze="B5")

r = header(ws, 4, ["介護サービスの種類", "事業所数", "従業員総数",
                   "東川町", "美瑛町", "東神楽町", "見方"])
for sv, tot in sorted(SV.items(), key=lambda x: -x[1]):
    per = {}
    for t in TOWNS:
        per[t] = sum(f2(rr[4]) or 0 for rr in S.SHIGEN[t] if rr[3] == sv)
    miss = [t for t in TOWNS if per[t] == 0]
    v = ("3町にある" if not miss else
         "%sにない" % "・".join(miss))
    r = body(ws, r, [sv, SVN[sv], round(tot, 1),
                     round(per["東川町"], 1) or None,
                     round(per["美瑛町"], 1) or None,
                     round(per["東神楽町"], 1) or None, v],
             {7: NG_O if miss else {}} if miss else {}, height=18,
             align={2: "center", 3: "right", 4: "right", 5: "right",
                    6: "right"})
r = body(ws, r, ["計（延べ）", N_ALL, round(sum(SUM_T.values()), 1),
                 round(SUM_T["東川町"], 1), round(SUM_T["美瑛町"], 1),
                 round(SUM_T["東神楽町"], 1), "―"],
         {1: GRAY}, height=20, bold=True,
         align={2: "center", 3: "right", 4: "right", 5: "right",
                6: "right"})

note(ws, r + 1,
     "注1）合計は延べ人数である。"
     "同一施設が複数のサービスで計上されている場合"
     "（介護老人福祉施設と短期入所生活介護など）、"
     "同じ職員が重複して数えられている。"
     "実人数を得るには施設単位での集約を要する［要確認］。"
     "注2）サービスの名称は本一覧の記載のままである。"
     "「通所リハビリテーションイケア」「地域老人福祉施設」など"
     "表記の揺れがあるが、原表のとおりとした。"
     "注3）本表は本一覧に載っている事業所のみである。"
     "区域外の事業所による給付は含まない。", 7)

# ============================================================ 04
ws = sheet("04_指定事業所一覧との網羅性",
           "指定事業所一覧（北海道）との網羅性",
           "北海道の介護保険事業所一覧により把握している"
           "区域内の指定事業所と、本一覧の網羅性を確認する。",
           [8, 34, 26, 14, 14, 30], freeze="C5")

SHITEI_N = sum(len(v) for v in H.SHITEI.values())
r = lead(ws, 4, "【1　件数の対照】", 6)
r = header(ws, r, ["", "出所", "件数", "単位", "時点", "内容"])
for a in [
    ("①", "北海道 介護保険事業所一覧", SHITEI_N, "サービス×事業所",
     "令和8年6月30日現在",
     "区域内の指定事業所。同一事業所が複数のサービスで計上される"),
    ("②", "3町の社会資源一覧（本一覧）", N_ALL, "サービス×事業所",
     "記載なし［要確認］",
     "介護保険外の事業所を含まない。従業員数を伴う"),
    ("③", "介護サービス情報公表システム", len(H.KOHYO), "事業所",
     "令和8年8月時点の公表内容",
     "個別公表画面を取得したもの。訪問介護13事業所すべてを含む"),
]:
    r = body(ws, r, list(a), {}, height=32,
             align={1: "center", 3: "right"})

r += 1
r = lead(ws, r, "【2　訪問介護での対照】", 6)
r = header(ws, r, ["", "出所", "事業所数", "従業員数", "", "備考"])
for a in [
    ("①", "北海道 介護保険事業所一覧",
     len(H.SHITEI.get("訪問介護", [])), "―", "",
     "指定事業所の数。従業員数は含まない"),
    ("②", "3町の社会資源一覧", len(PAIR), "%.1f人" % _t1, "",
     "従業員総数（全職種）"),
    ("③", "介護サービス情報公表システム", len(PAIR),
     "%.1f人／%.0f人" % (_t2, _t3), "",
     "総従業者数／訪問介護員等の実人数"),
]:
    r = body(ws, r, list(a), {}, height=24,
             align={1: "center", 3: "center", 4: "right"})
    ws.merge_cells(start_row=r - 1, start_column=4, end_row=r - 1, end_column=5)

note(ws, r + 1,
     "注1）3つの出所とも訪問介護は13事業所で一致する。"
     "区域内の訪問介護事業所は13で確定している。"
     "注2）本一覧の基準日が記載されていないため、"
     "指定事業所一覧（令和8年6月30日現在）及び"
     "公表画面（令和8年8月時点）との時点の違いは確認できない。"
     "従業員数は時点により動くため、基準日の確認をお願いする［要確認］。"
     "注3）本一覧に居宅介護支援事業所が含まれている一方、"
     "介護保険外の住まい（住宅型有料老人ホーム・"
     "サービス付き高齢者向け住宅）は含まれていない。"
     "これらの従業員数は介護人材実態調査によるほかない。", 6)

# ============================================================ 05
ws = sheet("05_反映する箇所",
           "計画本文及び成果品への反映",
           "本一覧の受領により更新する箇所と、"
           "更新しない箇所を示す。",
           [5, 26, 26, 34, 16], freeze="B5")

r = lead(ws, 4, "【1　更新する箇所】", 5)
r = header(ws, r, ["", "成果品", "箇所", "更新の内容", "作業量"])
for a in [
    ("①", "計画素案 第2章第6節", "介護事業所の現状",
     "3町74事業所の従業員数を表として追加する。"
     "訪問介護以外のサービスの従業員数を初めて示せる。", "中"),
    ("②", "計画素案 第2章第3節", "サービス提供体制の状況",
     "サービス別の従業員数を、事業所数・定員と並べて示す。", "小"),
    ("③", "住まいと施設の公表名簿との突合",
     "訪問介護13事業所の把握の状況",
     "第3の出所として社会資源一覧を加え、3出所の対照とする。", "小"),
    ("④", "アンケート調査の集計分析報告書",
     "09_供給構造④_従事者数、14_公表データによる補完",
     "介護人材実態調査の未回答分を補完できることを追記する。", "中"),
    ("⑤", "事業所調査の照会票と確定値管理表", "04_確定値表",
     "確定値表に本一覧の値を「参考値」として加える。"
     "重複計上13人の判断材料となる。", "小"),
    ("⑥", "エビデンス集2（事業所と公表情報）", "新規シート",
     "社会資源一覧74事業所を収録する。", "小"),
    ("⑦", "業務工程管理表 06_資料提供依頼一覧", "資料No.7の状態",
     "事業所・施設の定員と稼働の実績のうち、"
     "従業員数の部分が受領済みとなる。", "小"),
]:
    r = body(ws, r, list(a), {}, height=40, align={1: "center", 5: "center"})

r += 1
r = lead(ws, r, "【2　更新しない箇所】", 5)
r = header(ws, r, ["", "成果品", "箇所", "更新しない理由", ""])
for a in [
    ("①", "計画素案 第2章第6節", "訪問介護員等181人",
     "公表画面による職種を限った実人数であり、"
     "本一覧の従業員総数とは定義が異なる。"
     "供給力を示す数値としては公表画面を用いる", ""),
    ("②", "代表KPI H13（職種別従事者数）", "基準値535.17",
     "見える化M2系列（認定者1万対）によるものであり、"
     "本一覧では算定できない", ""),
    ("③", "計画素案 第2章第6節", "介護人材実態調査の集計値",
     "資格・年代・勤続年数・採用退職は本一覧に含まれない。"
     "調査の値をそのまま用いる", ""),
    ("④", "将来推計 第2段階", "サービス見込量",
     "従業員数は見込量の算定に用いていない。"
     "供給制約の検討に用いる余地はあるが、"
     "本一覧に定員・稼働の情報がない", ""),
]:
    r = body(ws, r, list(a), {}, height=40, align={1: "center"})
    ws.merge_cells(start_row=r - 1, start_column=4, end_row=r - 1, end_column=5)

note(ws, r + 1,
     "注1）本一覧の受領により、"
     "計画本文の主要な数値が変わることはない。"
     "訪問介護員等181人も、介護人材実態調査の集計値も、"
     "定義が異なるため置き換えない。"
     "注2）本一覧の価値は、"
     "これまで訪問介護でしか得られなかった全数の従業員数が"
     "全サービスについて得られたことにある。"
     "第2章第3節・第6節の記述を厚くできる。"
     "注3）基準日が確認できるまで、"
     "本一覧による記述には「令和8年8月20日受領の3町の社会資源一覧による」"
     "と付す。", 5)

del wb["Sheet"]
wb.save(OUT)
print("saved:", OUT)
for ws in wb:
    print("  -", ws.title, ws.max_row, "rows")
print()
print("社会資源一覧 %d事業所（東川%d・美瑛%d・東神楽%d）"
      % (N_ALL, len(S.SHIGEN["東川町"]), len(S.SHIGEN["美瑛町"]),
         len(S.SHIGEN["東神楽町"])))
print("従業員総数（延べ） %.1f人　東川%.0f／美瑛%.0f／東神楽%.0f"
      % (sum(SUM_T.values()), SUM_T["東川町"], SUM_T["美瑛町"],
         SUM_T["東神楽町"]))
print("訪問介護13事業所　社会資源%.1f人／公表総従業者%.1f人／訪問介護員等%.0f人"
      % (_t1, _t2, _t3))
print("  完全一致%d／差1人%d／差2人以上%d"
      % (sum(1 for d in _diff if d == 0), sum(1 for d in _diff if d == 1),
         sum(1 for d in _diff if d >= 2)))
