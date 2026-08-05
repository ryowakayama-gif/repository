# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　需要3シナリオ（低位・標準・高位）の感度表.

将来推計 第2段階（サービス見込量）の基本ケースに対し、
前提の置き方を変えた場合に見込量がどこまで動くかを示す。

3シナリオの設計
  標準  認定者数はトレンド継続（第1段階シナリオ②）、
        利用率及び1人当たり利用日数・回数は令和7年度で固定する。
        これは将来推計 第2段階の基本ケースと同じである。
  低位  各前提を需要が小さくなる側に置いた場合の下限。
  高位  各前提を需要が大きくなる側に置いた場合の上限。

低位・高位は、前提を同時に片側へ置いた場合の包絡であり、
確率的な区間（信頼区間）ではない。
実際にすべての前提が同じ側に振れる可能性は高くない。

レバー（前提）は4つである。
  ① 認定者数        第1段階の3シナリオ（現状固定・トレンド継続・令和元年水準）
  ② 区分別の利用率   令和7年度で固定 ／ 直近3年（令和5〜7年度）の趨勢を延長
  ③ 1人当たり日数    令和7年度で固定 ／ 同上
  ④ 要介護度別構成比 令和8年3月末で固定 ／ 要介護1への集中が続く

シート構成
  00_3シナリオの設計
  01_レバーごとの感度（1つずつ動かす）
  02_認定者数と区分別利用者数の3シナリオ
  03_サービス種別の見込量の3シナリオ
  04_年度別の3シナリオ
  05_施設・居住系の定員制約との突合
  06_調査結果による上振れの材料
  07_採用値の決定に要すること
"""

import io
import runpy
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_hokkaido_roster as R
import data_hokkaido_shitei as H
import data_survey2025 as S

OUT = ("/home/user/repository/output/"
       "第10期計画_将来推計_需要3シナリオの感度表.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------- 第2段階
_buf, _old = io.StringIO(), sys.stdout
sys.stdout = _buf
try:
    G = runpy.run_path("build_projection2.py")
finally:
    sys.stdout = _old

CARE = G["CARE"]
YEARS = G["YEARS"]
SVC = G["SVC"]
USE, SVCUSE, PER = G["USE"], G["SVCUSE"], G["PER"]
MIX = G["MIX"]
nin, total, SC = G["nin"], G["total"], G["SC"]
use_rate, per_val = G["use_rate"], G["per_val"]
TREND_USE, TREND_PER = G["TREND_USE"], G["TREND_PER"]
LAST3 = G["LAST3"]

# 定員（住まいと施設の公表名簿との突合による）
CAP_TOKUTEI = sum(y["定員"] for y in R.YU if y["類型"] == "介護付")
CAP_TOKUYO = sum(t["定員"] for t in H.TOKUYO if not t["地域密着型"])
CAP_CHITOKU = sum(t["定員"] for t in H.TOKUYO if t["地域密着型"])
CAP_ROKEN = sum(s["定員"] or 0 for s in S.SHI if s["種別"] == 7)
CAP_GH = sum(s["定員"] or 0 for s in S.SHI if s["種別"] == 4) + 18
CAP_SHISETSU = CAP_TOKUYO + CAP_CHITOKU + CAP_ROKEN
CAP_KYOJU = CAP_GH + CAP_TOKUTEI

# 要介護1への集中（第2段階 05シート③と同じ置き方。年13人を4年延長）
SHIFT = 13.0 * 4
_HEAVY = ["要介護2", "要介護3", "要介護4", "要介護5"]
_TOTMIX = sum(MIX[c] for c in _HEAVY)


def nin_mix(y, sc=2, shift=False):
    """要介護度別の認定者数。shift=True で要介護1への集中を織り込む。"""
    n = dict(nin(y, sc))
    if shift:
        d = SHIFT * (int(y) - 2025) / 4.0
        n["要介護1"] = n["要介護1"] + d
        for c in _HEAVY:
            n[c] = max(0.0, n[c] - d * MIX[c] / _TOTMIX)
    return n


def kubun(y, sc=2, trend_use=False, shift=False):
    """区分別（在宅・居住系・施設）の月平均利用者数。"""
    n = nin_mix(y, sc, shift)
    return {k: sum(n[c] * use_rate(k, c, y, trend_use) / 100 for c in CARE)
            for k in ["在宅", "居住系", "施設"]}


def svc_vol(nm, y, sc=2, trend_per=False, shift=False):
    """サービス種別の見込量（回・日／月）。"""
    n = nin_mix(y, sc, shift)
    return sum(n[c] * SVCUSE[nm][c] / 100 * (per_val(nm, c, y, trend_per) or 0)
               for c in CARE)


def tot_series(c46):
    for k, d in G["J"].D[c46]["系列"].items():
        if k.endswith("(合計)"):
            return d["値"]
    return {}


def svc_vol_flat(nm, c46, y, sc=2, shift=False):
    """1人当たりを「合計の趨勢」で一様に延長した場合の見込量。

    要介護度別に趨勢を延長すると、要介護度ごとの振れが増幅される。
    合計の1人当たりの推移を比率として全要介護度に一様に掛けることで、
    増幅を受けない置き方の見込量を得る。
    """
    tot = tot_series(c46)
    xs = [tot.get(k) for k in LAST3]
    xs = [x for x in xs if x is not None]
    if len(xs) < 2 or not xs[-1]:
        return None
    slope = (xs[-1] - xs[0]) / (len(xs) - 1)
    d = int(y) - 2025
    ratio = max(0.0, (xs[-1] + slope * d) / xs[-1])
    n = nin_mix(y, sc, shift)
    return sum(n[c] * SVCUSE[nm][c] / 100 * (PER[nm][c] or 0) * ratio
               for c in CARE)


def env_kubun(y, k):
    """区分別の低位・標準・高位。各レバーを片側に置いた包絡。"""
    cand = []
    for sc in (1, 2, 3):
        for tu in (False, True):
            for sh in (False, True):
                cand.append(kubun(y, sc, tu, sh)[k])
    std = kubun(y, 2, False, False)[k]
    return min(cand), std, max(cand)


C46 = {nm: c46 for _c, nm, c46, _u in SVC}


def env_svc(nm, y):
    """サービス種別の低位・標準・高位。

    1人当たりの延長は、要介護度別の趨勢による置き方と
    合計の趨勢を一様に掛ける置き方の両方を候補に含める。
    """
    cand = []
    for sc in (1, 2, 3):
        for tp in (False, True):
            for sh in (False, True):
                cand.append(svc_vol(nm, y, sc, tp, sh))
    c46 = C46.get(nm)
    if c46:
        for sc in (1, 2, 3):
            for sh in (False, True):
                f = svc_vol_flat(nm, c46, y, sc, sh)
                if f is not None:
                    cand.append(f)
    std = svc_vol(nm, y, 2, False, False)
    return min(cand), std, max(cand)


# ================================================================ 体裁
wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=len(widths))
    ws.row_dimensions[2].height = 52
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=20, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=104):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


RIGHT = {i: "right" for i in range(2, 26)}
Y29 = "2029"

# ================================================================ 00
ws = sheet("00_3シナリオの設計", "需要3シナリオ（低位・標準・高位）の設計",
           "サービス見込量は「要介護度別の認定者数 × 要介護度別の利用率 × "
           "受給者1人当たりの利用日数・回数」で算定する。"
           "この3つの前提と要介護度別の構成比を合わせた4つのレバーを"
           "どう置くかで見込量が変わる。"
           "本表は、レバーの置き方による見込量の幅を示すものである。",
           [4, 20, 30, 30, 30, 22], freeze="A5")

r = lead(ws, 4, "【1　4つのレバーと3シナリオでの置き方】", 6)
r = header(ws, r, ["No.", "レバー（前提）", "低位（需要が小さい側）",
                   "標準（基本ケース）", "高位（需要が大きい側）",
                   "根拠となる資料"])
for no, nm, lo, st, hi, src in [
    (1, "認定者数", "第1段階の3シナリオのうち最小のもの",
     "② トレンド継続（令和元年から令和8年の変化を延長）",
     "第1段階の3シナリオのうち最大のもの",
     "将来推計 第1段階、見える化B系列"),
    (2, "区分別・要介護度別の利用率",
     "令和7年度で固定と直近3年の趨勢延長のうち小さい方",
     "令和7年度（令和8年1月サービス提供分まで）で固定",
     "同じく大きい方", "見える化D45-a〜c"),
    (3, "受給者1人当たりの利用日数・回数",
     "令和7年度で固定、要介護度別の趨勢を延長、"
     "合計の趨勢を一様に掛けるの3つのうち小さいもの",
     "令和7年度で固定", "同じく大きいもの", "見える化D46-a〜n"),
    (4, "要介護度別の構成比",
     "令和8年3月末で固定と要介護1への集中のうち小さい方",
     "令和8年3月末で固定",
     "同じく大きい方", "見える化B3-a"),
]:
    r = body(ws, r, [no, nm, lo, st, hi, src],
             {3: MID_B, 4: OK_G, 5: NG_O}, height=44,
             align={1: "center"})

r += 1
r = lead(ws, r, "【2　低位・高位の読み方】", 6)
for txt in [
    "低位・高位は、4つのレバーを同時に片側へ置いた場合の包絡（上下限）である。"
    "統計的な信頼区間ではなく、すべてのレバーが同じ側に振れる可能性は高くない。",
    "レバーごとに振れる向きが異なる。"
    "たとえば利用率の趨勢は在宅では上向き、居住系・施設では下向きである"
    "（01シート）。このため低位・高位はサービスごとに"
    "異なるレバーの組合せから生じている。",
    "採用値は標準（基本ケース）とすることを受託者案とする。"
    "低位・高位は、給付費・保険料がどこまで振れるかを"
    "確認するために用いる（07シート）。",
    "本表の見込量は月平均の利用者数・回数・日数である。"
    "給付費への換算にはサービス単価を要し、単価は未受領である"
    "（資料提供依頼No.9・No.19）。",
]:
    r = body(ws, r, [None, txt, None, None, None, None], height=32)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1,
                   end_column=6)

r += 1
r = lead(ws, r, "【3　令和11年度の幅（要約）】", 6)
r = header(ws, r, ["区分", "低位", "標準", "高位", "標準からの幅", "見方"])
for k in ["在宅", "居住系", "施設"]:
    lo, st, hi = env_kubun(Y29, k)
    r = body(ws, r, [k + "サービスの利用者数（人／月）", "%.0f" % lo,
                     "%.0f" % st, "%.0f" % hi,
                     "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                            (hi / st - 1) * 100),
                     "定員制約は05シートで確認する"
                     if k != "在宅" else
                     "低位が標準と一致する。"
                     "在宅は4つのレバーがいずれも上向きに働く"],
             {3: OK_G}, height=20, align=RIGHT)
lo, st, hi = env_svc("訪問介護", Y29)
r = body(ws, r, ["訪問介護の見込量（回／月）", "%.0f" % lo, "%.0f" % st,
                 "%.0f" % hi,
                 "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                        (hi / st - 1) * 100),
                 "区域内で最も実数が大きく、幅も最も大きい"],
         {3: OK_G}, height=20, align=RIGHT)

note(ws, r + 1,
     "注1）本表は将来推計 第2段階（サービス見込量）の基本ケースを標準とし、"
     "同 05シートの感度分析を3シナリオに組み替えたものである。"
     "注2）認定者数の3シナリオは第1段階による。"
     "令和11年度は①現状固定2,044人、②トレンド継続2,005人、"
     "③令和元年水準2,112人であり、下限は②（基本ケース）と一致する。"
     "認定者数は上側に振れる余地の方が大きい。"
     "注3）令和7年度は11か月分（令和8年1月サービス提供分まで）である。"
     "確定値が公表された時点で再計算する。"
     "注4）本表はいずれかを採用値とするものではない。"
     "採用値は発注者のご意向と介護保険事業計画策定委員会の審議による。", 6)

# ================================================================ 01
ws = sheet("01_レバーごとの感度", "レバーごとの感度（1つずつ動かした場合）",
           "どのレバーが見込量をどれだけ動かすかを、1つずつ動かして確認する。"
           "3シナリオはこれらを組み合わせたものである。"
           "レバーによって振れる向きが異なることに留意する。",
           [26, 14, 14, 14, 16, 40], freeze="B5")

r = lead(ws, 4, "【① 認定者数のシナリオ（令和11年度）】", 6)
r = header(ws, r, ["シナリオ", "認定者数", "在宅", "居住系", "施設",
                   "基本ケースとの差"])
base_t = total(Y29, 2)
for sc in (1, 2, 3):
    kb = kubun(Y29, sc)
    r = body(ws, r, [SC[sc], "%.0f" % total(Y29, sc), "%.0f" % kb["在宅"],
                     "%.0f" % kb["居住系"], "%.0f" % kb["施設"],
                     "基本ケース" if sc == 2 else
                     "認定者数 %+.1f％" % ((total(Y29, sc) / base_t - 1) * 100)],
             {1: OK_G if sc == 2 else None}, height=20, align=RIGHT)
r += 1

r = lead(ws, r, "【② 区分別の利用率（令和11年度）】", 6)
r = header(ws, r, ["区分", "令和7年度で固定", "直近3年の趨勢を延長", "差",
                   "差の割合", "趨勢の向きと見方"])
for k in ["在宅", "居住系", "施設"]:
    b = kubun(Y29, 2, False)[k]
    t = kubun(Y29, 2, True)[k]
    tr = sum(TREND_USE[k][c] for c in CARE) / len(CARE)
    r = body(ws, r, [k, "%.0f" % b, "%.0f" % t, "%+.0f" % (t - b),
                     "%+.1f％" % ((t / b - 1) * 100),
                     "年平均%+.2fポイント。%s"
                     % (tr, "利用率が上がる方向" if tr > 0
                        else "利用率が下がる方向")],
             {5: NG_O if abs(t / b - 1) > 0.05 else None}, height=20,
             align=RIGHT)
r += 1

r = lead(ws, r, "【③ 受給者1人当たりの利用日数・回数（令和11年度）】", 6)
r = header(ws, r, ["サービス種別", "単位", "令和7年度で固定",
                   "要介護度別の趨勢を延長", "差の割合",
                   "1人当たり（合計）の推移 令和5→7年度と、"
                   "合計の趨勢で一様に延長した場合の差の割合"])
SMALL, GAP = [], []
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    b = svc_vol(nm, Y29, 2, False)
    if b == 0:
        continue
    t = svc_vol(nm, Y29, 2, True)
    tot = {}
    for k, d in G["J"].D[c46]["系列"].items():
        if k.endswith("(合計)"):
            tot = d["値"]
            break
    seq = "→".join("%g" % tot.get(y) if tot.get(y) is not None else "―"
                   for y in LAST3)
    dr = (t / b - 1) * 100
    fl = svc_vol_flat(nm, c46, Y29)
    frd = (fl / b - 1) * 100 if fl else None
    small = b < 300
    if small:
        SMALL.append(nm)
    gap = frd is not None and abs(dr - frd) >= 10
    if gap:
        GAP.append(nm)
    r = body(ws, r, [nm, unit + "／月", "%.0f" % b, "%.0f" % t,
                     "%+.1f％" % dr,
                     "%s（合計の趨勢で一様に延長すると%s）"
                     % (seq, "%+.1f％" % frd if frd is not None else "―")],
             {5: GRAY if small else (NG_O if abs(dr) >= 10 else None),
              6: IN_Y if gap else None},
             height=20, align=RIGHT)
r += 1

r = lead(ws, r, "【④ 要介護度別の構成比（令和11年度）】", 6)
r = header(ws, r, ["前提", "要介護1", "在宅", "施設", "訪問介護",
                   "見方"])
for sh, lab in [(False, "令和8年3月末の構成比で固定（基本ケース）"),
                (True, "要介護1への集中が続く（年13人・4年で52人）")]:
    n = nin_mix(Y29, 2, sh)
    kb = kubun(Y29, 2, False, sh)
    hv = svc_vol("訪問介護", Y29, 2, False, sh)
    r = body(ws, r, [lab, "%.0f" % n["要介護1"], "%.0f" % kb["在宅"],
                     "%.0f" % kb["施設"], "%.0f" % hv,
                     "" if not sh else
                     "在宅は増え、施設と訪問介護の見込量は減る"],
             {1: OK_G if not sh else IN_Y}, height=20, align=RIGHT)

note(ws, r + 1,
     "注1）見込量を最も動かすのは③の1人当たり利用日数・回数である。"
     "訪問介護は直近3年の趨勢を延長すると2割以上動く。"
     "注2）②の利用率の趨勢は、在宅では上向き、居住系・施設では下向きであり、"
     "向きが揃っていない。"
     "低位・高位はサービスごとに異なるレバーの組合せから生じる。"
     "注3）灰色に網掛けした種別（%s）は月平均の実数が300未満であり、"
     "要介護度別の趨勢の振れが増幅されやすい。"
     "注3-2）右端の列は、1人当たりの合計の推移と、"
     "その趨勢を全要介護度に一様に掛けた場合の差の割合である。"
     "要介護度別に延長した場合（左から2列目）との差が10ポイント以上ある種別"
     "（%s）は、2つの置き方が一致しない。"
     "要介護度別の趨勢は要介護度ごとの振れを増幅し、"
     "合計の趨勢は要介護度別の構成の変化を反映しない。"
     "これらの種別の低位・高位は置き方に依存するため、"
     "幅の広い方を上下限として読む（03シートは両方を候補に含めている）。"
     "注4）④の要介護1への集中は、要介護1の1人当たり利用回数が"
     "要介護5より大幅に少ないため、"
     "在宅の利用者数を増やす一方で訪問介護の見込量を減らす。"
     % ("・".join(SMALL), "・".join(GAP) or "なし"), 6, 132)

# ================================================================ 02
ws = sheet("02_認定者数と区分別利用者数",
           "認定者数と区分別利用者数の3シナリオ",
           "認定者数と、在宅・居住系・施設の区分別の月平均利用者数について、"
           "低位・標準・高位を示す。"
           "低位・高位は4つのレバーを同時に片側へ置いた場合の包絡である。",
           [22, 14, 14, 14, 14, 16, 30], freeze="B5")

r = lead(ws, 4, "【① 認定者数（人）】", 7)
r = header(ws, r, ["年度", "低位", "標準", "高位", "標準からの幅（人）",
                   "標準からの幅（％）", "備考"])
for y, lab in YEARS:
    vals = [total(y, sc) for sc in (1, 2, 3)]
    lo, st, hi = min(vals), total(y, 2), max(vals)
    r = body(ws, r, [lab, "%.0f" % lo, "%.0f" % st, "%.0f" % hi,
                     "%+.0f 〜 %+.0f" % (lo - st, hi - st),
                     "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                            (hi / st - 1) * 100),
                     "下限は標準と一致する" if abs(lo - st) < 0.5 else ""],
             {3: OK_G}, height=20, align=RIGHT)
r += 1

for k in ["在宅", "居住系", "施設"]:
    r = lead(ws, r, "【%sサービスの月平均利用者数（人／月）】" % k, 7)
    r = header(ws, r, ["年度", "低位", "標準", "高位", "標準からの幅（人）",
                       "標準からの幅（％）", "備考"])
    for y, lab in YEARS:
        lo, st, hi = env_kubun(y, k)
        r = body(ws, r, [lab, "%.0f" % lo, "%.0f" % st, "%.0f" % hi,
                         "%+.0f 〜 %+.0f" % (lo - st, hi - st),
                         "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                                (hi / st - 1) * 100), ""],
                 {3: OK_G}, height=20, align=RIGHT)
    r += 1

note(ws, r,
     "注1）認定者数の低位は第1段階のシナリオ②（トレンド継続）であり、"
     "標準と一致する。"
     "第1段階の3シナリオはいずれも標準以上の水準にあるため、"
     "認定者数は下側に振れる余地が小さい。"
     "注2）区分別の利用者数の幅は、認定者数のほか"
     "利用率の趨勢と要介護度別の構成比の影響を含む。"
     "注3）在宅・居住系・施設の合計は認定者数と一致しない。"
     "いずれのサービスも利用していない認定者がいるためである"
     "（令和7年度で24.5％）。", 7)

# ================================================================ 03
ws = sheet("03_サービス種別の見込量",
           "サービス種別の見込量の3シナリオ（令和11年度）",
           "サービス種別ごとの月平均の見込量について、低位・標準・高位を示す。"
           "1人当たりの利用日数・回数が得られない種別"
           "（居宅療養管理指導、福祉用具貸与、小規模多機能型居宅介護）は"
           "利用者数のみを算定するため、本表には含めない。",
           [24, 8, 13, 13, 13, 18, 12, 26], freeze="B5")

r = header(ws, 4, ["サービス種別", "単位", "低位", "標準", "高位",
                   "標準からの幅（％）", "実数の大小", "主に効いているレバー"])
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    lo, st, hi = env_svc(nm, Y29)
    if st == 0:
        continue
    small = st < 300
    # どのレバーが最も効いているか
    d_per = abs(svc_vol(nm, Y29, 2, True) / st - 1)
    d_nin = max(abs(svc_vol(nm, Y29, sc) / st - 1) for sc in (1, 3))
    d_mix = abs(svc_vol(nm, Y29, 2, False, True) / st - 1)
    lever = max([(d_per, "1人当たりの日数・回数"), (d_nin, "認定者数"),
                 (d_mix, "要介護度別の構成比")])[1]
    r = body(ws, r, [nm, unit + "／月", "%.0f" % lo, "%.0f" % st,
                     "%.0f" % hi,
                     "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                            (hi / st - 1) * 100),
                     "小（300未満）" if small else "―", lever],
             {4: OK_G, 7: GRAY if small else None}, height=20, align=RIGHT)

r += 1
r = lead(ws, r, "【1人当たりの日数・回数が得られない種別（利用者数のみ）】", 8)
r = header(ws, r, ["サービス種別", "単位", "低位", "標準", "高位",
                   "標準からの幅（％）", "", "備考"])
for _code, nm, c46, unit in SVC:
    if c46:
        continue
    cand = [sum(nin_mix(Y29, sc, sh)[c] * SVCUSE[nm][c] / 100 for c in CARE)
            for sc in (1, 2, 3) for sh in (False, True)]
    st = sum(nin_mix(Y29, 2)[c] * SVCUSE[nm][c] / 100 for c in CARE)
    lo, hi = min(cand), max(cand)
    if st == 0:
        continue
    r = body(ws, r, [nm, "人／月", "%.0f" % lo, "%.0f" % st, "%.0f" % hi,
                     "%+.1f％ 〜 %+.1f％" % ((lo / st - 1) * 100,
                                            (hi / st - 1) * 100), "",
                     "1人当たりの量は見込量の様式に含まれない"],
             {4: OK_G}, height=20, align=RIGHT)

note(ws, r + 1,
     "注1）「主に効いているレバー」は、レバーを1つずつ動かしたときの"
     "変化率が最も大きいものである。"
     "多くの種別で1人当たりの利用日数・回数が最も効いている。"
     "注2）実数が300未満の種別は、要介護度別の趨勢の振れが増幅されやすく、"
     "幅が過大に出る。01シートの合計の推移と併せて読む。"
     "注3）小規模多機能型居宅介護は、在宅生活改善調査で"
     "「より適切と思われるサービス」の最多（44件）に挙げられているが、"
     "この件数を見込量に加算していない。理由は06シートに示す。", 8)

# ================================================================ 04
ws = sheet("04_年度別の3シナリオ", "年度別の3シナリオ（主要サービス）",
           "計画期間の3年度について、主要なサービスの低位・標準・高位を示す。"
           "見込量の様式はサービス種別ごとに年度別の値を要するため、"
           "本シートがその原案となる。",
           [24, 8, 12, 12, 12, 12, 12, 12, 12, 12, 12], freeze="B5")

MAIN = ["訪問介護", "訪問看護", "通所介護", "通所リハビリテーション",
        "短期入所生活介護", "地域密着型通所介護"]
r = header(ws, 4,
           ["サービス種別", "単位"]
           + ["%s\n%s" % (lab, k) for lab in [y[1] for y in YEARS]
              for k in ["低位", "標準", "高位"]])
for nm in MAIN:
    vals = []
    for y, _lab in YEARS:
        lo, st, hi = env_svc(nm, y)
        vals += ["%.0f" % lo, "%.0f" % st, "%.0f" % hi]
    unit = [u for _c, n2, c46, u in SVC if n2 == nm][0]
    r = body(ws, r, [nm, unit + "／月"] + vals,
             {4: OK_G, 7: OK_G, 10: OK_G}, height=20, align=RIGHT)

r += 1
r = lead(ws, r, "【区分別の利用者数（人／月）】", 11)
r = header(ws, r, ["区分", "単位"]
           + ["%s\n%s" % (lab, k) for lab in [y[1] for y in YEARS]
              for k in ["低位", "標準", "高位"]])
for k in ["在宅", "居住系", "施設"]:
    vals = []
    for y, _lab in YEARS:
        lo, st, hi = env_kubun(y, k)
        vals += ["%.0f" % lo, "%.0f" % st, "%.0f" % hi]
    r = body(ws, r, [k, "人／月"] + vals, {4: OK_G, 7: OK_G, 10: OK_G},
             height=20, align=RIGHT)

note(ws, r + 1,
     "注1）計画本文の見込量の様式に転記するのは標準（緑の列）である。"
     "低位・高位は保険料の感応度の確認に用いる。"
     "注2）年度が進むほど幅が広がる。趨勢を延長する年数が増えるためである。"
     "注3）本シートに掲げていない種別は03シートによる。", 11)

# ================================================================ 05
ws = sheet("05_定員制約との突合", "施設・居住系の定員制約との突合",
           "施設・居住系サービスの見込量は定員を超えて実現しない。"
           "高位シナリオでも区域内定員に収まるかを確認する。"
           "定員は北海道が公表する名簿等による"
           "（住まいと施設の公表名簿との突合）。",
           [22, 12, 12, 12, 12, 14, 34], freeze="B5")

r = header(ws, 4, ["区分", "低位", "標準", "高位", "区域内定員",
                   "高位／定員", "判定"])
for k, cap, capnm in [("施設", CAP_SHISETSU, "特養160＋地域密着型特養62＋老健240"),
                      ("居住系", CAP_KYOJU, "グループホーム99＋特定施設156")]:
    lo, st, hi = env_kubun(Y29, k)
    rate = hi / cap * 100
    r = body(ws, r, [k + "サービス（令和11年度）", "%.0f" % lo, "%.0f" % st,
                     "%.0f" % hi, "%d" % cap, "%.1f％" % rate,
                     "高位でも定員内（%s）" % capnm if rate <= 100
                     else "高位は定員を超える"],
             {6: NG_O if rate > 100 else OK_G}, height=24, align=RIGHT)

r += 1
r = lead(ws, r, "【年度別（高位シナリオ）】", 7)
r = header(ws, r, ["区分", "令和9年度", "令和10年度", "令和11年度",
                   "区域内定員", "令和11年度の充足率", "備考"])
for k, cap in [("施設", CAP_SHISETSU), ("居住系", CAP_KYOJU)]:
    hs = [env_kubun(y, k)[2] for y, _l in YEARS]
    r = body(ws, r, [k, "%.0f" % hs[0], "%.0f" % hs[1], "%.0f" % hs[2],
                     "%d" % cap, "%.1f％" % (hs[2] / cap * 100),
                     "住所地特例により他の保険者の被保険者も入居するため、"
                     "差がそのまま空きを示すものではない"],
             {6: OK_G if hs[2] <= cap else NG_O}, height=32, align=RIGHT)

note(ws, r + 1,
     "注1）認知症対応型共同生活介護の区域内定員99人は、"
     "グループホームくるみの郷1事業所を含まない。"
     "同事業所は介護サービス情報公表システムに掲載がなく"
     "定員を確認できていないため、居住系の定員は過小である"
     "（資料提供依頼No.18、点検事項No.35）。"
     "注2）施設・居住系の見込量は定員により制約される。"
     "高位シナリオが定員を超える場合は、"
     "整備するか、区域外の施設の利用を見込むかの判断を要する。"
     "注3）特定施設は定員156人に対し入居者154人（98.7％）でほぼ満室である。"
     "居住系の見込量の上振れを区域内で受けられる余地は小さい。", 7)

# ================================================================ 06
ws = sheet("06_調査結果による上振れの材料",
           "調査結果による上振れの材料と、加算しない理由",
           "実施済み調査で挙げられた件数は、"
           "上記のレバーとは別に見込量を上振れさせる材料となり得る。"
           "ただし、そのまま需要人数として加算することはできない。"
           "加算するために必要な補正を整理する。",
           [4, 26, 12, 34, 34, 20], freeze="A5")

r = header(ws, 4, ["No.", "調査結果", "件数", "そのまま加算できない理由",
                   "加算するために必要な補正", "補正に要する資料"])
for no, res, cnt, why, fix, src in [
    (1, "より適切と思われるサービス（小規模多機能型居宅介護）", "44件",
     "利用者票99票のうち36票（36.4％）が同一法人の"
     "小規模多機能2事業所からの提出であり、"
     "44件のうち31件（70.5％）がこの2事業所からの回答である。"
     "回答事業所の構成による偏りを含む。",
     "提出元別の集計（実施済み）、重複の除去、"
     "現に利用しているサービスとの重なりの整理、利用意向の確認",
     "給付実績、町別の登録枠と稼働状況"),
    (2, "24時間対応の3サービスを必要とする回答", "延べ26件",
     "複数回答であり実人数ではない。"
     "現在利用しているサービスとの重なりも整理していない。",
     "実人数への換算、現在の利用との重なりの整理、夜間の要請頻度の把握",
     "利用者票の再集計、給付実績、事業所への聞き取り"),
    (3, "特別養護老人ホームの待機者", "82〜137人",
     "同一法人の2施設で同数が計上されており、"
     "個人単位の名寄せができていない。"
     "施設入所中の待機者と在宅の待機者も区分できていない。",
     "個人単位の名寄せ、申込の継続意思の確認、"
     "施設入所中と在宅の区分",
     "匿名化した申込者の照合キー、申込日、現在の居所"),
    (4, "在宅生活の維持が困難とされた方", "72人",
     "事業所が課題があると判断した利用者を抽出する設計であり、"
     "区域内の在宅利用者全体の割合ではない。",
     "母集団の推定、住み替え意向の確認、"
     "既にサービスを利用している方との重なりの整理",
     "居宅介護支援事業所の利用者名簿、給付実績"),
    (5, "認定を受けているがサービスを利用していない方", "486人",
     "未利用の理由が分かっていない。"
     "施策により利用率を上げる目標を立てる場合に、"
     "どの程度が実際に利用に結びつくかを置けない。",
     "未利用の理由の分解、利用に結びつく層の推定",
     "居宅サービス計画未作成者の内訳（資料提供依頼No.8）"),
]:
    r = body(ws, r, [no, res, cnt, why, fix, src], {3: IN_Y}, height=64,
             align={1: "center", 3: "right"})

note(ws, r + 1,
     "注1）上記5件はいずれも上振れの方向の材料である。"
     "補正を行わないまま高位シナリオに加算すると、"
     "見込量が過大になり保険料が過大に算定される。"
     "注2）本表の件数は「実施済み調査 結果報告書」による。"
     "同報告書は、調査結果を見込量の算定式の入力値としない旨を"
     "第8章第2節に明記している。"
     "注3）補正が済んだ場合は、高位シナリオに上乗せするのではなく、"
     "利用率又は1人当たりの利用日数・回数の置き方に反映する。"
     "算定式の外で人数を足すことはしない。", 6)

# ================================================================ 07
ws = sheet("07_採用値の決定に要すること",
           "採用値の決定に要すること",
           "計画本文に載せる見込量をどのシナリオとするかは、"
           "発注者のご意向と介護保険事業計画策定委員会の審議による。"
           "決定に要する事項を整理する。",
           [4, 26, 40, 30, 20, 12], freeze="A5")

r = header(ws, 4, ["No.", "決定すること", "受託者の案", "根拠・留意点",
                   "決まらないと確定しないもの", "時期"])
for no, item, plan, why, blk, when in [
    (1, "計画本文に載せる見込量のシナリオ",
     "標準（基本ケース）を採用する。"
     "利用率と1人当たりの利用日数・回数は令和7年度で固定する。",
     "令和7年度は11か月分であり、確定値が出た時点で再計算する。"
     "趨勢を延長する場合、種別により向きが異なり、"
     "実数の小さい種別で幅が過大に出る。",
     "第6章第2節の見込量の表", "R8.10"),
    (2, "認定者数のシナリオ",
     "② トレンド継続を採用する。",
     "3シナリオの幅は令和11年度で2,005〜2,112人（＋5.4％）である。"
     "下限は②と一致し、上側に振れる余地の方が大きい。",
     "第6章第1節の認定者数の表", "R8.10"),
    (3, "低位・高位の使い方",
     "計画本文には載せず、保険料の感応度の確認に用いる。"
     "委員会資料には幅として示す。",
     "低位・高位は前提を同時に片側へ置いた包絡であり、"
     "確率的な区間ではない。"
     "本文に3本並べると、いずれかが目標値と誤読されるおそれがある。",
     "委員会資料、第6章第6節の保険料の説明", "R8.11"),
    (4, "サービス単価",
     "受領後に見込量へ乗じて給付費を算定する。",
     "本表は月平均の利用者数・回数・日数までである。"
     "給付費への換算には単価を要する。",
     "第6章第3節の給付費、第6章第6節の保険料", "R8.10"),
    (5, "施策により利用率を上げる目標を立てるか",
     "未利用者486人（24.5％）の扱いを決める。"
     "利用率を固定するのが基本ケースである。",
     "第5章の施策により利用率を上げる目標を立てる場合は、"
     "見込量に上乗せする必要がある。"
     "未利用の理由は資料提供依頼No.8で照会中である。",
     "第6章第2節の見込量、第5章の達成目標", "R8.10"),
    (6, "調査結果の上振れ材料の扱い",
     "補正が済むまで加算しない。"
     "補正後は利用率又は1人当たりの量の置き方に反映する。",
     "06シートのとおり、5件はいずれも補正を要する。"
     "点検事項35件（うち重大6件）の取扱いが前提となる。",
     "第6章第2節・第4節、需要3シナリオの確定", "R8.9"),
    (7, "居住系の定員の確定",
     "グループホームくるみの郷の定員をご提供いただく。",
     "区域内定員255人は同事業所を含まない過小な値である。"
     "居住系はほぼ満室であり、上振れを区域内で受ける余地は小さい。",
     "第6章第4節・第5節の整備方針", "R8.9"),
]:
    r = body(ws, r, [no, item, plan, why, blk, when], {3: OK_G}, height=64,
             align={1: "center", 6: "center"})

note(ws, r + 1,
     "注1）本表の受託者案は、ヒアリングシート項目6（見込量の置き方）で"
     "ご意向を確認する事項である。"
     "注2）採用値が決まった後、将来推計 第3段階（給付費・保険料）に進む。"
     "第3段階にはサービス単価、介護給付費準備基金の残高、保険料収納率及び"
     "所得段階別被保険者数を要する（資料提供依頼No.9・No.11・No.19、"
     "確認事項No.16・No.36）。", 6)

wb.save(OUT)
print("saved:", OUT)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
