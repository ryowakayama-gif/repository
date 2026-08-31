# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画 図表集（白黒レイアウト）生成スクリプト.

数値の出所は 00_凡例・出典 シートに明記。第9期計画（令和6年3月）及び
計画素案に掲載された見える化データのみを使用し、推測値は使用していない。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (RichTextProperties, Paragraph, ParagraphProperties,
                                   CharacterProperties)

FONT = "Carlito"
IN_Y = "FFF2CC"

# 系列の塗り分け：ベタ塗り（濃淡）とパターン（斜線・格子・ドット）を交互に配置し、
# 1色印刷でも隣接系列を判別できるようにする（第9期計画の体裁に準拠）。
# (種別, 前景色, 背景色)  種別が "solid" の場合は前景色でベタ塗り
PATTERNS = [
    ("solid",    "000000", None),      # 1 黒ベタ
    ("lgGrid",   "000000", "FFFFFF"),  # 2 格子
    ("solid",    "808080", None),      # 3 中間グレー
    ("dkUpDiag", "000000", "FFFFFF"),  # 4 太斜線
    ("solid",    "C8C8C8", None),      # 5 淡グレー
    ("pct25",    "000000", "FFFFFF"),  # 6 ドット
    ("solid",    "FFFFFF", None),      # 7 白
    ("dkHorz",   "000000", "FFFFFF"),  # 8 横縞
    ("ltUpDiag", "000000", "FFFFFF"),  # 9 細斜線
]
MARK = ["square", "triangle", "circle", "diamond", "x", "star"]   # ■ ▲ ● ◆ × 
DASH = ["solid", "dash", "sysDot", "dashDot", "lgDash", "sysDashDot"]


def _fill(i):
    """系列番号に応じたパターン塗り／ベタ塗りのGraphicalPropertiesを返す。"""
    kind, fg, bg = PATTERNS[i % len(PATTERNS)]
    gp = GraphicalProperties()
    if kind == "solid":
        gp.solidFill = fg
    else:
        gp.pattFill = PatternFillProperties(prst=kind,
                                            fgClr=ColorChoice(srgbClr=fg),
                                            bgClr=ColorChoice(srgbClr=bg))
    gp.line = LineProperties(solidFill="000000", w=6350)
    return gp


def _txpr(size=800, rot=None):
    body = RichTextProperties(rot=rot, vert="horz") if rot is not None else RichTextProperties()
    cp = CharacterProperties(sz=size, latin=None)
    return RichText(bodyPr=body,
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp, r=[])])


def _label_style(i):
    """系列の塗りに応じたラベル文字色と背景色を返す。

    黒ベタ・中間グレーは白抜き文字。パターン塗り（格子・斜線・ドット・横縞）は
    柄の上に白文字を置くと白地の部分で消えるため、白の下地を敷いて黒文字とする。
    """
    kind, fg, _bg = PATTERNS[i % len(PATTERNS)]
    if kind == "solid":
        return ("FFFFFF", None) if fg in ("000000", "808080") else ("000000", None)
    return "000000", "FFFFFF"


def _dlbl(pos, size=800, numfmt=None, color="000000", bgfill=None):
    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showPercent = False
    dl.dLblPos = pos
    if numfmt:
        dl.numFmt = numfmt
    cp = CharacterProperties(sz=size, solidFill=color)
    dl.txPr = RichText(bodyPr=RichTextProperties(),
                       p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp, r=[])])
    if bgfill:
        gp = GraphicalProperties(solidFill=bgfill)
        gp.line = LineProperties(noFill=True)
        dl.spPr = gp
    return dl


def _dlabels(ch, pos="ctr", size=800, numfmt=None, per_series=True, line=False):
    """データラベルを付す。per_series=True で系列ごとに文字色・背景・位置を最適化する。"""
    if line:
        # 折れ線は系列ごとに上下へ振り分け、白の下地を敷いて重なりを避ける
        for i, sr in enumerate(ch.series):
            sr.dLbls = _dlbl("t" if i % 2 == 0 else "b", size, numfmt, "000000", "FFFFFF")
        return
    if not per_series:
        ch.dLbls = _dlbl(pos, size, numfmt)
        return
    for i, sr in enumerate(ch.series):
        col, bg = _label_style(i) if pos == "ctr" else ("000000", None)
        sr.dLbls = _dlbl(pos, size, numfmt, col, bg)


thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
wb = Workbook()


def sheet(name, title, subtitle, widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="000000")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor="F2F2F2")
    n = max(len(widths), 8)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    return ws


def table(ws, row, head, rows, numfmt=None, headfill="404040"):
    for i, h in enumerate(head, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=headfill)
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = BORDER
    r = row + 1
    for data in rows:
        for i, v in enumerate(data, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=9)
            c.border = BORDER
            if i == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")
                if numfmt:
                    c.number_format = numfmt
        r += 1
    return r - 1


def note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    ws.row_dimensions[row].height = 26


def mono_bar(ws, title, y_title, cats_ref, data_ref, anchor, stacked=False,
             width=22, height=11, gap=60, overlap=None, from_rows=False,
             labels=False, numfmt=None, cat_rot=None):
    ch = BarChart()
    ch.type = "col"
    ch.style = None
    if stacked:
        ch.grouping = "stacked"
        ch.overlap = 100
    elif overlap is not None:
        ch.overlap = overlap
    ch.title = title
    ch.y_axis.title = y_title
    ch.add_data(data_ref, titles_from_data=True, from_rows=from_rows)
    ch.set_categories(cats_ref)
    ch.gapWidth = gap
    ch.width, ch.height = width, height
    for i, s in enumerate(ch.series):
        s.graphicalProperties = _fill(i)
    _axis_mono(ch, cat_rot=cat_rot)
    if labels:
        _dlabels(ch, "ctr" if stacked else "outEnd", 800, numfmt)
    if len(ch.series) <= 1:
        ch.legend = None
    ws.add_chart(ch, anchor)
    return ch


def mono_hbar(ws, title, cats_ref, data_ref, anchor, width=20, height=10):
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "percentStacked"
    ch.overlap = 100
    ch.title = title
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ch.gapWidth = 60
    ch.width, ch.height = width, height
    for i, s in enumerate(ch.series):
        s.graphicalProperties = _fill(i)
    _axis_mono(ch)
    _dlabels(ch, "ctr", 800, "0.0")
    ws.add_chart(ch, anchor)
    return ch


def mono_line(ws, title, y_title, cats_ref, data_ref, anchor,
              width=22, height=11, min_=None, max_=None, from_rows=False,
              labels=False, numfmt=None, cat_rot=None):
    ch = LineChart()
    ch.style = None
    ch.title = title
    ch.y_axis.title = y_title
    ch.add_data(data_ref, titles_from_data=True, from_rows=from_rows)
    ch.set_categories(cats_ref)
    ch.width, ch.height = width, height
    if min_ is not None:
        ch.y_axis.scaling.min = min_
    if max_ is not None:
        ch.y_axis.scaling.max = max_
    for i, s in enumerate(ch.series):
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill="000000", w=19050,
                                 prstDash=DASH[i % len(DASH)])
        s.graphicalProperties = gp
        s.marker = Marker(symbol=MARK[i % len(MARK)], size=7)
        mfill = "000000" if i % 2 == 0 else "FFFFFF"
        s.marker.graphicalProperties = GraphicalProperties(solidFill=mfill)
        s.marker.graphicalProperties.line = LineProperties(solidFill="000000", w=9525)
        s.smooth = False
    _axis_mono(ch, cat_rot=cat_rot)
    if labels:
        _dlabels(ch, "t", 750, numfmt, line=True)
    ws.add_chart(ch, anchor)
    return ch


def _axis_mono(ch, cat_rot=None, tick_size=800):
    for ax in (ch.x_axis, ch.y_axis):
        ax.majorGridlines = None
        ax.delete = False
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill="000000", w=6350)
        ax.graphicalProperties = gp
        ax.txPr = _txpr(tick_size)
        if ax.title is not None:
            ax.title.overlay = False
    if cat_rot is not None:
        ch.x_axis.txPr = _txpr(tick_size, rot=cat_rot)
    if ch.title is not None:
        ch.title.overlay = False
    if ch.legend is not None:
        ch.legend.position = "b"
        ch.legend.overlay = False
        ch.legend.txPr = _txpr(850)
    ch.plotVisOnly = True
    ch.dispBlanksAs = "gap"



def _named_series(src_ws, row_from, row_to):
    """08_ニーズ調査データ の1行から、系列名=C列（地域名）、値=E～J列（全体・年齢階級）の系列を作る。

    D列（サンプル数）を値に含めないため add_data ではなく Series を直接構成する。
    """
    ref = Reference(src_ws, min_col=5, max_col=10, min_row=row_from, max_row=row_to)
    s = Series(ref, title_from_data=False)
    s.tx = SeriesLabel(strRef=StrRef(f"'{src_ws.title}'!$C${row_from}"))
    return s


YEARS = ["平成24年", "平成25年", "平成26年", "平成27年", "平成28年", "平成29年",
         "平成30年", "令和元年", "令和2年", "令和3年", "令和4年", "令和5年"]

# ============================================================ 00 凡例・出典
ws = sheet("00_凡例・出典", "第10期介護保険事業計画 図表集（白黒レイアウト）",
           "令和8（2026）年7月28日作成／第9期計画の図表構成・体裁（パターン塗り、マーカー形状、データラベル）に準拠し、確認済みの数値のみで作図",
           [6, 26, 34, 44, 30])
r = 4
ws.cell(row=r, column=1, value="1　本図表集の考え方").font = Font(name=FONT, size=11, bold=True)
r += 1
note(ws, r, "第9期介護保険事業計画（令和6年3月）の図表構成をそのまま踏襲し、白黒印刷（本文1色）を前提としたグレースケール・線種・マーカーで作図しています。"
            "数値は第9期計画本体及び第10期計画素案に掲載された見える化データのみを使用し、推測値・補間値は使用していません。"
            "令和6・7年度の実績は未受領のため、該当セルは淡黄色の入力欄としています。")
r += 2
ws.cell(row=r, column=1, value="2　シート構成と出典").font = Font(name=FONT, size=11, bold=True)
r += 1
end = table(ws, r, ["No.", "シート", "図表", "出典・基準時点", "更新時の差替え"], [
    (1, "01_人口推移", "人口の推移（広域連合）", "第9期計画 第2章第1節1／住民基本台帳 各年10月1日", "R6～R8実績を追加"),
    (2, "02_高齢化率推移", "高齢化率の推移（広域連合・3町）", "第9期計画 第2章第1節1／住民基本台帳 各年10月1日", "R6～R8実績を追加"),
    (3, "03_高齢者世帯", "高齢者を含む世帯／構成割合", "第9期計画 第2章第1節2／国勢調査（H22・H27・R2）", "令和7年国勢調査の公表後に更新"),
    (4, "04_認定者数推移", "認定者数の推移（要介護度別）", "第9期計画 第2章第1節3／介護保険事業状況報告 各年9月分", "R6～R8の9月分を追加"),
    (5, "05_認定者割合", "認定者割合の比較（全国・道・3町）", "第9期計画 第2章第1節3／介護保険事業状況報告 令和5年9月分", "最新年9月分で更新"),
    (6, "06_出現率推移", "出現率の推移／上川管内比較", "第9期計画 第2章第1節3／介護保険事業状況報告 各年9月分", "R6～R8を追加。管内比較は最新年で更新"),
    (7, "07_給付費推移", "サービス区分別給付費／地域支援事業費／中長期推計", "第9期計画 第2章第2節（見える化 R6.1.18参照）／計画素案 表39", "R6～R8実績の受領後に更新"),
    (8, "08_ニーズ調査データ", "介護予防・日常生活圏域ニーズ調査 14指標の集計値", "第9期計画 第2章第3節／令和4年11月調査（回収4,626票・63.9％）", "第10期ニーズ調査の集計受領後に全面差替え"),
    (9, "09_ニーズ調査グラフ", "同上の年齢階級別・町別グラフ", "同上", "同上"),
    (14, "14_年齢構成と85歳以上", "高齢者の年齢構成（5歳階級別）・85歳以上人口・前期後期別の推移",
     "見える化A3・A4（国勢調査＋社人研推計、取得日 令和8年7月22日）", "見える化の再出力時に更新"),
    (15, "15_サービス利用強度", "受給者1人あたり利用日数・回数の推移（5サービス・全国／北海道比較）",
     "見える化D31-a〜j（H26〜R7、R7はR8年1月サービス提供分まで）", "見える化の再出力時に更新"),
    (16, "16_町別将来推計", "町別の総人口・高齢化率の推移と将来推計（2000〜2050年）",
     "見える化A1・A2（国勢調査＋社人研推計、取得日 令和8年7月）", "見える化の再出力時に更新"),
    (17, "17_担い手の推移", "高齢者1人あたり現役世代数（15〜64歳÷65歳以上）の町別推移",
     "見える化A9（同上）", "同上"),
    (18, "18_受給率・利用率", "介護サービス利用率、在宅・居住系サービス利用者割合、施設・居住系の受給率",
     "見える化D38・D41-a・D42（H26〜R7）", "見える化の再出力時に更新"),
    (19, "19_給付月額の比較", "第1号被保険者1人あたり・受給者1人あたりの給付月額（全国／北海道比較）",
     "見える化D6-a・D6-b・D17-a〜k（H18〜R7）", "同上"),
    (20, "20_第9期の達成状況", "第9期計画の代表KPI4項目・給付費・地域支援事業費の計画値と実績",
     "第9期計画 第1章第6節・第6章／見える化B4-a・D48-b・D48-c／令和7年度JAGES調査",
     "令和6〜8年度の事業実績の受領後に確定"),
    (10, "10_在宅介護実態調査", "世帯類型・介護頻度・介護内容・離職・必要な支援・施設検討（令和2年度との比較）",
     "第9期計画 第2章第4節1／令和5年5月25日～6月30日・認定調査員の聞き取り", "第10期調査の集計受領後に差替え"),
    (11, "11_居所変更実態調査", "居所変更・死亡の割合／要介護度別人数／変更理由",
     "第9期計画 第2章第4節2／令和5年5月25日送付・21施設回答", "第10期調査の集計受領後に差替え"),
    (12, "12_在宅生活改善調査", "利用者の属性／本人の状態・意向／家族等介護者の意向・負担",
     "第9期計画 第2章第4節3／令和5年5月25日送付・12事業所91人", "第10期調査の集計受領後に差替え"),
    (13, "13_介護人材実態調査", "職員数・男女比・就業形態・開設後経過年数・年齢構成・在職年数・"
     "直前の職場・資格・採用離職・勤務時間", "第9期計画 第2章第5節／令和5年5月25日送付・27施設405人",
     "第10期事業所実態調査の集計受領後に差替え"),
])
r = end + 2
ws.cell(row=r, column=1, value="3　作図上の凡例（1色印刷対応）").font = Font(name=FONT, size=11, bold=True)
r += 1
end = table(ws, r, ["No.", "要素", "表現", "備考", ""], [
    (1, "積上げ棒・横棒", "ベタ塗り（黒・中間グレー・淡グレー・白）とパターン（格子／太斜線／ドット／横縞／細斜線）を交互に配置",
     "第9期計画と同じ体裁。1色印刷でも隣接系列を判別できるよう、濃淡と柄の両方で区別している"),
    (2, "折れ線", "すべて黒線。実線／破線／点線／一点鎖線で系列を区別",
     "マーカーは■（1系列目）▲（2系列目）●（3系列目）◆（4系列目）×（5系列目）。塗りは黒と白を交互にして重なりを回避"),
    (3, "データラベル", "主要グラフに数値を表示。積上げは中央、単独棒は外側、折れ線は系列ごとに上下へ振り分け",
     "文字色は塗りに応じて切替え。黒ベタ・中間グレーの系列は白抜き文字、パターン塗り（格子・斜線・ドット・横縞）は"
     "柄の上で白文字が消えるため白の下地を敷いた黒文字、淡グレー・白の系列は黒文字。フォントは7.5〜8ポイント"),
    (4, "目盛線", "非表示。軸線のみ黒で表示", "1色印刷時の視認性を優先"),
    (5, "凡例", "グラフ下部に配置し、プロット領域と重ならないよう設定", "タイトル・軸タイトルも重なり回避を設定済み"),
    (6, "カテゴリ名が多い場合", "軸ラベルを45度回転して表示", "出現率の管内比較（24区分）等"),
    (7, "入力欄", "淡黄色。R6・R7実績など未受領の値", "受領後に上書きするとグラフに反映"),
])
r = end + 2
ws.cell(row=r, column=1, value="4　作図にあたり確認が必要な事項").font = Font(name=FONT, size=11, bold=True)
r += 1
end = table(ws, r, ["No.", "事項", "内容", "対応", ""], [
    (1, "人口の年齢区分", "【解決】第9期計画 第2章第1節1の掲載表から0～39歳・40～64歳および町別人口を復元した。"
        "全12年で4区分の合計＝総人口、65～74歳＋75歳以上＝65歳以上、3町の合計＝総人口がすべて一致",
        "01シートを第9期計画と同じ年齢4区分で作図。町別総人口と町別0～39歳・40～64歳（平成24年・令和5年）も掲載",
        "対応済み。中間年の町別年齢別は原データ受領後に補う"),
    (2, "リスク点数の単位", "第9期計画 第2章第3節⑨は「当広域連合全体で15.8％」と記載しているが、"
        "実際は要支援・要介護リスク評価尺度による平均点（72,099点÷4,560人＝15.8点）", "本図表集では「点」として作図。第10期では単位表記を修正", "第10期計画で修正"),
    (3, "認定者数と出現率の基準月", "第9期計画は各年9月分、計画素案の直近値（認定者1,984人・21.8％）はR8年3月末",
        "基準月が異なるため同一系列に接続していない。R6～R8は9月分を受領して追加", "広域連合へ照会"),
    (4, "総給付費の定義", "第9期計画 第2章第2節の総給付費（R5 2,794,391千円）と、"
        "計画素案の保険給付費R5決算（2,940.4百万円）は定義が異なる", "別系列として併記。計画掲載時はいずれかに統一", "第10期計画で整理"),
    (11, "人口統計の基準", "第9期計画の図1・図2は住民基本台帳（各年10月1日）、14シートと素案第6章の推計は国勢調査＋社人研推計。"
        "令和5年で総人口が382人、高齢化率が1.1ポイント異なる", "01・02シートは住民基本台帳、14シートは国勢調査・社人研で作図し接続していない。"
        "第10期で採用基準を統一する", "広域連合と協議（修正指示書C-7）"),
    (12, "高齢夫婦世帯数", "見える化A8（H27 1,615・R2 1,773世帯）と第9期計画（同1,697・1,950世帯）が一致しない。"
        "一般世帯・高齢者を含む世帯・高齢独居世帯の3系列は完全に一致", "03シートは第9期計画の値を掲載。定義差の確認後に差替え",
        "広域連合へ照会（修正指示書C-6）"),
    (13, "町別データ", "受領した見える化データは広域連合・北海道・全国の3系列のみで町別の内訳がない",
        "02・03・06シートの町別値は第9期計画の掲載値のまま。町別データの受領後に更新", "広域連合へ照会（修正指示書C-8）"),
    (5, "ニーズ調査の基準", "掲載値は令和4年11月実施の第9期調査。第10期調査は実施済だが集計未受領",
        "第10期調査の集計受領後に全面差替え。設問・判定ロジックの一致を確認", "広域連合・3町へ照会"),
    (6, "在宅介護実態調査の回答数", "第9期計画に回収結果の表が掲載されているが本文から数値を復元できなかった。"
        "令和5年度の掲載値はすべて1/37の倍数で整合するため回答37人と判断（対象は65歳以上の要介護認定者79人）",
        "10シートに推定値として記載。回収票数・回収率は原資料で確認する", "広域連合へ照会"),
    (7, "居所変更実態調査の施設数", "回答施設の内訳（住宅型有料4・軽費1・サ高住0・GH5・特定施設2・老健3・特養2・地密特養3）の合計は20だが、"
        "第9期計画は「合計21」と記載", "11シートに原典どおり記載し差異を注記。内訳の欠落か合計の誤りかを確認", "広域連合へ照会"),
    (8, "介護人材実態調査の就業形態", "第9期計画 第2章第5節(1)④の本文は「施設・居住系サービスの正規職員が33.8％」とするが、"
        "グラフ値では正規職員（男）31.9＋正規職員（女）45.1＝77.0％であり、33.8％は男性職員の割合（31.9＋1.9）と一致する",
        "13シートはグラフ値どおり作図し、本文との差異を注記。第10期では正規職員割合と男性割合を区別して記載", "第10期計画で修正"),
    (9, "介護人材実態調査の職員数の分母", "職員総数421人、男女内訳の合計403人、該当者数405人と3種類の数値があり、"
        "1事業所あたり平均人数も405÷27＝15.0人で本文の16.0人と一致しない",
        "13シートは原典の3種類をそのまま併記し、比率算定の分母を明示。第10期では集計基準を統一", "第10期計画で整理"),
    (10, "直前の職場の一部区分", "第9期計画 第2章第5節(3)④のうち「訪問介護・入浴、夜間対応型」「小多機、看多機、定期巡回」の2区分は"
        "原典の図から数値を復元できなかった（他6区分の合計91.1％との差8.9％＝5人分）",
        "13シートで淡黄色の入力欄とし、原資料で確認のうえ入力する", "広域連合へ照会"),
])

# ============================================================ 01 人口推移
ws = sheet("01_人口推移", "図1　人口の推移（大雪地区広域連合・構成3町）",
           "資料：住民基本台帳 各年10月1日現在（第9期計画 第2章第1節1）／単位：人／令和6～8年は淡黄色欄に入力すると連動",
           [16] + [10] * 15)
POP = {
    "0～39歳": [10590, 10483, 10450, 10458, 10324, 10309, 10144, 9924, 9572, 9525, 9421, 9248],
    "40～64歳": [9903, 9832, 9740, 9706, 9629, 9589, 9540, 9479, 9433, 9450, 9443, 9448],
    "65～74歳": [3744, 3862, 4032, 4119, 4183, 4166, 4210, 4215, 4269, 4233, 4165, 3986],
    "75歳以上": [4398, 4491, 4578, 4667, 4768, 4865, 4933, 5006, 5014, 5039, 5109, 5205],
}
POP_TOTAL = [28635, 28668, 28800, 28950, 28904, 28929, 28827, 28624, 28288, 28247, 28138, 27887]
POP_65 = [8142, 8353, 8610, 8786, 8951, 9031, 9143, 9221, 9283, 9272, 9274, 9191]
TOWN_POP = {
    "東川町": [7952, 7944, 7946, 8115, 8130, 8312, 8406, 8373, 8295, 8428, 8542, 8558],
    "美瑛町": [10832, 10726, 10654, 10492, 10375, 10233, 10092, 9960, 9821, 9678, 9617, 9475],
    "東神楽町": [9851, 9998, 10200, 10343, 10399, 10384, 10329, 10291, 10172, 10141, 9979, 9854],
}
TOWN_0_39 = {"東川町": [2978, None, None, None, None, None, None, None, None, None, None, 3110],
             "美瑛町": [3514, None, None, None, None, None, None, None, None, None, None, 2674],
             "東神楽町": [4098, None, None, None, None, None, None, None, None, None, None, 3464]}
TOWN_40_64 = {"東川町": [2689, None, None, None, None, None, None, None, None, None, None, 2765],
              "美瑛町": [3681, None, None, None, None, None, None, None, None, None, None, 3126],
              "東神楽町": [3533, None, None, None, None, None, None, None, None, None, None, 3557]}

HD = ["区分"] + YEARS + ["令和6年", "令和7年", "令和8年"]
r = 4
ws.cell(row=r, column=1, value="（1）年齢4区分別人口（大雪地区広域連合）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
rows = [[k] + v + [None] * 3 for k, v in POP.items()] \
    + [["65歳以上（再掲）"] + POP_65 + [None] * 3, ["総人口"] + [None] * 15]
end = table(ws, hrow, HD, rows, numfmt="#,##0")
TOT_ROW = end
for c in range(2, 17):
    col = get_column_letter(c)
    ws.cell(row=TOT_ROW - 1, column=c).value = f"={col}{hrow+3}+{col}{hrow+4}"
    ws.cell(row=TOT_ROW, column=c).value = f"=SUM({col}{hrow+1}:{col}{hrow+4})"
    for rr in (TOT_ROW - 1, TOT_ROW):
        ws.cell(row=rr, column=c).number_format = "#,##0"
    ws.cell(row=TOT_ROW, column=c).font = Font(name=FONT, size=9, bold=True)
for rr in range(hrow + 1, hrow + 5):
    for c in range(14, 17):
        ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=IN_Y)
cats = Reference(ws, min_col=2, max_col=16, min_row=hrow)
data = Reference(ws, min_col=1, max_col=16, min_row=hrow + 1, max_row=hrow + 4)
mono_bar(ws, "人口の推移（年齢4区分・大雪地区広域連合）", "人口（人）", cats, data, "R4",
         stacked=True, width=28, height=13, from_rows=True, labels=True, numfmt="#,##0")
r = end + 2

ws.cell(row=r, column=1, value="（2）町別総人口").font = Font(name=FONT, size=10, bold=True)
hrow2 = r + 1
end2 = table(ws, hrow2, HD, [[k] + v + [None] * 3 for k, v in TOWN_POP.items()]
             + [["広域連合 計"] + [None] * 15], numfmt="#,##0")
for c in range(2, 17):
    col = get_column_letter(c)
    ws.cell(row=end2, column=c).value = f"=SUM({col}{hrow2+1}:{col}{end2-1})"
    ws.cell(row=end2, column=c).number_format = "#,##0"
    ws.cell(row=end2, column=c).font = Font(name=FONT, size=9, bold=True)
for rr in range(hrow2 + 1, end2):
    for c in range(14, 17):
        ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=IN_Y)
cats2 = Reference(ws, min_col=2, max_col=16, min_row=hrow2)
data2 = Reference(ws, min_col=1, max_col=16, min_row=hrow2 + 1, max_row=end2 - 1)
mono_line(ws, "町別総人口の推移", "人口（人）", cats2, data2, "R31", width=28, height=12,
          from_rows=True, labels=True, numfmt="#,##0")
r = end2 + 2

ws.cell(row=r, column=1, value="（3）町別 0～39歳・40～64歳人口（平成24年・令和5年）").font = Font(name=FONT, size=10, bold=True)
hrow3 = r + 1
end3 = table(ws, hrow3, ["区分", "平成24年 0～39歳", "令和5年 0～39歳", "平成24年 40～64歳", "令和5年 40～64歳"],
             [[t, TOWN_0_39[t][0], TOWN_0_39[t][11], TOWN_40_64[t][0], TOWN_40_64[t][11]]
              for t in ["東川町", "美瑛町", "東神楽町"]] + [["広域連合 計", None, None, None, None]],
             numfmt="#,##0")
for c in range(2, 6):
    col = get_column_letter(c)
    ws.cell(row=end3, column=c).value = f"=SUM({col}{hrow3+1}:{col}{end3-1})"
    ws.cell(row=end3, column=c).number_format = "#,##0"
    ws.cell(row=end3, column=c).font = Font(name=FONT, size=9, bold=True)
note(ws, end3 + 2,
     "注1）年齢4区分と町別人口は第9期計画 第2章第1節1の掲載表から復元した。全12年について"
     "「0～39歳＋40～64歳＋65～74歳＋75歳以上＝総人口」「65～74歳＋75歳以上＝65歳以上」"
     "「東川町＋美瑛町＋東神楽町＝総人口」がすべて一致することを確認している。"
     "注2）町別の高齢化率も逆算で一致する（令和5年：東川町2,683÷8,558＝31.4％、美瑛町3,675÷9,475＝38.8％、"
     "東神楽町2,833÷9,854＝28.7％）。"
     "注3）（3）は計画に平成24年と令和5年のみ掲載されているため2時点としている。中間年は原データの受領後に補う。"
     "注4）令和6～8年（淡黄色欄）は住民基本台帳の実績受領後に入力する。")


# ============================================================ 02 高齢化率
ws = sheet("02_高齢化率推移", "図2　高齢化率の推移（広域連合・構成3町）",
           "資料：住民基本台帳 各年10月1日現在（第9期計画 第2章第1節1）／単位：％",
           [16] + [10] * 15)
RATE = {
    "広域連合": [28.4, 29.1, 29.9, 30.3, 31.0, 31.2, 31.7, 32.2, 32.8, 32.8, 33.0, 33.0],
    "東川町": [28.7, 29.7, 31.1, 31.5, 32.1, 32.0, 32.2, 32.1, 32.8, 32.3, 31.8, 31.4],
    "美瑛町": [33.6, 34.4, 35.2, 35.9, 36.5, 36.7, 37.3, 37.9, 38.4, 38.6, 38.7, 38.8],
    "東神楽町": [22.5, 23.0, 23.4, 23.8, 24.5, 25.2, 25.8, 26.7, 27.4, 27.7, 28.4, 28.7],
}
end = table(ws, 4, HD, [[k] + v + [None] * 3 for k, v in RATE.items()], numfmt="0.0")
for r_ in range(5, 9):
    for c in range(14, 17):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)
note(ws, end + 1, "注）東川町の高齢化率は令和2年の32.8％をピークに緩やかな下降傾向、美瑛町・東神楽町は上昇傾向で推移している"
                  "（第9期計画 第2章第1節1）。令和6～8年（淡黄色欄）は実績受領後に入力する。")
cats = Reference(ws, min_col=2, max_col=16, min_row=4)
data = Reference(ws, min_col=1, max_col=16, min_row=5, max_row=8)
mono_line(ws, "高齢化率の推移", "高齢化率（％）", cats, data, "A13", width=28, height=13,
          min_=20, max_=42, from_rows=True, labels=True, numfmt="0.0")

# ============================================================ 03 世帯
ws = sheet("03_高齢者世帯", "図3　高齢者を含む世帯の状況",
           "資料：国勢調査（平成22年・平成27年・令和2年）（第9期計画 第2章第1節2）／単位：世帯・％",
           [22, 12, 12, 12, 12, 12, 12, 12, 12, 12])
r = 4
ws.cell(row=r, column=1, value="（1）世帯数").font = Font(name=FONT, size=10, bold=True)
r += 1
SETAI = [
    ["一般世帯　広域連合", 10536, 11050, 11426], ["　東川町", 2965, 3132, 3391],
    ["　美瑛町", 4289, 4274, 4205], ["　東神楽町", 3282, 3644, 3830],
    ["高齢者を含む世帯　広域連合", 4885, 5318, 5622], ["　東川町", 1316, 1506, 1620],
    ["　美瑛町", 2330, 2341, 2353], ["　東神楽町", 1239, 1471, 1649],
    ["高齢者夫婦世帯　広域連合", 1461, 1697, 1950], ["　東川町", 387, 504, 596],
    ["　美瑛町", 667, 739, 743], ["　東神楽町", 407, 454, 611],
    ["高齢者単身世帯　広域連合", 1076, 1331, 1638], ["　東川町", 295, 363, 490],
    ["　美瑛町", 531, 653, 714], ["　東神楽町", 250, 315, 434],
]
end = table(ws, r, ["区分", "平成22年", "平成27年", "令和2年"], SETAI, numfmt="#,##0")
r = end + 2
ws.cell(row=r, column=1, value="（2）高齢者を含む世帯の構成割合").font = Font(name=FONT, size=10, bold=True)
r += 1
KOSEI = [
    ["広域連合（平成22年）", 29.9, 22.0, 48.1], ["広域連合（平成27年）", 31.9, 25.0, 43.1],
    ["広域連合（令和2年）", 34.7, 29.1, 36.2], ["東川町（令和2年）", 36.8, 30.2, 33.0],
    ["美瑛町（令和2年）", 31.6, 30.3, 38.1], ["東神楽町（令和2年）", 37.1, 26.3, 36.6],
]
end2 = table(ws, r, ["区分", "高齢者夫婦世帯", "高齢者単身世帯", "その他の高齢者を含む世帯"], KOSEI, numfmt="0.0")
note(ws, end2 + 1, "注）構成割合は「高齢者を含む世帯」に占める割合。令和2年では高齢者単身世帯の割合は美瑛町が30.3％と最も高く、"
                   "高齢者夫婦世帯の割合は東神楽町が37.1％と最も高い（第9期計画 第2章第1節2）。次回は令和7年国勢調査の公表後に更新する。")
cats = Reference(ws, min_col=1, min_row=r + 1, max_row=r + 6)
data = Reference(ws, min_col=2, max_col=4, min_row=r, max_row=r + 6)
mono_hbar(ws, "高齢者を含む世帯の構成割合", cats, data, f"F{r}", width=22, height=12)
SETAI_HEAD = 5           # 世帯数テーブルの見出し行
FUFU_ROW = SETAI_HEAD + 9    # 高齢者夫婦世帯 広域連合
TANSHIN_ROW = SETAI_HEAD + 13  # 高齢者単身世帯 広域連合
assert ws.cell(row=FUFU_ROW, column=1).value.startswith("高齢者夫婦世帯")
assert ws.cell(row=TANSHIN_ROW, column=1).value.startswith("高齢者単身世帯")
# --- 町別世帯（見える化A5〜A8）と高齢夫婦世帯の定義差 ---
_r = end2 + 3
ws.cell(row=_r, column=1, value="（3）町別世帯数（見える化A5〜A8）と高齢夫婦世帯の定義差").font = Font(name=FONT, size=10, bold=True)
_h = _r + 1
_MIE = [
    ["一般世帯　東川町", 2965, 3132, 3391], ["　美瑛町", 4289, 4274, 4205], ["　東神楽町", 3282, 3644, 3830],
    ["高齢者を含む世帯　東川町", 1316, 1506, 1620], ["　美瑛町", 2330, 2341, 2353], ["　東神楽町", 1239, 1471, 1649],
    ["高齢独居世帯　東川町", 295, 363, 490], ["　美瑛町", 531, 653, 714], ["　東神楽町", 250, 315, 434],
    ["高齢夫婦世帯（見える化）　東川町", 387, 504, 554], ["　美瑛町", 667, 657, 676], ["　東神楽町", 407, 454, 543],
    ["高齢者夫婦世帯（第9期計画）　東川町", 387, 504, 596], ["　美瑛町", 667, 739, 743], ["　東神楽町", 407, 454, 611],
]
_e = table(ws, _h, ["区分", "平成22年", "平成27年", "令和2年"], _MIE, numfmt="#,##0")
note(ws, _e + 1,
     "注）一般世帯・高齢者を含む世帯・高齢独居世帯は、見える化A5〜A7と第9期計画の掲載値が町別まで完全に一致する。"
     "一方、高齢夫婦世帯のみ平成27年で82世帯、令和2年で177世帯の差があり、差は時点とともに拡大する。"
     "見える化の「高齢夫婦世帯」と第9期計画の「高齢者夫婦世帯」の定義差（夫婦双方が65歳以上か、いずれか一方が65歳以上か）"
     "によるものとみられる。第10期で採用する定義を決定する必要がある（修正指示書C-6）。")
_c = Reference(ws, min_col=1, min_row=_h + 10, max_row=_h + 15)
_d = Reference(ws, min_col=2, max_col=4, min_row=_h, max_row=_h + 15)
mono_bar(ws, "高齢夫婦世帯数の定義差（町別）", "世帯数（世帯）",
         Reference(ws, min_col=2, max_col=4, min_row=_h),
         Reference(ws, min_col=1, max_col=4, min_row=_h + 10, max_row=_h + 15),
         f"Q{_r}", width=20, height=11, from_rows=True, labels=True, numfmt="#,##0", gap=80)

cats2 = Reference(ws, min_col=2, max_col=4, min_row=SETAI_HEAD)
data2 = Reference(ws, min_col=1, max_col=4, min_row=FUFU_ROW, max_row=FUFU_ROW)
d3 = Reference(ws, min_col=1, max_col=4, min_row=TANSHIN_ROW, max_row=TANSHIN_ROW)
ch = BarChart()
ch.type = "col"
ch.title = "高齢者夫婦世帯・高齢者単身世帯の推移（広域連合）"
ch.y_axis.title = "世帯数（世帯）"
ch.add_data(data2, titles_from_data=True, from_rows=True)
ch.add_data(d3, titles_from_data=True, from_rows=True)
ch.set_categories(cats2)
ch.gapWidth = 80
ch.width, ch.height = 14, 9
for i, s in enumerate(ch.series):
    s.graphicalProperties = _fill(i * 2)
_axis_mono(ch)
_dlabels(ch, "outEnd", 800, "#,##0")
ws.add_chart(ch, f"F{r + 24}")

# ============================================================ 04 認定者数
ws = sheet("04_認定者数推移", "図4　要介護（要支援）認定者数の推移（大雪地区広域連合）",
           "資料：介護保険事業状況報告 各年9月分（第9期計画 第2章第1節3）／単位：人",
           [16] + [10] * 15)
NINTEI = {
    "要支援1": [169, 221, 264, 325, 297, 274, 289, 275, 271, 254, 239, 248],
    "要支援2": [243, 270, 282, 288, 263, 268, 256, 295, 302, 300, 313, 285],
    "要介護1": [336, 354, 396, 410, 415, 420, 447, 438, 424, 482, 470, 492],
    "要介護2": [283, 274, 301, 315, 306, 323, 320, 315, 323, 320, 350, 329],
    "要介護3": [222, 223, 196, 198, 220, 250, 228, 235, 237, 224, 229, 247],
    "要介護4": [175, 191, 190, 198, 211, 195, 182, 209, 210, 232, 213, 222],
    "要介護5": [181, 198, 188, 170, 177, 170, 179, 161, 154, 158, 158, 153],
}
rows = [[k] + v + [None] * 3 for k, v in NINTEI.items()] + [["合計"] + [None] * 15]
end = table(ws, 4, HD, rows, numfmt="#,##0")
TOT_ROW = end
for c in range(2, 17):
    col = get_column_letter(c)
    ws.cell(row=TOT_ROW, column=c).value = f"=SUM({col}5:{col}11)"
    ws.cell(row=TOT_ROW, column=c).number_format = "#,##0"
    ws.cell(row=TOT_ROW, column=c).font = Font(name=FONT, size=9, bold=True)
for r_ in range(5, 12):
    for c in range(14, 17):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)
note(ws, end + 1, "注）第1号・第2号被保険者の合計。令和5年の合計1,976人は第9期計画の記載と一致する。"
                  "要支援1～要介護4は増加傾向、要介護5は減少傾向で推移している（第9期計画 第2章第1節3）。"
                  "令和6～8年（淡黄色欄）は各年9月分の実績受領後に入力する。")
cats = Reference(ws, min_col=2, max_col=16, min_row=4)
data = Reference(ws, min_col=1, max_col=16, min_row=5, max_row=11)
mono_bar(ws, "認定者数の推移（要介護度別）", "認定者数（人）", cats, data, "A16", stacked=True,
         width=28, height=13, from_rows=True, labels=True, numfmt="#,##0")

# ============================================================ 05 認定者割合
ws = sheet("05_認定者割合", "図5　認定者割合の比較（令和5年9月）",
           "資料：介護保険事業状況報告 令和5年9月分（第9期計画 第2章第1節3）／単位：％",
           [16, 11, 11, 11, 11, 11, 11, 11])
WARIAI = [
    ["全国", 14.2, 13.9, 20.7, 16.7, 13.2, 12.7, 8.5],
    ["北海道", 18.4, 14.6, 23.1, 15.7, 10.6, 10.4, 7.2],
    ["広域連合", 12.6, 14.4, 24.9, 16.6, 12.5, 11.2, 7.7],
    ["東川町", 9.9, 14.8, 23.4, 18.1, 11.7, 13.1, 9.1],
    ["美瑛町", 14.2, 14.7, 23.6, 16.3, 14.3, 10.4, 6.4],
    ["東神楽町", 12.5, 13.6, 28.7, 15.7, 10.2, 10.7, 8.6],
]
end = table(ws, 4, ["区分", "要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"],
            WARIAI, numfmt="0.0")
note(ws, end + 1, "注）東川町は「要支援1」の割合が9.9％と全国・北海道・他町より低く、美瑛町は「要支援1」14.2％と「要介護3」14.3％が他町より高い。"
                  "東神楽町は「要介護1」の割合が28.7％と全国・北海道・他町より高い（第9期計画 第2章第1節3）。")
cats = Reference(ws, min_col=1, min_row=5, max_row=10)
data = Reference(ws, min_col=2, max_col=8, min_row=4, max_row=10)
mono_hbar(ws, "認定者割合の比較（令和5年9月）", cats, data, "A14", width=24, height=12)

# ============================================================ 06 出現率
ws = sheet("06_出現率推移", "図6　要支援・要介護認定者の出現率",
           "資料：介護保険事業状況報告 各年9月分（第9期計画 第2章第1節3）／出現率＝第1号認定者数÷第1号被保険者数・単位：％",
           [18] + [10] * 15)
r = 4
ws.cell(row=r, column=1, value="（1）出現率の推移（全国・北海道との比較）").font = Font(name=FONT, size=10, bold=True)
r += 1
SHUTSU = {
    "広域連合": [19.3, 20.2, 20.6, 21.2, 20.7, 20.7, 20.5, 20.7, 20.5, 21.0, 21.1, 21.2],
    "北海道": [18.5, 18.9, 19.2, 19.4, 19.5, 19.5, 19.7, 20.0, 20.2, 20.4, 20.6, 20.8],
    "全国": [17.5, 17.8, 17.9, 18.0, 18.0, 18.1, 18.3, 18.5, 18.6, 18.8, 19.1, 19.3],
}
end = table(ws, r, HD, [[k] + v + [None] * 3 for k, v in SHUTSU.items()], numfmt="0.0")
for r_ in range(r + 1, r + 4):
    for c in range(14, 17):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)
cats = Reference(ws, min_col=2, max_col=16, min_row=r)
data = Reference(ws, min_col=1, max_col=16, min_row=r + 1, max_row=r + 3)
mono_line(ws, "出現率の推移", "出現率（％）", cats, data, "R5", width=28, height=12,
          min_=15, max_=23, from_rows=True, labels=True, numfmt="0.0")

r = end + 2
ws.cell(row=r, column=1, value="（2）町別出現率の推移").font = Font(name=FONT, size=10, bold=True)
r += 1
SHUTSU_T = {
    "東川町": [18.7, 19.8, 19.2, 19.1, 17.5, 17.8, 18.4, 18.6, 19.4, 19.6, 19.9, 20.5],
    "美瑛町": [19.3, 20.6, 21.8, 23.0, 22.9, 23.3, 22.8, 23.2, 22.7, 23.4, 23.5, 24.0],
    "東神楽町": [19.8, 19.9, 20.3, 20.8, 20.7, 19.9, 19.3, 19.3, 18.5, 19.1, 18.9, 18.2],
}
end2 = table(ws, r, HD, [[k] + v + [None] * 3 for k, v in SHUTSU_T.items()], numfmt="0.0")
for r_ in range(r + 1, r + 4):
    for c in range(14, 17):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)
cats = Reference(ws, min_col=2, max_col=16, min_row=r)
data = Reference(ws, min_col=1, max_col=16, min_row=r + 1, max_row=r + 3)
mono_line(ws, "町別出現率の推移", "出現率（％）", cats, data, "R28", width=28, height=12,
          min_=15, max_=26, from_rows=True, labels=True, numfmt="0.0")

r = end2 + 26
ws.cell(row=r, column=1, value="（3）出現率の比較（令和5年 上川総合振興局管内）").font = Font(name=FONT, size=10, bold=True)
r += 1
KANNAI = [
    ("全国", 19.3), ("北海道", 20.8), ("上川総合振興局管内", 21.2), ("当麻町", 23.8), ("旭川市", 21.8),
    ("中川町", 21.5), ("比布町", 21.4), ("大雪地区広域連合", 21.2), ("和寒町", 21.2), ("上川町", 20.6),
    ("名寄市", 20.5), ("幌加内町", 20.2), ("愛別町", 20.2), ("富良野市", 19.8), ("中富良野町", 19.4),
    ("南富良野町", 19.3), ("美深町", 19.2), ("鷹栖町", 18.7), ("士別市", 18.7), ("剣淵町", 17.7),
    ("下川町", 17.3), ("占冠村", 17.2), ("上富良野町", 16.1), ("音威子府村", 10.8),
]
end3 = table(ws, r, ["区分", "出現率"], [list(x) for x in KANNAI], numfmt="0.0")
note(ws, end3 + 1, "注）当広域連合の出現率は全国・北海道を上回って推移し、平成25年以降は20％を超えている。"
                   "上川総合振興局管内の保険者では高い方から5番目（第9期計画 第2章第1節3）。")
cats = Reference(ws, min_col=1, min_row=r + 1, max_row=end3)
data = Reference(ws, min_col=2, min_row=r, max_row=end3)
ch = BarChart()
ch.type = "col"
ch.title = "出現率の比較（令和5年 上川総合振興局管内）"
ch.y_axis.title = "出現率（％）"
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.gapWidth = 40
ch.width, ch.height = 26, 11
ch.series[0].graphicalProperties = _fill(4)
_axis_mono(ch, cat_rot=-2700000, tick_size=750)
_dlabels(ch, "outEnd", 750, "0.0", per_series=False)
ch.legend = None
ws.add_chart(ch, "R51")

# ============================================================ 07 給付費
ws = sheet("07_給付費推移", "図7　介護給付費等の推移",
           "資料：第9期計画 第2章第2節（見える化システム 令和6年1月18日参照）／計画素案 表39（自然体推計）",
           [26, 14, 14, 14, 14, 14, 14, 14, 14])
r = 4
ws.cell(row=r, column=1, value="（1）サービス区分別給付費（実績・単位：千円）").font = Font(name=FONT, size=10, bold=True)
r += 1
KYUFU = [
    ["居宅サービス", 836496, 872021, 896194, None, None, None],
    ["地域密着型サービス", 765808, 718778, 743018, None, None, None],
    ["施設サービス", 973845, 924985, 941503, None, None, None],
    ["居宅介護支援", 111147, 115959, 115772, None, None, None],
    ["介護予防サービス", 68592, 64733, 66451, None, None, None],
    ["地域密着型介護予防サービス", 25160, 22391, 19802, None, None, None],
    ["介護予防支援", 12766, 12144, 11653, None, None, None],
    ["合計", None, None, None, None, None, None],
]
end = table(ws, r, ["区分", "令和3年度", "令和4年度", "令和5年度", "令和6年度", "令和7年度", "令和8年度"],
            KYUFU, numfmt="#,##0")
for c in range(2, 8):
    col = get_column_letter(c)
    ws.cell(row=end, column=c).value = f"=SUM({col}{r+1}:{col}{end-1})"
    ws.cell(row=end, column=c).number_format = "#,##0"
    ws.cell(row=end, column=c).font = Font(name=FONT, size=9, bold=True)
for r_ in range(r + 1, end):
    for c in range(5, 8):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)
cats = Reference(ws, min_col=2, max_col=7, min_row=r)
data = Reference(ws, min_col=1, max_col=7, min_row=r + 1, max_row=end - 1)
mono_bar(ws, "サービス区分別給付費の推移", "給付費（千円）", cats, data, "J4", stacked=True, width=20, height=11, from_rows=True)

r2 = end + 2
ws.cell(row=r2, column=1, value="（2）総給付費（在宅・居住系・施設別／実績・単位：千円）").font = Font(name=FONT, size=10, bold=True)
r2 += 1
SOU = [
    ["在宅サービス", 1191892, 1218034, 1250603, None, None, None],
    ["居住系サービス", 436392, 399288, 414922, None, None, None],
    ["施設サービス", 1165530, 1113689, 1128866, None, None, None],
    ["総給付費", None, None, None, None, None, None],
]
end2 = table(ws, r2, ["区分", "令和3年度", "令和4年度", "令和5年度", "令和6年度", "令和7年度", "令和8年度"],
             SOU, numfmt="#,##0")
for c in range(2, 8):
    col = get_column_letter(c)
    ws.cell(row=end2, column=c).value = f"=SUM({col}{r2+1}:{col}{end2-1})"
    ws.cell(row=end2, column=c).number_format = "#,##0"
    ws.cell(row=end2, column=c).font = Font(name=FONT, size=9, bold=True)
for r_ in range(r2 + 1, end2):
    for c in range(5, 8):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)

r3 = end2 + 2
ws.cell(row=r3, column=1, value="（3）地域支援事業費（実績・単位：円）").font = Font(name=FONT, size=10, bold=True)
r3 += 1
CHIIKI = [
    ["介護予防・日常生活支援総合事業費", 98828508, 100020350, 106455000, None, None, None],
    ["包括的支援事業（センター運営）及び任意事業費", 67884572, 71189693, 65532000, None, None, None],
    ["包括的支援事業（社会保障充実分）", 12746280, 15365016, 18387000, None, None, None],
    ["地域支援事業費 合計", None, None, None, None, None, None],
]
end3 = table(ws, r3, ["区分", "令和3年度", "令和4年度", "令和5年度", "令和6年度", "令和7年度", "令和8年度"],
             CHIIKI, numfmt="#,##0")
for c in range(2, 8):
    col = get_column_letter(c)
    ws.cell(row=end3, column=c).value = f"=SUM({col}{r3+1}:{col}{end3-1})"
    ws.cell(row=end3, column=c).number_format = "#,##0"
    ws.cell(row=end3, column=c).font = Font(name=FONT, size=9, bold=True)
for r_ in range(r3 + 1, end3):
    for c in range(5, 8):
        ws.cell(row=r_, column=c).fill = PatternFill("solid", fgColor=IN_Y)

r4 = end3 + 2
ws.cell(row=r4, column=1, value="（4）給付費の中長期見通し（自然体推計・参考／単位：百万円）").font = Font(name=FONT, size=10, bold=True)
r4 += 1
SHIZEN = [
    ["居宅サービス", 1537.4, 1559.6, 1589.9, 1652.4, 1738.6],
    ["居住系サービス", 380.6, 391.9, 391.9, 423.7, 441.5],
    ["施設サービス", 1138.7, 1138.7, 1138.7, 1265.5, 1312.4],
    ["合計", None, None, None, None, None],
]
end4 = table(ws, r4, ["区分", "令和9年度", "令和10年度", "令和11年度", "令和17年度", "令和22年度"],
             SHIZEN, numfmt="#,##0.0")
for c in range(2, 7):
    col = get_column_letter(c)
    ws.cell(row=end4, column=c).value = f"=SUM({col}{r4+1}:{col}{end4-1})"
    ws.cell(row=end4, column=c).number_format = "#,##0.0"
    ws.cell(row=end4, column=c).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=2, max_col=6, min_row=r4)
data = Reference(ws, min_col=1, max_col=6, min_row=r4 + 1, max_row=end4 - 1)
mono_bar(ws, "給付費の中長期見通し（自然体推計）", "給付費（百万円）", cats, data, "J27", stacked=True, width=20, height=11, from_rows=True)

note(ws, end4 + 2,
     "注1）（1）～（3）は第9期計画 第2章第2節の掲載値（見える化システム 令和6年1月18日参照）。令和6～8年度（淡黄色欄）は実績受領後に入力する。"
     "注2）（2）の総給付費は（1）の介護給付費と介護予防給付費の合計に相当し、令和5年度は2,794,391千円（在宅・居住系・施設の3区分で再集計した値）。"
     "計画素案に記載の保険給付費 令和5年度決算2,940.4百万円とは定義が異なるため、同一系列に接続していない（00_凡例・出典 4-4）。"
     "注3）（1）の合計（令和5年度2,794,393千円）と（2）の総給付費（同2,794,391千円）には2千円の差がある。"
     "これは第9期計画の介護予防給付費の合計欄（令和4年度99,267千円・令和5年度97,905千円）に内訳との±1千円の端数差があるためで、"
     "本図表集では内訳を原典どおり掲載し合計を数式で算出している。計画掲載時はいずれかに統一する。"
     "注4）（4）は計画素案 表39の自然体推計値で、報酬改定・供給制約・政策効果は未反映の参考値。採用値ではない。")

# ============================================================ 08 ニーズ調査データ
ws = sheet("08_ニーズ調査データ", "図8　健康とくらしの調査（JAGES調査）結果概要",
           "資料：令和7年度 健康とくらしの調査（3-19大雪地区広域連合）／令和7年11月17日〜12月8日実施・"
           "65歳以上7,121人対象・回収4,798票（67.4％）・集計4,729票／単位：％（リスク点数のみ点）",
           [6, 34, 9, 9, 9, 9, 9, 9, 9, 11, 11, 11, 11])

# (指標, 単位, 全体n, [全体,65-69,70-74,75-79,80-84,85+], [広域,東神楽,東川,美瑛], 令和4年値)
NEEDS = [
    ("①", "フレイルあり割合（基本チェックリスト8項目以上）", "％", 4719,
     [19.1, 11.8, 13.0, 18.0, 26.3, 38.7], [19.1, 17.0, 18.4, 21.6], 18.5),
    ("②", "運動機能低下者割合（基本チェックリスト）", "％", 4670,
     [9.1, 4.0, 5.6, 8.5, 14.3, 20.4], [9.1, 8.2, 9.0, 10.1], 9.7),
    ("③", "1年間の転倒あり割合", "％", 4656,
     [35.5, 33.5, 29.1, 38.0, 37.0, 45.3], [35.5, 32.6, 34.0, 39.3], 32.9),
    ("④", "物忘れが多い者の割合", "％", 4598,
     [41.3, 34.8, 38.7, 41.9, 46.6, 50.5], [41.3, 39.5, 40.3, 43.9], 43.4),
    ("⑤", "閉じこもり者割合", "％", 4651,
     [8.5, 3.7, 5.0, 8.2, 13.4, 19.1], [8.5, 7.2, 8.6, 9.7], 6.7),
    ("⑥", "うつ割合（基本チェックリスト）", "％", 4619,
     [28.2, 23.8, 23.9, 28.7, 30.8, 41.2], [28.2, 26.8, 25.0, 32.1], 28.9),
    ("⑦", "口腔機能低下者割合（基本チェックリスト）", "％", 4671,
     [24.7, 18.4, 20.7, 26.0, 29.9, 34.7], [24.7, 21.5, 24.8, 27.5], 23.4),
    ("⑧", "低栄養の傾向割合（BMI18.5未満）", "％", 4528,
     [7.4, 5.9, 6.2, 7.7, 9.0, 9.7], [7.4, 6.7, 8.1, 7.3], 6.0),
    ("⑨", "要支援・要介護リスク点数（平均点）", "点", 4729,
     [16.4, 5.1, 11.9, 18.5, 25.1, 30.0], [16.4, 15.5, 17.0, 16.9], 15.8),
    ("⑩", "認知機能低下者割合（基本チェックリスト）", "％", 4680,
     [34.7, 31.4, 32.6, 31.9, 38.0, 47.5], [34.7, 34.9, 34.6, 34.7], 34.5),
    ("⑪", "IADL（自立度）低下者（1項目以上）割合", "％", 4683,
     [9.8, 4.8, 5.1, 9.1, 14.9, 23.4], [9.8, 9.8, 10.1, 9.4], 9.8),
    ("⑫", "幸福感がある者の割合（8/10点以上）", "％", 4571,
     [50.6, 51.2, 48.7, 48.5, 51.5, 57.9], [50.6, 51.0, 51.9, 49.2], None),
    ("⑬", "就労していない者の割合", "％", 4274,
     [54.9, 33.7, 50.0, 61.7, 69.2, 74.4], [54.9, 56.8, 55.6, 52.6], None),
    ("⑭", "ボランティア参加者（月1回以上）割合", "％", 4524,
     [10.7, 8.0, 10.7, 11.0, 11.8, 13.8], [10.7, 9.3, 13.0, 10.2], 9.0),
    ("⑮", "スポーツの会参加者（月1回以上）割合", "％", 4537,
     [20.7, 16.1, 19.0, 20.4, 27.5, 23.8], [20.7, 21.0, 21.6, 19.6], None),
    ("⑯", "趣味の会参加者（月1回以上）割合", "％", 4533,
     [20.2, 13.8, 18.7, 21.4, 26.1, 24.4], [20.2, 20.4, 21.4, 18.9], None),
    ("⑰", "学習・教養サークル参加者（月1回以上）割合", "％", 4535,
     [6.0, 3.3, 5.2, 6.3, 7.8, 9.7], [6.0, 5.4, 6.8, 5.9], None),
    ("⑱", "通いの場参加者（月1回以上）割合", "％", 4559,
     [8.8, 4.1, 6.4, 8.5, 13.1, 17.6], [8.8, 7.5, 9.3, 9.5], 7.3),
    ("⑲", "特技や経験を他者に伝える活動参加者（月1回以上）割合", "％", 4531,
     [5.0, 4.9, 4.2, 5.2, 5.9, 5.6], [5.0, 3.8, 6.4, 5.0], None),
    ("⑳", "友人知人と会う頻度が高い（月1回以上）者の割合", "％", 4602,
     [63.0, 56.6, 62.8, 66.3, 67.5, 60.6], [63.0, 60.1, 63.1, 65.5], 64.3),
]

r = 4
ws.cell(row=r, column=1, value="（1）年齢階級別（大雪地区広域連合）").font = Font(name=FONT, size=10, bold=True)
HEAD_A = ["No.", "指標", "単位", "サンプル数", "全体", "65～69歳", "70～74歳", "75～79歳", "80～84歳", "85歳以上"]
hrowA = r + 1
endA = table(ws, hrowA, HEAD_A,
             [[no, nm, u, n] + a for no, nm, u, n, a, _t, _p in NEEDS], numfmt="0.0")
for rr in range(hrowA + 1, endA + 1):
    ws.cell(row=rr, column=4).number_format = "#,##0"
r = endA + 2

ws.cell(row=r, column=1, value="（2）町別（全年齢）と令和4年との比較").font = Font(name=FONT, size=10, bold=True)
hrowB = r + 1
HEAD_B = ["No.", "指標", "単位", "広域連合", "東神楽町", "東川町", "美瑛町",
          "令和4年（広域連合）", "差", "統計的判定"]
SIG = {"③": "有意", "④": "有意", "⑤": "有意", "⑧": "有意", "⑭": "有意", "⑱": "有意"}
rowsB = []
for no, nm, u, n, _a, t, p4 in NEEDS:
    rowsB.append([no, nm, u] + t + [p4, None,
                  (SIG.get(no, "有意でない") if p4 is not None else "第9期に該当なし")])
endB = table(ws, hrowB, HEAD_B, rowsB, numfmt="0.0")
for rr in range(hrowB + 1, endB + 1):
    ws.cell(row=rr, column=9).value = f'=IF(H{rr}="","-",D{rr}-H{rr})'
    ws.cell(row=rr, column=9).number_format = "+0.0;-0.0;0.0"
    if ws.cell(row=rr, column=10).value == "有意":
        ws.cell(row=rr, column=10).fill = PatternFill("solid", fgColor="FCE4D6")
# --- H05 社会参加の合成指標（個票CSVから算出） ---
_r = endB + 3
ws.cell(row=_r, column=1, value="（3）社会参加の合成指標（KPI H05の候補・個票データから算出）").font = Font(name=FONT, size=10, bold=True)
_h = _r + 1
H05 = [
    ["通いの場", 8.8, 7.5, 9.3, 9.5, 400, 4559],
    ["スポーツの会", 20.7, 21.0, 21.6, 19.6, 937, 4537],
    ["趣味の会", 20.2, 20.4, 21.4, 18.9, 915, 4533],
    ["ボランティア", 10.7, 9.3, 13.0, 10.2, 485, 4524],
    ["学習・教養サークル", 6.0, 5.4, 6.8, 5.9, 272, 4535],
    ["特技や経験を他者に伝える活動", 5.0, 3.8, 6.4, 5.0, 228, 4531],
    ["【合成】6区分のいずれかに月1回以上参加", 36.0, 34.8, 38.2, 35.4, 1661, 4611],
]
_e = table(ws, _h, ["区分", "広域連合", "東神楽町", "東川町", "美瑛町", "該当数（人）", "分母（人）"],
           H05, numfmt="0.0")
for cc in (6, 7):
    for rr in range(_h + 1, _e + 1):
        ws.cell(row=rr, column=cc).number_format = "#,##0"
for cc in range(1, 8):
    ws.cell(row=_e, column=cc).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=_e, column=cc).fill = PatternFill("solid", fgColor="E2EFDA")
_r2 = _e + 2
ws.cell(row=_r2, column=1, value="（4）合成指標の年齢階級別（広域連合）").font = Font(name=FONT, size=10, bold=True)
_h2 = _r2 + 1
_e2 = table(ws, _h2, ["区分", "全体", "65～69歳", "70～74歳", "75～79歳", "80～84歳", "85歳以上"],
            [["合成指標（6区分いずれか）", 36.0, 28.9, 35.2, 36.3, 43.1, 40.3],
             ["該当数（人）", 1661, 275, 416, 443, 324, 203],
             ["分母（人）", 4611, 953, 1182, 1221, 751, 504]], numfmt="#,##0.0")
mono_bar(ws, "社会参加の合成指標（年齢階級別・広域連合）", "割合（％）",
         Reference(ws, min_col=2, max_col=7, min_row=_h2),
         Reference(ws, min_col=1, max_col=7, min_row=_h2 + 1, max_row=_h2 + 1),
         f"M{_r}", width=20, height=10, from_rows=True, labels=True, numfmt="0.0")
note(ws, _e2 + 1,
     "注）個票データ（KK_2025_CSV）から算出。6区分の単独割合はすべて報告書の公表値と完全に一致することを確認している。"
     "合成指標は6区分のいずれかに月1回以上参加した人の割合で、全区分が無回答の票は分母から除外した（分母4,611人）。"
     "第9期の通いの場参加率7.3％を継承する場合は「通いの場」8.8％、社会参加の広がりを捉える場合は合成指標36.0％を採用する。"
     "年齢階級別では65〜69歳28.9％に対し80〜84歳43.1％と高齢層のほうが高く、通いの場単独（65〜69歳4.1％→80〜84歳13.1％）"
     "と同じ傾向を示す。")

note(ws, endB + 1,
     "注1）令和7年度 健康とくらしの調査（JAGES調査）の集計値。令和4年（広域連合）は第9期計画 第2章第3節の掲載値で、"
     "指標定義（基本チェックリストの該当項目数等）は両年で一致する。"
     "注2）統計的判定は2標本の母比率の差の検定（両側・正規近似、有意水準5％）による。"
     "「有意」は③1年間の転倒あり、④物忘れが多い、⑤閉じこもり、⑧低栄養の傾向、⑭ボランティア参加、⑱通いの場参加の6指標。"
     "①フレイル（＋0.6pt）、⑦口腔機能低下（＋1.3pt）、⑳友人知人と会う（－1.3pt）等は統計的な差が確認できない。"
     "注3）⑨は第9期計画に「15.8％」と記載されているが、令和7年度報告書では「16.4点」と単位が明示されている。"
     "注4）⑫⑬⑮⑯⑰⑲は令和7年度調査で新設された指標であり、第9期に対応する値がない。")

# ============================================================ 09 ニーズ調査グラフ
wsg = sheet("09_ニーズ調査グラフ", "図8-2　健康とくらしの調査 結果概要（グラフ）",
            "08_ニーズ調査データ を参照。左列＝年齢階級別（広域連合）、右列＝町別（全年齢）",
            [14] * 20)
anchor_r = 4
for k, (no, nm, u, n, _a, _t, _p) in enumerate(NEEDS):
    ra = hrowA + 1 + k
    rb = hrowB + 1 + k
    ytitle = "割合（％）" if u == "％" else "平均点（点）"
    ch = BarChart()
    ch.type = "col"
    ch.title = f"{no} {nm}（年齢階級別）"
    ch.y_axis.title = ytitle
    s1 = Series(Reference(ws, min_col=5, max_col=10, min_row=ra, max_row=ra), title_from_data=False)
    s1.tx = SeriesLabel(strRef=StrRef(f"'{ws.title}'!$B${ra}"))
    ch.append(s1)
    ch.set_categories(Reference(ws, min_col=5, max_col=10, min_row=hrowA))
    ch.gapWidth = 60
    ch.width, ch.height = 13, 8
    ch.series[0].graphicalProperties = _fill(3)
    _axis_mono(ch)
    _dlabels(ch, "outEnd", 750, "0.0", per_series=False)
    ch.legend = None
    wsg.add_chart(ch, f"A{anchor_r}")

    ch2 = BarChart()
    ch2.type = "col"
    ch2.title = f"{no} {nm}（町別）"
    ch2.y_axis.title = ytitle
    s2 = Series(Reference(ws, min_col=4, max_col=7, min_row=rb, max_row=rb), title_from_data=False)
    s2.tx = SeriesLabel(strRef=StrRef(f"'{ws.title}'!$B${rb}"))
    ch2.append(s2)
    ch2.set_categories(Reference(ws, min_col=4, max_col=7, min_row=hrowB))
    ch2.gapWidth = 60
    ch2.width, ch2.height = 13, 8
    ch2.series[0].graphicalProperties = _fill(5)
    _axis_mono(ch2)
    _dlabels(ch2, "outEnd", 750, "0.0", per_series=False)
    ch2.legend = None
    wsg.add_chart(ch2, f"J{anchor_r}")
    anchor_r += 17

# 令和4年→令和7年の比較グラフ
CMP = [(no, nm, p4, t[0]) for no, nm, u, n, _a, t, p4 in NEEDS if p4 is not None]
cr = anchor_r
wsg.cell(row=cr, column=1, value="令和4年→令和7年の比較").font = Font(name=FONT, size=10, bold=True)
endC = table(wsg, cr + 1, ["指標", "令和4年", "令和7年"],
             [[nm, p4, p7] for _no, nm, p4, p7 in CMP], numfmt="0.0")
cats = Reference(wsg, min_col=1, min_row=cr + 2, max_row=endC)
data = Reference(wsg, min_col=2, max_col=3, min_row=cr + 1, max_row=endC)
_ch = BarChart()
_ch.type = "bar"
_ch.grouping = "clustered"
_ch.title = "令和4年→令和7年の変化（広域連合）"
_ch.y_axis.title = "割合（％）／点"
_ch.add_data(data, titles_from_data=True)
_ch.set_categories(cats)
_ch.gapWidth = 60
_ch.width, _ch.height = 20, max(7, len(CMP) * 0.85 + 3)
for _i, _sr in enumerate(_ch.series):
    _sr.graphicalProperties = _fill(_i * 4)
_axis_mono(_ch)
_dlabels(_ch, "outEnd", 750, "0.0")
wsg.add_chart(_ch, f"F{cr}")


# ============================================================ 共通：横棒クラスター
def mono_hbar_cluster(ws_, title, cats_ref, data_ref, anchor, width=20, height=None,
                      n_cat=None, x_title="割合（％）"):
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "clustered"
    ch.title = title
    ch.y_axis.title = x_title
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ch.gapWidth = 60
    ch.width = width
    ch.height = height if height else max(7, (n_cat or 6) * 0.85 + 3)
    for i, sr in enumerate(ch.series):
        sr.graphicalProperties = _fill(i * 2)
    _axis_mono(ch)
    _dlabels(ch, "outEnd", 800, "0.0")
    if len(ch.series) <= 1:
        ch.legend = None
    ws_.add_chart(ch, anchor)
    return ch


def block(ws_, row, heading, head, rows, numfmt="0.0", chart_title=None,
          series_cols=None, anchor=None, width=20, x_title="割合（％）", input_cells=None):
    """見出し＋表を書き、横棒クラスターグラフを配置して次の開始行を返す。"""
    ws_.cell(row=row, column=1, value=heading).font = Font(name=FONT, size=10, bold=True)
    hrow = row + 1
    end = table(ws_, hrow, head, rows, numfmt=numfmt)
    if input_cells:
        for rr, cc in input_cells:
            ws_.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=IN_Y)
    if chart_title:
        cats = Reference(ws_, min_col=1, min_row=hrow + 1, max_row=end)
        data = Reference(ws_, min_col=series_cols[0], max_col=series_cols[1],
                         min_row=hrow, max_row=end)
        mono_hbar_cluster(ws_, chart_title, cats, data, anchor or f"H{row}",
                          width=width, n_cat=end - hrow, x_title=x_title)
    return end


# ============================================================ 10 在宅介護実態調査
ws = sheet("10_在宅介護実態調査", "図9　在宅介護実態調査 結果概要",
           "資料：第9期計画 第2章第4節1／令和5年5月25日～6月30日・認定調査員の聞き取り・回答37人（推定）／"
           "令和2年度調査との比較・単位：％",
           [44, 12, 12, 12, 12, 12, 12])
r = 4
r = block(ws, r, "① 世帯類型", ["区分", "令和5年度", "令和2年度"],
          [["単身世帯", 24.3, 26.3], ["夫婦のみ世帯", 18.9, 26.3],
           ["その他", 54.1, 47.4], ["無回答", 2.7, 0.0]],
          chart_title="① 世帯類型", series_cols=(2, 3), anchor="E4") + 2
note(ws, r, "注）令和2年度の無回答0.0％は、単身26.3＋夫婦のみ26.3＋その他47.4＝100.0％となることから復元した値。"
            "「夫婦のみ世帯」が7.4ポイント減少、「その他」が6.7ポイント増加している（第9期計画 第2章第4節1）。")
r += 2

r = block(ws, r, "② 家族等による介護の頻度", ["区分", "令和5年度", "令和2年度"],
          [["ない", 18.9, 11.8], ["家族・親族の介護はあるが、週に1日よりも少ない", 10.8, 10.5],
           ["週に1～2日ある", 13.5, 9.2], ["週に3～4日ある", 0.0, 0.0],
           ["ほぼ毎日ある", 56.8, 63.2], ["無回答", 0.0, 5.3]],
          chart_title="② 家族等による介護の頻度", series_cols=(2, 3), anchor="E23") + 2
note(ws, r, "注）令和2年度の無回答5.3％は、他5区分の合計94.7％との差から復元した値。"
            "「ほぼ毎日ある」は56.8％で令和2年度から6.4ポイント減少している。")
r += 2

CARE = ["日中の排泄", "夜間の排泄", "食事の介助（食べる時）", "入浴・洗身", "身だしなみ（洗顔・歯磨き等）",
        "衣服の着脱", "屋内の移乗・移動", "外出の付き添い、送迎等", "服薬", "認知症状への対応",
        "医療面での対応（経管栄養、ストーマ等）", "食事の準備（調理等）",
        "その他の家事（掃除、洗濯、買い物等）", "金銭管理や生活面に必要な諸手続き", "その他", "わからない", "無回答"]
C_R5 = [21.6, 16.2, 10.8, 21.6, 24.3, 27.0, 16.2, 67.6, 35.1, 18.9, 2.7, 62.2, 73.0, 59.5, 5.4, 0.0, 13.5]
C_R2 = [25.0, 18.8, 12.5, 25.0, 28.1, 31.3, 18.8, 78.1, 40.6, 21.9, 3.1, 71.9, 84.4, 68.8, 6.3, 0.0, None]
start = r + 1
r = block(ws, r, "③ 主な介護者が行っている介護（複数回答）", ["区分", "令和5年度", "令和2年度"],
          [[a, b, c] for a, b, c in zip(CARE, C_R5, C_R2)],
          chart_title="③ 主な介護者が行っている介護（複数回答）", series_cols=(2, 3),
          anchor="E45", input_cells=[(start + 17, 3)]) + 2
note(ws, r, "注）複数回答のため合計は100％にならない。令和2年度の「無回答」（淡黄色欄）は原典から数値を復元できなかったため未記載。"
            "「その他の家事（掃除、洗濯、買い物等）」は73.0％で最も高いが、令和2年度から11.4ポイント減少している。")
r += 2

r = block(ws, r, "④ 介護のための離職の有無", ["区分", "令和5年度", "令和2年度"],
          [["主な介護者が仕事を辞めた（転職除く）", 2.7, 1.5],
           ["主な介護者以外の家族・親族が仕事を辞めた（転職除く）", 0.0, 1.5],
           ["主な介護者が転職した", 2.7, 1.5], ["主な介護者以外の家族・親族が転職した", 0.0, 0.0],
           ["介護のために仕事を辞めた家族・親族はいない", 75.7, 83.6],
           ["わからない", 0.0, 0.0], ["無回答", 18.9, 11.9]],
          chart_title="④ 介護のための離職の有無", series_cols=(2, 3), anchor="E80") + 2
note(ws, r, "注）「介護のために仕事を辞めた家族・親族はいない」は75.7％で最も高いが、令和2年度から7.9ポイント減少している。"
            "介護離職（主な介護者・それ以外の家族親族が仕事を辞めた）は合計2.7％にとどまる。")
r += 2

SUP = ["配食", "調理", "掃除・洗濯", "買い物（宅配は含まない）", "ゴミ出し", "外出同行（通院、買い物など）",
       "移送サービス（介護・福祉等）", "見守り、声かけ", "サロンなどの定期的な通いの場", "その他", "特になし", "無回答"]
S_R5 = [13.5, 13.5, 21.6, 16.2, 16.2, 37.8, 18.9, 16.2, 5.4, 16.2, 27.0, 10.8]
S_R2 = [17.1, 14.5, 15.8, 13.2, 13.2, 21.1, 18.4, 18.4, 18.4, 2.7, 43.4, 10.0]
r = block(ws, r, "⑤ 在宅生活の継続のために充実が必要な支援・サービス（複数回答）",
          ["区分", "令和5年度", "令和2年度"], [[a, b, c] for a, b, c in zip(SUP, S_R5, S_R2)],
          chart_title="⑤ 在宅生活の継続のために充実が必要な支援・サービス", series_cols=(2, 3),
          anchor="E100") + 2
note(ws, r, "注）「外出同行（通院、買い物など）」が37.8％で最も高く、令和2年度から16.7ポイント上昇している。"
            "一方「特になし」は43.4％から27.0％へ減少しており、支援ニーズの顕在化がうかがえる。")
r += 2

r = block(ws, r, "⑥ 施設等検討の状況", ["区分", "令和5年度", "令和2年度"],
          [["入所・入居は検討していない", 73.0, 82.9], ["入所・入居を検討している", 13.5, 11.8],
           ["すでに入所・入居申し込みをしている", 13.5, 5.3], ["無回答", 0.0, 0.0]],
          chart_title="⑥ 施設等検討の状況", series_cols=(2, 3), anchor="E128")
note(ws, r + 2, "注）「入所・入居は検討していない」は73.0％で最も高いが令和2年度から9.9ポイント減少し、"
                "「すでに入所・入居申し込みをしている」が8.2ポイント上昇している。")

# ============================================================ 11 居所変更実態調査
ws = sheet("11_居所変更実態調査", "図10　居所変更実態調査 結果概要",
           "資料：第9期計画 第2章第4節2／令和5年5月25日に施設・居住系サービスの管理者へ書面送付・21施設回答／"
           "過去1年間に居所を変更又は死亡した利用者",
           [30, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11])
KYOSHO = [
    ("住宅型有料老人ホーム", 34, 3), ("軽費老人ホーム", 9, 3), ("グループホーム", 15, 5),
    ("特定施設", 13, 20), ("介護老人保健施設", 91, 6), ("特別養護老人ホーム", 41, 13),
    ("地域密着型特別養護老人ホーム", 25, 6),
]
r = 4
ws.cell(row=r, column=1, value="① 過去1年間の退居・退所者に占める居所変更・死亡").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
rows = [[n, a, b, None, None, None] for n, a, b in KYOSHO] + [["合計", None, None, None, None, None]]
end = table(ws, hrow, ["区分", "居所変更（件）", "死亡（件）", "合計（件）", "居所変更（％）", "死亡（％）"],
            rows, numfmt="#,##0")
TOT = end
for rr in range(hrow + 1, end + 1):
    ws.cell(row=rr, column=4).value = f"=B{rr}+C{rr}"
    ws.cell(row=rr, column=5).value = f'=IF($D{rr}=0,"-",B{rr}/$D{rr}*100)'
    ws.cell(row=rr, column=6).value = f'=IF($D{rr}=0,"-",C{rr}/$D{rr}*100)'
    ws.cell(row=rr, column=4).number_format = "#,##0"
    for cc in (5, 6):
        ws.cell(row=rr, column=cc).number_format = "0.0"
for cc in (2, 3):
    col = get_column_letter(cc)
    ws.cell(row=TOT, column=cc).value = f"=SUM({col}{hrow+1}:{col}{TOT-1})"
    ws.cell(row=TOT, column=cc).number_format = "#,##0"
    ws.cell(row=TOT, column=cc).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=TOT)
data = Reference(ws, min_col=5, max_col=6, min_row=hrow, max_row=TOT)
mono_hbar_cluster(ws, "① 退居・退所者に占める居所変更・死亡の割合（施設別）", cats, data, "M4",
                  width=20, n_cat=8)
r = end + 2
note(ws, r, "注）各施設の合計では80.3％が「居所変更」、19.7％が「死亡」。介護老人保健施設は「居所変更」が93.8％と高く、"
            "特定施設は「死亡」が60.6％と高い（第9期計画 第2章第4節2）。"
            "なお回答施設の内訳（住宅型有料4・軽費1・サ高住0・GH5・特定施設2・老健3・特養2・地密特養3）の合計は20だが、"
            "第9期計画は「合計21」と記載しており差異がある（00_凡例・出典 4-7）。")
r += 2

ws.cell(row=r, column=1, value="② 居所変更した人の要支援・要介護度（施設別・単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow2 = r + 1
DO = [
    ("住宅型有料老人ホーム", 0, 2, 4, 3, 1, 6, 6, 12, 0),
    ("軽費老人ホーム", 2, 1, 0, 3, 1, 2, 0, 0, 0),
    ("グループホーム", 0, 0, 0, 1, 2, 2, 2, 8, 0),
    ("特定施設", 0, 2, 3, 2, 2, 3, 1, 0, 0),
    ("介護老人保健施設", 0, 0, 0, 12, 13, 26, 22, 18, 0),
    ("特別養護老人ホーム", 0, 0, 0, 1, 1, 13, 9, 17, 0),
    ("地域密着型特別養護老人ホーム", 0, 0, 0, 1, 2, 8, 7, 7, 0),
]
HEAD_DO = ["区分", "自立", "要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5", "申請中", "合計"]
rows2 = [list(x) + [None] for x in DO] + [["合計"] + [None] * 10]
end2 = table(ws, hrow2, HEAD_DO, rows2, numfmt="#,##0")
for rr in range(hrow2 + 1, end2 + 1):
    ws.cell(row=rr, column=11).value = f"=SUM(B{rr}:J{rr})"
    ws.cell(row=rr, column=11).number_format = "#,##0"
for cc in range(2, 11):
    col = get_column_letter(cc)
    ws.cell(row=end2, column=cc).value = f"=SUM({col}{hrow2+1}:{col}{end2-1})"
    ws.cell(row=end2, column=cc).number_format = "#,##0"
    ws.cell(row=end2, column=cc).font = Font(name=FONT, size=9, bold=True)
ws.cell(row=end2, column=11).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=1, min_row=hrow2 + 1, max_row=end2 - 1)
data = Reference(ws, min_col=2, max_col=10, min_row=hrow2, max_row=end2 - 1)
mono_bar(ws, "② 居所変更した人の要支援・要介護度（施設別）", "人数（人）", cats, data,
         "M25", stacked=True, width=20, height=11)
r = end2 + 2
note(ws, r, "注）軽費老人ホームの「－」表記は0として集計している。合計では「要介護5」が62人と最も多く、"
            "次いで「要介護3」60人、「要介護4」47人となっている（第9期計画 第2章第4節2）。")
r += 2

ws.cell(row=r, column=1, value="③ 居所変更した理由（21施設・単位：件）").font = Font(name=FONT, size=10, bold=True)
hrow3 = r + 1
RIYU = [
    ("必要な支援の発生・増大", 2), ("必要な身体介護の発生・増大", 5), ("認知症の症状の悪化", 7),
    ("医療的ケア・医療処置の必要性の高まり", 19), ("上記以外の状態像が悪化", 8), ("状態等の改善", 5),
    ("必要な居宅サービスを望まなかったため", 1), ("費用負担が重くなった", 4), ("その他", 10),
]
end3 = table(ws, hrow3, ["区分", "件数"], [list(x) for x in RIYU], numfmt="#,##0")
cats = Reference(ws, min_col=1, min_row=hrow3 + 1, max_row=end3)
data = Reference(ws, min_col=2, min_row=hrow3, max_row=end3)
ch = mono_hbar_cluster(ws, "③ 居所変更した理由（21施設）", cats, data, "M48",
                       width=20, n_cat=9, x_title="件数（件）")
ch.legend = None
note(ws, end3 + 2, "注）上位3つは「医療的ケア・医療処置の必要性の高まり」19件、「その他」10件、「上記以外の状態像が悪化」8件。"
                   "一方で「状態等の改善」も5件あり、より良い方向への転換もみられる（第9期計画 第2章第4節2）。")

# ============================================================ 12 在宅生活改善調査
ws = sheet("12_在宅生活改善調査", "図11　在宅生活改善調査 結果概要",
           "資料：第9期計画 第2章第4節3／令和5年5月25日に居宅介護支援事業所のケアマネジャーへ書面送付・"
           "12事業所91人／自宅等から居所を変更した利用者・単位：％",
           [44, 12, 14, 14, 14, 12, 12])
r = 4
ws.cell(row=r, column=1, value="（参考）回答者の要介護度別内訳（単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["要介護度", "該当者数"],
            [["要支援1", 0], ["要支援2", 7], ["要介護1", 31], ["要介護2", 18],
             ["要介護3", 12], ["要介護4", 15], ["要介護5", 8], ["合計", None]], numfmt="#,##0")
ws.cell(row=end, column=2).value = f"=SUM(B{hrow+1}:B{end-1})"
ws.cell(row=end, column=2).number_format = "#,##0"
ws.cell(row=end, column=2).font = Font(name=FONT, size=9, bold=True)
r = end + 2

ATTR = ["独居", "夫婦のみ世帯", "単身の子供との同居", "その他世帯", "自宅等（持ち家）", "自宅等（借家）",
        "サ高住・住宅型有料・軽費", "要介護2以下", "要介護3以上"]
A_R5 = [46.7, 13.3, 0.0, 26.7, 62.2, 6.7, 17.8, 53.3, 33.4]
A_R2 = [27.1, 25.9, 11.1, 22.4, 77.7, 4.9, 4.9, 48.1, 39.4]
r = block(ws, r, "① 在宅での生活の維持が難しくなっている利用者の属性",
          ["区分", "令和5年度", "令和2年度"], [[a, b, c] for a, b, c in zip(ATTR, A_R5, A_R2)],
          chart_title="① 生活の維持が難しくなっている利用者の属性", series_cols=(2, 3), anchor="H4") + 2
note(ws, r, "注）「独居」が19.6ポイント増加、「自宅等（持ち家）」が15.5ポイント減少、"
            "「サ高住・住宅型有料老人ホーム・軽費老人ホーム」が12.9ポイント増加している（第9期計画 第2章第4節3）。"
            "世帯類型・居所・要介護度の各区分は無回答等を含むため合計は100％にならない。")
r += 2

JOTAI = ["必要な生活支援の発生・増大", "必要な身体介護の増大", "認知症の症状の悪化",
         "医療的ケア・医療処置の必要性の高まり", "その他、本人の状態等の悪化", "本人の状態等の改善",
         "その他", "無回答"]
J = [(48.9, 55.6, 38.9, 46.9), (51.1, 40.7, 66.7, 61.7), (66.7, 74.1, 55.6, 63.0),
     (8.9, 7.4, 11.1, 24.7), (22.2, 25.9, 16.7, 17.3), (4.4, 3.7, 5.6, 0.0),
     (4.4, 0.0, 11.1, 0.0), (0.0, 0.0, 0.0, 0.0)]
HD4 = ["区分", "合計（令和5年度）", "要支援1～要介護2", "要介護3～要介護5", "合計（令和2年度）"]
r = block(ws, r, "② 生活の維持が難しくなっている理由（本人の状態に属する理由）", HD4,
          [[a] + list(b) for a, b in zip(JOTAI, J)],
          chart_title="② 本人の状態に属する理由（要介護度別）", series_cols=(2, 4), anchor="H26")
cats = Reference(ws, min_col=1, min_row=r - len(JOTAI) + 1, max_row=r)
d2 = Reference(ws, min_col=2, min_row=r - len(JOTAI), max_row=r)
d5 = Reference(ws, min_col=5, min_row=r - len(JOTAI), max_row=r)
ch = BarChart(); ch.type = "bar"; ch.grouping = "clustered"
ch.title = "② 本人の状態に属する理由（令和2年度との比較）"
ch.y_axis.title = "割合（％）"
ch.add_data(d2, titles_from_data=True); ch.add_data(d5, titles_from_data=True)
ch.set_categories(cats); ch.gapWidth = 60; ch.width, ch.height = 20, 10
for i, sr in enumerate(ch.series):
    sr.graphicalProperties = _fill(i * 4)
_axis_mono(ch); _dlabels(ch, "outEnd", 750, "0.0"); ws.add_chart(ch, "H46")
r += 2
note(ws, r, "注）合計では「認知症の症状の悪化」が66.7％と最も高い一方、要介護3～要介護5では「必要な身体介護の増大」が66.7％と最も高い。"
            "令和2年度と比べ「必要な身体介護の増大」は10.6ポイント減少している（第9期計画 第2章第4節3）。")
r += 2

IKO = ["本人が、一部の居宅サービスの利用を望まないから", "生活の不安が大きいから", "居住環境が不便だから",
       "本人が介護者の負担軽減を望むから", "負担費用が重いから", "その他、本人の意向等があるから",
       "その他", "無回答"]
I2 = [(15.6, 18.5, 11.1, 32.1), (33.3, 33.3, 33.3, 29.6), (8.9, 7.4, 11.1, 9.9),
      (2.2, 0.0, 5.6, 4.9), (8.9, 11.1, 5.6, 6.2), (11.1, 3.7, 22.2, 23.5),
      (35.6, 40.7, 27.8, 25.9), (0.0, 0.0, 0.0, 3.7)]
r = block(ws, r, "③ 生活の維持が難しくなっている理由（本人の意向に属する理由）", HD4,
          [[a] + list(b) for a, b in zip(IKO, I2)],
          chart_title="③ 本人の意向に属する理由（要介護度別）", series_cols=(2, 4), anchor="H66")
cats = Reference(ws, min_col=1, min_row=r - len(IKO) + 1, max_row=r)
d2 = Reference(ws, min_col=2, min_row=r - len(IKO), max_row=r)
d5 = Reference(ws, min_col=5, min_row=r - len(IKO), max_row=r)
ch = BarChart(); ch.type = "bar"; ch.grouping = "clustered"
ch.title = "③ 本人の意向に属する理由（令和2年度との比較）"
ch.y_axis.title = "割合（％）"
ch.add_data(d2, titles_from_data=True); ch.add_data(d5, titles_from_data=True)
ch.set_categories(cats); ch.gapWidth = 60; ch.width, ch.height = 20, 10
for i, sr in enumerate(ch.series):
    sr.graphicalProperties = _fill(i * 4)
_axis_mono(ch); _dlabels(ch, "outEnd", 750, "0.0"); ws.add_chart(ch, "H86")
r += 2
note(ws, r, "注）合計では「その他」が35.6％と最も高い一方、要介護3～要介護5では「生活の不安が大きいから」が最も高い。"
            "令和2年度と比べ「本人が、一部の居宅サービスの利用を望まないから」が16.5ポイント減少、「その他」が9.7ポイント増加している。")
r += 2

KAZOKU = ["介護者の介護に係る不安・負担量の増大", "介護者が、一部の居宅サービスの利用を望まないから",
          "家族等の介護等技術では対応が困難", "費用負担が重いから", "家族等の就労継続が困難になり始めたから",
          "本人と家族等の関係性に課題があるから", "その他、家族等介護者の意向等があるから", "その他", "無回答"]
K2 = [(51.1, 44.4, 61.1, 66.7), (6.7, 3.7, 11.1, 17.3), (26.7, 11.1, 50.0, 34.6),
      (11.1, 7.4, 16.7, 12.3), (8.9, 3.7, 16.7, 17.3), (20.0, 22.2, 16.7, 19.8),
      (20.0, 18.5, 22.2, 23.5), (13.3, 18.5, 5.6, 3.7), (2.2, 0.0, 5.6, 3.7)]
r = block(ws, r, "④ 生活の維持が難しくなっている理由（家族等介護者の意向・負担等に属する理由）", HD4,
          [[a] + list(b) for a, b in zip(KAZOKU, K2)],
          chart_title="④ 家族等介護者の意向・負担等（要介護度別）", series_cols=(2, 4), anchor="H106")
cats = Reference(ws, min_col=1, min_row=r - len(KAZOKU) + 1, max_row=r)
d2 = Reference(ws, min_col=2, min_row=r - len(KAZOKU), max_row=r)
d5 = Reference(ws, min_col=5, min_row=r - len(KAZOKU), max_row=r)
ch = BarChart(); ch.type = "bar"; ch.grouping = "clustered"
ch.title = "④ 家族等介護者の意向・負担等（令和2年度との比較）"
ch.y_axis.title = "割合（％）"
ch.add_data(d2, titles_from_data=True); ch.add_data(d5, titles_from_data=True)
ch.set_categories(cats); ch.gapWidth = 60; ch.width, ch.height = 20, 10
for i, sr in enumerate(ch.series):
    sr.graphicalProperties = _fill(i * 4)
_axis_mono(ch); _dlabels(ch, "outEnd", 750, "0.0"); ws.add_chart(ch, "H130")
note(ws, r + 2, "注）いずれの介護度でも「介護者の介護に係る不安・負担量の増大」が最も高い。"
                "要支援1～要介護2では「本人と家族等の関係性に課題があるから」、要介護3～要介護5では「家族等の介護等技術では対応が困難」が次いで高い。"
                "令和2年度と比べ「介護者の介護に係る不安・負担量の増大」は15.6ポイント減少している（第9期計画 第2章第4節3）。"
                "②～④はいずれも複数回答のため合計は100％にならない。")


# ============================================================ 共通：100%積上げ横棒
def mono_pct_hbar(ws_, title, cats_ref, data_ref, anchor, n_cat=3, width=20, height=None):
    ch = BarChart()
    ch.type = "bar"
    ch.grouping = "percentStacked"
    ch.overlap = 100
    ch.title = title
    ch.y_axis.title = "構成比（％）"
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ch.gapWidth = 60
    ch.width = width
    ch.height = height if height else max(6, n_cat * 0.9 + 3.5)
    for i, sr in enumerate(ch.series):
        sr.graphicalProperties = _fill(i)
    _axis_mono(ch)
    _dlabels(ch, "ctr", 800, "0.0")
    if len(ch.series) <= 1:
        ch.legend = None
    ws_.add_chart(ch, anchor)
    return ch



_ca_row = [4]


def CA(height_cm, reset=None):
    """13シート用：グラフの高さに応じてF列の配置行を自動で送る。"""
    if reset is not None:
        _ca_row[0] = reset
    row = _ca_row[0]
    # 作図関数が既定の高さ（7.5cm）を用いる場合があるため、送り量の下限を7.5cmとする。
    # 標準の行高15pt＝0.529cmで換算し、切上げたうえで3行の余白を確保する。
    _ca_row[0] = row + int(-(-max(height_cm, 7.5) * 100 // 53)) + 3
    return f"J{row}"


# ============================================================ 13 介護人材実態調査
ws = sheet("13_介護人材実態調査", "図12　介護人材実態調査 結果概要",
           "資料：第9期計画 第2章第5節／令和5年5月25日に施設系サービスの管理者へ書面送付・"
           "27施設回答（該当職員405人）／対象35事業所・回収率77.1％",
           [40, 13, 13, 13, 13, 13, 13, 13])
r = 4
ws.cell(row=r, column=1, value="（参考）回答事業所の内訳（単位：施設）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["サービス種別", "施設数"],
            [["施設・居住系サービス", 18], ["通所系サービス", 3], ["不明", 6], ["合計", None]], numfmt="#,##0")
ws.cell(row=end, column=2).value = f"=SUM(B{hrow+1}:B{end-1})"
ws.cell(row=end, column=2).number_format = "#,##0"
ws.cell(row=end, column=2).font = Font(name=FONT, size=9, bold=True)
r = end + 2

ws.cell(row=r, column=1, value="（1）① 事業所の職員数（単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "男性", "女性", "合計"],
            [["正規職員", 134, 174, None], ["非正規職員", 8, 87, None], ["合計", None, None, None]],
            numfmt="#,##0")
for rr in range(hrow + 1, end + 1):
    ws.cell(row=rr, column=4).value = f"=B{rr}+C{rr}"
    ws.cell(row=rr, column=4).number_format = "#,##0"
for cc in (2, 3):
    col = get_column_letter(cc)
    ws.cell(row=end, column=cc).value = f"=SUM({col}{hrow+1}:{col}{end-1})"
    ws.cell(row=end, column=cc).number_format = "#,##0"
    ws.cell(row=end, column=cc).font = Font(name=FONT, size=9, bold=True)
ws.cell(row=end, column=4).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end - 1)
data = Reference(ws, min_col=2, max_col=3, min_row=hrow, max_row=end - 1)
mono_hbar_cluster(ws, "（1）① 事業所の職員数（就業形態別・男女別）", cats, data, CA(max(7, 2 * 0.85 + 3), reset=4),
                  width=17, n_cat=2, x_title="人数（人）")
r = end + 2
note(ws, r, "注）第9期計画には職員総数（正規325・非正規96・計421人）、男女内訳の合計403人、該当者数405人の3種類が併記されている。"
            "本表は男女内訳（403人）を掲載。原典の注記のとおり「施設に所属している介護職員が未記入の人は除いており、合計が総数と一致しない」。"
            "1事業所あたり平均人数も405÷27＝15.0人で本文の16.0人と一致しない（00_凡例・出典 4-9）。")
r += 2

r = block(ws, r, "（1）② 男女比較（該当者405人）", ["区分", "構成比"],
          [["男性", 35.3], ["女性", 64.4], ["無回答", 0.2]],
          chart_title="（1）② 男女比較", series_cols=(2, 2), anchor=CA(max(7, 3 * 0.85 + 3)), x_title="構成比（％）") + 2
r = block(ws, r, "（1）③ 就業形態比較（該当者405人）", ["区分", "構成比"],
          [["正規社員", 76.3], ["非正規社員", 23.5], ["無回答", 0.2]],
          chart_title="（1）③ 就業形態比較", series_cols=(2, 2), anchor=CA(max(7, 3 * 0.85 + 3)), x_title="構成比（％）") + 2

ws.cell(row=r, column=1, value="（1）④ 全職員の就業形態（介護保険サービス系列別・単位：％）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "正規職員（男）", "正規職員（女）", "非正規職員（男）", "非正規職員（女）",
                       "正規職員 計", "男性 計", "平均人数（人）"],
            [["全体", 31.2, 43.6, 1.8, 23.4, None, None, 16.0],
             ["施設・居住系サービス", 31.9, 45.1, 1.9, 21.1, None, None, 17.6],
             ["通所系サービス", 20.0, 20.0, 0.0, 60.0, None, None, 6.7]], numfmt="0.0")
for rr in range(hrow + 1, end + 1):
    ws.cell(row=rr, column=6).value = f"=B{rr}+C{rr}"
    ws.cell(row=rr, column=7).value = f"=B{rr}+D{rr}"
    for cc in (6, 7):
        ws.cell(row=rr, column=cc).number_format = "0.0"
        ws.cell(row=rr, column=cc).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end)
data = Reference(ws, min_col=2, max_col=5, min_row=hrow, max_row=end)
mono_pct_hbar(ws, "（1）④ 全職員の就業形態（サービス系列別）", cats, data, CA(max(6, 3 * 0.9 + 3.5)), n_cat=3, width=20)
r = end + 2
note(ws, r, "注）第9期計画の本文は「施設・居住系サービスの正規職員が33.8％」とするが、グラフ値では正規職員計は77.0％であり、"
            "33.8％は男性職員の割合（31.9＋1.9）と一致する。「通所系サービスの正規職員は40.0％」は正規職員計（20.0＋20.0）と一致する。"
            "本表はグラフ値どおり掲載し、正規職員計・男性計を数式で併記した（00_凡例・出典 4-8）。")
r += 2

ws.cell(row=r, column=1, value="（2）事業所の開設後経過年数（介護保険サービス系列別・単位：％）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "5年未満", "5年以上10年未満", "10年以上20年未満", "20年以上30年未満",
                       "30年以上", "平均経過年数"],
            [["全体", 3.7, 14.8, 37.0, 29.6, 14.8, "19年3か月"],
             ["施設・居住系サービス", 0.0, 16.7, 44.4, 27.8, 11.1, "19年8か月"],
             ["通所系サービス", 0.0, 0.0, 0.0, 66.7, 33.3, "25年7か月"]], numfmt="0.0")
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end)
data = Reference(ws, min_col=2, max_col=6, min_row=hrow, max_row=end)
mono_pct_hbar(ws, "（2）事業所の開設後経過年数（サービス系列別）", cats, data, CA(max(6, 3 * 0.9 + 3.5)), n_cat=3, width=20)
r = end + 2
note(ws, r, "注）全体では「10年以上20年未満」が37.0％で最も高く、次いで「20年以上30年未満」29.6％。"
            "通所系サービスは「20年以上30年未満」66.7％、「30年以上」33.3％で、平均経過年数25年7か月と長い。"
            "施設・居住系の「5年未満」0.0％は他4区分の合計100.0％から復元した値。")
r += 2

ws.cell(row=r, column=1, value="（3）① 全職員の年齢構成・在職年数（総数）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
AGE = [["20歳未満", 5, None], ["20代", 71, None], ["30代", 82, None], ["40代", 107, None],
       ["50代", 81, None], ["60代", 51, None], ["70代以上", 8, None], ["合計", None, None]]
end = table(ws, hrow, ["年代", "人数（人）", "構成比（％）"], AGE, numfmt="#,##0")
ws.cell(row=end, column=2).value = f"=SUM(B{hrow+1}:B{end-1})"
ws.cell(row=end, column=2).font = Font(name=FONT, size=9, bold=True)
for rr in range(hrow + 1, end + 1):
    ws.cell(row=rr, column=3).value = f"=B{rr}/$B${end}*100"
    ws.cell(row=rr, column=3).number_format = "0.0"
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end - 1)
data = Reference(ws, min_col=2, min_row=hrow, max_row=end - 1)
ch = mono_bar(ws, "（3）① 全職員の年齢構成（総数）", "人数（人）", cats, data, CA(9),
              width=16, height=9, labels=True, numfmt="#,##0")
ch.legend = None
r2 = end + 2
hrow2 = r2
end2 = table(ws, hrow2, ["在職年数", "人数（人）", "構成比（％）"],
             [["1年未満", 56, None], ["1年以上", 345, None], ["無回答", 4, None], ["合計", None, None]],
             numfmt="#,##0")
ws.cell(row=end2, column=2).value = f"=SUM(B{hrow2+1}:B{end2-1})"
ws.cell(row=end2, column=2).font = Font(name=FONT, size=9, bold=True)
for rr in range(hrow2 + 1, end2 + 1):
    ws.cell(row=rr, column=3).value = f"=B{rr}/$B${end2}*100"
    ws.cell(row=rr, column=3).number_format = "0.0"
r = end2 + 2

ws.cell(row=r, column=1, value="（3）② 全職員の年齢構成（介護保険サービス系列別・単位：％）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "20歳未満", "20代", "30代", "40代", "50代", "60代", "70代以上"],
            [["全体", 1.5, 17.8, 18.6, 26.9, 20.1, 13.0, 2.1],
             ["施設・居住系サービス", 1.6, 17.9, 18.2, 26.4, 20.1, 13.5, 2.2],
             ["通所系サービス", 0.0, 15.0, 25.0, 35.0, 20.0, 5.0, 0.0]], numfmt="0.0")
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end)
data = Reference(ws, min_col=2, max_col=8, min_row=hrow, max_row=end)
mono_pct_hbar(ws, "（3）② 全職員の年齢構成（サービス系列別）", cats, data, CA(max(6, 3 * 0.9 + 3.5)), n_cat=3, width=20)
r = end + 2
r = block(ws, r, "（3）③ 全職員の在職年数（介護保険サービス系列別・単位：％）",
          ["区分", "1年未満", "1年以上"],
          [["全体", 15.2, 84.8], ["施設・居住系サービス", 14.9, 85.1], ["通所系サービス", 20.0, 80.0]],
          chart_title="（3）③ 全職員の在職年数（サービス系列別）", series_cols=(2, 3),
          anchor=CA(max(7, 3 * 0.85 + 3)), x_title="構成比（％）") + 2
note(ws, r, "注）全体では「40代」が26.9％で最も高く、次いで「50代」20.1％。全てのサービス系列で「20歳未満」が最も少ない。"
            "在職年数は全てのサービス系列で「1年以上」が80％を超えている（第9期計画 第2章第5節(3)）。")
r += 2

PREV = ["現在の職場が初めての勤務先", "介護以外の職場",
        "特養、老健、療養型・介護医療院、ショートステイ、グループホーム、特定施設",
        "訪問介護・入浴、夜間対応型", "小多機、看多機、定期巡回", "通所介護、通所リハ、認知症デイ",
        "住宅型有料、サ高住（特定施設以外）", "その他の介護サービス"]
P_V = [16.1, 26.8, 37.5, None, None, 1.8, 7.1, 1.8]
start = r + 1
r = block(ws, r, "（3）④ 現在の施設等に勤務する直前の職場（在職1年未満の56人・単位：％）",
          ["区分", "構成比"], [[a, b] for a, b in zip(PREV, P_V)],
          chart_title="（3）④ 直前の職場（在職1年未満）", series_cols=(2, 2), anchor=CA(max(7, 8 * 0.85 + 3)),
          x_title="構成比（％）", input_cells=[(start + 4, 2), (start + 5, 2)]) + 2
note(ws, r, "注）「特養、老健、療養型・介護医療院、ショートステイ、グループホーム、特定施設」が37.5％で最も多く、"
            "次いで「介護以外の職場」26.8％。「現在の職場が初めての勤務先」16.1％と合わせると42.9％が介護職未経験からの入職。"
            "淡黄色の2区分は原典の図から数値を復元できなかった（他6区分の合計91.1％との差8.9％＝5人分・00_凡例・出典 4-10）。")
r += 2

ws.cell(row=r, column=1, value="（3）④ 直前が介護の職場だった方の内訳（32人・単位：％）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "同一", "別", "無回答"],
            [["市区町村", 28.1, 62.5, 9.4], ["法人・グループ", 25.0, 65.6, 9.4]], numfmt="0.0")
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end)
data = Reference(ws, min_col=2, max_col=4, min_row=hrow, max_row=end)
mono_pct_hbar(ws, "（3）④ 直前が介護の職場だった方の内訳", cats, data, CA(max(6, 2 * 0.9 + 3.5)), n_cat=2, width=18)
r = end + 2
note(ws, r, "注）直前の職場が介護の職場だった方のうち、市区町村は「別の市区町村内」62.5％が「同一の市区町村内」28.1％を上回り、"
            "法人も「別の法人・グループ」65.6％が「同一の法人・グループ」25.0％を上回る。圏域外・法人外からの人材流入が中心。")
r += 2

r = block(ws, r, "（4）資格の取得・研修の修了の状況（単位：％）", ["区分", "構成比"],
          [["介護福祉士（認定介護福祉士含む）", 64.0],
           ["介護職員実務者研修修了、または(旧)介護職員基礎研修修了、または(旧)ヘルパー1級", 3.0],
           ["介護職員初任者研修修了、または(旧)ヘルパー2級", 13.6],
           ["上記のいずれも該当しない", 19.5]],
          chart_title="（4）資格の取得・研修の修了の状況", series_cols=(2, 2), anchor=CA(max(7, 4 * 0.85 + 3)),
          x_title="構成比（％）") + 2

ws.cell(row=r, column=1, value="（5）① 過去1年間の採用者数・離職者数（単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "人数"], [["過去1年の介護職員の採用者数", 80], ["過去1年の介護職員の離職者数", 74]],
            numfmt="#,##0")
r = end + 2

ws.cell(row=r, column=1, value="（5）② サービス別の採用者数・退職者数（単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分", "採用（正規）", "採用（非正規）", "退職（正規）", "退職（非正規）"],
            [["施設・居住系サービス", 42, 15, 42, 15], ["通所系サービス", 4, 3, 3, 4],
             ["無回答", 6, 2, 8, 0], ["合計", None, None, None, None]], numfmt="#,##0")
for cc in range(2, 6):
    col = get_column_letter(cc)
    ws.cell(row=end, column=cc).value = f"=SUM({col}{hrow+1}:{col}{end-1})"
    ws.cell(row=end, column=cc).number_format = "#,##0"
    ws.cell(row=end, column=cc).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end - 1)
data = Reference(ws, min_col=2, max_col=5, min_row=hrow, max_row=end - 1)
mono_bar(ws, "（5）② サービス別の採用者数・退職者数", "人数（人）", cats, data, CA(9),
         width=17, height=9, gap=80)
r = end + 2

ws.cell(row=r, column=1, value="（5）③ 年齢別の採用者数・退職者数（単位：人）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
AGE5 = [["20歳未満", 3, 0, 1, 0], ["20代", 19, 4, 13, 2], ["30代", 11, 3, 10, 1],
        ["40代", 7, 6, 10, 8], ["50代", 8, 4, 12, 4], ["60代", 4, 3, 6, 4],
        ["70代以上", 0, 0, 0, 3], ["合計", None, None, None, None]]
end = table(ws, hrow, ["区分", "採用（正規）", "採用（非正規）", "退職（正規）", "退職（非正規）"],
            AGE5, numfmt="#,##0")
for cc in range(2, 6):
    col = get_column_letter(cc)
    ws.cell(row=end, column=cc).value = f"=SUM({col}{hrow+1}:{col}{end-1})"
    ws.cell(row=end, column=cc).number_format = "#,##0"
    ws.cell(row=end, column=cc).font = Font(name=FONT, size=9, bold=True)
ws.cell(row=end + 1, column=1, value="30代以下の正規採用者が正規採用者に占める割合（％）").font = Font(name=FONT, size=9, bold=True)
ws.cell(row=end + 1, column=2).value = f"=SUM(B{hrow+1}:B{hrow+3})/B{end}*100"
ws.cell(row=end + 1, column=2).number_format = "0.0"
ws.cell(row=end + 1, column=2).font = Font(name=FONT, size=9, bold=True)
ws.cell(row=end + 1, column=2).fill = PatternFill("solid", fgColor="E2EFDA")
cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=end - 1)
data = Reference(ws, min_col=2, max_col=5, min_row=hrow, max_row=end - 1)
mono_bar(ws, "（5）③ 年齢別の採用者数・退職者数", "人数（人）", cats, data, CA(9),
         width=17, height=9, gap=80)
r = end + 3
note(ws, r, "注）（5）②の退職者合計（正規53・非正規19）と（5）③の退職者合計（正規52・非正規22）は一致しない。"
            "原典の注記のとおり、内訳が不明な人を除いているため①の80人・74人とも一致しない。"
            "30代以下の正規採用者は33人で正規採用者52人の63.5％を占め、第9期計画 第2章第6節4の記載と一致する。")
r += 2

WORK = [["5時間未満", 1.5], ["5時間以上10時間未満", 2.7], ["10時間以上15時間未満", 3.0],
        ["15時間以上20時間未満", 3.0], ["20時間以上25時間未満", 7.4], ["25時間以上30時間未満", 1.7],
        ["30時間以上35時間未満", 9.1], ["35時間以上40時間未満", 11.6], ["40時間以上45時間未満", 46.9],
        ["45時間以上50時間未満", 4.9], ["50時間以上", 0.7], ["不明", 7.4]]
r = block(ws, r, "（6）過去1週間の勤務時間（単位：％）", ["区分", "構成比"], WORK,
          chart_title="（6）過去1週間の勤務時間", series_cols=(2, 2), anchor=CA(max(7, 12 * 0.85 + 3)),
          x_title="構成比（％）")
note(ws, r + 2, "注）「40時間以上45時間未満」が46.9％で最も高く、次いで「35時間以上40時間未満」11.6％。"
                "一方で30時間未満が19.3％を占めており、短時間勤務者の比率が高い（第9期計画 第2章第5節(6)）。")


# ============================================================ 14 年齢構成と85歳以上
ws = sheet("14_年齢構成と85歳以上", "図13　高齢者の年齢構成と85歳以上人口の推移",
           "資料：地域包括ケア「見える化」システム A3・A4（総務省国勢調査＋国立社会保障・人口問題研究所推計／取得日 令和8年7月22日）"
           "／単位：人。第9期計画の図1・図2は住民基本台帳ベースで基準が異なる（00_凡例・出典 4-11）",
           [18] + [10] * 9)
YR14 = ["2015年", "2020年", "2025年", "2030年", "2035年", "2040年", "2045年", "2050年"]
AGE14 = [
    ["65～69歳", 2289, 2062, 1673, 1784, 1884, 2082, 1778, 1548],
    ["70～74歳", 1843, 2208, 1949, 1589, 1701, 1803, 1998, 1713],
    ["75～79歳", 1700, 1702, 2044, 1815, 1482, 1592, 1695, 1884],
    ["80～84歳", 1496, 1488, 1490, 1843, 1638, 1346, 1455, 1560],
    ["85～89歳", 998, 1157, 1164, 1182, 1512, 1339, 1110, 1212],
    ["90歳以上", 629, 820, 1003, 1097, 1161, 1471, 1486, 1322],
    ["合計", None, None, None, None, None, None, None, None],
]
r = 4
ws.cell(row=r, column=1, value="（1）高齢者の年齢構成（5歳階級別）").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分"] + YR14, AGE14, numfmt="#,##0")
for c in range(2, 10):
    col = get_column_letter(c)
    ws.cell(row=end, column=c).value = f"=SUM({col}{hrow+1}:{col}{end-1})"
    ws.cell(row=end, column=c).number_format = "#,##0"
    ws.cell(row=end, column=c).font = Font(name=FONT, size=9, bold=True)
cats = Reference(ws, min_col=2, max_col=9, min_row=hrow)
data = Reference(ws, min_col=1, max_col=9, min_row=hrow + 1, max_row=end - 1)
mono_bar(ws, "高齢者の年齢構成（5歳階級別）の推移", "人口（人）", cats, data, "L4",
         stacked=True, width=22, height=12, from_rows=True)
r = end + 2

ws.cell(row=r, column=1, value="（2）85歳以上人口・前期後期別高齢者数").font = Font(name=FONT, size=10, bold=True)
hrow = r + 1
end = table(ws, hrow, ["区分"] + YR14,
            [["85歳以上（再掲）", None, None, None, None, None, None, None, None],
             ["前期高齢者（65～74歳）", 4132, 4270, 3622, 3373, 3585, 3885, 3776, 3261],
             ["後期高齢者（75歳以上）", 4823, 5167, 5701, 5937, 5793, 5748, 5746, 5978]],
            numfmt="#,##0")
for c in range(2, 10):
    col = get_column_letter(c)
    ws.cell(row=hrow + 1, column=c).value = f"={col}{hrow-5}+{col}{hrow-4}"
    ws.cell(row=hrow + 1, column=c).number_format = "#,##0"
cats = Reference(ws, min_col=2, max_col=9, min_row=hrow)
data = Reference(ws, min_col=1, max_col=9, min_row=hrow + 1, max_row=hrow + 1)
mono_line(ws, "85歳以上人口の推移", "人口（人）", cats, data, "L30", width=24, height=12,
          from_rows=True, labels=True, numfmt="#,##0")
cats2 = Reference(ws, min_col=2, max_col=9, min_row=hrow)
data2 = Reference(ws, min_col=1, max_col=9, min_row=hrow + 2, max_row=hrow + 3)
mono_bar(ws, "前期・後期別高齢者数の推移", "人口（人）", cats2, data2, "L56",
         stacked=True, width=24, height=12, from_rows=True, labels=True, numfmt="#,##0")
note(ws, end + 2,
     "注1）85歳以上は「85～89歳」と「90歳以上」の合計。"
     "2035年2,673人・2040年2,810人・2050年2,534人は社人研推計による値である。"
     "計画素案 第2章第8節の85歳以上（令和17年度2,550人・令和22年度2,681人・"
     "令和32年度2,417人）は、令和8年8月31日のご指示による人口の基礎の変更"
     "（総合戦略ベース）に従い、令和7年度の第1号被保険者数を起点として"
     "推計した値であり、本表とは基準が異なる。"
     "注2）総人口が減少する一方、85歳以上人口は2040年まで増加し2015年比で1.7倍となる。"
     "90歳以上は629人から1,471人へ2.3倍に増加する。"
     "注3）本表は国勢調査・社人研推計ベース。第9期計画の図1・図2（住民基本台帳ベース）とは基準が異なるため接続していない。")

# ============================================================ 15 サービス利用強度
ws = sheet("15_サービス利用強度", "図14　受給者1人あたり利用日数・回数の推移",
           "資料：地域包括ケア「見える化」システム D31-a〜j（取得日 令和8年7月22日）"
           "／R6は令和7年2月、R7は令和8年1月のサービス提供分まで／単位：訪問系は回、通所系は日",
           [22] + [8] * 13)
YR15 = ["H26", "H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
SVC = [
    ("訪問介護（回）", [22.1, 22.7, 25.5, 32.2, 38.8, 39.7, 42.4, 44.7, 45.6, 47.0, 50.9, 55.4],
     [13.4, 14.4, 15.6, 19.6, 23.4, 23.9, 25.2, 25.9, 26.4, 27.3, 28.7, 29.9],
     [14.6, 15.4, 16.8, 20.9, 23.6, 24.0, 25.1, 25.6, 26.2, 27.2, 28.4, 29.7]),
    ("訪問看護（回）", [6.2, 7.2, 8.5, 9.5, 7.2, 11.6, 9.1, 8.5, 7.9, 7.5, 7.8, 8.0],
     [6.2, 6.3, 6.5, 6.7, 6.7, 6.7, 6.7, 6.6, 6.6, 6.7, 6.7, 6.7],
     [8.0, 8.3, 8.5, 8.6, 8.7, 8.7, 8.8, 8.9, 8.9, 9.0, 9.0, 9.1]),
    ("通所介護（日）", [5.2, 5.0, 3.9, 5.4, 8.9, 8.9, 9.7, 9.7, 9.6, 9.6, 9.4, 8.8],
     [5.2, 5.2, 4.6, 6.0, 8.1, 8.2, 8.2, 8.3, 8.0, 8.2, 8.1, 8.1],
     [7.4, 7.6, 7.5, 9.2, 10.6, 10.7, 10.9, 11.0, 10.7, 10.8, 10.7, 10.7]),
    ("通所リハビリテーション（日）", [5.0, 4.7, 4.6, 4.4, 4.5, 4.6, 4.8, 4.8, 4.4, 4.8, 4.6, 5.0],
     [5.4, 5.4, 5.4, 5.3, 5.2, 5.1, 5.0, 5.0, 4.8, 4.9, 4.8, 4.8],
     [6.4, 6.4, 6.3, 6.2, 6.1, 5.9, 5.9, 5.9, 5.7, 5.7, 5.6, 5.6]),
    ("地域密着型通所介護（日）", [None, None, 6.9, 6.8, 6.8, 6.8, 6.6, 6.9, 6.8, 6.9, 7.1, 7.9],
     [None, None, 7.9, 8.0, 7.9, 8.0, 8.1, 8.1, 7.9, 7.9, 7.9, 7.8],
     [None, None, 9.6, 9.6, 9.5, 9.5, 9.7, 9.7, 9.4, 9.4, 9.2, 9.2]),
]
r = 4
anchor_row = 4
for name, o, hk, jp in SVC:
    ws.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
    hrow = r + 1
    end = table(ws, hrow, ["区分"] + YR15,
                [["大雪地区広域連合"] + o, ["北海道"] + hk, ["全国"] + jp], numfmt="0.0")
    cats = Reference(ws, min_col=2, max_col=13, min_row=hrow)
    data = Reference(ws, min_col=1, max_col=13, min_row=hrow + 1, max_row=end)
    mono_line(ws, f"{name}　受給者1人あたり利用日数・回数の推移", name.split("（")[1].rstrip("）"),
              cats, data, f"P{anchor_row}", width=24, height=11, from_rows=True,
              labels=True, numfmt="0.0")
    anchor_row += 21
    r = end + 2
note(ws, r,
     "注1）R7列（令和8年1月サービス提供分まで）の値は第10期計画素案 表5の掲載値と一致する。"
     "注2）訪問介護は平成26年度22.1回から令和7年度55.4回へ2.5倍に増加し、北海道（13.4→29.9回・2.2倍）を上回る伸びで推移している。"
     "令和7年度は北海道の1.85倍であり、需要特性、提供体制、算定・集計範囲の要因分解が必要である。"
     "注3）通所介護・通所リハビリテーション・地域密着型通所介護はいずれも全国を下回って推移している。")

# ============================================================ 16 町別将来推計
YR16 = ["2000年", "2005年", "2010年", "2015年", "2020年", "2025年",
        "2030年", "2035年", "2040年", "2045年", "2050年"]
POP16 = {
    "大雪地区広域連合": [27700, 28523, 28107, 28636, 28109, 27102, 25969, 24778, 23582, 22332, 21058],
    "東川町": [7671, 7701, 7859, 8111, 8314, 8213, 8059, 7849, 7607, 7339, 7088],
    "美瑛町": [11902, 11628, 10956, 10292, 9668, 8893, 8160, 7476, 6851, 6258, 5681],
    "東神楽町": [8127, 9194, 9292, 10233, 10127, 9996, 9750, 9453, 9124, 8735, 8289],
}
RATE16 = {
    "大雪地区広域連合": [23.3, 25.6, 28.3, 31.3, 33.6, 34.4, 35.9, 37.8, 40.8, 42.6, 43.9],
    "東川町": [23.2, 25.5, 28.0, 32.1, 33.2, 32.9, 33.4, 34.6, 37.3, 39.1, 40.1],
    "美瑛町": [26.5, 30.6, 33.5, 36.3, 38.8, 39.4, 40.8, 42.4, 45.1, 46.4, 47.5],
    "東神楽町": [18.6, 19.6, 22.5, 25.6, 28.9, 31.2, 33.7, 36.9, 40.6, 42.9, 44.7],
}
ws = sheet("16_町別将来推計", "図15　町別の総人口・高齢化率の推移と将来推計",
           "資料：地域包括ケア「見える化」システム A1・A2（総務省国勢調査＋国立社会保障・人口問題研究所"
           "「日本の地域別将来推計人口」）／2020年までが実績、2025年以降は推計",
           [18] + [10] * 11)
r = 4
ws.cell(row=r, column=1, value="（1）総人口（単位：人）").font = Font(name=FONT, size=10, bold=True)
h1 = r + 1
e1 = table(ws, h1, ["区分"] + YR16, [[k] + v for k, v in POP16.items()], numfmt="#,##0")
mono_line(ws, "町別総人口の推移と将来推計", "人口（人）",
          Reference(ws, min_col=2, max_col=12, min_row=h1),
          Reference(ws, min_col=1, max_col=12, min_row=h1 + 1, max_row=e1),
          "N4", width=26, height=12, from_rows=True, labels=True, numfmt="#,##0")
r = e1 + 2
ws.cell(row=r, column=1, value="（2）高齢化率（単位：％）").font = Font(name=FONT, size=10, bold=True)
h2 = r + 1
e2 = table(ws, h2, ["区分"] + YR16, [[k] + v for k, v in RATE16.items()], numfmt="0.0")
mono_line(ws, "町別高齢化率の推移と将来推計", "高齢化率（％）",
          Reference(ws, min_col=2, max_col=12, min_row=h2),
          Reference(ws, min_col=1, max_col=12, min_row=h2 + 1, max_row=e2),
          "N28", width=26, height=12, from_rows=True, labels=True, numfmt="0.0", min_=15, max_=52)
r = e2 + 2
ws.cell(row=r, column=1, value="（3）2020年から2050年の変化").font = Font(name=FONT, size=10, bold=True)
h3 = r + 1
e3 = table(ws, h3, ["区分", "総人口2020年", "総人口2050年", "増減数", "増減率（％）",
                    "高齢化率2020年", "高齢化率2050年", "増減（ポイント）"],
           [[k, POP16[k][4], POP16[k][10], None, None, RATE16[k][4], RATE16[k][10], None]
            for k in POP16], numfmt="#,##0")
for rr in range(h3 + 1, e3 + 1):
    ws.cell(row=rr, column=4).value = f"=C{rr}-B{rr}"
    ws.cell(row=rr, column=5).value = f"=(C{rr}-B{rr})/B{rr}*100"
    ws.cell(row=rr, column=8).value = f"=G{rr}-F{rr}"
    ws.cell(row=rr, column=4).number_format = "#,##0"
    for cc in (5, 6, 7, 8):
        ws.cell(row=rr, column=cc).number_format = "0.0"
note(ws, e3 + 2,
     "注1）2020年までが国勢調査の実績、2025年以降は国立社会保障・人口問題研究所の推計。"
     "第9期計画 第2章第1節1の図1・図2は住民基本台帳ベースであり、基準が異なるため接続していない（修正指示書C-7）。"
     "注2）2020年から2050年にかけて、総人口は美瑛町が41.2％減と最も大きく減少する一方、東川町は14.8％減にとどまる。"
     "高齢化率は東神楽町が15.8ポイント上昇と最も伸びが大きく、2050年には美瑛町47.5％・東神楽町44.7％・東川町40.1％となる。"
     "注3）東神楽町は2000年の高齢化率18.6％から2050年44.7％へ26.1ポイント上昇し、3町で最も変化が急である。")

# ============================================================ 17 担い手の推移
RATIO17 = {
    "大雪地区広域連合": [2.7, 2.3, 2.0, 1.8, 1.6, 1.6, 1.5, 1.4, 1.2, 1.1, 1.0],
    "東川町": [2.7, 2.4, 2.1, 1.7, 1.6, 1.7, 1.6, 1.6, 1.4, 1.3, 1.2],
    "美瑛町": [2.3, 1.9, 1.6, 1.5, 1.3, 1.3, 1.3, 1.2, 1.0, 1.0, 0.9],
    "東神楽町": [3.4, 3.2, 2.7, 2.3, 1.9, 1.8, 1.6, 1.4, 1.2, 1.1, 1.0],
}
ws = sheet("17_担い手の推移", "図16　高齢者1人あたり現役世代数の推移と将来推計",
           "資料：地域包括ケア「見える化」システム A9（15〜64歳人口÷65歳以上人口）／"
           "2020年までが実績、2025年以降は推計／単位：人",
           [18] + [10] * 11)
r = 4
h1 = r
e1 = table(ws, h1, ["区分"] + YR16, [[k] + v for k, v in RATIO17.items()], numfmt="0.0")
mono_line(ws, "高齢者1人あたり現役世代数の推移と将来推計", "現役世代数（人）",
          Reference(ws, min_col=2, max_col=12, min_row=h1),
          Reference(ws, min_col=1, max_col=12, min_row=h1 + 1, max_row=e1),
          "N4", width=26, height=12, from_rows=True, labels=True, numfmt="0.0", min_=0.5, max_=3.6)
note(ws, e1 + 2,
     "注1）高齢者1人を何人の現役世代（15〜64歳）で支えるかを示す。値が小さいほど支え手が少ない。"
     "注2）広域連合全体では2000年2.7人から2040年1.2人、2050年1.0人へ低下する。"
     "美瑛町は2040年に1.0人、2050年には0.9人となり、現役世代より高齢者が多い状態になる。"
     "注3）東神楽町は2000年3.4人と3町で最も余裕があったが、2050年1.0人まで低下し、低下幅（2.4人）が3町で最も大きい。"
     "注4）第10期計画では、サービス見込量を需要側だけで決めず、この担い手の減少を前提とした"
     "供給可能量・人材確保・共同化・代替提供の検討が必要となる（素案 第1章第3節「実行可能性」の視点）。")


# ============================================================ 18 受給率・利用率
YR18 = ["H26", "H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
ws = sheet("18_受給率・利用率", "図17　介護サービスの利用率・受給率",
           "資料：地域包括ケア「見える化」システム D38・D41-a・D42（取得日 令和8年7月）／単位：％",
           [22] + [9] * 12)
r = 4
ws.cell(row=r, column=1, value="（1）介護サービス利用率（認定者に占める受給者の割合）").font = Font(name=FONT, size=10, bold=True)
h1 = r + 1
e1 = table(ws, h1, ["区分"] + YR18[:11],
           [["介護サービス利用率", 76.9, 80.0, 82.2, 77.8, 76.1, 76.0, 77.6, 78.4, 75.9, 75.3, 74.3],
            ["合計認定者数（人）", 1835, 1869, 1842, 1882, 1859, 1901, 1878, 1903, 1921, 1931, 1962],
            ["合計受給者数（人）", 1411, 1496, 1515, 1465, 1414, 1444, 1457, 1492, 1458, 1455, 1458]],
           numfmt="#,##0.0")
mono_line(ws, "介護サービス利用率の推移（大雪地区広域連合）", "利用率（％）",
          Reference(ws, min_col=2, max_col=12, min_row=h1),
          Reference(ws, min_col=1, max_col=12, min_row=h1 + 1, max_row=h1 + 1),
          "P4", width=24, height=11, from_rows=True, labels=True, numfmt="0.0", min_=70, max_=85)
r = e1 + 2
ws.cell(row=r, column=1, value="（2）在宅・居住系サービス利用者割合（受給者に占める割合）").font = Font(name=FONT, size=10, bold=True)
h2 = r + 1
e2 = table(ws, h2, ["区分"] + YR18,
           [["在宅・居住系サービス利用者割合", 73.3, 74.2, 75.0, 74.7, 74.3, 74.8, 75.1, 76.1, 76.8, 76.1, 76.0, 77.5]],
           numfmt="0.0")
mono_line(ws, "在宅・居住系サービス利用者割合の推移", "割合（％）",
          Reference(ws, min_col=2, max_col=13, min_row=h2),
          Reference(ws, min_col=1, max_col=13, min_row=h2 + 1, max_row=e2),
          "P26", width=24, height=11, from_rows=True, labels=True, numfmt="0.0", min_=70, max_=80)
r = e2 + 2
ws.cell(row=r, column=1, value="（3）施設および居住系サービスの受給率（第1号被保険者に占める割合・年度平均）").font = Font(name=FONT, size=10, bold=True)
h3 = r + 1
e3 = table(ws, h3, ["区分"] + YR18,
           [["大雪地区広域連合", 6.0, 6.1, 6.0, 5.8, 5.6, 5.7, 5.8, 5.7, 5.6, 5.3, 5.4, 5.3],
            ["北海道", 4.7, 4.6, 4.5, 4.5, 4.5, 4.5, 4.5, 4.4, 4.4, 4.4, 4.4, 4.5],
            ["全国", 4.1, 4.0, 4.0, 4.1, 4.1, 4.1, 4.2, 4.2, 4.2, 4.2, 4.2, 4.3]],
           numfmt="0.0")
mono_line(ws, "施設・居住系サービス受給率の推移（全国・北海道との比較）", "受給率（％）",
          Reference(ws, min_col=2, max_col=13, min_row=h3),
          Reference(ws, min_col=1, max_col=13, min_row=h3 + 1, max_row=e3),
          "P48", width=24, height=11, from_rows=True, labels=True, numfmt="0.0", min_=3.5, max_=6.5)
note(ws, e3 + 2,
     "注1）（1）介護サービス利用率は令和6年度74.3％で、平成28年度の82.2％から7.9ポイント低下している。"
     "認定者は1,842人から1,962人へ増えた一方、受給者は1,515人から1,458人へ減っており、"
     "認定を受けても利用に至らない層が拡大している可能性がある。需要減か供給制約かの判別が必要である（修正指示書A-7）。"
     "注2）（3）施設・居住系サービスの受給率は令和7年度5.3％で、北海道4.5％の1.18倍、全国4.3％の1.23倍。"
     "平成27年度の6.1％からは低下しているが、依然として全国水準を上回る。"
     "注3）（3）は月次データの年度平均。原データは月別のため、年度の取り方により小数第1位が変動する。")

# ============================================================ 19 給付月額の比較
YR19 = ["H26", "H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
ws = sheet("19_給付月額の比較", "図18　給付月額の比較（全国・北海道との対比）",
           "資料：地域包括ケア「見える化」システム D6-a・D6-b・D17-a〜k（取得日 令和8年7月）／単位：円",
           [26] + [10] * 12)
r = 4
ws.cell(row=r, column=1, value="（1）第1号被保険者1人あたり給付月額（在宅サービス）").font = Font(name=FONT, size=10, bold=True)
h1 = r + 1
e1 = table(ws, h1, ["区分"] + YR19,
           [["大雪地区広域連合", 9620, 9699, 10051, 10258, 9840, 10068, 9991, 10743, 11020, 11342, 11429, 12558],
            ["北海道", 8869, 9083, 9298, 9284, 9051, 9407, 9592, 9948, 10164, 10618, 10966, 11276],
            ["全国", 11225, 11282, 11295, 11320, 11262, 11531, 11712, 12107, 12308, 12778, 13183, 13635]],
           numfmt="#,##0")
mono_line(ws, "第1号被保険者1人あたり給付月額（在宅サービス）", "給付月額（円）",
          Reference(ws, min_col=2, max_col=13, min_row=h1),
          Reference(ws, min_col=1, max_col=13, min_row=h1 + 1, max_row=e1),
          "P4", width=24, height=11, from_rows=True, labels=True, numfmt="#,##0")
r = e1 + 2
ws.cell(row=r, column=1, value="（2）第1号被保険者1人あたり給付月額（施設および居住系サービス）").font = Font(name=FONT, size=10, bold=True)
h2 = r + 1
e2 = table(ws, h2, ["区分"] + YR19,
           [["大雪地区広域連合", 14683, 14485, 14121, 13975, 14104, 14201, 14647, 14440, 13688, 13938, 14218, 14327],
            ["北海道", 11105, 10737, 10514, 10562, 10688, 10897, 11069, 11161, 11195, 11370, 11690, 11904],
            ["全国", 9920, 9779, 9709, 9912, 10151, 10393, 10633, 10758, 10863, 11082, 11502, 11773]],
           numfmt="#,##0")
mono_line(ws, "第1号被保険者1人あたり給付月額（施設および居住系サービス）", "給付月額（円）",
          Reference(ws, min_col=2, max_col=13, min_row=h2),
          Reference(ws, min_col=1, max_col=13, min_row=h2 + 1, max_row=e2),
          "P26", width=24, height=11, from_rows=True, labels=True, numfmt="#,##0")
r = e2 + 2
ws.cell(row=r, column=1, value="（3）受給者1人あたり給付月額（令和7年度・主要サービス）").font = Font(name=FONT, size=10, bold=True)
h3 = r + 1
e3 = table(ws, h3, ["サービス", "大雪地区広域連合", "北海道", "全国", "大雪／北海道", "大雪／全国"],
           [["訪問介護", 133685, 84606, 86541, None, None],
            ["訪問看護", 37385, 36323, 41701, None, None],
            ["通所介護", 62688, 58316, 84875, None, None],
            ["特定施設入居者生活介護", 173249, 174488, 191986, None, None]], numfmt="#,##0")
for rr in range(h3 + 1, e3 + 1):
    ws.cell(row=rr, column=5).value = f"=B{rr}/C{rr}"
    ws.cell(row=rr, column=6).value = f"=B{rr}/D{rr}"
    for cc in (5, 6):
        ws.cell(row=rr, column=cc).number_format = "0.00"
        ws.cell(row=rr, column=cc).font = Font(name=FONT, size=9, bold=True)
mono_bar(ws, "受給者1人あたり給付月額の比較（令和7年度）", "給付月額（円）",
         Reference(ws, min_col=1, min_row=h3 + 1, max_row=e3),
         Reference(ws, min_col=2, max_col=4, min_row=h3, max_row=e3),
         "P48", width=24, height=11, labels=True, numfmt="#,##0", gap=80)
note(ws, e3 + 2,
     "注1）在宅サービスの1人あたり給付月額は令和7年度12,558円で、北海道11,276円を上回るが全国13,635円は下回る。"
     "注2）施設・居住系サービスは14,327円で、北海道11,904円の1.20倍、全国11,773円の1.22倍と一貫して高い。"
     "在宅と施設・居住系の合計26,885円は計画素案 第2章第1節の掲載値と一致する。"
     "注3）（3）訪問介護は受給者1人あたり133,685円で、北海道の1.58倍、全国の1.54倍。"
     "受給者1人あたり利用回数も北海道の1.85倍（55.4回対29.9回・15シート参照）であり、"
     "利用量と単価の双方が突出している。需要特性、提供体制、算定・集計範囲の要因分解が必要である（素案 表5の重点論点）。"
     "注4）通所介護は62,688円で全国84,875円の0.74倍と低く、サービス種別によって全国との位置関係が逆転する。")


# ============================================================ 20 第9期の達成状況
ws = sheet("20_第9期の達成状況", "図19　第9期計画の達成状況（代表KPI・給付費）",
           "資料：第9期計画 第1章第6節（数値目標）・第6章／見える化B4-a・D48-b・D48-c／令和7年度 健康とくらしの調査"
           "／令和6〜8年度の事業実績は未受領のため淡黄色欄",
           [28, 13, 13, 13, 13, 13, 20])
r = 4
ws.cell(row=r, column=1, value="（1）代表KPI4項目の目標と実績").font = Font(name=FONT, size=10, bold=True)
h1 = r + 1
KPI9 = [
    ["① 要介護認定率（未調整）", 20.8, 20.8, 20.8, 20.8, None, "実績は3月末基準のため別掲（下表）"],
    ["② 重度要介護認定率（要介護3以上）", 6.7, 6.6, 6.5, 6.4, None, "実績6.2〜6.3％で目標を下回る"],
    ["③ フレイル該当割合", 18.5, None, 18.0, None, 19.1, "令和7年度調査19.1％で目標未達"],
    ["④ 通いの場参加率（月1回以上）", 7.3, 8.0, 9.0, 10.0, 8.8, "令和7年度調査8.8％。R7目標9.0％に0.2pt不足"],
]
e1 = table(ws, h1, ["指標", "基準値", "R6目標", "R7目標", "R8目標", "直近実績", "評価"], KPI9, numfmt="0.0")
r = e1 + 2

ws.cell(row=r, column=1, value="（2）認定率の実績（見える化・各年3月末基準）").font = Font(name=FONT, size=10, bold=True)
h2 = r + 1
e2 = table(ws, h2, ["区分", "R4年3月末", "R5年3月末", "R6年3月末", "R7年3月末", "R8年3月末", "第9期目標"],
           [["合計認定率", 20.6, 20.9, 21.2, 21.4, 21.8, 20.8],
            ["要介護3以上（重度）", 6.3, 6.3, 6.2, 6.3, 6.3, 6.4],
            ["第1号被保険者数（人）", 9245, 9211, 9130, 9158, 9090, None],
            ["合計認定者数（人）", 1903, 1921, 1931, 1962, 1984, None]], numfmt="#,##0.0")
mono_line(ws, "認定率の推移と第9期目標", "認定率（％）",
          Reference(ws, min_col=2, max_col=6, min_row=h2),
          Reference(ws, min_col=1, max_col=6, min_row=h2 + 1, max_row=h2 + 2),
          "I4", width=22, height=11, from_rows=True, labels=True, numfmt="0.0", min_=5, max_=23)
r = e2 + 2

ws.cell(row=r, column=1, value="（3）給付費の計画値と実績（単位：千円）").font = Font(name=FONT, size=10, bold=True)
h3 = r + 1
e3 = table(ws, h3, ["区分", "R6計画", "R6実績", "R7計画", "R7実績", "R8計画", "R8実績"],
           [["標準給付費見込額", 3127780, None, 3184593, None, 3206838, None],
            ["総給付費（保険給付費）", 2924324, None, 2979919, None, 3000991, None],
            ["地域支援事業費", 197886, None, 197886, None, 197886, None],
            ["合計（標準給付費＋地域支援事業費）", None, None, None, None, None, None]], numfmt="#,##0")
for c in (2, 3, 4, 5, 6, 7):
    col = get_column_letter(c)
    ws.cell(row=e3, column=c).value = f"={col}{h3+1}+{col}{h3+3}"
    ws.cell(row=e3, column=c).number_format = "#,##0"
    ws.cell(row=e3, column=c).font = Font(name=FONT, size=9, bold=True)
for rr in range(h3 + 1, e3):
    for c in (3, 5, 7):
        ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=IN_Y)
r = e3 + 2

ws.cell(row=r, column=1, value="（4）参考：保険給付費・地域支援事業費の決算推移（単位：千円）").font = Font(name=FONT, size=10, bold=True)
h4 = r + 1
e4 = table(ws, h4, ["区分", "R元年度", "R2年度", "R3年度", "R4年度", "R5年度", "R6年度以降"],
           [["保険給付費（決算）", 2886092, 2941727, 2992042, 2907204, 2940409, None],
            ["地域支援事業費（決算）", 176434, 183357, 177286, 182136, 177517, None]], numfmt="#,##0")
for rr in range(h4 + 1, e4 + 1):
    ws.cell(row=rr, column=7).fill = PatternFill("solid", fgColor=IN_Y)
mono_bar(ws, "保険給付費・地域支援事業費の決算推移", "決算額（千円）",
         Reference(ws, min_col=2, max_col=6, min_row=h4),
         Reference(ws, min_col=1, max_col=6, min_row=h4 + 1, max_row=e4),
         "I26", width=22, height=11, from_rows=True, labels=True, numfmt="#,##0", gap=80)
note(ws, e4 + 2,
     "注1）①要介護認定率は第9期計画が各年9月分を基準としているのに対し、見える化の公表値は各年3月末基準であるため、"
     "（2）に別掲した。令和8年3月末の21.8％は目標20.8％を1.0ポイント上回っており未達である。"
     "注2）②重度要介護認定率は見える化の要介護3〜5の認定率を合計した値で、令和8年3月末6.3％。"
     "第9期目標（R8 6.4％）を下回っており達成の見込みである。"
     "注3）③フレイル該当割合は令和7年度 健康とくらしの調査で19.1％。第9期目標（R7 18.0％）に対し1.1ポイント未達。"
     "ただし基準値18.5％との差0.6ポイントは統計的に有意ではない（08シート参照）。"
     "注4）④通いの場参加率は同調査で8.8％。基準値7.3％から1.5ポイント上昇し、この変化は統計的に有意である。"
     "R7目標9.0％には0.2ポイント届かず、R8目標10.0％の達成は難しい見通しである。"
     "注5）（3）の実績欄と（4）のR6年度以降は、令和6〜8年度の決算・事業実績の受領後に入力する（修正指示書 最優先項目）。"
     "注6）（4）の決算値は見える化D48-b・D48-cによる。第9期計画 第2章第2節の給付費（見える化の給付実績ベース）とは"
     "集計範囲が異なるため直接比較しない。")




# ============================================================ 21 受給率の内訳と定員
ws = sheet("21_受給率と定員", "図20　受給率の内訳と要支援・要介護者1人あたり定員",
           "出典：見える化D41（令和7年・地域別）、D41-a（平成26年4月〜令和8年1月の月次を年度平均に集計）、"
           "D28・D29・D30（介護サービス情報公表システム及び月報）。受給率の分母は第1号被保険者数。",
           [26, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10])

ws["A4"] = "（1）受給率の内訳（令和7年・％。分母は第1号被保険者数）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
e1 = table(ws, 5, ["区分", "全国", "北海道", "大雪地区広域連合"],
           [["在宅サービス", 11.0, 10.3, 11.2],
            ["施設および居住系サービス", 4.3, 4.5, 5.3]], numfmt="0.0")
mono_bar(ws, "受給率の内訳（令和7年）", "受給率（％）",
         Reference(ws, min_col=1, min_row=6, max_row=e1),
         Reference(ws, min_col=2, max_col=4, min_row=5, max_row=e1),
         "P4", width=15, height=9, labels=True, numfmt="0.0", gap=70)
note(ws, e1 + 1,
     "注）大雪の合計16.5％は全国15.3％の1.078倍で、見える化D49の地域差指数「受給率1.08」を再現する。"
     "在宅は全国比1.02とほぼ同水準で、超過分のほとんどは施設及び居住系の差である。")

r0 = e1 + 3
ws.cell(row=r0, column=1, value="（2）施設および居住系サービス受給率の推移（年度平均・％）").font = \
    Font(name=FONT, size=10, bold=True)
YY = ["H26", "H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
DAI = [6.02, 6.10, 5.92, 5.72, 5.66, 5.72, 5.72, 5.60, 5.28, 5.39, 5.31, 5.25]
HOK = [4.66, 4.57, 4.50, 4.47, 4.49, 4.48, 4.47, 4.44, 4.41, 4.44, 4.46, 4.49]
ZEN = [4.04, 4.04, 4.03, 4.08, 4.10, 4.15, 4.19, 4.18, 4.19, 4.20, 4.26, 4.30]
e2 = table(ws, r0 + 1, ["区分"] + YY,
           [["大雪地区広域連合"] + DAI, ["北海道"] + HOK, ["全国"] + ZEN], numfmt="0.00")
mono_line(ws, "施設および居住系サービス受給率の推移", "受給率（％）",
          Reference(ws, min_col=2, max_col=13, min_row=r0 + 1, max_row=r0 + 1),
          Reference(ws, min_col=1, max_col=13, min_row=r0 + 2, max_row=e2),
          "P22", width=21, height=11, from_rows=True, labels=True, numfmt="0.00",
          min_=3.5, max_=6.5)
note(ws, e2 + 1,
     "注）大雪は平成27年度6.10％をピークに0.85ポイント低下し、全国は平成28年度4.03％から0.27ポイント上昇した。"
     "全国との差は平成26年度1.98ポイントから令和7年度0.95ポイントへ半減している。"
     "令和7年度は令和8年1月サービス提供分までの10か月平均。")

r1 = e2 + 3
ws.cell(row=r1, column=1, value="（3）要支援・要介護者1人あたり定員の推移（人）").font = \
    Font(name=FONT, size=10, bold=True)
CY = ["H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
CAP = [
    ["介護老人福祉施設", 0.104, 0.094, 0.124, 0.126, 0.123, 0.085, 0.084, 0.083, 0.083, 0.082, 0.082],
    ["介護老人保健施設", 0.128, 0.130, 0.128, 0.129, 0.126, 0.128, 0.126, 0.125, 0.124, 0.122, 0.123],
    ["地域密着型介護老人福祉施設", None, None, None, 0.033, 0.033, 0.033, 0.033, 0.032, 0.032, 0.032, 0.032],
    ["特定施設入居者生活介護", 0.083, 0.085, 0.083, 0.084, 0.082, 0.083, 0.082, 0.082, 0.081, 0.080, 0.080],
    ["認知症対応型共同生活介護", 0.051, 0.062, 0.061, 0.062, 0.060, 0.061, 0.061, 0.052, 0.051, 0.050, 0.051],
]
e3 = table(ws, r1 + 1, ["サービス"] + CY, CAP, numfmt="0.000")
mono_line(ws, "要支援・要介護者1人あたり定員の推移", "定員（人）",
          Reference(ws, min_col=2, max_col=12, min_row=r1 + 1, max_row=r1 + 1),
          Reference(ws, min_col=1, max_col=12, min_row=r1 + 2, max_row=e3),
          "P46", width=21, height=11, from_rows=True, min_=0.0, max_=0.14)
e3b = table(ws, e3 + 2, ["区分"] + CY,
            [["施設サービス計", 0.232, 0.225, 0.252, 0.288, 0.282, 0.246, 0.243, 0.240, 0.239, 0.235, 0.237],
             ["居住系サービス計", 0.135, 0.147, 0.144, 0.146, 0.143, 0.144, 0.143, 0.134, 0.132, 0.130, 0.131],
             ["通所系サービス計", 0.140, 0.156, 0.147, 0.148, 0.145, 0.141, 0.141, 0.150, 0.150, 0.143, 0.144]],
            numfmt="0.000")
mono_line(ws, "要支援・要介護者1人あたり定員の推移（区分計）", "定員（人）",
          Reference(ws, min_col=2, max_col=12, min_row=e3 + 2, max_row=e3 + 2),
          Reference(ws, min_col=1, max_col=12, min_row=e3 + 3, max_row=e3b),
          "P70", width=21, height=11, from_rows=True, labels=True, numfmt="0.000",
          min_=0.10, max_=0.32)
r4 = e3b + 2
ws.cell(row=r4, column=1, value="（4）要介護度別の受給率の推移（％。分母は第1号被保険者数）").font = \
    Font(name=FONT, size=10, bold=True)
YD = ["H26", "H27", "H28", "H29", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"]
e4 = table(ws, r4 + 1, ["施設サービス"] + YD,
           [["要介護1", 0.4, 0.4, 0.4, 0.3, 0.4, 0.3, 0.3, 0.4, 0.3, 0.5, 0.5, 0.4],
            ["要介護2", 0.7, 0.8, 0.7, 0.7, 0.7, 0.7, 0.7, 0.6, 0.6, 0.7, 0.8, 0.7],
            ["要介護3", 0.9, 1.0, 1.0, 1.0, 0.9, 1.0, 1.1, 1.1, 1.0, 0.9, 0.9, 0.9],
            ["要介護4", 1.1, 1.2, 1.1, 1.1, 1.0, 1.1, 1.2, 1.1, 1.1, 1.0, 1.0, 1.1],
            ["要介護5", 1.2, 1.1, 1.0, 1.0, 1.0, 0.9, 0.7, 0.7, 0.7, 0.7, 0.6, 0.7]], numfmt="0.0")
mono_line(ws, "受給率の推移（施設サービス・要介護度別）", "受給率（％）",
          Reference(ws, min_col=2, max_col=13, min_row=r4 + 1, max_row=r4 + 1),
          Reference(ws, min_col=1, max_col=13, min_row=r4 + 2, max_row=e4),
          "P94", width=21, height=11, from_rows=True, min_=0.0, max_=1.4)
e5 = table(ws, e4 + 2, ["在宅サービス"] + YD,
           [["要支援1", 1.6, 2.0, 2.0, 1.5, 1.1, 1.1, 1.1, 1.1, 1.0, 1.0, 1.0, 1.2],
            ["要支援2", 2.1, 2.2, 2.1, 1.8, 1.5, 1.6, 1.7, 1.8, 1.7, 1.6, 1.8, 1.8],
            ["要介護1", 3.0, 3.1, 3.2, 3.2, 3.3, 3.3, 3.3, 3.5, 3.4, 3.4, 3.5, 3.7],
            ["要介護2", 2.1, 2.0, 2.0, 2.1, 2.2, 2.1, 2.2, 2.3, 2.3, 2.3, 2.1, 2.1],
            ["要介護3", 0.9, 0.8, 1.0, 1.2, 1.0, 1.1, 0.9, 0.8, 1.0, 1.2, 1.1, 1.0],
            ["要介護4", 0.5, 0.5, 0.5, 0.5, 0.4, 0.5, 0.5, 0.7, 0.7, 0.7, 0.7, 0.8],
            ["要介護5", 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.5, 0.4, 0.4, 0.6]], numfmt="0.0")
mono_line(ws, "受給率の推移（在宅サービス・要介護度別）", "受給率（％）",
          Reference(ws, min_col=2, max_col=13, min_row=e4 + 2, max_row=e4 + 2),
          Reference(ws, min_col=1, max_col=13, min_row=e4 + 3, max_row=e5),
          "P118", width=21, height=11, from_rows=True, min_=0.0, max_=4.0)
note(ws, e5 + 1,
     "注1）施設サービスの受給率の低下は要介護5に集中している（1.2％→0.7％）。要介護3・4は横ばい。"
     "注2）在宅サービスは要介護5が0.3％→0.6％、要介護4が0.5％→0.8％と重度者で上昇している。"
     "施設の要介護5の減少と在宅の要介護4・5の増加が対応しており、重度者の施設から在宅への移行を示す。"
     "注3）要支援1・2の低下は、平成29年度から平成30年度にかけての介護予防訪問介護・介護予防通所介護の"
     "介護予防・日常生活支援総合事業への移行による。")

r5 = e5 + 3
ws.cell(row=r5, column=1, value="（5）年齢層別にみた認定率の推移（各年3月末・％）").font = \
    Font(name=FONT, size=10, bold=True)
YA = ["H30", "R2", "R5", "R6", "R7", "R8"]
e6 = table(ws, r5 + 1, ["区分"] + YA,
           [["合計認定率（第1号被保険者）", 20.8, 20.7, 20.9, 21.2, 21.4, 21.8],
            ["合計認定率（75歳以上）", 35.3, 34.8, 33.9, 33.7, 33.4, 33.5],
            ["合計認定率（85歳以上）", None, 63.7, 60.1, 60.6, 61.1, 61.6]], numfmt="0.0")
mono_line(ws, "年齢層別にみた認定率の推移", "認定率（％）",
          Reference(ws, min_col=2, max_col=7, min_row=r5 + 1, max_row=r5 + 1),
          Reference(ws, min_col=1, max_col=7, min_row=r5 + 2, max_row=e6),
          "P142", width=21, height=11, from_rows=True, labels=True, numfmt="0.0",
          min_=15.0, max_=70.0)
note(ws, e6 + 1,
     "注）全体の認定率が平成30年3月末20.8％から令和8年3月末21.8％へ1.0ポイント上昇する一方、"
     "75歳以上は35.3％から33.5％へ1.8ポイント、85歳以上は令和2年3月末63.7％から61.6％へ2.1ポイント低下している。"
     "全体の上昇は第1号被保険者に占める75歳以上・85歳以上の構成比の高まりによるものであり、"
     "同一年齢層でみれば認定率は改善している。出典：見える化B4-a・B4-d・B4-e。")

note(ws, e3b + 1,
     "注1）定員そのものが変わらなくても認定者が増えれば低下する。実数の推移は06シート（定員）を参照。"
     "注2）施設サービス計は平成30年度0.288人から令和7年度0.237人へ17.7％減、うち介護老人福祉施設は34.9％減。"
     "居住系サービス計は10.3％減。"
     "注3）介護療養型医療施設・介護医療院・地域密着型特定施設入居者生活介護は当圏域に事業所がない。")


# ============================================================ 22 サービス別の給付動向
ws = sheet("22_サービス別給付動向", "図21　サービス種類別　第1号被保険者1人あたり給付月額の推移",
           "出典：見える化D13-a〜z（時系列）及びD13（令和7年・サービス種類別）。"
           "令和6年度は令和7年2月、令和7年度は令和8年1月のサービス提供分までの暫定値。",
           [24, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11])

SY = ["H26", "H30", "R2", "R5", "R6", "R7"]
SVC = [
    ["訪問介護", 2317, 2602, 2706, 3287, 3421, 3976],
    ["訪問看護", 314, 328, 463, 451, 431, 555],
    ["通所リハビリテーション", 1378, 1209, 1328, 1425, 1578, 1676],
    ["小規模多機能型居宅介護", 1549, 1802, 1731, 1989, 1996, 2155],
    ["福祉用具貸与", 429, 516, 518, 662, 653, 704],
    ["介護予防支援・居宅介護支援", 899, 1004, 1023, 1163, 1168, 1229],
    ["通所介護", 1619, 694, 697, 669, 678, 675],
    ["地域密着型通所介護", None, 777, 623, 719, 538, 567],
    ["介護老人福祉施設", 5094, 3952, 3943, 4039, 4086, 4105],
    ["介護老人保健施設", 4562, 4547, 4672, 4359, 4762, 4440],
    ["地域密着型介護老人福祉施設", 674, 1662, 1715, 1712, 1777, 1890],
    ["特定施設入居者生活介護", 1032, 985, 1128, 1139, 1061, 1188],
    ["認知症対応型共同生活介護", 2549, 2750, 2947, 2445, 2288, 2461],
    ["定期巡回・随時対応型訪問介護看護", 0, 4, 33, 0, 0, 15],
]
ws["A4"] = "（1）サービス種類別の推移（円／月）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
e1 = table(ws, 5, ["サービス"] + SY, SVC, numfmt="#,##0")
note(ws, e1 + 1,
     "注）平成28年度の制度改正で定員18人以下の通所介護が地域密着型通所介護へ移行したため、"
     "通所介護と地域密着型通所介護は合算して読む（平成30年度1,471円→令和7年度1,242円で15.6％減）。")

r0 = e1 + 3
ws.cell(row=r0, column=1, value="（2）平成30年度から令和7年度の増減率（％）").font = \
    Font(name=FONT, size=10, bold=True)
# 平成30年度の水準が100円未満のサービス（定期巡回等）は、増減率が実態を表さないため除外する
CH = sorted([[x[0], round((x[6] / x[2] - 1) * 100, 1)] for x in SVC
             if x[2] is not None and x[2] >= 100], key=lambda v: -v[1])
e2 = table(ws, r0 + 1, ["サービス", "増減率（％）"], CH, numfmt="0.0")
mono_hbar(ws, "平成30年度→令和7年度の増減率", 
          Reference(ws, min_col=1, min_row=r0 + 2, max_row=e2),
          Reference(ws, min_col=2, min_row=r0 + 1, max_row=e2),
          "P4", width=18, height=13)
note(ws, e2 + 1,
     "注1）訪問看護＋69.2％、訪問介護＋52.8％、通所リハビリテーション＋38.6％、福祉用具貸与＋36.4％と"
     "訪問系・用具系が伸び、地域密着型通所介護▲27.0％、認知症対応型共同生活介護▲10.5％と"
     "通所系・居住系が縮小している。"
     "注2）定期巡回・随時対応型訪問介護看護は平成30年度4円・令和7年度15円と実質的に未整備であり、"
     "増減率は掲載していない。夜間対応型訪問介護と看護小規模多機能型居宅介護は全期間0円。"
     "注3）要介護5の在宅・居住系割合が54.6％に達するなかで24時間対応サービスがほぼ存在しないことは、"
     "第6章の整備方針の主要論点となる（修正指示書A-13）。")

r1 = e2 + 3
ws.cell(row=r1, column=1, value="（3）受給者1人あたりの比較（令和7年・全国／北海道との対比）").font = \
    Font(name=FONT, size=10, bold=True)
e3 = table(ws, r1 + 1, ["指標", "大雪地区広域連合", "北海道", "全国", "全国比"],
           [["訪問介護　受給者1人あたり利用回数（回／月）", 55.4, 29.9, 29.7, 1.87],
            ["訪問介護　受給者1人あたり給付月額（円）", 133685, 84606, 86541, 1.54],
            ["通所介護　受給者1人あたり給付月額（円）", 62688, 58316, 84875, 0.74]],
           numfmt="#,##0.00")
note(ws, e3 + 1,
     "注）訪問介護は利用回数が全国の1.87倍、給付月額が1.54倍と突出する一方、"
     "通所介護の給付月額は全国の0.74倍にとどまる。受給率の超過が施設・居住系に集中する（21シート）のに対し、"
     "受給者1人あたり単価の超過は訪問介護に集中しており、要因が分かれている。")



# ============================================================ 23 上川中部圏域との比較
ws = sheet("23_上川中部圏域比較", "図22　上川中部圏域における大雪地区広域連合の位置",
           "出典：第9期北海道高齢者保健福祉計画・介護保険事業支援計画 第7章第12節（令和2年国勢調査、"
           "地域包括ケア「見える化」システム）／見える化B6・D8。"
           "圏域は旭川市・鷹栖町・東神楽町・当麻町・比布町・愛別町・上川町・東川町・美瑛町・幌加内町の10市町村。",
           [16, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）圏域内市町村の人口構造（令和2年国勢調査）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
TOWN = [
    ["旭川市", 325162, 112411, 34.57, 17.85, 50.89],
    ["鷹栖町", 6567, 2272, 34.60, 18.00, 50.83],
    ["東神楽町", 10112, 2929, 28.97, 15.27, 47.77],
    ["当麻町", 6319, 2659, 42.08, 24.24, 54.79],
    ["比布町", 3520, 1461, 41.51, 24.09, 54.85],
    ["愛別町", 2605, 1206, 46.30, 26.99, 56.80],
    ["上川町", 3500, 1550, 44.29, 25.14, 55.81],
    ["東川町", 8313, 2759, 33.19, 17.92, 49.30],
    ["美瑛町", 9668, 3749, 38.78, 22.06, 53.49],
    ["幌加内町", 1370, 559, 40.80, 25.11, 54.44],
    ["圏域計", 377136, 131555, 34.88, 18.21, 52.90],
]
e1 = table(ws, 5, ["市町村", "総人口", "高齢者人口", "高齢化率（％）", "75歳以上割合（％）", "平均年齢（歳）"],
           TOWN, numfmt="#,##0.00")
mono_bar(ws, "圏域内市町村の高齢化率と75歳以上割合（令和2年）", "割合（％）",
         Reference(ws, min_col=1, min_row=6, max_row=e1),
         Reference(ws, min_col=4, max_col=5, min_row=5, max_row=e1),
         "N4", width=20, height=11, labels=True, numfmt="0.0", gap=60, cat_rot=-45)
note(ws, e1 + 1,
     "注）構成3町の計は28,093人で圏域の7.4％。圏域の86.2％は旭川市が占める。"
     "3町のうち東神楽町の高齢化率28.97％は圏域で最も低く、美瑛町38.78％は圏域平均34.88％を上回る。"
     "平均年齢も東神楽町47.77歳から美瑛町53.49歳まで5.7歳の開きがあり、3町の人口構造は大きく異なる。")

r0 = e1 + 3
ws.cell(row=r0, column=1, value="（2）調整済み第1号被保険者1人あたり給付月額の比較（円）").font = \
    Font(name=FONT, size=10, bold=True)
e2 = table(ws, r0 + 1, ["区分", "在宅サービス", "施設及び居住系サービス"],
           [["全国（参考・D8の基準）", None, None],
            ["北海道", 8918, 10504],
            ["大雪地区広域連合", 9164, 13097]], numfmt="#,##0")
mono_bar(ws, "調整済み第1号1人あたり給付月額（北海道との比較）", "給付月額（円）",
         Reference(ws, min_col=1, min_row=r0 + 3, max_row=e2),
         Reference(ws, min_col=2, max_col=3, min_row=r0 + 1, max_row=e2),
         "N26", width=16, height=10, labels=True, numfmt="#,##0", gap=60)
note(ws, e2 + 1,
     "注1）北海道の値は北海道第9期計画 第7章第12節の分布図に付された注記（在宅8,918円、"
     "施設及び居住系10,504円）。大雪の値は見える化D8（令和5年）。"
     "注2）大雪は施設及び居住系が北海道を2,593円（24.7％）上回る一方、在宅は246円（2.8％）上回る程度で、"
     "超過分のほとんどは施設及び居住系の差である。21シートの受給率の分解と同じ結果である。"
     "注3）北海道計画の分布図によると、旭川市は逆に在宅が約12,000円と高く施設及び居住系が約9,700円と低い。"
     "生活圏が重なる旭川市と構造が逆転している点は、住民説明で問われやすい。"
     "注4）全国の調整済み値は受領データにないため空欄とした。見える化の再出力後に補う。")

r1 = e2 + 3
ws.cell(row=r1, column=1, value="（3）24時間対応サービスの圏域と大雪の比較（人／月）").font = \
    Font(name=FONT, size=10, bold=True)
e3 = table(ws, r1 + 1, ["サービス", "圏域R4実績", "圏域R6見込", "圏域R8見込", "圏域R22見込", "大雪の状況"],
           [["定期巡回・随時対応型訪問介護看護", 59.1, 63.0, 68.0, 84.0, "給付月額15円（R7）で実質未整備"],
            ["夜間対応型訪問介護", 70.8, 64.0, 70.0, 58.0, "給付月額0円で事業所なし"],
            ["看護小規模多機能型居宅介護", 1.0, 64.0, 101.0, 86.0, "給付月額0円で事業所なし"]],
           numfmt="#,##0.0")
mono_bar(ws, "24時間対応サービスの圏域見込量", "利用者数（人／月）",
         Reference(ws, min_col=1, min_row=r1 + 2, max_row=e3),
         Reference(ws, min_col=2, max_col=5, min_row=r1 + 1, max_row=e3),
         "N48", width=17, height=10, labels=True, numfmt="#,##0", gap=60)
note(ws, e3 + 1,
     "注1）圏域の見込量は北海道第9期計画 第7章第12節4「老人福祉サービスの目標（介護サービス）」。"
     "注2）圏域は看護小規模多機能型居宅介護を令和4年度実績1.0人/月から令和8年度101.0人/月へ大幅に増やす計画。"
     "大雪は3サービスとも実質ゼロで、圏域の定期巡回はほぼ旭川市に集中していると考えられる。"
     "注3）大雪の要介護5の在宅・居住系割合は令和7年度54.6％（18シート）に達しており、"
     "重度在宅者を支える24時間対応サービスの欠落は第6章の整備方針の中心論点となる（修正指示書A-13・B-22）。")



# ============================================================ 24 通いの場（総合事業ベース）
ws = sheet("24_通いの場_総合事業", "図23　通いの場の状況（介護予防・日常生活支援総合事業ベース）",
           "出典：厚生労働省「介護予防事業及び介護予防・日常生活支援総合事業（地域支援事業）の実施状況に関する調査」"
           "／見える化F1〜F9。第9期計画の代表KPI④が用いた日常生活圏域ニーズ調査（自己申告）とは"
           "分母・分子の定義が異なる別の統計であり、直接比較できない。",
           [24, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10])
FY = ["H25", "H26", "H27", "H28", "H29", "H30", "R元", "R2", "R6"]

ws["A4"] = "（1）月1回以上の通いの場の参加率（％）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
e1 = table(ws, 5, ["区分"] + FY,
           [["大雪地区広域連合", 1.1, 1.4, 1.4, 1.9, 3.1, 3.0, 2.7, 1.6, 3.53],
            ["北海道", 1.9, 1.9, 2.6, 2.8, 3.8, 3.9, 4.8, 3.9, None],
            ["全国", 2.7, 3.2, 3.9, 4.2, 4.9, 5.7, 6.7, 5.2, None]], numfmt="0.0")
mono_line(ws, "月1回以上の通いの場の参加率", "参加率（％）",
          Reference(ws, min_col=2, max_col=10, min_row=5, max_row=5),
          Reference(ws, min_col=1, max_col=10, min_row=6, max_row=e1),
          "P4", width=20, height=11, from_rows=True, labels=True, numfmt="0.0", min_=0.0, max_=8.0)
note(ws, e1 + 1,
     "注1）大雪は平成29年度3.1％をピークに令和2年度1.6％まで低下したが、"
     "令和6年度は3.53％（参加者330人）へ回復した。"
     "令和2年度の151人からは2.19倍、平成29年度の277人も上回る。"
     "注2）北海道・全国は令和2年度が最新である。"
     "北海道の第9期の評価指標は同じ統計で3.94％（令和3年度）→5.37％以上（令和8年度）であり、"
     "大雪の令和6年度3.53％は1.84ポイント下回る。国の目標値8％も同じ統計による。"
     "注3）令和3年度から令和5年度は本調査の当該年度分を受領していないため空欄である。")

r0 = e1 + 3
ws.cell(row=r0, column=1, value="（2）週1回以上の通いの場の参加率（％）").font = Font(name=FONT, size=10, bold=True)
e2 = table(ws, r0 + 1, ["区分"] + FY,
           [["大雪地区広域連合", 1.0, 0.0, 0.3, 0.5, 0.9, 0.6, 0.6, 0.7, 1.28],
            ["北海道", 0.7, 0.7, 1.0, 1.0, 1.3, 1.3, 1.8, 1.6, None],
            ["全国", 0.7, 0.9, 1.1, 1.4, 1.7, 2.2, 2.6, 2.1, None]], numfmt="0.0")
mono_line(ws, "週1回以上の通いの場の参加率", "参加率（％）",
          Reference(ws, min_col=2, max_col=10, min_row=r0 + 1, max_row=r0 + 1),
          Reference(ws, min_col=1, max_col=10, min_row=r0 + 2, max_row=e2),
          "P26", width=20, height=11, from_rows=True, labels=True, numfmt="0.0", min_=0.0, max_=3.0)
note(ws, e2 + 1,
     "注）平成25年度は大雪1.0％が北海道0.7％・全国0.7％を上回っていたが、"
     "令和元年度には大雪0.6％と低下し逆転した。"
     "令和6年度は1.28％（120人）へ回復し、北海道の令和2年度1.6％に近づいている。")

r1 = e2 + 3
ws.cell(row=r1, column=1, value="（3）週1回以上の通いの場の箇所数（箇所）").font = Font(name=FONT, size=10, bold=True)
e3 = table(ws, r1 + 1, ["区分"] + FY,
           [["大雪地区広域連合", 2, 0, 2, 5, 7, 6, 4, 3, 11],
            ["北海道", 490, 598, 920, 924, 1200, 1273, 1720, 1766, None],
            ["全国", 11712, 15477, 20336, 25266, 33461, 41509, 51032, 47181,
             None]], numfmt="#,##0")
e3b = table(ws, e3 + 2, ["区分", "箇所数", "65歳以上人口", "1万人あたり箇所数",
                         "全国（R2）比"],
            [["大雪地区広域連合（R2）", 3, 9251, 3.24, 0.25],
             ["大雪地区広域連合（R6）", 11, 9346, 11.77, 0.90],
             ["北海道（R2）", 1766, 1685744, 10.48, 0.80],
             ["全国（R2）", 47181, 35974231, 13.12, 1.00]], numfmt="#,##0.00")
mono_bar(ws, "65歳以上1万人あたりの週1回以上の通いの場の箇所数", "箇所数",
         Reference(ws, min_col=1, min_row=e3 + 3, max_row=e3b),
         Reference(ws, min_col=4, min_row=e3 + 2, max_row=e3b),
         "P48", width=14, height=9, labels=True, numfmt="0.00", gap=80)
note(ws, e3b + 1,
     "注1）65歳以上人口は、大雪の令和2年度は見える化F4の掲載値、"
     "令和6年度は見える化A系列の令和6年（2024年）の値、"
     "北海道・全国は参加者数と参加率から逆算した推計値。"
     "注2）大雪の箇所数の密度は令和2年度に全国の25％であったが、"
     "令和6年度は11.77箇所で全国の令和2年度水準の90％にあたる。"
     "注3）令和6年度の週1回以上11箇所はすべて東神楽町である。"
     "東川町・美瑛町には週1回以上の通いの場がない。")

r2 = e3b + 3
ws.cell(row=r2, column=1,
        value="（4）大雪地区広域連合の通いの場の内訳（令和6年度・町別）").font = \
    Font(name=FONT, size=10, bold=True)
e4a = table(ws, r2 + 1,
            ["区分", "東川町", "美瑛町", "東神楽町", "3町計"],
            [["箇所数", 2, 8, 15, 25],
             ["　うち週1回以上", 0, 0, 11, 11],
             ["参加者実人数（人）", 40, 133, 157, 330],
             ["　うち週1回以上（人）", 0, 0, 120, 120],
             ["【運営主体】住民団体", 0, 3, 14, 17],
             ["【運営主体】社会福祉協議会", 2, 3, 1, 6],
             ["【運営主体】介護関係施設・事業所", 0, 2, 0, 2],
             ["【運営主体】住民個人", 0, 0, 0, 0]], numfmt="#,##0")
mono_bar(ws, "通いの場の箇所数と参加者実人数（令和6年度・町別）", "箇所・人",
         Reference(ws, min_col=2, max_col=4, min_row=r2 + 1, max_row=r2 + 1),
         Reference(ws, min_col=1, max_col=4, min_row=r2 + 2, max_row=r2 + 5),
         "P70", width=16, height=9, from_rows=True, labels=True,
         numfmt="#,##0", gap=60)
note(ws, e4a + 1,
     "注1）令和6年度の25箇所は、運営主体が住民団体17・社会福祉協議会6・"
     "介護関係施設事業所2で、住民個人が運営する通いの場は0箇所である。"
     "住民個人による運営は全国19.2％・北海道28.8％（令和2年度）あり、"
     "当広域連合はこの類型を欠いている。"
     "注2）町別の差が大きい。箇所数は東川町2・美瑛町8・東神楽町15であり、"
     "週1回以上開催はすべて東神楽町である。")

r2 = e4a + 3
ws.cell(row=r2, column=1,
        value="（5）令和2年度の週1回以上の通いの場3箇所の内訳（参考）").font = \
    Font(name=FONT, size=10, bold=True)
e4 = table(ws, r2 + 1, ["区分", "大雪（箇所）", "大雪（％）", "北海道（％）", "全国（％）"],
           [["【運営主体】住民団体", 2, 66.7, 61.4, 73.1],
            ["【運営主体】社会福祉協議会", 1, 33.3, 1.3, 2.1],
            ["【運営主体】住民個人", 0, 0.0, 28.8, 19.2],
            ["【活動場所】公民館・自治会館・集会所", 1, 33.3, 87.2, 80.3],
            ["【活動場所】その他", 2, 66.7, 4.5, 5.4],
            ["【活動内容】体操（運動）", 1, 33.3, 57.2, 79.8],
            ["【活動内容】認知症予防", 1, 33.3, 2.9, 1.4],
            ["【活動内容】農作業", 1, 33.3, 0.1, 0.1]], numfmt="0.0")
note(ws, e4 + 1,
     "注1）令和2年度の3箇所は、運営主体が住民団体2・社会福祉協議会1、活動場所が公民館等1・その他2、"
     "活動内容が体操（運動）1・認知症予防1・農作業1であった。"
     "注2）農作業を主な活動内容とする通いの場は全国で59箇所（0.1％）、北海道で1箇所（0.1％）と稀であり、"
     "大雪は3箇所中1箇所（33.3％）を占める。圏域の地域特性を反映していると考えられ、"
     "第10期の介護予防・社会参加の施策で活かせる可能性がある。"
     "注3）住民個人が運営する通いの場が全国19.2％・北海道28.8％あるのに対し、大雪は0箇所である。"
     "注4）活動内容の内訳は令和2年度までしか得られていない。"
     "令和6年度の調査には活動内容の設問がない。")


# ============================================================ 25 事業所数の推移
ws = sheet("25_事業所数の推移", "図24　サービス種別ごとの事業所数と供給の空白",
           "出典：厚生労働省「介護保険総合データベース」及び総務省「住民基本台帳に基づく人口、人口動態及び"
           "世帯数調査」／見える化K1-a〜e・K2-a〜c・K3-a〜t（全28サービス種別、平成24年度〜令和6年度）。"
           "第10期基本指針の新別表 七「介護保険施設・事業所数」に対応する。",
           [26, 8, 8, 8, 8, 8, 8, 8, 8, 10, 10, 10, 10])
KY = ["H24", "H29", "H30", "R元", "R3", "R4", "R5", "R6"]

ws["A4"] = "（1）増減のあったサービスの事業所数（箇所）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
k1 = table(ws, 5, ["サービス"] + KY,
           [["訪問看護", 2, 4, 4, 4, 5, 6, 7, 7],
            ["居宅療養管理指導", 0, 3, 3, 3, 4, 6, 8, 10],
            ["訪問介護", 8, 10, 10, 10, 10, 10, 10, 13],
            ["訪問リハビリテーション", 1, 2, 2, 2, 3, 3, 3, 3],
            ["通所介護", 4, 5, 2, 2, 2, 2, 2, 2],
            ["認知症対応型共同生活介護", 6, 6, 6, 6, 6, 5, 5, 5],
            ["居宅介護支援", 11, 10, 10, 11, 11, 10, 10, 10],
            ["認知症対応型通所介護", 2, 0, 0, 0, 0, 0, 0, 0]], numfmt="#,##0")
mono_line(ws, "増減のあったサービスの事業所数の推移", "事業所数（箇所）",
          Reference(ws, min_col=2, max_col=9, min_row=5, max_row=5),
          Reference(ws, min_col=1, max_col=9, min_row=6, max_row=k1),
          "N4", width=22, height=12, from_rows=True, labels=False, min_=0.0, max_=14.0)
note(ws, k1 + 1,
     "注）訪問看護は平成24年度2事業所から令和6年度7事業所へ3.5倍、居宅療養管理指導は0から10へ増加した。"
     "一方、認知症対応型通所介護は平成27年度に消滅（2→0）、通所介護は平成30年度に5→2、"
     "認知症対応型共同生活介護は令和4年度に6→5となっている。"
     "掲載していない20種別は全期間で事業所数が変わらないか、全期間ゼロである。")

r0 = k1 + 3
ws.cell(row=r0, column=1, value="（2）域内に事業所が存在しないサービス（令和6年度・人口10万対）").font = \
    Font(name=FONT, size=10, bold=True)
k2 = table(ws, r0 + 1, ["サービス", "大雪", "北海道", "全国", "空白の期間"],
           [["定期巡回・随時対応型訪問介護看護", 0.0, 2.6, 1.2, "平成24年度〜（13年）"],
            ["看護小規模多機能型居宅介護", 0.0, 1.6, 0.9, "平成24年度〜（13年）"],
            ["夜間対応型訪問介護", 0.0, 0.1, 0.2, "平成24年度〜（13年）"],
            ["認知症対応型通所介護", 0.0, 2.8, 2.3, "平成27年度〜（10年）"],
            ["福祉用具貸与", 0.0, 6.0, 6.0, "平成24年度〜（13年）"],
            ["訪問入浴介護", 0.0, 1.1, 1.3, "平成24年度〜（13年）"],
            ["介護医療院", 0.0, 1.0, 0.8, "平成30年度〜（7年）"],
            ["地域密着型特定施設入居者生活介護", 0.0, 0.6, 0.3, "平成24年度〜（13年）"],
            ["短期入所療養介護（病院等）", 0.0, 0.1, 0.1, "平成24年度〜（13年）"],
            ["短期入所療養介護（介護医療院）", 0.0, 0.1, 0.1, "平成30年度〜（7年）"],
            ["介護療養型医療施設", 0.0, 0.0, 0.0, "平成24年度〜（13年）"]], numfmt="0.0")
mono_bar(ws, "域内に事業所が存在しないサービス（令和6年度・人口10万対）", "事業所数（人口10万対）",
         Reference(ws, min_col=1, min_row=r0 + 2, max_row=k2),
         Reference(ws, min_col=2, max_col=4, min_row=r0 + 1, max_row=k2),
         "N30", width=22, height=12, labels=False, numfmt="0.0", gap=60, cat_rot=-45)
note(ws, k2 + 1,
     "注）全28サービス種別のうち11種別（39.3％）で域内の事業所数がゼロである。"
     "うち定期巡回・随時対応型訪問介護看護、夜間対応型訪問介護、看護小規模多機能型居宅介護の3種別は、"
     "重度者の在宅生活を24時間支える機能を担うもので、制度創設以来13年間ゼロである。"
     "一方、要介護5の54.6％・要介護4の49.1％が在宅・居住系にいる。"
     "上川中部圏域では定期巡回が令和8年に68.0人/月、看護小規模多機能型居宅介護が101.0人/月の見込みがあり、"
     "圏域内の他保険者では機能している。")

r1 = k2 + 3
ws.cell(row=r1, column=1, value="（3）人口10万対の事業所数（令和6年度）").font = Font(name=FONT, size=10, bold=True)
k3 = table(ws, r1 + 1, ["サービス", "大雪", "北海道", "全国", "大雪／全国"],
           [["小規模多機能型居宅介護", 18.0, 7.2, 4.5, 4.00],
            ["短期入所療養介護（老健）", 10.8, 3.4, 3.1, 3.48],
            ["介護老人保健施設", 10.8, 3.7, 3.4, 3.18],
            ["特定施設入居者生活介護", 10.8, 6.0, 4.9, 2.20],
            ["短期入所生活介護", 18.0, 8.7, 9.2, 1.96],
            ["訪問看護", 25.2, 15.7, 14.7, 1.71],
            ["介護老人福祉施設", 10.8, 7.6, 6.9, 1.57],
            ["訪問介護", 46.9, 33.0, 30.1, 1.56],
            ["認知症対応型共同生活介護", 18.0, 19.8, 11.7, 1.54],
            ["居宅介護支援", 36.1, 29.4, 30.3, 1.19],
            ["地域密着型通所介護", 14.4, 16.2, 15.5, 0.93],
            ["居宅療養管理指導", 36.1, 41.8, 48.3, 0.75],
            ["通所介護", 7.2, 14.6, 20.3, 0.35]], numfmt="0.00")
mono_bar(ws, "人口10万対の事業所数　大雪／全国（令和6年度）", "全国＝1.00",
         Reference(ws, min_col=1, min_row=r1 + 2, max_row=k3),
         Reference(ws, min_col=5, min_row=r1 + 1, max_row=k3),
         "N56", width=22, height=12, labels=True, numfmt="0.00", gap=60, cat_rot=-45)
note(ws, k3 + 1,
     "注）大雪の人口は約2.78万人であり、1事業所が人口10万対3.6に相当する。"
     "人口が小さいほど1事業所あたりの寄与が大きくなるため、人口10万対の比較は大雪の値が高く出やすい。"
     "その補正を加えてもなお低いのは通所介護（0.35）と居宅療養管理指導（0.75）である。"
     "通所介護は地域密着型通所介護と合わせても21.6で全国35.8の0.60にとどまる。")


# ============================================================ 26 従事者数の推移
ws = sheet("26_従事者数の推移", "図25　サービス別の従事者数の推移",
           "出典：厚生労働省「介護サービス施設・事業所調査」／見える化M2-a〜m（13サービス、"
           "平成29年度〜令和6年度）。令和2年度は同調査が実施されなかったため全国・北海道を含め欠測である。"
           "第10期基本指針の新別表 七「従事者数」に対応する。",
           [26, 9, 9, 9, 9, 9, 9, 9, 10, 10, 10, 10, 10])
MY = ["H29", "H30", "R元", "R3", "R4", "R5", "R6"]

ws["A4"] = "（1）サービス別の従事者数（実数・人）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
m1 = table(ws, 5, ["サービス"] + MY,
           [["介護老人保健施設", 90, 158, 165, 162, 160, 113, 124],
            ["介護老人福祉施設", 116, 104, 123, 119, 115, 119, 105],
            ["地域密着型介護老人福祉施設", 24, 40, 59, 84, 77, 67, 50],
            ["訪問看護", 15, 22, 23, 19, 21, 29, 37],
            ["地域密着型通所介護", 0, 21, 12, 22, 26, 23, 25],
            ["通所リハビリテーション（老健）", 23, 41, 33, 31, 35, 20, 22]], numfmt="#,##0")
mono_line(ws, "サービス別の従事者数の推移", "従事者数（人）",
          Reference(ws, min_col=2, max_col=8, min_row=5, max_row=5),
          Reference(ws, min_col=1, max_col=8, min_row=6, max_row=m1),
          "N4", width=22, height=12, from_rows=True, labels=False, min_=0.0)
note(ws, m1 + 1,
     "注1）訪問看護は平成29年度15人から令和6年度37人へ2.5倍。事業所数2→7、給付月額＋69.2％と符合する"
     "明確な拡大であり、第9期で最も伸びた供給基盤である。"
     "注2）地域密着型介護老人福祉施設は令和3年度84人をピークに令和6年度50人へ40.5％減少している。"
     "事業所数は3で不変であり、1事業所あたりの人員が減っている。"
     "注3）地域密着型通所介護の平成29年度は制度上該当がないため0としている。"
     "注4）訪問介護（平成29年度59人）・通所介護（同13人）・居宅介護支援（同32人）は"
     "平成29年度のみの収録のため掲載していない。")

r0 = m1 + 3
ws.cell(row=r0, column=1, value="（2）認定者1万対でみた全国・北海道との比較（令和6年度）").font = \
    Font(name=FONT, size=10, bold=True)
m2 = table(ws, r0 + 1, ["サービス", "大雪", "北海道", "全国", "大雪／全国"],
           [["地域密着型介護老人福祉施設", 254.84, 68.23, 77.94, 3.27],
            ["介護老人保健施設", 632.01, 281.41, 312.22, 2.02],
            ["通所リハビリテーション（老健）", 112.13, 59.55, 65.55, 1.71],
            ["介護老人福祉施設", 535.17, 485.72, 540.03, 0.99],
            ["訪問看護", 188.58, 157.15, 207.04, 0.91],
            ["地域密着型通所介護", 127.42, 139.23, 160.27, 0.80]], numfmt="#,##0.00")
mono_bar(ws, "従事者数（認定者1万対）の比較（令和6年度）", "従事者数（認定者1万対）",
         Reference(ws, min_col=1, min_row=r0 + 2, max_row=m2),
         Reference(ws, min_col=2, max_col=4, min_row=r0 + 1, max_row=m2),
         "N30", width=22, height=12, labels=False, numfmt="#,##0", gap=60, cat_rot=-45)
note(ws, m2 + 1,
     "注）分母は要介護（要支援）認定者数。全国・北海道と直接比較できる。"
     "介護老人福祉施設は平成29年度616.37（全国520.67の1.18倍）から令和6年度535.17（同0.99倍）へ低下し、"
     "8年間で全国並みとなった。訪問看護は逆に平成29年度79.70（全国105.15の0.76倍）から"
     "令和6年度188.58（同0.91倍）へ全国水準に近づいている。")

r1 = m2 + 3
ws.cell(row=r1, column=1, value="（3）職種別の従事者数（介護老人福祉施設・実数）").font = Font(name=FONT, size=10, bold=True)
m3 = table(ws, r1 + 1, ["職種"] + MY,
           [["介護職員", 80, 66, 78, 74, 77, 74, 70],
            ["准看護師", 7, 7, 6, 6, 5, 5, 3],
            ["看護師", 5, 7, 8, 7, 7, 6, 7],
            ["生活相談員", 5, 5, 5, 6, 5, 5, 3],
            ["介護支援専門員", 3, 3, 3, 3, 3, 3, 3],
            ["機能訓練指導員", 2, 3, 2, 2, 2, 2, 1],
            ["合計", 116, 104, 123, 119, 115, 119, 105]], numfmt="#,##0")
mono_line(ws, "職種別の従事者数の推移（介護老人福祉施設）", "従事者数（人）",
          Reference(ws, min_col=2, max_col=8, min_row=r1 + 1, max_row=r1 + 1),
          Reference(ws, min_col=1, max_col=8, min_row=r1 + 2, max_row=m3 - 1),
          "N56", width=22, height=12, from_rows=True, labels=True, min_=0.0, max_=90.0)
note(ws, m3 + 1,
     "注）准看護師は平成29年度7人から令和6年度3人へ57.1％減少しており、看護師の増加（5→7人）では"
     "補いきれていない。生活相談員も5→3人、機能訓練指導員も2→1人へ減少している。"
     "定員は平成29年度234人から令和2年度以降160人へ31.6％減少しており、"
     "従事者総数の減少（▲9.5％）はこれを下回るため、利用者あたりの手厚さ自体は増している。"
     "介護人材の必要数（見える化H1・H2）は大雪についてデータ登録がなく出力できないため、"
     "職種別欠員率は算定できない。")


# ============================================================ 27 交付金評価
ws = sheet("27_交付金評価", "図26　保険者機能強化推進交付金等の評価得点",
           "出典：厚生労働省「保険者機能強化推進交付金・介護保険保険者努力支援交付金に係る評価指標"
           "（市町村分）」／見える化W126〜W145。調査実施年は令和5年（2023年）で、令和6年度分の指標として"
           "用いられる。全国・北海道は全保険者の平均値。"
           "第10期基本指針の新別表 九「インセンティブ交付金における評価」に対応する。",
           [30, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10])

ws["A4"] = "（1）総合得点（点）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
w1 = table(ws, 5, ["区分", "大雪地区広域連合", "北海道", "全国"],
           [["保険者機能強化推進交付金", 164.7, 210.4, 205.6],
            ["介護保険保険者努力支援交付金", 167.3, 203.3, 216.7],
            ["合計", 332.0, 413.6, 422.4]], numfmt="#,##0.0")
mono_bar(ws, "保険者機能強化推進交付金等の総合得点（令和5年調査）", "得点（点）",
         Reference(ws, min_col=1, min_row=6, max_row=w1),
         Reference(ws, min_col=2, max_col=4, min_row=5, max_row=w1),
         "N4", width=20, height=11, labels=True, numfmt="#,##0.0", gap=60)
note(ws, w1 + 1,
     "注）総合得点332.0点は全国平均422.4点の78.6％で、90.4点（21.4％）低い。北海道平均も81.6点下回る。"
     "交付金の交付額は得点に応じて配分されるため、この差はそのまま財源の差となる。"
     "第9期計画は交付金の評価結果に言及していない。")

r0 = w1 + 3
ws.cell(row=r0, column=1, value="（2）指標群別の得点（点）").font = Font(name=FONT, size=10, bold=True)
w2 = table(ws, r0 + 1, ["区分", "大雪地区広域連合", "北海道", "全国", "大雪／全国"],
           [["推進　取組・体制指標群", 96.7, 121.4, 122.5, 0.789],
            ["推進　活動指標群", 8.0, 33.2, 34.5, 0.232],
            ["推進　成果指標群", 60.0, 55.8, 48.6, 1.235],
            ["支援　取組・体制指標群", 82.3, 109.7, 123.1, 0.669],
            ["支援　活動指標群", 25.0, 37.8, 45.0, 0.556],
            ["支援　成果指標群", 60.0, 55.8, 48.6, 1.235]], numfmt="#,##0.000")
mono_bar(ws, "指標群別の得点（令和5年調査）", "得点（点）",
         Reference(ws, min_col=1, min_row=r0 + 2, max_row=w2),
         Reference(ws, min_col=2, max_col=4, min_row=r0 + 1, max_row=w2),
         "N28", width=20, height=11, labels=False, numfmt="#,##0.0", gap=60, cat_rot=-45)
note(ws, w2 + 1,
     "注）成果（アウトカム）は全国の123.5％である一方、活動（アウトプット）は全国の23.2〜55.6％にとどまる。"
     "成果と活動の得点に差がある状態（原因は評価調書の受領後に確認）である。原因は未確認である。"
     "第9期計画の第5章がアウトプット指標を持たないこと、代表KPIの検証が期末1回のみであることと"
     "同じ構造であり、国の評価がこれを独立に裏付けている。")

r1 = w2 + 3
ws.cell(row=r1, column=1, value="（3）目標別の得点（点）").font = Font(name=FONT, size=10, bold=True)
w3 = table(ws, r1 + 1, ["目標", "大雪地区広域連合", "北海道", "全国", "大雪／全国"],
           [["推進Ⅰ 持続可能な地域のあるべき姿", 40.3, 55.0, 56.2, 0.717],
            ["推進Ⅱ 公正・公平な給付体制", 23.3, 57.9, 59.8, 0.390],
            ["推進Ⅲ 介護人材の確保・基盤整備", 41.0, 41.7, 41.0, 1.000],
            ["支援Ⅰ 介護予防／日常生活支援", 47.3, 49.7, 51.5, 0.918],
            ["支援Ⅱ 認知症総合支援", 36.3, 48.5, 54.5, 0.666],
            ["支援Ⅲ 在宅医療・在宅介護連携", 23.7, 49.3, 62.1, 0.382],
            ["共通Ⅳ 自立した日常生活（成果）", 60.0, 55.8, 48.6, 1.235]], numfmt="#,##0.000")
mono_bar(ws, "目標別の得点　大雪／全国（令和5年調査）", "全国＝1.000",
         Reference(ws, min_col=1, min_row=r1 + 2, max_row=w3),
         Reference(ws, min_col=5, min_row=r1 + 1, max_row=w3),
         "N52", width=20, height=11, labels=True, numfmt="#,##0.000", gap=60, cat_rot=-45)
note(ws, w3 + 1,
     "注）最も低いのは支援Ⅲ（在宅医療・在宅介護連携）の全国比0.382と推進Ⅱ（公正・公平な給付体制）の0.390である。"
     "支援Ⅲは入退院支援・人生の最終段階における支援がいずれも0点、"
     "推進Ⅱはケアプラン点検・医療情報との突合がいずれも0点である。"
     "後者は北海道の第9期評価指標⑧「給付適正化主要3事業の実施率100％」に反する。")

r2 = w3 + 3
ws.cell(row=r2, column=1, value="（4）性・年齢調整済み要介護2以上認定率（目標Ⅳ-5）").font = \
    Font(name=FONT, size=10, bold=True)
w4 = table(ws, r2 + 1, ["区分", "認定率（％）", "変化率（％）"],
           [["大雪地区広域連合", 9.25, 1.76],
            ["北海道", 8.59, -1.71],
            ["全国", 8.99, -0.37]], numfmt="0.00")
mono_bar(ws, "性・年齢調整済み要介護2以上認定率（令和5年調査）", "認定率（％）",
         Reference(ws, min_col=1, min_row=r2 + 2, max_row=w4),
         Reference(ws, min_col=2, min_row=r2 + 1, max_row=w4),
         "N76", width=14, height=9, labels=True, numfmt="0.00", gap=80)
note(ws, w4 + 1,
     "注1）本指標は国が全保険者について共通の方法で毎年算定しているため、第10期のH01"
     "（年齢調整済要介護認定率）のデータ源として最も確実である。"
     "注2）大雪9.25％は全国8.99％の102.9％、北海道8.59％の107.7％である。"
     "変化率は全国▲0.37・北海道▲1.71が低下する中で大雪のみ＋1.76と上昇している。"
     "注3）75歳以上の粗認定率が平成30年3月末35.3％から令和8年3月末33.5％へ低下していることと"
     "矛盾するものではない。前者は要介護2以上に限った性年齢調整済みの値、"
     "後者は全要介護度を含む年齢層別の粗率であり、"
     "合わせると「軽度は改善したが中重度は悪化している」という構造が読み取れる。")


# ============================================================ 28 特定地域と地域指定
ws = sheet("28_特定地域と地域指定", "図27　特定地域（中山間・人口減少地域）の基準と3町の位置",
           "出典：第135回社会保障審議会介護保険部会 資料3「特定地域（中山間・人口減少地域）の"
           "考え方について」（令和8年6月29日）／北海道 総合政策部地域創生局地域政策課"
           "「地域指定一覧（令和8年4月1日現在）」／国土交通省「豪雪地帯・特別豪雪地帯の指定"
           "（令和8年4月1日現在）」。",
           [26, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10])

ws["A4"] = "（1）3町の地域指定（令和8年4月1日現在）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
t1 = table(ws, 5, ["町", "過疎", "辺地", "山村", "離島", "半島", "特別豪雪地帯", "豪雪地帯"],
           [["東川町", "―", "○", "―", "―", "―", "○", "○"],
            ["美瑛町", "○", "○", "―", "―", "―", "○", "○"],
            ["東神楽町", "―", "○", "―", "―", "―", "―", "○"]])
note(ws, t1 + 1,
     "注）○は当該市町村が指定地域を有すること、△は市町村内の一部の区域が指定されていることを示す。"
     "豪雪地帯は道内全市町村が指定されている。"
     "3町とも辺地を有し、東川町・美瑛町は特別豪雪地帯を有する。"
     "これにより3町とも介護報酬上の「中山間地域等」（中山間地域等小規模事業所加算・"
     "中山間地域等居住者サービス提供加算の対象）に該当する見込みである。")

r0 = t1 + 3
ws.cell(row=r0, column=1, value="（2）特定地域の基準②（75歳以上人口密度）による3町の位置").font = \
    Font(name=FONT, size=10, bold=True)
t2 = table(ws, r0 + 1, ["町", "面積（km²）", "5人／km²の閾値（人）", "75歳以上人口（推計）",
                        "密度（人／km²）", "基準②の判定"],
           [["美瑛町", 676.78, 3384, 2065, 3.05, "該当する見込み"],
            ["東川町", 247.30, 1236, 1528, 6.18, "非該当の見込み"],
            ["東神楽町", 68.50, 342, 1548, 22.60, "非該当"]], numfmt="#,##0.00")
mono_bar(ws, "75歳以上人口密度（人／km²）", "人／km²",
         Reference(ws, min_col=1, min_row=r0 + 2, max_row=t2),
         Reference(ws, min_col=5, min_row=r0 + 1, max_row=t2),
         "N4", width=14, height=9, labels=True, numfmt="0.00", gap=80)
note(ws, t2 + 1,
     "注1）75歳以上人口は見える化A3（後期高齢者数）の広域連合計（令和2年5,167人）を"
     "第9期計画の町別高齢化率で按分した推計値。3町の確定値による検証を要する。"
     "注2）原典は令和2年国勢調査及び社人研「日本の地域別将来推計人口 令和5(2023)年推計」に"
     "基づくこととしている。"
     "注3）基準②のもう一方の要件「75歳以上人口1,000人未満かつ減少」は、"
     "広域連合の75歳以上人口が令和2年5,167人から令和7年5,701人へ10.3％増加しているため、"
     "3町とも該当しないと考えられる。")

r1 = t2 + 3
ws.cell(row=r1, column=1, value="（3）特定地域の基準の3段階構造").font = \
    Font(name=FONT, size=10, bold=True)
t3 = table(ws, r1 + 1, ["段階", "基準", "指定の単位", "大雪地区広域連合への適用"],
           [["①", "特別地域加算・離島等相当サービスの対象地域"
             "（市町村の全部指定・一部指定いずれも含む）", "市町村の全部又は一部",
             "東川町・美瑛町は特別豪雪地帯を有し、美瑛町は過疎地域でもあるため対象地域である"
             "可能性がある。告示の確認を要する。"],
            ["②", "①に該当しない場合、75歳以上人口密度5人／km²未満"
             "又は75歳以上人口1,000人未満かつ減少", "市町村の全域",
             "美瑛町は密度要件に該当する見込み。"],
            ["③", "①②に該当しない場合でも、⑴市町村内の一部地域が②の基準に該当する地域、"
             "⑵特定のサービス類型の事業所が地域に存在しない等、"
             "サービス基盤の維持が困難である地域",
             "市町村の全域又は一部地域（旧市町村単位、行政区単位、日常生活圏域単位）",
             "全28サービス種別のうち11種別で域内事業所がゼロであり⑵に直接該当する。"
             "①②に該当しない町についても指定の対象となり得る。"]])
note(ws, t3 + 1,
     "注）指定は「市町村が特定地域としての指定の要否を検討し、その意向を踏まえて"
     "都道府県が対象地域を定める」こととされている。市町村側の意向表明が起点となる。")


# ============================================================ 29 美瑛町の交通
ws = sheet("29_美瑛町の交通", "図28　美瑛町の公共交通とスクールバスの運行状況",
           "出典：美瑛町過疎地域持続的発展市町村計画（令和3年度〜令和7年度）。"
           "第10期基本指針案の別表 一「特に訪問・通所困難地域」への該当性の検討資料。",
           [30, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）スクールバス路線別乗車人員（人／年）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
b1 = table(ws, 5, ["路線名", "主たる通過集落", "平成30年度", "令和元年度", "令和2年度"],
           [["俵真布線", "美瑛〜朗根内〜俵真布", 7554, 7453, 8188],
            ["宇莫別線", "美瑛〜下宇莫別〜上宇莫別", 5215, 4714, 6165],
            ["二股線", "美瑛〜ルベシベ〜二股", 6531, 5795, 6268],
            ["美田・五稜線", "美瑛〜美田〜五稜", 4362, 4059, 4817],
            ["旭線", "美瑛〜北瑛〜旭", 5403, 5025, 4612],
            ["置杵牛線", "美瑛〜置杵牛", 3177, 2245, 2888],
            ["水沢線", "美瑛〜春日台〜千代田", 3224, 2683, 2643],
            ["美馬牛線", "美瑛〜美馬牛", 3652, 2948, 2408]], numfmt="#,##0")
mono_bar(ws, "スクールバス路線別乗車人員（令和2年度）", "人／年",
         Reference(ws, min_col=1, min_row=6, max_row=b1),
         Reference(ws, min_col=5, min_row=5, max_row=b1),
         "N4", width=20, height=11, labels=True, numfmt="#,##0", gap=60, cat_rot=-45)
note(ws, b1 + 1,
     "注）美瑛町過疎地域持続的発展市町村計画は「町内には、国鉄バスの路線廃止により、"
     "昭和60年から8路線でスクールバスの運行が開始され、平成23年からは、"
     "地域の利便性を図ることを目的に、10路線が運行されている。"
     "今日では、学校の統廃合による児童生徒の遠距離通学、"
     "高齢者の通院や日常の買い物など交通弱者の貴重な公共交通として、"
     "スクールバスは欠くことのできない交通手段となっている。」と記載している。"
     "すなわち市街地・周辺地区以外の集落には定期路線バスがない。"
     "うち朗根内地区を通過する俵真布線の乗車人員が10路線中最多である。")

r2 = b1 + 3
ws.cell(row=r2, column=1, value="（2）公共交通と冬期の道路状況").font = \
    Font(name=FONT, size=10, bold=True)
b2 = table(ws, r2 + 1, ["区分", "内容", "訪問・通所困難地域の判定への意味"],
           [["鉄道", "ＪＲ富良野線　美瑛〜旭川間1日19往復、美瑛〜富良野間1日12往復。"
             "町内に美瑛・美馬牛・北美瑛の3駅",
             "鉄道駅があるのは市街地・周辺地区と美馬牛地区のみ"],
            ["民間バス", "白金温泉〜美瑛〜旭川間の1系統",
             "朗根内地区・旭・北西地区・美馬牛地区は経路上にない"],
            ["スクールバス", "国鉄バスの路線廃止により昭和60年から運行。現在10路線",
             "定期路線バスの代替であり、高齢者の通院・買物の手段となっている"],
            ["冬期の除雪", "道道7路線99.8kmで除雪率88％、町道414路線346.6kmで除雪率52.9％、"
             "歩道51路線39.60kmで除雪率51.0％",
             "町道の約47％は除雪されない。冬期のアクセスが制約される"],
            ["今後の方針", "「走行距離や耐用年数に応じたスクールバスの計画的な更新、"
             "デマンド型交通を含めた輸送サービスの充実を図る必要がある」",
             "美瑛町自身がデマンド型交通の導入を課題としている"]])

r3 = b2 + 3
ws.cell(row=r3, column=1, value="（3）美瑛町4地区の状況と判定").font = \
    Font(name=FONT, size=10, bold=True)
b3 = table(ws, r3 + 1, ["地区", "人口", "高齢者人口", "高齢化率", "鉄道駅", "民間バス",
                        "スクールバス", "判定"],
           [["市街地・周辺地区", 7869, "―", "―", "あり", "あり", "―", "非該当"],
            ["美馬牛地区", "―", "―", "―", "あり", "なし", "美馬牛線", "要確認"],
            ["旭・北西地区", 684, 308, "45.0％", "なし", "なし", "旭線", "該当の可能性がある"],
            ["朗根内地区", 233, 82, "―", "なし", "なし", "俵真布線",
             "該当の可能性が最も高い"]], numfmt="#,##0")
note(ws, b3 + 1,
     "注1）人口・高齢者人口は第9期計画 第1章第7節の掲載値。"
     "注2）判定を確定するには、区域内で訪問系・通所系サービスを提供する全事業所の"
     "運営規程における「通常の事業の実施地域」の収集を要する。"
     "居宅サービスは北海道、地域密着型サービスは広域連合が保有している。"
     "注3）特定地域の基準③は一部地域の指定を認めており、"
     "その範囲として旧市町村単位・行政区単位・日常生活圏域単位が挙げられている（図27）。"
     "訪問・通所困難地域として特定した地区を、そのまま特定地域として指定することを検討できる。")


# ============================================================ 30 保険料の分析
ws = sheet("30_保険料の分析", "図29　保険料の水準と管内比較",
           "出典：北海道『第9期介護保険料（第1号被保険者分）各保険者の保険料等』／"
           "介護保険事業状況報告等から算定した推計値。",
           [24, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）必要保険料月額と条例上の基準額の対比（第9期）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
h1 = table(ws, 5, ["年度", "必要保険料月額（推計）", "条例上の基準額", "差"],
           [["令和6年度", 6063, 6400, 337],
            ["令和7年度", 6352, 6400, 48]], numfmt="#,##0")
mono_bar(ws, "必要保険料月額と条例上の基準額（第9期）", "月額（円）",
         Reference(ws, min_col=1, min_row=6, max_row=h1),
         Reference(ws, min_col=2, max_col=3, min_row=5, max_row=h1),
         "N4", width=14, height=9, labels=True, numfmt="#,##0", gap=70)
note(ws, h1 + 1,
     "注）必要保険料月額は当該年度の保険給付費及び地域支援事業費から逆算した概算値。"
     "第7期は累計で月額159円、第8期は同66円の剰余であり、"
     "第9期は令和6・7年度の累計で月額385円（約42百万円）の剰余となる見込み。"
     "基金残高及び取崩実績と併せて確認する必要がある。")

r4 = h1 + 3
ws.cell(row=r4, column=1, value="（2）上川管内21保険者の第9期保険料基準額（月額）").font = \
    Font(name=FONT, size=10, bold=True)
KANNAI = [["当麻町", 6800], ["愛別町", 6706], ["鷹栖町", 6700], ["中川町", 6550],
          ["大雪地区広域連合", 6400], ["比布町", 6300], ["旭川市", 6190],
          ["富良野市", 6000], ["上川町", 6000], ["剣淵町", 6000], ["下川町", 6000],
          ["和寒町", 5950], ["美深町", 5900], ["中富良野町", 5700], ["南富良野町", 5700],
          ["名寄市", 5400], ["上富良野町", 5400], ["占冠村", 5100], ["士別市", 5025],
          ["幌加内町", 5000], ["音威子府村", 3600]]
h2 = table(ws, r4 + 1, ["保険者", "月額基準額（円）"], KANNAI, numfmt="#,##0")
mono_bar(ws, "上川管内21保険者の第9期保険料基準額（月額）", "円",
         Reference(ws, min_col=1, min_row=r4 + 2, max_row=h2),
         Reference(ws, min_col=2, min_row=r4 + 1, max_row=h2),
         "N28", width=22, height=11, labels=False, numfmt="#,##0", gap=40, cat_rot=-45)
note(ws, h2 + 1,
     "注）単純平均は約5,830円、中央値は6,000円。大雪地区広域連合の6,400円は高い方から5番目で、"
     "単純平均との差は約570円（約9.8％）。ただし人口規模等を加重していない単純比較である。")


# ============================================================ 31 KPIのデータ源
ws = sheet("31_KPIのデータ源", "図30　代表KPI16項目のデータ源の確保状況",
           "出典：本計画 資料1（代表KPI16項目の定義）。"
           "令和8年7月29日の見える化システムの追加受領（W144・M2系列）により"
           "H01・H13のデータ源が確定した。",
           [26, 12, 12, 12, 12, 40, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）データ源の確保状況の内訳"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
k1 = table(ws, 5, ["区分", "項目数", "該当ID", "内容"],
           [["確保（基準値算定済み）", 6, "H01・H02・H04・H05・H13・H15",
             "第1回委員会の資料に基準値を掲載できる"],
            ["確保（第10期調査の受領待ち）", 2, "H06・H11",
             "業務仕様書の対象4調査に含まれる。集計データの受領後に基準値を算定する"],
            ["様式整備が必要", 2, "H09・H10",
             "3町の地域包括支援センターの記録様式の統一が前提。令和10年度から実測"],
            ["データ源が未確保", 4, "H07・H08・H12・H16",
             "在宅介護実態調査・事業所実態調査が業務仕様書の対象外。"
             "業務範囲の決定又は代理指標への振替が必要"],
            ["抽出可否の確認が必要", 2, "H03・H14",
             "H03は認定台帳からの新規認定者数の抽出、"
             "H14は介護人材実態調査の採用者数・退職者数からの算定"]], numfmt="#,##0")
mono_bar(ws, "代表KPI16項目のデータ源の確保状況", "項目数",
         Reference(ws, min_col=1, min_row=6, max_row=k1),
         Reference(ws, min_col=2, min_row=5, max_row=k1),
         "N4", width=18, height=9, labels=True, numfmt="0", gap=60, cat_rot=-30)
note(ws, k1 + 1,
     "注）令和9年度当初から実測できる見通しが立っているのは8項目"
     "（H01・H02・H04・H05・H06・H11・H13・H15）。"
     "うちH01（性・年齢調整済み要介護2以上認定率＝見える化W144）と"
     "H13（職種別従事者数の推移＝見える化M2系列）は、"
     "令和8年7月29日の追加受領により確定した。"
     "残る8項目は、業務範囲の決定（4項目）、様式整備（2項目）、"
     "抽出可否の確認（2項目）を要する。")


# ============================================================ 32 第9期の対計画比
ws = sheet("32_第9期の対計画比", "図31　第9期計画の計画値と実績値の対比",
           "出典：地域包括ケア「見える化」システム 総括表（令和8年8月31日受領）。"
           "実績値は厚生労働省「介護保険事業状況報告」"
           "（被保険者数・認定者数・認定率は各年9月月報、給付費は年報）、"
           "計画値は介護保険事業計画にかかる保険者からの報告値。"
           "令和8年度は実績が未登録のため令和6・7年度の2か年で示す。",
           [26, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）主要指標の対計画比（％）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
k1 = table(ws, 5, ["指標", "令和6年度", "令和7年度"],
           [["第1号被保険者数", 99.3, 98.5],
            ["要介護認定者数", 97.7, 96.0],
            ["要介護認定率", 98.4, 97.4],
            ["総給付費", 96.3, 97.8],
            ["　施設サービス給付費", 104.3, 100.8],
            ["　居住系サービス給付費", 83.8, 88.3],
            ["　在宅サービス給付費", 93.7, 98.4],
            ["第1号被保険者1人あたり給付費", 97.0, 99.3]], numfmt="0.0")
mono_bar(ws, "第9期の対計画比（令和6・7年度）", "対計画比（％）",
         Reference(ws, min_col=1, min_row=6, max_row=k1),
         Reference(ws, min_col=2, max_col=3, min_row=5, max_row=k1),
         "P4", width=20, height=11, labels=True, numfmt="0.0", gap=60,
         cat_rot=-45)
note(ws, k1 + 1,
     "注1）100％が計画どおりである。総給付費は令和6年度96.3％・令和7年度97.8％で"
     "いずれも計画の範囲内にある。"
     "注2）施設サービスは唯一計画を上回り、居住系サービスは2か年とも"
     "10ポイント以上下回る。"
     "注3）総括表の「第9期累計」欄は3か年の計画値に2か年の実績値を対比したもので"
     "あり達成率ではない。本図は年度別の対計画比による。")

r0 = k1 + 3
ws.cell(row=r0, column=1,
        value="（2）総給付費の計画値と実績値（百万円）").font = \
    Font(name=FONT, size=10, bold=True)
k2 = table(ws, r0 + 1, ["区分", "令和6年度", "令和7年度"],
           [["計画値", 2924.3, 2979.9],
            ["実績値", 2817.2, 2915.3],
            ["差", -107.1, -64.6]], numfmt="#,##0.0")
mono_bar(ws, "総給付費の計画値と実績値", "給付費（百万円）",
         Reference(ws, min_col=2, max_col=3, min_row=r0 + 1, max_row=r0 + 1),
         Reference(ws, min_col=1, max_col=3, min_row=r0 + 2, max_row=r0 + 3),
         "P26", width=16, height=10, from_rows=True, labels=True,
         numfmt="#,##0", gap=60)
note(ws, k2 + 1,
     "注）代表KPI H15（給付費の計画乖離率）は令和6年度▲3.66％・令和7年度▲2.17％で、"
     "いずれも目標の±5％以内である。"
     "実績値は介護保険事業状況報告年報の給付費合計と1円まで一致する。")

r1 = k2 + 3
ws.cell(row=r1, column=1,
        value="（3）給付費3区分の対計画比の推移（％）").font = \
    Font(name=FONT, size=10, bold=True)
k3 = table(ws, r1 + 1, ["区分", "H30", "R元", "R2", "R3", "R4", "R5", "R6", "R7"],
           [["施設サービス", 96, 94, 95, 95, 90, 90, 104, 101],
            ["居住系サービス", 99, 101, 105, 94, 85, 82, 84, 88],
            ["在宅サービス", 95, 97, 95, 96, 96, 96, 94, 98]], numfmt="0")
mono_line(ws, "給付費3区分の対計画比の推移", "対計画比（％）",
          Reference(ws, min_col=2, max_col=9, min_row=r1 + 1, max_row=r1 + 1),
          Reference(ws, min_col=1, max_col=9, min_row=r1 + 2, max_row=k3),
          "P50", width=20, height=11, from_rows=True, labels=True,
          numfmt="0", min_=70, max_=110)
note(ws, k3 + 1,
     "注）第8期（令和3〜5年度）は3区分とも計画を下回っていたが、"
     "第9期に入って施設サービスが計画を上回るようになった。"
     "居住系サービスは第8期の令和5年度82％を底に回復しているが、"
     "なお計画を10ポイント以上下回る。")


# ============================================================ 33 サービス別の乖離
ws = sheet("33_サービス別の乖離", "図32　計画との乖離が大きいサービス（第9期）",
           "出典：地域包括ケア「見える化」システム 総括表詳細（利用者数）（給付費）。"
           "給付費の対計画比が令和6年度又は令和7年度に±20ポイントを超えるものを"
           "掲げた。利用者数は延べ人数（年間の月計）である。",
           [30, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）給付費の対計画比（％）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
s1 = table(ws, 5, ["サービス", "令和6年度", "令和7年度"],
           [["定期巡回・随時対応型訪問介護看護", 0.0, 19.5],
            ["地域密着型通所介護", 66.5, 66.6],
            ["短期入所生活介護", 67.7, 65.0],
            ["訪問看護", 76.5, 95.6],
            ["特定施設入居者生活介護", 79.0, 84.8],
            ["特定福祉用具販売", 106.3, 128.5],
            ["居宅療養管理指導", 122.7, 154.9],
            ["住宅改修", 125.4, 126.3],
            ["訪問入浴介護", 142.9, 166.9]], numfmt="0.0")
mono_bar(ws, "計画との乖離が大きいサービスの対計画比", "対計画比（％）",
         Reference(ws, min_col=1, min_row=6, max_row=s1),
         Reference(ws, min_col=2, max_col=3, min_row=5, max_row=s1),
         "P4", width=20, height=12, labels=True, numfmt="0", gap=60,
         cat_rot=-45)
note(ws, s1 + 1,
     "注1）100％が計画どおりである。"
     "注2）定期巡回・随時対応型訪問介護看護は区域内に事業所がない。"
     "第8期も3か年とも計画72人に対し実績0人であり、"
     "2期続けて計画に位置づけながら整備に至っていない。"
     "注3）地域密着型通所介護は、同じ通所系の通所介護88.7％・"
     "通所リハビリテーション110.5％（令和7年度）と差が大きい。")

r0 = s1 + 3
ws.cell(row=r0, column=1,
        value="（2）延べ利用者数の計画値と実績値（令和7年度・人）").font = \
    Font(name=FONT, size=10, bold=True)
s2 = table(ws, r0 + 1, ["サービス", "計画値", "実績値", "対計画比（％）"],
           [["定期巡回・随時対応型訪問介護看護", 60, 7, 11.7],
            ["地域密着型通所介護", 1512, 914, 60.4],
            ["短期入所生活介護", 456, 318, 69.7],
            ["訪問看護", 1488, 1622, 109.0],
            ["特定施設入居者生活介護", 828, 738, 89.1],
            ["特定福祉用具販売", 84, 129, 153.6],
            ["居宅療養管理指導", 1356, 2319, 171.0],
            ["住宅改修", 96, 101, 105.2],
            ["訪問入浴介護", 72, 63, 87.5]], numfmt="#,##0.0")
mono_bar(ws, "延べ利用者数の計画値と実績値（令和7年度）", "延べ利用者数（人）",
         Reference(ws, min_col=1, min_row=r0 + 2, max_row=s2),
         Reference(ws, min_col=2, max_col=3, min_row=r0 + 1, max_row=s2),
         "P28", width=20, height=12, labels=True, numfmt="#,##0", gap=60,
         cat_rot=-45)
note(ws, s2 + 1,
     "注1）給付費と延べ利用者数とで向きが異なるサービスがある。"
     "訪問入浴介護は延べ利用者数が計画を下回る（87.5％）一方、"
     "給付費は166.9％である。1人1月あたりの利用回数が計画の1.8倍であるためである。"
     "注2）居宅療養管理指導は延べ利用者数・給付費とも大きく上回る。"
     "第8期から連続しており、在宅医療・介護連携の進展を示すものとみられる。")

r1 = s2 + 3
ws.cell(row=r1, column=1,
        value="（3）計画値を0としながら実績のあるサービス").font = \
    Font(name=FONT, size=10, bold=True)
s3 = table(ws, r1 + 1,
           ["サービス", "令和6年度\n計画（人）", "令和6年度\n実績（人）",
            "令和6年度\n給付費（円）", "令和7年度\n実績（人）",
            "令和7年度\n給付費（円）"],
           [["認知症対応型通所介護", 0, 12, 1243134, 12, 1175328],
            ["介護療養型医療施設", 0, 1, 325178, 0, 0]], numfmt="#,##0")
note(ws, s3 + 1,
     "注）計画値が0であるため対計画比が算出されず、評価の対象から漏れている。"
     "認知症対応型通所介護は平成27年度に区域内の事業所がなくなったが、"
     "給付実績は続いている。"
     "介護療養型医療施設は令和5年度末で廃止されており、"
     "令和6年度の実績は経過的なものとみられる。"
     "第10期では、実績のあるサービスは計画値を置く扱いとする。")


# ============================================================ 34 保険料の計画と実績
ws = sheet("34_保険料の計画と実績", "図33　保険料基準額の計画と実績・第10期の暫定算定",
           "出典：地域包括ケア「見える化」システム 管理指標グラフ（保険料基準額）、"
           "将来推計 第3段階（給付費と保険料）。"
           "第10期の値は暫定の算定であり、条例で採用する基準額ではない。",
           [26, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])

ws["A4"] = "（1）期別の保険料基準額（月額・円）"
ws["A4"].font = Font(name=FONT, size=10, bold=True)
h1 = table(ws, 5, ["区分", "第7期", "第8期", "第9期"],
           [["計画値（準備基金取崩前）", 6233, 6330, 6553],
            ["計画値（準備基金取崩後）", 6077, 6237, 6428],
            ["実績値（準備基金取崩前）", 6100, 6334, 6196],
            ["実績値（準備基金取崩後）", 6021, 6334, 6069]], numfmt="#,##0")
mono_bar(ws, "期別の保険料基準額（月額）", "基準額（円）",
         Reference(ws, min_col=2, max_col=4, min_row=5, max_row=5),
         Reference(ws, min_col=1, max_col=4, min_row=6, max_row=h1),
         "P4", width=18, height=11, from_rows=True, labels=True,
         numfmt="#,##0", gap=60)
note(ws, h1 + 1,
     "注1）第8期は実績が計画を上回った（取崩後102％）が、"
     "第9期は下回っている（取崩後94％）。"
     "注2）第9期の実績は令和8年度分が未登録の状態での値である。"
     "注3）見える化システムの計画値は算定上の月額であり、"
     "条例で採用した基準額（第7期6,077円・第8期6,300円・第9期6,400円）とは"
     "百円未満の端数処理により一致しない場合がある。")

r0 = h1 + 3
ws.cell(row=r0, column=1,
        value="（2）第9期から第10期への要因分解（月額・円）").font = \
    Font(name=FONT, size=10, bold=True)
h2 = table(ws, r0 + 1, ["要因", "寄与額", "累計"],
           [["第9期の再現", 0, 6428],
            ["給付費の水準", 339, 6767],
            ["地域支援事業費", -29, 6738],
            ["基金取崩額（40百万円→0円）", 125, 6863],
            ["補正後被保険者数の係数", -153, 6710],
            ["被保険者数", 45, 6755]], numfmt="#,##0")
mono_line(ws, "第9期から第10期への要因分解", "月額基準額（円）",
          Reference(ws, min_col=1, min_row=r0 + 2, max_row=h2),
          Reference(ws, min_col=3, min_row=r0 + 1, max_row=h2),
          "P26", width=18, height=10, labels=True, numfmt="#,##0",
          min_=6300, max_=6900, cat_rot=-45)
note(ws, h2 + 1,
     "注1）第9期の前提をそのまま入れると6,428円となり、"
     "第9期計画の公表値と一致する。"
     "注2）置き換える順序により各要因の寄与額は変わる。"
     "注3）給付費の水準には、令和8年8月31日のご指示による人口の基礎の変更"
     "（総合戦略ベース）に伴う置換え（9,267百万円→9,484百万円）を含む。"
     "注4）第10期の基本ケース6,755円は暫定の算定である。")

r1 = h2 + 3
ws.cell(row=r1, column=1,
        value="（3）前提を動かした場合の幅（月額・円）").font = \
    Font(name=FONT, size=10, bold=True)
h3 = table(ws, r1 + 1, ["ケース", "月額", "基本との差"],
           [["給付費を第9期の実績年率で置く", 6159, -596],
            ["財政調整基金を全額取り崩す", 6040, -715],
            ["補正後被保険者数の係数を令和7年度実績とする", 6592, -163],
            ["予定収納率を実績（99.87％）とする", 6695, -60],
            ["基本ケース", 6755, 0],
            ["調整交付金を第8期の交付実績の水準とする", 7037, 282],
            ["第1号被保険者負担割合を24％とする", 7081, 326]], numfmt="#,##0")
mono_bar(ws, "前提を動かした場合の月額基準額", "月額基準額（円）",
         Reference(ws, min_col=1, min_row=r1 + 2, max_row=h3),
         Reference(ws, min_col=2, min_row=r1 + 1, max_row=h3),
         "P50", width=20, height=11, labels=True, numfmt="#,##0", gap=60,
         cat_rot=-45)
note(ws, h3 + 1,
     "注1）最も大きいレバーは給付費の水準で、月額に596円の幅がある。"
     "注2）見える化システムの自然体推計6,238円は、"
     "給付費を第9期の実績年率で置いた場合の6,159円に近い。"
     "注3）第9期の条例基準額は6,400円である。")


# ============================================================ 35 図表番号一覧
# 各シートに配置済みのグラフを走査し、計画本文への差し込み位置と通し番号を付与する。
# 第9期計画は図表に番号を付けておらず、【 】括弧のキャプションのみで管理していた
# （第2章第3節では丸数字⑪の次が（2）に変わるなど、採番自体が途中で崩れている）。
# 第10期では「図N-M」の通し番号を付し、本文・図表集・修正指示書で同じ番号を用いる。

# シート名 → (差し込み先の章・節, 見出し・段落, 資料（本文に付す出典表記）, 既定の掲載区分)
PLACE = {
    "01_人口推移": ("第2章第1節", "1 第9期策定時の基礎状況",
                 "資料：住民基本台帳　各年10月1日現在", "本文"),
    "02_高齢化率推移": ("第2章第1節", "1 第9期策定時の基礎状況",
                   "資料：住民基本台帳　各年10月1日現在", "本文"),
    "03_高齢者世帯": ("第2章第1節", "1 第9期策定時の基礎状況",
                 "資料：国勢調査（平成22年・平成27年・令和2年）", "本文"),
    "04_認定者数推移": ("第2章第1節", "1 第9期策定時の基礎状況",
                  "資料：介護保険事業状況報告　各年9月分", "本文"),
    "05_認定者割合": ("第2章第1節", "1 第9期策定時の基礎状況",
                 "資料：介護保険事業状況報告　令和5年9月分", "本文"),
    "06_出現率推移": ("第2章第1節", "1 第9期策定時の基礎状況",
                 "資料：介護保険事業状況報告　各年9月分", "本文"),
    "07_給付費推移": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                 "資料：地域包括ケア「見える化」システム／介護保険特別会計決算", "本文"),
    "08_ニーズ調査データ": ("第2章第3節", "高齢者の生活実態",
                     "資料：令和7年度 健康とくらしの調査（JAGES調査）", "本文"),
    "09_ニーズ調査グラフ": ("第2章第3節", "高齢者の生活実態",
                     "資料：令和7年度 健康とくらしの調査（JAGES調査）", "本文"),
    "10_在宅介護実態調査": ("第2章第4節", "1 在宅介護実態調査",
                     "資料：在宅介護実態調査（令和5年5月25日〜6月30日・認定調査員の聞き取り）", "本文"),
    "11_居所変更実態調査": ("第2章第4節", "2 居所変更実態調査",
                     "資料：居所変更実態調査（令和5年5月25日送付・21施設回答）", "本文"),
    "12_在宅生活改善調査": ("第2章第4節", "3 在宅生活改善調査",
                     "資料：在宅生活改善調査（令和5年5月25日送付・12事業所91人）", "本文"),
    "13_介護人材実態調査": ("第2章第5節", "介護事業所の現状",
                     "資料：介護人材実態調査（令和5年5月25日送付・27施設405人）", "本文"),
    "14_年齢構成と85歳以上": ("第2章第1節", "2 見える化システムで確認した直近の状況",
                       "資料：地域包括ケア「見える化」システム（国勢調査及び社人研推計）", "本文"),
    "15_サービス利用強度": ("第2章第1節", "3 主要サービスの受給者1人当たり利用日数・回数",
                     "資料：地域包括ケア「見える化」システム", "本文"),
    "16_町別将来推計": ("第2章第7節", "中長期推計からみた需要と財政",
                   "資料：地域包括ケア「見える化」システム（国勢調査及び社人研推計）", "本文"),
    "17_担い手の推移": ("第2章第7節", "中長期推計からみた需要と財政",
                   "資料：地域包括ケア「見える化」システム（国勢調査及び社人研推計）", "本文"),
    "18_受給率・利用率": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                    "資料：地域包括ケア「見える化」システム", "本文"),
    "19_給付月額の比較": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                    "資料：地域包括ケア「見える化」システム", "本文"),
    "21_受給率と定員": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                 "資料：地域包括ケア「見える化」システム", "本文"),
    "22_サービス別給付動向": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                      "資料：地域包括ケア「見える化」システム", "本文"),
    "23_上川中部圏域比較": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                    "資料：第9期北海道高齢者保健福祉計画・介護保険事業支援計画／"
                    "地域包括ケア「見える化」システム", "本文"),
    "24_通いの場_総合事業": ("第2章第3節", "高齢者の生活実態",
                     "資料：厚生労働省「介護予防事業及び介護予防・日常生活支援総合事業の実施状況に関する調査」",
                     "本文"),
    "25_事業所数の推移": ("第2章第2節", "保険給付や地域支援事業の実態把握と分析",
                   "資料：地域包括ケア「見える化」システム（厚生労働省「介護保険総合データベース」）", "本文"),
    "26_従事者数の推移": ("第2章第5節", "介護事業所の現状",
                   "資料：地域包括ケア「見える化」システム（厚生労働省「介護サービス施設・事業所調査」）", "本文"),
    "27_交付金評価": ("第3章第4節", "第9期計画の評価と第10期への課題",
                 "資料：地域包括ケア「見える化」システム（厚生労働省「保険者機能強化推進交付金・"
                 "介護保険保険者努力支援交付金に係る評価指標（市町村分）」）", "本文"),
    "28_特定地域と地域指定": ("第1章第9節", "1 国の第10期基本指針（案）の動向",
                    "資料：第135回社会保障審議会介護保険部会 資料3／北海道「地域指定一覧」／"
                    "国土交通省「豪雪地帯・特別豪雪地帯の指定」", "本文"),
    "29_美瑛町の交通": ("第1章第7節", "日常生活圏域の設定",
                  "資料：美瑛町過疎地域持続的発展市町村計画（令和3年度〜令和7年度）", "本文"),
    "30_保険料の分析": ("第6章第6節", "1 管内保険料の比較",
                  "資料：北海道『第9期介護保険料（第1号被保険者分）各保険者の保険料等』", "本文"),
    "31_KPIのデータ源": ("第4章第3節", "施策の体系と成果指標",
                    "資料：本計画 資料1（代表KPI16項目の定義）", "本文"),
    "20_第9期の達成状況": ("第3章第4節", "第9期計画の評価と第10期への課題",
                     "資料：第9期介護保険事業計画／地域包括ケア「見える化」システム／"
                     "令和7年度 健康とくらしの調査", "本文"),
    "32_第9期の対計画比": ("第3章第3節", "1 主要指標の計画値と実績値",
                     "資料：地域包括ケア「見える化」システム 総括表", "本文"),
    "33_サービス別の乖離": ("第3章第3節", "2 計画との乖離が大きいサービス",
                      "資料：地域包括ケア「見える化」システム 総括表詳細", "本文"),
    "34_保険料の計画と実績": ("第6章第6節", "2 見える化自然体推計（基準ケース）",
                       "資料：地域包括ケア「見える化」システム／"
                       "将来推計 第3段階（給付費と保険料）", "本文"),
}

# 掲載区分・備考の個別指定  キー＝(シート名, グラフタイトルの先頭一致)
OVERRIDE = {
    ("03_高齢者世帯", "高齢夫婦世帯数の定義差"):
        ("参考", "見える化A8と第9期計画の差の検証用。本文には掲載せず、修正指示書C-6の確認資料とする"),
    ("06_出現率推移", "町別出現率の推移"):
        ("本文", "町別データが未受領のため第9期計画の掲載値のまま（修正指示書C-8）"),
    ("07_給付費推移", "給付費の中長期見通し"):
        ("本文", "第2章第7節へ移す案もある。掲載箇所を委員会で決定する"),
    ("08_ニーズ調査データ", "社会参加の合成指標"):
        ("本文", "H05の算定根拠。第3章第1節（第9期KPIの評価）にも再掲する"),
    ("09_ニーズ調査グラフ", "令和4年→令和7年の変化"):
        ("本文", "第2章第6節「各調査結果のまとめ」にも再掲する"),
    ("11_居所変更実態調査", "① 退居・退所者に占める"):
        ("本文", "回答施設の内訳合計20と本文記載21が一致しない（修正指示書 確認事項）"),
    ("13_介護人材実態調査", "（1）③ 就業形態比較"):
        ("本文", "本文の33.8％とグラフ値77.0％が一致しない（修正指示書 確認事項）"),
    ("13_介護人材実態調査", "（3）④ 直前が介護の職場"):
        ("本文", "2区分が原典から復元できず入力欄としている（修正指示書 確認事項）"),
    ("25_事業所数の推移", "域内に事業所が存在しない"):
        ("本文", "第6章第4節（サービス見込量・整備方針）にも再掲する。修正指示書A-13の根拠図表"),
    ("26_従事者数の推移", "職種別の従事者数の推移"):
        ("本文", "H13の定義変更（職種別欠員率→職種別従事者数）の根拠。修正指示書C-16"),
    ("27_交付金評価", "指標群別の得点"):
        ("本文", "第5章のロジックモデル（アウトプット指標の整備）の根拠図表"),
    ("27_交付金評価", "性・年齢調整済み要介護2以上認定率"):
        ("本文", "H01のデータ源の確定根拠。第2章第1節にも再掲する。修正指示書A-10"),
}
# 第9期計画が掲載していなかった追加指標（社会参加6区分のうち第9期に無いもの）
EXTRA_NEEDS = ("⑫ 幸福感", "⑬ 就労していない", "⑮ スポーツ", "⑯ 趣味",
               "⑰ 学習・教養", "⑲ 特技や経験")


def _chart_title(_ch):
    try:
        return _ch.title.tx.rich.p[0].r[0].t
    except Exception:
        return ""


def _group_no(_ws):
    """シート見出し（A1）の「図N　…」から通し番号のグループ番号を取り出す。"""
    t = str(_ws["A1"].value or "")
    if t.startswith("図"):
        n = ""
        for c in t[1:]:
            if c.isdigit():
                n += c
            else:
                break
        if n:
            return int(n)
    return None


_rows = []
for _ws in wb.worksheets:
    if not _ws._charts:
        continue
    _g = _group_no(_ws)
    _pl = PLACE.get(_ws.title)
    if _g is None or _pl is None:
        continue
    _sec, _head, _src, _kind0 = _pl
    _seq = sum(1 for r in _rows if r[0].startswith("図%d-" % _g))
    for _ch in _ws._charts:
        _seq += 1
        _t = _chart_title(_ch)
        _kind, _memo = _kind0, ""
        for (_k_sheet, _k_pre), (_k, _m) in OVERRIDE.items():
            if _k_sheet == _ws.title and _t.startswith(_k_pre):
                _kind, _memo = _k, _m
        _sec_i, _head_i = _sec, _head
        if _ws.title == "09_ニーズ調査グラフ" and _t.startswith(EXTRA_NEEDS):
            _kind = "資料編"
            _sec_i, _head_i = "資料編", "資料6 調査結果の詳細（新設案）"
            _memo = ("第9期計画は14指標を掲載しており本指標は含まれない。"
                     "本文（第2章第3節）へ移すか資料編に置くかを委員会で決定する")
        if isinstance(_ch.anchor, str):
            _cell = _ch.anchor
        else:
            _cell = "%s%d" % (get_column_letter(_ch.anchor._from.col + 1),
                              _ch.anchor._from.row + 1)
        _rows.append(["図%d-%d" % (_g, _seq), "【%s】" % _t, _sec_i, _head_i,
                      _kind, _ws.title, _cell, _src, _memo])

ws = sheet("35_図表番号一覧", "図表番号一覧（計画本文への差し込み位置）",
           "図表集に収録した全%d点のグラフに通し番号を付し、計画本文（計画素案の章立て）"
           "のどの節・見出しに差し込むかを対応させたもの。"
           "第9期計画は図表に番号を付けていないため、第10期から新たに採番する。" % len(_rows),
           [10, 46, 12, 30, 10, 20, 7, 44, 40])

ws["A4"] = "1　採番と表記の規則"
ws["A4"].font = Font(name=FONT, size=11, bold=True)
_rules = [
    ["1", "通し番号", "「図N-M」。Nは図表集のシート単位（図1〜図19）、Mはシート内の掲載順。"
     "本文・図表集・修正指示書で同じ番号を用いる"],
    ["2", "本文のキャプション", "第9期計画と同じく図の上に【　】で囲んで表示する。"
     "番号は【　】の前に置き「図1-1【人口の推移（大雪地区広域連合・構成3町）】」とする"],
    ["3", "出典表記", "図の下に「資料：〜」の1行を右寄せで置く。第9期計画の体裁に合わせる"],
    ["4", "掲載区分", "本文＝計画本文に掲載／資料編＝資料編に掲載／"
     "参考＝数値検証用で計画には掲載しない"],
    ["5", "第9期の採番", "第9期計画 第2章第3節は丸数字①〜⑪の次が（2）（3）（4）に変わっており、"
     "同一節の中で採番方式が途中で切り替わっている。第10期では図表番号に一本化する"],
]
_r = table(ws, 5, ["No.", "項目", "内容"], _rules, headfill="404040")
for _rr in range(6, _r + 1):
    for _cc in (2, 3):
        ws.cell(row=_rr, column=_cc).alignment = Alignment(
            wrap_text=True, horizontal="left", vertical="top")
    ws.row_dimensions[_rr].height = 30

_r += 2
ws.cell(row=_r, column=1, value="2　図表番号と差し込み位置").font = Font(
    name=FONT, size=11, bold=True)
_hd = ["図表番号", "本文キャプション案", "差し込み先", "見出し・段落", "掲載区分",
       "図表集シート", "位置", "資料（出典表記）", "備考"]
_end = table(ws, _r + 1, _hd, _rows, headfill="404040")
for _rr in range(_r + 2, _end + 1):
    for _cc in range(1, 10):
        _c = ws.cell(row=_rr, column=_cc)
        _c.alignment = Alignment(wrap_text=True, vertical="top",
                                 horizontal="center" if _cc in (1, 5, 7) else "left")
    ws.row_dimensions[_rr].height = 26
ws.freeze_panes = ws.cell(row=_r + 2, column=1)
ws.auto_filter.ref = "A%d:I%d" % (_r + 1, _end)

_n_body = sum(1 for x in _rows if x[4] == "本文")
_n_app = sum(1 for x in _rows if x[4] == "資料編")
_n_ref = sum(1 for x in _rows if x[4] == "参考")
note(ws, _end + 2,
     "注1）収録点数は全%d点（本文%d点、資料編%d点、参考%d点）。"
     "第9期計画は第2章の分析を14指標のニーズ調査と3つの実態調査で構成していたが、"
     "第10期は見える化データによる分析（13〜19シート）を加えたため点数が増えている。"
     "成果品仕様（本文約100頁）に対する配分は、修正指示書B-4「章別ページ配分」と併せて委員会で決定する。"
     "注2）「差し込み先」は計画素案の章立てによる。章構成が変わった場合は本シートを先に更新し、"
     "図表集・本文・修正指示書へ反映する。"
     "注3）掲載区分「資料編」は、第9期計画に対応する図表が無い6指標"
     "（幸福感・就労状況・スポーツの会・趣味の会・学習教養サークル・特技を伝える活動）の"
     "年齢階級別・町別あわせて%d点。うち4指標は第10期の代表KPI H05の構成指標であるため、"
     "本文に何点まで掲載するかを委員会で決定する。"
     "注4）掲載区分「参考」は数値の検証用で計画には掲載しない。"
     % (len(_rows), _n_body, _n_app, _n_ref, _n_app))


del wb["Sheet"]
wb.save("/home/user/repository/output/第10期計画_図表集_白黒.xlsx")
print("saved:", len(wb.sheetnames), "sheets")
