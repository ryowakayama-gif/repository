"""
川崎町 第9期介護保険事業計画 実績一覧 町記入用フォーマット

第3章「第9期計画の取組実績と評価」の根拠となる実績データを、
町担当課（保健福祉課・大宮様）に記入していただくためのExcelフォーマット。

7シート構成：
00_使い方           : 記入ガイド・対象期間・担当割り当て
01_KPI一覧          : 基本目標別 目標値・実績・達成率
02_介護予防_健康    : 通いの場・サポーター・健診関連
03_在宅生活支援     : 紙おむつ・配食・緊急通報・移動支援
04_認知症_包括      : 認知症サポーター・カフェ・初期集中支援
05_介護サービス_人材: 認定者・サービス利用・事業所人材
06_第10期反映方針   : 継続・新規・廃止・見直しの分類整理

設計方針：
- 黄色＝記入欄、緑＝既知の確定値（KO・各種データ）、グレー＝項目
- 各シートに優先度（A/B）と備考列を設置
- 実績データ確認サマリーで既知の数値は事前投入
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# カラーパレット（計画書と統一）
NAVY = "1F3864"
BLUE = "2F5597"
LBLUE = "DAE3F3"
ORANGE = "ED7D31"
LORANGE = "FCE4D6"
GREEN = "548235"
LGREEN = "E2EFDA"
GRAY = "808080"
LGRAY = "F2F2F2"
INPUT_YELLOW = "FFFFCC"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def F(c): return PatternFill("solid", fgColor=c)
cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
cl = Alignment(horizontal="left", vertical="center", wrap_text=True)
cr = Alignment(horizontal="right", vertical="center", wrap_text=True)

f_title = Font(name="游ゴシック", size=13, bold=True, color=WHITE)
f_sub   = Font(name="游ゴシック", size=10, italic=True, color=WHITE)
f_head  = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
f_section = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
f_body  = Font(name="游ゴシック", size=10)
f_input = Font(name="游ゴシック", size=10, color="0000FF")
f_known = Font(name="游ゴシック", size=10, bold=True, color=GREEN)
f_note  = Font(name="游ゴシック", size=9, italic=True, color="595959")
f_pri_a = Font(name="游ゴシック", size=10, bold=True, color="C00000")
f_pri_b = Font(name="游ゴシック", size=10, bold=True, color="2F5597")

wb = Workbook()
wb.remove(wb.active)

def ms(ws, rng, val, font, fill, align):
    """セル結合＋スタイル適用"""
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = val
    c.font = font
    if fill is not None:
        c.fill = fill
    c.alignment = align
    from openpyxl.utils.cell import range_boundaries
    a, b, d, e = range_boundaries(rng)
    for r in range(b, e+1):
        for col in range(a, d+1):
            ws.cell(row=r, column=col).border = border

def setup_page(ws, orientation="landscape"):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orientation == "landscape" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

# ===========================================================
# 00_使い方
# ===========================================================
ws = wb.create_sheet("00_使い方")
ms(ws, "A1:G1", "川崎町 第9期介護保険事業計画 実績一覧（町記入用フォーマット）",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 28
ms(ws, "A2:G2", "─ 第10期計画素案 第3章「第9期の取組実績と評価」のための実績収集シート ─",
   f_sub, F(BLUE), cc)
ws.row_dimensions[2].height = 20

# 1. 目的
ms(ws, "A4:G4", "1．フォーマットの目的", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22

purposes = [
    "本シートは、川崎町第10期介護保険事業計画の素案 第3章「第9期計画の取組実績と評価」の根拠データとして、町担当課に記入していただくためのフォーマットです。",
    "対象期間は、第9期計画期間中の令和6年度（R6.4〜R7.3）の実績を基本としています。令和7年度（R7.4以降）の最新値があれば併せてご記入ください。",
    "本シートの記入結果は、計画素案Ver.2.0（令和8年8月以降）の第3章に反映し、第1回・第2回策定委員会で報告します。",
]
r = 5
for t in purposes:
    ms(ws, f"A{r}:G{r}", t, f_body, None, cl)
    ws.row_dimensions[r].height = 30
    r += 1

# 2. シート構成
r += 1
ms(ws, f"A{r}:G{r}", "2．シート構成と担当割り当て（案）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

# ヘッダ行
ws.cell(row=r, column=1, value="シート").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY)
ws.cell(row=r, column=1).alignment = cc
ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "内容", f_head, F(NAVY), cc)
ws.cell(row=r, column=4, value="想定担当").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY)
ws.cell(row=r, column=4).alignment = cc
ws.cell(row=r, column=4).border = border
ms(ws, f"E{r}:F{r}", "主な項目", f_head, F(NAVY), cc)
ws.cell(row=r, column=7, value="優先度").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY)
ws.cell(row=r, column=7).alignment = cc
ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 22
r += 1

sheets_info = [
    ("01_KPI一覧", "基本目標別 KPI目標値・実績・達成率", "保健福祉課", "5基本目標×3〜5指標", "A"),
    ("02_介護予防_健康", "通いの場・サポーター・健診関連", "保健福祉課", "ユニバーサルサポーター等", "A"),
    ("03_在宅生活支援", "紙おむつ・配食・緊急通報・移動支援", "保健福祉課\n地域振興課", "独自施策実績", "A"),
    ("04_認知症_包括", "認知症サポーター・カフェ・初期集中支援", "保健福祉課\n包括センター", "認知症基本法対応", "A"),
    ("05_介護サービス_人材", "認定者・サービス利用・事業所人材", "保健福祉課", "保険者データ", "B"),
    ("06_第10期反映方針", "継続/新規/廃止/見直しの分類", "保健福祉課", "事業棚卸し", "B"),
]
for sh, na, ta, item, pri in sheets_info:
    ws.cell(row=r, column=1, value=sh).font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=1).alignment = cl
    ws.cell(row=r, column=1).fill = F(LBLUE)
    ws.cell(row=r, column=1).border = border
    ms(ws, f"B{r}:C{r}", na, f_body, None, cl)
    ws.cell(row=r, column=4, value=ta).font = f_body
    ws.cell(row=r, column=4).alignment = cl
    ws.cell(row=r, column=4).border = border
    ms(ws, f"E{r}:F{r}", item, f_body, None, cl)
    pri_font = f_pri_a if pri == "A" else f_pri_b
    pri_fill = F(LORANGE) if pri == "A" else F(LBLUE)
    ws.cell(row=r, column=7, value=pri).font = pri_font
    ws.cell(row=r, column=7).alignment = cc
    ws.cell(row=r, column=7).fill = pri_fill
    ws.cell(row=r, column=7).border = border
    ws.row_dimensions[r].height = 32
    r += 1

# 3. 凡例
r += 1
ms(ws, f"A{r}:G{r}", "3．凡例", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

legends = [
    (INPUT_YELLOW, "記入欄（町担当者にご記入いただくセル）"),
    (LGREEN, "確認済み（実績データ確認サマリー等で既知の値・参考表示）"),
    (LGRAY, "項目見出し（記入不要）"),
    ("FFF2CC", "注記・コメント枠"),
]
for color, mean in legends:
    ws.cell(row=r, column=1).fill = F(color)
    ws.cell(row=r, column=1).border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 4. 記入のお願い
r += 1
ms(ws, f"A{r}:G{r}", "4．記入のお願い", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

notes = [
    "①  数値は原則「令和6年度」を基本とし、R7年度の最新値があれば併せてご記入ください。",
    "②  不明・該当なしの場合は「─」または「0」とご記入ください。空欄のままでも結構です。",
    "③  優先度A（赤）は計画素案の根幹となる項目です。優先的にご対応ください。",
    "④  記入後、東京コンサルティング（若山）までデータでご返送ください。",
    "⑤  数値の根拠（出典・計算方法等）があれば、各シートの備考欄にご記入ください。",
]
for n in notes:
    ms(ws, f"A{r}:G{r}", n, f_body, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

# 列幅
widths = [22, 18, 18, 18, 14, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

setup_page(ws)

# ===========================================================
# 共通ヘルパー：各実績シート用
# ===========================================================
def make_record_sheet(name, title, subtitle, columns, col_widths, rows_data, notes=None):
    """実績シート作成（INPUT=黄、既知=緑、項目=グレーで色分け）"""
    ws = wb.create_sheet(name)
    lc = get_column_letter(len(columns))
    
    ms(ws, f"A1:{lc}1", title, f_title, F(NAVY), cc)
    ws.row_dimensions[1].height = 26
    ms(ws, f"A2:{lc}2", subtitle, f_sub, F(BLUE), cl)
    ws.row_dimensions[2].height = 20
    
    # ヘッダ行
    for i, (h, w) in enumerate(zip(columns, col_widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = f_head
        c.fill = F(BLUE)
        c.alignment = cc
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30
    
    # データ行
    r = 4
    for row in rows_data:
        if isinstance(row, str):
            # セクション見出し行
            ms(ws, f"A{r}:{lc}{r}", row, Font(name="游ゴシック", size=11, bold=True, color=NAVY), F(LBLUE), cl)
            ws.row_dimensions[r].height = 22
            r += 1
            continue
        # 通常行：(項目名, 値1, 値2, ...)
        item = row[0]
        vals = row[1:]
        c = ws.cell(row=r, column=1, value=item)
        c.font = Font(name="游ゴシック", size=10, bold=True)
        c.alignment = cl
        c.fill = F(LGRAY)
        c.border = border
        for j, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=j)
            cell.border = border
            if v == "INPUT":
                cell.value = ""
                cell.fill = F(INPUT_YELLOW)
                cell.font = f_input
                cell.alignment = cc
            elif isinstance(v, tuple) and len(v) == 2 and v[0] == "KNOWN":
                cell.value = v[1]
                cell.fill = F(LGREEN)
                cell.font = f_known
                cell.alignment = cc
            elif v == "":
                cell.fill = F(LGRAY)
            else:
                cell.value = v
                cell.font = f_body
                cell.alignment = cl
        ws.row_dimensions[r].height = 26
        r += 1
    
    # 注記
    if notes:
        r += 1
        for note in notes:
            ms(ws, f"A{r}:{lc}{r}", "※ " + note, f_note, F("FFF2CC"), cl)
            ws.row_dimensions[r].height = 24
            r += 1
    
    setup_page(ws)
    return ws

# ===========================================================
# 01_KPI一覧
# ===========================================================
make_record_sheet(
    "01_KPI一覧",
    "01　第9期計画 KPI目標値・実績・達成率",
    "■第9期計画で設定されたKPI指標について、目標値・実績値・達成率をご記入ください",
    ["基本目標", "KPI項目", "単位", "第9期目標値\n(R6設定)", "R6実績", "R7実績\n(最新)", "達成率", "備考"],
    [16, 28, 8, 14, 12, 12, 10, 22],
    [
        "【基本目標1：健康づくりと介護予防の推進】",
        ("基本目標1", "通いの場参加率（65歳以上）", "%", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標1", "一般介護予防事業 年間実施回数", "回/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標1", "特定健診受診率（65歳以上）", "%", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標1", "ユニバーサルサポーター活動回数", "回/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        "【基本目標2：高齢者が安心して暮らせる仕組みづくり】",
        ("基本目標2", "ふれあいネットワーク活動員数", "人", "INPUT", ("KNOWN", "145名(参考)"), "INPUT", "INPUT", "活動員15+協力員130"),
        ("基本目標2", "緊急通報装置 設置件数", "件", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標2", "高齢者紙おむつ等支給件数", "件/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標2", "福祉移送サービス利用者数", "延べ/年", "INPUT", "INPUT", "INPUT", "INPUT", "社協・NPO"),
        "【基本目標3：在宅生活の継続支援】",
        ("基本目標3", "在宅医療・介護連携会議 開催回数", "回/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標3", "家族介護教室 開催回数", "回/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標3", "ケアプラン外部委託件数", "件", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        "【基本目標4：介護サービスの質の確保と提供体制】",
        ("基本目標4", "町内介護事業所数", "か所", "INPUT", "INPUT", "INPUT", "INPUT", "種別別"),
        ("基本目標4", "ケアプラン点検実施件数", "件/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標4", "介護人材の不足感", "%", "INPUT", "INPUT", "INPUT", "INPUT", "把握範囲で"),
        "【基本目標5：地域包括ケアシステムの深化】",
        ("基本目標5", "地域包括支援センター 総合相談件数", "件/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標5", "自立支援型地域ケア会議 開催回数", "回/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標5", "成年後見制度 利用件数", "件", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
        ("基本目標5", "高齢者虐待 相談・通報件数", "件/年", "INPUT", "INPUT", "INPUT", "INPUT", "INPUT"),
    ],
    notes=[
        "達成率 = 実績 ÷ 目標値 × 100。実績が目標を上回る場合は100%を超えます。",
        "第9期計画でKPIが設定されていない項目は、空欄または「未設定」とご記入ください。",
        "本シートで把握できないKPIは、別途追加でご相談ください（第10期で新規KPIを設定する場合の検討材料）。",
    ]
)

# ===========================================================
# 02_介護予防_健康
# ===========================================================
make_record_sheet(
    "02_介護予防_健康",
    "02　介護予防・健康づくり事業 実績",
    "■通いの場・サポーター活動・健診関連の実績をご記入ください",
    ["事業区分", "事業名", "指標", "R6実績", "R7実績", "備考"],
    [16, 30, 14, 12, 12, 16],
    [
        "【一般介護予防事業】",
        ("一般介護予防", "元気まんてん教室", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("一般介護予防", "元気まんてん教室", "延べ参加者数(人)", "INPUT", "INPUT", "INPUT"),
        ("一般介護予防", "スマイル教室", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("一般介護予防", "スマイル教室", "延べ参加者数(人)", "INPUT", "INPUT", "INPUT"),
        ("一般介護予防", "パドル運動教室", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("一般介護予防", "パドル運動教室", "延べ参加者数(人)", "INPUT", "INPUT", "INPUT"),
        "【通いの場（サロン等）】",
        ("通いの場", "介護予防サロン", "開催箇所数(か所)", "INPUT", "INPUT", "INPUT"),
        ("通いの場", "介護予防サロン", "年間開催延べ回数(回)", "INPUT", "INPUT", "INPUT"),
        ("通いの場", "介護予防サロン", "延べ参加者数(人)", "INPUT", "INPUT", "INPUT"),
        "【ユニバーサルサポーター制度（川崎町独自）】",
        ("UV-S", "介護予防サロン サポーター", "登録者数(人)", ("KNOWN", "84名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "スマイルサポーター", "登録者数(人)", ("KNOWN", "40名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "レクリエーションサポーター", "登録者数(人)", ("KNOWN", "29名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "傾聴サポーター", "登録者数(人)", ("KNOWN", "24名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "生活支援サポーター(SC)", "登録者数(人)", ("KNOWN", "25名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "ふれあいNW 活動員", "登録者数(人)", ("KNOWN", "15名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "ふれあいNW 協力員", "登録者数(人)", ("KNOWN", "130名(地福資料)"), "INPUT", "最新値を→"),
        ("UV-S", "やすらぎ・パドル等 その他", "登録者数(人)", "INPUT", "INPUT", "全種別の合計把握"),
        ("UV-S", "サポーター実活動回数(全種別計)", "活動回数(回/年)", "INPUT", "INPUT", "INPUT"),
        "【健診・健康管理】",
        ("健診", "特定健診", "対象者数(人)", "INPUT", "INPUT", "INPUT"),
        ("健診", "特定健診", "受診者数(人)", "INPUT", "INPUT", "INPUT"),
        ("健診", "特定健診", "受診率(%)", "INPUT", "INPUT", "INPUT"),
        ("健診", "後期高齢者健診", "受診者数(人)", "INPUT", "INPUT", "INPUT"),
        ("健診", "後期高齢者健診", "受診率(%)", "INPUT", "INPUT", "INPUT"),
        ("健診", "口腔ケア・歯科健診", "実施回数(回/年)", "INPUT", "INPUT", "オーラルフレイル予防"),
        ("健診", "栄養指導・配食サービス", "対象者数(人/年)", "INPUT", "INPUT", "INPUT"),
    ],
    notes=[
        "ユニバーサルサポーター（UV-S）の値は第2期地域福祉計画資料の値を参考表示しています。最新値があればご記入ください。",
        "通いの場の集計は介護予防の重要指標です。市町村介護予防事業状況調査（厚労省）に合わせた集計が望ましいです。",
        "健診受診率は健康かわさき21計画・データヘルス計画との整合を確認ください。",
    ]
)

# ===========================================================
# 03_在宅生活支援
# ===========================================================
make_record_sheet(
    "03_在宅生活支援",
    "03　在宅生活支援事業 実績",
    "■紙おむつ・配食・緊急通報・移動支援・透析助成等の独自施策の実績をご記入ください",
    ["事業区分", "事業名", "指標", "R6実績", "R7実績", "備考"],
    [16, 30, 14, 12, 12, 16],
    [
        "【独自施策（紙おむつ等）】",
        ("独自施策", "高齢者紙おむつ等支給事業", "対象者数(人/年)", "INPUT", "INPUT", "INPUT"),
        ("独自施策", "高齢者紙おむつ等支給事業", "支給件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("独自施策", "高齢者紙おむつ等支給事業", "事業費(円/年)", "INPUT", "INPUT", "INPUT"),
        ("独自施策", "高齢者世帯エアコン購入支援", "件数(件/年)", "INPUT", "INPUT", "R7.10開始"),
        ("独自施策", "高齢者世帯エアコン購入支援", "事業費(円/年)", "INPUT", "INPUT", "R7.10開始"),
        ("独自施策", "人工透析患者通院交通費助成", "対象者数(人/年)", "INPUT", "INPUT", "INPUT"),
        ("独自施策", "人工透析患者通院交通費助成", "事業費(円/年)", "INPUT", "INPUT", "INPUT"),
        "【見守り・緊急支援】",
        ("見守り", "緊急通報装置設置事業", "設置件数(件・累計)", "INPUT", "INPUT", "INPUT"),
        ("見守り", "緊急通報装置設置事業", "新規設置件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("見守り", "緊急通報装置設置事業", "通報対応件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("見守り", "ふれあいネットワーク活動", "訪問件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("見守り", "民生委員による高齢者訪問", "件数(件/年)", "INPUT", "INPUT", "INPUT"),
        "【配食サービス】",
        ("配食", "高齢者配食サービス", "対象者数(人)", "INPUT", "INPUT", "INPUT"),
        ("配食", "高齢者配食サービス", "提供食数(食/年)", "INPUT", "INPUT", "INPUT"),
        ("配食", "高齢者配食サービス", "事業費(円/年)", "INPUT", "INPUT", "INPUT"),
        "【移動支援（R7.3でタクシー助成終了→3層構造へ移行）】",
        ("移動支援", "(旧)高齢者外出タクシー助成", "対象者数(人/年)", "INPUT", ("KNOWN", "R7.3終了"), "R7.3で終了"),
        ("移動支援", "福祉移送サービス(社協)", "利用者数(延べ/年)", "INPUT", "INPUT", "社協運営"),
        ("移動支援", "福祉移送サービス(社協)", "うち高齢者割合(%)", "INPUT", "INPUT", "把握範囲で"),
        ("移動支援", "福祉移送サービス(NPO)", "利用者数(延べ/年)", "INPUT", "INPUT", "NPO運営"),
        ("移動支援", "デマンドバス", "利用者数(延べ/年)", "INPUT", "INPUT", "町民生活課"),
        ("移動支援", "デマンドバス", "うち高齢者割合(%)", "INPUT", "INPUT", "把握範囲で"),
        ("移動支援", "町民バス", "利用者数(延べ/年)", "INPUT", "INPUT", "地域振興課"),
    ],
    notes=[
        "移動支援は所管が分かれています（地域振興課・町民生活課・社協・NPO）。各所管からの情報集約をお願いします。",
        "高齢者世帯エアコン購入支援は令和7年10月開始の新規事業です。R7実績のみの記入で結構です。",
        "ふれあいネットワーク活動は社協所管。協力員130名・活動員15名による地域見守りの実績を把握ください。",
    ]
)

# ===========================================================
# 04_認知症_包括
# ===========================================================
make_record_sheet(
    "04_認知症_包括",
    "04　認知症施策・地域包括支援センター 実績",
    "■認知症基本法対応の中核データ。サポーター・カフェ・初期集中支援等の実績をご記入ください",
    ["事業区分", "事業名", "指標", "R6実績", "R7実績", "備考"],
    [16, 30, 14, 12, 12, 16],
    [
        "【認知症サポーター養成】",
        ("認知症", "認知症サポーター養成講座", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症サポーター養成講座", "新規受講者数(人/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症サポーター", "累計養成数(人)", ("KNOWN", "550名(地福資料)"), "INPUT", "最新累計を→"),
        ("認知症", "認知症キャラバンメイト", "新規養成数(人/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症キャラバンメイト", "累計数(人)", ("KNOWN", "73名(地福資料)"), "INPUT", "最新累計を→"),
        ("認知症", "企業・学校サポーター", "養成数(人/年)", "INPUT", "INPUT", "基本法対応"),
        "【認知症カフェ・本人活動】",
        ("認知症", "認知症カフェ「喫茶みかん」", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症カフェ「喫茶みかん」", "延べ参加者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症本人ミーティング", "開催回数(回/年)", "INPUT", "INPUT", "基本法新設"),
        ("認知症", "チームオレンジ", "整備状況", "INPUT", "INPUT", "未整備なら『未整備』"),
        "【早期発見・支援】",
        ("認知症", "もの忘れ相談", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "もの忘れ相談", "延べ相談件数(件)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症初期集中支援チーム", "チーム員会議(回/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症初期集中支援チーム", "対象者数(人/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症初期集中支援チーム", "訪問件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("認知症", "認知症地域支援推進員", "配置人数(人)", "INPUT", "INPUT", "INPUT"),
        "【地域包括支援センター業務】",
        ("包括", "地域包括支援センター職員数", "保健師(人)", ("KNOWN", "3名(KO確認)"), "INPUT", "最新体制を→"),
        ("包括", "地域包括支援センター職員数", "認定調査員(人)", ("KNOWN", "1名(KO確認)"), "INPUT", "最新体制を→"),
        ("包括", "地域包括支援センター職員数", "ケアマネ(人)", ("KNOWN", "1名(KO確認・会計年度任用)"), "INPUT", "最新体制を→"),
        ("包括", "総合相談支援業務", "件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("包括", "総合相談支援業務", "延べ件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("包括", "自立支援型地域ケア会議", "開催数(回/年)", "INPUT", "INPUT", "INPUT"),
        "【権利擁護】",
        ("権利擁護", "成年後見制度利用", "件数(件)", "INPUT", "INPUT", "INPUT"),
        ("権利擁護", "成年後見市町村長申立て", "件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("権利擁護", "日常生活自立支援事業", "利用件数(件)", "INPUT", "INPUT", "社協"),
        ("権利擁護", "高齢者虐待相談・通報", "件数(件/年)", "INPUT", "INPUT", "INPUT"),
        ("権利擁護", "高齢者虐待 認定件数", "件数(件/年)", "INPUT", "INPUT", "INPUT"),
        "【医療介護連携】",
        ("医療介護", "在宅医療・介護連携講話", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("医療介護", "保健医療福祉介護連携会議", "開催回数(回/年)", "INPUT", "INPUT", "INPUT"),
        ("医療介護", "国保川崎病院との連携", "実施内容", "INPUT", "INPUT", "町内拠点病院"),
    ],
    notes=[
        "認知症サポーター累計550名・キャラバンメイト累計73名は第2期地域福祉計画資料の値です。最新累計をご記入ください。",
        "地域包括支援センターの体制（保健師3＋認定調査1＝計4名）はキックオフ会議での確認値です。最新体制をご記入ください。",
        "認知症本人ミーティングは認知症基本法第3条「本人の意思を尊重」に対応する新規取組です。第10期で新設を検討します。",
        "チームオレンジは認知症サポーターのうちステップアップ研修を修了したメンバーで編成。整備状況をご確認ください。",
    ]
)

# ===========================================================
# 05_介護サービス_人材
# ===========================================================
make_record_sheet(
    "05_介護サービス_人材",
    "05　介護サービス利用実績・事業所人材",
    "■認定者数・サービス利用・町内事業所の実態をご記入ください",
    ["事業区分", "事業名・項目", "指標", "R6実績", "R7実績", "備考"],
    [16, 30, 14, 12, 12, 16],
    [
        "【認定者数（年度末）】",
        ("認定者", "要支援1", "認定者数(人)", "INPUT", ("KNOWN", "13名（R7.6受給）"), "受給者数で参考"),
        ("認定者", "要支援2", "認定者数(人)", "INPUT", ("KNOWN", "50名（R7.6受給）"), "受給者数で参考"),
        ("認定者", "要介護1", "認定者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "要介護2", "認定者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "要介護3", "認定者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "要介護4", "認定者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "要介護5", "認定者数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "認定者計", "総数(人)", "INPUT", "INPUT", "INPUT"),
        ("認定者", "第1号被保険者数", "総数(人)", "INPUT", ("KNOWN", "3,244人（R7.6）"), "保険者データ"),
        ("認定者", "認定率", "%", "INPUT", "INPUT", "認定者÷被保険者"),
        "【施設サービス利用】",
        ("施設利用", "特別養護老人ホーム", "町内利用者数(人)", "INPUT", "INPUT", "INPUT"),
        ("施設利用", "特別養護老人ホーム", "町外利用者数(人)", "INPUT", "INPUT", "住所地特例含む"),
        ("施設利用", "介護老人保健施設", "町内利用者数(人)", "INPUT", "INPUT", "INPUT"),
        ("施設利用", "介護老人保健施設", "町外利用者数(人)", "INPUT", "INPUT", "住所地特例含む"),
        ("施設利用", "住所地特例該当者", "総数(人)", "INPUT", ("KNOWN", "24名（R7.6）"), "保険者データ"),
        "【町内介護事業所の状況】",
        ("事業所", "町内介護事業所数(種別計)", "か所", "INPUT", "INPUT", "訪問/通所/施設/地密"),
        ("事業所", "うち訪問系", "か所", "INPUT", "INPUT", "INPUT"),
        ("事業所", "うち通所系", "か所", "INPUT", "INPUT", "INPUT"),
        ("事業所", "うち施設系", "か所", "INPUT", "INPUT", "INPUT"),
        ("事業所", "うち地域密着型", "か所", "INPUT", "INPUT", "GH等"),
        "【介護人材の実態（事業所照会で把握可能な範囲で）】",
        ("人材", "町内介護職員数", "常勤換算(人)", "INPUT", "INPUT", "把握範囲で"),
        ("人材", "介護職員の過不足感", "充足/やや不足/大幅不足", "INPUT", "INPUT", "事業者連絡会照会"),
        ("人材", "直近1年の離職率", "%", "INPUT", "INPUT", "把握範囲で"),
        ("人材", "処遇改善加算 取得状況", "取得事業所数(Ⅰ〜Ⅴ別)", "INPUT", "INPUT", "事業者連絡会照会"),
        ("人材", "外国人介護人材の受入", "人数(人)", "INPUT", "INPUT", "EPA・技能実習等"),
        ("人材", "介護ロボット・ICT導入", "導入事業所数(か所)", "INPUT", "INPUT", "INPUT"),
    ],
    notes=[
        "認定者数は介護保険事業状況報告（保険者データ）から抽出可能です。年度末（3月末時点）の値が基本です。",
        "施設利用の町内/町外区分は、住所地特例該当者の所在自治体内訳を把握できれば、広域連携の検討材料になります。",
        "介護人材の実態は事業所照会または事業者連絡会等で把握可能な範囲でご記入ください。新規調査は不要です。",
    ]
)

# ===========================================================
# 06_第10期反映方針
# ===========================================================
ws = wb.create_sheet("06_第10期反映方針")
ms(ws, "A1:F1", "06　第9期事業の第10期への反映方針（事業棚卸し）",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "■第9期で実施した各事業について、第10期での扱い（継続/拡充/縮小/廃止）と理由をご記入ください",
   f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# ヘッダ
heads = ["事業区分", "事業名", "第9期の状況", "第10期方針", "見直し理由・課題", "優先度"]
widths = [14, 28, 18, 16, 26, 10]
for i, (h, w) in enumerate(zip(heads, widths), 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = f_head
    c.fill = F(BLUE)
    c.alignment = cc
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[3].height = 30

# 第10期方針の選択肢説明
ms(ws, "A4:F4", "【第10期方針の記入例：継続／拡充／縮小／廃止／新規／統合】",
   Font(name="游ゴシック", size=9, italic=True, color=ORANGE), F("FFF2CC"), cl)
ws.row_dimensions[4].height = 22

# 主要事業のリスト（第9期計画でわかっている事業を網羅）
events = [
    "【健康づくり・介護予防】",
    ("健康・予防", "元気まんてん教室", "実施中"),
    ("健康・予防", "スマイル教室", "実施中"),
    ("健康・予防", "パドル運動教室", "実施中"),
    ("健康・予防", "やすらぎサロン", "実施中"),
    ("健康・予防", "通いの場（地区サロン）", "実施中"),
    ("健康・予防", "ユニバーサルサポーター養成", "実施中(独自)"),
    "【見守り・地域支援】",
    ("地域支援", "ふれあいネットワーク事業", "実施中(社協)"),
    ("地域支援", "民生委員による訪問", "実施中"),
    ("地域支援", "緊急通報装置設置事業", "実施中"),
    "【在宅生活支援（独自施策）】",
    ("独自施策", "高齢者紙おむつ等支給", "実施中"),
    ("独自施策", "高齢者世帯エアコン購入支援", "R7.10開始(新規)"),
    ("独自施策", "人工透析患者通院交通費助成", "実施中"),
    ("独自施策", "配食サービス", "実施中"),
    "【移動支援】",
    ("移動支援", "高齢者外出タクシー助成", "R7.3 終了"),
    ("移動支援", "福祉移送サービス(社協)", "実施中"),
    ("移動支援", "福祉移送サービス(NPO)", "実施中"),
    ("移動支援", "デマンドバス", "実施中"),
    ("移動支援", "町民バス", "実施中"),
    "【認知症施策】",
    ("認知症", "認知症サポーター養成", "実施中"),
    ("認知症", "キャラバンメイト養成", "実施中"),
    ("認知症", "認知症カフェ「喫茶みかん」", "実施中"),
    ("認知症", "もの忘れ相談", "実施中"),
    ("認知症", "認知症初期集中支援チーム", "実施中"),
    ("認知症", "認知症地域支援推進員", "実施中"),
    ("認知症", "チームオレンジ整備", "未整備"),
    ("認知症", "認知症本人ミーティング", "未実施"),
    "【在宅医療・介護連携】",
    ("医療介護", "在宅医療・介護連携推進事業", "実施中"),
    ("医療介護", "国保川崎病院との連携", "実施中"),
    ("医療介護", "広域医療連携(県南中核病院等)", "実施中"),
    "【介護人材・サービス基盤】",
    ("人材・基盤", "介護人材確保・育成", "実施中"),
    ("人材・基盤", "ケアマネ研修・ケアプラン点検", "実施中"),
    ("人材・基盤", "事業者連絡会", "実施中"),
    "【家族介護者支援】",
    ("家族支援", "家族介護教室", "実施中"),
    ("家族支援", "レスパイト(短期入所)活用支援", "実施中"),
    ("家族支援", "介護離職防止支援", "新規検討"),
]

r = 5
for ev in events:
    if isinstance(ev, str):
        # セクション見出し
        ms(ws, f"A{r}:F{r}", ev, Font(name="游ゴシック", size=11, bold=True, color=NAVY), F(LBLUE), cl)
        ws.row_dimensions[r].height = 22
        r += 1
        continue
    kbn, name, status = ev
    # 区分
    c = ws.cell(row=r, column=1, value=kbn)
    c.font = Font(name="游ゴシック", size=10, bold=True)
    c.alignment = cl
    c.fill = F(LGRAY)
    c.border = border
    # 事業名
    c = ws.cell(row=r, column=2, value=name)
    c.font = f_body
    c.alignment = cl
    c.border = border
    # 第9期の状況
    c = ws.cell(row=r, column=3, value=status)
    c.font = f_body
    c.alignment = cl
    c.fill = F(LGREEN)
    c.border = border
    # 第10期方針（INPUT）
    c = ws.cell(row=r, column=4)
    c.fill = F(INPUT_YELLOW)
    c.font = f_input
    c.alignment = cc
    c.border = border
    # 見直し理由・課題（INPUT）
    c = ws.cell(row=r, column=5)
    c.fill = F(INPUT_YELLOW)
    c.font = f_input
    c.alignment = cl
    c.border = border
    # 優先度（INPUT）
    c = ws.cell(row=r, column=6)
    c.fill = F(INPUT_YELLOW)
    c.font = f_input
    c.alignment = cc
    c.border = border
    ws.row_dimensions[r].height = 26
    r += 1

# 注記
r += 1
for note in [
    "事業の取捨選択は計画素案 第3章「第10期に向けた課題」と連動します。事業棚卸しを通じて、川崎町独自の重点施策を整理します。",
    "「廃止」を選ぶ場合は、住民への影響と代替策の検討が必要です。代替策があれば「見直し理由・課題」欄にご記入ください。",
    "「新規」事業を提案される場合は、本シート末尾に追記ください。認知症本人ミーティング・チームオレンジ整備は基本法対応の候補です。",
    "本シートの結果は、計画素案Ver.2.0 第3章 3-4「第10期に向けた課題」及び第5章「施策の展開」の根拠データとなります。",
]:
    ms(ws, f"A{r}:F{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 28
    r += 1

setup_page(ws)

# 保存
out = "/home/claude/kawasaki_work/川崎町_第9期実績一覧_町記入用.xlsx"
wb.save(out)
print(f"作成完了: {out}")
print(f"シート数: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws_s = wb[s]
    print(f"  - {s}: {ws_s.max_row}行 x {ws_s.max_column}列")
