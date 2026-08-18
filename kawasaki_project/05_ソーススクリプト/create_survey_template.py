"""
川崎町第10期 アンケート集計・分析テンプレート

目的：令和8年7月末のアンケート回収後、すぐに集計を開始できるよう
事前に集計表・クロス集計枠・自由記述分類・計画反映ガイドを準備する。

5シート構成：
00_使い方            : 集計手順・凡例・スケジュール
01_一般高齢者_単純集計: 国標準＋追加5問の単純集計（選択肢別件数・構成比）
02_認定者_単純集計   : 国標準＋追加6問の単純集計
03_クロス集計        : 地区×移動手段、年齢×外出困難等の重要クロス
04_自由記述_5分類    : 移動／介護／見守り／サービス不足／医療連携
05_計画反映ガイド    : 各設問結果を計画素案のどの章に反映するか

設計方針：
- 黄色=集計値記入欄、緑=既知の確定値、グレー=項目見出し
- 各問の選択肢数・サンプル数枠を明示
- 計画素案v1.3の章・節と対応付け（ガイド機能）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# カラーパレット（既存成果物と統一）
NAVY = "1F3864"; BLUE = "2F5597"; LBLUE = "DAE3F3"
ORANGE = "ED7D31"; LORANGE = "FCE4D6"
GREEN = "548235"; LGREEN = "E2EFDA"
GRAY = "808080"; LGRAY = "F2F2F2"
INPUT = "FFFFCC"; WHITE = "FFFFFF"
PURPLE = "7030A0"; LPURPLE = "E4D6F0"

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def F(c): return PatternFill("solid", fgColor=c)
cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
cl = Alignment(horizontal="left", vertical="center", wrap_text=True)
cr = Alignment(horizontal="right", vertical="center", wrap_text=True)

f_title = Font(name="游ゴシック", size=13, bold=True, color=WHITE)
f_sub   = Font(name="游ゴシック", size=10, italic=True, color=WHITE)
f_head  = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
f_section = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
f_q      = Font(name="游ゴシック", size=10, bold=True, color="C00000")
f_body  = Font(name="游ゴシック", size=10)
f_input = Font(name="游ゴシック", size=10, color="0000FF")
f_known = Font(name="游ゴシック", size=10, bold=True, color=GREEN)
f_note  = Font(name="游ゴシック", size=9, italic=True, color="595959")

wb = Workbook()
wb.remove(wb.active)

def ms(ws, rng, val, font, fill, align):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = val; c.font = font
    if fill is not None: c.fill = fill
    c.alignment = align
    from openpyxl.utils.cell import range_boundaries
    a, b, d, e = range_boundaries(rng)
    for r in range(b, e+1):
        for col in range(a, d+1):
            ws.cell(row=r, column=col).border = border

def setup_page(ws, orient="landscape"):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orient == "landscape" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

# ===========================================================
# 00_使い方
# ===========================================================
ws = wb.create_sheet("00_使い方")
ms(ws, "A1:G1", "川崎町第10期 アンケート集計・分析テンプレート", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 28
ms(ws, "A2:G2", "─ 令和8年7月末回収→8月集計→第1回策定委員会報告に対応する事前テンプレート ─", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# 1. 目的・前提
ms(ws, "A4:G4", "1．本テンプレートの目的", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22

ms(ws, "A5:G5", "本テンプレートは、令和8年7月末のアンケート回収後、すぐに集計作業を開始できるよう事前に集計表を整備するものです。回収後約2週間で第1回策定委員会（令和8年8月中旬）に集計結果を報告する予定です。", f_body, None, cl)
ws.row_dimensions[5].height = 36

# 2. アンケート概要
ms(ws, "A7:G7", "2．アンケート概要", f_head, F(BLUE), cl)
ws.row_dimensions[7].height = 22

survey_info = [
    ("対象", "一般高齢者調査：65歳以上の高齢者 1,000名／認定者調査：要支援・要介護認定者 300名"),
    ("発送", "令和8年6月下旬発送"),
    ("回収", "令和8年7月末締切（約1ヶ月）"),
    ("集計", "回収後 約2週間（弊社で実施）"),
    ("報告", "令和8年8月中旬 第1回策定委員会"),
    ("反映", "計画素案Ver.2.0 第3章・第5章・第6章・第7章に反映"),
]
r = 8
for k, v in survey_info:
    c = ws.cell(row=r, column=1, value=k)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY)
    c.alignment = cc; c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:G{r}", v, f_body, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

# 3. シート構成
r += 1
ms(ws, f"A{r}:G{r}", "3．シート構成と作業順序", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

# ヘッダ
ws.cell(row=r, column=1, value="No").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="シート").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ms(ws, f"C{r}:D{r}", "内容", f_head, F(NAVY), cc)
ms(ws, f"E{r}:F{r}", "作業内容", f_head, F(NAVY), cc)
ws.cell(row=r, column=7, value="所要").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 22
r += 1

work_order = [
    ("1", "01_一般高齢者_単純集計", "国標準＋追加問の単純集計", "選択肢別件数・構成比を入力", "3〜5日"),
    ("2", "02_認定者_単純集計", "国標準＋追加問の単純集計", "選択肢別件数・構成比を入力", "2〜3日"),
    ("3", "03_クロス集計", "重要クロス集計", "地区×移動・年齢×外出困難等", "2〜3日"),
    ("4", "04_自由記述_5分類", "自由記述の5分類集計", "テキスト分類・代表的意見抽出", "3〜4日"),
    ("5", "05_計画反映ガイド", "計画素案への反映マトリクス", "各設問結果と章・節の対応確認", "1日"),
]

for no, sh, na, work, span in work_order:
    ws.cell(row=r, column=1, value=no).font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).fill = F(LORANGE); ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=sh).font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=2).alignment = cl; ws.cell(row=r, column=2).fill = F(LBLUE); ws.cell(row=r, column=2).border = border
    ms(ws, f"C{r}:D{r}", na, f_body, None, cl)
    ms(ws, f"E{r}:F{r}", work, f_body, None, cl)
    ws.cell(row=r, column=7, value=span).font = f_body
    ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
    ws.row_dimensions[r].height = 28
    r += 1

# 4. 凡例
r += 1
ms(ws, f"A{r}:G{r}", "4．凡例", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

legends = [
    (INPUT, "集計値記入欄（件数・構成比・自由記述等）"),
    (LGREEN, "確定値（事前にわかっている数値）"),
    (LGRAY, "項目見出し（記入不要）"),
    (LPURPLE, "計画素案への反映先（章・節）"),
]
for color, mean in legends:
    ws.cell(row=r, column=1).fill = F(color)
    ws.cell(row=r, column=1).border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 5. 注意事項
r += 1
ms(ws, f"A{r}:G{r}", "5．集計時の注意事項", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

notes = [
    "① 構成比は「回答者数」を母数として計算（無回答を除く）。N（回答者数）を必ず明記すること。",
    "② 複数回答可の設問は「該当者数÷回答者数」で計算（合計は100%を超える）。設問ごとにこれを明記。",
    "③ クロス集計のセルが10件未満となる場合は、参考値として扱い、計画に反映する際は注意書きを付す。",
    "④ 自由記述の分類は本テンプレートの5分類を基本とし、それに該当しないものは「その他」に集約。",
    "⑤ 全ての集計表は、第1回策定委員会（令和8年8月中旬）で報告し、計画素案Ver.2.0に反映する。",
]
for n in notes:
    ms(ws, f"A{r}:G{r}", n, f_body, None, cl)
    ws.row_dimensions[r].height = 26
    r += 1

# 列幅
widths = [6, 22, 14, 14, 14, 14, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

setup_page(ws)

# ===========================================================
# 共通ヘルパー：単純集計シート用の設問ブロック生成
# ===========================================================
def add_question_block(ws, start_row, q_no, q_text, choices, n_input=True):
    """1問分の集計表ブロックを追加。choices は選択肢のリスト。
    Returns: 次に書き込む行番号
    """
    r = start_row
    # 設問タイトル
    ms(ws, f"A{r}:F{r}", f"Q{q_no}  {q_text}", f_q, F(LORANGE), cl)
    ws.row_dimensions[r].height = 24
    r += 1
    
    # N（回答者数）
    ws.cell(row=r, column=1, value="N（回答者数）").font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).fill = F(LGRAY); ws.cell(row=r, column=1).border = border
    c = ws.cell(row=r, column=2)
    c.fill = F(INPUT); c.font = f_input; c.alignment = cc; c.border = border
    ms(ws, f"C{r}:F{r}", "→ 回答があった人数を入力（無回答を除く）", f_note, F(LGRAY), cl)
    ws.row_dimensions[r].height = 22
    r += 1
    
    # ヘッダ
    headers = ["選択肢", "件数", "構成比(%)", "備考", "", ""]
    widths_h = [30, 12, 14, 30, 0, 0]
    for i, h in enumerate(headers[:4], 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
    ms(ws, f"D{r}:F{r}", "備考", f_head, F(BLUE), cc)
    ws.row_dimensions[r].height = 22
    r += 1
    
    # 選択肢行
    for choice in choices:
        c = ws.cell(row=r, column=1, value=choice)
        c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
        c = ws.cell(row=r, column=2)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
        c = ws.cell(row=r, column=3)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
        c.number_format = "0.0"
        ms(ws, f"D{r}:F{r}", "", f_body, F(INPUT), cl)
        ws.row_dimensions[r].height = 22
        r += 1
    
    # 合計行
    c = ws.cell(row=r, column=1, value="合計")
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl; c.fill = F(LBLUE); c.border = border
    c = ws.cell(row=r, column=2)
    c.fill = F(LBLUE); c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cr; c.border = border
    c.value = "= " + " + ".join([f"B{rr}" for rr in range(r-len(choices), r)])
    c = ws.cell(row=r, column=3, value=100.0)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cr; c.fill = F(LBLUE); c.border = border
    c.number_format = "0.0"
    ms(ws, f"D{r}:F{r}", "※複数回答可の設問は100%超になる場合あり", f_note, F(LBLUE), cl)
    ws.row_dimensions[r].height = 22
    r += 2  # 設問間のスペース
    
    return r

# ===========================================================
# 01_一般高齢者_単純集計
# ===========================================================
ws = wb.create_sheet("01_一般高齢者_単純集計")
ms(ws, "A1:F1", "01　一般高齢者ニーズ調査 単純集計（N=対象1,000名）",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "■国標準設問＋川崎町追加5問の単純集計表",
   f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# 列幅設定
widths = [30, 12, 14, 16, 16, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 基本属性
r = 4
ms(ws, f"A{r}:F{r}", "■基本属性", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "0-1", "性別",
    ["男性", "女性", "答えたくない"])
r = add_question_block(ws, r, "0-2", "年齢",
    ["65〜69歳", "70〜74歳", "75〜79歳", "80〜84歳", "85〜89歳", "90歳以上"])
r = add_question_block(ws, r, "0-3", "お住まいの地区",
    ["裏丁・上下・本荒町・中新町", "前川・青根", "今宿", "川内", "本砂金", "小野",
     "小沢・支倉・碁石・支倉台"])
r = add_question_block(ws, r, "0-4", "世帯類型",
    ["独居", "夫婦のみ", "子と同居（未婚・既婚問わず）", "子の家族と同居（孫含む）",
     "その他", "答えたくない"])

# 健康・介護予防
ms(ws, f"A{r}:F{r}", "■健康・介護予防", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "1-1", "現在の健康状態",
    ["とても健康", "まあ健康", "あまり健康でない", "健康でない", "わからない"])
r = add_question_block(ws, r, "1-2", "週何回外出していますか",
    ["毎日", "週4〜6回", "週2〜3回", "週1回", "ほぼ外出しない"])
r = add_question_block(ws, r, "1-3", "外出時の主な交通手段【複数回答可】",
    ["自家用車（自分で運転）", "家族の車", "町民バス", "デマンドバス",
     "社協・NPO移送サービス", "タクシー", "徒歩・自転車", "その他", "外出しない"])
r = add_question_block(ws, r, "1-4", "通いの場や集まりへの参加状況",
    ["週1回以上", "月1〜3回", "年数回", "参加していない", "あれば参加したい"])

# 移動・買い物・社会参加
ms(ws, f"A{r}:F{r}", "■移動・買い物・社会参加", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "2-1", "外出に困っていることがありますか",
    ["特に困っていない", "公共交通機関が少ない", "運転が不安・できない",
     "近くに目的地がない", "体力的に難しい", "その他"])
r = add_question_block(ws, r, "2-2", "町の移動支援を知っていますか【複数回答可】",
    ["町民バス", "デマンドバス", "社協・NPO移送サービス（福祉移送）", "いずれも知らない"])
r = add_question_block(ws, r, "2-3", "買い物先は主にどこですか",
    ["町内（コンビニ・スーパー等）", "白石市", "大河原町・柴田町・村田町",
     "仙台方面", "宅配・通販", "家族に依頼"])

# 認知症
ms(ws, f"A{r}:F{r}", "■認知症（基本法対応・追加設問）", f_section, F(LPURPLE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "3-1", "認知症についてどのような不安をお持ちですか【複数回答可】",
    ["自分自身がなることが不安", "家族がなることが不安", "ご近所の方への対応",
     "情報が少なくわからない", "特に不安はない"])
r = add_question_block(ws, r, "3-2", "認知症の相談窓口を知っていますか",
    ["地域包括支援センター", "認知症カフェ（喫茶みかん）", "国保川崎病院",
     "知っている窓口はない", "相談したことがある"])
r = add_question_block(ws, r, "3-3", "認知症サポーター養成講座を受講したことがありますか",
    ["はい", "いいえ", "受講したい", "わからない"])

# 介護・在宅
ms(ws, f"A{r}:F{r}", "■介護・在宅サービス", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "4-1", "将来介護が必要になった時、どのような生活を望みますか",
    ["自宅で家族の介護を受けたい", "自宅で介護サービスを利用したい",
     "施設（特養・老健等）に入りたい", "わからない"])
r = add_question_block(ws, r, "4-2", "介護や生活で困りごとがあった時の相談先【複数回答可】",
    ["家族・親族", "ご近所・友人", "民生委員", "地域包括支援センター",
     "ケアマネジャー", "町役場", "誰にも相談しない"])

# 川崎町追加（重点）
ms(ws, f"A{r}:F{r}", "■川崎町追加設問（重点）", f_section, F(LORANGE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "K-1", "町外医療機関を利用していますか",
    ["みやぎ県南中核病院（大河原）", "刈田綜合病院（白石）",
     "仙台市内の病院", "国保川崎病院のみ", "病院にかかっていない"])
r = add_question_block(ws, r, "K-2", "町独自支援（紙おむつ・エアコン購入等）を知っていますか",
    ["知っていて利用している", "知っているが利用していない", "知らない"])
r = add_question_block(ws, r, "K-3", "災害時の避難について不安がありますか",
    ["不安がある", "やや不安がある", "あまり不安はない", "全く不安はない"])

setup_page(ws, "portrait")

# ===========================================================
# 02_認定者_単純集計
# ===========================================================
ws = wb.create_sheet("02_認定者_単純集計")
ms(ws, "A1:F1", "02　要支援・要介護認定者調査 単純集計（N=対象300名）",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "■国標準設問＋川崎町追加6問の単純集計表",
   f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

widths = [30, 12, 14, 16, 16, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 4
ms(ws, f"A{r}:F{r}", "■基本属性", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "0-1", "性別",
    ["男性", "女性", "答えたくない"])
r = add_question_block(ws, r, "0-2", "年齢",
    ["65〜74歳", "75〜84歳", "85歳以上"])
r = add_question_block(ws, r, "0-3", "要介護度",
    ["要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"])
r = add_question_block(ws, r, "0-4", "現在の主な生活場所",
    ["自宅（自分で生活）", "自宅（家族と同居）", "サ高住・有料老人ホーム",
     "施設（特養・老健等）", "病院"])

# 介護サービス利用
ms(ws, f"A{r}:F{r}", "■介護サービス利用状況", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "1-1", "利用しているサービス【複数回答可】",
    ["訪問介護", "訪問看護", "通所介護（デイサービス）", "通所リハビリ（デイケア）",
     "短期入所（ショートステイ）", "福祉用具貸与・購入", "住宅改修", "施設サービス"])
r = add_question_block(ws, r, "1-2", "サービスへの満足度",
    ["とても満足", "まあ満足", "あまり満足でない", "不満", "わからない"])
r = add_question_block(ws, r, "1-3", "サービス利用で困っていること【複数回答可】",
    ["費用が高い", "希望のサービスがない", "事業者が遠い", "曜日・時間が合わない",
     "情報が少ない", "特に困っていない"])

# 家族介護
ms(ws, f"A{r}:F{r}", "■家族介護者の状況", f_section, F(LBLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "2-1", "主な介護者の年齢",
    ["50歳未満", "50〜64歳", "65〜74歳", "75歳以上", "介護者なし"])
r = add_question_block(ws, r, "2-2", "介護者の心身の負担感",
    ["とても負担", "やや負担", "あまり負担でない", "負担ではない"])
r = add_question_block(ws, r, "2-3", "介護と仕事の両立状況【介護者本人回答】",
    ["仕事を続けている", "仕事を辞めた・転職した", "労働時間を減らした",
     "もともと仕事をしていない"])

# 認知症
ms(ws, f"A{r}:F{r}", "■認知症対応（基本法対応）", f_section, F(LPURPLE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "3-1", "認知症の診断を受けていますか",
    ["受けている", "受けていないが疑いあり", "受けていない", "わからない"])
r = add_question_block(ws, r, "3-2", "認知症の人としての社会参加の希望",
    ["デイサービスへの参加", "認知症カフェへの参加", "本人ミーティングへの参加",
     "ボランティア活動", "希望しない"])

# 川崎町追加（重点）
ms(ws, f"A{r}:F{r}", "■川崎町追加設問（重点）", f_section, F(LORANGE), cl)
ws.row_dimensions[r].height = 22
r += 1

r = add_question_block(ws, r, "K-1", "町外施設を利用していますか",
    ["町内施設のみ利用", "町外施設も利用", "町外施設のみ利用", "施設は利用していない"])
r = add_question_block(ws, r, "K-2", "町外施設の場合、所在地",
    ["柴田町", "大河原町", "村田町", "仙台市", "その他", "町内のみ"])
r = add_question_block(ws, r, "K-3", "町外医療機関の利用状況",
    ["みやぎ県南中核病院（大河原）", "刈田綜合病院（白石）",
     "仙台市内", "国保川崎病院のみ", "病院にかかっていない"])
r = add_question_block(ws, r, "K-4", "現在の生活で最も不安なこと【1つ選択】",
    ["健康状態の悪化", "介護者の負担", "経済的なこと", "将来の住まい",
     "認知症の進行", "災害時の避難", "その他"])

setup_page(ws, "portrait")

# ===========================================================
# 03_クロス集計
# ===========================================================
ws = wb.create_sheet("03_クロス集計")
ms(ws, "A1:H1", "03　重要クロス集計", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:H2", "■計画策定で重要となる5つのクロス集計", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

widths = [16, 12, 12, 12, 12, 12, 12, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# クロス1: 地区 × 移動手段
r = 4
ms(ws, f"A{r}:H{r}", "クロス1：地区 × 主な交通手段（Q1-3）  → 移動支援の地域偏在を把握", f_q, F(LORANGE), cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="地区").font = f_head
ws.cell(row=r, column=1).fill = F(BLUE); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
cross1_cols = ["自家用車", "家族の車", "町民バス", "デマンドバス", "社協・NPO移送", "タクシー", "徒歩・自転車"]
for i, c_name in enumerate(cross1_cols, 2):
    c = ws.cell(row=r, column=i, value=c_name)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

cross1_areas = ["裏丁・上下・本荒町・中新町", "前川・青根", "今宿", "川内", "本砂金", "小野", "小沢・支倉・碁石・支倉台"]
for area in cross1_areas:
    c = ws.cell(row=r, column=1, value=area)
    c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
    for i in range(2, 9):
        c = ws.cell(row=r, column=i)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

# 反映先
ms(ws, f"A{r}:H{r}", "→ 計画素案v1.3 第5章 5-2「高齢者が安心して暮らせるまちづくり」の重点施策2「移動支援の再構築」に反映", f_note, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24
r += 2

# クロス2: 年齢 × 外出困難理由
ms(ws, f"A{r}:H{r}", "クロス2：年齢階級 × 外出に困っている理由（Q2-1）  → 年齢別の外出支援ニーズ", f_q, F(LORANGE), cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="年齢階級").font = f_head
ws.cell(row=r, column=1).fill = F(BLUE); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
cross2_cols = ["特に困らない", "公共交通少", "運転不安", "目的地ない", "体力的に難", "その他"]
for i, c_name in enumerate(cross2_cols, 2):
    c = ws.cell(row=r, column=i, value=c_name)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

cross2_ages = ["65〜69歳", "70〜74歳", "75〜79歳", "80〜84歳", "85〜89歳", "90歳以上"]
for age in cross2_ages:
    c = ws.cell(row=r, column=1, value=age)
    c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
    for i in range(2, 8):
        c = ws.cell(row=r, column=i)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

ms(ws, f"A{r}:H{r}", "→ 計画素案v1.3 第5章 5-2 と 第5章 5-1「健康づくり・介護予防」（通いの場参加への移動支援）に反映", f_note, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24
r += 2

# クロス3: 世帯類型 × 介護不安
ms(ws, f"A{r}:H{r}", "クロス3：世帯類型 × 介護への不安（Q3-1認知症含む）  → 8050・老老介護対応", f_q, F(LORANGE), cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="世帯類型").font = f_head
ws.cell(row=r, column=1).fill = F(BLUE); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
cross3_cols = ["自身がなる不安", "家族がなる不安", "ご近所対応", "情報少不安", "特に不安なし", "回答者数"]
for i, c_name in enumerate(cross3_cols, 2):
    c = ws.cell(row=r, column=i, value=c_name)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

cross3_types = ["独居", "夫婦のみ", "子と同居（未婚）", "子の家族と同居", "その他"]
for tp in cross3_types:
    c = ws.cell(row=r, column=1, value=tp)
    c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
    for i in range(2, 8):
        c = ws.cell(row=r, column=i)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

ms(ws, f"A{r}:H{r}", "→ 計画素案v1.3 第5章 5-3「在宅生活継続の支援」の重点施策3「地域包括ケアの深化と家族介護者支援」、第6章「認知症施策推進計画」に反映", f_note, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24
r += 2

# クロス4: 要介護度 × サービス満足度
ms(ws, f"A{r}:H{r}", "クロス4：要介護度 × サービス満足度（認定者調査Q1-2）  → サービス改善優先度", f_q, F(LORANGE), cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="要介護度").font = f_head
ws.cell(row=r, column=1).fill = F(BLUE); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
cross4_cols = ["とても満足", "まあ満足", "あまり満足でない", "不満", "わからない", "回答者数"]
for i, c_name in enumerate(cross4_cols, 2):
    c = ws.cell(row=r, column=i, value=c_name)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

cross4_deg = ["要支援1・2", "要介護1・2", "要介護3", "要介護4・5"]
for deg in cross4_deg:
    c = ws.cell(row=r, column=1, value=deg)
    c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
    for i in range(2, 8):
        c = ws.cell(row=r, column=i)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

ms(ws, f"A{r}:H{r}", "→ 計画素案v1.3 第5章 5-4「介護サービスの質の確保と人材確保」、第7章 7-1 サービス見込量に反映", f_note, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24
r += 2

# クロス5: 地区 × 認知症相談窓口認知度
ms(ws, f"A{r}:H{r}", "クロス5：地区 × 認知症相談窓口の認知度（Q3-2）  → 地区別の認知症啓発ニーズ", f_q, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="地区").font = f_head
ws.cell(row=r, column=1).fill = F(BLUE); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
cross5_cols = ["包括センター", "認知症カフェ", "国保川崎病院", "窓口知らない", "相談経験あり", "回答者数"]
for i, c_name in enumerate(cross5_cols, 2):
    c = ws.cell(row=r, column=i, value=c_name)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

for area in cross1_areas:
    c = ws.cell(row=r, column=1, value=area)
    c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
    for i in range(2, 8):
        c = ws.cell(row=r, column=i)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

ms(ws, f"A{r}:H{r}", "→ 計画素案v1.3 第6章 6-3 重点施策（J-4「早期発見・早期対応の体制強化」）の地区別展開に反映", f_note, F(LPURPLE), cl)
ws.row_dimensions[r].height = 24

setup_page(ws)

# ===========================================================
# 04_自由記述_5分類
# ===========================================================
ws = wb.create_sheet("04_自由記述_5分類")
ms(ws, "A1:G1", "04　自由記述の5分類集計", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:G2", "■自由記述欄を5分類で集計し、代表的意見を抽出。計画素案への反映先を明示。",
   f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

widths = [8, 14, 12, 10, 30, 22, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ヘッダ
r = 4
heads = ["No", "分類", "サブ分類", "件数", "代表的意見（要約）", "計画反映先", "優先度"]
for i, h in enumerate(heads, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 26
r += 1

# 5分類の枠を準備
categories = [
    ("①移動", [
        ("バス・デマンド改善", "計画v1.3 5-2", "A"),
        ("通院支援強化", "計画v1.3 5-3", "A"),
        ("除雪・冬季交通", "計画v1.3 5-2", "B"),
        ("免許返納後の対応", "計画v1.3 5-2", "A"),
    ]),
    ("②介護", [
        ("施設不足・順番待ち", "計画v1.3 5-4", "A"),
        ("家族介護の負担軽減", "計画v1.3 5-3", "A"),
        ("ショートステイ拡充", "計画v1.3 5-3", "B"),
        ("ケアマネへの不安・要望", "計画v1.3 5-4", "B"),
    ]),
    ("③見守り・地域", [
        ("独居高齢者見守り", "計画v1.3 5-2", "A"),
        ("近所付き合い希薄化", "計画v1.3 5-2", "B"),
        ("緊急通報・連絡体制", "計画v1.3 5-2", "B"),
        ("民生委員・サポーター活動", "計画v1.3 5-2", "B"),
    ]),
    ("④サービス不足", [
        ("サービス情報がない・分かりにくい", "計画v1.3 全般", "A"),
        ("町外施設の情報不足", "計画v1.3 5-4", "B"),
        ("配食・買物支援の充実", "計画v1.3 5-3", "B"),
        ("認知症対応サービス不足", "計画v1.3 第6章", "A"),
    ]),
    ("⑤医療連携", [
        ("町内医療機関の体制", "計画v1.3 5-3", "A"),
        ("町外通院の負担", "計画v1.3 5-3", "A"),
        ("在宅医療・看取り", "計画v1.3 5-3", "B"),
        ("広域救急の不安", "計画v1.3 5-3", "B"),
    ]),
]

no = 1
for cat, items in categories:
    # カテゴリ見出し行
    ms(ws, f"A{r}:G{r}", f"【{cat}】", f_section, F(LBLUE), cl)
    ws.row_dimensions[r].height = 22
    r += 1
    
    for sub, ref, pri in items:
        # No
        c = ws.cell(row=r, column=1, value=no)
        c.font = f_body; c.alignment = cc; c.border = border
        # 分類
        c = ws.cell(row=r, column=2, value=cat)
        c.font = f_body; c.alignment = cl; c.fill = F(LGRAY); c.border = border
        # サブ分類
        c = ws.cell(row=r, column=3, value=sub)
        c.font = Font(name="游ゴシック", size=10, bold=True)
        c.alignment = cl; c.fill = F(LGRAY); c.border = border
        # 件数（INPUT）
        c = ws.cell(row=r, column=4)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cr; c.border = border
        # 代表的意見（INPUT・長文）
        c = ws.cell(row=r, column=5)
        c.fill = F(INPUT); c.font = f_input; c.alignment = cl; c.border = border
        # 計画反映先
        c = ws.cell(row=r, column=6, value=ref)
        c.font = f_body; c.alignment = cl; c.fill = F(LPURPLE); c.border = border
        # 優先度
        pri_fill = F("FFCCCC") if pri == "A" else F("CCDDFF")
        pri_font = Font(name="游ゴシック", size=10, bold=True, color="C00000" if pri == "A" else "2F5597")
        c = ws.cell(row=r, column=7, value=pri)
        c.font = pri_font; c.alignment = cc; c.fill = pri_fill; c.border = border
        ws.row_dimensions[r].height = 32
        r += 1
        no += 1

# その他カテゴリ用の空行
ms(ws, f"A{r}:G{r}", "【⑥その他（上記5分類に該当しないもの）】", f_section, F(LGRAY), cl)
ws.row_dimensions[r].height = 22
r += 1
for _ in range(3):
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = F(INPUT) if col in [3,4,5] else F(LGRAY)
        c.font = f_input; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 28
    r += 1

# 注記
r += 1
notes = [
    "件数欄には、各サブ分類に該当する自由記述の数を入力してください。",
    "「代表的意見」欄には、原文の要約（30〜50字程度）を1〜3件記入してください。",
    "優先度A（赤）は計画素案Ver.2.0で本文反映が必要、B（青）は施策の論拠補強として活用します。",
    "5分類に該当しない意見は「⑥その他」欄に集約し、必要に応じて新規分類を追加してください。",
]
for note in notes:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

setup_page(ws)

# ===========================================================
# 05_計画反映ガイド
# ===========================================================
ws = wb.create_sheet("05_計画反映ガイド")
ms(ws, "A1:G1", "05　アンケート結果の計画反映ガイド", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:G2", "■各設問・クロス集計の結果を計画素案v1.3のどの章・節に反映するかのマトリクス",
   f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

widths = [12, 30, 14, 22, 14, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ヘッダ
r = 4
heads = ["設問No", "設問内容", "対象シート", "反映先（計画素案v1.3）", "プレースホルダー", "Ver.2.0反映時期", "優先度"]
for i, h in enumerate(heads, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 36
r += 1

# 反映マトリクス
reflect = [
    # 一般高齢者
    ("一般 Q1-1", "現在の健康状態", "01", "第2章 2-2 高齢者の状況", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q1-3", "外出時の主な交通手段", "01・03", "第5章 5-2 移動支援の再構築", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q1-4", "通いの場参加状況", "01", "第5章 5-1 KPI「通いの場参加率」", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q2-1", "外出に困っていること", "01・03", "第5章 5-2 KPI「外出困難割合」", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q2-2", "町の移動支援を知っているか", "01", "第5章 5-2 KPI「移動支援3制度の周知度」", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q3-1", "認知症への不安", "01・03", "第6章 6-3 KPI「認知症本人と家族の地域生活満足度」", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q3-2", "認知症の相談窓口認知", "01・03", "第6章 6-3 KPI「相談窓口を知っている人の割合」", "【調査後設定】", "Ver.2.0", "A"),
    ("一般 Q3-3", "サポーター養成受講経験", "01", "第6章 6-3 重点施策J-1の根拠", "─", "Ver.2.0", "B"),
    ("一般 Q4-1", "将来の介護希望", "01", "第5章 5-3 在宅生活継続支援の根拠", "─", "Ver.2.0", "A"),
    ("一般 K-1", "町外医療機関の利用", "01", "第5章 5-3 重点施策2 広域医療連携の根拠", "─", "Ver.2.0", "A"),
    ("一般 K-2", "町独自支援の認知", "01", "第5章 5-2 KPI 独自施策の根拠", "─", "Ver.2.0", "B"),
    # 認定者
    ("認定 Q0-3", "要介護度分布", "02", "第2章 2-3 受給者数の傍証", "─（実績データで既知）", "Ver.2.0", "B"),
    ("認定 Q1-1", "利用サービスの種類", "02", "第7章 7-1 サービス見込量算定の補正", "【見込量算定後反映】", "Ver.2.0", "A"),
    ("認定 Q1-2", "サービス満足度", "02・03", "第5章 5-4 介護サービスの質の確保", "【調査後設定】", "Ver.2.0", "A"),
    ("認定 Q1-3", "サービス利用の困りごと", "02", "第5章 5-4 第7章 7-1 補正要素", "【調査後設定】", "Ver.2.0", "A"),
    ("認定 Q2-2", "介護者の負担感", "02", "第5章 5-3 重点施策3「家族介護者支援」", "【調査後設定】", "Ver.2.0", "A"),
    ("認定 Q2-3", "介護離職状況", "02", "第5章 5-3 介護離職防止支援の根拠", "─", "Ver.2.0", "A"),
    ("認定 Q3-1", "認知症の診断状況", "02", "第6章 6-1 川崎町の認知症対応方針", "【調査後設定】", "Ver.2.0", "A"),
    ("認定 Q3-2", "認知症の人の社会参加希望", "02", "第6章 6-3 重点施策J-3「認知症カフェ・本人M」", "─", "Ver.2.0", "A"),
    ("認定 K-1", "町外施設の利用状況", "02", "第7章 7-1 施設サービス・住所地特例", "【見込量算定後反映】", "Ver.2.0", "A"),
    ("認定 K-2", "町外施設の所在地", "02", "第5章 5-4 重点施策2「町外施設・住所地特例の整理」", "─", "Ver.2.0", "B"),
    ("認定 K-4", "現在の生活で最も不安なこと", "02・04", "第3章 3-4 第10期に向けた課題の補強", "─", "Ver.2.0", "A"),
    # クロス・自由記述
    ("クロス1", "地区 × 移動手段", "03", "第5章 5-2 移動支援の地区別ニーズ", "─", "Ver.2.0", "A"),
    ("クロス2", "年齢 × 外出困難", "03", "第5章 5-1 介護予防（通いの場）の対象拡大", "─", "Ver.2.0", "A"),
    ("クロス3", "世帯類型 × 介護不安", "03", "第5章 5-3 第6章 8050問題・老老介護対応", "─", "Ver.2.0", "A"),
    ("クロス4", "要介護度 × サービス満足度", "03", "第5章 5-4 第7章 7-1 サービス改善", "─", "Ver.2.0", "A"),
    ("クロス5", "地区 × 認知症窓口認知", "03", "第6章 6-3 地区別認知症啓発", "─", "Ver.2.0", "A"),
    ("自由①", "移動関連", "04", "第5章 5-2 重点施策の論拠補強", "─", "Ver.2.0", "A"),
    ("自由②", "介護関連", "04", "第5章 5-3・5-4 補強", "─", "Ver.2.0", "A"),
    ("自由③", "見守り・地域", "04", "第5章 5-2 補強", "─", "Ver.2.0", "A"),
    ("自由④", "サービス不足", "04", "第3章 3-4 第10期課題の補強", "─", "Ver.2.0", "A"),
    ("自由⑤", "医療連携", "04", "第5章 5-3 重点施策2「広域医療連携」", "─", "Ver.2.0", "A"),
]

for q_no, q_text, sh, ref, ph, ver, pri in reflect:
    # 設問No
    c = ws.cell(row=r, column=1, value=q_no)
    c.font = Font(name="游ゴシック", size=9, bold=True); c.alignment = cc
    c.fill = F(LORANGE if q_no.startswith("一般") else (LBLUE if q_no.startswith("認定") else LPURPLE)); c.border = border
    # 設問内容
    c = ws.cell(row=r, column=2, value=q_text)
    c.font = f_body; c.alignment = cl; c.border = border
    # 対象シート
    c = ws.cell(row=r, column=3, value=sh)
    c.font = f_body; c.alignment = cc; c.fill = F(LGRAY); c.border = border
    # 反映先
    c = ws.cell(row=r, column=4, value=ref)
    c.font = Font(name="游ゴシック", size=9, bold=True, color=NAVY)
    c.alignment = cl; c.fill = F(LPURPLE); c.border = border
    # プレースホルダー
    c = ws.cell(row=r, column=5, value=ph)
    if "【" in ph: c.font = Font(name="游ゴシック", size=9, color=ORANGE, italic=True)
    else: c.font = f_note
    c.alignment = cc; c.border = border
    # Ver.2.0反映時期
    c = ws.cell(row=r, column=6, value=ver)
    c.font = f_body; c.alignment = cc; c.border = border
    # 優先度
    pri_fill = F("FFCCCC") if pri == "A" else F("CCDDFF")
    pri_font = Font(name="游ゴシック", size=10, bold=True, color="C00000" if pri == "A" else "2F5597")
    c = ws.cell(row=r, column=7, value=pri)
    c.font = pri_font; c.alignment = cc; c.fill = pri_fill; c.border = border
    ws.row_dimensions[r].height = 28
    r += 1

# 注記
r += 1
notes = [
    "この反映ガイドは、計画素案v1.3の【調査後設定】等のプレースホルダーを埋めるためのナビゲーションです。",
    "優先度A（赤）：計画素案Ver.2.0で本文反映が必須／優先度B（青）：施策の論拠補強として活用",
    "「対象シート」は本テンプレートの集計シート番号、「反映先」は計画素案v1.3の章・節を示します。",
    "本ガイドの結果をもとに、計画素案v1.3を Ver.2.0 に更新し、第2回策定委員会（令和8年11月）で素案として審議します。",
]
for note in notes:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 26
    r += 1

setup_page(ws)

# 保存
out = "/home/claude/kawasaki_work/川崎町_アンケート集計分析テンプレート.xlsx"
wb.save(out)
print(f"作成完了: {out}")
print(f"シート数: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws_s = wb[s]
    print(f"  - {s}: {ws_s.max_row}行 x {ws_s.max_column}列")
