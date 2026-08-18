"""
川崎町第10期計画 委員意見反映管理シート

第1回策定委員会で出された委員意見を記録し、
計画素案Ver.2.0のどこにどう反映するかを管理するためのExcel

8シート構成：
00_使い方
01_意見ダッシュボード（カテゴリ別・反映状況サマリー）
02_協議事項1意見集（基本理念サブタイトル）
03_協議事項2意見集（基本目標6・第6章独立化）
04_協議事項3意見集（移動支援3層構造）
05_協議事項4意見集（認知症5本柱・3層KPI）
06_協議事項5意見集（保険料試算3パターン）
07_その他意見・付帯事項
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# カラーパレット
NAVY = "1F3864"; BLUE = "2F5597"; LBLUE = "DAE3F3"
ORANGE = "ED7D31"; LORANGE = "FCE4D6"
GREEN = "548235"; LGREEN = "E2EFDA"
RED = "C00000"; LRED = "FFE4E4"
GRAY = "808080"; LGRAY = "F2F2F2"
PURPLE = "9333B0"; LPURPLE = "EAD5F0"
INPUT_Y = "FFFFCC"; WHITE = "FFFFFF"

# ステータスカラー
ACCEPT = "C6EFCE"      # 反映採用 緑
PARTIAL = "FFEB9C"     # 部分反映 黄
RESERVE = "FFE4B0"     # 保留検討 橙
REJECT = "FFC7CE"      # 反映なし 赤
PENDING = "F2F2F2"     # 未判断 灰

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def F(c): return PatternFill("solid", fgColor=c)
cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
cl = Alignment(horizontal="left", vertical="center", wrap_text=True)
cr = Alignment(horizontal="right", vertical="center", wrap_text=True)

f_title = Font(name="游ゴシック", size=13, bold=True, color=WHITE)
f_sub   = Font(name="游ゴシック", size=10, italic=True, color=WHITE)
f_head  = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
f_section = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
f_body  = Font(name="游ゴシック", size=10)
f_note  = Font(name="游ゴシック", size=9, italic=True, color="595959")
f_input = Font(name="游ゴシック", size=10, color="0000FF")

wb = Workbook()
wb.remove(wb.active)

def ms(ws, rng, val, font, fill, align):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = val; c.font = font
    if fill is not None: c.fill = fill
    c.alignment = align
    from openpyxl.utils.cell import range_boundaries
    a, b, d, e = range_boundaries(rng)
    for r in range(b, e+1):
        for col in range(a, d+1):
            ws.cell(row=r, column=col).border = border

def setup_page(ws, orient="landscape"):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orient == "landscape" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

# ===========================================================
# 00_使い方
# ===========================================================
ws = wb.create_sheet("00_使い方")
ms(ws, "A1:G1", "川崎町第10期計画 委員意見反映管理シート", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 28
ms(ws, "A2:G2", "── 第1回策定委員会(R8.8中旬)以降の意見集約と素案Ver.2.0への反映管理 ──",
   f_sub, F(BLUE), cc)
ws.row_dimensions[2].height = 20

# 1. 目的
ms(ws, "A4:G4", "1．本シートの目的", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22
ms(ws, "A5:G5",
   "第1回策定委員会で委員から出された意見を5協議事項別に記録し、計画素案Ver.2.0のどこにどう反映するかを管理します。委員会議事録の作成と並行して使用し、第2回策定委員会(R8.11)までに全意見の対応方針を確定させます。",
   f_body, None, cl)
ws.row_dimensions[5].height = 50

# 2. シート構成
ms(ws, "A7:G7", "2．シート構成", f_head, F(BLUE), cl)
ws.row_dimensions[7].height = 22

ws.cell(row=8, column=1, value="シート").font = f_head
ws.cell(row=8, column=1).fill = F(NAVY); ws.cell(row=8, column=1).alignment = cc; ws.cell(row=8, column=1).border = border
ms(ws, "B8:D8", "対象協議事項", f_head, F(NAVY), cc)
ms(ws, "E8:G8", "用途", f_head, F(NAVY), cc)
ws.row_dimensions[8].height = 22

sheets = [
    ("01_意見ダッシュボード", "全協議事項のサマリー", "全意見の進捗・カテゴリ別状況の把握"),
    ("02_協議事項1意見集", "基本理念サブタイトル付加", "サブタイトル付加可否・別表現案の集約"),
    ("03_協議事項2意見集", "基本目標6新設・第6章独立化", "独立章化の是非・章構成への意見"),
    ("04_協議事項3意見集", "移動支援3層構造の整理", "3制度の統合・周知強化への意見"),
    ("05_協議事項4意見集", "認知症5本柱・3層KPI", "チームオレンジ・本人ミーティング等への意見"),
    ("06_協議事項5意見集", "保険料試算3パターン", "保険料水準・基金活用・13段階区分への意見"),
    ("07_その他意見・付帯事項", "5協議事項以外の意見", "全章共通・運営方法への意見集約"),
]
r = 9
for sh, target, use in sheets:
    c = ws.cell(row=r, column=1, value=sh)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:D{r}", target, f_body, None, cl)
    ms(ws, f"E{r}:G{r}", use, f_body, None, cl)
    ws.row_dimensions[r].height = 26
    r += 1

# 3. 反映状況凡例
r += 1
ms(ws, f"A{r}:G{r}", "3．反映状況凡例", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

statuses = [
    (ACCEPT, "反映採用", "委員意見を素案Ver.2.0に反映する方針が確定"),
    (PARTIAL, "部分反映", "意見の一部を反映する方針が確定"),
    (RESERVE, "保留検討", "第2回委員会で再協議またはVer.2.1以降で対応検討"),
    (REJECT, "反映なし", "意見を聴くが計画には反映しない（理由を明記）"),
    (PENDING, "未判断", "事務局・町担当課・弊社で対応方針を検討中"),
]
for color, status, mean in statuses:
    c = ws.cell(row=r, column=1, value=status)
    c.fill = F(color)
    c.font = Font(name="游ゴシック", size=10, bold=True)
    c.alignment = cc; c.border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 4. 運用フロー
r += 1
ms(ws, f"A{r}:G{r}", "4．運用フロー（第1回委員会後の作業）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

flow = [
    ("Step1", "委員会当日", "委員発言を録音・要点筆記。協議事項別に発言を分類"),
    ("Step2", "委員会翌日〜1週間", "本シートに全意見を転記。発言者属性も記録"),
    ("Step3", "委員会後2週間", "事務局・町担当・弊社で対応方針(採用/部分/保留/不採用)を協議"),
    ("Step4", "委員会後3-4週間", "対応方針を本シートに記入。計画素案Ver.2.0の反映先(章節)を確定"),
    ("Step5", "委員会後1ヶ月", "計画素案Ver.2.0更新作業開始。本シートを参照して反映"),
    ("Step6", "第2回委員会前", "Ver.2.0反映状況を委員に事前提示。継続協議事項を整理"),
]
for step, when, work in flow:
    c = ws.cell(row=r, column=1, value=step)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", when, f_body, None, cl)
    ms(ws, f"D{r}:G{r}", work, f_body, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

# 5. 注意事項
r += 1
ms(ws, f"A{r}:G{r}", "5．注意事項", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

notes = [
    "委員氏名は個人情報のため、本シートでは委員番号(C-01〜C-XX)で管理し、別途委員名簿で対応関係を管理してください。",
    "発言要旨は委員確認後の議事録の文言を採用し、本人の意図と異なる解釈にならないよう注意します。",
    "反映方針は事務局・町担当・弊社の3者協議で決定し、必要に応じて第2回委員会で再協議します。",
    "「保留検討」項目は次回委員会の継続協議事項として明示し、対応の透明性を確保します。",
]
for note in notes:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 26
    r += 1

# 列幅
widths = [22, 14, 12, 16, 12, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 01_意見ダッシュボード
# ===========================================================
ws = wb.create_sheet("01_意見ダッシュボード")
ms(ws, "A1:H1", "01　全体ダッシュボード", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:H2", "── 5協議事項の意見集約状況・反映方針サマリー ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# KPI（カードレイアウト）
ms(ws, "A4:H4", "■ 集約状況サマリー", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22

kpis = [
    ("総意見数", "─", "件"),
    ("反映採用", "─", "件"),
    ("部分反映", "─", "件"),
    ("保留検討", "─", "件"),
]
for i, (k, v, u) in enumerate(kpis):
    col = 1 + i * 2
    c_k = ws.cell(row=5, column=col, value=k)
    c_k.font = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
    c_k.alignment = cc; c_k.fill = F(NAVY); c_k.border = border
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    
    c_v = ws.cell(row=6, column=col, value=v)
    c_v.font = Font(name="游ゴシック", size=24, bold=True, color=NAVY)
    c_v.alignment = cc; c_v.border = border
    
    c_u = ws.cell(row=6, column=col+1, value=u)
    c_u.font = f_body; c_u.alignment = cl; c_u.border = border

ws.row_dimensions[5].height = 24
ws.row_dimensions[6].height = 36

# 協議事項別サマリー
r = 8
ms(ws, f"A{r}:H{r}", "■ 協議事項別 意見集約サマリー", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

headers = ["No", "協議事項", "意見数", "採用", "部分", "保留", "不採用", "Ver.2.0反映先"]
widths = [6, 30, 10, 8, 8, 8, 8, 22]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(NAVY); c.alignment = cc; c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[r].height = 24
r += 1

summary = [
    ("1", "基本理念サブタイトル付加（〜認知症になっても誰もが…〜）", 
     "─", "─", "─", "─", "─", "第4章 4-1基本理念"),
    ("2", "基本目標6（認知症施策）の新設と第6章独立章化", 
     "─", "─", "─", "─", "─", "第4章 4-2 / 第6章全体"),
    ("3", "移動支援3層構造の整理と住民周知強化", 
     "─", "─", "─", "─", "─", "第5章 5-2"),
    ("4", "認知症施策5本柱(J-1〜J-5)・KPI3層構造", 
     "─", "─", "─", "─", "─", "第6章 6-3"),
    ("5", "介護保険料試算3パターン(A/B/C)・13段階区分", 
     "─", "─", "─", "─", "─", "第7章 7-3"),
    ("付帯", "その他意見・全章共通事項", 
     "─", "─", "─", "─", "─", "各章 適宜"),
]
for no, name, total, acc, par, res, rej, ref in summary:
    c = ws.cell(row=r, column=1, value=no)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc; c.border = border
    c.fill = F(LBLUE)
    c = ws.cell(row=r, column=2, value=name)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl; c.border = border
    c = ws.cell(row=r, column=3, value=total)
    c.font = f_input; c.alignment = cc; c.fill = F(INPUT_Y); c.border = border
    c = ws.cell(row=r, column=4, value=acc)
    c.font = f_input; c.alignment = cc; c.fill = F(ACCEPT); c.border = border
    c = ws.cell(row=r, column=5, value=par)
    c.font = f_input; c.alignment = cc; c.fill = F(PARTIAL); c.border = border
    c = ws.cell(row=r, column=6, value=res)
    c.font = f_input; c.alignment = cc; c.fill = F(RESERVE); c.border = border
    c = ws.cell(row=r, column=7, value=rej)
    c.font = f_input; c.alignment = cc; c.fill = F(REJECT); c.border = border
    c = ws.cell(row=r, column=8, value=ref)
    c.font = Font(name="游ゴシック", size=10, color=NAVY, bold=True); c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 32
    r += 1

# Ver.2.0更新タイムライン
r += 1
ms(ws, f"A{r}:H{r}", "■ Ver.2.0更新タイムライン", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

headers2 = ["時期", "マイルストーン", "対応事項", "担当", "成果物"]
widths2 = [12, 26, 36, 10, 16]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(NAVY); c.alignment = cc; c.border = border
    if i in [5]: ws.column_dimensions[get_column_letter(i+3)].width = w  # E列以降は使わず
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
ws.cell(row=r, column=4, value="対応事項").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc
ws.cell(row=r, column=7, value="担当").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.cell(row=r, column=8, value="成果物").font = f_head
ws.cell(row=r, column=8).fill = F(NAVY); ws.cell(row=r, column=8).alignment = cc; ws.cell(row=r, column=8).border = border
ws.row_dimensions[r].height = 24
r += 1

timeline = [
    ("R8.8中旬", "第1回策定委員会", "委員意見の収集・本シートへの転記開始", "弊社/町", "委員会議事録"),
    ("R8.8末", "意見集約完了", "全意見の本シートへの転記完了", "弊社/町", "本シート Ver.1"),
    ("R8.9上旬", "対応方針協議", "事務局・町担当・弊社で対応方針協議(週1回x3回)", "3者", "対応方針確定"),
    ("R8.9中旬", "アンケート集計完了", "R8.7末回収分の集計完了・結果反映準備", "弊社", "集計結果"),
    ("R8.9下旬〜10", "Ver.2.0更新作業", "委員意見＋アンケート結果を計画素案に反映", "弊社", "計画素案Ver.2.0素案"),
    ("R8.11上旬", "Ver.2.0完成", "Red Team検証・町担当課確認", "弊社/町", "計画素案Ver.2.0"),
    ("R8.11", "第2回策定委員会", "Ver.2.0審議・サービス見込量方向性確定", "委員会", "委員会議事録"),
]
for t, ms_name, work, owner, output in timeline:
    c = ws.cell(row=r, column=1, value=t)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", ms_name, Font(name="游ゴシック", size=10, bold=True, color=NAVY), None, cl)
    ms(ws, f"D{r}:F{r}", work, f_body, None, cl)
    c = ws.cell(row=r, column=7, value=owner)
    c.font = f_body; c.alignment = cc
    if "弊社" in owner: c.fill = F(LORANGE)
    elif "町" in owner: c.fill = F(LBLUE)
    else: c.fill = F(LGREEN)
    c.border = border
    c = ws.cell(row=r, column=8, value=output)
    c.font = f_body; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 26
    r += 1

setup_page(ws)

# ===========================================================
# 共通：協議事項別意見集シート生成ヘルパー
# ===========================================================
def make_opinion_sheet(name, title, agenda_desc, theme_color=None):
    """協議事項別意見集シート
    各意見について:
    - 意見No, 委員番号, 委員属性, 発言要旨, 計画への影響, 対応方針, 反映先, 状況, 備考
    """
    theme_color = theme_color or NAVY
    ws = wb.create_sheet(name)
    
    ms(ws, "A1:I1", title, f_title, F(theme_color), cc)
    ws.row_dimensions[1].height = 26
    ms(ws, "A2:I2", agenda_desc, f_sub, F(BLUE), cl)
    ws.row_dimensions[2].height = 20
    
    # 協議事項の説明（リマインダー）
    ms(ws, "A4:I4", "■ 協議事項の概要", f_head, F(BLUE), cl)
    ws.row_dimensions[4].height = 22
    
    return ws

def write_opinion_table(ws, start_row, agenda_summary, sample_opinions=None):
    """意見集約表のヘッダと記入枠を生成"""
    # 概要説明
    ms(ws, f"A5:I5", agenda_summary, f_body, F(LGRAY), cl)
    ws.row_dimensions[5].height = 60
    
    # 意見テーブルのヘッダ
    r = 7
    ms(ws, f"A{r}:I{r}", "■ 委員意見集約と対応方針", f_head, F(BLUE), cl)
    ws.row_dimensions[r].height = 22
    r += 1
    
    headers = ["No", "委員番号", "委員属性", "発言要旨", "計画への影響", "対応方針", "反映先(章節)", "状況", "備考"]
    widths = [6, 10, 16, 30, 18, 18, 14, 12, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = f_head; c.fill = F(NAVY); c.alignment = cc; c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 30
    r += 1
    
    # 記入枠（10行分）
    for i in range(1, 11):
        c = ws.cell(row=r, column=1, value=i)
        c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc
        c.fill = F(LBLUE); c.border = border
        for col in range(2, 10):
            cell = ws.cell(row=r, column=col)
            cell.fill = F(INPUT_Y); cell.font = f_input
            cell.alignment = cl; cell.border = border
        ws.row_dimensions[r].height = 36
        r += 1
    
    # 注釈
    r += 1
    notes = [
        "意見Noは委員発言順に1から連番で記入してください。",
        "状況欄は「反映採用」「部分反映」「保留検討」「反映なし」「未判断」のいずれかを記入してください。",
        "反映先は計画素案v1.5の章節番号(例：第5章 5-2)を記入してください。",
        "対応方針が決まり次第、本シートを更新し、Ver.2.0更新作業の参照資料とします。",
    ]
    for note in notes:
        ms(ws, f"A{r}:I{r}", "※ " + note, f_note, F("FFF2CC"), cl)
        ws.row_dimensions[r].height = 22
        r += 1
    
    return ws

# ===========================================================
# 02_協議事項1：基本理念サブタイトル
# ===========================================================
ws = make_opinion_sheet(
    "02_協議事項1意見集",
    "02　協議事項1：基本理念サブタイトル付加について",
    "■基本理念「住民が住み慣れた地域で安心して暮らせるまちづくり」に認知症基本法対応サブタイトル付加の是非"
)
write_opinion_table(ws, 7,
    "第9期から継承する基本理念に「〜認知症になっても誰もが自分らしく暮らせる地域共生社会の実現〜」を新たに付加することの是非を協議。委員のご意見によっては保留・別表現への変更も可能。"
)
setup_page(ws)

# ===========================================================
# 03_協議事項2：基本目標6・第6章独立化
# ===========================================================
ws = make_opinion_sheet(
    "03_協議事項2意見集",
    "03　協議事項2：基本目標6新設と第6章独立化について",
    "■認知症基本法対応として基本目標6を新設し、第6章として独立章化する方向性の是非",
    PURPLE
)
write_opinion_table(ws, 7,
    "認知症基本法第14条対応として、第9期の5基本目標に「基本目標6：認知症施策の総合的推進」を追加し、第6章として独立章化(基本法第15条〜21条の7基本的施策に対応した町施策体系を構築)。本日協議。"
)
setup_page(ws)

# ===========================================================
# 04_協議事項3：移動支援3層構造
# ===========================================================
ws = make_opinion_sheet(
    "04_協議事項3意見集",
    "04　協議事項3：移動支援3層構造の整理について",
    "■令和7年3月タクシー助成終了後の3制度(社協NPO移送/デマンドバス/町民バス)の整理・住民周知強化"
)
write_opinion_table(ws, 7,
    "アンケートで認知度の低さが確認された場合、第10期計画では3制度の整理・住民周知の強化を重点施策(J-B)として明示する方向性の是非。所管調整(地域振興課・町民生活課・社協・NPO)の進め方も含めて協議。"
)
setup_page(ws)

# ===========================================================
# 05_協議事項4：認知症5本柱・3層KPI
# ===========================================================
ws = make_opinion_sheet(
    "05_協議事項4意見集",
    "05　協議事項4：認知症施策の重点5本柱とKPI3層構造について",
    "■J-1サポーター/J-2チームオレンジ/J-3本人ミーティング/J-4早期発見/J-5医療連携の方向性",
    PURPLE
)
write_opinion_table(ws, 7,
    "認知症施策5本柱の方向性、特にJ-2チームオレンジとJ-3本人ミーティングの新規取組の進め方、KPI3層構造(プロセス・アウトプット・アウトカム)の設定方針について協議。"
)
setup_page(ws)

# ===========================================================
# 06_協議事項5：保険料試算3パターン
# ===========================================================
ws = make_opinion_sheet(
    "06_協議事項5意見集",
    "06　協議事項5：介護保険料試算3パターン(A/B/C)について",
    "■3パターン(A:取崩なし/B:50%取崩/C:全額取崩)の検討方向性・13段階区分への見直し"
)
write_opinion_table(ws, 7,
    "第10期保険料は基金取崩額により3パターン試算。確定値は介護給付費準備基金残高(R8.6時点・現在確認中)の確定後、第3回策定委員会(R9.1中旬)で協議。本日は方向性へのご意見を頂戴。所得段階9→13段階見直しの検討方針も合わせて協議。"
)
setup_page(ws)

# ===========================================================
# 07_その他意見・付帯事項
# ===========================================================
ws = wb.create_sheet("07_その他意見・付帯事項")
ms(ws, "A1:I1", "07　その他意見・付帯事項", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:I2", "── 5協議事項以外の意見・全章共通事項・運営方法等 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# 想定される付帯意見カテゴリ
ms(ws, "A4:I4", "■ 想定される付帯意見カテゴリ", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22

cats = [
    ("カテゴリA", "計画全体の構成・章立てへの意見", "第1-3章・第8章"),
    ("カテゴリB", "アンケート結果の解釈・追加分析希望", "第2-3章・別添"),
    ("カテゴリC", "個別施策(認知症以外)への意見", "第5章 5-1〜5-5"),
    ("カテゴリD", "見込量・サービス量への意見", "第7章 7-1〜7-2"),
    ("カテゴリE", "推進体制・評価への意見", "第8章"),
    ("カテゴリF", "委員会運営・スケジュール調整", "運営"),
    ("カテゴリG", "次回委員会への持ち越し議題", "次回協議"),
]
r = 5
ws.cell(row=r, column=1, value="カテゴリ").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:G{r}", "想定意見", f_head, F(NAVY), cl)
ms(ws, f"H{r}:I{r}", "反映先候補", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1
for cat, theme, ref in cats:
    c = ws.cell(row=r, column=1, value=cat)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:G{r}", theme, f_body, None, cl)
    ms(ws, f"H{r}:I{r}", ref, f_body, None, cc)
    ws.row_dimensions[r].height = 24
    r += 1

# 記入枠
r += 1
ms(ws, f"A{r}:I{r}", "■ 委員意見記入欄", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

headers = ["No", "委員番号", "委員属性", "カテゴリ", "発言要旨", "対応方針", "反映先(章節)", "状況", "備考"]
widths = [6, 10, 16, 12, 28, 16, 12, 12, 14]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(NAVY); c.alignment = cc; c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[r].height = 30
r += 1

# 記入枠（15行）
for i in range(1, 16):
    c = ws.cell(row=r, column=1, value=i)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    for col in range(2, 10):
        cell = ws.cell(row=r, column=col)
        cell.fill = F(INPUT_Y); cell.font = f_input
        cell.alignment = cl; cell.border = border
    ws.row_dimensions[r].height = 32
    r += 1

# 注釈
r += 1
notes = [
    "カテゴリ欄には「A〜G」のいずれかを記入してください（不明な場合は空欄）。",
    "「次回委員会への持ち越し議題」(カテゴリG)は、第2回策定委員会の議事として整理します。",
    "委員会の進行方法・配布資料についての意見は、カテゴリFとして整理します。",
]
for note in notes:
    ms(ws, f"A{r}:I{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

setup_page(ws)

# 保存
out = "/home/claude/kawasaki_work/川崎町_委員意見反映管理シート.xlsx"
wb.save(out)
print(f"作成完了: {out}")
print(f"シート数: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws_s = wb[s]
    print(f"  - {s}: {ws_s.max_row}行 x {ws_s.max_column}列")
