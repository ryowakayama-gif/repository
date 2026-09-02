# -*- coding: utf-8 -*-
"""
業務工程管理表の更新（交付金評価データの照合結果を踏まえた更新）

・03_確認事項一覧に No.57〜59 を追加
・06_資料提供依頼一覧 No.18（交付金の全国集計結果）に再送分の照合結果を追記
"""
import copy

import openpyxl

WB = "川崎町_業務工程管理表.xlsx"


def unmerge_below(ws, at_row):
    kept = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= at_row:
            kept.append((m.min_row, m.max_row, m.min_col, m.max_col))
            ws.unmerge_cells(str(m))
    return kept


def remerge(ws, kept, n):
    for r1, r2, c1, c2 in kept:
        ws.merge_cells(start_row=r1 + n, end_row=r2 + n,
                       start_column=c1, end_column=c2)


def copy_style(ws, src_row, dst_row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(dst_row, c)._style = copy.copy(ws.cell(src_row, c)._style)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


KAKUNIN = [
    ("57", "(4)", "R8.9",
     "第1回策定委員会資料の交付金評価 目標別内訳の訂正",
     "資料4の「令和8年度評価の目標別内訳」は7行の表としていますが、"
     "支援交付金の目標Ⅳ（アウトカム）40点が脱落しています。"
     "アウトカム指標は推進交付金と支援交付金の双方に同じ40点が計上されるため正しくは8行で、"
     "7行の合計488点は表に記載した合計528点と一致していません。"
     "第2回策定委員会資料で8行に訂正します（推進計252点＋支援計276点＝528点）。",
     "議事録\n第2回委員会資料\n素案 第3章（第9期の評価）",
     "受託者", "実施中", "R8.9",
     "全国集計結果（令和8年度）の全項目照合により判明。"
     "アウトカムは800点中80点（10％）であり、資料は重みを過小に見せていた。"),

    ("58", "(4)", "R8.9",
     "交付金評価「評価結果の活用」が16点から0点となった理由",
     "保険者機能強化推進交付金の目標Ⅰ（ⅰ）「4 評価結果の活用」のア〜エは、"
     "令和7年度評価では各4点の満点16点でしたが、令和8年度評価では4項目とも0点となっています。"
     "令和7年度から令和8年度への合計得点の減少37点のうち、単独で最大の要因です。"
     "提出内容や様式の変更の有無を含め、理由をご確認ください。"
     "第10期計画のKPI管理表及びサービス別の計画値・実績値対比表により回復可能な項目です。",
     "第9期の評価（仕様書6（4））\n第2回委員会資料\n素案 第9章（KPI管理表）",
     "発注者", "確認待ち", "R8.10",
     "目標Ⅰは75点→57点（▲18点）。ほかに活動指標群で▲6点。"),

    ("59", "(4)(6)", "R8.9",
     "交付金評価「給付費適正化事業の取組状況」が6点から0点となった理由",
     "保険者機能強化推進交付金の目標Ⅱ（ⅰ）「給付費適正化事業の取組状況 ア」は、"
     "令和7年度評価の6点から令和8年度評価で0点となっています。"
     "目標Ⅱは88点から82点に低下しました。理由をご確認ください。",
     "第9期の評価（仕様書6（4））\n給付実績分析（仕様書6（6））\n第2回委員会資料",
     "発注者", "確認待ち", "R8.10",
     "医療情報との突合の実施状況を含む指標群。"),
]


def main():
    wb = openpyxl.load_workbook(WB)
    ws = wb["03_確認事項一覧"]

    at = 61                      # No.56（r60）の直後
    n = len(KAKUNIN)
    kept = unmerge_below(ws, at)
    ws.insert_rows(at, n)
    remerge(ws, kept, n)
    for i, rec in enumerate(KAKUNIN):
        r = at + i
        copy_style(ws, 60, r, 10)
        for c, v in enumerate(rec, start=1):
            ws.cell(r, c).value = v
    last = 60 + n
    for r in range(last + 2, last + 8):
        cell = ws.cell(r, 2)
        if isinstance(cell.value, str) and cell.value.startswith("=COUNTIF"):
            cell.value = cell.value.replace("H5:H60", f"H5:H{last}")

    # 06_資料提供依頼一覧 No.18 に再送分の照合結果を追記
    ws6 = wb["06_資料提供依頼一覧"]
    cur = ws6.cell(22, 10).value or ""
    ws6.cell(22, 10).value = (
        cur + "／令和8年9月2日に同じ資料（令和6・7・8年度の評価結果xlsx 3件、"
        "令和8年度評価指標pdf 2件）を再度ご提供いただいた。川崎町の全項目を照合し、"
        "8月28日受領分と内容が一致することを確認済み（pdf 2件は相互に同一）。追加の格納は不要。")

    wb.save(WB)
    print(f"03_確認事項一覧：No.57〜59 を追加（{n}件）／06_資料提供依頼一覧 No.18 を更新")


if __name__ == "__main__":
    main()
