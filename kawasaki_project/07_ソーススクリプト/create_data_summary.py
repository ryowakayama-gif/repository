"""
川崎町第10期計画策定 実績データ確認サマリー
保険者データ202506（R7.6月時点）と年報データ2021（R3年度実績）から
計画策定に必要な数値を抽出・整理。

構成（6シート）：
00_サマリー：全体俯瞰
01_被保険者数：人口・年齢構成
02_認定者・受給者：要介護度別・サービス区分別
03_給付費・財政：年間給付費・収納実績・特別会計
04_所得段階別：保険料区分別人口分布
05_利活用ガイド：計画策定でのデータ活用方針
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

NAVY="1F3864"; BLUE="2F5597"; LIGHT_BLUE="DAE3F3"; ORANGE="ED7D31"
LIGHT_ORANGE="FCE4D6"; LIGHT_GREEN="E2EFDA"; GRAY_LIGHT="F2F2F2"; YELLOW="FFFF99"

thin=Side(style="thin",color="BFBFBF")
border=Border(top=thin,bottom=thin,left=thin,right=thin)

f_title=Font(name="游ゴシック",size=14,bold=True,color="FFFFFF")
f_head=Font(name="游ゴシック",size=11,bold=True,color="FFFFFF")
f_sec=Font(name="游ゴシック",size=11,bold=True,color="1F3864")
f_body=Font(name="游ゴシック",size=10)
f_num=Font(name="游ゴシック",size=10,bold=True,color="C00000")
f_note=Font(name="游ゴシック",size=9,italic=True,color="595959")

def F(color): return PatternFill("solid",fgColor=color)
cc=Alignment(horizontal="center",vertical="center",wrap_text=True)
cl=Alignment(horizontal="left",vertical="center",wrap_text=True)
cr=Alignment(horizontal="right",vertical="center")

def ms(ws,rng,val,font,fill,align):
    ws.merge_cells(rng); c=ws[rng.split(":")[0]]; c.value=val
    c.font=font; c.alignment=align
    if fill is not None: c.fill=fill
    from openpyxl.utils.cell import range_boundaries
    a,b,d,e=range_boundaries(rng)
    for r in range(b,e+1):
        for col in range(a,d+1): ws.cell(row=r,column=col).border=border

def setup_page(ws):
    ws.page_setup.orientation=ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
    ws.page_margins=PageMargins(left=0.4,right=0.4,top=0.5,bottom=0.5,header=0.3,footer=0.3)

wb=Workbook(); wb.remove(wb.active)

# ===================================================
# 00_サマリー
# ===================================================
ws=wb.create_sheet("00_サマリー")
ms(ws,"A1:F1","川崎町第10期介護保険事業計画策定　実績データ確認サマリー",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=28
ms(ws,"A2:F2","保険者データ202506（R7.6月時点）＋年報データ2021（R3年度実績）からの抽出整理",Font(name="游ゴシック",size=9,italic=True,color="FFFFFF"),F(BLUE),cl)
ws.row_dimensions[2].height=20

ms(ws,"A4:F4","1．確認した資料",f_head,F(BLUE),cl); ws.row_dimensions[4].height=22
sources=[
 ("保険者データ202506","_変更後様式_保険者テ_ータ_202506川崎町.xlsx","R7.6月時点（単月）","32シート","令和7年度の最新被保険者・受給者数"),
 ("年報データ2021","年報テ_ータ_2021_川崎町_修正_.xlsx","R3年度（通年）","47シート","R3年度の年間給付費・収納・財政"),
]
ws.cell(row=5,column=1,value="名称").font=f_head; ws.cell(row=5,column=1).fill=F(NAVY); ws.cell(row=5,column=1).alignment=cc; ws.cell(row=5,column=1).border=border
ms(ws,"B5:C5","ファイル名",f_head,F(NAVY),cc)
ws.cell(row=5,column=4,value="対象期間").font=f_head; ws.cell(row=5,column=4).fill=F(NAVY); ws.cell(row=5,column=4).alignment=cc; ws.cell(row=5,column=4).border=border
ws.cell(row=5,column=5,value="シート数").font=f_head; ws.cell(row=5,column=5).fill=F(NAVY); ws.cell(row=5,column=5).alignment=cc; ws.cell(row=5,column=5).border=border
ws.cell(row=5,column=6,value="主な内容").font=f_head; ws.cell(row=5,column=6).fill=F(NAVY); ws.cell(row=5,column=6).alignment=cc; ws.cell(row=5,column=6).border=border
ws.row_dimensions[5].height=22
r=6
for nm,fl,kk,sh,nai in sources:
    ws.cell(row=r,column=1,value=nm).font=Font(name="游ゴシック",size=10,bold=True)
    ws.cell(row=r,column=1).alignment=cl; ws.cell(row=r,column=1).fill=F(LIGHT_BLUE); ws.cell(row=r,column=1).border=border
    ms(ws,f"B{r}:C{r}",fl,f_body,None,cl)
    ws.cell(row=r,column=4,value=kk).font=f_body; ws.cell(row=r,column=4).alignment=cc; ws.cell(row=r,column=4).border=border
    ws.cell(row=r,column=5,value=sh).font=f_body; ws.cell(row=r,column=5).alignment=cc; ws.cell(row=r,column=5).border=border
    ws.cell(row=r,column=6,value=nai).font=f_body; ws.cell(row=r,column=6).alignment=cl; ws.cell(row=r,column=6).border=border
    ws.row_dimensions[r].height=28; r+=1

r+=1
ms(ws,f"A{r}:F{r}","2．重要数値ハイライト（計画策定で核となる8つの数字）",f_head,F(BLUE),cl); ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="No").font=f_head; ws.cell(row=r,column=1).fill=F(NAVY); ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).border=border
ws.cell(row=r,column=2,value="項目").font=f_head; ws.cell(row=r,column=2).fill=F(NAVY); ws.cell(row=r,column=2).alignment=cc; ws.cell(row=r,column=2).border=border
ws.cell(row=r,column=3,value="数値").font=f_head; ws.cell(row=r,column=3).fill=F(NAVY); ws.cell(row=r,column=3).alignment=cc; ws.cell(row=r,column=3).border=border
ws.cell(row=r,column=4,value="時点").font=f_head; ws.cell(row=r,column=4).fill=F(NAVY); ws.cell(row=r,column=4).alignment=cc; ws.cell(row=r,column=4).border=border
ms(ws,f"E{r}:F{r}","計画策定での意味",f_head,F(NAVY),cc); ws.row_dimensions[r].height=22; r+=1

highlights=[
 (1,"第1号被保険者数","3,244人","R7.6月","保険料基準額算定の分母"),
 (2,"うち後期高齢者(75歳以上)","1,675人 (51.6%)","R7.6月","認定率上昇要因・調整交付金算定"),
 (3,"住所地特例被保険者","24人","R7.6月","町外施設利用の実数（KO確認の裏付け）"),
 (4,"サービス受給者総数","466人(居宅276+地密56+施設134)","R7.6月","認定者の利用実態"),
 (5,"R3年度 年間給付費総額","960,499,764円(約9.6億円)","R3年度","保険料算定の基礎・第10期試算の出発点"),
 (6,"R3年度 保険料収納率","96.3%","R3年度","予定収納率の根拠（過去5年平均で確定）"),
 (7,"第8期 保険料基準額(月額)","6,380円","R3年度","第9期6,500円・第10期との比較基準"),
 (8,"R3年度 保険給付支払総額","1,042,179,056円(約10.4億円)","R3年度","給付費＋高額・特定入所者を含む総額"),
]
for h in highlights:
    no,kou,suu,jik,imi=h
    ws.cell(row=r,column=1,value=no).font=Font(name="游ゴシック",size=10,bold=True)
    ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).fill=F(LIGHT_ORANGE); ws.cell(row=r,column=1).border=border
    ws.cell(row=r,column=2,value=kou).font=Font(name="游ゴシック",size=10,bold=True)
    ws.cell(row=r,column=2).alignment=cl; ws.cell(row=r,column=2).border=border
    ws.cell(row=r,column=3,value=suu).font=f_num
    ws.cell(row=r,column=3).alignment=cr; ws.cell(row=r,column=3).fill=F(LIGHT_GREEN); ws.cell(row=r,column=3).border=border
    ws.cell(row=r,column=4,value=jik).font=f_body
    ws.cell(row=r,column=4).alignment=cc; ws.cell(row=r,column=4).border=border
    ms(ws,f"E{r}:F{r}",imi,f_body,None,cl)
    ws.row_dimensions[r].height=26; r+=1

for col,w in zip("ABCDEF",[6,32,22,12,18,28]): ws.column_dimensions[col].width=w
setup_page(ws)

# ===================================================
# 01_被保険者数
# ===================================================
ws=wb.create_sheet("01_被保険者数")
ms(ws,"A1:F1","01　第1号被保険者数（人口・年齢構成）",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=26

ms(ws,"A3:F3","1-1．年齢階級別被保険者数（R3年度末 vs R7.6月）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[3].height=22
heads=["年齢階級","R3年度末\n(R4.3月末)","R7.6月時点","増減","増減率","備考"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=4,column=c,value=h)
    cell.font=f_head; cell.fill=F(NAVY); cell.alignment=cc; cell.border=border
ws.row_dimensions[4].height=32

age_data=[
 ("65-75歳未満",1745,1569,-176,-10.1,"前期高齢者：減少傾向（人口減）"),
 ("75-85歳未満",905,1070,165,18.2,"後期高齢者前段：大幅増（団塊世代の移行）"),
 ("85歳以上",605,605,0,0.0,"後期高齢者後段：横ばい"),
 ("計",3255,3244,-11,-0.3,"全体はほぼ横ばい"),
 ("(再掲)後期高齢者75歳以上",1510,1675,165,10.9,"★認定率上昇・施設需要増加の要因"),
 ("(再掲)住所地特例",34,24,-10,-29.4,"町外施設入所者（KO確認事項の裏付け）"),
 ("(再掲)外国人被保険者",5,9,4,80.0,"少数だが増加"),
]
r=5
for ag,v1,v2,sa,rate,bik in age_data:
    is_total = ag=="計"
    is_recap = ag.startswith("(再掲)")
    ws.cell(row=r,column=1,value=ag).font=Font(name="游ゴシック",size=10,bold=is_total)
    ws.cell(row=r,column=1).alignment=cl
    if is_total: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    elif is_recap: ws.cell(row=r,column=1).fill=F(GRAY_LIGHT)
    ws.cell(row=r,column=1).border=border
    for c,v in enumerate([v1,v2,sa],2):
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(name="游ゴシック",size=10,bold=is_total)
        cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
        if is_total: cell.fill=F(LIGHT_BLUE)
    cell=ws.cell(row=r,column=5,value=rate/100); cell.font=Font(name="游ゴシック",size=10,bold=is_total)
    cell.alignment=cr; cell.border=border; cell.number_format="0.0%"
    if is_total: cell.fill=F(LIGHT_BLUE)
    ws.cell(row=r,column=6,value=bik).font=Font(name="游ゴシック",size=9,italic=True,color="595959")
    ws.cell(row=r,column=6).alignment=cl; ws.cell(row=r,column=6).border=border
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:F{r}","1-2．人口動態（増減内訳）R7.6月単月",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="区分").font=f_head; ws.cell(row=r,column=1).fill=F(NAVY); ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).border=border
ms(ws,f"B{r}:F{r}","内訳",f_head,F(NAVY),cc); ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="増（11人）").font=Font(name="游ゴシック",size=10,bold=True)
ws.cell(row=r,column=1).alignment=cl; ws.cell(row=r,column=1).fill=F(LIGHT_GREEN); ws.cell(row=r,column=1).border=border
ms(ws,f"B{r}:F{r}","転入2人／65歳到達8人／その他1人　→ 主因は65歳到達（72.7%）",f_body,None,cl)
ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="減（15人）").font=Font(name="游ゴシック",size=10,bold=True)
ws.cell(row=r,column=1).alignment=cl; ws.cell(row=r,column=1).fill=F(LIGHT_ORANGE); ws.cell(row=r,column=1).border=border
ms(ws,f"B{r}:F{r}","転出3人／死亡12人　→ 主因は死亡（80.0%）",f_body,None,cl)
ws.row_dimensions[r].height=22; r+=1
r+=1
ms(ws,f"A{r}:F{r}","【分析メモ】R7.6月単月で65歳到達<死亡となっており、第1号被保険者は微減局面に入っている。ただし75歳以上の後期高齢者は引き続き増加し、認定率上昇・施設需要は当面継続する見込み。",f_note,F("FFF2CC"),cl)
ws.row_dimensions[r].height=36

for col,w in zip("ABCDEF",[22,16,16,12,12,34]): ws.column_dimensions[col].width=w
setup_page(ws)

# ===================================================
# 02_認定者・受給者
# ===================================================
ws=wb.create_sheet("02_認定者_受給者")
ms(ws,"A1:J1","02　サービス受給者数（R7.6月時点・要介護度別）",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=26

ms(ws,"A3:J3","2-1．居宅・地域密着型・施設サービス受給者数（様式1の6）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[3].height=22

heads=["サービス区分","要支援1","要支援2","要支援計","要介護1","要介護2","要介護3","要介護4","要介護5","受給者計"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=4,column=c,value=h)
    cell.font=f_head; cell.fill=F(NAVY); cell.alignment=cc; cell.border=border
ws.row_dimensions[4].height=26

usage=[
 ("居宅(介護予防)サービス",13,50,63,39,77,66,23,8,276),
 ("地域密着型サービス",0,0,0,9,11,17,12,7,56),
 ("施設サービス計",0,0,0,6,15,40,43,30,134),
 ("　うち特養",0,0,0,1,2,23,21,21,68),
 ("　うち老健",0,0,0,5,13,17,22,9,66),
 ("　うち療養型・医療院",0,0,0,0,0,0,0,0,0),
 ("延べ合計（重複あり）",13,50,63,54,103,123,78,45,466),
]
r=5
for row in usage:
    is_total = row[0]=="延べ合計（重複あり）"
    is_sub = row[0].startswith("　")
    ws.cell(row=r,column=1,value=row[0]).font=Font(name="游ゴシック",size=10,bold=(is_total or not is_sub))
    ws.cell(row=r,column=1).alignment=cl
    if is_total: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    elif is_sub: ws.cell(row=r,column=1).fill=F(GRAY_LIGHT)
    ws.cell(row=r,column=1).border=border
    for c,v in enumerate(row[1:],2):
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(name="游ゴシック",size=10,bold=is_total)
        cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
        if is_total: cell.fill=F(LIGHT_BLUE)
        elif is_sub: cell.fill=F(GRAY_LIGHT)
        if c==10 and v>0: cell.fill=F(LIGHT_GREEN if not (is_total or is_sub) else cell.fill.fgColor.rgb)
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:J{r}","2-2．サービス利用構成比（受給者ベース）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="区分").font=f_head; ws.cell(row=r,column=1).fill=F(NAVY); ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).border=border
ws.cell(row=r,column=2,value="人数").font=f_head; ws.cell(row=r,column=2).fill=F(NAVY); ws.cell(row=r,column=2).alignment=cc; ws.cell(row=r,column=2).border=border
ws.cell(row=r,column=3,value="構成比").font=f_head; ws.cell(row=r,column=3).fill=F(NAVY); ws.cell(row=r,column=3).alignment=cc; ws.cell(row=r,column=3).border=border
ms(ws,f"D{r}:J{r}","計画策定での意味",f_head,F(NAVY),cc); ws.row_dimensions[r].height=22; r+=1

comp=[
 ("居宅サービス",276,0.592,"在宅介護の主軸。アンケート認定者票で深掘り"),
 ("地域密着型",56,0.120,"小規模・夜間対応等。GHが主要(町内施設の中心)"),
 ("施設サービス",134,0.288,"特養68・老健66。介護療養型・医療院は0"),
 ("延べ合計",466,1.000,"※複数サービス併用者は重複カウント"),
]
for kbn,n,r_rate,imi in comp:
    is_t = kbn=="延べ合計"
    ws.cell(row=r,column=1,value=kbn).font=Font(name="游ゴシック",size=10,bold=is_t)
    ws.cell(row=r,column=1).alignment=cl
    if is_t: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    ws.cell(row=r,column=1).border=border
    cell=ws.cell(row=r,column=2,value=n); cell.font=Font(name="游ゴシック",size=10,bold=is_t)
    cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
    if is_t: cell.fill=F(LIGHT_BLUE)
    cell=ws.cell(row=r,column=3,value=r_rate); cell.font=Font(name="游ゴシック",size=10,bold=is_t)
    cell.alignment=cr; cell.border=border; cell.number_format="0.0%"
    if is_t: cell.fill=F(LIGHT_BLUE)
    ms(ws,f"D{r}:J{r}",imi,f_body,F(LIGHT_BLUE) if is_t else None,cl)
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:J{r}","【特記】R7.6月単月では介護療養型医療施設・介護医療院の受給者は0人。第10期の見込量算定では、これらは0据置でよい（廃止・転換の影響なし）。",f_note,F("FFF2CC"),cl); ws.row_dimensions[r].height=30

for col,w in zip("ABCDEFGHIJ",[24,9,9,9,9,9,9,9,9,11]): ws.column_dimensions[col].width=w
setup_page(ws)

# ===================================================
# 03_給付費・財政
# ===================================================
ws=wb.create_sheet("03_給付費_財政")
ms(ws,"A1:F1","03　R3年度 給付費・財政（保険料算定の基礎）",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=26

ms(ws,"A3:F3","3-1．サービス種類別 年間給付費（R3年度・円）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[3].height=22
heads=["サービス区分","サービス種類","要支援計","要介護計","合計","構成比"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=4,column=c,value=h)
    cell.font=f_head; cell.fill=F(NAVY); cell.alignment=cc; cell.border=border
ws.row_dimensions[4].height=26

kyufu=[
 ("居宅","居宅サービス計",26821755,339234895,366056650,0.381),
 ("居宅","　訪問サービス",2490773,54605942,57096715,0.059),
 ("居宅","　通所サービス",16765707,173710493,190476200,0.198),
 ("居宅","　短期入所サービス",393812,41858752,42252564,0.044),
 ("居宅","　福祉用具・住宅改修",3850663,24723944,28574607,0.030),
 ("居宅","　介護予防支援・居宅介護支援",3320800,44335764,47656564,0.050),
 ("地域密着","地域密着型計",0,152764596,152764596,0.159),
 ("地域密着","　認知症対応型共同生活介護",0,82494459,82494459,0.086),
 ("地域密着","　地域密着型介護老人福祉施設",0,70270137,70270137,0.073),
 ("施設","施設サービス計",0,441678518,441678518,0.460),
 ("施設","　介護老人福祉施設(特養)",0,193053976,193053976,0.201),
 ("施設","　介護老人保健施設(老健)",0,248624542,248624542,0.259),
 ("施設","　介護療養型・介護医療院",0,0,0,0.000),
 ("総計","総計",26821755,933678009,960499764,1.000),
]
r=5
for kbn,nm,si,ka,go,kos in kyufu:
    is_total = nm=="総計"
    is_sum = nm.endswith("計") and not is_total
    is_sub = nm.startswith("　")
    ws.cell(row=r,column=1,value=kbn).font=Font(name="游ゴシック",size=10,bold=(is_total or is_sum))
    ws.cell(row=r,column=1).alignment=cl
    if is_total: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    elif is_sum: ws.cell(row=r,column=1).fill=F(LIGHT_GREEN)
    elif is_sub: ws.cell(row=r,column=1).fill=F(GRAY_LIGHT)
    ws.cell(row=r,column=1).border=border
    ws.cell(row=r,column=2,value=nm).font=Font(name="游ゴシック",size=10,bold=(is_total or is_sum))
    ws.cell(row=r,column=2).alignment=cl
    if is_total: ws.cell(row=r,column=2).fill=F(LIGHT_BLUE)
    elif is_sum: ws.cell(row=r,column=2).fill=F(LIGHT_GREEN)
    elif is_sub: ws.cell(row=r,column=2).fill=F(GRAY_LIGHT)
    ws.cell(row=r,column=2).border=border
    for c,v in [(3,si),(4,ka),(5,go)]:
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(name="游ゴシック",size=10,bold=(is_total or is_sum))
        cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
        if is_total: cell.fill=F(LIGHT_BLUE)
        elif is_sum: cell.fill=F(LIGHT_GREEN)
        elif is_sub: cell.fill=F(GRAY_LIGHT)
    cell=ws.cell(row=r,column=6,value=kos); cell.font=Font(name="游ゴシック",size=10,bold=(is_total or is_sum))
    cell.alignment=cr; cell.border=border; cell.number_format="0.0%"
    if is_total: cell.fill=F(LIGHT_BLUE)
    elif is_sum: cell.fill=F(LIGHT_GREEN)
    elif is_sub: cell.fill=F(GRAY_LIGHT)
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:F{r}","3-2．R3年度 保険料収納実績",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[r].height=22; r+=1
ws.cell(row=r,column=1,value="区分").font=f_head; ws.cell(row=r,column=1).fill=F(NAVY); ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).border=border
ws.cell(row=r,column=2,value="調定額(円)").font=f_head; ws.cell(row=r,column=2).fill=F(NAVY); ws.cell(row=r,column=2).alignment=cc; ws.cell(row=r,column=2).border=border
ws.cell(row=r,column=3,value="収納額(円)").font=f_head; ws.cell(row=r,column=3).fill=F(NAVY); ws.cell(row=r,column=3).alignment=cc; ws.cell(row=r,column=3).border=border
ws.cell(row=r,column=4,value="収納率").font=f_head; ws.cell(row=r,column=4).fill=F(NAVY); ws.cell(row=r,column=4).alignment=cc; ws.cell(row=r,column=4).border=border
ms(ws,f"E{r}:F{r}","備考",f_head,F(NAVY),cc); ws.row_dimensions[r].height=22; r+=1

shunou=[
 ("現年度分 特別徴収",211826640,211826640,1.000,"年金天引・100%収納"),
 ("現年度分 普通徴収",18949580,16538210,0.873,"自主納付・約13%が未納"),
 ("現年度分 計",230776220,228364850,0.990,"99.0%（高水準）"),
 ("滞納繰越分",7060233,771720,0.109,"回収困難（10.9%）"),
 ("合計",237836453,229136570,0.963,"★予定収納率の根拠"),
]
for row in shunou:
    is_t = row[0]=="合計"
    is_sum = row[0]=="現年度分 計"
    ws.cell(row=r,column=1,value=row[0]).font=Font(name="游ゴシック",size=10,bold=(is_t or is_sum))
    ws.cell(row=r,column=1).alignment=cl
    if is_t: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    elif is_sum: ws.cell(row=r,column=1).fill=F(LIGHT_GREEN)
    ws.cell(row=r,column=1).border=border
    for c,v in [(2,row[1]),(3,row[2])]:
        cell=ws.cell(row=r,column=c,value=v); cell.font=Font(name="游ゴシック",size=10,bold=(is_t or is_sum))
        cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
        if is_t: cell.fill=F(LIGHT_BLUE)
        elif is_sum: cell.fill=F(LIGHT_GREEN)
    cell=ws.cell(row=r,column=4,value=row[3]); cell.font=Font(name="游ゴシック",size=10,bold=(is_t or is_sum))
    cell.alignment=cr; cell.border=border; cell.number_format="0.0%"
    if is_t: cell.fill=F(LIGHT_BLUE)
    elif is_sum: cell.fill=F(LIGHT_GREEN)
    ms(ws,f"E{r}:F{r}",row[4],f_body,F(LIGHT_BLUE) if is_t else (F(LIGHT_GREEN) if is_sum else None),cl)
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:F{r}","3-3．R3年度 介護保険特別会計の経理（抜粋）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[r].height=22; r+=1
keiri=[
 ("【歳入】","",""),
 ("　保険料(介護保険料)",229292110,"第1号被保険者保険料"),
 ("　国庫支出金 介護給付費負担金",200296679,"国20%"),
 ("　国庫支出金 調整交付金",59117000,"後期高齢者比率による調整・約6.1%相当"),
 ("　支払基金交付金 介護給付費",284142000,"第2号被保険者27%"),
 ("　都道府県負担金",157438000,"県12.5%"),
 ("　一般会計繰入金(12.5%)",130357012,"町12.5%"),
 ("　保険者機能強化推進交付金",2616000,""),
 ("　保険者努力支援交付金",2615000,""),
 ("【歳出】","",""),
 ("　保険給付費 計",1042179056,"★最重要：R3年度総支払額"),
 ("　　うち介護サービス等諸費",933658461,""),
 ("　　うち介護予防サービス等諸費",26841303,""),
 ("　　うち高額介護サービス等費",26815719,""),
 ("　　うち高額医療合算",2395731,""),
 ("　　うち特定入所者介護サービス等費",52467842,"補足給付"),
 ("　地域支援事業 計",32071763,"5,650+1,950+24,471千円"),
 ("　　うち介護予防・生活支援",5650417,""),
 ("　　うち一般介護予防",1950101,""),
 ("　　うち包括的支援・任意",24471245,""),
]
for row in keiri:
    nm=row[0]; val=row[1]; bik=row[2]
    is_h = nm.startswith("【")
    is_imp = "★" in bik
    ms(ws,f"A{r}:B{r}",nm,Font(name="游ゴシック",size=10,bold=is_h),F(LIGHT_BLUE) if is_h else None,cl)
    if val=="":
        ws.cell(row=r,column=3).fill=F(LIGHT_BLUE) if is_h else PatternFill("solid",fgColor="FFFFFF")
        ws.cell(row=r,column=3).border=border
        ws.cell(row=r,column=4).fill=F(LIGHT_BLUE) if is_h else PatternFill("solid",fgColor="FFFFFF")
        ws.cell(row=r,column=4).border=border
    else:
        ms(ws,f"C{r}:D{r}",val,Font(name="游ゴシック",size=10,bold=is_imp,color="C00000" if is_imp else "000000"),F(LIGHT_GREEN) if is_imp else None,cr)
        ws[f"C{r}"].number_format="#,##0"
    ms(ws,f"E{r}:F{r}",bik,f_body,F(LIGHT_BLUE) if is_h else None,cl)
    ws.row_dimensions[r].height=22; r+=1

for col,w in zip("ABCDEF",[14,32,14,12,14,16]): ws.column_dimensions[col].width=w
setup_page(ws)

# ===================================================
# 04_所得段階別
# ===================================================
ws=wb.create_sheet("04_所得段階別")
ms(ws,"A1:G1","04　R3年度末 所得段階別第1号被保険者数（保険料区分）",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=26

ms(ws,"A3:G3","4-1．所得段階別の人口分布（保険料基準額算定の根拠）",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[3].height=22
heads=["段階","区分(R3時点)","保険料率","月額(円)","R3年度末人数","構成比","年額(円)"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=4,column=c,value=h)
    cell.font=f_head; cell.fill=F(NAVY); cell.alignment=cc; cell.border=border
ws.row_dimensions[4].height=26

# 第8期月額基準額6,380円
base=6380
dankai=[
 ("第1段階","世帯非課税(年金80万円以下)",0.50,3190,431,431/3255),
 ("第2段階","世帯非課税(年金80-120万円)",0.75,4785,273,273/3255),
 ("第3段階","世帯非課税(年金120万円超)",0.75,4785,275,275/3255),
 ("第4段階","本人非課税・他課税(年金80万円以下)",0.90,5742,466,466/3255),
 ("第5段階","本人非課税・他課税(年金80万円超)",1.00,6380,697,697/3255),
 ("第6段階","本人課税(合計所得120万円未満)",1.20,7656,449,449/3255),
 ("第7段階","本人課税(120-210万円未満)",1.30,8294,385,385/3255),
 ("第8段階","本人課税(210-320万円未満)",1.50,9570,173,173/3255),
 ("第9段階","本人課税(320万円以上)",1.70,10846,106,106/3255),
 ("計","─",None,None,3255,1.000),
]
r=5
for d in dankai:
    is_t = d[0]=="計"
    for c,v in enumerate(d,1):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="游ゴシック",size=10,bold=is_t)
        cell.border=border
        if c==1 or c==2:
            cell.alignment=cl
            if is_t: cell.fill=F(LIGHT_BLUE)
        elif c==3:
            cell.alignment=cr
            if v is not None: cell.number_format="0.00"
            if is_t: cell.fill=F(LIGHT_BLUE)
        elif c==4 or c==5:
            cell.alignment=cr
            if v is not None: cell.number_format="#,##0"
            if is_t: cell.fill=F(LIGHT_BLUE)
        elif c==6:
            cell.alignment=cr; cell.number_format="0.0%"
            if is_t: cell.fill=F(LIGHT_BLUE)
    # 年額（月額×12）
    if not is_t and d[3]:
        nen=d[3]*12
        cell=ws.cell(row=r,column=7,value=nen)
        cell.font=Font(name="游ゴシック",size=10); cell.alignment=cr; cell.border=border; cell.number_format="#,##0"
    elif is_t:
        ws.cell(row=r,column=7,value="─").font=Font(name="游ゴシック",size=10,bold=True)
        ws.cell(row=r,column=7).alignment=cc; ws.cell(row=r,column=7).fill=F(LIGHT_BLUE); ws.cell(row=r,column=7).border=border
    ws.row_dimensions[r].height=24; r+=1

r+=1
ms(ws,f"A{r}:G{r}","4-2．課税区分別の構成",f_sec,F(LIGHT_BLUE),cl); ws.row_dimensions[r].height=22; r+=1
group=[
 ("非課税層(1-3段階)",431+273+275,(431+273+275)/3255,"低所得：低い段階の負担軽減対象"),
 ("住民税本人非課税(4-5段階)",466+697,(466+697)/3255,"中間層：第5段階が標準"),
 ("住民税本人課税(6-9段階)",449+385+173+106,(449+385+173+106)/3255,"課税者：相対的負担大"),
 ("計",3255,1.000,""),
]
ws.cell(row=r,column=1,value="区分").font=f_head; ws.cell(row=r,column=1).fill=F(NAVY); ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).border=border
ms(ws,f"B{r}:C{r}","人数",f_head,F(NAVY),cc)
ws.cell(row=r,column=4,value="構成比").font=f_head; ws.cell(row=r,column=4).fill=F(NAVY); ws.cell(row=r,column=4).alignment=cc; ws.cell(row=r,column=4).border=border
ms(ws,f"E{r}:G{r}","計画策定での意味",f_head,F(NAVY),cc); ws.row_dimensions[r].height=22; r+=1
for g in group:
    is_t = g[0]=="計"
    ws.cell(row=r,column=1,value=g[0]).font=Font(name="游ゴシック",size=10,bold=is_t)
    ws.cell(row=r,column=1).alignment=cl
    if is_t: ws.cell(row=r,column=1).fill=F(LIGHT_BLUE)
    ws.cell(row=r,column=1).border=border
    ms(ws,f"B{r}:C{r}",g[1],Font(name="游ゴシック",size=10,bold=is_t),F(LIGHT_BLUE) if is_t else None,cr)
    ws[f"B{r}"].number_format="#,##0"
    cell=ws.cell(row=r,column=4,value=g[2]); cell.font=Font(name="游ゴシック",size=10,bold=is_t)
    cell.alignment=cr; cell.border=border; cell.number_format="0.0%"
    if is_t: cell.fill=F(LIGHT_BLUE)
    ms(ws,f"E{r}:G{r}",g[3],f_body,F(LIGHT_BLUE) if is_t else None,cl)
    ws.row_dimensions[r].height=22; r+=1

r+=1
ms(ws,f"A{r}:G{r}","【分析メモ】第5段階(標準)が697人で最多。非課税層(1-3段階)が979人(30.1%)を占め、低所得対策が重要。本人課税者層は1,113人(34.2%)で、保険料負担余力のある層も一定規模。",f_note,F("FFF2CC"),cl)
ws.row_dimensions[r].height=36

for col,w in zip("ABCDEFG",[12,30,10,10,12,10,12]): ws.column_dimensions[col].width=w
setup_page(ws)

# ===================================================
# 05_利活用ガイド
# ===================================================
ws=wb.create_sheet("05_利活用ガイド")
ms(ws,"A1:E1","05　計画策定での実績データ利活用ガイド",f_title,F(NAVY),cc)
ws.row_dimensions[1].height=26
ms(ws,"A2:E2","各データを第10期計画のどの場面で使うかを整理しました",Font(name="游ゴシック",size=9,italic=True,color="FFFFFF"),F(BLUE),cl)
ws.row_dimensions[2].height=20

heads=["フェーズ","データ項目","活用場面","必要追加データ","出典"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=4,column=c,value=h)
    cell.font=f_head; cell.fill=F(NAVY); cell.alignment=cc; cell.border=border
ws.row_dimensions[4].height=26

guide=[
 ("F5見込量算定","第1号被保険者数(年齢別)","人口推計の出発点(社人研データと突合)","R4-R6の年度末値","保険者データ202506・年報2021"),
 ("F5見込量算定","受給者数(要介護度別)","認定率・サービス利用率算定","R4-R7各年度の認定者数","保険者データ様式1の6"),
 ("F5見込量算定","施設受給者134人(うち町外含む)","施設サービス需要見込・住所地特例反映","R7年度の施設別利用者推移","保険者データ・KO確認"),
 ("F5施策反映","住所地特例24人","町外施設利用の実態(KO論点)","町外施設の所在自治体内訳","保険者データ様式1"),
 ("F5保険料算定","R3年度給付費 9.6億円","給付費伸び率算定の基準値","R4-R7各年度の給付費","年報2021様式2"),
 ("F5保険料算定","保険給付支払総額10.4億円","保険料収納必要額の出発点","R4-R7各年度の決算","年報2021様式4"),
 ("F5保険料算定","収納率96.3%","予定収納率の根拠(過去5年平均で確定)","R4-R7各年度の収納率","年報2021様式3"),
 ("F5保険料算定","所得段階別人口分布","13段階区分の見直し検討","R7年度の最新分布","年報2021様式1所得段階別"),
 ("F5保険料算定","調整交付金 59,117千円(約6.1%)","調整交付金見込・後期高齢者割合の影響","R4-R7各年度の交付金実績","年報2021様式4"),
 ("F5保険料算定","第8期保険料6,380円→第9期6,500円","保険料水準推移(+1.9%)","第10期試算3パターン作成","年報2021・キックオフ"),
 ("F2施策評価","食費居住費負担認定127件","低所得施設利用者の補足給付実態","認定者全体に占める割合","保険者データ様式1の2"),
 ("F4アンケート","受給者数のサービス区分構成","認定者300名抽出時の層別配分の参考","─","保険者データ様式1の6"),
 ("F1前提条件","保険者番号04324","所定様式作成時の必須項目","─","両ファイル共通"),
]
r=5
for ph,dt,ka,tu,sh in guide:
    ws.cell(row=r,column=1,value=ph).font=Font(name="游ゴシック",size=9,bold=True)
    ws.cell(row=r,column=1).alignment=cc; ws.cell(row=r,column=1).fill=F(LIGHT_ORANGE); ws.cell(row=r,column=1).border=border
    ws.cell(row=r,column=2,value=dt).font=f_body
    ws.cell(row=r,column=2).alignment=cl; ws.cell(row=r,column=2).border=border
    ws.cell(row=r,column=3,value=ka).font=f_body
    ws.cell(row=r,column=3).alignment=cl; ws.cell(row=r,column=3).border=border
    ws.cell(row=r,column=4,value=tu).font=Font(name="游ゴシック",size=9,italic=True,color="C00000")
    ws.cell(row=r,column=4).alignment=cl; ws.cell(row=r,column=4).fill=F(LIGHT_GREEN); ws.cell(row=r,column=4).border=border
    ws.cell(row=r,column=5,value=sh).font=Font(name="游ゴシック",size=8,italic=True,color="595959")
    ws.cell(row=r,column=5).alignment=cl; ws.cell(row=r,column=5).border=border
    ws.row_dimensions[r].height=30; r+=1

r+=2
ms(ws,f"A{r}:E{r}","【追加で町からご提供いただきたいデータ】",f_sec,F(LIGHT_ORANGE),cl); ws.row_dimensions[r].height=22; r+=1
adds=[
 "①R4年度・R5年度・R6年度の年報データ（給付費・収納率・財政の経年推移把握）",
 "②R7年度の月次保険者データ（4月・5月・7月以降）または年度末確定値",
 "③介護給付費準備基金の残高推移（R3-R7各年度末）→6月確定値で最終確認",
 "④地域支援事業費の詳細内訳（総合事業・包括的支援事業・任意事業）",
 "⑤所得段階別人口の最新値（R7時点）→13段階区分の見直し検討用",
 "⑥認定者数の経年推移（要介護度別・R4-R7各年度3月末）",
]
for a in adds:
    ms(ws,f"A{r}:E{r}",a,Font(name="游ゴシック",size=10),F("FFF2CC"),cl); ws.row_dimensions[r].height=24; r+=1

for col,w in zip("ABCDE",[16,28,28,28,16]): ws.column_dimensions[col].width=w
setup_page(ws)

out="/home/claude/kawasaki_work/川崎町_実績データ確認サマリー.xlsx"
wb.save(out)
print("作成完了:",out)
print("シート数:",len(wb.sheetnames))
for s in wb.sheetnames: print("  -",s)
