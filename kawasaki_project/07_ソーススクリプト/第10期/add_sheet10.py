# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb=openpyxl.load_workbook('川崎町_保険料試算ワークブック.xlsx')
if '10_長期推計サマリー' in wb.sheetnames:
    del wb['10_長期推計サマリー']
ws=wb.create_sheet('10_長期推計サマリー')
NAVY='1F3864'; BLUE='2E75B6'; LBLUE='DEEBF7'; ORANGE='C55A11'; LORANGE=PatternFill('solid',fgColor='FCE4D6'); YEL=PatternFill('solid',fgColor='FFF2CC'); GRAY='808080'
H=Font(bold=True,color='FFFFFF',size=11); HF=PatternFill('solid',fgColor=NAVY)
SUB=Font(bold=True,color=NAVY,size=11); SUBF=PatternFill('solid',fgColor=LBLUE)
thin=Side(style='thin',color='BFBFBF'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
def w(addr,val,font=None,fill=None,fmt=None,align=None,border=True):
    c=ws[addr]; c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=Alignment(horizontal=align,vertical='center')
    if border:c.border=BORD
ws.column_dimensions['A'].width=26
for col in ['B','C','D','E','F']: ws.column_dimensions[col].width=13

w('A1','10　長期推計サマリー（人口・高齢化率・給付費・保険料）',Font(bold=True,size=13,color=NAVY),border=False)
ws.merge_cells('A1:F1')
w('A2','── 社人研R5推計・第9期計画の確定データに基づく ──',Font(italic=True,color=GRAY,size=9),border=False)
ws.merge_cells('A2:F2')

# A. 人口・高齢化率
r=4; w(f'A{r}','A. 人口・高齢化率の長期推移（社人研R5推計）',SUB,SUBF); ws.merge_cells(f'A{r}:F{r}')
r=5
for col,v in zip(['A','B','C','D','E','F'],['年次','2020','2025','2030','2040','2050']): w(f'{col}{r}',v,H,HF,align='center')
data=[('総人口(人)',[8345,8161,7029,5776,4525]),
      ('65歳以上(人)',[3210,3219,3149,2835,2494]),
      ('15-64歳(人)',[4381,4394,3424,2605,1795]),
      ('高齢化率(%)',[38.6,38.6,44.8,49.1,55.1])]
for i,(lbl,vals) in enumerate(data):
    r=6+i; w(f'A{r}',lbl,Font(bold=True))
    for col,v in zip(['B','C','D','E','F'],vals):
        w(f'{col}{r}',v,fmt='#,##0' if '%' not in lbl else '0.0',align='center')
r=10; w(f'A{r}','◆ 高齢者人口ピークは2025年(令和7年)頃。高齢化率は2050年55.1%まで上昇継続(全国37.1%)。',Font(bold=True,color=ORANGE,size=9),LORANGE); ws.merge_cells(f'A{r}:F{r}')

# B. 認定者推計
r=12; w(f'A{r}','B. 要支援・要介護認定者の推計',SUB,SUBF); ws.merge_cells(f'A{r}:F{r}')
r=13
for col,v in zip(['A','B','C','D','E','F'],['区分','2023実績','R9(2027)','R11(2029)','2035','2040']): w(f'{col}{r}',v,H,HF,align='center')
w('A14','認定者数(人)',Font(bold=True))
for col,v in zip(['B','C','D','E','F'],[578,584,598,605,604]): w(f'{col}14',v,fmt='#,##0',align='center')
w('A15','認定率(%)',Font(bold=True))
for col,v in zip(['B','C','D','E','F'],[17.6,18.3,18.9,20.4,21.3]): w(f'{col}15',v,fmt='0.0',align='center')
r=16; w(f'A{r}','◆ 高齢者数は減るが後期高齢化・重度化で認定率上昇。認定者は横ばい〜微増、給付費は当面増加圧力。',Font(bold=True,color=ORANGE,size=9),LORANGE); ws.merge_cells(f'A{r}:F{r}')

# C. 基金
r=18; w(f'A{r}','C. 介護給付費準備基金の確認',SUB,SUBF); ws.merge_cells(f'A{r}:F{r}')
rows=[('第9期 基金残高(算定時)','131,700千円','第9期計画 保険料算出より'),
      ('第9期 予定取崩額','78,000千円','6,500円実現のため取崩'),
      ('第10期開始時 基金残高(計画ベース)','53,700千円','131,700−78,000。R8.6確定値は町確認待ち')]
for i,(a,b,c) in enumerate(rows):
    r=19+i; w(f'A{r}',a,Font(bold=True)); w(f'B{r}',b,align='center',fill=YEL if i==2 else None); w(f'C{r}',c,Font(size=9)); ws.merge_cells(f'C{r}:F{r}')

# D. 保険料試算(06から参照)
r=23; w(f'A{r}','D. 第10期 保険料基準額の試算（3パターン・シート06から参照）',SUB,SUBF); ws.merge_cells(f'A{r}:F{r}')
r=24
for col,v in zip(['A','B','C','D','E','F'],['','第8期','第9期','第10期A','第10期B','第10期C']): w(f'{col}{r}',v,H,HF,align='center')
w('A25','基準月額(円)',Font(bold=True))
w('B25',6380,fmt='#,##0',align='center'); w('C25',6500,fmt='#,##0',align='center')
w('D25',"='06_Step7-8_収納額_基準額'!C18",Font(bold=True,color='C00000'),fmt='#,##0',align='center')
w('E25',"='06_Step7-8_収納額_基準額'!D18",Font(bold=True,color='C00000'),fmt='#,##0',align='center')
w('F25',"='06_Step7-8_収納額_基準額'!E18",Font(bold=True,color='C00000'),fmt='#,##0',align='center')
w('A26','取崩想定',Font(size=9))
for col,v in zip(['D','E','F'],['取崩なし','50%取崩','全額取崩']): w(f'{col}26',v,Font(size=9),align='center')
r=27; w(f'A{r}','◆ 第9期6,500円は基金131,700千円から78,000千円取崩で実現。第10期は基金枯渇で抑制余地が小さく、',Font(bold=True,color=ORANGE,size=9),LORANGE); ws.merge_cells(f'A{r}:F{r}')
r=28; w(f'A{r}','　 給付費増(+2.9%)も加わり7,000円台への上昇圧力。6,500円据置では基金が第10期中に枯渇。',Font(bold=True,color=ORANGE,size=9),LORANGE); ws.merge_cells(f'A{r}:F{r}')

# 出典
r=30; w(f'A{r}','【出典】人口・高齢化率:国立社会保障・人口問題研究所「日本の地域別将来推計人口(令和5年推計)」／川崎町第9期計画。',Font(size=8,color=GRAY),border=False); ws.merge_cells(f'A{r}:F{r}')
r=31; w(f'A{r}','　　　　給付費・保険料・基金枯渇:前提を置いた参考試算。確定値は町提供データ(基金R8.6・収納率・所得段階別人口)で算定。',Font(size=8,color=GRAY),border=False); ws.merge_cells(f'A{r}:F{r}')

wb.save('川崎町_保険料試算ワークブック.xlsx')
print("sheet 10 added. sheets:", wb.sheetnames)
