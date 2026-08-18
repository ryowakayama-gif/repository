# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb=openpyxl.load_workbook('wb_work.xlsx')
YEL=PatternFill('solid',fgColor='FFF2CC')
BLU=Font(color='0000FF')           # input
GRN=Font(color='008000')           # cross-sheet
BLK=Font(color='000000')
NUMFMT='#,##0'
def setnum(ws,addr,val,fmt=NUMFMT,font=None,fill=None,note=None):
    c=ws[addr]; c.value=val; c.number_format=fmt
    if font: c.font=font
    if fill: c.fill=fill
    if note: c.comment=Comment(note,"ビズアップ")

# ============================================================
# 02_Step1-2: サービス別給付費(R9-R11)＋補完項目行を追加、合計をF11へ
# ============================================================
ws=wb['02_Step1-2_標準給付費']
# サービス別(千円) R9,R10,R11
svc={
 6:[390630,395318,400061],   # 居宅
 7:[162673,164625,166601],   # 地密
 8:[471967,477630,483362],   # 施設
 9:[3058,3094,3131],         # 住改
}
for r,vals in svc.items():
    for col,v in zip(['C','D','E'],vals):
        setnum(ws,f'{col}{r}',v,font=BLU)
# 行10を「特定入所者・高額・審査支払手数料 等(補完項目)」に変更
ws['A10']='特定入所者・高額・審査支払手数料 等'
setnum(ws,'B10',87402,font=BLK)  # R8参考
for col,v in zip(['C','D','E'],[88451,89512,90586]):
    setnum(ws,f'{col}10',v,font=BLU)
ws['F10']='=SUM(C10:E10)'; ws['F10'].number_format=NUMFMT
ws['G10']='第9期R8実績(特定入所者55,714+高額28,475+高額医療2,505+審査708)×年1.2%'
# 行11: Step2 標準給付費合計
ws['A11']='Step2: 標準給付費合計'; ws['A11'].font=Font(bold=True)
for col in ['B','C','D','E','F']:
    ws[f'{col}11']=f'=SUM({col}6:{col}10)'; ws[f'{col}11'].number_format=NUMFMT; ws[f'{col}11'].font=Font(bold=True)
ws['G11']='Step3-5で使用(総給付額+補完項目=標準給付費見込額)'
# B6-B9のR3実績参考はそのまま。G注記更新
ws['A12']='※ R9-R11は第9期計画R8実績(総給付額1,016,134千円)を基に年1.2%(後期高齢化・重度化・R9報酬改定)で推計し、サービス構成比(R3実績)で配分。'
ws['A13']='※ 補完項目=特定入所者介護サービス費・高額介護サービス費・高額医療合算・審査支払手数料。標準給付費見込額=総給付額+補完項目。'
ws['A14']='※ 確定値はアンケート結果反映後の見込量(Phase2 W8-W9)で更新。給付費の伸び率は前提値であり町データで精緻化。'

# ============================================================
# 03_Step3-4: 地域支援事業費(R9-R11)、参照をF11へ
# ============================================================
ws=wb['03_Step3-4_地域支援等']
chiiki={6:5377,7:1856,8:23288}  # 千円/年
for r,v in chiiki.items():
    for col in ['C','D','E']:
        setnum(ws,f'{col}{r}',v,font=BLU)
ws['B14']="='02_Step1-2_標準給付費'!F11"; ws['B14'].font=GRN   # 標準給付費合計の新位置
ws['A10']='※ 地域支援事業費は第9期水準(30,520千円/年)を構成比配分。認知症基本法対応で上方圧力あり、確定はPhase2。'

# ============================================================
# 04_Step5: 調整交付金 相当額(5%)＋見込額(実績3.9%)に再構成
# ============================================================
ws=wb['04_Step5_調整交付金']
ws['A7']='標準調整交付金率'; ws['B7']='5.0%'; ws['D7']='国基準値'
ws['A8']='川崎町の調整交付金見込率(実績)'; setnum(ws,'B8',0.039,'0.0%',font=BLU,fill=YEL,
    note='第9期実績:調整交付金見込額129,349千円÷標準給付費3,294,553千円≒3.9%。第10期は国通知で確定。')
ws['D8']='第9期実績ベース(3.9%)。川崎町は標準5%を下回る→保険料は増要因'
ws['A9']='標準給付費合計(Step2)'; ws['B9']="='02_Step1-2_標準給付費'!F11"; ws['B9'].font=GRN; ws['B9'].number_format=NUMFMT
ws['A10']='調整交付金相当額(標準5%)'; ws['B10']='=B9*0.05'; ws['B10'].number_format=NUMFMT; ws['D10']='標準給付費×5%。Step7で加算'
ws['A11']='調整交付金見込額(実績率)'; ws['B11']='=B9*B8'; ws['B11'].number_format=NUMFMT; ws['D11']='標準給付費×実績率。Step7で減算'
ws['A12']='■ 保険料への影響(相当額−見込額)'
ws['A13']='相当額−見込額'; ws['B13']='=B10-B11'; ws['B13'].number_format=NUMFMT
ws['D13']='プラス=川崎町は標準5%より調整交付金が少なく、保険料の増要因(第1号負担に上乗せ)'
ws['A14']='※ 介護保険財政では第1号負担分(23%)に標準5%相当を加算し実際の調整交付金見込額を減算する(第9期計画と同方式)。'

# ============================================================
# 05_Step6: 基金残高参照はそのまま(01!B15)。注記更新
# ============================================================
ws=wb['05_Step6_基金3パターン']
ws['A13']='※ 基金残高(B5)はシート01のB15(第10期開始時・計画ベース53,700千円)を参照。R8.6確定値受領後に更新。'

# ============================================================
# 06_Step7-8: 正しい8ステップに全面修正(自己参照・参照ズレを解消)
# ============================================================
ws=wb['06_Step7-8_収納額_基準額']
for rng in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(rng))
# クリア(値のある範囲を一旦再設定)
for r in range(6,24):
    for col in ['A','B','C','D','E','F']:
        ws[f'{col}{r}'].value=None
ws['A6']='項目'; ws['B6']='共通値'; ws['C6']='パターンA'; ws['D6']='パターンB'; ws['E6']='パターンC'; ws['F6']='備考'
for col in ['A','B','C','D','E','F']: ws[f'{col}6'].font=Font(bold=True,color='FFFFFF'); ws[f'{col}6'].fill=PatternFill('solid',fgColor='1F3864')
ws['A7']='Step4 第1号負担分(23%)'; ws['B7']="='03_Step3-4_地域支援等'!B17"; ws['B7'].font=GRN; ws['B7'].number_format=NUMFMT
ws['C7']='=$B$7'; ws['D7']='=$B$7'; ws['E7']='=$B$7'; ws['F7']='3パターン共通'
ws['A8']='調整交付金相当額(5%・加算)'; ws['B8']="='04_Step5_調整交付金'!B10"; ws['B8'].font=GRN; ws['B8'].number_format=NUMFMT
ws['C8']='=$B$8'; ws['D8']='=$B$8'; ws['E8']='=$B$8'; ws['F8']='3パターン共通(加算)'
ws['A9']='調整交付金見込額(実績・減算)'; ws['B9']="='04_Step5_調整交付金'!B11"; ws['B9'].font=GRN; ws['B9'].number_format=NUMFMT
ws['C9']='=$B$9'; ws['D9']='=$B$9'; ws['E9']='=$B$9'; ws['F9']='3パターン共通(減算)'
ws['A10']='保険者機能強化交付金(減算)'; setnum(ws,'B10',10800,font=BLU,fill=YEL,note='第9期実績10,800千円(3年計)。第10期は国通知で確定。')
ws['C10']='=$B$10'; ws['D10']='=$B$10'; ws['E10']='=$B$10'; ws['F10']='3パターン共通(減算)'
ws['A11']='基金取崩額(パターン別・減算)'; ws['B11']='(パターン別)'
ws['C11']="='05_Step6_基金3パターン'!C9"; ws['D11']="='05_Step6_基金3パターン'!C10"; ws['E11']="='05_Step6_基金3パターン'!C11"
for col in ['C','D','E']: ws[f'{col}11'].font=GRN; ws[f'{col}11'].number_format=NUMFMT
ws['F11']='シート05から参照'
ws['A12']='Step7 保険料収納必要額'; ws['A12'].font=Font(bold=True)
ws['C12']='=$B$7+$B$8-$B$9-$B$10-C11'; ws['D12']='=$B$7+$B$8-$B$9-$B$10-D11'; ws['E12']='=$B$7+$B$8-$B$9-$B$10-E11'
for col in ['C','D','E']: ws[f'{col}12'].font=Font(bold=True); ws[f'{col}12'].number_format=NUMFMT
ws['F12']='第1号負担分+調整交付金相当額−見込額−機能強化−基金取崩'
# Step8
ws['A14']='Step8：保険料基準月額の算定（円）'; ws['A14'].font=Font(bold=True)
ws['A15']='第1号被保険者数 3年合計'; ws['B15']="='01_入力_基礎データ'!B11"; ws['B15'].font=GRN; ws['B15'].number_format=NUMFMT; ws['D15']='シート01から自動参照(R9-R11合計)'
ws['A16']='予定収納率'; setnum(ws,'B16',0.96,'0.0%',font=BLU,fill=YEL,note='第9期と同じ96%を暫定使用。確定は過去5年(R4-R7)平均(町確認待ち)。シート01のC節参照。')
ws['D16']='第9期と同水準96%(暫定)'
ws['A17']='補正係数(所得段階分布)'; setnum(ws,'B17',1.0,'0.00',font=BLU,fill=YEL,note='所得段階別加入割合の加重平均料率。暫定1.0。確定は所得段階別人口(町確認待ち)で精緻化。')
ws['D17']='暫定1.0。所得段階別人口で精緻化'
ws['A18']='Step8 保険料基準月額(円)'; ws['A18'].font=Font(bold=True)
for col in ['C','D','E']:
    ws[f'{col}18']=f'=IFERROR({col}12*1000/($B$15*12*$B$16*$B$17),"町データ入力後算定")'
    ws[f'{col}18'].number_format='#,##0"円"'; ws[f'{col}18'].font=Font(bold=True,color='C00000')
ws['F18']='第10期基準月額(3パターン)'
# 比較
ws['A20']='■ 第9期との比較'; ws['A20'].font=Font(bold=True)
ws['A21']='比較項目'; ws['B21']='第8期'; ws['C21']='第9期'; ws['D21']='第10期A'; ws['E21']='第10期B'; ws['F21']='第10期C'
for col in ['A','B','C','D','E','F']: ws[f'{col}21'].font=Font(bold=True)
ws['A22']='月額(円)'; setnum(ws,'B22',6380,'#,##0"円"'); setnum(ws,'C22',6500,'#,##0"円"')
ws['D22']='=IFERROR(C18,"未算定")'; ws['E22']='=IFERROR(D18,"未算定")'; ws['F22']='=IFERROR(E18,"未算定")'
for col in ['D','E','F']: ws[f'{col}22'].number_format='#,##0"円"'
ws['A23']='第9期比増減率(%)'
ws['D23']='=IFERROR((D18/$C$22-1)*100,"未算定")'; ws['E23']='=IFERROR((E18/$C$22-1)*100,"未算定")'; ws['F23']='=IFERROR((F18/$C$22-1)*100,"未算定")'
for col in ['D','E','F']: ws[f'{col}23'].number_format='+0.0;-0.0'

for rng in ['A1:F1','A2:F2','A4:F4','A5:F5','A20:F20']:
    ws.merge_cells(rng)

# ============================================================
# 01_入力: 被保険者数・基金・収納率平均の修正
# ============================================================
ws=wb['01_入力_基礎データ']
setnum(ws,'B7',3296,font=BLU,note='第9期計画コーホート推計(R8・住基9月末)')
setnum(ws,'B8',3262,font=BLU,note='第9期計画コーホート推計(R9・住基9月末)')
setnum(ws,'B9',3234,font=BLU,note='社人研R5推計の減少率(年-0.43%)でR9から補外')
setnum(ws,'B10',3206,font=BLU,note='社人研R5推計の減少率でR10から補外')
# 基金: 計画ベース第10期開始値
setnum(ws,'B15',53700,font=BLU,fill=YEL,note='第9期計画の基金残高131,700千円−予定取崩78,000千円=53,700千円(計画ベース理論値)。R8.6確定値は町確認待ち。第9期に給付が計画を下回れば残高はこれより大きくなる。')
ws['D15']='第10期開始時(R8末)・計画ベース。R8.6確定値は町確認待ち'
# 第9期基金の確定値を参考行として追記
ws['A16']='(参考)第9期 基金残高/予定取崩'; ws['B16']='131,700 / 78,000'; ws['D16']='第9期計画 保険料算出より(千円)。第9期は6,500円を実現'
# 収納率5年平均の範囲修正(B19はヘッダのため除外)
ws['B25']='=IFERROR(AVERAGE(B20:B24),"R4-R7入力後算定")'
ws['D25']='=IFERROR(AVERAGE(D20:D24),"R4-R7入力後算定")'
ws['F25']='予定収納率(R4-R7町確認後に確定)。暫定はStep8で96%使用'

wb.save('川崎町_保険料試算ワークブック.xlsx')
print("saved. now recalc...")
