# -*- coding: utf-8 -*-
"""
業務工程管理表の更新（令和8年9月3日）

・03_確認事項一覧に No.60〜62（自然体推計・第1号負担割合・施策反映）を追加
・No.48（所得段階の記述の是正）を、素案Ver.1.9で反映済みとして完了に更新
・No.9（準備基金の残高）の反映先を素案Ver.1.9に追記
"""
import copy

import openpyxl

WB = "川崎町_業務工程管理表.xlsx"


def unmerge_below(ws, at_row):
    kept = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= at_row:
            kept.append((m.min_row, m.max_row, m.min_col, m.max_col))
            ws.unmerge_cells(str(m))
    return kept


def remerge(ws, kept, n):
    for r1, r2, c1, c2 in kept:
        ws.merge_cells(start_row=r1 + n, end_row=r2 + n,
                       start_column=c1, end_column=c2)


def copy_style(ws, src_row, dst_row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(dst_row, c)._style = copy.copy(ws.cell(src_row, c)._style)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


KAKUNIN = [
    ("60", "(5)", "R8.9",
     "地域包括ケア「見える化」システムの推計ツールの利用",
     "第10期のサービス見込量は、地域包括ケア「見える化」システムの推計ツール（自然体推計）"
     "により算定し、同システムへ登録することが前提となります。"
     "厚生労働省『自然体推計の計算過程確認シートのガイド』により推計の枠組みを確認しました。"
     "町のシステム利用の可否、操作を担当される方、及び受託者による支援の範囲について"
     "ご確認をお願いします。",
     "サービス見込量の推計（仕様書6（5））\n保険料試算\n素案 第7章",
     "発注者", "確認待ち", "R8.10",
     "第3回で答申を確定する段取りのため、令和8年12月までに推計を終える必要がある。"),

    ("61", "(5)", "R8.9",
     "第10期の第1号被保険者負担割合",
     "自然体推計ガイドの係数表では、第1号被保険者負担割合は第9期が23.00％、"
     "令和12年度が24.00％とされています。第10期（令和9〜11年度）の割合は政令で定められます。"
     "当社の概算試算は23％を用いており、24％となった場合は保険料基準額が"
     "月額約322円上昇します（中位・基金50％取崩で7,023円→約7,345円）。"
     "国の政令の内容が判明次第、試算に反映します。",
     "保険料試算\n素案 第7章\n第3回委員会資料",
     "受託者・発注者", "実施中", "R8.12",
     "受託者が随時確認する。資料提供依頼No.14（第10期の政令改正の内容）と一体。"),

    ("62", "(5)(7)", "R8.9",
     "自然体推計に対する施策反映の方針",
     "自然体推計は、要介護認定率・サービス利用率・1人1月あたり利用回（日）数について"
     "直近年度の実績が継続すると仮定するものです。第9期計画の見込量が3か年ほぼ同値なのは"
     "この仕様によるもので、施策反映を加えていないためです。"
     "第10期では、①介護予防の強化による認定率の抑制、②在宅サービスの利用勧奨、"
     "③施設・居住系の整備量の変更、④サービス提供体制の制約 を"
     "どこまで見込量に反映させるかが論点となります。とくに④は自然体推計では表現できず、"
     "事業所ヒアリングによる供給可能量の把握が前提となります。",
     "サービス見込量の推計\n素案 第7章\n第2回委員会資料（選択事項2・3）",
     "発注者", "確認待ち", "R8.11",
     "確認事項No.50・51（事業所ヒアリング）と一体。第2回策定委員会でお諮りする。"),
]


def main():
    wb = openpyxl.load_workbook(WB)
    ws = wb["03_確認事項一覧"]

    # No.48（所得段階の記述の是正）を完了に
    ws.cell(52, 8).value = "完了"
    cur = ws.cell(52, 10).value or ""
    ws.cell(52, 10).value = (
        cur + "／素案Ver.1.9で是正済み。第2章の記述と第7章の該当節（見出しを含む5箇所）を、"
        "「令和6年度に13段階へ移行済みであり、第10期の論点は乗率の設定である」と改めた。")

    # No.9（準備基金）の反映先を追記
    cur9 = ws.cell(13, 10).value or ""
    ws.cell(13, 10).value = cur9 + "／素案Ver.1.9の第7章に残高（R7末148,100千円・R8末見込152,500千円）を記載済み。"

    at = 64                      # No.59（r63）の直後
    n = len(KAKUNIN)
    kept = unmerge_below(ws, at)
    ws.insert_rows(at, n)
    remerge(ws, kept, n)
    for i, rec in enumerate(KAKUNIN):
        r = at + i
        copy_style(ws, 63, r, 10)
        for c, v in enumerate(rec, start=1):
            ws.cell(r, c).value = v
    last = 63 + n
    for r in range(last + 2, last + 8):
        cell = ws.cell(r, 2)
        if isinstance(cell.value, str) and cell.value.startswith("=COUNTIF"):
            cell.value = cell.value.replace("H5:H63", f"H5:H{last}")

    wb.save(WB)
    print(f"03_確認事項一覧：No.60〜62 を追加（{n}件）／No.48を完了に更新、No.9に反映先を追記")


if __name__ == "__main__":
    main()
