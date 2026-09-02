# -*- coding: utf-8 -*-
"""
町ご提供の「12〜14 サービス見込量」シートと、第9期計画書の見込量表との突合

第9期計画書は各サービスについて
  「令和5年度／令和6年度／令和7年度／令和8年度」の
  「延利用人数（人／年）」「給付費（千円／年）」を掲げている。
本スクリプトは計画書PDFからこれを抽出し、シートのR6〜R8と照合する。
"""
import re

import openpyxl
import pypdf

PDF = "09_元資料/川崎町_第9期計画_04324.pdf"
SRC = "09_元資料/R8実績データ/R8.9.1受領版/川崎町_町提供実績データ_R8.9.1受領.xlsx"


def plan_values():
    """計画書から {サービス名: {'人': [R5,R6,R7,R8], '費': [...]}} を作る。"""
    r = pypdf.PdfReader(PDF)
    text = "\n".join((p.extract_text() or "") for p in r.pages)
    text = re.sub(r"[ \t]+", " ", text)
    out = {}
    for m in re.finditer(r"■(.+?)見込量(.*?)(?=■|【|第\s*\d+\s*節|\Z)", text, re.S):
        name = m.group(1).strip().replace("\n", "")
        body = m.group(2)
        nums = re.findall(r"(?:延利用人数|延利用回数|利用人数)[^0-9\-]*((?:[\d,]+\s+){3}[\d,]+)", body)
        cost = re.findall(r"給付費[^0-9\-]*((?:[\d,]+\s+){3}[\d,]+)", body)
        d = {}
        if nums:
            d["人"] = [int(x.replace(",", "")) for x in nums[0].split()]
        if cost:
            d["費"] = [int(x.replace(",", "")) for x in cost[0].split()]
        if d:
            out[name] = d
    return out


def sheet_values():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = {}
    for s in ["12_サービス見込量（居宅サービス）",
              "13_サービス見込量（地域密着型サービス）",
              "14_サービス見込量（施設サービス）"]:
        ws = wb[s]
        cur = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if isinstance(a, str) and a.startswith("■"):
                cur = a.replace("■", "").replace("の実績と見込量", "").strip()
                out[cur] = {}
            if cur and ws.cell(r, 3).value == "見込量":
                key = "人" if isinstance(a, str) and "利用" in a else ("費" if isinstance(a, str) and "給付費" in a else None)
                if key:
                    v = [ws.cell(r, c).value for c in (4, 5, 6)]
                    if all(isinstance(x, (int, float)) for x in v):
                        out[cur][key] = [int(x) for x in v]
    return out


def norm(s):
    s = re.sub(r"[（(].*?[）)]", "", s)
    return s.replace("・", "").replace(" ", "").replace("　", "").strip()


def main():
    plan, sheet = plan_values(), sheet_values()
    pmap = {norm(k): (k, v) for k, v in plan.items()}
    print(f"計画書から抽出したサービス：{len(plan)}件　／　シートのサービス：{len(sheet)}件\n")
    ok = ng = nf = 0
    for name, sv in sheet.items():
        hit = pmap.get(norm(name))
        if not hit:
            for k, val in pmap.items():
                if norm(name) and (norm(name) in k or k in norm(name)):
                    hit = val
                    break
        if not hit:
            if sv:
                nf += 1
                print(f"  ─ 計画書に対応が見つからない：{name}")
            continue
        pname, pv = hit
        for key, label in (("人", "延利用人数"), ("費", "給付費")):
            if key not in sv or key not in pv:
                continue
            s3, p3 = sv[key], pv[key][1:]   # 計画書はR5〜R8。R6〜R8を比較
            if s3 == p3:
                ok += 1
            else:
                ng += 1
                print(f"  ★{name}／{label}")
                print(f"      シート R6-R8: {s3}")
                print(f"      計画書 R6-R8: {p3}")
    print(f"\n一致 {ok}項目 ／ 不一致 {ng}項目 ／ 計画書に対応なし {nf}サービス")


if __name__ == "__main__":
    main()
