# -*- coding: utf-8 -*-
"""
業務工程管理表の更新（令和8年9月3日・素案Ver.1.12の作成に伴うもの）

03_確認事項一覧に No.69・70（介護給付適正化事業の区分、KPI管理表の集計体制）を追加する。
根拠：01_第10期_最新版成果品/川崎町_計画素案_修正記録_v1.12.md
"""
import copy

import openpyxl

WB = "川崎町_業務工程管理表.xlsx"
AT = 73          # No.68（r72）の直後
TPL = 72         # 書式のひな形とする行


def unmerge_below(ws, at_row):
    kept = []
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= at_row:
            kept.append((m.min_row, m.max_row, m.min_col, m.max_col))
            ws.unmerge_cells(str(m))
    return kept


def remerge(ws, kept, n):
    for r1, r2, c1, c2 in kept:
        ws.merge_cells(start_row=r1 + n, end_row=r2 + n,
                       start_column=c1, end_column=c2)


def copy_style(ws, src_row, dst_row, ncol):
    for c in range(1, ncol + 1):
        ws.cell(dst_row, c)._style = copy.copy(ws.cell(src_row, c)._style)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


KAKUNIN = [
    ("69", "(7)", "R8.9",
     "介護給付適正化事業の区分と実施の考え方",
     "素案Ver.1.12の第8章8-2に、介護給付適正化の5事業（要介護認定の適正化、"
     "ケアプランの点検、医療情報との突合・縦覧点検、福祉用具・住宅改修等の点検、"
     "介護給付費通知）を掲載しました。国は事業区分の整理を進めており、"
     "第10期の区分は国の基本指針の確定を待つ必要があります。"
     "あわせて、①現在実施している事業、②ケアプラン点検の年間件数、"
     "③医療情報との突合の実施状況（交付金の減点項目）をご確認ください。"
     "令和8年度の交付金評価では「給付費適正化事業の取組状況」が6点から0点となっています。",
     "素案 第8章8-2\n第2回委員会資料",
     "発注者", "確認待ち", "R8.10",
     "国の基本指針が確定した段階で受託者が区分を整理する。"),

    ("70", "(7)", "R8.9",
     "KPI管理表の集計担当・集計方法・頻度",
     "素案Ver.1.12の第10章10-4にKPI管理表を新設し、主要13指標について"
     "区分・現状値・目標値・集計担当／方法・頻度を掲載しました。"
     "第9期計画の指標は集計担当・方法・頻度・出典が定められておらず、"
     "実績を確認できない指標が残りました。"
     "各指標の集計担当（保健福祉課・地域包括支援センター・社会福祉協議会の別）と、"
     "集計方法・頻度をご確認ください。"
     "現状値が【町確認】の指標については、記録様式を整備する年度もあわせてご検討ください。",
     "素案 第10章10-4\n第2回委員会資料\n第9期実績一覧",
     "発注者", "確認待ち", "R8.11",
     "確認事項No.8・34（第9期の事業実績）と一体。"
     "保険者機能強化推進交付金の「評価結果の活用」（令和8年度16点→0点）の回復に直結する。"),
]


def main():
    wb = openpyxl.load_workbook(WB)
    ws = wb["03_確認事項一覧"]

    n = len(KAKUNIN)
    kept = unmerge_below(ws, AT)
    ws.insert_rows(AT, n)
    remerge(ws, kept, n)
    for i, rec in enumerate(KAKUNIN):
        r = AT + i
        copy_style(ws, TPL, r, 10)
        for c, v in enumerate(rec, start=1):
            ws.cell(r, c).value = v

    last = TPL + n
    for r in range(last + 2, last + 9):
        cell = ws.cell(r, 2)
        if isinstance(cell.value, str) and cell.value.startswith("=COUNTIF"):
            cell.value = cell.value.replace(f"H5:H{TPL}", f"H5:H{last}")

    wb.save(WB)
    print(f"03_確認事項一覧：No.69・70 を追加（{n}件）／集計式の範囲を H5:H{last} に更新")


if __name__ == "__main__":
    main()
