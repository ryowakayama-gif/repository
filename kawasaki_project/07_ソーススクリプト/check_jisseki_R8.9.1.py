# -*- coding: utf-8 -*-
"""
町ご提供の実績データ（給付・事業実績）の内容確認

・年度列に同一値が連続する箇所（見込＝実績のコピーが疑われる箇所）の抽出
・既存の確定資料（介護保険事業状況報告・第9期計画書・第1回策定委員会資料）との突合
・単位・桁・内訳の整合確認
"""
import re
from pathlib import Path

import openpyxl

SRC = "09_元資料/R8実績データ/R8.9.1受領版/川崎町_町提供実績データ_R8.9.1受領.xlsx"
YEARS = ["R2", "R3", "R4", "R5", "R6", "R7見込", "R8当初・見込"]


def num(v):
    return v if isinstance(v, (int, float)) else None


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def scan_same(ws, name_cols, first_year_col, label_prefix=""):
    """年度列で3年以上同一値が続く行を抽出する。"""
    out = []
    for r in range(5, ws.max_row + 1):
        labels = [ws.cell(r, c).value for c in name_cols]
        if all(v in (None, "") for v in labels):
            continue
        vals = [num(ws.cell(r, first_year_col + i).value) for i in range(len(YEARS))]
        if sum(1 for v in vals if v is not None) < 3:
            continue
        # 末尾から連続して同じ値が何年続くか
        run, last = 1, None
        best = 1
        for v in vals:
            if v is None:
                run, last = 1, None
                continue
            if last is not None and v == last:
                run += 1
                best = max(best, run)
            else:
                run = 1
            last = v
        if best >= 3:
            lab = " / ".join(str(v) for v in labels if v not in (None, ""))
            out.append((r, label_prefix + lab,
                        " ".join("―" if v is None else f"{v:g}" for v in vals)))
    return out


def main():
    wb = load(SRC)
    print("=" * 78)
    print("■ 1. 年度列に3年以上の同一値が続く行（見込＝実績のコピーが疑われる箇所）")
    print("=" * 78)
    targets = [
        ("02_総合事業_介護予防", [1, 2], 4),
        ("03_認定者数", [1, 2], 4),
        ("04_サービス利用給付", [2, 3], 5),
        ("05_基金保険料", [1], 3),
        ("06_移動支援", [1, 2], 4),
        ("07_生活支援見守り", [1, 2], 4),
        ("08_包括認知症権利擁護", [2, 3], 5),
    ]
    total = 0
    for sheet, cols, fy in targets:
        rows = scan_same(wb[sheet], cols, fy)
        if not rows:
            continue
        print(f"\n【{sheet}】")
        for r, lab, vals in rows:
            print(f"  r{r:<4}{lab[:44]:46s} {vals}")
            total += 1
    print(f"\n  → 該当 {total}行")

    print("\n" + "=" * 78)
    print("■ 2. サービス見込量シート（12〜14）の実績欄の記入状況")
    print("=" * 78)
    for sheet in ["12_サービス見込量（居宅サービス）",
                  "13_サービス見込量（地域密着型サービス）",
                  "14_サービス見込量（施設サービス）"]:
        ws = wb[sheet]
        blocks, filled_m, filled_j, filled_10 = 0, 0, 0, 0
        cur = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if isinstance(a, str) and a.startswith("■"):
                blocks += 1
                cur = a
            c3 = ws.cell(r, 3).value
            if c3 == "見込量":
                if any(num(ws.cell(r, c).value) is not None for c in (4, 5, 6)):
                    filled_m += 1
                if any(num(ws.cell(r, c).value) is not None for c in (7, 8, 9)):
                    filled_10 += 1
            if c3 == "実績":
                if any(num(ws.cell(r, c).value) is not None for c in (4, 5, 6, 7, 8, 9)):
                    filled_j += 1
        print(f"  {sheet}")
        print(f"    サービス数(■) {blocks} ／ 第9期見込量の記入行 {filled_m} ／ "
              f"実績の記入行 {filled_j} ／ 第10期見込量の記入行 {filled_10}")

    print("\n" + "=" * 78)
    print("■ 3. 見込量シートで第9期3か年（R6〜R8）が同値のサービス")
    print("=" * 78)
    for sheet in ["12_サービス見込量（居宅サービス）",
                  "13_サービス見込量（地域密着型サービス）",
                  "14_サービス見込量（施設サービス）"]:
        ws = wb[sheet]
        cur = None
        same, diff = [], []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if isinstance(a, str) and a.startswith("■"):
                cur = a.replace("■", "").replace("の実績と見込量", "")
            if ws.cell(r, 3).value == "見込量" and ws.cell(r, 1).value == "延利用者数":
                v = [num(ws.cell(r, c).value) for c in (4, 5, 6)]
                if None in v:
                    continue
                (same if v[0] == v[1] == v[2] else diff).append((cur, v))
        print(f"\n  {sheet}")
        print(f"    R6=R7=R8 が同値：{len(same)}サービス ／ 年度で変動：{len(diff)}サービス")
        for c, v in same[:40]:
            print(f"      同値 {c[:34]:36s} {v[0]:g}")

    print("\n" + "=" * 78)
    print("■ 4. 個別の確認事項（既存資料との突合・単位・内訳）")
    print("=" * 78)
    ws = wb["03_認定者数"]
    v = [num(ws.cell(7, 4 + i).value) for i in range(7)]
    print(f"  75歳以上人口の推移: {v}")
    tot = [num(ws.cell(16, 4 + i).value) for i in range(7)]
    p1 = [num(ws.cell(5, 4 + i).value) for i in range(7)]
    print("  認定率の再計算（認定者合計÷第1号被保険者数）:")
    for i, y in enumerate(YEARS):
        if tot[i] and p1[i]:
            rate = tot[i] / p1[i] * 100
            shown = num(ws.cell(17, 4 + i).value)
            print(f"    {y:8s} {tot[i]:>4}/{p1[i]:>5} = {rate:5.2f}%   表の記載 {shown}%")
    # 内訳の合計チェック
    for i, y in enumerate(YEARS):
        s = sum(num(ws.cell(r, 4 + i).value) or 0 for r in range(9, 16))
        if tot[i] is not None and s != tot[i]:
            print(f"    ★{y} 要介護度別の合計 {s} ≠ 記載の合計 {tot[i]}")

    ws = wb["04_サービス利用給付"]
    print("\n  給付費の内訳合計と総給付費の比較（千円）:")
    for i, y in enumerate(YEARS):
        parts = sum(num(ws.cell(r, 5 + i).value) or 0
                    for r in (6, 8, 10, 12, 14, 16, 18, 20, 22, 24))
        tot2 = num(ws.cell(26, 5 + i).value)
        if tot2:
            print(f"    {y:8s} 掲載サービスの計 {parts:>9,} ／ 総給付費 {tot2:>9,} "
                  f"／ 差 {tot2 - parts:>9,}（掲載外サービス分を含む）")

    ws5 = wb["05_基金保険料"]
    print("\n  05シートと04シートの給付費の突合（千円）:")
    for i, y in enumerate(YEARS):
        a = num(ws5.cell(6, 3 + i).value)
        b = num(ws.cell(26, 5 + i).value)
        if a is not None and b is not None and a != b:
            print(f"    ★{y:8s} 05保険給付費 {a:,} ≠ 04総給付費 {b:,} （差 {a - b:,}）")

    ws7 = wb["07_生活支援見守り"]
    print("\n  紙おむつ等支給事業:")
    for i, y in enumerate(YEARS):
        n = num(ws7.cell(11, 4 + i).value)
        c = num(ws7.cell(12, 4 + i).value)
        if n and c:
            print(f"    {y:8s} 利用実人数 {n:>3}人 事業費 {c:>9,}（表の単位は千円）"
                  f" → 1人当たり {c / n:,.0f}")

    ws8 = wb["08_包括認知症権利擁護"]
    print("\n  認知症サポーター（累計と年間受講者の整合）:")
    cum = [num(ws8.cell(23, 5 + i).value) for i in range(7)]
    yr = [num(ws8.cell(10, 5 + i).value) for i in range(7)]
    print(f"    累計   {cum}")
    print(f"    受講者 {yr}")
    for i in range(1, 7):
        if cum[i] is not None and cum[i - 1] is not None:
            d = cum[i] - cum[i - 1]
            print(f"    {YEARS[i]:8s} 累計の増 {d:>4}人 ／ 受講者の記載 {yr[i]}人"
                  f"{'　★不一致' if yr[i] is not None and d != yr[i] else ''}")


if __name__ == "__main__":
    main()
