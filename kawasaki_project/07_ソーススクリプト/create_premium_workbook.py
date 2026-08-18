"""
川崎町第10期 保険料試算ワークブック

8ステップで第10期保険料基準額(月額)を試算するExcel
3パターン(A: 取崩なし / B: 50%取崩 / C: 全額取崩)を並列計算

10シート構成：
00_使い方
01_入力_基礎データ（人口・認定者・収納率等の入力）
02_Step1-2_標準給付費（サービス見込量×単価）
03_Step3-4_地域支援事業費・第1号負担分相当額
04_Step5_調整交付金
05_Step6_3パターン基金取崩設定
06_Step7-8_保険料収納必要額・基準額算定
07_所得段階_9段階版（現行）
08_所得段階_13段階版（協議事項5）
09_比較分析（3パターン×9/13段階）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# カラーパレット
NAVY = "1F3864"; BLUE = "2F5597"; LBLUE = "DAE3F3"
ORANGE = "ED7D31"; LORANGE = "FCE4D6"
GREEN = "548235"; LGREEN = "E2EFDA"
RED = "C00000"; LRED = "FFE4E4"
GRAY = "808080"; LGRAY = "F2F2F2"
PURPLE = "9333B0"; LPURPLE = "EAD5F0"
INPUT_Y = "FFFFCC"; KNOWN_G = "E2EFDA"; CALC_B = "DAE3F3"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def F(c): return PatternFill("solid", fgColor=c)
cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
cl = Alignment(horizontal="left", vertical="center", wrap_text=True)
cr = Alignment(horizontal="right", vertical="center", wrap_text=True)

f_title = Font(name="游ゴシック", size=13, bold=True, color=WHITE)
f_sub   = Font(name="游ゴシック", size=10, italic=True, color=WHITE)
f_head  = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
f_section = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
f_body  = Font(name="游ゴシック", size=10)
f_note  = Font(name="游ゴシック", size=9, italic=True, color="595959")
f_input = Font(name="游ゴシック", size=10, color="0000FF", bold=True)
f_calc = Font(name="游ゴシック", size=10, color=NAVY, bold=True)
f_known = Font(name="游ゴシック", size=10, color="006100", bold=True)
f_result = Font(name="游ゴシック", size=11, color=RED, bold=True)

wb = Workbook()
wb.remove(wb.active)

def ms(ws, rng, val, font, fill, align):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = val; c.font = font
    if fill is not None: c.fill = fill
    c.alignment = align
    from openpyxl.utils.cell import range_boundaries
    a, b, d, e = range_boundaries(rng)
    for r in range(b, e+1):
        for col in range(a, d+1):
            ws.cell(row=r, column=col).border = border

def setup_page(ws, orient="landscape"):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orient == "landscape" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

# ===========================================================
# 00_使い方
# ===========================================================
ws = wb.create_sheet("00_使い方")
ms(ws, "A1:G1", "川崎町第10期 保険料試算ワークブック",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 28
ms(ws, "A2:G2", "── 8ステップ準拠・3パターン(A/B/C)試算・9段階/13段階比較 ──",
   f_sub, F(BLUE), cc)
ws.row_dimensions[2].height = 20

ms(ws, "A4:G4", "1．本ワークブックの目的", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22
ms(ws, "A5:G5",
   "国基本指針に準拠した8ステップで川崎町第10期介護保険料(第1号被保険者基準月額)を試算します。基金取崩額により3パターン(A:取崩なし/B:50%取崩/C:全額取崩)を並列計算し、9段階と13段階区分の比較も実施。第3回策定委員会(R9.1中旬)で確定値を提示するための作業用シートです。",
   f_body, None, cl)
ws.row_dimensions[5].height = 60

ms(ws, "A7:G7", "2．シート構成", f_head, F(BLUE), cl)
ws.row_dimensions[7].height = 22
ws.cell(row=8, column=1, value="シート").font = f_head
ws.cell(row=8, column=1).fill = F(NAVY); ws.cell(row=8, column=1).alignment = cc; ws.cell(row=8, column=1).border = border
ms(ws, "B8:E8", "内容", f_head, F(NAVY), cc)
ms(ws, "F8:G8", "対応ステップ", f_head, F(NAVY), cc)
ws.row_dimensions[8].height = 22

sheets = [
    ("01_入力_基礎データ", "人口・認定者・収納率・基金残高等の基礎数値", "全Step共通"),
    ("02_Step1-2_標準給付費", "サービス見込量と単価から3年間の標準給付費を算定", "Step1-2"),
    ("03_Step3-4_地域支援等", "地域支援事業費と第1号負担分相当額(23%)を算定", "Step3-4"),
    ("04_Step5_調整交付金", "国の調整交付金(5%基本+変動)を算定", "Step5"),
    ("05_Step6_基金3パターン", "基金取崩額の3パターン設定(A/B/C)", "Step6"),
    ("06_Step7-8_収納額・基準額", "保険料収納必要額と基準額(月額)の算定", "Step7-8"),
    ("07_所得段階_9段階版", "現行9段階区分での所得段階別保険料", "Step8拡張"),
    ("08_所得段階_13段階版", "13段階区分での所得段階別保険料(協議事項5)", "Step8拡張"),
    ("09_比較分析", "3パターン×9/13段階のクロス比較・近隣比較", "総合"),
]
r = 9
for sh, cont, step in sheets:
    c = ws.cell(row=r, column=1, value=sh)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:E{r}", cont, f_body, None, cl)
    ms(ws, f"F{r}:G{r}", step, f_body, None, cc)
    ws.row_dimensions[r].height = 26
    r += 1

# 入力凡例
r += 1
ms(ws, f"A{r}:G{r}", "3．セルの色分け凡例", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

legends = [
    (INPUT_Y, "入力欄(黄色)", "町担当課・弊社で入力する数値。基礎データ・基金残高等"),
    (KNOWN_G, "既知数値(緑)", "確定済の数値。R3給付実績・R7.6人口等"),
    (CALC_B, "計算結果(青)", "数式により自動算出される数値。手入力不要"),
    (LRED, "重要結果(赤)", "保険料基準額等の最重要算定結果"),
]
for color, label, mean in legends:
    c = ws.cell(row=r, column=1, value=label)
    c.fill = F(color)
    c.font = Font(name="游ゴシック", size=10, bold=True)
    c.alignment = cc; c.border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 8ステップの全体像
r += 1
ms(ws, f"A{r}:G{r}", "4．保険料算定の8ステップ全体像", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="Step").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:D{r}", "内容", f_head, F(NAVY), cc)
ms(ws, f"E{r}:F{r}", "数式・元データ", f_head, F(NAVY), cc)
ws.cell(row=r, column=7, value="シート").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 22
r += 1

steps = [
    ("Step1", "サービス見込量×単価=標準給付費(年度別)", "見込量(第7章7-1) × サービス単価", "02"),
    ("Step2", "標準給付費3年間総額", "Step1の3年分(R9-R11)を合計", "02"),
    ("Step3", "地域支援事業費(包括的・任意事業)", "Step1の3〜4%相当(国基準)", "03"),
    ("Step4", "第1号負担分相当額", "(Step2+Step3)×23%(第1号負担割合)", "03"),
    ("Step5", "調整交付金額", "Step4×(調整交付金率5%-実績)補正", "04"),
    ("Step6", "基金取崩額の設定(3パターン)", "A:0/B:基金×50%/C:基金×100%", "05"),
    ("Step7", "保険料収納必要額", "Step4-Step5+Step6", "06"),
    ("Step8", "保険料基準月額", "Step7÷(被保険者数×12×収納率×補正)", "06"),
]
for sp, cont, formula, sh in steps:
    c = ws.cell(row=r, column=1, value=sp)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:D{r}", cont, f_body, None, cl)
    ms(ws, f"E{r}:F{r}", formula, f_note, None, cl)
    c = ws.cell(row=r, column=7, value=sh)
    c.font = f_body; c.alignment = cc; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

# 注意事項
r += 1
ms(ws, f"A{r}:G{r}", "5．運用上の注意事項", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

notes = [
    "本ワークブックは作業用です。確定値は介護給付費準備基金残高(R8.6時点)の入手後にシート01で入力してください。",
    "サービス見込量(Step1)は計画素案v1.5第7章7-1の暫定値を使用。Phase 2でアンケート補正後に更新します。",
    "見える化システムの仙南圏域他自治体保険料との比較は、町担当課が見える化システム登録完了後に実施します。",
    "13段階区分(シート08)は国推奨基準率に基づく試算。確定段階率は委員会協議事項5で確定します。",
    "3パターン試算の結果を第3回策定委員会(R9.1中旬)で提示し、委員協議の上、第10期保険料パターンを決定します。",
]
for note in notes:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 26
    r += 1

widths = [22, 18, 12, 14, 14, 14, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 01_入力_基礎データ
# ===========================================================
ws = wb.create_sheet("01_入力_基礎データ")
ms(ws, "A1:F1", "01　基礎データ入力シート", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── 8ステップ算定の前提となる基礎数値を入力 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# A. 第1号被保険者数
r = 4
ms(ws, f"A{r}:F{r}", "A. 第1号被保険者数の推計（Step8で使用）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="年度").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "65歳以上人口(人)", f_head, F(NAVY), cc)
ms(ws, f"D{r}:E{r}", "備考", f_head, F(NAVY), cc)
ws.cell(row=r, column=6, value="出典").font = f_head
ws.cell(row=r, column=6).fill = F(NAVY); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.row_dimensions[r].height = 22
r += 1

pop_data = [
    ("R7(2025)", 3244, "実績(R7.6時点)", "町保険者データ", KNOWN_G),
    ("R8(2026)", None, "推計値(町担当課入力)", "町推計", INPUT_Y),
    ("R9(2027)", None, "第10期初年度推計", "社人研推計+補正", INPUT_Y),
    ("R10(2028)", None, "第10期中年度推計", "社人研推計+補正", INPUT_Y),
    ("R11(2029)", None, "第10期最終年度推計", "社人研推計+補正", INPUT_Y),
]
for year, val, note, source, fill in pop_data:
    c = ws.cell(row=r, column=1, value=year)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", val if val else "[入力]",
       f_known if fill == KNOWN_G else f_input, F(fill), cc)
    ms(ws, f"D{r}:E{r}", note, f_body, None, cl)
    c = ws.cell(row=r, column=6, value=source)
    c.font = f_note; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

# 3年間合計
ws.cell(row=r, column=1, value="3年合計").font = Font(name="游ゴシック", size=10, bold=True, color=NAVY)
ws.cell(row=r, column=1).fill = F(CALC_B); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "=SUM(B8:B10)", f_calc, F(CALC_B), cc)
ms(ws, f"D{r}:E{r}", "R9-R11の合計(Step8で使用)", f_note, None, cl)
c = ws.cell(row=r, column=6, value="計算")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 24
r += 1

# B. 介護給付費準備基金残高
r += 1
ms(ws, f"A{r}:F{r}", "B. 介護給付費準備基金（Step6で使用）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="項目").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "金額(千円)", f_head, F(NAVY), cc)
ms(ws, f"D{r}:F{r}", "備考・時点", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="基金残高(R8.6時点)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "[町担当課入力]", f_input, F(INPUT_Y), cc)
ms(ws, f"D{r}:F{r}", "R8.6確定時点の残高(町担当課に依頼中)", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="計画期間中の予定取崩額").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "[3パターン算定]", f_calc, F(CALC_B), cc)
ms(ws, f"D{r}:F{r}", "シート05_Step6で算定", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# C. 収納率（過去5年実績）
r += 1
ms(ws, f"A{r}:F{r}", "C. 保険料収納率の過去5年実績（Step8で使用）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="年度").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "総合収納率(%)", f_head, F(NAVY), cc)
ms(ws, f"D{r}:E{r}", "現年度分(%)", f_head, F(NAVY), cc)
ws.cell(row=r, column=6, value="出典").font = f_head
ws.cell(row=r, column=6).fill = F(NAVY); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.row_dimensions[r].height = 22
r += 1

rates = [
    ("R3", 96.3, 99.0, "実績(MECEデータ既知)", KNOWN_G),
    ("R4", None, None, "町担当課入力", INPUT_Y),
    ("R5", None, None, "町担当課入力", INPUT_Y),
    ("R6", None, None, "町担当課入力", INPUT_Y),
    ("R7", None, None, "町担当課入力(R7.3末)", INPUT_Y),
]
for year, total, current, source, fill in rates:
    c = ws.cell(row=r, column=1, value=year)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", total if total else "[入力]",
       f_known if fill == KNOWN_G else f_input, F(fill), cc)
    ms(ws, f"D{r}:E{r}", current if current else "[入力]",
       f_known if fill == KNOWN_G else f_input, F(fill), cc)
    c = ws.cell(row=r, column=6, value=source)
    c.font = f_note; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

# 5年平均(予定収納率)
ws.cell(row=r, column=1, value="5年平均").font = Font(name="游ゴシック", size=10, bold=True, color=NAVY)
ws.cell(row=r, column=1).fill = F(CALC_B); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "=AVERAGE(B19:B23)", f_calc, F(CALC_B), cc)
ms(ws, f"D{r}:E{r}", "=AVERAGE(D19:D23)", f_calc, F(CALC_B), cc)
c = ws.cell(row=r, column=6, value="予定収納率(Step8で使用)")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 24
r += 1

# D. 第10期 第1号負担割合
r += 1
ms(ws, f"A{r}:F{r}", "D. 第10期の制度パラメータ（国基準値）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="パラメータ").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "値", f_head, F(NAVY), cc)
ms(ws, f"D{r}:F{r}", "備考", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1

params = [
    ("第1号被保険者負担割合", "23%", "Step4で使用。第10期も第9期と同じ23%(国基準)", KNOWN_G),
    ("第2号被保険者負担割合", "27%", "参考。Step4対象外", KNOWN_G),
    ("公費負担割合(国)", "25%", "参考。調整交付金5%含む", KNOWN_G),
    ("公費負担割合(都道府県)", "12.5%", "参考", KNOWN_G),
    ("公費負担割合(市町村)", "12.5%", "参考", KNOWN_G),
    ("標準調整交付金率", "5%", "Step5で使用。実際は補正あり", KNOWN_G),
    ("地域支援事業費率(対給付費)", "3〜4%", "Step3で使用。上限額管理あり", KNOWN_G),
]
for k, v, note, fill in params:
    c = ws.cell(row=r, column=1, value=k)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", v, f_known if fill == KNOWN_G else f_input, F(fill), cc)
    ms(ws, f"D{r}:F{r}", note, f_note, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

widths = [26, 14, 14, 14, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws, "portrait")

# ===========================================================
# 02_Step1-2_標準給付費
# ===========================================================
ws = wb.create_sheet("02_Step1-2_標準給付費")
ms(ws, "A1:G1", "02　Step1-2 標準給付費の算定", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:G2", "── サービス見込量×単価で年度別給付費を算定 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# Step1: サービス別見込量×単価
r = 4
ms(ws, f"A{r}:G{r}", "Step1：サービス区分別の3年間給付費見込み（千円）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="サービス区分").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="R3実績(参考)").font = f_head
ws.cell(row=r, column=2).fill = F(GREEN); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="R9見込み").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="R10見込み").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ws.cell(row=r, column=5, value="R11見込み").font = f_head
ws.cell(row=r, column=5).fill = F(NAVY); ws.cell(row=r, column=5).alignment = cc; ws.cell(row=r, column=5).border = border
ws.cell(row=r, column=6, value="3年合計").font = f_head
ws.cell(row=r, column=6).fill = F(BLUE); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.cell(row=r, column=7, value="備考").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 28
r += 1

service_data = [
    ("居宅サービス", 365000, "R3実績9.6億×38.1%(参考)"),
    ("地域密着型サービス", 152000, "R3実績9.6億×15.9%"),
    ("施設サービス", 441000, "R3実績9.6億×46.0%(特養68・老健66)"),
    ("住宅改修・福祉用具", 2857, "R3実績(参考値)"),
]
for sv, r3_val, note in service_data:
    c = ws.cell(row=r, column=1, value=sv)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    c = ws.cell(row=r, column=2, value=r3_val)
    c.font = f_known; c.alignment = cr; c.fill = F(KNOWN_G); c.border = border
    c.number_format = "#,##0"
    for col in range(3, 6):
        c = ws.cell(row=r, column=col)
        c.fill = F(INPUT_Y); c.font = f_input; c.alignment = cr; c.border = border
        c.number_format = "#,##0"
    # 合計（数式）
    c = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
    c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
    c.number_format = "#,##0"
    c = ws.cell(row=r, column=7, value=note)
    c.font = f_note; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

# 標準給付費合計（Step2）
ms(ws, f"A{r}:A{r}", "Step2: 標準給付費合計", Font(name="游ゴシック", size=11, bold=True, color=WHITE), F(NAVY), cl)
c = ws.cell(row=r, column=2, value=f"=SUM(B6:B9)")
c.font = f_known; c.fill = F(KNOWN_G); c.alignment = cr; c.border = border
c.number_format = "#,##0"
for col in range(3, 6):
    c = ws.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}9)")
    c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
    c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value=f"=SUM(F6:F9)")
c.font = f_result; c.fill = F(LRED); c.alignment = cr; c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=7, value="Step3で使用")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 32
r += 1

# 注釈
r += 1
notes_step12 = [
    "R3実績は参考値です。R9-R11の見込み値は計画素案v1.5 第7章7-1のサービス見込量(6ステップ算定)に基づき入力します。",
    "住宅改修・福祉用具は便宜上1区分にまとめています。詳細区分が必要な場合は別途展開します。",
    "見込量はR8.7末アンケート結果による補正後の値を入力してください(Phase 2のW8-W9で実施)。",
]
for note in notes_step12:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

widths_s2 = [22, 14, 14, 14, 14, 16, 22]
for i, w in enumerate(widths_s2, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 03_Step3-4_地域支援・第1号負担分
# ===========================================================
ws = wb.create_sheet("03_Step3-4_地域支援等")
ms(ws, "A1:F1", "03　Step3-4 地域支援事業費と第1号負担分相当額", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── 地域支援事業費(Step3)と第1号負担分23%(Step4)を算定 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# Step3
r = 4
ms(ws, f"A{r}:F{r}", "Step3：地域支援事業費（千円）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="区分").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
for i, label in enumerate(["R3実績", "R9見込み", "R10見込み", "R11見込み", "3年合計"]):
    c = ws.cell(row=r, column=i+2, value=label)
    c.font = f_head
    c.fill = F(GREEN) if i == 0 else (F(BLUE) if i == 4 else F(NAVY))
    c.alignment = cc; c.border = border
ws.row_dimensions[r].height = 24
r += 1

local_support = [
    ("総合事業(介護予防・生活支援)", 5650, "R3実績"),
    ("一般介護予防事業", 1950, "R3実績"),
    ("包括的支援事業・任意事業", 24471, "R3実績(包括センター運営等)"),
]
for region, r3_val, note in local_support:
    c = ws.cell(row=r, column=1, value=region)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    c = ws.cell(row=r, column=2, value=r3_val)
    c.font = f_known; c.alignment = cr; c.fill = F(KNOWN_G); c.border = border
    c.number_format = "#,##0"
    for col in range(3, 6):
        c = ws.cell(row=r, column=col)
        c.fill = F(INPUT_Y); c.font = f_input; c.alignment = cr; c.border = border
        c.number_format = "#,##0"
    c = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
    c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
    c.number_format = "#,##0"
    ws.row_dimensions[r].height = 24
    r += 1

# Step3合計
ms(ws, f"A{r}:A{r}", "Step3 合計", Font(name="游ゴシック", size=11, bold=True, color=WHITE), F(NAVY), cl)
c = ws.cell(row=r, column=2, value=f"=SUM(B6:B8)")
c.font = f_known; c.fill = F(KNOWN_G); c.alignment = cr; c.border = border
c.number_format = "#,##0"
for col in range(3, 6):
    c = ws.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}8)")
    c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
    c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value=f"=SUM(F6:F8)")
c.font = f_result; c.fill = F(LRED); c.alignment = cr; c.border = border
c.number_format = "#,##0"
ws.row_dimensions[r].height = 28
r += 1

# Step4: 第1号負担分相当額
r += 1
ms(ws, f"A{r}:F{r}", "Step4：第1号被保険者負担分相当額（千円・国基準23%）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

# 計算ロジック説明
ms(ws, f"A{r}:F{r}", "計算式：(Step2:標準給付費合計 + Step3:地域支援事業費合計) × 23%", f_note, F(LGRAY), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="項目").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "金額(千円・3年合計)", f_head, F(BLUE), cc)
ms(ws, f"D{r}:F{r}", "備考", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1

# 標準給付費合計
ws.cell(row=r, column=1, value="Step2 標準給付費合計").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "='02_Step1-2_標準給付費'!F10", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "シート02から自動参照", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 地域支援事業費合計
ws.cell(row=r, column=1, value="Step3 地域支援事業費合計").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "=F9", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "本シート上部から自動参照", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 合計
ws.cell(row=r, column=1, value="保険給付費等合計").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", f"=B{r-2}+B{r-1}", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "Step2 + Step3", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# Step4 結果
ws.cell(row=r, column=1, value="Step4 第1号負担分相当額").font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", f"=B{r-1}*0.23", f_result, F(LRED), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "× 23%(第1号負担割合)= Step5で使用", f_note, None, cl)
ws.row_dimensions[r].height = 32
r += 1

widths_s3 = [26, 16, 16, 14, 14, 14]
for i, w in enumerate(widths_s3, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws, "portrait")

# ===========================================================
# 04_Step5_調整交付金
# ===========================================================
ws = wb.create_sheet("04_Step5_調整交付金")
ms(ws, "A1:F1", "04　Step5 調整交付金", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── 国の調整交付金(基準5%+変動)を算定 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

r = 4
ms(ws, f"A{r}:F{r}", "Step5：調整交付金額の算定（千円）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

# 解説
ms(ws, f"A{r}:F{r}", "解説：調整交付金は標準調整交付金率5%を基本に、後期高齢者割合・所得段階分布等で町別に補正されます。実際の交付金率は5%±αで、町の状況により4-6%程度になります。", f_note, F(LGRAY), cl)
ws.row_dimensions[r].height = 50
r += 1

ws.cell(row=r, column=1, value="項目").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "値", f_head, F(BLUE), cc)
ms(ws, f"D{r}:F{r}", "備考", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1

# 標準調整交付金率
ws.cell(row=r, column=1, value="標準調整交付金率").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "5.0%", f_known, F(KNOWN_G), cc)
ms(ws, f"D{r}:F{r}", "国基準値", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 実際の交付金率(町補正)
ws.cell(row=r, column=1, value="川崎町の交付金率(暫定)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "5.0%", f_input, F(INPUT_Y), cc)
ms(ws, f"D{r}:F{r}", "町担当課で国通知に基づき調整(暫定値5.0%)", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 標準給付費(Step2)
ws.cell(row=r, column=1, value="標準給付費合計(Step2)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "='02_Step1-2_標準給付費'!F10", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "シート02から参照(千円)", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 調整交付金額(Step5結果)
ws.cell(row=r, column=1, value="Step5 調整交付金額").font = Font(name="游ゴシック", size=11, bold=True, color=WHITE)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", f"=B{r-1}*B{r-2}", f_result, F(LRED), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "標準給付費×交付金率(千円)。Step7で減算", f_note, None, cl)
ws.row_dimensions[r].height = 32
r += 1

# 差額(基準5%との差)
r += 1
ms(ws, f"A{r}:F{r}", "■ 標準5%との差額(調整交付金不足額)", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="標準5%額").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", f"=B8*0.05", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "標準給付費×5%(国標準額)", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

ws.cell(row=r, column=1, value="差額(町実額-標準5%額)").font = Font(name="游ゴシック", size=10, bold=True, color=NAVY)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", f"=B9-B{r-1}", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "差額がプラス→保険料軽減効果あり、マイナス→保険料追加負担", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

widths_s5 = [26, 16, 16, 14, 14, 14]
for i, w in enumerate(widths_s5, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws, "portrait")

# ===========================================================
# 05_Step6_基金3パターン
# ===========================================================
ws = wb.create_sheet("05_Step6_基金3パターン")
ms(ws, "A1:F1", "05　Step6 介護給付費準備基金取崩額の3パターン設定", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── A:取崩なし / B:50%取崩 / C:全額取崩 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# 基金残高入力
r = 4
ms(ws, f"A{r}:F{r}", "■ 基金残高（シート01から自動参照）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="基金残高(R8.6時点・千円)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "='01_入力_基礎データ'!B15", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "シート01のB15から自動参照", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 3パターン
r += 1
ms(ws, f"A{r}:F{r}", "■ 3パターン基金取崩額設定", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="パターン").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="取崩割合").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="取崩額(千円)").font = f_head
ws.cell(row=r, column=3).fill = F(BLUE); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="残高(千円)").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ms(ws, f"E{r}:F{r}", "効果・想定", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 24
r += 1

# パターンA
c = ws.cell(row=r, column=1, value="A 取崩なし")
c.font = Font(name="游ゴシック", size=11, bold=True, color=NAVY); c.alignment = cc
c.fill = F(LBLUE); c.border = border
c = ws.cell(row=r, column=2, value=0)
c.font = f_known; c.alignment = cc; c.fill = F(LBLUE); c.border = border
c.number_format = "0%"
c = ws.cell(row=r, column=3, value=f"=$B$5*B{r}")
c.font = f_result; c.fill = F(LRED); c.alignment = cr; c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=4, value=f"=$B$5-C{r}")
c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
c.number_format = "#,##0"
ms(ws, f"E{r}:F{r}", "基金温存・第11期負担緩和・給付費増を保険料に直接反映(最高水準)", f_body, None, cl)
ws.row_dimensions[r].height = 28
r += 1

# パターンB
c = ws.cell(row=r, column=1, value="B 50%取崩")
c.font = Font(name="游ゴシック", size=11, bold=True, color=C.orange if False else "ED7D31"); c.alignment = cc
c.fill = F(LORANGE); c.border = border
c = ws.cell(row=r, column=2, value=0.5)
c.font = f_known; c.alignment = cc; c.fill = F(LORANGE); c.border = border
c.number_format = "0%"
c = ws.cell(row=r, column=3, value=f"=$B$5*B{r}")
c.font = f_result; c.fill = F(LRED); c.alignment = cr; c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=4, value=f"=$B$5-C{r}")
c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
c.number_format = "#,##0"
ms(ws, f"E{r}:F{r}", "給付費増を一部相殺・次期負担との均衡(中位水準)", f_body, None, cl)
ws.row_dimensions[r].height = 28
r += 1

# パターンC
c = ws.cell(row=r, column=1, value="C 全額取崩")
c.font = Font(name="游ゴシック", size=11, bold=True, color="548235"); c.alignment = cc
c.fill = F(LGREEN); c.border = border
c = ws.cell(row=r, column=2, value=1.0)
c.font = f_known; c.alignment = cc; c.fill = F(LGREEN); c.border = border
c.number_format = "0%"
c = ws.cell(row=r, column=3, value=f"=$B$5*B{r}")
c.font = f_result; c.fill = F(LRED); c.alignment = cr; c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=4, value=f"=$B$5-C{r}")
c.font = f_calc; c.fill = F(CALC_B); c.alignment = cr; c.border = border
c.number_format = "#,##0"
ms(ws, f"E{r}:F{r}", "住民負担最小化・第11期負担増のリスクあり(最低水準)", f_body, None, cl)
ws.row_dimensions[r].height = 28
r += 1

# 注釈
r += 1
notes_s6 = [
    "基金残高(B5)はシート01の確定値を参照しています。基金残高未確定の場合は本シートも未確定となります。",
    "パターンBの50%は標準的な選択肢ですが、25%/75%等の中間値も検討可能です(シート05_2を別途作成可)。",
    "基金取崩は計画期間(3年)の累計額です。年度別配分は別途検討します。",
    "国は基金残高の概ね半分以下を取崩水準として推奨しているため、パターンBが標準的とされます。",
]
for note in notes_s6:
    ms(ws, f"A{r}:F{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 26
    r += 1

widths_s6 = [16, 12, 16, 16, 18, 22]
for i, w in enumerate(widths_s6, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws, "portrait")

# ===========================================================
# 06_Step7-8_収納額・基準額
# ===========================================================
ws = wb.create_sheet("06_Step7-8_収納額_基準額")
ms(ws, "A1:F1", "06　Step7-8 保険料収納必要額と基準月額の算定", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── 3パターン(A/B/C)で保険料基準月額を算定 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# Step7: 保険料収納必要額
r = 4
ms(ws, f"A{r}:F{r}", "Step7：保険料収納必要額（3パターン・千円）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ms(ws, f"A{r}:F{r}", "計算式：Step4(第1号負担分) - Step5(調整交付金) - Step6(基金取崩額)", f_note, F(LGRAY), cl)
ws.row_dimensions[r].height = 22
r += 1

# 必要額算定（3パターン横並び）
ws.cell(row=r, column=1, value="項目").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="共通値").font = f_head
ws.cell(row=r, column=2).fill = F(BLUE); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="パターンA").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="パターンB").font = f_head
ws.cell(row=r, column=4).fill = F(ORANGE); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ws.cell(row=r, column=5, value="パターンC").font = f_head
ws.cell(row=r, column=5).fill = F(GREEN); ws.cell(row=r, column=5).alignment = cc; ws.cell(row=r, column=5).border = border
ws.cell(row=r, column=6, value="備考").font = f_head
ws.cell(row=r, column=6).fill = F(NAVY); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.row_dimensions[r].height = 24
r += 1

# Step4
c = ws.cell(row=r, column=1, value="Step4 第1号負担分")
c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
c.fill = F(LBLUE); c.border = border
c = ws.cell(row=r, column=2, value="='03_Step3-4_地域支援等'!B17")
c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
c.number_format = "#,##0"
for col in range(3, 6):
    c = ws.cell(row=r, column=col, value="=$B$8")
    c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
    c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value="3パターン共通")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 24
r += 1

# Step5
c = ws.cell(row=r, column=1, value="Step5 調整交付金(減算)")
c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
c.fill = F(LBLUE); c.border = border
c = ws.cell(row=r, column=2, value="='04_Step5_調整交付金'!B9")
c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
c.number_format = "#,##0"
for col in range(3, 6):
    c = ws.cell(row=r, column=col, value="=$B$9")
    c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
    c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value="3パターン共通")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 24
r += 1

# Step6 - パターン別
c = ws.cell(row=r, column=1, value="Step6 基金取崩額(減算)")
c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
c.fill = F(LBLUE); c.border = border
c = ws.cell(row=r, column=2, value="(パターン別)")
c.font = f_note; c.alignment = cc; c.fill = F(LGRAY); c.border = border
c = ws.cell(row=r, column=3, value="='05_Step6_基金3パターン'!C9")
c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=4, value="='05_Step6_基金3パターン'!C10")
c.font = f_calc; c.alignment = cr; c.fill = F(LORANGE); c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=5, value="='05_Step6_基金3パターン'!C11")
c.font = f_calc; c.alignment = cr; c.fill = F(LGREEN); c.border = border
c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value="シート05から参照")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 24
r += 1

# Step7: 必要額
c = ws.cell(row=r, column=1, value="Step7 保険料収納必要額")
c.font = Font(name="游ゴシック", size=11, bold=True, color=WHITE); c.alignment = cl
c.fill = F(NAVY); c.border = border
c = ws.cell(row=r, column=2, value="(計算式)")
c.font = f_note; c.alignment = cc; c.fill = F(LGRAY); c.border = border
for col in range(3, 6):
    col_l = get_column_letter(col)
    c = ws.cell(row=r, column=col, value=f"={col_l}8-{col_l}9-{col_l}10")
    c.font = f_result; c.alignment = cr; c.fill = F(LRED); c.border = border
    c.number_format = "#,##0"
c = ws.cell(row=r, column=6, value="Step4-Step5-Step6")
c.font = f_note; c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 32
r += 1

# Step8: 基準月額算定
r += 1
ms(ws, f"A{r}:F{r}", "Step8：保険料基準月額の算定（円）", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ms(ws, f"A{r}:F{r}", "計算式：Step7 ÷ (第1号被保険者数3年合計 × 12ヶ月 × 予定収納率 × 補正係数1.0)", f_note, F(LGRAY), cl)
ws.row_dimensions[r].height = 36
r += 1

# 被保険者数3年合計
ws.cell(row=r, column=1, value="第1号被保険者数 3年合計").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "='01_入力_基礎データ'!B11", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "#,##0"
ms(ws, f"D{r}:F{r}", "シート01から自動参照", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 予定収納率
ws.cell(row=r, column=1, value="予定収納率(5年平均)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "='01_入力_基礎データ'!B24/100", f_calc, F(CALC_B), cc)
ws[f"B{r}"].number_format = "0.0%"
ms(ws, f"D{r}:F{r}", "シート01から自動参照(%→比率変換)", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# 補正係数（所得段階分布による補正）
ws.cell(row=r, column=1, value="補正係数(所得段階分布)").font = Font(name="游ゴシック", size=10, bold=True)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(LBLUE); ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", "1.0", f_input, F(INPUT_Y), cc)
ws[f"B{r}"].number_format = "0.00"
ms(ws, f"D{r}:F{r}", "9段階・13段階の平均係数。後で精緻化", f_note, None, cl)
ws.row_dimensions[r].height = 24
r += 1

# Step8結果
r += 1
ws.cell(row=r, column=1, value="Step8 保険料基準月額(円)").font = Font(name="游ゴシック", size=12, bold=True, color=WHITE)
ws.cell(row=r, column=1).alignment = cl; ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="計算結果").font = f_note
ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).fill = F(LGRAY); ws.cell(row=r, column=2).border = border
for col in range(3, 6):
    col_l = get_column_letter(col)
    # 月額 = Step7 ÷ (3年人口 × 12 × 収納率 × 補正)
    # Step7: row=11, 3年人口: row=15, 収納率: row=16, 補正: row=17
    c = ws.cell(row=r, column=col, value=f'=IFERROR({col_l}11*1000/($B$15*12*$B$16*$B$17),"町データ入力後算定")')
    c.font = Font(name="游ゴシック", size=14, bold=True, color=RED)
    c.alignment = cc; c.fill = F(LRED); c.border = border
    c.number_format = "¥#,##0"
c = ws.cell(row=r, column=6, value="第10期基準月額")
c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cl; c.border = border
ws.row_dimensions[r].height = 40
r += 1

# 第9期比較
r += 1
ms(ws, f"A{r}:F{r}", "■ 第9期との比較", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="比較項目").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="第8期").font = f_head
ws.cell(row=r, column=2).fill = F(GREEN); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="第9期").font = f_head
ws.cell(row=r, column=3).fill = F(GREEN); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="第10期A").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ws.cell(row=r, column=5, value="第10期B").font = f_head
ws.cell(row=r, column=5).fill = F(ORANGE); ws.cell(row=r, column=5).alignment = cc; ws.cell(row=r, column=5).border = border
ws.cell(row=r, column=6, value="第10期C").font = f_head
ws.cell(row=r, column=6).fill = F(GREEN); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.row_dimensions[r].height = 22
r += 1

# 月額
c = ws.cell(row=r, column=1, value="月額(円)")
c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
c.fill = F(LBLUE); c.border = border
c = ws.cell(row=r, column=2, value=6380)
c.font = f_known; c.alignment = cr; c.fill = F(KNOWN_G); c.border = border
c.number_format = "¥#,##0"
c = ws.cell(row=r, column=3, value=6500)
c.font = f_known; c.alignment = cr; c.fill = F(KNOWN_G); c.border = border
c.number_format = "¥#,##0"
for col in range(4, 7):
    col_l = get_column_letter(col-1)
    c = ws.cell(row=r, column=col, value=f'=IFERROR({col_l}18,"未算定")')
    c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
    c.number_format = "¥#,##0"
ws.row_dimensions[r].height = 24
r += 1

# 第9期比増減
c = ws.cell(row=r, column=1, value="第9期比増減率(%)")
c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
c.fill = F(LBLUE); c.border = border
ws.cell(row=r, column=2).fill = F(LGRAY); ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3).fill = F(LGRAY); ws.cell(row=r, column=3).border = border
for col in range(4, 7):
    c = ws.cell(row=r, column=col, value=f'=IFERROR(({get_column_letter(col)}{r-1}/$C${r-1}-1)*100,"未算定")')
    c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
    c.number_format = "+0.0%;-0.0%;0.0%"
ws.row_dimensions[r].height = 24
r += 1

widths_s7 = [22, 14, 14, 14, 14, 22]
for i, w in enumerate(widths_s7, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 07_所得段階_9段階版
# ===========================================================
ws = wb.create_sheet("07_所得段階_9段階")
ms(ws, "A1:G1", "07　所得段階別保険料（9段階版・現行）", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:G2", "── 現行9段階区分での所得段階別年間保険料(月額×12) ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

r = 4
ws.cell(row=r, column=1, value="段階").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="対象").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="係数").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="月額A(円)").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ws.cell(row=r, column=5, value="月額B(円)").font = f_head
ws.cell(row=r, column=5).fill = F(ORANGE); ws.cell(row=r, column=5).alignment = cc; ws.cell(row=r, column=5).border = border
ws.cell(row=r, column=6, value="月額C(円)").font = f_head
ws.cell(row=r, column=6).fill = F(GREEN); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.cell(row=r, column=7, value="該当人数").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 26
r += 1

stages9 = [
    ("第1段階", "生活保護・老齢福祉年金受給者・市町村民税非課税世帯+本人非課税", 0.300),
    ("第2段階", "市町村民税非課税世帯+本人非課税(年金収入80万円超120万円以下)", 0.500),
    ("第3段階", "市町村民税非課税世帯+本人非課税(年金収入120万円超)", 0.700),
    ("第4段階", "本人市町村民税非課税(世帯課税)・課税年金収入80万円以下", 0.900),
    ("第5段階(基準)", "本人市町村民税非課税(世帯課税)・課税年金収入80万円超", 1.000),
    ("第6段階", "本人市町村民税課税・合計所得120万円未満", 1.200),
    ("第7段階", "本人市町村民税課税・合計所得120万円以上210万円未満", 1.300),
    ("第8段階", "本人市町村民税課税・合計所得210万円以上320万円未満", 1.500),
    ("第9段階", "本人市町村民税課税・合計所得320万円以上", 1.700),
]

for stage, target, ratio in stages9:
    c = ws.cell(row=r, column=1, value=stage)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    c = ws.cell(row=r, column=2, value=target)
    c.font = f_body; c.alignment = cl; c.border = border
    c = ws.cell(row=r, column=3, value=ratio)
    c.font = f_known; c.alignment = cc; c.fill = F(KNOWN_G); c.border = border
    c.number_format = "0.000"
    # パターン別月額
    for col, base_col in [(4, "C"), (5, "D"), (6, "E")]:
        c = ws.cell(row=r, column=col, value=f"=IFERROR('06_Step7-8_収納額_基準額'!{base_col}18*$C{r},\"未算定\")")
        c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
        c.number_format = "¥#,##0"
    # 該当人数(R3末・参考)
    c = ws.cell(row=r, column=7)
    c.fill = F(INPUT_Y); c.font = f_input; c.alignment = cr; c.border = border
    c.number_format = "#,##0"
    ws.row_dimensions[r].height = 36
    r += 1

# 注釈
r += 1
notes_s8 = [
    "係数は国基準値です。市町村は条例により調整可能(±10%程度)。",
    "該当人数の合計は所得段階別R3末3,255人(第5基準697人)が参考値です。R7時点の確定値を入力してください。",
    "本シートはステップ8の算定結果を所得段階別に展開した参考表です。条例改正案作成の元データとなります。",
]
for note in notes_s8:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

widths_s8 = [16, 36, 10, 14, 14, 14, 12]
for i, w in enumerate(widths_s8, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 08_所得段階_13段階版
# ===========================================================
ws = wb.create_sheet("08_所得段階_13段階")
ms(ws, "A1:G1", "08　所得段階別保険料（13段階版・協議事項5）", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:G2", "── 国推奨13段階区分での所得段階別保険料 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

r = 4
ws.cell(row=r, column=1, value="段階").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="対象").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="係数").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="月額A(円)").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ws.cell(row=r, column=5, value="月額B(円)").font = f_head
ws.cell(row=r, column=5).fill = F(ORANGE); ws.cell(row=r, column=5).alignment = cc; ws.cell(row=r, column=5).border = border
ws.cell(row=r, column=6, value="月額C(円)").font = f_head
ws.cell(row=r, column=6).fill = F(GREEN); ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
ws.cell(row=r, column=7, value="該当人数").font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.row_dimensions[r].height = 26
r += 1

stages13 = [
    ("第1段階", "生活保護受給者・老齢福祉年金受給者・住民税非課税世帯本人収入80万円以下", 0.285),
    ("第2段階", "住民税非課税世帯+本人収入80万円超120万円以下", 0.485),
    ("第3段階", "住民税非課税世帯+本人収入120万円超", 0.685),
    ("第4段階", "本人非課税(世帯課税)+課税年金収入80万円以下", 0.880),
    ("第5段階(基準)", "本人非課税(世帯課税)+課税年金収入80万円超", 1.000),
    ("第6段階", "本人課税・合計所得120万円未満", 1.200),
    ("第7段階", "本人課税・合計所得120万円以上210万円未満", 1.300),
    ("第8段階", "本人課税・合計所得210万円以上320万円未満", 1.500),
    ("第9段階", "本人課税・合計所得320万円以上420万円未満", 1.700),
    ("第10段階(新)", "本人課税・合計所得420万円以上520万円未満", 1.900),
    ("第11段階(新)", "本人課税・合計所得520万円以上620万円未満", 2.100),
    ("第12段階(新)", "本人課税・合計所得620万円以上720万円未満", 2.300),
    ("第13段階(新)", "本人課税・合計所得720万円以上", 2.400),
]

for stage, target, ratio in stages13:
    c = ws.cell(row=r, column=1, value=stage)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc
    fill_color = LORANGE if "(新)" in stage else LBLUE
    c.fill = F(fill_color); c.border = border
    c = ws.cell(row=r, column=2, value=target)
    c.font = f_body; c.alignment = cl; c.border = border
    c = ws.cell(row=r, column=3, value=ratio)
    c.font = f_known; c.alignment = cc; c.fill = F(KNOWN_G); c.border = border
    c.number_format = "0.000"
    for col, base_col in [(4, "C"), (5, "D"), (6, "E")]:
        c = ws.cell(row=r, column=col, value=f"=IFERROR('06_Step7-8_収納額_基準額'!{base_col}18*$C{r},\"未算定\")")
        c.font = f_calc; c.alignment = cr; c.fill = F(CALC_B); c.border = border
        c.number_format = "¥#,##0"
    c = ws.cell(row=r, column=7)
    c.fill = F(INPUT_Y); c.font = f_input; c.alignment = cr; c.border = border
    c.number_format = "#,##0"
    ws.row_dimensions[r].height = 32
    r += 1

# 注釈
r += 1
notes_s13 = [
    "(新)マークは9段階版にはなかった新設段階(第10〜13段階)です。高所得層への応分負担で全体最適化。",
    "係数は国推奨値ですが、市町村は条例で調整可能。委員会協議で確定値を決定します。",
    "13段階への移行は、低所得層(1〜3段階・約30%)の負担軽減と、高所得層への応分負担拡大の効果を狙います。",
    "該当人数のR7時点確定値は町担当課入力後に9段階版と13段階版で再分布を計算します。",
]
for note in notes_s13:
    ms(ws, f"A{r}:G{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

widths_s13 = [16, 36, 10, 14, 14, 14, 12]
for i, w in enumerate(widths_s13, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# ===========================================================
# 09_比較分析
# ===========================================================
ws = wb.create_sheet("09_比較分析")
ms(ws, "A1:F1", "09　3パターン×9/13段階 比較分析", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:F2", "── 委員会協議に向けた総合比較表・推奨パターンの検討材料 ──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

# 基準月額比較
r = 4
ms(ws, f"A{r}:F{r}", "■ 第10期基準月額の3パターン比較", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="パターン").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="月額").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="第9期比").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="基金活用").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ms(ws, f"E{r}:F{r}", "推奨度・特徴", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 24
r += 1

patterns = [
    ("A 取崩なし", "=IFERROR('06_Step7-8_収納額_基準額'!C18,\"未算定\")", "高", "なし", "基金温存・第11期負担緩和の安定性", LBLUE),
    ("B 50%取崩", "=IFERROR('06_Step7-8_収納額_基準額'!D18,\"未算定\")", "中", "50%", "中庸選択・住民負担と次期均衡(国推奨水準)", LORANGE),
    ("C 全額取崩", "=IFERROR('06_Step7-8_収納額_基準額'!E18,\"未算定\")", "低", "100%", "住民負担最小化・次期負担増リスク", LGREEN),
]
for ptn, formula, level, fund, feature, color in patterns:
    c = ws.cell(row=r, column=1, value=ptn)
    c.font = Font(name="游ゴシック", size=11, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(color); c.border = border
    c = ws.cell(row=r, column=2, value=formula)
    c.font = f_result; c.alignment = cr; c.fill = F(LRED); c.border = border
    c.number_format = "¥#,##0"
    c = ws.cell(row=r, column=3, value=level)
    c.font = f_body; c.alignment = cc; c.border = border
    c = ws.cell(row=r, column=4, value=fund)
    c.font = f_body; c.alignment = cc; c.border = border
    ms(ws, f"E{r}:F{r}", feature, f_body, None, cl)
    ws.row_dimensions[r].height = 30
    r += 1

# 仙南圏域・宮城県平均比較枠
r += 1
ms(ws, f"A{r}:F{r}", "■ 仙南圏域・宮城県との比較(見える化システム参考値)", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="自治体").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="第9期月額").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ws.cell(row=r, column=3, value="第10期月額").font = f_head
ws.cell(row=r, column=3).fill = F(NAVY); ws.cell(row=r, column=3).alignment = cc; ws.cell(row=r, column=3).border = border
ws.cell(row=r, column=4, value="高齢化率").font = f_head
ws.cell(row=r, column=4).fill = F(NAVY); ws.cell(row=r, column=4).alignment = cc; ws.cell(row=r, column=4).border = border
ms(ws, f"E{r}:F{r}", "備考", f_head, F(NAVY), cc)
ws.row_dimensions[r].height = 22
r += 1

others = [
    ("川崎町(本町)", 6500, "（試算）", 41.4, "本町", LORANGE),
    ("柴田町", "[入力]", "[入力]", "[入力]", "見える化システムから入力", LBLUE),
    ("大河原町", "[入力]", "[入力]", "[入力]", "見える化システムから入力", LBLUE),
    ("村田町", "[入力]", "[入力]", "[入力]", "見える化システムから入力", LBLUE),
    ("蔵王町", "[入力]", "[入力]", "[入力]", "見える化システムから入力", LBLUE),
    ("仙南圏域平均", "[入力]", "[入力]", "[入力]", "見える化システムから入力", LGRAY),
    ("宮城県平均", "[入力]", "[入力]", 28.5, "宮城県計画から入力", LGRAY),
    ("全国平均", 6225, "[入力]", 29.1, "全国平均(第9期)", LGREEN),
]
for area, p9, p10, age, note, color in others:
    c = ws.cell(row=r, column=1, value=area)
    font = Font(name="游ゴシック", size=10, bold=True, color=NAVY) if "(本町)" in area else f_body
    c.font = font; c.alignment = cl
    c.fill = F(color); c.border = border
    fill_p9 = KNOWN_G if isinstance(p9, (int, float)) else INPUT_Y
    c = ws.cell(row=r, column=2, value=p9)
    c.font = f_known if isinstance(p9, (int, float)) else f_input
    c.alignment = cr; c.fill = F(fill_p9); c.border = border
    if isinstance(p9, (int, float)): c.number_format = "¥#,##0"
    fill_p10 = INPUT_Y
    c = ws.cell(row=r, column=3, value=p10)
    c.font = f_input; c.alignment = cr; c.fill = F(fill_p10); c.border = border
    c = ws.cell(row=r, column=4, value=age)
    c.font = f_known if isinstance(age, (int, float)) else f_input
    c.alignment = cc; c.fill = F(fill_p9); c.border = border
    if isinstance(age, (int, float)): c.number_format = "0.0"
    ms(ws, f"E{r}:F{r}", note, f_note, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

# 委員会向けまとめ
r += 1
ms(ws, f"A{r}:F{r}", "■ 第3回策定委員会向け 委員協議のポイント", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

points = [
    "①住民負担vs持続可能性のバランス：基金温存(A)か、住民負担軽減(C)か、中庸(B)か。",
    "②国推奨水準との比較：国はパターンB(50%取崩)を標準的水準として推奨。",
    "③仙南圏域・宮城県平均との水準感：本町は宮城県平均より若干上の水準で推移する見込み。",
    "④第11期(R12〜)の予測：給付費上昇傾向の中で、第11期負担増のリスクを勘案。",
    "⑤所得段階区分(9→13段階)の影響：低所得層負担軽減と高所得層応分負担のバランス。",
    "⑥委員会では各委員にA/B/Cの選好を伺い、論点整理→次回再協議の流れを推奨。",
]
for pt in points:
    ms(ws, f"A{r}:F{r}", pt, f_body, F(LBLUE), cl)
    ws.row_dimensions[r].height = 24
    r += 1

widths_s9 = [22, 14, 14, 14, 18, 18]
for i, w in enumerate(widths_s9, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
setup_page(ws)

# 保存
out = "/home/claude/kawasaki_work/川崎町_保険料試算ワークブック.xlsx"
wb.save(out)
print(f"作成完了: {out}")
print(f"シート数: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws_s = wb[s]
    print(f"  - {s}: {ws_s.max_row}行 x {ws_s.max_column}列")
