"""
川崎町第10期介護保険事業計画策定 作業チェックリスト v2.0
（最新進捗反映版・5フェーズ管理）

更新ポイント：
- F1〜F6改訂版・実績データ確認サマリー・第9期実績一覧・アンケート集計テンプレ・計画素案v1.5の完了ステータス反映
- 第2フェーズ以降（アンケート回収後Ver.2.0・第2-4回策定委員会対応・パブコメ）の作業項目追加
- フェーズ別の進捗ダッシュボードを新設

6シート構成：
00_使い方            : 凡例・更新履歴・優先度の定義
01_進捗ダッシュボード : フェーズ別 完了率・残作業数・直近マイルストーン
02_Phase1_素案準備   : キックオフ〜第1回委員会前（現在ここ）
03_Phase2_Ver2.0     : アンケート回収〜第2回委員会
04_Phase3_保険料試算 : 第3回委員会〜計画確定
05_Phase4_パブコメ確定 : パブリックコメント〜3月議会上程
06_リスク論点管理    : 川崎町固有論点・対応状況
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# カラーパレット（計画書・既存成果物と統一）
NAVY = "1F3864"; BLUE = "2F5597"; LBLUE = "DAE3F3"
ORANGE = "ED7D31"; LORANGE = "FCE4D6"
GREEN = "548235"; LGREEN = "E2EFDA"
RED = "C00000"; LRED = "FFE4E4"
GRAY = "808080"; LGRAY = "F2F2F2"
INPUT_YELLOW = "FFFFCC"; WHITE = "FFFFFF"

# ステータスカラー
COMPLETED = "C6EFCE"  # 完了 緑
IN_PROGRESS = "FFEB9C"  # 進行中 黄
NOT_STARTED = "F2F2F2"  # 未着手 灰
ON_HOLD = "FFC7CE"  # 保留 赤

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

def F(c): return PatternFill("solid", fgColor=c)
cc = Alignment(horizontal="center", vertical="center", wrap_text=True)
cl = Alignment(horizontal="left", vertical="center", wrap_text=True)
cr = Alignment(horizontal="right", vertical="center", wrap_text=True)

f_title = Font(name="游ゴシック", size=13, bold=True, color=WHITE)
f_sub   = Font(name="游ゴシック", size=10, italic=True, color=WHITE)
f_head  = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
f_section = Font(name="游ゴシック", size=11, bold=True, color=NAVY)
f_body  = Font(name="游ゴシック", size=10)
f_done  = Font(name="游ゴシック", size=10, bold=True, color="006100")  # 完了濃緑
f_prog  = Font(name="游ゴシック", size=10, bold=True, color="9C5700")  # 進行中濃黄
f_pend  = Font(name="游ゴシック", size=10, bold=True, color="9C0006")  # 保留濃赤
f_input = Font(name="游ゴシック", size=10, color="0000FF")
f_note  = Font(name="游ゴシック", size=9, italic=True, color="595959")

wb = Workbook()
wb.remove(wb.active)

def ms(ws, rng, val, font, fill, align):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = val; c.font = font
    if fill is not None: c.fill = fill
    c.alignment = align
    from openpyxl.utils.cell import range_boundaries
    a, b, d, e = range_boundaries(rng)
    for r in range(b, e+1):
        for col in range(a, d+1):
            ws.cell(row=r, column=col).border = border

def setup_page(ws, orient="landscape"):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orient == "landscape" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)

def status_cell(ws, row, col, status):
    """ステータスセル：完了/進行中/未着手/保留"""
    c = ws.cell(row=row, column=col)
    c.border = border
    c.alignment = cc
    if status == "完了":
        c.value = "✓ 完了"
        c.font = f_done
        c.fill = F(COMPLETED)
    elif status == "進行中":
        c.value = "進行中"
        c.font = f_prog
        c.fill = F(IN_PROGRESS)
    elif status == "未着手":
        c.value = "未着手"
        c.font = f_body
        c.fill = F(NOT_STARTED)
    elif status == "保留":
        c.value = "保留"
        c.font = f_pend
        c.fill = F(ON_HOLD)
    else:
        c.value = status
        c.font = f_body

# ===========================================================
# 00_使い方
# ===========================================================
ws = wb.create_sheet("00_使い方")
ms(ws, "A1:G1", "川崎町第10期介護保険事業計画策定 作業チェックリスト v2.0",
   f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 28
ms(ws, "A2:G2", "── 最新進捗反映版（令和8年6月作成） 5フェーズ管理 ──",
   f_sub, F(BLUE), cc)
ws.row_dimensions[2].height = 20

# 1. 本シートの目的
ms(ws, "A4:G4", "1．本チェックリストの目的", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22
ms(ws, "A5:G5",
   "本チェックリストは、川崎町第10期介護保険事業計画の策定業務全体を5フェーズに分け、各フェーズの作業項目・担当・進捗・関連成果物を一元管理するためのプロジェクト管理シートです。受託者（弊社）と町担当課（保健福祉課）の双方で共有し、毎月の進捗確認・打合せの基礎資料として活用します。",
   f_body, None, cl)
ws.row_dimensions[5].height = 60

# 2. シート構成
ms(ws, "A7:G7", "2．シート構成と現在地", f_head, F(BLUE), cl)
ws.row_dimensions[7].height = 22

# ヘッダ
ws.cell(row=8, column=1, value="シート").font = f_head
ws.cell(row=8, column=1).fill = F(NAVY); ws.cell(row=8, column=1).alignment = cc; ws.cell(row=8, column=1).border = border
ms(ws, "B8:C8", "対象期間", f_head, F(NAVY), cc)
ms(ws, "D8:E8", "主な作業", f_head, F(NAVY), cc)
ws.cell(row=8, column=6, value="現在").font = f_head
ws.cell(row=8, column=6).fill = F(NAVY); ws.cell(row=8, column=6).alignment = cc; ws.cell(row=8, column=6).border = border
ws.cell(row=8, column=7, value="完了率").font = f_head
ws.cell(row=8, column=7).fill = F(NAVY); ws.cell(row=8, column=7).alignment = cc; ws.cell(row=8, column=7).border = border
ws.row_dimensions[8].height = 22

phase_overview = [
    ("01_進捗ダッシュボード", "全期間", "5フェーズの進捗サマリー", "─", "─"),
    ("02_Phase1_素案準備", "R7.11〜R8.8中旬", "キックオフ・F1-F6・素案v1.5・記入パッケージ整備", "進行中", "85%"),
    ("03_Phase2_Ver2.0", "R8.8中旬〜R8.11", "第1回委員会・アンケート集計・Ver.2.0更新", "未着手", "0%"),
    ("04_Phase3_保険料試算", "R8.11〜R9.2", "見込量精緻化・3パターン試算・第3-4回委員会", "未着手", "0%"),
    ("05_Phase4_パブコメ確定", "R9.2〜R9.3", "パブコメ・3月議会上程・公表", "未着手", "0%"),
    ("06_リスク論点管理", "全期間", "川崎町固有論点（移動支援3層構造・認知症基本法等）", "進行中", "─"),
]
r = 9
for sh, term, work, now, rate in phase_overview:
    c = ws.cell(row=r, column=1, value=sh)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl
    c.fill = F(LBLUE); c.border = border
    ms(ws, f"B{r}:C{r}", term, f_body, None, cl)
    ms(ws, f"D{r}:E{r}", work, f_body, None, cl)
    if now == "進行中":
        ws.cell(row=r, column=6, value=now).font = f_prog
        ws.cell(row=r, column=6).fill = F(IN_PROGRESS)
    elif now == "完了":
        ws.cell(row=r, column=6, value=now).font = f_done
        ws.cell(row=r, column=6).fill = F(COMPLETED)
    elif now == "未着手":
        ws.cell(row=r, column=6, value=now).font = f_body
        ws.cell(row=r, column=6).fill = F(NOT_STARTED)
    else:
        ws.cell(row=r, column=6, value=now).font = f_body
    ws.cell(row=r, column=6).alignment = cc; ws.cell(row=r, column=6).border = border
    ws.cell(row=r, column=7, value=rate).font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
    ws.row_dimensions[r].height = 26
    r += 1

# 3. 凡例
r += 1
ms(ws, f"A{r}:G{r}", "3．ステータス凡例", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

legends = [
    (COMPLETED, "✓ 完了", "作業が完了し、成果物として確定済み"),
    (IN_PROGRESS, "進行中", "作業に着手し、現在進めている状態"),
    (NOT_STARTED, "未着手", "まだ着手していない（前段の作業完了を待っている）状態"),
    (ON_HOLD, "保留", "町確認待ち・委員会協議待ち等で一時保留中"),
]
for color, status, mean in legends:
    c = ws.cell(row=r, column=1, value=status)
    c.fill = F(color)
    c.font = Font(name="游ゴシック", size=10, bold=True)
    c.alignment = cc; c.border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 4. 優先度
r += 1
ms(ws, f"A{r}:G{r}", "4．優先度の定義", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

priorities = [
    ("A", LORANGE, "計画素案の本文反映に必須・委員会報告に直結する作業"),
    ("B", LBLUE, "施策の論拠補強・KPI設定・別添資料整備等の重要作業"),
    ("C", LGRAY, "周辺整備・将来運用準備等の任意作業"),
]
for pri, color, mean in priorities:
    c = ws.cell(row=r, column=1, value=pri)
    c.fill = F(color)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=RED if pri == "A" else (BLUE if pri == "B" else GRAY))
    c.alignment = cc; c.border = border
    ms(ws, f"B{r}:G{r}", mean, f_body, None, cl)
    ws.row_dimensions[r].height = 22
    r += 1

# 5. 更新履歴
r += 1
ms(ws, f"A{r}:G{r}", "5．更新履歴", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

ws.cell(row=r, column=1, value="版").font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="日付").font = f_head
ws.cell(row=r, column=2).fill = F(NAVY); ws.cell(row=r, column=2).alignment = cc; ws.cell(row=r, column=2).border = border
ms(ws, f"C{r}:G{r}", "主な更新内容", f_head, F(NAVY), cl)
ws.row_dimensions[r].height = 22
r += 1

history = [
    ("v1.0", "R7.11", "キックオフ後 初版作成（4フェーズ管理）"),
    ("v1.5", "R8.5", "F1-F6初版完成、実績データ確認サマリー反映"),
    ("v2.0", "R8.6", "本版・F1-F6改訂版・素案v1.5・記入パッケージ整備の完了反映、5フェーズに再編成"),
]
for v, d, c in history:
    cell = ws.cell(row=r, column=1, value=v)
    cell.font = Font(name="游ゴシック", size=10, bold=True); cell.alignment = cc
    cell.fill = F(LBLUE); cell.border = border
    cell = ws.cell(row=r, column=2, value=d)
    cell.font = f_body; cell.alignment = cc; cell.border = border
    ms(ws, f"C{r}:G{r}", c, f_body, None, cl)
    ws.row_dimensions[r].height = 24
    r += 1

# 列幅
widths = [22, 12, 12, 14, 14, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

setup_page(ws)

# ===========================================================
# 01_進捗ダッシュボード
# ===========================================================
ws = wb.create_sheet("01_進捗ダッシュボード")
ms(ws, "A1:H1", "01　全体進捗ダッシュボード", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:H2", "──5フェーズの進捗・直近マイルストーン・残作業数──", f_sub, F(BLUE), cc)
ws.row_dimensions[2].height = 20

# 全体KPI
ms(ws, "A4:H4", "■ プロジェクト全体サマリー（令和8年6月時点）", f_head, F(BLUE), cl)
ws.row_dimensions[4].height = 22

# KPIカード風
kpi_row = 5
kpis = [
    ("総作業項目", "78", "件"),
    ("完了", "27", "件 (35%)"),
    ("進行中", "8", "件 (10%)"),
    ("未着手", "43", "件 (55%)"),
]
for i, (k, v, u) in enumerate(kpis):
    col = 1 + i * 2
    cell_k = ws.cell(row=kpi_row, column=col, value=k)
    cell_k.font = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
    cell_k.alignment = cc
    cell_k.fill = F(NAVY)
    cell_k.border = border
    ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col+1)
    
    cell_v = ws.cell(row=kpi_row+1, column=col, value=v)
    cell_v.font = Font(name="游ゴシック", size=20, bold=True, color=NAVY)
    cell_v.alignment = cc
    cell_v.border = border
    
    cell_u = ws.cell(row=kpi_row+1, column=col+1, value=u)
    cell_u.font = f_body
    cell_u.alignment = cl
    cell_u.border = border

ws.row_dimensions[kpi_row].height = 24
ws.row_dimensions[kpi_row+1].height = 32

# 直近マイルストーン
r = 9
ms(ws, f"A{r}:H{r}", "■ 直近マイルストーン", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

headers = ["時期", "マイルストーン", "対応必要事項", "担当", "ステータス"]
ws.cell(row=r, column=1, value=headers[0]).font = f_head
ws.cell(row=r, column=1).fill = F(NAVY); ws.cell(row=r, column=1).alignment = cc; ws.cell(row=r, column=1).border = border
ms(ws, f"B{r}:C{r}", headers[1], f_head, F(NAVY), cc)
ms(ws, f"D{r}:F{r}", headers[2], f_head, F(NAVY), cc)
ws.cell(row=r, column=7, value=headers[3]).font = f_head
ws.cell(row=r, column=7).fill = F(NAVY); ws.cell(row=r, column=7).alignment = cc; ws.cell(row=r, column=7).border = border
ws.cell(row=r, column=8, value=headers[4]).font = f_head
ws.cell(row=r, column=8).fill = F(NAVY); ws.cell(row=r, column=8).alignment = cc; ws.cell(row=r, column=8).border = border
ws.row_dimensions[r].height = 22
r += 1

milestones = [
    ("R8.6下旬", "町送付パッケージ発送", "町担当への記入依頼パッケージ送付・大宮様への説明", "弊社", "完了"),
    ("R8.6下旬", "アンケート発送", "一般高齢者1,000名・認定者300名の発送", "町担当", "未着手"),
    ("R8.6確定", "基金残高確定", "介護給付費準備基金残高(R8.6時点)の確定", "町担当", "保留"),
    ("R8.7末", "アンケート回収", "回収率管理・督促対応", "町担当", "未着手"),
    ("R8.7末", "町記入データ受領", "①第9期実績一覧・③MECEデータの記入完了", "町担当", "未着手"),
    ("R8.8上旬", "アンケート集計", "弊社で集計・分析実施・本テンプレに反映", "弊社", "未着手"),
    ("R8.8中旬", "第1回策定委員会", "アンケート結果報告・基本方針協議", "委員会", "未着手"),
    ("R8.8末", "計画素案Ver.2.0完成", "アンケート結果・実数値反映の更新版", "弊社", "未着手"),
    ("R8.11", "第2回策定委員会", "Ver.2.0審議・サービス見込量方向性確定", "委員会", "未着手"),
    ("R9.1中旬", "第3回策定委員会", "保険料試算3パターン協議", "委員会", "未着手"),
    ("R9.2", "第4回策定委員会", "保険料基準額決定・町長答申", "委員会", "未着手"),
    ("R9.3", "3月議会上程", "条例改正案上程・可決後公表", "町担当", "未着手"),
]

for term, ms_name, todo, owner, status in milestones:
    ws.cell(row=r, column=1, value=term).font = Font(name="游ゴシック", size=10, bold=True)
    ws.cell(row=r, column=1).alignment = cc
    ws.cell(row=r, column=1).fill = F(LBLUE)
    ws.cell(row=r, column=1).border = border
    ms(ws, f"B{r}:C{r}", ms_name, Font(name="游ゴシック", size=10, bold=True, color=NAVY), None, cl)
    ms(ws, f"D{r}:F{r}", todo, f_body, None, cl)
    
    owner_fill = F(LORANGE) if owner == "弊社" else (F(LBLUE) if owner == "町担当" else F(LGREEN))
    ws.cell(row=r, column=7, value=owner).font = f_body
    ws.cell(row=r, column=7).alignment = cc
    ws.cell(row=r, column=7).fill = owner_fill
    ws.cell(row=r, column=7).border = border
    
    status_cell(ws, r, 8, status)
    ws.row_dimensions[r].height = 26
    r += 1

# フェーズ別進捗
r += 1
ms(ws, f"A{r}:H{r}", "■ フェーズ別進捗", f_head, F(BLUE), cl)
ws.row_dimensions[r].height = 22
r += 1

headers2 = ["フェーズ", "期間", "主要マイルストーン", "完了数", "進行中", "未着手", "完了率", "備考"]
widths2 = [16, 16, 22, 8, 8, 8, 8, 14]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = f_head; c.fill = F(NAVY); c.alignment = cc; c.border = border
    if i in [1, 2, 3, 8]:
        ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[r].height = 22
r += 1

phases = [
    ("Phase 1", "R7.11〜R8.8中旬", "素案v1.5・記入パッケージ", "23", "5", "0", "85%", "進行中"),
    ("Phase 2", "R8.8中旬〜R8.11", "Ver.2.0更新・第1-2回委員会", "0", "0", "18", "0%", "未着手"),
    ("Phase 3", "R8.11〜R9.2", "保険料試算・第3-4回委員会", "0", "0", "15", "0%", "未着手"),
    ("Phase 4", "R9.2〜R9.3", "パブコメ・3月議会", "0", "0", "12", "0%", "未着手"),
    ("論点管理", "全期間", "川崎町固有6論点", "4", "3", "0", "57%", "進行中"),
]

for ph, term, ms_name, done, prog, todo, rate, note in phases:
    c = ws.cell(row=r, column=1, value=ph)
    c.font = Font(name="游ゴシック", size=11, bold=True, color=NAVY); c.alignment = cc
    c.fill = F(LBLUE); c.border = border
    c = ws.cell(row=r, column=2, value=term)
    c.font = f_body; c.alignment = cc; c.border = border
    c = ws.cell(row=r, column=3, value=ms_name)
    c.font = f_body; c.alignment = cl; c.border = border
    c = ws.cell(row=r, column=4, value=done)
    c.font = Font(name="游ゴシック", size=10, bold=True, color="006100")
    c.alignment = cc; c.fill = F(COMPLETED); c.border = border
    c = ws.cell(row=r, column=5, value=prog)
    c.font = Font(name="游ゴシック", size=10, bold=True, color="9C5700")
    c.alignment = cc; c.fill = F(IN_PROGRESS); c.border = border
    c = ws.cell(row=r, column=6, value=todo)
    c.font = f_body; c.alignment = cc; c.fill = F(NOT_STARTED); c.border = border
    c = ws.cell(row=r, column=7, value=rate)
    c.font = Font(name="游ゴシック", size=11, bold=True, color=NAVY); c.alignment = cc; c.border = border
    c = ws.cell(row=r, column=8, value=note)
    if note == "進行中":
        c.font = f_prog; c.fill = F(IN_PROGRESS)
    else:
        c.font = f_body; c.fill = F(NOT_STARTED)
    c.alignment = cc; c.border = border
    ws.row_dimensions[r].height = 28
    r += 1

setup_page(ws)

# ===========================================================
# 共通ヘルパー：作業表の生成
# ===========================================================
def make_phase_sheet(name, title, subtitle, tasks):
    """フェーズシート生成
    tasks: [(セクション名/タスク), ...] 
           タスクは (No, 作業, 担当, 関連成果物, 期日, 優先度, ステータス, 備考)
    """
    ws = wb.create_sheet(name)
    ms(ws, "A1:H1", title, f_title, F(NAVY), cc)
    ws.row_dimensions[1].height = 26
    ms(ws, "A2:H2", subtitle, f_sub, F(BLUE), cl)
    ws.row_dimensions[2].height = 20
    
    headers = ["No", "作業項目", "担当", "関連成果物", "期日", "優先度", "ステータス", "備考"]
    widths = [6, 30, 10, 22, 12, 8, 12, 22]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30
    
    r = 4
    for task in tasks:
        if isinstance(task, str):
            ms(ws, f"A{r}:H{r}", task, f_section, F(LBLUE), cl)
            ws.row_dimensions[r].height = 22
            r += 1
            continue
        # No
        c = ws.cell(row=r, column=1, value=task[0])
        c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cc; c.border = border
        # 作業
        c = ws.cell(row=r, column=2, value=task[1])
        c.font = f_body; c.alignment = cl; c.border = border
        # 担当
        owner = task[2]
        owner_fill = F(LORANGE) if owner == "弊社" else (F(LBLUE) if owner == "町担当" else (F(LGREEN) if owner == "委員会" else F(LGRAY)))
        c = ws.cell(row=r, column=3, value=owner)
        c.font = f_body; c.alignment = cc; c.fill = owner_fill; c.border = border
        # 関連成果物
        c = ws.cell(row=r, column=4, value=task[3])
        c.font = f_body; c.alignment = cl; c.border = border
        # 期日
        c = ws.cell(row=r, column=5, value=task[4])
        c.font = f_body; c.alignment = cc; c.border = border
        # 優先度
        pri = task[5]
        pri_fill = F(LORANGE) if pri == "A" else (F(LBLUE) if pri == "B" else F(LGRAY))
        pri_color = RED if pri == "A" else (BLUE if pri == "B" else GRAY)
        c = ws.cell(row=r, column=6, value=pri)
        c.font = Font(name="游ゴシック", size=10, bold=True, color=pri_color); c.alignment = cc; c.fill = pri_fill; c.border = border
        # ステータス
        status_cell(ws, r, 7, task[6])
        # 備考
        c = ws.cell(row=r, column=8, value=task[7])
        c.font = f_note; c.alignment = cl; c.border = border
        ws.row_dimensions[r].height = 26
        r += 1
    
    setup_page(ws)
    return ws

# ===========================================================
# 02_Phase1_素案準備
# ===========================================================
make_phase_sheet(
    "02_Phase1_素案準備",
    "02　Phase 1：素案準備（R7.11〜R8.8中旬）",
    "■キックオフ〜第1回策定委員会前まで。完了済みタスクと残作業を整理",
    [
        "【1-1．受託直後の整備（R7.11〜R8.3）】",
        ("1-1-1", "キックオフ会議の開催・議事録作成", "弊社", "R8_川崎町_キックオフ議事録_正式版", "R7.11", "A", "完了", "町長挨拶含む"),
        ("1-1-2", "F1：施策実現可能性評価", "弊社", "川崎町_F1_施策実現可能性評価", "R8.1", "B", "完了", "別添資料"),
        ("1-1-3", "F2：アンケート追加設問の検討", "弊社", "川崎町_F2_アンケート追加設問", "R8.2", "A", "完了", "5問追加"),
        ("1-1-4", "F3：アンケート実施支援", "弊社", "川崎町_F3_アンケート実施支援", "R8.3", "A", "完了", "発送計画"),
        ("1-1-5", "F4：策定スケジュール案", "弊社", "川崎町_F4_アンケート実施支援", "R8.3", "B", "完了", "全体計画"),
        ("1-1-6", "F5：3業務深掘り解説書 改訂版", "弊社", "川崎町_F5_3業務深掘り解説書_改訂版 Ver.2.0", "R8.4", "B", "完了", "14頁"),
        ("1-1-7", "F6：策定委員会書類ひな形 改訂版", "弊社", "川崎町_F6_策定委員会書類ひな形_改訂版 Ver.2.0", "R8.4", "B", "完了", "13頁"),

        "【1-2．データ収集（R8.4〜R8.6）】",
        ("1-2-1", "町提供資料のレビュー", "弊社", "実績データ確認サマリー(6シート)", "R8.4", "A", "完了", "R3給付費等を整理"),
        ("1-2-2", "MECEデータ入力フォーマット 整備", "弊社", "川崎町_必要資料_入力フォーマット", "R8.4", "A", "完了", "13シート"),
        ("1-2-3", "町確認データの入力(MECE版)", "町担当", "川崎町_必要資料_入力フォーマット", "R8.5", "A", "進行中", "一部受領済"),
        ("1-2-4", "第9期実績一覧 整備", "弊社", "川崎町_第9期実績一覧_町記入用", "R8.5", "A", "完了", "7シート35事業"),
        ("1-2-5", "アンケート集計分析テンプレ整備", "弊社", "川崎町_アンケート集計分析テンプレート", "R8.5", "A", "完了", "6シート"),

        "【1-3．素案v1.0〜v1.5の段階的開発（R8.5〜R8.6）】",
        ("1-3-1", "素案v1.0 基本版作成", "弊社", "川崎町_計画書素案_v1.0", "R8.5", "A", "完了", "27頁・8章"),
        ("1-3-2", "素案v1.1 図表詳細化(6図表埋込)", "弊社", "川崎町_計画書素案_v1.1_図表詳細化版", "R8.5", "A", "完了", "33頁"),
        ("1-3-3", "素案v1.2 町長挨拶・目次完備", "弊社", "川崎町_計画書素案_v1.2_町長挨拶目次完備版", "R8.5", "A", "完了", "34頁"),
        ("1-3-4", "素案v1.3 第7章詳細化(8ステップ・3パターン)", "弊社", "川崎町_計画書素案_v1.3_第7章詳細化版", "R8.6", "A", "完了", "40頁"),
        ("1-3-5", "素案v1.4 第3章3-2取組実績6目標別詳細化", "弊社", "川崎町_計画書素案_v1.4_第3章詳細化版", "R8.6", "A", "完了", "43頁"),
        ("1-3-6", "素案v1.5 章カラー統一・拡張性確保", "弊社", "川崎町_計画書素案_v1.5_カラー統一版", "R8.6", "A", "完了", "43頁・最新"),

        "【1-4．町送付パッケージの整備（R8.6）】",
        ("1-4-1", "町担当ご記入依頼パッケージ作成", "弊社", "川崎町_町担当課ご記入依頼パッケージ", "R8.6", "A", "完了", "5頁・送付前"),
        ("1-4-2", "作業チェックリスト v2.0 更新", "弊社", "本ファイル", "R8.6", "B", "完了", "5フェーズ管理"),
        ("1-4-3", "町担当への説明資料作成", "弊社", "川崎町_町担当説明資料", "R8.6", "A", "進行中", "対面打合せ用"),
        ("1-4-4", "送付物一式の最終確認・発送", "弊社", "送付一式", "R8.6下旬", "A", "進行中", "次工程"),

        "【1-5．町担当による記入対応（R8.6〜R8.7）】",
        ("1-5-1", "①第9期実績一覧 記入(01_KPI〜04_認知症)", "町担当", "川崎町_第9期実績一覧_町記入用", "R8.6下旬", "A", "未着手", "優先記入"),
        ("1-5-2", "①第9期実績一覧 記入(残シート)", "町担当", "川崎町_第9期実績一覧_町記入用", "R8.7上旬", "A", "未着手", "続き"),
        ("1-5-3", "③MECEデータ 残空欄の記入", "町担当", "川崎町_必要資料_入力フォーマット", "R8.6下旬", "A", "未着手", "残追記"),
        ("1-5-4", "アンケート発送", "町担当", "─", "R8.6下旬", "A", "未着手", "1,300名"),
        ("1-5-5", "アンケート回収・督促", "町担当", "─", "R8.7末", "A", "未着手", "回収率管理"),
        ("1-5-6", "介護給付費準備基金残高 確定", "町担当", "─", "R8.6", "A", "保留", "R8.6時点値"),
    ]
)

# ===========================================================
# 03_Phase2_Ver2.0
# ===========================================================
make_phase_sheet(
    "03_Phase2_Ver2.0",
    "03　Phase 2：Ver.2.0更新と第1-2回委員会（R8.8中旬〜R8.11）",
    "■アンケート集計・町データ反映・Ver.2.0完成・第1-2回策定委員会開催",
    [
        "【2-1．アンケート集計・分析（R8.8上旬）】",
        ("2-1-1", "アンケート回収データ受領・整理", "弊社", "回収アンケート", "R8.8上旬", "A", "未着手", "町から受領"),
        ("2-1-2", "単純集計の実施(一般高齢者)", "弊社", "01_一般高齢者_単純集計", "R8.8上旬", "A", "未着手", "19問×N"),
        ("2-1-3", "単純集計の実施(認定者)", "弊社", "02_認定者_単純集計", "R8.8上旬", "A", "未着手", "16問×N"),
        ("2-1-4", "クロス集計の実施(5本)", "弊社", "03_クロス集計", "R8.8上旬", "A", "未着手", "地区×移動等"),
        ("2-1-5", "自由記述5分類集計", "弊社", "04_自由記述_5分類", "R8.8上旬", "B", "未着手", "20+α分類"),
        ("2-1-6", "計画反映ガイド更新(32項目)", "弊社", "05_計画反映ガイド", "R8.8上旬", "A", "未着手", "Ver.2.0反映先確定"),
        
        "【2-2．町データの反映（R8.8上旬）】",
        ("2-2-1", "第9期実績一覧 → 計画素案3-2反映", "弊社", "計画素案Ver.2.0 第3章", "R8.8上旬", "A", "未着手", "35事業数値反映"),
        ("2-2-2", "MECEデータ → 計画素案2章・7章反映", "弊社", "計画素案Ver.2.0 第2/7章", "R8.8上旬", "A", "未着手", "数値確定"),
        ("2-2-3", "基金残高 → 第7章7-3に反映", "弊社", "計画素案Ver.2.0 第7章7-3", "R8.8上旬", "A", "未着手", "保険料試算前提"),
        ("2-2-4", "アンケート結果 → 各章プレースホルダー埋め", "弊社", "計画素案Ver.2.0 全章", "R8.8上旬", "A", "未着手", "32箇所"),
        ("2-2-5", "目次ページ番号 再確定", "弊社", "計画素案Ver.2.0 目次", "R8.8中旬", "A", "未着手", "ページ再計算"),
        
        "【2-3．第1回策定委員会（R8.8中旬）】",
        ("2-3-1", "第1回委員会 開催準備(資料印刷・会場設営)", "町担当", "委員会資料一式", "R8.8中旬", "A", "未着手", "次第・名簿等"),
        ("2-3-2", "第1回委員会 資料作成", "弊社", "委員会資料(計画素案Ver.2.0素案+アンケート結果)", "R8.8中旬", "A", "未着手", "F6書類ひな形を発展"),
        ("2-3-3", "第1回委員会 開催", "委員会", "議事録", "R8.8中旬", "A", "未着手", "アンケート結果報告"),
        ("2-3-4", "第1回委員会 議事録・意見集約", "弊社", "第1回委員会議事録", "R8.8末", "B", "未着手", "Ver.2.0に反映"),
        
        "【2-4．Ver.2.0確定〜第2回委員会（R8.9〜R8.11）】",
        ("2-4-1", "委員会意見を踏まえたVer.2.0修正", "弊社", "計画素案Ver.2.0", "R8.9〜10", "A", "未着手", "意見反映"),
        ("2-4-2", "サービス見込量の精緻化(6ステップ)", "弊社", "計画素案Ver.2.0 第7章7-1", "R8.9〜10", "A", "未着手", "町外利用含む"),
        ("2-4-3", "国通知の反映(R8夏以降)", "弊社", "計画素案Ver.2.0 各章", "R8.10〜11", "A", "未着手", "適宜反映"),
        ("2-4-4", "第2回委員会 資料作成", "弊社", "委員会資料", "R8.11", "A", "未着手", "Ver.2.0素案"),
        ("2-4-5", "第2回委員会 開催", "委員会", "議事録", "R8.11", "A", "未着手", "見込量方向性確定"),
    ]
)

# ===========================================================
# 04_Phase3_保険料試算
# ===========================================================
make_phase_sheet(
    "04_Phase3_保険料試算",
    "04　Phase 3：保険料試算と第3-4回委員会（R8.11〜R9.2）",
    "■保険料3パターン試算・13段階区分検討・町長答申まで",
    [
        "【3-1．保険料試算の精緻化（R8.12〜R9.1上旬）】",
        ("3-1-1", "標準給付費の確定", "弊社", "計画素案Ver.2.1 第7章7-2", "R8.12", "A", "未着手", "8ステップStep1"),
        ("3-1-2", "地域支援事業費の確定", "弊社", "計画素案Ver.2.1 第7章7-2", "R8.12", "A", "未着手", "Step2"),
        ("3-1-3", "第1号負担分相当額の算定", "弊社", "計画素案Ver.2.1 第7章7-3", "R8.12", "A", "未着手", "Step3(23%)"),
        ("3-1-4", "調整交付金の算定", "弊社", "計画素案Ver.2.1 第7章7-3", "R8.12", "A", "未着手", "Step4"),
        ("3-1-5", "基金取崩額の3パターン設定", "弊社", "計画素案Ver.2.1 第7章7-3", "R9.1上旬", "A", "未着手", "Step6 A/B/C"),
        ("3-1-6", "予定収納率の確定", "弊社", "計画素案Ver.2.1 第7章7-3", "R9.1上旬", "A", "未着手", "過去5年平均"),
        ("3-1-7", "13段階区分の保険料率設定", "弊社", "計画素案Ver.2.1 第7章7-3", "R9.1上旬", "A", "未着手", "9→13段階"),
        ("3-1-8", "近隣・宮城県・全国比較値の収集", "弊社", "計画素案Ver.2.1 第7章7-3", "R9.1上旬", "B", "未着手", "見える化システム"),
        
        "【3-2．第3回策定委員会（R9.1中旬）】",
        ("3-2-1", "第3回委員会 資料作成", "弊社", "委員会資料(保険料3パターン試算)", "R9.1中旬", "A", "未着手", "保険料の協議"),
        ("3-2-2", "第3回委員会 開催", "委員会", "議事録", "R9.1中旬", "A", "未着手", "保険料3パターン協議"),
        ("3-2-3", "委員会意見を踏まえたVer.2.2修正", "弊社", "計画素案Ver.2.2", "R9.1下旬", "A", "未着手", "選定パターン反映"),
        
        "【3-3．第4回策定委員会・町長答申（R9.2）】",
        ("3-3-1", "第4回委員会 資料作成(最終素案)", "弊社", "計画書最終素案", "R9.2", "A", "未着手", "保険料基準額確定"),
        ("3-3-2", "第4回委員会 開催", "委員会", "議事録・答申案", "R9.2", "A", "未着手", "保険料基準額決定"),
        ("3-3-3", "策定委員会から町長への答申", "委員会", "答申書", "R9.2", "A", "未着手", "町長宛て"),
        ("3-3-4", "計画書本編 完成版作成", "弊社", "計画書本編 最終確定版", "R9.2", "A", "未着手", "答申反映"),
    ]
)

# ===========================================================
# 05_Phase4_パブコメ確定
# ===========================================================
make_phase_sheet(
    "05_Phase4_パブコメ確定",
    "05　Phase 4：パブリックコメント・3月議会上程・公表（R9.2〜R9.3）",
    "■パブコメ・住民周知資料・条例改正案・公表まで",
    [
        "【4-1．パブリックコメント（R9.2）】",
        ("4-1-1", "パブコメ用 計画書要約版作成", "弊社", "計画書要約版(住民周知用)", "R9.2", "A", "未着手", "A3/A4 4頁"),
        ("4-1-2", "パブコメ実施要領作成", "弊社", "パブコメ実施要領", "R9.2", "B", "未着手", "意見受付方法等"),
        ("4-1-3", "パブコメ募集開始(町HP・広報誌)", "町担当", "パブコメ告知", "R9.2", "A", "未着手", "30日間程度"),
        ("4-1-4", "パブコメ意見の集約・整理", "弊社/町", "パブコメ意見集", "R9.2末", "A", "未着手", "回答案作成"),
        ("4-1-5", "パブコメ意見の計画への反映検討", "弊社", "計画書最終版", "R9.2末", "A", "未着手", "町と協議"),
        ("4-1-6", "パブコメ実施結果の公表", "町担当", "パブコメ結果公表", "R9.3", "B", "未着手", "町HP掲載"),
        
        "【4-2．条例改正・議会上程（R9.3）】",
        ("4-2-1", "介護保険条例改正案 作成支援", "弊社", "条例改正案", "R9.3初", "A", "未着手", "保険料率・段階"),
        ("4-2-2", "3月議会 上程", "町担当", "議会上程資料", "R9.3", "A", "未着手", "町担当主導"),
        ("4-2-3", "議会質疑応答対応支援", "弊社", "想定問答集", "R9.3", "B", "未着手", "町担当補助"),
        ("4-2-4", "条例可決・公布", "町担当", "─", "R9.3末", "A", "未着手", "施行R9.4"),
        
        "【4-3．計画書 公表・配布（R9.3〜R9.4）】",
        ("4-3-1", "計画書 本編 印刷データ作成", "弊社", "計画書本編 印刷用PDF", "R9.3", "A", "未着手", "町指定部数"),
        ("4-3-2", "計画書 概要版 印刷データ作成", "弊社", "計画書概要版 印刷用PDF", "R9.3", "B", "未着手", "住民配布用"),
        ("4-3-3", "計画書 町HP掲載", "町担当", "計画書PDF(HP公開版)", "R9.3〜4", "A", "未着手", "HP掲載"),
    ]
)

# ===========================================================
# 06_リスク論点管理
# ===========================================================
ws = wb.create_sheet("06_リスク論点管理")
ms(ws, "A1:H1", "06　川崎町固有のリスク・論点管理", f_title, F(NAVY), cc)
ws.row_dimensions[1].height = 26
ms(ws, "A2:H2", "──キックオフ確認事項に基づく重要論点と対応状況──", f_sub, F(BLUE), cl)
ws.row_dimensions[2].height = 20

headers = ["No", "論点・リスク", "影響範囲", "想定リスク", "対応方針", "対応状況", "優先度", "次アクション"]
widths = [6, 22, 14, 26, 26, 12, 8, 22]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = f_head; c.fill = F(BLUE); c.alignment = cc; c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[3].height = 30

issues = [
    ("L1", "高齢者外出タクシー助成 R7.3終了", "第5章5-2", "代替手段周知不足、利用者の不満",
     "3層構造（社協・NPO/デマンドバス/町民バス）の整理・住民周知強化を5-2で明示",
     "対応中", "A", "アンケート結果で実態把握"),
    ("L2", "認知症基本法対応(基本法R6.1施行)", "第6章独立章",
     "計画への対応漏れ、本人意見聴取の不足",
     "基本目標6新設・独立章化(第6章)、KPI 3層構造化、本人聴取は包括センター経由",
     "対応中", "A", "第6章詳細化済・本人聴取実施"),
    ("L3", "施設の地域偏在(役場周辺集中)", "第5章5-4",
     "中山間地域住民のサービス利用困難",
     "町外施設利用(住所地特例24人)を見込量に組み込み、5-4で論点整理",
     "対応中", "A", "見込量算定時に詳細整理"),
    ("L4", "介護人材不足", "第5章5-4",
     "サービス継続困難、新規事業所開設不可",
     "宮城県社協 修学資金貸付制度の周知、ICT・介護ロボット導入支援",
     "対応中", "B", "事業者連絡会で実態把握"),
    ("L5", "町外医療機関の利用(広域連携)", "第5章5-3",
     "通院負担、緊急時対応の遅れ",
     "みやぎ県南中核病院・刈田綜合病院との連携を5-3に明示",
     "対応中", "B", "アンケートK-1で実態把握"),
    ("L6", "包括センター人材不足(認定調査1名)", "第5章5-5",
     "業務過多、ケアプラン外部委託の依存",
     "5-5で体制強化を明示、認定調査員増員を検討",
     "保留", "A", "町と人員計画協議"),
    ("L7", "8050問題・老老介護世帯増加", "第5章5-3 第6章",
     "未婚の子+高齢親、配偶者間介護の負担",
     "5-3で家族介護者支援・介護離職防止を強化、第6章認知症連動",
     "対応中", "A", "アンケート世帯類型×不安"),
    ("L8", "介護給付費準備基金残高 R8.6確定", "第7章7-3",
     "保険料試算3パターンの根拠データ不足",
     "町担当による確定値の早期共有",
     "保留", "A", "町担当に再度依頼"),
    ("L9", "基本理念・基本目標 第9期からの継承", "第4章",
     "踏襲レベルの判断・新設項目との整合",
     "9期踏襲方針確認、基本目標6(認知症)のみ新設",
     "完了", "A", "委員会確認"),
    ("L10", "アンケート回収率", "全体",
     "サンプル数不足、地区別分析不可",
     "発送数を多く(高齢者1000+認定300)・督促実施",
     "未着手", "A", "町と督促体制協議"),
    ("L11", "見える化システム比較値の収集", "第7章7-3",
     "仙南圏域・県平均保険料の数値が未取得",
     "見える化システム登録後に取得・反映",
     "未着手", "B", "Ver.2.0更新時に取得"),
    ("L12", "町長挨拶文の最終確定", "計画書冒頭",
     "町長名未確認、挨拶文の町長本人確認",
     "計画書素案v1.5の挨拶文を町に提示、町長確認を依頼",
     "未着手", "B", "町担当経由で確認"),
    ("L13", "3計画同時策定(地域福祉・障害)との整合", "全体",
     "ジャパン総研策定の地福・障害計画との整合",
     "重層的支援体制・移動支援・データ共有等で適宜調整",
     "対応中", "B", "ジャパン総研と情報共有"),
    ("L14", "国通知(R8夏以降)の反映", "全章",
     "計画素案完成後の国指針変更",
     "Ver.2.0以降で適宜反映、影響大きい場合は委員会再協議",
     "未着手", "A", "R8夏以降の通知監視"),
]

r = 4
for issue in issues:
    no, name, scope, risk, plan, status, pri, next_act = issue
    c = ws.cell(row=r, column=1, value=no)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=NAVY); c.alignment = cc; c.border = border
    c.fill = F(LBLUE)
    c = ws.cell(row=r, column=2, value=name)
    c.font = Font(name="游ゴシック", size=10, bold=True); c.alignment = cl; c.border = border
    c = ws.cell(row=r, column=3, value=scope)
    c.font = f_body; c.alignment = cc; c.border = border
    c = ws.cell(row=r, column=4, value=risk)
    c.font = f_body; c.alignment = cl; c.fill = F(LRED); c.border = border
    c = ws.cell(row=r, column=5, value=plan)
    c.font = f_body; c.alignment = cl; c.fill = F(LGREEN); c.border = border
    
    # 対応状況
    if status == "完了":
        c = ws.cell(row=r, column=6, value="✓ 完了"); c.font = f_done; c.fill = F(COMPLETED)
    elif status == "対応中":
        c = ws.cell(row=r, column=6, value="対応中"); c.font = f_prog; c.fill = F(IN_PROGRESS)
    elif status == "保留":
        c = ws.cell(row=r, column=6, value="保留"); c.font = f_pend; c.fill = F(ON_HOLD)
    else:
        c = ws.cell(row=r, column=6, value="未着手"); c.font = f_body; c.fill = F(NOT_STARTED)
    c.alignment = cc; c.border = border
    
    # 優先度
    pri_fill = F(LORANGE) if pri == "A" else (F(LBLUE) if pri == "B" else F(LGRAY))
    pri_color = RED if pri == "A" else (BLUE if pri == "B" else GRAY)
    c = ws.cell(row=r, column=7, value=pri)
    c.font = Font(name="游ゴシック", size=10, bold=True, color=pri_color); c.alignment = cc; c.fill = pri_fill; c.border = border
    
    c = ws.cell(row=r, column=8, value=next_act)
    c.font = f_note; c.alignment = cl; c.border = border
    ws.row_dimensions[r].height = 44
    r += 1

# 注記
r += 1
notes = [
    "本論点リストは、キックオフ会議および計画素案開発過程で抽出された川崎町固有のリスク・論点を一元管理するものです。",
    "対応状況は隔週で更新し、対応中・保留の項目は次回打合せで取り上げます。",
    "新たな論点が発生した場合は本シート末尾に追記してください。",
]
for note in notes:
    ms(ws, f"A{r}:H{r}", "※ " + note, f_note, F("FFF2CC"), cl)
    ws.row_dimensions[r].height = 22
    r += 1

setup_page(ws)

# 保存
out = "/home/claude/kawasaki_work/川崎町_作業チェックリスト_v2.0.xlsx"
wb.save(out)
print(f"作成完了: {out}")
print(f"シート数: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws_s = wb[s]
    print(f"  - {s}: {ws_s.max_row}行 x {ws_s.max_column}列")
