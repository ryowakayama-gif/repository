# -*- coding: utf-8 -*-
"""
保険者機能強化推進交付金等 全国集計結果の照合（令和8年9月2日 追加受領分）

・追加受領の5ファイルが、令和8年8月28日受領のZIP（09_元資料/交付金評価/）と
  同一内容であるかを、川崎町の全項目で照合する
・3か年（令和6〜8年度交付）の目標別得点を抽出する
・令和7年度→令和8年度の項目別の増減を抽出する

各年度のファイルは、先頭に「前年度の評価指標 合計得点・順位」を参考として持ち、
当年度の合計得点・順位は末尾（推進・支援合計の直後）にある。先頭の列を当年度の値と
読み違えないよう、合計は「推進・支援合計」等の見出しを持つ列から取得する。
"""
import re

import openpyxl

U = "/root/.claude/uploads/4be8f82c-e2c9-52ac-b2a9-bad5154d0b13/"
B = "09_元資料/交付金評価/"
PAIRS = [
    ("令和6年度交付", U + "15644dcb-001474341.xlsx",
     B + "①令和６年度保険者機能強化推進交付金等（市町村）に係る全国集計結果/"
         "01_（掲載用）令和６年度評価結果（市町村分）【交付見込額確定版】.xlsx"),
    ("令和7年度交付", U + "4761d46e-001732645.xlsx",
     B + "①令和７年度保険者機能強化推進交付金等（市町村）に係る全国集計結果/"
         "01_（掲載用）令和７年度評価結果（市町村分）.xlsx"),
    ("令和8年度交付", U + "d98e56eb-001732614_2.xlsx",
     B + "①令和８年度保険者機能強化推進交付金等（市町村）に係る全国集計結果/"
         "01_（掲載用）令和８年度評価結果（市町村分）.xlsx"),
]
NUM = re.compile(r"^[\d,.\-％%]+$")
TOTALS = ["Ⅰ合計", "Ⅱ合計", "Ⅲ合計", "Ⅳ合計", "推進合計", "支援合計", "推進・支援合計"]


def load(path):
    """川崎町の行番号と、前方補完したヘッダを返す。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    kr = None
    for r in range(1, ws.max_row + 1):
        for c in range(6, 10):
            if ws.cell(r, c).value == "川崎町":
                kr = r
                break
        if kr:
            break
    nrow = kr - 1
    H = [[""] * (ws.max_column + 2) for _ in range(nrow + 1)]
    for hr in range(1, nrow + 1):
        last = ""
        for c in range(1, ws.max_column + 1):
            v = ws.cell(hr, c).value
            v = "" if v is None else str(v).replace("\n", "").strip()
            if v:
                last = v
            H[hr][c] = last
    return ws, kr, H, nrow


def totals(path):
    """目標別・区分別の合計得点を {ラベル: [推進, 支援]} で返す。

    ヘッダを前方補完すると「Ⅳ合計」が右隣の「推進合計」列にも波及するため、
    合計列の判定には前方補完前の生のセルを用いる。
    """
    ws, kr, H, nrow = load(path)
    out = {}
    for c in range(1, ws.max_column + 1):
        raw = [ws.cell(hr, c).value for hr in range(1, kr)]
        raw = [str(v).replace("\n", "").strip() for v in raw if v is not None]
        for key in TOTALS:
            if any(x == key for x in raw):
                v = ws.cell(kr, c).value
                if isinstance(v, (int, float)):
                    out.setdefault(key, []).append(v)
                break
    return out


def items(path):
    """項目別の得点を {ラベル: 得点} で返す（推進・支援の別を含む）。

    推進交付金と支援交付金の境目は、前方補完前のセルに
    「介護保険保険者努力支援交付金」が現れる列とする。
    """
    ws, kr, H, nrow = load(path)
    # 1行目は資料全体の標題（推進・支援の両方を含む）なので2行目以降で判定する
    border = None
    for c in range(1, ws.max_column + 1):
        for hr in range(2, kr):
            v = ws.cell(hr, c).value
            if v and str(v).replace("\n", "").strip().startswith("介護保険保険者努力支援交付金"):
                border = c
                break
        if border:
            break
    d = {}
    for c in range(1, ws.max_column + 1):
        head = [H[hr][c] for hr in range(1, nrow + 1)]
        if "交付金" not in " ".join(head[:4]):
            continue
        kind = "支援" if border and c >= border else "推進"
        segs = [x for x in dict.fromkeys(head[1:]) if x and not NUM.match(x)]
        segs = [x for x in segs if "交付金" not in x][:5]
        if not segs:
            continue
        v = ws.cell(kr, c).value
        if not isinstance(v, (int, float)):
            continue
        # 列番号は年度により異なるため鍵に含めない。同名が続く場合は出現順で区別する。
        key = f"{kind} / " + " / ".join(segs)
        n = 2
        while key in d:
            key = f"{kind} / " + " / ".join(segs) + f" #{n}"
            n += 1
        d[key] = v
    return d


def main():
    print("■ 1. 追加受領分と既存分の照合（川崎町）")
    print("-" * 76)
    for label, new, old in PAIRS:
        tn, to = totals(new), totals(old)
        same = tn == to
        print(f"  {label}: {'一致' if same else '★不一致'}")
        if not same:
            for k in TOTALS:
                if tn.get(k) != to.get(k):
                    print(f"      {k}: 添付{tn.get(k)} / 既存{to.get(k)}")

    print("\n■ 2. 3か年の目標別得点（川崎町）")
    print("-" * 76)
    print(f"  {'目標':44s} {'R6交付':>8s} {'R7交付':>8s} {'R8交付':>8s}")
    series = {label: totals(old) for label, _, old in PAIRS}
    labels = [("推進 Ⅰ 持続可能な地域", "Ⅰ合計", 0), ("推進 Ⅱ 公正・公平な給付", "Ⅱ合計", 0),
              ("推進 Ⅲ 介護人材・提供基盤", "Ⅲ合計", 0),
              ("推進 Ⅳ アウトカム", "Ⅳ合計", 0), ("推進 計（400点）", "推進合計", 0),
              ("支援 Ⅰ 介護予防／日常生活支援", "Ⅰ合計", 1), ("支援 Ⅱ 認知症総合支援", "Ⅱ合計", 1),
              ("支援 Ⅲ 在宅医療・介護連携", "Ⅲ合計", 1),
              ("支援 Ⅳ アウトカム", "Ⅳ合計", 1), ("支援 計（400点）", "支援合計", 0),
              ("合計（800点）", "推進・支援合計", 0)]
    for name, key, idx in labels:
        row = []
        for label, _, _ in PAIRS:
            v = series[label].get(key, [])
            row.append(v[idx] if len(v) > idx else "―")
        print(f"  {name:44s} {str(row[0]):>8s} {str(row[1]):>8s} {str(row[2]):>8s}")

    print("\n■ 3. 令和7年度→令和8年度の項目別の増減（川崎町）")
    print("-" * 76)
    a, b = items(PAIRS[1][2]), items(PAIRS[2][2])
    for k in dict.fromkeys(list(a) + list(b)):
        x, y = a.get(k), b.get(k)
        if x is None or y is None or x == y:
            continue
        if any(t in k for t in TOTALS):
            continue
        print(f"  {k[:72]:74s} {x:>4} → {y:<4} ({y - x:+})")


if __name__ == "__main__":
    main()
