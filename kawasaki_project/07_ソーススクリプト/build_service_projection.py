# -*- coding: utf-8 -*-
"""
川崎町 第10期　サービス見込量・給付費推計ワークブック ジェネレータ

委託仕様書（保介第41号）6（5）が求める
　・要介護認定者数の推計
　・サービス事業量の推計
に対応するため、給付費を「認定者数 × 利用率 × 単価」に分解して積み上げる。

実績はすべて第9期計画書（09_元資料/川崎町_第9期計画_04324.pdf）による。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "川崎町_サービス見込量・給付費推計ワークブック.xlsx")

FONT = "游ゴシック"
C_TITLE, C_HEAD, C_LEAD, C_BAND = "1F4E78", "5B9BD5", "F2F2F2", "DDEBF7"
YELLOW, GREEN, PINK = "FFF2CC", "E2EFDA", "FCE4D6"
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, lead, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    n = len(widths)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=lead)
    c.font = Font(name=FONT, size=9)
    c.fill = PatternFill("solid", fgColor=C_LEAD)
    c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 40
    ws.sheet_view.showGridLines = False
    return ws


def put(ws, r, c, v, *, align="left", fill=None, bold=False, fmt=None, size=10, wrap=True):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=FONT, size=size, bold=bold)
    x.alignment = Alignment(vertical="center", horizontal=align, wrap_text=wrap,
                            indent=1 if align == "left" else 0)
    x.border = BORDER
    if fill:
        x.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        x.number_format = fmt
    return x


def hdr(ws, r, labels, fill=C_HEAD):
    for i, h in enumerate(labels, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[r].height = 30


def note(ws, r, n, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT, size=8)
    c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 24


def build():
    wb = Workbook(); wb.remove(wb.active)

    # ============================================== 00_使い方
    ws = sheet(wb, "00_使い方", "第10期　サービス見込量・給付費推計ワークブック",
               "委託仕様書（保介第41号）6（5）が求める「要介護認定者数の推計」及び「サービス事業量の推計」に対応し、"
               "給付費を『認定者数 × 利用率 × 単価』に分解して積み上げる。"
               "黄色セルが入力項目。値を変更すると下流のシートと保険料が自動で再計算される。",
               [26, 60, 34])
    r = 4
    hdr(ws, r, ["シート", "内容", "入力の要否"]); r += 1
    for n, d, i in [
        ("01_実績データ", "第9期計画書による認定者数・受給者数・給付費の実績（R2〜R5）", "入力不要（実績・固定）"),
        ("02_認定者数の推計", "高齢者人口 × 認定率 により第10期の認定者数を推計", "黄色：認定率"),
        ("03_利用率の設定", "認定者数に対するサービス区分別の受給者割合", "黄色：第10期の利用率"),
        ("04_単価の設定", "受給者1人あたり月額給付費。介護報酬改定率を乗じる", "黄色：報酬改定率"),
        ("05_見込量・給付費", "受給者数と給付費の年度別見込み（02×03×04）", "入力不要（自動計算）"),
        ("06_施設シナリオ", "施設入所申込47人の反映度合いによる3シナリオ", "黄色：入所増加人数"),
        ("07_保険料への反映", "標準給付費を保険料8ステップに接続し基準額を算定", "黄色：他パラメータ"),
        ("08_必要資料", "推計の精度を上げるために提供を要する資料の一覧", "─"),
    ]:
        put(ws, r, 1, n, size=9); put(ws, r, 2, d, size=9); put(ws, r, 3, i, size=9)
        ws.row_dimensions[r].height = 22; r += 1
    r += 1
    for head, body in [
        ("【推計の考え方】",
         "給付費 ＝ Σ（サービス区分別 受給者数 × 1人あたり月額単価 × 12か月）\n"
         "　　受給者数 ＝ 認定者数 × サービス区分別利用率\n"
         "　　認定者数 ＝ 第1号被保険者数（高齢者人口）× 認定率\n"
         "従来の「総給付額を一律の伸び率で延伸する方法」と異なり、人口・認定率・利用率・単価のどれが"
         "給付費を動かしているかが分解して見える。委員会での説明と政策判断に用いる。"),
        ("【実績の出典】",
         "認定者数・認定率・要介護度別内訳：第9期計画書 p.12（介護保険事業状況報告、各年9月末）\n"
         "受給者数（居宅・地域密着型・施設）：同 p.13（介護保険事業報告年報、各年度1か月平均）\n"
         "サービス種類別給付費：同 p.14（各年度介護状況報告年報）"),
        ("【現時点の限界】",
         "国の推計ガイドは「年齢階級別・性別の人口 × 年齢階級別・性別の認定率」による推計を求めているが、"
         "年齢階級別の認定者数が未入手のため、本ワークブックは全体の認定率による簡便法を用いている。"
         "また利用率・単価はサービス3区分の粗い単位であり、仕様書6（5）が求めるサービス種類別"
         "（訪問介護・通所介護等）の事業量推計には08シートの資料が必要である。"),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=head + body)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
        c.border = BORDER
        ws.row_dimensions[r].height = 76; r += 1

    # ============================================== 01_実績データ
    ws = sheet(wb, "01_実績データ", "01　実績データ（第9期計画書より）",
               "第9期計画書に掲載された実績。入力不要。02〜04シートの設定はこの実績を出発点とする。",
               [22, 13, 13, 13, 13, 30])
    r = 4
    put(ws, r, 1, "A. 要支援・要介護認定者数と認定率（各年9月末）", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["区分", "R2(2020)", "R3(2021)", "R4(2022)", "R5(2023)", "出典・備考"]); r += 1
    for lab, v, memo in [
        ("高齢者人口（人）", [3207, 3250, 3273, 3286], "第9期計画 p.11"),
        ("認定者数（人）", [599, 568, 575, 578], "第9期計画 p.11・p.12"),
        ("認定率", [None, None, None, None], "＝認定者数÷高齢者人口"),
    ]:
        put(ws, r, 1, lab, size=9)
        for i in range(4):
            if v[i] is None:
                put(ws, r, 2+i, f"={get_column_letter(2+i)}{r-1}/{get_column_letter(2+i)}{r-2}",
                    align="center", fmt="0.0%")
            else:
                put(ws, r, 2+i, v[i], align="center", fmt="#,##0")
        put(ws, r, 6, memo, size=8); r += 1
    r += 1
    put(ws, r, 1, "B. サービス区分別 受給者数（各年度1か月平均）", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["サービス区分", "R2年度", "R3年度", "R4年度", "（R4を基準）", "出典・備考"]); r += 1
    JU = {"居宅サービス": [281, 279, 278], "地域密着型サービス": [46, 47, 46], "施設サービス": [130, 124, 131]}
    juk_first = r
    for k, v in JU.items():
        put(ws, r, 1, k, size=9)
        for i, x in enumerate(v): put(ws, r, 2+i, x, align="center", fmt="#,##0")
        put(ws, r, 5, "◀ 基準年", align="center", size=8, fill=GREEN)
        put(ws, r, 6, "第9期計画 p.13", size=8); r += 1
    put(ws, r, 1, "受給者 計", size=9, bold=True)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=SUM({L}{juk_first}:{L}{r-1})", align="center", fmt="#,##0", bold=True)
    put(ws, r, 6, "※延べではなく区分別の実人数", size=8); r += 2

    put(ws, r, 1, "C. サービス区分別 給付費（千円）", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["サービス区分", "R2年度", "R3年度", "R4年度", "（R4を基準）", "出典・備考"]); r += 1
    KY = {"居宅サービス": [367393, 366043, 339473], "地域密着型サービス": [149355, 152764, 153142],
          "施設サービス": [443089, 441677, 469181]}
    kyu_first = r
    for k, v in KY.items():
        put(ws, r, 1, k, size=9)
        for i, x in enumerate(v): put(ws, r, 2+i, x, align="center", fmt="#,##0")
        put(ws, r, 5, "◀ 基準年", align="center", size=8, fill=GREEN)
        put(ws, r, 6, "第9期計画 p.14", size=8); r += 1
    put(ws, r, 1, "総給付費 計", size=9, bold=True)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=SUM({L}{kyu_first}:{L}{r-1})", align="center", fmt="#,##0", bold=True)
    put(ws, r, 6, "計画表の総給付費と一致", size=8); r += 2
    note(ws, r, 6, "注）R4年度を推計の基準年とする。R5〜R8年度の年報実績は未入手であり、"
                   "受領後は基準年をR7年度に更新することが望ましい（08シート No.1）。")

    # ============================================== 02_認定者数の推計
    ws = sheet(wb, "02_認定者数の推計", "02　認定者数の推計（人口 × 認定率）",
               "第1号被保険者数に認定率を乗じて認定者数を推計する。認定率（黄色）が入力項目。"
               "第9期計画の認定率実績は R3=17.5% R4=17.6% R5=17.5%、見える化による推計は R6=17.7% R7=17.6% R8=18.1%。"
               "高齢者数は減少するが後期高齢者の増加により認定率は上昇する。",
               [22, 15, 15, 15, 15, 34])
    r = 4
    hdr(ws, r, ["区分", "R9(2027)", "R10(2028)", "R11(2029)", "3年計", "備考"]); r += 1
    pop_row = r
    put(ws, r, 1, "第1号被保険者数（人）", size=9)
    for i, v in enumerate([3262, 3234, 3206]):
        put(ws, r, 2+i, v, align="center", fmt="#,##0")
    put(ws, r, 5, f"=SUM(B{r}:D{r})", align="center", fmt="#,##0", bold=True)
    put(ws, r, 6, "保険料試算ワークブック 01!B8:B10と同値", size=8); r += 1
    rate_row = r
    put(ws, r, 1, "認定率", size=9)
    for i, v in enumerate([0.183, 0.186, 0.189]):
        put(ws, r, 2+i, v, align="center", fmt="0.0%", fill=YELLOW)
    put(ws, r, 5, "【入力】", align="center", size=8, fill=YELLOW)
    put(ws, r, 6, "R8実績見込18.1%からの上昇を見込む。年齢階級別認定率が入手できれば置換", size=8); r += 1
    nin_row = r
    put(ws, r, 1, "認定者数（人）", size=9, bold=True)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=ROUND({L}{pop_row}*{L}{rate_row},0)", align="center", fmt="#,##0", bold=True, fill=GREEN)
    put(ws, r, 5, f"=SUM(B{r}:D{r})", align="center", fmt="#,##0", bold=True, fill=GREEN)
    put(ws, r, 6, "＝第1号被保険者数×認定率", size=8); r += 2

    put(ws, r, 1, "（参考）要介護度別の構成比（第9期計画 p.12 のR8推計）", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["要介護度", "構成比", "R9見込（人）", "R10見込（人）", "R11見込（人）", "備考"]); r += 1
    kousei_first = r
    for lab, ratio in [("要支援1", 0.057), ("要支援2", 0.158), ("要介護1", 0.104), ("要介護2", 0.241),
                       ("要介護3", 0.152), ("要介護4", 0.177), ("要介護5", 0.111)]:
        put(ws, r, 1, lab, size=9)
        put(ws, r, 2, ratio, align="center", fmt="0.0%", fill=YELLOW)
        for i in range(3):
            L = get_column_letter(2+i)
            put(ws, r, 3+i, f"=ROUND({L}{nin_row}*$B{r},0)", align="center", fmt="#,##0")
        put(ws, r, 6, "第9期計画R8推計の構成比を据置", size=8); r += 1
    put(ws, r, 1, "計", size=9, bold=True)
    put(ws, r, 2, f"=SUM(B{kousei_first}:B{r-1})", align="center", fmt="0.0%", bold=True)
    for i in range(3):
        L = get_column_letter(3+i)
        put(ws, r, 3+i, f"=SUM({L}{kousei_first}:{L}{r-1})", align="center", fmt="#,##0", bold=True)
    r += 2
    note(ws, r, 6, "注）要介護3以上は第10期計画期間で約44%。特別養護老人ホームの入所要件を満たす層であり、"
                   "06シートの施設シナリオの基礎となる。")

    # ============================================== 03_利用率の設定
    ws = sheet(wb, "03_利用率の設定", "03　利用率の設定（受給者数 ÷ 認定者数）",
               "認定者数に対するサービス区分別の受給者の割合。R4年度実績を基準とし、第10期の設定値（黄色）を入力する。"
               "利用率を据え置けば現状維持、引き上げれば在宅・施設サービスの利用促進を織り込むことになる。",
               [24, 13, 13, 13, 15, 38])
    r = 4
    hdr(ws, r, ["サービス区分", "R2年度", "R3年度", "R4年度", "第10期設定", "設定の考え方"]); r += 1
    riyou_first = r
    for k, jj, jissekiJ, memo in [
        ("居宅サービス", [281, 279, 278], [599, 568, 575],
         "R4実績を据置。在宅生活継続の施策効果は当面見込まない"),
        ("地域密着型サービス", [46, 47, 46], [599, 568, 575],
         "R4実績を据置。定員に空きがなく大幅な増は見込みにくい"),
        ("施設サービス", [130, 124, 131], [599, 568, 575],
         "R4実績を据置。入所申込47人の反映は06シートで別途検討"),
    ]:
        put(ws, r, 1, k, size=9)
        for i in range(3):
            put(ws, r, 2+i, jj[i]/jissekiJ[i], align="center", fmt="0.0%")
        put(ws, r, 5, jj[2]/jissekiJ[2], align="center", fmt="0.0%", fill=YELLOW, bold=True)
        put(ws, r, 6, memo, size=8); r += 1
    put(ws, r, 1, "受給率（合計）", size=9, bold=True)
    for i in range(4):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=SUM({L}{riyou_first}:{L}{r-1})", align="center", fmt="0.0%", bold=True)
    put(ws, r, 6, "第9期計画 p.13 の受給率（R4=79.0%）と一致", size=8); r += 2
    note(ws, r, 6, "注1）利用率は認定者数に対する割合であり、1人が複数区分を利用する場合は重複する。"
                   "第9期計画 p.13 の受給実人数（R4=454人）とは概念が異なる。")
    r += 1
    note(ws, r, 6, "注2）在宅介護実態調査では、在宅の認定者142人のうち介護保険サービス利用ありは41.3%にとどまる。"
                   "居宅利用率48.3%との差は要確認（08シート No.5）。")

    # ============================================== 04_単価の設定
    ws = sheet(wb, "04_単価の設定", "04　単価の設定（受給者1人あたり月額給付費）",
               "給付費 ÷ 受給者数 ÷ 12 により1人あたり月額単価を算出する。"
               "第10期は令和9年度の介護報酬改定が見込まれるため、改定率（黄色）を乗じて設定する。",
               [24, 15, 15, 15, 14, 34])
    r = 4
    hdr(ws, r, ["サービス区分", "R2年度", "R3年度", "R4年度", "報酬改定率", "第10期単価（千円/月）"]); r += 1
    tanka_first = r
    for k, kg, ju in [("居宅サービス", [367393, 366043, 339473], [281, 279, 278]),
                      ("地域密着型サービス", [149355, 152764, 153142], [46, 47, 46]),
                      ("施設サービス", [443089, 441677, 469181], [130, 124, 131])]:
        put(ws, r, 1, k, size=9)
        for i in range(3):
            put(ws, r, 2+i, kg[i]/ju[i]/12, align="center", fmt="#,##0.0")
        put(ws, r, 5, 1.000, align="center", fmt="0.0%", fill=YELLOW)
        put(ws, r, 6, f"=D{r}*E{r}", align="center", fmt="#,##0.0", bold=True, fill=GREEN)
        r += 1
    r += 1
    note(ws, r, 6, "注1）報酬改定率は現在1.000（改定なし）で仮置きしている。令和9年度の介護報酬改定率が"
                   "示された段階で入力する。改定率1%につき標準給付費は約1%、保険料基準額は約1%動く。")
    r += 1
    note(ws, r, 6, "注2）単価は区分内のサービス構成が一定であることを前提とする。"
                   "施設サービスは特別養護老人ホームと介護老人保健施設の構成比により単価が変動する。")

    # ============================================== 05_見込量・給付費
    ws = sheet(wb, "05_見込量・給付費", "05　サービス見込量と給付費の見込み",
               "02（認定者数）×03（利用率）＝受給者数、受給者数×04（単価）×12＝給付費。すべて自動計算。",
               [24, 15, 15, 15, 17, 30])
    r = 4
    put(ws, r, 1, "A. 受給者数の見込み（人）＝ 認定者数 × 利用率", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["サービス区分", "R9(2027)", "R10(2028)", "R11(2029)", "3年計", "備考"]); r += 1
    juk2_first = r
    for i, k in enumerate(["居宅サービス", "地域密着型サービス", "施設サービス"]):
        put(ws, r, 1, k, size=9)
        for j in range(3):
            L = get_column_letter(2+j)
            put(ws, r, 2+j, f"=ROUND('02_認定者数の推計'!{L}{nin_row}*'03_利用率の設定'!$E${riyou_first+i},0)",
                align="center", fmt="#,##0")
        put(ws, r, 5, f"=SUM(B{r}:D{r})", align="center", fmt="#,##0")
        put(ws, r, 6, "02×03", size=8); r += 1
    put(ws, r, 1, "計", size=9, bold=True)
    for j in range(4):
        L = get_column_letter(2+j)
        put(ws, r, 2+j, f"=SUM({L}{juk2_first}:{L}{r-1})", align="center", fmt="#,##0", bold=True)
    r += 2

    put(ws, r, 1, "B. 給付費の見込み（千円）＝ 受給者数 × 単価 × 12か月", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["サービス区分", "R9(2027)", "R10(2028)", "R11(2029)", "3年計", "備考"]); r += 1
    kyu2_first = r
    for i in range(3):
        put(ws, r, 1, ["居宅サービス", "地域密着型サービス", "施設サービス"][i], size=9)
        for j in range(3):
            L = get_column_letter(2+j)
            put(ws, r, 2+j, f"=ROUND({L}{juk2_first+i}*'04_単価の設定'!$F${tanka_first+i}*12,0)",
                align="center", fmt="#,##0")
        put(ws, r, 5, f"=SUM(B{r}:D{r})", align="center", fmt="#,##0")
        put(ws, r, 6, "受給者数×単価×12", size=8); r += 1
    sougyufu_row = r
    put(ws, r, 1, "総給付額 計", size=9, bold=True)
    for j in range(4):
        L = get_column_letter(2+j)
        put(ws, r, 2+j, f"=SUM({L}{kyu2_first}:{L}{r-1})", align="center", fmt="#,##0", bold=True, fill=GREEN)
    r += 1
    hokan_row = r
    put(ws, r, 1, "補完項目", size=9)
    for j, v in enumerate([88451, 89512, 90586]):
        put(ws, r, 2+j, v, align="center", fmt="#,##0", fill=YELLOW)
    put(ws, r, 5, f"=SUM(B{r}:D{r})", align="center", fmt="#,##0")
    put(ws, r, 6, "特定入所者・高額介護・高額医療合算・審査手数料", size=8); r += 1
    hyoujun_row = r
    put(ws, r, 1, "標準給付費見込額", size=9, bold=True)
    for j in range(4):
        L = get_column_letter(2+j)
        put(ws, r, 2+j, f"={L}{sougyufu_row}+{L}{hokan_row}", align="center", fmt="#,##0", bold=True, fill=GREEN)
    put(ws, r, 6, "07シートで保険料に接続", size=8); r += 2
    note(ws, r, 6, "注）補完項目は現行の保険料試算ワークブック 02シートの値をそのまま用いている。"
                   "特定入所者介護サービス費は施設・短期入所の利用量に連動するため、"
                   "本来は施設受給者数に比例させるべきである（08シート No.4）。")

    # ============================================== 06_施設シナリオ
    ws = sheet(wb, "06_施設シナリオ", "06　施設入所申込の反映シナリオ",
               "在宅介護実態調査（n=139）では、施設等への入所・入居を『すでに申し込み』33.8%（142件換算で約47人）、"
               "『検討している』14.4%（約20人）であった。要介護3以上は58.3%を占める。"
               "この需要をどこまで見込量に反映するかは委員会の政策判断事項であり、保険料に直結する。",
               [30, 16, 16, 16, 16, 26])
    r = 4
    hdr(ws, r, ["シナリオ", "施設受給者の増", "施設受給者数(R11)", "標準給付費3年計(千円)", "保険料基準額A(円)", "考え方"]); r += 1
    scen_first = r
    base_shisetsu = 131/575
    for lab, add, memo in [
        ("① 現状維持（利用率据置）", 0, "町内定員を増やさず、待機は解消しない"),
        ("② 一部反映（申込の2割）", 10, "町外施設の活用等により段階的に対応"),
        ("③ 相当程度の反映（申込の4割）", 20, "地域密着型の整備等を伴う"),
        ("④ 大幅反映（申込の6割）", 30, "施設整備を要し、供給側の制約が課題"),
    ]:
        put(ws, r, 1, lab, size=9)
        put(ws, r, 2, add, align="center", fmt="#,##0", fill=YELLOW)
        put(ws, r, 3, f"=ROUND('02_認定者数の推計'!D{nin_row}*'03_利用率の設定'!$E${riyou_first+2},0)+B{r}",
            align="center", fmt="#,##0")
        # 標準給付費 = 05の標準給付費 + 増分×施設単価×12×3年
        put(ws, r, 4, f"='05_見込量・給付費'!E{hyoujun_row}+B{r}*'04_単価の設定'!$F${tanka_first+2}*12*3",
            align="center", fmt="#,##0")
        put(ws, r, 5, f"=ROUND((D{r}*(1+'07_保険料への反映'!$B$8)*0.23"
                      f"+D{r}*'07_保険料への反映'!$B$9-D{r}*'07_保険料への反映'!$B$10"
                      f"-'07_保険料への反映'!$B$11)*1000"
                      f"/('07_保険料への反映'!$B$12*12*'07_保険料への反映'!$B$13*'07_保険料への反映'!$B$14),0)",
            align="center", fmt="#,##0", bold=True, fill=GREEN)
        put(ws, r, 6, memo, size=8); r += 1
    r += 1
    note(ws, r, 6, "注1）保険料基準額は基金取崩なし（パターンA）で算定している。"
                   "基金を取り崩す場合はこの額から下がる。")
    r += 1
    note(ws, r, 6, "注2）施設サービスは介護保険法上、計画に定める定員が都道府県の指定に係る協議の基礎となる。"
                   "シナリオ②以降を採る場合は、計画本文に整備方針と定員を明記する必要がある。")

    # ============================================== 07_保険料への反映
    ws = sheet(wb, "07_保険料への反映", "07　保険料への反映（8ステップへの接続）",
               "05シートの標準給付費を保険料算定の8ステップに接続し、基準額を算定する。"
               "現行の保険料試算ワークブック（一律1.2%/年で延伸）との差も示す。",
               [30, 20, 16, 16, 16, 30])
    r = 4
    put(ws, r, 1, "A. 算定パラメータ", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["項目", "値", "", "", "", "備考"]); r += 1
    par_first = r
    for lab, v, fmt, fill, memo in [
        ("標準給付費（3年計・千円）", f"='05_見込量・給付費'!E{hyoujun_row}", "#,##0", GREEN, "05シートから自動参照"),
        ("地域支援事業費／標準給付費", 91563/3390699, "0.00%", YELLOW, "第1号負担分の算定基礎に加算する比率"),
        ("調整交付金相当額率", 0.05, "0.0%", YELLOW, "国基準5%"),
        ("調整交付金見込率", 0.039, "0.0%", YELLOW, "第9期実績3.9%。標準5%との乖離は要確認"),
        ("保険者機能強化交付金（千円）", 10800, "#,##0", YELLOW, "第9期実績"),
        ("第1号被保険者数 3年計（人）", f"='02_認定者数の推計'!E{pop_row}", "#,##0", GREEN, "02シートから自動参照"),
        ("予定収納率", 0.96, "0.0%", YELLOW, "第9期と同水準（暫定）"),
        ("補正係数（所得段階別加重平均料率）", 0.9734, "0.0000", YELLOW,
         "R3年度末の所得段階別人数×第9期軽減前料率で算定（実績データ確認サマリー04シート）"),
    ]:
        put(ws, r, 1, lab, size=9)
        put(ws, r, 2, v, align="center", fmt=fmt, fill=fill, bold=True)
        put(ws, r, 6, memo, size=8); r += 1
    r += 1
    put(ws, r, 1, "B. 保険料基準額の算定", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["項目", "パターンA", "パターンB", "パターンC", "", "備考"]); r += 1
    B = par_first
    kikin_row = r
    put(ws, r, 1, "基金取崩額（千円）", size=9)
    for i, v in enumerate([0, 26850, 53700]):
        put(ws, r, 1+i+1, v, align="center", fmt="#,##0", fill=YELLOW)
    put(ws, r, 6, "取崩なし／50%／全額（残高53,700千円）", size=8); r += 1
    need_row = r
    put(ws, r, 1, "保険料収納必要額（千円）", size=9)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=$B${B}*(1+$B${B+1})*0.23+$B${B}*$B${B+2}-$B${B}*$B${B+3}-$B${B+4}-{L}{kikin_row}",
            align="center", fmt="#,##0")
    put(ws, r, 6, "第1号負担分＋調整交付金相当額−見込額−機能強化−基金取崩", size=8); r += 1
    put(ws, r, 1, "保険料基準額（月額・円）", size=9, bold=True)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"=ROUND({L}{need_row}*1000/($B${B+5}*12*$B${B+6}*$B${B+7}),0)",
            align="center", fmt="#,##0", bold=True, fill=GREEN, size=12)
    put(ws, r, 6, "第9期は6,508円（年額78,000円）", size=8); r += 1
    kijun_row = r - 1
    put(ws, r, 1, "第9期比", size=9)
    for i in range(3):
        L = get_column_letter(2+i)
        put(ws, r, 2+i, f"={L}{kijun_row}/6508-1", align="center", fmt="+0.0%;-0.0%")
    r += 2
    put(ws, r, 1, "C. 現行の保険料試算ワークブックとの比較", bold=True, fill=C_BAND)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
    hdr(ws, r, ["推計方法", "標準給付費3年計", "基準額A", "差", "", "備考"]); r += 1
    put(ws, r, 1, "本ワークブック（認定者数×利用率×単価）", size=9)
    put(ws, r, 2, f"=$B${B}", align="center", fmt="#,##0")
    put(ws, r, 3, f"=B{kijun_row}", align="center", fmt="#,##0")
    put(ws, r, 4, "─", align="center")
    put(ws, r, 6, "仕様書6（5）に対応", size=8); r += 1
    put(ws, r, 1, "現行（総給付額を一律1.2%/年で延伸）", size=9)
    put(ws, r, 2, 3390699, align="center", fmt="#,##0")
    put(ws, r, 3, 7632, align="center", fmt="#,##0")
    put(ws, r, 4, f"=C{r}-C{r-1}", align="center", fmt="+#,##0;-#,##0")
    put(ws, r, 6, "保険料試算ワークブックの現行値", size=8); r += 2
    note(ws, r, 6, "注）差は、現行が総給付額の実績横引きであるのに対し、本ワークブックが"
                   "人口減少（第1号被保険者数の減）と認定率上昇を分けて織り込んだことによる。")

    # ============================================== 08_必要資料
    ws = sheet(wb, "08_必要資料", "08　推計の精度を上げるために提供を要する資料",
               "本ワークブックは第9期計画書に掲載された実績のみで構成している。"
               "以下の資料により、簡便法から国の推計ガイドに沿った方法へ移行できる。",
               [5, 30, 44, 26, 12, 22])
    r = 4
    hdr(ws, r, ["No", "資料", "内容", "これにより可能になること", "優先度", "入手先"]); r += 1
    for no, shiryo, naiyo, kouka, pri, saki in [
        (1, "介護保険事業状況報告 年報（R5〜R7年度）",
         "サービス種類別・要介護度別の受給者数と給付費。第9期計画はR4年度までしか掲載していない。",
         "推計の基準年をR4からR7に更新できる。3年古いデータで推計している現状を解消", "最高", "発注者"),
        (2, "年齢階級別・性別の認定者数（R4〜R8年度）",
         "5歳階級別（65-69、70-74、75-79、80-84、85-89、90以上）の認定者数。性別内訳を含む。",
         "国の推計ガイドが求める「年齢階級別・性別の認定率×人口」による推計に移行できる。"
         "後期高齢者の増加が認定者数に与える影響を分離できる", "最高", "発注者"),
        (3, "サービス種類別の給付実績（R5〜R7年度）",
         "訪問介護・通所介護・短期入所・地域密着型各種・施設種別ごとの件数と給付費。",
         "仕様書6（5）が求めるサービス種類別の事業量推計ができる。"
         "現在の3区分（居宅・地密・施設）の粗い推計を細分化できる", "高", "発注者・国保連"),
        (4, "特定入所者介護サービス費等の実績（R4〜R7年度）",
         "特定入所者介護サービス費・高額介護サービス費・高額医療合算・審査支払手数料の年度別実績。",
         "補完項目を施設利用量に連動させられる。現在は固定値で置いている", "高", "発注者"),
        (5, "介護保険サービス未利用の認定者の状況",
         "認定を受けながらサービスを利用していない者の数と要介護度別内訳。",
         "在宅介護実態調査の「サービス利用あり41.3%」と居宅利用率48.3%の差を説明できる。"
         "利用率の設定根拠が確かになる", "高", "発注者"),
        (6, "町内事業所・施設の定員と稼働状況",
         "特別養護老人ホーム・介護老人保健施設・地域密着型各事業所の定員、利用者数、待機者数。",
         "06シートの施設シナリオを供給側の制約と突き合わせられる。"
         "計画に定員を記載するための基礎となる", "高", "発注者・3事業所"),
        (7, "所得段階別被保険者数（R7またはR8時点・13段階）",
         "第9期の13段階区分による段階別人数。手元にあるのはR3年度末の9段階（第8期区分）のみ。",
         "補正係数を最新分布で確定できる。現在はR3の分布から0.9734を算定して用いている",
         "高", "発注者（賦課台帳）"),
        (8, "住所地特例者の状況（R4〜R7年度）",
         "町外施設利用者24人の施設種別・所在自治体・要介護度の内訳。",
         "施設サービスの給付費のうち町外分を分離でき、町内定員の議論と切り分けられる", "中", "発注者"),
        (9, "介護報酬改定率（令和9年度）",
         "国が示す改定率。サービス種類別の改定内容を含む。",
         "04シートの単価に反映できる。改定率1%につき保険料基準額は約1%動く", "中", "国の情報"),
        (10, "日常生活圏域別の認定者数・受給者数",
         "圏域別（または行政区別）の認定者数、サービス受給者数、事業所所在地。",
         "施設の地域偏在という論点を数値で示せる。ニーズ調査に圏域IDがないため現在は検証できない",
         "中", "発注者"),
    ]:
        put(ws, r, 1, no, align="center", size=9)
        put(ws, r, 2, shiryo, size=9)
        put(ws, r, 3, naiyo, size=8)
        put(ws, r, 4, kouka, size=8)
        put(ws, r, 5, pri, align="center", size=9, fill=PINK if pri == "最高" else None,
            bold=(pri == "最高"))
        put(ws, r, 6, saki, align="center", size=8)
        ws.row_dimensions[r].height = 62; r += 1
    r += 1
    note(ws, r, 6, "注）No.1・No.2があれば、本ワークブックは国の推計ガイドに沿った方法へ移行できる。"
                   "この2点が最優先である。")

    wb.save(OUT)
    print("作成:", OUT)


if __name__ == "__main__":
    build()
