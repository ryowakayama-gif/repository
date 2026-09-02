# -*- coding: utf-8 -*-
"""
サービス別 計画値・実績値 対比表の作成

第9期計画（令和6年3月）の各サービス・事業の見込量（計画値）を計画書から全数抽出し、
実績と対比できる様式を作る。仕様書6（4）の現行施策の検証、及び保険者機能強化推進交付金
目標Ⅰ-2「介護保険事業計画の進捗状況（計画値と実績値の乖離状況）を分析しているか」に対応する。

計画値の出典
  09_元資料/川崎町_第9期計画_04324.pdf（第2部 各論の「■○○見込量」表・全59ブロック）

実績の出典
  09_元資料/R8実績データ/R8.9.1受領版/川崎町_町提供実績データ_R8.9.1受領.xlsx
  ・04シート（サービス別の利用者数・給付費）
  ・06〜08シート（移動支援・生活支援・包括／認知症／権利擁護）
  ※02シートの一般介護予防は出典欄が「第9期計画」であり計画値が転記されているため、
  　実績としては扱わず記入欄のままとする（確認事項No.42）。
"""
import json
import re
import subprocess
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PDF = "09_元資料/川崎町_第9期計画_04324.pdf"
JISSEKI_XLSX = ("09_元資料/R8実績データ/R8.9.1受領版/"
                "川崎町_町提供実績データ_R8.9.1受領.xlsx")
OUT = "05_試算・管理シート/川崎町_サービス別_計画値実績対比表_R8.9.xlsx"

NAVY = "1F3864"
HEAD = "2F5597"
PLAN = "E2EFDA"      # 計画値（確定）
INPUT_ = "FFFFCC"    # 記入欄
GRAY = "F2F2F2"
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [("区分", 12), ("サービス・事業名", 34), ("指標", 16), ("単位", 9),
        ("R6計画", 11), ("R6実績", 11), ("達成率", 9),
        ("R7計画", 11), ("R7実績", 11), ("達成率", 9),
        ("R8計画", 11), ("R8見込", 11), ("達成率", 9),
        ("乖離の要因・対応", 34), ("出典・記入担当", 18)]


# ------------------------------------------------------------ 計画値の抽出
def extract_plan():
    """第9期計画書から「■○○見込量」表を全数抽出する。"""
    import pypdf
    r = pypdf.PdfReader(PDF)
    pages = [re.sub(r"[ \t]+", " ", (r.pages[i].extract_text() or ""))
             for i in range(56, 102)]
    whole = "\n".join(pages)
    whole = re.sub(r"\n※\d+ [^\n]*", "\n", whole)      # 脚注を除去
    out = []
    for part in re.split(r"\n?■", whole)[1:]:
        head = part.split("\n")[0].strip()
        body = re.sub(r"\s+", " ", part)
        if body.count("令和") < 4:
            continue
        rows, seen = [], set()
        for m in re.finditer(
                r"([^\d（）]{1,26}?)\s*（([^）]{1,10}?)）\s*((?:[\d,]+\s+){3}[\d,]+)", body):
            lab = m.group(1).strip()
            if "年度" in lab or not lab or lab in seen:
                continue
            seen.add(lab)
            rows.append((lab, m.group(2), m.group(3).split()))
        for m in re.finditer(r"([^\d（）\s][^（）\d]{0,20}?)\s+((?:[\d,]+\s+){3}[\d,]+)", body):
            lab = m.group(1).strip()
            if "年度" in lab or not lab or lab in seen:
                continue
            seen.add(lab)
            rows.append((lab, "", m.group(2).split()))
        rows = [x for x in rows if valid_label(x[0])]
        if rows:
            out.append({"name": head, "rows": rows})
    return out


def valid_label(lab):
    """指標名として妥当か。表の外の文字列を巻き込んだ行を落とす。

    「キャラバンメイト 研修修了者数」のように区切りの空白を1つ含む指標名があるため、
    空白そのものは許し、空白が2つ以上あるものと単位記号を含むものだけを落とす。
    """
    if not lab or len(lab) > 20:
        return False
    if any(ch in lab for ch in "％%"):
        return False
    if lab.replace("―", "").strip() == "":
        return False
    return lab.count(" ") + lab.count("　") < 2


def num(s):
    s = str(s).strip().translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    s = s.replace(",", "")
    return int(s) if s.isdigit() else None


# ------------------------------------------------------------ 実績の読込み
def load_jisseki():
    """実績データから {(シート,行): [R6, R7見込, R8]} を作る。"""
    wb = openpyxl.load_workbook(JISSEKI_XLSX, data_only=True)
    def take(sheet, row, first_col):
        ws = wb[sheet]
        return [ws.cell(row, first_col + i).value for i in range(3)]
    return wb, take


# サービス名（計画書の見出し）→ 実績データの参照先
# (シート名, 行, 先頭列)。先頭列は R6 の列。
JMAP = {
    "訪問介護（ホームヘルプサービス）見込量": {
        "延利用人数": ("04_サービス利用給付", 5, 9),
        "給付費": ("04_サービス利用給付", 6, 9)},
    "訪問看護見込量": {
        "延利用人数": ("04_サービス利用給付", 7, 9),
        "給付費": ("04_サービス利用給付", 8, 9)},
    "通所介護（デイサービス）見込量": {
        "延利用人数": ("04_サービス利用給付", 9, 9),
        "給付費": ("04_サービス利用給付", 10, 9)},
    "短期入所生活介護（ショートステイ）見込量": {
        "延利用人数": ("04_サービス利用給付", 11, 9),
        "給付費": ("04_サービス利用給付", 12, 9)},
    "認知症対応型共同生活介護見込量": {
        "延利用人数": ("04_サービス利用給付", 13, 9),
        "給付費": ("04_サービス利用給付", 14, 9)},
    "介護老人福祉施設（特別養護老人施設）見込量": {
        "延利用人数": ("04_サービス利用給付", 19, 9),
        "給付費": ("04_サービス利用給付", 20, 9)},
    "介護老人保健施設（老人保健施設）見込量": {
        "延利用人数": ("04_サービス利用給付", 21, 9),
        "給付費": ("04_サービス利用給付", 22, 9)},
    "移送サービス事業見込量": {
        "実利用人数": ("06_移動支援", 5, 8),
        "延利用人数": ("06_移動支援", 6, 8),
        "給付費": ("06_移動支援", 7, 8)},
    "「食」の自立支援事業（配食サービス）見込量": {
        "延利用人数": ("07_生活支援見守り", 5, 8),
        "利用実人数": ("07_生活支援見守り", 6, 8)},
    "緊急通報システム事業": {
        "実利用人数": ("07_生活支援見守り", 7, 8)},
    "認知症高齢者見守りＱＲコード活用事業見込量": {
        "延利用人数": ("07_生活支援見守り", 9, 8)},
    "もの忘れ相談事業見込み量": {
        "開催数": ("08_包括認知症権利擁護", 5, 9)},
    "認知症初期集中支援事業見込み量": {
        "チーム員会議開催数": ("08_包括認知症権利擁護", 7, 9)},
    "認知症サポーター養成講座事業見込み量": {
        "受講者数": ("08_包括認知症権利擁護", 10, 9)},
    "認知症カフェ（喫茶みかん）開催事業見込み量": {
        "カフェ開催数": ("08_包括認知症権利擁護", 12, 9),
        "延利用者数": ("08_包括認知症権利擁護", 13, 9)},
}

# 分野（シート）への割り当て。計画書のブロック名で明示的に指定する
# （部分一致では「訪問介護」が「訪問型サービス（指定事業所による訪問介護サービス）」に
#   当たるなどの取り違えが起きるため）。
ASSIGN = {
    "01": ["訪問介護（ホームヘルプサービス）", "訪問入浴介護", "介護予防訪問入浴介護",
           "訪問看護", "介護予防訪問看護", "訪問リハビリテーション",
           "介護予防訪問リハビリテーション", "居宅療養管理指導", "介護予防居宅療養管理指導",
           "通所介護（デイサービス）", "通所リハビリテーション",
           "介護予防通所リハビリテーション", "短期入所生活介護（ショートステイ）",
           "介護予防短期入所生活介護（ショートステイ）", "短期入所療養介護（ショートステイ）",
           "介護予防短期療養生活介護（ショートステイ）", "特定施設入居者生活介護",
           "介護予防特定施設入居者生活介護", "福祉用具貸与", "介護予防福祉用具貸与",
           "特定福祉用具購入費", "介護予防特定福祉用具購入費", "住宅改修", "介護予防住宅改修",
           "居宅介護支援", "介護予防支援"],
    "02": ["認知症対応型共同生活介護", "介護予防認知症対応型共同生活介護",
           "地域密着型介護老人福祉施設入居者生活介護"],
    "03": ["介護老人福祉施設（特別養護老人施設）", "介護老人保健施設（老人保健施設）"],
    "04": ["訪問型サービス（指定事業所による訪問介護サービス）",
           "訪問型サービス B（住民主体のサービス）",
           "訪問型サービス C（短期集中予防サービス）",
           "通所型サービス（指定事業所による通所サービス）",
           "通所型サービス B（住民主体のサービス）", "介護予防支援事業（ケアマネジメント）",
           "介護予防フェスティバル", "地域リハビリテーション活動支援事業"],
    "05": ["もの忘れ相談事業", "認知症初期集中支援事業", "認知症サポーター養成講座事業",
           "認知症地域支援推進員・キャラバンメイト", "認知症カフェ（喫茶みかん）開催事業",
           "高齢者見守りネットワーク協力機関事業所",
           "介護用品支給（非課税世帯 要介護度１～３）",
           "介護用品支給（非課税世帯 要介護度４・５）",
           "介護用品支給（課税世帯要 介護度４・５）",
           "「食」の自立支援事業（配食サービス）", "認知症高齢者見守りＱＲコード活用事業"],
    "06": ["移送サービス事業", "タクシー利用助成事業", "会食サービス事業",
           "緊急通報システム事業", "老人クラブ助成事業", "敬老会開催事業"],
}

SHEETS = [("01_居宅サービス", "01", "居宅"),
          ("02_地域密着型サービス", "02", "地域密着型"),
          ("03_施設サービス", "03", "施設"),
          ("04_総合事業・一般介護予防", "04", "地域支援事業"),
          ("05_包括的支援事業・任意事業", "05", "地域支援事業"),
          ("06_高齢者福祉サービス", "06", "高齢者福祉")]

# 計画値と実績で定義が一致しない可能性が高い箇所への注記
NOTES = {
    ("移送サービス事業", "延利用人数"):
        "計画は「延利用人数（人）」、実績データは「延利用回数（回）」。定義の確認が必要",
    ("緊急通報システム事業", "実利用人数"):
        "計画40人に対し実績393人。実績は設置台数の累計の可能性。定義の確認が必要",
    ("訪問介護（ホームヘルプサービス）", "給付費"):
        "令和6年度40,912千円から令和7年度27,779千円へ急減。訪問介護の実働ヘルパー数"
        "（登録21名・実働4名）との関係を含め確認が必要",
    ("訪問型サービス（指定事業所による訪問介護サービス）", "延利用人数"):
        "実績データ02シートに訪問型サービスの利用実人数481〜534人の記載があるが、"
        "計画の延利用人数96人と定義が一致しないため転記していない",
    ("移送サービス事業", "給付費"):
        "計画3,150千円に対し実績1,551千円（令和6年度・49％）。事業費の範囲の確認が必要",
    ("「食」の自立支援事業（配食サービス）", "延利用人数"):
        "計画3,972人に対し実績3,060人（令和6年度・77％）",
    ("認知症カフェ（喫茶みかん）開催事業", "カフェ開催数"):
        "計画書の単位表記は「人／年」だが、開催数であり「回／年」の誤記と思われる",
    ("通所型サービス（指定事業所による通所サービス）", "延利用人数"):
        "実績データ02シートに通所型サービスの利用実人数647〜786人の記載があるが、"
        "計画の延利用人数300人と定義が一致しないため転記していない",
}

# 介護予防教室は、計画書の表が複数の教室を1ブロックにまとめており指標名から教室を
# 特定できないため、実績データ02シート（出典欄が「第9期計画」＝計画値）の事業名で行を起こす。
# 計画値が同シートのR6〜R8列、実績は未記入のため空欄とする。
KYOSHITSU = [
    ("元気まんてん介護予防教室", "開催回数", "回／年", 5),
    ("元気まんてん介護予防教室", "延参加者", "人／年", 6),
    ("元気いきいきセミナー", "開催回数", "回／年", 7),
    ("元気いきいきセミナー", "延参加者", "人／年", 8),
    ("パドル体操教室", "開催回数", "回／年", 9),
    ("パドル体操教室", "延参加者", "人／年", 10),
    ("ノルディックウォーキング教室", "開催回数", "回／年", 11),
    ("ノルディックウォーキング教室", "延参加者", "人／年", 12),
    ("介護予防サロン", "設置か所", "か所", 13),
    ("ヨーガ教室", "開催回数", "回／年", 14),
    ("ヨーガ教室", "延参加者", "人／年", 15),
]

# 第9期計画に見込量表がないサービス（「今後のニーズ動向を勘案し検討」とされたもの）
NO_PLAN = [
    ("地域密着型", "定期巡回・随時対応型訪問介護看護",
     "第9期計画に見込量表なし（「今後のサービスニーズの動向を勘案し検討」）"),
    ("地域密着型", "夜間対応型訪問介護", "同上"),
    ("地域密着型", "認知症対応型通所介護・介護予防認知症対応型通所介護", "同上"),
    ("地域密着型", "小規模多機能型居宅介護・介護予防小規模多機能型居宅介護",
     "同上。ただし実績データでは令和6年度以降に給付実績がある"),
    ("地域密着型", "地域密着型特定施設入居者生活介護", "同上"),
    ("地域密着型", "看護小規模多機能型居宅介護（複合型サービス）", "同上"),
    ("地域密着型", "地域密着型通所介護（小規模デイ）",
     "同上。ただし実績データでは令和4年度以降に給付実績がある"),
    ("施設", "介護療養型医療施設", "令和5年度末に制度廃止"),
    ("施設", "介護医療院", "第9期計画に記載なし。実績データにも計上なし"),
]
# 見込量表がないが実績のあるサービスの実績
NO_PLAN_JISSEKI = {
    "小規模多機能型居宅介護・介護予防小規模多機能型居宅介護": [
        ("延利用人数", "人／年", ("04_サービス利用給付", 15, 9)),
        ("給付費", "千円／年", ("04_サービス利用給付", 16, 9))],
    "地域密着型通所介護（小規模デイ）": [
        ("延利用人数", "人／年", ("04_サービス利用給付", 17, 9)),
        ("給付費", "千円／年", ("04_サービス利用給付", 18, 9))],
}


def style_header(ws, row=4):
    for c, (name, width) in enumerate(COLS, start=1):
        cell = ws.cell(row, c, name)
        cell.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = f"C{row + 1}"


def title(ws, text, sub):
    ws.cell(1, 1, text).font = Font(name="游ゴシック", size=13, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(COLS))
    ws.cell(1, 1).alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 24
    ws.cell(2, 1, sub).font = Font(name="游ゴシック", size=9, color="595959")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(COLS))


def put_row(ws, r, kubun, name, shihyo, tani, plan3, jis3, biko, shutten):
    """1行を書き出す。plan3/jis3 は [R6,R7,R8]（None可）。"""
    vals = [kubun, name, shihyo, tani]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name="游ゴシック", size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 2))
        cell.border = BOX
    for i in range(3):
        cp, cj, ca = 5 + i * 3, 6 + i * 3, 7 + i * 3
        p = ws.cell(r, cp, plan3[i] if plan3 else None)
        p.font = Font(name="游ゴシック", size=9)
        p.fill = PatternFill("solid", fgColor=PLAN)
        p.number_format = "#,##0"
        p.border = BOX
        j = ws.cell(r, cj, jis3[i] if jis3 else None)
        j.font = Font(name="游ゴシック", size=9, color="0000FF")
        j.fill = PatternFill("solid", fgColor=INPUT_)
        j.number_format = "#,##0"
        j.border = BOX
        a = ws.cell(r, ca)
        a.value = (f"=IF(OR({get_column_letter(cp)}{r}=\"\","
                   f"{get_column_letter(cp)}{r}=0,{get_column_letter(cj)}{r}=\"\"),\"\","
                   f"{get_column_letter(cj)}{r}/{get_column_letter(cp)}{r})")
        a.font = Font(name="游ゴシック", size=9)
        a.number_format = "0.0%"
        a.alignment = Alignment(horizontal="center")
        a.border = BOX
    b = ws.cell(r, 14, biko)
    b.font = Font(name="游ゴシック", size=9, color="0000FF")
    b.fill = PatternFill("solid", fgColor=INPUT_)
    b.alignment = Alignment(vertical="center", wrap_text=True)
    b.border = BOX
    s = ws.cell(r, 15, shutten)
    s.font = Font(name="游ゴシック", size=8, color="595959")
    s.alignment = Alignment(vertical="center", wrap_text=True)
    s.border = BOX


def main():
    plan = extract_plan()
    wb_j, take = load_jisseki()
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------- 00 使い方
    ws = wb.active
    ws.title = "00_使い方"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 92
    title(ws, "川崎町　サービス別 計画値・実績値 対比表", "")
    rows = [
        ("", ""),
        ("1．この表の目的", ""),
        ("", "第9期計画（令和6年3月）に定めた各サービス・事業の見込量（計画値）と、実績値を対比する表です。"),
        ("", "仕様書6（4）の現行施策の検証に用いるとともに、保険者機能強化推進交付金の評価指標"),
        ("", "目標Ⅰ-2「介護保険事業計画の進捗状況（計画値と実績値の乖離状況）を分析しているか」に対応します。"),
        ("", "同指標は、単に数値を把握しただけでは非該当とされ、乖離の要因の考察までが求められます。"),
        ("", "このため N列「乖離の要因・対応」の記入が評価上も重要です。"),
        ("", ""),
        ("2．色の意味", ""),
        ("緑（計画値）", "第9期計画書から転記した確定値です。編集しないでください。"),
        ("黄（記入欄）", "町にご記入いただく欄です。青字で入力してください。"),
        ("白（達成率）", "実績÷計画で自動計算されます。"),
        ("", ""),
        ("3．シート構成", ""),
        ("01_居宅サービス", "訪問・通所・短期入所・福祉用具・住宅改修・居宅介護支援（介護予防を含む）"),
        ("02_地域密着型サービス", "認知症対応型共同生活介護ほか。第9期に見込量表がないサービスも一覧に含めています。"),
        ("03_施設サービス", "介護老人福祉施設・介護老人保健施設ほか"),
        ("04_総合事業・一般介護予防", "訪問型／通所型サービス、介護予防教室、地域リハビリテーション活動支援"),
        ("05_包括的支援事業・任意事業", "認知症関係、介護用品の支給、配食サービス、見守り"),
        ("06_高齢者福祉サービス", "移送・会食・緊急通報・老人クラブ・敬老会"),
        ("07_記入状況", "各シートの記入状況を自動集計します。"),
        ("", ""),
        ("4．実績を記入済みの箇所", ""),
        ("", "令和8年9月1日にご提供いただいた実績データから転記できたものは、あらかじめ記入しています。"),
        ("", "ただし次のものは実績として扱っていません。"),
        ("", "・02シート（総合事業・介護予防）の教室関係は、出典欄が「第9期計画」であり計画値が転記されているため"),
        ("", "　（確認事項No.42）。空欄のままとしています。"),
        ("", "・12〜14シート（サービス見込量）の実績欄はほぼ全面的に未記入のため（確認事項No.40）。"),
        ("", ""),
        ("5．留意点", ""),
        ("", "・計画値の「延利用人数」と実績データの「利用者数・利用量」は定義が一致するか確認が必要です。"),
        ("", "・実績データのR7は見込、R8は当初・見込であり、確定値ではありません（仮試算扱い）。"),
        ("", "・確定値をいただき次第、本表を更新します。"),
    ]
    for i, (a, b) in enumerate(rows, start=4):
        c1 = ws.cell(i, 2, a)
        c1.font = Font(name="游ゴシック", size=10,
                       bold=a.endswith("目的") or a.startswith(("1．", "2．", "3．", "4．", "5．")),
                       color=NAVY if a.startswith(("1．", "2．", "3．", "4．", "5．")) else "000000")
        c2 = ws.cell(i, 3, b)
        c2.font = Font(name="游ゴシック", size=10)
        c2.alignment = Alignment(wrap_text=True, vertical="center")
    for a, fill in (("緑（計画値）", PLAN), ("黄（記入欄）", INPUT_)):
        for i in range(4, 4 + len(rows)):
            if ws.cell(i, 2).value == a:
                ws.cell(i, 2).fill = PatternFill("solid", fgColor=fill)

    # ---------------------------------------------------------- 各シート
    used = set()
    counts = {}
    for sheet_name, key, kubun in SHEETS:
        ws = wb.create_sheet(sheet_name)
        title(ws, f"サービス別 計画値・実績値 対比表　{sheet_name[3:]}",
              "計画値：第9期計画書（令和6年3月）／実績：町提供実績データ（令和8年9月1日受領・仮試算）")
        style_header(ws)
        r = 5
        n_plan = n_fill = 0
        for blk in plan:
            nm = blk["name"]
            base = re.sub(r"(の)?見込(み)?量", "", nm).strip()
            if base not in ASSIGN[key]:
                continue
            if nm in used:
                continue
            used.add(nm)
            jm = JMAP.get(nm, {})
            for lab, tani, vals in blk["rows"]:
                if lab.startswith(base):          # 「介護予防フェスティバル 参加人数」等
                    lab = lab[len(base):].strip() or lab
                plan3 = [num(v) for v in vals[1:4]]      # 計画書はR5〜R8。R6〜R8を取る
                ref = jm.get(lab)
                jis3 = None
                shutten = "第9期計画書"
                if ref:
                    jis3 = [v if isinstance(v, (int, float)) else None
                            for v in take(*ref)]
                    shutten = f"計画書／実績データ {ref[0][:2]}シート"
                    n_fill += 1
                n_plan += 1
                put_row(ws, r, kubun, base, lab, tani, plan3, jis3,
                        NOTES.get((base, lab)), shutten)
                r += 1
        # 介護予防教室（実績データ02シートの事業名で起こす）
        if key == "04":
            ws_k = wb_j["02_総合事業_介護予防"]
            for nm, lab, tani, row in KYOSHITSU:
                plan3 = [ws_k.cell(row, 8 + i).value for i in range(3)]
                plan3 = [v if isinstance(v, (int, float)) else None for v in plan3]
                put_row(ws, r, kubun, nm, lab, tani, plan3, None,
                        "実績データ02シートの出典欄は「第9期計画」であり計画値。"
                        "実績のご記入が必要（確認事項No.42）",
                        "実績データ 02シート（計画値）")
                n_plan += 1
                r += 1

        # 見込量表がないサービス
        for k, nm, note in NO_PLAN:
            if k != kubun:
                continue
            extra = NO_PLAN_JISSEKI.get(nm)
            if extra:
                for lab, tani, ref in extra:
                    jis3 = [v if isinstance(v, (int, float)) else None for v in take(*ref)]
                    put_row(ws, r, kubun, nm, lab, tani, None, jis3, note,
                            "実績データ 04シート")
                    n_plan += 1
                    n_fill += 1
                    r += 1
            else:
                put_row(ws, r, kubun, nm, "―", "―", None, None, note, "第9期計画書")
                r += 1
        counts[sheet_name] = (n_plan, n_fill, r - 5)
        ws.auto_filter.ref = f"A4:{get_column_letter(len(COLS))}{r - 1}"
        # 達成率が80％未満・120％超の欄を目立たせる
        if r > 5:
            from openpyxl.formatting.rule import CellIsRule
            warn = PatternFill("solid", fgColor="FCE4E4")
            red = Font(name="游ゴシック", size=9, color="C00000", bold=True)
            for col in ("G", "J", "M"):
                rng = f"{col}5:{col}{r - 1}"
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="lessThan", formula=["0.8"],
                                    fill=warn, font=red))
                ws.conditional_formatting.add(
                    rng, CellIsRule(operator="greaterThan", formula=["1.2"],
                                    fill=warn, font=red))

    # ---------------------------------------------------------- 07 記入状況
    ws = wb.create_sheet("07_記入状況")
    title(ws, "記入状況", "実績欄（F・I・L列）の記入済み件数を自動集計します。")
    hdr = ["シート", "行数", "計画値のある行", "実績を記入済み（当社転記分）", "R6実績の記入済み",
           "R7実績の記入済み", "R8見込の記入済み"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(4, c, h)
        cell.font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(c)].width = 22 if c == 1 else 16
    ws.row_dimensions[4].height = 30
    r = 5
    for sheet_name, _, _ in SHEETS:
        n_plan, n_fill, n_rows = counts[sheet_name]
        last = 4 + n_rows
        ws.cell(r, 1, sheet_name).font = Font(name="游ゴシック", size=9)
        ws.cell(r, 2, n_rows).font = Font(name="游ゴシック", size=9)
        ws.cell(r, 3, n_plan).font = Font(name="游ゴシック", size=9)
        ws.cell(r, 4, n_fill).font = Font(name="游ゴシック", size=9)
        for i, col in enumerate("FIL"):
            c = ws.cell(r, 5 + i, f"=COUNT('{sheet_name}'!{col}5:{col}{last})")
            c.font = Font(name="游ゴシック", size=9)
        for c in range(1, 8):
            ws.cell(r, c).border = BOX
        r += 1
    ws.cell(r + 1, 1, "※「実績を記入済み（当社転記分）」は、令和8年9月1日提供の実績データから"
                      "転記できた指標の数です。").font = Font(name="游ゴシック", size=9, color="595959")
    ws.cell(r + 2, 1, "※ 空欄の実績欄は、町にご記入いただく必要があります"
                      "（確認事項No.40・No.42）。").font = Font(name="游ゴシック", size=9, color="C00000")

    wb.save(OUT)
    print(f"保存：{OUT}")
    for k, (a, b, c) in counts.items():
        print(f"  {k:28s} 行数{c:>4} ／ 計画値のある行{a:>4} ／ 実績を転記{b:>4}")


if __name__ == "__main__":
    main()
