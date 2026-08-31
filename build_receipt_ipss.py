# -*- coding: utf-8 -*-
"""社人研「日本の地域別将来推計人口」結果表3の受領点検.

令和8年8月31日に受領した4ファイル（結果表3-1〜3-4）の点検結果。
確認事項No.62（社人研推計の町別・男女5歳階級別表の入手）に対する
ご対応であり、これにより計画素案 第2章第1節の町別75歳以上人口を
公表値により算定できるようになった。

従来は、令和2年国勢調査の町別75歳以上を起点に3町計の後期高齢者比の
推移で按分していた。按分値と公表値の差は最大122人で、
町別の補正方向が4件変わった。3町計の値と補正係数は変わらない。

シート構成
  00_受領資料の概要    4ファイルの内容と、この資料で確定すること
  01_3町の年齢別人口割合  結果表3-1〜3-4の3町の値
  02_按分値との差      従来の按分値と公表値の対比
  03_補正方向の変化     町別の補正方向が変わった4件
  04_上川管内との比較    21保険者の構成市町村における位置
  05_成果品への反映     更新した箇所と、なお残ること
"""

import io
import runpy
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_ipss_town as IP

OUT = ("/home/user/repository/output/"
       "第10期計画_社人研推計の町別データの受領点検.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F4E78", "5B9BD5"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# build_pop_adjust から町別総人口・公表ベースの値・従来の按分値を取り出す
_buf, _old = io.StringIO(), sys.stdout
sys.stdout = _buf
G = runpy.run_path("build_pop_adjust.py")
sys.stdout = _old
POP16, K65, K75, ANBUN = G["POP16"], G["K65"], G["K75"], G["ANBUN"]
TOWNS, YS, CMP = G["TOWNS"], G["YS"], G["CMP"]

WA = {2020: "令和2年", 2025: "令和7年", 2030: "令和12年", 2035: "令和17年",
      2040: "令和22年", 2045: "令和27年", 2050: "令和32年"}
BANDS = [("0-14", "0〜14歳"), ("15-64", "15〜64歳"),
         ("65+", "65歳以上"), ("75+", "75歳以上")]


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 56
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center",
                                vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=22, align=None, bold=False,
         numfmt=None):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.border = BORDER
        ha = (align or {}).get(i, "left" if isinstance(v, str) else "right")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if numfmt and not isinstance(v, str):
            c.number_format = numfmt
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8, height=88):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ============================================================ 00
ws = sheet("00_受領資料の概要",
           "社人研「日本の地域別将来推計人口」結果表3の受領点検",
           "令和8年8月31日に受領した4ファイルの点検結果です。"
           "国立社会保障・人口問題研究所「日本の地域別将来推計人口"
           "（令和5（2023）年推計）」の結果表3（年齢別人口割合）で、"
           "市区町村ごとの総人口に占める年齢区分別人口の割合が"
           "令和2年から令和32年まで5年ごとに収録されています。"
           "確認事項No.62に対するご対応です。",
           [4, 24, 30, 44, 30, 12], freeze="A5")

r = lead(ws, 4, "【1　受領した4ファイル】", 6)
r = header(ws, r, ["No.", "ファイル", "内容", "点検した結果", "扱い", "判定"])
for no, fn, cont, res, atsu, hh in [
    (1, "kekkahyo3_1.xlsx", "結果表3-1　年齢別人口割合　0〜14歳人口",
     "全国の市区町村・都道府県1,951行が収録されています。"
     "北海道と上川管内21保険者の構成市町村の値を確認しました。",
     "収録。第2章第1節の参考", "受領"),
    (2, "kekkahyo3_2.xlsx", "結果表3-2　年齢別人口割合　15〜64歳人口",
     "同上。担い手（生産年齢人口）の割合です。",
     "収録。第2章第1節の担い手の推移に用います", "受領"),
    (3, "kekkahyo3_3.xlsx", "結果表3-3　年齢別人口割合　65歳以上人口",
     "同上。町別高齢化率です。"
     "従来用いていた小数第1位の値（東川町令和2年33.2％等）と"
     "小数第6位まで一致します。",
     "採用。町別65歳以上人口の算定に用います", "受領"),
    (4, "kekkahyo3_4.xlsx", "結果表3-4　年齢別人口割合　75歳以上人口",
     "同上。これが最も重要です。"
     "従来は町別の公表値がなく、按分により推計していました。",
     "採用。町別75歳以上人口の算定に用います", "受領"),
]:
    r = body(ws, r, [no, fn, cont, res, atsu, hh], height=54,
             fills={6: OK_G}, align={1: "center", 6: "center"})

r += 1
r = lead(ws, r, "【2　この資料により確定すること】", 6)
r = header(ws, r, ["No.", "確定する事項", "従来", "確定した内容", "反映先", ""])
for no, ko, jurai, naiyo, han in [
    (1, "町別の75歳以上人口",
     "令和2年国勢調査の町別75歳以上を起点に、"
     "3町計の後期高齢者比の推移で按分した推計値",
     "結果表3-4の町別割合に町別総人口を乗じた公表ベースの値。"
     "按分値との差は最大122人（東川町・令和12年）です。",
     "計画素案 第2章第1節\n人口推計の補正 05シート"),
    (2, "町別の補正方向",
     "按分値による判定（暫定）",
     "公表値による判定。4件が変わりました（03シート）。",
     "人口推計の補正 05シート"),
    (3, "自己点検の指摘No.5（重大）",
     "「町別の判定は按分値に依存している」として"
     "重大の指摘としていた",
     "按分をやめたことにより解消しました。",
     "人口推計の補正 07シート"),
    (4, "計画素案の［要確認］2件",
     "「町別の補正の向きは、社人研の市区町村別・"
     "男女5歳階級別表の取得までは暫定」と本文に付していた",
     "公表値に置き換えたため外せます。",
     "計画素案 第2章第1節"),
    (5, "上川管内における位置",
     "町別の75歳以上割合を管内と比較できなかった",
     "21保険者の構成市町村23団体との比較ができます（04シート）。",
     "計画素案 第2章第1節\n3町別の論点整理"),
]:
    r = body(ws, r, [no, ko, jurai, naiyo, han, ""], height=56,
             align={1: "center"})

note(ws, r + 1,
     "注1）結果表3は割合（％）であり実数ではありません。"
     "実数は町別の総人口に割合を乗じて求めます。"
     "町別の総人口は既に収録している社人研の値を用いています。\n"
     "注2）3町計の値と補正係数（65〜74歳1.0377・75歳以上0.9927）は"
     "変わりません。保険料の試算にも影響しません。"
     "変わるのは町別の内訳と補正方向です。\n"
     "注3）男女別・5歳階級別の表ではありませんが、"
     "本業務で必要なのは65〜74歳と75歳以上の切り分けであるため、"
     "結果表3で足ります。確認事項No.62は完了として扱います。", 6, 76)


# ============================================================ 01
ws = sheet("01_3町の年齢別人口割合",
           "3町の年齢別人口割合（結果表3-1〜3-4）",
           "総人口に占める年齢区分別人口の割合（％）です。"
           "令和2年は令和2年国勢調査による実績、令和7年以降が推計です。",
           [16, 14, 11, 11, 11, 11, 11, 11, 11], freeze="C5")

r = lead(ws, 4, "【1　年齢区分別の割合（％）】", 9)
r = header(ws, r, ["町", "区分"] + [WA[y] for y in YS])
for t in TOWNS:
    for b, nm in BANDS:
        v = [IP.rate(t, b, y) for y in YS]
        r = body(ws, r, [t if b == "0-14" else "", nm]
                 + [round(x, 1) for x in v],
                 {2: IN_Y if b == "75+" else None}, height=18,
                 numfmt="0.0")

r += 1
r = lead(ws, r, "【2　割合に町別総人口を乗じた実数（人）】", 9)
r = header(ws, r, ["町", "区分"] + [WA[y] for y in YS])
for t in TOWNS:
    r = body(ws, r, [t, "総人口"] + [POP16[t][y] for y in YS],
             {2: GRAY}, height=18, numfmt="#,##0")
    r = body(ws, r, ["", "65歳以上"] + [round(K65[t][y]) for y in YS],
             height=18, numfmt="#,##0")
    r = body(ws, r, ["", "　うち75歳以上"]
             + [round(K75[t][y]) for y in YS],
             {2: IN_Y}, height=18, numfmt="#,##0")
    r = body(ws, r, ["", "　うち65〜74歳"]
             + [round(K65[t][y] - K75[t][y]) for y in YS],
             height=18, numfmt="#,##0")
r = body(ws, r, ["3町計", "65歳以上"]
         + [round(sum(K65[t][y] for t in TOWNS)) for y in YS],
         {1: MID_B, 2: MID_B}, bold=True, height=18, numfmt="#,##0")
r = body(ws, r, ["", "　うち75歳以上"]
         + [round(sum(K75[t][y] for t in TOWNS)) for y in YS],
         {2: MID_B}, bold=True, height=18, numfmt="#,##0")

note(ws, r + 1,
     "注1）町別の総人口は社人研の値です（既に収録しているもの）。"
     "結果表3は割合のみで、実数は収録されていません。\n"
     "注2）3町計の75歳以上は、見える化システムのA3系列"
     "（社人研推計の後期高齢者数）と令和7年以降のすべての時点で"
     "一致します（差0人）。令和2年のみ13人多くなります。"
     "国勢調査の総人口の取り方の違いによるものです。\n"
     "注3）65〜74歳は65歳以上から75歳以上を差し引いて求めています。", 9, 68)


# ============================================================ 02
ws = sheet("02_按分値との差",
           "従来の按分値と公表値の対比（75歳以上）",
           "令和8年8月30日までは、令和2年国勢調査の町別75歳以上を起点に、"
           "3町計の後期高齢者比の推移で按分していました。"
           "その按分値と、今回受領した公表値との差です。",
           [14, 14, 14, 14, 12, 14, 40], freeze="A5")

r = lead(ws, 4, "【1　町別・時点別の差（75歳以上・人）】", 7)
r = header(ws, r, ["町", "時点", "従来の按分値", "公表ベースの値", "差",
                   "差の率", "所見"])
mx = (0, None)
for t in TOWNS:
    for y in YS:
        a, b = ANBUN[t][y], K75[t][y]
        d = b - a
        if abs(d) > abs(mx[0]):
            mx = (d, (t, y))
        sho = ""
        if y == 2025:
            sho = "補正方向の判定に用いる時点"
        r = body(ws, r, [t if y == 2020 else "", WA[y], round(a), round(b),
                         round(d), "%+.1f％" % (d / a * 100), sho],
                 {5: (NG_O if abs(d) >= 50 else
                      IN_Y if abs(d) >= 20 else None)},
                 height=18, numfmt="#,##0")

r += 1
r = lead(ws, r, "【2　3町計での検算】", 7)
r = header(ws, r, ["時点", "按分値の3町計", "公表ベースの3町計",
                   "見える化A3", "差", "", "判定"])
import data_population as P                                   # noqa: E402
IX = G["IX"]
for y in YS:
    a = sum(ANBUN[t][y] for t in TOWNS)
    b = sum(K75[t][y] for t in TOWNS)
    m = P.A3["後期高齢者数"][IX[str(y)]]
    r = body(ws, r, [WA[y], round(a), round(b), round(m), round(b - m), "",
                     "一致" if abs(b - m) < 1 else "差あり"],
             {7: OK_G if abs(b - m) < 1 else IN_Y}, height=18,
             numfmt="#,##0", align={7: "center"})

note(ws, r + 1,
     "注1）按分は3町計が見える化A3に一致するよう調整していたため、"
     "3町計では按分値も公表ベースの値も同じです。"
     "町別の内訳のみが異なります。\n"
     "注2）差が最も大きいのは%s（%s）の%+d人です。\n"
     "注3）令和2年に13人の差が生じるのは、"
     "町別総人口と見える化A2の総人口とで"
     "国勢調査の値の取り方が異なるためです。"
     "令和7年以降は差0人で一致します。"
     % (mx[1][0], WA[mx[1][1]], round(mx[0])), 7, 68)


# ============================================================ 03
ws = sheet("03_補正方向の変化",
           "町別の補正方向の変化",
           "公表値に置き換えたことにより、町別の補正方向が4件変わりました。"
           "3町計の補正係数（65〜74歳1.0377・75歳以上0.9927）は"
           "変わりません。",
           [14, 14, 20, 20, 20, 34], freeze="A5")

r = lead(ws, 4, "【1　変わった4件】", 6)
r = header(ws, r, ["町", "区分", "従来（按分値）", "公表値", "変化",
                   "内容"])
KAWATTA = [
    ("東川町", "75歳以上", "上方に補正（3年で＋3.2％）", "補正しない",
     "反転",
     "社人研の令和7年の75歳以上が1,632人から1,731人に変わり、"
     "住民基本台帳の実績1,637人との関係が逆転しました。"
     "年平均の差は▲1.08ポイントから＋0.13ポイントになり、"
     "判定基準（±0.20ポイント）の内側に入りました。"),
    ("東川町", "65〜74歳", "補正しない", "上方に補正（3年で＋6.3％）",
     "新たに補正",
     "65歳以上は変わらず75歳以上が増えたため、"
     "差し引きの65〜74歳が減りました。"
     "社人研の令和7年が1,070人から972人になり、"
     "住民基本台帳の1,085人との差が広がりました。"),
    ("美瑛町", "75歳以上", "補正しない", "上方に補正（3年で＋0.9％）",
     "新たに補正",
     "社人研の令和7年の75歳以上が2,230人から2,176人に変わり、"
     "年平均の差が＋0.20ポイントから▲0.29ポイントになりました。"),
    ("東神楽町", "65歳以上", "下方に補正", "補正しない", "反転",
     "65歳以上は内訳の合計であり、補正の対象ではありません。"
     "参考として掲げている行の判定が変わったものです。"),
]
for t, b, jurai, kohyo, henka, naiyo in KAWATTA:
    r = body(ws, r, [t, b, jurai, kohyo, henka, naiyo], height=56,
             fills={5: NG_O if henka == "反転" else IN_Y})

r += 1
r = lead(ws, r, "【2　方向は変わらないが幅が変わった3件】", 6)
r = header(ws, r, ["町", "区分", "従来（按分値）", "公表値", "変化", "内容"])
for t, b, jurai, kohyo, naiyo in [
    ("美瑛町", "65〜74歳", "上方に補正（3年で＋6.2％）",
     "上方に補正（3年で＋3.6％）",
     "補正の幅が約6割に縮みました。"),
    ("東神楽町", "65〜74歳", "上方に補正（3年で＋4.2％）",
     "上方に補正（3年で＋2.3％）",
     "同上。"),
    ("東神楽町", "75歳以上", "下方に補正（3年で▲4.2％）",
     "下方に補正（3年で▲2.2％）",
     "社人研の令和7年が1,839人から1,794人に変わり、"
     "住民基本台帳1,552人との差が287人から242人に縮みました。"),
]:
    r = body(ws, r, [t, b, jurai, kohyo, "幅が縮小", naiyo], height=34,
             fills={5: GRAY})

r += 1
r = lead(ws, r, "【3　変わらないこと】", 6)
r = header(ws, r, ["No.", "項目", "値", "", "", "内容"])
for no, ko, atai, naiyo in [
    (1, "3町計の補正係数（65〜74歳）", "1.0377",
     "3町計は見える化A2・A3を用いており、町別の内訳によりません"),
    (2, "3町計の補正係数（75歳以上）", "0.9927", "同上"),
    (3, "令和11年度の65歳以上人口", "9,072人→9,159人", "同上"),
    (4, "令和11年度の認定者数", "2,004人→1,996人", "同上"),
    (5, "保険料基準額への影響", "6,238円→約6,153円", "同上"),
]:
    r = body(ws, r, [no, ko, atai, "", "", naiyo], height=22,
             align={1: "center"})

note(ws, r + 1,
     "注1）補正方向は、年平均の伸び率の差（社人研－住民基本台帳）が"
     "±0.20ポイント未満のときは「補正しない」、"
     "社人研が速いときは「下方に補正」、"
     "住民基本台帳が速いときは「上方に補正」としています"
     "（人口推計の補正 07シートの判定ルール）。\n"
     "注2）町別の補正係数は算定していますが、適用は保留しています。"
     "町別の基準人口（65〜74歳・75〜84歳・85歳以上の第1号被保険者数）が"
     "未受領のためです。\n"
     "注3）第9期計画の町別の調整との対照（人口推計の補正 05シート【4】）は、"
     "本受領により判定の根拠が公表値に変わりました。", 6, 68)


# ============================================================ 04
ws = sheet("04_上川管内との比較",
           "上川管内における3町の位置（75歳以上人口割合）",
           "上川管内21保険者の構成市町村23団体について、"
           "結果表3-4の75歳以上人口割合を並べたものです。"
           "保険料の管内比較（計画素案 第6章第6節1）と同じ21保険者の"
           "構成市町村です。",
           [16, 12, 12, 12, 12, 12, 12, 34], freeze="A5")

KANNAI = [k for k in IP.RATE if k != "北海道"]
r = lead(ws, 4, "【1　75歳以上人口割合（％）】", 8)
r = header(ws, r, ["市町村", WA[2020], WA[2025], WA[2030], WA[2040],
                   WA[2050], "令和2→32年\nの上昇", "所見"])
rows = sorted(KANNAI, key=lambda k: -IP.rate(k, "75+", 2025))
for k in rows:
    v = [IP.rate(k, "75+", y) for y in (2020, 2025, 2030, 2040, 2050)]
    sho = ""
    if k in TOWNS:
        sho = "大雪地区広域連合の構成町"
    r = body(ws, r, [k] + [round(x, 1) for x in v]
             + [round(v[-1] - v[0], 1), sho],
             {1: IN_Y if k in TOWNS else None}, height=18, numfmt="0.0")
v = [IP.rate("北海道", "75+", y) for y in (2020, 2025, 2030, 2040, 2050)]
r = body(ws, r, ["北海道"] + [round(x, 1) for x in v]
         + [round(v[-1] - v[0], 1), "参考"], {1: MID_B}, bold=True,
         height=18, numfmt="0.0")

r += 1
r = lead(ws, r, "【2　3町の位置】", 8)
r = header(ws, r, ["町", "令和7年の割合", "23団体中の順位", "北海道との差",
                   "", "", "", "所見"])
for t in TOWNS:
    v = IP.rate(t, "75+", 2025)
    rank = sorted(KANNAI, key=lambda k: -IP.rate(k, "75+", 2025)).index(t) + 1
    hk = IP.rate("北海道", "75+", 2025)
    sho = {"東川町": "管内では低い方に位置する",
           "美瑛町": "3町で最も高い",
           "東神楽町": "管内で最も低い水準にある"}[t]
    r = body(ws, r, [t, round(v, 1), "%d位" % rank,
                     "%+.1fpt" % (v - hk), "", "", "", sho],
             height=20, numfmt="0.0", align={3: "center"})

note(ws, r + 1,
     "注1）順位は23団体を割合の高い順に並べたものです。"
     "1位が最も高齢化が進んでいます。\n"
     "注2）大雪地区広域連合は1保険者ですが、"
     "本表は構成市町村単位の推計であるため3町を分けて掲げています。\n"
     "注3）保険料の管内比較（計画素案 第6章第6節1）は"
     "21保険者の単位です。本表の23団体とは単位が異なります。", 8, 56)


# ============================================================ 05
ws = sheet("05_成果品への反映",
           "更新した箇所と、なお残ること",
           "本受領により更新した成果品と、なお残る事項です。",
           [4, 30, 44, 26, 14, 12], freeze="A5")

r = lead(ws, 4, "【1　更新した箇所】", 6)
r = header(ws, r, ["No.", "成果品・箇所", "更新した内容", "根拠", "時期", ""])
for no, saki, naiyo, kon, ji in [
    (1, "人口推計の補正 05シート",
     "社人研側の町別65歳以上・75歳以上を、按分値から"
     "結果表3-3・3-4の公表値に置き換えた。"
     "町別の補正方向が4件変わった。",
     "結果表3-3・3-4", "令和8年8月31日"),
    (2, "人口推計の補正 07シート",
     "自己点検の指摘No.5（重大・町別の判定が按分値に依存）を"
     "「解消」に改めた。",
     "同上", "令和8年8月31日"),
    (3, "計画素案 第2章第1節",
     "町別75歳以上が按分による推計値である旨の［要確認］2件を外した。",
     "同上", "令和8年8月31日"),
    (4, "data_ipss_town.py（新規）",
     "北海道と上川管内23団体の年齢別人口割合（4区分×7時点）を収録した。",
     "同上", "令和8年8月31日"),
    (5, "業務工程管理表 確認事項No.62",
     "「完了」に改めた。",
     "同上", "令和8年8月31日"),
]:
    r = body(ws, r, [no, saki, naiyo, kon, ji, ""], height=44,
             align={1: "center"})

r += 1
r = lead(ws, r, "【2　なお残ること】", 6)
r = header(ws, r, ["No.", "事項", "内容", "解消の条件", "優先", ""])
for no, ko, naiyo, joken, yu in [
    (1, "町別の補正の適用は保留のまま",
     "町別の補正係数は算定できるが、当方の推計は基準人口"
     "（第1号被保険者数）に伸び率を掛ける構造であるため、"
     "町別に適用するには町別の基準人口が必要である。",
     "町別の65〜74歳・75〜84歳・85歳以上の第1号被保険者数"
     "（令和8年3月末）の受領", "中"),
    (2, "町別総人口は既収録の値による",
     "結果表3は割合のみで実数を収録していない。"
     "町別総人口は既に収録している社人研の値を用いている。",
     "結果表1（総人口）の受領。ただし現在の値で"
     "3町計が見える化A2・A3と一致しているため急がない", "低"),
    (3, "男女別・5歳階級別の内訳はない",
     "結果表3は4区分（0〜14歳・15〜64歳・65歳以上・75歳以上）である。"
     "85歳以上の町別の内訳は得られない。",
     "本業務では65〜74歳と75歳以上の切り分けで足りるため、"
     "必要になった時点で改めて依頼する", "低"),
]:
    r = body(ws, r, [no, ko, naiyo, joken, yu, ""], height=48,
             align={1: "center", 5: "center"},
             fills={5: IN_Y if yu == "中" else GRAY})

note(ws, r + 1,
     "注1）確認事項No.62は完了として扱います。"
     "受託者の作業環境からは外部サイトへ接続できないため、"
     "発注者にご取得いただいたものです。\n"
     "注2）本受領は町別の内訳の精度を上げるものであり、"
     "3町計の推計・保険料の試算には影響しません。\n"
     "注3）今後、社人研の推計に関する資料が必要となった場合も、"
     "同じ形でのご提供をお願いすることになります。", 6, 56)


wb.save(OUT)
print("saved:", OUT)
for ws in wb.worksheets:
    print("  -", ws.title, ws.max_row, "rows")
print()
print("按分値との最大差 %s %s %+d人" % (mx[1][0], WA[mx[1][1]], round(mx[0])))
print("補正方向が変わった 4件／幅が変わった 3件")
