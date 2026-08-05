# -*- coding: utf-8 -*-
"""
北塩原村 第8期障がい福祉計画・第4期障がい児福祉計画
圏域サービス資源（事業所数）管理ブック ジェネレータ

入力: source/障害福祉資源/ReMHRAD_障害福祉資源_*_2020-2025.xlsx
      ReMHRAD（地域精神保健福祉資源分析データベース）の障害福祉資源タブから
      サービスごとに市町村別事業所数（2020年→2025年）をダウンロードしたもの。

出力: output/北塩原村_圏域サービス資源.xlsx

用途
- 基本指針 第三の二の2(二)「各市町村において事業を実施する事業所を最低一箇所
  確保できるよう努める」に対する本村の現状把握
- 同 第一の一の3「中山間・人口減少地域における提供体制の維持・確保」の裏づけ
- サービス見込量ブック 08_供給制約 の入力材料
- 骨子案修正版 第5章8（必要な見込量の確保のための方策）の根拠

注意
- 出典は「障害福祉サービス等情報公表システム（障害種を問わない全施設）」2025年度分。
  同システムに公表されていない事業所は計上されない。第8期基本指針が公表率・更新率
  100%を成果目標に新設したのは、この網羅性の問題に対処するためである（第二の八）。
- 事業所数は「所在地の市町村」でカウントされる。本村の住民が圏域外の事業所を
  利用している分は、この表には現れない。利用可能量の把握には別途照会が必要。
"""

import os
import re
import warnings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from kitashiobara_common import (
    COLORS, FONT, OUT_DIR, add_sheet, ensure_out_dir, style_data_cell,
    style_header_row, style_note, style_title, write_row,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

SRC_DIR = "source/障害福祉資源"
OUT_FILE = f"{OUT_DIR}/北塩原村_圏域サービス資源.xlsx"

VILLAGE = "北塩原村"
# 会津北部圏域（地域生活支援拠点等を共同整備する4町村）
KEN_HOKUBU = ["北塩原村", "猪苗代町", "磐梯町", "湯川村"]
# 実際に利用実績のある近隣市を含めた広域
AIZU_WIDE = KEN_HOKUBU + ["会津若松市", "喜多方市", "会津坂下町", "西会津町"]

# 掲載順（訪問系→日中活動系→居住系）
SERVICE_ORDER = [
    "居宅介護", "重度訪問介護", "同行援護", "行動援護", "重度障害者等包括支援",
    "生活介護", "自立訓練（機能訓練）", "自立訓練（生活訓練）", "宿泊型自立訓練",
    "就労移行支援", "就労継続支援（A型）", "就労継続支援（B型）", "就労定着支援",
    "療養介護", "短期入所",
    "自立生活援助", "共同生活援助", "施設入所支援",
]

INT = "#,##0"
PCT1 = "+0.0%;-0.0%;0.0%"


def load_all(src_dir=SRC_DIR):
    """サービス名 -> {市町村: (2020年事業所数, 2025年事業所数)} を返す。"""
    data = {}
    if not os.path.isdir(src_dir):
        raise SystemExit(f"入力ディレクトリがありません: {src_dir}")
    for name in sorted(os.listdir(src_dir)):
        if not name.endswith(".xlsx"):
            continue
        wb = load_workbook(os.path.join(src_dir, name), data_only=True, read_only=True)
        title = wb.sheetnames[0]
        m = re.search(r"（(.+?)）の状況", title)
        svc = (m.group(1) if m else title).replace("(", "（").replace(")", "）")
        ws = wb[title]
        vals = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            try:
                vals[str(r[0])] = (int(r[2]), int(r[4]))
            except (TypeError, ValueError):
                continue
        wb.close()
        data[svc] = vals
    return data


def sheet_overview(wb, data):
    ws = add_sheet(
        wb, "00_概要", "会津北部圏域の障がい福祉サービス資源（事業所数）",
        "ReMHRAD 障害福祉資源による市町村別事業所数の2020年から2025年への推移。"
        "出典は障害福祉サービス等情報公表システム（障害種を問わない全施設）2025年度分。"
        "事業所数は所在地の市町村でカウントされるため、本村の住民が圏域外の事業所を"
        "利用している分はこの表には現れない。",
        [26, 20, 20, 66])
    r = 5
    style_header_row(ws, r, ["シート", "対象", "対応する基本指針の規定", "内容"])
    r += 1
    rows = [
        ("01_会津北部圏域", "4町村×18サービス", "第三の二の2(二)",
         "地域生活支援拠点等を共同整備する猪苗代町・磐梯町・湯川村・北塩原村の事業所数"),
        ("02_広域", "8市町村×18サービス", "第三の二の2(四)",
         "会津若松市・喜多方市・会津坂下町・西会津町を加えた広域。実際の利用先を含む"),
        ("03_村の状況", "本村18サービス", "第三の二の2(二)",
         "本村の事業所数と、告示が求める「最低一箇所確保」への到達状況"),
        ("04_県内比較", "18サービス", "別表第五・見直し事項⑩",
         "県内59市町村のうち事業所0か所の市町村数。人口減少地域に共通の課題であることの確認"),
    ]
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_title(ws.cell(row=r, column=1), "この表から言えること",
                fill=COLORS["subhead"], size=11)
    r += 1
    zero25 = [s for s in SERVICE_ORDER
              if s in data and data[s].get(VILLAGE, (0, 0))[1] == 0]
    has25 = [s for s in SERVICE_ORDER
             if s in data and data[s].get(VILLAGE, (0, 0))[1] > 0]
    notes = [
        f"本村に事業所があるサービスは{len(has25)}種類"
        + (f"（{'・'.join(has25)}）" if has25 else "")
        + f"、事業所が無いサービスは{len(zero25)}種類です。",
        "基本指針 第三の二の2(二)は、訪問系サービス及び指定通所支援について"
        "「各市町村において事業を実施する事業所を最低一箇所確保できるよう努める」としています。",
        "同項は、小規模町村等で確保できない場合の工夫として、介護保険の訪問介護事業所・"
        "居宅介護支援事業所への居宅介護指定の働きかけと、共生型サービスの指定制度の周知を挙げています。",
        "事業所数は所在地ベースのため、圏域外の事業所を利用している分は現れません。"
        "実際に利用可能な量は、08_供給制約シート（サービス見込量ブック）で事業所へ照会して把握します。",
        "出典の障害福祉サービス等情報公表システムに公表されていない事業所は計上されません。"
        "第8期基本指針が公表率・更新率100%を成果目標に新設したのは、この網羅性の問題への対処です（第二の八）。"
        "村の指定・委託の記録と突き合わせて確認してください。",
    ]
    for t in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        style_note(ws.cell(row=r, column=1), t)
        ws.row_dimensions[r].height = 26
        r += 1
    return ws


def _matrix_sheet(wb, title, heading, note, munis, data):
    widths = [24] + [11] * (len(munis) * 2) + [14]
    ws = add_sheet(wb, title, heading, note, widths)
    r = 5
    head = ["サービス"]
    for m in munis:
        head += [f"{m}\n2020", f"{m}\n2025"]
    head += ["圏域計\n2025"]
    style_header_row(ws, r, head)
    r += 1
    for i, svc in enumerate(SERVICE_ORDER):
        vals = data.get(svc, {})
        row = [svc]
        total25 = 0
        for m in munis:
            a, b = vals.get(m, (None, None))
            row += [a, b]
            total25 += (b or 0)
        row.append(total25)
        n = len(munis) * 2 + 2
        fills = [None]
        for m in munis:
            b = vals.get(m, (None, None))[1]
            f = COLORS["input"] if (m == VILLAGE and b == 0) else None
            fills += [None, f]
        fills.append(COLORS["calc"] if total25 == 0 else None)
        write_row(ws, r, row, alt=(i % 2 == 1),
                  aligns=["left"] + ["right"] * (n - 1),
                  numfmts=[None] + [INT] * (n - 1),
                  fills=fills)
        r += 1
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=len(widths))
    style_note(ws.cell(row=r + 1, column=1),
               "網掛けは本村に事業所が無いサービス。圏域計が0のサービスは、"
               "圏域内に事業所が1か所も無いため、圏域外の事業所への依存または"
               "確保方策の検討が必要。")
    ws.row_dimensions[r + 1].height = 28
    return ws


def sheet_village(wb, data):
    ws = add_sheet(
        wb, "03_村の状況", "北塩原村の障がい福祉サービス事業所数",
        "基本指針 第三の二の2(二)は、訪問系サービス及び指定通所支援について"
        "「各市町村において事業を実施する事業所を最低一箇所確保できるよう努める」としている。"
        "本表はその到達状況を示す。",
        [26, 14, 14, 14, 18, 60])
    style_header_row(ws, 5, ["サービス", "2020年", "2025年", "増減",
                             "最低1か所の努力規定", "会津北部圏域の状況（2025年）"])
    r = 6
    for i, svc in enumerate(SERVICE_ORDER):
        vals = data.get(svc, {})
        a, b = vals.get(VILLAGE, (None, None))
        # 訪問系サービスは告示が名指しで最低1か所の努力を求める
        houmon = svc in ("居宅介護", "重度訪問介護", "同行援護", "行動援護",
                         "重度障害者等包括支援")
        target = "対象（訪問系）" if houmon else "―"
        kenhokubu = "／".join(
            f"{m}{vals.get(m, (0, 0))[1]}" for m in KEN_HOKUBU if m != VILLAGE)
        wide = "／".join(
            f"{m}{vals.get(m, (0, 0))[1]}" for m in ["会津若松市", "喜多方市"])
        write_row(ws, r, [svc, a, b, (b - a) if (a is not None and b is not None) else None,
                          target, f"圏域内 {kenhokubu}　／　近隣市 {wide}"],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "center", "left"],
                  numfmts=[None, INT, INT, "+0;-0;0", None, None],
                  fills=[None, None,
                         COLORS["input"] if b == 0 else COLORS["calc"],
                         None,
                         "FFD9D9" if (houmon and b == 0) else None,
                         None])
        if houmon and b == 0:
            c = ws.cell(row=r, column=5)
            c.font = Font(name=FONT, size=10, bold=True, color="C00000")
        r += 1
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    style_note(ws.cell(row=r + 1, column=1),
               "赤色は、告示が名指しで「最低一箇所確保できるよう努める」としている訪問系サービスのうち、"
               "本村に事業所が無いもの。骨子案修正版 第5章8 に掲げた確保方策"
               "（介護保険事業所への居宅介護指定の働きかけ、共生型サービスの周知、"
               "基準該当障害福祉サービス、多機能型・従たる事業所）の対象となる。")
    ws.row_dimensions[r + 1].height = 40
    return ws


def sheet_kennai(wb, data):
    ws = add_sheet(
        wb, "04_県内比較", "県内市町村における事業所0か所の状況（2025年）",
        "本村だけの課題か、人口減少地域に共通の課題かを確認する。"
        "基本指針の見直し事項⑩「人口減少地域におけるサービスの維持・確保」の裏づけになる。",
        [26, 16, 16, 16, 58])
    style_header_row(ws, 5, ["サービス", "県内市町村数", "うち0か所",
                             "0か所の割合", "本村の状況"])
    r = 6
    for i, svc in enumerate(SERVICE_ORDER):
        vals = data.get(svc, {})
        n = len(vals)
        zero = sum(1 for v in vals.values() if v[1] == 0)
        b = vals.get(VILLAGE, (None, None))[1]
        state = "0か所" if b == 0 else f"{b}か所"
        write_row(ws, r, [svc, n, zero, (zero / n if n else None), state],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "center"],
                  numfmts=[None, INT, INT, "0.0%", None],
                  fills=[None, None, None, None,
                         COLORS["input"] if b == 0 else COLORS["calc"]])
        r += 1
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    style_note(ws.cell(row=r + 1, column=1),
               "0か所の割合が高いサービスは、県内の多くの市町村で事業所が確保できていない。"
               "本村単独での確保を目標に掲げるのではなく、圏域確保・共生型サービス・"
               "基準該当障害福祉サービス等の活用を前提に方策を組むのが実際的。")
    ws.row_dimensions[r + 1].height = 30
    return ws


def verify(data):
    missing = [s for s in SERVICE_ORDER if s not in data]
    if missing:
        raise SystemExit(f"入力に無いサービスがあります: {missing}")
    extra = [s for s in data if s not in SERVICE_ORDER]
    if extra:
        raise SystemExit(f"掲載順に無いサービスがあります: {extra}")
    for svc, vals in data.items():
        if VILLAGE not in vals:
            raise SystemExit(f"{svc} に {VILLAGE} の行がありません")
    n = {len(v) for v in data.values()}
    print(f"  自己検証: {len(data)}サービス／市町村数 {sorted(n)}／"
          f"{VILLAGE} 全サービス分の値を確認")


def main():
    data = load_all()
    verify(data)
    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb, data)
    _matrix_sheet(wb, "01_会津北部圏域", "会津北部圏域4町村の事業所数（2020→2025）",
                  "地域生活支援拠点等を共同整備する猪苗代町・磐梯町・湯川村・北塩原村。"
                  "圏域計が0のサービスは圏域外への依存となる。",
                  KEN_HOKUBU, data)
    _matrix_sheet(wb, "02_広域", "会津北部圏域＋近隣市町村の事業所数（2020→2025）",
                  "本村の利用者が実際に通所している会津若松市・喜多方市等を含む広域。",
                  AIZU_WIDE, data)
    sheet_village(wb, data)
    sheet_kennai(wb, data)
    wb.save(OUT_FILE)
    print(f"作成: {OUT_FILE}")

    has = [s for s in SERVICE_ORDER if data[s].get(VILLAGE, (0, 0))[1] > 0]
    print(f"  {VILLAGE} 2025年に事業所があるサービス: {has if has else 'なし'}")


if __name__ == "__main__":
    main()
