#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""階上町 下水道使用料 改定検討用 集計スクリプト（レビュー反映版 v2）.

使い方:
    python3 analyze_choutei.py <集計ブック.xlsx> [調定明細.xlsx ...]

  第1引数 : 【R7】使用料集計ブック（「R7漁集 (件数)」「R7公共 (件数) 」シートを持つもの）
  第2引数以降（任意）: 調定簿明細のブック。現行料金体系との整合を検証する。

調定Excelは個人情報を含むためリポジトリには含めていない。

v2での修正（2026-08-31）
  1. 金額からの水量逆算を修正。基本料金2,217円は1〜10㎥のいずれでも同額のため
     水量が一意に定まらない。v1は「1㎥」として集計していた（過少計上）。
     v2では「水量不詳（1〜10㎥）」として下限・上限を併記する。
  2. 調定を「通常算定」「基本料金帯（水量不詳）」「特殊算定（日割・異動等）」に分類。
     特殊算定分は水量推計の対象外とし、増収試算では現行調定額×改定率で扱う。
  3. 実効単価は、金額と水量が同一の母集団から求まる区分のみで算出する。
"""
import math
import sys

import openpyxl

# --- 現行使用料体系（税抜・1か月あたり） -----------------------------------
# 基本使用料 5㎥まで 1,008円 / 6〜10㎥ 37円 / 11〜50㎥ 174円 / 51㎥〜 200円
# 調定は隔月検針のため、2か月分の水量 V に対して区分境界を2倍(10/20/100㎥)して適用する。
CURRENT = (2016, 37, 174, 200)

# 新料金案1（税込・1か月あたり 1,220 / 45.5 / 211.5 / 242円）を
# 2か月分の水量 V に直接適用できる形にしたもの
PLAN1 = (2440, 45.5, 211.5, 242)


def charge(volume, base, r1, r2, r3):
    """2か月分の使用水量から料金を計算する（単価の積み上げ）。"""
    amount = base
    if volume > 10:
        amount += r1 * min(volume - 10, 10)
    if volume > 20:
        amount += r2 * min(volume - 20, 80)
    if volume > 100:
        amount += r3 * (volume - 100)
    return amount


def current_amount(volume):
    """現行の2か月分請求額（税込・1円未満切捨て）。"""
    return math.floor(charge(volume, *CURRENT) * 1.1)


def plan1_amount(volume):
    """案1の2か月分請求額。単価が税込のため消費税の再乗算は行わない。"""
    return math.floor(charge(volume, *PLAN1))


def uniform_plan(ratio):
    """現行単価を一律 ratio 倍した料金案。"""
    params = (math.floor(2217 * ratio), round(40.7 * ratio, 1),
              round(191.4 * ratio, 1), round(220.0 * ratio, 1))
    return lambda volume: math.floor(charge(volume, *params))


# 11㎥以上は金額と水量が1対1で対応する。1〜10㎥は基本料金2,217円に縮退するため
# 金額から水量を特定できない（v1の誤りの原因）。
EXACT_VOLUME = {current_amount(v): v for v in range(11, 4000)}
BASIC_AMOUNT = current_amount(10)

NORMAL, BASIC, SPECIAL = '通常算定', '基本料金帯', '特殊算定'


def load_counts(worksheet):
    """（件数）シートを (区分, 水量, 件数, 現行調定額) のリストに変換する。

    水量は通常算定のみ確定値、基本料金帯と特殊算定は None。
    """
    records = []
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        if str(row[0]).strip() == '合計':
            continue
        amount, count = row[1], row[8]
        if not isinstance(count, (int, float)) or count == 0:
            continue
        count = int(count)
        if not isinstance(amount, (int, float)):
            # 「～2,216」区分（日割・休止等）。金額は集計ブックの計算値をそのまま使う
            records.append((SPECIAL, None, count, int(row[9] or 0)))
            continue
        amount = int(amount)
        if amount == BASIC_AMOUNT:
            records.append((BASIC, None, count, amount * count))
        elif amount in EXACT_VOLUME:
            records.append((NORMAL, EXACT_VOLUME[amount], count, amount * count))
        else:
            records.append((SPECIAL, None, count, amount * count))
    return records


def volume_bounds(records):
    """(下限, 上限) の水量。特殊算定は推計不可のため含めない。"""
    low = high = 0
    for kind, volume, count, _ in records:
        if kind == NORMAL:
            low += volume * count
            high += volume * count
        elif kind == BASIC:
            low += 1 * count      # 1〜10㎥のいずれかで、金額からは特定できない
            high += 10 * count
    return low, high


def revenue(records, amount_of, ratio):
    """料金体系を当てはめた年間調定額。

    通常算定・基本料金帯は水量から再計算する（基本料金帯は水量が1でも10でも
    基本料金のみのため、水量不詳でも金額は一意に定まる）。
    特殊算定は日割等で本来の水量が不明なため、現行調定額×改定率で扱う。
    """
    total = 0
    for kind, volume, count, current in records:
        if kind == NORMAL:
            total += amount_of(volume) * count
        elif kind == BASIC:
            total += amount_of(10) * count
        else:
            total += round(current * ratio)
    return total


def report_summary(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    grand = {}
    for sheet, label in [('R7漁集 (件数)', '漁業集落排水'), ('R7公共 (件数) ', '公共下水道')]:
        records = load_counts(workbook[sheet])
        count = sum(c for _, _, c, _ in records)
        current = sum(a for _, _, _, a in records)
        low, high = volume_bounds(records)
        # 実効単価は水量を推計できる区分（通常算定＋基本料金帯）のみで算出する
        metered = sum(a for k, _, _, a in records if k != SPECIAL)
        special_count = sum(c for k, _, c, _ in records if k == SPECIAL)
        special_amount = sum(a for k, _, _, a in records if k == SPECIAL)

        print('=== %s' % label)
        print('  調定件数        : %s 件（奇数月6回調定分）' % format(count, ','))
        print('  調定額（税込）  : %s 円' % format(current, ','))
        print('  内訳            : 通常算定 %s件 / 基本料金帯 %s件 / 特殊算定 %s件（%s円）'
              % (format(sum(c for k, _, c, _ in records if k == NORMAL), ','),
                 format(sum(c for k, _, c, _ in records if k == BASIC), ','),
                 format(special_count, ','), format(special_amount, ',')))
        print('  推計水量        : %s 〜 %s ㎥（基本料金帯は水量不詳。特殊算定 %s件は推計対象外）'
              % (format(low, ','), format(high, ','), format(special_count, ',')))
        print('  実効単価        : %.1f 〜 %.1f 円/㎥（税込・水量を推計できる区分のみ）'
              % (metered / high, metered / low))
        plan = revenue(records, plan1_amount, 1.10)
        print('  案1（約+10%%）   : %s 円 (%+.2f%%, 増収 %s 円)'
              % (format(plan, ','), (plan / current - 1) * 100, format(plan - current, ',')))
        for ratio in (1.15, 1.20):
            total = revenue(records, uniform_plan(ratio), ratio)
            print('  参考 一律%+.0f%%   : %s 円 (%+.2f%%, 増収 %s 円)'
                  % ((ratio - 1) * 100, format(total, ','), (total / current - 1) * 100,
                     format(total - current, ',')))
        grand[label] = (current, plan)

    current_total = sum(v[0] for v in grand.values())
    plan_total = sum(v[1] for v in grand.values())
    print('=== 2事業合計')
    print('  現行 %s 円 → 案1 %s 円（増収 %s 円 / %+.2f%%）'
          % (format(current_total, ','), format(plan_total, ','),
             format(plan_total - current_total, ','), (plan_total / current_total - 1) * 100))


def verify_tariff(paths):
    """調定明細の実額が現行料金体系で再現できるか検証する。"""
    for path in paths:
        workbook = openpyxl.load_workbook(path, data_only=True)
        matched = under_base = mismatched = 0
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                if not (row[1] and str(row[1]).startswith('階上') and row[2]):
                    continue
                if not isinstance(row[6], (int, float)):
                    continue
                amount = int(row[6])
                if amount == BASIC_AMOUNT or amount in EXACT_VOLUME:
                    matched += 1
                elif amount < BASIC_AMOUNT:
                    under_base += 1
                else:
                    mismatched += 1
        total = matched + under_base + mismatched
        print('=== 料金体系の検証: %s' % path)
        print('  対象 %s 件 / 体系と一致 %s 件 (%.1f%%) / 基本料金未満（日割等） %s 件 / 不一致 %s 件'
              % (format(total, ','), format(matched, ','), matched / total * 100,
                 format(under_base, ','), format(mismatched, ',')))


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    report_summary(sys.argv[1])
    if len(sys.argv) > 2:
        verify_tariff(sys.argv[2:])
    return 0


if __name__ == '__main__':
    sys.exit(main())
