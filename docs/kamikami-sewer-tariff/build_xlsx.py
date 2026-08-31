#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""試算エビデンスCSV一式を単一のExcelブックに統合する。"""
import csv, io, os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

EV = '/tmp/claude-0/-home-user-repository/670c168c-8281-57ba-9df0-b54358bb5879/scratchpad/evidence/data'
OUT = '/tmp/claude-0/-home-user-repository/670c168c-8281-57ba-9df0-b54358bb5879/scratchpad/階上町下水道使用料改定_試算エビデンス.xlsx'

BASE   = Font(name='Arial', size=10)
BOLD   = Font(name='Arial', size=10, bold=True)
NOTE   = Font(name='Arial', size=9, color='6B8189', italic=True)
TITLE  = Font(name='Arial', size=13, bold=True)
HEADFILL  = PatternFill('solid', fgColor='DCEAEA')
TOTALFILL = PatternFill('solid', fgColor='F0E4D8')
THIN = Side(style='thin', color='BFCFCF')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEETS = [
    ('01_現行使用料体系',      '01_現行使用料体系.csv'),
    ('02_隔月請求額 早見表',    '02_隔月請求額_早見表.csv'),
    ('03_度数分布 漁業集落排水', '03_金額別度数分布_漁業集落排水.csv'),
    ('04_度数分布 公共下水道',  '04_金額別度数分布_公共下水道.csv'),
    ('05_水量区分別集計',      '05_水量区分別集計.csv'),
    ('06_料金体系の検証',      '06_料金体系の検証_R8.3月調定.csv'),
    ('07_増収試算',           '07_増収試算.csv'),
    ('08_R7調定実績サマリ',    '08_R7調定実績サマリ.csv'),
    ('09_モデルケース',        '09_モデルケース.csv'),
    ('10_経営戦略 抜粋数値',    '10_経営戦略_抜粋数値.csv'),
    ('11_元ファイル一覧',      '11_元ファイル一覧.csv'),
    ('12_水量推計の内訳',      '12_水量推計の内訳.csv'),
    ('13_訂正履歴',           '13_訂正履歴.csv'),
    ('14_R6決算統計 32表40表',  '14_R6決算統計_32表40表.csv'),
    ('15_料金表の比較',        '15_料金表の比較.csv'),
    ('16_増収と経費回収率',     '16_増収と経費回収率の比較.csv'),
    ('17_モデルケースと単価',   '17_モデルケースと単価の公平性.csv'),
]

def read_csv(path):
    with io.open(os.path.join(EV, path), encoding='utf-8-sig') as fh:
        return [r for r in csv.reader(fh)]

def as_number(s):
    if s is None or s == '':
        return ''
    t = s.replace(',', '')
    try:
        return int(t)
    except ValueError:
        pass
    try:
        return float(t)
    except ValueError:
        return s

wb = openpyxl.Workbook()

# ---------------- 目次 ----------------
toc = wb.active
toc.title = '目次'
toc.sheet_view.showGridLines = False
rows = [
    ('階上町 下水道使用料改定　試算エビデンス（v3）', ''),
    ('', ''),
    ('作成日', '2026-08-19 ／ 改訂 2026-08-31（v3）'),
    ('対象', '令和7年度 調定データ／経営戦略（令和7年2月）／【R7】使用料集計ブック／令和6年度決算統計（32表・40表）'),
    ('集計範囲', 'R7年度奇数月6回調定集計ベース。通常の隔月検針者は概ね12か月相当だが、偶数月調定及び毎月検針者の一部を含まない'),
    ('個人情報', '氏名・調定明細の個票は収録していません。金額別に集約した度数分布と、その集計値・試算値のみです。'),
    ('', ''),
    ('シート', '内容 ／ どの数値の根拠か'),
    ('01_現行使用料体系', '条例上の使用料体系（一般汚水・公衆浴場、1か月あたり税込／税抜換算）'),
    ('02_隔月請求額 早見表', '2か月水量ごとの税抜額・現行請求額・案1請求額・差額'),
    ('03_度数分布 漁業集落排水', '★増収試算の直接の計算元。合計欄＝現行 7,373,849円／案1 8,136,688円'),
    ('04_度数分布 公共下水道', '★増収試算の直接の計算元。合計欄＝現行 35,613,283円／案1 39,278,957円'),
    ('05_水量区分別集計', '2か月水量区分ごとの件数・金額と構成比'),
    ('06_料金体系の検証', 'R8.3月調定1,519件を体系式と照合した結果（判定別件数＋不一致の金額内訳）'),
    ('07_増収試算', '現行／案1／一律+15%／+20% の事業別・合計の調定額と増収額'),
    ('08_R7調定実績サマリ', '件数・調定額・推計水量の下限上限・実効単価の下限上限'),
    ('09_モデルケース', '月使用量別の現行／案1の月額・2か月請求額・差額'),
    ('10_経営戦略 抜粋数値', '経営指標（R4実績・R16目標）、改善方策ケース、使用料収入・有収水量の推計値'),
    ('11_元ファイル一覧', '元ファイルのファイル名・サイズ・SHA-256ハッシュ・シート構成'),
    ('12_水量推計の内訳', '通常算定／基本料金帯／特殊算定の件数・調定額・水量下限上限（v2で追加）'),
    ('13_訂正履歴', 'v1→v2の訂正対照表（v2で追加）'),
    ('14_R6決算統計 32表40表', 'R6決算統計の32表・40表の対応検証と経費回収率（v3で追加）'),
    ('15_料金表の比較', '現行／案A（体系維持）／案B（体系見直し）の料金表（v3で追加）'),
    ('16_増収と経費回収率', '各案の調定額・増収額・経費回収率・基準内繰入金（v3で追加）'),
    ('17_モデルケースと単価', '月使用量別の負担額と単価。負担の公平性の比較（v3で追加）'),
    ('', ''),
    ('経費回収率の定義', '使用料収入 ÷ 汚水処理費。汚水処理費は32表の維持管理費（汚水分）のみ'),
    ('', '資本費（減価償却費−長期前受金戻入＋支払利息）は全額が「分流式下水道等に要する経費」'),
    ('', 'として基準内繰入金で措置されるため、汚水処理費には含まれない'),
    ('', 'R6実績：公共 33,652÷67,464＝49.9%／漁集 7,229÷19,614＝36.9%'),
    ('', '※水洗便所等普及費330千円（漁集は不明水処理費2,318千円）は現行額が継続する前提'),
    ('', ''),
    ('案A／案Bの違い', '案A＝現行3区分を維持し単価を一律引上げ（経営戦略ケース①準拠）'),
    ('', '案B＝区分を4つに増やし、6〜10㎥の単価を引上げて11㎥での4.7倍の段差を1.6倍に緩和'),
    ('', '増収額・経費回収率はほぼ同等。負担の分布が異なる（案Bは月6〜10㎥層が重く、月20㎥以上が軽い）'),
    ('', ''),
    ('現行使用料の算定式', '隔月検針のため1請求＝2か月分。2か月水量Vに対し区分境界を2倍(10/20/100㎥)して適用'),
    ('', '2か月分請求額（税込） = FLOOR( 税抜額 × 1.1 )'),
    ('', '税抜額 = 2,016 ＋ 37×(V−10)[10<V≦20] ＋ 174×(V−20)[20<V≦100] ＋ 200×(V−100)[100<V]'),
    ('', 'R8.3月調定1,519件のうち1,428件（94.0%）がこの式と完全一致（シート06）'),
    ('', ''),
    ('調定の3区分', '通常算定＝金額から水量が一意に定まる（2か月11㎥以上）'),
    ('', '基本料金帯＝2,217円。2か月1〜10㎥のいずれでも同額のため水量不詳（下限1㎥／上限10㎥）'),
    ('', '特殊算定＝日割・月中異動等で料金式に一致しない。水量推計は対象外、増収試算は現行×改定率'),
    ('', ''),
    ('v2での主な訂正', '基本料金帯を1㎥として集計していた誤りを修正（漁集285件・公共1,657件、幅は最大17,478㎥）'),
    ('', '増収試算への影響は軽微。2か月10㎥以下は基本料金のみで金額が決まるため水量に依存しない'),
    ('', '特殊算定の扱い変更による差は2事業計 +4,019円（増収額の0.09%）'),
    ('', ''),
    ('案1の端数処理', '集計ブックの案1集計値 8,218,319円（+11.45%）は奇数㎥を2㎥ブロックに切り上げているため過大。'),
    ('', '現行と同じ算定方法で再計算すると 8,136,688円（+10.35%）。本ブックは後者を採用。'),
    ('', ''),
    ('既知の限界', '① 集計ブックは奇数月6回分のみ対象。偶数月の調定（漁集で7件・43,091円）が未計上'),
    ('', '② 毎月検針の大口分は年12回のうち奇数月6回しか捕捉できていない。増収額は過小の可能性'),
    ('', '③ 公共下水道の月別明細は3月分のみ入手済み。他は金額からの逆算'),
    ('', '④ 漁集2ブック間で1月分に1件差異（オリジナル側に3,800円が1件多い）。本試算は236件側に準拠'),
    ('', '⑤ 特殊算定492件（3,045,941円＝全体の7.1%）は現行×改定率の簡便法によっている'),
]
for i, (a, b) in enumerate(rows, 1):
    toc.cell(i, 1, a).font = TITLE if i == 1 else (BOLD if b and i in (7,) or (a and not b) else BOLD if a else BASE)
    toc.cell(i, 2, b).font = BASE
    toc.cell(i, 2).alignment = Alignment(wrap_text=False, vertical='center')
toc.cell(7, 1).fill = HEADFILL; toc.cell(7, 2).fill = HEADFILL
toc.column_dimensions['A'].width = 26
toc.column_dimensions['B'].width = 110
toc.freeze_panes = 'A8'

# ---------------- データシート ----------------
FORMULA_SHEETS = {'03_度数分布 漁業集落排水', '04_度数分布 公共下水道'}

for name, csvfile in SHEETS:
    data = read_csv(csvfile)
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    r = 1
    header_row = None
    for row in data:
        if row and row[0].startswith('#'):
            ws.cell(r, 1, (row[1] if len(row) > 1 else row[0].lstrip('# ')) or row[0])
            ws.cell(r, 1).font = NOTE
            r += 1
            continue
        if not any(row):
            r += 1
            continue
        is_header = header_row is None and row[0] not in ('', '合計')
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v if is_header else as_number(v))
            cell.font = BOLD if is_header else BASE
            if is_header:
                cell.fill = HEADFILL
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = BOX
            if not is_header and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0' if float(cell.value) == int(cell.value) and abs(cell.value) >= 1000 else '#,##0.0#'
            if row[0] == '合計':
                cell.fill = TOTALFILL
                cell.font = BOLD
        if is_header:
            header_row = r
        r += 1

    # 度数分布シート：ウィンドウ枠固定のみ（値は検算済み。数式は使わない）
    if header_row:
        ws.freeze_panes = ('C%d' % (header_row + 1)) if name in FORMULA_SHEETS else ('A%d' % (header_row + 1))

    # 列幅
    for c in range(1, ws.max_column + 1):
        width = 10
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(rr, c).value
            if v is None:
                continue
            s = str(v)
            if s.startswith('='):
                continue
            ln = sum(2 if ord(ch) > 0x2000 else 1 for ch in s)
            width = max(width, min(ln + 3, 62))
        ws.column_dimensions[get_column_letter(c)].width = width

wb.save(OUT)
print('saved:', OUT)
print('sheets:', wb.sheetnames)
