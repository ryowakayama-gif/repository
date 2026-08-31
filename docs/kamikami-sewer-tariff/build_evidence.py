#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""試算エビデンス一式を生成する。個人情報（氏名・使用者番号）は一切出力しない。"""
import csv, hashlib, io, math, os, shutil, sys
from collections import Counter
import openpyxl

SRC = '/root/.claude/uploads/670c168c-8281-57ba-9df0-b54358bb5879'
OUT = '/tmp/claude-0/-home-user-repository/670c168c-8281-57ba-9df0-b54358bb5879/scratchpad/evidence'
FILES = {
    'shukei':   ('90c4df88-__R7________.xlsx', '【R7】使用料集計ブック'),
    'gyoshu_a': ('b7fef742-R7__.xlsx',        'R7漁集 調定簿明細（集計シート付）'),
    'gyoshu_b': ('37ccc0e0-R7___.xlsx',       'R7漁集 調定簿明細（オリジナル）'),
    'r83':      ('6ff71ec0-R8.3_.xlsx',       'R8.3月 調定簿明細（公共・漁集）'),
    'senryaku': ('39af4522-20250424093158.pdf','階上町下水道事業経営戦略（改定）令和7年2月'),
}

def w(path, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with io.open(path, 'w', encoding='utf-8-sig', newline='') as fh:
        csv.writer(fh).writerows(rows)
    print('  ->', os.path.relpath(path, OUT), '(%d行)' % len(rows))

# --- 現行料金体系 ------------------------------------------------------------
CUR = (2016, 37, 174, 200)
PLAN1 = (2440, 45.5, 211.5, 242)

def charge(v, base, r1, r2, r3):
    n = base
    if v > 10:  n += r1 * min(v - 10, 10)
    if v > 20:  n += r2 * min(v - 20, 80)
    if v > 100: n += r3 * (v - 100)
    return n

def cur_amt(v):   return math.floor(charge(v, *CUR) * 1.1)
def plan1_amt(v): return math.floor(charge(v, *PLAN1))

def uniform(ratio):
    p = (math.floor(2217 * ratio), round(40.7 * ratio, 1),
         round(191.4 * ratio, 1), round(220.0 * ratio, 1))
    return (lambda v: math.floor(charge(v, *p))), p

# 11㎥以上は金額と水量が1対1。1〜10㎥は基本料金2,217円に縮退し特定できない。
EXACT = {cur_amt(v): v for v in range(11, 4000)}
BASIC_AMT = cur_amt(10)
NORMAL, BASIC, SPECIAL = '通常算定', '基本料金帯', '特殊算定'

def classify(amount):
    """請求金額から (区分, 水量) を返す。水量は通常算定のみ確定。"""
    if amount == BASIC_AMT: return BASIC, None
    if amount in EXACT:     return NORMAL, EXACT[amount]
    return SPECIAL, None

shutil.rmtree(OUT, ignore_errors=True)
os.makedirs(OUT, exist_ok=True)
wb = openpyxl.load_workbook(os.path.join(SRC, FILES['shukei'][0]), data_only=True)

# === 01 現行料金表 ===========================================================
print('01 現行料金表')
w(OUT + '/data/01_現行使用料体系.csv', [
    ['#', '出典：階上町下水道事業経営戦略（改定）表2.7（出典 階上町HP）／調定実額で検証済み'],
    ['汚水種別', '区分', '範囲', '税込単価(円)', '税抜換算(円)', '備考'],
    ['一般汚水', '基本使用料', '5㎥まで',  '1108.8', '1008', '1か月あたり'],
    ['一般汚水', '超過使用料', '6〜10㎥',  '40.7',   '37',   '1㎥につき'],
    ['一般汚水', '超過使用料', '11〜50㎥', '191.4',  '174',  '1㎥につき'],
    ['一般汚水', '超過使用料', '51㎥〜',   '220.0',  '200',  '1㎥につき'],
    ['公衆浴場', '基本使用料', '5㎥まで',  '1108.8', '1008', '1か月あたり'],
    ['公衆浴場', '超過使用料', '6㎥〜',    '62.7',   '57',   '1㎥につき'],
])

# === 02 隔月請求額の算定式と早見表 ==========================================
print('02 隔月請求額 早見表')
rows = [['#', '隔月検針のため1請求＝2か月分。2か月水量Vに対し区分境界を2倍(10/20/100㎥)して適用し、税込額は1円未満切捨て'],
        ['2か月水量(㎥)', '税抜額(円)', '現行 請求額(円・税込)', '案1 請求額(円・税込)', '差額(円)', '増減率(%)']]
for v in list(range(1, 51)) + [60, 70, 80, 90, 100, 101, 110, 150, 200, 300, 378, 500, 766]:
    c, p = cur_amt(v), plan1_amt(v)
    rows.append([v, int(charge(v, *CUR)), c, p, p - c, '%.2f' % ((p / c - 1) * 100)])
w(OUT + '/data/02_隔月請求額_早見表.csv', rows)

# === 03/04 金額別度数分布 ====================================================
def distribution(sheet, path, title):
    ws = wb[sheet]
    rows = [['#', title + '／出典：【R7】使用料集計ブック ' + sheet],
            ['#', '区分：通常算定＝金額から水量が一意に定まる／基本料金帯＝1〜10㎥のいずれでも同額で水量不詳／特殊算定＝日割・異動等で料金式に一致しない'],
            ['#', '案1調定額は、通常算定・基本料金帯は水量から再計算、特殊算定は現行調定額×1.10で算定'],
            ['2か月水量(㎥)', '現行請求額(円)', '区分', '5月', '7月', '9月', '11月', '1月', '3月',
             '件数計', '現行調定額(円)', '案1請求額(円)', '案1調定額(円)']]
    tc = ta = tp = 0
    tally = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if str(r[0]).strip() == '合計': continue
        amt, cnt = r[1], r[8]
        if not isinstance(cnt, (int, float)) or cnt == 0: continue
        cnt = int(cnt)
        months = [int(x) if isinstance(x, (int, float)) else '' for x in r[2:8]]
        if not isinstance(amt, (int, float)):
            cur_a = int(r[9] or 0); new_a = round(cur_a * 1.10)
            rows.append([r[0] or '', '〜2,216', SPECIAL] + months +
                        [cnt, cur_a, '現行×1.10', new_a])
            kind, vol = SPECIAL, None
        else:
            amt = int(amt); kind, vol = classify(amt)
            cur_a = amt * cnt
            if kind == SPECIAL:
                new_unit, new_a = '現行×1.10', round(cur_a * 1.10)
            else:
                new_unit = plan1_amt(vol if kind == NORMAL else 10)
                new_a = new_unit * cnt
            rows.append([r[0] if r[0] is not None else (vol or ''), amt, kind] + months +
                        [cnt, cur_a, new_unit, new_a])
        tc += cnt; ta += cur_a; tp += new_a
        t = tally.setdefault(kind, [0, 0, 0, 0])
        t[0] += cnt; t[1] += cur_a
        if kind == NORMAL:
            t[2] += vol * cnt; t[3] += vol * cnt
        elif kind == BASIC:
            t[2] += 1 * cnt; t[3] += 10 * cnt
    rows.append(['合計', '', '', '', '', '', '', '', '', tc, ta, '', tp])
    w(path, rows)
    return tc, ta, tp, tally

print('03/04 金額別度数分布')
g_cnt, g_cur, g_plan, g_tally = distribution(
    'R7漁集 (件数)', OUT + '/data/03_金額別度数分布_漁業集落排水.csv', '漁業集落排水 R7年度 金額別度数分布')
k_cnt, k_cur, k_plan, k_tally = distribution(
    'R7公共 (件数) ', OUT + '/data/04_金額別度数分布_公共下水道.csv', '公共下水道 R7年度 金額別度数分布')

# === 05 水量区分別集計 =======================================================
print('05 水量区分別集計')
rows = [['#', '出典：【R7】使用料集計ブック（件数）シート右表。2か月分の水量ベース'],
        ['事業', '2か月水量区分', '件数', '件数割合(%)', '金額(円)', '金額割合(%)']]
for sheet, label in [('R7漁集 (件数)', '漁業集落排水'), ('R7公共 (件数) ', '公共下水道')]:
    ws = wb[sheet]
    for r in range(3, 20):
        lab = ws.cell(r, 13).value
        if lab is None: continue
        c, cr, a, ar = (ws.cell(r, i).value for i in (14, 15, 16, 17))
        rows.append([label, lab, c,
                     '%.2f' % (cr * 100) if isinstance(cr, float) else '',
                     a, '%.2f' % (ar * 100) if isinstance(ar, float) else ''])
w(OUT + '/data/05_水量区分別集計.csv', rows)

# === 06 料金体系の検証（R8.3月調定・匿名集計） ==============================
print('06 料金体系の検証')
wb83 = openpyxl.load_workbook(os.path.join(SRC, FILES['r83'][0]), data_only=True)
tally = Counter(); miss = Counter()
for ws in wb83.worksheets:
    for r in ws.iter_rows(values_only=True):
        if not (r[1] and str(r[1]).startswith('階上') and r[2]): continue
        if not isinstance(r[6], (int, float)): continue
        a = int(r[6]); k = str(r[1])
        if a == BASIC_AMT or a in EXACT: tally[(k, '体系と完全一致')] += 1
        elif a < BASIC_AMT: tally[(k, '基本料金未満（日割・休止等）')] += 1
        else:               tally[(k, '不一致（月中異動の日割等）')] += 1; miss[(k, a)] += 1
rows = [['#', '出典：R8.3月 調定簿明細（令和7年度3月分）。氏名・使用者番号は出力していない'],
        ['下水区分', '判定', '件数']]
for (k, j), n in sorted(tally.items()):
    rows.append([k, j, n])
rows.append(['合計', '', sum(tally.values())])
rows.append([])
rows.append(['# 不一致の内訳（金額と件数のみ）'])
rows.append(['下水区分', '調定額(円)', '件数'])
for (k, a), n in sorted(miss.items()):
    rows.append([k, a, n])
w(OUT + '/data/06_料金体系の検証_R8.3月調定.csv', rows)

# === 07 増収試算 =============================================================
print('07 増収試算')
def load(sheet):
    ws = wb[sheet]; recs = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if str(r[0]).strip() == '合計': continue
        amt, cnt = r[1], r[8]
        if not isinstance(cnt, (int, float)) or cnt == 0: continue
        cnt = int(cnt)
        if not isinstance(amt, (int, float)):
            recs.append((SPECIAL, None, cnt, int(r[9] or 0))); continue
        amt = int(amt); kind, vol = classify(amt)
        recs.append((kind, vol, cnt, amt * cnt))
    return recs

def revenue(recs, fn, ratio):
    """通常算定・基本料金帯は水量から再計算、特殊算定は現行調定額×改定率。"""
    total = 0
    for kind, vol, cnt, cur in recs:
        if kind == NORMAL:  total += fn(vol) * cnt
        elif kind == BASIC: total += fn(10) * cnt
        else:               total += round(cur * ratio)
    return total

gr, kr = load('R7漁集 (件数)'), load('R7公共 (件数) ')
base_total = g_cur + k_cur
rows = [['#', '試算方法：R7年度奇数月6回調定の金額別度数分布に各料金案を当てはめて再計算（税込・年額）'],
        ['#', '通常算定・基本料金帯は水量から再計算、特殊算定（日割・異動等）は現行調定額×改定率'],
        ['料金案', '基本(5㎥まで)', '6〜10㎥', '11〜50㎥', '51㎥〜',
         '漁業集落排水(円)', '公共下水道(円)', '合計(円)', '対現行 増収額(円)', '増減率(%)']]
rows.append(['現行', '1108.8', '40.7', '191.4', '220.0', g_cur, k_cur, base_total, 0, '0.00'])
gp, kp = revenue(gr, plan1_amt, 1.10), revenue(kr, plan1_amt, 1.10)
rows.append(['案1（集計ブック／約+10%）', '1220', '45.5', '211.5', '242', gp, kp, gp + kp,
             gp + kp - base_total, '%.2f' % ((gp + kp) / base_total * 100 - 100)])
for ratio in (1.15, 1.20):
    fn, p = uniform(ratio)
    g2, k2 = revenue(gr, fn, ratio), revenue(kr, fn, ratio)
    rows.append(['参考 一律%+.0f%%' % ((ratio - 1) * 100), p[0] / 2, p[1], p[2], p[3],
                 g2, k2, g2 + k2, g2 + k2 - base_total, '%.2f' % ((g2 + k2) / base_total * 100 - 100)])
rows.append([])
rows.append(['# 参考：集計ブック上の案1集計値（奇数㎥を2㎥ブロックに切上げ）'])
rows.append(['漁業集落排水 案1（ブック値）', '', '', '', '', 8218319, '', '', 8218319 - g_cur,
             '%.2f' % (8218319 / g_cur * 100 - 100)])
rows.append(['漁業集落排水 案1（同一算定方法で再計算）', '', '', '', '', gp, '', '', gp - g_cur,
             '%.2f' % (gp / g_cur * 100 - 100)])
w(OUT + '/data/07_増収試算.csv', rows)

# === 08 実績サマリ ===========================================================
print('08 実績サマリ')
def bounds(recs):
    lo = sum((v if k == NORMAL else 1) * c for k, v, c, _ in recs if k != SPECIAL)
    hi = sum((v if k == NORMAL else 10) * c for k, v, c, _ in recs if k != SPECIAL)
    return lo, hi
glo, ghi = bounds(gr); klo, khi = bounds(kr)
gmet = sum(a for k, _, _, a in gr if k != SPECIAL)
kmet = sum(a for k, _, _, a in kr if k != SPECIAL)
gsp = sum(c for k, _, c, _ in gr if k == SPECIAL); ksp = sum(c for k, _, c, _ in kr if k == SPECIAL)
w(OUT + '/data/08_R7調定実績サマリ.csv', [
    ['#', 'R7年度奇数月6回調定集計ベース。通常の隔月検針者については概ね12か月相当だが、'
          '偶数月調定及び毎月検針者の一部を含まない'],
    ['#', '水量は請求金額から算定式を逆引きした推計値。基本料金帯（2,217円）は1〜10㎥のいずれでも'
          '同額のため水量が特定できず、下限・上限で示す。特殊算定分は推計対象外'],
    ['項目', '漁業集落排水', '公共下水道', '合計'],
    ['調定件数(件)', g_cnt, k_cnt, g_cnt + k_cnt],
    ['調定額(円・税込)', g_cur, k_cur, g_cur + k_cur],
    ['推計水量 下限(㎥)', glo, klo, glo + klo],
    ['推計水量 上限(㎥)', ghi, khi, ghi + khi],
    ['実効単価 下限(円/㎥・税込)', '%.1f' % (gmet / ghi), '%.1f' % (kmet / khi),
     '%.1f' % ((gmet + kmet) / (ghi + khi))],
    ['実効単価 上限(円/㎥・税込)', '%.1f' % (gmet / glo), '%.1f' % (kmet / klo),
     '%.1f' % ((gmet + kmet) / (glo + klo))],
    ['1件平均調定額(円)', '%.0f' % (g_cur / g_cnt), '%.0f' % (k_cur / k_cnt), ''],
    ['水量推計対象外の件数(特殊算定)', gsp, ksp, gsp + ksp],
])

# === 12 水量推計の内訳 =======================================================
print('12 水量推計の内訳')
rows = [['#', '調定を3区分に分けた内訳。実効単価は水量を推計できる区分（通常算定＋基本料金帯）のみで算出'],
        ['事業', '区分', '件数', '調定額(円)', '水量 下限(㎥)', '水量 上限(㎥)', '備考']]
for recs, label in [(gr, '漁業集落排水'), (kr, '公共下水道')]:
    for kind, note in [(NORMAL, '金額から水量が一意に定まる'),
                       (BASIC, '基本料金2,217円。1〜10㎥のいずれでも同額のため水量不詳'),
                       (SPECIAL, '日割・異動等で料金式に一致しない。水量推計は対象外')]:
        c = sum(x[2] for x in recs if x[0] == kind)
        a = sum(x[3] for x in recs if x[0] == kind)
        if kind == NORMAL:
            lo = hi = sum(x[1] * x[2] for x in recs if x[0] == kind)
        elif kind == BASIC:
            lo, hi = c * 1, c * 10
        else:
            lo = hi = ''
        rows.append([label, kind, c, a, lo, hi, note])
w(OUT + '/data/12_水量推計の内訳.csv', rows)

# === 13 v1からの訂正 =========================================================
print('13 訂正履歴')
w(OUT + '/data/13_訂正履歴.csv', [
    ['#', '2026-08-31 レビュー指摘を受けた訂正。増収試算への影響は軽微だが、水量・実効単価は訂正が必要'],
    ['項目', 'v1（2026-08-19）', 'v2（2026-08-31）', '訂正理由'],
    ['漁集 推計水量(㎥)', '41,303', '%s〜%s' % (format(glo, ','), format(ghi, ',')),
     '基本料金帯285件を1㎥として計上していた。1〜10㎥で特定不能のため範囲表示に変更'],
    ['公共 推計水量(㎥)', '193,707', '%s〜%s' % (format(klo, ','), format(khi, ',')),
     '基本料金帯1,657件を1㎥として計上。併せて特殊算定487件を推計対象外とした'],
    ['漁集 実効単価(円/㎥)', '178.5', '%.1f〜%.1f' % (gmet / ghi, gmet / glo), '上記水量の訂正を反映'],
    ['公共 実効単価(円/㎥)', '183.9', '%.1f〜%.1f' % (kmet / khi, kmet / klo), '上記水量の訂正を反映'],
    ['漁集 案1調定額(円)', '8,136,229', format(gp, ','), '特殊算定5件を現行×1.10に統一'],
    ['公共 案1調定額(円)', '39,275,397', format(kp, ','), '特殊算定487件を現行×1.10に統一'],
    ['2事業計 増収額(円)', '4,424,494', format(gp + kp - base_total, ','), '同上（差 %s円）'
     % format(gp + kp - base_total - 4424494, ',')],
    ['調定の表現', 'R7年度年間実績（隔月6回＝12か月分）',
     'R7年度奇数月6回調定集計ベース', '偶数月調定・毎月検針者の一部を含まないため'],
    ['経営戦略との収入比較', '公共 R7予測38,724千円に対し実績35,613千円＝▲8%',
     '単純比較を保留', '税込／税抜の別、調定と決算上の収入の別、集計範囲の欠落が未整理のため'],
])

# === 09 モデルケース =========================================================
print('09 モデルケース')
def monthly(v, b, u1, u2, u3):
    n = b
    if v > 5:  n += u1 * min(v - 5, 5)
    if v > 10: n += u2 * min(v - 10, 40)
    if v > 50: n += u3 * (v - 50)
    return n
rows = [['#', '一般汚水。1か月あたりの月額と、隔月検針による2か月分請求額'],
        ['月使用量(㎥)', '現行 月額(円)', '案1 月額(円)', '差(円)',
         '現行 2か月請求(円)', '案1 2か月請求(円)', '差(円)', '増減率(%)']]
for v in [5, 8, 10, 12, 15, 20, 25, 30, 40, 50, 60, 80, 100]:
    c = monthly(v, 1108.8, 40.7, 191.4, 220.0); p = monthly(v, 1220, 45.5, 211.5, 242)
    c2, p2 = math.floor(c * 2), math.floor(p * 2)
    rows.append([v, '%.1f' % c, '%.1f' % p, '%+.1f' % (p - c), c2, p2, p2 - c2, '%.2f' % ((p / c - 1) * 100)])
w(OUT + '/data/09_モデルケース.csv', rows)

# === 10 経営戦略 抜粋 ========================================================
print('10 経営戦略 抜粋')
w(OUT + '/data/10_経営戦略_抜粋数値.csv', [
    ['#', '出典：階上町下水道事業経営戦略（改定）令和7年2月'],
    ['区分', '項目', '公共下水道', '漁業集落排水', '出典'],
    ['現状(R4)', '収益的収支比率(%)', '88.3', '84.5', '表2.15/2.16'],
    ['現状(R4)', '経費回収率(%)', '35.0', '25.8', '表2.15/2.16'],
    ['現状(R4)', '汚水処理原価(円/㎥)', '496.84', '662.35', '表2.15/2.16'],
    ['現状(R4)', '水洗化率(%)', '63.9', '69.3', '表2.15/2.16'],
    ['現状(R4)', '類似団体平均 経費回収率(%)', '48.9', '36.0', '表2.15/2.16'],
    ['目標(R16)', '収益的収支比率(%)', '100', '100', '表5.19'],
    ['目標(R16)', '経費回収率(%)', '40', '26', '表5.19'],
    ['目標(R16)', '水洗化率(%)', '75', '80', '表5.19'],
    ['改善方策', '採用ケース', 'ケース①', 'ケース②', '5.6'],
    ['改善方策', 'ケース①の内容', '5年毎に10%ずつ使用料を上昇', '同左', '5.6'],
    ['改善方策', 'ケース②の内容', '—', 'ケース①＋R16水洗化率80%', '5.6'],
    ['改善方策', '改定後 経費回収率(R16,%)', '47.7', '26.4', '表5.17/5.18'],
    ['単価前提', '使用料単価 採用値(円/㎥)', '173.9', '171.1', '表3.5（R1〜R4実績平均）'],
    [],
    ['# 使用料収入の推移（千円）／表5.13・5.22（公共）、表5.14・5.25（漁集）'],
    ['系列', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13', 'R14', 'R15', 'R16'],
    ['公共 現況継続', 35257, 38724, 40421, 40468, 40484, 40484, 40453, 40421, 40358, 40280, 40185],
    ['公共 採用(ケース①)', 35257, 38724, 44463, 44515, 44532, 44532, 44498, 48505, 48430, 48336, 48222],
    ['漁集 現況継続', 7697, 7014, 6843, 6672, 6501, 6330, 5987, 5816, 5645, 5474, 5303],
    ['漁集 採用(ケース②)', 7697, 7197, 7810, 7686, 7544, 7420, 7278, 7765, 7610, 7436, 7281],
    [],
    ['# 年間有収水量の推計（㎥/年）／表3.4'],
    ['系列', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13', 'R14', 'R15', 'R16'],
    ['公共下水道', 213135, 222618, 232372, 232642, 232733, 232733, 232553, 232372, 232011, 231559, 231017],
    ['漁業集落排水', 42167, 41035, 39903, 38771, 37639, 36601, 35469, 34337, 33206, 32074, 31036],
])

# === 11 元ファイル一覧（ハッシュ） ==========================================
print('11 元ファイル一覧')
rows = [['#', '試算に用いた元ファイル。個人情報を含むため本zipには同梱していない'],
        ['区分', 'ファイル名', 'サイズ(byte)', 'SHA-256', 'シート構成']]
for key, (fn, desc) in FILES.items():
    path = os.path.join(SRC, fn)
    h = hashlib.sha256(io.open(path, 'rb').read()).hexdigest()
    sheets = ''
    if fn.endswith('.xlsx'):
        sheets = ' / '.join(openpyxl.load_workbook(path, read_only=True).sheetnames)
    else:
        sheets = '80ページ'
    rows.append([desc, fn, os.path.getsize(path), h, sheets])
w(OUT + '/data/11_元ファイル一覧.csv', rows)

print('done')
