# -*- coding: utf-8 -*-
"""地域包括ケア「見える化」システム出力（北塩原村）Excel群 → tidy形式CSV
   ※ 大雪地区広域連合（比較対象の設定残り）は全面的に除外する。
"""
import openpyxl, glob, os, csv, re, sys

EXCLUDE_REGION = ("大雪地区広域", "大雪地区")
SRC = "mieruka8"
ERA = {"令和": 2018, "平成": 1988, "昭和": 1925}
ERA_ABBR = {"R": 2018, "H": 1988, "S": 1925}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "-", "－", "―", "‐"):
            return None
    return v


def is_filler(s):
    return bool(s) and bool(re.fullmatch(r"[行列]\d+", str(s).strip()))


def parse_period(raw):
    """ヘッダ文字列 → (label, 西暦年, 種別)。判定できなければ (label, None, None)。"""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or is_filler(s):
        return None
    label = re.sub(r"\s+", "", s)

    # 2000 / 2000.0
    m = re.fullmatch(r"(\d{4})(?:\.0)?", label)
    if m and 1990 < int(m.group(1)) < 2100:
        return (m.group(1), int(m.group(1)), "暦年")

    # 令和2年3月末 / 令和元年3月末 / 令和7年3月末時点 / 平成26年4月
    m = re.match(r"(令和|平成|昭和)(元|\d+)年(\d+)月(末)?", label)
    if m:
        era, y, mo, sue = m.group(1), m.group(2), int(m.group(3)), m.group(4)
        yr = ERA[era] + (1 if y == "元" else int(y))
        return (label, yr, "月末時点" if sue else "各月")

    # 令和2年度 / 平成24年度 / 令和元年度
    m = re.match(r"(令和|平成|昭和)(元|\d+)年度", label)
    if m:
        era, y = m.group(1), m.group(2)
        yr = ERA[era] + (1 if y == "元" else int(y))
        return (label, yr, "年度")

    # R6（R7/2月サービス提供分まで） / R2 / H26 / R元
    m = re.match(r"([RHS])(元|\d+)", label)
    if m:
        ab, y = m.group(1), m.group(2)
        yr = ERA_ABBR[ab] + (1 if y == "元" else int(y))
        return (label, yr, "年度")

    return None


def excluded(name):
    return any(x in str(name) for x in EXCLUDE_REGION)


rows, inventory, warns = [], [], []

for path in sorted(glob.glob(f"{SRC}/*/*.xlsx")):
    folder, fname = os.path.basename(os.path.dirname(path)), os.path.basename(path)
    code = fname.split("_")[0]
    wb = openpyxl.load_workbook(path, data_only=True)
    kind = "時系列" if "表形式（時系列）" in wb.sheetnames else "地域別"
    ws = wb[f"表形式（{kind}）"]
    grid = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    wb.close()

    out_date = None
    if grid and grid[0] and grid[0][0]:
        m = re.search(r"出力日[:：](\S+)", str(grid[0][0]))
        if m:
            out_date = m.group(1)

    hdr = grid[1] if len(grid) > 1 else ()
    n_excl = 0

    if kind == "時系列":
        # B列=地域, C列=指標, D列=単位, E列以降=期間
        periods = []
        for ci in range(4, len(hdr)):
            p = parse_period(clean(hdr[ci]))
            if p:
                periods.append((ci, *p))
        if not periods:
            warns.append(f"[期間ヘッダ検出0] {fname}")
        for r in grid[2:]:
            region = clean(r[1]) if len(r) > 1 else None
            indicator = clean(r[2]) if len(r) > 2 else None
            unit = clean(r[3]) if len(r) > 3 else None
            if not region or is_filler(region):
                continue
            if excluded(region):
                n_excl += 1
                continue
            for ci, label, yr, ptype in periods:
                val = clean(r[ci]) if ci < len(r) else None
                if val is None:
                    continue
                rows.append(dict(folder=folder, file=fname, code=code, kind=kind,
                                 indicator=indicator, unit=unit, region=region,
                                 period=label, year=yr, period_type=ptype, value=val))
    else:
        # B列=指標, C列=単位, D列以降=地域
        regions = []
        for ci in range(3, len(hdr)):
            h = clean(hdr[ci])
            if h and not is_filler(h):
                regions.append((ci, str(h).strip()))
        yr = None
        m = re.search(r"_(\d{4})_地域別", fname)
        if m:
            yr = int(m.group(1))
        for r in grid[2:]:
            indicator = clean(r[1]) if len(r) > 1 else None
            unit = clean(r[2]) if len(r) > 2 else None
            if not indicator or is_filler(indicator):
                continue
            for ci, reg in regions:
                if excluded(reg):
                    n_excl += 1
                    continue
                val = clean(r[ci]) if ci < len(r) else None
                if val is None:
                    continue
                rows.append(dict(folder=folder, file=fname, code=code, kind=kind,
                                 indicator=indicator, unit=unit, region=reg,
                                 period=str(yr) if yr else "", year=yr,
                                 period_type="地域比較", value=val))

    mine = [r for r in rows if r["file"] == fname]
    yrs = sorted({r["year"] for r in mine if r["year"]})
    inventory.append(dict(
        folder=folder, code=code, file=fname, kind=kind, out_date=out_date,
        n_indicators=len({r["indicator"] for r in mine}),
        indicators="／".join(sorted({str(r["indicator"]) for r in mine})),
        regions="／".join(sorted({r["region"] for r in mine})),
        year_min=yrs[0] if yrs else "", year_max=yrs[-1] if yrs else "",
        excluded_series=n_excl, n_records=len(mine)))

FIELDS = ["folder", "file", "code", "kind", "indicator", "unit", "region", "period", "year", "period_type", "value"]
with open("mieruka8_tidy.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=FIELDS)
    w.writeheader(); w.writerows(rows)

with open("mieruka8_inventory.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=["folder", "code", "file", "kind", "out_date", "n_indicators",
                                      "indicators", "regions", "year_min", "year_max",
                                      "excluded_series", "n_records"])
    w.writeheader(); w.writerows(inventory)

print(f"files={len(inventory)}  records={len(rows)}  除外系列(大雪地区広域)={sum(i['excluded_series'] for i in inventory)}")
print("採用地域:", sorted({r['region'] for r in rows}))
empty = [i['file'] for i in inventory if i['n_records'] == 0]
print(f"データ0件のファイル: {len(empty)}")
for e in empty: print("   -", e)
for w_ in warns: print(w_)
