import openpyxl, json
FILES={
 '令和6年度':'/root/.claude/uploads/134138ca-61f7-57d3-9e9b-5f081a1a345d/00bbc0fc-001474341.xlsx',
 '令和7年度':'/root/.claude/uploads/134138ca-61f7-57d3-9e9b-5f081a1a345d/8283fa24-001732645.xlsx',
 '令和8年度':'/root/.claude/uploads/134138ca-61f7-57d3-9e9b-5f081a1a345d/45c5049a-001732614_2.xlsx',
}
STATS=('配点','全国合計','平均点','項目平均','平均得点率','中央値','標準偏差','該当市町村数','該当率')
def norm(v):
    return str(v).replace('　',' ').replace('\n',' ').strip() if v is not None else None
def num(v):
    try: return float(v)
    except: return None
def load(year):
    wb=openpyxl.load_workbook(FILES[year],data_only=True); ws=wb[wb.sheetnames[0]]
    labcol=None; stat={}
    for c in range(5,12):
        s={}
        for r in range(1,30):
            l=norm(ws.cell(r,c).value)
            if l in STATS: s[l]=r
        if len(s)>len(stat): stat=s; labcol=c
    hdr_end=max(stat.values())
    hier_rows=list(range(2,min(stat.values())))
    tgt=None; namecol=None
    for r in range(hdr_end,ws.max_row+1):
        for c in range(6,10):
            if norm(ws.cell(r,c).value)=='北塩原村': tgt=r; namecol=c
        if tgt: break
    # データ開始列 = 交付金名(保険者機能強化推進交付金)が現れる列
    start=None
    for c in range(8,40):
        for r in hier_rows:
            if norm(ws.cell(r,c).value)=='保険者機能強化推進交付金': start=c
        if start: break
    ff={r:None for r in hier_rows}
    out=[]
    for c in range(start,ws.max_column+1):
        for r in hier_rows:
            v=norm(ws.cell(r,c).value)
            if v:
                ff[r]=v
                for r2 in hier_rows:
                    if r2>r: ff[r2]=None
        rec={'col':c,'path':[ff[r] for r in hier_rows if ff[r]],'val':num(ws.cell(tgt,c).value)}
        for k in STATS:
            rec[k]=num(ws.cell(stat[k],c).value) if k in stat else None
        out.append(rec)
    return out,tgt,start,stat
if __name__=='__main__':
    for y in FILES:
        out,tgt,start,stat=load(y)
        json.dump(out,open(f'/tmp/kofukin_detail_{y}.json','w'),ensure_ascii=False)
        print(y,'村行',tgt,'開始列',start,'項目数',len(out),'統計行',stat)
