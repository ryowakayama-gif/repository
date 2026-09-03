# -*- coding: utf-8 -*-
"""第10期北塩原村 アンケート調査 集計仕様書（xlsx）
   shukei_data.py から生成。"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from shukei_data import N, Z, DERIVED, CROSS, AXES

OUT = "/home/user/repository/output/07_北塩原村第10期_集計仕様書.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)
F = "游ゴシック"
C = {"header":"1F3864","sub":"2E75B6","band":"DDEBF7","alt":"F7FAFC",
     "note":"FFF3F3","key":"FCE4D6","white":"FFFFFF","own":"FFF2CC","chk":"E2EFDA"}
THIN = Side(border_style="thin", color="BFBFBF")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def title_bar(ws, rng, text, size=14, h=30):
    ws.merge_cells(rng); c = ws[rng.split(":")[0]]
    c.value = text; c.font = Font(name=F, size=size, bold=True, color=C["white"])
    c.fill = PatternFill("solid", fgColor=C["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int("".join(x for x in rng.split(":")[0] if x.isdigit()))].height = h

def sub_bar(ws, rng, text, h=20, fill=None):
    ws.merge_cells(rng); c = ws[rng.split(":")[0]]
    c.value = text; c.font = Font(name=F, size=9)
    c.fill = PatternFill("solid", fgColor=fill or C["band"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[int("".join(x for x in rng.split(":")[0] if x.isdigit()))].height = h

def head_row(ws, row, heads, widths, h=34):
    for i, hh in enumerate(heads, 1):
        c = ws.cell(row=row, column=i, value=hh)
        c.font = Font(name=F, size=10, bold=True, color=C["white"])
        c.fill = PatternFill("solid", fgColor=C["sub"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BD
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = h

def put(ws, r, vals, wraps=(), center=(), fill=None, bold=()):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=F, size=9, bold=(i in bold))
        c.border = BD
        c.alignment = Alignment(vertical="top", wrap_text=(i in wraps),
                                horizontal="center" if i in center else "left")
        if fill: c.fill = PatternFill("solid", fgColor=fill)

wb = Workbook()

# ── Sheet1 方針・凡例 ───────────────────────────
ws = wb.active; ws.title = "方針・凡例"
ws.column_dimensions["A"].width = 3
for col, w in zip("BCDEF", (16, 22, 46, 30, 26)): ws.column_dimensions[col].width = w
title_bar(ws, "B1:F1", "第10期北塩原村 アンケート調査　集計仕様書")
sub_bar(ws, "B2:F2", "作成：令和8年9月3日／根拠：調査票 第1稿（0828版）＋doc23の修正案＋doc24 骨子案／"
                     "設問番号は doc23 の修正（ニーズ問7(7)の復活・在宅問13の復活）を反映した確定案です。", h=32)
r = 4
sub_bar(ws, f"B{r}:F{r}", "■ 集計の基本方針", h=22, fill=C["key"]); r += 1
POLICY = [
 ("1", "見える化への登録を最優先", "必須項目・オプション項目は設問文・選択肢・順序を標準どおりとし、集計区分も標準に合わせます。独自の区分を作ると全国・県との比較ができなくなります。"),
 ("2", "全設問を同じ形式で出す", "全体／地区別／年齢階級別／性別／前回比較（可能なもの）の5点セットを全設問で統一します。"),
 ("3", "地区別集計を必須とする", "4地区は地理的に分断されており、移動手段・外出抑制・除雪は地区差が大きいと想定されます。施策の地区別展開の根拠になります。"),
 ("4", "端数処理の統一", "構成比は小数第1位まで（第2位を四捨五入）。合計が100.0%にならない場合は「四捨五入のため合計が100%にならない場合があります」と注記します。"),
 ("5", "n数の明示", "すべての図表にn数を記載します。n<30 の区分は参考値である旨を明示し、単独の根拠としません。"),
 ("6", "複数回答の扱い", "回答者数を分母とした％で表示し、「複数回答のため合計が100%を超えます」と注記します。"),
 ("7", "無回答の扱い", "無回答は集計対象に含め、分母から除外しません。ただし前回比較の際は前回の扱いに合わせます。"),
 ("8", "認定データとの紐付け", "管理番号により認定状況を付与します。これがないと分析編の中核（X-02）が作れません。"),
]
head_row(ws, r, ["", "#", "方針", "内容", "", ""], [3, 6, 22, 70, 2, 2], h=22)
r += 1
for no, t, d in POLICY:
    ws.cell(row=r, column=2, value=no).font = Font(name=F, size=9, bold=True)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r, column=2).border = BD
    ws.cell(row=r, column=3, value=t).font = Font(name=F, size=9, bold=True)
    ws.cell(row=r, column=3).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=r, column=3).border = BD
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    c = ws.cell(row=r, column=4, value=d)
    c.font = Font(name=F, size=9); c.alignment = Alignment(vertical="top", wrap_text=True)
    for i in range(4, 7): ws.cell(row=r, column=i).border = BD
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
sub_bar(ws, f"B{r}:F{r}", "■ 集計軸（クロス集計の区分）", h=22, fill=C["key"]); r += 1
head_row(ws, r, ["", "コード", "軸", "区分", "適用", ""], [3, 10, 22, 46, 30, 2], h=22)
r += 1
for code, name, div, apply in AXES:
    put(ws, r, [None, code, name, div, apply], wraps=(4, 5), center=(2,))
    ws.row_dimensions[r].height = 20
    r += 1

r += 1
sub_bar(ws, f"B{r}:F{r}", "■ シートの構成", h=22, fill=C["key"]); r += 1
SHEETS = [
 ("ニーズ調査 集計一覧", f"{len(N)}設問。設問ごとの集計方法とクロス軸"),
 ("在宅調査 集計一覧", f"{len(Z)}設問。A票・B票"),
 ("派生変数", f"{len(DERIVED)}件。リスク判定・合成変数の定義"),
 ("クロス集計定義", f"{len(CROSS)}件。分析編で作成するクロス表"),
 ("データレイアウト", "入力データの列構成"),
]
for nm, ds in SHEETS:
    put(ws, r, [None, nm, ds], wraps=(3,))
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for i in range(3, 7): ws.cell(row=r, column=i).border = BD
    ws.row_dimensions[r].height = 20
    r += 1

r += 1
NOTES = [
 "※ 設問数の内訳：ニーズ調査 必須34・オプション32・村独自14（枝番を1設問として数えた場合）。手引きの「必須35問／オプション30問」とは数え方が異なります。",
 "※ 派生変数のリスク判定の該当基準には【要確認】を付しています。手引きに明記があるのはBMI 18.5以下のみで、他は見える化システムの算出方法との突合が必要です（確認事項K-91）。",
 "※ ★印は分析編（doc24 第Ⅱ部）で中心的に用いる設問・クロス表です。",
 "※ 村独自設問は見える化への登録対象外（「−」）です。集計は行いますが全国・県との比較はできません。",
]
for nt in NOTES:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=nt); c.font = Font(name=F, size=8)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 24
    r += 1

# ── Sheet2/3 集計一覧 ────────────────────────────
def sheet_list(title, rows, heads, widths, is_needs):
    w = wb.create_sheet(title)
    w.column_dimensions["A"].width = 3
    title_bar(w, f"A1:{get_column_letter(len(heads))}1", title)
    sub_bar(w, f"A2:{get_column_letter(len(heads))}2",
            "★＝分析編で中心的に用いる設問／黄色＝村独自設問（見える化への登録対象外）／"
            "クロス軸のコードは「方針・凡例」シートを参照", h=22)
    head_row(w, 3, heads, widths)
    rr = 4
    for i, row in enumerate(rows, 1):
        if is_needs:
            _, q, e, text, kb, fm, sh, cr, mr, note = row
            vals = [i, q, e, text, kb, fm, sh, cr, mr, note]
        else:
            _, hyo, q, text, kb, fm, sh, cr, mr, note = row
            vals = [i, hyo, q, text, kb, fm, sh, cr, mr, note]
        fill = C["own"] if kb == "村独自" else (C["alt"] if i % 2 == 0 else None)
        put(w, rr, vals, wraps=(4, 7, 10), center=(1, 2, 3, 5, 6, 8, 9), fill=fill)
        w.row_dimensions[rr].height = 30
        rr += 1
    w.freeze_panes = "A4"
    w.auto_filter.ref = f"A3:{get_column_letter(len(heads))}{rr-1}"
    return w

sheet_list("ニーズ調査 集計一覧",
           N,
           ["No.", "問", "枝番", "設問文", "区分", "形式", "集計方法", "クロス軸", "見える化", "備考・ICFの領域"],
           [5, 6, 7, 40, 8, 8, 30, 14, 7, 44], True)
sheet_list("在宅調査 集計一覧",
           Z,
           ["No.", "票", "問", "設問文", "区分", "形式", "集計方法", "クロス軸", "見える化", "備考"],
           [5, 6, 8, 40, 8, 8, 30, 12, 7, 44], False)

# ── Sheet4 派生変数 ────────────────────────────
w = wb.create_sheet("派生変数")
w.column_dimensions["A"].width = 3
title_bar(w, "A1:F1", "派生変数の定義（リスク判定・合成変数）")
sub_bar(w, "A2:F2", "【要確認】を付した該当基準は、手引きに明記がなく見える化システムの算出方法との突合が必要です（確認事項K-91）。"
                    "独自の基準を作ると全国・県との比較ができなくなるため、必ず見える化の定義に合わせます。", h=32)
head_row(w, 3, ["No.", "コード", "変数名", "用いる設問", "該当基準・算出方法", "区分"], [5, 10, 24, 22, 60, 12])
rr = 4
for i, (code, name, src, rule, kind) in enumerate(DERIVED, 1):
    fill = C["note"] if "【要確認】" in rule else (C["alt"] if i % 2 == 0 else None)
    put(w, rr, [i, code, name, src, rule, kind], wraps=(3, 4, 5), center=(1, 2, 6), fill=fill)
    w.row_dimensions[rr].height = 28
    rr += 1
w.freeze_panes = "A4"

# ── Sheet5 クロス集計定義 ──────────────────────────
w = wb.create_sheet("クロス集計定義")
w.column_dimensions["A"].width = 3
title_bar(w, "A1:G1", "分析編で作成するクロス表の定義")
sub_bar(w, "A2:G2", "★＝doc24 第Ⅱ部（分析編）の中核。X-01・X-02・X-03・X-04 は第2回策定委員会の資料の骨格になります。", h=22)
head_row(w, 3, ["No.", "コード", "クロス表", "行（表側）", "列（表頭）", "調査", "用途・出典"],
         [5, 8, 34, 22, 22, 8, 46])
rr = 4
for i, (code, name, row_v, col_v, survey, use) in enumerate(CROSS, 1):
    fill = C["key"] if use.startswith("★") else (C["alt"] if i % 2 == 0 else None)
    put(w, rr, [i, code, name, row_v, col_v, survey, use], wraps=(3, 4, 5, 7), center=(1, 2, 6), fill=fill)
    w.row_dimensions[rr].height = 26
    rr += 1
w.freeze_panes = "A4"

# ── Sheet6 データレイアウト ─────────────────────────
w = wb.create_sheet("データレイアウト")
w.column_dimensions["A"].width = 3
title_bar(w, "A1:F1", "入力データのレイアウト")
sub_bar(w, "A2:F2", "紙回答の入力データとWeb回答のエクスポートを同一のレイアウトに揃えます。"
                    "見える化システムへの登録時は標準様式の設問順に並べ替え直す必要があります（手引き Ⅲ）。", h=32)
head_row(w, 3, ["列", "項目", "型", "値", "出所", "備考"], [8, 26, 10, 34, 18, 46])
LAYOUT = [
 ("1", "管理番号", "文字", "村が付番。氏名は含めない", "村（宛名ラベル）", "★被保険者番号との紐付けの鍵。調査票・Webフォームの双方に印字／入力"),
 ("2", "回答方法", "文字", "紙／Web", "システム", "重複回答の排除に使用"),
 ("3", "受付日", "日付", "YYYY/MM/DD", "受託者", "小分け送付の受渡し記録と突合"),
 ("4", "調査種別", "文字", "ニーズ／在宅", "―", ""),
 ("5", "地区", "文字", "北山／大塩／桧原／裏磐梯", "名簿・問0", "集計軸B"),
 ("6", "性別", "文字", "男性／女性", "名簿", "集計軸D"),
 ("7", "生年月／年齢", "数値", "基準日（R8.9.1）時点の年齢", "名簿", "集計軸C。年齢階級を導出"),
 ("8", "認定状況", "文字", "認定なし／要支援1／要支援2／要介護1〜5", "★認定データ", "集計軸E・I。K-87で提供を依頼中"),
 ("9", "認定日", "日付", "YYYY/MM/DD", "★認定データ", "新規／更新の別の把握に使用"),
 ("10〜", "各設問の回答", "数値・文字", "単一＝選択肢番号／複数＝設問ごとに0-1のダミー列／数値＝実数／自由記述＝原文", "調査票・Webフォーム", "複数回答は選択肢ごとに列を立てる（横持ち）"),
 ("末尾", "派生変数", "数値・文字", "R01〜R10・R99・D01・P01・S01・S02・S99・L01・B01", "算出", "「派生変数」シートの定義による"),
]
rr = 4
for i, row in enumerate(LAYOUT, 1):
    fill = C["note"] if "★" in row[4] or "★" in row[5] else (C["alt"] if i % 2 == 0 else None)
    put(w, rr, list(row), wraps=(4, 6), center=(1, 3), fill=fill)
    w.row_dimensions[rr].height = 30
    rr += 1

rr += 1
sub_bar(w, f"A{rr}:F{rr}", "■ 入力上の留意点", h=22, fill=C["key"]); rr += 1
TIPS = [
 "複数回答は「選択肢ごとに1列（0／1）」の横持ちとします。カンマ区切りの1列にまとめると集計時に分解が必要になり、誤りが混入します。",
 "自由記述は原文をそのまま保持し、分類コードは別列に付与します。原文を書き換えないでください。",
 "身長・体重（問3(1)）は単位を統一し、明らかな入力誤り（身長20cm等）は欠測として扱い、その旨を記録します。",
 "幸福度（問8(2)）は0〜10の整数。範囲外は欠測とします。",
 "Web回答は必須入力の設定により無回答が発生しにくく、紙回答と無回答率が構造的に異なります。回答方法別の集計を必ず1回は行い、差が大きい設問は本文で注記します。",
 "小分け送付（村が100件程度ごとに送付）に対応するため、受付日とロット番号を記録し、入力の進捗を管理します。",
]
for t in TIPS:
    w.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    c = w.cell(row=rr, column=1, value="・" + t)
    c.font = Font(name=F, size=9); c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    w.row_dimensions[rr].height = 26
    rr += 1

wb.save(OUT)
print(f"保存: {OUT}")
print(f"ニーズ {len(N)}設問 / 在宅 {len(Z)}設問 / 派生変数 {len(DERIVED)}件 / クロス表 {len(CROSS)}件")
