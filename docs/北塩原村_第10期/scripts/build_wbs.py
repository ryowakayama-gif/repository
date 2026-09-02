# -*- coding: utf-8 -*-
"""第10期北塩原村高齢者福祉計画・介護保険事業計画策定業務 WBS／進捗管理表（Ver.2）
   wbs_data.W（仕様書の分解）＋ wbs_progress.P（実績反映）＋ wbs_kakunin.K（村への確認事項）から生成。
   基準日：令和8年8月31日"""
import os, sys, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from wbs_data import W
from wbs_progress import P, BASE_DATE
from wbs_kakunin import K, SOLVED

OUT = "/home/user/repository/output/05_北塩原村第10期_WBS進捗管理表.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)
F = "游ゴシック"
C = {"header":"1F3864","sub":"2E75B6","band":"DDEBF7","alt":"F7FAFC",
     "note":"FFF3F3","vill":"FFF2CC","key":"FCE4D6","white":"FFFFFF",
     "done":"E2EFDA","doing":"FFF2CC","chk":"FCE4D6","prioS":"FF5050","prioA":"FFC000",
     "prioB":"FFE699","prioC":"D9D9D9","solved":"E7E6E6"}
THIN = Side(border_style="thin", color="BFBFBF")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
D = lambda s: datetime.datetime.strptime(s, "%Y/%m/%d") if s else None
TANTO = {"受":"受託者","村":"北塩原村","双":"双方"}

def title_bar(ws, rng, text, size=14, h=30):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = text
    c.font = Font(name=F, size=size, bold=True, color=C["white"])
    c.fill = PatternFill("solid", fgColor=C["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int("".join(ch for ch in rng.split(":")[0] if ch.isdigit()))].height = h

def sub_bar(ws, rng, text, h=20, fill=None):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = text
    c.font = Font(name=F, size=9)
    c.fill = PatternFill("solid", fgColor=fill or C["band"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[int("".join(ch for ch in rng.split(":")[0] if ch.isdigit()))].height = h

def head_row(ws, row, heads, widths, h=34):
    for i, hh in enumerate(heads, 1):
        c = ws.cell(row=row, column=i, value=hh)
        c.font = Font(name=F, size=10, bold=True, color=C["white"])
        c.fill = PatternFill("solid", fgColor=C["sub"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BD
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = h

wb = Workbook()

# ══════════════════ Sheet1: WBS・進捗管理 ══════════════════
ws = wb.active; ws.title = "WBS・進捗管理"
HEAD = ["No.","大分類","中分類","作業項目","仕様書該当","担当","成果物・アウトプット","前提・依存",
        "想定時期","予定開始","予定完了","実績開始","実績完了","進捗率","ステータス","備考"]
WID  = [9,16,17,42,11,9,26,34,11,12,12,12,12,8,10,52]

title_bar(ws, "A1:P1", "第10期北塩原村高齢者福祉計画・第10期北塩原村介護保険事業計画　策定業務　WBS／進捗管理表")
sub_bar(ws, "A2:P2",
    "委託第27号／業務期間：契約締結の日〜令和9年3月31日／計画期間：令和9〜11年度　"
    "｜　出典：仕様書（8北保福第483号 別紙）4「業務の内容」を作業レベルに分解　"
    f"｜　進捗基準日：令和8年8月31日")

HR = 4
head_row(ws, HR, HEAD, WID)

r = HR + 1
prev_major = None
rows_by_no = {}
for major, mid, task, spec, tanto, deliv, dep, timing in W:
    no = f"{major.split()[0]}-{r-HR:02d}"
    st, pct, s_st, s_ed, note = P.get(task, ("未着手", 0.0, None, None, ""))
    vals = [no, major, mid, task, spec, TANTO[tanto], deliv, dep, timing,
            None, None, D(s_st), D(s_ed), pct, st, note]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=F, size=9)
        c.border = BD
        c.alignment = Alignment(vertical="top", wrap_text=(i in (4,7,8,16)),
                                horizontal="center" if i in (1,5,6,9,14,15) else "left")
    # 行の塗り：ステータス優先、次いで村担当、次いで縞
    if st == "完了":
        fill = C["done"]
    elif st in ("着手", "確認中"):
        fill = C["doing"] if tanto != "村" else C["vill"]
    elif tanto == "村":
        fill = C["vill"]
    elif (r - HR) % 2 == 0:
        fill = C["alt"]
    else:
        fill = None
    if fill:
        for i in range(1, 17):
            ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=fill)
    if major != prev_major:
        ws.cell(row=r, column=2).font = Font(name=F, size=9, bold=True, color=C["header"])
        prev_major = major
    ws.cell(row=r, column=14).number_format = "0%"
    ws.cell(row=r, column=15).font = Font(name=F, size=9, bold=(st != "未着手"))
    for i in (10, 11, 12, 13):
        ws.cell(row=r, column=i).number_format = "yyyy/mm/dd"
    ws.row_dimensions[r].height = 30
    rows_by_no[no] = r
    r += 1
LAST = r - 1

dv = DataValidation(type="list", formula1='"未着手,着手,確認中,完了,保留,対象外"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"O{HR+1}:O{LAST}")
dv2 = DataValidation(type="list", formula1='"受託者,北塩原村,双方"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"F{HR+1}:F{LAST}")

ws.freeze_panes = "E5"
ws.auto_filter.ref = f"A{HR}:P{LAST}"
ws.sheet_view.zoomScale = 90

# ══════════════════ Sheet2: 村への確認事項 ══════════════════
wk = wb.create_sheet("村への確認事項")
KH = ["優先度","区分","確認事項","なぜ必要か／影響","関連WBS No.","出所","状況","回答内容・回答日"]
KW = [8, 12, 46, 56, 14, 16, 10, 30]
title_bar(wk, "A1:H1", "北塩原村への確認事項　一覧（基準日：令和8年8月31日）")
sub_bar(wk, "A2:H2",
    "優先度S＝9月3日の調査票第1稿提出・9月9日の校了・9月18日の発送を守るために決着が必要（キックオフ議事録で未決着のもの）／"
    "A＝令和8年9月中に確定が必要／B＝第1回策定委員会（11月）まで／C＝第2回以降　"
    "｜　出所の doc番号は docs/北塩原村_第10期/ 配下の分析資料に対応します")
KHR = 4
head_row(wk, KHR, KH, KW)
kr = KHR + 1
for kubun, item, why, wbsno, src, prio, sts in sorted(K, key=lambda x: ("SABC".index(x[5]), x[0])):
    vals = [prio, kubun, item, why, wbsno, src, sts, ""]
    for i, v in enumerate(vals, 1):
        c = wk.cell(row=kr, column=i, value=v)
        c.font = Font(name=F, size=9)
        c.border = BD
        c.alignment = Alignment(vertical="top", wrap_text=(i in (3,4,8)),
                                horizontal="center" if i in (1,5,6,7) else "left")
    wk.cell(row=kr, column=1).fill = PatternFill("solid", fgColor=C["prio"+prio])
    wk.cell(row=kr, column=1).font = Font(name=F, size=10, bold=True)
    if prio in ("S", "A"):
        for i in range(2, 9):
            wk.cell(row=kr, column=i).fill = PatternFill("solid", fgColor=C["note"])
    wk.row_dimensions[kr].height = 34
    kr += 1
KLAST = kr - 1

# 解決済みブロック
kr += 1
sub_bar(wk, f"A{kr}:H{kr}", "■ 解決済み（記録）", fill=C["sub"])
wk[f"A{kr}"].font = Font(name=F, size=10, bold=True, color=C["white"])
kr += 1
for kubun, item, why, wbsno, src, prio, sts in SOLVED:
    for i, v in enumerate([prio, kubun, item, why, wbsno, src, sts, ""], 1):
        c = wk.cell(row=kr, column=i, value=v)
        c.font = Font(name=F, size=9)
        c.border = BD
        c.fill = PatternFill("solid", fgColor=C["solved"])
        c.alignment = Alignment(vertical="top", wrap_text=(i in (3,4,8)),
                                horizontal="center" if i in (1,5,6,7) else "left")
    wk.row_dimensions[kr].height = 34
    kr += 1

dv3 = DataValidation(type="list", formula1='"未依頼,依頼済,回答待ち,回答済,解決済"', allow_blank=True)
dv4 = DataValidation(type="list", formula1='"S,A,B,C"', allow_blank=True)
wk.add_data_validation(dv4); dv4.add(f"A{KHR+1}:A{KLAST}")
wk.add_data_validation(dv3); dv3.add(f"G{KHR+1}:G{KLAST}")
wk.freeze_panes = "C5"
wk.auto_filter.ref = f"A{KHR}:H{KLAST}"
wk.sheet_view.zoomScale = 90

# ══════════════════ Sheet3: 受領資料・データ管理 ══════════════════
wd = wb.create_sheet("受領資料・データ管理")
DH = ["区分","資料・データ名","内容・数量","受領日","整理先","状況"]
DW = [16, 44, 46, 12, 40, 10]
title_bar(wd, "A1:F1", "受領資料・データ管理（基準日：令和8年8月31日）")
sub_bar(wd, "A2:F2", "仕様書4Ⅲ「資料授受の管理」に対応。★＝未受領で業務進行に影響するもの")
DHR = 4
head_row(wd, DHR, DH, DW, h=24)
DOCS = [
 ("仕様書","委託第27号 仕様書（8北保福第483号 別紙）","業務内容・成果品・その他条件","2026/08/25","docs/…/01_業務内容整理.md","受領済"),
 ("現行計画","第9期北塩原村高齢者福祉計画・介護保険事業計画","本文104頁・全6章＋資料編","2026/08/25","docs/…/01_業務内容整理.md","受領済"),
 ("現行計画","第9期計画 概要版（表・裏の2ページ）","基本理念・基本目標・数値①〜⑦・施策体系図","2026/08/31","docs/…/08_第9期計画概要_計画実績突合.md","受領済"),
 ("現行計画","★北塩原村第五次総合振興計画（2017〜2026）","上位計画。第六次の策定状況も要確認","―","―","未受領"),
 ("現行計画","★第3次健康きたしおばら21","関連計画","―","―","未受領"),
 ("調査票","キックオフ資料（0828版・082802版）","2版をレビュー済","2026/08/25","docs/…/03,06","受領済"),
 ("調査票","ニーズ調査・在宅介護実態調査 調査票案（資料B・C・C'）","修正案まで受領","2026/08/25","docs/…/03,04","受領済"),
 ("国様式","令和7年8月版 標準調査票（ニーズ調査・在宅介護実態調査）","逐条突合の基礎。要確認0件で完了","2026/08/27","docs/…/05_令和7年8月版との逐条突合.md","受領済"),
 ("見える化","A系 人口・世帯／B系 認定者","第1号被保険者数・認定者数・認定率ほか","2026/08/25","data/mieruka_tidy.csv","受領済"),
 ("見える化","C1 保険料／D系 給付費・受給者・利用回数","第1号被保険者1人あたり給付月額ほか","2026/08/31","data/mieruka_tidy.csv","受領済"),
 ("見える化","K系 施設数／D13系 給付費","村内施設ゼロを確認","2026/08/31","data/mieruka_tidy.csv","受領済"),
 ("見える化","D25〜D30 定員／F15〜F25 地域包括支援センター","定員は平成30年度以降据置","2026/08/31","docs/…/02 追記3","受領済"),
 ("見える化","F1〜F14・F28〜F40 日常生活支援総合事業","17系列が実データなし（＝未実施）","2026/08/31","docs/…/02 追記3","受領済"),
 ("見える化","D47・D48 介護保険特別会計 歳入・歳出","平成26〜令和5年度決算","2026/08/31","docs/…/02 追記3","受領済"),
 ("見える化","W系 保険者機能強化推進交付金 評価指標（77ファイル）","令和3〜5年度。0点項目21件を抽出","2026/08/31","docs/…/07","受領済"),
 ("見える化","W138〜W155 アウトプット・アウトカム指標（18ファイル）","令和5年度","2026/08/31","docs/…/07","受領済"),
 ("見える化","地域分析 P1〜P4","県平均・全国平均・県内順位・全国順位","2026/08/31","docs/…/09","受領済"),
 ("村データ","★介護給付費準備基金の現在高","令和6年度末実績・令和7年度末見込","―","―","未受領"),
 ("村データ","★令和6年度決算・令和7年度給付費見込","第9期の乖離確定に必要","―","―","未受領"),
 ("村データ","★通いの場の実数（令和3〜7年度）","見える化は令和2年度で更新停止","―","―","未受領"),
 ("村データ","★高齢者福祉事業（村単独事業）の実績","現行計画の評価に必要","―","―","未受領"),
 ("村データ","★成年後見制度の利用実態・市民後見人養成実績","見える化に該当データなし","―","―","未受領"),
 ("国資料","令和8年度 交付金評価指標（市町村分）配点表","全17頁。推進400点・支援400点、目標Ⅰ〜Ⅳ各100点","2026/09/02","docs/…/20_交付金評価結果分析.md","受領済"),
 ("国資料","令和6年度 該当状況調査票集計表（全国一覧表）","全国1,741市町村。村は第398行。合計295点","2026/09/02","docs/…/20_交付金評価結果分析.md","受領済"),
 ("国資料","令和7年度 該当状況調査票集計表（全国集計・市町村）","合計339点。全国平均435.0点","2026/09/02","docs/…/20_交付金評価結果分析.md","受領済"),
 ("国資料","令和8年度 該当状況調査票集計表（全国集計・市町村）","合計447点。全国平均455.1点／全国1,005位／県内34位","2026/09/02","docs/…/20_交付金評価結果分析.md","受領済"),
 ("関連計画","第四次北塩原村生涯学習推進計画","全112頁。2018〜2027年度。7z分割書庫4分割で受領","2026/09/02","docs/…/21_関連計画_第四次生涯学習推進計画.md","受領済"),
 ("村データ","★要介護認定データ（調査票との関連付け用）","仕様書4Ⅰ(5)。調査回収後に必要","―","―","未受領"),
]
dr = DHR + 1
for kubun, name, cont, day, dest, sts in DOCS:
    for i, v in enumerate([kubun, name, cont, D(day) if day != "―" else "―", dest, sts], 1):
        c = wd.cell(row=dr, column=i, value=v)
        c.font = Font(name=F, size=9)
        c.border = BD
        c.alignment = Alignment(vertical="top", wrap_text=(i in (2,3,5)),
                                horizontal="center" if i in (1,4,6) else "left")
        if i == 4 and day != "―": c.number_format = "yyyy/mm/dd"
    fill = C["note"] if sts == "未受領" else C["done"]
    for i in range(1, 7):
        wd.cell(row=dr, column=i).fill = PatternFill("solid", fgColor=fill)
    wd.row_dimensions[dr].height = 28
    dr += 1
DLAST = dr - 1
wd.freeze_panes = "C5"
wd.auto_filter.ref = f"A{DHR}:F{DLAST}"
wd.sheet_view.zoomScale = 90

# ══════════════════ Sheet4: 凡例・集計 ══════════════════
ws2 = wb.create_sheet("凡例・集計")
ws2.column_dimensions["A"].width = 3
for col, w in zip("BCDEFGH", [24, 13, 11, 11, 11, 11, 40]):
    ws2.column_dimensions[col].width = w

title_bar(ws2, "B2:H2", "凡例・入力ルール／進捗集計", size=13, h=26)

def blk(row, title):
    ws2.merge_cells(f"B{row}:H{row}")
    c = ws2[f"B{row}"]; c.value = title
    c.font = Font(name=F, size=10, bold=True, color=C["white"])
    c.fill = PatternFill("solid", fgColor=C["sub"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

blk(4, "■ 入力する列（ここだけ編集してください）")
rules = [
 ("予定開始・予定完了", "日付", "yyyy/mm/dd 形式で入力します。村との協議で工程が固まり次第、入力してください。"),
 ("実績開始・実績完了", "日付", "着手日・完了日を実績として入力します。"),
 ("進捗率", "％", "0〜100% を入力します。完了時は 100%。"),
 ("ステータス", "選択", "未着手／着手／確認中／完了／保留／対象外 から選択します。"),
 ("備考", "自由", "課題・懸案・村への確認事項などを記入します。"),
]
rr = 5
for a, b, cc in rules:
    ws2.cell(row=rr, column=2, value=a).font = Font(name=F, size=9, bold=True)
    ws2.cell(row=rr, column=3, value=b).font = Font(name=F, size=9)
    ws2.merge_cells(f"D{rr}:H{rr}")
    ws2.cell(row=rr, column=4, value=cc).font = Font(name=F, size=9)
    for i in range(2, 9): ws2.cell(row=rr, column=i).border = BD
    rr += 1

blk(rr + 1, "■ 行の塗り分け")
legend = [
 (C["done"], "薄い緑", "完了した作業です。"),
 (C["doing"], "薄い黄", "着手中・確認中の作業、または北塩原村が担当する作業です。"),
 (C["note"], "薄い赤", "「村への確認事項」シートの優先度S・A、および未受領の資料です。"),
]
lr = rr + 2
for fill, nm, desc in legend:
    c = ws2.cell(row=lr, column=2, value=nm)
    c.font = Font(name=F, size=9, bold=True); c.border = BD
    c.fill = PatternFill("solid", fgColor=fill)
    ws2.merge_cells(f"C{lr}:H{lr}")
    ws2.cell(row=lr, column=3, value=desc).font = Font(name=F, size=9)
    for i in range(2, 9): ws2.cell(row=lr, column=i).border = BD
    lr += 1

# 進捗集計（COUNTIF）
SUM_R = lr + 1
blk(SUM_R, "■ 進捗集計（WBSシートから自動集計されます）")
sh = "'WBS・進捗管理'"
sts_list = ["未着手", "着手", "確認中", "完了", "保留", "対象外"]
hdr = ["大分類", "件数"] + sts_list
for i, hh in enumerate(hdr):
    c = ws2.cell(row=SUM_R + 1, column=2 + i, value=hh)
    c.font = Font(name=F, size=9, bold=True, color=C["white"])
    c.fill = PatternFill("solid", fgColor=C["sub"]); c.border = BD
    c.alignment = Alignment(horizontal="center")
majors = []
for m, *_ in W:
    if m not in majors: majors.append(m)
rw = SUM_R + 2
for m in majors:
    ws2.cell(row=rw, column=2, value=m).font = Font(name=F, size=9)
    ws2.cell(row=rw, column=3, value=f'=COUNTIF({sh}!$B${HR+1}:$B${LAST},$B{rw})')
    for j, st in enumerate(sts_list):
        ws2.cell(row=rw, column=4 + j,
                 value=f'=COUNTIFS({sh}!$B${HR+1}:$B${LAST},$B{rw},'
                       f'{sh}!$O${HR+1}:$O${LAST},{get_column_letter(4+j)}${SUM_R+1})')
    for i in range(2, 10):
        c = ws2.cell(row=rw, column=i); c.border = BD
        c.font = Font(name=F, size=9)
        if i >= 3: c.alignment = Alignment(horizontal="center")
    rw += 1
ws2.cell(row=rw, column=2, value="合計").font = Font(name=F, size=9, bold=True)
for i in range(3, 10):
    col = get_column_letter(i)
    c = ws2.cell(row=rw, column=i, value=f"=SUM({col}{SUM_R+2}:{col}{rw-1})")
    c.font = Font(name=F, size=9, bold=True); c.border = BD
    c.fill = PatternFill("solid", fgColor=C["band"])
    c.alignment = Alignment(horizontal="center")
ws2.cell(row=rw, column=2).border = BD
ws2.cell(row=rw, column=2).fill = PatternFill("solid", fgColor=C["band"])
TOT_R = rw

# 全体進捗率・確認事項集計
kpi = TOT_R + 2
blk(kpi, "■ 主要指標")
kpis = [
 ("全体進捗率（対象外を除く）", f'=AVERAGEIF({sh}!$O${HR+1}:$O${LAST},"<>対象外",{sh}!$N${HR+1}:$N${LAST})', "0.0%"),
 ("作業総数", f'=COUNTA({sh}!$A${HR+1}:$A${LAST})', "0"),
 ("完了した作業", f'=COUNTIF({sh}!$O${HR+1}:$O${LAST},"完了")', "0"),
 ("着手・確認中の作業", f'=COUNTIF({sh}!$O${HR+1}:$O${LAST},"着手")+COUNTIF({sh}!$O${HR+1}:$O${LAST},"確認中")', "0"),
 ("未着手の作業", f'=COUNTIF({sh}!$O${HR+1}:$O${LAST},"未着手")', "0"),
 ("対象外の作業", f'=COUNTIF({sh}!$O${HR+1}:$O${LAST},"対象外")', "0"),
 ("村への確認事項（優先度S）", '=COUNTIF(\'村への確認事項\'!$A:$A,"S")', "0"),
 ("村への確認事項（優先度A）", '=COUNTIF(\'村への確認事項\'!$A:$A,"A")', "0"),
 ("村への確認事項（未依頼）", '=COUNTIF(\'村への確認事項\'!$G:$G,"未依頼")', "0"),
 ("未受領の資料・データ", '=COUNTIF(\'受領資料・データ管理\'!$F:$F,"未受領")', "0"),
]
kr2 = kpi + 1
for nm, fml, fmt in kpis:
    ws2.cell(row=kr2, column=2, value=nm).font = Font(name=F, size=10, bold=True)
    c = ws2.cell(row=kr2, column=3, value=fml)
    c.font = Font(name=F, size=12, bold=True, color=C["header"])
    if fmt: c.number_format = fmt
    c.border = BD
    c.fill = PatternFill("solid", fgColor=C["key"])
    c.alignment = Alignment(horizontal="center")
    ws2.cell(row=kr2, column=2).border = BD
    kr2 += 1

# 直近マイルストーン
ms = kr2 + 1
blk(ms, "■ マイルストーン（令和8年9月1日 キックオフ議事録により確定）")
MILESTONES = [
 ("令和8年9月1日（火）", "キックオフ会議", "完了。議事録を受領（doc22）。S項目16件のうち9件が決着", ""),
 ("令和8年9月3日（木）", "調査票 第1稿の提出", "★議事録で確定。未決事項は選択肢併記で提出", "★"),
 ("令和8年9月9日（水）頃", "調査票の最終校了・印刷依頼", "★Webフォームの実装方法が未決だと間に合わない恐れ", "★"),
 ("令和8年9月14日（月）", "宛名ラベル・配布用封筒の提供", "村。9/11への前倒しを要請", ""),
 ("令和8年9月16〜17日", "宛名貼付・封入・封緘", "受託者", ""),
 ("令和8年9月18日（金）頃", "アンケート発送", "★5連休（9/19〜9/23）の直前。9/17前倒しを協議", "★"),
 ("令和8年9月30日（水）", "策定委員会 現委員の任期満了", "★議事録で触れられず。第2回（11月）には改選が前提", "★"),
 ("令和8年10月", "回収・データ入力", "村が100件程度集まるごとに随時送付", ""),
 ("令和8年11月", "第2回策定委員会", "★分析結果・効果検証・骨子・体系・見込量・保険料概算・素案（完成度7割）", "★"),
 ("令和8年11月〜12月", "パブリックコメント", "第2回委員会の後に実施", ""),
 ("令和9年1月下旬〜2月上旬", "第3回策定委員会", "パブコメ結果と反映内容を報告。ほぼ完成版の素案で最終合意", ""),
 ("令和9年2月末頃", "議会全員協議会", "概要版（A3両面程度）を用いて説明", ""),
 ("令和9年3月31日", "履行期限（納品・検収）", "", ""),
]
mh = ms + 1
for j, h in enumerate(["", "期日", "内容", "留意点"], 1):
    c = ws2.cell(row=mh, column=j+1, value=h)
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C["header"]); c.border = BD
    c.alignment = Alignment(horizontal="center")
mr = mh + 1
for day, cont, note, mark in MILESTONES:
    vals = [mark, day, cont, note]
    for j, v in enumerate(vals, 2):
        c = ws2.cell(row=mr, column=j, value=v)
        c.font = Font(name=F, size=9, bold=bool(mark))
        c.border = BD
        c.alignment = Alignment(vertical="top", wrap_text=(j == 5),
                                horizontal="center" if j == 2 else "left")
        if mark:
            c.fill = PatternFill("solid", fgColor=C["note"])
    mr += 1

notes = [
 "※「仕様書該当」欄の凡例：仕様書＝8北保福第483号 別紙／手引き＝厚生労働省の実施の手引き（令和7年8月版）／確認＝本業務での確認事項",
 "※ 想定時期は仕様書の履行期限（令和9年3月31日）から逆算した目安です。村との協議結果により確定します。",
 "※ 令和8年9月19日（土）〜23日（水）は5連休（敬老の日9/21・国民の休日9/22・秋分の日9/23）です。発送・回収工程に影響します。",
 "※ 進捗基準日は令和8年9月2日（キックオフ議事録の反映後）です。備考欄の doc番号は docs/北塩原村_第10期/ 配下の分析資料に対応します。",
 "※ マイルストーンの★印は、遅れると後続工程が連鎖的に破綻する項目です。",
 "※ 第1回策定委員会は受託前に開催済みのため、関連する2作業はステータス「対象外」としています。全体進捗率は対象外を除いて算出しています。",
]
nr = mr + 1
for n in notes:
    ws2.cell(row=nr, column=2, value=n).font = Font(name=F, size=8)
    nr += 1

wb.save(OUT)
print(f"保存: {OUT}")
print(f"WBS: {LAST-HR}件 / 確認事項: {len(K)}件（解決済{len(SOLVED)}件） / 受領資料: {DLAST-DHR}件")
print(f"進捗反映: {len(P)}件")
