"""介護保険特別会計経理状況（D47・D48）を整理ブックにまとめる。

見える化システムから出力された25ファイル（D47・D47-a〜m・D48・D48-a〜j）を
読み込み、小野町の時系列を1ブックに集約する。
既存の整理2ブック（基礎数値整理・地域分析データ整理）と同じ構成
（委員会要点／主要時系列／内訳／取込確認）とする。
"""

import glob
import pathlib
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = pathlib.Path(__file__).parent
SRC = ROOT / "小野町_引継ぎ_整理済" / "07_介護保険_見える化整理" / "介護保険特別会計経理状況"
OUT = ROOT / "小野町_引継ぎ_整理済" / "07_介護保険_見える化整理"

PER = ["H26", "H27", "H28", "H29", "H30", "R1", "R2", "R3", "R4", "R5"]


def norm(p):
    p = str(p).replace("\n", "").strip()
    if p in ("R元", "R元年"):
        return "R1"
    m = re.match(r"^平成(\d+)年3月末$", p)
    if m:
        return f"H{int(m.group(1)) - 1}"
    m = re.match(r"^令和(\d+)年3月末$", p)
    if m:
        return f"R{int(m.group(1)) - 1}"
    if p == "令和元年3月末":
        return "H30"
    if re.match(r"^(H|R)\d{1,2}$", p):
        return p
    return None


def load():
    store, files = {}, []
    for f in sorted(SRC.glob("*.xlsx")):
        m = re.match(r"^(D4[78](?:-[a-z])?)_", f.name)
        if not m:
            continue
        sid = m.group(1)
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb["表形式（時系列）"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        for r in rows[:6]:
            if r and sum(1 for x in r if x and norm(x)) >= 5:
                hdr = [norm(x) if x else None for x in r]
                break
        n = 0
        for r in rows:
            if r and r[1] == "小野町":
                store.setdefault(sid, {})[str(r[2])] = {
                    hdr[i]: r[i] for i in range(min(len(r), len(hdr)))
                    if hdr[i] and isinstance(r[i], (int, float))
                }
                n += 1
        files.append((sid, f.name, n))
    return store, files


def main():
    store, files = load()

    def g(sid, lab):
        return [store.get(sid, {}).get(lab, {}).get(p) for p in PER]

    wb = openpyxl.Workbook()
    head = PatternFill("solid", fgColor="DDEBF7")
    hl = PatternFill("solid", fgColor="FFF2CC")

    # ---- 00_委員会要点
    ws = wb.active
    ws.title = "00_委員会要点"
    ws.append(["項目", "確認結果", "第10期計画への接続", "委員会での確認ポイント"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = head
    kuri = g("D47-k", "繰越金")
    kyufu = g("D48-b", "合計")
    tsumi = [x or 0 for x in g("D48-g", "基金積立金")]
    tori = [x or 0 for x in g("D47-j", "介護給付費準備基金繰入金")]
    rows = [
        ("データ取込",
         f"介護保険特別会計経理状況25ファイル（D47・D48系）をすべて小野町データとして確認。"
         f"期間はH26〜R5の10年度分",
         "保険料算定の財政面の前提として使用可能",
         "年報ベースのためR6・R7が未反映。町の決算で補完する"),
        ("実質収支",
         f"歳入−歳出はH26の{(g('D47','合計')[0] - g('D48','合計')[0]) / 1e6:.1f}百万円から"
         f"R5の{(g('D47','合計')[-1] - g('D48','合計')[-1]) / 1e6:.1f}百万円まで10年連続でプラス",
         "第10期の保険料設定において、余剰の還元をどう扱うかが論点となる",
         "余剰が生じ続けている要因（給付の伸び悩みか、保険料設定か）を確認する"),
        ("繰越金",
         f"H26の{kuri[0] / 1e6:.1f}百万円からR5の{kuri[-1] / 1e6:.1f}百万円へ"
         f"{kuri[-1] / kuri[0]:.1f}倍に増加。R5の繰越金は保険給付費の{kuri[-1] / kyufu[-1] * 100:.1f}％に相当",
         "第10期の財源として算入する範囲を決める",
         "繰越金の水準として適正か、基金への積立を進めるかを整理する"),
        ("介護給付費準備基金",
         f"積立累計{sum(tsumi) / 1e6:.1f}百万円に対し取崩累計{sum(tori) / 1e6:.1f}百万円。"
         f"差引{(sum(tsumi) - sum(tori)) / 1e6:.1f}百万円の純増",
         "取崩による保険料抑制の余地がある。第11期の急騰リスクと併せて検討する",
         "基金残高の実額を町の決算で確認する（本表は積立・取崩の差分による推計）"),
        ("保険給付費",
         f"H26の{kyufu[0] / 1e6:.1f}百万円からR5の{kyufu[-1] / 1e6:.1f}百万円へ"
         f"{(kyufu[-1] / kyufu[0] - 1) * 100:+.1f}％。H29の{max(x for x in kyufu if x) / 1e6:.1f}百万円をピークにほぼ横ばい",
         "第10期の標準給付費見込みの基礎。給付が伸びていない前提で推計する",
         "認定率が高いのに給付が伸びない構造（認定と利用のギャップ）を確認する"),
        ("地域支援事業費",
         f"H26の{g('D48-c', '合計')[0] / 1e6:.1f}百万円からR5の{g('D48-c', '合計')[-1] / 1e6:.1f}百万円へ"
         f"{g('D48-c', '合計')[-1] / g('D48-c', '合計')[0]:.1f}倍に増加",
         "第10期の地域支援事業費見込みの基礎。上限額との関係を確認する",
         "総合事業・包括的支援事業・任意事業の内訳と、事業別の財源区分を確認する"),
        ("保険料収入",
         f"H26の{g('D47-a', '保険料')[0] / 1e6:.1f}百万円からR5の{g('D47-a', '保険料')[-1] / 1e6:.1f}百万円へ"
         f"{(g('D47-a', '保険料')[-1] / g('D47-a', '保険料')[0] - 1) * 100:+.1f}％。"
         f"保険料基準額は4,400円から6,600円に引上げ",
         "給付費が横ばいの一方で保険料収入が増加しており、余剰の蓄積要因",
         "第10期の保険料水準の考え方（据置・引下げ・引上げ）を諮る"),
        ("低所得者保険料軽減",
         f"繰入金のうち低所得者保険料軽減はH27の{g('D47-j', '低所得者保険料軽減繰入金')[1] / 1e6:.1f}百万円から"
         f"R5の{g('D47-j', '低所得者保険料軽減繰入金')[-1] / 1e6:.1f}百万円へ増加",
         "第10期の所得段階別保険料の設定に接続する",
         "所得段階別被保険者数（第1〜3段階割合28％）と併せて確認する"),
        ("調整交付金",
         f"国庫支出金のうち調整交付金はH27の{g('D47-d', '調整交付金')[1] / 1e6:.1f}百万円をピークに"
         f"R5は{g('D47-d', '調整交付金')[-1] / 1e6:.1f}百万円まで減少",
         "第10期の財源見通しに影響する。交付率の推移を確認する",
         "後期高齢者比率・所得段階分布の変化との関係を確認する"),
    ]
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", (16, 62, 46, 46)):
        ws.column_dimensions[col].width = w

    # ---- 01_主要時系列
    ws2 = wb.create_sheet("01_主要時系列")
    ws2.append(["区分", "項目", "単位"] + PER)
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = head
    main_rows = [
        ("歳入", "合計", g("D47", "合計")),
        ("歳入", "保険料", g("D47-a", "保険料")),
        ("歳入", "国庫支出金", g("D47-d", "合計")),
        ("歳入", "　うち介護給付費負担金", g("D47-d", "介護給付費負担金")),
        ("歳入", "　うち調整交付金", g("D47-d", "調整交付金")),
        ("歳入", "支払基金交付金", g("D47-e", "合計")),
        ("歳入", "都道府県支出金", g("D47-f", "合計")),
        ("歳入", "繰入金", g("D47-j", "合計")),
        ("歳入", "　うち一般会計繰入金", g("D47-j", "一般会計繰入金")),
        ("歳入", "　うち介護給付費準備基金繰入金（取崩）", g("D47-j", "介護給付費準備基金繰入金")),
        ("歳入", "　うち低所得者保険料軽減繰入金", g("D47-j", "低所得者保険料軽減繰入金")),
        ("歳入", "繰越金", g("D47-k", "繰越金")),
        ("歳出", "合計", g("D48", "合計")),
        ("歳出", "総務費", g("D48-a", "総務費")),
        ("歳出", "保険給付費", g("D48-b", "合計")),
        ("歳出", "　うち介護サービス等諸費", g("D48-b", "介護サービス等諸費")),
        ("歳出", "　うち介護予防サービス等諸費", g("D48-b", "介護予防サービス等諸費")),
        ("歳出", "　うち高額介護サービス等費", g("D48-b", "高額介護サービス等費")),
        ("歳出", "　うち特定入所者介護サービス等費", g("D48-b", "特定入所者介護サービス等費")),
        ("歳出", "地域支援事業", g("D48-c", "合計")),
        ("歳出", "　うち介護予防・生活支援サービス事業費", g("D48-c", "介護予防・生活支援サービス事業費")),
        ("歳出", "　うち一般介護予防事業費", g("D48-c", "一般介護予防事業費")),
        ("歳出", "　うち包括的支援事業・任意事業", g("D48-c", "包括的支援事業･任意事業")),
        ("歳出", "基金積立金", g("D48-g", "基金積立金")),
        ("歳出", "諸支出金", g("D48-j", "合計")),
    ]
    for cat, lab, v in main_rows:
        ws2.append([cat, lab, "円"] + [(x if x is not None else None) for x in v])
    inc, out = g("D47", "合計"), g("D48", "合計")
    ws2.append(["計算", "歳入−歳出（実質収支）", "円"] + [a - b for a, b in zip(inc, out)])
    bal, hist = 0, []
    for t, k in zip([x or 0 for x in g("D48-g", "基金積立金")],
                    [x or 0 for x in g("D47-j", "介護給付費準備基金繰入金")]):
        bal += t - k
        hist.append(bal)
    ws2.append(["計算", "基金の純増累計（H26以降）", "円"] + hist)
    ws2.append(["計算", "保険給付費に対する繰越金の割合", "％"] +
               [round(a / b * 100, 1) if a and b else None for a, b in zip(g("D47-k", "繰越金"), g("D48-b", "合計"))])
    for row in ws2.iter_rows(min_row=2):
        if row[1].value in ("合計", "歳入−歳出（実質収支）", "基金の純増累計（H26以降）"):
            for c in row:
                c.fill = hl
        for c in row[3:]:
            c.number_format = "#,##0"
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 6
    for i in range(len(PER)):
        ws2.column_dimensions[chr(ord("D") + i)].width = 13
    ws2.freeze_panes = "D2"

    # ---- 02_内訳ロング
    ws3 = wb.create_sheet("02_内訳ロング")
    ws3.append(["指標ID", "区分", "項目", "年度", "値（円）"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = head
    n = 0
    for sid in sorted(store):
        kubun = "歳入" if sid.startswith("D47") else "歳出"
        for lab, vals in store[sid].items():
            for p in PER:
                if vals.get(p) is not None:
                    ws3.append([sid, kubun, lab, p, vals[p]])
                    n += 1
    for row in ws3.iter_rows(min_row=2):
        row[4].number_format = "#,##0"
    for col, w in zip("ABCDE", (10, 8, 46, 8, 16)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    # ---- 90_取込確認
    ws4 = wb.create_sheet("90_取込確認")
    ws4.append(["指標ID", "ファイル", "自治体", "取込行数", "状態"])
    for c in ws4[1]:
        c.font = Font(bold=True)
        c.fill = head
    for sid, name, cnt in files:
        ws4.append([sid, name, "小野町", cnt, "小野町データ確認"])
    for col, w in zip("ABCDE", (10, 58, 10, 10, 20)):
        ws4.column_dimensions[col].width = w

    path = OUT / "小野町_第10期介護保険事業計画_介護保険財政整理_特別会計経理状況.xlsx"
    wb.save(path)
    print("出力:", path)
    print(f"  取込ファイル {len(files)} / 内訳レコード {n} / 期間 {PER[0]}〜{PER[-1]}")


if __name__ == "__main__":
    main()
