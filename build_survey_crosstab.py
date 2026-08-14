# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　調査結果のクロス集計・分析.

業務仕様書「４．業務内容（3）実施済み調査結果の集計・分析」に対応する。
対象4調査のうち、健康とくらしの調査（JAGES調査・令和7年度）の個票データにより
単純集計及びクロス集計を行い、地域包括ケア「見える化」システム、介護給付実績及び
人口推計と組み合わせて分析する。

個票データは個人情報を含むため本リポジトリには格納しない。
SRC で指定するディレクトリに配置して実行する。

シート構成
  00_目次と分析の前提
  01_地区別の主要指標
  02_年齢調整による地域比較
  03_外出手段と移動制約
  04_外出の抑制と生活動作の困りごと
  05_社会参加とリスクの関連
  06_世帯構成とリスク
  07_経済状況とリスク
  08_在宅生活の課題と解決手段
  09_住み続ける意向と転居理由
  10_介護の担い手と介護の意向
  11_KPIのデータ源の再検討
  12_主要所見
"""

import os
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = ("/tmp/claude-0/-home-user-repository/54f527c9-842b-534d-b822-e5a6c91f837c/scratchpad/"
       "z4/3-19大雪地区広域連合_2025健康とくらしの調査_個票データ・集計表_030")
CSV = os.path.join(SRC, "KK_2025_CSV_3-19大雪地区広域連合.csv")
OUT = "/home/user/repository/output/第10期計画_調査クロス集計・分析.xlsx"

FONT = "游ゴシック"
NAVY, HEAD = "1F4E78", "5B9BD5"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============================================================ データの読込み
df = pd.read_csv(CSV, encoding="cp932", low_memory=False)
A = df[df["uncounted25"].isna()].copy()          # 集計対象 4,729票

TOWN = {1453: "東神楽町", 1458: "東川町", 1459: "美瑛町"}
AREA = {1453001: ("東神楽町", "東聖小学校区"), 1453002: ("東神楽町", "東神楽小学校区"),
        1453004: ("東神楽町", "志比内小学校区"),
        1458001: ("東川町", "東川第1小学校区"), 1458002: ("東川町", "東川第2小学校区"),
        1458003: ("東川町", "東川第3小学校区"), 1458004: ("東川町", "東川小学校区"),
        1459001: ("美瑛町", "美瑛小学校区"), 1459003: ("美瑛町", "美馬牛小学校区"),
        1459004: ("美瑛町", "美沢小学校区"), 1459005: ("美瑛町", "明徳小学校区"),
        1459007: ("美瑛町", "美瑛東小学校区")}
AGES = ["65～69歳", "70～74歳", "75～79歳", "80～84歳", "85歳以上"]

A["町"] = A["kcode25"].map(TOWN)
A["地区"] = A["scode1_25"].map(lambda v: AREA.get(v, ("―", "―"))[1])
A["年齢階級"] = pd.cut(pd.to_numeric(A["age_m25"], errors="coerce"),
                       bins=[65, 70, 75, 80, 85, 200], labels=AGES, right=False)


def num(col):
    return pd.to_numeric(A[col], errors="coerce")


def flag_in(col, vals):
    """単一選択の設問について、指定した選択肢に該当するかの0/1系列を返す。"""
    s = num(col)
    return s.isin(vals).astype(float).where(s.notna())


def any_of(cols, vals=(1,)):
    """複数回答の設問群について、いずれかに該当するかの0/1系列を返す。
    分母は当該設問群に1つでも回答のあった者とする。"""
    sub = A[list(cols)].apply(pd.to_numeric, errors="coerce")
    ok = sub.notna().any(axis=1)
    return sub.isin(vals).any(axis=1).astype(float).where(ok)


# 派生変数
SOC6 = ["cmnt6vl25", "cmnt6sp25", "cmnt6hb25", "cmnt6le25", "cmnt6sl25", "cmnt6sk25"]
SOC6 = [c for c in SOC6 if c in A.columns]
A["社会参加"] = any_of(SOC6, vals=(1, 2, 3, 4))          # 月1回以上
A["通いの場"] = flag_in("cmnt6sl25", [1, 2, 3, 4])
A["運転あり"] = flag_in("gout2dv25", [1])
A["同乗のみ"] = ((flag_in("gout2dv25", [1]) == 0) & (flag_in("gout2rd25", [1]) == 1)).astype(float)
GOUT = [c for c in A.columns if c.startswith("gout2") and c != "gout2ref25"]
A["自力移動なし"] = ((num("gout2dv25") != 1) & (num("gout2bi25") != 1)
                     & (num("gout2mo25") != 1)).astype(float).where(A[GOUT].notna().any(axis=1))
A["外出低頻度"] = flag_in("gout7fq25", [5, 6, 7])          # 月1～3回以下
A["外出減少"] = flag_in("gout4la25", [1, 2])
A["独居"] = pd.to_numeric(A["alone2gp_25"], errors="coerce")
A["夫婦のみ"] = ((num("hous2sp25") == 1) & (pd.to_numeric(A["mebr2nb25"], errors="coerce") == 2)
                 ).astype(float).where(num("hous2no25").notna())
A["暮らし苦しい"] = flag_in("sfs5_25", [1, 2])
A["介助不要"] = flag_in("adl3ra25", [1])
A["要介助・未受給"] = flag_in("adl3ra25", [2])
A["要介助・受給中"] = flag_in("adl3ra25", [3])
A["住み続けたい"] = flag_in("taisetsu_q3_1", [1])
A["転居も考える"] = flag_in("taisetsu_q3_1", [2])
Q21 = [c for c in A.columns if c.startswith("taisetsu_q2_1_s")]
Q21_NEED = [c for c in Q21 if not c.endswith("_s21")]          # 「特になし」を除く
A["困りごとあり"] = any_of(Q21_NEED)
A["解決できず困っている"] = flag_in("taisetsu_q2_2_s10", [1])
A["世話をする人なし"] = flag_in("taisetsu_q4_1_s9", [1])

IND = [
    ("frail2gp_25", "フレイルあり割合"),
    ("undo2gp_25", "運動機能低下者割合"),
    ("tojikomori2gp_25", "閉じこもり者割合"),
    ("utsucheck2gp_25", "うつ割合"),
    ("ninchicheck2gp_25", "認知機能低下者割合"),
    ("koukukinou2gp_25", "口腔機能低下者割合"),
    ("iadl1_2gp_25", "IADL低下者割合"),
    ("tento2gp_25", "1年間の転倒あり割合"),
    ("alone2gp_25", "独居者割合"),
    ("社会参加", "社会参加あり割合（月1回以上）"),
    ("通いの場", "通いの場参加割合（月1回以上）"),
]


def series(name):
    return pd.to_numeric(A[name], errors="coerce")


def rate(mask, name):
    """割合（％）と分母を返す。"""
    s = series(name)[mask]
    n = int(s.notna().sum())
    if n == 0:
        return None, 0
    return round(100 * s.sum() / n, 1), n


def wilson(p, n, z=1.96):
    """Wilson法による95％信頼区間（p は0〜1）。"""
    if n == 0:
        return None, None
    c = p + z * z / (2 * n)
    d = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return (c - d) / (1 + z * z / n), (c + d) / (1 + z * z / n)


def ci_text(mask, name):
    r, n = rate(mask, name)
    if r is None:
        return "―"
    lo, hi = wilson(r / 100, n)
    return "%.1f〜%.1f" % (100 * lo, 100 * hi)


def ztest(m1, m2, name):
    """2標本の母比率の差の検定（両側・正規近似）。判定文字列を返す。"""
    s1, s2 = series(name)[m1].dropna(), series(name)[m2].dropna()
    n1, n2 = len(s1), len(s2)
    if n1 < 10 or n2 < 10:
        return "判定不能"
    p1, p2 = s1.mean(), s2.mean()
    p = (s1.sum() + s2.sum()) / (n1 + n2)
    se = math.sqrt(p * (1 - p) * (1 / n1 + 1 / n2))
    if se == 0:
        return "判定不能"
    z = abs(p1 - p2) / se
    return "有意" if z >= 1.96 else "有意でない"


# 年齢調整（直接法・基準人口＝広域連合全体の年齢階級別回答者数）
STD = A["年齢階級"].value_counts().reindex(AGES).fillna(0)
STD_TOTAL = STD.sum()


def adjusted(mask, name):
    """直接法による年齢調整済み割合（％）。"""
    s = series(name)
    tot, wt = 0.0, 0.0
    for g in AGES:
        sub = s[mask & (A["年齢階級"] == g)].dropna()
        if len(sub) < 5:
            continue
        tot += sub.mean() * STD[g]
        wt += STD[g]
    if wt < STD_TOTAL * 0.8:
        return None
    return round(100 * tot / wt, 1)


# ============================================================ 出力
wb = Workbook()


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    n = max(len(widths), 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 50
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=32):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=18, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.border = BORDER
        ha = (align or {}).get(i, "left" if isinstance(v, str) else "right")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal=ha)
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def note(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = max(28, 12 * (len(text) // (span * 14) + 1))
    return row + 2


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 20
    return row + 1


ALL = pd.Series(True, index=A.index)
N_ALL = len(A)

# ============================================================ 00 目次
ws = sheet("00_目次と分析の前提", "調査結果のクロス集計・分析",
           "業務仕様書４（3）実施済み調査結果の集計・分析に対応する成果物。"
           "健康とくらしの調査（JAGES調査）令和7年度の個票データ4,729票により、"
           "単純集計、クロス集計及び見える化システム・給付実績・人口推計と組み合わせた分析を行う。"
           "作成日：令和8年7月29日。",
           [4, 26, 58, 34, 16], freeze="A4")

r = header(ws, 4, ["No.", "シート", "内容", "計画への反映先", "分母"])
for i, (nm, cont, dst, den) in enumerate([
    ("01_地区別の主要指標", "小学校区12地区別に主要11指標を集計。美瑛町の美馬牛・美沢・"
     "明徳・美瑛東の各小学校区が独立して取得できる。", "第1章第7節／第2章第4節", "地区別"),
    ("02_年齢調整による地域比較", "町間・地区間の差が年齢構成の違いによるものか、"
     "地域固有の差かを直接法の年齢調整により判別する。", "第2章第4節", "町・地区別"),
    ("03_外出手段と移動制約", "外出時の交通手段14区分の利用状況を町・地区別に集計。"
     "自分で運転できない層の分布を示す。", "第1章第7節／第5章基本目標2", "4,680票"),
    ("04_外出の抑制と生活動作の困りごと", "外出を控えている者の割合とその理由、"
     "外出頻度の低い者の割合。", "第2章第4節", "設問により異なる"),
    ("05_社会参加とリスクの関連", "社会参加の有無別のリスク指標。年齢階級で層化し、"
     "年齢の交絡を除いて比較する。", "第5章基本目標1", "4,611票"),
    ("06_世帯構成とリスク", "独居・夫婦のみ・その他の別によるリスク指標の差。",
     "第2章第1節／第5章基本目標2", "4,652票"),
    ("07_経済状況とリスク", "暮らし向き5段階別のリスク指標。経済的理由による外出抑制。",
     "第2章第2節3", "4,596票"),
    ("08_在宅生活の課題と解決手段", "大雪独自設問。生活動作の困りごと20項目と"
     "解決手段10項目。「解決できず、困っている」を含む。", "第2章第5節／第5章基本目標2", "4,128票"),
    ("09_住み続ける意向と転居理由", "大雪独自設問。住み続ける意向と、転居を考える理由"
     "（通院・買い物の困難、施設に入所できない、生活支援サービスがない）。",
     "第2章第6節／第6章第2節", "4,446票"),
    ("10_介護の担い手と介護の意向", "大雪独自設問。世話をしてくれる家族の有無、"
     "介護を受けたい場所、人生会議・死後事務委任の認知。", "第5章基本目標2・5", "4,599票"),
    ("11_KPIのデータ源の再検討", "算定不可としていた代表KPIのうち、本調査の大雪独自設問"
     "により算定できるものを整理する。", "第4章第3節／資料1", "―"),
    ("12_主要所見", "本分析から得られた所見と、計画への反映方針。", "各章", "―"),
    ("13_課題と施策への対応レビュー", "第9期の評価から得られた課題20件及び第10期の施策に対し、"
     "本分析がどこまで根拠を提供できているかをレビューする。"
     "「不足」と判定した項目は15〜18シートで追加分析した。", "各章", "―"),
    ("14_統計的妥当性の検証", "調査対象者の範囲、町別の回収率と非回答の偏り、"
     "主要指標の95％信頼区間、地区別分析の標本規模、多重比較の扱い。", "第2章第4節", "―"),
    ("15_介護・介助が必要な層", "介護・介助が必要と回答した404人の世帯・支援者・"
     "移動制約・困りごと。24時間対応サービスの需要側の根拠。",
     "第2章第2節／第6章第4節", "4,645票"),
    ("16_認知症の状況と相談窓口", "本人又は家族の認知症症状の有無と相談窓口の認知度。"
     "見える化に認知症指標のデータ登録がないため本調査が唯一の数量的根拠。",
     "第5章 基本目標2", "4,634票"),
    ("17_施設志向と在宅資源", "施設入所志向のある層の在宅資源（支援者・移動・世帯）の状況。"
     "第6章の整備方針の需要側の根拠。", "第6章第4・5節", "4,634票"),
    ("18_地域重点度と社会参加の障壁", "JAGESの重点対象地域ランク（70指標・小学校区別）との突合と、"
     "通いの場に参加していない層の特性。", "第1章第7節／第5章 基本目標1", "地区別"),
    ("19_同規模保険者との比較", "人口5万未満の40保険者（集計数82,106票）との比較。"
     "20指標を全体及び年齢階級別に対比し、調査対象者の範囲の違いによる偏りを検証する。",
     "第2章第4節", "82,106票"),
    ("20_参加自治体における順位", "健康とくらしの調査に参加した64保険者74市町村における"
     "3町の順位（主要12指標）。", "第2章第4節", "74市町村"),
    ("21_他団体との比較の総括", "健康とくらしの調査、見える化システム、交付金評価及び"
     "北海道の資料による20項目の比較。仕様書４（4）の類似保険者比較に対応する。",
     "第2章第2・4節／第3章", "―"),
], start=1):
    r = body(ws, r, [i, nm, cont, dst, den], height=40)

r += 1
r = lead(ws, r, "【分析の前提】")
r = header(ws, r, ["項目", "内容", "", "", ""])
for k, vtxt in [
    ("調査名", "令和7年度 健康とくらしの調査（日常生活圏域ニーズ調査／JAGES調査）"),
    ("実施期間", "令和7年11月17日〜12月8日"),
    ("対象", "大雪地区広域連合に居住する65歳以上の方 7,121人"
     "（東神楽町2,319・東川町2,074・美瑛町2,728）"),
    ("対象者の属性", "一般高齢者＋総合事業対象者。要支援者・要介護者は含まれていない"
     "（調査報告書「４（５）各保険者の調査対象者」）。"
     "この制約は代表KPIのデータ源の選択に影響する（14シート）"),
    ("回収", "4,798票（回収率67.4％）。うち集計対象4,729票"),
    ("除外", "ID切り取り及び対象者名簿65歳未満の69票（uncounted25）を除外している"),
    ("分母の扱い", "指標ごとに有効回答者を分母とする。複数回答の設問は、当該設問群に"
     "1つ以上の回答があった者を分母とする。すべての表に分母を併記する"),
    ("調査票の版", "本調査は8種類の調査票（A〜H）を無作為に配布しており、"
     "各版に固有の設問（サブコア・セクションA〜H）の分母は550〜616票となる。"
     "共通設問及び大雪独自設問の分母は4,100〜4,700票である"),
    ("信頼区間", "母比率の95％信頼区間はWilson法による"),
    ("差の検定", "2群の比率の差は正規近似による両側検定（有意水準5％）。"
     "いずれかの群の分母が10未満の場合は判定不能とする"),
    ("年齢調整", "直接法。基準人口は広域連合全体の年齢階級別回答者数"
     "（65〜69歳968・70〜74歳1,204・75〜79歳1,246・80〜84歳784・85歳以上527）"),
    ("突合の確認", "本分析による広域連合全体及び町別の値のうち、"
     "主要20指標について調査報告書の公表値と一致することを確認している。"
     "全設問について照合したものではない"),
]:
    r = body(ws, r, [k, vtxt, "", "", ""], height=32)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=5)

note(ws, r + 1,
     "注1）個票データは個人情報を含むため、成果物には集計値のみを収録している。"
     "注2）他の3調査（在宅生活改善調査、居所変更実態調査、介護人材実態調査）については、"
     "第10期分の調査票・回答データ・集計資料が未受領のため、本表ではクロス集計を行っていない。"
     "第9期計画に掲載された令和5年調査分の再整理は、図表集11〜13シートに収録している。"
     "注3）本分析により、大雪独自設問を用いれば、これまで算定不可としていた代表KPIの一部が"
     "算定できることが判明した。詳細は11シートに示す。", 5)

# ============================================================ 01 地区別
ws = sheet("01_地区別の主要指標", "小学校区別の主要指標",
           "分析地域1（小学校区）別に主要11指標を集計する。"
           "美瑛町については美瑛・美馬牛・美沢・明徳・美瑛東の5小学校区が独立して取得できるため、"
           "日常生活圏域別の分析及び訪問・通所困難地域の判定の基礎資料となる。"
           "分母が100票未満の地区は網掛けとし、値の変動が大きいことに留意する。",
           [16, 18, 8, 9] + [9] * 11)

cols = ["町", "地区", "回答者数", "75歳以上\n構成比"] + [t for _, t in IND]
r = header(ws, 4, cols, height=44)

rows = [("大雪地区広域連合", "全体", ALL)]
for t in ["東川町", "美瑛町", "東神楽町"]:
    rows.append((t, "町計", A["町"] == t))
    for code, (tn, an) in sorted(AREA.items()):
        if tn == t:
            rows.append(("", an, A["地区"] == an))

for tname, aname, mask in rows:
    n = int(mask.sum())
    o75 = A.loc[mask, "年齢階級"].isin(["75～79歳", "80～84歳", "85歳以上"]).mean()
    vals = [tname, aname, n, round(100 * o75, 1)]
    for key, _ in IND:
        v, _n = rate(mask, key)
        vals.append(v)
    small = n < 100
    fills = {3: NG_O} if small else ({2: MID_B} if aname == "町計" else None)
    if aname == "全体":
        fills = {i: GRAY for i in range(1, 16)}
    r = body(ws, r, vals, fills, height=18, align={1: "left", 2: "left"},
             bold=(aname in ("全体", "町計")))

note(ws, r + 1,
     "注1）単位は％。回答者数のみ人。「75歳以上構成比」は回答者に占める75歳以上の割合で、"
     "地区間の年齢構成の差を示す。地区間の比較は年齢調整後の値（02シート）による。"
     "注2）分母が100票未満の地区（志比内小学校区21票、明徳小学校区49票、美沢小学校区57票）は"
     "回答者数を網掛けとした。これらの地区の値は95％信頼区間が広く、単独では判断材料としない。"
     "注3）小学校区と第9期計画の日常生活圏域（美瑛町の旭・北西地区、美馬牛地区、朗根内地区、"
     "市街地・周辺地区）との対応関係は確定していない。"
     "美馬牛小学校区は美馬牛地区に対応すると考えられるが、"
     "他の小学校区との対応は美瑛町への確認を要する（確認事項No.12）。"
     "注4）東川小学校区は1,040票と東川町の回答の73.5％を占める。"
     "東川第1〜第3小学校区は旧小学校区であり、統合後の区域との関係の確認を要する。", 15)

# ============================================================ 02 年齢調整
ws = sheet("02_年齢調整による地域比較", "年齢調整による町間・地区間の比較",
           "町別・地区別の指標の差が年齢構成の違いによるものか、地域固有の差かを判別する。"
           "直接法により、広域連合全体の年齢階級別回答者数を基準人口として調整した。"
           "粗率と調整率の差が大きい地域は、年齢構成の影響が大きいことを示す。",
           [16, 18, 8] + [11] * 10)

SEL = ["frail2gp_25", "tojikomori2gp_25", "iadl1_2gp_25", "utsucheck2gp_25", "社会参加"]
SELN = ["フレイル", "閉じこもり", "IADL低下", "うつ", "社会参加あり"]
hdr = ["町", "地区", "回答者数"]
for nm in SELN:
    hdr += [nm + "\n粗率", nm + "\n年齢調整"]
r = header(ws, 4, hdr, height=40)

for tname, aname, mask in rows:
    n = int(mask.sum())
    vals = [tname, aname, n]
    for key in SEL:
        v, _ = rate(mask, key)
        vals += [v, adjusted(mask, key)]
    fills = {i: GRAY for i in range(1, 14)} if aname == "全体" else (
        {2: MID_B} if aname == "町計" else None)
    r = body(ws, r, vals, fills, height=18, align={1: "left", 2: "left"},
             bold=(aname in ("全体", "町計")))

r += 1
r = lead(ws, r, "【町間の差の検定（粗率）】")
r = header(ws, r, ["指標", "東川町", "美瑛町", "東神楽町", "最大差",
                   "美瑛町と東神楽町の差の検定", "", "", "", "", "", "", ""])
mB, mH, mHG = A["町"] == "美瑛町", A["町"] == "東川町", A["町"] == "東神楽町"
for key, nm in IND:
    vH, _ = rate(mH, key)
    vB, _ = rate(mB, key)
    vG, _ = rate(mHG, key)
    if None in (vH, vB, vG):
        continue
    r = body(ws, r, [nm, vH, vB, vG, round(max(vH, vB, vG) - min(vH, vB, vG), 1),
                     ztest(mB, mHG, key), "", "", "", "", "", "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％。年齢調整は直接法（基準人口＝広域連合全体の年齢階級別回答者数）による。"
     "いずれかの年齢階級の分母が5票未満の場合は当該階級を除外し、"
     "基準人口の8割に満たない場合は算定していない。"
     "注2）美瑛町のフレイル該当割合21.6％は東神楽町17.0％より4.6ポイント高いが、"
     "美瑛町は回答者に占める75歳以上の割合が高い。"
     "年齢調整後の値を比較することで、年齢構成の影響を除いた地域差を確認できる。"
     "注3）検定は2標本の母比率の差の検定（両側・有意水準5％）による。", 13)

# ============================================================ 03 外出手段
ws = sheet("03_外出手段と移動制約", "外出時の交通手段と移動制約",
           "外出時に利用している交通手段（複数回答）を町別・地区別に集計する。"
           "美瑛町では国鉄バスの路線廃止によりスクールバスが交通弱者の公共交通となっており、"
           "町道414路線346.6kmの除雪率は52.9％である。"
           "自分で運転できない層の分布は、訪問・通所困難地域の判定及び"
           "通所系サービスの利用可能性の検討に直結する。",
           [16, 18, 8] + [9] * 12)

MEANS = [("gout2wk25", "徒歩"), ("gout2bi25", "自転車"), ("gout2dv25", "自動車\n（自分で運転）"),
         ("gout2rd25", "自動車\n（乗せてもらう）"), ("gout2pb25", "路線バス"),
         ("gout2cb25", "コミュニティ\nバス"), ("gout2tx25", "タクシー"),
         ("gout2fc25", "病院や施設\nのバス"), ("gout2tr25", "電車"),
         ("gout2sc25", "歩行器・\nシルバーカー")]
r = header(ws, 4, ["町", "地区", "回答者数"] + [t for _, t in MEANS]
           + ["自力移動の\n手段なし", "外出頻度が\n月1〜3回以下"], height=44)

for tname, aname, mask in rows:
    n = int(mask.sum())
    vals = [tname, aname, n]
    for key, _ in MEANS:
        v, _ = rate(mask, key)
        vals.append(v)
    vals.append(rate(mask, "自力移動なし")[0])
    vals.append(rate(mask, "外出低頻度")[0])
    fills = {i: GRAY for i in range(1, 16)} if aname == "全体" else (
        {2: MID_B} if aname == "町計" else None)
    r = body(ws, r, vals, fills, height=18, align={1: "left", 2: "left"},
             bold=(aname in ("全体", "町計")))

r += 1
r = lead(ws, r, "【自分で運転しない層のリスク指標】")
r = header(ws, r, ["区分", "回答者数"] + [t for _, t in IND[:8]] + ["", "", "", "", ""])
for lb, m in [("自分で運転する", series("運転あり") == 1),
              ("自分では運転しない", series("運転あり") == 0),
              ("（再掲）運転せず、他者の運転で移動する", (series("運転あり") == 0)
               & (num("gout2rd25") == 1)),
              ("（再掲）自力移動の手段なし（運転・自転車・バイクのいずれもなし）",
               series("自力移動なし") == 1)]:
    vals = [lb, int(m.sum())] + [rate(m, k)[0] for k, _ in IND[:8]] + ["", "", "", "", ""]
    r = body(ws, r, vals, height=18)
r = body(ws, r, ["差の検定（運転する／しない）", ""]
         + [ztest(series("運転あり") == 1, series("運転あり") == 0, k) for k, _ in IND[:8]]
         + ["", "", "", "", ""], {1: IN_Y}, height=18)

note(ws, r + 1,
     "注1）単位は％。交通手段は複数回答のため合計は100％を超える。"
     "分母は交通手段の設問に1つ以上回答した4,680票。"
     "注2）「自力移動の手段なし」は、自分での自動車運転・自転車・バイクのいずれも"
     "利用していない者の割合である。"
     "注3）自分で運転しない層はフレイル・閉じこもり・IADL低下の割合が高いが、"
     "この差には年齢の交絡が含まれる。運転の可否は年齢と強く相関するため、"
     "因果関係の解釈には年齢調整又は年齢階級別の比較を要する。"
     "注4）路線バス・コミュニティバスの利用割合が低い地区は、"
     "公共交通による通所が現実的でないことを示唆する。"
     "第10期基本指針案の別表 一が求める「訪問・通所困難地域」の判定においては、"
     "事業所の通常の事業の実施地域（確認事項No.13）と本表を併せて用いる。"
     "注5）「（再掲）」の2行は「自分では運転しない」の内訳であるが、"
     "他者の運転による移動と自力移動の手段なしは重複するため、合計は一致しない。"
     "注6）志比内小学校区は21票であり、徒歩0％・路線バス0％等の値は標本の偏りによる。"
     "同地区単独では判断材料としない。", 15)

# ============================================================ 04 外出の抑制
ws = sheet("04_外出の抑制と生活動作の困りごと", "外出の抑制とその理由",
           "外出を控えているかどうか、及びその理由を集計する。"
           "外出の抑制に関する設問はセクションAの調査票にのみ含まれるため分母が小さい。"
           "外出頻度及び前年比の変化は共通設問であり分母が大きい。",
           [30, 10, 10, 10, 10, 12, 12, 12])

r = lead(ws, 4, "【外出頻度（共通設問・全数）】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "回答者数", "", ""])
FQ = [("週に5回以上", [1]), ("週4回", [2]), ("週2〜3回", [3]), ("週1回", [4]),
      ("月1〜3回", [5]), ("年に数回", [6]), ("していない", [7])]
for lb, vs in FQ:
    A["_t"] = flag_in("gout7fq25", vs)
    r = body(ws, r, [lb, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], "", ""], height=18)
A["_t"] = flag_in("gout7fq25", [5, 6, 7])
r = body(ws, r, ["【再掲】月1〜3回以下", rate(ALL, "_t")[0], rate(mH, "_t")[0],
                 rate(mB, "_t")[0], rate(mHG, "_t")[0], rate(ALL, "_t")[1], "", ""],
         {1: IN_Y}, height=18, bold=True)

r += 1
r = lead(ws, r, "【前年と比べた外出回数（共通設問・全数）】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "回答者数", "", ""])
for lb, vs in [("とても減っている", [1]), ("減っている", [2]),
               ("あまり減っていない", [3]), ("減っていない", [4])]:
    A["_t"] = flag_in("gout4la25", vs)
    r = body(ws, r, [lb, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], "", ""], height=18)

r += 1
r = lead(ws, r, "【外出を控えているか（セクションAの調査票のみ）】")
r = header(ws, r, ["区分", "割合", "該当数", "分母", "", "", "", ""])
for lb, vs in [("控えている", [1]), ("控えていない", [2])]:
    A["_t"] = flag_in("gout2ref25", vs)
    v, n = rate(ALL, "_t")
    r = body(ws, r, [lb, v, int(series("_t").sum()), n, "", "", "", ""], height=18)

r += 1
r = lead(ws, r, "【外出を控えている理由（外出を控えている者が分母）】")
r = header(ws, r, ["理由", "割合", "該当数", "分母", "計画上の意味", "", "", ""])
GORE = [("gore2pa25", "足腰などの痛み", "介護予防・リハビリテーション"),
        ("gore2dis25", "病気", "医療との連携"),
        ("gore2tr25", "交通手段がない", "移動支援・訪問通所困難地域"),
        ("gore2fu25", "外での楽しみがない", "通いの場・社会参加"),
        ("gore2re25", "トイレの心配（失禁など）", "在宅生活の支援"),
        ("gore2ec25", "経済的に出られない", "低所得層の支援"),
        ("gore2ea25", "耳の障害", "コミュニケーション支援"),
        ("gore2ey25", "目の障害", "コミュニケーション支援"),
        ("gore2st25", "障害（脳卒中の後遺症など）", "医療との連携"),
        ("gore2ot25", "その他", "―")]
mref = series("gout2ref25_flag") if "gout2ref25_flag" in A.columns else (num("gout2ref25") == 1)
for key, lb, mean in GORE:
    s = pd.to_numeric(A[key], errors="coerce")[mref]
    n = int(s.notna().sum())
    v = round(100 * s.sum() / n, 1) if n else None
    r = body(ws, r, [lb, v, int(s.sum()) if n else 0, n, mean, "", "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％。"
     "注2）外出の抑制に関する設問（Ａ【問19】6）7））はセクションAの調査票にのみ含まれ、"
     "外出を控えていると回答した者は111人、理由の回答は106票である。"
     "分母が小さいため、理由別の割合は参考値として扱い、"
     "町別・地区別の内訳は算定していない。"
     "注3）外出頻度及び前年比の変化は全数の共通設問であり、"
     "町別・地区別の比較に用いることができる。"
     "注4）「交通手段がない」を理由とする者の割合は、"
     "移動支援施策及び通所系サービスの送迎体制の検討に用いる。", 8)

# ============================================================ 05 社会参加
ws = sheet("05_社会参加とリスクの関連", "社会参加の有無とリスク指標の関連",
           "社会参加（ボランティア・スポーツの会・趣味の会・学習教養サークル・通いの場・"
           "特技を伝える活動の6区分のいずれかに月1回以上参加）の有無別にリスク指標を比較する。"
           "社会参加は年齢と相関するため、年齢階級で層化した比較を併せて示す。",
           [26, 10, 10, 10, 10, 10, 10, 12, 12])

msoc1, msoc0 = series("社会参加") == 1, series("社会参加") == 0
r = lead(ws, 4, "【社会参加の有無別（全年齢）】")
r = header(ws, r, ["指標", "参加あり", "参加なし", "差\n（ポイント）", "検定",
                   "参加あり\n分母", "参加なし\n分母", "参加あり\n95％信頼区間",
                   "参加なし\n95％信頼区間"], height=40)
for key, nm in [("frail2gp_25", "フレイルあり割合"), ("undo2gp_25", "運動機能低下者割合"),
                ("tojikomori2gp_25", "閉じこもり者割合"), ("utsucheck2gp_25", "うつ割合"),
                ("ninchicheck2gp_25", "認知機能低下者割合"),
                ("koukukinou2gp_25", "口腔機能低下者割合"),
                ("iadl1_2gp_25", "IADL低下者割合"), ("tento2gp_25", "1年間の転倒あり割合"),
                ("happy2gp_25", "幸福感がある者の割合"), ("eatalone2gp_25", "孤食者割合")]:
    v1, n1 = rate(msoc1, key)
    v0, n0 = rate(msoc0, key)
    if v1 is None or v0 is None:
        continue
    r = body(ws, r, [nm, v1, v0, round(v1 - v0, 1), ztest(msoc1, msoc0, key), n1, n0,
                     ci_text(msoc1, key), ci_text(msoc0, key)], height=18)
A["_rs"] = num("riskscore_25")
r = body(ws, r, ["要支援・要介護リスク点数（平均点）",
                 round(series("_rs")[msoc1].mean(), 1), round(series("_rs")[msoc0].mean(), 1),
                 round(series("_rs")[msoc1].mean() - series("_rs")[msoc0].mean(), 1),
                 "―", int(series("_rs")[msoc1].notna().sum()),
                 int(series("_rs")[msoc0].notna().sum()), "―", "―"], {1: IN_Y}, height=18)

r += 1
r = lead(ws, r, "【年齢階級で層化したフレイル該当割合】")
r = header(ws, r, ["年齢階級", "参加あり", "参加なし", "差\n（ポイント）", "検定",
                   "参加あり\n分母", "参加なし\n分母", "", ""])
for g in AGES:
    m1 = msoc1 & (A["年齢階級"] == g)
    m0 = msoc0 & (A["年齢階級"] == g)
    v1, n1 = rate(m1, "frail2gp_25")
    v0, n0 = rate(m0, "frail2gp_25")
    r = body(ws, r, [g, v1, v0, round(v1 - v0, 1) if None not in (v1, v0) else None,
                     ztest(m1, m0, "frail2gp_25"), n1, n0, "", ""], height=18)

r += 1
r = lead(ws, r, "【参加の区分別の参加率（月1回以上）】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母", "", "", ""])
SOCN = [("cmnt6vl25", "ボランティアのグループ"), ("cmnt6le25", "学習・教養サークル"),
        ("cmnt6sl25", "介護予防のための通いの場"), ("cmnt6sk25", "特技や経験を伝える活動")]
for key, nm in SOCN:
    if key not in A.columns:
        continue
    A["_t"] = flag_in(key, [1, 2, 3, 4])
    r = body(ws, r, [nm, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], "", "", ""], height=18)
r = body(ws, r, ["【合成】6区分のいずれかに参加", rate(ALL, "社会参加")[0],
                 rate(mH, "社会参加")[0], rate(mB, "社会参加")[0], rate(mHG, "社会参加")[0],
                 rate(ALL, "社会参加")[1], "", "", ""], {1: IN_Y}, height=18, bold=True)

note(ws, r + 1,
     "注1）単位は％（リスク点数のみ点）。"
     "注2）社会参加ありの群はフレイル・閉じこもり・うつ・IADL低下の割合がいずれも低い。"
     "ただし、これは社会参加が心身の状態を保っている効果と、"
     "心身の状態が良いから参加できているという逆方向の関係の両方を含む。"
     "本調査は横断調査であるため、因果関係を示すものではない。"
     "注3）年齢階級で層化しても同じ方向の差がみられることは、"
     "年齢の交絡だけでは説明できないことを示す。"
     "注4）代表KPI H05（社会参加率）の基準値としては、"
     "第9期の通いの場参加率7.3％を継承する場合は「通いの場」の値を、"
     "社会参加の広がりを捉える場合は合成指標を用いる。"
     "注5）要支援・要介護リスク点数のみ参加あり群の方がわずかに高いが、"
     "これは社会参加の割合が65〜69歳28.9％に対し80〜84歳43.1％と高齢層で高く、"
     "参加あり群の年齢構成が高いことによる。"
     "同点数は年齢と強く相関するため（65〜69歳5.1点、85歳以上30.0点）、"
     "年齢調整を行わない群間比較には適さない。"
     "年齢階級で層化した比較では、他の指標と同様に参加あり群が良好である。", 9)

# ============================================================ 06 世帯構成
ws = sheet("06_世帯構成とリスク", "世帯構成別のリスク指標",
           "独居・夫婦のみ・その他の別にリスク指標を比較する。"
           "令和2年国勢調査では、高齢者を含む世帯に占める高齢者単身世帯の割合は"
           "美瑛町が30.3％と最も高く、高齢者夫婦世帯の割合は東神楽町が37.1％と最も高い。",
           [26, 10, 10, 10, 10, 10, 12, 12, 12])

mal, mcp = series("独居") == 1, series("夫婦のみ") == 1
moth = (series("独居") == 0) & (series("夫婦のみ") != 1)
r = lead(ws, 4, "【世帯構成別のリスク指標】")
r = header(ws, r, ["指標", "独居", "夫婦のみ", "その他", "独居とその他の差",
                   "検定", "独居\n分母", "夫婦のみ\n分母", "その他\n分母"], height=40)
for key, nm in [("frail2gp_25", "フレイルあり割合"), ("tojikomori2gp_25", "閉じこもり者割合"),
                ("utsucheck2gp_25", "うつ割合"), ("ninchicheck2gp_25", "認知機能低下者割合"),
                ("iadl1_2gp_25", "IADL低下者割合"), ("eatalone2gp_25", "孤食者割合"),
                ("社会参加", "社会参加あり割合"), ("通いの場", "通いの場参加割合"),
                ("happy2gp_25", "幸福感がある者の割合"),
                ("困りごとあり", "生活動作の困りごとあり割合"),
                ("世話をする人なし", "世話をしてくれる人がいない割合")]:
    v1, n1 = rate(mal, key)
    v2, n2 = rate(mcp, key)
    v3, n3 = rate(moth, key)
    if v1 is None or v3 is None:
        continue
    r = body(ws, r, [nm, v1, v2, v3, round(v1 - v3, 1), ztest(mal, moth, key),
                     n1, n2, n3], height=18)

r += 1
r = lead(ws, r, "【町別・地区別の独居者割合】")
r = header(ws, r, ["町", "地区", "回答者数", "独居者割合", "夫婦のみ割合",
                   "世話をしてくれる\n人がいない割合", "", "", ""], height=36)
for tname, aname, mask in rows:
    r = body(ws, r, [tname, aname, int(mask.sum()), rate(mask, "独居")[0],
                     rate(mask, "夫婦のみ")[0], rate(mask, "世話をする人なし")[0],
                     "", "", ""], height=18,
             bold=(aname in ("全体", "町計")))

note(ws, r + 1,
     "注1）単位は％。「独居」はリスク指標58（独居者割合）による。"
     "「夫婦のみ」は同居者が配偶者のみで世帯人員が2人の者とした。"
     "注2）独居者はフレイル・閉じこもり・うつ・孤食の割合が高く、"
     "社会参加の割合も低い傾向がある。"
     "注3）「世話をしてくれる人がいない割合」は大雪独自設問【問4】1）による。"
     "この指標は、在宅生活の継続可能性及び身寄りのない高齢者への支援の"
     "検討に用いることができる（第5章 基本目標2・3）。", 9)

# ============================================================ 07 経済状況
ws = sheet("07_経済状況とリスク", "経済状況とリスク指標",
           "暮らし向き（経済的にみた現在の暮らしの状況）の5段階別にリスク指標を比較する。"
           "認定を受けながらサービスを利用していない方が504人（認定者の25.7％）おり、"
           "そのうちに経済的理由によるものが含まれるかどうかは第10期の論点である。",
           [26, 11, 11, 11, 11, 11, 11, 12, 12])

SFS = [("大変苦しい", [1]), ("やや苦しい", [2]), ("ふつう", [3]),
       ("ややゆとりがある", [4]), ("大変ゆとりがある", [5])]
r = lead(ws, 4, "【暮らし向き別のリスク指標】")
r = header(ws, r, ["指標"] + [n for n, _ in SFS] + ["苦しい計\n（1・2）", "ゆとり計\n（4・5）",
                                                     "苦しい計と\nふつうの差の検定"], height=40)
msfs = {n: flag_in("sfs5_25", v) == 1 for n, v in SFS}
mhard = series("暮らし苦しい") == 1
mnorm = flag_in("sfs5_25", [3]) == 1
measy = flag_in("sfs5_25", [4, 5]) == 1
for key, nm in [("frail2gp_25", "フレイルあり割合"), ("tojikomori2gp_25", "閉じこもり者割合"),
                ("utsucheck2gp_25", "うつ割合"), ("iadl1_2gp_25", "IADL低下者割合"),
                ("teieiyo2gp_25", "低栄養者割合"), ("eatalone2gp_25", "孤食者割合"),
                ("社会参加", "社会参加あり割合"), ("通いの場", "通いの場参加割合"),
                ("happy2gp_25", "幸福感がある者の割合"), ("独居", "独居者割合"),
                ("困りごとあり", "生活動作の困りごとあり割合"),
                ("解決できず困っている", "解決できず困っている割合")]:
    if key not in A.columns:
        continue
    vs = [rate(msfs[n], key)[0] for n, _ in SFS]
    if all(v is None for v in vs):
        continue
    r = body(ws, r, [nm] + vs + [rate(mhard, key)[0], rate(measy, key)[0],
                                 ztest(mhard, mnorm, key)], height=18)
r = body(ws, r, ["回答者数（人）"] + [int(msfs[n].sum()) for n, _ in SFS]
         + [int(mhard.sum()), int(measy.sum()), ""], {1: GRAY}, height=18)

r += 1
r = lead(ws, r, "【町別の経済状況】")
r = header(ws, r, ["町", "大変苦しい", "やや苦しい", "苦しい計", "ふつう",
                   "ゆとり計", "回答者数", "", ""])
for tname, m in [("大雪地区広域連合", ALL), ("東川町", mH), ("美瑛町", mB), ("東神楽町", mHG)]:
    A["_h"] = flag_in("sfs5_25", [1])
    A["_h2"] = flag_in("sfs5_25", [2])
    A["_n"] = flag_in("sfs5_25", [3])
    A["_e"] = flag_in("sfs5_25", [4, 5])
    r = body(ws, r, [tname, rate(m, "_h")[0], rate(m, "_h2")[0], rate(m, "暮らし苦しい")[0],
                     rate(m, "_n")[0], rate(m, "_e")[0], rate(m, "暮らし苦しい")[1],
                     "", ""], height=18, bold=(tname == "大雪地区広域連合"))

note(ws, r + 1,
     "注1）単位は％。分母は暮らし向きの設問に回答した4,596票。"
     "注2）暮らし向きが「苦しい」層はフレイル・閉じこもり・うつ・孤食の割合が高く、"
     "社会参加の割合が低い。経済状況と心身の状態及び社会参加は関連している。"
     "注3）認定を受けながらサービスを利用していない504人について、"
     "経済的理由によるものがあるかどうかは本調査からは判別できない。"
     "本調査の対象者は一般高齢者及び総合事業対象者であり、"
     "要支援者・要介護者を含まないためである（14シート）。"
     "未利用の要因の判別には、居宅サービス計画未作成者の抽出等の代理指標を用いる"
     "（素案 第2章第2節3）。", 9)

# ============================================================ 08 在宅生活の課題
ws = sheet("08_在宅生活の課題と解決手段", "生活動作の困りごとと解決手段（大雪独自設問）",
           "【大雪－問2】生活動作の中で不安や困っていると感じていることと、"
           "その解決手段を集計する。本設問は当広域連合が独自に追加したものであり、"
           "在宅生活改善調査が把握しようとしている在宅生活の継続困難要因に対応する。",
           [30, 10, 10, 10, 10, 10, 12, 12, 12])

Q21LAB = [("taisetsu_q2_1_s12", "除雪"), ("taisetsu_q2_1_s4", "庭の手入れ"),
          ("taisetsu_q2_1_s13", "災害時の避難"), ("taisetsu_q2_1_s15", "布団干し"),
          ("taisetsu_q2_1_s3", "風呂やトイレの掃除"), ("taisetsu_q2_1_s2", "部屋の掃除や片付け"),
          ("taisetsu_q2_1_s14", "簡単な修繕や電球替え"), ("taisetsu_q2_1_s16", "季節の衣服の入れ替え"),
          ("taisetsu_q2_1_s1", "食事の準備や片付け"), ("taisetsu_q2_1_s7", "買い物"),
          ("taisetsu_q2_1_s8", "通院"), ("taisetsu_q2_1_s11", "外出"),
          ("taisetsu_q2_1_s6", "ゴミの分別やゴミ出し"), ("taisetsu_q2_1_s5", "衣服の洗濯や片付け"),
          ("taisetsu_q2_1_s9", "薬の管理"), ("taisetsu_q2_1_s10", "預貯金の管理"),
          ("taisetsu_q2_1_s17", "犬の散歩などペットの世話"),
          ("taisetsu_q2_1_s18", "話し相手がいない"), ("taisetsu_q2_1_s19", "趣味や役割がない"),
          ("taisetsu_q2_1_s20", "その他"), ("taisetsu_q2_1_s21", "特になし")]
Q21LAB = [(k, n) for k, n in Q21LAB if k in A.columns]

r = lead(ws, 4, "【生活動作の困りごと（複数回答・全数）】")
r = header(ws, r, ["項目", "広域連合", "東川町", "美瑛町", "東神楽町", "該当数", "独居者",
                   "75歳以上", "美瑛町と東神楽町\nの差の検定"], height=40)
m75 = A["年齢階級"].isin(["75～79歳", "80～84歳", "85歳以上"])
res = []
for key, nm in Q21LAB:
    if nm == "特になし":
        continue
    A["_t"] = pd.to_numeric(A[key], errors="coerce")
    v, n = rate(ALL, "_t")
    res.append((v if v is not None else -1, key, nm))
res = sorted(res, reverse=True)
res += [(-1, k, n) for k, n in Q21LAB if n == "特になし"]
for _, key, nm in res:
    A["_t"] = pd.to_numeric(A[key], errors="coerce")
    v, n = rate(ALL, "_t")
    fill = {1: IN_Y} if nm == "特になし" else None
    r = body(ws, r, [nm, v, rate(mH, "_t")[0], rate(mB, "_t")[0], rate(mHG, "_t")[0],
                     int(series("_t").sum()), rate(mal, "_t")[0], rate(m75, "_t")[0],
                     ztest(mB, mHG, "_t")], fill, height=18)
r = body(ws, r, ["【再掲】いずれかの困りごとあり", rate(ALL, "困りごとあり")[0],
                 rate(mH, "困りごとあり")[0], rate(mB, "困りごとあり")[0],
                 rate(mHG, "困りごとあり")[0], int(series("困りごとあり").sum()),
                 rate(mal, "困りごとあり")[0], rate(m75, "困りごとあり")[0],
                 ztest(mB, mHG, "困りごとあり")], {1: OK_G}, height=18, bold=True)

r += 1
r = lead(ws, r, "【困りごとの解決手段（困りごとがある者が分母）】")
r = header(ws, r, ["解決手段", "割合", "該当数", "分母", "独居者", "75歳以上",
                   "計画上の意味", "", ""], height=32)
Q22 = [("taisetsu_q2_2_s8", "自力で何とかしている", "支援につながっていない可能性"),
       ("taisetsu_q2_2_s1", "家族や親族の手助け", "家族介護力への依存"),
       ("taisetsu_q2_2_s2", "近所に住む方々の手助け", "地域の支え合い"),
       ("taisetsu_q2_2_s3", "友人や知人のサポート", "地域の支え合い"),
       ("taisetsu_q2_2_s6", "介護保険などのサービス", "保険給付でのカバー"),
       ("taisetsu_q2_2_s7", "民間業者のサービス", "インフォーマルサービス"),
       ("taisetsu_q2_2_s5", "シルバー人材センターなどのサービス", "生活支援体制整備"),
       ("taisetsu_q2_2_s4", "ボランティアのサポート", "生活支援体制整備"),
       ("taisetsu_q2_2_s10", "解決できず、困っている", "未充足ニーズ（H06の候補）"),
       ("taisetsu_q2_2_s9", "その他", "―")]
mneed = series("困りごとあり") == 1
for key, nm, mean in Q22:
    if key not in A.columns:
        continue
    s = pd.to_numeric(A[key], errors="coerce")[mneed]
    n = int(s.notna().sum())
    v = round(100 * s.sum() / n, 1) if n else None
    A["_t"] = pd.to_numeric(A[key], errors="coerce")
    fill = {1: NG_O} if nm == "解決できず、困っている" else None
    r = body(ws, r, [nm, v, int(s.sum()) if n else 0, n,
                     rate(mneed & mal, "_t")[0], rate(mneed & m75, "_t")[0], mean, "", ""],
             fill, height=18)

note(ws, r + 1,
     "注1）単位は％。複数回答のため合計は100％を超える。"
     "困りごとの分母は当該設問群に回答した4,128票、"
     "解決手段の分母は困りごとがあると回答した者である。"
     "注2）「除雪」が最上位となることは、当広域連合の3町がいずれも豪雪地帯対策特別措置法の"
     "特別豪雪地帯又は豪雪地帯に指定されていること、"
     "美瑛町の町道除雪率が52.9％であることと整合する。"
     "除雪は介護保険の給付対象外であり、生活支援体制整備事業及び3町の施策で対応する領域である。"
     "注3）「解決できず、困っている」と回答した者の割合は、"
     "在宅生活改善調査が把握しようとしている在宅生活の継続困難要因に対応しており、"
     "代表KPI H06（在宅生活継続困難割合）の基準値として用いることができる（11シート）。"
     "注4）「自力で何とかしている」の割合が高いことは、"
     "支援につながっていない層が一定数存在することを示唆する。", 9)

# ============================================================ 09 住み続ける意向
ws = sheet("09_住み続ける意向と転居理由", "住み続ける意向と転居を考える理由（大雪独自設問）",
           "【大雪－問3】現在住んでいる地域に住み続けたいかどうかと、転居を考える理由を集計する。"
           "転居理由には「今住んでいる地域では通院や買い物などが困難なため」"
           "「希望する介護施設に入所できないため」"
           "「希望する生活支援サービスが受けることができないため」が含まれ、"
           "居所変更実態調査が把握しようとしている供給不足を理由とする住替えに対応する。",
           [30, 10, 10, 10, 10, 10, 12, 12, 12])

r = lead(ws, 4, "【住み続ける意向】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "75歳以上", "自力移動の\n手段なし"], height=36)
mnodrive = series("自力移動なし") == 1
for lb, vs in [("住み続けたい", [1]), ("転居も少し考えている", [2]), ("わからない", [3])]:
    A["_t"] = flag_in("taisetsu_q3_1", vs)
    r = body(ws, r, [lb, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], rate(mal, "_t")[0],
                     rate(m75, "_t")[0], rate(mnodrive, "_t")[0]], height=18)

r += 1
r = lead(ws, r, "【地区別の「転居も少し考えている」割合】")
r = header(ws, r, ["町", "地区", "回答者数", "住み続けたい", "転居も少し\n考えている",
                   "わからない", "", "", ""], height=32)
A["_stay"] = flag_in("taisetsu_q3_1", [1])
A["_move"] = flag_in("taisetsu_q3_1", [2])
A["_unk"] = flag_in("taisetsu_q3_1", [3])
for tname, aname, mask in rows:
    r = body(ws, r, [tname, aname, int(mask.sum()), rate(mask, "_stay")[0],
                     rate(mask, "_move")[0], rate(mask, "_unk")[0], "", "", ""],
             height=18, bold=(aname in ("全体", "町計")))

r += 1
r = lead(ws, r, "【転居を考える理由（転居も少し考えていると回答した者が分母）】")
r = header(ws, r, ["理由", "割合", "該当数", "分母", "計画上の意味", "", "", "", ""], height=32)
mmove = series("_move") == 1
Q32 = [("taisetsu_q3_2_s1", "子などと一緒に暮らすため", "家族との同居（供給とは無関係）"),
       ("taisetsu_q3_2_s2", "今住んでいる地域では通院や買い物などが困難なため",
        "移動支援・訪問通所困難地域"),
       ("taisetsu_q3_2_s3", "希望する介護施設に入所できないため", "施設整備（H11の候補）"),
       ("taisetsu_q3_2_s4", "希望する生活支援サービスが受けることができないため",
        "サービス供給不足（H11の候補）"),
       ("taisetsu_q3_2_s5", "その他", "―")]
for key, nm, mean in Q32:
    if key not in A.columns:
        continue
    s = pd.to_numeric(A[key], errors="coerce")[mmove]
    n = int(s.notna().sum())
    v = round(100 * s.sum() / n, 1) if n else None
    fill = {1: NG_O} if "希望する" in nm or "困難" in nm else None
    r = body(ws, r, [nm, v, int(s.sum()) if n else 0, n, mean, "", "", "", ""],
             fill, height=18)
SUP3 = [c for c in ["taisetsu_q3_2_s2", "taisetsu_q3_2_s3", "taisetsu_q3_2_s4"]
        if c in A.columns]
A["_supply"] = any_of(SUP3)                       # 分母＝転居理由に回答した者
# 分母を住み続ける意向の設問に回答した者（4,446票）に広げた系列
A["_supply_all"] = (series("_supply") == 1).astype(float).where(num("taisetsu_q3_1").notna())
s = series("_supply")[mmove]
n = int(s.notna().sum())
r = body(ws, r, ["【再掲】供給・アクセスを理由とする転居意向（2〜4のいずれか）",
                 round(100 * s.sum() / n, 1) if n else None, int(s.sum()) if n else 0, n,
                 "分母＝転居理由の設問に回答した者", "", "", "", ""],
         {1: MID_B}, height=18, bold=True)
sa = series("_supply_all")
na = int(sa.notna().sum())
r = body(ws, r, ["【再掲】回答者全体に対する割合",
                 round(100 * sa.sum() / na, 1) if na else None, int(sa.sum()) if na else 0,
                 na, "H11（供給不足を理由とする住替え割合）の候補", "", "", "", ""],
         {1: OK_G}, height=18, bold=True)

note(ws, r + 1,
     "注1）単位は％。転居理由は複数回答のため合計は100％を超える。"
     "注2）転居を考える理由のうち、「通院や買い物などが困難」"
     "「希望する介護施設に入所できない」「希望する生活支援サービスが受けることができない」の"
     "3つは、いずれもサービス供給又はアクセスの不足を理由とするものであり、"
     "居所変更実態調査が把握しようとしている供給不足を理由とする住替えに対応する。"
     "注3）代表KPI H11の基準値としては、転居意向がある者を分母とする割合と、"
     "回答者全体を分母とする割合の両方を示した。"
     "計画では母集団の解釈が明確な後者を用いることを提案する。"
     "注4）本設問は意向であり、実際の住替えの実績ではない。"
     "実績の把握は居所変更実態調査（第10期分）による。"
     "注5）地区別では美馬牛小学校区の「転居も少し考えている」が23.4％と高い。"
     "同地区は自分で運転する者の割合が83.8％と高く、徒歩22.3％・タクシー2.0％と低いことから、"
     "運転できなくなった後の生活の見通しが転居意向に表れている可能性がある。"
     "志比内小学校区（21票）の値は標本が小さく、単独では判断材料としない。", 9)

# ============================================================ 10 介護の担い手
ws = sheet("10_介護の担い手と介護の意向", "介護の担い手と介護に関する意向（大雪独自設問）",
           "【大雪－問4】日常的に支障が生じた場合に世話をしてくれる家族等の有無、"
           "介護を受けたい場所、人生会議（ACP）及び死後事務委任契約の認知度を集計する。"
           "在宅介護実態調査が把握しようとしている家族の介護力に対応する。",
           [30, 10, 10, 10, 10, 10, 12, 12, 12])

r = lead(ws, 4, "【世話をしてくれる家族等（複数回答）】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "75歳以上", "85歳以上"], height=32)
m85 = A["年齢階級"] == "85歳以上"
Q41 = [("taisetsu_q4_1_s1", "配偶者"), ("taisetsu_q4_1_s2", "子"),
       ("taisetsu_q4_1_s3", "子の配偶者"), ("taisetsu_q4_1_s4", "孫"),
       ("taisetsu_q4_1_s5", "兄弟・姉妹"), ("taisetsu_q4_1_s6", "近所の人"),
       ("taisetsu_q4_1_s7", "友人"), ("taisetsu_q4_1_s8", "その他の人"),
       ("taisetsu_q4_1_s9", "そのような人はいない")]
for key, nm in Q41:
    if key not in A.columns:
        continue
    A["_t"] = pd.to_numeric(A[key], errors="coerce")
    fill = {1: NG_O} if nm == "そのような人はいない" else None
    r = body(ws, r, [nm, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], rate(mal, "_t")[0],
                     rate(m75, "_t")[0], rate(m85, "_t")[0]], fill, height=18)

r += 1
r = lead(ws, r, "【介護が必要になった場合に介護を受けたい場所】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "75歳以上", "計画上の意味"], height=32)
Q42 = [("なるべく家族のみで、自宅で介護してほしい", [1], "家族介護力への依存"),
       ("介護保険のサービスや福祉サービスを利用しながら、自宅で介護してほしい", [2],
        "在宅サービスの需要"),
       ("老人ホームなどの施設に入所したい", [3], "施設サービスの需要"),
       ("その他", [4], "―"), ("わからない", [5], "意思決定支援・ACP")]
for nm, vs, mean in Q42:
    A["_t"] = flag_in("taisetsu_q4_2", vs)
    r = body(ws, r, [nm, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], rate(mal, "_t")[0],
                     rate(m75, "_t")[0], mean], height=30)
A["_home"] = flag_in("taisetsu_q4_2", [1, 2])
r = body(ws, r, ["【再掲】自宅での介護を希望（1・2の計）", rate(ALL, "_home")[0],
                 rate(mH, "_home")[0], rate(mB, "_home")[0], rate(mHG, "_home")[0],
                 rate(ALL, "_home")[1], rate(mal, "_home")[0], rate(m75, "_home")[0],
                 "在宅サービスの見込量（第6章第2節）"], {1: OK_G}, height=18, bold=True)

r += 1
r = lead(ws, r, "【意思決定支援に関する認知度】")
r = header(ws, r, ["区分", "既に実施\nしている", "知っている", "言葉だけは\n知っている",
                   "知らない", "分母", "75歳以上で\n「知らない」", "", ""], height=32)
for key, nm in [("taisetsu_q4_4", "人生会議（ACP）"), ("taisetsu_q4_5", "死後事務委任契約")]:
    if key not in A.columns:
        continue
    vs = []
    for v in [1, 2, 3, 4]:
        A["_t"] = flag_in(key, [v])
        vs.append(rate(ALL, "_t")[0])
    A["_t"] = flag_in(key, [4])
    r = body(ws, r, [nm] + vs + [rate(ALL, "_t")[1], rate(m75, "_t")[0], "", ""], height=18)

r += 1
r = lead(ws, r, "【おひとりになった場合の終のすみか】")
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "", ""], height=32)
Q46 = [("現在のまま、自宅に住み続けたい", [1]), ("改築の上、自宅に住み続けたい", [6]),
       ("高齢者用住宅へ引っ越したい", [2]), ("老人ホームへ入居したい", [3]),
       ("子どもの住宅へ引っ越したい", [5]), ("病院に入院したい", [7]),
       ("実家（田舎）に引っ越したい", [4]), ("その他", [8])]
for nm, vs in Q46:
    A["_t"] = flag_in("taisetsu_q4_6", vs)
    r = body(ws, r, [nm, rate(ALL, "_t")[0], rate(mH, "_t")[0], rate(mB, "_t")[0],
                     rate(mHG, "_t")[0], rate(ALL, "_t")[1], rate(mal, "_t")[0], "", ""],
             height=18)

note(ws, r + 1,
     "注1）単位は％。世話をしてくれる家族等は複数回答のため合計は100％を超える。"
     "注2）「そのような人はいない」と回答した者の割合は、独居者及び85歳以上で高い。"
     "身寄りのない高齢者の入居・入所、保証、死後事務等の支援体制"
     "（第5章 基本目標3）の対象規模を示す指標となる。"
     "注3）介護を受けたい場所について「介護保険のサービスや福祉サービスを利用しながら、"
     "自宅で介護してほしい」が最も多いことは、"
     "在宅サービスの見込量（第6章第2節）の前提となる。"
     "一方、当広域連合には定期巡回・随時対応型訪問介護看護、夜間対応型訪問介護、"
     "看護小規模多機能型居宅介護の事業所が制度創設以来存在せず、"
     "重度者の在宅生活を24時間支える体制がない。"
     "注4）人生会議（ACP）を「知らない」と回答した者が多数を占めることは、"
     "第10期における意思決定支援の普及啓発の必要性を示す。", 9)

# ============================================================ 11 KPIのデータ源
ws = sheet("11_KPIのデータ源の再検討", "本分析により算定可能となった代表KPI",
           "第10稿では、事業所実態調査及び在宅介護実態調査を当広域連合が実施していないことから、"
           "H07・H08・H12・H16の4項目を「算定不可」又は「データ源の確保が前提」としていた。"
           "本分析の結果、大雪独自設問により一部の指標が算定できることが判明したため、"
           "データ源の再検討結果を示す。",
           [7, 24, 22, 26, 30, 12, 12, 12])

r = header(ws, 4, ["ID", "指標", "第10稿での扱い", "本分析による代替データ源",
                   "算定方法", "基準値\n（令和7年度）", "分母", "判定"], height=40)


def val_of(name, mask=None):
    v, n = rate(mask if mask is not None else ALL, name)
    return v, n


kpi_rows = []
A["_unsolved_all"] = ((series("解決できず困っている") == 1).astype(float)
                      .where(series("困りごとあり").notna()))
sv = series("解決できず困っている")[mneed]
n_sub = int(sv.notna().sum())
v_sub = round(100 * sv.sum() / n_sub, 1) if n_sub else None
sa2 = series("_unsolved_all")
n_all = int(sa2.notna().sum())
v_all = round(100 * sa2.sum() / n_all, 1) if n_all else None
kpi_rows.append(("H06", "在宅生活継続困難割合", "第10期調査で算定（未受領）",
                 "健康とくらしの調査【大雪－問2】1）2）（補完）",
                 "生活動作の困りごとについて「解決できず、困っている」と回答した者を"
                 "分子とする。分母は回答者全体（%d票）とする案と、"
                 "困りごとがある者（%d票）とする案がある。前者では%.1f％、後者では%.1f％となる"
                 % (n_all, n_sub, v_all, v_sub), v_all, n_all, "補完指標として算定可能"))
sup_all = series("_supply_all")
n_h = int(sup_all.notna().sum())
v_h = round(100 * sup_all.sum() / n_h, 1) if n_h else None
sup_sub = series("_supply")[mmove]
n_h2 = int(sup_sub.notna().sum())
v_h2 = round(100 * sup_sub.sum() / n_h2, 1) if n_h2 else None
kpi_rows.append(("H11", "供給不足を理由とする住替え意向割合", "第10期調査で算定（未受領）",
                 "健康とくらしの調査【大雪－問3】1）2）（補完）",
                 "転居を考える理由として「通院や買い物などが困難」「希望する介護施設に"
                 "入所できない」「希望する生活支援サービスが受けることができない」の"
                 "いずれかを選んだ者を分子とする。分母を回答者全体（%d票）とすると%.1f％、"
                 "転居理由に回答した者（%d票）とすると%.1f％となる"
                 % (n_h, v_h, n_h2, v_h2), v_h, n_h, "補完指標として算定可能"))
A["_care"] = flag_in("care3fam25", [1])
v_c, n_c = val_of("_care")
kpi_rows.append(("H07", "主介護者の高負担割合", "算定不可（在宅介護実態調査が未実施）",
                 "健康とくらしの調査 サブコア1【問17】4）",
                 "「主に介護をしている」と回答した者÷回答者×100。"
                 "負担感そのものではなく主介護者の割合であるため、指標の定義変更を要する",
                 v_c, n_c, "代替指標として算定可能"))
A["_nohelp"] = flag_in("taisetsu_q4_1_s9", [1])
v_x, n_x = val_of("_nohelp")
kpi_rows.append(("H08", "家族介護者の支援充足割合", "算定不可（在宅介護実態調査が未実施）",
                 "健康とくらしの調査【大雪－問4】1）",
                 "「日常的に支障が生じた場合に世話をしてくれる人がいない」と回答した者÷"
                 "回答者×100。支援の充足ではなく担い手の不在を測る指標に変更する",
                 v_x, n_x, "代替指標として算定可能"))
kpi_rows.append(("H12", "必要サービス未充足率", "算定不可（事業所実態調査が未実施）",
                 "①在宅生活改善調査 問3（本調査では算定できない）",
                 "本調査は一般高齢者の調査であり、事業所側の受入れ可否を把握できない。"
                 "①在宅生活改善調査の問3で、区域内に事業所が存在しない4サービス"
                 "（夜間対応型訪問介護・定期巡回・随時対応型訪問介護看護・"
                 "看護小規模多機能型居宅介護・介護医療院）を"
                 "「より適切と思われるサービス」として1つ以上選んだ票を分子とする。"
                 "98票中20票で20.4％。町別は東川町33.3％・美瑛町18.8％・東神楽町17.6％",
                 20.4, 98, "別調査により算定可能"))
kpi_rows.append(("H16", "災害・感染症時の必須サービス継続率", "算定不可（事業所実態調査が未実施）",
                 "国民健康保険団体連合会の給付実績（本調査では算定できない）",
                 "令和6年度の報酬改定で業務継続計画未策定減算が新設されており"
                 "（施設系3％・その他1％）、"
                 "同減算を算定していない事業所の割合として算定できる。"
                 "災害・感染症が発生していないため実際の継続実績は測定できず、"
                 "「備えの整備率」への改称を要する",
                 None, None, "給付実績により算定可能"))
A["_frail"] = pd.to_numeric(A["frail2gp_25"], errors="coerce")
v_f, n_f = val_of("_frail")
kpi_rows.append(("H04", "フレイル該当割合", "19.1％（R7）で確定済み",
                 "健康とくらしの調査（基本チェックリスト8項目以上）",
                 "基本チェックリスト該当項目数8以上の者÷判定可能な有効回答者×100",
                 v_f, n_f, "確定済み"))
v_s, n_s = val_of("社会参加")
kpi_rows.append(("H05", "社会参加率", "8.8％（通いの場）で確定済み",
                 "健康とくらしの調査【問5】1）",
                 "6区分（ボランティア・スポーツの会・趣味の会・学習教養サークル・"
                 "通いの場・特技を伝える活動）のいずれかに月1回以上参加した者÷回答者×100",
                 v_s, n_s, "合成指標も選択可能"))
A["_need_nouse"] = flag_in("adl3ra25", [2])
v_nn, n_nn = val_of("_need_nouse")
kpi_rows.append(("―", "介護・介助が必要だが受けていない者の割合（新規提案）", "第10稿では設定なし",
                 "健康とくらしの調査【問1】2）",
                 "「何らかの介護・介助が必要だが、現在は受けていない」と回答した者÷"
                 "回答者×100。未利用認定者504人（認定者の25.7％）の論点に対応する",
                 v_nn, n_nn, "新規に設定を提案"))

for row in kpi_rows:
    fill = {8: OK_G} if "算定可能" in str(row[7]) or row[7] == "確定済み" else {8: NG_O}
    r = body(ws, r, list(row), fill, height=64, align={1: "center", 6: "center", 8: "center"})

note(ws, r + 1,
     "注1）単位は％。基準値は令和7年度 健康とくらしの調査による。"
     "注2）H06・H11については、当初は在宅生活改善調査及び居所変更実態調査を"
     "データ源としていたが、本調査の大雪独自設問により基準値を設定できる。"
     "第10期分の各調査を受領した後に、両者の値を突合して整合を確認する。"
     "注3）H07・H08については、指標の定義を変更したうえで代替データ源を用いる案である。"
     "定義変更の可否は発注者の決定による（確認事項No.4）。"
     "注4）H12・H16は事業所側の情報を要するため、本調査では算定できない。"
     "ただしH12は①在宅生活改善調査から、H16は国民健康保険団体連合会の"
     "給付実績から算定できることが確認できた。"
     "H12の20.4％は、母集団が「在宅生活の維持が難しい利用者」の抽出であるため、"
     "区域内の在宅利用者全体の未充足率ではない。"
     "この留保を指標の定義に明記する必要がある。"
     "注5）「介護・介助が必要だが受けていない者の割合」は、"
     "認定を受けていない層の潜在的な支援ニーズを示す指標として新規に設定することを提案する。"
     "本調査の対象者は一般高齢者及び総合事業対象者であり要支援者・要介護者を含まないため、"
     "この7.0％は、認定を受けながらサービスを利用していない504人"
     "（認定者の25.7％）とは別の集団である。両者を合算してはならない。"
     "注6）本調査の対象者の制約により、H06・H11については"
     "在宅生活改善調査・居所変更実態調査の代替ではなく補完として位置づける。"
     "在宅生活改善調査は居宅介護支援事業所のケアマネジャーを通じた認定者の調査、"
     "居所変更実態調査は施設・居住系サービスの管理者を通じた実績の調査であり、"
     "いずれも母集団が異なる。第10期分を受領した後に両者を併記する。", 8)

# ============================================================ 12 主要所見
ws = sheet("12_主要所見", "クロス集計・分析の主要所見",
           "本分析から得られた所見と、計画素案への反映方針を示す。"
           "所見は根拠となる数値とシートを併記する。",
           [4, 26, 44, 30, 20, 14])

r = header(ws, 4, ["No.", "所見", "根拠となる数値", "計画への反映", "反映先", "シート"])


def g(name, mask=None):
    v, n = rate(mask if mask is not None else ALL, name)
    return "―" if v is None else "%.1f％（n=%d）" % (v, n)


FIND = [
    ("小学校区別の分析が可能",
     "個票データの分析地域1により、東神楽町3・東川町4・美瑛町5の計12小学校区別の集計ができる。"
     "美瑛町では美馬牛小学校区（149票）、美沢小学校区（57票）、明徳小学校区（49票）、"
     "美瑛東小学校区（533票）が独立して取得できる。",
     "日常生活圏域別の分析を、見える化システムのデータを待たずに着手できる。"
     "小学校区と第9期の6圏域の対応関係の確認を要する。",
     "第1章第7節\n第2章第4節", "01・02"),
    ("美瑛町の指標の差には年齢構成の影響がある",
     "美瑛町のフレイル該当割合は21.6％で東神楽町17.0％より4.6ポイント高いが、"
     "美瑛町は回答者に占める75歳以上の割合が高い。年齢調整後の値を02シートに示した。",
     "町間比較は粗率ではなく年齢調整後の値によることとし、"
     "本文の記述を「高齢化の進行度の差を含む」と明示する。", "第2章第4節", "02"),
    ("除雪が生活動作の困りごとの最上位",
     "生活動作の困りごとのうち除雪が最も多い（%s）。"
     "美瑛町の町道除雪率は52.9％であり、3町はいずれも豪雪地帯又は特別豪雪地帯である。"
     % g("taisetsu_q2_1_s12"),
     "除雪は介護保険の給付対象外であり、生活支援体制整備事業及び3町の施策で対応する。"
     "第5章 基本目標2の主な事業に位置づける。", "第5章 基本目標2", "08"),
    ("困りごとを解決できない層が存在する",
     "生活動作の困りごとがある者のうち「解決できず、困っている」と回答した者は%s、"
     "回答者全体に対する割合は%s である。"
     % (g("解決できず困っている", mneed), g("_unsolved_all")),
     "代表KPI H06（在宅生活継続困難割合）の基準値として設定できる。"
     "在宅生活改善調査（第10期分）の受領後に突合する。", "第4章第3節\n第5章 基本目標2", "08・11"),
    ("供給・アクセスを理由とする転居意向が把握できる",
     "転居を考える理由として、通院・買い物の困難、希望する介護施設に入所できない、"
     "希望する生活支援サービスが受けられないのいずれかを挙げた者は、"
     "住み続ける意向の設問に回答した者に対して%s である。"
     % g("_supply_all"),
     "代表KPI H11（供給不足を理由とする住替え）の基準値として設定できる。"
     "居所変更実態調査（第10期分）の実績と突合する。", "第4章第3節\n第6章第2節", "09・11"),
    ("介護・介助が必要だが受けていない層がいる",
     "「何らかの介護・介助が必要だが、現在は受けていない」と回答した者は%s である。"
     % g("_need_nouse"),
     "認定を受けながらサービスを利用していない504人（認定者の25.7％）の論点に対応する"
     "指標として、新規に設定することを提案する。", "第2章第2節3\n第4章第3節", "11"),
    ("世話をしてくれる人がいない層がいる",
     "日常的に支障が生じた場合に世話をしてくれる人が「いない」と回答した者は%s であり、"
     "独居者では%s と高い。"
     % (g("世話をする人なし"), g("世話をする人なし", mal)),
     "身寄りのない高齢者の入居・入所、保証、死後事務等の支援体制の"
     "対象規模を示す指標として用いる。", "第5章 基本目標3", "06・10"),
    ("在宅での介護を希望する者が多数",
     "介護が必要になった場合に自宅での介護を希望する者は%s である。" % g("_home"),
     "在宅サービスの見込量の前提となる。"
     "一方、24時間対応の3サービスが区域内に存在しないことと合わせて論じる。",
     "第6章第2節・第4節", "10"),
    ("社会参加とリスク指標に一貫した関連",
     "社会参加あり群はフレイル・閉じこもり・うつ・IADL低下の割合がいずれも低く、"
     "年齢階級で層化しても同じ方向の差がみられる。",
     "介護予防・社会参加の施策（基本目標1）の根拠として用いる。"
     "横断調査であり因果関係を示すものではないことを注記する。", "第5章 基本目標1", "05"),
    ("経済状況による差が最も大きい",
     "フレイル該当割合は暮らし向きが「大変苦しい」層で42.7％、"
     "「大変ゆとりがある」層で6.8％と6.3倍の差がある。"
     "うつ割合（49.0％対12.6％）、生活動作の困りごと（60.9％対20.0％）も同様であり、"
     "本分析で確認した属性のうち経済状況による差が最も大きい。",
     "低所得層への配慮を保険料段階の設定、相談支援及び介護予防の"
     "対象者の把握に反映する。", "第2章第2節3\n第5章 基本目標1\n第6章第6節", "07"),
    ("独居者の社会参加は高い",
     "独居者の社会参加あり割合は41.5％で、その他の世帯30.5％より11.0ポイント高く、"
     "通いの場参加割合も14.8％対7.1％と高い。"
     "一方、独居者は孤食16.9％、生活動作の困りごと57.7％、"
     "世話をしてくれる人がいない32.4％といずれも高い。",
     "独居であることが直ちに社会的孤立を意味しないため、"
     "施策の対象は世帯類型ではなく、困りごとの有無及び支援者の有無で捉える。"
     "通いの場は独居者の参加が多く、既存の資源が機能している。",
     "第5章 基本目標1・2", "06"),
    ("自力移動の手段がない層の分布",
     "自分での運転・自転車・バイクのいずれも利用していない者の割合を町別・地区別に算定した。"
     "路線バス及びコミュニティバスの利用割合は低い。",
     "訪問・通所困難地域の判定において、事業所の通常の事業の実施地域と併せて用いる。",
     "第1章第7節", "03"),
    ("人生会議（ACP）の認知度が低い",
     "人生会議（ACP）を「知らない」と回答した者が多数を占める。",
     "意思決定支援の普及啓発を基本目標5の主な事業に位置づける。", "第5章 基本目標5", "10"),
    ("本調査は要支援者・要介護者を含まない",
     "調査対象者は一般高齢者及び総合事業対象者であり、"
     "要支援者・要介護者は対象から外れている（調査報告書「４（５）各保険者の調査対象者」）。"
     "対象者7,121人と65歳以上人口9,191人の差、"
     "及び個票データの要介護度欄がすべて空欄であることとも整合する。",
     "H06・H11は在宅生活改善調査・居所変更実態調査の代替ではなく補完として位置づける。"
     "本調査の値と第10期分の各調査の値は母集団が異なるため、併記して比較する。",
     "第4章第3節\n第2章第4節", "13・14"),
    ("施設入所志向は支え手の有無と強く結びついている",
     "施設入所を希望する層と自宅での介護を希望する層とでは、"
     "生活動作の困りごと（42.2％対41.2％）、自力移動の手段なし（20.9％対20.4％）、"
     "暮らし向きが苦しい（26.0％対27.3％）、フレイル（19.4％対17.8％）に有意な差がない。"
     "有意な差があるのは独居（21.1％対13.6％）と"
     "世話をしてくれる人がいない（10.2％対5.4％）である。",
     "施設入所志向は、在宅の支え手の有無と強く結びついている。"
     "心身の状態や経済状況については有意な差を確認できないが、"
     "差がないことは関係がないことを意味しないため、"
     "17シートに差の大きさと信頼区間を併記する。"
     "施設整備の判断において、在宅の支援体制の強化が需要を変えうることを"
     "第6章第5節の整備方針に明記する。", "第6章第4・5節", "17"),
    ("認知症の相談窓口を知らない者が3分の2",
     "認知症に関する相談窓口を「知らない」と回答した者は64.3％である。"
     "本人又は家族に認知症の症状がある者に限っても47.3％が知らない。",
     "見える化に認知症指標のデータ登録がないため、本設問がH10の補完指標となる。"
     "相談窓口の周知を基本目標2の主な事業に位置づける。", "第5章 基本目標2\n第4章第3節", "16"),
    ("介護・介助が必要な層が公的サービスにつながっていない",
     "介護・介助が必要と回答した404人のうち、困りごとの解決手段として"
     "「介護保険などのサービス」を挙げた者は7.8％にとどまり、"
     "「自力で何とかしている」が49.3％、「家族や親族の手助け」が58.9％である。"
     "この404人は要支援・要介護認定を受けていない層である。",
     "認定申請の手前にある層への相談支援と総合事業の接続を"
     "基本目標1・2の主な事業に位置づける。"
     "未利用認定者504人とは別の集団であり、合算しない。", "第2章第2節3\n第5章 基本目標1・2", "15"),
    ("通いの場は既にハイリスク層に届いている",
     "通いの場の参加者は、独居30.3％（不参加16.7％）、"
     "生活動作の困りごとあり54.6％（同38.7％）と、支援の必要度が高い層の割合が高い。"
     "自分で運転する者の割合は58.1％（同72.9％）と低く、"
     "移動手段が限られる層も参加できている。",
     "通いの場の量的拡大だけでなく、既存の場が届いている層と届いていない層を分けて"
     "施策を組み立てる。総合事業ベースの参加率が全国の3分の1である"
     "（所見12）こととの整合を第3章で説明する。", "第5章 基本目標1", "18"),
    ("同規模保険者との比較で外出と社会参加が最大の課題",
     "同規模保険者40（人口5万未満・集計数82,106票）と比べ、"
     "友人知人と会う頻度が高い者の割合が8.2ポイント低く（63.0％対71.2％）、"
     "通いの場参加者割合が4.4ポイント低い（8.8％対13.2％、同規模の3分の2）。"
     "閉じこもり者割合は1.4倍（8.5％対6.0％）である。"
     "20指標の判定は良好5・同等7・課題7。",
     "外出と社会参加を基本目標1・2の重点領域とする。"
     "通いの場については、総合事業ベースで全国の3分の1（所見12）という"
     "別系列の結果とも一致しており、第3章の評価に反映する。",
     "第2章第4節\n第5章 基本目標1・2", "19・21"),
    ("転倒と口腔機能低下が同規模を明確に上回る",
     "1年間の転倒あり割合35.5％（同規模30.0％）、"
     "口腔機能低下者割合24.7％（同規模21.8％）。"
     "転倒は65〜69歳で同規模を9.6ポイント上回る。"
     "美瑛町は口腔機能低下・閉じこもりとも参加74市町村中73位、"
     "東川町は両指標とも70位である。",
     "転倒予防と口腔機能の維持を基本目標1の主な事業に明示する。"
     "転倒が若い年齢層で顕著なことから、冬期の路面条件との関連を"
     "第2章の地域特性の記述に接続する。", "第5章 基本目標1\n第2章第4節", "19・20"),
    ("他団体比較には調査対象者の範囲による偏りがある",
     "同規模保険者40のうち31（77.5％）は要支援者を調査対象に含むが、"
     "当広域連合は含まない（一般高齢者＋総合事業対象者）。"
     "要支援者を含む集団はリスク指標が高く出るため、当広域連合の値は良好側へ偏る。"
     "運動機能低下（▲3.2ポイント）、IADL低下（▲2.6ポイント）、"
     "85歳以上のフレイル（▲3.6ポイント）は、この偏りで説明できる可能性がある。",
     "「良好」と判定した指標は断定せず、対象者範囲の違いを注記する。"
     "「課題」と判定した指標は、偏りが良好側に働いてもなお下回るため確度が高い。"
     "計画本文でも同じ整理で記述する。", "第2章第4節", "19"),
    ("健康指標は悪いが地域への連帯感は高い",
     "ソーシャルキャピタル得点の連帯感は美瑛町163.3（参加74市町村中14位）、"
     "東川町162.7（15位）と上位にある。"
     "幸福感がある者の割合も東川町14位、東神楽町27位である。"
     "一方、同じ美瑛町・東川町は口腔機能低下・閉じこもりで70位以下である。",
     "地域への信頼と連帯感の高さを、通いの場の担い手確保及び"
     "生活支援体制整備の資源として位置づける。"
     "健康指標の課題と地域資源の強みを分けて記述する。",
     "第5章 基本目標1・2\n第2章第4節", "20"),
    ("仕様書が求める類似保険者比較の一部を実施できた",
     "仕様書４（4）は全国平均、北海道平均及び類似保険者との比較分析を求めているが、"
     "見える化システムの類似保険者比較機能は比較対象の選定条件が未決定であった。"
     "健康とくらしの調査の同規模保険者比較により、"
     "住民の心身の状態・社会参加の領域については比較分析ができた。",
     "見える化システムによる給付・供給面の類似保険者比較は引き続き必要である"
     "（確認事項No.3）。本分析で対応できた範囲を明示する。",
     "第2章第2・4節", "19・21"),
    ("地区別の重点度が独立した2つの方法で一致する",
     "JAGESの重点対象地域選定シート（70指標のランク評価）による健康コア11指標の"
     "ランク平均は、明徳小学校区4.27、志比内小学校区4.27、美瑛小学校区4.00、"
     "美瑛東小学校区3.82が高く、本分析の地区別集計と同じ方向を示す。"
     "美瑛東小学校区は社会参加コア15指標のランク平均も4.07と高い。",
     "独立に算定した2つの方法が一致する地区は、重点化の根拠として確度が高い。"
     "日常生活圏域の設定及び介護予防の重点地区の選定に用いる。",
     "第1章第7節\n第5章 基本目標1", "18"),
]
for i, (t, ev, act, dst, sh) in enumerate(FIND, start=1):
    r = body(ws, r, [i, t, ev, act, dst, sh], height=64, align={1: "center", 6: "center"})

note(ws, r + 1,
     "注1）本シートの所見は、健康とくらしの調査（令和7年度）の個票データの分析によるものである。"
     "他の3調査（在宅生活改善調査、居所変更実態調査、介護人材実態調査）の第10期分は未受領であり、"
     "受領後に本分析の結果と突合して整合を確認する。"
     "注2）所見2〜12は計画素案（第10稿）への反映を要する。"
     "反映は図表レイアウトの確認結果と併せて次稿で行う。"
     "注3）業務工程管理表の01シート（3）実施済み調査結果の集計・分析について、"
     "本分析の完了により作業項目①⑤⑥が完了する。", 6)



# ============================================================ 13 妥当性レビュー
ws = sheet("13_課題と施策への対応レビュー", "第9期の課題・第10期の施策と本分析の対応関係",
           "妥当性検証報告書の主要所見20件（第9期計画の評価から得られた課題）及び"
           "第10期の5基本目標・施策体系に対して、本分析がどの程度の根拠を提供できているかを"
           "レビューする。「不足」と判定した項目については、"
           "15〜18シートに追加のクロス集計を行った。",
           [5, 30, 26, 30, 12, 30, 10])

r = header(ws, 4, ["所見", "第9期の評価から得られた課題", "対応する第10期の施策",
                   "本分析による根拠の提供状況", "判定", "追加した分析又は残る限界", "シート"],
           height=44)

REVIEW = [
    (1, "給付水準の高さは施設・居住系サービスの受給率と受給者単価がともに高いことによる"
     "（在宅は全国比1.02、施設・居住系は1.08）",
     "基本目標3 持続可能なサービス提供・住まい／第6章の整備方針",
     "本調査は給付実績を扱わないため直接の根拠にならない。"
     "ただし施設入所志向（31.0％）と在宅介護希望（49.2％）の分布は需要側の根拠となる。",
     "一部不足", "施設志向を支援者の有無・移動制約・世帯構成と交差させる分析を追加した。"
     "給付側の分析は見える化データによる（素案第2章第2節）。", "17"),
    (2, "介護サービス利用率が7年間で7.9ポイント低下し、"
     "未利用認定者が504人（認定者の25.7％）に達している",
     "基本目標3／H12（必要サービス未充足率）／第2章第2節3",
     "「介護・介助が必要だが受けていない」7.0％を把握した。"
     "ただし本調査の対象は要支援・要介護認定を受けていない一般高齢者及び総合事業対象者であり、"
     "未利用認定者504人とは母集団が異なる。",
     "限界あり", "認定を受けていない層の潜在ニーズとして別に位置づける分析を追加した。"
     "未利用認定者の要因分解は保険者の統計によるほかない。", "14・15"),
    (3, "重度者の在宅化が進む一方（要介護5の在宅・居住系割合31.3％→54.6％）、"
     "施設定員が234人→160人へ減少している",
     "基本目標3／第6章第4節 整備方針／H07・H08（家族介護者支援）",
     "第10稿では家族介護力の根拠がなかった。"
     "本分析で世話をしてくれる人がいない者8.7％（独居者32.4％）を把握した。",
     "対応済＋追加", "介護・介助が必要な層（404人）に限定した世帯構成・支援者・"
     "困りごとの分析を追加した。", "10・15"),
    (5, "3町の高齢者福祉計画に数値目標が1件も設定されていない",
     "資料2 3町との役割分担・共通指標",
     "本分析は小学校区12地区別の集計を可能にした。"
     "町別・地区別の共通指標を3町の計画に掲載する際の基礎データとなる。",
     "対応済", "小学校区と日常生活圏域の対応関係の確認が残る（確認事項No.12）。", "01・02"),
    (6, "代表KPI③フレイルの目標設定が2時点の外挿に依存していた",
     "H04（フレイル該当割合）の目標設定",
     "基準値19.1％に95％信頼区間（17.9〜20.2％）を付し、"
     "令和4年18.5％との差が統計的に有意でないことを確認した。",
     "対応済", "―", "00・14"),
    (7, "認定率は全体では上昇したが、年齢層をそろえると低下している",
     "H01（年齢調整済認定率）／第3章第4節の4軸評価",
     "本分析でも直接法による年齢調整を全指標に適用し、"
     "町間の差が年齢構成によるものかを判別できるようにした。",
     "対応済", "―", "02"),
    (8, "代表KPI④（通いの場参加率）は国の目標水準を達成し、自主目標のみ僅差で未達",
     "H05（社会参加率）／基本目標1",
     "通いの場8.8％と社会参加6区分の合成36.0％の双方を算定し、"
     "年齢階級別・町別・地区別の内訳を示した。",
     "対応済", "―", "05"),
    (12, "総合事業ベースでみると通いの場の水準は全国の3分の1で、低下傾向にある",
     "基本目標1／施策2-2 総合事業の推進",
     "本調査（ニーズ調査ベース）は8.8％で、総合事業ベースの1.6％と大きく異なる。"
     "参加していない層の特性を分析していなかった。",
     "不足", "通いの場に参加していない層を移動手段・経済状況・地区・"
     "困りごとと交差させる分析を追加した。", "18"),
    (13, "通いの場の担い手と内容に圏域固有の特徴がある（3箇所のみ）",
     "基本目標1／施策2-2",
     "本調査は参加者側の情報であり、箇所数・運営主体・活動内容は把握できない。"
     "参加者の特性（独居30.3％、困りごとあり54.6％）と、"
     "同規模保険者との差（8.8％対13.2％）は把握した。",
     "一部不足", "通いの場の箇所数・運営主体・活動内容は3町からの情報による。"
     "参加率の他団体比較を19シートに追加した。", "18・19"),
    ("―", "仕様書４（4）　類似保険者との比較分析（未実施）",
     "第2章第2節／計画全体",
     "見える化システムの類似保険者比較機能は比較対象の選定条件が未決定である。"
     "健康とくらしの調査の同規模保険者40（集計数82,106票）との比較により、"
     "住民の心身の状態と社会参加の領域は比較分析ができた。",
     "一部不足", "同規模保険者との比較（19シート）、参加74市町村の順位（20シート）、"
     "他団体比較の総括（21シート）を追加した。"
     "給付・供給面の類似保険者比較は見える化システムによる（確認事項No.3）。",
     "19・20・21"),
    (16, "日常生活圏域の設定に人口規模の著しい偏りがある"
     "（朗根内233人と市街地・周辺7,869人で33.8倍）",
     "第1章第7節 圏域の設定／統計単位と整備単位の分離",
     "小学校区12地区別の集計により、統計・分析単位としての地区別分析が可能となった。"
     "JAGESの重点対象地域選定シートによる70指標のランク評価とも突合した。",
     "対応済", "小学校区と第9期の6圏域の対応関係の確認が残る（確認事項No.12）。", "01・18"),
    (17, "24時間対応の3サービスは制度創設以来13年間、域内に1事業所も存在しない",
     "第6章第4節 確保方策／基本目標3",
     "第10稿では夜間・深夜のニーズを示す根拠がなかった。",
     "不足", "介護・介助が必要な者のうち独居又は支援者がいない層の規模を算定した。"
     "本調査に認定者が含まれないため、これは重度者ではなく"
     "「認定前の段階で支援者がいない層」の規模である点に留意する。", "15"),
    (19, "成果は全国トップ級だが、活動の記録・報告が全国の23.2％にとどまる",
     "第5章の5層構成／アウトプット指標の設定",
     "本分析はアウトカム側の根拠を提供するが、活動（アウトプット）の記録は対象外。",
     "対象外", "アウトプット指標は事業実績から設定する（確認事項No.5）。", "―"),
    (20, "性・年齢調整済みの要介護2以上認定率は全国を上回り、しかも上昇している",
     "H01のデータ源をW144に確定",
     "本調査は認定率を扱わない。",
     "対象外", "―", "―"),
    ("―", "施策2-3 認知症施策推進大綱を踏まえた施策の推進（指標不足）",
     "基本目標2／H10（認知症相談から支援接続までの適時率）",
     "第10稿では認知症の指標が見える化のデータ登録がないため空白であった。"
     "本調査には本人又は家族の認知症症状の有無、"
     "認知症相談窓口の認知度の設問が全数で含まれている。",
     "不足", "認知症の状況と相談窓口の認知度の分析を追加した。"
     "相談窓口を知らない層の分布はH10の補完指標となる。", "16"),
    ("―", "施策1-2 在宅療養者支援に向けた医療機関との連携体制の構築（指標不足）",
     "基本目標1／H09（退院時支援調整実施率）",
     "かかりつけ医の有無はセクションBの調査票のみ（556票）で、"
     "町別・地区別の分析に耐えない。",
     "限界あり", "医療側の情報は上川中部圏域の医療機関又は北海道の資料による。", "―"),
    ("―", "施策4-4 人材確保・育成の推進（最重点化）",
     "基本目標4／H13・H14",
     "本調査は住民調査であり、事業所・従事者の情報を含まない。",
     "対象外", "介護人材実態調査（第10期分）による（確認事項No.8・10）。", "―"),
    ("―", "施策5-1 災害や感染症対策に係る体制整備（指標不足）",
     "基本目標5／H16",
     "「災害時の避難」を困りごととする者8.1％（独居者13.8％）を把握した。"
     "住民側の備えに関する設問はセクションHのみで分母が小さい。",
     "一部不足", "事業所のBCPは指定台帳及び集団指導により把握する。", "08"),
]
for row in REVIEW:
    j = {"対応済": OK_G, "対応済＋追加": OK_G, "一部不足": IN_Y,
         "不足": NG_O, "限界あり": NG_O, "対象外": GRAY}.get(row[4], GRAY)
    r = body(ws, r, list(row), {5: j}, height=64, align={1: "center", 5: "center", 7: "center"})

note(ws, r + 1,
     "注1）所見番号は妥当性検証報告書 01シートの主要所見20件による。"
     "「―」は施策側から見た不足の指摘である。"
     "注2）判定の意味は次のとおり。対応済＝本分析が計画に必要な根拠を提供している。"
     "対応済＋追加＝提供しているが、レビューを受けて分析を追加した。"
     "一部不足・不足＝レビューを受けて分析を追加した。"
     "限界あり＝本調査の設計上、必要な根拠が得られない部分がある。"
     "対象外＝本調査の対象範囲外であり、他のデータ源による。"
     "注3）本レビューにより、本調査の対象者が「一般高齢者＋総合事業対象者」であり、"
     "要支援者・要介護者を含まないことが判明した（14シート）。"
     "この制約は代表KPIのデータ源の選択に影響するため、11シートの記述を改めた。", 7)

# ============================================================ 14 統計的妥当性
ws = sheet("14_統計的妥当性の検証", "調査の代表性と統計処理の妥当性",
           "本分析の結果を計画の根拠として用いるにあたり、"
           "母集団の範囲、回収率の地域差、標本規模、多重比較の扱いを検証する。"
           "特に、本調査の対象者が要支援者・要介護者を含まないことは、"
           "代表KPIのデータ源の選択に直接影響する。",
           [26, 16, 16, 16, 16, 16, 30])

r = lead(ws, 4, "【1　調査対象者の範囲】", 7)
r = header(ws, r, ["項目", "内容", "", "", "", "", "計画上の意味"], height=28)
for k, v_, mean in [
    ("調査対象者の属性", "一般高齢者＋総合事業対象者",
     "要支援者・要介護者を含まない。"
     "認定者を対象とする指標の代替データ源にはならない"),
    ("対象者数", "7,121人（東神楽町2,319・東川町2,074・美瑛町2,728）",
     "基準日は令和7年10月17日"),
    ("参考：65歳以上人口", "9,191人（令和5年10月1日・住民基本台帳）",
     "対象者数との差2,070人が認定者等に相当すると考えられる"),
    ("参考：認定者数", "1,976人（令和5年10月1日・要介護1,443／要支援533）",
     "認定者はほぼ全数が調査対象から外れている"),
    ("事業対象者", "69人（回答者中）",
     "総合事業対象者は対象に含まれている"),
    ("要介護度の記録", "個票データの要介護度欄はすべて空欄",
     "対象者に認定者が含まれていないことと整合する"),
]:
    r = body(ws, r, [k, v_, "", "", "", "", mean], {2: IN_Y}, height=32)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=6)

r += 1
r = lead(ws, r, "【2　町別の回収率と非回答の偏り】", 7)
r = header(ws, r, ["町", "対象者数", "回収数", "回収率", "集計数", "回答者に占める\n75歳以上の割合",
                   "留意点"], height=32)
RECOV = [("東神楽町", 2319, 1586, mHG), ("東川町", 2074, 1416, mH), ("美瑛町", 2728, 1727, mB)]
for nm, tgt, rec, m in RECOV:
    o75 = round(100 * A.loc[m, "年齢階級"].isin(["75～79歳", "80～84歳", "85歳以上"]).mean(), 1)
    note_t = ("回収率が他の2町より5ポイント低い。回答しにくい層ほど"
              "心身の状態が良くない傾向があるため、美瑛町の指標は実態より良い方向に"
              "偏っている可能性がある" if nm == "美瑛町" else "―")
    r = body(ws, r, [nm, tgt, rec, round(100 * rec / tgt, 1), int(m.sum()), o75, note_t],
             {4: NG_O} if nm == "美瑛町" else None, height=44)
r = body(ws, r, ["計", 7121, 4798, 67.4, N_ALL,
                 round(100 * A["年齢階級"].isin(["75～79歳", "80～84歳", "85歳以上"]).mean(), 1),
                 "ID切り取り69票を集計から除外している"], {1: GRAY}, height=32, bold=True)

r += 1
r = lead(ws, r, "【3　主要指標の95％信頼区間（広域連合全体）】", 7)
r = header(ws, r, ["指標", "推定値", "分母", "95％信頼区間", "区間の幅",
                   "令和4年の値", "第9期からの変化の判定"], height=32)
R4 = {"frail2gp_25": 18.5, "undo2gp_25": 9.7, "tojikomori2gp_25": 6.7,
      "utsucheck2gp_25": 28.9, "ninchicheck2gp_25": 34.5, "koukukinou2gp_25": 23.4,
      "iadl1_2gp_25": 9.8, "tento2gp_25": 32.9, "通いの場": 7.3}
N4 = 4700          # 令和4年調査の有効回答数の推定値


def two_sample(p1, n1, p0, n0=N4):
    """2標本の母比率の差の検定（両側・正規近似）。"""
    p1_, p0_ = p1 / 100.0, p0 / 100.0
    pp = (p1_ * n1 + p0_ * n0) / (n1 + n0)
    se = math.sqrt(pp * (1 - pp) * (1.0 / n1 + 1.0 / n0))
    if se == 0:
        return 0.0
    return abs(p1_ - p0_) / se


for key, nm in IND:
    v, n = rate(ALL, key)
    if v is None:
        continue
    lo, hi = wilson(v / 100, n)
    prev = R4.get(key)
    if prev is None:
        judge, diff = "第9期に対応する値がない", None
    else:
        z = two_sample(v, n, prev)
        diff = round(v - prev, 1)
        judge = "有意（z=%.2f）" % z if z >= 1.96 else "有意でない（z=%.2f）" % z
    r = body(ws, r, [nm, v, n, "%.1f〜%.1f" % (100 * lo, 100 * hi),
                     round(100 * (hi - lo), 1), prev, judge],
             {7: NG_O if judge.startswith("有意（") else None}, height=18)

r += 1
r = lead(ws, r, "【4　地区別分析における標本規模】", 7)
r = header(ws, r, ["地区", "回答者数", "割合10％の場合の\n95％信頼区間の幅",
                   "割合30％の場合の\n95％信頼区間の幅", "判定", "", "留意点"], height=44)
for tname, aname, mask in rows:
    if aname in ("全体", "町計"):
        continue
    n = int(mask.sum())
    w10 = 2 * 1.96 * math.sqrt(0.1 * 0.9 / n) * 100
    w30 = 2 * 1.96 * math.sqrt(0.3 * 0.7 / n) * 100
    jd = "十分" if n >= 300 else ("参考値" if n >= 100 else "単独では用いない")
    fl = {5: OK_G} if jd == "十分" else ({5: IN_Y} if jd == "参考値" else {5: NG_O})
    r = body(ws, r, [aname, n, round(w10, 1), round(w30, 1), jd, "",
                     "±%.1fポイント程度の幅がある" % (w30 / 2)], fl, height=18)

r += 1
r = lead(ws, r, "【5　多重比較と解釈上の注意】", 7)
r = header(ws, r, ["論点", "内容", "", "", "", "", "本分析での扱い"], height=28)
for k, v_, mean in [
    ("多重比較", "01シートは12地区×11指標＝132の値を並べており、"
     "有意水準5％で検定すればおよそ7件が偶然に有意となる",
     "01シートには検定を付さず、地区間の傾向の把握に用いる。"
     "個別の差の判定は02シートの町間検定及び本シートの信頼区間による"),
    ("因果の解釈", "本調査は横断調査であり、社会参加とリスク指標の関連は"
     "「参加が状態を保つ効果」と「状態が良いから参加できる」の両方を含む",
     "05シートに明記し、年齢階級で層化した比較を併記している"),
    ("年齢の交絡", "運転の可否、社会参加、リスク指標はいずれも年齢と強く相関する",
     "02シートで直接法による年齢調整を行い、粗率と併記している"),
    ("小標本", "志比内小学校区21票、明徳小学校区49票、美沢小学校区57票",
     "01・14シートで網掛けとし、単独では判断材料としない旨を注記している"),
    ("分母の変動", "調査票が8種類（A〜H）に分かれ、"
     "版固有の設問の分母は550〜616票にとどまる",
     "すべての表に分母を併記している。"
     "外出の抑制の理由（106票）等は参考値として扱う"),
    ("母集団の制約", "要支援者・要介護者を含まないため、"
     "認定者を対象とする指標の代替にはならない",
     "11シートのH06・H11の位置づけを「代替」から「補完」に改めた"),
]:
    r = body(ws, r, [k, v_, "", "", "", "", mean], height=44)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=6)

note(ws, r + 1,
     "注1）信頼区間はWilson法（両側95％）による。区間の幅は上限と下限の差である。"
     "注2）令和4年の値は第9期計画 第2章第3節の掲載値である。"
     "指標の定義（基本チェックリストの該当項目数等）は両年で一致する。"
     "注2-2）第9期からの変化の判定は2標本の母比率の差の検定（両側・正規近似・"
     "有意水準5％）による。令和4年調査の有効回答数は受領していないため4,700と仮定した。"
     "4,000から5,000の範囲ではいずれの指標も判定が変わらないことを確認している。"
     "本判定は調査報告書及び図表集08シートの統計的判定と13指標すべてで一致する。"
     "なお、令和7年の信頼区間に令和4年の値が含まれるかどうかで判定すると、"
     "令和4年側の標本誤差を無視するため有意と判定されやすくなる（口腔機能低下者割合が該当）。"
     "本シートは2標本の検定によっている。"
     "注3）美瑛町の回収率63.3％は他の2町より約5ポイント低い。"
     "一般に、心身の状態が良くない層ほど回答しにくいため、"
     "美瑛町の指標は実態より良い方向に偏っている可能性がある。"
     "美瑛町の値が3町で最も悪い指標が多いことは、この偏りを考慮してもなお差があることを示す。"
     "注4）調査対象者の属性は調査報告書「４（５）各保険者の調査対象者」による。", 7)


# ============================================================ 15 介護・介助が必要な層
ws = sheet("15_介護・介助が必要な層", "介護・介助が必要な者の状況（認定を受けていない層）",
           "【問1】2）で「何らかの介護・介助が必要」と回答した404人（回答者の8.7％）を分析する。"
           "本調査の対象者は一般高齢者及び総合事業対象者であり、要支援者・要介護者を含まない。"
           "したがってこの404人は、認定を受けていないが介護・介助を必要としている層であり、"
           "未利用認定者504人とは別の集団である。"
           "重度在宅化と施設定員の縮小が同時に進むなかで、"
           "認定の手前にある層の規模と支援状況を示す。",
           [30, 11, 11, 11, 11, 11, 12, 12, 12])

mneed_all = flag_in("adl3ra25", [2, 3]) == 1
mneed_no = flag_in("adl3ra25", [2]) == 1        # 必要だが受けていない
mneed_yes = flag_in("adl3ra25", [3]) == 1       # 必要で受けている
mok = flag_in("adl3ra25", [1]) == 1

r = lead(ws, 4, "【介護・介助の要否（回答者全体）】", 9)
r = header(ws, r, ["区分", "人数", "割合", "75歳以上の\n割合", "85歳以上の\n割合",
                   "東川町", "美瑛町", "東神楽町", "分母"], height=36)
for lb, m, vname in [("介護・介助は必要ない", mok, "介助不要"),
                     ("何らかの介護・介助が必要だが、現在は受けていない", mneed_no, "要介助・未受給"),
                     ("介護・介助が必要で、介護・介助を受けている", mneed_yes, "要介助・受給中")]:
    v, n = rate(ALL, vname)
    fill = {1: NG_O} if vname == "要介助・未受給" else None
    r = body(ws, r, [lb, int(m.sum()), v, rate(m75, vname)[0], rate(m85, vname)[0],
                     rate(mH, vname)[0], rate(mB, vname)[0], rate(mHG, vname)[0], n],
             fill, height=30)
A["_need_any"] = flag_in("adl3ra25", [2, 3])
r = body(ws, r, ["【再掲】何らかの介護・介助が必要（計）", int(mneed_all.sum()),
                 rate(ALL, "_need_any")[0], rate(m75, "_need_any")[0],
                 rate(m85, "_need_any")[0], rate(mH, "_need_any")[0],
                 rate(mB, "_need_any")[0], rate(mHG, "_need_any")[0],
                 rate(ALL, "_need_any")[1]], {1: OK_G}, height=18, bold=True)

r += 1
r = lead(ws, r, "【介護・介助が必要な者の状況（404人が分母）】", 9)
r = header(ws, r, ["指標", "介護・介助が\n必要（計）", "うち受けて\nいない", "うち受けて\nいる",
                   "介護・介助は\n必要ない", "差の検定\n（必要／不要）", "必要（計）\n分母",
                   "受けていない\n分母", ""], height=44)
HH = [("独居", "独居者割合"), ("世話をする人なし", "世話をしてくれる人がいない割合"),
      ("困りごとあり", "生活動作の困りごとあり割合"),
      ("解決できず困っている", "解決できず困っている割合"),
      ("自力移動なし", "自力移動の手段なし割合"), ("外出低頻度", "外出頻度が月1〜3回以下"),
      ("暮らし苦しい", "暮らし向きが苦しい割合"), ("社会参加", "社会参加あり割合"),
      ("通いの場", "通いの場参加割合"), ("転居も考える", "転居も少し考えている割合"),
      ("frail2gp_25", "フレイルあり割合"), ("iadl1_2gp_25", "IADL低下者割合"),
      ("ninchicheck2gp_25", "認知機能低下者割合"), ("utsucheck2gp_25", "うつ割合")]
for key, nm in HH:
    v1, n1 = rate(mneed_all, key)
    v2, n2 = rate(mneed_no, key)
    v3, _ = rate(mneed_yes, key)
    v0, _ = rate(mok, key)
    if v1 is None:
        continue
    r = body(ws, r, [nm, v1, v2, v3, v0, ztest(mneed_all, mok, key), n1, n2, ""],
             height=18)

r += 1
r = lead(ws, r, "【支援の状況の組合せ（404人が分母）】", 9)
r = header(ws, r, ["区分", "人数", "割合", "計画上の意味", "", "", "", "", ""], height=28)
A["_alone_need"] = ((series("独居") == 1) & mneed_all).astype(float).where(mneed_all)
A["_nohelp_need"] = ((series("世話をする人なし") == 1) & mneed_all).astype(float).where(mneed_all)
A["_alone_or_nohelp"] = (((series("独居") == 1) | (series("世話をする人なし") == 1))
                         & mneed_all).astype(float).where(mneed_all)
A["_nodrive_need"] = ((series("自力移動なし") == 1) & mneed_all).astype(float).where(mneed_all)
for key, nm, mean in [
    ("_alone_need", "独居である", "夜間・深夜の支援の担い手がいない"),
    ("_nohelp_need", "世話をしてくれる人がいない", "在宅生活の継続が困難になりやすい"),
    ("_alone_or_nohelp", "独居又は世話をしてくれる人がいない",
     "24時間対応サービスの必要性を検討する対象規模"),
    ("_nodrive_need", "自力移動の手段がない", "通所系サービスの送迎又は訪問系への依存"),
]:
    v, n = rate(mneed_all, key)
    cnt = int(series(key).sum()) if n else 0
    fill = {1: NG_O} if key == "_alone_or_nohelp" else None
    r = body(ws, r, [nm, cnt, v, mean, "", "", "", "", ""], fill, height=18)

r += 1
r = lead(ws, r, "【介護・介助が必要な者の困りごとと解決手段（上位）】", 9)
r = header(ws, r, ["項目", "介護・介助が\n必要（計）", "介護・介助は\n必要ない", "差",
                   "検定", "分母", "", "", ""], height=36)
for key, nm in [("taisetsu_q2_1_s12", "困りごと：除雪"), ("taisetsu_q2_1_s7", "困りごと：買い物"),
                ("taisetsu_q2_1_s8", "困りごと：通院"), ("taisetsu_q2_1_s11", "困りごと：外出"),
                ("taisetsu_q2_1_s1", "困りごと：食事の準備や片付け"),
                ("taisetsu_q2_1_s2", "困りごと：部屋の掃除や片付け"),
                ("taisetsu_q2_1_s13", "困りごと：災害時の避難"),
                ("taisetsu_q2_1_s18", "困りごと：話し相手がいない"),
                ("taisetsu_q2_2_s8", "解決：自力で何とかしている"),
                ("taisetsu_q2_2_s1", "解決：家族や親族の手助け"),
                ("taisetsu_q2_2_s6", "解決：介護保険などのサービス"),
                ("taisetsu_q2_2_s5", "解決：シルバー人材センター等")]:
    if key not in A.columns:
        continue
    A["_t"] = pd.to_numeric(A[key], errors="coerce")
    v1, n1 = rate(mneed_all, "_t")
    v0, _ = rate(mok, "_t")
    if v1 is None or v0 is None:
        continue
    r = body(ws, r, [nm, v1, v0, round(v1 - v0, 1), ztest(mneed_all, mok, "_t"), n1,
                     "", "", ""], height=18)

r += 1
r = lead(ws, r, "【介護を受けたい場所（介護・介助の要否別）】", 9)
r = header(ws, r, ["区分", "介護・介助が\n必要（計）", "介護・介助は\n必要ない", "独居かつ\n必要",
                   "分母", "", "", "", ""], height=36)
for nm, vs, _mean in Q42:
    A["_t"] = flag_in("taisetsu_q4_2", vs)
    r = body(ws, r, [nm, rate(mneed_all, "_t")[0], rate(mok, "_t")[0],
                     rate(mneed_all & (series("独居") == 1), "_t")[0],
                     rate(mneed_all, "_t")[1], "", "", "", ""], height=30)

note(ws, r + 1,
     "注1）単位は％（人数欄のみ人）。"
     "注2）本調査の対象者は一般高齢者及び総合事業対象者であり、"
     "要支援者・要介護者を含まない（14シート）。"
     "このため本シートの404人は、認定を受けていないが介護・介助を必要としている層である。"
     "介護サービス利用率の低下（平成28年度82.2％→令和6年度74.3％）及び"
     "未利用認定者504人とは別の集団であり、両者を合算してはならない。"
     "注3）「介護・介助が必要で、介護・介助を受けている」81人には、"
     "総合事業対象者及び家族等による介助を受けている者が含まれると考えられる。"
     "注4）「独居又は世話をしてくれる人がいない」層の規模は、"
     "24時間対応サービス（定期巡回・随時対応型訪問介護看護、夜間対応型訪問介護、"
     "看護小規模多機能型居宅介護）の確保方策を検討する際の需要側の根拠となる"
     "（素案 第6章第4節）。ただし本調査には認定者が含まれないため、"
     "重度者の夜間・深夜のニーズそのものではなく、"
     "認定の手前の段階で支援者を欠く層の規模である点に留意する。", 9)

# ============================================================ 16 認知症
ws = sheet("16_認知症の状況と相談窓口", "認知症の状況と相談窓口の認知度",
           "本人又は家族に認知症の症状がある者の割合と、認知症に関する相談窓口の認知度を"
           "集計する。いずれも全数の共通設問である。"
           "見える化システムの認知症関連指標（J系列）は当広域連合のデータ登録がなく"
           "取得できないため、本調査が認知症施策の数量的根拠となる。",
           [30, 11, 11, 11, 11, 11, 12, 12, 12])

A["_dem_fa"] = flag_in("dem2fa25", [1])          # 本人又は家族に認知症の症状
A["_dem_co"] = flag_in("dem2co25", [1])          # 相談窓口を知っている
A["_dem_nc"] = flag_in("dem2co25", [2])          # 相談窓口を知らない
A["_dem_tx"] = pd.to_numeric(A["dgns2dt25"], errors="coerce")   # 認知症で治療中

r = lead(ws, 4, "【認知症の状況（全数）】", 9)
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "75歳以上", "85歳以上"], height=32)
for key, nm in [("_dem_fa", "本人又は家族に認知症の症状がある"),
                ("_dem_tx", "認知症（アルツハイマー病等）を治療中又は後遺症がある"),
                ("ninchicheck2gp_25", "認知機能低下者割合（基本チェックリスト）"),
                ("ninchiscore2gp_25", "認知症リスク者割合（7/15点以上）"),
                ("wasure2gp_25", "物忘れが多い者の割合")]:
    if key not in A.columns:
        continue
    r = body(ws, r, [nm, rate(ALL, key)[0], rate(mH, key)[0], rate(mB, key)[0],
                     rate(mHG, key)[0], rate(ALL, key)[1], rate(mal, key)[0],
                     rate(m75, key)[0], rate(m85, key)[0]], height=18)

r += 1
r = lead(ws, r, "【認知症に関する相談窓口の認知度（全数）】", 9)
r = header(ws, r, ["区分", "広域連合", "東川町", "美瑛町", "東神楽町", "分母",
                   "独居者", "75歳以上", "症状がある者"], height=32)
mdem = series("_dem_fa") == 1
for key, nm in [("_dem_co", "相談窓口を知っている"), ("_dem_nc", "相談窓口を知らない")]:
    fill = {1: NG_O} if key == "_dem_nc" else None
    r = body(ws, r, [nm, rate(ALL, key)[0], rate(mH, key)[0], rate(mB, key)[0],
                     rate(mHG, key)[0], rate(ALL, key)[1], rate(mal, key)[0],
                     rate(m75, key)[0], rate(mdem, key)[0]], fill, height=18)
r = body(ws, r, ["差の検定（症状がある者／ない者）", "", "", "", "", "", "", "",
                 ztest(mdem, series("_dem_fa") == 0, "_dem_co")], {1: IN_Y}, height=18)

r += 1
r = lead(ws, r, "【地区別の認知症の状況と相談窓口の認知度】", 9)
r = header(ws, r, ["町", "地区", "回答者数", "本人又は家族に\n症状がある",
                   "相談窓口を\n知っている", "認知機能\n低下者割合",
                   "認知症リスク者\n割合", "", ""], height=44)
for tname, aname, mask in rows:
    r = body(ws, r, [tname, aname, int(mask.sum()), rate(mask, "_dem_fa")[0],
                     rate(mask, "_dem_co")[0], rate(mask, "ninchicheck2gp_25")[0],
                     rate(mask, "ninchiscore2gp_25")[0], "", ""], height=18,
             bold=(aname in ("全体", "町計")))

r += 1
r = lead(ws, r, "【症状がある者の状況（本人又は家族に認知症の症状がある者が分母）】", 9)
r = header(ws, r, ["指標", "症状がある", "症状がない", "差", "検定", "分母", "", "", ""],
           height=28)
mnodem = series("_dem_fa") == 0
for key, nm in [("_dem_co", "相談窓口を知っている"), ("困りごとあり", "生活動作の困りごとあり"),
                ("解決できず困っている", "解決できず困っている"),
                ("世話をする人なし", "世話をしてくれる人がいない"),
                ("_need_any", "何らかの介護・介助が必要"), ("独居", "独居"),
                ("utsucheck2gp_25", "うつ割合"), ("社会参加", "社会参加あり")]:
    v1, n1 = rate(mdem, key)
    v0, _ = rate(mnodem, key)
    if v1 is None or v0 is None:
        continue
    r = body(ws, r, [nm, v1, v0, round(v1 - v0, 1), ztest(mdem, mnodem, key), n1,
                     "", "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％。"
     "注2）見える化システムの認知症関連指標（認知症初期集中支援チームの訪問実績、"
     "認知症地域支援推進員の配置人数、認知症カフェの箇所数等）は"
     "当広域連合のデータ登録がなく取得できない（確認事項No.26）。"
     "本調査の2つの設問は全数で回答があり、認知症施策の数量的根拠として使える。"
     "注3）代表KPI H10（認知症相談から支援接続までの適時率）は"
     "地域包括支援センターの業務記録をデータ源とするが、記録様式が統一されていない可能性が高い。"
     "「認知症に関する相談窓口を知っている者の割合」は、"
     "H10が実測できるようになるまでの補完指標として設定できる。"
     "注4）本人又は家族に症状がある者と、それ以外の者とで"
     "相談窓口の認知度に差があるかどうかは、普及啓発の対象設定に用いる。", 9)


# ============================================================ 17 施設志向と在宅資源
ws = sheet("17_施設志向と在宅資源", "施設入所志向と在宅の支援資源の関係",
           "第9期の評価では、給付水準の高さが施設・居住系サービスの受給率と受給者単価がともに高いことによるものであること"
           "（全国比1.08）、重度者の在宅化が進む一方で介護老人福祉施設の定員が"
           "234人から160人へ減少していることが課題として示された。"
           "本シートは、施設入所を希望する層がどのような在宅資源の状況にあるかを分析し、"
           "第6章の整備方針の需要側の根拠とする。",
           [30, 11, 11, 11, 11, 11, 12, 12, 12])

A["_want_fac"] = flag_in("taisetsu_q4_2", [3])          # 老人ホーム等に入所したい
A["_want_home"] = flag_in("taisetsu_q4_2", [1, 2])      # 自宅での介護を希望
A["_want_unk"] = flag_in("taisetsu_q4_2", [5])          # わからない
A["_end_fac"] = flag_in("taisetsu_q4_6", [2, 3])        # 終のすみか＝高齢者住宅・老人ホーム
A["_end_home"] = flag_in("taisetsu_q4_6", [1, 6])       # 終のすみか＝自宅

r = lead(ws, 4, "【属性別の施設入所志向】", 9)
r = header(ws, r, ["区分", "回答者数", "施設に入所\nしたい", "自宅での介護\nを希望",
                   "わからない", "終のすみかが\n施設", "終のすみかが\n自宅",
                   "施設志向の差\nの検定", ""], height=44)
GRPS = [("回答者全体", ALL, None),
        ("独居", series("独居") == 1, series("独居") == 0),
        ("世話をしてくれる人がいない", series("世話をする人なし") == 1,
         series("世話をする人なし") == 0),
        ("自力移動の手段がない", series("自力移動なし") == 1, series("自力移動なし") == 0),
        ("生活動作の困りごとあり", series("困りごとあり") == 1, series("困りごとあり") == 0),
        ("何らかの介護・介助が必要", mneed_all, mok),
        ("暮らし向きが苦しい", mhard, mnorm),
        ("75歳以上", m75, ~m75),
        ("85歳以上", m85, ~m85)]
for nm, m, mref2 in GRPS:
    jd = ztest(m, mref2, "_want_fac") if mref2 is not None else "―"
    r = body(ws, r, [nm, int(m.sum()), rate(m, "_want_fac")[0], rate(m, "_want_home")[0],
                     rate(m, "_want_unk")[0], rate(m, "_end_fac")[0],
                     rate(m, "_end_home")[0], jd, ""],
             {1: GRAY} if nm == "回答者全体" else None, height=18,
             bold=(nm == "回答者全体"))

r += 1
r = lead(ws, r, "【地区別の施設入所志向】", 9)
r = header(ws, r, ["町", "地区", "回答者数", "施設に入所\nしたい", "自宅での介護\nを希望",
                   "終のすみかが\n施設", "転居も少し\n考えている",
                   "自力移動の\n手段なし", ""], height=44)
for tname, aname, mask in rows:
    r = body(ws, r, [tname, aname, int(mask.sum()), rate(mask, "_want_fac")[0],
                     rate(mask, "_want_home")[0], rate(mask, "_end_fac")[0],
                     rate(mask, "_move")[0], rate(mask, "自力移動なし")[0], ""],
             height=18, bold=(aname in ("全体", "町計")))

r += 1
r = lead(ws, r, "【施設入所志向のある者の状況】", 9)
r = header(ws, r, ["指標", "施設に入所\nしたい", "自宅での介護\nを希望", "差", "検定",
                   "施設志向\n分母", "自宅志向\n分母", "", ""], height=40)
mfac, mhome = series("_want_fac") == 1, series("_want_home") == 1
for key, nm in [("独居", "独居者割合"), ("世話をする人なし", "世話をしてくれる人がいない割合"),
                ("困りごとあり", "生活動作の困りごとあり割合"),
                ("自力移動なし", "自力移動の手段なし割合"),
                ("暮らし苦しい", "暮らし向きが苦しい割合"),
                ("_need_any", "何らかの介護・介助が必要な割合"),
                ("frail2gp_25", "フレイルあり割合"), ("社会参加", "社会参加あり割合"),
                ("_move", "転居も少し考えている割合")]:
    v1, n1 = rate(mfac, key)
    v0, n0 = rate(mhome, key)
    if v1 is None or v0 is None:
        continue
    r = body(ws, r, [nm, v1, v0, round(v1 - v0, 1), ztest(mfac, mhome, key), n1, n0,
                     "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％。「施設に入所したい」は【大雪－問4】2）の選択肢3、"
     "「自宅での介護を希望」は同選択肢1・2の計である。"
     "「終のすみか」は同問6）による別設問である。"
     "注2）本調査の対象者は一般高齢者及び総合事業対象者であり、"
     "現に施設への入所を検討している段階の層を代表するものではない。"
     "本シートの値は将来の意向であり、需要の推計には用いない。"
     "注3）施設入所志向のある層に独居又は支援者不在が多い場合、"
     "施設整備ではなく在宅の支援体制の強化によって志向が変わる可能性がある。"
     "介護老人福祉施設の定員が234人から160人へ減少していること"
     "（要支援・要介護者1人当たり定員は平成30年比34.9％減）と併せて、"
     "第6章第5節の整備方針で論じる。"
     "注4）「わからない」と回答した層は意思決定支援（人生会議）の対象となる。", 9)

# ============================================================ 18 地域重点度
ws = sheet("18_地域重点度と社会参加の障壁", "JAGESの重点対象地域ランクと社会参加の障壁",
           "JAGESが提供する重点対象地域選定シートは、70指標について小学校区別に"
           "5段階のランクを付している（1が良好、5が要重点）。"
           "本シートでは健康コア11指標と社会参加コア15指標のランク平均を算定し、"
           "本分析の地区別集計及び通いの場への参加の障壁と併せて示す。",
           [16, 20, 8, 11, 11, 11, 11, 12, 26])

RANK1 = os.path.join(
    os.path.dirname(SRC),
    "3-19_大雪地区広域連合_2025JAGES調査・重点対象地域選定シート_26",
    "3-19_大雪地区広域連合_2025健康とくらしの調査・重点対象地域選定シート_1.xlsx")
rank_rows = []
if os.path.exists(RANK1):
    rk = pd.read_excel(RANK1, sheet_name="重点対象地域選定シート", header=None)
    for i in range(7, 22):
        code = rk.iat[i, 0]
        name = rk.iat[i, 1]
        if pd.isna(code):
            continue
        h = [pd.to_numeric(rk.iat[i, c], errors="coerce") for c in range(2, 13)]
        s = [pd.to_numeric(rk.iat[i, c], errors="coerce") for c in range(13, 28)]
        h = [x for x in h if pd.notna(x)]
        s = [x for x in s if pd.notna(x)]
        rank_rows.append((int(code), str(name),
                          round(sum(h) / len(h), 2) if h else None,
                          round(sum(s) / len(s), 2) if s else None))

r = lead(ws, 4, "【重点対象地域ランクの平均（1が良好、5が要重点）】", 9)
r = header(ws, r, ["町", "地区", "回答者数", "健康コア11指標\nランク平均",
                   "社会参加コア15指標\nランク平均", "通いの場\n参加割合",
                   "社会参加あり\n割合", "自力移動の\n手段なし", "判定"], height=48)
NAME2AREA = {"東聖小学校": "東聖小学校区", "東神楽小学校": "東神楽小学校区",
             "志比内小学校": "志比内小学校区", "東川第１小学校": "東川第1小学校区",
             "東川第２小学校": "東川第2小学校区", "東川第３小学校": "東川第3小学校区",
             "東川小学校": "東川小学校区", "美瑛小学校": "美瑛小学校区",
             "美馬牛小学校": "美馬牛小学校区", "美沢小学校": "美沢小学校区",
             "明徳小学校": "明徳小学校区", "美瑛東小学校": "美瑛東小学校区",
             "東神楽町全体": "町計", "東川町全体": "町計", "美瑛町全体": "町計"}
TOWN_OF = {1453: "東神楽町", 1458: "東川町", 1459: "美瑛町"}
if rank_rows:
    for code, name, hv, sv in rank_rows:
        an = NAME2AREA.get(name, name)
        tn = TOWN_OF.get(code if code < 10000 else code // 1000, "")
        mask = (A["町"] == tn) if an == "町計" else (A["地区"] == an)
        n = int(mask.sum())
        jd = ""
        if hv is not None and sv is not None:
            if hv >= 4 and sv >= 4:
                jd = "健康・社会参加とも要重点"
            elif hv >= 4:
                jd = "健康面が要重点"
            elif sv >= 4:
                jd = "社会参加面が要重点"
        fill = {9: NG_O} if "要重点" in jd else None
        r = body(ws, r, [tn if an == "町計" else "", an if an != "町計" else "町計", n, hv, sv,
                         rate(mask, "通いの場")[0], rate(mask, "社会参加")[0],
                         rate(mask, "自力移動なし")[0], jd], fill, height=18,
                 bold=(an == "町計"))
else:
    r = body(ws, r, ["重点対象地域選定シートが見つかりません", "", "", "", "", "", "", "", ""],
             {1: NG_O}, height=18)

r += 1
r = lead(ws, r, "【通いの場に参加していない層の特性】", 9)
r = header(ws, r, ["指標", "通いの場に\n参加", "参加して\nいない", "差", "検定",
                   "参加\n分母", "不参加\n分母", "", ""], height=40)
msal1, msal0 = series("通いの場") == 1, series("通いの場") == 0
for key, nm in [("自力移動なし", "自力移動の手段なし割合"), ("運転あり", "自分で運転する割合"),
                ("暮らし苦しい", "暮らし向きが苦しい割合"), ("独居", "独居者割合"),
                ("外出低頻度", "外出頻度が月1〜3回以下の割合"),
                ("困りごとあり", "生活動作の困りごとあり割合"),
                ("frail2gp_25", "フレイルあり割合"), ("tojikomori2gp_25", "閉じこもり者割合"),
                ("utsucheck2gp_25", "うつ割合"), ("happy2gp_25", "幸福感がある者の割合"),
                ("_dem_co", "認知症相談窓口を知っている割合")]:
    v1, n1 = rate(msal1, key)
    v0, n0 = rate(msal0, key)
    if v1 is None or v0 is None:
        continue
    r = body(ws, r, [nm, v1, v0, round(v1 - v0, 1), ztest(msal1, msal0, key), n1, n0,
                     "", ""], height=18)

r += 1
r = lead(ws, r, "【年齢階級別の通いの場参加割合と社会参加の広がり】", 9)
r = header(ws, r, ["年齢階級", "通いの場", "社会参加\n（6区分合成）", "スポーツの会",
                   "趣味の会", "ボランティア", "回答者数", "", ""], height=36)
for g in AGES:
    m = A["年齢階級"] == g
    A["_sp"] = flag_in("cmnt6sp25", [1, 2, 3, 4])
    A["_hb"] = flag_in("cmnt6hb25", [1, 2, 3, 4])
    A["_vl"] = flag_in("cmnt6vl25", [1, 2, 3, 4])
    r = body(ws, r, [g, rate(m, "通いの場")[0], rate(m, "社会参加")[0], rate(m, "_sp")[0],
                     rate(m, "_hb")[0], rate(m, "_vl")[0], int(m.sum()), "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％（ランク平均のみ1〜5の値）。"
     "注2）重点対象地域選定シートのランクは、JAGESが参加保険者の地区を"
     "指標ごとに5段階に分けたものである。1が良好、5が要重点を示す。"
     "健康コア11指標は幸福感・リスク点数・フレイル・運動機能・転倒・認知症リスク・"
     "認知症得点・物忘れ・口腔機能・残歯数・うつの11指標、"
     "社会参加コア15指標は閉じこもり・スポーツの会・趣味の会・ボランティア・"
     "学習教養・特技を伝える活動・友人と会う頻度・交流する友人数・情緒的サポート・"
     "手段的サポート・ソーシャルキャピタル得点等の15指標である。"
     "注3）本分析の地区別集計（01シート）とランク評価は独立に算定したものであり、"
     "両者が同じ方向を示す地区は、重点化の根拠として確度が高い。"
     "注4）通いの場に参加していない層の特性は、"
     "総合事業ベースの参加率が全国の3分の1にとどまり低下傾向にあること"
     "（妥当性検証報告書 所見12）を踏まえ、参加の障壁を特定するために分析した。", 9)




# ============================================================ 19 同規模保険者との比較
ws = sheet("19_同規模保険者との比較", "同規模保険者（人口5万未満・40保険者）との比較",
           "健康とくらしの調査は令和7年度に64介護保険者74市町村が参加しており、"
           "調査報告書は人口5万未満の40保険者を「同規模保険者」として比較の対象としている"
           "（同規模保険者の集計数82,106票）。"
           "仕様書４（4）が求める類似保険者との比較分析の一部を、"
           "見える化システムの類似保険者比較機能を待たずに行うことができる。"
           "ただし同規模保険者40はこの調査の参加団体であって"
           "介護保険上の類似保険者ではないため、本比較は暫定的なものである"
           "（第9期評価のレッドチームレビュー No.14、確認事項No.3）。",
           [30, 10, 10, 10, 10, 10, 11, 11, 26])

PEER = [
    ("フレイルあり割合（基本チェックリスト8項目以上）",
     [19.1, 11.8, 13.0, 18.0, 26.3, 38.7], [17.0, 10.0, 11.1, 16.5, 24.5, 42.9],
     [18.4, 13.6, 11.1, 17.1, 24.1, 35.1], [21.6, 12.7, 16.2, 20.1, 29.7, 38.9],
     [19.4, 10.7, 13.1, 17.0, 25.8, 42.3], "低い方が良い"),
    ("運動機能低下者割合（基本チェックリスト）",
     [9.1, 4.0, 5.6, 8.5, 14.3, 20.4], [8.2, 2.1, 4.8, 7.5, 13.5, 28.7],
     [9.0, 6.7, 4.7, 8.3, 13.0, 16.9], [10.1, 4.3, 7.1, 9.6, 16.1, 17.9],
     [12.3, 4.5, 6.8, 10.3, 17.7, 33.2], "低い方が良い"),
    ("１年間の転倒あり割合",
     [35.5, 33.5, 29.1, 38.0, 37.0, 45.3], [32.6, 30.5, 25.8, 35.2, 34.4, 48.9],
     [34.0, 35.7, 26.3, 36.5, 32.8, 42.4], [39.3, 35.5, 34.5, 41.9, 42.8, 45.3],
     [30.0, 23.9, 25.7, 29.4, 34.3, 43.9], "低い方が良い"),
    ("物忘れが多い者の割合",
     [41.3, 34.8, 38.7, 41.9, 46.6, 50.5], [39.5, 34.4, 37.4, 41.2, 42.2, 51.5],
     [40.3, 31.9, 33.7, 42.5, 47.3, 50.0], [43.9, 37.3, 43.7, 42.1, 49.8, 50.2],
     [39.0, 31.9, 36.0, 38.6, 43.8, 51.4], "低い方が良い"),
    ("閉じこもり者割合",
     [8.5, 3.7, 5.0, 8.2, 13.4, 19.1], [7.2, 2.3, 4.8, 6.5, 9.9, 25.5],
     [8.6, 5.0, 4.2, 9.6, 12.6, 14.5], [9.7, 4.3, 5.8, 8.4, 17.1, 18.6],
     [6.0, 2.4, 3.5, 4.8, 8.1, 16.7], "低い方が良い"),
    ("うつ割合（基本チェックリスト）",
     [28.2, 23.8, 23.9, 28.7, 30.8, 41.2], [26.8, 22.7, 23.7, 28.1, 29.9, 39.7],
     [25.0, 18.8, 22.5, 24.3, 28.4, 35.5], [32.1, 28.8, 25.2, 33.2, 33.7, 46.5],
     [28.9, 22.7, 24.1, 27.6, 34.5, 44.0], "低い方が良い"),
    ("口腔機能低下者割合（基本チェックリスト）",
     [24.7, 18.4, 20.7, 26.0, 29.9, 34.7], [21.5, 16.3, 17.2, 21.5, 28.7, 36.7],
     [24.8, 18.0, 22.3, 25.2, 29.6, 31.7], [27.5, 21.1, 22.8, 30.7, 31.2, 35.7],
     [21.8, 16.1, 18.6, 22.2, 26.1, 31.1], "低い方が良い"),
    ("低栄養の傾向割合",
     [7.4, 5.9, 6.2, 7.7, 9.0, 9.7], [6.7, 5.8, 4.7, 8.6, 8.1, 8.0],
     [8.1, 6.3, 8.4, 7.0, 11.6, 8.0], [7.3, 5.5, 6.0, 7.5, 7.7, 11.9],
     [7.5, 6.4, 6.5, 7.5, 8.5, 10.5], "低い方が良い"),
    ("認知機能低下者割合（基本チェックリスト）",
     [34.7, 31.4, 32.6, 31.9, 38.0, 47.5], [34.9, 30.0, 35.5, 30.5, 39.1, 51.4],
     [34.6, 34.3, 28.4, 33.2, 39.1, 44.0], [34.7, 30.9, 33.1, 31.9, 36.2, 47.7],
     [35.1, 28.8, 31.8, 33.3, 38.9, 50.1], "低い方が良い"),
    ("IADL（自立度）低下者（1項目以上）割合",
     [9.8, 4.8, 5.1, 9.1, 14.9, 23.4], [9.8, 5.9, 6.0, 8.3, 13.6, 30.2],
     [10.1, 4.2, 5.0, 12.2, 12.9, 19.6], [9.4, 4.0, 4.4, 7.1, 17.7, 22.0],
     [12.4, 4.9, 7.2, 10.3, 17.2, 32.8], "低い方が良い"),
    ("幸福感がある者の割合（8/10点以上）",
     [50.6, 51.2, 48.7, 48.5, 51.5, 57.9], [51.0, 49.3, 50.1, 50.5, 49.6, 63.3],
     [51.9, 60.2, 45.7, 48.4, 53.6, 58.6], [49.2, 46.5, 49.7, 46.9, 51.5, 54.0],
     [47.2, 45.2, 45.6, 47.1, 49.0, 51.8], "高い方が良い"),
    ("就労していない者の割合",
     [54.9, 33.7, 50.0, 61.7, 69.2, 74.4], [56.8, 37.7, 56.1, 63.6, 72.7, 72.8],
     [55.6, 32.2, 48.6, 61.5, 69.0, 74.8], [52.6, 30.0, 45.2, 60.1, 66.4, 75.1],
     [57.5, 36.5, 52.1, 64.3, 72.0, 75.0], "低い方が良い"),
    ("ボランティア参加者（月1回以上）割合",
     [10.7, 8.0, 10.7, 11.0, 11.8, 13.8], [9.3, 8.0, 8.7, 11.2, 9.7, 9.2],
     [13.0, 12.1, 13.0, 12.5, 12.4, 16.6], [10.2, 5.0, 11.0, 9.4, 13.2, 14.7],
     [12.2, 9.6, 12.4, 13.7, 13.7, 11.2], "高い方が良い"),
    ("スポーツの会参加者（月1回以上）割合",
     [20.7, 16.1, 19.0, 20.4, 27.5, 23.8], [21.0, 19.3, 20.7, 21.0, 25.0, 19.2],
     [21.6, 17.2, 18.7, 21.1, 30.7, 22.4], [19.6, 11.4, 17.6, 19.3, 27.1, 27.9],
     [19.1, 15.4, 19.0, 20.5, 22.8, 17.4], "高い方が良い"),
    ("趣味の会参加者（月1回以上）割合",
     [20.2, 13.8, 18.7, 21.4, 26.1, 24.4], [20.4, 15.2, 21.5, 21.7, 24.6, 21.2],
     [21.4, 15.1, 14.7, 23.1, 33.6, 23.5], [18.9, 11.1, 19.1, 19.5, 21.0, 27.3],
     [20.4, 15.7, 20.0, 22.8, 24.3, 18.4], "高い方が良い"),
    ("学習・教養サークル参加者（月1回以上）割合",
     [6.0, 3.3, 5.2, 6.3, 7.8, 9.7], [5.4, 2.6, 5.5, 7.0, 6.8, 6.1],
     [6.8, 2.9, 4.2, 6.3, 10.8, 13.5], [5.9, 4.3, 5.6, 5.8, 6.3, 9.0],
     [6.2, 4.3, 5.7, 7.0, 7.9, 6.9], "高い方が良い"),
    ("通いの場参加者（月1回以上）割合",
     [8.8, 4.1, 6.4, 8.5, 13.1, 17.6], [7.5, 4.6, 6.3, 8.8, 10.1, 11.3],
     [9.3, 6.7, 7.0, 7.8, 12.8, 17.2], [9.5, 1.5, 6.1, 9.0, 15.9, 22.1],
     [13.2, 5.5, 9.9, 14.5, 19.2, 22.3], "高い方が良い"),
    ("特技や経験を他者に伝える活動参加者（月1回以上）割合",
     [5.0, 4.9, 4.2, 5.2, 5.9, 5.6], [3.8, 4.1, 2.4, 4.5, 4.2, 4.5],
     [6.4, 6.3, 5.1, 7.3, 6.8, 6.5], [5.0, 4.7, 5.2, 3.9, 6.8, 5.5],
     [4.4, 3.7, 4.4, 4.7, 4.8, 4.1], "高い方が良い"),
    ("友人知人と会う頻度が高い（月1回以上）者の割合",
     [63.0, 56.6, 62.8, 66.3, 67.5, 60.6], [60.1, 58.5, 62.4, 60.5, 62.7, 51.9],
     [63.1, 51.9, 64.3, 67.9, 66.0, 61.5], [65.5, 57.8, 62.0, 70.0, 73.0, 65.4],
     [71.2, 66.2, 71.3, 74.5, 74.5, 68.1], "高い方が良い"),
]

r = lead(ws, 4, "【1　全体の比較（大雪と同規模保険者）】", 9)
r = header(ws, r, ["指標", "大雪地区\n広域連合", "東川町", "美瑛町", "東神楽町",
                   "同規模保険者", "同規模との差\n（ポイント）", "評価の向き", "判定"],
           height=44)
peer_sum = {"良好": 0, "同等": 0, "課題": 0}
for nm, dz, hg, hk, bi, pe, direc in PEER:      # hg=東神楽町, hk=東川町, bi=美瑛町
    d = round(dz[0] - pe[0], 1)
    if abs(d) < 1.0:
        jd = "同等"
    elif (d < 0) == (direc == "低い方が良い"):
        jd = "良好"
    else:
        jd = "課題"
    peer_sum[jd] += 1
    fl = {9: OK_G if jd == "良好" else (NG_O if jd == "課題" else GRAY)}
    r = body(ws, r, [nm, dz[0], hk[0], bi[0], hg[0], pe[0], d, direc, jd], fl,
             height=18, align={9: "center"})
r = body(ws, r, ["【集計】良好%d・同等%d・課題%d" % (peer_sum["良好"], peer_sum["同等"],
                                                    peer_sum["課題"]),
                 "", "", "", "", "", "", "", ""], {1: IN_Y}, height=18, bold=True)

r += 1
r = lead(ws, r, "【2　同規模保険者との差が大きい指標（年齢階級別）】", 9)
r = header(ws, r, ["指標", "区分", "全体", "65～69歳", "70～74歳", "75～79歳",
                   "80～84歳", "85歳以上", "所見"], height=32)
BIG = ["友人知人と会う頻度が高い（月1回以上）者の割合", "通いの場参加者（月1回以上）割合",
       "１年間の転倒あり割合", "閉じこもり者割合", "口腔機能低下者割合（基本チェックリスト）",
       "IADL（自立度）低下者（1項目以上）割合", "運動機能低下者割合（基本チェックリスト）"]
NOTE_BIG = {
    "友人知人と会う頻度が高い（月1回以上）者の割合":
        "全年齢階級で同規模を下回る。社会的ネットワークの薄さは年齢によらない",
    "通いの場参加者（月1回以上）割合":
        "65〜79歳で同規模の6〜7割にとどまる。85歳以上ではほぼ同水準",
    "１年間の転倒あり割合": "65〜69歳で同規模を9.6ポイント上回る。冬期の路面条件の影響が疑われる",
    "閉じこもり者割合": "80〜84歳で同規模を5.3ポイント上回る",
    "口腔機能低下者割合（基本チェックリスト）": "75歳以上で同規模を3.8〜3.8ポイント上回る",
    "IADL（自立度）低下者（1項目以上）割合":
        "85歳以上で同規模を9.4ポイント下回る。対象者範囲の違いの影響が大きい可能性",
    "運動機能低下者割合（基本チェックリスト）":
        "85歳以上で同規模を12.8ポイント下回る。対象者範囲の違いの影響が大きい可能性",
}
for nm, dz, hg, hk, bi, pe, direc in PEER:
    if nm not in BIG:
        continue
    r = body(ws, r, [nm, "大雪"] + dz + [NOTE_BIG.get(nm, "")], height=18)
    r = body(ws, r, ["", "同規模保険者"] + pe + [""], {2: GRAY}, height=18)
    r = body(ws, r, ["", "差"] + [round(a - b, 1) for a, b in zip(dz, pe)] + [""],
             {2: IN_Y}, height=18)

r += 1
r = lead(ws, r, "【3　比較の妥当性　調査対象者の範囲の違い】", 9)
r = header(ws, r, ["対象者属性", "保険者数", "構成比", "大雪の該当", "比較への影響",
                   "", "", "", ""], height=32)
ATTR = [("一般高齢者＋総合事業対象者＋要支援者", 29, "72.5％", "―",
         "要支援者を含むため、リスク指標は高く出る"),
        ("一般高齢者＋総合事業対象者", 5, "12.5％", "○（大雪はこの区分）",
         "要支援者を含まないため、リスク指標は低く出る"),
        ("一般高齢者のみ", 4, "10.0％", "―", "最もリスク指標が低く出る"),
        ("一般高齢者＋総合事業対象者＋要支援者＋要介護者", 1, "2.5％", "―",
         "最もリスク指標が高く出る"),
        ("同＋要介護1・2のみ", 1, "2.5％", "―", "同上")]
for a, c_, p_, o, imp in ATTR:
    fl = {4: IN_Y} if o.startswith("○") else None
    r = body(ws, r, [a, c_, p_, o, imp, "", "", "", ""], fl, height=18)
r = body(ws, r, ["計", 40, "100.0％", "", "要支援者を含む保険者が31（77.5％）を占める",
                 "", "", "", ""], {1: GRAY}, height=18, bold=True)

note(ws, r + 1,
     "注1）単位は％。同規模保険者は人口5万未満の40保険者で、集計数は82,106票。"
     "広域連合の場合は構成市町村の人口の和を自治体数で除した平均による。"
     "注2）判定は同規模保険者との差が1.0ポイント未満を「同等」、"
     "それ以上で望ましい方向にあるものを「良好」、逆方向にあるものを「課題」とした。"
     "この閾値は本表独自のものであり、他の成果物の定性的な判定とは基準が異なる"
     "（レッドチームレビュー No.17）。"
     "注3）比較の妥当性について。同規模保険者40のうち31（77.5％）は要支援者を"
     "調査対象に含んでいるのに対し、当広域連合は含んでいない（一般高齢者＋総合事業対象者）。"
     "要支援者を含む集団はリスク指標が高く出るため、"
     "当広域連合の値は構造的に良好側へ偏る。したがって、"
     "運動機能低下（▲3.2ポイント）、IADL低下（▲2.6ポイント）、"
     "85歳以上のフレイル（▲3.6ポイント）などの良好な結果は、"
     "対象者範囲の違いで説明できる可能性があり、良好と断定できない。"
     "一方、転倒（＋5.5ポイント）、閉じこもり（＋2.5ポイント）、"
     "口腔機能低下（＋2.9ポイント）、通いの場参加（▲4.4ポイント）、"
     "友人知人と会う頻度（▲8.2ポイント）は、"
     "偏りが良好側に働いてもなお同規模を下回っており、課題としての確度が高い。"
     "注4）年齢階級別の比較は、この偏りの影響を年齢層ごとに確認するために示している。"
     "偏りは高齢層ほど大きくなるため、65〜74歳の比較が最も信頼できる。", 9)

# ============================================================ 20 参加自治体順位
ws = sheet("20_参加自治体における順位", "健康とくらしの調査 参加74市町村における3町の順位",
           "健康とくらしの調査には令和7年度に64介護保険者74市町村が参加している。"
           "JAGESが提供する地域診断書は、主要12指標について参加市町村中の順位を示している"
           "（順位が大きいほど当該指標の状況が悪い）。"
           "同規模保険者との比較（19シート）とは母集団が異なるため、両者を併せて確認する。",
           [30, 14, 11, 9, 11, 11, 11, 11, 26])

RANK = [
    ("フレイルあり割合", "東神楽町", 16.9, 39, 16.9, 11.7, 21.6),
    ("フレイルあり割合", "東川町", 18.3, 56, 16.9, 11.7, 21.6),
    ("フレイルあり割合", "美瑛町", 21.1, 71, 16.9, 11.7, 21.6),
    ("生活機能低下者割合", "東神楽町", 3.6, 45, 3.3, 1.5, 6.3),
    ("生活機能低下者割合", "東川町", 2.9, 31, 3.3, 1.5, 6.3),
    ("生活機能低下者割合", "美瑛町", 3.9, 53, 3.3, 1.5, 6.3),
    ("運動機能低下者割合", "東神楽町", 7.9, 19, 9.4, 4.8, 14.0),
    ("運動機能低下者割合", "東川町", 8.8, 29, 9.4, 4.8, 14.0),
    ("運動機能低下者割合", "美瑛町", 9.7, 41, 9.4, 4.8, 14.0),
    ("低栄養者割合", "東神楽町", 1.7, 23, 1.9, 0.8, 3.2),
    ("低栄養者割合", "東川町", 2.3, 59, 1.9, 0.8, 3.2),
    ("低栄養者割合", "美瑛町", 1.7, 28, 1.9, 0.8, 3.2),
    ("口腔機能低下者割合", "東神楽町", 21.4, 52, 20.6, 15.4, 27.1),
    ("口腔機能低下者割合", "東川町", 24.8, 70, 20.6, 15.4, 27.1),
    ("口腔機能低下者割合", "美瑛町", 27.1, 73, 20.6, 15.4, 27.1),
    ("閉じこもり者割合", "東神楽町", 7.1, 60, 4.7, 2.0, 9.8),
    ("閉じこもり者割合", "東川町", 8.6, 70, 4.7, 2.0, 9.8),
    ("閉じこもり者割合", "美瑛町", 9.8, 73, 4.7, 2.0, 9.8),
    ("認知機能低下者割合", "東神楽町", 34.7, 48, 33.5, 28.2, 38.1),
    ("認知機能低下者割合", "東川町", 34.4, 46, 33.5, 28.2, 38.1),
    ("認知機能低下者割合", "美瑛町", 34.1, 43, 33.5, 28.2, 38.1),
    ("うつがある者の割合", "東神楽町", 26.9, 33, 27.2, 20.3, 33.5),
    ("うつがある者の割合", "東川町", 24.9, 17, 27.2, 20.3, 33.5),
    ("うつがある者の割合", "美瑛町", 31.5, 70, 27.2, 20.3, 33.5),
    ("幸福感がある者の割合", "東神楽町", 50.9, 27, 48.5, 38.0, 56.3),
    ("幸福感がある者の割合", "東川町", 52.0, 14, 48.5, 38.0, 56.3),
    ("幸福感がある者の割合", "美瑛町", 49.2, 37, 48.5, 38.0, 56.3),
    ("ソーシャルキャピタル得点・社会参加", "東神楽町", 46.1, 55, 52.2, 19.0, 81.9),
    ("ソーシャルキャピタル得点・社会参加", "東川町", 51.2, 41, 52.2, 19.0, 81.9),
    ("ソーシャルキャピタル得点・社会参加", "美瑛町", 42.5, 60, 52.2, 19.0, 81.9),
    ("ソーシャルキャピタル得点・連帯感", "東神楽町", 153.9, 40, 154.8, 128.9, 174.5),
    ("ソーシャルキャピタル得点・連帯感", "東川町", 162.7, 15, 154.8, 128.9, 174.5),
    ("ソーシャルキャピタル得点・連帯感", "美瑛町", 163.3, 14, 154.8, 128.9, 174.5),
    ("ソーシャルキャピタル得点・助け合い", "東神楽町", 197.8, 47, 198.3, 193.7, 202.4),
    ("ソーシャルキャピタル得点・助け合い", "東川町", 196.1, 65, 198.3, 193.7, 202.4),
    ("ソーシャルキャピタル得点・助け合い", "美瑛町", 195.6, 68, 198.3, 193.7, 202.4),
]
r = header(ws, 4, ["指標", "町", "値", "順位\n（74市町村中）", "参加市町村\n平均",
                   "最小値", "最大値", "順位の位置", "所見"], height=44)
for nm, tn, v, rk, av, mn, mx in RANK:
    pos = "下位4分の1（要重点）" if rk >= 56 else (
        "中位より下" if rk >= 38 else ("中位より上" if rk >= 19 else "上位4分の1"))
    fl = {4: NG_O} if rk >= 56 else ({4: OK_G} if rk <= 18 else None)
    obs = ""
    if rk >= 70:
        obs = "参加74市町村中で最も悪い水準"
    elif rk <= 15:
        obs = "参加74市町村中で最も良い水準"
    r = body(ws, r, [nm, tn, v, rk, av, mn, mx, pos, obs], fl, height=18,
             align={4: "center", 8: "center"})

r += 1
r = lead(ws, r, "【町別の順位の総括】", 9)
r = header(ws, r, ["町", "上位4分の1\n（1〜18位）", "中位より上\n（19〜37位）",
                   "中位より下\n（38〜55位）", "下位4分の1\n（56〜74位）",
                   "最悪水準の指標", "", "", ""], height=44)
for tn in ["東川町", "美瑛町", "東神楽町"]:
    sub = [x for x in RANK if x[1] == tn]
    q1 = len([x for x in sub if x[3] <= 18])
    q2 = len([x for x in sub if 19 <= x[3] <= 37])
    q3 = len([x for x in sub if 38 <= x[3] <= 55])
    q4 = len([x for x in sub if x[3] >= 56])
    worst = "、".join(x[0] for x in sub if x[3] >= 65) or "―"
    r = body(ws, r, [tn, q1, q2, q3, q4, worst, "", "", ""],
             {5: NG_O} if q4 >= 4 else None, height=32)

note(ws, r + 1,
     "注1）順位は健康とくらしの調査に参加した74市町村の中での順位で、"
     "数字が大きいほど当該指標の状況が悪いことを示す。"
     "ソーシャルキャピタル得点は数字が大きいほど良好であるが、順位の向きは統一されている。"
     "注2）参加市町村には人口規模の異なる自治体が含まれ、"
     "調査対象者の範囲も保険者ごとに異なる（19シート）。"
     "順位は目安であり、同規模保険者との比較と併せて解釈する。"
     "注3）美瑛町は口腔機能低下者割合73位、閉じこもり者割合73位、"
     "フレイルあり割合71位、うつがある者の割合70位と、"
     "健康関連の4指標で参加74市町村中の下位に位置する。"
     "東川町も口腔機能低下者割合70位、閉じこもり者割合70位である。"
     "注4）一方、幸福感がある者の割合は東川町14位、東神楽町27位と良好であり、"
     "ソーシャルキャピタル得点の連帯感は美瑛町14位、東川町15位と参加市町村の上位にある。"
     "健康指標が悪い一方で、地域への信頼や連帯感は高いという構造がみられる。", 9)

# ============================================================ 21 他団体比較の総括
ws = sheet("21_他団体との比較の総括", "他団体との比較による当広域連合の位置",
           "健康とくらしの調査、地域包括ケア「見える化」システム、"
           "保険者機能強化推進交付金等及び北海道の資料により、"
           "当広域連合が他団体と比べてどの位置にあるかを整理する。"
           "仕様書４（4）が求める全国平均、北海道平均及び類似保険者との比較分析に対応する。"
           "ただし比較対象は入手可能なもの（同規模保険者40、参加74市町村、全国・北海道）であり、"
           "介護保険上の類似保険者の選定条件は未決定である"
           "（第9期評価のレッドチームレビュー No.14、確認事項No.3）。"
           "本表の比較は暫定的なものである。",
           [7, 26, 22, 14, 14, 14, 12, 34])

r = header(ws, 4, ["No.", "比較の観点", "指標", "大雪", "比較対象の値",
                   "比較対象", "判定", "第10期での意味"], height=44)

CMP = [
    (1, "住民の心身の状態", "フレイルあり割合", "19.1％", "19.4％",
     "同規模保険者40", "同等", "同規模と同水準。ただし当広域連合は要支援者を"
     "調査対象に含まないため、実態はやや悪い可能性がある"),
    (2, "住民の心身の状態", "1年間の転倒あり割合", "35.5％", "30.0％",
     "同規模保険者40", "課題", "同規模を5.5ポイント上回る。65〜69歳では9.6ポイント上回り、"
     "冬期の路面条件との関連が疑われる。介護予防の重点課題"),
    (3, "住民の心身の状態", "口腔機能低下者割合", "24.7％", "21.8％",
     "同規模保険者40", "課題", "同規模を2.9ポイント上回る。"
     "美瑛町は参加74市町村中73位。歯科口腔保健との連携を要する"),
    (4, "外出と社会参加", "閉じこもり者割合", "8.5％", "6.0％",
     "同規模保険者40", "課題", "同規模の1.4倍。美瑛町・東川町は参加74市町村中70位以下。"
     "移動支援と通いの場の配置が論点"),
    (5, "外出と社会参加", "通いの場参加者割合", "8.8％", "13.2％",
     "同規模保険者40", "課題", "同規模の3分の2。65〜79歳では6〜7割にとどまる。"
     "総合事業ベースでは全国の3分の1（所見12）であり、両者が同じ方向を示す"),
    (6, "外出と社会参加", "友人知人と会う頻度が高い者の割合", "63.0％", "71.2％",
     "同規模保険者40", "課題", "同規模を8.2ポイント下回り、全年齢階級で下回る。"
     "本調査で最も差が大きい指標。社会的ネットワークの薄さは年齢によらない"),
    (7, "外出と社会参加", "ソーシャルキャピタル得点・連帯感", "美瑛163.3／東川162.7",
     "154.8", "参加74市町村", "良好", "美瑛町14位、東川町15位。"
     "健康指標が悪い一方で地域への信頼・連帯感は高い。地域づくりの資源となる"),
    (8, "要介護認定", "性・年齢調整済み要介護2以上認定率", "9.25％", "全国8.99％／北海道8.59％",
     "全国・北海道（W144）", "課題", "全国を0.26ポイント上回る。"
     "H01のデータ源。年齢構成を調整してもなお全国を上回る"),
    (9, "要介護認定", "同上の変化率", "＋1.76", "全国▲0.37",
     "全国（W145）", "課題", "全国が改善する中で悪化している。"
     "年齢構成を調整すると中重度は低下が続き軽度が直近で上昇に転じている"
     "（認定率の年齢調整分析 04シート）"),
    (10, "給付水準", "第1号被保険者1人あたり給付月額の地域差指数", "1.08", "1.00",
     "全国（D49）", "課題", "内訳は認定率1.02・受給率1.08・受給者単価1.06。"
     "給付水準の高さは受給率と単価がともに高いことによる"),
    (11, "給付水準", "施設・居住系サービスの受給率", "全国比1.08", "1.00",
     "全国（D41）", "課題", "在宅は全国比1.02で全国並み。"
     "給付水準の高さは施設・居住系の受給率と単価がともに高いことによる（所見1）"),
    (12, "給付水準", "調整済み給付月額（施設・居住系）", "13,097円", "北海道10,504円",
     "北海道（第9期道計画）", "課題", "北海道平均を2,593円上回る。"
     "上川中部圏域の中でも高い水準"),
    (13, "サービス利用", "介護サービス利用率", "74.3％", "―",
     "経年（平成28年度82.2％）", "課題", "7年間で7.9ポイント低下。"
     "未利用認定者504人（認定者の25.7％）。比較可能な他団体の値は未入手"),
    (14, "サービス利用", "訪問介護の利用回数", "55.4回／月", "全国29.7回／北海道29.9回",
     "全国・北海道（D31）", "要因分析", "全国の1.87倍。"
     "通所の供給が薄いことによる代替の可能性（第2章第2節）"),
    (15, "サービス利用", "通所介護の受給者1人あたり給付月額", "62,688円", "全国84,875円",
     "全国（D31）", "課題", "全国の0.74倍。通所の利用強度が低い"),
    (16, "供給基盤", "通所介護の事業所数（人口10万対）", "21.6", "全国35.8",
     "全国（K3）", "課題", "全国の0.60倍。地域密着型と合わせても同水準"),
    (17, "供給基盤", "24時間対応3サービスの事業所数", "0", "上川中部圏域は令和4年実績あり",
     "圏域（北海道第9期計画）", "課題", "制度創設以来13年間ゼロ。"
     "圏域では定期巡回59.1人／月、夜間対応70.8人／月の実績がある（所見17）"),
    (18, "保険者機能", "保険者機能強化推進交付金等の総合得点", "332.0点",
     "全国422.4点／北海道413.6点", "全国・北海道（W126）", "課題",
     "全国平均の78.6％。交付金額は得点に応じて配分されるため財源差に直結"),
    (19, "保険者機能", "同　活動指標群の得点", "全国の23.2％", "―",
     "全国（W130）", "課題", "成果指標群は全国の123.5％。"
     "成果は出ているが活動を国の様式で記録・報告できていない（所見19）"),
    (20, "保険料", "第9期の保険料基準額", "6,400円", "上川管内21保険者",
     "北海道（管内保険料一覧）", "妥当", "必要保険料月額は令和6年度6,063円・"
     "令和7年度6,352円で基準額を下回り、剰余方向で推移している（所見9）"),
]
for row in CMP:
    fl = {7: NG_O if row[6] == "課題" else (OK_G if row[6] in ("良好", "妥当") else GRAY)}
    r = body(ws, r, list(row), fl, height=44, align={1: "center", 7: "center"})

r += 1
r = lead(ws, r, "【総括】", 8)
r = header(ws, r, ["観点", "他団体との比較でみた当広域連合の位置", "", "", "", "", "",
                   "第10期での対応"], height=32)
for k, v_, mean in [
    ("心身の状態", "同規模保険者と同水準だが、転倒と口腔機能低下が明確に悪い。"
     "調査対象者に要支援者を含まないことを考慮すると、実態はさらに悪い可能性がある",
     "基本目標1に転倒予防と口腔機能の維持を明示する"),
    ("外出と社会参加", "本調査で最も差が大きい領域。通いの場参加は同規模の3分の2、"
     "友人知人と会う頻度は8.2ポイント下、閉じこもりは1.4倍。"
     "一方で地域への連帯感は参加74市町村の上位",
     "基本目標1・2の重点領域とする。連帯感の高さを地域づくりの資源として活用する"),
    ("要介護認定と給付", "年齢調整後の認定率は全国並みだが、"
     "給付水準は施設・居住系の受給率と単価により全国を8％上回る。"
     "性・年齢調整済み要介護2以上認定率は全国を上回り、しかも悪化している",
     "H01・H02の目標設定と、第6章の整備方針に反映する"),
    ("供給基盤", "通所介護の事業所数は全国の0.6倍、24時間対応3サービスは13年間ゼロ。"
     "圏域には実績があり、区域外利用の可否が論点",
     "第6章第4節の確保方策で決定する（確認事項No.24）"),
    ("保険者機能", "交付金の総合得点は全国平均の78.6％。"
     "成果指標群は全国の123.5％である一方、活動指標群は23.2％",
     "第5章を5層構成としアウトプット指標を設定する。"
     "交付金の評価項目に対応する記録の様式を整える"),
    ("保険料", "第9期は剰余方向で推移しており、設定は妥当",
     "基金残高と取崩実績を確認のうえ第10期の算定を行う（確認事項No.16）"),
]:
    r = body(ws, r, [k, v_, "", "", "", "", "", mean], height=48)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=7)

note(ws, r + 1,
     "注1）比較対象の欄に示した略号は次のとおり。"
     "同規模保険者40＝健康とくらしの調査に参加した人口5万未満の40保険者。"
     "参加74市町村＝同調査に参加した64介護保険者74市町村。"
     "D・K・W＝地域包括ケア「見える化」システムの系列番号。"
     "注2）健康とくらしの調査による比較は、当広域連合が調査対象者に"
     "要支援者を含まない一方、同規模保険者40のうち31（77.5％）は含んでいるため、"
     "当広域連合の値が良好側へ偏る（19シート）。"
     "「課題」と判定した指標は、この偏りを考慮してもなお他団体を下回るものである。"
     "注3）仕様書４（4）が求める類似保険者との比較分析について、"
     "本シートは健康とくらしの調査の同規模保険者及び見える化の全国・北海道値によっている。"
     "見える化システムの類似保険者比較機能による出力は、"
     "比較対象の選定条件の決定を要する（確認事項No.3）。"
     "注4）判定は次の意味である。良好＝他団体より望ましい方向にある。"
     "同等＝差が小さい。課題＝他団体より望ましくない方向にある。"
     "要因分析＝差が大きいが、良否の判断に要因の分解を要する。"
     "妥当＝設定や運営が適切と評価できる。", 8)



# ============================================================ 22 性別による差
ws = sheet("22_性別による差", "性別によるリスク・社会参加・移動の差",
           "調査報告書及びJAGESのクロス集計表は性別を集計軸の一つとしているが、"
           "本分析の01〜21シートでは性別を用いていなかった。"
           "介護予防・社会参加の施策は対象の性別によって設計が変わるため、"
           "主要指標を性別に集計する。回答者は男性2,233人・女性2,496人。",
           [26, 10, 10, 10, 10, 10, 10, 12, 30])

SEXV = pd.to_numeric(A["sex_m25"], errors="coerce")
mM, mF = SEXV == 1, SEXV == 2
A["孤食"] = pd.to_numeric(A["eatalone2gp_25"], errors="coerce")

r = lead(ws, 4, "【1　性別の主要指標】", 9)
r = header(ws, r, ["指標", "男性", "女性", "差\n（ポイント）", "検定",
                   "男性\n分母", "女性\n分母", "男女比", "所見"], height=40)
SEXIND = [
    ("frail2gp_25", "フレイルあり割合", ""),
    ("undo2gp_25", "運動機能低下者割合", ""),
    ("tojikomori2gp_25", "閉じこもり者割合", ""),
    ("utsucheck2gp_25", "うつ割合", ""),
    ("ninchicheck2gp_25", "認知機能低下者割合", ""),
    ("koukukinou2gp_25", "口腔機能低下者割合", ""),
    ("iadl1_2gp_25", "IADL低下者割合",
     "男性が5.7ポイント高い。IADLは買い物・食事の用意・請求書の支払い等を含み、"
     "生活技能の差が表れている"),
    ("tento2gp_25", "1年間の転倒あり割合", "女性が2.8ポイント高い"),
    ("孤食", "孤食者割合", "男性が3.6ポイント高い"),
    ("独居", "独居者割合", "女性が9.3ポイント高い"),
    ("社会参加", "社会参加あり割合（6区分）", "女性が8.3ポイント高い"),
    ("通いの場", "通いの場参加割合",
     "女性13.3％に対し男性3.8％と3.5倍の差。通いの場が同規模保険者の3分の2に"
     "とどまる要因は、実質的に男性の不参加である"),
    ("運転あり", "自分で運転する割合",
     "男性87.1％に対し女性57.1％。女性の移動制約が大きい"),
    ("自力移動なし", "自力移動の手段なし割合", "女性が高い"),
    ("困りごとあり", "生活動作の困りごとあり割合", "女性が9.4ポイント高い"),
    ("世話をする人なし", "世話をしてくれる人がいない割合", "差は有意でない"),
    ("happy2gp_25", "幸福感がある者の割合", "女性が6.1ポイント高い"),
]
for key, nm, obs in SEXIND:
    vm, nm_ = rate(mM, key)
    vf, nf = rate(mF, key)
    if vm is None or vf is None:
        continue
    d = round(vm - vf, 1)
    jd = ztest(mM, mF, key)
    fl = {4: NG_O if jd == "有意" and abs(d) >= 5 else (IN_Y if jd == "有意" else None)}
    r = body(ws, r, [nm, vm, vf, d, jd, nm_, nf,
                     round(vm / vf, 2) if vf else "―", obs], fl, height=30)

r += 1
r = lead(ws, r, "【2　社会参加の区分別（性・年齢階級別）】", 9)
r = header(ws, r, ["区分", "男性\n65〜74歳", "男性\n75歳以上", "女性\n65〜74歳",
                   "女性\n75歳以上", "男性計", "女性計", "男女差", "所見"], height=44)
mY = A["年齢階級"].isin(["65～69歳", "70～74歳"])
mO = ~mY
for key, nm in [("cmnt6vl25", "ボランティア"), ("cmnt6sp25", "スポーツの会"),
                ("cmnt6hb25", "趣味の会"), ("cmnt6le25", "学習・教養サークル"),
                ("cmnt6sl25", "通いの場"), ("cmnt6sk25", "特技を伝える活動")]:
    if key not in A.columns:
        continue
    A["_t"] = flag_in(key, [1, 2, 3, 4])
    vm, _ = rate(mM, "_t")
    vf, _ = rate(mF, "_t")
    obs = "男性の参加が女性の3分の1以下" if key == "cmnt6sl25" else ""
    r = body(ws, r, [nm, rate(mM & mY, "_t")[0], rate(mM & mO, "_t")[0],
                     rate(mF & mY, "_t")[0], rate(mF & mO, "_t")[0], vm, vf,
                     round(vm - vf, 1), obs], height=18)
r = body(ws, r, ["【合成】6区分のいずれか", rate(mM & mY, "社会参加")[0],
                 rate(mM & mO, "社会参加")[0], rate(mF & mY, "社会参加")[0],
                 rate(mF & mO, "社会参加")[0], rate(mM, "社会参加")[0],
                 rate(mF, "社会参加")[0],
                 round(rate(mM, "社会参加")[0] - rate(mF, "社会参加")[0], 1), ""],
         {1: IN_Y}, height=18, bold=True)

r += 1
r = lead(ws, r, "【3　男性の非参加層の特性（社会参加なしの群の中での比較）】", 9)
r = header(ws, r, ["指標", "男性・参加なし", "女性・参加なし", "男性・参加あり",
                   "女性・参加あり", "", "", "", "所見"], height=40)
for key, nm in [("frail2gp_25", "フレイルあり割合"), ("tojikomori2gp_25", "閉じこもり者割合"),
                ("孤食", "孤食者割合"), ("utsucheck2gp_25", "うつ割合"),
                ("happy2gp_25", "幸福感がある者の割合"),
                ("困りごとあり", "生活動作の困りごとあり割合")]:
    r = body(ws, r, [nm, rate(mM & msoc0, key)[0], rate(mF & msoc0, key)[0],
                     rate(mM & msoc1, key)[0], rate(mF & msoc1, key)[0],
                     "", "", "", ""], height=18)

note(ws, r + 1,
     "注1）単位は％。性別は調査対象者名簿による。"
     "注2）本シートは、JAGESが提供するクロス集計表（CROSS_P・CROSS_N・CROSS_MIX）が"
     "性別を集計軸としていることを踏まえ、本分析に欠けていた性別の視点を補うものである"
     "（23シート）。"
     "注3）最も差が大きいのは通いの場参加割合（男性3.8％・女性13.3％）である。"
     "通いの場参加率が同規模保険者13.2％の3分の2にとどまる（19シート）ことの実質は、"
     "男性の不参加である。女性13.3％は同規模保険者の全体値に近く、"
     "男性を対象とした場づくりが第10期の課題となる。"
     "注4）自分で運転する割合は男性87.1％・女性57.1％で、女性の移動制約が大きい。"
     "一方、通いの場への参加は女性の方が高く、"
     "移動手段の有無だけでは参加の差を説明できない。"
     "注5）IADL低下者割合は男性が高い（12.8％対7.1％）。"
     "IADLの設問は買い物・食事の用意・請求書の支払い等を含むため、"
     "生活技能の差が表れていると考えられる。"
     "単身化が進むと男性の生活支援ニーズが顕在化する。", 9)

# ============================================================ 23 提供済み集計表
ws = sheet("23_提供済み集計表との関係", "JAGESが提供する集計表と本分析の関係",
           "健康とくらしの調査については、JAGESから単純集計表及びクロス集計表が"
           "既に提供されている。本分析はこれを代替するものではなく、"
           "提供済みの集計では得られない属性の掛け合わせを行うものである。"
           "両者の役割を整理する。",
           [26, 14, 40, 30, 30])

r = lead(ws, 4, "【1　提供済みの集計表】", 5)
r = header(ws, r, ["資料名", "形式", "内容", "集計軸（表側）", "本分析での扱い"])
for nm, fm, cont, ax, use in [
    ("KK_2025_GT", "Excel 1シート", "全設問の単純集計（実数と％）", "総数のみ",
     "本分析の集計値の照合に用いた。広域連合全体の主要20指標の値は一致する"),
    ("KK_2025_CROSS_P", "Excel 489シート", "全設問のクロス集計（％）",
     "①総数②性別③年齢5区分\n④性・年齢5歳刻み⑤性・年齢10歳刻み⑥分析地域",
     "町別・地区別・年齢階級別の値の照合に用いた。"
     "性別は本分析で用いていなかったため22シートを追加した"),
    ("KK_2025_CROSS_N", "Excel 489シート", "同（実数）", "同上", "同上"),
    ("KK_2025_CROSS_MIX", "Excel 489シート", "同（実数と％の併記）", "同上", "同上"),
    ("KK_2025_2次変数指標", "Excel 5ファイル", "リスク指標等の2次変数を地域別・年齢別に集計",
     "地域別（全体・3町）\n年齢別", "指標の定義の確認に用いた"),
    ("集計値一覧表", "Excel 3シート", "約130指標を小地域別に2025年・2022年・改善状況で表示",
     "小地域（小学校区）\n高齢者全体・前期・後期",
     "18シートの地域重点度の分析に用いた。2022年との比較が可能"),
    ("重点対象地域選定シート", "Excel 2ファイル", "70指標を小学校区別に5段階ランクで評価",
     "小地域（小学校区）", "18シートで本分析の地区別集計と突合した"),
    ("報告書", "PDF 8分冊", "調査の概要、年齢比較、地域比較、保険者独自項目、基礎集計表",
     "―", "同規模保険者との比較（19シート）、"
     "調査対象者の範囲の確認（14シート）に用いた"),
    ("サマリー（地域診断書）", "Excel 3ファイル", "主要12指標の参加自治体中の順位・平均・最小・最大",
     "町別", "20シートの順位分析に用いた"),
]:
    r = body(ws, r, [nm, fm, cont, ax, use], height=48)

r += 1
r = lead(ws, r, "【2　提供済み集計表で得られること・得られないこと】", 5)
r = header(ws, r, ["区分", "内容", "", "", "対応"])
for k, v, act in [
    ("得られること", "全設問の単純集計と、性別・年齢・地域を軸とした基本クロス集計。"
     "小地域別の指標値と2022年からの変化、参加自治体中の順位",
     "計画本文の記述と図表の大半はこれで足りる。"
     "本分析の01〜04・22シートは、この範囲を再整理し計画の章立てに合わせたものである"),
    ("得られないこと", "属性を掛け合わせたクロス集計。"
     "例えば「社会参加の有無別のフレイル該当割合」「介護・介助が必要な者の世帯構成」"
     "「施設入所志向のある者の支援者の有無」は、提供済みの集計表からは算定できない",
     "本分析の05〜10・15〜18シートはこれに当たり、"
     "個票データからでなければ得られない"),
    ("得られないこと（2）", "年齢調整、信頼区間、群間差の検定、合成指標",
     "本分析の02・14シートで算定した。"
     "特に年齢調整は町間比較の妥当性に関わる"),
    ("留意点", "提供済みのクロス集計表は「無回答」を分母に含む形式であり、"
     "本分析の分母（有効回答者）と一致しない設問がある",
     "本分析はすべての表に分母を併記している。"
     "計画本文で数値を引用する際は、分母の定義を統一する"),
]:
    r = body(ws, r, [k, v, "", "", act], height=64)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=4)

note(ws, r + 1,
     "注1）提供済みの集計表は成果品「分析資料、推計資料、図表、グラフ、クロス集計表等"
     "電子データ」の一部として、そのまま納品物に含めることができる。"
     "本分析はこれに追加する分析資料として位置づける。"
     "注2）本分析の作成にあたり、集計値一覧表及び重点対象地域選定シートの活用が"
     "不十分であった。集計値一覧表には2022年との比較（改善状況）が"
     "約130指標について小地域別・年齢層別に収録されているため、"
     "第9期からの変化の分析に活用できる。今後の分析で反映する。"
     "注3）性別を集計軸とした分析が本分析に欠けていたため、22シートを追加した。", 5)


del wb["Sheet"]
wb.save(OUT)
print("saved:", OUT, "sheets=%d" % len(wb.sheetnames))
for _s in wb.sheetnames:
    print("  -", _s, wb[_s].max_row, "rows")
