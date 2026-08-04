"""見える化システムからの追加出力依頼リストを生成する。

小野町 第10期介護保険事業計画の策定に必要な指標を、見える化システムの
指標リスト（目的別・905指標）と、既に取得済みの指標との突合により特定する。

入力:
  - 指標リスト（目的別）……見える化システムの指標一覧
  - 07_介護保険_見える化整理 の整理2ブック（90_取込確認シート）
出力:
  - 11_見える化出力依頼/小野町_見える化システム出力依頼リスト_20260804.xlsx
"""

import pathlib
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent
SRC_LIST = ROOT / "小野町_引継ぎ_整理済" / "11_見える化出力依頼" / "見える化_指標リスト_目的別.xlsx"
MIERUKA = ROOT / "小野町_引継ぎ_整理済" / "07_介護保険_見える化整理"
OUT = ROOT / "小野町_引継ぎ_整理済" / "11_見える化出力依頼"


def load_master():
    wb = openpyxl.load_workbook(SRC_LIST, data_only=True)
    ws = wb["指標リスト"]
    rows = list(ws.iter_rows(values_only=True))
    h = next(i for i, r in enumerate(rows) if r and r[0] == "目的")
    recs, cur = [], ["", "", "", "", ""]
    for r in rows[h + 1:]:
        if not r or all(x is None for x in r):
            continue
        v = [("" if x is None else str(x).strip()) for x in r[:7]]
        for j in range(5):
            if v[j]:
                cur[j] = v[j]
        if v[6]:
            recs.append({"目的": cur[0], "観点": cur[1], "大分類": cur[2],
                         "中分類": cur[3], "小分類": cur[4], "名称": v[5], "ID": v[6]})
    return recs


def load_held():
    """整理2ブックの取込確認シートから、取込済みの指標IDを取得する。"""
    held = set()
    for f in MIERUKA.glob("*.xlsx"):
        wb = openpyxl.load_workbook(f, data_only=True)
        if "90_取込確認" not in wb.sheetnames:
            continue
        for row in wb["90_取込確認"].iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.endswith(".xlsx"):
                    m = re.match(r"^([A-Zδ][0-9]{1,2}(?:-[a-z])?)_", pathlib.Path(v).name)
                    if m:
                        held.add(m.group(1))
    return held


def load_memo_confirmed():
    """確認メモで小野町データと確認済みだが、整理ブックには未取込の系列。"""
    txt = "".join(p.read_text() for p in MIERUKA.glob("*.md"))
    return set(re.findall(r"(?<![A-Za-z])([BCDKδ][0-9]{1,2}(?:-[a-z])?)(?![0-9])", txt))


# 依頼する指標の定義（優先度、用途、対象IDの選び方）
REQUESTS = [
    # (優先度, 区分, 用途, マッチ条件)
    ("S", "認定データの検証",
     "令和8年3月末の認定者数が前年比▲16.3％と実態として説明困難な減少を示している。"
     "認定率（B4系）とは別系列の認定者数実数（B3系）及び調整済み認定率（B5系）を取得し、"
     "突合して異常の有無を判定する。第10期の認定者推計・給付費推計・保険料算定の出発点",
     lambda r: r["ID"] in {"B3-a", "B3-b", "B3-c", "B3-d", "B3-e", "B5-a", "B5-c"}),

    ("A", "保険料算定・介護保険財政",
     "第10期保険料の算定に必要。歳入歳出の実績から、保険料収納、国庫支出金、支払基金交付金、"
     "都道府県支出金、繰入金、基金積立金の推移を把握する。介護給付費準備基金の残高推移の"
     "確認にも用いる。現在まったく取得していない",
     lambda r: r["大分類"] == "介護保険特別会計経理状況"),

    ("A", "供給体制・施設整備の検討",
     "施設・居住系・通所系の定員と、要支援・要介護者1人あたり定員を把握する。"
     "保険料算定の施設整備ケースの前提、待機・供給不足の判定に用いる。"
     "施設・居住系は給付の49.9％を占め自給率も59.74％と最低であり、"
     "供給体制の把握が第10期の最重要論点のひとつ。現在まったく取得していない",
     lambda r: r["大分類"] == "入所（利用）定員"),

    ("A", "介護予防・総合事業（第9期評価の核心）",
     "通いの場の参加率・箇所数・運営主体・活動内容、訪問型／通所型サービスA〜D、見守り・配食、"
     "介護予防ケアマネジメントの実績。調整済み軽度認定率が10.6％から15.2％へ上昇しており、"
     "第9期の介護予防施策の到達状況を評価するために必須。現在まったく取得していない",
     lambda r: r["大分類"] == "介護予防・日常生活支援総合事業"),

    ("A", "令和7年度 在宅介護実態調査",
     "仕様書4(1)③が反映を求める調査。特にZ53「介護保険サービスの未利用の理由」は、"
     "認定率の地域差指数1.24に対し受給率1.07にとどまる「認定と利用のギャップ」の"
     "要因特定に直結する。主な介護者の年齢・就労継続・介護離職は家族介護者支援の根拠となる。"
     "町から生データの提供を受ける場合でも、見える化上の集計値と突合する",
     lambda r: r["大分類"] == "在宅介護実態調査"),

    ("A", "令和7年度 介護予防・日常生活圏域ニーズ調査",
     "仕様書4(1)③が反映を求める調査。各種リスクを有する割合（運動器機能低下、転倒、閉じこもり、"
     "認知機能低下、社会的役割の低下等）は、介護予防施策の対象把握と成果指標の設定に用いる。"
     "まず E1〜E7 系（各種リスクを有する割合）を優先し、回答項目別は必要に応じて追加する",
     lambda r: (r["大分類"] == "介護予防・日常生活圏域ニーズ調査"
                and r["中分類"] == "各種リスクを有する割合")),

    ("B", "認知症施策（仕様書が計画への包含を要求）",
     "仕様書2(3)は認知症施策推進基本計画に基づく市町村計画の包含を求めている。"
     "初期集中支援チームの訪問実績、地域支援推進員の配置、認知症カフェの設置箇所数、"
     "各種研修の実施状況を把握する。認知症高齢者自立度Ⅱa は68人から148人、"
     "M は7人から29人に増加しており、施策の柱として位置づける必要がある",
     lambda r: r["大分類"] == "認知症施策"),

    ("B", "地域包括支援センター・生活支援体制整備",
     "センターの設置数・人員体制（3職種）、地域ケア会議の開催回数、生活支援コーディネーター、"
     "協議体の状況。居宅介護支援事業所が5から3に減少しており、"
     "相談支援・ケアマネジメント基盤の評価に用いる",
     lambda r: (r["大分類"] in {"地域包括支援センター", "包括的支援事業（社会保障充実分）"})),

    ("B", "認定の詳細（新規認定・自立度）",
     "認知症高齢者自立度・障害高齢者自立度の状況、新規認定者の要介護度別分布・平均要介護度・"
     "年齢階級別分布・平均年齢。確認メモに値の記載はあるが整理ブックに未取込であり、"
     "認定者推計と介護予防施策の設計に用いる",
     lambda r: r["ID"] in {"B7", "B8", "B9", "B10", "B11", "B13", "B14"}),

    ("B", "介護人材",
     "介護人材の必要数と、需要見込みに対する供給見込みの割合。"
     "訪問介護2事業所・訪問看護1事業所・居宅介護支援3事業所という供給状況のもとで、"
     "第10期のサービス見込量の実現可能性を検討するために用いる",
     lambda r: r["大分類"] == "介護人材の必要数"),

    ("C", "在宅医療・介護連携",
     "在宅医療・介護連携推進事業の実施状況。居宅（医療系）サービス事業所数の偏差値が41.49と"
     "低く、訪問看護は1事業所である。在宅医療・看取りの体制を記載するために用いる。"
     "指標数が多いため、必要な項目を選択して出力する",
     lambda r: r["大分類"] == "在宅医療・介護連携推進事業"),

    ("C", "医療提供体制",
     "医療機関数、医師数、患者数。町内の医療資源が限られ町外通院が前提となる状況の"
     "裏付けに用いる。指標数が多いため、必要な項目を選択して出力する",
     lambda r: r["大分類"] == "医療機関、医師、患者数"),

    ("C", "リハビリテーション提供体制",
     "リハビリ専門職の関与状況。介護予防・重度化防止の施策設計に用いる。"
     "指標数が多いため、必要な項目を選択して出力する",
     lambda r: r["大分類"] == "リハビリテーション提供体制"),

    ("C", "保険者機能強化推進交付金の評価指標",
     "交付金の評価指標は、第9期の取組状況を国の評価軸で点検できる。"
     "第9期評価と第10期の重点施策の妥当性検証に用いる。"
     "指標数が多いため、評価点の低い項目を中心に選択して出力する",
     lambda r: r["大分類"] == "保険者機能強化推進交付金 ・介護保険保険者努力支援交付金 に係る評価指標"),
]

NOTE_ROWS = [
    ("表示モード", "「注目する地域を時系列で見る」を基本とする。全国・福島県・県中圏域・"
                "類似団体との比較が必要な指標は「注目する地域と他を時系列で見る」も併せて出力する"),
    ("調整済み指標の注意", "B5・B6・B12等の調整済み指標は、表示モードにより標準化に用いる"
                   "性・年齢構成が異なる。出力時にモードをファイル名または備考に明記する"),
    ("期間", "第9期評価に用いるため、取得可能な最新年度までの全期間を時系列で出力する"),
    ("ファイル名", "指標IDのプレフィックスを保持する。同一指標の重複出力（(1)(2)等）は避け、"
              "やむを得ず複数になる場合は採用版を明示する"),
    ("格納", "大分類ごとにフォルダを分けて格納する。Downloads直下に置かない"),
    ("自治体の確認", "出力前に対象自治体が「小野町」（地域コード07522）であることを確認する。"
               "過去に金ヶ崎町のデータが混在した経緯がある"),
]


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    recs = load_master()
    held = load_held()
    memo = load_memo_confirmed()

    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    s_fill = PatternFill("solid", fgColor="FCE4E4")
    a_fill = PatternFill("solid", fgColor="FFF2CC")

    # ---- 00_依頼概要
    ws = wb.active
    ws.title = "00_依頼概要"
    ws.append(["小野町 第10期介護保険事業計画　見える化システム 出力依頼リスト"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["作成日", "2026-08-04"])
    ws.append(["対象", "地域包括ケア「見える化」システム　小野町（地域コード07522）"])
    ws.append(["突合方法", "見える化システムの指標リスト（目的別）905指標と、"
                      "07_介護保険_見える化整理の整理2ブックの取込確認シートを突合"])
    ws.append([])
    ws.append(["優先度", "区分", "依頼指標数", "用途"])
    for c in ws[7]:
        c.font = Font(bold=True)
        c.fill = head_fill

    detail = []
    for pri, cat, use, cond in REQUESTS:
        hit = [r for r in recs if cond(r)]
        need = [r for r in hit if r["ID"] not in held]
        ws.append([pri, cat, len(need), use])
        for r in need:
            detail.append({
                "優先度": pri, "区分": cat, "ID": r["ID"], "名称": r["名称"],
                "大分類": r["大分類"], "中分類": r["中分類"], "小分類": r["小分類"],
                "保有状況": ("確認メモで確認済（整理ブック未取込）"
                         if r["ID"].split("-")[0] in memo else "未取得"),
                "用途": use,
            })
    ws.append([])
    ws.append(["合計", "", len(detail), ""])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        if row[0].value == "S":
            for c in row:
                c.fill = s_fill
        elif row[0].value == "A":
            for c in row:
                c.fill = a_fill
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", (8, 34, 12, 96)):
        ws.column_dimensions[col].width = w

    # ---- 01_依頼指標一覧
    ws2 = wb.create_sheet("01_依頼指標一覧")
    cols = ["優先度", "区分", "指標ID", "指標名", "大分類", "中分類", "小分類", "保有状況"]
    ws2.append(cols)
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
    for d in detail:
        ws2.append([d["優先度"], d["区分"], d["ID"], d["名称"],
                    d["大分類"], d["中分類"], d["小分類"], d["保有状況"]])
    for row in ws2.iter_rows(min_row=2):
        if row[0].value == "S":
            for c in row:
                c.fill = s_fill
        elif row[0].value == "A":
            for c in row:
                c.fill = a_fill
    for col, w in zip("ABCDEFGH", (8, 34, 11, 56, 32, 30, 26, 30)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws2.max_row}"

    # ---- 02_取得済み指標
    ws3 = wb.create_sheet("02_取得済み指標")
    ws3.append(["指標ID", "指標名", "大分類", "中分類", "備考"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
    byid = {r["ID"]: r for r in recs}
    for i in sorted(held):
        r = byid.get(i, {})
        ws3.append([i, r.get("名称", "（指標リストに該当なし）"),
                    r.get("大分類", ""), r.get("中分類", ""), "整理ブックに取込済"])
    for col, w in zip("ABCDE", (11, 56, 32, 30, 22)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    # ---- 03_出力の留意点
    ws4 = wb.create_sheet("03_出力の留意点")
    ws4.append(["項目", "内容"])
    for c in ws4[1]:
        c.font = Font(bold=True)
        c.fill = head_fill
    for k, v in NOTE_ROWS:
        ws4.append([k, v])
    for row in ws4.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 104

    path = OUT / "小野町_見える化システム出力依頼リスト_20260804.xlsx"
    wb.save(path)
    print("出力:", path)
    print(f"  依頼指標 {len(detail)}件 / 取得済 {len(held)}件 / マスタ {len(recs)}件")
    from collections import Counter
    for k, v in sorted(Counter(d["優先度"] for d in detail).items()):
        print(f"    優先度{k}: {v}件")


if __name__ == "__main__":
    main()
