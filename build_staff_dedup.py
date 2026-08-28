# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　従業員数の重複計上の整理.

令和8年8月26日の進捗確認の打合せでのご指示に対応する。

  同一の施設が複数のサービスで登録されている場合、同じ職員が
  サービスごとに計上されている。供給制約の推定に用いる従業員数を
  どの単位で数えるかを整理する。

  打合せでは次の2つの方法が挙げられた。
    ① 受託者が重複しない方法を提案し、発注者がご了承する
    ② 重複の候補の一覧を受託者が示し、発注者が実態をご記入いただく

  本表は①と②の両方を用意し、②の記入欄を設けたものである。

シート構成
  00_この表について
  01_重複の候補（ご確認をお願いする組）
  02_数え方の3案と比較
  03_供給制約の推定への影響
"""

import os
import re
import sys
from collections import defaultdict
from itertools import combinations

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

ODIR = "/home/user/repository/output"
SRC = os.path.join(ODIR, "第10期計画_3町の社会資源一覧との突合.xlsx")
OUT = os.path.join(ODIR, "第10期計画_従業員数の重複計上の整理.xlsx")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 50
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=30):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=26, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and i in fills:
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=84):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


# ---------------------------------------------------------------- 元データ
def load():
    src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = src["01_受領した一覧"]
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not isinstance(row[0], str):
            continue
        if row[0] not in ("東川町", "美瑛町", "東神楽町"):
            continue
        town, _no, name, houjin, svc, tot, jo, hi = row[:8]
        try:
            tot = float(tot)
        except (TypeError, ValueError):
            tot = None
        out.append({"町": town, "事業所名": (name or "").strip(),
                    "法人": (houjin or "").strip(), "サービス": (svc or "").strip(),
                    "総数": tot, "常勤": jo, "非常勤": hi})
    return out


ROWS = load()
NOBE = sum(r["総数"] or 0 for r in ROWS)

# 同一町・同一法人で従業員総数が一致する組を重複の候補とする
PAIRS = []
g = defaultdict(list)
for r in ROWS:
    g[(r["町"], r["法人"])].append(r)


def norm(nm):
    """施設名の表記ゆれを吸収する（空白・全角半角・付属語）。"""
    nm = re.sub(r"[\s　]", "", nm)
    for w in ("指定", "事業所", "ステーション", "センター",
              "（介護予防）", "介護予防"):
        nm = nm.replace(w, "")
    return nm


for (t, h), v in sorted(g.items()):
    used = set()
    for a, b in combinations(v, 2):
        if not a["総数"] or a["総数"] != b["総数"]:
            continue
        k = (a["事業所名"], b["事業所名"], a["総数"])
        if k in used:
            continue
        used.add(k)
        na, nb = norm(a["事業所名"]), norm(b["事業所名"])
        # 短い方の名称が長い方に含まれるとき、同一施設とみなす
        # （例「東川町羽衣園」と「特別養護老人ホーム東川町羽衣園」）
        short, long = sorted((na, nb), key=len)
        same = len(short) >= 4 and short in long
        PAIRS.append({"町": t, "法人": h, "A": a, "B": b,
                      "人数": a["総数"], "同名": same})

DUP_CLEAR = [p for p in PAIRS if p["同名"]]
DUP_CHECK = [p for p in PAIRS if not p["同名"]]
N_CLEAR = sum(p["人数"] for p in DUP_CLEAR)
N_CHECK = sum(p["人数"] for p in DUP_CHECK)

# ============================================================ 00
ws = sheet("00_この表について", "従業員数の重複計上の整理",
           "3町からご提供いただいた社会資源一覧（74事業所）は、"
           "同一の施設が複数のサービスで登録されている場合、"
           "同じ職員がサービスごとに計上されています。"
           "供給制約の推定に用いる従業員数をどの単位で数えるかを整理します。"
           "令和8年8月26日の進捗確認の打合せでのご指示に対応するものです。",
           [16, 22, 56, 42])

r = lead(ws, 4, "【1　何が起きているか】", 4)
r = note(ws, r,
         "介護保険の事業所は、一つの建物（施設）で複数のサービスの指定を"
         "受けていることがあります。"
         "例えば、介護老人保健施設と短期入所療養介護、"
         "介護老人福祉施設と短期入所生活介護は、"
         "同じ建物・同じ職員でサービスを提供しています。\n"
         "社会資源一覧はサービスごとに1行で記載されているため、"
         "こうした施設では同じ職員が2回計上されます。\n"
         "このまま合計すると延べ %.1f 人となり、"
         "実際に働いている人数（実人数）より多くなります。"
         % NOBE, 4, height=110)

r += 1
r = lead(ws, r, "【2　確認した結果】", 4)
r = header(ws, r, ["区分", "組数", "人数", "内容"])
for a in [
    ("重複が明らかな組", "%d組" % len(DUP_CLEAR), "%.0f人" % N_CLEAR,
     "同一法人・同一施設名で、従業員総数が一致するもの。"
     "介護老人保健施設と短期入所療養介護、"
     "介護老人福祉施設と短期入所生活介護の組合せ。"
     "同じ建物・同じ職員と考えられます。"),
    ("確認を要する組", "%d組" % len(DUP_CHECK), "%.0f人" % N_CHECK,
     "同一法人だが施設名が異なり、従業員総数が一致するもの。"
     "同じ職員が兼務しているのか、たまたま同数なのかを"
     "実態でご確認いただく必要があります。"),
    ("計", "%d組" % len(PAIRS), "%.0f人" % (N_CLEAR + N_CHECK),
     "延べ %.1f 人の %.1f％に相当します。" % (NOBE, (N_CLEAR + N_CHECK) / NOBE * 100)),
]:
    r = body(ws, r, list(a), {1: NG_O if "明らか" in a[0] else
                              (IN_Y if "確認" in a[0] else MID_B)},
             height=72, align={2: "center", 3: "center"},
             bold=(a[0] == "計"))

r += 1
r = lead(ws, r, "【3　受託者の案】", 4)
r = note(ws, r,
         "供給制約の推定には、施設（建物）を単位とする実人数を用いることを"
         "案といたします。\n"
         "理由は次のとおりです。\n"
         "① 供給の制約は、指定の数ではなく、実際に働いている職員の数で決まります。"
         "同じ職員が入所・通所・短期入所を担当している場合、"
         "3サービス分の受入れが同時にできるわけではありません。\n"
         "② サービス別の従業員数を単純に合計すると、"
         "職員が実際より多くいるように見え、"
         "供給に余力があるという誤った推定につながります。\n"
         "③ 一方、サービス別の分析（どのサービスに何人配置されているか）には、"
         "登録どおりの数（延べ）が必要です。"
         "したがって延べと実人数の両方を保持し、用途によって使い分けます。\n"
         "\n"
         "01シートに重複の候補を掲げています。"
         "「確認を要する組」4組について、"
         "同じ職員かどうかを「ご確認欄」にご記入いただければ、"
         "実人数を確定できます。"
         "ご記入が難しい場合は、受託者の案（同じ職員として扱う）で"
         "進めさせていただきます。", 4, height=190)

r += 1
r = note(ws, r,
         "注1）本表は3町からご提供いただいた社会資源一覧によるものです。"
         "介護サービス情報公表システム及び北海道の介護保険事業所一覧とは"
         "「従業員」の範囲が異なります"
         "（社会資源一覧203.5人／公表システムの総従業者184.5人／"
         "計画本文の訪問介護員等181人。いずれも訪問介護）。\n"
         "注2）本表は重複の可能性を機械的に抽出したものであり、"
         "重複であると判定したものではありません。", 4, height=76)

# ============================================================ 01
ws = sheet("01_重複の候補", "重複の候補（ご確認をお願いする組）",
           "同一町・同一法人で従業員総数が一致する組を機械的に抽出しました。"
           "「ご確認欄」に、同じ職員かどうかをご記入ください。"
           "「同じ」の場合は片方を実人数から除きます。",
           [5, 7, 22, 30, 22, 7, 12, 24, 16])

r = header(ws, 4, ["No.", "町", "法人等の名称", "事業所名", "介護サービスの種類",
                   "従業員\n総数", "受託者の判定",
                   "判定の理由", "ご確認欄\n（同じ／別）"])
n = 0
for p in DUP_CLEAR + DUP_CHECK:
    n += 1
    hantei = "重複（同一施設）" if p["同名"] else "要確認"
    riyu = ("同一法人・同一施設名で、入所系と短期入所系の組合せ。"
            "同じ建物・同じ職員と考えられます。" if p["同名"] else
            "同一法人だが施設名が異なります。"
            "兼務によるものか、たまたま同数かの判別ができません。")
    fill = {7: NG_O if p["同名"] else IN_Y, 9: OK_G}
    for k, rec in (("A", p["A"]), ("B", p["B"])):
        first = k == "A"
        r = body(ws, r, [n if first else "", p["町"] if first else "",
                         p["法人"] if first else "",
                         rec["事業所名"], rec["サービス"], rec["総数"],
                         hantei if first else "", riyu if first else "", ""],
                 fill if first else {9: OK_G}, height=30,
                 align={1: "center", 2: "center", 6: "center", 7: "center"})

r += 1
r = body(ws, r, ["", "", "計", "%d組" % len(PAIRS), "",
                 N_CLEAR + N_CHECK, "", "", ""],
         {j: GRAY for j in range(1, 10)}, height=20, bold=True,
         align={4: "center", 6: "center"})

r += 1
r = note(ws, r,
         "注1）「ご確認欄」には「同じ」又は「別」とご記入ください。"
         "「同じ」の場合、実人数の算定において片方を除きます。\n"
         "注2）「重複（同一施設）」と判定した組についても、"
         "実態と異なる場合はご指摘ください。\n"
         "注3）本表に挙げていない組でも、"
         "従業員総数が一致しないだけで同じ職員が兼務している場合があります。"
         "お気づきの点があればご教示ください。\n"
         "注4）東神楽町の「東神楽町ケアプラン相談センター」（居宅介護支援10人）と"
         "「東神楽町ホームヘルプサービスセンター」（訪問介護10人）は、"
         "サービスの性質が異なるため、"
         "同数であっても別の職員である可能性があります。", 9, height=110)

# ============================================================ 02
ws = sheet("02_数え方の3案", "従業員数の数え方の3案",
           "用途によって適切な数え方が異なります。"
           "3案を比較し、用途ごとの使い分けを示します。",
           [6, 20, 12, 44, 44])

r = header(ws, 4, ["案", "数え方", "人数", "長所", "短所"])
for a in [
    ("案1", "延べ（登録どおり）", "%.1f人" % NOBE,
     "社会資源一覧の記載をそのまま用いるため、加工を要しません。"
     "サービス別の配置状況の分析に適します。",
     "同じ職員が複数回数えられます。"
     "実際に働いている人数より多くなり、"
     "供給に余力があるという誤った推定につながります。"),
    ("案2", "施設（建物）単位の実人数\n〈受託者の案〉",
     "%.1f人（推計）" % (NOBE - N_CLEAR - N_CHECK),
     "実際に働いている人数に近づきます。"
     "供給制約の推定（受け入れられる利用者数の上限）に適します。",
     "どの組が同じ職員かの確認を要します。"
     "サービス別の内訳が得られなくなります。"),
    ("案3", "案分（サービス数で割る）",
     "%.1f人" % (NOBE - (N_CLEAR + N_CHECK) / 2),
     "サービス別の内訳を保ったまま、総数を実人数に近づけられます。",
     "案分の根拠がありません。"
     "実際の配分（入所8割・短期入所2割など）は施設により異なり、"
     "均等に割る前提が成り立ちません。"),
]:
    r = body(ws, r, list(a), {1: MID_B, 3: OK_G if a[0] == "案2" else GRAY},
             height=86, align={1: "center", 3: "center"})

r += 1
r = lead(ws, r, "【用途ごとの使い分け（受託者の案）】", 5)
r = header(ws, r, ["用途", "用いる数え方", "理由", "", ""])
for a, b, c in [
    ("供給制約の推定（第6章第2節・第4節）", "案2　施設単位の実人数",
     "同じ職員が複数のサービスを担当している場合、"
     "サービス分の受入れが同時にできるわけではないため。"),
    ("サービス別の配置状況の分析（第2章第3節）", "案1　延べ",
     "どのサービスに何人が登録されているかを示すため。"
     "延べであることを注記する。"),
    ("代表KPI H13（職種別従事者数の推移）", "見える化M2系列",
     "全国・北海道と比較するため、"
     "国が共通の方法で集計した系列を用いる。"
     "社会資源一覧とは範囲が異なるため置き換えない。"),
    ("介護人材の確保に関する記述（第5章基本目標4）", "案2　施設単位の実人数",
     "人材の不足は実際に働いている人数で測るため。"),
]:
    r = body(ws, r, [a, b, c, "", ""], {2: OK_G}, height=50)
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=5)

r += 1
r = note(ws, r,
         "注）案分（案3）は採用しないことを案とします。"
         "施設ごとに配分の実態が異なり、"
         "均等に割る前提を置く根拠がないためです。"
         "配分の実態をご提供いただける場合は、案3も選択できます。", 5, height=48)

# ============================================================ 03
ws = sheet("03_推定への影響", "供給制約の推定への影響",
           "数え方により、供給制約の推定がどれだけ変わるかを示します。"
           "第6章のサービス見込量の算定に用いる値に影響します。",
           [24, 14, 14, 14, 46])

r = header(ws, 4, ["区分", "案1\n延べ", "案2\n実人数", "差", "推定への影響"])
BYTOWN = defaultdict(float)
for x in ROWS:
    BYTOWN[x["町"]] += x["総数"] or 0
DUPTOWN = defaultdict(float)
for p in PAIRS:
    DUPTOWN[p["町"]] += p["人数"]
for t in ("東川町", "美瑛町", "東神楽町"):
    a, d = BYTOWN[t], DUPTOWN[t]
    r = body(ws, r, [t, "%.1f人" % a, "%.1f人" % (a - d), "▲%.0f人" % d,
                     "延べの %.1f％が重複の候補です。"
                     "実人数を用いると、受け入れられる利用者数の上限が"
                     "小さくなります。" % (d / a * 100 if a else 0)],
             {}, height=44, align={2: "right", 3: "right", 4: "right"})
r = body(ws, r, ["3町計", "%.1f人" % NOBE,
                 "%.1f人" % (NOBE - N_CLEAR - N_CHECK),
                 "▲%.0f人" % (N_CLEAR + N_CHECK),
                 "延べの %.1f％。"
                 % ((N_CLEAR + N_CHECK) / NOBE * 100)],
         {j: GRAY for j in range(1, 6)}, height=26, bold=True,
         align={2: "right", 3: "right", 4: "right"})

r += 1
r = lead(ws, r, "【この整理が影響する箇所】", 5)
r = header(ws, r, ["箇所", "現在の記載", "整理後", "", ""])
for a, b, c in [
    ("計画素案 第2章第3節（サービス提供体制）",
     "社会資源一覧による従業員総数（延べ）",
     "延べと実人数を併記し、延べである旨を注記する"),
    ("計画素案 第6章第2節（サービス見込量）",
     "定員による制約のみを考慮",
     "実人数による職員の制約を加えて検討する"),
    ("3町の社会資源一覧との突合 03シート",
     "「合計は延べ人数である…実人数を得るには施設単位での集約を要する"
     "［要確認］」",
     "本表により［要確認］を解消する"),
    ("事業所調査の照会票と確定値管理表",
     "確定値表No.1・No.4（介護職員の総数）は重複計上13人の扱いが未確定",
     "本表の整理と併せて確定する"),
]:
    r = body(ws, r, [a, b, c, "", ""], {}, height=48)
    ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=5)

r += 1
r = note(ws, r,
         "注1）実人数は、01シートの「要確認」4組をすべて重複とみなした場合の値です。"
         "ご確認の結果により変わります。\n"
         "注2）本表の重複は、社会資源一覧における従業員総数の一致により"
         "抽出したものです。従業員総数が一致しない場合でも、"
         "一部の職員が兼務している場合があります。"
         "その場合の把握には、施設ごとの勤務実態の確認を要します。\n"
         "注3）供給制約の推定には、職員数のほか、定員、稼働率、"
         "受入困難の有無（資料提供依頼No.7）を併せて用います。", 5, height=94)

wb.save(OUT)
print("saved:", os.path.basename(OUT), "sheets=%d" % len(wb.sheetnames))
print("事業所 %d／延べ %.1f人" % (len(ROWS), NOBE))
print("重複が明らかな組 %d組 %.0f人／確認を要する組 %d組 %.0f人"
      % (len(DUP_CLEAR), N_CLEAR, len(DUP_CHECK), N_CHECK))
print("実人数（推計）%.1f人（延べの %.1f％減）"
      % (NOBE - N_CLEAR - N_CHECK, (N_CLEAR + N_CHECK) / NOBE * 100))
