# -*- coding: utf-8 -*-
"""
北塩原村 第8期障がい福祉計画・第4期障がい児福祉計画
管理ブック共通モジュール（書式・基礎データ）

- build_kitashiobara_projection.py（将来推計ブック）
- build_kitashiobara_finance.py（財源構成案ブック）
から参照する。
"""

import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/home/user/repository/output"

# ============================================================
# 配色（build_excel.py の規約に準拠。障がい計画＝オレンジ）
# ============================================================
COLORS = {
    "障がい":  "FF7F0E",
    "header":  "1F3864",
    "subhead": "2E75B6",
    "band":    "DDEBF7",
    "alt":     "F7FAFC",
    "note":    "FFF3F3",
    "input":   "FFF2CC",   # 村データ入力欄
    "calc":    "E2EFDA",   # 自動計算欄
    "white":   "FFFFFF",
}

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT = "游ゴシック"

# 状態ラベルの色分け
STATUS_FILL = {
    "確定":       "2CA02C",
    "暫定":       "ED7D31",
    "村資料待ち": "C00000",
    "対応済み":   "2CA02C",
    "一部対応":   "ED7D31",
    "要確認":     "C00000",
}


# ============================================================
# 書式ヘルパー
# ============================================================
def style_title(cell, text, fill=COLORS["header"], font_color="FFFFFF", size=14):
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)


def style_note(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=9, color="404040")
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)


def style_header_row(ws, row, headers, fill=COLORS["subhead"]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def style_data_cell(cell, alt=False, align="left", numfmt=None, fill=None):
    cell.font = Font(name=FONT, size=10)
    cell.alignment = Alignment(vertical="center", horizontal=align,
                               wrap_text=(align == "left"), indent=(1 if align == "left" else 0))
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    elif alt:
        cell.fill = PatternFill("solid", fgColor=COLORS["alt"])
    if numfmt:
        cell.number_format = numfmt


def write_row(ws, row, values, alt=False, aligns=None, numfmts=None, fills=None):
    """1行分をまとめて書き込む。aligns/numfmts/fills は列数分のリストまたは None。"""
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        style_data_cell(
            c,
            alt=alt,
            align=(aligns[col - 1] if aligns else "left"),
            numfmt=(numfmts[col - 1] if numfmts else None),
            fill=(fills[col - 1] if fills else None),
        )
    return row + 1


def style_status(cell):
    """状態セルを色分けする。"""
    fill = STATUS_FILL.get(str(cell.value))
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_sheet(wb, title, heading, note, widths):
    """見出し・説明・列幅を設定したシートを作る。データ開始行を返す。"""
    ws = wb.create_sheet(title)
    set_col_widths(ws, widths)
    ncol = len(widths)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    style_title(ws["A1"], heading)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    style_note(ws["A2"], note)
    ws.row_dimensions[2].height = 30
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    return ws


def ensure_out_dir():
    os.makedirs(OUT_DIR, exist_ok=True)


# ============================================================
# 基礎データ
# ============================================================

# --- 人口（住民基本台帳、各年4月1日現在）------------------------------
# 令和11年は北塩原村こども・子育て計画のコーホート変化率法による公式推計。
# 令和2〜4年、令和7〜10年は空欄（村資料受領後に入力／線形補間）。
POPULATION = [
    # 和暦, 西暦, 区分, 総人口, 年少, 生産年齢, 老年, 出典・備考
    ("平成31年", 2019, "実績", 2743, 284, 1495, 964,
     "現行計画本編・関連計画より"),
    ("令和2年", 2020, "実績", None, None, None, None,
     "村資料待ち"),
    ("令和3年", 2021, "実績", None, None, None, 1006,
     "老年人口はこの年（1,006人）をピークに減少に転じている"),
    ("令和4年", 2022, "実績", None, None, None, None,
     "村資料待ち"),
    ("令和5年", 2023, "実績", 2443, 205, 1241, 997,
     "現行計画本編より。手帳所持率の基準年"),
    ("令和6年", 2024, "実績", 2394, 180, 1223, 991,
     "こども・子育て計画 第2章"),
    ("令和7年", 2025, "補間推計", None, None, None, None,
     "令和6年実績と令和11年公式推計の線形補間"),
    ("令和8年", 2026, "補間推計", None, None, None, None,
     "同上"),
    ("令和9年", 2027, "補間推計", None, None, None, None,
     "同上／次期計画1年目"),
    ("令和10年", 2028, "補間推計", None, None, None, None,
     "同上／次期計画2年目"),
    ("令和11年", 2029, "公式推計", 2185, 111, 1023, 1051,
     "こども・子育て計画 コーホート変化率法／次期計画目標年度"),
]

# --- 障がい者手帳所持者数（各年4月1日現在）----------------------------
# 平成31年〜令和5年は現行計画本編。令和6〜8年は村資料待ち（入力欄）。
TEGATA_YEARS = [("平成31年", 2019), ("令和2年", 2020), ("令和3年", 2021),
                ("令和4年", 2022), ("令和5年", 2023), ("令和6年", 2024),
                ("令和7年", 2025), ("令和8年", 2026)]

TEGATA = {
    "身体障害者手帳":         [126, 121, 128, 122, 114, None, None, None],
    "療育手帳":               [13, 13, 14, 12, 12, None, None, None],
    "精神障害者保健福祉手帳": [24, 28, 27, 26, 30, None, None, None],
}

# --- 給付実績（受領データ「【北塩原村】障がいサービス給付実績.xlsx」整理版）---
BENEFIT_YEARS = [("令和2年度", 2), ("令和3年度", 3), ("令和4年度", 4),
                 ("令和5年度", 5), ("令和6年度", 6), ("令和7年度", 7)]

# サービス名: {令和年度: (件数, 金額)}
KAIGO_KYUFU = {
    "居宅介護":                     {2: (27, 1387313), 3: (24, 1294783), 4: (24, 1016801),
                                     5: (27, 1095062), 6: (43, 1155018), 7: (44, 1713336)},
    "生活介護":                     {2: (58, 9144480), 3: (64, 10414930), 4: (83, 11977750),
                                     5: (84, 12373933), 6: (84, 13040653), 7: (79, 13209810)},
    "生活介護（基準該当）":         {6: (11, 99627), 7: (7, 55760)},
    "短期入所":                     {6: (2, 26000), 7: (3, 54630)},
    "自立訓練（生活訓練）":         {2: (23, 2615274), 3: (21, 3023251), 4: (11, 1973204),
                                     5: (7, 1183986), 6: (0, 0), 7: (2, 382238)},
    "宿泊型自立訓練":               {2: (11, 689860), 3: (0, 0), 4: (0, 0), 5: (0, 0)},
    "就労移行支援":                 {4: (2, 40690), 5: (5, 744891)},
    "就労定着支援":                 {2: (12, 295250), 3: (2, 53230)},
    "就労継続支援（B型）":          {2: (74, 9458880), 3: (89, 10911480), 4: (106, 12180133),
                                     5: (114, 14141251), 6: (140, 19829973), 7: (143, 19707121)},
    "共同生活援助":                 {2: (64, 6413031), 3: (108, 13103695), 4: (83, 9680695),
                                     5: (75, 8186883), 6: (82, 9927798), 7: (105, 9663491)},
    "共同生活援助（特定障害者特別給付費）": {6: (81, 810000), 7: (105, 1026928)},
    "施設入所支援":                 {2: (46, 4993225), 3: (50, 5747691), 4: (48, 5829436),
                                     5: (48, 5696159), 6: (48, 5845600), 7: (48, 5993690)},
    "施設入所支援（特定障害者特別給付費）": {6: (43, 269576), 7: (12, 206084)},
    "計画相談支援":                 {2: (61, 1004961), 3: (81, 1442575), 4: (83, 1535312),
                                     5: (55, 997809), 6: (71, 1694563), 7: (76, 1863988)},
}

KAIGO_TOTAL = {2: (376, 36002274), 3: (439, 45991635), 4: (440, 44234021),
               5: (415, 44419974), 6: (605, 52698808), 7: (624, 53877076)}

JIDO_KYUFU = {
    "児童発達支援":       {2: (7, 308236), 3: (22, 3032590), 4: (24, 3715790),
                           5: (23, 3338200), 6: (13, 1452980), 7: (1, 124060)},
    "放課後等デイサービス": {2: (11, 161658), 3: (11, 174771), 4: (12, 176913),
                           5: (12, 128322), 6: (12, 111195), 7: (27, 1075951)},
    "保育所等訪問支援":   {7: (3, 105232)},
    "障害児相談支援":     {2: (11, 276170), 3: (3, 47260), 4: (6, 137720),
                           5: (5, 111310), 6: (3, 77100), 7: (8, 330460)},
}

JIDO_TOTAL = {2: (29, 746064), 3: (36, 3254621), 4: (42, 4030423),
              5: (40, 3577832), 6: (28, 1641275), 7: (39, 1635703)}

# --- 財源の負担区分 ---------------------------------------------------
# 障害者総合支援法・児童福祉法に基づく法定負担割合。
# 実際の決算額は村の歳入歳出で確認する（09_村確認事項参照）。
FUNDING_RULES = [
    ("介護給付費・訓練等給付費（自立支援給付）", 0.50, 0.25, 0.25, "義務的経費",
     "障害者総合支援法第92条〜第95条。国庫負担基準額の範囲内。"),
    ("計画相談支援給付費", 0.50, 0.25, 0.25, "義務的経費",
     "障害者総合支援法に基づく自立支援給付。"),
    ("障害児通所給付費", 0.50, 0.25, 0.25, "義務的経費",
     "児童福祉法第57条の2等。"),
    ("障害児相談支援給付費", 0.50, 0.25, 0.25, "義務的経費",
     "児童福祉法に基づく給付。"),
    ("自立支援医療（更生医療・育成医療）", 0.50, 0.25, 0.25, "義務的経費",
     "精神通院医療は都道府県の実施事務のため村負担なし。"),
    ("補装具費", 0.50, 0.25, 0.25, "義務的経費",
     "障害者総合支援法第76条。"),
    ("地域生活支援事業", 0.50, 0.25, 0.25, "裁量的経費",
     "国1/2以内・県1/4以内の統合補助金。補助基準額を超える分は村単独負担となる。"),
    ("村単独事業（重度心身障害者医療費助成等）", 0.00, None, None, "村単独",
     "一部に県補助金あり。県補助率は事業ごとに確認が必要。"),
]
