import openpyxl, warnings
warnings.filterwarnings('ignore')
SP="/tmp/claude-0/-home-user-repository/05ae8dfb-f2f7-5c38-805c-45df053cf245/scratchpad"
SRC=f"{SP}/zipB/zip_work/01_final_deliverables/02_mutsu_toushizaisei_plan_v7_final.xlsx"
wv=openpyxl.load_workbook(SRC,data_only=True)
o=wv["02_R8予算_年次計画置換"]; sim=wv["SIM_R8予算置換_0%"]
COLS=list(range(2,12))              # B..K = R7..R16
YEARS=[7,8,9,10,11,12,13,14,15,16]
def orow(r): return [o.cell(row=r,column=c).value or 0 for c in COLS]
def srow(r): return [sim.cell(row=r,column=c).value for c in COLS]

IN={k:orow(v) for k,v in dict(給水収益=7,その他営収=8,営業外収益=9,繰入負担金=11,長期前受金戻入=12,
    特別利益=14,営業費用=18,減価償却費=25,資産減耗=26,営業外費用=28,支払利息=29,特別損失=31,
    資本的収入=34,資本的支出=44).items()}
B6_ORIG=sim["B6"].value if isinstance(sim["B6"].value,float) else 0.0726393943654804
R7_GENKIN=242588.0; R7_HOTEN=183463.0

def run(rate, fix_double):
    """rate=料金改定率(R11以降)、fix_double=True で r9 の二重計上を是正"""
    out={k:[] for k in ["r9","r10","r11","r28","r31","r32","r54","r57","r59","r72"]}
    genkin, hoten = R7_GENKIN, R7_HOTEN
    for i,y in enumerate(YEARS):
        mult = 1+rate if y>=11 else 1.0
        r10 = IN["給水収益"][i]*mult
        r11 = r10 - IN["給水収益"][i]
        r12 = IN["その他営収"][i]
        r9  = (r10+r12) if fix_double else (r10+r11+r12)
        r13,r14,r15,r16 = IN["営業外収益"][i],IN["長期前受金戻入"][i],IN["繰入負担金"][i],IN["特別利益"][i]
        r20,r21,r22 = IN["営業費用"][i],IN["減価償却費"][i],IN["資産減耗"][i]
        r23,r25 = IN["営業外費用"][i],IN["特別損失"][i]
        r27 = (r9+r13)-(r20+r23)
        r28 = r27+r16-r25
        r31 = (r9+r13)/(r20+r23)*100
        r34 = (r20+r23)-r14
        r32 = r10/r34*100
        r54 = IN["資本的収入"][i]-IN["資本的支出"][i]
        r57 = r28+r21+r22-r14
        if i==0:
            r59, r72 = R7_GENKIN, R7_HOTEN          # R7は起点（定数）
        else:
            genkin += r57+r54; hoten += r57+r54
            r59, r72 = genkin, hoten
        if i==0: genkin, hoten = R7_GENKIN, R7_HOTEN
        for k,v in zip(out,[r9,r10,r11,r28,r31,r32,r54,r57,r59,r72]): out[k].append(v)
    return out

# --- 検証：現行モデル（改定あり・二重計上そのまま）が納品値を再現するか ---
base=run(B6_ORIG, fix_double=False)
print("【検証】Python再実装 vs 納品Excelキャッシュ値")
ok=True
for key,rw in [("r28",28),("r31",31),("r32",32),("r54",54),("r59",59),("r72",72)]:
    exp=srow(rw)
    for i,y in enumerate(YEARS):
        a,b=base[key][i],exp[i]
        if isinstance(b,(int,float)) and abs(a-b)>0.6:
            print(f"  ★不一致 {key} R{y}: py={a:,.2f} xls={b:,.2f}"); ok=False
print("  → 全項目一致" if ok else "  → 差異あり")

YRS=["R7","R8","R9","R10","R11","R12","R13","R14","R15","R16"]
A=run(B6_ORIG, False)   # 納品版（改定あり・二重計上そのまま）
B=run(0.0,      False)  # 料金改定なし（rate=0 → r11=0 なので二重計上は発生しない）
C=run(B6_ORIG, True)    # 改定あり・二重計上を是正

def show(title,key,unit="百万円",div=1000.0,dec=0):
    print(f"\n=== {title} ===")
    print("        "+"".join(f"{y:>10}" for y in YRS))
    for nm,d in [("改定なし",B),("改定あり(納品)",A),("改定あり(是正)",C)]:
        print(f"{nm:<14}"+"".join(f"{d[key][i]/div:>10,.{dec}f}" for i in range(10)))

show("補填財源残高（期末）","r72")
show("年度末現預金残高","r59")
show("当年度純損益","r28")
show("経常収支比率（％）","r31",div=1.0,dec=1)
show("料金回収率（％）","r32",div=1.0,dec=1)

print("\n=== 二重計上の影響額（千円）改定あり：納品 − 是正 ===")
print("        "+"".join(f"{y:>10}" for y in YRS))
for key,nm in [("r28","純損益"),("r72","補填財源"),("r59","現預金")]:
    print(f"{nm:<14}"+"".join(f"{A[key][i]-C[key][i]:>10,.0f}" for i in range(10)))
