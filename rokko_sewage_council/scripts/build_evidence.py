# -*- coding: utf-8 -*-
"""エビデンス集に第2回審議会資料の算定根拠（パターン別シミュレーション結果）を追加する."""
import json
import pathlib
import shutil

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent.parent
SRC = HERE / "source" / "六戸町_使用料改定_成果物エビデンス集.xlsx"
DST = HERE / "output" / "六戸町_使用料改定_成果物エビデンス集.xlsx"
M = json.loads((HERE / "data" / "metrics.json").read_text(encoding="utf-8"))

YEARS = ["令和7年度", "令和8年度", "令和9年度", "令和10年度", "令和11年度",
         "令和12年度", "令和13年度", "令和14年度", "令和15年度", "令和16年度"]
ROWS = [("現行（改定なし）", "現行"), ("パターン①（標準型）2か年", "①2"), ("パターン①（標準型）3か年", "①3"),
        ("パターン②（家庭軽減型）2か年", "②2"), ("パターン②（家庭軽減型）3か年", "②3"),
        ("パターン④（段階累進型）2か年", "④2"), ("パターン④（段階累進型）3か年", "④3")]

RATES = {
    "現行":   [1000, 120, 120, 130, 130, 140, 140, 140, 160],
    "①中間": [1250, 135, 138, 145, 148, 158, 163, 173, 193],
    "①最終": [1500, 150, 155, 160, 165, 175, 185, 205, 225],
    "②中間": [1100, 150, 153, 160, 163, 173, 178, 188, 208],
    "②最終": [1200, 180, 185, 190, 195, 205, 215, 235, 255],
    "④中間": [1200, 140, 143, 150, 153, 163, 168, 178, 198],
    "④最終": [1400, 160, 165, 170, 175, 185, 195, 215, 235],
}
BLOCKS = ["基本(0〜10㎥)", "11〜20㎥", "21〜30㎥", "31〜40㎥", "41〜50㎥",
          "51〜70㎥", "71〜100㎥", "101〜150㎥", "151㎥〜"]
CAPS = [20, 30, 40, 50, 70, 100, 150]


def fee(rate, vol):
    total, prev = rate[0], 10
    for i, cap in enumerate(CAPS, start=1):
        if vol <= prev:
            break
        total += (min(vol, cap) - prev) * rate[i]
        prev = cap
    if vol > 150:
        total += (vol - 150) * rate[8]
    return total


NAVY = PatternFill("solid", fgColor="1E3A8A")
GREY = PatternFill("solid", fgColor="F1F5F9")
BAND = PatternFill("solid", fgColor="F8FAFC")
THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
BODY = Font(size=10)

DST.parent.mkdir(parents=True, exist_ok=True)
shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

# ------------------------------------------------ 成果物一覧の更新
ws = wb["成果物一覧"]
for row in ws.iter_rows():
    if row[1].value and "第2回審議会資料" in str(row[1].value):
        row[1].value = "六戸町_第2回審議会資料.pptx"
        row[2].value = "PPTX（37スライド）"
        row[4].value = ("第2回審議会資料。改定しない場合の影響（01-1）・パターン①②④それぞれの"
                        "収支シミュレーション（01-3〜01-5）・近隣自治体比較（01-8）・"
                        "段階単価一覧（02-2〜02-4）・経常収支比率見込み（03-1〜03-3）・"
                        "家庭/事業所別影響額（04-1〜04-3）を3パターン並列で収録。"
                        "指標値は各使用料改定.xlsx「パターン別比較」（社人研人口推計）から転記。")
    if row[0].value == "No.":
        hdr_row = row[0].row
last = ws.max_row
ws.cell(last + 1, 1, len(list(ws.iter_rows(min_row=hdr_row + 1, max_row=last))) + 1)
ws.cell(last + 1, 2, "本ブック「パターン別SIM結果」シート")
ws.cell(last + 1, 3, "XLSX（算定根拠）")
ws.cell(last + 1, 4, "✅ 追加")
ws.cell(last + 1, 5, "第2回審議会資料に掲載した経常収支比率・経費回収率・モデルケース別月額の元数値。")
for c in range(1, 6):
    ws.cell(last + 1, c).font = BODY
    ws.cell(last + 1, c).alignment = Alignment(vertical="center", wrap_text=True)

# ------------------------------------------------ 料金データ確認の補正
ws = wb["料金データ確認"]
for r, blk in enumerate(BLOCKS, start=4):
    for c, key in enumerate(["現行", "①中間", "①最終", "②中間", "②最終", "④中間", "④最終"], start=2):
        ws.cell(r, c).value = RATES[key][r - 4]
ws.cell(13, 1).value = ("※ R8中間単価は算定ブックの数式どおり（現行＋最終）÷2の四捨五入。"
                        "第1回審議会資料は切り捨てで提示しており、①71〜100㎥・101〜150㎥・151㎥〜、"
                        "②21〜30㎥・41〜50㎥の5区分で1円の差があります（R8単年の収入影響は約16万円）。")
ws.cell(13, 1).font = Font(size=9, color="B45309")
ws.cell(16, 1).value = ("20㎥時の月額：現行 税抜2,200円／税込2,420円　→　改定後いずれも 税抜3,000円／税込3,300円"
                        "（税抜＋800円・税込＋880円・＋36.4%）")
for r, key in ((18, "①最終"), (19, "②最終"), (20, "④最終")):
    ws.cell(r, 2).value = f"{fee(RATES[key], 20):,}円"
    ws.cell(r, 4).value = f"税込 {int(round(fee(RATES[key], 20) * 1.1)):,}円"
ws.cell(17, 4).value = "税込 2,420円"

# ------------------------------------------------ 汚水処理費の算定シート
if "汚水処理費の算定" in wb.sheetnames:
    del wb["汚水処理費の算定"]
oc = wb.create_sheet("汚水処理費の算定")
oc.sheet_view.showGridLines = False
r = 1
oc.cell(r, 1, "経費回収率の分母（汚水処理費）の算定　― 維持管理費ベース ―").font = Font(bold=True, size=12)
r += 1
for line in ("算定式：汚水処理費（維持管理費分）＝ 経常費用 − 減価償却費 − 支払利息 − その他営業外費用",
             "　　　　資本費 ＝ 減価償却費 − 長期前受金戻入 ＋ 支払利息"
             "　…　分流式下水道等に要する経費（一般会計繰入金）でカバーされるため、分母から全額控除",
             "出典：六戸町_公共／農集_使用料改定.xlsx「財政計画（ベース）」（社人研人口推計・現行料金ベース）"):
    oc.cell(r, 1, line).font = Font(size=9, color="475569")
    r += 1
r += 1

CALC = [("経常費用（経常支出D）", "経常支出", "#,##0", False),
        ("　− 減価償却費", "減価償却費", "#,##0", False),
        ("　− 支払利息", "支払利息", "#,##0", False),
        ("　− その他営業外費用", "営業外費用その他", "#,##0", False),
        ("▶ 汚水処理費（維持管理費分）", None, "#,##0", True),
        ("　【内訳】職員給与費", "職員給与費", "#,##0", False),
        ("　【内訳】動力費", "動力費", "#,##0", False),
        ("　【内訳】修繕費", "修繕費", "#,##0", False),
        ("　【内訳】材料費", "材料費", "#,##0", False),
        ("　【内訳】その他（主に委託料）", "経費その他", "#,##0", False),
        ("（参考）長期前受金戻入", "長期前受金戻入", "#,##0", False),
        ("（参考）資本費", None, "#,##0", False),
        ("（参考）一般会計繰入金（補助金）", "補助金", "#,##0", False),
        ("（参考）繰入金 ÷ 資本費（倍）", None, "0.00", False),
        ("使用料収入（現行・改定なし）", None, "#,##0", False),
        ("経費回収率（現行・改定なし）", None, "0.0%", True)]

for biz in ("公共", "農集"):
    label = "公共下水道事業" if biz == "公共" else "農業集落排水事業"
    fp = M[biz]["財政計画"]
    cap = [fp["減価償却費"][i] - fp["長期前受金戻入"][i] + fp["支払利息"][i] for i in range(10)]
    mnt = [fp["経常支出"][i] - fp["減価償却費"][i] - fp["支払利息"][i] - fp["営業外費用その他"][i]
           for i in range(10)]
    rev = M[biz]["使用料収入"]["現行"]
    oc.cell(r, 1, f"■ {label}（単位：千円）").font = BOLD
    r += 1
    oc.cell(r, 1, "項目").fill, oc.cell(r, 1).font, oc.cell(r, 1).border = NAVY, HEAD, BOX
    for j, y in enumerate(YEARS, start=2):
        c = oc.cell(r, j, y)
        c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
    r += 1
    for i, (name, key, fmt, strong) in enumerate(CALC):
        vals = (fp[key] if key else
                mnt if name.startswith("▶") else
                cap if "資本費" in name and "÷" not in name else
                [fp["補助金"][k] / cap[k] for k in range(10)] if "÷" in name else
                rev if "使用料収入" in name else
                [rev[k] / mnt[k] for k in range(10)])
        c = oc.cell(r, 1, name)
        c.font, c.border = (BOLD if strong else BODY), BOX
        if strong:
            c.fill = GREY
        elif i % 2:
            c.fill = BAND
        for j, v in enumerate(vals, start=2):
            c = oc.cell(r, j, round(v, 4) if fmt == "0.0%" else v)
            c.number_format, c.border = fmt, BOX
            c.alignment = Alignment(horizontal="right")
            c.font = BOLD if strong else BODY
            if strong:
                c.fill = GREY
            elif i % 2:
                c.fill = BAND
        r += 1
    r += 1

# 令和6年度決算との接続
R6 = {"公共": {"経常費用": 497395, "汚水処理費": 119503, "使用料収入": 56302, "回収率": 0.4711},
      "農集": {"経常費用": 106700, "汚水処理費": 32870, "使用料収入": 11831, "回収率": 0.3599}}
oc.cell(r, 1, "■ 令和6年度決算（第1回審議会資料）との接続（単位：千円）").font = BOLD
r += 1
for j, h in enumerate(["事業／項目", "R6決算", "R7見込", "差", "備考"], start=1):
    c = oc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
for biz in ("公共", "農集"):
    fp, a = M[biz]["財政計画"], R6[biz]
    mnt7 = fp["経常支出"][0] - fp["減価償却費"][0] - fp["支払利息"][0] - fp["営業外費用その他"][0]
    resid7 = fp["減価償却費"][0] + fp["支払利息"][0] + fp["営業外費用その他"][0]
    rows = [(f"{biz}　経常費用", a["経常費用"], fp["経常支出"][0], ""),
            (f"{biz}　汚水処理費（維持管理費分）", a["汚水処理費"], mnt7, ""),
            (f"{biz}　差引（減価償却費＋支払利息＋その他）", a["経常費用"] - a["汚水処理費"], resid7,
             "R6決算とR7見込がほぼ一致 → 同一基準で算定されていることの確認")]
    for i, (name, v6, v7, memo) in enumerate(rows):
        oc.cell(r, 1, name).font = BODY
        for j, v in enumerate((v6, v7, v7 - v6), start=2):
            c = oc.cell(r, j, v)
            c.number_format, c.alignment, c.font = "#,##0;△#,##0", Alignment(horizontal="right"), BODY
        oc.cell(r, 5, memo).font = Font(size=9, color="475569")
        for j in range(1, 6):
            oc.cell(r, j).border = BOX
            if i == 2:
                oc.cell(r, j).fill = GREY
        r += 1
r += 1
for line in ("※ 令和6年度決算の汚水処理費（公共119,503千円・農集32,870千円）も、"
             "経常費用から減価償却費・支払利息を控除した維持管理費ベースで算定されており、本シミュレーションと同一基準。",
             "※【内訳】その他は、職員給与費・動力費・修繕費・材料費以外の経費（主に委託料）。"
             "公共下水道ではR7の汚水処理費168,331千円のうち155,280千円（92%）を占める。",
             "※ 公共下水道はR6→R7で維持管理費が119,503千円→168,331千円（＋40.9%）と増加する見込みのため、"
             "経費回収率は47.11%→34.4%に低下する。どの費目の増加によるものかは要確認（確認事項A）。"):
    oc.cell(r, 1, line).font = Font(size=9, color="B45309")
    r += 1

oc.column_dimensions["A"].width = 34
for j in range(2, 12):
    oc.column_dimensions[get_column_letter(j)].width = 12
oc.freeze_panes = "B1"

# ------------------------------------------------ R7決算突合シート
# 令和7年度決算（合計残高試算表・貸借対照表／令和8年3月31日）の実績を data/r7_actual.json から
# 読み込み、経営戦略の推計値と突合する。
if "R7決算突合" in wb.sheetnames:
    del wb["R7決算突合"]
rc = wb.create_sheet("R7決算突合")
rc.sheet_view.showGridLines = False
ACT = json.loads((HERE / "data" / "r7_actual.json").read_text(encoding="utf-8"))
GAP = PatternFill("solid", fgColor="FEF2F2")            # 乖離が大きい行
r = 1
rc.cell(r, 1, "令和7年度　推計値と決算実績の突合").font = Font(bold=True, size=12)
r += 1
for line in ("・「R7推計値」は経営戦略（六戸町_公共／農集_使用料改定.xlsx「財政計画（ベース）」）の値。",
             "・「R7決算実績」の出所：R7公共／農集 合計残高試算表（令和8年3月31日・出力R8.6.16）、"
             "R7公共／農集 貸借対照表。経常収入＝下水道事業収益−特別利益、経常支出＝下水道事業費用−特別損失。",
             "・乖離が10%以上の行は赤色で網掛けしています。"):
    rc.cell(r, 1, line).font = Font(size=9, color="475569")
    r += 1
r += 1

for biz in ("公共", "農集"):
    label = "公共下水道事業" if biz == "公共" else "農業集落排水事業"
    fp = M[biz]["財政計画"]
    a = {k: v / 1000 for k, v in ACT[biz]["損益"].items()}
    ai = {k: v / 1000 for k, v in ACT[biz]["費目"].items()}
    v = lambda k: fp.get(k, [0] * 10)[0]
    oth_op = v("営業収益") - v("使用料収入") - v("雨水処理負担金")
    oth_no = v("営業外収益") - v("補助金") - v("長期前受金戻入")
    mnt = v("経常支出") - v("減価償却費") - v("支払利息") - v("営業外費用その他")
    a_in = a["下水道事業収益"] - a["特別利益"]
    a_out = a["下水道事業費用"] - a["特別損失"]
    a_mnt = a_out - a["減価償却費"] - a["支払利息"] - a["その他営業外費用"]
    a_oth = a_mnt - sum(ai.get(k, 0) for k in ("職員給与費", "動力費", "修繕費", "材料費"))
    FIN = [
        ("【収入】", "使用料収入", v("使用料収入"), a["下水道使用料"], "確認事項M・Fの検証に直結"),
        ("", "雨水処理負担金", v("雨水処理負担金"), a["雨水処理負担金"], "農集の推計799千円は実績に計上なし"),
        ("", "その他営業収益", oth_op, a["その他営業収益"], "督促手数料等"),
        ("", "営業収益 (A)", v("営業収益"), a["営業収益"], ""),
        ("", "補助金（一般会計繰入金）", v("補助金"), a["他会計補助金"], "農集は推計の約2.7倍"),
        ("", "長期前受金戻入", v("長期前受金戻入"), a["長期前受金戻入"], ""),
        ("", "その他営業外収益", oth_no, a["営業外収益"] - a["他会計補助金"] - a["長期前受金戻入"],
         "実績は受取利息＋雑収益。公共は別に特別利益8,306千円あり（経常外）"),
        ("", "営業外収益", v("営業外収益"), a["営業外収益"], ""),
        ("", "経常収入 (C)", v("経常収入"), a_in, "＝下水道事業収益−特別利益"),
        ("【支出】", "職員給与費", v("職員給与費"), ai.get("職員給与費", 0), "総係費の給料・手当・法定福利費等"),
        ("", "経費", v("経費"), a_mnt - ai.get("職員給与費", 0), ""),
        ("", "　うち動力費", v("動力費"), ai.get("動力費", 0), ""),
        ("", "　うち修繕費", v("修繕費"), ai.get("修繕費", 0), ""),
        ("", "　うち材料費", v("材料費"), ai.get("材料費", 0), ""),
        ("", "　うちその他（主に委託料）", v("経費その他"), a_oth,
         "実績の内訳は下表参照。確認事項Aの主因" if biz == "公共" else "実績の内訳は下表参照"),
        ("", "減価償却費", v("減価償却費"), a["減価償却費"], ""),
        ("", "営業費用", v("営業費用"), a["営業費用"], ""),
        ("", "支払利息", v("支払利息"), a["支払利息"], ""),
        ("", "その他営業外費用", v("営業外費用その他"), a["その他営業外費用"], ""),
        ("", "営業外費用", v("営業外費用"), a["営業外費用"], ""),
        ("", "経常支出 (D)", v("経常支出"), a_out, "＝下水道事業費用−特別損失"),
        ("【指標】", "経常損益 (C)-(D)", v("経常損益"), a_in - a_out, ""),
        ("", "汚水処理費（維持管理費分）", mnt, a_mnt, "＝経常支出−減価償却費−支払利息−その他営業外費用"),
        ("", "資本費", v("減価償却費") - v("長期前受金戻入") + v("支払利息"),
         a["減価償却費"] - a["長期前受金戻入"] + a["支払利息"], "＝減価償却費−長期前受金戻入＋支払利息"),
    ]
    rc.cell(r, 1, f"■ {label}（単位：千円）").font = BOLD
    r += 1
    for j, h in enumerate(["区分", "科目", "R7推計値", "R7決算実績", "乖離額", "乖離率", "備考"], start=1):
        c = rc.cell(r, j, h)
        c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
    r += 1
    for i, (sec, name, est, act, memo) in enumerate(FIN):
        strong = name in ("経常収入 (C)", "経常支出 (D)", "汚水処理費（維持管理費分）", "使用料収入")
        big = est > 0 and abs(act / est - 1) >= 0.10
        rc.cell(r, 1, sec).font = BOLD
        rc.cell(r, 2, name).font = BOLD if strong else BODY
        rc.cell(r, 3, round(est)).number_format = "#,##0;△#,##0"
        rc.cell(r, 4, round(act)).number_format = "#,##0;△#,##0"
        rc.cell(r, 5, round(act - est)).number_format = "+#,##0;△#,##0"
        # 推計値がマイナス（経常損益）の行は乖離率が符号を反転して読めるため出さない
        rc.cell(r, 6, round(act / est - 1, 4) if est > 0 else None).number_format = "+0.0%;△0.0%"
        for j in (3, 4, 5, 6):
            rc.cell(r, j).font = BOLD if strong else BODY
            rc.cell(r, j).alignment = Alignment(horizontal="right")
        rc.cell(r, 7, memo).font = Font(size=9, color="475569")
        for j in range(1, 8):
            rc.cell(r, j).border = BOX
            rc.cell(r, j).fill = GREY if strong else (GAP if big else (BAND if i % 2 else PatternFill()))
        r += 1
    cr_e, er_e = M[biz]["経常収支比率"]["現行"][0], M[biz]["経費回収率"]["現行"][0]
    for name, est, act, memo in (
            ("経常収支比率", cr_e, a_in / a_out, "経常収入(C)÷経常支出(D)"),
            ("経費回収率", er_e, a["下水道使用料"] / a_mnt, "使用料収入÷汚水処理費（維持管理費分）")):
        rc.cell(r, 2, name).font = BOLD
        rc.cell(r, 3, round(est, 4)).number_format = "0.0%"
        rc.cell(r, 4, round(act, 4)).number_format = "0.0%"
        rc.cell(r, 5, round(act - est, 4)).number_format = "+0.0%;△0.0%"
        for j in (3, 4, 5):
            rc.cell(r, j).font = BOLD
            rc.cell(r, j).alignment = Alignment(horizontal="right")
        rc.cell(r, 7, memo).font = Font(size=9, color="475569")
        for j in range(1, 8):
            rc.cell(r, j).border, rc.cell(r, j).fill = BOX, GREY
        r += 1
    r += 1

# 「その他」の実績内訳と、貸借対照表から読み取れる情報
rc.cell(r, 1, "■ 「その他（主に委託料）」の実績内訳（千円）").font = BOLD
r += 1
for j, h in enumerate(["費目", "公共下水道", "農業集落排水", "備考"], start=1):
    c = rc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
DETAIL = [("流域下水道管理運営費負担金", 63706663, 0, "公共の最大費目。実績の51%"),
          ("委託料", 24284955, 13640109, "管渠・処理場・業務・総係の合計"),
          ("工事請負費", 14480000, 2270000, ""),
          ("光熱水費・燃料費", 7285678, 1669521, ""),
          ("手数料", 2361589, 6739571, ""),
          ("通信運搬費・備消品費・保険料ほか", 2113111, 1332999, "残差")]
for i, (name, ko, no, memo) in enumerate(DETAIL):
    rc.cell(r, 1, name).font = BODY
    for j, val in ((2, ko), (3, no)):
        c = rc.cell(r, j, round(val / 1000))
        c.number_format, c.alignment, c.font = "#,##0", Alignment(horizontal="right"), BODY
    rc.cell(r, 4, memo).font = Font(size=9, color="475569")
    for j in range(1, 5):
        rc.cell(r, j).border = BOX
        if i % 2:
            rc.cell(r, j).fill = BAND
    r += 1
r += 1

rc.cell(r, 1, "■ 貸借対照表（令和8年3月31日）から読み取れる財政指標（確認事項Q）").font = BOLD
r += 1
for j, h in enumerate(["項目", "公共下水道", "農業集落排水", "合計"], start=1):
    c = rc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
BS = [("企業債残高（固定＋流動）", "企業債残高"), ("　うち固定負債", "企業債_固定"),
      ("　うち流動負債（1年以内償還）", "企業債_流動"), ("現金預金", "現金預金"),
      ("基金", "基金"), ("未収金", "未収金"), ("当年度未処分利益剰余金", "当年度未処分利益剰余金")]
for i, (name, key) in enumerate(BS):
    ko, no = ACT["公共"]["貸借"][key], ACT["農集"]["貸借"][key]
    rc.cell(r, 1, name).font = BOLD if i == 0 else BODY
    for j, val in ((2, ko), (3, no), (4, ko + no)):
        c = rc.cell(r, j, round(val / 1000))
        c.number_format, c.alignment = "#,##0;△#,##0", Alignment(horizontal="right")
        c.font = BOLD if i == 0 else BODY
    for j in range(1, 5):
        rc.cell(r, j).border = BOX
        if i % 2:
            rc.cell(r, j).fill = BAND
    r += 1
r += 1

for line in ("※ 有収水量・調定口数の実績は未受領のため、使用料単価・汚水処理原価は"
             "経営戦略の推計水量（公共452,992㎥・農集94,354㎥）で試算しています。",
             "※ 一般会計繰入金の繰入基準別内訳（基準内／基準外）は未受領です。"
             "農業集落排水の実績51,777千円は資本費19,185千円の約2.7倍で、維持管理費分への充当が含まれるとみられます。"):
    rc.cell(r, 1, line).font = Font(size=9, color="B45309")
    r += 1

rc.column_dimensions["A"].width = 30
rc.column_dimensions["B"].width = 26
for col in ("C", "D", "E", "F"):
    rc.column_dimensions[col].width = 15
rc.column_dimensions["G"].width = 46

# ------------------------------------------------ パターン別SIM結果シート
if "パターン別SIM結果" in wb.sheetnames:
    del wb["パターン別SIM結果"]
sim = wb.create_sheet("パターン別SIM結果")
sim.sheet_view.showGridLines = False
r = 1
sim.cell(r, 1, "第2回審議会資料　掲載数値の算定根拠（パターン別シミュレーション結果）").font = Font(bold=True, size=12)
r += 1
sim.cell(r, 1, "出典：六戸町_公共_使用料改定.xlsx／六戸町_農集_使用料改定.xlsx「パターン別比較」シート"
               "（人口推計＝社人研／使用料収入以外の収支項目は全パターン共通）").font = Font(size=9, color="475569")
r += 2

for biz in ("公共", "農集"):
    label = "公共下水道事業" if biz == "公共" else "農業集落排水事業"
    for metric, goal in (("経常収支比率", "目標：100%以上（単年度黒字）"),
                         ("経費回収率", "目標：47%以上")):
        sim.cell(r, 1, f"■ {label}　{metric}　{goal}").font = BOLD
        r += 1
        sim.cell(r, 1, "パターン／年度").fill = NAVY
        sim.cell(r, 1).font = HEAD
        for j, y in enumerate(YEARS, start=2):
            c = sim.cell(r, j, y)
            c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
        sim.cell(r, 1).border = BOX
        r += 1
        for i, (name, key) in enumerate(ROWS):
            c = sim.cell(r, 1, name)
            c.font, c.border = (BOLD if "②" in name else BODY), BOX
            if i % 2:
                c.fill = BAND
            for j, v in enumerate(M[biz][metric][key], start=2):
                c = sim.cell(r, j, round(v, 4))
                c.number_format = "0.0%"
                c.alignment, c.border = Alignment(horizontal="center"), BOX
                c.font = BOLD if "②" in name else BODY
                if i % 2:
                    c.fill = BAND
            r += 1
        r += 1

    sim.cell(r, 1, f"■ {label}　使用料収入（千円）").font = BOLD
    r += 1
    sim.cell(r, 1, "パターン／年度").fill = NAVY
    sim.cell(r, 1).font = HEAD
    sim.cell(r, 1).border = BOX
    for j, y in enumerate(YEARS, start=2):
        c = sim.cell(r, j, y)
        c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
    r += 1
    for i, (name, key) in enumerate(ROWS):
        c = sim.cell(r, 1, name)
        c.font, c.border = (BOLD if "②" in name else BODY), BOX
        if i % 2:
            c.fill = BAND
        for j, v in enumerate(M[biz]["使用料収入"][key], start=2):
            c = sim.cell(r, j, v)
            c.number_format = "#,##0"
            c.alignment, c.border = Alignment(horizontal="right"), BOX
            c.font = BOLD if "②" in name else BODY
            if i % 2:
                c.fill = BAND
        r += 1
    r += 1

# モデルケース別月額
sim.cell(r, 1, "■ モデルケース別　月額使用料（円）　※各区分の単価を段階累進で積み上げて算定").font = BOLD
r += 1
hdr = ["使用水量", "現行(税抜)", "現行(税込)"]
for p in ("①", "②", "④"):
    hdr += [f"{p}最終(税抜)", f"{p}最終(税込)", f"{p}増加額(税込)"]
for j, h in enumerate(hdr, start=1):
    c = sim.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center", wrap_text=True), BOX
r += 1
for i, vol in enumerate([10, 15, 20, 30, 40, 50, 100, 200, 500]):
    cur = fee(RATES["現行"], vol)
    curt = int(round(cur * 1.1))
    vals = [f"{vol}㎥", cur, curt]
    for p in ("①", "②", "④"):
        v = fee(RATES[p + "最終"], vol)
        vals += [v, int(round(v * 1.1)), int(round(v * 1.1)) - curt]
    for j, v in enumerate(vals, start=1):
        c = sim.cell(r, j, v)
        c.border = BOX
        c.font = BOLD if vol == 20 else BODY
        if j > 1:
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")
        if vol == 20:
            c.fill = GREY
        elif i % 2:
            c.fill = BAND
    r += 1
r += 1
sim.cell(r, 1, "※ 20㎥（一般家庭の標準使用量）は3パターンとも税抜3,000円／税込3,300円で共通（現行比＋880円・＋36.4%）。").font = Font(size=9, color="475569")

sim.column_dimensions["A"].width = 26
for j in range(2, len(hdr) + 1):
    sim.column_dimensions[get_column_letter(j)].width = 12
sim.freeze_panes = "B1"

# ------------------------------------------------ 改定なしとの比較シート
# 「使用料改定を行わない場合」の影響（資料01-1）と、目標到達に必要な増収額の根拠。
# 経常損益は 財政計画（ベース）の使用料収入をパターン別の値に差し替えて算定する。
if "改定なしとの比較" in wb.sheetnames:
    del wb["改定なしとの比較"]
nc = wb.create_sheet("改定なしとの比較")
nc.sheet_view.showGridLines = False
YI = {y: i for i, y in enumerate(
    ["R7", "R8", "R9", "R10", "R11", "R12", "R13", "R14", "R15", "R16"])}
CMP = [("現行（改定なし）", "現行"), ("パターン①（標準型）2か年", "①2"),
       ("パターン②（家庭軽減型）2か年", "②2"), ("パターン④（段階累進型）2か年", "④2")]
SPAN = ["R8", "R9", "R10", "R11", "R12", "R13", "R14", "R15", "R16"]   # 改定効果が及ぶ期間


def _plan(biz, item, i):
    return M[biz]["財政計画"][item][i]


def _pl(biz, key, i):
    """経常損益（千円）＝ 経常収入 − 使用料収入(ベース) ＋ 使用料収入(パターン) − 経常支出."""
    return (_plan(biz, "経常収入", i) - _plan(biz, "使用料収入", i)
            + M[biz]["使用料収入"][key][i] - _plan(biz, "経常支出", i))


r = 1
nc.cell(r, 1, "使用料改定を行わない場合との比較（資料01-1・想定問答Q4の算定根拠）").font = Font(bold=True, size=12)
r += 1
nc.cell(r, 1, "経常損益＝経常収入−経常支出。使用料収入以外の収支項目は「財政計画（ベース）」の値を"
              "全ケース共通で用い、使用料収入のみパターン別比較シートの値に差し替えて算定。"
              ).font = Font(size=9, color="475569")
r += 2

for biz in ("公共", "農集"):
    label = "公共下水道事業" if biz == "公共" else "農業集落排水事業"
    nc.cell(r, 1, f"■ {label}　経常損益の見込み（千円）").font = BOLD
    r += 1
    c = nc.cell(r, 1, "ケース／年度")
    c.fill, c.font, c.border = NAVY, HEAD, BOX
    for j, y in enumerate(YEARS, start=2):
        c = nc.cell(r, j, y)
        c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
    c = nc.cell(r, len(YEARS) + 2, "R8〜R16 累計")
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center", wrap_text=True), BOX
    r += 1
    for i, (name, key) in enumerate(CMP):
        emph = "②" in name or "現行" in name
        c = nc.cell(r, 1, name)
        c.font, c.border = (BOLD if emph else BODY), BOX
        if i % 2:
            c.fill = BAND
        for j, y in enumerate(["R7"] + SPAN, start=2):
            c = nc.cell(r, j, round(_pl(biz, key, YI[y])))
            c.number_format = "#,##0;△#,##0"
            c.alignment, c.border = Alignment(horizontal="right"), BOX
            c.font = BOLD if emph else BODY
            if i % 2:
                c.fill = BAND
        c = nc.cell(r, len(YEARS) + 2, round(sum(_pl(biz, key, YI[y]) for y in SPAN)))
        c.number_format = "#,##0;△#,##0"
        c.alignment, c.border, c.font = Alignment(horizontal="right"), BOX, BOLD
        c.fill = GREY
        r += 1
    r += 1

nc.cell(r, 1, "■ 両事業合計　R8〜R16 経常損益の累計（千円）").font = BOLD
r += 1
for j, h in enumerate(["ケース", "公共下水道", "農業集落排水", "合計", "改定なしとの差"], start=1):
    c = nc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
_base = sum(sum(_pl(b, "現行", YI[y]) for y in SPAN) for b in ("公共", "農集"))
for i, (name, key) in enumerate(CMP):
    ko = sum(_pl("公共", key, YI[y]) for y in SPAN)
    no = sum(_pl("農集", key, YI[y]) for y in SPAN)
    for j, v in enumerate([name, round(ko), round(no), round(ko + no), round(ko + no - _base)], start=1):
        c = nc.cell(r, j, v)
        c.border, c.font = BOX, (BOLD if ("②" in name or "現行" in name) else BODY)
        if j > 1:
            c.number_format = "#,##0;△#,##0"
            c.alignment = Alignment(horizontal="right")
        if i % 2:
            c.fill = BAND
    r += 1
r += 1

nc.cell(r, 1, "■ 使用料単価（円/㎥）＝ 使用料収入 ÷ 有収水量　"
              "※ 交付要件の除外判定（150円/㎥未満）に用いる指標").font = BOLD
r += 1
for j, h in enumerate(["事業／ケース", "R7", "R9", "R10", "R12"], start=1):
    c = nc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
for biz in ("公共", "農集"):
    for i, (name, key) in enumerate(CMP):
        c = nc.cell(r, 1, f"{biz}　{name}")
        c.border, c.font = BOX, (BOLD if "②" in name else BODY)
        for j, y in enumerate(["R7", "R9", "R10", "R12"], start=2):
            k = YI[y]
            c = nc.cell(r, j, round(M[biz]["使用料収入"][key][k] * 1000 / _plan(biz, "有収水量", k), 1))
            c.number_format = "#,##0.0"
            c.alignment, c.border = Alignment(horizontal="right"), BOX
            c.font = BOLD if "②" in name else BODY
        r += 1
r += 1

nc.cell(r, 1, "■ 目標水準に到達するために必要な使用料収入（千円）").font = BOLD
r += 1
for j, h in enumerate(["事業／指標", "年度", "目標", "汚水処理費等",
                       "必要な使用料収入", "②改定後の見込み", "不足額（△は超過）", "必要な増収率"], start=1):
    c = nc.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center", wrap_text=True), BOX
r += 1
NEEDS = [("公共　経費回収率", "R9", 0.47), ("公共　経費回収率", "R10", 0.50),
         ("公共　経費回収率", "R12", 0.55), ("農集　経費回収率", "R9", 0.47)]
for i, (lab, y, tgt) in enumerate(NEEDS):
    biz = lab[:2]
    k = YI[y]
    base = _plan(biz, "汚水処理費", k)
    need = base * tgt
    now = M[biz]["使用料収入"]["②2"][k]
    for j, v in enumerate([lab, y, f"{tgt*100:.0f}%", round(base), round(need), now,
                           round(need - now), round(need / now - 1, 4)], start=1):
        c = nc.cell(r, j, v)
        c.border, c.font = BOX, BODY
        if j >= 4:
            c.number_format = "0.0%" if j == 8 else "#,##0;△#,##0"
            c.alignment = Alignment(horizontal="right")
        if i % 2:
            c.fill = BAND
    r += 1
# 農集の経常収支比率100%（目標①）
for i, y in enumerate(["R9", "R10"]):
    k = YI[y]
    out = _plan("農集", "経常支出", k)
    other = _plan("農集", "経常収入", k) - _plan("農集", "使用料収入", k)
    need = out - other
    now = M["農集"]["使用料収入"]["②2"][k]
    for j, v in enumerate(["農集　経常収支比率", y, "100%", round(out), round(need), now,
                           round(need - now), round(need / now - 1, 4)], start=1):
        c = nc.cell(r, j, v)
        c.border, c.font = BOX, BOLD
        if j >= 4:
            c.number_format = "0.0%" if j == 8 else "#,##0;△#,##0"
            c.alignment = Alignment(horizontal="right")
        c.fill = GREY
    r += 1
r += 1
nc.cell(r, 1, "※「汚水処理費等」は経費回収率の行は汚水処理費（維持管理費分）、"
              "経常収支比率の行は経常支出。必要な使用料収入は前者が汚水処理費×目標、"
              "後者が経常支出−使用料収入以外の経常収入。").font = Font(size=9, color="475569")

nc.column_dimensions["A"].width = 30
for j in range(2, len(YEARS) + 3):
    nc.column_dimensions[get_column_letter(j)].width = 13

# ------------------------------------------------ R7実績ベース試算シート
# 令和7年度決算が確定したため、シミュレーションの起点をR7実績に置き換えた場合の試算。
# 使用料収入は基本使用料の年間乗数を公共・農集とも6回として再計算し、R7実績で補正する。
# 汚水処理費・減価償却費・支払利息・長期前受金戻入・補助金は R7実績÷R7推計 の比率で全年度をスケールする。
if "R7実績ベース試算" in wb.sheetnames:
    del wb["R7実績ベース試算"]
rb = wb.create_sheet("R7実績ベース試算")
rb.sheet_view.showGridLines = False
YKEY = ["R7", "R8", "R9", "R10", "R11", "R12", "R13", "R14", "R15", "R16"]
# 使用料改定ブック「使用料算定」シートの調定口数と、R6実績で固定した区分別水量の構成比
KUCHI = {"公共": [2563, 2545, 2527, 2508, 2490, 2472, 2452, 2432, 2412, 2391],
         "農集": [524, 520, 517, 513, 509, 506, 501, 497, 493, 489]}
SHARE = {"公共": [0.03184, 0.254673, 0.18661, 0.115417, 0.063915, 0.043817, 0.026321, 0.00904, 0.021102],
         "農集": [0.295786, 0.249802, 0.180081, 0.113113, 0.068836, 0.054322, 0.026809, 0.005093, 0.003074]}
STEP = {"①": [1150, 130, 130, 140, 140, 150, 155, 160, 180],
        "②": [1050, 140, 140, 150, 150, 160, 165, 170, 190],
        "④": [1100, 130, 135, 140, 145, 155, 155, 165, 185]}
STEP2 = {"①": [1300, 140, 140, 150, 150, 160, 170, 180, 200],
         "②": [1100, 160, 160, 170, 170, 180, 190, 200, 220],
         "④": [1200, 140, 150, 150, 160, 170, 170, 190, 210]}


def _rate(p, span, i):
    if p == "現行":
        return RATES["現行"]
    if i == 0:
        return RATES["現行"]
    if span == "2":
        return RATES[p + "中間"] if i == 1 else RATES[p + "最終"]
    return STEP[p] if i == 1 else (STEP2[p] if i == 2 else RATES[p + "最終"])


def _gross(biz, rate, i):
    """基本使用料（年6回）＋超過使用料の算定額（千円・補正前）."""
    vol = M[biz]["財政計画"]["有収水量"][i]
    q = [round(vol * s) for s in SHARE[biz]]
    base = round(rate[0] * KUCHI[biz][i] * 6 / 1000, 1)
    over = sum(round(rate[j] * q[j] / 1000, 1) for j in range(1, 9))
    return round(base + over)


SC = {}
for biz in ("公共", "農集"):
    fp, L = M[biz]["財政計画"], {k: v / 1000 for k, v in ACT[biz]["損益"].items()}
    a_mnt = (L["下水道事業費用"] - L["特別損失"] - L["減価償却費"]
             - L["支払利息"] - L["その他営業外費用"])
    SC[biz] = {"補正率": L["下水道使用料"] / _gross(biz, RATES["現行"], 0),
               "汚水処理費": a_mnt / fp["汚水処理費"][0],
               "減価償却費": L["減価償却費"] / fp["減価償却費"][0],
               "支払利息": L["支払利息"] / fp["支払利息"][0],
               "補助金": L["他会計補助金"] / fp["補助金"][0],
               "長期前受金戻入": L["長期前受金戻入"] / fp["長期前受金戻入"][0],
               "その他営業収益": L["その他営業収益"],
               "その他営業外収益": L["営業外収益"] - L["他会計補助金"] - L["長期前受金戻入"]}

r = 1
rb.cell(r, 1, "令和7年度決算を起点に組み直した場合の試算（暫定）").font = Font(bold=True, size=12)
r += 1
for line in ("・本シートは決算突合の結果を機械的に将来へ反映した「試算」です。第2回審議会資料（37ページ）には反映していません。",
             "・使用料収入：基本使用料の年間乗数を公共・農集とも6回（隔月請求）として再計算し、"
             "R7実績÷R7算定値の補正率（公共1.005・農集1.015）を全年度に乗じています。",
             "・費用・繰入金等：R7実績÷R7推計の比率を全年度に乗じています（伸び率は経営戦略の想定を踏襲）。",
             "・経常収支比率は一般会計繰入金の前提で大きく変わるため、2ケースを併記しています。"):
    rb.cell(r, 1, line).font = Font(size=9, color="475569")
    r += 1
r += 1

rb.cell(r, 1, "■ 推計値と実績から得た換算率").font = BOLD
r += 1
for j, h in enumerate(["項目", "公共下水道", "農業集落排水", "内容"], start=1):
    c = rb.cell(r, j, h)
    c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
r += 1
for i, (name, memo) in enumerate((("補正率", "R7実績使用料 ÷ R7算定値（基本×6）"),
                                  ("汚水処理費", "R7実績 ÷ R7推計"),
                                  ("減価償却費", "同上"), ("支払利息", "同上"),
                                  ("補助金", "同上（農集は基準外繰入を含むとみられる）"),
                                  ("長期前受金戻入", "同上"))):
    rb.cell(r, 1, name).font = BODY
    for j, biz in ((2, "公共"), (3, "農集")):
        c = rb.cell(r, j, round(SC[biz][name], 4))
        c.number_format, c.alignment, c.font = "0.000", Alignment(horizontal="right"), BODY
    rb.cell(r, 4, memo).font = Font(size=9, color="475569")
    for j in range(1, 5):
        rb.cell(r, j).border = BOX
        if i % 2:
            rb.cell(r, j).fill = BAND
    r += 1
r += 1

CASES = [("現行（改定なし）", "現行", "2"), ("パターン①2か年", "①", "2"),
         ("パターン②2か年", "②", "2"), ("パターン④2か年", "④", "2"),
         ("パターン①3か年", "①", "3"), ("パターン②3か年", "②", "3"),
         ("パターン④3か年", "④", "3")]
SHOW = ["R7", "R8", "R9", "R10", "R12"]

for biz in ("公共", "農集"):
    label = "公共下水道事業" if biz == "公共" else "農業集落排水事業"
    fp, s = M[biz]["財政計画"], SC[biz]
    rb.cell(r, 1, f"■ {label}　経費回収率　※（ ）内は現行資料の値").font = BOLD
    r += 1
    for j, h in enumerate(["ケース"] + SHOW, start=1):
        c = rb.cell(r, j, h)
        c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
    r += 1
    for i, (lab, p, span) in enumerate(CASES):
        key = "現行" if p == "現行" else p + span
        rb.cell(r, 1, lab).font = BOLD if p == "②" else BODY
        for j, y in enumerate(SHOW, start=2):
            k = YKEY.index(y)
            inc = _gross(biz, _rate(p, span, k), k) * s["補正率"]
            oz = fp["汚水処理費"][k] * s["汚水処理費"]
            c = rb.cell(r, j, f"{inc/oz*100:.1f}%　({M[biz]['経費回収率'][key][k]*100:.1f}%)")
            c.alignment, c.font = Alignment(horizontal="right"), (BOLD if p == "②" else BODY)
        for j in range(1, len(SHOW) + 2):
            rb.cell(r, j).border = BOX
            if i % 2:
                rb.cell(r, j).fill = BAND
        r += 1
    r += 1

    for mode, sub in (("計画どおり（資本費相当）", 1.0), ("R7実績水準を維持", s["補助金"])):
        rb.cell(r, 1, f"■ {label}　経常収支比率　一般会計繰入金＝{mode}　※（ ）内は現行資料の値").font = BOLD
        r += 1
        for j, h in enumerate(["ケース"] + SHOW, start=1):
            c = rb.cell(r, j, h)
            c.fill, c.font, c.alignment, c.border = NAVY, HEAD, Alignment(horizontal="center"), BOX
        r += 1
        for i, (lab, p, span) in enumerate(CASES[:4]):
            key = "現行" if p == "現行" else p + span
            rb.cell(r, 1, lab).font = BOLD if p == "②" else BODY
            for j, y in enumerate(SHOW, start=2):
                k = YKEY.index(y)
                cin = (_gross(biz, _rate(p, span, k), k) * s["補正率"] + s["その他営業収益"]
                       + fp["補助金"][k] * sub + fp["長期前受金戻入"][k] * s["長期前受金戻入"]
                       + s["その他営業外収益"])
                cout = (fp["汚水処理費"][k] * s["汚水処理費"] + fp["減価償却費"][k] * s["減価償却費"]
                        + fp["支払利息"][k] * s["支払利息"])
                c = rb.cell(r, j, f"{cin/cout*100:.1f}%　({M[biz]['経常収支比率'][key][k]*100:.1f}%)")
                c.alignment, c.font = Alignment(horizontal="right"), (BOLD if p == "②" else BODY)
            for j in range(1, len(SHOW) + 2):
                rb.cell(r, j).border = BOX
                if i % 2:
                    rb.cell(r, j).fill = BAND
            r += 1
        r += 1

for line in ("※ R7列が決算実績（公共 経費回収率45.3%・経常収支比率104.1%／農集 37.2%・112.6%）を再現することで、"
             "換算方法の妥当性を確認しています。",
             "※ 有収水量・調定口数の実績が未受領のため、水量・口数は経営戦略の推計値を使用しています。"
             "実績を受領し次第、補正率を再計算する必要があります。"):
    rb.cell(r, 1, line).font = Font(size=9, color="B45309")
    r += 1

rb.column_dimensions["A"].width = 22
for j in range(2, 8):
    rb.column_dimensions[get_column_letter(j)].width = 18

wb.save(DST)
print("saved", DST)
