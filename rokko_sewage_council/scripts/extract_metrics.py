# -*- coding: utf-8 -*-
"""使用料改定シミュレーション（公共・農集）から data/metrics.json を再生成する.

各ブックの「パターン別比較」シートは数式で組まれているため、まず LibreOffice で
再計算させてから値を読み出す。人口推計は両事業とも「社人研」に揃える
（原本は公共＝社人研・農集＝人口ビジョン。差は0.1ポイント以内）。

必要なもの: libreoffice-calc, openpyxl
"""
import json
import pathlib
import subprocess
import tempfile

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent.parent
BOOKS = {"公共": HERE / "source" / "六戸町_公共_使用料改定.xlsx",
         "農集": HERE / "source" / "六戸町_農集_使用料改定.xlsx"}
OUT = HERE / "data" / "metrics.json"
POPULATION = "社人研"

KEYMAP = {"現行（改定なし）": "現行",
          "パターン①（標準型）2か年": "①2", "パターン①（標準型）3か年": "①3",
          "パターン②（家庭軽減型）2か年": "②2", "パターン②（家庭軽減型）3か年": "②3",
          "パターン④（段階累進型）2か年": "④2", "パターン④（段階累進型）3か年": "④3"}

# 「財政計画（ベース）」から拾う費用項目（左端の見出し → JSONのキー）
COSTMAP = {"経常支出 (D)": "経常支出", "(1) 職員給与費": "職員給与費", "(2) 経費": "経費", "(1) 経費": "経費",
           "(3) 減価償却費": "減価償却費", "(2) 減価償却費": "減価償却費",
           "(1) 支払利息": "支払利息", "(2) その他": "営業外費用その他",
           "(2) 長期前受金戻入": "長期前受金戻入", "(1) 補助金": "補助金",
           "(2) 雨水処理負担金": "雨水処理負担金",
           "汚水処理費": "汚水処理費", "有収水量（㎥）": "有収水量"}

# xlsx を読み込むとき LibreOffice は既定で再計算しない。強制するプロファイルを作る。
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
  xmlns:xs="http://www.w3.org/2001/XMLSchema"
  xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
<item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""


def recalculate(paths, work):
    """LibreOffice で再計算させた xlsx を work/out に書き出す."""
    profile = work / "lo_profile" / "user"
    profile.mkdir(parents=True)
    (profile / "registrymodifications.xcu").write_text(RECALC_XCU, encoding="utf-8")
    subprocess.run(
        ["soffice", "--headless", "--norestore",
         f"-env:UserInstallation=file://{profile.parent}",
         "--convert-to", "xlsx", "--outdir", str(work / "out"), *map(str, paths)],
        check=True, capture_output=True)
    return work / "out"


def main():
    with tempfile.TemporaryDirectory() as tmp:
        work = pathlib.Path(tmp)
        staged = []
        for label, src in BOOKS.items():
            wb = openpyxl.load_workbook(src)
            ws = wb["設定"]
            for row in ws.iter_rows():                     # 人口推計を揃える
                for cell in row:
                    if cell.value == "① 人口推計":
                        ws.cell(cell.row, cell.column + 1).value = POPULATION
            path = work / f"{label}.xlsx"
            wb.save(path)
            staged.append(path)

        out_dir = recalculate(staged, work)
        metrics = {}
        for label in BOOKS:
            ws = openpyxl.load_workbook(out_dir / f"{label}.xlsx", data_only=True)["パターン別比較"]
            section, table = None, {"経常収支比率": {}, "経費回収率": {}, "使用料収入": {}}
            for row in ws.iter_rows():
                head = (row[0].value or "")
                if "【経常収支比率】" in head:
                    section = "経常収支比率"
                elif "【経費回収率】" in head:
                    section = "経費回収率"
                elif "【参考】" in head and "使用料収入" in head:
                    section = "使用料収入"
                key = KEYMAP.get(head.strip()) if head else None
                if section and key:
                    table[section][key] = [c.value for c in row[1:11]]   # 令和7〜16年度

            # 汚水処理費の算定根拠となる費用項目（パターンによらず共通）
            fp = openpyxl.load_workbook(out_dir / f"{label}.xlsx", data_only=True)["財政計画（ベース）"]
            plan = {}
            for row in fp.iter_rows():
                head = row[0].value
                if isinstance(head, str):
                    name = COSTMAP.get(head.strip())
                    if name and name not in plan:
                        plan[name] = [c.value for c in row[1:11]]
            plan.setdefault("職員給与費", [0] * 10)          # 農集は職員給与費の行を持たない
            table["財政計画"] = plan
            metrics[label] = table

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(metrics, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"saved {OUT}")


if __name__ == "__main__":
    main()
