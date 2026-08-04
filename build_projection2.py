# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画　将来推計 第2段階（サービス見込量）.

第1段階（人口と認定者数）に続き、サービス見込量を算定する。

算定式
  ① 要介護度別の認定者数 ＝ 認定者数（第1段階）× 要介護度別の構成比
  ② 区分別の利用者数     ＝ 要介護度別の認定者数 × 区分別・要介護度別の利用率
  ③ 種別別の利用者数     ＝ 要介護度別の認定者数 × 種別別・要介護度別の利用率
  ④ 見込量               ＝ 種別別の利用者数 × 受給者1人当たりの利用日数・回数

置き方（発注者のご意向を令和8年8月4日に確認）
  基本ケース   利用率及び1人当たり利用日数・回数は令和7年度の値で固定する
  感度分析     直近3年（令和5〜7年度）の趨勢を第10期の3年間に延長する
  認定者数     第1段階の②トレンド継続を基本とし、①③を併記する
  要介護度別の構成比は令和8年3月末の実績で固定する（趨勢は感度分析で示す）

データの出所
  認定者数        将来推計 第1段階（build_projection.py）
  要介護度別構成比 見える化B3-a（令和8年3月末）
  利用率          見える化D45-a〜c（令和7年度＝令和8年1月サービス提供分まで）
  種別別受給率    見える化D32-a〜s（同上）
  1人当たり日数   見える化D46-a〜n（同上）

シート構成
  00_算定の方法と前提
  01_要介護度別の認定者数の見込み
  02_区分別の利用者数の見込み
  03_サービス種別の利用者数の見込み
  04_サービス見込量
  05_感度分析
  06_施設・居住系の定員との突合
  07_残る作業と確認事項
"""

import io
import runpy
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import data_jukyu as J
import data_nintei as N
import data_survey2025 as S

OUT = "/home/user/repository/output/第10期計画_将来推計_第2段階_サービス見込量.xlsx"

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

Y7 = "R7（R8/1月サービス提供分まで）"
CARE = ["要支援1", "要支援2", "要介護1", "要介護2", "要介護3", "要介護4",
        "要介護5"]
YEARS = [("2027", "令和9年度"), ("2028", "令和10年度"), ("2029", "令和11年度")]
POP1 = 9082                      # 令和7年度の第1号被保険者数（D2〜D4の分母）

# ---------------------------------------------------------------- 第1段階
_buf, _old = io.StringIO(), sys.stdout
sys.stdout = _buf
try:
    _g = runpy.run_path("build_projection.py")
finally:
    sys.stdout = _old
total = _g["total"]
SC = _g["SC"]

# ---------------------------------------------------------------- 構成比
NIN8 = {}
for k, v in N.B["B3-a"]["系列"].items():
    if not k.startswith("認定者数（"):
        continue
    lab = (k.replace("認定者数（", "").replace("）", "")
           .replace("１", "1").replace("２", "2").replace("３", "3")
           .replace("４", "4").replace("５", "5"))
    NIN8[lab] = v["令和8年3月末"]
BASE_N = sum(NIN8[c] for c in CARE)
MIX = {c: NIN8[c] / BASE_N for c in CARE}


def jv(code, part, year=Y7):
    for k, d in J.D[code]["系列"].items():
        if part in k:
            return d["値"].get(year)
    return None


def jyears(code, part):
    for k, d in J.D[code]["系列"].items():
        if part in k:
            return d["値"]
    return {}


# 区分別・要介護度別の利用率（％。分母は認定者数）
USE = {}
for code, kub in [("D45-a", "在宅"), ("D45-b", "居住系"), ("D45-c", "施設")]:
    USE[kub] = {c: (jv(code, "(%s)" % c) or 0.0) for c in CARE}

# 種別別・要介護度別の利用率（％。受給率×第1号被保険者数÷認定者数）
SVC = [
    ("D32-a", "訪問介護", "D46-a", "回"),
    ("D32-b", "訪問入浴介護", "D46-b", "回"),
    ("D32-c", "訪問看護", "D46-c", "回"),
    ("D32-d", "訪問リハビリテーション", "D46-d", "回"),
    ("D32-e", "居宅療養管理指導", None, ""),
    ("D32-f", "通所介護", "D46-e", "日"),
    ("D32-g", "通所リハビリテーション", "D46-f", "日"),
    ("D32-h", "短期入所生活介護", "D46-g", "日"),
    ("D32-i", "短期入所療養介護", "D46-h", "日"),
    ("D32-j", "福祉用具貸与", None, ""),
    ("D32-n", "小規模多機能型居宅介護", None, ""),
    ("D32-s", "地域密着型通所介護", "D46-j", "回"),
]


def svc_rate(code, c):
    """種別別・要介護度別の月平均受給者数（人）。"""
    d = J.D[code]["系列"]
    for k in d:
        if k == c or k.endswith("（%s）" % c):
            v = d[k]["値"].get(Y7)
            return (v or 0.0) * POP1 / 100
    return 0.0


SVCUSE = {}
for code, nm, _c46, _u in SVC:
    SVCUSE[nm] = {c: (svc_rate(code, c) / NIN8[c] * 100 if NIN8[c] else 0.0)
                  for c in CARE}

# 1人当たり利用日数・回数（要介護度別）
PER = {}
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    PER[nm] = {}
    for c in CARE:
        v = None
        for k, d in J.D[c46]["系列"].items():
            if k.endswith("(%s)" % c):
                v = d["値"].get(Y7)
                break
        PER[nm][c] = v


def nin(y, sc=2):
    """年度・シナリオ別の認定者数（要介護度別）。"""
    t = total(y, sc)
    return {c: t * MIX[c] for c in CARE}


# ---------------------------------------------------------------- 感度（趨勢）
def trend3(vals, keys):
    """直近3時点の年平均変化。"""
    xs = [vals.get(k) for k in keys]
    xs = [x for x in xs if x is not None]
    if len(xs) < 2:
        return 0.0
    return (xs[-1] - xs[0]) / (len(xs) - 1)


YR = list(J.D["D45-a"]["系列"]["サービス利用率（在宅サービス）(合計)"]["値"])
LAST3 = YR[-3:]
TREND_USE = {}
for code, kub in [("D45-a", "在宅"), ("D45-b", "居住系"), ("D45-c", "施設")]:
    TREND_USE[kub] = {c: trend3(jyears(code, "(%s)" % c), LAST3) for c in CARE}
TREND_PER = {}
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    TREND_PER[nm] = {}
    for c in CARE:
        d = {}
        for k, dd in J.D[c46]["系列"].items():
            if k.endswith("(%s)" % c):
                d = dd["値"]
                break
        TREND_PER[nm][c] = trend3(d, LAST3)


def use_rate(kub, c, y, sens=False):
    r = USE[kub][c]
    if sens:
        d = int(y) - 2025
        r = max(0.0, r + TREND_USE[kub][c] * d)
    return r


def per_val(nm, c, y, sens=False):
    v = PER.get(nm, {}).get(c)
    if v is None:
        return None
    if sens:
        d = int(y) - 2025
        v = max(0.0, v + TREND_PER[nm][c] * d)
    return v


# ================================================================ 体裁
wb = Workbook()


def sheet(name, title, subtitle, widths, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 22
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(widths))
    ws.row_dimensions[2].height = 48
    ws.freeze_panes = freeze
    return ws


def header(ws, row, cols, height=30):
    for i, v in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=HEAD, end_color=HEAD, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def body(ws, row, vals, fills=None, height=20, align=None, bold=False):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal=(align or {}).get(i, "left"))
        c.border = BORDER
        if fills and i in fills and fills[i]:
            c.fill = PatternFill(start_color=fills[i], end_color=fills[i],
                                 fill_type="solid")
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span, height=104):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = height
    return row + 1


RIGHT = {i: "right" for i in range(2, 26)}

# ================================================================ 00
ws = sheet("00_算定の方法と前提", "将来推計 第2段階（サービス見込量）",
           "第1段階（人口と認定者数）に続き、サービス見込量を算定します。"
           "利用率及び受給者1人当たりの利用日数・回数は令和7年度の値で固定することを"
           "基本ケースとし、直近3年の趨勢を延長した場合を感度分析として併記します"
           "（令和8年8月4日に発注者のご意向を確認）。"
           "作成日：令和8年8月4日。",
           [4, 24, 52, 48, 16], freeze="A4")

r = header(ws, 4, ["No.", "段階", "算定の内容", "置き方", "出所"])
for i, (k, cont, how, src) in enumerate([
    ("① 認定者数",
     "第1段階により令和9〜11年度の認定者数を3シナリオで推計済み。"
     "②トレンド継続では令和9年度1,995人・令和10年度2,000人・"
     "令和11年度2,004人。",
     "②トレンド継続を基本とし、①現状固定・③令和元年水準へ回帰を併記する。"
     "シナリオの選択は発注者のご意向による（ヒアリングシート項目6-2）。",
     "将来推計 03・04シート"),
    ("② 要介護度別の構成比",
     "認定者数を要介護度別に配分する。"
     "令和8年3月末の実績（要支援1 15.0％・要支援2 14.7％・要介護1 26.0％・"
     "要介護2 15.1％・要介護3 10.7％・要介護4 10.7％・要介護5 7.8％）による。",
     "令和8年3月末の構成比で固定する。"
     "要介護1への集中の趨勢を見込む場合は感度分析（05シート）で示す。",
     "見える化B3-a"),
    ("③ 区分別の利用率",
     "在宅・居住系・施設の別に、要介護度別の利用率を乗じて利用者数を求める。"
     "令和7年度は在宅51.2％・居住系7.3％・施設17.0％。",
     "令和7年度の値で固定する（基本ケース）。"
     "直近3年の趨勢を延長した場合を感度分析で示す。",
     "見える化D45-a〜c"),
    ("④ 種別別の利用率",
     "在宅サービスを12種別に分解する。"
     "種別別・要介護度別の受給率（第1号被保険者を分母）を"
     "認定者を分母とする利用率に換算して用いる。",
     "同上。",
     "見える化D32-a〜s"),
    ("⑤ 1人当たり利用日数・回数",
     "利用者数に受給者1人当たりの月間の利用日数又は回数を乗じて"
     "見込量を求める。訪問介護は令和7年度55回、通所介護は9日。",
     "令和7年度の値で固定する（基本ケース）。"
     "直近3年の趨勢を延長した場合を感度分析で示す。"
     "訪問介護は11年で2.5倍になっており、感度分析の差が最も大きい。",
     "見える化D46-a〜n"),
    ("⑥ 定員による上限",
     "施設・居住系サービスは定員を超えて利用できない。"
     "居所変更実態調査により定員652人・入所者536人を把握している。",
     "見込量が定員を超える場合は、定員を上限とするか整備を見込むかを"
     "第6章第4節で判断する（06シート）。",
     "3調査の点検 05シート"),
], start=1):
    r = body(ws, r, [str(i), k, cont, how, src], height=98,
             align={1: "center"})

r += 1
r = lead(ws, r, "【基本ケースの結果（シナリオ②トレンド継続）】", 5)
r = header(ws, r, ["区分", "令和7年度\n（実績）", "令和9年度", "令和10年度",
                   "令和11年度"])
ACT = {"在宅": 1012.8, "居住系": 143.8, "施設": 336.4}
for kub in ["在宅", "居住系", "施設"]:
    vals = []
    for y, _lab in YEARS:
        n = nin(y)
        vals.append(sum(n[c] * use_rate(kub, c, y) / 100 for c in CARE))
    r = body(ws, r, [kub + "サービスの利用者数（人／月）",
                     "%.0f" % ACT[kub]] + ["%.0f" % v for v in vals],
             height=20, align=RIGHT)
tot_act = sum(ACT.values())
tot_y = []
for y, _lab in YEARS:
    n = nin(y)
    tot_y.append(sum(n[c] * use_rate(k, c, y) / 100
                     for c in CARE for k in ["在宅", "居住系", "施設"]))
r = body(ws, r, ["計", "%.0f" % tot_act] + ["%.0f" % v for v in tot_y],
         {1: GRAY}, height=20, align=RIGHT, bold=True)
r = body(ws, r, ["認定者数（シナリオ②）", "1,984"]
         + ["%.0f" % total(y, 2) for y, _ in YEARS], {1: GRAY}, height=20,
         align=RIGHT)
r = body(ws, r, ["未利用者", "%.0f" % (1984 - tot_act)]
         + ["%.0f" % (total(y, 2) - v) for (y, _), v in zip(YEARS, tot_y)],
         {1: GRAY}, height=20, align=RIGHT)

note(ws, r + 1,
     "注1）認定者数がほぼ横ばい（令和9年度1,995人→令和11年度2,004人）であるため、"
     "利用率を固定した基本ケースでは利用者数もほぼ横ばいとなる。"
     "見込量が動くのは主として1人当たり利用日数・回数の趨勢による（05シート）。"
     "注2）本推計は第1段階のシナリオ②（トレンド継続）を基本としている。"
     "シナリオ①（現状固定）では令和11年度の認定者数が2,044人、"
     "シナリオ③（令和元年水準へ回帰）では2,112人となり、"
     "利用者数もその分増える。"
     "注3）要介護度別の構成比を令和8年3月末で固定しているため、"
     "要介護1への集中（7年間の増加125人のうち91人）が続く場合は"
     "在宅の利用者数が上振れする。05シートで感度を示している。"
     "注4）本書は第2段階の試算であり、決定した見込量ではない。"
     "利用率と1人当たり利用日数・回数の置き方は"
     "ヒアリングシート項目6-3・6-4でご意向を確認する。", 5)

# ================================================================ 01
ws = sheet("01_要介護度別の認定者数の見込み", "要介護度別の認定者数の見込み",
           "第1段階で推計した認定者数を、令和8年3月末の要介護度別構成比により"
           "配分する。3シナリオを併記する。",
           [16, 12, 12, 12, 12, 12, 12, 12, 12, 12], freeze="B5")

r = lead(ws, 4, "【構成比（令和8年3月末の実績）】", 10)
r = header(ws, r, ["要介護度"] + CARE + ["計", ""])
r = body(ws, r, ["認定者数（人）"] + [int(NIN8[c]) for c in CARE]
         + [int(BASE_N), ""], height=20, align=RIGHT)
r = body(ws, r, ["構成比（％）"] + ["%.1f" % (MIX[c] * 100) for c in CARE]
         + ["100.0", ""], {1: GRAY}, height=20, align=RIGHT)
r += 1

for sc in (1, 2, 3):
    r = lead(ws, r, "【シナリオ%s】" % SC[sc], 10)
    r = header(ws, r, ["年度"] + CARE + ["計", ""])
    for y, lab in [("2025", "令和7年度（実績）")] + YEARS:
        n = nin(y, sc)
        r = body(ws, r, [lab] + ["%.0f" % n[c] for c in CARE]
                 + ["%.0f" % sum(n.values()), ""],
                 {1: GRAY if "実績" in lab else (OK_G if sc == 2 else None)},
                 height=20, align=RIGHT)
    r += 1

note(ws, r,
     "注1）令和7年度は令和8年3月末の実績（1,984人）である。"
     "注2）シナリオ②（トレンド継続）を基本とする。"
     "第1段階の算定方法は将来推計 03・04シートに示している。"
     "注3）要介護度別の構成比は令和8年3月末で固定している。"
     "実際には要介護1が増え続けており（7年間で91人増、増加分の72.8％）、"
     "この趨勢が続く場合は要介護1の人数が上振れする。"
     "05シートで感度を示している。"
     "注4）本表の認定者数は第1号被保険者のみである。"
     "第2号被保険者（令和8年3月末36人・1.8％）は含まない。"
     "見込量に第2号を含める場合は1.8％を上乗せする。", 10)

# ================================================================ 02
ws = sheet("02_区分別の利用者数の見込み", "区分別（在宅・居住系・施設）の利用者数の見込み",
           "要介護度別の認定者数に、区分別・要介護度別の利用率（令和7年度・固定）を"
           "乗じて利用者数を求める。単位は人／月。",
           [16, 12] + [11] * 8 + [12], freeze="B5")

r = lead(ws, 4, "【利用率（令和7年度・％。分母は認定者数）】", 11)
r = header(ws, r, ["区分"] + CARE + ["合計", "", ""])
for kub, code in [("在宅", "D45-a"), ("居住系", "D45-b"), ("施設", "D45-c")]:
    r = body(ws, r, [kub] + [USE[kub][c] for c in CARE]
             + [jv(code, "(合計)"), "", ""], height=20, align=RIGHT)
r = body(ws, r, ["3区分の計"]
         + ["%.1f" % sum(USE[k][c] for k in USE) for c in CARE]
         + ["%.1f" % sum(jv(cd, "(合計)") for cd in
                         ["D45-a", "D45-b", "D45-c"]), "", ""],
         {1: GRAY}, height=20, align=RIGHT, bold=True)
r += 1

for kub in ["在宅", "居住系", "施設"]:
    r = lead(ws, r, "【%sサービスの利用者数（人／月）】" % kub, 11)
    r = header(ws, r, ["年度"] + CARE + ["計", "", ""])
    for y, lab in YEARS:
        n = nin(y)
        vs = [n[c] * USE[kub][c] / 100 for c in CARE]
        r = body(ws, r, [lab] + ["%.1f" % v for v in vs]
                 + ["%.1f" % sum(vs), "", ""], height=20, align=RIGHT)
    r += 1

note(ws, r,
     "注1）利用率は令和7年度（令和7年4月〜令和8年1月）の値で固定している。"
     "注2）3区分の利用率の合計は75.5％であり、"
     "認定者の24.5％はいずれのサービスも利用していない。"
     "この未利用率も令和7年度の水準で固定している。"
     "第5章の施策により利用率を上げる目標を立てる場合は、"
     "見込量に上乗せする必要がある。"
     "注3）在宅と居住系は重複しうる（居住系の入居者が訪問系を利用する等）。"
     "見える化の定義に従い、そのまま合計している。"
     "注4）施設サービスは要支援では0である。", 11)

# ================================================================ 03
ws = sheet("03_サービス種別の利用者数の見込み", "サービス種別の利用者数の見込み",
           "在宅サービスを12種別に分解する。"
           "種別別・要介護度別の受給率（見える化D32）を"
           "認定者を分母とする利用率に換算して用いる。単位は人／月。",
           [26, 12] + [10] * 7 + [12, 12, 12], freeze="B5")

r = lead(ws, 4, "【種別別・要介護度別の利用率（令和7年度・％。分母は認定者数）】", 13)
r = header(ws, r, ["サービス種別"] + CARE + ["令和7年度の\n利用者数（人／月）",
                                            "", "", ""])
for _code, nm, _c46, _u in SVC:
    vs = [SVCUSE[nm][c] for c in CARE]
    act = sum(NIN8[c] * SVCUSE[nm][c] / 100 for c in CARE)
    r = body(ws, r, [nm] + ["%.1f" % v for v in vs] + ["%.1f" % act, "", "",
                                                       ""],
             height=20, align=RIGHT)
r += 1

for y, lab in YEARS:
    r = lead(ws, r, "【%s の利用者数（人／月）】" % lab, 13)
    r = header(ws, r, ["サービス種別"] + CARE + ["計", "", "", ""])
    n = nin(y)
    for _code, nm, _c46, _u in SVC:
        vs = [n[c] * SVCUSE[nm][c] / 100 for c in CARE]
        r = body(ws, r, [nm] + ["%.1f" % v for v in vs]
                 + ["%.1f" % sum(vs), "", "", ""], height=20, align=RIGHT)
    r += 1

note(ws, r,
     "注1）受給率は小数第1位までしか公表されていない。"
     "第1号被保険者9,082人に対し0.1％は9.1人であるため、"
     "種別別の利用者数には±5人程度の丸めの影響がある。"
     "訪問介護・訪問看護・通所介護は受給者数の実数が公表されているため、"
     "換算値と実数が一致することを確認している。"
     "注2）定期巡回・随時対応型訪問介護看護、夜間対応型訪問介護、"
     "認知症対応型通所介護、看護小規模多機能型居宅介護は"
     "受給率0.0％のため本表に含めていない。"
     "第6章第4節で確保方策を検討する対象である。"
     "注3）居住系・施設サービスの種別別（グループホーム・特定施設・"
     "特別養護老人ホーム等）の内訳は見える化に系列がないため、"
     "02シートの区分別と06シートの定員により見込む。"
     "注4）在宅サービスの種別を合計しても02シートの在宅利用者数とは一致しない。"
     "1人が複数の種別を利用するためである。", 13)

# ================================================================ 04
ws = sheet("04_サービス見込量", "サービス見込量（基本ケース）",
           "利用者数に受給者1人当たりの月間の利用日数又は回数を乗じて求める。"
           "1人当たりの値は令和7年度で固定している。",
           [26, 8, 14, 14, 14, 14, 12, 12, 12, 12], freeze="B5")

r = header(ws, 4, ["サービス種別", "単位", "令和7年度\n（実績相当）",
                   "令和9年度", "令和10年度", "令和11年度",
                   "令和7→11の\n増減率", "備考", "", ""])
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    act = sum(NIN8[c] * SVCUSE[nm][c] / 100 * (PER[nm][c] or 0) for c in CARE)
    if act == 0:
        r = body(ws, r, [nm, unit + "／月", "―", "―", "―", "―", "―",
                         "要介護度別の受給率がいずれも0.0％のため算定できない",
                         "", ""], {3: GRAY}, height=20, align=RIGHT)
        continue
    vals = []
    for y, _lab in YEARS:
        n = nin(y)
        vals.append(sum(n[c] * SVCUSE[nm][c] / 100 * (PER[nm][c] or 0)
                        for c in CARE))
    rate = (vals[-1] / act - 1) * 100 if act else 0
    r = body(ws, r, [nm, unit + "／月", "%.0f" % act]
             + ["%.0f" % v for v in vals] + ["%+.1f％" % rate, "", "", ""],
             height=20, align=RIGHT)

r += 1
r = lead(ws, r, "【1人当たりの利用日数・回数（令和7年度・固定）】", 10)
r = header(ws, r, ["サービス種別", "単位"] + CARE + ["合計"])
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    tot = None
    for k, d in J.D[c46]["系列"].items():
        if k.endswith("(合計)"):
            tot = d["値"].get(Y7)
            break
    r = body(ws, r, [nm, unit]
             + ["―" if PER[nm][c] is None else PER[nm][c] for c in CARE]
             + [tot], height=20, align=RIGHT)

note(ws, r + 1,
     "注0）「令和7年度（実績相当）」は、受給率（小数第1位）から逆算した"
     "要介護度別の利用者数に1人当たりの値を乗じたものである。"
     "受給者数の実数から直接計算した値（訪問介護の場合、"
     "延べ2,971人÷11か月×55回＝14,850回／月）とは数％の差がある。"
     "丸めと要介護度別の重みづけの違いによるものであり、"
     "第6章第2節に掲載する際は実数のある3種別は実数により算定する。"
     "注1）本表は基本ケース（利用率・1人当たりの値をいずれも令和7年度で固定）である。"
     "認定者数がほぼ横ばいであるため、見込量もほぼ横ばいとなる。"
     "注2）1人当たりの値の趨勢を見込む場合は05シートの感度分析による。"
     "訪問介護は平成26年度22回から令和7年度55回へ2.5倍になっており、"
     "趨勢を延長すると見込量が大きく変わる。"
     "注3）居宅療養管理指導・福祉用具貸与・小規模多機能型居宅介護は"
     "1人当たりの利用日数・回数の系列がないため、利用者数のみを示す（03シート）。"
     "計画本文では、福祉用具貸与は利用者数、小規模多機能は登録者数により見込む。"
     "注4）短期入所生活介護・短期入所療養介護は、"
     "1人当たりの利用日数に利用者数を乗じた延べ日数である。"
     "第6章第2節では「延べ利用日数／月」として掲載する。", 10)

# ================================================================ 05
ws = sheet("05_感度分析", "感度分析（趨勢を見込んだ場合）",
           "基本ケース（令和7年度で固定）に対し、直近3年（令和5〜7年度）の"
           "年平均変化を第10期の3年間に延長した場合を示す。"
           "どの前提が見込量を最も動かすかを確認する。",
           [26, 14, 14, 14, 14, 14, 34], freeze="B5")

r = lead(ws, 4, "【① 区分別の利用者数（人／月・令和11年度）】", 7)
r = header(ws, r, ["区分", "基本ケース", "利用率に趨勢",
                   "差", "差の割合", "直近3年の年平均変化", "見方"])
for kub in ["在宅", "居住系", "施設"]:
    n = nin("2029")
    base = sum(n[c] * USE[kub][c] / 100 for c in CARE)
    sens = sum(n[c] * use_rate(kub, c, "2029", True) / 100 for c in CARE)
    tr = sum(TREND_USE[kub][c] for c in CARE) / len(CARE)
    r = body(ws, r, [kub, "%.0f" % base, "%.0f" % sens,
                     "%+.0f" % (sens - base),
                     "%+.1f％" % ((sens / base - 1) * 100) if base else "―",
                     "%+.2fポイント／年" % tr,
                     "利用率が上がる方向" if tr > 0 else "利用率が下がる方向"],
             {4: NG_O if abs(sens / base - 1) > 0.05 else None}, height=20,
             align=RIGHT)
r += 1

r = lead(ws, r, "【② サービス見込量（令和11年度）】", 7)
r = header(ws, r, ["サービス種別", "単位", "基本ケース",
                   "1人当たりに趨勢", "差の割合",
                   "1人当たり（合計）の推移\n令和5→7年度", "見方"])
for _code, nm, c46, unit in SVC:
    if not c46:
        continue
    n = nin("2029")
    base = sum(n[c] * SVCUSE[nm][c] / 100 * (PER[nm][c] or 0) for c in CARE)
    if base == 0:
        continue
    sens = sum(n[c] * SVCUSE[nm][c] / 100 * (per_val(nm, c, "2029", True) or 0)
               for c in CARE)
    tot = {}
    for k, d in J.D[c46]["系列"].items():
        if k.endswith("(合計)"):
            tot = d["値"]
            break
    seq = [tot.get(y) for y in LAST3]
    trs = "→".join("%g" % v if v is not None else "―" for v in seq)
    dr = (sens / base - 1) * 100 if base else 0
    small = base < 300
    r = body(ws, r, [nm, unit + "／月", "%.0f" % base, "%.0f" % sens,
                     "%+.1f％" % dr, trs,
                     "実数が小さく趨勢が振れやすい" if small
                     else ("感度が大きい" if abs(dr) >= 10 else "")],
             {5: NG_O if abs(dr) >= 10 and not small else
              (GRAY if small else None)}, height=20, align=RIGHT)
r += 1

r = lead(ws, r, "【③ 要介護度別の構成比（要介護1への集中）】", 7)
r = header(ws, r, ["前提", "要介護1の\n認定者数", "在宅の利用者数",
                   "訪問介護の見込量", "差の割合", "", "見方"])
n29 = nin("2029")
base_k1 = n29["要介護1"]
base_zai = sum(n29[c] * USE["在宅"][c] / 100 for c in CARE)
base_hou = sum(n29[c] * SVCUSE["訪問介護"][c] / 100 * (PER["訪問介護"][c] or 0)
               for c in CARE)
# 要介護1が7年間で91人増（年13人）の趨勢を4年延長し、要介護2〜5から按分して移す
SHIFT = 13.0 * 4
mix2 = dict(MIX)
tot_mix = sum(MIX[c] for c in ["要介護2", "要介護3", "要介護4", "要介護5"])
t29 = total("2029", 2)
n29b = dict(n29)
n29b["要介護1"] = n29["要介護1"] + SHIFT
for c in ["要介護2", "要介護3", "要介護4", "要介護5"]:
    n29b[c] = n29[c] - SHIFT * MIX[c] / tot_mix
sens_zai = sum(n29b[c] * USE["在宅"][c] / 100 for c in CARE)
sens_hou = sum(n29b[c] * SVCUSE["訪問介護"][c] / 100 * (PER["訪問介護"][c] or 0)
               for c in CARE)
r = body(ws, r, ["令和8年3月末の構成比で固定（基本）", "%.0f" % base_k1,
                 "%.0f" % base_zai, "%.0f" % base_hou, "―", "", ""],
         height=20, align=RIGHT)
r = body(ws, r, ["要介護1への集中が続く（年13人・4年で52人）",
                 "%.0f" % n29b["要介護1"], "%.0f" % sens_zai,
                 "%.0f" % sens_hou,
                 "在宅%+.1f％／訪問介護%+.1f％"
                 % ((sens_zai / base_zai - 1) * 100,
                    (sens_hou / base_hou - 1) * 100), "",
                 "在宅は増え、訪問介護の見込量は減る"], {5: IN_Y},
         height=20, align=RIGHT)
r += 1

r = lead(ws, r, "【④ 認定者数のシナリオ（令和11年度）】", 7)
r = header(ws, r, ["シナリオ", "認定者数", "在宅の利用者数", "施設の利用者数",
                   "訪問介護の見込量", "", "見方"])
for sc in (1, 2, 3):
    n = nin("2029", sc)
    z = sum(n[c] * USE["在宅"][c] / 100 for c in CARE)
    s = sum(n[c] * USE["施設"][c] / 100 for c in CARE)
    h = sum(n[c] * SVCUSE["訪問介護"][c] / 100 * (PER["訪問介護"][c] or 0)
            for c in CARE)
    r = body(ws, r, [SC[sc], "%.0f" % total("2029", sc), "%.0f" % z,
                     "%.0f" % s, "%.0f" % h, "",
                     "基本ケース" if sc == 2 else ""],
             {1: OK_G if sc == 2 else None}, height=20, align=RIGHT)

note(ws, r + 1,
     "注1）見込量を最も動かすのは1人当たりの利用日数・回数の趨勢である（②）。"
     "訪問介護は直近3年で年平均4回増えており、"
     "これを4年延長すると令和11年度の見込量が大きく変わる。"
     "注2）利用率の趨勢（①）と認定者数のシナリオ（④）の影響は"
     "いずれも数％にとどまる。"
     "注3）要介護度別の構成比（③）は、要介護1が増えると"
     "在宅の利用者数は増えるが訪問介護の見込量は減る。"
     "要介護1の1人当たり利用回数（24回）が要介護5（114回）より"
     "大幅に少ないためである。"
     "注3-2）1人当たりの趨勢は要介護度別に延長している。"
     "このため、実数の小さい種別（短期入所療養介護46日／月、"
     "通所介護688日／月等）では、要介護度別の振れが増幅され、"
     "合計の推移（短期入所療養介護は令和5〜7年度とも6日で横ばい）と"
     "乖離することがある。"
     "感度分析は実数の大きい訪問介護（15,645回／月）で読む。"
     "注4）本シートの趨勢は直近3年（令和5〜7年度）の年平均変化による。"
     "令和7年度は11か月分であるため、確定値が出た時点で再計算する。"
     "注5）感度分析は前提の影響の大きさを確認するためのものであり、"
     "いずれかを採用値とするものではない。"
     "採用値は発注者のご意向と委員会の審議による。", 7)

# ================================================================ 06
ws = sheet("06_施設・居住系の定員との突合", "施設・居住系サービスの定員との突合",
           "施設・居住系の見込量は定員を超えて実現しない。"
           "居所変更実態調査で把握した定員及び入所者数と突合する。"
           "特定施設の定員は、令和8年8月4日に受領した北海道の"
           "届出済有料老人ホーム一覧（令和8年7月1日現在）により"
           "58人から156人に改めた"
           "（住まいと施設の公表名簿との突合 03シート）。",
           [26, 12, 12, 12, 12, 12, 40], freeze="A5")

CAP = {}
for s_ in S.SHI:
    g = {1: "住宅型有料・サ高住", 3: "住宅型有料・サ高住", 4: "グループホーム",
         5: "特定施設", 7: "介護老人保健施設", 9: "特養・地域密着型特養",
         10: "特養・地域密着型特養"}.get(s_["種別"], "その他")
    d = CAP.setdefault(g, {"施設数": 0, "定員": 0, "入所": 0})
    d["施設数"] += 1
    d["定員"] += s_["定員"] or 0
    d["入所"] += s_["入所"] or 0

# 名簿による区域内の定員（住まいと施設の公表名簿との突合 02シート）
CAP_TOKUTEI = 156          # 特定施設　3施設（届出済有料老人ホーム一覧）
CAP_HOKENGAI = 223 + 66 + 50   # 住宅型有料223人＋サ高住66戸＋軽費50人

r = header(ws, 4, ["区分", "調査の\n回答数", "調査の\n定員", "名簿等による\n定員",
                   "調査の\n入所者", "空き", "備考"])
CAP_TOKUYO = 160 + 62      # 特養160人＋地域密着型特養62人（特別養護老人ホーム名簿）

ROWS6 = [
    ("特養・地域密着型特養", 4, 6, CAP_TOKUYO,
     "北海道の特別養護老人ホーム名簿（令和8年7月1日現在）による"
     "介護老人福祉施設3施設160人＋地域密着型3施設62人。"
     "見える化K1a・K1bの各3事業所と一致する。"
     "調査には東神楽町特別養護老人ホームアゼリアハイツ"
     "（広域型50人・ユニット型20人）が回答していない"),
    ("介護老人保健施設", 3, None, None,
     "見える化K1c（3事業所）と一致し、区域内のすべてを把握している"),
    ("グループホーム", 5, 7, None,
     "見える化K2c（令和6年度5事業所）とは一致するが、"
     "北海道の介護保険事業所一覧（令和8年6月30日現在）は7事業所である。"
     "新たな2事業所（ファミリー・くるみの郷。いずれも東川町）は"
     "定員が不明であり、調査の81人に算入していない"),
    ("特定施設", 2, 3, CAP_TOKUTEI,
     "北海道の届出済有料老人ホーム一覧による3施設の定員。"
     "調査に回答したのは2施設（58人）で、"
     "さわやか東神楽館（東神楽町・100人）が回答していない。"
     "入所者数は回答した2施設分のみであり、入所率は算定できない"),
    ("住宅型有料・サ高住・軽費", 4, 12, CAP_HOKENGAI,
     "北海道の名簿による住宅型有料9施設223人・"
     "サービス付き高齢者向け住宅2件66戸・軽費老人ホーム1施設50人。"
     "調査に回答したのは4施設で、入所率は算定できない。"
     "介護保険の指定サービスではないため見込量には含まれない"),
]
tot_cap = tot_res = 0
for g, n_sur, n_ros, cap_ros, memo in ROWS6:
    d = CAP.get(g if g in CAP else "住宅型有料・サ高住")
    res = d["入所"] + (58 if g == "特養・地域密着型特養" else 0)
    cap = cap_ros if cap_ros is not None else d["定員"]
    if cap_ros is None:
        tot_res += res
    tot_cap += cap
    r = body(ws, r, [g, n_sur, d["定員"], cap_ros or "調査による",
                     res if cap_ros is None else "%d（%d施設分）"
                     % (res, n_sur),
                     cap - res if cap_ros is None else "―", memo],
             {4: OK_G if cap_ros else None}, height=44, align=RIGHT)
r = body(ws, r, ["計", 18, sum(d["定員"] for d in CAP.values()), tot_cap,
                 "―", "―",
                 "名簿等による定員は、介護保険の指定を受ける施設・居住系"
                 "（特養・老健・グループホーム・特定施設）と"
                 "指定を受けない住まいを合わせたものである"],
         {1: GRAY}, height=32, align=RIGHT, bold=True)

r += 1
r = lead(ws, r, "【見込量と定員の対照（令和11年度・基本ケース）】", 7)
r = header(ws, r, ["区分", "令和7年度\n（実績）", "令和11年度\n（見込み）",
                   "定員", "定員に対する割合", "判定", "見方"])
n29 = nin("2029")
for kub, cap in [("施設", CAP["介護老人保健施設"]["定員"] + CAP_TOKUYO),
                 ("居住系", CAP["グループホーム"]["定員"] + CAP_TOKUTEI)]:
    act = ACT[kub]
    v = sum(n29[c] * USE[kub][c] / 100 for c in CARE)
    r = body(ws, r, [kub + "サービス", "%.0f" % act, "%.0f" % v, cap,
                     "%.1f％" % (v / cap * 100),
                     "定員内" if v <= cap else "定員超過",
                     "定員に余裕がある" if v <= cap * 0.95
                     else "整備又は上限の判断を要する"],
             {6: OK_G if v <= cap else NG_O}, height=20, align=RIGHT)

note(ws, r + 1,
     "注0）居住系の定員には、令和8年6月30日現在で7事業所となっている"
     "グループホームのうち、調査で把握した5事業所81人のみを算入している。"
     "新たな2事業所（ファミリー・くるみの郷）の定員が不明であるため、"
     "定員の余裕は本表よりさらに大きい。"
     "注1）調査で把握した定員652人・入所者594人（美瑛慈光園の58人を含む）は、"
     "居所変更実態調査に回答した18施設分である。"
     "見える化K系列では特別養護老人ホーム3・地域密着型特別養護老人ホーム3・"
     "介護老人保健施設3・特定施設3・グループホーム5であり、"
     "特養と地域密着型特養で計2施設が未回答である（3調査の点検 10シート）。"
     "定員の全体像は未回答2施設を加える必要がある。"
     "注2）居住系サービスの区域内定員は237人"
     "（グループホーム81人＋特定施設156人）であり、"
     "令和11年度の見込量145人は定員内に収まる。"
     "特定施設の定員は当初、居所変更実態調査に回答した2施設分の58人と"
     "していたが、北海道の届出済有料老人ホーム一覧により"
     "3施設156人であることが確認された"
     "（さわやか東神楽館100人が調査に回答していない）。"
     "見える化K2a（特定施設入居者生活介護の事業所数3）とも一致する。"
     "定員に余裕があるため、居住系は定員による制約を受けない。"
     "なお区域内の施設には住所地特例により他の保険者の被保険者も入居しており、"
     "定員と見込量の差がそのまま空きを示すものではない。"
     "居所変更実態調査では特定施設の新規入所の84.2％が"
     "区域外からの入所であった。"
     "施設サービスは定員462人（特養160人・地域密着型特養62人・老健240人）に"
     "対し339人であり余裕がある。"
     "特養の定員は北海道の特別養護老人ホーム名簿により確定した"
     "（住まいと施設の公表名簿との突合 09シート）。"
     "注3）住宅型有料老人ホーム・サービス付高齢者向け住宅・軽費老人ホームは"
     "介護保険の指定サービスではないため、"
     "見込量（居住系サービス）には含まれない。"
     "北海道の名簿により、区域内には住宅型有料老人ホーム9施設・定員223人、"
     "サービス付き高齢者向け住宅2件・66戸、"
     "軽費老人ホーム1施設・50人があることが確認された。"
     "住まいの供給量として第2章第3節に掲載する"
     "（住まいと施設の公表名簿との突合 04シート）。"
     "注4）第6章第4節（施設等の整備の方針）では、"
     "待機者数（重複を除くと82人・3調査の点検No.13）と"
     "本表の空き（58人）を対照して判断する。", 7)

# ================================================================ 07
ws = sheet("07_残る作業と確認事項", "残る作業と確認事項",
           "第2段階の試算は完了したが、第6章第2節の見込量として確定するには"
           "残る作業と確認事項がある。",
           [4, 26, 50, 46, 12], freeze="A5")

r = header(ws, 4, ["No.", "項目", "内容", "取扱い", "区分"])
for no, k, cont, act, kd in [
    (1, "利用率と1人当たりの置き方",
     "基本ケースは令和7年度で固定、感度分析は直近3年の趨勢の延長としている。"
     "受託者は基本ケースを採用値とすることを案とする。",
     "ヒアリングシート項目6-3（利用率）・6-4（1人当たり利用日数・回数）で"
     "ご意向を確認する。認定率のシナリオ（項目6-2）と併せて確認する。",
     "確認"),
    (2, "居住系・施設サービスの種別別内訳",
     "グループホーム・特定施設・特別養護老人ホーム・介護老人保健施設の"
     "種別別の利用者数は、見える化に系列がないため算定していない。",
     "定員と入所者数により見込む（06シート）。"
     "第6章第2節の様式が種別別の見込量を求める場合は、"
     "介護保険事業状況報告の月報又は広域連合の給付実績による。", "確認"),
    (3, "介護予防・日常生活支援総合事業",
     "訪問型サービス・通所型サービスの実績が見える化にない。",
     "広域連合の実績によるほかない。"
     "第6章第2節に総合事業の見込量を掲載する場合はご提供いただく。",
     "確認"),
    (4, "第2号被保険者",
     "本推計は第1号被保険者のみである。"
     "第2号被保険者は令和8年3月末で36人（1.8％）である。",
     "見込量に第2号を含める場合は1.8％を上乗せする。"
     "第6章第1節に前提として明記する。", "留意"),
    (5, "未利用者の扱い",
     "認定者の24.5％はいずれのサービスも利用していない。"
     "基本ケースではこの割合を固定している。",
     "第5章の施策により利用率を上げる目標を立てる場合は、"
     "見込量に上乗せする必要がある。"
     "未利用の理由は資料依頼No.7で照会中である。", "確認"),
    (6, "令和7年度が11か月分であること",
     "利用率・1人当たり利用日数はいずれも令和7年度"
     "（令和7年4月〜令和8年1月）の値である。",
     "令和7年度の確定値が公表された時点で再計算する。"
     "第6章第2節の確定版（令和8年12月）までに間に合う見込みである。",
     "留意"),
    (7, "第3段階（給付費・保険料）",
     "見込量にサービス種別の単価を乗じて給付費を算定し、"
     "給付費から保険料を算定する。",
     "単価は報酬改定率の告示待ちである（告示待ちNo.4）。"
     "地域区分・処遇改善加算の取扱いも確認を要する。", "留意"),
    (8, "定員による制約",
     "基本ケースでは施設・居住系の見込量が定員内に収まるため、"
     "定員を制約として反映する必要は生じていない。",
     "認定率のシナリオ③（令和元年水準へ回帰）を採用する場合は"
     "定員に近づくため、改めて確認する。", "留意"),
]:
    fl = {"確認": IN_Y, "留意": MID_B}[kd]
    r = body(ws, r, [no, k, cont, act, kd], {5: fl}, height=88,
             align={1: "center", 5: "center"})

note(ws, r + 1,
     "注1）確認4件（No.1・No.2・No.3・No.5）のうち、"
     "No.1は見込量の値そのものを左右する。"
     "キックオフ会議でご意向を確認する。"
     "注2）本書は第2段階の試算であり、決定した見込量ではない。"
     "仕様書の定めにより、業務工程の都度発注者のご意向を確認しながら進める。"
     "注3）第3段階（給付費・保険料）は、"
     "本書の見込量に単価を乗じて算定する。"
     "報酬改定率の告示（令和8年秋）を待って着手する。", 5)

del wb["Sheet"]
wb.save(OUT)
print("saved:", OUT)
for ws in wb:
    print("  -", ws.title, ws.max_row, "rows")
