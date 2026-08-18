# -*- coding: utf-8 -*-
"""
業務進捗管理ブック（高齢者／障害の2計画）ジェネレータ

出力: output/業務進捗管理_高齢者・障害.xlsx

目的
  受託中の計画策定業務を、仕様書に記載された業務（＝各段階）の単位で
  1冊にまとめて管理する。計画分野ごとにシートを分け、
  サマリーで2計画を並べて見られるようにする。

2つの track
  障害   ： 委託第24号 第8期北塩原村障がい福祉計画・第4期北塩原村障がい児福祉計画策定業務
           仕様書（8北保福第366号 別紙）の条項に対応。
           build_kitashiobara_progress.py の管理データをそのまま取り込むため、
           本ブックと個別ブック（北塩原村_業務進捗管理.xlsx）で数値が食い違わない。
  高齢者 ： 介護保険事業計画・高齢者福祉計画の策定支援業務。
           本ワークスペースに当該業務の仕様書・進捗資料が存在しないため、
           介護保険法及び老人福祉法に基づく標準的な業務構成を骨格として置き、
           状態はすべて「仕様書未受領」としている。
           仕様書を受領しだい KOUREI_ITEMS を差し替える。
           推測で進捗率や実績を埋めることはしない。

注意
  高齢者シートの各行は「一般に計画策定支援業務に含まれる工程」であって、
  実際の契約が何を求めているかを示すものではない。
  進捗率0％・状態「仕様書未受領」は、作業が遅れているという意味ではなく、
  当方が契約内容を把握していないという意味である。
"""

import importlib.util
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from kitashiobara_common import (
    COLORS, FONT, OUT_DIR, add_sheet, ensure_out_dir,
    style_header_row, style_note, style_title, write_row,
)

OUT_FILE = f"{OUT_DIR}/業務進捗管理_高齢者・障害.xlsx"

PCT = "0%"
DATE = "yyyy/mm/dd"

TRACK_KOUREI = "高齢者"
TRACK_SHOGAI = "障害"

TRACK_COLOR = {
    TRACK_KOUREI: "2E75B6",   # 青
    TRACK_SHOGAI: "FF7F0E",   # オレンジ（障がい計画の配色規約）
}

# 進捗状態と色（障害トラックは個別ブックと同じ語彙を使う）
PROGRESS_FILL = {
    "完了": "2CA02C",
    "概ね完了": "70AD47",
    "作業中": "ED7D31",
    "継続実施": "2E75B6",
    "未着手": "808080",
    "村資料待ち": "C00000",
    "要協議": "7030A0",
    "村側作業": "9E480E",
    "仕様書未受領": "404040",
}


def _load_shogai():
    """北塩原村の進捗管理データを取り込む（二重管理を避けるため import する）。"""
    spec = importlib.util.spec_from_file_location(
        "kitashiobara_progress", "/home/user/repository/build_kitashiobara_progress.py")
    mod = importlib.util.module_from_spec(spec)
    sys.modules["kitashiobara_progress"] = mod
    spec.loader.exec_module(mod)
    return mod


SHOGAI = _load_shogai()

# ---------------------------------------------------------------------------
# 契約の基本情報
# 分野, 委託名, 発注者, 計画期間, 業務期間, 根拠法, 出典
# ---------------------------------------------------------------------------
CONTRACTS = [
    (TRACK_SHOGAI,
     "委託第24号　第8期北塩原村障がい福祉計画・第4期北塩原村障がい児福祉計画策定業務委託",
     "北塩原村（保健福祉課 福祉係）", "令和9年度〜令和11年度（3年間）",
     "契約締結の日から令和9年3月31日まで",
     "障害者総合支援法第88条（市町村障害福祉計画）／児童福祉法第33条の20（市町村障害児福祉計画）",
     "業務委託仕様書（8北保福第366号 別紙）"),
    (TRACK_KOUREI,
     "【要記入】高齢者分野の委託名",
     "【要記入】発注者", "【要確認】介護保険事業計画は3年1期。第9期が令和6〜8年度のため、"
     "次期は第10期＝令和9〜11年度",
     "【要記入】業務期間",
     "介護保険法第117条（市町村介護保険事業計画）／老人福祉法第20条の8（市町村老人福祉計画）",
     "仕様書未受領。受領しだい記入する"),
]

# ---------------------------------------------------------------------------
# 高齢者トラックの工程（介護保険法・老人福祉法に基づく標準的な業務構成）
#
# 【重要】これは契約内容ではなく、一般に計画策定支援業務に含まれる工程の骨格である。
#         仕様書を受領しだい、実際の条項に差し替える。
#         状態・進捗率は当方が契約内容を把握していないことを示すもので、
#         作業の遅れを意味しない。
#
# 区分, 項番, 業務の内容（標準構成）, 想定成果物, 状態, 進捗率, 主担当, 期限, 備考・確認事項
# ---------------------------------------------------------------------------
MIJU = "仕様書未受領"
KOUREI_ITEMS = [
    # ---------------- Ⅰ 基礎調査・分析 ----------------
    ("Ⅰ 基礎調査・分析", "Ⅰ(1)",
     "各種統計データの収集・整理（総人口、高齢者人口、高齢化率、要支援・要介護認定者数、"
     "認定率、被保険者数、サービス受給者数、給付費実績）",
     "基礎データ集計表", MIJU, 0.00, "当社", None,
     "広域連合の場合は構成市町村別の内訳が必要。どの単位まで分析するかを確認する"),
    ("Ⅰ 基礎調査・分析", "Ⅰ(2)",
     "制度動向の整理（介護保険法改正、国の基本指針、都道府県高齢者保健福祉計画、"
     "認知症施策推進基本計画、医療計画・地域医療構想との整合）",
     "制度動向整理資料", MIJU, 0.00, "当社", None,
     "次期（第10期）基本指針の告示時期を確認する。第9期の基本指針は令和6年1月に"
     "告示されており（日付は原典で要確認）、同様であれば令和9年1月頃。"
     "告示前に着手する部分と告示後に確定する部分を分ける"),
    ("Ⅰ 基礎調査・分析", "Ⅰ(3)",
     "現行計画（第9期・令和6〜8年度）の評価・分析（サービス見込量と実績の乖離、"
     "重点施策の進捗、給付費・保険料の実績と計画値の差）",
     "現行計画評価表", MIJU, 0.00, "当社", None,
     "障害トラックと同じ構造の評価ブックを流用できる。"
     "令和6・7年度実績と令和8年度見込の受領時期を確認する"),

    # ---------------- Ⅱ 実態調査 ----------------
    ("Ⅱ 実態調査", "Ⅱ(1)",
     "調査設計（介護予防・日常生活圏域ニーズ調査、在宅介護実態調査。"
     "厚生労働省の標準調査票を基本とし、独自設問を協議のうえ決定）",
     "調査設計書・調査票案", MIJU, 0.00, "当社", None,
     "国の標準様式の版を確認する。在宅介護実態調査は認定調査・更新申請時の"
     "聞き取りで実施する自治体もあり、実施方法（郵送／窓口）を確認する"),
    ("Ⅱ 実態調査", "Ⅱ(2)", "調査票・封筒の印刷",
     "調査票・封筒", MIJU, 0.00, "当社", None,
     "判型・頁数・色数・部数の仕様を確認する"),
    ("Ⅱ 実態調査", "Ⅱ(3)", "発送・回収管理（宛名の抽出主体、発送費用の負担、回収先）",
     "発送記録・回収管理表", MIJU, 0.00, "―", None,
     "宛名抽出が委託者側作業か受託者側作業かを確認する"),
    ("Ⅱ 実態調査", "Ⅱ(4)", "データ入力（国の集計ファイル様式に準拠）",
     "入力データ", MIJU, 0.00, "当社", None,
     "国の集計ソフト（在宅介護実態調査 集計・分析ツール等）を用いるかを確認する"),
    ("Ⅱ 実態調査", "Ⅱ(5)", "集計・分析（基本集計、圏域別・要介護度別・年齢別クロス集計）",
     "集計表", MIJU, 0.00, "当社", None,
     "小規模自治体・広域連合では圏域別の母数が小さくなる。"
     "秘匿基準（n<10 等）を委託者と決める必要がある"),
    ("Ⅱ 実態調査", "Ⅱ(6)", "調査結果報告書の作成",
     "調査結果報告書", MIJU, 0.00, "当社", None, "成果品としての判型・部数を確認する"),

    # ---------------- Ⅲ 計画策定業務 ----------------
    ("Ⅲ 計画策定業務", "Ⅲ(1)",
     "将来推計（人口・高齢者人口・認定者数）。"
     "地域包括ケア「見える化」システムの推計機能を用いる",
     "将来推計表", MIJU, 0.00, "当社", None,
     "見える化システムの推計値をそのまま用いるか、独自補正を加えるかを"
     "委託者と決める。小規模自治体では自然体推計と実勢が乖離しやすい"),
    ("Ⅲ 計画策定業務", "Ⅲ(2)",
     "サービス見込量の推計（居宅・地域密着型・施設・介護予防）",
     "サービス見込量表", MIJU, 0.00, "当社", None,
     "施設整備の予定（特養・老健・地域密着型）の有無で結果が大きく動く。"
     "整備計画の有無を早期に確認する"),
    ("Ⅲ 計画策定業務", "Ⅲ(3)",
     "第1号被保険者の介護保険料の算定（所得段階別、"
     "介護給付費準備基金の取崩し、財政安定化基金の償還）",
     "保険料算定ワークシート", MIJU, 0.00, "当社", None,
     "本業務の最大の山場。基金残高・取崩し方針は委託者の政策判断であり、"
     "複数パターンの試算を求められることが多い。"
     "国の算定ワークシートの配布時期が工程を規定する"),
    ("Ⅲ 計画策定業務", "Ⅲ(4)",
     "地域支援事業・介護予防・日常生活支援総合事業の計画",
     "地域支援事業計画", MIJU, 0.00, "当社", None,
     "総合事業の単価・サービス類型の見直し状況を確認する"),
    ("Ⅲ 計画策定業務", "Ⅲ(5)",
     "施策の検討（地域包括ケアシステムの深化、認知症施策、介護人材の確保、"
     "在宅医療・介護連携、生活支援体制整備、高齢者虐待防止・権利擁護）",
     "施策検討資料", MIJU, 0.00, "当社", None,
     "第10期の重点事項は基本指針の告示で確定する。告示前は前期の重点事項と"
     "審議会資料をもとに仮置きする"),
    ("Ⅲ 計画策定業務", "Ⅲ(6)", "計画案（素案）の作成",
     "計画素案", MIJU, 0.00, "当社", None, "章立ての合意時期を確認する"),
    ("Ⅲ 計画策定業務", "Ⅲ(7)",
     "策定委員会（介護保険運営協議会等）の運営支援（資料作成、出席、会議録作成）",
     "委員会資料・会議録", MIJU, 0.00, "当社", None,
     "開催回数と、受託者の役割（資料作成のみか、説明も行うか）を確認する"),
    ("Ⅲ 計画策定業務", "Ⅲ(8)", "パブリックコメントの支援（資料作成、意見の整理・回答案）",
     "パブコメ資料・意見整理表", MIJU, 0.00, "当社", None, "実施時期を確認する"),
    ("Ⅲ 計画策定業務", "Ⅲ(9)", "計画書の作成（最終取りまとめ・印刷・製本）",
     "計画書", MIJU, 0.00, "当社", None, "判型・頁数・色数・部数を確認する"),

    # ---------------- Ⅳ 打合せ等業務 ----------------
    ("Ⅳ 打合せ等業務", "Ⅳ(1)", "打合せ（対面・オンライン・電話）、議事録の作成",
     "議事録", MIJU, 0.00, "当社", None, "回数と開催形態を確認する"),

    # ---------------- 契約管理 ----------------
    ("契約管理", "管理1", "仕様書・契約書の受領と受注範囲の整理",
     "受注内容整理", MIJU, 0.00, "当社", None,
     "最優先。これが無いと本シートの各行を確定できない"),
    ("契約管理", "管理2", "個人情報の取扱い（実態調査の回収物、認定情報）",
     "取扱い手順", MIJU, 0.00, "当社", None,
     "介護保険の実態調査は要介護認定情報を含むため、"
     "障害トラックより厳格な管理が必要になる場合がある"),
]

# ---------------------------------------------------------------------------
# 高齢者トラックで受領が必要な資料
# 優先度, 資料, 内容, 用途, 状態
# ---------------------------------------------------------------------------
KOUREI_DOCS = [
    ("最優先", "業務委託仕様書・契約書",
     "業務の範囲、成果品、期限、打合せ回数、費用負担の区分",
     "本ブック 01 の各行を確定する", "未受領"),
    ("最優先", "現行計画（第9期介護保険事業計画・高齢者福祉計画）本編",
     "計画期間、サービス見込量、保険料、重点施策、圏域設定",
     "現行計画の評価・分析", "未受領"),
    ("最優先", "要介護認定者数・被保険者数の推移",
     "年度別・要介護度別・年齢階級別。広域連合の場合は構成市町村別",
     "将来推計", "未受領"),
    ("最優先", "介護給付費の実績",
     "年度別・サービス種別別の給付費と受給者数（介護保険事業状況報告ベース）",
     "サービス見込量・保険料算定", "未受領"),
    ("高", "介護給付費準備基金の残高と取崩し方針",
     "各年度末残高、第9期の取崩し実績、次期の取崩し方針",
     "保険料算定", "未受領"),
    ("高", "施設・居住系サービスの整備計画",
     "特養・老健・介護医療院・地域密着型サービスの整備予定と定員",
     "サービス見込量", "未受領"),
    ("高", "地域支援事業・総合事業の実績",
     "事業別の実施状況、利用者数、事業費",
     "地域支援事業計画", "未受領"),
    ("高", "地域包括支援センターの体制",
     "設置数、職員配置、相談件数、ブランチの有無",
     "施策検討", "未受領"),
    ("中", "前回の実態調査結果",
     "介護予防・日常生活圏域ニーズ調査、在宅介護実態調査の報告書",
     "調査設計・経年比較", "未受領"),
    ("中", "策定委員会の委員名簿・開催実績",
     "任期、氏名、所属、前期の開催回数",
     "委員会運営支援", "未受領"),
]

# ---------------------------------------------------------------------------
# 高齢者トラックの論点（仕様書がなくても先に押さえられること）
# 論点, 内容, 障害トラックとの関係
# ---------------------------------------------------------------------------
KOUREI_RONTEN = [
    ("計画期間が障害計画と完全に一致する",
     "介護保険事業計画は介護保険法第117条により3年1期。第9期が令和6〜8年度であるため、"
     "次期は第10期＝令和9〜11年度となり、第8期障がい福祉計画（令和9〜11年度）と"
     "計画期間が完全に一致する。",
     "2計画の工程を揃えられる。基礎データの収集、将来推計の人口前提、"
     "策定委員会・パブコメの時期を共通化すると、両計画で二重に村・連合へ"
     "照会する手間が省ける"),
    ("基本指針の告示時期が両計画のクリティカルパスになる",
     "障害福祉計画の基本指針は令和8年3月31日に告示済み（こども家庭庁・厚生労働省告示第4号）。"
     "介護保険事業計画の基本指針は、第9期が令和6年1月に告示されたことから、"
     "第10期は令和9年1月頃の告示が見込まれる。"
     "第9期の告示日及び第10期の告示予定時期は原典で要確認。",
     "障害は指針が確定済みで先行できるが、高齢者は指針前に固められる部分と"
     "固められない部分の切り分けが必要。告示前は前期の指針と社会保障審議会"
     "介護保険部会の資料をもとに仮置きし、告示後に差し替える工程を組む"),
    ("保険料算定が高齢者側の最大の山場",
     "第1号被保険者の保険料は、サービス見込量→給付費→標準給付費見込額→"
     "第1号被保険者負担分→準備基金取崩し→所得段階別保険料という順に積み上げる。"
     "国の算定ワークシートの配布時期が工程を規定する。",
     "障害計画には対応する工程がない（自立支援給付は義務的経費で"
     "国1/2・県1/4・市町村1/4の法定負担）。高齢者側だけ年明けに山が来る"),
    ("実態調査の設計思想が異なる",
     "介護保険は国が標準調査票（介護予防・日常生活圏域ニーズ調査、在宅介護実態調査）を"
     "示しており、経年比較・全国比較のため標準設問を維持する要請が強い。",
     "障害は「厚生労働省が示す調査項目及び過去調査内容を踏まえて設計し、"
     "委託者との協議の上決定」（仕様書Ⅰ(2)）であり自由度が高い。"
     "調査票の作り込みにかける工数配分が両者で異なる"),
    ("共生型サービス・65歳到達が両計画の接点",
     "障害福祉サービス利用者が65歳に到達すると介護保険優先原則（障害者総合支援法第7条）が"
     "適用される。共生型サービス（平成30年創設）は介護保険の訪問介護・通所介護等が"
     "障害福祉の指定を受けやすくする仕組み。",
     "両計画で同じ人を扱う。障害トラックでは"
     "docs/北塩原村_年齢到達による給付移行_調査報告.md に整理済み。"
     "高齢者側の計画にも共生型サービスの記載を求められる可能性がある"),
]


# ============================================================
# 書式ヘルパー
# ============================================================
def _fill_status(cell):
    fill = PROGRESS_FILL.get(str(cell.value))
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")


def _fill_priority(cell):
    v = str(cell.value)
    color = {"最優先": "C00000", "高": "ED7D31"}.get(v)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")


def _track_badge(cell, track):
    cell.fill = PatternFill("solid", fgColor=TRACK_COLOR[track])
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center", horizontal="center")


def _section(ws, row, ncol, text, fill=COLORS["subhead"]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    style_title(ws.cell(row=row, column=1), text, fill=fill, size=11)
    return row + 1


# ============================================================
# 00_サマリー
# ============================================================
def sheet_summary(wb, shogai_cats, kourei_cats):
    ws = add_sheet(
        wb, "00_サマリー", "業務進捗管理　高齢者／障害",
        "受託中の計画策定業務を、仕様書に記載された業務（各段階）の単位で管理します。"
        "障害の数値は 北塩原村_業務進捗管理.xlsx と同一のデータから生成しており、"
        "2つのブックで食い違いません。",
        [22, 40, 12, 12, 12, 46])

    r = 5
    r = _section(ws, r, 6, "1　契約の基本情報")
    style_header_row(ws, r, ["分野", "委託名", "計画期間", "", "業務期間", "根拠法・出典"])
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r += 1
    for i, (track, name, hacchu, kikan, gyomu, kyoho, shutten) in enumerate(CONTRACTS):
        row = r
        r = write_row(ws, r, [track, f"{name}\n発注者：{hacchu}", kikan, "", gyomu,
                              f"{kyoho}\n出典：{shutten}"], alt=(i % 2 == 1))
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        _track_badge(ws.cell(row=row, column=1), track)
        ws.row_dimensions[row].height = 58

    r += 1
    r = _section(ws, r, 6, "2　区分別の進捗")
    style_header_row(ws, r, ["分野", "区分", "項目数", "完了・概ね完了", "平均進捗率", "状況"])
    r += 1
    for track, cats, items, note_fn in (
        (TRACK_SHOGAI, shogai_cats, SHOGAI.SPEC_ITEMS, None),
        (TRACK_KOUREI, kourei_cats, KOUREI_ITEMS, None),
    ):
        for i, cat in enumerate(cats):
            rows = [x for x in items if x[0] == cat]
            done = len([x for x in rows if x[4] in ("完了", "概ね完了")])
            avg = sum(x[5] for x in rows) / len(rows) if rows else 0
            if track == TRACK_KOUREI:
                comment = "仕様書未受領のため進捗を評価できない"
            else:
                pending = [x for x in rows if x[4] in ("村資料待ち", "要協議", "村側作業")]
                comment = (f"待ち・協議中 {len(pending)}件" if pending else "自走可能")
            row = r
            r = write_row(ws, r, [track, cat, len(rows), done, avg, comment],
                          alt=(i % 2 == 1),
                          aligns=["center", "left", "center", "center", "center", "left"],
                          numfmts=[None, None, "#,##0", "#,##0", PCT, None])
            _track_badge(ws.cell(row=row, column=1), track)

    r += 1
    r = _section(ws, r, 6, "3　いま押さえるべきこと", fill="C00000")
    style_header_row(ws, r, ["分野", "事項", "", "期限", "", "内容"])
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    r += 1
    critical = [
        (TRACK_KOUREI, "業務委託仕様書の受領", "至急",
         "高齢者分野の仕様書が当方に無く、01シートの各行は介護保険法・老人福祉法に基づく"
         "標準構成を骨格として置いたものです。受領しだい実際の条項に差し替えます。"),
        (TRACK_KOUREI, "大雪広域の業務進捗管理ブック", "至急",
         "本ブックは体裁を合わせる元のファイルが無いため、"
         "障害分野で運用中の管理構造（仕様書条項別・工程表・資料受領・課題）を"
         "そのまま2分野へ拡張したものです。"),
        (TRACK_SHOGAI, "アンケートの印刷仕様の確定と発送", "2026-08-05超過",
         "仕様書は約24頁だが現行の調査票は障がい者版17頁・障がい児版19頁。"
         "村ラベルの提供と村封筒の手配も未了で、発送のクリティカルパス上にあります。"),
        (TRACK_SHOGAI, "発送対象者数の確定", "2026-08-05超過",
         "仕様書の約190人と現行計画の156人の差の理由が未確認。"
         "回収率の分母と管理番号の設計に直結します。"),
        (TRACK_SHOGAI, "令和6・7年度実績の受領", "2026-09-30",
         "成果目標・見込量・給付費のすべての推計がここで止まっています。"),
    ]
    for i, (track, item, due, body) in enumerate(critical):
        row = r
        r = write_row(ws, r, [track, item, "", due, "", body], alt=(i % 2 == 1),
                      aligns=["center", "left", "left", "center", "center", "left"])
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        _track_badge(ws.cell(row=row, column=1), track)
        ws.row_dimensions[row].height = 40

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_note(ws.cell(row=r, column=1),
               "高齢者分野の進捗率0％は、作業が遅れているという意味ではありません。"
               "当方が契約内容を把握していないため評価できない、という意味です。"
               "仕様書と現行計画（第9期）を受領しだい、01シートを実際の条項に差し替え、"
               "着手済みの工程に進捗率を入れます。")
    ws.row_dimensions[r].height = 46
    ws.freeze_panes = "A5"
    return ws


# ============================================================
# 仕様書別進捗（共通レンダラ）
# ============================================================
def sheet_items(wb, title, heading, note, items, cats, track, last_col_head):
    ws = add_sheet(wb, title, heading, note, [18, 8, 52, 34, 12, 8, 10, 12, 50])
    style_header_row(ws, 5, ["区分", "項番", "仕様書の要求／業務の内容", last_col_head,
                             "状態", "進捗率", "主担当", "期限", "次の一手・確認事項"])
    ws.cell(row=5, column=1).fill = PatternFill("solid", fgColor=TRACK_COLOR[track])

    r = 6
    cat_rows = {}
    for cat in cats:
        rows = [x for x in items if x[0] == cat]
        start = r
        for i, rec in enumerate(rows):
            vals = list(rec)
            write_row(ws, r, vals, alt=(i % 2 == 1),
                      aligns=["left", "center", "left", "left", "center", "center",
                              "center", "center", "left"],
                      numfmts=[None, None, None, None, None, PCT, None, DATE, None])
            _fill_status(ws.cell(row=r, column=5))
            ws.row_dimensions[r].height = 42
            r += 1
        cat_rows[cat] = (start, r - 1)
        # 区分の小計
        write_row(ws, r, [f"{cat}　小計", "", "", f"{len(rows)}項目", "",
                          f"=AVERAGE(F{start}:F{r - 1})" if rows else 0, "", "", ""],
                  aligns=["left", "center", "left", "center", "center", "center",
                          "center", "center", "left"],
                  numfmts=[None, None, None, None, None, PCT, None, None, None],
                  fills=[COLORS["band"]] * 9)
        for c in range(1, 10):
            ws.cell(row=r, column=c).font = Font(name=FONT, size=10, bold=True)
        r += 1

    last = r - 1
    ws.auto_filter.ref = f"A5:I{last}"
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    total = sum(x[5] for x in items) / len(items) if items else 0
    if track == TRACK_KOUREI:
        style_note(ws.cell(row=r, column=1),
                   "【本シートの位置づけ】高齢者分野の業務委託仕様書が当方に無いため、"
                   "介護保険法第117条（市町村介護保険事業計画）及び老人福祉法第20条の8"
                   "（市町村老人福祉計画）に基づく標準的な業務構成を骨格として置いています。"
                   "各行は「一般に計画策定支援業務に含まれる工程」であって、"
                   "実際の契約が何を求めているかを示すものではありません。"
                   "状態「仕様書未受領」・進捗率0％は、作業が遅れているという意味ではなく、"
                   "当方が契約内容を把握していないという意味です。"
                   "仕様書を受領しだい実際の条項に差し替えます。")
        ws.row_dimensions[r].height = 62
    else:
        style_note(ws.cell(row=r, column=1),
                   f"全{len(items)}項目・平均進捗率 {total:.0%}。"
                   "データは build_kitashiobara_progress.py の SPEC_ITEMS と同一で、"
                   "北塩原村_業務進捗管理.xlsx と食い違いません。"
                   "状態が「村資料待ち」「要協議」「村側作業」の項目は当方だけでは進められません。")
        ws.row_dimensions[r].height = 40
    ws.freeze_panes = "C6"
    return ws


# ============================================================
# 03_工程表
# ============================================================
def sheet_schedule(wb):
    ws = add_sheet(
        wb, "03_工程表", "工程表（2計画）",
        "障害は令和8年7月7日打合せの工程案。高齢者は仕様書未受領のため、"
        "介護保険事業計画の一般的な策定サイクルから逆算した目安を灰色で置いています。",
        [10, 14, 14, 44, 20, 12, 44])
    style_header_row(ws, 5, ["分野", "開始", "終了", "工程", "担当", "状態", "備考"])
    r = 6
    for i, (st, en, name, tan, state, note) in enumerate(SHOGAI.SCHEDULE):
        row = r
        r = write_row(ws, r, [TRACK_SHOGAI, st, en, name, tan, state, note],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "center", "left", "left", "center", "left"],
                      numfmts=[None, DATE, DATE, None, None, None, None])
        _track_badge(ws.cell(row=row, column=1), TRACK_SHOGAI)
        _fill_status(ws.cell(row=row, column=6))
        for c in (2, 3):
            ws.cell(row=row, column=c).number_format = DATE

    kourei_sched = [
        ("2026-09", "2026-11", "基礎調査・現行計画（第9期）の評価",
         "当社", "基本指針の告示前に着手できる部分"),
        ("2026-09", "2026-12", "実態調査（設計・印刷・発送・回収・集計）",
         "当社", "国の標準調査票を用いる場合は設計工数が小さい"),
        ("2026-12", "2027-02", "将来推計・サービス見込量の算定",
         "当社", "地域包括ケア「見える化」システムを使用"),
        ("2027-01", "2027-03", "介護保険料の算定",
         "当社", "国の算定ワークシートの配布時期が工程を規定する"),
        ("2027-01", "2027-02", "基本指針の告示を受けた計画内容の確定",
         "当社", "第9期は令和6年1月に告示。第10期も同時期と見込まれる（要確認）"),
        ("2027-02", "2027-03", "計画素案の作成・策定委員会",
         "当社／発注者", "障害計画の協議会と時期が重なる"),
        ("2027-03", "2027-04", "パブリックコメント",
         "発注者／当社", "障害計画のパブコメと同時期"),
        ("2027-04", "2027-06", "計画書の作成・納品",
         "当社", "年度をまたぐ場合は契約年度の扱いを確認する"),
    ]
    r += 1
    for i, (st, en, name, tan, note) in enumerate(kourei_sched):
        row = r
        r = write_row(ws, r, [TRACK_KOUREI, st, en, name, tan, "仕様書未受領", note],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "center", "left", "left", "center", "left"])
        _track_badge(ws.cell(row=row, column=1), TRACK_KOUREI)
        _fill_status(ws.cell(row=row, column=6))

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "高齢者の日付は年月のみの目安であり、契約上の期限ではありません。"
               "介護保険事業計画は3年1期（介護保険法第117条）で、第9期が令和6〜8年度のため、"
               "次期は第10期＝令和9〜11年度となり、第8期障がい福祉計画と計画期間が一致します。"
               "策定委員会とパブリックコメントの時期が両計画で重なるため、"
               "資料作成と説明の負荷が令和9年2〜4月に集中します。")
    ws.row_dimensions[r].height = 46
    return ws


# ============================================================
# 04_成果品
# ============================================================
def sheet_deliverables(wb):
    ws = add_sheet(
        wb, "04_成果品", "成果品一覧（2計画）",
        "障害は仕様書5に明記された3点。高齢者は仕様書未受領のため、"
        "一般的な成果品の構成を置いています。",
        [10, 34, 44, 14, 14, 44])
    style_header_row(ws, 5, ["分野", "成果品", "仕様", "状態", "期限", "備考"])
    r = 6
    items = [
        (TRACK_SHOGAI, "①計画書",
         "A4判・両面 約60頁・モノクロ・コピー・くるみ製本　30部", "作業中", "2027-03-31",
         "骨子案修正版が原型。全8章＋新設5節。分量配分計画は 08 で管理（個別ブック）"),
        (TRACK_SHOGAI, "②アンケート調査報告書",
         "A4判・両面 約60頁・モノクロ・コピー・ホチキス製本　10部", "未着手", "2027-03-31",
         "設問ごとにグラフを作成。集計後に着手"),
        (TRACK_SHOGAI, "③電子媒体",
         "①②の電子データ一式", "未着手", "2027-03-31", "編集可能形式とPDFの別を確認する"),
        (TRACK_KOUREI, "①計画書", "【要確認】判型・頁数・色数・製本・部数",
         "仕様書未受領", None, "介護保険事業計画と高齢者福祉計画の一体作成が通例"),
        (TRACK_KOUREI, "②実態調査報告書", "【要確認】判型・頁数・部数",
         "仕様書未受領", None,
         "介護予防・日常生活圏域ニーズ調査と在宅介護実態調査を1冊にまとめるか別冊かを確認"),
        (TRACK_KOUREI, "③概要版", "【要確認】作成の要否",
         "仕様書未受領", None, "介護保険事業計画では概要版・住民向けリーフレットを求められることがある"),
        (TRACK_KOUREI, "④電子媒体", "【要確認】",
         "仕様書未受領", None, ""),
    ]
    for i, (track, name, spec, state, due, note) in enumerate(items):
        row = r
        r = write_row(ws, r, [track, name, spec, state, due, note], alt=(i % 2 == 1),
                      aligns=["center", "left", "left", "center", "center", "left"],
                      numfmts=[None, None, None, None, DATE, None])
        _track_badge(ws.cell(row=row, column=1), track)
        _fill_status(ws.cell(row=row, column=4))
        ws.row_dimensions[row].height = 34

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_note(ws.cell(row=r, column=1),
               "障害①の計画書について、仕様書はアンケート調査を「A4判、約24ページ、黒1色」と"
               "していますが、現行の調査票は障がい者版17頁・障がい児版19頁です。"
               "印刷部数・郵送費の負担区分とあわせて村と協議中です（06 の R-2）。")
    ws.row_dimensions[r].height = 40
    return ws


# ============================================================
# 05_受領資料
# ============================================================
def sheet_docs(wb):
    ws = add_sheet(
        wb, "05_受領資料", "資料の受領状況（2計画）",
        "計画分野ごとに、発注者から受領が必要な資料と受領状況を管理します。"
        "障害は 北塩原村_業務進捗管理.xlsx の 04 と同一データです。",
        [10, 10, 34, 50, 30, 12, 12])
    style_header_row(ws, 5, ["分野", "優先度", "資料", "内容", "反映先", "状態", "希望期限"])
    r = 6
    for i, (pri, name, cont, dest, st, due) in enumerate(SHOGAI.VILLAGE_DOCS):
        row = r
        r = write_row(ws, r, [TRACK_SHOGAI, pri, name, cont, dest, st, due],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "left", "left", "left", "center", "center"],
                      numfmts=[None, None, None, None, None, None, DATE])
        _track_badge(ws.cell(row=row, column=1), TRACK_SHOGAI)
        _fill_priority(ws.cell(row=row, column=2))
        cs = ws.cell(row=row, column=6)
        color = {"受領済": "2CA02C", "一部受領": "70AD47",
                 "県照会中": "2E75B6", "未受領": "C00000"}.get(cs.value)
        if color:
            cs.fill = PatternFill("solid", fgColor=color)
            cs.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[row].height = 32

    for i, (pri, name, cont, dest, st) in enumerate(KOUREI_DOCS):
        row = r
        r = write_row(ws, r, [TRACK_KOUREI, pri, name, cont, dest, st, None],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "left", "left", "left", "center", "center"],
                      numfmts=[None, None, None, None, None, None, DATE])
        _track_badge(ws.cell(row=row, column=1), TRACK_KOUREI)
        _fill_priority(ws.cell(row=row, column=2))
        cs = ws.cell(row=row, column=6)
        cs.fill = PatternFill("solid", fgColor="C00000")
        cs.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[row].height = 32

    last = r - 1
    ws.auto_filter.ref = f"A5:G{last}"
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               f"障害 {len(SHOGAI.VILLAGE_DOCS)}件、高齢者 {len(KOUREI_DOCS)}件。"
               "高齢者は仕様書そのものが未受領のため、"
               "ここに挙げた10件は介護保険事業計画の策定に一般に必要となる資料です。"
               "実際に必要な資料は仕様書の受領後に確定します。")
    ws.row_dimensions[r].height = 40
    return ws


# ============================================================
# 06_課題・リスク
# ============================================================
def sheet_risks(wb):
    ws = add_sheet(
        wb, "06_課題・リスク", "課題・リスク管理（2計画）",
        "障害は 北塩原村_業務進捗管理.xlsx の 05 と同一データです。"
        "高齢者は仕様書未受領に起因するもののみを掲げています。",
        [10, 8, 14, 8, 40, 50, 44, 12])
    style_header_row(ws, 5, ["分野", "#", "区分", "影響度", "リスク", "内容", "対応", "状態"])
    r = 6
    for i, rec in enumerate(SHOGAI.RISKS):
        row = r
        r = write_row(ws, r, [TRACK_SHOGAI] + list(rec), alt=(i % 2 == 1),
                      aligns=["center", "center", "center", "center", "left",
                              "left", "left", "center"])
        _track_badge(ws.cell(row=row, column=1), TRACK_SHOGAI)
        imp = ws.cell(row=row, column=4)
        if imp.value == "高":
            imp.fill = PatternFill("solid", fgColor="C00000")
            imp.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[row].height = 44

    kourei_risks = [
        ("K-1", "契約", "高", "高齢者分野の業務委託仕様書が当方の手元にない",
         "受注範囲、成果品、期限、打合せ回数、費用負担の区分がいずれも不明。"
         "本ブックの01シートは介護保険法・老人福祉法に基づく標準構成を置いたものであり、"
         "実際の契約と一致する保証がない",
         "仕様書・契約書の写しをご提供ください。受領しだい01シートを差し替えます", "未着手"),
        ("K-2", "工程", "高", "国の基本指針の告示前に着手できる範囲が未確定",
         "介護保険事業計画の基本指針は第9期が令和6年1月に告示（日付は要確認）。"
         "第10期も同時期であれば、計画内容の確定は令和9年1〜2月以降となる。"
         "告示前に固められる部分（基礎調査・実態調査・現行計画評価）と"
         "固められない部分（重点施策・成果目標）の切り分けが必要",
         "第10期基本指針の告示予定時期を確認し、告示前後で工程を二段構えにする", "未着手"),
        ("K-3", "工程", "高", "策定委員会とパブリックコメントが障害計画と同時期に重なる",
         "介護保険事業計画も障がい福祉計画も計画期間が令和9〜11年度で一致するため、"
         "策定委員会・パブコメ・計画書作成が令和9年1〜3月に集中する。"
         "同一担当者が両方を抱えると資料作成が輻輳する",
         "両計画の会議日程を早期に把握し、資料作成の山を前倒しする。"
         "基礎データ・人口推計の前提は共通化して二度手間を避ける", "未着手"),
        ("K-4", "算定", "中", "保険料算定は国のワークシート配布まで着手できない",
         "第1号被保険者保険料の算定は国が配布する算定ワークシートを用いるのが通例で、"
         "配布時期が年明けになると年度末に作業が集中する",
         "準備基金の残高・取崩し方針を先に発注者と詰め、"
         "ワークシート配布後は入力と試算に専念できる状態にしておく", "未着手"),
        ("K-5", "管理", "中", "本ブックの体裁を合わせる元のファイル（大雪広域の進捗管理）が無い",
         "「大雪広域の業務進捗と同様に」というご指示に対し、"
         "当該ファイルが本ワークスペースに存在しないため、"
         "障害分野で運用中の管理構造（仕様書条項別・工程表・成果品・資料受領・課題）を"
         "そのまま2分野へ拡張した",
         "大雪広域の進捗管理ブックをご提供いただければ、列構成・状態区分・"
         "集計方法をそちらに合わせて作り直します", "対応中"),
    ]
    r += 1
    for i, rec in enumerate(kourei_risks):
        row = r
        r = write_row(ws, r, [TRACK_KOUREI] + list(rec), alt=(i % 2 == 1),
                      aligns=["center", "center", "center", "center", "left",
                              "left", "left", "center"])
        _track_badge(ws.cell(row=row, column=1), TRACK_KOUREI)
        imp = ws.cell(row=row, column=4)
        if imp.value == "高":
            imp.fill = PatternFill("solid", fgColor="C00000")
            imp.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[row].height = 44

    last = r - 1
    ws.auto_filter.ref = f"A5:H{last}"
    return ws


# ============================================================
# 07_2計画の関係
# ============================================================
def sheet_relation(wb):
    ws = add_sheet(
        wb, "07_2計画の関係", "高齢者計画と障害計画の関係",
        "2つの計画を1人で見ることの利点と注意点を整理しました。"
        "仕様書が無くても先に押さえられる論点です。",
        [34, 56, 56])
    style_header_row(ws, 5, ["論点", "内容", "2計画にまたがる意味"])
    r = 6
    for i, rec in enumerate(KOUREI_RONTEN):
        r = write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r - 1].height = 70
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_note(ws.cell(row=r, column=1),
               "介護保険事業計画の期は介護保険法第117条により3年1期です。"
               "第9期が令和6〜8年度であることから次期は第10期＝令和9〜11年度となり、"
               "第8期障がい福祉計画（令和9〜11年度）と計画期間が完全に一致します。"
               "第9期基本指針の告示日及び第10期の告示予定時期は本ワークスペースで"
               "原典を確認できていないため、類推として扱ってください。")
    ws.row_dimensions[r].height = 52
    return ws


def verify():
    """データの内部整合を検証する。"""
    # 障害トラックは個別ブックと同じデータであること
    assert len(SHOGAI.SPEC_ITEMS) == 33, f"障害の仕様書項目が33件でない: {len(SHOGAI.SPEC_ITEMS)}"
    assert all(len(x) == 9 for x in SHOGAI.SPEC_ITEMS), "障害の仕様書項目の列数が9でない"

    # 高齢者トラックは列数を障害トラックに揃えること
    assert all(len(x) == 9 for x in KOUREI_ITEMS), "高齢者の工程の列数が9でない"

    # 高齢者トラックに推測の進捗を入れていないこと
    for rec in KOUREI_ITEMS:
        assert rec[4] == MIJU, f"高齢者の状態が「{MIJU}」でない: {rec[1]}"
        assert rec[5] == 0.00, f"高齢者に進捗率が入っている: {rec[1]}"
        assert rec[7] is None, f"高齢者に期限が入っている: {rec[1]}"

    # 高齢者の受領資料はすべて未受領であること
    for rec in KOUREI_DOCS:
        assert rec[4] == "未受領", f"高齢者の資料状態が未受領でない: {rec[1]}"

    # 契約情報は2分野
    assert {c[0] for c in CONTRACTS} == {TRACK_KOUREI, TRACK_SHOGAI}


def main():
    verify()
    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)

    shogai_cats = SHOGAI.CATEGORIES
    kourei_cats = []
    for rec in KOUREI_ITEMS:
        if rec[0] not in kourei_cats:
            kourei_cats.append(rec[0])

    sheet_summary(wb, shogai_cats, kourei_cats)
    sheet_items(
        wb, "01_高齢者_工程別進捗",
        "高齢者　介護保険事業計画・高齢者福祉計画　工程別進捗",
        "【仕様書未受領】本シートは介護保険法第117条及び老人福祉法第20条の8に基づく"
        "標準的な業務構成を骨格として置いたものです。実際の契約内容ではありません。"
        "仕様書を受領しだい差し替えます。",
        KOUREI_ITEMS, kourei_cats, TRACK_KOUREI, "想定成果物")
    sheet_items(
        wb, "02_障害_工程別進捗",
        "障害　第8期北塩原村障がい福祉計画・第4期障がい児福祉計画　仕様書条項別進捗",
        "業務委託仕様書（8北保福第366号 別紙）の条項に対応させた管理表です。"
        "北塩原村_業務進捗管理.xlsx の 01 と同一データから生成しています。",
        SHOGAI.SPEC_ITEMS, shogai_cats, TRACK_SHOGAI, "現在の成果物・根拠")
    sheet_schedule(wb)
    sheet_deliverables(wb)
    sheet_docs(wb)
    sheet_risks(wb)
    sheet_relation(wb)

    wb.save(OUT_FILE)

    s_avg = sum(x[5] for x in SHOGAI.SPEC_ITEMS) / len(SHOGAI.SPEC_ITEMS)
    print(f"作成: {OUT_FILE}")
    print(f"  シート数: {len(wb.sheetnames)}")
    print(f"  障害　： 仕様書項目{len(SHOGAI.SPEC_ITEMS)}件・平均進捗{s_avg:.0%}／"
          f"資料{len(SHOGAI.VILLAGE_DOCS)}件・リスク{len(SHOGAI.RISKS)}件")
    print(f"  高齢者： 標準構成{len(KOUREI_ITEMS)}件（すべて仕様書未受領）／"
          f"必要資料{len(KOUREI_DOCS)}件・リスク5件")


if __name__ == "__main__":
    main()
