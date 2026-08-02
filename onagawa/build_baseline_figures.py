# -*- coding: utf-8 -*-
"""
女川町 上水道事業　数値定義と確定値一覧
─────────────────────────────────────────────
成果物間で数値が食い違わないようにするための基準表。
すべての成果物（料金改定シミュレーション、経営戦略、中間報告書）は
本ファイルの定義・確定値に従うこととする。

出力: onagawa/output/女川町_水道_数値定義と確定値一覧.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "output")
os.makedirs(OUT_DIR, exist_ok=True)

C = {"header": "1F3864", "subhead": "2E75B6", "orange": "C55A11", "alt": "F7FAFC",
     "white": "FFFFFF", "gray": "808080", "bad": "FCE4E4", "warn": "FFF2CC", "good": "E2F0D9"}
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT = "Meiryo UI"

# ============================================================
# 確定値（令和７年度決算）
# ============================================================
R7 = {
    "給水収益": 119_480_750, "営業収益": 119_522_350,
    "経常収益": 655_339_107, "経常費用": 674_928_751,
    "長期前受金戻入益": 334_252_462, "減価償却費": 393_745_802,
    "支払利息等": 10_426_807, "営業外費用": 14_641_896,
    "有収水量": 1_024_576, "年間総配水量": 1_391_383, "給水人口": 5_656,
    "一過性_江島": 128_789_977, "基準外補助金": 166_434_074,
}
# 過年度の有収水量・給水人口
JISSEKI = [("令和５年度", 1_039_302, 5_842), ("令和６年度", 1_019_321, 5_769),
           ("令和７年度", 1_024_576, 5_656)]
# 経営戦略（案）の推計
AN = {"R8_有収水量": 901_000, "R8_給水人口": 5_625,
      "R17_有収水量": 723_000, "R17_給水人口": 4_510}
# 用途区分別の有収水量シェア（調定データ1か月分）
YOUTO = [("家事用", 2471, 34_668), ("工業用", 66, 28_067), ("団体用", 195, 7_498),
         ("営業用", 298, 6_908), ("湯屋用", 2, 662), ("臨時用", 26, 33),
         ("委託船舶給水用", 1, 11)]
# 引継ぎメモ記載の総括原価の費目（R8〜R12の5年計・千円）
SOKATSU = {"維持管理費": 672_327, "減価償却費": 2_032_440,
           "既往債利息": 50_585, "新規債利息": 127_402,
           "資産維持費": 134_445,          # 対象資産896,300千円 × 3％ × 5年
           "控除_長期前受金戻入益_メモ": -450_000,
           "控除_他会計補助金等": -15_000}


def f(size=10, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)


def fl(h):
    return PatternFill("solid", fgColor=h)


class Sheet:
    def __init__(self, wb, name, ncol, widths):
        self.ws = wb.create_sheet(name)
        self.ws.sheet_view.showGridLines = False
        self.n = ncol
        self.r = 1
        for col, w in widths.items():
            self.ws.column_dimensions[col].width = w

    def title(self, text, sub=""):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
        c = ws.cell(row=self.r, column=1, value=text)
        c.font = f(13, True, C["white"]); c.fill = fl(C["header"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[self.r].height = 26
        self.r += 1
        if sub:
            ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
            c = ws.cell(row=self.r, column=1, value=sub)
            c.font = f(8.5, color=C["gray"])
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            self.r += 1
        self.r += 1

    def sub(self, text, orange=False):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
        c = ws.cell(row=self.r, column=1, value=text)
        c.font = f(10.5, True, C["white"]); c.fill = fl(C["orange"] if orange else C["subhead"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[self.r].height = 20
        self.r += 1

    def table(self, heads, data, aligns=None, flags=None, h=28, fmt=None):
        ws = self.ws
        for i, hh in enumerate(heads):
            c = ws.cell(row=self.r, column=i + 1, value=hh)
            c.font = f(9, True, C["white"]); c.fill = fl(C["subhead"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[self.r].height = 28
        self.r += 1
        for j, row in enumerate(data):
            base = C["alt"] if j % 2 else C["white"]
            if flags and j < len(flags) and flags[j]:
                base = {"bad": C["bad"], "warn": C["warn"], "good": C["good"]}[flags[j]]
            for i, v in enumerate(row):
                c = ws.cell(row=self.r, column=i + 1, value=v)
                bold = isinstance(v, str) and (v.startswith("★") or v in ("○", "×"))
                c.font = f(9, bold=bold, color="C00000" if (bold and v != "○") else "000000")
                c.fill = fl(base)
                al = (aligns or ["left"] * len(heads))[i]
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        wrap_text=True, indent=(1 if al == "left" else 0))
                c.border = BORDER
                if fmt and i in fmt and isinstance(v, (int, float)):
                    c.number_format = fmt[i]
            ws.row_dimensions[self.r].height = h
            self.r += 1
        self.r += 1

    def note(self, lines):
        ws = self.ws
        for t in lines:
            ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
            c = ws.cell(row=self.r, column=1, value=t)
            c.font = f(9)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[self.r].height = 16
            self.r += 1
        self.r += 1


# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 派生値
    kyokyu = R7["給水収益"] / R7["有収水量"]
    kaisyu_a = R7["経常費用"] - R7["長期前受金戻入益"]
    heijo = R7["経常費用"] - R7["一過性_江島"]
    kaisyu_b = heijo - R7["長期前受金戻入益"]
    genka_a = kaisyu_a / R7["有収水量"]
    genka_b = kaisyu_b / R7["有収水量"]
    genka_nis = R7["経常費用"] / R7["有収水量"]

    # ---------- 00 数値の使用ルール ----------
    s = Sheet(wb, "00_数値の使用ルール", 4, {"A": 24, "B": 50, "C": 50, "D": 40})
    s.title("女川町 上水道事業　数値定義と確定値一覧",
            "すべての成果物（料金改定シミュレーション・経営戦略・中間報告書）は本ファイルの定義と確定値に従うこととする　／　令和８年８月")
    s.sub("■　なぜこの表が必要か", orange=True)
    s.note([
        "令和７年度決算の確定前に作成された資料との突合で、税抜と税込の混在、令和５年度と令和７年度の指標の混在が5点確認されました。",
        "同じ「供給単価」「給水原価」「料金回収率」という名称で異なる値が並存すると、住民・議会説明で説明が破綻します。",
        "以下のルールを全成果物に適用し、数値には必ず「年度」「税抜／税込」「算式の出典」を併記してください。",
    ])
    s.sub("■　ルール")
    s.table(["区　分", "ルール", "理　由", "例外・注意"],
            [["消費税", "★原則は税抜。総務省指標・経営比較分析表との比較、原価計算、収支計画はすべて税抜",
              "地方公営企業決算状況調査（法適用）は税抜。損益計算書も税抜",
              "住民・議会説明で実際の支払額を示す場合のみ税込。その場合は必ず「（税込）」と明記"],
             ["年度", "★指標は「令和○年度」を必ず併記。異なる年度の指標を組み合わせて逆算しない",
              "料金回収率52.87％は令和５年度、35.07％は令和７年度。混在すると値の意味が失われる",
              "他団体比較値は令和６年度が最新（総務省 水道事業経営指標）。1年ずれる旨を注記"],
             ["給水原価", "★算式が5通りある。用途に応じて使い分け、必ず出典を明記",
              "経営比較分析表・日本水道協会・総括原価で分母分子の定義が異なる",
              "シート02の使い分け表に従う"],
             ["有収水量", "令和７年度実績は1,024,576㎥。将来推計は見直しが必要",
              "経営戦略（案）の令和８年度推計901千㎥は実績を12％下回る",
              "シート03の見直し案を参照"],
             ["単位", "金額は円単位を原則とし、表示のみ千円・百万円に丸める",
              "千円単位で保持すると端数処理の累積で不整合が生じる", "－"],
             ["出典", "すべての数値に出典（決算書の該当表・シート名）を付す",
              "後から検証できない数値は住民・議会説明で使えない", "－"]],
            aligns=["center", "left", "left", "left"], flags=["bad", "bad", "bad", "warn", None, None], h=52)

    # ---------- 01 確定値一覧 ----------
    s = Sheet(wb, "01_確定値一覧_R7決算", 4, {"A": 34, "B": 20, "C": 12, "D": 46})
    s.title("１　確定値一覧（令和７年度決算）", "出典：令和７年度 上水道事業会計決算書（損益計算書・貸借対照表・企業債明細書・注記）")
    s.sub("■　損益")
    s.table(["項　目", "金額（円）", "税区分", "出典・備考"],
            [["給水収益", R7["給水収益"], "税抜", "損益計算書 1 営業収益(1)。★税込は131,428,825円"],
             ["営業収益", R7["営業収益"], "税抜", "損益計算書 1"],
             ["経常収益", R7["経常収益"], "税抜", "営業収益＋営業外収益535,816,757円"],
             ["経常費用", R7["経常費用"], "税抜", "営業費用660,286,855円＋営業外費用14,641,896円"],
             ["　うち減価償却費", R7["減価償却費"], "税抜", "損益計算書 2 営業費用(4)"],
             ["　うち支払利息及び企業債取扱諸費", R7["支払利息等"], "税抜", "損益計算書 4 営業外費用(1)"],
             ["長期前受金戻入益", R7["長期前受金戻入益"], "税抜", "損益計算書 3 営業外収益(4)　★原価算定で控除する項目"],
             ["経常損失", -19_589_644, "税抜", "経常収益−経常費用"],
             ["特別利益（過年度損益修正益）", 774_867_329, "税抜", "★一過性。過年度会計処理の一括修正"],
             ["特別損失", 2_333_576_494, "税抜", "★一過性。うち過年度損益修正損2,333,473,005円"],
             ["当年度純損失", 1_578_298_809, "税抜", "損益計算書"],
             ["当年度未処理欠損金", 2_732_933_550, "税抜", "貸借対照表 7 剰余金(2)"]],
            aligns=["left", "right", "center", "left"], fmt={1: "#,##0"},
            flags=[None] * 6 + ["warn"] + [None] + ["warn", "warn", "bad", "bad"])
    s.sub("■　財政状態・資金")
    s.table(["項　目", "金額（円）", "税区分", "出典・備考"],
            [["企業債残高", 1_094_345_662, "－", "企業債明細書（発行総額1,187,600,000−償還高累計93,254,338）"],
             ["　うち固定負債", 1_063_463_872, "－", "貸借対照表 3 固定負債(1)"],
             ["　うち流動負債（1年以内償還）", 30_881_790, "－", "貸借対照表 4 流動負債(1)"],
             ["流動資産", 156_291_331, "－", "貸借対照表 2"],
             ["流動負債", 110_678_659, "－", "貸借対照表 4"],
             ["長期前受金", 11_060_263_895, "－", "貸借対照表 5 繰延収益(1)"],
             ["　収益化累計額", -3_568_580_403, "－", "同(2)。未収益化残高は7,491,683,492円"],
             ["有形固定資産", 9_159_891_337, "－", "貸借対照表 1(1)"],
             ["資金残高（年度末）", 50_295_933, "－", "残高月報・キャッシュフロー計算書。前年度末164,922,940円"]],
            aligns=["left", "right", "center", "left"], fmt={1: "#,##0"},
            flags=[None] * 8 + ["bad"])
    s.sub("■　業務量・繰入金")
    s.table(["項　目", "数　値", "単位", "出典・備考"],
            [["有収水量", R7["有収水量"], "㎥", "決算書"],
             ["年間総配水量", R7["年間総配水量"], "㎥", "有収率73.64％から検算。無収水量366,807㎥"],
             ["一日平均配水量", 3_812, "㎥", "決算書"],
             ["一日配水能力", 12_381, "㎥", "経営戦略（案）3-2"],
             ["給水人口（年度末）", R7["給水人口"], "人", "決算書。前年度5,769人"],
             ["他会計補助金（収益的収入分）", 166_987_093, "円", "決算書 注記5"],
             ["　うち繰出基準内", 553_019, "円", "児童手当240,000＋旧簡水起債利子313,019"],
             ["　うち繰出基準外", R7["基準外補助金"], "円",
              "★職員給与費28,566,469＋保険料9,077,628＋江島応急修繕等128,789,977"],
             ["　　うち江島海底送水管応急修繕等", R7["一過性_江島"], "円", "★一過性。平常時原価の算定で控除"]],
            aligns=["left", "right", "center", "left"], fmt={1: "#,##0"},
            flags=[None] * 5 + [None, None, "bad", "bad"])

    # ---------- 02 指標の算式と値 ----------
    s = Sheet(wb, "02_指標の算式と値", 5, {"A": 26, "B": 52, "C": 16, "D": 14, "E": 46})
    s.title("２　指標の算式と値　－　同じ名称で複数の値が存在するもの",
            "★成果物では必ず「どの算式か」を明記すること")
    s.sub("■　給水原価（5通り）", orange=True)
    s.table(["区　分", "算　式", "値（円/㎥）", "年度", "使う場面"],
            [["経営比較分析表（総務省）", "（経常費用 − 長期前受金戻入益）÷ 年間総有収水量",
              round(genka_a, 2), "令和７年度", "★他団体比較・経営戦略の経営分析。R5公表値218.64円との経年比較に使う"],
             ["経営比較分析表（総務省）", "同　上", 218.64, "令和５年度", "公表値。本セッションで算式を検証・再現済み"],
             ["日本水道協会 経営指標", "経常費用 ÷ 年間総有収水量", round(genka_nis, 2), "令和７年度",
              "決算審査意見書が採用。全国平均231.15円（給水人口5千人〜1万人）との比較"],
             ["平常時ベース（当方算定）", "（経常費用 − 江島応急給水128,790千円 − 長期前受金戻入益）÷ 有収水量",
              round(genka_b, 2), "令和７年度", "一過性要因を除いた実力値。料金改定率の下限の根拠"],
             ["総括原価ベース", "（維持管理費＋減価償却費＋支払利息＋資産維持費 − 控除項目）÷ 有収水量",
              "シート04", "R8〜R12", "★料金算定。控除項目の設定に論点あり（シート04参照）"]],
            aligns=["left", "left", "right", "center", "left"], flags=["good", None, None, "good", "warn"], h=46)
    s.sub("■　供給単価")
    s.table(["区　分", "算　式", "値（円/㎥）", "年度", "備　考"],
            [["★正（税抜）", "給水収益（税抜）119,480,750 ÷ 有収水量1,024,576", round(kyokyu, 2), "令和７年度",
              "総務省指標・経営比較分析表との比較はこの値を使う"],
             ["税込", "給水収益（税込）131,428,825 ÷ 有収水量1,024,576", 128.28, "令和７年度",
              "★引継ぎメモの128.2円はこれ。他団体比較には使えない"]],
            aligns=["left", "left", "right", "center", "left"], flags=["good", "bad"], h=32)
    s.sub("■　料金回収率")
    s.table(["区　分", "算　式", "値（％）", "年度", "備　考"],
            [["★令和７年度", "給水収益 ÷（経常費用 − 長期前受金戻入益）",
              round(R7["給水収益"] / kaisyu_a * 100, 2), "令和７年度", "決算書掲載値と一致"],
             ["平常時ベース", "給水収益 ÷（平常時経常費用 − 長期前受金戻入益）",
              round(R7["給水収益"] / kaisyu_b * 100, 2), "令和７年度", "江島応急給水を除く"],
             ["令和５年度", "同　上", 52.87, "令和５年度", "★引継ぎメモが「R7実績」欄に記載していた値"],
             ["令和６年度", "同　上", 47.11, "令和６年度", "決算書掲載値"]],
            aligns=["left", "left", "right", "center", "left"], flags=["good", "good", "bad", None], h=30)
    s.note([
        "※ 経常収支比率も同様に注意が必要です。令和７年度の公表ベースは97.10％ですが、繰出基準外の他会計補助金166,434,074円を除いた実質は72.44％です。",
        "※ 住民・議会説明では「現行料金は必要な費用の35％しか賄えていない（令和７年度）」が正しい表現です。「半分」ではありません。",
    ])

    # ---------- 03 有収水量推計の見直し ----------
    s = Sheet(wb, "03_有収水量推計の見直し", 6, {"A": 26, "B": 16, "C": 14, "D": 16, "E": 14, "F": 48})
    s.title("３　有収水量推計の見直し", "経営戦略（案）の推計方法と令和７年度実績との乖離")
    s.sub("■　実績（1人1日あたり有収水量）")
    rows = []
    for name, vol, pop in JISSEKI:
        rows.append([name, vol, pop, round(vol / pop / 365 * 1000, 1), "", "決算書・経営比較分析表"])
    avg = sum(v / p / 365 * 1000 for _, v, p in JISSEKI) / len(JISSEKI)
    rows.append(["直近3か年平均", "", "", round(avg, 1), "", "★見直し推計の基礎とする"])
    rows.append(["経営戦略（案）の前提", AN["R8_有収水量"], AN["R8_給水人口"],
                 round(AN["R8_有収水量"] / AN["R8_給水人口"] / 365 * 1000, 1), "令和８年度",
                 "★過去10年平均。震災復興期（有収水量が低く給水人口が多い時期）を含むため過小"])
    s.table(["区　分", "有収水量（㎥）", "給水人口（人）", "1人1日（L）", "年度", "備　考"],
            rows, aligns=["left", "right", "right", "right", "center", "left"],
            flags=[None, None, "good", "good", "bad"], fmt={1: "#,##0", 2: "#,##0", 3: "0.0"})
    s.sub("■　見直し案")
    r8p, r17p = AN["R8_給水人口"], AN["R17_給水人口"]
    a_r8 = r8p * avg / 1000 * 365
    a_r17 = r17p * avg / 1000 * 365
    s.table(["推計方法", "令和８年度（㎥）", "令和17年度（㎥）", "案との差", "評　価"],
            [["経営戦略（案）＝給水人口 × 439L", AN["R8_有収水量"], AN["R17_有収水量"], "－",
              "★令和７年度実績1,024,576㎥を12％下回る。初年度から実績と乖離する推計"],
             [f"見直し案Ａ＝給水人口 × {avg:.0f}L（直近3か年平均）", round(a_r8), round(a_r17),
              f"＋{(a_r8/AN['R8_有収水量']-1)*100:.1f}％", "実績の水準を反映。最も簡便で説明しやすい"],
             ["見直し案Ｂ＝用途別（家事用は人口比例、非家事用は別トレンド）", "要試算", "要試算", "－",
              "★理論的には最も妥当。ただし非家事用のトレンド設定に事業所側の情報が必要"]],
            aligns=["left", "right", "right", "right", "left"],
            flags=["bad", "good", "warn"], fmt={1: "#,##0", 2: "#,##0"}, h=40)
    s.sub("■【重要】用途区分別の有収水量シェア　－　給水人口比例の推計が成り立つのは44.5％だけ", orange=True)
    tot = sum(v for _, _, v in YOUTO)
    rows = [[k, n, v, round(v / tot * 100, 1), ""] for k, n, v in YOUTO]
    rows.append(["合　計", sum(n for _, n, _ in YOUTO), tot, 100.0, ""])
    s.table(["用途区分", "件数", "水量（㎥／月）", "シェア（％）", "備　考"],
            rows, aligns=["left", "right", "right", "right", "left"],
            fmt={1: "#,##0", 2: "#,##0", 3: "0.0"},
            flags=["good", "bad", None, None, None, None, None, None], h=24)
    s.note([
        "※ 調定データ（1か月分3,060件）より集計。",
        "★ 家事用は有収水量の44.5％にすぎず、工業用が36.1％（66件）を占めます。工業用の大半は水産加工業です。",
        "★ 経営戦略（案）は有収水量の全量を「給水人口 × 1人1日あたり有収水量」で推計していますが、",
        "　 55.5％を占める非家事用は給水人口に連動しません。推計方法そのものに構造的な弱点があります。",
        "※ 見直し案Ｂを採る場合、水産加工業の生産動向・事業所数の見通しについて町・事業者への確認が必要です。",
        "※ なお、給水人口の推計値（令和８年度5,625人）は令和７年度実績5,656人と整合しており、こちらは差し替え不要です。",
    ])

    # ---------- 04 総括原価の再計算 ----------
    s = Sheet(wb, "04_総括原価の再計算", 5, {"A": 34, "B": 20, "C": 20, "D": 18, "E": 50})
    s.title("４　総括原価の再計算　－　控除項目に論点",
            "引継ぎメモ記載の費目（令和８〜12年度の5年計）に基づく再計算。原資料の確認が必要")
    s.sub("■【重要】長期前受金戻入益の控除額", orange=True)
    s.table(["区　分", "5年計（千円）", "年平均（千円）", "対比", "根　拠"],
            [["引継ぎメモの控除額", -450_000, -90_000, "－", "「推計値・要確定後更新」と記載されている"],
             ["★令和７年度実績ベース", -1_671_262, -334_252, "3.7倍",
              "令和７年度実績334,252,462円×5年。未収益化残高7,491,683千円に対し年334,252千円なので、今後20年以上この水準が続く"]],
            aligns=["left", "right", "right", "center", "left"],
            flags=["bad", "good"], fmt={1: "#,##0", 2: "#,##0"}, h=48)
    s.note([
        "★ 控除額が3.7倍違うため、総括原価はほぼ倍の差が出ます。原資料（onagawa_cost_sim_v1.xlsx 等）での確認が最優先です。",
    ])

    sub_total = sum(v for k, v in SOKATSU.items() if not k.startswith("控除"))
    memo_total = sub_total + SOKATSU["控除_長期前受金戻入益_メモ"] + SOKATSU["控除_他会計補助金等"]
    fix_total = sub_total - 1_671_262 + SOKATSU["控除_他会計補助金等"]

    s.sub("■　総括原価（5年計・千円）")
    s.table(["費　目", "引継ぎメモ", "★実績ベースに是正", "差", "備　考"],
            [["維持管理費", SOKATSU["維持管理費"], SOKATSU["維持管理費"], 0,
              "年平均134,465千円。令和７年度の平常時（江島分控除後）137,751千円とほぼ一致"],
             ["減価償却費", SOKATSU["減価償却費"], SOKATSU["減価償却費"], 0,
              "年平均406,488千円。令和７年度実績393,746千円"],
             ["既往債利息", SOKATSU["既往債利息"], SOKATSU["既往債利息"], 0, "年平均10,117千円"],
             ["新規債利息", SOKATSU["新規債利息"], SOKATSU["新規債利息"], 0,
              "年平均25,480千円。★令和７年度の新規借入利率は加重平均2.58％"],
             ["資産維持費（3％）", SOKATSU["資産維持費"], SOKATSU["資産維持費"], 0,
              "対象資産896,300千円×3％×5年。★対象資産の範囲の根拠確認が必要"],
             ["小　計", sub_total, sub_total, 0, ""],
             ["控除　長期前受金戻入益", SOKATSU["控除_長期前受金戻入益_メモ"], -1_671_262,
              -1_671_262 - SOKATSU["控除_長期前受金戻入益_メモ"], "★令和７年度実績334,252千円×5年"],
             ["控除　他会計補助金等", SOKATSU["控除_他会計補助金等"], SOKATSU["控除_他会計補助金等"], 0, ""],
             ["★総括原価", memo_total, fix_total, fix_total - memo_total, ""]],
            aligns=["left", "right", "right", "right", "left"],
            flags=[None] * 5 + [None, "bad", None, "good"],
            fmt={1: "#,##0", 2: "#,##0", 3: "#,##0"}, h=32)

    # 有収水量5年計
    an_5 = 0.0
    slope = (AN["R17_有収水量"] - AN["R8_有収水量"]) / 9
    for i in range(5):
        an_5 += AN["R8_有収水量"] + slope * i
    rev_5 = an_5 * (a_r8 / AN["R8_有収水量"])

    s.sub("■　給水原価（総括原価ベース）と必要改定率")
    s.table(["前　提", "総括原価\n5年計（千円）", "有収水量\n5年計（千㎥）", "給水原価\n（円/㎥）", "料金回収率100％に必要な改定率"],
            [["引継ぎメモのまま", memo_total, round(an_5 / 1000), round(memo_total * 1000 / an_5, 1),
              f"＋{(memo_total*1000/5/R7['給水収益']-1)*100:.1f}％"],
             ["控除額のみ是正（有収水量は案のまま）", fix_total, round(an_5 / 1000),
              round(fix_total * 1000 / an_5, 1), f"＋{(fix_total*1000/5/R7['給水収益']-1)*100:.1f}％"],
             ["★控除額の是正＋有収水量の見直し", fix_total, round(rev_5 / 1000),
              round(fix_total * 1000 / rev_5, 1), f"＋{(fix_total*1000/5/R7['給水収益']-1)*100:.1f}％"],
             ["（参考）資産維持費を算入しない場合", fix_total - SOKATSU["資産維持費"], round(rev_5 / 1000),
              round((fix_total - SOKATSU["資産維持費"]) * 1000 / rev_5, 1),
              f"＋{((fix_total-SOKATSU['資産維持費'])*1000/5/R7['給水収益']-1)*100:.1f}％"]],
            aligns=["left", "right", "right", "right", "right"],
            flags=["bad", "warn", "good", None], fmt={1: "#,##0", 2: "#,##0", 3: "#,##0.0"}, h=34)
    s.note([
        "※ 有収水量5年計は、経営戦略（案）の令和８年度901千㎥・令和17年度723千㎥を線形補間して令和８〜12年度分を積み上げたものです（4,307千㎥）。",
        "　 引継ぎメモの給水原価572円は約4,462千㎥を前提とした値と推定され、当方の補間値との差（約4％）が592.5円との差になっています。原価側の差ではありません。",
        "※ 必要改定率は「総括原価の年平均 ÷ 令和７年度の給水収益119,480,750円」で算定しています（有収水量の増減は考慮していない簡易計算）。",
        "★ 引継ぎメモの給水原価572円に対し、控除額を実績ベースに是正すると300円前後まで下がります。",
        "　 「1回の改定では不十分」という結論自体は変わりませんが、住民・議会に示す数値としては大きな差です。",
        "★ 本シートは引継ぎメモ記載の費目を前提とした再計算です。原資料（onagawa_cost_sim_v1.xlsx、onagawa_tariff_cost_calc.xlsx）での確認が必要です。",
    ])

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = C["subhead"]
    wb.worksheets[0].sheet_properties.tabColor = C["header"]
    path = os.path.join(OUT_DIR, "女川町_水道_数値定義と確定値一覧.xlsx")
    wb.save(path)
    print("saved:", path)
    print(f"  供給単価(税抜) {kyokyu:.2f}円 / 給水原価(経営比較) {genka_a:.2f}円 / 平常時 {genka_b:.2f}円")
    print(f"  1人1日 直近3年平均 {avg:.1f}L （案の前提 439L）")
    print(f"  総括原価 メモ {memo_total:,}千円 → 是正後 {fix_total:,}千円")
    print(f"  給水原価 メモ {memo_total*1000/an_5:.1f}円 → 是正後(有収水量見直し) {fix_total*1000/rev_5:.1f}円")
    print(f"  必要改定率 メモ +{(memo_total*1000/5/R7['給水収益']-1)*100:.1f}% → 是正後 +{(fix_total*1000/5/R7['給水収益']-1)*100:.1f}%")


if __name__ == "__main__":
    main()
