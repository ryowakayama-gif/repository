# -*- coding: utf-8 -*-
"""
女川町 上下水道事業　支援業務 工程表ジェネレータ
─────────────────────────────────────────────
１　経営戦略の更新・再評価（水道・下水道）
２　浄化槽事業　公営企業版経営戦略の策定
３　財務・経営指標分析による課題抽出と改善案作成（料金改定提案を含む）
４　令和７年度決算サマリと経営戦略（案）との差異
５　会計システムの見直し（ビズアップの提供範囲・考え方）

・工程は「月単位／上旬・中旬・下旬」の粒度で表示
・令和８年８月（現在）を起点とする
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# 配色
# ============================================================
C = {
    "header":   "1F3864",   # 濃紺（大見出し）
    "subhead":  "2E75B6",   # 青（小見出し）
    "band":     "DDEBF7",   # 淡青（帯）
    "alt":      "F7FAFC",   # ゼブラ
    "white":    "FFFFFF",
    "gray":     "808080",
    # 工程バー
    "bar_main": "2E75B6",   # 主作業（ビズアップ）
    "bar_town": "70AD47",   # 町の作業
    "bar_prep": "BDD7EE",   # 準備・調整
    "bar_opt":  "FFD966",   # 任意／条件付き
    "ms":       "C00000",   # マイルストーン
    # 年度帯
    "fy8":      "1F3864",
    "fy9":      "375623",
    # 注意喚起
    "warn":     "FFF2CC",
    "bad":      "FCE4E4",
    "good":     "E2F0D9",
}

THIN = Side(border_style="thin", color="BFBFBF")
HAIR = Side(border_style="hair", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_H = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

FONT = "Meiryo UI"


def font(size=10, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)


def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


# ============================================================
# タイムライン定義（令和８年８月 ～ 令和10年３月）
# ============================================================
JUN = ["上", "中", "下"]


def build_months():
    """(年度ラベル, 和暦年, 月) のリストを返す"""
    ms = []
    # 令和８年度：R8.8 ～ R9.3
    for m in range(8, 13):
        ms.append(("令和８年度", 8, m))
    for m in range(1, 4):
        ms.append(("令和８年度", 9, m))
    # 令和９年度：R9.4 ～ R10.3
    for m in range(4, 13):
        ms.append(("令和９年度", 9, m))
    for m in range(1, 4):
        ms.append(("令和９年度", 10, m))
    return ms


MONTHS_ALL = build_months()                 # 20か月
MONTHS_R8 = [m for m in MONTHS_ALL if m[0] == "令和８年度"]   # 8か月

# 旬インデックス（"R8.9中" のようなキー → 0起点の列オフセット）
def jun_index(months, wareki, month, jun):
    for i, (_fy, w, mo) in enumerate(months):
        if w == wareki and mo == month:
            return i * 3 + JUN.index(jun)
    raise KeyError(f"R{wareki}.{month}{jun} は対象期間外です")


# ============================================================
# 共通：工程表シートの描画
# ============================================================
LABEL_COLS = 4          # No / 区分 / 工程 / 主担当
FIRST_BAR_COL = LABEL_COLS + 1


def draw_gantt(ws, title, subtitle, months, rows, notes=None, memo_col=True):
    """
    rows: list of dict
        kind : "group" | "task"
        no   : 番号（task のみ）
        cat  : 区分
        name : 工程名
        who  : 主担当
        bars : [(開始キー, 終了キー, 種別)]  種別: main/town/prep/opt
        ms   : [(キー, 表示文字)]  マイルストーン
        memo : 備考
    """
    ncols = LABEL_COLS + len(months) * 3 + (1 if memo_col else 0)

    # ---- タイトル -------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = font(14, True, C["white"])
    c.fill = fill(C["header"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=2 - 1, value=subtitle)
    c.font = font(9, False, "1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ---- 凡例 -----------------------------------------------------
    legend = [
        ("■", C["bar_main"], "ビズアップ主担当"),
        ("■", C["bar_town"], "町（上下水道課）主担当"),
        ("■", C["bar_prep"], "準備・調整・待機"),
        ("■", C["bar_opt"], "任意／条件付き工程"),
        ("●", C["ms"], "マイルストーン（議会・審議会等）"),
    ]
    col = 1
    for mark, color, text in legend:
        c = ws.cell(row=3, column=col, value=mark)
        c.font = font(10, True, color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=3, start_column=col + 1, end_row=3, end_column=col + 5)
        c = ws.cell(row=3, column=col + 1, value=text)
        c.font = font(8.5)
        c.alignment = Alignment(horizontal="left", vertical="center")
        col += 6
    ws.row_dimensions[3].height = 16

    # ---- ヘッダー（年度／月／旬）---------------------------------
    hr1, hr2, hr3 = 4, 5, 6
    for r in (hr1, hr2, hr3):
        for cc in range(1, ncols + 1):
            cell = ws.cell(row=r, column=cc)
            cell.border = BORDER

    labels = ["No.", "区　分", "工　程", "主担当"]
    for i, lab in enumerate(labels):
        ws.merge_cells(start_row=hr1, start_column=i + 1, end_row=hr3, end_column=i + 1)
        c = ws.cell(row=hr1, column=i + 1, value=lab)
        c.font = font(9.5, True, C["white"])
        c.fill = fill(C["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 年度帯
    i = 0
    while i < len(months):
        fy = months[i][0]
        j = i
        while j < len(months) and months[j][0] == fy:
            j += 1
        s = FIRST_BAR_COL + i * 3
        e = FIRST_BAR_COL + j * 3 - 1
        ws.merge_cells(start_row=hr1, start_column=s, end_row=hr1, end_column=e)
        c = ws.cell(row=hr1, column=s, value=fy)
        c.font = font(10, True, C["white"])
        c.fill = fill(C["fy8"] if "８" in fy else C["fy9"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        i = j

    # 月・旬
    for i, (_fy, w, mo) in enumerate(months):
        s = FIRST_BAR_COL + i * 3
        ws.merge_cells(start_row=hr2, start_column=s, end_row=hr2, end_column=s + 2)
        c = ws.cell(row=hr2, column=s, value=f"R{w}.{mo}")
        c.font = font(8.5, True, C["white"])
        c.fill = fill(C["subhead"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        for k, j in enumerate(JUN):
            c = ws.cell(row=hr3, column=s + k, value=j)
            c.font = font(8)
            c.fill = fill(C["band"])
            c.alignment = Alignment(horizontal="center", vertical="center")

    if memo_col:
        mc = ncols
        ws.merge_cells(start_row=hr1, start_column=mc, end_row=hr3, end_column=mc)
        c = ws.cell(row=hr1, column=mc, value="備　考／成果物")
        c.font = font(9.5, True, C["white"])
        c.fill = fill(C["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[hr1].height = 18
    ws.row_dimensions[hr2].height = 16
    ws.row_dimensions[hr3].height = 14

    # ---- 明細 -----------------------------------------------------
    r = hr3 + 1
    zebra = False
    for row in rows:
        if row["kind"] == "group":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            c = ws.cell(row=r, column=1, value=row["name"])
            c.font = font(10.5, True, C["white"])
            c.fill = fill(C["subhead"])
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[r].height = 20
            for cc in range(1, ncols + 1):
                ws.cell(row=r, column=cc).border = BORDER
            zebra = False
            r += 1
            continue

        base = C["alt"] if zebra else C["white"]
        zebra = not zebra

        vals = [row.get("no", ""), row.get("cat", ""), row.get("name", ""), row.get("who", "")]
        aligns = ["center", "center", "left", "center"]
        for i, (v, al) in enumerate(zip(vals, aligns)):
            c = ws.cell(row=r, column=i + 1, value=v)
            c.font = font(9, bold=(i == 2 and row.get("bold")))
            c.fill = fill(base)
            c.alignment = Alignment(horizontal=al, vertical="center",
                                    wrap_text=(i == 2), indent=(1 if al == "left" else 0))
            c.border = BORDER

        for k in range(len(months) * 3):
            c = ws.cell(row=r, column=FIRST_BAR_COL + k)
            c.fill = fill(base)
            c.border = BORDER_H

        # バー
        for (s_key, e_key, kind) in row.get("bars", []):
            s = jun_index(months, *s_key)
            e = jun_index(months, *e_key)
            color = {"main": C["bar_main"], "town": C["bar_town"],
                     "prep": C["bar_prep"], "opt": C["bar_opt"]}[kind]
            for k in range(s, e + 1):
                c = ws.cell(row=r, column=FIRST_BAR_COL + k)
                c.fill = fill(color)
                c.border = BORDER_H

        # マイルストーン（濃色バーの上では白抜きにして視認性を確保）
        dark = {C["bar_main"], C["bar_town"]}
        for (key, text) in row.get("ms", []):
            k = jun_index(months, *key)
            c = ws.cell(row=r, column=FIRST_BAR_COL + k, value="●")
            cur = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else ""
            on_dark = any(cur and cur.endswith(d) for d in dark)
            c.font = font(11, True, C["white"] if on_dark else C["ms"])
            c.alignment = Alignment(horizontal="center", vertical="center")

        if memo_col:
            c = ws.cell(row=r, column=ncols, value=row.get("memo", ""))
            c.font = font(8.5)
            c.fill = fill(base)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border = BORDER

        ws.row_dimensions[r].height = row.get("h", 28)
        r += 1

    # ---- 注記 -----------------------------------------------------
    if notes:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(ncols, 20))
        c = ws.cell(row=r, column=1, value="【前提・留意事項】")
        c.font = font(10, True, C["header"])
        r += 1
        for n in notes:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(ncols, 20))
            c = ws.cell(row=r, column=1, value="　" + n)
            c.font = font(9)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 16
            r += 1

    # ---- 列幅 -----------------------------------------------------
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 11
    for k in range(len(months) * 3):
        ws.column_dimensions[get_column_letter(FIRST_BAR_COL + k)].width = 2.6
    if memo_col:
        ws.column_dimensions[get_column_letter(ncols)].width = 40

    ws.freeze_panes = f"{get_column_letter(FIRST_BAR_COL)}{hr3 + 1}"
    ws.sheet_view.showGridLines = False
    return r


# ============================================================
# シート01　経営戦略の更新・再評価
# ============================================================
def sheet_strategy_update(wb):
    ws = wb.create_sheet("01_経営戦略の更新")
    B = "bars"

    rows = [
        {"kind": "group", "name": "Ⅰ　決算数値の反映（インプット整備）"},
        {"kind": "task", "no": 1, "cat": "決算反映", "who": "町",
         "name": "令和７年度決算の内部決定・決算書一式の提供\n（決算書／附属書類／固定資産明細／企業債明細）",
         B: [((8, 8, "上"), (8, 8, "中"), "town")],
         "ms": [((8, 8, "上"), "内部決定")],
         "memo": "決算書PDFは受領済。Excel版・決算統計様式を追加提供いただきたい", "h": 34},
        {"kind": "task", "no": 2, "cat": "決算反映", "who": "ビズアップ",
         "name": "試算シートへの決算数値入力（水道PL/BS・下水道予測シート）",
         B: [((8, 8, "中"), (8, 8, "下"), "main")],
         "memo": "水道は R6・R7 の２か年分をまとめて追加入力（現行シートは R5 まで）"},
        {"kind": "task", "no": 3, "cat": "決算反映", "who": "ビズアップ",
         "name": "実績期間・平均値の再設定（R1〜R5 → R3〜R7）",
         B: [((8, 8, "下"), (8, 9, "上"), "main")],
         "memo": "費用推計は「直近５か年平均×物価上昇率」方式。基準年の入替で全推計が変動"},
        {"kind": "task", "no": 4, "cat": "決算反映", "who": "ビズアップ",
         "name": "異常値（過年度修正損益・江島応急給水）の分離処理方針の決定",
         B: [((8, 9, "上"), (8, 9, "中"), "main")],
         "memo": "特損23億・特益7.7億は一過性。推計母数から除外する整理が必須", "h": 30},
        {"kind": "task", "no": 5, "cat": "決算反映", "who": "両者",
         "name": "前提条件の再設定（給水人口・有収水量・物価上昇率・企業債利率）",
         B: [((8, 9, "上"), (8, 9, "下"), "main")],
         "memo": "有収率73.64%への急落を織り込むか（捨水・応急給水の解消時期）を要協議", "h": 30},

        {"kind": "group", "name": "Ⅱ　再計算・差異分析"},
        {"kind": "task", "no": 6, "cat": "再計算", "who": "町",
         "name": "議会認定（令和７年度決算）",
         "ms": [((8, 9, "中"), "認定")],
         B: [((8, 9, "上"), (8, 9, "中"), "prep")],
         "memo": "令和７年度実績：９月定例会に提出→中旬認定。認定後に指標を確定"},
        {"kind": "task", "no": 7, "cat": "再計算", "who": "ビズアップ",
         "name": "決算統計・経営比較分析表（R7）の指標算定",
         B: [((8, 9, "中"), (8, 10, "上"), "main")],
         "memo": "経常収支比率／料金・経費回収率／給水・汚水処理原価／企業債残高対給水収益比率 ほか", "h": 30},
        {"kind": "task", "no": 8, "cat": "再計算", "who": "ビズアップ",
         "name": "成行シミュレーション（30年）の再計算",
         B: [((8, 9, "下"), (8, 10, "中"), "main")],
         "memo": "水道：資金残高／下水道：予測シート・経常収支・経費回収率"},
        {"kind": "task", "no": 9, "cat": "再計算", "who": "ビズアップ",
         "name": "現行案との差異分析・影響評価（レポート化）",
         B: [((8, 10, "中"), (8, 10, "下"), "main")],
         "memo": "★成果物：差異分析表。資金残高・累積欠損金の見通しが大きく変わる見込み", "h": 30},

        {"kind": "group", "name": "Ⅲ　投資・財源計画の見直し"},
        {"kind": "task", "no": 10, "cat": "投資計画", "who": "両者",
         "name": "投資計画への未反映事業の追加（鷲神浄水場高度処理・江島海底送水管本復旧）",
         B: [((8, 10, "下"), (8, 11, "下"), "main")],
         "memo": "現行案の建設改良費約45億円に未計上。事業費・財源・年割の確認が必要", "h": 34},
        {"kind": "task", "no": 11, "cat": "投資計画", "who": "両者",
         "name": "繰入金の整理（基準内／基準外の区分・繰出基準の再点検）",
         B: [((8, 10, "下"), (8, 11, "下"), "main")],
         "memo": "水道R7は基準外補助金1.67億円で収支を補填。一般会計との協議材料を整理", "h": 30},
        {"kind": "task", "no": 12, "cat": "料金", "who": "ビズアップ",
         "name": "料金・使用料改定パターンの再検証",
         B: [((8, 11, "上"), (8, 12, "上"), "main")],
         "memo": "水道：基本料金×1.5案の妥当性／下水道：パターン❷（R14・25%）の再評価", "h": 30},

        {"kind": "group", "name": "Ⅳ　本文修正・確定手続"},
        {"kind": "task", "no": 13, "cat": "本文修正", "who": "ビズアップ",
         "name": "経営戦略（案）本文・図表の修正（水道・下水道）",
         B: [((8, 11, "下"), (8, 12, "下"), "main")],
         "memo": "水道：3-2給水状況、3-8経営分析、資料４〜13を全面差替", "h": 30},
        {"kind": "task", "no": 14, "cat": "本文修正", "who": "町",
         "name": "庁内協議・課内確認",
         B: [((8, 12, "中"), (9, 1, "上"), "town")],
         "memo": ""},
        {"kind": "task", "no": 15, "cat": "確定手続", "who": "町",
         "name": "上下水道審議会への説明・意見聴取",
         B: [((9, 1, "上"), (9, 1, "下"), "town")],
         "ms": [((9, 1, "中"), "審議会")],
         "memo": "下水道戦略の推進体制に位置づけあり。水道も併せて付議"},
        {"kind": "task", "no": 16, "cat": "確定手続", "who": "町",
         "name": "パブリックコメント（実施する場合）",
         B: [((9, 1, "下"), (9, 2, "中"), "opt")],
         "memo": "実施要否は町の判断。実施する場合は約30日間を確保"},
        {"kind": "task", "no": 17, "cat": "確定手続", "who": "町",
         "name": "議会（全員協議会）への報告",
         B: [((9, 2, "中"), (9, 2, "下"), "town")],
         "ms": [((9, 2, "下"), "議会報告")],
         "memo": "３月定例会前の説明を想定"},
        {"kind": "task", "no": 18, "cat": "確定手続", "who": "両者",
         "name": "最終版の作成・公表・総務省報告",
         B: [((9, 3, "上"), (9, 3, "下"), "main")],
         "ms": [((9, 3, "下"), "策定・公表")],
         "memo": "★成果物：経営戦略（確定版）水道・下水道", "h": 30},
    ]

    notes = [
        "① 現行の水道経営戦略（案）は「令和８年３月」付・計画期間 令和８〜17年度。本文中に『令和７年度決算見込みの数値に変動予定あり！再計算の上、入替すること！』の注記があり、決算確定を待って確定版とする前提です。",
        "② 水道の試算シート（20251209版）の実績期間は R1〜R5 までです。令和６年度決算が未反映のため、今回は R6・R7 の２か年分をまとめて反映します。",
        "③ 令和７年度は、過年度会計処理の一括修正（特別損失23.3億円・特別利益7.7億円）と江島海底送水管破損への応急給水という２つの一過性要因が重なっています。推計の母数からこれらを分離しないと、将来推計が過大に悪化します。",
        "④ 令和７年度末の資金残高は上水道50,296千円・下水道49,983千円（前年度末はそれぞれ164,923千円・76,235千円）。水道の現行案は「令和17年度に約4千万円」と見込んでいましたが、実績が既にその水準に到達しています。資金計画の前提の見直しが必要です。",
        "⑤ 審議会・議会の開催時期は町の議事日程により変動します。確定後にスケジュールを調整します。",
    ]

    draw_gantt(ws,
               "１　経営戦略の更新・再評価　工程表（水道事業・下水道事業）",
               "令和７年度決算の確定を受けた経営戦略（案）の数値更新・再評価　／　対象期間：令和８年８月〜令和９年３月",
               MONTHS_R8, rows, notes)
    return ws


# ============================================================
# シート02　浄化槽事業　経営戦略の策定
# ============================================================
def sheet_johkasou(wb):
    ws = wb.create_sheet("02_浄化槽_経営戦略策定")
    B = "bars"

    rows = [
        {"kind": "group", "name": "Ⅰ　令和８年度：下準備と予算獲得（パターンＢ＝推奨ルート）"},
        {"kind": "task", "no": 1, "cat": "現況整理", "who": "両者",
         "name": "浄化槽事業の現況データ棚卸\n（設置基数・設置年度・区域・耐用年数・維持管理委託内容）",
         B: [((8, 8, "中"), (8, 9, "中"), "main")],
         "memo": "経営戦略の骨格を決めるインプット。資産台帳と現地基数の突合を先行", "h": 34},
        {"kind": "task", "no": 2, "cat": "現況整理", "who": "ビズアップ",
         "name": "セグメント財務情報の整理（R4〜R7の４か年）",
         B: [((8, 9, "上"), (8, 9, "下"), "main")],
         "memo": "法適用（R4.4.1〜）以降の実績が揃う。R7：収益22,313千円／費用22,524千円", "h": 30},
        {"kind": "task", "no": 3, "cat": "方針決定", "who": "両者",
         "name": "策定方式の決定（単独の経営戦略とするか／下水道経営戦略に「浄化槽事業編」として統合するか）",
         B: [((8, 9, "中"), (8, 10, "上"), "main")],
         "ms": [((8, 10, "上"), "方針決定")],
         "memo": "★要決定。統合方式なら期間・費用とも大幅圧縮が可能", "h": 34},
        {"kind": "task", "no": 4, "cat": "予算", "who": "ビズアップ",
         "name": "業務範囲・仕様書案の作成、見積提出",
         B: [((8, 9, "下"), (8, 10, "中"), "main")],
         "memo": "予算要求資料として町へ提出"},
        {"kind": "task", "no": 5, "cat": "予算", "who": "町",
         "name": "令和９年度当初予算要求",
         B: [((8, 10, "下"), (8, 11, "中"), "town")],
         "ms": [((8, 11, "上"), "予算要求")],
         "memo": "★要確認：町の予算要求締切。10月下旬〜11月上旬を仮置き", "h": 30},
        {"kind": "task", "no": 6, "cat": "予算", "who": "町",
         "name": "査定・内示・予算案調製",
         B: [((8, 11, "下"), (9, 2, "上"), "prep")],
         "memo": ""},
        {"kind": "task", "no": 7, "cat": "予算", "who": "町",
         "name": "３月定例会　当初予算議決",
         B: [((9, 3, "上"), (9, 3, "中"), "town")],
         "ms": [((9, 3, "中"), "予算議決")],
         "memo": ""},
        {"kind": "task", "no": 8, "cat": "契約", "who": "両者",
         "name": "契約手続・キックオフ",
         B: [((9, 4, "上"), (9, 4, "下"), "main")],
         "memo": ""},

        {"kind": "group", "name": "Ⅱ　令和９年度：経営戦略の策定（総務省「経営戦略策定・改定マニュアル」準拠）"},
        {"kind": "task", "no": 9, "cat": "現状分析", "who": "ビズアップ",
         "name": "事業の現況整理（施設・組織・使用料体系・民間活力）",
         B: [((9, 5, "上"), (9, 6, "上"), "main")],
         "memo": "戦略「Ⅱ 現状」に相当"},
        {"kind": "task", "no": 10, "cat": "現状分析", "who": "ビズアップ",
         "name": "経営状況分析（決算推移・経営指標・類似団体比較）",
         B: [((9, 5, "下"), (9, 6, "下"), "main")],
         "memo": "R7実績：経常収支比率99.06%／経費回収率64.40%（R6:74.48%から急落）", "h": 30},
        {"kind": "task", "no": 11, "cat": "予測", "who": "ビズアップ",
         "name": "人口・接続基数・有収水量の予測（社人研推計ベース）",
         B: [((9, 6, "下"), (9, 7, "下"), "main")],
         "memo": "下水道戦略と同一の人口推計を使用し整合を確保"},
        {"kind": "task", "no": 12, "cat": "予測", "who": "両者",
         "name": "投資試算（浄化槽の更新・新設、下水道区域縮小に伴う転換分）",
         B: [((9, 7, "下"), (9, 9, "上"), "main")],
         "memo": "令和２年度の下水道区域縮小で浄化槽区域へ転換した分の整備見込みを反映", "h": 30},
        {"kind": "task", "no": 13, "cat": "予測", "who": "ビズアップ",
         "name": "財源試算（使用料・一般会計繰入金・企業債・国庫補助）",
         B: [((9, 8, "下"), (9, 10, "上"), "main")],
         "memo": "R7他会計繰入金5,928千円。繰出基準に照らした基準内／外の整理を含む"},
        {"kind": "task", "no": 14, "cat": "収支計画", "who": "ビズアップ",
         "name": "投資・財政計画（30年収支シミュレーション）の作成",
         B: [((9, 10, "上"), (9, 11, "中"), "main")],
         "memo": "★成果物：収支シミュレーション"},
        {"kind": "task", "no": 15, "cat": "料金", "who": "ビズアップ",
         "name": "使用料水準の検証（総括原価による原価計算・改定パターン試算）",
         B: [((9, 11, "上"), (9, 12, "中"), "main")],
         "memo": "経費回収率100%に向けた改定率・時期の複数案を提示"},
        {"kind": "task", "no": 16, "cat": "戦略", "who": "両者",
         "name": "経営目標・基本方針・ロードマップの設定",
         B: [((9, 12, "上"), (9, 12, "下"), "main")],
         "memo": "下水道戦略（目標：経常収支比率100%／経費回収率80%）との整合を確保"},
        {"kind": "task", "no": 17, "cat": "とりまとめ", "who": "ビズアップ",
         "name": "素案の作成",
         B: [((10, 1, "上"), (10, 1, "下"), "main")],
         "memo": "★成果物：経営戦略（素案）"},
        {"kind": "task", "no": 18, "cat": "確定手続", "who": "町",
         "name": "庁内協議・上下水道審議会への付議",
         B: [((10, 1, "下"), (10, 2, "中"), "town")],
         "ms": [((10, 2, "上"), "審議会")],
         "memo": ""},
        {"kind": "task", "no": 19, "cat": "確定手続", "who": "町",
         "name": "パブリックコメント・議会報告",
         B: [((10, 2, "上"), (10, 3, "上"), "opt")],
         "memo": "実施要否は町の判断"},
        {"kind": "task", "no": 20, "cat": "確定手続", "who": "両者",
         "name": "策定・公表・総務省報告",
         B: [((10, 3, "上"), (10, 3, "下"), "main")],
         "ms": [((10, 3, "下"), "策定・公表")],
         "memo": "★成果物：浄化槽事業経営戦略（確定版）"},

        {"kind": "group", "name": "【参考】パターンＡ：令和８年度内に策定する場合（補正予算・流用等で財源措置できる場合）"},
        {"kind": "task", "no": "A1", "cat": "短縮案", "who": "両者",
         "name": "現況整理・経営状況分析",
         B: [((8, 9, "上"), (8, 10, "下"), "opt")],
         "memo": "パターンＢの No.1・2・9・10 を前倒しで実施"},
        {"kind": "task", "no": "A2", "cat": "短縮案", "who": "ビズアップ",
         "name": "予測・投資／財源試算・収支シミュレーション",
         B: [((8, 10, "下"), (8, 12, "下"), "opt")],
         "memo": "下水道の予測シートを流用し工数を圧縮"},
        {"kind": "task", "no": "A3", "cat": "短縮案", "who": "ビズアップ",
         "name": "使用料水準の検証・素案作成",
         B: [((9, 1, "上"), (9, 2, "上"), "opt")],
         "memo": ""},
        {"kind": "task", "no": "A4", "cat": "短縮案", "who": "町",
         "name": "審議会・議会報告・策定・公表",
         B: [((9, 2, "上"), (9, 3, "下"), "opt")],
         "ms": [((9, 3, "下"), "策定・公表")],
         "memo": "下水道経営戦略の確定と同時期になるため、町側の負荷が集中する点に留意", "h": 30},
    ]

    notes = [
        "① 浄化槽事業は令和４年４月１日から地方公営企業法（全部適用）が適用され、下水道事業会計の報告セグメントとして区分経理されています。「公営企業版の経営戦略」を策定する前提条件は既に整っています。",
        "② 現行の下水道事業経営戦略（案）は対象を「女川町公共下水道事業」に限定しており、浄化槽事業は対象外です。この空白を埋めるのが本業務の位置づけです。",
        "③ 令和７年度実績（セグメント情報）：営業収益8,804千円／営業費用22,451千円／経常損益△211千円／セグメント資産156,130千円／セグメント負債147,411千円／減価償却費9,200千円／他会計繰入金5,928千円。",
        "④ 経費回収率は R5:71.66% → R6:74.48% → R7:64.40% と直近で急落しています。使用料水準の検証は優先度の高い論点です。",
        "⑤ パターンＡ（令和８年度内）／パターンＢ（令和９年度）はいずれも実行可能ですが、令和８年度は水道・下水道の経営戦略確定と決算対応が重なるため、品質確保の観点からパターンＢを推奨します。",
        "⑥ 予算要求の締切・議会日程は町の実際のスケジュールに合わせて確定させてください（本表は10月下旬要求を仮置き）。",
    ]

    draw_gantt(ws,
               "１（２）　浄化槽事業　公営企業版　経営戦略　策定スケジュール",
               "対象：女川町下水道事業会計　浄化槽事業（報告セグメント）　／　対象期間：令和８年８月〜令和10年３月",
               MONTHS_ALL, rows, notes)
    return ws


# ============================================================
# シート03　財務・経営指標分析
# ============================================================
def sheet_analysis(wb):
    ws = wb.create_sheet("03_財務経営指標分析")
    B = "bars"

    rows = [
        {"kind": "group", "name": "Ⅰ　基礎データの整備と指標算定"},
        {"kind": "task", "no": 1, "cat": "データ整備", "who": "町",
         "name": "令和７年度決算数値の提供（上水道・公共下水道・浄化槽の３区分）",
         B: [((8, 8, "上"), (8, 8, "中"), "town")],
         "ms": [((8, 8, "上"), "決算内部決定")],
         "memo": "決算書・決算統計・セグメント情報・企業債明細・繰入金内訳", "h": 30},
        {"kind": "task", "no": 2, "cat": "データ整備", "who": "ビズアップ",
         "name": "経年データベースの作成（H26〜R7）",
         B: [((8, 8, "中"), (8, 9, "上"), "main")],
         "memo": "水道は R6・R7 の２か年分を新規追加"},
        {"kind": "task", "no": 3, "cat": "指標算定", "who": "ビズアップ",
         "name": "経営指標の算定（経常収支比率・料金／経費回収率・給水／汚水処理原価・有収率ほか）",
         B: [((8, 8, "下"), (8, 9, "中"), "main")],
         "memo": "経営比較分析表の全指標＋資金残高・累積欠損金の推移", "h": 30},
        {"kind": "task", "no": 4, "cat": "指標算定", "who": "ビズアップ",
         "name": "類似団体比較（経営比較分析表・団体類型別平均・経営指標）",
         B: [((8, 9, "上"), (8, 9, "下"), "main")],
         "memo": "水道：総務省・日本水道協会「水道事業経営指標」／下水道：類似団体Cd2"},

        {"kind": "group", "name": "Ⅱ　課題の抽出"},
        {"kind": "task", "no": 5, "cat": "課題抽出", "who": "ビズアップ",
         "name": "課題の抽出・構造化（収益／費用／投資／財源／組織の５視点）",
         B: [((8, 9, "下"), (8, 10, "中"), "main")],
         "memo": "★成果物：課題一覧（原因・影響度・対応難易度で整理）"},
        {"kind": "task", "no": 6, "cat": "課題抽出", "who": "両者",
         "name": "中間打合せ①　課題の共有と優先順位づけ",
         B: [((8, 10, "中"), (8, 10, "中"), "town")],
         "ms": [((8, 10, "中"), "打合せ①")],
         "memo": "町のご意向を確認のうえ改善案の検討範囲を確定"},

        {"kind": "group", "name": "Ⅲ　改善案の作成"},
        {"kind": "task", "no": 7, "cat": "改善案", "who": "ビズアップ",
         "name": "改善案①　費用面（委託の包括化・動力費・修繕費・施設統廃合）",
         B: [((8, 10, "下"), (8, 11, "下"), "main")],
         "memo": "水道：針浜／御前浄水場、尾浦送水槽・総合配水池の統廃合検討を含む", "h": 30},
        {"kind": "task", "no": 8, "cat": "改善案", "who": "ビズアップ",
         "name": "改善案②　収益面（未収金対策・資産活用・接続率／水洗化率向上）",
         B: [((8, 11, "上"), (8, 12, "上"), "main")],
         "memo": "水道R7未収額104,000千円（うち補助金等83,973千円）の回収管理を含む", "h": 30},
        {"kind": "task", "no": 9, "cat": "改善案", "who": "両者",
         "name": "改善案③　繰入金の適正化（基準内／基準外の整理・一般会計との協議材料）",
         B: [((8, 11, "上"), (8, 12, "中"), "main")],
         "memo": "水道R7の他会計補助金166,987千円は大半が繰出基準外。基準内算定の余地を精査", "h": 34},

        {"kind": "group", "name": "Ⅳ　料金・使用料改定の提案"},
        {"kind": "task", "no": 10, "cat": "料金改定", "who": "ビズアップ",
         "name": "総括原価計算（水道・公共下水道・浄化槽）",
         B: [((8, 11, "下"), (8, 12, "下"), "main")],
         "memo": "総務省様式に基づく原価計算。資産維持費の計上要否も併せて検討"},
        {"kind": "task", "no": 11, "cat": "料金改定", "who": "ビズアップ",
         "name": "改定パターンの試算（改定率×実施時期×料金体系の組合せ）",
         B: [((8, 12, "中"), (9, 1, "中"), "main")],
         "memo": "水道：基本料金中心か口径別へ移行するか／メーター使用料の改定を含む", "h": 30},
        {"kind": "task", "no": 12, "cat": "料金改定", "who": "両者",
         "name": "中間打合せ②　改定パターンの絞り込み",
         B: [((9, 1, "中"), (9, 1, "中"), "town")],
         "ms": [((9, 1, "中"), "打合せ②")],
         "memo": "施行目標時期の決定が必要（要確認事項）"},
        {"kind": "task", "no": 13, "cat": "料金改定", "who": "ビズアップ",
         "name": "料金改定提案書の作成（影響額試算・使用者負担のモデルケース比較）",
         B: [((9, 1, "下"), (9, 2, "中"), "main")],
         "memo": "★成果物：料金改定提案書。一般家庭・水産加工・大口需要者別の影響額を提示", "h": 34},

        {"kind": "group", "name": "Ⅴ　とりまとめ"},
        {"kind": "task", "no": 14, "cat": "報告", "who": "ビズアップ",
         "name": "報告書（課題と改善案）のとりまとめ",
         B: [((9, 2, "中"), (9, 3, "上"), "main")],
         "memo": "★成果物：財務・経営指標分析報告書"},
        {"kind": "task", "no": 15, "cat": "報告", "who": "両者",
         "name": "最終報告会・成果物納品",
         B: [((9, 3, "中"), (9, 3, "下"), "main")],
         "ms": [((9, 3, "中"), "最終報告")],
         "memo": ""},
        {"kind": "task", "no": 16, "cat": "後続", "who": "町",
         "name": "（後続）審議会への諮問・条例改正・住民説明",
         B: [((9, 4, "上"), (9, 4, "下"), "opt")],
         "memo": "改定を実施する場合の後続工程。施行時期から逆算して別途工程を作成", "h": 30},
    ]

    notes = [
        "① 本工程は「１ 経営戦略の更新・再評価」と同一の決算数値・試算シートを共用します。指標算定（No.3）までは両業務の共通作業です。",
        "② 令和７年度の主要指標（決算書記載値）　上水道：経常収支比率97.10%／料金回収率35.07%（R6:47.11%）／有収率73.64%（前年比△7.62pt）／給水原価658.74円・供給単価116.61円。",
        "③ 公共下水道：経常収支比率69.96%／経費回収率76.38%（R6:83.70%）。浄化槽：経常収支比率99.06%／経費回収率64.40%（R6:74.48%）。３事業とも回収率が前年から低下しています。",
        "④ 料金改定の「施行目標時期」が決まると、審議会・条例改正・システム改修・住民周知から逆算した確定工程を作成できます。現時点では未定のため、本表は分析・提案までを対象としています。",
        "⑤ 下水道経営戦略（案）のロードマップは令和13年度に使用料改定を検討し令和14年度に改定する想定です。今回の分析結果によっては前倒しの要否を検証する必要があります。",
    ]

    draw_gantt(ws,
               "２　財務・経営指標分析による課題抽出と改善案作成（料金改定提案を含む）　日程表",
               "対象：上水道事業／公共下水道事業／浄化槽事業　／　対象期間：令和８年８月〜令和９年４月",
               MONTHS_R8 + [m for m in MONTHS_ALL if m[1] == 9 and m[2] == 4],
               rows, notes)
    return ws


# ============================================================
# シート04　令和７年度決算サマリと差異
# ============================================================
def sheet_summary(wb):
    ws = wb.create_sheet("04_R7決算サマリと差異")
    ws.sheet_view.showGridLines = False

    def title(r, text, width=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        c = ws.cell(row=r, column=1, value=text)
        c.font = font(13, True, C["white"])
        c.fill = fill(C["header"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 26

    def subtitle(r, text, width=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        c = ws.cell(row=r, column=1, value=text)
        c.font = font(10.5, True, C["white"])
        c.fill = fill(C["subhead"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    def table(r, headers, data, widths=None, aligns=None, flags=None):
        for i, h in enumerate(headers):
            c = ws.cell(row=r, column=i + 1, value=h)
            c.font = font(9, True, C["white"])
            c.fill = fill(C["subhead"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1
        for j, row in enumerate(data):
            base = C["alt"] if j % 2 else C["white"]
            if flags and flags[j]:
                base = {"bad": C["bad"], "warn": C["warn"], "good": C["good"]}[flags[j]]
            for i, v in enumerate(row):
                c = ws.cell(row=r, column=i + 1, value=v)
                c.font = font(9)
                c.fill = fill(base)
                al = (aligns or ["left"] * len(headers))[i]
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        wrap_text=(al == "left"), indent=(1 if al == "left" else 0))
                c.border = BORDER
            ws.row_dimensions[r].height = 26
            r += 1
        return r

    r = 1
    title(r, "令和７年度決算　主要数値サマリと経営戦略（案）との差異"); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1,
                value="出典：令和７年度 上水道事業会計決算書／下水道事業会計決算書、令和７年度水道分決算審査意見書（案・初稿）、令和７年度末残高月報")
    c.font = font(8.5, color=C["gray"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 2

    # --- 上水道 -----------------------------------------------------
    subtitle(r, "■　上水道事業　令和７年度決算（単位：千円）"); r += 1
    r = table(r,
              ["区　分", "令和７年度", "令和６年度", "増　減", "コメント"],
              [
                  ["給水収益", "119,481", "118,769", "+712", "有収水量は+5,255㎥。営業用・工業用の増による"],
                  ["営業外収益", "535,817", "397,779", "+138,038", "うち他会計補助金166,987千円（大半が繰出基準外）"],
                  ["経常費用", "674,929", "582,892", "+92,037", "江島海底送水管破損に伴う応急給水・復旧費が主因"],
                  ["経常損失", "△19,590", "－", "－", ""],
                  ["特別利益", "774,867", "－", "－", "過年度の会計処理誤りの一括修正（一過性）"],
                  ["特別損失", "2,333,576", "－", "－", "同上。うち過年度損益修正損2,333,473千円"],
                  ["当年度純損失", "1,578,299", "－", "－", "★一過性要因が大半。ただし欠損金として累積"],
                  ["当年度未処理欠損金", "2,732,934", "－", "－", "★累積欠損金が27億円規模に"],
                  ["資金残高（年度末）", "50,296", "164,923", "△114,627", "★１年で約1.1億円減。戦略案のR17見込（約4千万円）に既に接近"],
                  ["建設改良費", "149,117", "－", "－", "女川出島線送水管布設、旭が丘地区老朽管布設替ほか"],
                  ["企業債借入", "122,900", "－", "－", ""],
                  ["企業債元金償還金", "18,171", "－", "－", ""],
              ],
              aligns=["left", "right", "right", "right", "left"],
              flags=[None, None, None, None, None, None, "bad", "bad", "bad", None, None, None])
    r += 1

    subtitle(r, "■　上水道事業　経営指標の推移（決算書掲載値）"); r += 1
    r = table(r,
              ["指　標", "令和５年", "令和６年", "令和７年", "評　価"],
              [
                  ["経常収支比率（％）", "89.45", "88.12", "97.10", "改善だが基準外補助金による嵩上げ。実質は未改善"],
                  ["料金回収率（％）", "52.87", "47.11", "35.07", "★急落。100%が目安"],
                  ["有収率（％）", "－", "81.26", "73.64", "★捨水（水質基準超過対応）と応急給水で悪化"],
                  ["管路経年化率（％）", "15.60", "23.74", "23.16", "全国水準51.38%は下回る"],
                  ["管路更新率（％）", "0.00", "1.45", "0.75", ""],
                  ["有形固定資産減価償却率（％）", "17.21", "19.43", "22.07", ""],
                  ["給水原価（円）", "－", "－", "658.74", "★経営指標231.15円の約2.9倍"],
                  ["供給単価（円）", "－", "－", "116.61", "★経営指標160.62円を44.01円下回る"],
              ],
              aligns=["left", "right", "right", "right", "left"],
              flags=[None, "bad", "bad", None, None, None, "bad", "bad"])
    r += 1

    # --- 下水道 -----------------------------------------------------
    subtitle(r, "■　下水道事業　令和７年度決算　セグメント別（単位：千円）"); r += 1
    r = table(r,
              ["区　分", "公共下水道事業", "浄化槽事業", "合　計", "コメント"],
              [
                  ["営業収益", "187,824", "8,804", "196,628", "R6合計159,198千円から+37,430千円"],
                  ["うち使用料", "118,450", "8,804", "127,254", "★R6は130,349千円。使用料は△3,095千円の減"],
                  ["営業費用", "638,636", "22,451", "661,087", ""],
                  ["営業損益", "△450,812", "△13,647", "△464,459", ""],
                  ["経常損益", "△143,497", "△211", "△143,708", ""],
                  ["セグメント資産", "17,084,429", "156,130", "17,240,559", ""],
                  ["セグメント負債", "15,309,313", "147,411", "15,456,724", ""],
                  ["減価償却費", "427,303", "9,200", "436,503", "公共下水道は費用の56.19%を占める"],
                  ["他会計繰入金", "235,383", "5,928", "241,311", ""],
                  ["当年度純損失", "－", "－", "228,419", "特別損失84,982千円（過年度修正）を含む"],
                  ["資金残高（年度末）", "－", "－", "49,983", "R6末76,235千円から△26,252千円"],
              ],
              aligns=["left", "right", "right", "right", "left"],
              flags=[None, "warn", None, None, None, None, None, None, None, None, "warn"])
    r += 1

    subtitle(r, "■　下水道事業　経営指標の推移（決算書掲載値）"); r += 1
    r = table(r,
              ["指　標", "令和５年", "令和６年", "令和７年", "評　価"],
              [
                  ["【公共】経常収支比率（％）", "78.69", "77.97", "69.96", "★低下。減価償却費が費用の過半"],
                  ["【公共】経費回収率（％）", "88.37", "83.70", "76.38", "★低下傾向が継続"],
                  ["【浄化槽】経常収支比率（％）", "97.36", "97.14", "99.06", "ほぼ均衡"],
                  ["【浄化槽】経費回収率（％）", "71.66", "74.48", "64.40", "★急落。使用料水準の検証が必要"],
                  ["【浄化槽】有形固定資産減価償却率（％）", "9.72", "14.27", "19.17", ""],
                  ["有収率（％）", "－", "－", "109.4", "測定方法の差異により100%超"],
                  ["処理原価（円）", "－", "－", "1,219.85", "経営指標263.79円"],
                  ["使用料単価（円）", "－", "－", "189.99", "経営指標165.57円を上回る"],
              ],
              aligns=["left", "right", "right", "right", "left"],
              flags=["bad", "bad", None, "bad", None, None, None, "good"])
    r += 1

    # --- 差異 -------------------------------------------------------
    subtitle(r, "■　経営戦略（案）の記載と令和７年度決算との主な差異（要修正箇所）"); r += 1
    r = table(r,
              ["対象", "経営戦略（案）の記載", "令和７年度決算の実績", "対応方針"],
              [
                  ["水道 3-2 給水の状況",
                   "給水人口5,769人（R6末）／有収水量1,019.32千㎥／有収率は平均87%で推移",
                   "給水人口5,656人／有収水量1,024.58千㎥／有収率73.64%",
                   "R7末時点の数値へ差替。有収率の低下要因（捨水・応急給水）を本文に追記"],
                  ["水道 3-8 経営分析",
                   "令和５年度決算の経営比較分析表に基づく分析（料金回収率52.87%）",
                   "料金回収率35.07%（R7）／経常収支比率97.10%",
                   "R7実績まで延伸。基準外繰入による経常収支比率の嵩上げを注記"],
                  ["水道 4-2 投資・財政計画",
                   "「令和７年度決算見込みの数値に変動予定あり！再計算の上、入替すること！」と注記",
                   "決算確定済",
                   "★試算シート再計算のうえ資料12を全面差替"],
                  ["水道 4-3 計画期間内の経営状況",
                   "資金残高は令和17年度に約4千万円",
                   "令和７年度末で既に50,296千円",
                   "★資金計画の前提を根本から見直し。対策の前倒しが不可避"],
                  ["水道 投資計画",
                   "計画期間中の建設改良費 約45億円",
                   "鷲神浄水場高度処理施設・江島海底送水管本復旧が別途必要",
                   "★未計上事業の事業費・年割・財源を確認のうえ追加"],
                  ["水道 3-10 組織の見通し",
                   "水道事業会計所属職員は３名を基本",
                   "決算審査で内部けん制の不備を指摘（改善済）",
                   "人材育成・内部統制の記載を補強"],
                  ["下水道 Ⅱ-2 経営状況分析",
                   "「使用料収入は過去３か年で増加傾向」",
                   "使用料は R6:130,349千円 → R7:127,254千円（△3,095千円）",
                   "★増加傾向の前提が反転。有収水量も△16,246㎥"],
                  ["下水道 Ⅳ-2 シミュレーション",
                   "パターン❷（令和14年度に25%引上げ）で目標達成",
                   "経費回収率は R6:83.70% → R7:76.38% と低下",
                   "★改定率・実施時期の再検証が必要"],
                  ["下水道 全体",
                   "対象は「女川町公共下水道事業」のみ",
                   "浄化槽事業は法適用済の報告セグメント",
                   "★浄化槽事業の経営戦略が未策定（シート02で対応）"],
              ],
              aligns=["left", "left", "left", "left"],
              flags=[None, None, "warn", "bad", "bad", None, "bad", "warn", "warn"])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 46
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    return ws


# ============================================================
# シート05　会計システムの見直し
# ============================================================
def sheet_system(wb):
    ws = wb.create_sheet("05_会計システム見直し")
    ws.sheet_view.showGridLines = False

    def title(r, text, width=6):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        c = ws.cell(row=r, column=1, value=text)
        c.font = font(13, True, C["white"])
        c.fill = fill(C["header"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 26

    def subtitle(r, text, width=6):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
        c = ws.cell(row=r, column=1, value=text)
        c.font = font(10.5, True, C["white"])
        c.fill = fill(C["subhead"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    def table(r, headers, data, aligns=None, flags=None, h=30):
        for i, hh in enumerate(headers):
            c = ws.cell(row=r, column=i + 1, value=hh)
            c.font = font(9, True, C["white"])
            c.fill = fill(C["subhead"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1
        for j, row in enumerate(data):
            base = C["alt"] if j % 2 else C["white"]
            if flags and flags[j]:
                base = {"bad": C["bad"], "warn": C["warn"], "good": C["good"]}[flags[j]]
            for i, v in enumerate(row):
                c = ws.cell(row=r, column=i + 1, value=v)
                c.font = font(9)
                c.fill = fill(base)
                al = (aligns or ["left"] * len(headers))[i]
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        wrap_text=True, indent=(1 if al == "left" else 0))
                c.border = BORDER
            ws.row_dimensions[r].height = h
            r += 1
        return r

    r = 1
    title(r, "３　会計システムの見直し　－　課題整理とビズアップの提供範囲・考え方"); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1,
                value="令和７年度決算審査で顕在化した事象を出発点に、システム要件と予算獲得の論拠を整理したものです。")
    c.font = font(8.5, color=C["gray"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 2

    subtitle(r, "■　Ⅰ　現状の課題　－　令和７年度決算審査で顕在化した事象"); r += 1
    r = table(r,
              ["No.", "顕在化した事象", "決算への影響", "根本原因（システム面）"],
              [
                  ["1", "平成23年度会計基準見直し（平成26年度決算適用）が未対応。資本剰余金5,573,284,917円を繰延収益（長期前受金）へ振り替える必要があった",
                   "資本剰余金 5,812,974千円 → 239,689千円\n長期前受金 5,477,657千円 → 11,050,942千円",
                   "会計基準改正時の移行処理がシステム側で担保されず、10年以上是正されないまま推移した"],
                  ["2", "固定資産の減価償却および長期前受金収益化の処理誤り（令和３年度）。決算書の固定資産明細書と貸借対照表が不一致",
                   "構築物減価償却累計額 △1,564,522千円 → △3,031,943千円 ほか計４科目を修正",
                   "★固定資産台帳と財務会計（仕訳・BS）が連携しておらず、整合チェック機能が働いていない"],
                  ["3", "量水器の個数管理不十分。除却時の台帳未修正、現地未撤去、減価償却方法の誤り",
                   "固定資産（量水器）14,172千円 → 10,543千円\n貯蔵品 5,303千円 → 1,680千円",
                   "★貯蔵品（量水器）の受払と固定資産計上が別管理。取得〜設置〜除却の一連が追跡できない"],
                  ["4", "その他流動資産169,933千円の内容不明（H29〜R2の仮払消費税取崩し・還付金の消込漏れ）",
                   "全額を特別損失として処理",
                   "★消費税申告と会計処理が連動せず、未収金の消込が手作業"],
                  ["5", "【下水道】無形固定資産（施設利用権）の減価償却漏れ（令和４〜５年度）",
                   "施設利用権 1,187,112千円 → 1,018,242千円",
                   "法適用移行時に取り込んだ資産の償却設定が未整備"],
                  ["6", "伝票処理で担当者欄と責任者欄が同一者となる事例が多数。内部けん制が機能していなかった",
                   "誤りや処理漏れの発生要因と指摘（指摘後は複数人確認により改善）",
                   "★システム上の承認ワークフロー（起票者と承認者の分離）が未設定"],
              ],
              aligns=["center", "left", "left", "left"],
              flags=[None, "bad", "bad", "bad", None, "bad"], h=58)
    r += 1

    subtitle(r, "■　Ⅱ　見直しの考え方　－　「システムを入れ替える」ではなく「決算を作れる仕組みにする」"); r += 1
    r = table(r,
              ["視　点", "考え方", "具体的な打ち手"],
              [
                  ["① 台帳と会計の一元化",
                   "計数不一致の根本原因は、固定資産台帳・貯蔵品・企業債・長期前受金がそれぞれ別管理となっていること。単一のマスタから財務諸表が生成される構造にする",
                   "固定資産台帳と財務会計の連携／長期前受金の自動収益化／企業債の元利償還スケジュール自動連携"],
                  ["② セグメント会計への対応",
                   "下水道事業会計は公共下水道・浄化槽の２セグメント。浄化槽事業の経営戦略策定を見据え、セグメント別の財務情報が決算時に自動集計される必要がある",
                   "勘定科目にセグメント区分を保持／セグメント別PL・BS・キャッシュフローの自動作成"],
                  ["③ 内部統制の作り込み",
                   "「複数人で確認する」という運用ルールだけでは再発する。システムで起票者と承認者を分離し、承認履歴を残す",
                   "承認ワークフロー設定／権限マスタの整備／修正仕訳の履歴保全"],
                  ["④ 経営分析への接続",
                   "決算値の入力が終わった時点で経営指標・経営比較分析表・経営戦略の試算シートに流し込める形にする。毎年の更新作業を定型化する",
                   "決算統計様式・経営比較分析表の出力／経営戦略試算シートへのデータ連携"],
                  ["⑤ 移行リスクの最小化",
                   "過年度の修正が令和７年度に集中したことを踏まえ、移行時に過年度データの検証を組み込む",
                   "移行前データの検証・クレンジング／並行稼働期間の設定"],
              ],
              aligns=["left", "left", "left"], h=64)
    r += 1

    subtitle(r, "■　Ⅲ　ビズアップができること（提供範囲の候補）"); r += 1
    r = table(r,
              ["区　分", "内　容", "位置づけ", "備　考"],
              [
                  ["Ａ 現状診断", "現行会計システムの機能・運用の棚卸、決算業務フローの可視化、課題の特定",
                   "予算要求の根拠づくり", "令和８年度中に実施可能。要求資料の中核となる"],
                  ["Ｂ 要件定義支援", "求める機能の整理、要求水準書・仕様書案の作成、概算費用の算定",
                   "予算要求の根拠づくり", "Ａとセットで実施することで要求額の説明力が高まる"],
                  ["Ｃ 固定資産台帳の再整備", "資産の実在性確認、取得価額・耐用年数・償却方法の検証、台帳の再構築",
                   "システム導入の前提条件", "★今回の計数不一致の再発防止に直結。単独でも効果が大きい"],
                  ["Ｄ 調達支援", "事業者選定の支援（プロポーザル要領・評価基準の作成、審査補助）",
                   "導入年度", ""],
                  ["Ｅ 導入・移行支援", "移行データの検証、並行稼働時の突合、初年度決算の伴走",
                   "導入年度", ""],
                  ["Ｆ 決算・経営分析支援", "決算統計・経営比較分析表の作成支援、経営指標の算定、経営戦略との連動",
                   "導入後の継続支援", "１・２の業務と一体で提供可能"],
                  ["Ｇ 職員研修", "公営企業会計の基礎、決算実務、内部統制の考え方",
                   "全期間", "決算審査で人材育成の改善が求められている"],
              ],
              aligns=["center", "left", "center", "left"], h=40)
    r += 1

    subtitle(r, "■　Ⅳ　予算獲得に向けた考え方"); r += 1
    r = table(r,
              ["論　点", "整理の方向"],
              [
                  ["① 要求の必要性をどう説明するか",
                   "「システムが古いから」ではなく「決算審査で指摘を受けた計数不一致の再発防止」を前面に出す。監査委員の決算審査意見書という第三者文書が根拠として使える点が強い"],
                  ["② いつ要求するか",
                   "令和９年度当初予算での要求を基本線とする。要求資料の作成には現状診断（Ａ）が前提となるため、令和８年８〜10月に着手する必要がある"],
                  ["③ どの会計で持つか",
                   "上水道事業会計・下水道事業会計の双方に関わるため、費用按分の考え方を先に整理する。一般会計繰入の対象となるかも併せて確認"],
                  ["④ 財源をどう確保するか",
                   "公営企業の DX・デジタル化に係る国の支援措置、地方財政措置（公営企業会計の適用・経営基盤強化に係るもの）の活用可能性を調査する"],
                  ["⑤ 段階的に進める選択肢",
                   "システム更改を一度に行わず、令和８年度＝現状診断・固定資産台帳再整備、令和９年度＝要件定義・調達、令和10年度＝導入 と段階化すれば、単年度の要求額を抑えられる"],
                  ["⑥ 上下水道課の体制との関係",
                   "水道事業会計所属職員は３名。システム導入の負荷を職員が吸収できないため、外部支援の必要性そのものが要求理由になる"],
              ],
              aligns=["left", "left"], h=46)
    r += 1

    subtitle(r, "■　Ⅴ　当面のアクション（令和８年８月〜10月）"); r += 1
    r = table(r,
              ["時　期", "アクション", "主担当"],
              [
                  ["令和８年８月　中旬〜下旬", "現行会計システムの契約内容・機能・保守条件の確認（資料提供）", "町"],
                  ["令和８年８月　下旬〜９月上旬", "決算業務フローのヒアリング（誰が・何を・どの帳票で）", "両者"],
                  ["令和８年９月　中旬〜下旬", "課題の整理と見直し方針案の提示", "ビズアップ"],
                  ["令和８年10月　上旬〜中旬", "業務範囲・概算費用の提示、予算要求資料案の作成", "ビズアップ"],
                  ["令和８年10月　下旬", "令和９年度当初予算要求", "町"],
              ],
              aligns=["center", "left", "center"], h=26)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    return ws


# ============================================================
# シート00　表紙・全体像
# ============================================================
def sheet_cover(wb):
    ws = wb.create_sheet("00_全体像", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "女川町　上下水道事業　支援業務　全体工程"
    c.font = font(18, True, C["white"])
    c.fill = fill(C["header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "作成日：令和８年８月　／　対象：上水道事業・公共下水道事業・浄化槽事業"
    c.font = font(10, color=C["gray"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    rows = [
        ("１", "経営戦略の更新・再評価",
         "令和７年度決算を反映した経営戦略（案）の数値更新と再評価。水道・下水道の２事業。",
         "令和８年８月〜令和９年３月", "01_経営戦略の更新"),
        ("１(2)", "浄化槽事業　経営戦略の策定",
         "下水道事業会計の報告セグメントである浄化槽事業について、公営企業版の経営戦略を新規策定。",
         "令和８年８月〜令和10年３月", "02_浄化槽_経営戦略策定"),
        ("２", "財務・経営指標分析による課題抽出と改善案作成",
         "経営指標の算定・類似団体比較から課題を抽出し、改善案と料金改定提案を作成。",
         "令和８年８月〜令和９年４月", "03_財務経営指標分析"),
        ("－", "令和７年度決算サマリと差異",
         "決算の主要数値と、経営戦略（案）の記載との差異一覧。上記１・２の共通インプット。",
         "－", "04_R7決算サマリと差異"),
        ("３", "会計システムの見直し",
         "決算審査で顕在化した課題の整理、ビズアップの提供範囲、予算獲得に向けた考え方。",
         "令和８年８月〜（令和９年度当初予算要求）", "05_会計システム見直し"),
    ]

    r = 4
    headers = ["項番", "業務", "内　容", "期　間", "シート"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = font(10, True, C["white"])
        c.fill = fill(C["subhead"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[r].height = 24
    r += 1
    for j, row in enumerate(rows):
        base = C["alt"] if j % 2 else C["white"]
        for i, v in enumerate(row):
            c = ws.cell(row=r, column=i + 1, value=v)
            c.font = font(10, bold=(i == 1))
            c.fill = fill(base)
            c.alignment = Alignment(horizontal="center" if i in (0, 3, 4) else "left",
                                    vertical="center", wrap_text=True, indent=(1 if i == 2 else 0))
            c.border = BORDER
        ws.row_dimensions[r].height = 44
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="■　直近で決めていただきたい事項")
    c.font = font(11, True, C["white"])
    c.fill = fill(C["subhead"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1

    decisions = [
        ("A", "料金改定の施行目標時期（水道）",
         "審議会・条例改正・システム改修・住民周知から逆算するため。決まればシート03の後続工程を確定できます。", "令和８年10月まで"),
        ("B", "浄化槽事業の経営戦略：単独策定か、下水道経営戦略への統合か",
         "統合方式なら期間・費用とも圧縮できます。予算要求の内容が変わるため先に決定が必要です。", "令和８年10月上旬"),
        ("C", "浄化槽事業の実施年度（パターンＡ：令和８年度／パターンＢ：令和９年度）",
         "令和８年度内に行う場合は補正予算等の財源措置が必要です。", "令和８年９月中"),
        ("D", "令和９年度当初予算の要求締切",
         "会計システム・浄化槽経営戦略とも、この日程から逆算します。", "至急"),
        ("E", "経営戦略（案）に未計上の投資事業の事業費・年割",
         "鷲神浄水場高度処理施設、江島海底送水管本復旧。収支計画に直結します。", "令和８年10月まで"),
    ]
    headers = ["", "決定事項", "理　由", "期　限"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = font(9.5, True, C["white"])
        c.fill = fill(C["header"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1
    for j, row in enumerate(decisions):
        for i, v in enumerate(row):
            c = ws.cell(row=r, column=i + 1, value=v)
            c.font = font(9.5, bold=(i == 1))
            c.fill = fill(C["warn"])
            c.alignment = Alignment(horizontal="center" if i in (0, 3) else "left",
                                    vertical="center", wrap_text=True, indent=(1 if i in (1, 2) else 0))
            c.border = BORDER
        ws.row_dimensions[r].height = 36
        r += 1

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 24
    return ws


# ============================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_strategy_update(wb)
    sheet_johkasou(wb)
    sheet_analysis(wb)
    sheet_summary(wb)
    sheet_system(wb)
    sheet_cover(wb)

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = C["subhead"]
    wb.worksheets[0].sheet_properties.tabColor = C["header"]

    path = os.path.join(OUT_DIR, "女川町_上下水道事業_支援業務工程表.xlsx")
    wb.save(path)
    print("saved:", path)


if __name__ == "__main__":
    main()
