# -*- coding: utf-8 -*-
"""
女川町 上水道事業　料金の妥当性検証と改定試算
─────────────────────────────────────────────
入力
  ・経営戦略試算シート「有収水量別」（顧客別の月次調定データ 3,060件）
  ・令和７年度 上水道事業会計決算書の数値（本ファイル内に定数として保持）
  ・総務省 水道事業経営指標（令和６年度）の団体累計別平均

出力
  onagawa/output/女川町_水道料金_妥当性検証と改定試算.xlsx

留意
  ・入力の顧客別データは個人が特定されうるため、出力には一切含めない。
    集計（用途別・口径別・水量帯別）のみを出力する。
  ・入力ファイルのパスは環境変数 ONAGAWA_SIM_XLSX で指定する。
"""

import os
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "output")
os.makedirs(OUT_DIR, exist_ok=True)

SRC = os.environ.get(
    "ONAGAWA_SIM_XLSX",
    "/tmp/claude-0/-home-user-repository/827b157d-89c7-5513-9a28-6465ad6d529d/scratchpad/"
    "uploads/gmail11/女川町_水道 経営戦略試算シート20251209版.xlsx",
)

# ============================================================
# 令和７年度決算の実績値（円）
# ============================================================
R7 = {
    "給水収益": 119_480_750,
    "営業収益": 119_522_350,
    "経常収益": 655_339_107,
    "経常費用": 674_928_751,
    "長期前受金戻入益": 334_252_462,
    "有収水量": 1_024_576,          # ㎥
    "年間総配水量": 1_391_383,       # ㎥
    # 江島海底送水管応急修繕等（繰出基準外の他会計補助金が全額充当された修繕費等）
    "一過性費用_江島": 128_789_977,
    # 繰出基準外の他会計補助金
    "基準外補助金": 166_434_074,
}

# 経常費用の推移（円）　※R7は経常費用（特別損失を除く）
KEIJO_HIYO = {"R3": 543_887_736, "R4": 694_576_672, "R5": 564_428_515,
              "R6": 582_892_184, "R7": 674_928_751}

# 総務省 水道事業経営指標（令和６年度）給水人口5千人以上1万人未満
BENCH = {
    "水源区分総合計・全平均（190事業）": {"給水原価": 240.31, "供給単価": 195.74,
                                "料金回収率": 81.45, "20㎥料金": 3822.92,
                                "10㎥料金": 1887.42, "有収率": 75.37},
    "表流水・有収水量密度 全国平均以上（3事業）": {"給水原価": 231.15, "供給単価": 160.62,
                                    "料金回収率": 69.49, "20㎥料金": 3127.67,
                                    "10㎥料金": 1624.33, "有収率": 81.47},
    "表流水・有収水量密度 全国平均未満（46事業）": {"給水原価": 265.44, "供給単価": 215.16,
                                     "料金回収率": 81.06, "20㎥料金": 4207.80,
                                     "10㎥料金": 2120.11, "有収率": 77.37},
    "全国計・全平均（1,230事業）": {"給水原価": 181.66, "供給単価": 177.28,
                          "料金回収率": 97.59, "20㎥料金": 3373.90,
                          "10㎥料金": 1629.95, "有収率": 89.21},
}

# 経営比較分析表（令和５年度決算）の類似団体平均（区分A8）
BENCH_HIKAKU = {"給水原価": 230.21, "料金回収率": 84.16, "有収率": 76.64,
                "20㎥料金": None, "全国平均_給水原価": 177.56, "全国平均_料金回収率": 97.82}

# 現行料金（税抜）
CUR = {
    "基本_5㎥まで": 900,
    "基本_6-10㎥": 1100,
    "従量_10超100以下": 110,
    "従量_100超": 100,
    "湯屋_500㎥まで": 33000,
    "湯屋_超過": 100,
    "船舶委託_1㎥": 110,
    "メーター": {13: 50, 20: 70, 25: 100, 40: 200, 50: 800, 75: 1000, 100: 1200, 150: 3000},
}

TAX = 1.10


# ============================================================
# データ読込
# ============================================================
def load_records():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["有収水量別"]
    recs = []
    for r in range(3, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if v[0] is None:
            continue
        use = v[2] or "不明"
        dia = v[3]
        try:
            dia = int(dia)
        except (TypeError, ValueError):
            dia = None
        recs.append({
            "用途": use,
            "口径": dia,
            "水量": float(v[5] or 0),
            "基本": float(v[6] or 0),
            "超過": float(v[7] or 0),
            "メーター": float(v[9] or 0),
        })
    return recs


# ============================================================
# 料金計算（体系変更シナリオ用）
# ============================================================
def calc_charge(use, vol, dia, p):
    """税抜の 基本／従量／メーター を返す"""
    if use == "湯屋用" and vol > 0:
        base = p["湯屋_500㎥まで"]
        over = max(0.0, vol - 500) * p["湯屋_超過"]
    elif use == "委託船舶給水用":
        base = 0.0
        over = vol * p["船舶委託_1㎥"]
    else:
        free = p.get("基本水量", 10)
        base = p["基本_5㎥まで"] if vol <= 5 else p["基本_6-10㎥"]
        o1 = max(0.0, min(vol, 100) - free) * p["従量_10超100以下"]
        o2 = max(0.0, vol - 100) * p["従量_100超"]
        over = o1 + o2
    meter = p["メーター"].get(dia, 0) if dia else 0
    return base, over, meter


# ============================================================
# シナリオ（率改定：実績調定額に係数を乗じる）
# ============================================================
def rate_scenarios(recs):
    cur = {"基本": sum(r["基本"] for r in recs),
           "超過": sum(r["超過"] for r in recs),
           "メーター": sum(r["メーター"] for r in recs)}
    cur_total = sum(cur.values())

    def mult_for_uplift(target_rate, on):
        """指定の料金項目のみで全体をtarget_rate倍にするための倍率"""
        need = cur_total * target_rate - cur_total
        return 1 + need / cur[on]

    scen = [
        # (ID, 名称, 基本倍率, 従量倍率, メーター倍率, 分類)
        ("S0", "現行料金", 1.00, 1.00, 1.00, "基準"),
        ("S1", "経営戦略（案）　基本料金×1.5・従量据置", 1.50, 1.00, 1.00, "既存案"),
        ("S3", "一律 ＋30％", 1.30, 1.30, 1.30, "率改定"),
        ("S4", "一律 ＋50％", 1.50, 1.50, 1.50, "率改定"),
        ("S5", "一律 ＋77.3％（料金回収率100％・平常時ベース）", 1.7734, 1.7734, 1.7734, "目標達成"),
        ("S6", "一律 ＋185.1％（料金回収率100％・R7実績ベース）", 2.8513, 2.8513, 2.8513, "目標達成"),
        ("S7", "段階改定　第１段階 ＋35％", 1.35, 1.35, 1.35, "段階案"),
        ("S8", "段階改定　第２段階（累積 ＋82.3％）", 1.8225, 1.8225, 1.8225, "段階案"),
        ("S9", "＋30％を基本料金のみで実現", mult_for_uplift(1.30, "基本"), 1.00, 1.00, "配分比較"),
        ("S10", "＋30％を従量料金のみで実現", 1.00, mult_for_uplift(1.30, "超過"), 1.00, "配分比較"),
    ]
    return cur, cur_total, scen


def apply_rate_scenario(recs, mb, mt, mm):
    out = defaultdict(float)
    for r in recs:
        b = r["基本"] * mb
        o = r["超過"] * mt
        m = r["メーター"] * mm
        out["基本"] += b
        out["超過"] += o
        out["メーター"] += m
        out["計"] += b + o + m
    return out


# ============================================================
# 集計
# ============================================================
def aggregate(recs, mb, mt, mm, key):
    agg = defaultdict(lambda: {"件数": 0, "水量": 0.0, "現行": 0.0, "改定後": 0.0})
    for r in recs:
        k = r[key] if r[key] not in (None, "") else "不明"
        cur = r["基本"] + r["超過"] + r["メーター"]
        new = r["基本"] * mb + r["超過"] * mt + r["メーター"] * mm
        a = agg[k]
        a["件数"] += 1
        a["水量"] += r["水量"]
        a["現行"] += cur
        a["改定後"] += new
    return agg


def volume_band(v):
    for hi, lab in [(5, "　0〜5㎥"), (10, "　6〜10㎥"), (20, "　11〜20㎥"),
                    (30, "　21〜30㎥"), (50, "　31〜50㎥"), (100, "　51〜100㎥"),
                    (500, "101〜500㎥")]:
        if v <= hi:
            return lab
    return "500㎥超"


# ============================================================
# Excel 出力
# ============================================================
C = {"header": "1F3864", "subhead": "2E75B6", "band": "DDEBF7", "alt": "F7FAFC",
     "white": "FFFFFF", "gray": "808080", "bad": "FCE4E4", "warn": "FFF2CC", "good": "E2F0D9"}
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT = "Meiryo UI"


def f(size=10, bold=False, color="000000"):
    return Font(name=FONT, size=size, bold=bold, color=color)


def fl(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


class Sheet:
    def __init__(self, wb, name, ncol):
        self.ws = wb.create_sheet(name)
        self.ws.sheet_view.showGridLines = False
        self.n = ncol
        self.r = 1

    def title(self, text, sub=""):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
        c = ws.cell(row=self.r, column=1, value=text)
        c.font = f(13, True, C["white"]); c.fill = fl(C["header"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[self.r].height = 26
        self.r += 1
        if sub:
            ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
            c = ws.cell(row=self.r, column=1, value=sub)
            c.font = f(8.5, color=C["gray"])
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            self.r += 1
        self.r += 1

    def sub(self, text):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
        c = ws.cell(row=self.r, column=1, value=text)
        c.font = f(10.5, True, C["white"]); c.fill = fl(C["subhead"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[self.r].height = 20
        self.r += 1

    def table(self, heads, data, aligns=None, flags=None, h=26, fmt=None):
        ws = self.ws
        for i, hh in enumerate(heads):
            c = ws.cell(row=self.r, column=i + 1, value=hh)
            c.font = f(9, True, C["white"]); c.fill = fl(C["subhead"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[self.r].height = 28
        self.r += 1
        for j, row in enumerate(data):
            base = C["alt"] if j % 2 else C["white"]
            if flags and j < len(flags) and flags[j]:
                base = {"bad": C["bad"], "warn": C["warn"], "good": C["good"]}[flags[j]]
            for i, v in enumerate(row):
                c = ws.cell(row=self.r, column=i + 1, value=v)
                c.font = f(9, bold=(i == 0 and str(v).startswith("★")))
                c.fill = fl(base)
                al = (aligns or ["left"] * len(heads))[i]
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        wrap_text=(al == "left"), indent=(1 if al == "left" else 0))
                c.border = BORDER
                if fmt and i in fmt and isinstance(v, (int, float)):
                    c.number_format = fmt[i]
            ws.row_dimensions[self.r].height = h
            self.r += 1
        self.r += 1

    def note(self, lines):
        ws = self.ws
        for t in lines:
            ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.n)
            c = ws.cell(row=self.r, column=1, value=t)
            c.font = f(9)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[self.r].height = 16
            self.r += 1
        self.r += 1

    def widths(self, ws_widths):
        for col, w in ws_widths.items():
            self.ws.column_dimensions[col].width = w


# ============================================================
def main():
    recs = load_records()
    cur, cur_total, scen = rate_scenarios(recs)
    n_rec = len(recs)
    mon_vol = sum(r["水量"] for r in recs)

    # 年換算の補正係数（月次データ×12 を R7 実績の給水収益に合わせる）
    annual_factor = R7["給水収益"] / (cur_total * 12)

    # ---- 妥当性検証の基礎数値 -------------------------------
    kesson_a = R7["経常費用"] - R7["長期前受金戻入益"]                       # 340,676,289
    heijo_hiyo = R7["経常費用"] - R7["一過性費用_江島"]                      # 546,138,774
    kesson_b = heijo_hiyo - R7["長期前受金戻入益"]                          # 211,886,312
    heijo_5yr = (KEIJO_HIYO["R3"] + KEIJO_HIYO["R4"] + KEIJO_HIYO["R5"]
                 + KEIJO_HIYO["R6"] + heijo_hiyo) / 5
    kesson_c = heijo_5yr - R7["長期前受金戻入益"]

    ratio_a = R7["給水収益"] / kesson_a
    ratio_b = R7["給水収益"] / kesson_b
    ratio_c = R7["給水収益"] / kesson_c
    kyokyu_tanka = R7["給水収益"] / R7["有収水量"]
    genka_hikaku = kesson_a / R7["有収水量"]
    genka_nissui = R7["経常費用"] / R7["有収水量"]

    wb = Workbook()
    wb.remove(wb.active)

    # ========================================================
    # 01 現行料金体系
    # ========================================================
    s = Sheet(wb, "01_現行料金体系", 6)
    s.title("１　現行の料金体系（令和８年８月時点）",
            "出典：女川町上水道事業経営戦略（第二期）（案）3-3 料金体系、経営戦略試算シート「有収水量別」（顧客別調定データ）")
    s.sub("■　水道料金（一般用）")
    s.table(["区　分", "税抜", "税込", "備　考"],
            [["基本料金　５㎥まで", 900, 990, "使用水量0㎥でも課金"],
             ["基本料金　６〜10㎥", 1100, 1210, "10㎥までの基本水量を含む"],
             ["従量料金　10㎥超100㎥以下（1㎥）", 110, 121, ""],
             ["従量料金　100㎥超（1㎥）", 100, 110, "★逓減制（大口ほど単価が下がる）"]],
            aligns=["left", "right", "right", "left"], fmt={1: "#,##0", 2: "#,##0"})
    s.sub("■　水道料金（特殊用）")
    s.table(["区　分", "税抜", "税込", "備　考"],
            [["湯屋　500㎥まで", 33000, 36300, "対象2件"],
             ["湯屋　超過（1㎥）", 100, 110, ""],
             ["船舶（委託）　1㎥につき", 110, 121, "基本料金・メーター使用料なし。対象1件"],
             ["船舶（直接）　10㎥まで", 2200, 2420, "対象なし"],
             ["船舶（直接）　超過（1㎥）", 220, 242, "対象なし"]],
            aligns=["left", "right", "right", "left"], fmt={1: "#,##0", 2: "#,##0"})
    s.sub("■　メーター使用料（月額）")
    meter_rows = [[f"{d}㎜", v, int(round(v * TAX))] for d, v in sorted(CUR["メーター"].items())]
    s.table(["口　径", "税抜", "税込"], meter_rows,
            aligns=["center", "right", "right"], fmt={1: "#,##0", 2: "#,##0"})
    s.sub("■　需要家の構成（調定データ1か月分・件数）")
    by_use = aggregate(recs, 1, 1, 1, "用途")
    rows = []
    for k, v in sorted(by_use.items(), key=lambda x: -x[1]["件数"]):
        rows.append([k, v["件数"], round(v["件数"] / n_rec * 100, 1), round(v["水量"]),
                     round(v["水量"] / mon_vol * 100, 1), round(v["現行"]),
                     round(v["現行"] / cur_total * 100, 1)])
    rows.append(["合　計", n_rec, 100.0, round(mon_vol), 100.0, round(cur_total), 100.0])
    s.table(["用途区分", "件数", "件数構成比(％)", "水量(㎥)", "水量構成比(％)", "調定額(円・税抜)", "金額構成比(％)"],
            rows, aligns=["left"] + ["right"] * 6,
            fmt={1: "#,##0", 2: "0.0", 3: "#,##0", 4: "0.0", 5: "#,##0", 6: "0.0"})
    by_dia = aggregate(recs, 1, 1, 1, "口径")
    rows = []
    for k, v in sorted(by_dia.items(), key=lambda x: (not isinstance(x[0], int), x[0] if isinstance(x[0], int) else 0)):
        rows.append([f"{k}㎜" if isinstance(k, int) else "不明", v["件数"],
                     round(v["件数"] / n_rec * 100, 1), round(v["水量"]), round(v["現行"])])
    s.table(["口　径", "件数", "件数構成比(％)", "水量(㎥)", "調定額(円・税抜)"], rows,
            aligns=["center"] + ["right"] * 4,
            fmt={1: "#,##0", 2: "0.0", 3: "#,##0", 4: "#,##0"})
    s.note([
        "※ 顧客別の調定データ（3,060件・1か月分）から集計。個人が特定されうるため、明細は本ファイルに収録していません。",
        "※ 料金の内訳（1か月）：基本料金 3,213,650円（34.9％）／従量料金 5,754,420円（62.5％）／メーター使用料 237,670円（2.6％）　いずれも税抜。",
    ])
    s.widths({"A": 34, "B": 16, "C": 16, "D": 16, "E": 16, "F": 18, "G": 16})

    # ========================================================
    # 02 妥当性検証
    # ========================================================
    s = Sheet(wb, "02_料金の妥当性検証", 6)
    s.title("２　現行料金の妥当性検証",
            "令和７年度決算および総務省「水道事業経営指標（令和６年度）」・経営比較分析表（令和５年度決算）による")

    s.sub("■　(1) 料金回収率　－　使用料で回収すべき費用をどれだけ賄えているか")
    s.table(["費用ベース", "経常費用（円）", "長期前受金戻入益（円）", "回収すべき費用（円）", "料金回収率（％）", "100％に必要な改定率"],
            [["Ａ　令和７年度実績（一過性費用込み）", R7["経常費用"], R7["長期前受金戻入益"],
              kesson_a, round(ratio_a * 100, 2), f"＋{(1/ratio_a-1)*100:.1f}％"],
             ["Ｂ　平常時ベース（江島応急給水128,790千円を控除）", round(heijo_hiyo), R7["長期前受金戻入益"],
              round(kesson_b), round(ratio_b * 100, 2), f"＋{(1/ratio_b-1)*100:.1f}％"],
             ["Ｃ　R3〜R7平均（江島分控除後・物価上昇なし）", round(heijo_5yr), R7["長期前受金戻入益"],
              round(kesson_c), round(ratio_c * 100, 2), f"＋{(1/ratio_c-1)*100:.1f}％"]],
            aligns=["left", "right", "right", "right", "right", "right"],
            flags=["bad", "bad", "bad"],
            fmt={1: "#,##0", 2: "#,##0", 3: "#,##0", 4: "0.00"}, h=32)
    s.note([
        "※ 算式は総務省 経営比較分析表に準拠：料金回収率 ＝ 給水収益 ÷（経常費用 − 長期前受金戻入益）×100。",
        "※ ベースＡの35.07％は令和７年度決算書の掲載値と一致します。",
        "※ ベースＢは、繰出基準外の他会計補助金128,789,977円が全額充当された江島海底送水管応急修繕等の費用を控除したものです。",
        "※ いずれのベースでも100％には遠く、現行料金は「使用料で回収すべき費用」を賄えていません。",
    ])

    s.sub("■　(2) 単価水準の他団体比較　－　総務省 水道事業経営指標（令和６年度）／給水人口5千人以上1万人未満")
    rows = [["女川町（令和７年度）", round(genka_nissui, 2), round(kyokyu_tanka, 2),
             round(kyokyu_tanka / genka_nissui * 100, 2), 2475, 1265, 73.64]]
    flags = ["bad"]
    for k, v in BENCH.items():
        rows.append([k, v["給水原価"], v["供給単価"], v["料金回収率"],
                     round(v["20㎥料金"]), round(v["10㎥料金"]), v["有収率"]])
        flags.append(None)
    s.table(["区　分", "給水原価(円)", "供給単価(円)", "料金回収率(％)",
             "20㎥料金(円・税込)", "10㎥料金(円・税込)", "有収率(％)"],
            rows, aligns=["left"] + ["right"] * 6, flags=flags,
            fmt={1: "#,##0.00", 2: "#,##0.00", 3: "0.00", 4: "#,##0", 5: "#,##0", 6: "0.00"}, h=30)
    s.note([
        "※ 女川町の給水原価658.74円は「経常費用÷有収水量」（日本水道協会の経営指標と同じ算式）。",
        f"　 経営比較分析表の算式（長期前受金戻入益を控除）では {genka_hikaku:.2f}円で、類似団体平均（R5）230.21円を上回ります。",
        "※ 20㎥料金2,475円は 基本料金1,100円＋従量10㎥×110円＋メーター使用料（13㎜）50円 の税込額。経営比較分析表の公表値（R5）は2,470円。",
        "※【要確認】決算審査意見書（案）が引用している比較値（給水原価231.15円・供給単価160.62円・有収率81.47％・営業収支比率42.05％）は、",
        "　 「表流水を主とするもの × 有収水量密度が全国平均以上」の区分で、対象がわずか3事業です。",
        "　 女川町の有収水量密度は1.30千㎥/ha（給水人口密度7.4人/ha）と極めて低く、「全国平均未満」（46事業）に該当する可能性があります。",
        "　 その場合の比較値は 給水原価265.44円・供給単価215.16円・20㎥料金4,207.80円 となり、料金水準の乖離はさらに大きくなります。",
    ])

    s.sub("■　(3) 妥当性の判定")
    s.table(["観　点", "判　定", "根　拠"],
            [["使用料で回収すべき費用を賄えているか", "×",
              f"料金回収率 {ratio_a*100:.2f}％（平常時ベースでも {ratio_b*100:.2f}％）。全国平均97.82％（R5）"],
             ["単価水準は他団体と比べて妥当か", "×",
              f"供給単価 {kyokyu_tanka:.2f}円は、給水人口5千人〜1万人の平均195.74円の {kyokyu_tanka/195.74*100:.0f}％"],
             ["家庭料金の水準は妥当か", "×",
              "20㎥ 2,475円は同区分平均3,822.92円の65％、全国計平均3,373.90円の73％"],
             ["独立採算が成り立っているか", "×",
              "繰出基準外の他会計補助金166,434千円で収支を補填。これを除いた経常収支比率は72.44％"],
             ["料金体系は妥当か", "△",
              "①基本水量10㎥は使用実態（中央値14㎥）に対し大きい　②100㎥超の逓減制は大口優遇　"
              "③口径別基本料金がなく固定費の負担配分が不公平　④メーター使用料が別建てで分かりにくい"]],
            aligns=["left", "center", "left"], flags=["bad", "bad", "bad", "bad", "warn"], h=40)

    s.widths({"A": 46, "B": 20, "C": 66, "D": 20, "E": 20, "F": 18})

    # ========================================================
    # 03 改定シナリオ
    # ========================================================
    s = Sheet(wb, "03_改定シナリオ比較", 9)
    s.title("３　料金改定シナリオの比較",
            "顧客別調定データ（1か月分3,060件）に各シナリオの料金を適用し、令和７年度の給水収益に換算したもの")
    rows, flags = [], []
    for sid, name, mb, mt, mm, cat in scen:
        res = apply_rate_scenario(recs, mb, mt, mm)
        annual = res["計"] * 12 * annual_factor
        up = annual - R7["給水収益"]
        rate_a = annual / kesson_a * 100
        rate_b = annual / kesson_b * 100
        # 20㎥ 家庭モデル（13㎜）の税込月額
        m20 = (CUR["基本_6-10㎥"] * mb + 10 * CUR["従量_10超100以下"] * mt
               + CUR["メーター"][13] * mm) * TAX
        rows.append([f"{sid}　{name}", cat, round(annual), round(up),
                     round((annual / R7["給水収益"] - 1) * 100, 1),
                     round(rate_a, 1), round(rate_b, 1), round(m20), round(m20 - 2475)])
        flags.append({"基準": None, "既存案": "warn", "率改定": None,
                      "目標達成": "good", "段階案": "good", "配分比較": None}.get(cat))
    s.table(["シナリオ", "分類", "年間給水収益（円）", "増収額（円）", "増収率(％)",
             "料金回収率\nベースＡ(％)", "料金回収率\nベースＢ(％)",
             "20㎥家庭料金\n(円・税込)", "20㎥の増額\n(円)"],
            rows, aligns=["left", "center"] + ["right"] * 7, flags=flags,
            fmt={2: "#,##0", 3: "#,##0", 4: "0.0", 5: "0.0", 6: "0.0", 7: "#,##0", 8: "#,##0"}, h=30)
    s.note([
        "※ ベースＡ＝令和７年度実績の経常費用（回収すべき費用340,676千円）／ベースＢ＝江島応急給水を除く平常時（同211,886千円）。",
        "※ 20㎥家庭料金は口径13㎜・家事用の月額（税込、メーター使用料を含む）。現行は2,475円。",
        "※【重要】経営戦略（案）のS1（基本料金×1.5）では増収率は＋17.5％にとどまり、料金回収率はベースＡで41.2％、ベースＢでも66.2％です。",
        "　 現行案の改定率では、計画期間中に料金回収率100％に到達しません。",
        "※ S9・S10は「同じ＋30％の増収を、基本料金だけ／従量料金だけで実現した場合」の比較です。配分によって需要家ごとの負担が大きく変わります。",
    ])
    s.widths({"A": 46, "B": 12, "C": 18, "D": 16, "E": 12, "F": 14, "G": 14, "H": 14, "I": 13})

    # ========================================================
    # 04 用途別・口径別の影響
    # ========================================================
    s = Sheet(wb, "04_用途別_口径別の影響", 8)
    s.title("４　改定シナリオ別の影響　－　用途区分・口径別",
            "代表3シナリオ（S1 経営戦略案／S3 一律＋30％／S7 段階第1段階＋35％）で比較")
    for sid, name, mb, mt, mm in [("S1", "経営戦略（案）基本料金×1.5", 1.50, 1.00, 1.00),
                                  ("S3", "一律＋30％", 1.30, 1.30, 1.30),
                                  ("S7", "段階改定 第1段階＋35％", 1.35, 1.35, 1.35)]:
        s.sub(f"■　{sid}　{name}")
        agg = aggregate(recs, mb, mt, mm, "用途")
        rows = []
        for k, v in sorted(agg.items(), key=lambda x: -x[1]["現行"]):
            up = v["改定後"] - v["現行"]
            rows.append([k, v["件数"], round(v["現行"]), round(v["改定後"]), round(up),
                         round(up / v["現行"] * 100, 1) if v["現行"] else 0,
                         round(v["現行"] / v["件数"]), round(v["改定後"] / v["件数"])])
        tot_c = sum(v["現行"] for v in agg.values()); tot_n = sum(v["改定後"] for v in agg.values())
        rows.append(["合　計", n_rec, round(tot_c), round(tot_n), round(tot_n - tot_c),
                     round((tot_n / tot_c - 1) * 100, 1), round(tot_c / n_rec), round(tot_n / n_rec)])
        s.table(["用途区分", "件数", "現行（円／月）", "改定後（円／月）", "増減（円／月）",
                 "増減率(％)", "1件平均\n現行(円)", "1件平均\n改定後(円)"],
                rows, aligns=["left"] + ["right"] * 7,
                fmt={1: "#,##0", 2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "0.0", 6: "#,##0", 7: "#,##0"})
    s.note([
        "※ 金額は税抜・1か月分。1件平均は各区分の平均であり、実際の負担は使用水量によって大きく異なります。",
        "※ S1（基本料金のみ×1.5）は、使用水量が少ない需要家ほど増減率が高くなります（基本料金の比重が大きいため）。",
    ])
    s.widths({"A": 22, "B": 12, "C": 18, "D": 18, "E": 16, "F": 12, "G": 14, "H": 14})

    # ========================================================
    # 05 モデルケース
    # ========================================================
    s = Sheet(wb, "05_モデルケース別の影響", 10)
    s.title("５　使用水量別のモデルケース　－　1か月あたりの水道料金（税込・メーター使用料を含む）",
            "口径13㎜（家事用の76％が20㎜、24％が13㎜）を基本とし、大口は実際の口径で試算")
    cases = [("家事用（単身・節水）", 13, 5), ("家事用（少量）", 13, 10),
             ("家事用（標準・2〜3人）", 13, 20), ("家事用（多め・4人以上）", 13, 30),
             ("家事用（20㎜・標準）", 20, 20), ("営業用（小規模店舗）", 20, 50),
             ("営業用（中規模）", 25, 100), ("工業用（水産加工・中）", 40, 250),
             ("工業用（水産加工・大）", 40, 1300), ("湯屋", 40, 662)]
    show = [("S0", "現行", 1.0, 1.0, 1.0), ("S1", "経営戦略案\n基本×1.5", 1.50, 1.00, 1.00),
            ("S3", "一律\n＋30％", 1.30, 1.30, 1.30), ("S7", "段階1\n＋35％", 1.35, 1.35, 1.35),
            ("S5", "回収率100％\n＋77.3％", 1.7734, 1.7734, 1.7734)]
    heads = ["モデルケース", "口径", "使用水量\n(㎥)"] + [f"{sid}\n{nm}" for sid, nm, *_ in show] + ["S3の増額", "S5の増額"]
    rows = []
    for label, dia, vol in cases:
        use = "湯屋用" if "湯屋" in label else "家事用"
        vals = []
        for sid, nm, mb, mt, mm in show:
            b, o, m = calc_charge(use, vol, dia, CUR)
            vals.append(round((b * mb + o * mt + m * mm) * TAX))
        rows.append([label, f"{dia}㎜", vol] + vals + [vals[2] - vals[0], vals[4] - vals[0]])
    s.table(heads, rows, aligns=["left", "center", "right"] + ["right"] * 7,
            fmt={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0", 6: "#,##0", 7: "#,##0",
                 8: "#,##0", 9: "#,##0"}, h=30)
    s.note([
        "※ 家事用の使用水量の中央値は14㎥、平均は20㎥です（調定データより）。",
        "※ 経営戦略（案）のS1は、20㎥世帯で月額＋605円（＋24.4％）ですが、100㎥超の大口には増額がほとんど及びません。",
        "　 現行の逓減制（100㎥超は110円→100円）と併せて、大口需要者への配分が課題です。",
    ])
    s.widths({"A": 26, "B": 8, "C": 10, "D": 12, "E": 13, "F": 12, "G": 12, "H": 14, "I": 12, "J": 12})

    # ========================================================
    # 06 体系見直しの選択肢
    # ========================================================
    s = Sheet(wb, "06_料金体系の見直し", 6)
    s.title("６　率の改定に加えて検討すべき「体系」の見直し",
            "同じ増収額でも、体系の設計によって需要家間の負担配分と将来の収入安定性が変わります")
    # 逓減制廃止・基本水量縮小の効果
    opts = []
    p2 = dict(CUR); p2["従量_100超"] = 110
    tot2 = sum(sum(calc_charge(r["用途"], r["水量"], r["口径"], p2)) for r in recs)
    p3 = dict(CUR); p3["基本水量"] = 5
    tot3 = sum(sum(calc_charge(r["用途"], r["水量"], r["口径"], p3)) for r in recs)
    base_calc = sum(sum(calc_charge(r["用途"], r["水量"], r["口径"], CUR)) for r in recs)
    opts.append(["① 逓減制の廃止（100㎥超も110円／㎥に統一）",
                 round(base_calc), round(tot2), round(tot2 - base_calc),
                 round((tot2 / base_calc - 1) * 100, 2),
                 "大口需要者のみに影響。水需要の減少局面では逓減制の合理性が乏しい"])
    opts.append(["② 基本水量の縮小（10㎥→5㎥。超過分を従量課金）",
                 round(base_calc), round(tot3), round(tot3 - base_calc),
                 round((tot3 / base_calc - 1) * 100, 2),
                 "使用水量の少ない層の負担が増える。低所得者への配慮が必要"])
    s.sub("■　単独の体系変更による増収効果（1か月・税抜）")
    s.table(["見直しの選択肢", "現行（円）", "見直し後（円）", "増収（円）", "増収率(％)", "評　価"],
            opts, aligns=["left", "right", "right", "right", "right", "left"],
            fmt={1: "#,##0", 2: "#,##0", 3: "#,##0", 4: "0.00"}, h=36)
    s.sub("■　体系上の論点")
    s.table(["論　点", "現　状", "考えられる方向"],
            [["基本水量", "10㎥（5㎥まで900円、6〜10㎥1,100円の2段階）",
              "家事用の中央値は14㎥。基本水量を縮小し、使用量に応じた負担に近づける。ただし少量利用者への影響に配慮"],
             ["口径別基本料金", "設定なし（用途別・水量別のみ）",
              "★固定費（施設の維持管理費）は口径に応じて負担するのが原則。口径別基本料金への移行は、"
              "水需要が減っても収入が安定するという利点もある"],
             ["逓減制", "100㎥超は110円→100円に低減",
              "水需要の抑制と原価回収の観点からは逓減制の廃止または縮小。大口需要者（水産加工業）への影響は要調整"],
             ["メーター使用料", "口径別に別建て（13㎜55円〜150㎜3,300円・税込）",
              "口径別基本料金に統合すれば体系が簡素化する。試算シートには単独での見直し案（13㎜418円等）も作成されている"],
             ["用途区分", "一般用／湯屋／船舶の3区分",
              "家事用・営業用・工業用・団体用・臨時用は同一料金。用途別料金の必要性は要検討"],
             ["改定の時期・方法", "未定",
              "★一度に大幅改定は困難。段階改定（例：3〜4年ごとに30〜35％）とし、改定のたびに検証する方式を推奨"]],
            aligns=["left", "left", "left"], h=48)
    s.widths({"A": 34, "B": 40, "C": 66, "D": 16, "E": 14, "F": 46})

    # ========================================================
    # 00 サマリー
    # ========================================================
    s = Sheet(wb, "00_結論サマリー", 5)
    s.ws.sheet_state = "visible"
    s.title("女川町 上水道事業　料金の妥当性検証と改定試算　－　結論",
            "令和８年８月　／　令和７年度決算・総務省 水道事業経営指標（令和６年度）・顧客別調定データ（3,060件）に基づく")
    s.sub("■　結論１　現行料金は、どの尺度でも明確に不足している")
    s.table(["指　標", "女川町（R7）", "比較対象", "水　準"],
            [["料金回収率", f"{ratio_a*100:.2f}％", "全国平均 97.82％（R5経営比較分析表）", "★36％の水準"],
             ["料金回収率（平常時ベース）", f"{ratio_b*100:.2f}％", "同　上", "★58％の水準"],
             ["供給単価", f"{kyokyu_tanka:.2f}円", "給水人口5千人〜1万人 平均 195.74円", "★60％の水準"],
             ["20㎥家庭料金（税込）", "2,475円", "同区分 平均 3,822.92円", "★65％の水準"],
             ["経常収支比率（基準外繰入を除く）", "72.44％", "100％以上が目安", "★27.6pt不足"]],
            aligns=["left", "right", "left", "center"], flags=["bad"] * 5, h=28)
    s.sub("■　結論２　経営戦略（案）の改定率では目標に到達しない")
    res_s1 = apply_rate_scenario(recs, 1.50, 1.00, 1.00)
    a_s1 = res_s1["計"] * 12 * annual_factor
    s.table(["項　目", "経営戦略（案）＝基本料金×1.5", "料金回収率100％に必要な水準"],
            [["増収率", f"＋{(a_s1/R7['給水収益']-1)*100:.1f}％", "＋77.3％（平常時）〜 ＋185.1％（R7実績）"],
             ["年間給水収益", f"{a_s1:,.0f}円", "211,886,312円 〜 340,676,289円"],
             ["料金回収率（ベースＡ）", f"{a_s1/kesson_a*100:.1f}％", "100.0％"],
             ["料金回収率（ベースＢ）", f"{a_s1/kesson_b*100:.1f}％", "100.0％"],
             ["20㎥家庭料金（税込）", "3,080円", "4,388円 〜 7,058円"]],
            aligns=["left", "right", "right"], flags=[None, None, "warn", "warn", None], h=26)
    s.note([
        "※ 現行案（基本料金×1.5・従量据置）の増収率は＋17.5％です。料金回収率は改善しますが、目標には遠く及びません。",
        "※ 一方で、＋77.3％や＋185.1％を一度に実施することは現実的ではありません。段階改定の設計が必要です。",
    ])
    s.sub("■　結論３　推奨する進め方")
    s.table(["No.", "内　容", "時　期"],
            [["1", "料金回収率100％は中長期目標と位置づけ、計画期間内の到達目標は「経常収支比率100％（基準外繰入を除く）」に設定する", "令和８年度中"],
             ["2", "段階改定（例：第1段階＋35％、4年後に第2段階＋35％＝累積＋82.3％）を基本線として設計する", "令和８年度中"],
             ["3", "率の改定と併せて体系を見直す（口径別基本料金の導入、逓減制の縮小、メーター使用料の統合）", "令和８〜９年度"],
             ["4", "捨水・応急給水の解消時期（鷲神浄水場高度処理施設の完成、江島海底送水管の本復旧）を確定させ、平常時の原価を確定する", "★最優先"],
             ["5", "繰出基準外の他会計補助金166,434千円の扱いを一般会計と整理する。料金で回収するのか、基準内繰入に組み替えるのかで必要改定率が変わる", "★最優先"],
             ["6", "大口需要者（水産加工業）への影響を個別に試算し、産業政策との調整を行う", "令和８年度中"]],
            aligns=["center", "left", "center"], h=40)
    s.widths({"A": 34, "B": 62, "C": 40, "D": 18, "E": 18})
    wb.move_sheet("00_結論サマリー", offset=-6)

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = C["subhead"]
    wb.worksheets[0].sheet_properties.tabColor = C["header"]

    path = os.path.join(OUT_DIR, "女川町_水道料金_妥当性検証と改定試算.xlsx")
    wb.save(path)
    print("saved:", path)

    # 標準出力に主要数値を出す（検証用）
    print(f"  調定データ {n_rec}件 / 月次水量 {mon_vol:,.0f}㎥ / 月次調定額(税抜) {cur_total:,.0f}円")
    print(f"  年換算補正係数 {annual_factor:.4f}")
    print(f"  料金回収率 A={ratio_a*100:.2f}%  B={ratio_b*100:.2f}%  C={ratio_c*100:.2f}%")
    print(f"  供給単価 {kyokyu_tanka:.2f}円 / 給水原価(日水協) {genka_nissui:.2f}円 / 給水原価(経営比較) {genka_hikaku:.2f}円")


if __name__ == "__main__":
    sys.exit(main())
