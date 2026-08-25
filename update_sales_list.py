# -*- coding: utf-8 -*-
"""
環境基本計画 営業先リストの更新（2026-08-25）

やること:
  1. 06_町村改訂時期_20260825 の調査結果のうち、01_R9満了候補へ未反映だった
     芽室町（令和8年度 策定手続中）と津別町（令和5年度 満了済）を反映する
  2. 文字コード調査報告（encodinginvestigation.md）の結論を適用する
     - 07_提出仕様_文字コード シートを追加し、提出時の規則を明文化
     - 提出用CSVを output/csv/ へ出力（既定 cp932・CRLF）
     - --ascii で半角英数ファイル名モード
  3. 00_概要・05_更新メモ を今回の更新に合わせて追記する

使い方:
  python3 update_sales_list.py                 # 日本語ファイル名 + cp932 CSV
  python3 update_sales_list.py --ascii         # 半角英数ファイル名 + cp932 CSV
  CSV_ENCODING=utf-8-sig python3 update_sales_list.py --ascii

注意:
  本ブックは build_hokkaido_tohoku_list.py が生成した初版に対し、
  その後の調査結果が直接編集で積み上がっている。初版生成スクリプトを
  再実行すると調査結果が失われるため、更新は本スクリプトで行う。
"""

import copy as _copy
import os
import shutil
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import naming

OUT_DIR = "/home/user/repository/output"
CSV_DIR = os.path.join(OUT_DIR, "csv")
BOOK_STEM_JP = "環境基本計画_営業先リスト_北海道東北"
SRC = os.path.join(OUT_DIR, f"{BOOK_STEM_JP}.xlsx")

C_TITLE, C_SUBHEAD, C_BAND, C_ALT, C_NOTE, C_WHITE = "1F3864", "2E75B6", "DDEBF7", "F7FAFC", "FFF3F3", "FFFFFF"
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name="Yu Gothic", size=14, bold=True, color=C_WHITE)
F_HEAD = Font(name="Yu Gothic", size=10, bold=True, color=C_WHITE)
F_BODY = Font(name="Yu Gothic", size=10)
F_BOLD = Font(name="Yu Gothic", size=10, bold=True)
F_NOTE = Font(name="Yu Gothic", size=9, color="7F3F3F")
AL_WRAP = Alignment(vertical="top", wrap_text=True)
AL_CTR = Alignment(vertical="center", horizontal="center", wrap_text=True)

ASCII = naming.ascii_mode()
ENC = naming.csv_encoding()


# ------------------------------------------------------------
# 1. 01_R9満了候補 への反映（芽室町・津別町）
# ------------------------------------------------------------
NEW_01_ROWS = [
    ["◆別扱い", "北海道", "芽室町", "努力義務", "クリーンめむろ環境基本計画",
     "令和8年度(2026) 策定手続中", "確認済",
     "令和8年度の町民参加予定に環境審議会・パブリックコメント・策定予定が掲げられており、まさに策定手続の最中。終期を見るだけでは拾えないが、当年度に動いている数少ない案件",
     "最優先。今すぐ接触し、既発注か自前策定か、支援できる範囲が残っているかを確認する",
     "①既発注・自前策定・支援範囲の確認 ②審議会・パブコメ工程に合わせた短期支援メニュー ③素案レビュー・アンケート二次分析の提案",
     "芽室町 令和8年度町民参加予定 https://www.memuro.net/administration/sanka/yotei/R8.html（06_町村改訂時期_20260825 行164）"],
    ["◆別扱い", "北海道", "津別町", "努力義務", "津別町環境基本計画",
     "平成26年度〜令和5年度（10年間・満了済）", "確認済",
     "令和5年度末で満了済。令和6年度に検証と今後の課題の協議を行っている一方、後継計画の策定状況は未確認。満了したまま次期が立ち上がっていない可能性がある",
     "後継計画の有無を早急に確認する。未策定なら令和8年度補正または令和9年度当初での策定を提案できる",
     "①後継計画の策定状況の確認 ②満了後の空白期間のリスク整理 ③町村向けの軽量な策定メニューと概算見積",
     "津別町 環境基本計画ページ https://www.town.tsubetsu.hokkaido.jp/soshiki/juminkikaku/14/859.html（06_町村改訂時期_20260825 行125）"],
]

# 期間要確認のまま残っている町村（01の末尾に注意喚起として置く）
NEW_01_CHECK_ROWS = [
    ["？要確認", "北海道", "上士幌町", "努力義務", "第2次上士幌町環境基本計画", "令和6年度(2024)策定・終期未確認", "一部確認",
     "令和6年度策定のため、R9満了ではなくR6策定直後の可能性が高い。ただしHTML上で終期が読めず確定していない",
     "PDF本文の計画期間だけ確認する", "—",
     "https://www.kamishihoro.jp/page/00000416（06シート 行160）"],
    ["？要確認", "北海道", "新得町", "努力義務", "新得町環境基本計画・地球温暖化対策実行計画(区域施策編)・気候変動適応計画", "策定済・終期未確認", "一部確認",
     "環境基本計画と温対・適応計画を一体で掲載している。一体計画のため改定時の業務範囲が広く、単価が伸びやすい",
     "PDF本文で改訂時期・終期を確認する", "—",
     "https://www.shintoku-town.jp/gyousei/koukai_kouhyou/tyoumin/kankyoukihonkeikaku/（06シート 行162）"],
    ["？要確認", "北海道", "池田町", "努力義務", "池田町環境基本計画（第3次改訂）", "令和8年度(2026)改訂・終期未確認", "一部確認",
     "令和8年度に改訂済み。短期は進捗管理・施策評価・アンケート分析の支援余地",
     "PDF本文で計画期間を確認する", "—",
     "https://www.town.hokkaido-ikeda.lg.jp/kurashi/kankyo/430.html（06シート 行170）"],
    ["？要確認", "北海道", "厚岸町", "努力義務", "第2期厚岸町豊かな環境を守り育てる基本計画", "終期未確認", "一部確認",
     "年次報告が2024年まで公表されており現行性は高い。計画名が『環境基本計画』でないため名称検索から漏れやすい",
     "PDF本文で計画期間だけ確認する", "—",
     "https://www.akkeshi-town.jp/gyosei/seisaku/kankyo/kankyo19/（06シート 行178）"],
    ["？要確認", "北海道", "鶴居村", "努力義務", "第2次鶴居村環境基本計画", "終期未確認", "一部確認",
     "村公式の計画一覧で第2次計画を確認。HTML上では期間が読めない",
     "PDF本文で計画期間だけ確認する", "—",
     "https://www.vill.tsurui.lg.jp/soshikikarasagasu/kikakuzaiseika/muranokeikaku/804.html（06シート 行182）"],
    ["？要確認", "北海道", "弟子屈町", "努力義務", "弟子屈町環境基本計画（外部DBで存在確認）", "公式終期未確認", "一部確認",
     "外部データベースで計画名は確認できるが、公式ページ・期間は未確認。外部DB情報だけで終期を埋めないこと",
     "公式PDFの有無と計画期間のみ確認する", "—",
     "外部DB（06シート 行181）"],
]


def copy_style(dst_cell, src_cell):
    dst_cell.font = _copy.copy(src_cell.font)
    dst_cell.fill = _copy.copy(src_cell.fill)
    dst_cell.border = _copy.copy(src_cell.border)
    dst_cell.alignment = _copy.copy(src_cell.alignment)


def find_row(ws, col, needle):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and needle in str(v):
            return r
    return None


def classify(row, ncol):
    """行の種別を返す。band=見出し帯 / note=注記 / data=データ / blank=空行。"""
    a = row[0]
    rest = [v for v in row[1:ncol] if v not in (None, "")]
    if a in (None, "") and not rest:
        return "blank"
    if isinstance(a, str) and a.startswith("■") and not rest:
        return "band"
    if not rest:
        return "note"
    return "data"


def rebuild_01(ws, rows, ncol, widths):
    """
    01_R9満了候補 を values から作り直す。
    openpyxl の insert_rows は既存セルを取りこぼすことがあるため、行挿入は使わない。
    """
    title = ws.cell(row=1, column=1).value
    headers = [ws.cell(row=2, column=c).value for c in range(1, ncol + 1)]
    ws_parent = ws.parent
    idx = ws_parent.sheetnames.index(ws.title)
    name = ws.title
    del ws_parent[name]
    new_ws = ws_parent.create_sheet(name, idx)

    new_ws.cell(row=1, column=1, value=title).font = F_TITLE
    new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    for c in range(1, ncol + 1):
        new_ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=C_TITLE)
    new_ws.row_dimensions[1].height = 26
    for c, h in enumerate(headers, start=1):
        cell = new_ws.cell(row=2, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = F_HEAD, PatternFill("solid", fgColor=C_SUBHEAD), AL_CTR, BORDER
    new_ws.row_dimensions[2].height = 32

    r = 3
    last_data = 2
    for row in rows:
        kind = classify(row, ncol)
        if kind == "blank":
            r += 1
            continue
        if kind == "band":
            new_ws.cell(row=r, column=1, value=row[0]).font = F_BOLD
            new_ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                new_ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_BAND)
                new_ws.cell(row=r, column=c).border = BORDER
            r += 1
            continue
        if kind == "note":
            new_ws.cell(row=r, column=1, value=row[0]).font = F_NOTE
            new_ws.cell(row=r, column=1).alignment = AL_WRAP
            new_ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                new_ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_NOTE)
            r += 1
            continue
        fill = PatternFill("solid", fgColor=C_ALT) if (r % 2 == 1) else None
        for c in range(1, ncol + 1):
            cell = new_ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else None)
            cell.font, cell.alignment, cell.border = F_BODY, AL_WRAP, BORDER
            if fill:
                cell.fill = fill
        last_data = r
        r += 1

    for c, w in enumerate(widths, start=1):
        new_ws.column_dimensions[get_column_letter(c)].width = w
    new_ws.freeze_panes = "A3"
    new_ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{last_data}"
    new_ws.sheet_view.zoomScale = 90
    return new_ws


def update_01(wb):
    ws = wb["01_R9満了候補"]
    if find_row(ws, 3, "芽室町") is not None:
        print("  01_R9満了候補: 反映済みのためスキップ")
        return 0
    ncol = ws.max_column
    widths = [ws.column_dimensions[get_column_letter(c)].width for c in range(1, ncol + 1)]

    rows = [list(r) for r in ws.iter_rows(min_row=3, values_only=True)]
    before = sum(1 for r in rows if classify(r, ncol) == "data")

    anchor = None
    for i, row in enumerate(rows):
        if classify(row, ncol) == "band" and "期間満了済・改定進行中" in str(row[0]):
            anchor = i
            break
    if anchor is None:
        raise SystemExit("エラー: 01_R9満了候補 に『■ 期間満了済・改定進行中』の帯が見つかりません。")

    spliced = rows[: anchor + 1] + [list(x) for x in NEW_01_ROWS] + rows[anchor + 1 :]
    spliced.append(["■ 2026-08-25 町村調査：計画はあるが期間が確認できていない先（PDFで終期だけ確認する）"])
    spliced.extend([list(x) for x in NEW_01_CHECK_ROWS])

    new_ws = rebuild_01(ws, spliced, ncol, widths)
    after = sum(1 for r in new_ws.iter_rows(min_row=3, values_only=True) if classify(list(r), ncol) == "data")
    added = len(NEW_01_ROWS) + len(NEW_01_CHECK_ROWS)
    if after != before + added:
        raise SystemExit(f"エラー: 行数が合いません（更新前{before} + 追加{added} != 更新後{after}）。データを失っている可能性があります。")
    print(f"  01_R9満了候補: データ行 {before} -> {after}（+{added}）")
    return added


# ------------------------------------------------------------
# 2. 07_提出仕様_文字コード シートの追加
# ------------------------------------------------------------
ENC_HEADERS = ["区分", "項目", "内容", "根拠・補足"]
ENC_WIDTHS = [12, 30, 92, 60]
ENC_ROWS = [
    ["■ 結論（文字コード調査報告より）"],
    ["結論", "このリポジトリのエンコード", "ソース・実行時・出力xlsxの全レイヤーがUTF-8であり、UTF-8化のための修正対象は存在しなかった",
     "encodinginvestigation.md 第1章・第2章"],
    ["結論", "行政系で弾かれる要因", "①出力ファイル名が日本語（非ASCII）②CSV提出時の文字コード（行政系はUTF-8ではなくShift_JIS(CP932)指定が多い）の2点",
     "同 第3章。「UTF-8へ直す」ことがかえって弾かれる原因になり得る"],
    ["結論", "xlsxの内部XML", "OOXML仕様上UTF-8固定。openpyxlに選択の余地がなく、変更対象にならない", "同 第2.3節"],
    ["■ 本ブックのCP932変換可否（2026-08-25 検査）"],
    ["検査", "検査対象", "全7シートの文字列セル 6,685件を1文字ずつ検査", "naming.check_encodable() による"],
    ["検査", "CP932で表現できない文字", "'—'（U+2014 EM DASH）が26箇所。ほかは無し", "「―」（U+2015 HORIZONTAL BAR）へ置換して出力する"],
    ["検査", "紛らわしいが問題なかった文字", "★ ▲ ◆ ■ ○ △ × ？ ｜ ＝ ① 〜 はいずれもCP932で表現できる",
     "〜 は U+301C（WAVE DASH）だが Python の cp932 コーデックで変換可能"],
    ["検査", "xlsx本体は書き換えていない", "'—' の置換はCSV書き出し時のみ行う。xlsxはUTF-8なので表示上の問題がなく、原文を保つ",
     "encodinginvestigation.md 第5.3節と同じ方針"],
    ["■ 提出用の出力（output/csv/）"],
    ["出力", "文字コード", "既定 cp932（Shift_JIS）。環境変数 CSV_ENCODING で変更する", "行政系システムで最も多い指定"],
    ["出力", "CSV_ENCODING=cp932", "Shift_JIS指定の行政系システム向け（既定）", "—"],
    ["出力", "CSV_ENCODING=utf-8-sig", "UTF-8指定かつExcelで開く場合。BOM付き", "—"],
    ["出力", "CSV_ENCODING=utf-8", "BOMなしUTF-8", "—"],
    ["出力", "改行コード", "CRLF固定", "Excel・行政系システムの慣例"],
    ["出力", "セル内改行の扱い", "CSVでは半角スペースへ畳む。引用符で囲めばRFC4180上は正しいが、行政系の取り込みで行数がずれる原因になりやすいため",
     "xlsx側の改行は保持している。畳みたくない場合は naming.write_csv(flatten_newlines=False)"],
    ["出力", "変換不能文字の扱い", "naming.CP932_SUBSTITUTIONS で代替文字へ置換し、置換内容を標準出力へ警告する", "代替表に無い文字は '?' へ落とす"],
    ["■ ファイル名"],
    ["ファイル名", "既定", "日本語ファイル名（環境基本計画_営業先リスト_北海道東北.xlsx）", "社内利用はこちら"],
    ["ファイル名", "半角英数モード", "--ascii オプション、または環境変数 ASCII_FILENAMES=1 で切り替える", "行政系へ提出する場合はこちら"],
    ["ファイル名", "対応表（ブック）", "環境基本計画_営業先リスト_北海道東北 → env_plan_sales_list_hokkaido_tohoku ／ 環境基本計画_営業整理_v8 → env_plan_sales_pack_v8 ／ 士幌町環境基本計画_レビューと提案 → shihoro_env_plan_review",
     "naming.py の BOOKS"],
    ["ファイル名", "対応表（シート→CSV）", "02_営業先マスタ → 02_target_master.csv など。未登録シートも機械生成のため非ASCIIにはならない", "naming.py の SHEET_SLUGS / sheet_slug()"],
    ["ファイル名", "実例", "本ブックをアップロードした際、日本語ファイル名がアンダースコアへ置換されて届いた（_____________________________20260825.xlsx）",
     "日本語ファイル名が経路上で壊れる実例。半角英数モードを用意した理由そのもの"],
    ["■ 運用"],
    ["運用", "社内", "python3 update_sales_list.py", "日本語ファイル名 + cp932 CSV"],
    ["運用", "行政系へ提出", "python3 update_sales_list.py --ascii", "半角英数ファイル名 + cp932 CSV"],
    ["運用", "提出先がUTF-8指定", "CSV_ENCODING=utf-8-sig python3 update_sales_list.py --ascii", "BOM付きUTF-8"],
    ["運用", "テキスト出力を追加する場合", "必ず naming.write_csv() を通すか、open() に encoding= を明示する",
     "Windows日本語環境では locale.getpreferredencoding() が cp932 を返すため"],
    ["■ 未確認事項"],
    ["未確認", "提出先システムの実際の仕様", "システム名・実際のエラーメッセージ・要求されるファイル形式／文字コード／ファイル名の制約は未確認",
     "encodinginvestigation.md 第7.2節。原因の特定にはこの3点が必要"],
    ["未確認", "ファイル形式の制限", "xlsxを受け付けず xls / csv / pdf のみ、という制限の有無は未確認", "提出先への確認が必要"],
]


def add_encoding_sheet(wb):
    name = "07_提出仕様_文字コード"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ncol = len(ENC_HEADERS)
    ws.cell(row=1, column=1, value="⑦ 提出仕様（ファイル名・文字コード）｜ 文字コード調査報告の結論を本ブックへ適用したもの").font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=C_TITLE)
    ws.row_dimensions[1].height = 26
    for c, h in enumerate(ENC_HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = F_HEAD, PatternFill("solid", fgColor=C_SUBHEAD), AL_CTR, BORDER
    ws.row_dimensions[2].height = 30
    r = 3
    for row in ENC_ROWS:
        if len(row) == 1:
            ws.cell(row=r, column=1, value=row[0]).font = F_BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            for c in range(1, ncol + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=C_BAND)
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            continue
        fill = PatternFill("solid", fgColor=C_ALT) if (r % 2 == 1) else None
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c, value=row[c - 1] if c - 1 < len(row) else None)
            cell.font, cell.alignment, cell.border = F_BODY, AL_WRAP, BORDER
            if fill:
                cell.fill = fill
        r += 1
    note = ("提出前に必ず確認すること: 提出先システム名／アップロード時の実際のエラーメッセージ／"
            "要求されるファイル形式・文字コード・ファイル名の制約。この3点が揃うまで、原因を文字コードと断定しないこと。")
    ws.cell(row=r + 1, column=1, value=note).font = F_NOTE
    ws.cell(row=r + 1, column=1).alignment = AL_WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(row=r + 1, column=c).fill = PatternFill("solid", fgColor=C_NOTE)
    for c, w in enumerate(ENC_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}{r - 1}"
    ws.sheet_view.zoomScale = 90
    return ws


# ------------------------------------------------------------
# 3. 00_概要 / 05_更新メモ の追記
# ------------------------------------------------------------
def append_styled_row(ws, values, template_row):
    ncol = ws.max_column
    r = ws.max_row + 1
    for c in range(1, ncol + 1):
        src = ws.cell(row=template_row, column=c)
        cell = ws.cell(row=r, column=c)
        copy_style(cell, src)
        cell.value = values[c - 1] if c - 1 < len(values) else None
    return r


def update_00(wb, added_rows):
    ws = wb["00_概要"]
    if find_row(ws, 2, "07_提出仕様_文字コード") is not None:
        print("  00_概要: 反映済みのためスキップ")
        return
    # 構成に07シートを追記
    tmpl = find_row(ws, 1, "構成") or 3
    append_styled_row(ws, ["構成", "07_提出仕様_文字コード",
                           "ファイル名と文字コードの規則。行政系へ提出する際の出力方法を明文化",
                           "文字コード調査報告（encodinginvestigation.md）の結論を適用"], tmpl)
    append_styled_row(ws, ["更新", "2026-08-25 の更新",
                           f"06_町村改訂時期の調査結果のうち01へ未反映だった芽室町・津別町を反映（追加{added_rows}行）。提出用CSV出力と半角英数ファイル名モードを追加",
                           "05_更新メモ・07_提出仕様_文字コード を参照"], tmpl)
    append_styled_row(ws, ["注意", "本ブックの更新方法",
                           "本ブックは初版生成後、調査結果が直接編集で積み上がっている。初版生成スクリプト（build_hokkaido_tohoku_list.py）を再実行すると調査結果が失われるため、更新は update_sales_list.py で行う",
                           "—"], tmpl)


def update_05(wb, added_rows):
    ws = wb["05_更新メモ"]
    if find_row(ws, 1, "2026-08-25 提出仕様") is not None:
        print("  05_更新メモ: 反映済みのためスキップ")
        return
    tmpl = ws.max_row
    append_styled_row(ws, [
        "2026-08-25 反映漏れ解消",
        "芽室町・津別町",
        "06_町村改訂時期の最優先・優先先が01_R9満了候補に載っていなかったため反映。併せて期間未確認の町村6件を要確認ブロックとして追加",
        "芽室町はR8策定手続中で今すぐ接触。津別町はR5満了済で後継の有無を早急に確認",
        "06_町村改訂時期_20260825 行164・行125ほか",
    ], tmpl)
    append_styled_row(ws, [
        "2026-08-25 提出仕様",
        "文字コード・ファイル名",
        "文字コード調査報告の結論を適用。07_提出仕様_文字コードを追加し、提出用CSV（既定cp932・CRLF）と半角英数ファイル名モード（--ascii）を実装",
        "行政系へ提出する場合は --ascii で出力する。CP932非対応は '—' 26箇所のみで、CSV書き出し時に '―' へ置換",
        "encodinginvestigation.md ／ naming.py",
    ], tmpl)


# ------------------------------------------------------------
# 4. 出力（xlsx / CSV）
# ------------------------------------------------------------
def export_csv(wb, ascii_names):
    os.makedirs(CSV_DIR, exist_ok=True)
    written = []
    for i, ws in enumerate(wb.worksheets):
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        stem = naming.sheet_slug(ws.title, i) if ascii_names else ws.title
        path = os.path.join(CSV_DIR, f"{stem}.csv")
        naming.write_csv(path, rows, encoding=ENC)
        written.append(path)
    return written


def audit(wb):
    values = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values.extend(v for v in row if isinstance(v, str))
    bad = naming.check_encodable(values, ENC)
    print(f"\n=== {ENC} エンコード検査（xlsx本体） ===")
    if not bad:
        print("  変換不能文字: なし")
    for ch, n in sorted(bad.items(), key=lambda x: -x[1]):
        alt = naming.CP932_SUBSTITUTIONS.get(ch, "?")
        print(f"  {ch!r} U+{ord(ch):04X} {n}箇所 -> CSV出力時に {alt!r} へ置換")
    return bad


def main():
    if not os.path.exists(SRC):
        print(f"エラー: 入力ブックがありません -> {SRC}")
        print(f"       （現在のモード … {naming.mode_label(ASCII)}）")
        sys.exit(1)

    wb = load_workbook(SRC)
    added = update_01(wb)
    add_encoding_sheet(wb)
    update_00(wb, added)
    update_05(wb, added)

    out_stem = naming.book_stem(BOOK_STEM_JP, ASCII)
    out_path = os.path.join(OUT_DIR, f"{out_stem}.xlsx")
    wb.save(out_path)
    print(f"保存: {out_path}")

    # 日本語名で更新した場合でも、提出用に半角英数名の複製を残す
    if not ASCII:
        ascii_path = os.path.join(OUT_DIR, f"{naming.book_stem(BOOK_STEM_JP, True)}.xlsx")
        shutil.copyfile(out_path, ascii_path)
        print(f"保存: {ascii_path}（提出用・半角英数名）")

    audit(wb)
    print(f"\n=== 提出用CSV（{ENC} / CRLF） ===")
    for p in export_csv(wb, ascii_names=True):
        print(f"  {p}")
    print(f"\nモード … {naming.mode_label(ASCII)}")
    print(f"01_R9満了候補への追加行数: {added}")


if __name__ == "__main__":
    main()
