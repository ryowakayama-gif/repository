# -*- coding: utf-8 -*-
"""
北塩原村 第8期障がい福祉計画・第4期障がい児福祉計画
サービス見込量（個別積上げ）管理ブック ジェネレータ

出力: output/北塩原村_サービス見込量.xlsx

推計方式
    翌年度利用者数 ＝ 前年度継続利用者
                    ＋ 新規支給決定見込
                    ＋ 年齢・進学・卒業等による流入
                    ＋ 施設・病院等からの地域移行
                    － 介護保険等への移行
                    － 一般就労・サービス終了
                    － 死亡・転出・長期入院等
    月間サービス量 ＝ 利用者数 × 1人当たり月利用量

    ＋ 未特定の潜在需要（第3層）／－ 供給制約により利用困難な量

採用理由（令和8年3月31日 こども家庭庁・厚生労働省告示第4号による改正後の基本指針で確認済み）
- 第三の二の2(一)：別表第一を参考としつつ、現在の利用実績等に関する分析、障がい者等の
  サービスの利用に関する意向、心身の状況等を勘案し、地域の実情を踏まえて設定する。
- 別表第一：現に利用している者の数、ニーズ、地域移行者数、平均的な一人当たり利用量等を
  勘案して「利用者数及び量」の見込みを設定する。本ブックの利用者数×月利用量はこれに従う。
- 福島県計画は各市町村の見込量を積み上げる構造であり、県値を人口比で按分する方法は適合しない。
- 本村は利用者1人が生活介護で16.7%、居宅介護・施設入所支援で25%を占め、
  年率2〜8%の伸び率を大きく上回る。過去実績の伸び率・据置では表現できない。

留意
- 全サービスの個別積上げが国により義務付けられているわけではない。
  第8期基本指針が重度障がい者について個別利用者数の把握を強めた考え方を、
  本村の規模に応じて全サービスへ応用するもの。
- 第8期で新設された別表第五は、全部過疎市町村でなく、かつサービス利用者割合が
  上位25%以内の市町村に対し、全国の伸び率を用いる算定方法を基本と定めた。
  本村は過疎法第2条により全域が過疎地域に該当するため適用対象外であることを確認済み
  （福島県「県内の過疎・中山間地域の指定状況」令和7年4月1日現在。12_別表第五判定シート）。
- 65歳到達・40歳到達は一律に減算しない。個別判定の結果のみを反映する。
  告示は65歳到達による減算を求める規定を置いておらず、別表第一の勘案要素にも現れない。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from kitashiobara_common import (
    COLORS, FONT, FLOW_ROWS_ADULT, FLOW_ROWS_CHILD, KAIGO_SOTO, OUT_DIR,
    SERVICES_ADULT, SERVICES_CHILD, SERVICES_ZERO,
    add_sheet, ensure_out_dir, set_col_widths, style_data_cell,
    style_header_row, style_note, style_status, style_title, write_row,
)

OUT_FILE = f"{OUT_DIR}/北塩原村_サービス見込量.xlsx"

INT = "#,##0"
DEC1 = "#,##0.0"
SIGNED = "+0;-0;－"

YEARS = ["令和8年度\n（基準）", "令和9年度", "令和10年度", "令和11年度"]
YCOL = ["C", "D", "E", "F"]   # 年度の列


# ============================================================
# 00_概要
# ============================================================
def sheet_overview(wb):
    ws = wb.create_sheet("00_概要")
    set_col_widths(ws, [34, 18, 22, 14, 44])
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:E1")
    style_title(ws["A1"], "北塩原村 サービス見込量 管理ブック（個別積上げ方式）",
                fill=COLORS["障がい"])
    ws.merge_cells("A2:E2")
    style_note(ws["A2"],
               "第8期北塩原村障がい福祉計画・第4期北塩原村障がい児福祉計画（令和9〜11年度）／"
               "業務仕様書4-Ⅱ(3)②〜④「障がい福祉サービス・障がい児通所支援等・地域生活支援事業の見込量の推計」に対応。")
    ws.row_dimensions[2].height = 28

    r = 4
    style_header_row(ws, r, ["シート", "役割", "推計方式上の位置", "状態", "備考"])
    r += 1
    rows = [
        ("01_推計方式", "算式と3層構造の定義", "全体の前提", "確定",
         "計画本文・協議会資料にそのまま転記できる形で記載"),
        ("02_年齢到達者一覧", "18歳・40歳・65歳到達予定者の匿名台帳", "第1層の入口", "村資料待ち",
         "支給決定者台帳から作成する。本ブックの起点"),
        ("03_積上げ_障害福祉サービス", "サービス別の増減要因の積上げ", "第1層＋第2層", "村資料待ち",
         "12サービス。基準値は現行計画の令和8年度見込量"),
        ("04_積上げ_障害児通所支援", "同上（障がい児）", "第1層＋第2層", "村資料待ち",
         "6サービス。就学・18歳移行を明示的に扱う"),
        ("05_実績ゼロサービス", "実績ゼロのサービスの確認欄", "第3層", "村資料待ち",
         "実績ゼロをニーズゼロと判断しないための確認"),
        ("06_65歳移行判定", "介護保険との相当関係と個別判定", "第1層の判定", "村資料待ち",
         "一律移行ではなく個別判断。制度整理は確定"),
        ("07_潜在需要", "アンケート・相談実績による補正", "第3層", "集計後",
         "伸び率を用いるのはこの層に限定する"),
        ("08_供給制約", "圏域事業所の定員・空き・送迎", "利用可能量の上限", "村資料待ち",
         "利用希望量と実際に利用可能な量を分離する"),
        ("09_地域生活支援事業", "村事業の見込量", "第1層＋第3層", "村資料待ち",
         "移動支援0人の再検証を含む"),
        ("10_村確認事項", "本方式に必要な村資料", "―", "村資料待ち",
         "年齢別・サービス別の匿名利用者一覧が最優先"),
        ("11_告示別表第一根拠", "サービス別に告示が勘案を求める要素", "全体の前提", "確定",
         "令和8年3月31日告示第4号 別表第一と本ブックの積上げ項目の対照表"),
        ("12_別表第五判定", "地域差是正規定の適用判定", "算定方法の前提", "確定",
         "全域が過疎（過疎法第2条）のため適用対象外。個別積上げ方式を採用できる"),
    ]
    for i, row in enumerate(rows):
        write_row(ws, r, list(row), alt=(i % 2 == 1),
                  aligns=["left", "left", "left", "center", "left"])
        style_status(ws.cell(row=r, column=4))
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "作業の順序", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "1. 村から年齢別・サービス別の匿名利用者一覧を受領し、02_年齢到達者一覧を作成します。",
        "2. 令和9〜11年度に18歳・40歳・65歳へ到達する方を抽出し、到達年度と現在の利用サービスを記入します。",
        "3. 65歳到達者は06_65歳移行判定で個別に移行可否を判定します。一律に減算してはいけません。",
        "4. 判定結果を03・04の該当する増減行に転記します。継続利用者と利用者数は数式のため触りません。",
        "5. 圏域事業所へ利用可能性を照会し、08_供給制約に記入します。利用希望量が供給を超える場合は差を明示します。",
        "6. ここまでで得られた値を基本ケースとし、未特定需要（07）のみを低位・中位・高位で補正します。",
        "7. 確定した利用者数・月間サービス量を計画書 第5章の見込量表と、財源構成案ブックの08へ転記します。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 22
        r += 1
    return ws


# ============================================================
# 01_推計方式
# ============================================================
def sheet_method(wb):
    ws = add_sheet(
        wb, "01_推計方式", "見込量の推計方式",
        "国は全国一律の算定式を定めておらず、地域の実情に即した見込量設定を求めています。"
        "福島県計画は各市町村の見込量を積み上げる構造のため、県値を人口比で按分する方法は適合しません。"
        "本村は利用者1人の増減がサービスによって10〜25%を占めるため、個別積上げを基本とします。",
        [26, 62, 52])

    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_title(ws.cell(row=r, column=1), "1. 算式", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "算式", "備考"])
    r += 1
    for i, (k, f, n) in enumerate([
        ("利用者数",
         "翌年度利用者数 ＝ 前年度継続利用者 ＋ 新規支給決定見込 ＋ 年齢・進学・卒業等による流入 "
         "＋ 施設・病院等からの地域移行 － 介護保険等への移行 － 一般就労・サービス終了 "
         "－ 死亡・転出・長期入院等",
         "各項を1人単位で加減算する。03・04の各サービスブロックがこの構造"),
        ("サービス量",
         "月間サービス量 ＝ 利用者ごとの月間予定利用量の合計 ＋ 未特定の潜在需要分 "
         "－ 供給制約により利用困難な量",
         "潜在需要（07）と実際の利用見込量は分けて記載する"),
    ]):
        write_row(ws, r, [k, f, n], alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 60
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_title(ws.cell(row=r, column=1), "2. 推計対象の3層", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["層", "対象", "扱い"])
    r += 1
    layers = [
        ("第1層\n確定・高確度案件",
         "特別支援学校高等部の卒業予定者／障害児通所支援利用者の就学予定／"
         "令和9〜11年度の65歳到達者／施設退所・病院退院予定者／一般就労予定者／"
         "転出予定者／事業所から利用開始相談を受けている方",
         "1人単位で加減算する。02_年齢到達者一覧と06_65歳移行判定で個別に確定させる"),
        ("第2層\n既存利用者の利用量変化",
         "障害支援区分の変更／利用日数の増減／家族介護力の低下／親亡き後への対応／"
         "送迎条件の変化／事業所の変更",
         "利用者ごとの月利用日数・時間を調整する。03・04の「1人当たり月利用量」行"),
        ("第3層\n未特定の潜在需要",
         "アンケートの利用希望／相談支援事業所への相談件数／未利用者のニーズ／"
         "新規診断・手帳取得の傾向／圏域のサービス整備予定",
         "過去実績の伸び率を用いるのはこの層に限定する。07_潜在需要で管理する"),
    ]
    for i, rec in enumerate(layers):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 66
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_title(ws.cell(row=r, column=1), "3. 年齢到達の扱い", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年齢", "制度上の内容", "推計上の扱い"])
    r += 1
    ages = [
        ("6歳（就学）",
         "児童発達支援から放課後等デイサービスへ移行する。",
         "04で児童発達支援の流出と放課後等デイサービスの流入を同一人数で対応させる。"
         "令和6・7年度の給付実績にこの動きが現れている可能性がある（受給者単位の確認が必要）"),
        ("18歳到達・高校卒業",
         "障害児通所支援が終了し、成人の障害福祉サービスへ移行する。"
         "障害児相談支援は計画相談支援へ切り替わる。介護給付の利用には障害支援区分の認定が必要。",
         "卒業年度により扱いが異なる。令和8年度卒業なら令和9年度の流入に計上する。"
         "令和6〜7年度卒業なら既に移行済みのため令和8年度の基準人数に含める"),
        ("18〜22歳（入所児）",
         "障害児入所施設からの移行。令和6年4月施行の改正児童福祉法により、"
         "移行調整の責任主体は都道府県。22歳満了時までの入所継続が可能。"
         "福島県は令和5年度に移行調整の協議の場を設置している。",
         "村単独で推計せず、県・入所施設・相談支援事業所・成人サービス事業所と"
         "移行年度と移行先を確認したうえで計上する"),
        ("20歳到達",
         "特別児童扶養手当から障害基礎年金へ切り替わる。放課後等デイサービスの特例利用の上限。",
         "サービス見込量への直接の影響は小さい。手当・年金の切替として整理する"),
        ("40歳到達",
         "介護保険第2号被保険者となる。ただし介護保険給付を受けるには、"
         "16の特定疾病に起因する要介護・要支援状態であることが必要。",
         "40歳到達人数を一律に減算しない。特定疾病に該当し、実際に介護保険の認定申請・"
         "移行予定がある方のみを個別に反映する"),
        ("65歳到達",
         "介護保険優先原則（障害者総合支援法第7条）が適用される。"
         "ただし平成19年通知（平成27年改正）により一律適用ではなく市町村が個別に判断する。"
         "最高裁も、近隣に利用可能な事業所がない等の事情があれば介護保険の認定を経ずに"
         "給付決定し得ることを認めている。",
         "一律に減算しない。06_65歳移行判定でサービスを「相当あり」「相当なし」「一部あり」に"
         "分け、到達者ごとに移行予定・障害福祉継続・月量への影響を判定して反映する"),
    ]
    for i, rec in enumerate(ages):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 76
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_title(ws.cell(row=r, column=1), "4. 県内他団体の推計方法との関係",
                fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["類型", "県内で確認された方法", "本村への適合性"])
    r += 1
    types = [
        ("①実績傾向型", "過去の利用実績・増減傾向を延長する（川俣町の手帳所持者数推計等）",
         "低い。1人の増減が10〜25%を占める本村では傾向線が意味を持たない"),
        ("②実績＋ニーズ型",
         "実績に加え、アンケート・利用者ニーズを考慮する（湯川村が算式ではなく方針として明記）",
         "中程度。県内で最も一般的だが、増減の理由が1人単位で説明できない"),
        ("③個別積上げ型",
         "現利用者・新規利用予定者・進学・卒業・移行予定を積み上げる",
         "高い。本ブックが採用する方式。県内で算式まで明示している団体は限定的"),
    ]
    for i, rec in enumerate(types):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 44
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    style_note(ws.cell(row=r, column=1),
               "県内計画に共通する弱点として、算式が明記されていない、1人単位の増減理由が分からない、"
               "児童発達支援→放課後等デイサービス→成人サービスの連動が見えない、"
               "65歳到達による介護保険移行が見込量表から判別できない、"
               "利用希望量と圏域の供給可能量が分離されていない、の5点があります。"
               "県内他団体と同じ方法に合わせるだけでは本村の課題は解消しません。")
    ws.row_dimensions[r].height = 44
    return ws


# ============================================================
# 02_年齢到達者一覧
# ============================================================
def sheet_age_roster(wb):
    ws = add_sheet(
        wb, "02_年齢到達者一覧", "年齢到達者一覧（匿名台帳）",
        "村の支給決定者台帳から作成します。氏名は不要で、利用者番号等の匿名の識別子で管理してください。"
        "本ブックの起点となるシートです。個人が特定されるため、公表用の報告書には掲載しません。",
        [12, 12, 12, 12, 12, 26, 14, 14, 16, 16, 30])
    style_header_row(ws, 5, ["利用者\n番号", "到達区分", "到達\n年度", "到達時\n年齢",
                             "障害\n支援区分", "現在利用しているサービス", "介護保険\n相当の有無",
                             "移行\n予定", "障害福祉\n継続", "月量への\n影響", "確認状況・備考"])

    aligns = ["center", "center", "center", "center", "center", "left",
              "center", "center", "center", "right", "left"]
    r = 6
    # 記入例（1行のみ。実データではないことを明示）
    write_row(ws, r, ["（例）001", "65歳", "令和10年度", "65", "区分4",
                      "生活介護、居宅介護", "あり", "一部", "生活介護は継続",
                      "-10人日", "例示。実データではありません"],
              aligns=aligns, fills=[COLORS["note"]] * 11)
    ws.cell(row=r, column=1).font = Font(name=FONT, size=10, italic=True)
    r += 1
    for i in range(24):
        write_row(ws, r, [None] * 11, alt=(i % 2 == 1), aligns=aligns,
                  fills=[COLORS["input"]] * 11)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    style_title(ws.cell(row=r, column=1), "記入のしかた", fill=COLORS["subhead"], size=11)
    r += 1
    for txt in [
        "・到達区分：18歳／40歳／65歳のいずれか。就学（6歳）や20歳も必要に応じて追加してください。",
        "・到達年度：令和9〜11年度に到達する方が対象です。令和8年度に到達済みの方は基準人数に含まれるため、"
        "03・04の令和8年度基準へ反映します。",
        "・介護保険相当の有無：06_65歳移行判定の区分表に従って「あり」「なし」「一部」を記入します。",
        "・移行予定：「する」「しない」「一部」「未定」。40歳到達者は、特定疾病に該当し要介護認定を"
        "受ける見込みがある場合のみ「する」とし、それ以外は「しない」とします。",
        "・月量への影響：人日／月または時間／月で、増減を符号付きで記入します。03・04の該当行に転記します。",
        "・18歳到達者は、卒業年度と卒業後の進路（生活介護、就労継続支援B型、就労移行支援等）、"
        "成人サービスの利用開始日、月利用日数、計画相談への切替状況まで確認してください。",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        style_note(ws.cell(row=r, column=1), txt)
        ws.row_dimensions[r].height = 22
        r += 1
    return ws


# ============================================================
# 03・04 個別積上げ
# ============================================================
def _buildup_sheet(wb, title, heading, note, services, flow_rows, kind):
    ws = add_sheet(wb, title, heading, note, [30, 12, 14, 14, 14, 14, 46])
    r = 5
    for si, (name, unit_p, unit_v, r8_users, r8_vol, per_user, svc_note) in enumerate(services):
        # サービス見出し
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        style_title(ws.cell(row=r, column=1), f"{si + 1}　{name}",
                    fill=COLORS["障がい"] if kind == "child" else COLORS["subhead"], size=11)
        r += 1
        if svc_note:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            style_note(ws.cell(row=r, column=1), f"※ {svc_note}")
            ws.row_dimensions[r].height = 20
            r += 1
        style_header_row(ws, r, ["増減要因", "単位"] + YEARS + ["備考"])
        r += 1
        first = r
        rowmap = {}
        for label, mode, note_txt in flow_rows:
            rowmap[label] = r
            r += 1
        users_row = rowmap["＝　利用者数"]
        per_row = rowmap["　　1人当たり月利用量"]
        vol_row = rowmap["＝　月間サービス量"]
        inflow = [rowmap[k] for k, m, n in flow_rows if k.startswith("＋")]
        outflow = [rowmap[k] for k, m, n in flow_rows if k.startswith("－")]
        cont_row = rowmap["継続利用者"]

        for label, mode, note_txt in flow_rows:
            rr = rowmap[label]
            is_vol = label in ("　　1人当たり月利用量", "＝　月間サービス量")
            unit = (unit_v if is_vol else unit_p)
            vals = []
            fills = [None, None]
            for ci, col in enumerate(YCOL):
                if label == "継続利用者":
                    if ci == 0:
                        vals.append(r8_users)
                        fills.append(COLORS["input"])
                    else:
                        vals.append(f"={YCOL[ci - 1]}{users_row}")
                        fills.append(COLORS["calc"])
                elif label == "＝　利用者数":
                    inc = f"SUM({col}{min(inflow)}:{col}{max(inflow)})"
                    dec = f"SUM({col}{min(outflow)}:{col}{max(outflow)})"
                    vals.append(f"={col}{cont_row}+{inc}-{dec}" if ci > 0
                                else f"={col}{cont_row}")
                    fills.append(COLORS["calc"])
                elif label == "　　1人当たり月利用量":
                    vals.append(per_user if per_user else None)
                    fills.append(COLORS["input"])
                elif label == "＝　月間サービス量":
                    vals.append(f"=ROUND({col}{users_row}*{col}{per_row},1)")
                    fills.append(COLORS["calc"])
                else:
                    vals.append(None)
                    fills.append(COLORS["input"] if ci > 0 else None)
            fills.append(None)
            emphasise = label in ("＝　利用者数", "＝　月間サービス量")
            write_row(ws, rr, [label, unit] + vals + [note_txt],
                      aligns=["left", "center", "right", "right", "right", "right", "left"],
                      numfmts=[None, None] + ([DEC1] * 4 if is_vol else [INT] * 4) + [None],
                      fills=fills)
            if emphasise:
                for c in range(1, 7):
                    ws.cell(row=rr, column=c).font = Font(name=FONT, size=10, bold=True)
            if label.startswith("－"):
                for c in range(3, 7):
                    ws.cell(row=rr, column=c).number_format = "-#,##0;-#,##0;－"
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "黄色は入力欄、緑は数式です。継続利用者は前年度の利用者数を自動で繰り越します。"
               "流出行はマイナスの人数ではなく、減少する人数を正の数で入力してください（表示は自動でマイナスになります）。"
               "令和8年度の基準人数は、村の支給決定者台帳による実利用者数に置き換えてください。"
               "現在入っているのは現行計画（第7期）の令和8年度見込量です。")
    ws.row_dimensions[r].height = 40
    return ws


def sheet_buildup_adult(wb):
    return _buildup_sheet(
        wb, "03_積上げ_障害福祉サービス", "障害福祉サービスの個別積上げ",
        "サービスごとに、継続・流入・流出を1人単位で積み上げます。"
        "65歳到達・40歳到達による減算は、06_65歳移行判定で個別に判定した結果のみを反映し、"
        "到達人数を一律に減算してはいけません。",
        SERVICES_ADULT, FLOW_ROWS_ADULT, "adult")


def sheet_buildup_child(wb):
    return _buildup_sheet(
        wb, "04_積上げ_障害児通所支援", "障がい児通所支援等の個別積上げ",
        "就学（児童発達支援→放課後等デイサービス）と18歳到達（成人サービスへ）を明示的に扱います。"
        "「－ 18歳到達・高校卒業による移行」の人数は、03の「＋ 特別支援学校卒業・18歳到達」と一致させてください。",
        SERVICES_CHILD, FLOW_ROWS_CHILD, "child")


# ============================================================
# 05_実績ゼロサービス
# ============================================================
def sheet_zero(wb):
    ws = add_sheet(
        wb, "05_実績ゼロサービス", "実績ゼロのサービスの確認",
        "実績ゼロをニーズゼロと判断しないための確認欄です。制度が未周知である、相談につながっていない、"
        "村内・圏域に供給がない、移動手段がない、家族が代替している、のいずれに該当するかを確認してください。",
        [28, 14, 14, 14, 14, 20, 44])
    style_header_row(ws, 5, ["サービス", "令和8年度\n見込", "令和9年度",
                             "令和10年度", "令和11年度", "実績ゼロの理由", "確認の観点"])
    r = 6
    reasons = "制度未周知／相談未接続／供給なし／移動手段なし／家族が代替／ニーズ不在"
    for i, (name, note) in enumerate(SERVICES_ZERO):
        write_row(ws, r, [name, 0, None, None, None, None, note],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "center", "left"],
                  numfmts=[None, INT, INT, INT, INT, None, None],
                  fills=[None, None, COLORS["input"], COLORS["input"], COLORS["input"],
                         COLORS["input"], None])
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               f"「実績ゼロの理由」欄には次のいずれかを記入します：{reasons}。"
               "同行援護・行動援護・重度訪問介護の移動支援部分は介護保険に相当サービスがないため、"
               "65歳以上の方でも障害福祉サービスとして利用できます。高齢の視覚障がい者・"
               "強度行動障害のある方の有無を確認してください。")
    ws.row_dimensions[r].height = 40
    return ws


# ============================================================
# 06_65歳移行判定
# ============================================================
def sheet_age65(wb):
    ws = add_sheet(
        wb, "06_65歳移行判定", "65歳到達に伴う介護保険への移行判定",
        "介護保険優先原則は一律適用ではありません。平成19年通知（平成27年改正）により、"
        "市町村は本人の意向を聴取し、介護保険サービスで必要な支援が受けられるかを個別に確認したうえで"
        "支給決定を行うこととされています。到達人数に移行率を掛ける機械的な推計は行いません。",
        [26, 14, 26, 52, 20])
    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "1. サービス区分（制度整理・確定）",
                fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["障害福祉サービス", "区分", "介護保険の相当サービス", "推計上の扱い", ""])
    r += 1
    for i, (svc, kubun, kaigo, atsukai) in enumerate(KAIGO_SOTO):
        write_row(ws, r, [svc, kubun, kaigo, atsukai, ""], alt=(i % 2 == 1),
                  aligns=["left", "center", "left", "left", "left"])
        c = ws.cell(row=r, column=2)
        if kubun == "相当なし":
            c.fill = PatternFill("solid", fgColor="2CA02C")
            c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        elif kubun == "相当あり":
            c.fill = PatternFill("solid", fgColor="ED7D31")
            c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "2. 到達者別の判定（入力）", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["利用者番号", "到達年度", "現在のサービス・月量",
                             "判定と根拠", "月量への影響"])
    r += 1
    for i in range(12):
        write_row(ws, r, [None] * 5, alt=(i % 2 == 1),
                  aligns=["center", "center", "left", "left", "right"],
                  fills=[COLORS["input"]] * 5)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "3. 判定にあたって確認する事項",
                fill=COLORS["subhead"], size=11)
    r += 1
    checks = [
        ("介護保険で必要な支援が受けられるか",
         "支援内容・時間・専門性が介護保険サービスで代替できるか。代替できない場合は障害福祉を継続する"),
        ("圏域に利用可能な事業所があるか",
         "村内に事業所がないため、圏域の介護保険事業所が受け入れ可能かを確認する。"
         "共生型サービスの指定を受けている事業所があれば、同一事業所での継続利用が可能"),
        ("新高額障害福祉サービス等給付費の要件を満たすか",
         "①65歳に達する日前5年間の支給決定 ②65歳に達する日の前日の属する年度の市町村民税非課税または"
         "生活保護 ③65歳に達する日の前日に障害支援区分2以上 ④65歳まで介護保険サービス未利用。"
         "該当すれば介護保険移行後の利用者負担が償還される"),
        ("介護保険では不足する部分があるか",
         "介護保険の支給限度基準額では必要量に満たない場合、不足分を障害福祉サービスで上乗せできる。"
         "この場合は障害福祉の見込量から全額を減らさない"),
        ("本人の意向", "本人・家族の意向を聴取したうえで支給決定を行う"),
    ]
    style_header_row(ws, r, ["確認事項", "内容", "", "", ""])
    r += 1
    for i, (k, v) in enumerate(checks):
        write_row(ws, r, [k, v, "", "", ""], alt=(i % 2 == 1))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 40
        r += 1
    return ws


# ============================================================
# 07_潜在需要
# ============================================================
def sheet_latent(wb):
    ws = add_sheet(
        wb, "07_潜在需要", "未特定の潜在需要（第3層）",
        "第1層・第2層で個別に把握できない需要をここで補正します。"
        "過去実績の伸び率を用いるのは、この層に限定してください。"
        "アンケート集計後に実数を入力します。",
        [26, 16, 16, 16, 16, 20, 40])
    style_header_row(ws, 5, ["需要の源泉", "対応するサービス", "低位", "中位", "高位",
                             "根拠となる資料", "備考"])
    r = 6
    rows = [
        ("アンケートの利用希望", "全サービス", None, None, None, "アンケート問29-2・問33",
         "利用希望をそのまま見込量に置かない。実績・支給決定・供給体制で補正する"),
        ("相談支援事業所への相談件数", "計画相談支援、各サービス", None, None, None, "相談実績（村資料待ち）",
         "相談から支給決定に至る割合を確認する"),
        ("未利用者のニーズ", "移動支援、日中一時支援等", None, None, None, "アンケート自由記述",
         "実績ゼロのサービス（05）と対応させる"),
        ("新規診断・手帳取得の傾向", "障害児通所支援、精神関係", None, None, None, "手帳交付実績（村資料待ち）",
         "精神障害者保健福祉手帳は増加傾向にある"),
        ("圏域のサービス整備予定", "全サービス", None, None, None, "圏域4町村の協議",
         "新設事業所があれば利用可能量が増える"),
        ("親亡き後・家族介護力の低下", "短期入所、共同生活援助、施設入所支援", None, None, None,
         "アンケート問8〜問11", "第2層の利用量変化としても現れる"),
    ]
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "right", "right", "right", "left", "left"],
                  numfmts=[None, None, DEC1, DEC1, DEC1, None, None],
                  fills=[None, None, COLORS["input"], COLORS["input"], COLORS["input"], None, None])
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "低位・中位・高位のシナリオはこの層にのみ適用します。"
               "第1層（個別に把握できる案件）と第2層（既存利用者の利用量変化）は、"
               "シナリオではなく確定値として扱ってください。"
               "計画書には基本ケース（第1層＋第2層＋潜在需要の中位）を記載し、"
               "低位・高位は幅として注記します。")
    ws.row_dimensions[r].height = 40
    return ws


# ============================================================
# 08_供給制約
# ============================================================
def sheet_supply(wb):
    ws = add_sheet(
        wb, "08_供給制約", "供給制約（利用希望量と利用可能量の分離）",
        "本村には障害福祉サービス事業所がなく、圏域事業所の受入可能性が見込量の上限を決めます。"
        "利用希望量が供給可能量を超える場合は、差を明示したうえで確保方策に記載してください。",
        [26, 22, 14, 14, 14, 16, 40])
    style_header_row(ws, 5, ["サービス", "利用している事業所の所在",
                             "定員", "空き", "送迎の\n可否", "受入可否", "確認事項"])
    r = 6
    rows = [
        ("生活介護", "圏域（喜多方市・会津若松市等）", None, None, None, None,
         "特別支援学校卒業者の受入枠。送迎範囲に北塩原村が含まれるか"),
        ("就労継続支援B型", "喜多方市・会津若松市等", None, None, None, None,
         "給付費が令和2年度比2.08倍。定員に対する余力を確認"),
        ("就労選択支援", "圏域（協議会設置圏域ごとに設置）", None, None, None, None,
         "圏域内に事業所があるか。ない場合の確保方策"),
        ("共同生活援助", "圏域", None, None, None, None, "空室の有無。親亡き後の受入余力"),
        ("短期入所", "圏域", None, None, None, None, "緊急時の受入可否。地域生活支援拠点と連動"),
        ("居宅介護", "圏域", None, None, None, None, "訪問可能地区。冬季の提供体制"),
        ("計画相談支援", "猪苗代町・喜多方市等", None, None, None, None,
         "相談支援専門員の担当可能件数。セルフプラン解消の可否"),
        ("児童発達支援", "圏域（村内0か所）", None, None, None, None, "通所距離と保護者の送迎負担"),
        ("放課後等デイサービス", "圏域", None, None, None, None, "長期休暇中の受入。送迎"),
        ("保育所等訪問支援", "圏域", None, None, None, None, "訪問可能範囲"),
        ("介護保険 通所介護・訪問介護", "圏域", None, None, None, None,
         "65歳移行後の受け皿。共生型サービスの指定を受けているか"),
    ]
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "right", "right", "center", "center", "left"],
                  numfmts=[None, None, INT, INT, None, None, None],
                  fills=[None, None] + [COLORS["input"]] * 4 + [None])
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    style_note(ws.cell(row=r, column=1),
               "人口減少地域では需要が減るとは限りません。利用者総数は少なくても1人当たりの支援量が大きく、"
               "事業所撤退の影響が大きく、村外・圏域外利用への依存度が高く、"
               "送迎・人材・移動時間が供給制約になります。"
               "「利用人数の推計」と「実際に利用可能な供給量」を分けて管理してください。")
    ws.row_dimensions[r].height = 40
    return ws


# ============================================================
# 09_地域生活支援事業
# ============================================================
def sheet_chiiki(wb):
    ws = add_sheet(
        wb, "09_地域生活支援事業", "地域生活支援事業の見込量",
        "村が実施する事業です。移動支援事業は令和3〜8年度の全期間で見込量0人・実績0人であり、"
        "地区分散・冬季移動という本村の地域特性と整合しません。"
        "地域活動支援センター機能強化事業は実績3人が見込量2人を上回っています。",
        [30, 12, 14, 14, 14, 14, 44])
    style_header_row(ws, 5, ["事業", "単位", "令和5年度\n実績", "令和8年度\n見込（現行）",
                             "令和9〜11年度\n実施目標", "", "確認事項"])
    r = 6
    rows = [
        ("障がい者相談支援事業", "か所", 2, 2, None, "", "村住民課・地域生活支援センターいなわしろ"),
        ("成年後見制度利用支援事業", "人", 1, 5, None, "", "見込5人に対し実績1人。乖離の要因を確認"),
        ("意思疎通支援事業", "人", 0, 0, None, "", "手話通訳者派遣の需要の有無"),
        ("日常生活用具給付等事業", "件", 69, 72, None, "", "排泄管理支援用具が大半。増加傾向"),
        ("移動支援事業", "人", 0, 0, None, "",
         "制度未周知／家族送迎等での代替／ニーズ不在のいずれかを確認。実績ゼロで据え置かない"),
        ("日中一時支援事業", "人", 2, 2, None, "", "家族レスパイト。障がい児と接続。増加傾向"),
        ("地域活動支援センター機能強化事業", "人", 3, 2, None, "",
         "実績が見込量を上回っている。次期見込量に実績を反映する"),
        ("訪問入浴サービス事業", "人", 0, 0, None, "", "重度者の在宅生活支援"),
        ("理解促進研修・啓発事業", "件", 0, 0, None, "", "実施方法と財源を確認"),
        ("自発的活動支援事業", "件", 0, 0, None, "", "障がい者団体・家族会の活動支援"),
    ]
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "center", "right", "right", "right", "left", "left"],
                  numfmts=[None, None, INT, INT, INT, None, None],
                  fills=[None, None, None, None, COLORS["input"], None, None])
        ws.row_dimensions[r].height = 28
        r += 1
    return ws


# ============================================================
# 10_村確認事項
# ============================================================
def sheet_confirm(wb):
    ws = add_sheet(
        wb, "10_村確認事項", "村への確認・依頼事項（見込量推計関係）",
        "個別積上げ方式に切り替えるために必要な資料です。1番が揃わないと本ブックは動きません。",
        [30, 52, 24, 12, 16])
    style_header_row(ws, 5, ["確認事項", "確認したい内容", "反映先", "優先度", "状態"])
    rows = [
        ("年齢別・サービス別の匿名利用者一覧",
         "利用者番号、生年月（年度でも可）、障害支援区分、利用中のサービス、月利用日数・時間。氏名は不要。",
         "02_年齢到達者一覧／03・04", "最優先", "村資料待ち"),
        ("令和9〜11年度の65歳到達者",
         "到達年度と現在利用しているサービス、月量。",
         "02／06_65歳移行判定", "高", "村資料待ち"),
        ("新高額障害福祉サービス等給付費の要件該当見込み",
         "65歳到達予定者について、5年要件・市町村民税非課税・障害支援区分2以上の該当見込み。",
         "06_65歳移行判定", "中", "村資料待ち"),
        ("特別支援学校・特別支援学級の在籍者",
         "学年、卒業予定年度、卒業後の進路の見通し。令和6〜7年度に卒業済みの方の現在の利用状況も含む。",
         "02／03_生活介護・就労系", "高", "村資料待ち"),
        ("障害児通所支援の利用児童の学年",
         "児童発達支援・放課後等デイサービスの利用児童の学年と、就学予定年度。",
         "04_積上げ_障害児通所支援", "高", "村資料待ち"),
        ("令和6・7年度の障害児給付費の変動理由",
         "児童発達支援が23件から1件へ減り、放課後等デイサービスが12件から27件へ増えた理由。"
         "受給者番号ベースで同一児童の就学移行かどうかを確認する。",
         "04／計画本文の記述", "高", "村資料待ち"),
        ("障害児入所施設の利用者の有無",
         "18歳以上の在籍を含む。県の移行調整の協議の場での取扱い。",
         "02／03_施設入所支援", "中", "村資料待ち"),
        ("40歳以上65歳未満の特定疾病該当者",
         "16の特定疾病に該当し、要介護認定を受けているまたは申請予定の方の有無。",
         "02／03", "中", "村資料待ち"),
        ("圏域事業所の受入可能性",
         "定員、空き、送迎範囲、受入可否。介護保険事業所の共生型サービス指定の有無を含む。",
         "08_供給制約", "高", "村資料待ち"),
        ("過去3年間の65歳到達による移行実績",
         "人数、移行先、障害福祉を継続した方の人数と理由。",
         "06_65歳移行判定", "中", "村資料待ち"),
        ("支給決定者数（サービス別・実人数）",
         "請求件数ではなく実人数。計画相談支援は請求が計画作成・モニタリング時のみのため特に必要。",
         "03・04の令和8年度基準", "最優先", "村資料待ち"),
    ]
    r = 6
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1),
                  aligns=["left", "left", "left", "center", "center"])
        style_status(ws.cell(row=r, column=5))
        c = ws.cell(row=r, column=4)
        if c.value == "最優先":
            c.fill = PatternFill("solid", fgColor="C00000")
            c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 42
        r += 1
    return ws


# ============================================================
# 11_告示別表第一根拠
# ============================================================
# (サービス, 別表第一が勘案を求める要素, 本ブックでの反映先)
BEPPYO1_ROWS = [
    ("居宅介護", "現に利用している者の数／ニーズ／施設入所者の地域生活への移行者数／"
     "入院中の精神障がい者のうち移行後に利用が見込まれる者の数／平均的な一人当たり利用量",
     "03の継続・新規・地域移行行、1人当たり月利用量"),
    ("重度訪問介護", "同上", "03"),
    ("同行援護", "現に利用している者の数／ニーズ／平均的な一人当たり利用量"
     "（地域移行・精神科病院からの流入は勘案要素に含まれない）", "03"),
    ("行動援護", "居宅介護と同じ", "03"),
    ("重度障害者等包括支援", "居宅介護と同じ", "03"),
    ("生活介護", "居宅介護と同じ。加えて強度行動障害の状態にある者・高次脳機能障害を有する者・"
     "医療的ケアを必要とする者について個別に利用者数の見込みを設定するよう努める",
     "03。重度者は内数として別掲（村資料受領後）"),
    ("自立訓練（機能訓練）", "現に利用している者の数／ニーズ／施設入所者の地域移行者数／"
     "平均的な一人当たり利用量", "05_実績ゼロサービス"),
    ("自立訓練（生活訓練）", "居宅介護と同じ", "05_実績ゼロサービス"),
    ("就労選択支援", "ニーズ／特別支援学校卒業者数／就労移行支援・A型・B型を新たに利用する者の数／"
     "現に利用している者の数", "02_年齢到達者一覧（18歳）／03"),
    ("就労移行支援", "居宅介護と同じ。加えて一般就労への移行者数／特別支援学校卒業者／"
     "復職を希望する休職者等", "02（18歳）／03"),
    ("就労継続支援A型", "居宅介護と同じ。加えて一般就労への移行者数／地域の雇用情勢", "03"),
    ("就労継続支援B型", "居宅介護と同じ。加えて一般就労への移行者数。"
     "区域内のB型事業所の平均工賃月額について区域ごとの目標水準を設定することが望ましい",
     "03。工賃目標は圏域協議事項"),
    ("就労定着支援", "現に利用している者の数／ニーズ／一般就労への移行者数", "03"),
    ("療養介護", "現に利用している者の数／ニーズ", "05_実績ゼロサービス"),
    ("短期入所（福祉型・医療型）", "居宅介護と同じ。加えて重度障がい者について個別に"
     "利用者数の見込みを設定するよう努める", "03。重度者は内数として別掲"),
    ("自立生活援助", "現に利用している者の数／同居している家族による支援を受けられない障がい者の数／"
     "施設入所者の地域移行者数／入院中の精神障がい者のうち移行後の利用見込", "05_実績ゼロサービス"),
    ("共同生活援助", "居宅介護と同じ。加えて一人暮らしや家庭からGHに入所する者の数／"
     "GHから退所する者の数／重度障がい者の個別見込み", "03。退所者は流出行"),
    ("施設入所支援", "令和7年度末時点の施設入所者数を基礎とし、地域生活への移行者数を控除の上、"
     "GH等での対応が困難な者の利用といった真に必要と判断される数を加えた数。"
     "5%以上削減と未達分の上乗せ。居室の個室化等の取組状況の把握", "03。基準はR7年度末"),
    ("計画相談支援", "現に利用している者の数／ニーズ／入院中の精神障がい者のうち移行後の利用見込",
     "03。請求件数ではなく実人数"),
    ("地域移行支援", "現に利用している者の数／ニーズ／施設入所者の地域移行者数／"
     "入院中の精神障がい者のうち移行後の利用見込。"
     "入所又は入院前の居住地を有する市町村が対象者数の見込みを設定する",
     "05_実績ゼロサービス。帰属ルールに注意"),
    ("地域定着支援", "現に利用している者の数／単身世帯である障がい者の数／"
     "同居している家族による支援を受けられない障がい者の数／施設入所者の地域移行者数",
     "05_実績ゼロサービス"),
    ("児童発達支援", "地域における児童の数の推移／現に利用している障がい児の数／ニーズ／"
     "重症心身障がい児等のニーズ／医療的ケア児等のニーズ／保育所・認定こども園・幼稚園での受入状況／"
     "入所施設から退所した後に利用が見込まれる障がい児の数／平均的な一人当たり利用量",
     "04。就学による流出は02（6歳）"),
    ("放課後等デイサービス", "同上（受入状況は放課後児童健全育成事業・小中学校・特別支援学校）",
     "04。就学による流入は02（6歳）"),
    ("保育所等訪問支援", "同上（受入状況は保育所・認定こども園・幼稚園・小学校・特別支援学校）",
     "05_実績ゼロサービス"),
    ("居宅訪問型児童発達支援", "地域における児童の数の推移／現に利用している障がい児の数／"
     "ニーズ／重症心身障がい児等のニーズ／医療的ケア児等のニーズ／平均的な一人当たり利用量",
     "05_実績ゼロサービス"),
    ("障害児入所支援（福祉型・医療型）", "地域における児童の数の推移／現に利用している障がい児の数／"
     "ニーズ／重症心身障がい児等のニーズ／医療的ケア児等のニーズ", "10_村確認事項（県所管）"),
    ("障害児相談支援", "同上", "04"),
    ("医療的ケア児等コーディネーター（市町村）", "地域における医療的ケア児等のニーズ等を勘案して"
     "必要となる配置人数の見込みを設定する", "計画本文 第4章(4)"),
]


def sheet_beppyo1(wb):
    ws = add_sheet(
        wb, "11_告示別表第一根拠", "基本指針 別表第一が勘案を求める要素（サービス別）",
        "令和8年3月31日 こども家庭庁・厚生労働省告示第4号による改正後の別表第一による。"
        "本ブックの積上げ項目が告示の要求を満たしているかの対照表。"
        "「努める」とあるものは努力義務、それ以外は算定に当たって勘案すべき要素。",
        [26, 62, 30])
    style_header_row(ws, 5, ["サービス", "別表第一が勘案を求める要素", "本ブックでの反映先"])
    r = 6
    for i, rec in enumerate(BEPPYO1_ROWS):
        write_row(ws, r, list(rec), alt=(i % 2 == 1), aligns=["left", "left", "left"])
        ws.row_dimensions[r].height = 42
        r += 1
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
    style_note(ws.cell(row=r + 1, column=1),
               "※ 別表第一の標準文言は「…等を勘案して、利用者数及び量の見込みを設定する」であり、"
               "利用者数と量（＝利用者数×平均的な一人当たり利用量）を分けて設定することを求めている。"
               "本ブック03・04の構成はこれに対応している。")
    ws.row_dimensions[r + 1].height = 30
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=3)
    style_note(ws.cell(row=r + 2, column=1),
               "※ 重度障がい者（強度行動障害の状態にある者・高次脳機能障害を有する者・"
               "医療的ケアを必要とする者等）の個別見込みを求めているのは、"
               "生活介護・短期入所・共同生活援助の3サービスに限られる（いずれも努力義務）。")
    ws.row_dimensions[r + 2].height = 30
    return ws


# ============================================================
# 12_別表第五判定
# ============================================================
def sheet_beppyo5(wb):
    ws = add_sheet(
        wb, "12_別表第五判定", "基本指針 別表第五（地域差是正）の適用判定",
        "第8期で新設。要件に該当する市町村は、全国の伸び率を用いた算定方法を基本とすることになる。"
        "本村は過疎法第2条により全域が過疎地域に該当するため（福島県「県内の過疎・中山間地域の指定状況」"
        "令和7年4月1日現在）、同表一の項⑴を満たさず適用対象外。個別積上げ方式を原則どおり採用できる。",
        [28, 24, 34, 52])

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    style_title(ws.cell(row=5, column=1), "■ 条文（第三の二の2(一)ただし書及び別表第五）",
                fill=COLORS["subhead"], size=11)
    texts = [
        "別表第五 一の項　⑴ 全部過疎市町村ではないこと　かつ　"
        "⑵ 当該サービスの、人口に占めるサービス利用者割合が、⑴を満たす市町村の上位25%の範囲内であること",
        "別表第五 二の項　当該市町村の当該サービスの利用者数の伸び率が全国の伸び率を上回る場合は、"
        "令和9〜11年度の見込みを、令和8年度の見込みと全国の伸び率を用いて定める",
        "備考　全部過疎市町村＝過疎地域の持続的発展の支援に関する特別措置法（令和3年法律第19号）の"
        "過疎地域のうち過疎区分が全部過疎である市町村／人口＝特定障害福祉サービスは18歳以上人口、"
        "特定障害児通所支援は18歳未満人口／伸び率＝令和5年度から令和7年度までの幾何平均／"
        "この方法は全国の伸び率が正の値であるときに限る",
        "第三の二の2(一)　⑵ 地域の具体的なニーズ調査を踏まえ、市町村障害福祉計画等において、"
        "⑴と異なる算定方法をその必要性及び根拠を添えて示す場合には、当該算定方法",
    ]
    r = 6
    for t in texts:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        style_note(ws.cell(row=r, column=1), t)
        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_title(ws.cell(row=r, column=1), "■ 判定", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["判定項目", "入力", "判定", "根拠・備考"])
    head = r
    r += 1

    # ⑴ 過疎区分
    write_row(ws, r, ["⑴ 本村の過疎区分", "全部過疎（確定）", None,
                      "福島県「県内の過疎・中山間地域の指定状況」（令和7年4月1日現在）において、"
                      "北塩原村は過疎法第2条により全域が過疎地域に該当する市町村（◎）と表示されている。"
                      "旧北山村・旧大塩村・旧檜原村の3区域とも過疎地域・振興山村地域・特定農山村地域に該当。"],
              aligns=["left", "center", "center", "left"])
    ws.cell(row=r, column=3).value = f'=IF(B{r}="全部過疎（確定）","要件⑴を満たさない→別表第五 適用外","要確認")'
    style_data_cell(ws.cell(row=r, column=3), align="center")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=COLORS["calc"])
    ws.row_dimensions[r].height = 46
    row_kaso = r
    r += 1

    # ⑵ サービス別の利用者割合
    write_row(ws, r, ["⑵ サービス利用者割合", "判定不要", "―",
                      "18歳以上人口（障害児通所支援は18歳未満人口）に占める利用者割合が、"
                      "全部過疎でない市町村の上位25%以内かどうか。"
                      "⑴を満たさないため判定不要。"],
              alt=True, aligns=["left", "center", "center", "left"])
    ws.row_dimensions[r].height = 46
    r += 1

    write_row(ws, r, ["結論", None, None,
                      "⑴かつ⑵の連言であるため、⑴を満たさない本村には別表第五 二の項の算定方法は適用されない。"
                      "第三の二の2(一)本文の一般則により、実績分析・意向・心身の状況を勘案して設定する。"],
              aligns=["left", "center", "center", "left"])
    ws.cell(row=r, column=3).value = (
        f'=IF(C{row_kaso}="要件⑴を満たさない→別表第五 適用外",'
        f'"個別積上げ方式を採用可（第三の二の2(一)本文による）","要判定")')
    style_data_cell(ws.cell(row=r, column=3), align="center")
    ws.cell(row=r, column=3).font = Font(name=FONT, size=10, bold=True)
    ws.row_dimensions[r].height = 46
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_note(ws.cell(row=r, column=1),
               "※ あわせて本村は福島県過疎・中山間地域振興条例の中山間地域（全域）にも該当する。"
               "基本指針 第一の一の3は「中山間・人口減少地域においては、共生型サービスや基準該当障害福祉サービス、"
               "多機能型、従たる事業所等の現行制度の活用等も図りつつ、サービス提供体制を維持・確保していくことが重要」"
               "としており、見直し事項⑩「人口減少地域におけるサービスの維持・確保」に正面から該当する。")
    ws.row_dimensions[r].height = 34
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_note(ws.cell(row=r, column=1),
               "※ 別表第五が対象とするのは「特定障害福祉サービス」及び「特定障害児通所支援」であり、"
               "全サービスではない。該当サービスの範囲は障害者総合支援法第88条第10項及び"
               "児童福祉法第21条の5の15第2項の定義による。")
    ws.row_dimensions[r].height = 34

    assert ws.cell(row=head, column=1).value == "判定項目", "12_別表第五判定 の見出し位置がずれています"
    return ws


def verify():
    """行構成とサービス定義の整合を確認する。"""
    for rows, label in ((FLOW_ROWS_ADULT, "成人"), (FLOW_ROWS_CHILD, "障がい児")):
        labels = [k for k, m, n in rows]
        for required in ("継続利用者", "＝　利用者数", "　　1人当たり月利用量", "＝　月間サービス量"):
            assert required in labels, f"{label}の行構成に{required}がありません"
        assert any(k.startswith("＋") for k in labels), f"{label}に流入行がありません"
        assert any(k.startswith("－") for k in labels), f"{label}に流出行がありません"
    names = [s[0] for s in SERVICES_ADULT] + [s[0] for s in SERVICES_CHILD] + \
            [s[0] for s in SERVICES_ZERO]
    assert len(names) == len(set(names)), "サービス名が重複しています"
    # 65歳移行判定の区分が想定の3種類に収まっているか
    kubun = {k for _, k, _, _ in KAIGO_SOTO}
    assert kubun <= {"相当あり", "相当なし", "一部あり"}, f"想定外の区分: {kubun}"
    print(f"  自己検証: 増減行構成・サービス名の一意性（{len(names)}件）・65歳区分 いずれも整合")


def main():
    verify()
    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb)
    sheet_method(wb)
    sheet_age_roster(wb)
    sheet_buildup_adult(wb)
    sheet_buildup_child(wb)
    sheet_zero(wb)
    sheet_age65(wb)
    sheet_latent(wb)
    sheet_supply(wb)
    sheet_chiiki(wb)
    sheet_confirm(wb)
    sheet_beppyo1(wb)
    sheet_beppyo5(wb)
    wb.save(OUT_FILE)
    print(f"作成: {OUT_FILE}")


if __name__ == "__main__":
    main()
