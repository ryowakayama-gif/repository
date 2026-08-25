# -*- coding: utf-8 -*-
"""発注者ご作成のキックオフ会議議事録の校正.

原本
  99672316-_________10_____________________1.odt
  （大雪地区広域連合 第10期 介護保険事業計画策定支援業務　キックオフ会議
    令和8年8月6日開催）

誤字脱字及び明らかな表記の誤りを反映した版を作成する。
原本の書式を保つため、ODF の content.xml の文字列のみを置換し、
段落・表・スタイルには手を加えない。

判断を要する事項（文体・会議名の呼称・実質の論点）は反映せず、
校正結果の一覧に「提案」「要協議」として掲げる。

出力
  output/第10期計画_キックオフ会議議事録_校正反映.odt
  output/第10期計画_キックオフ会議議事録の校正結果.xlsx
"""

import os
import shutil
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = ("/root/.claude/uploads/54f527c9-842b-534d-b822-e5a6c91f837c/"
       "99672316-_________10_____________________1.odt")
OUT = ("/home/user/repository/output/"
       "第10期計画_キックオフ会議議事録_校正反映.odt")
OUTX = ("/home/user/repository/output/"
        "第10期計画_キックオフ会議議事録の校正結果.xlsx")

# ------------------------------------------------------------ 反映する校正
# (区分, 箇所, 置換前, 置換後, 理由)
FIX = [
    ("誤字", "2. 基礎データ",
     "前回の保険慮", "前回の保険料",
     "「保険慮」は「保険料」の誤りである。"),
    ("誤字", "2. 既存調査／4. 令和8年9月",
     "居宅変更実態調査", "居所変更実態調査",
     "調査の正式名称は「居所変更実態調査」である。2か所。"),
    ("表記", "2. 計画の位置付け／基礎データ、3-2",
     "大雪広域連合", "大雪地区広域連合",
     "正式名称は「大雪地区広域連合」である。3か所。"),
]

# 構造を伴う置換（XMLの断片を直接置き換える）
FIX_XML = [
    ("表記", "3. 議題別整理",
     '<text:span text:style-name="T128">3. 議題別整理</text:span>'
     '<text:span text:style-name="T129">。</text:span>',
     '<text:span text:style-name="T128">3. 議題別整理</text:span>',
     "見出しの末尾に句点が付いている。他の見出しには付いていない。"),
    ("脱字", "4. 令和8年10月　主な成果・判断",
     '<text:span text:style-name="T390">案、骨子・KPI目標、町別論点</text:span>',
     '<text:span text:style-name="T390">案）、骨子・KPI目標、町別論点</text:span>',
     "「推計（案」の括弧が閉じていない。"),
    ("誤記", "3-4. 最終報告・答申　想定内容",
     '<text:p text:style-name="P306">パブコメ反映、/計画原案・保険料方向性'
     '</text:p><text:p text:style-name="P307">第10計画（最終案）作成</text:p>',
     '<text:p text:style-name="P306">パブコメ反映、計画原案・保険料の方向性'
     '</text:p><text:p text:style-name="P307">第10期計画（最終案）作成</text:p>',
     "不要な「/」が混入している。また「第10計画」は「第10期計画」の脱字である。"),
    ("誤記", "2. 既存調査",
     '<text:span text:style-name="T59">令和7年度の介護予防・'
     '日常生活圏域ニーズ調査及び</text:span>'
     '<text:span text:style-name="T60">、実施済みの</text:span>',
     '<text:span text:style-name="T59">令和7年度に実施した</text:span>'
     '<text:span text:style-name="T60"></text:span>',
     "「介護予防・日常生活圏域ニーズ調査」と「健康とくらしの調査"
     "（JAGES調査）」は同一の調査である。"
     "「及び」で並べると別の調査を2件実施したと読めるため、"
     "④に統合する。"),
    ("誤記", "2. 既存調査",
     '<text:span text:style-name="T63">生活改善調査、②居所変更実態調査、'
     '③介護人材実態調査、④健康とくらしの調査（JAGES調査）</text:span>',
     '<text:span text:style-name="T63">生活改善調査、②居所変更実態調査、'
     '③介護人材実態調査及び④介護予防・日常生活圏域ニーズ調査'
     '（健康とくらしの調査、JAGES調査）</text:span>',
     "上に同じ。④に調査名を統合する。"),
    ("表記", "3-4. 最終報告・答申　時期・状態",
     '<text:span text:style-name="T314">19日事務管理者・主監</text:span>',
     '<text:span text:style-name="T314">19日、事務管理者・主監</text:span>',
     "日付と会議名が続けて書かれており、区切りがない。"),
    ("表記", "3-4. 成案完成　時期・状態",
     '<text:p text:style-name="P324">令和9年3月2日全員協議会情報提供</text:p>'
     '<text:p text:style-name="P325">令和9年3月3日第10期介護保険事業計画決定'
     '（条例改正提出準備）</text:p>'
     '<text:p text:style-name="P326">令和9年3月23日広域連合会議</text:p>',
     '<text:p text:style-name="P324">令和9年3月2日　全員協議会情報提供</text:p>'
     '<text:p text:style-name="P325">令和9年3月3日　第10期介護保険事業計画決定'
     '（条例改正提出準備）</text:p>'
     '<text:p text:style-name="P326">令和9年3月23日　広域連合会議</text:p>',
     "日付と会議名が続けて書かれており、区切りがない。3行とも。"),

    # ---------------------------------------- 令和8年8月6日のご指示による反映
    ("表記", "2. 計画の位置付け",
     "介護人材、保険事業運営を持続可能なものとする",
     "介護人材、介護保険事業の運営を持続可能なものとする",
     "「保険事業運営」は生命保険等で用いる語である。"
     "「福祉の中でも上位に位置する計画」はご指示によりそのままとした。"),
    ("表記", "5. ③ KPI・役割分担",
     "本文KPI12＋補完4、5基本目標、広域連合と3町の責任分担",
     "代表KPI16項目（うち補完4項目）、5基本目標、"
     "広域連合と3町の責任分担。"
     "アンケートは現在実施したもので完了のため、"
     "現在のデータでは算定できない4項目（H07・H08・H12・H16）は、"
     "代理指標への振替え又は項目の削除を検討する",
     "計画骨子案及びKPI定義書の表記に揃えた。"
     "あわせて、算定できない4項目の扱いをご指示により追記した。"),
    ("表記", "2. 重点論点",
     "３町の議員が同席する中で報告を行うため、３町の特異性を平準化させる"
     "ことに留意する。",
     "３町の議員が同席する中で報告を行うため、"
     "３町の差を把握したうえで、広域として共通の施策体系に整理する"
     "ことに留意する。",
     "「特異性を平準化」は、地域差を均すとも様式を揃えるとも読める。"
     "地域差は把握して施策に反映する対象であるため、意味を特定した。"),
    ("要協議①", "3-1. 現状分析・将来推計",
     '<text:span text:style-name="T136">人口推計は</text:span>'
     '<text:span text:style-name="T137">地方創生計画による人口推計の値を使用。'
     '社人研の自然推計を元にした</text:span>',
     '<text:span text:style-name="T136">人口推計は</text:span>'
     '<text:span text:style-name="T137">、３町の地方創生総合戦略の人口推計を'
     'ベースに補正を行う。社人研の自然推計を元にした</text:span>',
     "会議での発言をそのまま記載したものであり、"
     "方針は「総合戦略をベースに補正を行う」であるとのご指示による。"
     "2. 基礎データの「人口推計のカーブを参考に補正を行うことを確認」"
     "と表現を揃えた。"),
    ("要協議②", "3-1. 現状分析・将来推計",
     '<text:span text:style-name="T139">との乖離を確認する。</text:span>',
     '<text:span text:style-name="T139">との乖離を確認する。'
     '乖離の確認は、総人口ではなく65歳以上及び75歳以上について行う。'
     '</text:span>',
     "給付費に効くのは65歳以上・75歳以上である。"
     "総人口では住民基本台帳が社人研推計を622人上回る一方、"
     "65歳以上では144人、75歳以上では256人下回っており、向きが逆である。"),
    ("要協議③", "3-3. 本文",
     '<text:span text:style-name="T154">地方創生計画の人口推計を元に、'
     '</text:span><text:span text:style-name="T155">令和8年実績で再基準化'
     '</text:span><text:span text:style-name="T156">します。その上で、'
     '需要上振れ、供給制約、基金取崩し、報酬・制度改正等の複数ケースを設定し、'
     '採用値を協議します。\u00a0</text:span>',
     '<text:span text:style-name="T154">地方創生総合戦略の人口推計を'
     'ベースに、</text:span><text:span text:style-name="T155">年報・月報の'
     '実績で再基準化</text:span><text:span text:style-name="T156">する。'
     'その上で、需要上振れ、供給制約、基金取崩し、報酬・制度改正等の'
     '複数ケースを設定し、採用値を協議する。\u00a0</text:span>',
     "「令和8年実績」を「年報・月報」に改めるとのご指示による。"
     "あわせて、3-1・3-2が常体であるため文体を揃えた。"),
    ("要協議③", "3-3. 手順2",
     "令和8年実績で人口・認定・受給・給付を再基準化",
     "年報・月報の実績で人口・認定・受給・給付を再基準化",
     "上に同じ。"),
    ("要協議③", "4. 令和8年10月",
     "令和8年実績で再基準化、供給制約分析",
     "年報・月報の実績で再基準化、供給制約分析",
     "上に同じ。"),
    ("要協議③", "5. ④ 再基準化・シナリオ",
     "令和8年実績を起点に自然体・需要上振れ・供給制約・基金反映を比較",
     "年報・月報の実績を起点に自然体・需要上振れ・供給制約・基金反映を比較",
     "上に同じ。"),
    ("要協議④", "2. 保険料",
     '<text:p text:style-name="P105">米の価格が上昇し、農業収入が高かったため、'
     '想定以上の基金の取り崩しはなかった。前期計画を受け、'
     'それほど基金を下げることはないと現段階で推察。</text:p>',
     '<text:p text:style-name="P105">米の価格が上昇し、農業収入は高かった。'
     'また、基金の取り崩しは想定を下回った。前期計画を受け、'
     'それほど基金を下げることはないと現段階で推察。</text:p>',
     "農業収入が高かったことは事実である一方、"
     "それが基金の取り崩しが少なかった理由であるかは確かめていないため、"
     "因果を切り離し、事実の記載にとどめるとのご指示による。"),
    ("要協議⑥", "5. ① 調査データの受渡し",
     "事業所実態調査の回答データの受渡し方法、集計単位、公表単位",
     "事業所実態調査の回答データは受領済み。集計単位、公表単位を確認する",
     "令和8年8月6日に「現時点で提出いただいている調査結果のすべて」として"
     "受領済みであるとのご指示による。"),
    ("表記", "3-4. 10月構成町意見照会",
     '<text:p text:style-name="P267">10月</text:p>'
     '<text:p text:style-name="P268">'
     '<text:span text:style-name="T269">構成町意見照会</text:span></text:p>',
     '<text:p text:style-name="P267">第1回協議会</text:p>'
     '<text:p text:style-name="P268">'
     '<text:span text:style-name="T269">構成町意見照会</text:span></text:p>',
     "段階名が「10月」、時期欄が「10〜11月頃」で食い違っていた。"
     "また第2回の前に第1回の記載がなかった。"
     "調査結果報告・現状分析・骨子案を扱う回であり、"
     "第1回協議会に当たるものとして段階名を改めた。"),
    ("表記", "3-4. パブリックコメント",
     '<text:span text:style-name="T292">概要版を用い、パブリックコメント'
     '</text:span><text:span text:style-name="T293">を令和9年１月に、2週間、'
     '</text:span><text:span text:style-name="T294">WEBにて</text:span>'
     '<text:span text:style-name="T295">実施。その他、</text:span>',
     '<text:span text:style-name="T292">概要版を用いたパブリックコメント'
     '</text:span><text:span text:style-name="T293">の実施。'
     '</text:span><text:span text:style-name="T294"></text:span>'
     '<text:span text:style-name="T295">あわせて、</text:span>',
     "想定内容欄と時期・状態欄に「令和9年1月に概要版を用いWEBにて2週間実施」"
     "が重複して記載されていた。想定内容欄から日程を外した。"),
    ("表記", "4. 令和8年8月上旬",
     "計画素案、第1回委員会資料、KPI定義書",
     "計画素案、第1回協議会資料、KPI定義書",
     "「委員会」と「協議会」が混在していたため「協議会」に統一した。"
     "正式名称は要確認である（当方の成果品では「策定委員会」と表記している）。"),
    ("表記", "5. ⑤ 会議・公開",
     "委員会の審議順、町別値・事業所情報の公開範囲",
     "協議会の審議順、町別値・事業所情報の公開範囲",
     "上に同じ。"),
    ("表記", "2. 基礎データ",
     "３町の地方創生計画にある人口推計のカーブ",
     "３町の地方創生総合戦略にある人口推計のカーブ",
     "3-1及び3-3を「地方創生総合戦略」に改めたため、"
     "文書内の呼称を揃えた。"
     "3町の計画の正式名称は「第3期（写真文化首都／美瑛町）"
     "まち・ひと・しごと創生総合戦略」である。"),
]

# ------------------------------------------------------------ 反映しない指摘
PROPOSE = [
    ("見送り", "2. 計画の位置付け",
     "福祉の中でも上位に位置する計画",
     "（ご指示によりそのまま）",
     "介護保険事業計画は、老人福祉計画と一体のものとして作成し、"
     "市町村地域福祉計画等と調和が保たれたものとされる"
     "（介護保険法第117条）。"
     "「福祉の中でも上位」とすると3町の地域福祉計画との関係を"
     "説明しにくくなるが、令和8年8月6日のご指示により原本のままとした。"
     "計画本文（第1章第1節）に同じ表現を用いる場合は、"
     "改めてご相談したい。"),
    ("解決済", "3-4. 最終報告・答申",
     "「事務管理者・主監会議」",
     "原本のまま（修正不要）",
     "令和8年8月6日のご指示により、発注者側の資料からの抜粋であることを"
     "確認した。表記はこのままとする。"),
    ("要確認", "3-4／4／5",
     "「協議会」と「委員会」の呼称",
     "議事録内は「協議会」に統一",
     "議事録内では3-4の表題（推進協議会）に合わせて「協議会」に統一した。"
     "ただし当方の成果品では「策定委員会」と表記しており、"
     "正式名称の確認を要する。"
     "確定後、当方の成果品側も揃える。"),
]

KYOGI = [
    ("反映済①", "3-1. 現状分析・将来推計",
     "総合戦略をベースに補正を行う",
     "ご指示のとおり「３町の地方創生総合戦略の人口推計をベースに"
     "補正を行う」に改めた。"
     "町ごとの扱いは、補正の作業の中で次のとおり整理する。"
     "第3期総合戦略の総人口目標は東神楽町が令和11年度9,500人維持、"
     "東川町が8,635人、美瑛町は設定なしである。"
     "令和7年国勢調査（速報）は東神楽町9,588人・東川町8,726人・"
     "美瑛町9,337人で、東川町は目標を91人上回る一方、"
     "東神楽町は年▲1.09％のペースで令和8年中に9,500人を割る見込みである。"
     "美瑛町はベースとする目標値がないため、国勢調査及び社人研推計による。"),
    ("反映済②", "3-1. 現状分析・将来推計",
     "乖離の確認は65歳以上・75歳以上で行う",
     "ご指示のとおり本文に追記した。"
     "社人研推計（見える化A1）の令和7年値と国勢調査（速報）の実績の差は、"
     "東川町＋513人・美瑛町＋444人・東神楽町▲408人で町ごとに向きが異なる。"
     "一方、住民基本台帳（令和8年）との突合では、"
     "総人口は住基が622人多いのに、65歳以上は144人、75歳以上は256人少ない。"
     "（第10期計画_世帯構成の突合.xlsx 06シート）"),
    ("反映済③", "3-3／4／5",
     "「令和8年実績」を「年報・月報」に",
     "ご指示のとおり4か所を改めた。"
     "令和8年度は第9期の最終年度であり実績の確定は令和9年3月末となるため、"
     "8〜9月の再基準化には年報・月報の直近実績を用いる。"),
    ("反映済④", "2. 保険料",
     "農業収入と基金の因果を切り離す",
     "ご指示のとおり「米の価格が上昇し、農業収入は高かった。"
     "また、基金の取り崩しは想定を下回った。」に改めた。"
     "なお現行（第9期）の基準額は月額6,400円、"
     "見える化システムの自然体推計による第10期の3年平均は6,238円"
     "（基金取崩0円・予定収納率99.0％）で、現行を162円下回る。"
     "据え置きは現時点で無理のない目標である。"
     "第3段階（給付費・保険料）の推計は、単価・基金残高・収納率・"
     "所得段階別被保険者数の受領後に確定する。"),
    ("反映済⑤", "5. ③ KPI・役割分担",
     "代理指標への振替え又は項目の削除を検討",
     "ご指示のとおり追記した。"
     "対象は代表KPIのうちH07（主介護者の高負担割合）・"
     "H08（介護離職懸念）・H12（必要サービス未充足率）・"
     "H16（災害・感染症時のサービス継続率）の4項目である。"
     "H07・H08は在宅介護実態調査、"
     "H12は事業所調査の受入困難の設問、"
     "H16は同調査のBCPの設問を要するが、いずれも実施していない。"
     "振替えの候補は、H12を居所変更実態調査の"
     "「供給不足を理由とする住替え」、"
     "H16をBCP策定率（公表システムから把握可）とすることが考えられる。"
     "次回の打合せまでに振替え案を整理して提示する。"),
    ("反映済⑥", "5. ① 調査データの受渡し",
     "受領済みへ修正",
     "ご指示のとおり「回答データは受領済み。集計単位、公表単位を確認する」"
     "に改めた。"
     "受領した回答は事業所票27件・職員個票317人・訪問系職員票26件である。"
     "訪問系は13事業所のうち3事業所からの回答であり、"
     "残る10事業所は介護サービス情報公表システムの個別公表画面により"
     "把握済みである（訪問介護員等は13事業所で181人）。"),
]


def main():
    with zipfile.ZipFile(SRC) as z:
        names = z.namelist()
        data = {n: z.read(n) for n in names}

    x = data["content.xml"].decode("utf-8")
    done = {}

    for _k, _w, old, new, _r in FIX:
        n = x.count(old)
        if n == 0:
            raise RuntimeError("該当なし: %r" % old)
        x = x.replace(old, new)
        done[old] = n

    for _k, _w, old, new, _r in FIX_XML:
        n = x.count(old)
        if n == 0:
            raise RuntimeError("該当なし（XML）: %r" % old[:60])
        x = x.replace(old, new)
        done[old[:40]] = n

    data["content.xml"] = x.encode("utf-8")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    if os.path.exists(OUT):
        os.remove(OUT)
    # mimetype は無圧縮で先頭に置く（ODF の要件）
    with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(zipfile.ZipInfo("mimetype"), data["mimetype"],
                   compress_type=zipfile.ZIP_STORED)
        for n in names:
            if n == "mimetype":
                continue
            z.writestr(n, data[n])

    # 原本から引き継いだ文書プロパティ（他団体名・作成ソフト名）を整える
    from docmeta import clean_odt
    clean_odt(OUT,
              title="大雪地区広域連合　第10期介護保険事業計画"
                    "　キックオフ会議議事録（令和8年8月6日開催）",
              subject="キックオフ会議の議事録")

    print("saved:", OUT)
    for k, v in done.items():
        print("  OK %d件  %s" % (v, k.replace("\n", "/")[:60]))

    # ------------------------------------------------------ 校正結果の一覧
    FONT = "游ゴシック"
    NAVY, HEAD = "1F3864", "4472C4"
    OK_G, IN_Y, NG_O, MID_B = "E2EFDA", "FFF2CC", "FCE4D6", "DEEBF7"
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = Workbook()
    ws = wb.active
    ws.title = "校正結果"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([6, 9, 26, 26, 26, 56], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1,
                value="キックオフ会議議事録（令和8年8月6日開催）の校正結果")
    c.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    c = ws.cell(row=2, column=1,
                value="誤字脱字及び明らかな表記の誤りは"
                      "「第10期計画_キックオフ会議議事録_校正反映.odt」に"
                      "反映済みである。"
                      "「提案」及び「要協議」は反映していない。")
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A5"

    r = 4
    for i, v in enumerate(["No.", "区分", "箇所", "現在の記載",
                           "修正後／提案", "理由"], start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[r].height = 26
    r += 1

    FILL = {"誤字": OK_G, "脱字": OK_G, "誤記": OK_G, "表記": OK_G,
            "見送り": IN_Y, "要確認": IN_Y, "解決済": OK_G}
    no = 0
    rows = ([(k, w, o, n, why) for k, w, o, n, why in FIX]
            + [(k, w, (o if "<" not in o else "XMLの断片（%s）" % w),
                (n if "<" not in n else "反映済み"), why)
               for k, w, o, n, why in FIX_XML]
            + [(k, w, o, n, why) for k, w, o, n, why in PROPOSE]
            + [(k, w, o, "―", why) for k, w, o, why in KYOGI])
    for k, w, o, n, why in rows:
        no += 1
        for i, v in enumerate([no, k, w, o, n, why], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if i in (1, 2)
                                    else "left")
            c.border = BORDER
            if i == 2:
                c.fill = PatternFill("solid",
                                     fgColor=FILL.get(k, MID_B))
        ws.row_dimensions[r].height = max(30, 13 * (len(why) // 34 + 1))
        r += 1

    wb.save(OUTX)
    print("saved:", OUTX)
    print("  校正 %d件（原本への反映 %d件・見送り／要確認／解決済 %d件・"
          "ご指示への回答 %d件）"
          % (no, len(FIX) + len(FIX_XML), len(PROPOSE), len(KYOGI)))


if __name__ == "__main__":
    main()
