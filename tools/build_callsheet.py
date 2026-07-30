# -*- coding: utf-8 -*-
"""
要個別確認リスト（架電・現物確認用シート）の生成

research/findings.jsonl の確度「低」の残件を、そのまま電話確認に使える様式で出力する。
- 想定担当課は調査メモから自動抽出
- 計画期間・策定年月などは記入欄として空欄で用意
"""

import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from senryaku_db import load  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE_DIR, "output", "要個別確認リスト_架電用.xlsx")

AREA_ORDER = ["石狩", "渡島", "檜山", "後志", "胆振", "日高", "空知", "上川",
              "留萌", "宗谷", "オホーツク", "十勝", "釧路", "根室", "一部事務組合等",
              "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DEPT = re.compile(r"([一-龥ぁ-んァ-ヶー]{2,12}(?:課|係|グループ|室|局|部|企業団))")
TEL = re.compile(r"([0-9]{2,5}-[0-9]{2,4}-[0-9]{3,4})")

COLS = [
    ("No", 5), ("エリア", 10), ("自治体", 13), ("事業区分", 22), ("計画名", 30),
    ("想定担当課", 16), ("電話", 14), ("確認事項", 30),
    ("計画期間 開始年度", 15), ("計画期間 終了年度", 15), ("策定年月", 13),
    ("直近改定年月", 13), ("次期改定予定", 15), ("確認日", 11), ("確認者", 10),
    ("メモ", 34), ("これまでに判明している事実", 46), ("出典URL", 58),
]

ASK = "①現行の計画期間（開始・終了年度）②策定年月 ③直近改定年月 ④次期改定の予定時期"


def main():
    recs = [r for r in load() if r.get("confidence") == "低"]
    wb = Workbook()
    ws = wb.active
    ws.title = "要個別確認リスト"
    ws.freeze_panes = "D3"

    ws["A1"] = ("水道・下水道「経営戦略」改訂時期調査　要個別確認リスト"
                "（黄色セルが記入欄／確認事項は各行共通）")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    for i, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    def sort_key(r):
        a = r.get("area", "")
        return (AREA_ORDER.index(a) if a in AREA_ORDER else 99, r.get("muni", ""), r.get("jigyo", ""))

    row = 3
    for n, r in enumerate(sorted(recs, key=sort_key), start=1):
        note = r.get("note", "")
        dept = ""
        m = DEPT.search(note)
        if m:
            cand = m.group(1)
            if not cand.startswith(("検索", "町村", "個別")):
                dept = cand
        tel = TEL.search(note).group(1) if TEL.search(note) else ""
        vals = [n, r.get("area", ""), r.get("muni", ""), r.get("jigyo", ""),
                r.get("plan_name", "") or "（計画名不明）", dept, tel, ASK,
                "", "", "", "", "", "", "", "", note, r.get("source", "")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in (5, 8, 17, 18)))
            if 9 <= i <= 16:
                c.fill = INPUT_FILL
        row += 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"{OUT}  （{len(recs)}件）")


if __name__ == "__main__":
    main()
