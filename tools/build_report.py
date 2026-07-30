# -*- coding: utf-8 -*-
"""
経営戦略（水道・下水道）改訂時期調査 レポート生成
- research/findings.jsonl → Excel（一覧＋抽出シート）＋ Markdown サマリ
"""

import os
import sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from senryaku_db import load, fy_label, BASE_FY, BASE_FY_LABEL  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUT_DIR, exist_ok=True)

AREA_ORDER = ["石狩", "渡島", "檜山", "後志", "胆振", "日高", "空知", "上川",
              "留萌", "宗谷", "オホーツク", "十勝", "釧路", "根室", "一部事務組合等",
              "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"]

TOHOKU = ["青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
PRI_FILL = {
    "A": PatternFill("solid", fgColor="FFC7CE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="E2EFDA"),
    "-": PatternFill("solid", fgColor="F2F2F2"),
}
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("振興局", 10), ("自治体", 12), ("事業区分", 22), ("計画名", 34),
    ("策定年度", 15), ("直近改定年度", 15), ("計画期間", 22),
    ("満了判定", 22), ("経過年数", 9), ("経過判定", 16),
    ("該当区分", 24), ("優先度", 18), ("確度", 6), ("備考", 52), ("出典URL", 60),
]


def period(r):
    s, e = r.get("start_fy"), r.get("end_fy")
    if s and e:
        return f"{fy_label(s)}～{fy_label(e)}"
    if e:
        return f"～{fy_label(e)}"
    if s:
        return f"{fy_label(s)}～"
    return "不明"


def row_of(r):
    return [
        r.get("area", ""), r.get("muni", ""), r.get("jigyo", ""), r.get("plan_name", ""),
        fy_label(r.get("made_fy")), fy_label(r.get("revised_fy")), period(r),
        r.get("expiry_judge", ""),
        r.get("elapsed_years") if r.get("elapsed_years") is not None else "",
        r.get("elapsed_judge", ""), r.get("hit", ""), r.get("priority", ""),
        r.get("confidence", ""), r.get("note", ""), r.get("source", ""),
    ]


def sort_key(r):
    a = r.get("area", "")
    return (AREA_ORDER.index(a) if a in AREA_ORDER else 99, r.get("muni", ""), r.get("jigyo", ""))


def write_sheet(ws, records, title_note=""):
    ws.freeze_panes = "A3"
    ws["A1"] = ws.title + ("　" + title_note if title_note else "")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    for i, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    for j, r in enumerate(sorted(records, key=sort_key)):
        for i, v in enumerate(row_of(r), start=1):
            c = ws.cell(row=3 + j, column=i, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in (4, 14, 15)))
        pri = (r.get("priority") or "-")[:1]
        ws.cell(row=3 + j, column=12).fill = PRI_FILL.get(pri, PRI_FILL["-"])
    n = len(records)
    if n:
        ref = f"A2:{get_column_letter(len(COLS))}{2 + n}"
        t = Table(displayName=f"T{abs(hash(ws.title)) % 100000}", ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        try:
            ws.add_table(t)
        except ValueError:
            pass


def build(region="北海道"):
    recs = [r for r in load()
            if (r.get("pref", "北海道") == "北海道") == (region == "北海道")]
    if not recs:
        return
    wb = Workbook()

    # --- 表紙 ---
    ws = wb.active
    ws.title = "調査概要"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    rows = [
        ("調査名", f"{region}　市町村 水道・下水道事業「経営戦略」改訂時期調査"),
        ("調査基準日", f"2026年7月30日（{BASE_FY_LABEL}）"),
        ("抽出条件①", "来年で計画期間がおわるもの＝計画終期が令和8年度末（2027年3月）または令和9年度末（2028年3月）"),
        ("　（参考）", "既に計画期間が満了しているもの（令和7年度末以前）も改訂対象として併記"),
        ("抽出条件②", "作成から5年経過する計画＝直近の策定・改定年度が令和3年度以前（令和8年度時点で5年以上経過）"),
        ("優先度A", "計画期間が満了済／今年度末（R8年度末）満了／直近策定・改定から6年以上経過"),
        ("優先度B", "来年度末（R9年度末）満了／直近策定・改定からちょうど5年経過"),
        ("優先度C", "上記以外（翌年度以降に検討）"),
        ("国の要請", "総務省は経営戦略の計画期間を10年以上とし、3～5年ごとの見直し（改定）を要請。"
                    "下水道事業は令和7年度末までの改定率100%を目標としている。"),
        ("確度の凡例", "高＝一次資料（経営戦略本文PDF・公表ページ）に計画期間が明記／"
                     "中＝公式サイトの記載から判断／低＝公表の事実のみ確認で年度は要確認"),
        ("調査方法", "各自治体公式サイトおよび公表PDFのウェブ検索による。"
                   "確度「低」および空欄は各団体への個別確認（電話・HP精査）が必要。"),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=k)
        a.font = Font(bold=True, color="FFFFFF")
        a.fill = SUB_FILL
        a.alignment = Alignment(vertical="center", wrap_text=True)
        b = ws.cell(row=i, column=2, value=v)
        b.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 30

    write_sheet(wb.create_sheet("全件一覧"), recs)
    hit_expiry = [r for r in recs if r.get("end_fy") in (BASE_FY, BASE_FY + 1)]
    write_sheet(wb.create_sheet("①来年で期間満了"), hit_expiry,
                "計画終期が令和8年度末または令和9年度末")
    expired = [r for r in recs if r.get("end_fy") is not None and r["end_fy"] < BASE_FY]
    write_sheet(wb.create_sheet("①-2 満了済"), expired, "計画期間が既に満了（要改訂）")
    hit_5y = [r for r in recs if (r.get("elapsed_years") or 0) >= 5]
    write_sheet(wb.create_sheet("②策定から5年経過"), hit_5y,
                "直近の策定・改定から5年以上経過")
    just5 = [r for r in recs if r.get("elapsed_years") == 5]
    write_sheet(wb.create_sheet("②-2 ちょうど5年"), just5,
                "令和3年度に策定・改定＝令和8年度で5年目")

    # --- 自治体別サマリ ---
    ws2 = wb.create_sheet("自治体別サマリ")
    ws2.freeze_panes = "A3"
    cols2 = [("振興局", 12), ("自治体", 14), ("水道系の状況", 46), ("下水道系の状況", 46),
             ("最優先度", 18), ("該当", 26)]
    ws2["A1"] = "自治体別サマリ（水道系／下水道系の直近状況）"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = HEADER_FILL
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols2))
    for i, (name, w) in enumerate(cols2, start=1):
        c = ws2.cell(row=2, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = SUB_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws2.column_dimensions[get_column_letter(i)].width = w

    def is_water(j):
        return ("水道" in j and "下水" not in j) or "上下水道" in j

    munis = {}
    for r in recs:
        munis.setdefault((r.get("area", ""), r["muni"]), []).append(r)

    def brief(r):
        p = period(r)
        j = r.get("hit") or r.get("expiry_judge") or "－"
        return f"{r.get('jigyo','')}：{p}／{j}"

    order = {"A": 0, "B": 1, "C": 2, "－": 3}
    row = 3
    for (area, muni) in sorted(munis, key=lambda k: (AREA_ORDER.index(k[0]) if k[0] in AREA_ORDER else 99, k[1])):
        rs = munis[(area, muni)]
        w = [brief(r) for r in rs if is_water(r.get("jigyo", ""))]
        s = [brief(r) for r in rs if not is_water(r.get("jigyo", ""))]
        pris = sorted({(r.get("priority") or "－")[:1] for r in rs}, key=lambda x: order.get(x, 9))
        top = [r for r in rs if (r.get("priority") or "－").startswith(pris[0])]
        hits = "／".join(sorted({r.get("hit", "") for r in top if r.get("hit")}))
        vals = [area, muni, "\n".join(w) or "－", "\n".join(s) or "－",
                top[0].get("priority", ""), hits]
        for i, v in enumerate(vals, start=1):
            c = ws2.cell(row=row, column=i, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws2.cell(row=row, column=5).fill = PRI_FILL.get(pris[0], PRI_FILL["-"])
        row += 1
    pri_a = [r for r in recs if (r.get("priority") or "").startswith("A")]
    write_sheet(wb.create_sheet("優先度A"), pri_a, "改訂着手が必要")
    todo = [r for r in recs if r.get("confidence") == "低" or (not r.get("end_fy") and not r.get("made_fy") and not r.get("revised_fy"))]
    write_sheet(wb.create_sheet("要個別確認"), todo, "年度が特定できず個別確認が必要")

    path = os.path.join(OUT_DIR, f"{region}_上下水道経営戦略_改訂時期一覧.xlsx")
    wb.save(path)

    # --- Markdown ---
    md = []
    md.append(f"# {region}　市町村 水道・下水道事業「経営戦略」改訂時期調査\n")
    md.append(f"- 調査基準日：2026年7月30日（{BASE_FY_LABEL}）")
    md.append("- 抽出条件①：**来年で計画期間がおわるもの**＝計画終期が令和8年度末（2027年3月）／令和9年度末（2028年3月）")
    md.append("- 抽出条件②：**作成から5年経過する計画**＝直近の策定・改定が令和3年度以前")
    md.append(f"- 収録件数：{len(recs)}件（{len({r['muni'] for r in recs})}自治体）\n")

    def table(rs):
        out = ["| 振興局 | 自治体 | 事業 | 計画期間 | 策定/改定 | 判定 | 確度 |",
               "|---|---|---|---|---|---|---|"]
        for r in sorted(rs, key=sort_key):
            made = fy_label(r.get("revised_fy")) or fy_label(r.get("made_fy")) or "－"
            out.append(f"| {r.get('area','')} | {r.get('muni','')} | {r.get('jigyo','')} | "
                       f"{period(r)} | {made} | {r.get('hit','')} | {r.get('confidence','')} |")
        return out

    md.append("## ① 来年で計画期間がおわる計画\n")
    md += table(hit_expiry) + [""]
    md.append("## ①-2 すでに計画期間が満了している計画（要改訂）\n")
    md += table(expired) + [""]
    md.append("## ② 策定・改定から5年以上経過している計画\n")
    md += table(hit_5y) + [""]
    md.append("## エリア別の進捗\n")
    by_area = defaultdict(set)
    for r in recs:
        by_area[r.get("area", "")].add(r["muni"])
    md.append("| エリア | 調査済自治体数 |")
    md.append("|---|---|")
    for a in AREA_ORDER:
        if a in by_area:
            md.append(f"| {a} | {len(by_area[a])} |")
    md.append("")
    md.append("## 要個別確認（年度が特定できなかったもの）\n")
    md += table(todo) + [""]

    mdpath = os.path.join(BASE_DIR, "research", f"{region}_経営戦略_改訂時期サマリ.md")
    with open(mdpath, "w", encoding="utf-8") as f:
        f.write("\n".join(md))

    print(f"Excel: {path}")
    print(f"MD   : {mdpath}")
    print(f"件数 {len(recs)} / 自治体 {len({r['muni'] for r in recs})} / "
          f"①来年満了 {len(hit_expiry)} / 満了済 {len(expired)} / ②5年経過 {len(hit_5y)} / 要確認 {len(todo)}")


if __name__ == "__main__":
    build("北海道")
    build("東北")
