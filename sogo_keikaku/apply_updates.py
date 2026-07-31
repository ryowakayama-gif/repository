# -*- coding: utf-8 -*-
"""
総合計画シートへの更新反映

- 「総合計画」シートは D(始期)/E(終期)/H(根拠URL)/I(確認メモ)/J(調査状況) のみ書き換え、
  F列の数式 =IF(E行="","",E行-1) は触らない。
- 「総合計画_個別確認ログ」に今回の確認行を追記。
- 「優先確認リスト」「確認対象サマリ」は最新状態から再構築する。
  （旧版はヘッダ等がASCIIの'?'に化けていたため、正しい日本語ヘッダで作り直す）
"""

import json
import os
import sys
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "各種計画_周期表_北海道東北_総合計画_作業中.xlsx")
UPD = os.path.join(BASE, "updates.jsonl")
CHECK_DATE = "2026-07-31"

AREA_ORDER = ["北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"]
REMAIN = {"確認継続：最新期間未確認", "確認済：掲載計画終期超過", "要確認：終期超過"}
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", name="Arial")


def load_updates():
    out = OrderedDict()
    with open(UPD, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                u = json.loads(line)
                out[u["row"]] = u  # 同一行は後勝ち
    return list(out.values())


def write_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def main(dst):
    wb = load_workbook(SRC)  # 数式を保持したまま読む
    ws = wb["総合計画"]
    updates = load_updates()

    # --- 1. 総合計画シートへ反映 ---
    for u in updates:
        r = u["row"]
        actual = ws.cell(r, 3).value
        assert actual == u["muni"], f"row{r} 自治体不一致: {actual} != {u['muni']}"
        if u.get("start") is not None:
            ws.cell(r, 4).value = u["start"]
        if u.get("end") is not None:
            ws.cell(r, 5).value = u["end"]
        if u.get("url"):
            ws.cell(r, 8).value = u["url"]
        if u.get("memo"):
            ws.cell(r, 9).value = u["memo"]
        if u.get("status"):
            ws.cell(r, 10).value = u["status"]

    # --- 2. 個別確認ログへ追記 ---
    log = wb["総合計画_個別確認ログ"]
    row = log.max_row + 1
    for u in updates:
        r = u["row"]
        vals = [CHECK_DATE, ws.cell(r, 2).value, u["muni"], "公式ページ個別確認", "総合計画",
                ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 8).value, u.get("memo", "")]
        for i, v in enumerate(vals, start=1):
            log.cell(row, i, value=v)
        row += 1

    # --- 3. 優先確認リストを再構築 ---
    idx = wb.sheetnames.index("優先確認リスト")
    del wb["優先確認リスト"]
    pl = wb.create_sheet("優先確認リスト", idx)
    write_header(pl,
                 ["No", "エリア", "自治体名", "始期", "終期", "調査状況", "根拠URL", "確認メモ", "次のアクション"],
                 [5, 10, 14, 8, 8, 24, 52, 70, 30])
    n = 0
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 3).value is None:
            continue
        stat = ws.cell(r, 10).value
        if stat not in REMAIN and ws.cell(r, 4).value is not None and ws.cell(r, 5).value is not None:
            continue
        n += 1
        action = ("計画期間の年度を自治体へ個別確認" if stat == "確認継続：最新期間未確認"
                  else "次期計画の策定状況を確認")
        vals = [n, ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value,
                ws.cell(r, 5).value, stat, ws.cell(r, 8).value, ws.cell(r, 9).value, action]
        for i, v in enumerate(vals, start=1):
            c = pl.cell(n + 1, i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=(i in (7, 8, 9)))

    # --- 4. 確認対象サマリを再構築 ---
    idx = wb.sheetnames.index("確認対象サマリ")
    del wb["確認対象サマリ"]
    sm = wb.create_sheet("確認対象サマリ", idx)
    write_header(sm,
                 ["エリア", "対象行数", "始期・終期とも入力済", "未入力", "終期2025年以前",
                  "要継続確認（残件）", "公式確認済・個別確認済"],
                 [12, 10, 18, 8, 14, 16, 20])
    stats = OrderedDict((a, dict(total=0, filled=0, blank=0, past=0, remain=0, done=0)) for a in AREA_ORDER)
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 3).value is None:
            continue
        area = ws.cell(r, 2).value
        if area not in stats:
            continue
        s, e, stat = ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 10).value
        st = stats[area]
        st["total"] += 1
        if s is not None and e is not None:
            st["filled"] += 1
        else:
            st["blank"] += 1
        if isinstance(e, int) and e <= 2025:
            st["past"] += 1
        if stat in REMAIN or s is None or e is None:
            st["remain"] += 1
        if stat in ("公式確認済", "個別確認済"):
            st["done"] += 1
    i = 2
    for area, st in stats.items():
        for j, v in enumerate([area, st["total"], st["filled"], st["blank"],
                               st["past"], st["remain"], st["done"]], start=1):
            sm.cell(i, j, value=v)
        i += 1
    for j, v in enumerate(["合計", sum(s["total"] for s in stats.values()),
                           sum(s["filled"] for s in stats.values()),
                           sum(s["blank"] for s in stats.values()),
                           sum(s["past"] for s in stats.values()),
                           sum(s["remain"] for s in stats.values()),
                           sum(s["done"] for s in stats.values())], start=1):
        c = sm.cell(i, j, value=v)
        c.font = Font(bold=True, name="Arial")

    wb.save(dst)
    print(f"反映 {len(updates)}件 / 残件 {n}件 → {dst}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "out.xlsx"))
