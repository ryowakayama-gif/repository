# -*- coding: utf-8 -*-
"""公共施設等総合管理計画シートを複製し「個別施設計画」シートを作成する"""
import os, sys, shutil
from copy import copy
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "outputs", "各種計画_周期表_北海道東北_総合計画_秋田確認版18.xlsx")
DST = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "outputs", "各種計画_周期表_北海道東北_個別施設計画_版19.xlsx")

wb = load_workbook(SRC)
if "個別施設計画" in wb.sheetnames:
    del wb["個別施設計画"]
src = wb["公共施設等総合管理計画"]
dst = wb.create_sheet("個別施設計画", wb.sheetnames.index("公共施設等総合管理計画") + 1)

# 列幅・行高
for k, v in src.column_dimensions.items():
    dst.column_dimensions[k].width = v.width
dst.freeze_panes = src.freeze_panes

for row in src.iter_rows(min_row=1, max_row=src.max_row, max_col=src.max_column):
    for c in row:
        n = dst.cell(c.row, c.column)
        if c.has_style:
            n.font, n.fill, n.border = copy(c.font), copy(c.fill), copy(c.border)
            n.alignment, n.number_format = copy(c.alignment), c.number_format
        r, col = c.row, c.column
        if r <= 3:                       # タイトル・ヘッダ・例示行はそのまま
            n.value = c.value
        elif col in (1, 2, 3):           # No・エリア・自治体名は引き継ぐ
            n.value = c.value
        elif col == 6:                   # F列は数式を張り直す
            n.value = f'=IF(E{r}="","",E{r}-1)' if src.cell(r, 3).value else None
        elif col == 10:                  # 調査状況は未着手で初期化
            n.value = "未着手" if src.cell(r, 3).value else None
        else:                            # 始期・終期・営業担当・URL・メモは空に
            n.value = None

dst["A1"] = "計画周期表（個別施設計画）"
wb.save(DST)
print(f"作成 → {DST}")
rows = sum(1 for r in range(4, dst.max_row + 1) if dst.cell(r, 3).value)
print(f"個別施設計画シート データ行: {rows}")
