# -*- coding: utf-8 -*-
"""個別施設計画シートへの更新反映（自治体名で行を特定）"""
import json, os, sys
from collections import OrderedDict
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "outputs", "各種計画_周期表_北海道東北_個別施設計画_版19.xlsx")
UPD = os.path.join(BASE, "updates_kobetsu.jsonl")

def main(dst):
    wb = load_workbook(SRC); ws = wb["個別施設計画"]
    idx = {}
    for r in range(4, ws.max_row + 1):
        m = ws.cell(r, 3).value
        if m:
            idx.setdefault((ws.cell(r, 2).value, m), r)
    ups = OrderedDict()
    with open(UPD, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                u = json.loads(line)
                ups[(u["area"], u["muni"])] = u
    n = 0
    for key, u in ups.items():
        r = idx.get(key)
        assert r, f"行が見つからない: {key}"
        if u.get("start") is not None: ws.cell(r, 4).value = u["start"]
        if u.get("end") is not None:   ws.cell(r, 5).value = u["end"]
        if u.get("url"):               ws.cell(r, 8).value = u["url"]
        if u.get("memo"):              ws.cell(r, 9).value = u["memo"]
        if u.get("status"):            ws.cell(r, 10).value = u["status"]
        n += 1
    wb.save(dst)
    done = sum(1 for r in range(4, ws.max_row + 1)
               if ws.cell(r, 3).value and ws.cell(r, 10).value != "未着手")
    print(f"反映 {n}件 / 着手済 {done}件 → {dst}")

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else SRC)
