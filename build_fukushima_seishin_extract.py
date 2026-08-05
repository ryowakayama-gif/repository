# -*- coding: utf-8 -*-
"""
福島県 精神医療指標（NDB）抽出ツール

入力: 国立精神・神経医療研究センター「精神保健医療福祉に関する指標（NDB集計）」
      Tables_f2013_2023_ver1.0.xlsx      通常集計（2013〜2023年度）
      Tables_f2023_kohi_ver1.0.xlsx      公費負担医療を含む集計
      いずれも 19〜21MB あり、リポジトリには格納しない。
      入手先は docs/北塩原村_長期入院患者基盤整備量_資料充足性確認.md を参照。

出力: output/福島県_精神医療指標抽出.xlsx
      基本指針 第二の二（成果目標）及び別表第四に必要な福島県分のみを抜き出す。

抽出する付表
  付表1.1 精神病床退院患者における地域平均生活日数   → 国基準 319.3日 の県実績
  付表1.2 特定時点の再入院患者割合                   → 国基準 10.3/17.4/25.7% の県実績
  付表3.3 精神病床在院患者延数(年齢×在院日数×認知症) → 別表第四 A1・A2・B1・B2 の実績側

使い方
  python3 build_fukushima_seishin_extract.py [通常集計.xlsx] [公費含む集計.xlsx]
  引数を省略した場合は SRC_DEFAULT を参照する。
"""

import os
import sys
import warnings

from openpyxl import Workbook, load_workbook

from kitashiobara_common import (
    COLORS, FONT, OUT_DIR, add_sheet, ensure_out_dir, style_header_row,
    style_note, style_title, write_row,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

OUT_FILE = f"{OUT_DIR}/福島県_精神医療指標抽出.xlsx"
SRC_DEFAULT = [
    "/root/.claude/uploads/6ba69703-c476-5e43-a82e-705be12592ce/a988ab72-Tables_f2013_2023_ver1.0.xlsx",
    "/root/.claude/uploads/6ba69703-c476-5e43-a82e-705be12592ce/f5e69661-Tables_f2023_kohi_ver1.0.xlsx",
]
# 630調査 従来ベース集計（リポジトリに格納済み。20MB級のNDBと違い軽い）
SRC_630 = "source/精神保健/630調査/630調査_令和5年度_従来ベース集計.xlsx"
PREF = "福島県"

SHEET_SEIKATSU = "付表1.1 精神病床退院患者における地域平均生活日数"
SHEET_SAINYUIN_PREFIX = "付表1.2"
SHEET_ZAIIN = "付表3.3 精神病床在院患者延数(年齢×在院日数区分×認知症)"
SHEET_ZAIIN_ALT = "付表3.3 精神病床在院患者延数(年齢×在院日数×認知症)"

# 別表第四の記号 → (年齢区分の集合, 認知症区分)
BEPPYO4_MAP = [
    ("A1", "65歳以上・認知症を除く", {"65~74歳", "75~84歳", "85歳以上"}, "なし"),
    ("A2", "65歳以上・認知症である者", {"65~74歳", "75~84歳", "85歳以上"}, "あり"),
    ("B1", "65歳未満・認知症を除く", {"0~64歳"}, "なし"),
    ("B2", "65歳未満・認知症である者", {"0~64歳"}, "あり"),
]
AGE75 = {"75~84歳", "85歳以上"}

INT = "#,##0"
DEC1 = "#,##0.0"
PCT1 = "0.0"


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _sheet(wb, *names):
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    for n in wb.sheetnames:
        if n.startswith(names[0][:6]):
            return wb[n]
    raise KeyError(f"シートが見つかりません: {names}")


def read_source(path):
    """1 ファイルから福島県分の 3 付表を読み出す。"""
    wb = load_workbook(path, data_only=True, read_only=True)

    seikatsu = []
    ws = _sheet(wb, SHEET_SEIKATSU)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] != PREF:
            continue
        seikatsu.append({
            "年度": r[13], "医療機関数": r[2], "退院患者数": r[3],
            "全病床": _num(r[5]), "精神病床": _num(r[7]),
            "年齢調整全病床": _num(r[10]), "年齢調整精神病床": _num(r[11]),
        })

    sainyuin = []
    ws = _sheet(wb, SHEET_SAINYUIN_PREFIX)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] != PREF:
            continue
        sainyuin.append({
            "年度": r[-1], "在院日数区分": r[2], "生存退院患者数": _num(r[6]),
            "90日": _num(r[10]), "180日": _num(r[11]), "365日": _num(r[12]),
            "90日年齢調整": _num(r[13]), "180日年齢調整": _num(r[14]),
            "365日年齢調整": _num(r[15]),
        })

    zaiin = []
    ws = _sheet(wb, SHEET_ZAIIN, SHEET_ZAIIN_ALT)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[1] != PREF:
            continue
        zaiin.append({
            "年度": r[8], "年齢区分": r[2], "在院日数区分": r[3], "認知症区分": r[4],
            "延数": _num(r[5]), "1日平均": _num(r[6]), "受療率": _num(r[7]),
        })

    wb.close()
    return {"生活日数": seikatsu, "再入院": sainyuin, "在院": zaiin}


def read_630(path):
    """630調査 従来ベース集計から福島県分の在院患者数を読み出す。

    Ⅲ.2.(11) 在院患者数（入院期間×住所地・施設所在地×年齢）
        別表第四のCは「当該都道府県の区域に住所を有する者」を前提とするため、
        住所地ベースの列を採る。
    Ⅲ.2.(4)  在院患者数（年齢階級・入院形態×性）
        75歳以上の在院患者数（ただし在院期間との交差はない）。
    Ⅲ.4.(1)  認知症治療病棟の在院患者数（在院期間別）
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {}

    ws = wb["Ⅲ.2.(11)"]
    for r in ws.iter_rows(min_row=7, max_row=80, max_col=15, values_only=True):
        if r[0] != PREF:
            continue
        out["期間別"] = [
            ("3ヶ月未満", _num(r[1]), _num(r[2]), _num(r[3]), _num(r[4])),
            ("3ヶ月以上12ヶ月未満", _num(r[5]), _num(r[6]), _num(r[7]), _num(r[8])),
            ("1年以上", _num(r[9]), _num(r[10]), _num(r[11]), _num(r[12])),
        ]
        break

    ws = wb["Ⅲ.2.(4)"]
    labels = ["20歳未満", "20歳以上40歳未満", "40歳以上65歳未満",
              "65歳以上75歳未満", "75歳以上", "不明"]
    for r in ws.iter_rows(min_row=7, max_row=80, max_col=20, values_only=True):
        if r[0] != PREF:
            continue
        rows = []
        for i, lb in enumerate(labels):
            m, f, u = _num(r[2 + i * 3]), _num(r[3 + i * 3]), _num(r[4 + i * 3])
            rows.append((lb, m, f, u, (m or 0) + (f or 0) + (u or 0)))
        out["年齢階級"] = rows
        out["総数"] = _num(r[1])
        break

    ws = wb["Ⅲ.4.(1)"]
    heads = ["合計", "1ヶ月未満", "1ヶ月以上3ヶ月未満", "3ヶ月以上6ヶ月未満",
             "6ヶ月以上1年未満", "1年以上5年未満", "5年以上10年未満",
             "10年以上20年未満", "20年以上", "不明"]
    for r in ws.iter_rows(min_row=6, max_row=80, max_col=11, values_only=True):
        if r[0] != PREF:
            continue
        out["認知症病棟"] = [(h, _num(r[1 + i])) for i, h in enumerate(heads)]
        break

    wb.close()
    return out


def latest_year(rows):
    years = [r["年度"] for r in rows if r["年度"] is not None]
    return max(years) if years else None


def beppyo4(zaiin, year):
    """別表第四 A1・A2・B1・B2 及び 75歳以上 を年度指定で集計する。"""
    sel = [r for r in zaiin
           if r["年度"] == year and r["在院日数区分"] == "366日以上"]
    out = {}
    for code, label, ages, dem in BEPPYO4_MAP:
        vals = [r["1日平均"] for r in sel
                if r["年齢区分"] in ages and r["認知症区分"] == dem
                and r["1日平均"] is not None]
        out[code] = (label, round(sum(vals), 1) if vals else None, len(vals))
    v75 = [r["1日平均"] for r in sel
           if r["年齢区分"] in AGE75 and r["認知症区分"] in ("あり", "なし")
           and r["1日平均"] is not None]
    out["75"] = ("75歳以上（認知症の有無を問わない）",
                 round(sum(v75), 1) if v75 else None, len(v75))
    total = [v[1] for k, v in out.items() if k != "75" and v[1] is not None]
    out["C"] = ("合計＝1年以上入院患者数（別表第四のCに相当）",
                round(sum(total), 1) if total else None, len(total))
    return out


# ============================================================
# シート生成
# ============================================================
def sheet_overview(wb, sources):
    ws = add_sheet(
        wb, "00_概要", "福島県 精神医療指標（NDB集計）抽出",
        "国立精神・神経医療研究センター「精神保健医療福祉に関する指標」から福島県分のみを抽出したもの。"
        "第8期基本指針 第二の二の成果目標及び別表第四の算定に用いる。"
        "原典は1ファイル20MB前後のためリポジトリには格納していない。",
        [30, 24, 24, 60])
    r = 5
    style_header_row(ws, r, ["シート", "対応する国基準", "原典の付表", "内容"])
    r += 1
    rows = [
        ("01_平均生活日数", "第二の二の1（319.3日以上）", "付表1.1",
         "精神病床から退院後1年以内の地域における平均生活日数。全病床・精神病床・年齢調整済みの4系列"),
        ("02_再入院率", "第二の二の3（90日10.3%・180日17.4%・365日25.7%以下）", "付表1.2",
         "退院患者の30日以上の再入院率。特定時点3種と年齢調整済み"),
        ("03_長期入院患者", "第二の二の2、別表第四", "付表3.3",
         "精神病床在院患者延数（年齢区分×在院日数区分×認知症区分）。366日以上がA1・A2・B1・B2の実績側"),
        ("05_630調査R5", "第二の二の2、別表第四のC", "630調査 Ⅲ.2.(11)・Ⅲ.2.(4)・Ⅲ.4.(1)",
         "令和5年6月30日現在の在院患者数。住所地ベースの1年以上入院患者数がCに相当"),
        ("04_別表第四", "別表第四 一〜三の項", "付表3.3から集計",
         "別表第四の記号への当てはめと、算定に不足している要素の一覧"),
    ]
    for i, rec in enumerate(rows):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 34
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_title(ws.cell(row=r, column=1), "抽出元", fill=COLORS["subhead"], size=11)
    r += 1
    for label, path in sources:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        style_note(ws.cell(row=r, column=1), f"{label}：{os.path.basename(path)}")
        r += 1
    return ws


def sheet_seikatsu(wb, data):
    ws = add_sheet(
        wb, "01_平均生活日数", "精神病床退院患者における地域平均生活日数（福島県）",
        "国の基本指針 第二の二の1は319.3日以上を基本とし、設定時点で既に319.3日以上である場合は"
        "その時点の平均生活日数以上とすることを基本としている。判定にはこの県実績を用いる。単位：日。",
        [12, 14, 14, 16, 16, 18, 18, 14])
    style_header_row(ws, 5, ["年度", "医療機関数", "退院患者数", "全病床", "精神病床",
                             "年齢調整（全病床）", "年齢調整（精神病床）", "集計区分"])
    r = 6
    for label, rows in data:
        for i, rec in enumerate(sorted(rows, key=lambda x: x["年度"] or 0)):
            write_row(ws, r, [rec["年度"], rec["医療機関数"], rec["退院患者数"],
                              rec["全病床"], rec["精神病床"],
                              rec["年齢調整全病床"], rec["年齢調整精神病床"], label],
                      alt=(i % 2 == 1),
                      aligns=["center", "right", "right", "right", "right", "right",
                              "right", "center"],
                      numfmts=[None, INT, INT, DEC1, DEC1, DEC1, DEC1, None])
            r += 1
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
    style_note(ws.cell(row=r + 1, column=1),
               "※「公費含む」は自立支援医療等の公費負担医療を含む集計であり、対象患者数が多くなる。"
               "どちらを国基準との比較に用いるかは、国が示す算定要領の確認が必要。")
    ws.row_dimensions[r + 1].height = 28
    return ws


def sheet_sainyuin(wb, data):
    ws = add_sheet(
        wb, "02_再入院率", "退院患者の精神病床への30日以上の再入院率（福島県）",
        "国の基本指針 第二の二の3は、退院後90日時点10.3%以下、180日時点17.4%以下、"
        "365日時点25.7%以下を基本としている。単位：％。",
        [12, 14, 16, 12, 12, 12, 16, 16, 16, 14])
    style_header_row(ws, 5, ["年度", "在院日数区分", "生存退院患者数",
                             "90日", "180日", "365日",
                             "90日（年齢調整）", "180日（年齢調整）", "365日（年齢調整）",
                             "集計区分"])
    r = 6
    for label, rows in data:
        for i, rec in enumerate(sorted(rows, key=lambda x: (x["年度"] or 0, str(x["在院日数区分"])))):
            write_row(ws, r, [rec["年度"], rec["在院日数区分"], rec["生存退院患者数"],
                              rec["90日"], rec["180日"], rec["365日"],
                              rec["90日年齢調整"], rec["180日年齢調整"], rec["365日年齢調整"],
                              label],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "right"] + ["right"] * 6 + ["center"],
                      numfmts=[None, None, INT] + [PCT1] * 6 + [None])
            r += 1
    return ws


def sheet_zaiin(wb, data):
    ws = add_sheet(
        wb, "03_長期入院患者", "精神病床在院患者延数（年齢区分×在院日数区分×認知症区分・福島県）",
        "在院日数区分「366日以上」が基本指針でいう1年以上長期入院患者。"
        "366日以上は集約年齢区分（0〜64歳／65〜74歳／75〜84歳／85歳以上）×認知症の有無でのみ公表されており、"
        "5歳刻みの区分は1〜90日と91〜365日にしか存在しない。",
        [12, 14, 16, 12, 16, 16, 18, 14])
    style_header_row(ws, 5, ["年度", "年齢区分", "在院日数区分", "認知症区分",
                             "在院患者延数", "1日平均在院患者数", "入院受療率（人口10万対）",
                             "集計区分"])
    r = 6
    for label, rows in data:
        keep = [x for x in rows if x["在院日数区分"] == "366日以上"]
        for i, rec in enumerate(sorted(keep, key=lambda x: (x["年度"] or 0, str(x["認知症区分"]), str(x["年齢区分"])))):
            write_row(ws, r, [rec["年度"], rec["年齢区分"], rec["在院日数区分"],
                              rec["認知症区分"], rec["延数"], rec["1日平均"], rec["受療率"],
                              label],
                      alt=(i % 2 == 1),
                      aligns=["center", "center", "center", "center", "right", "right",
                              "right", "center"],
                      numfmts=[None, None, None, None, INT, DEC1, DEC1, None])
            r += 1
    return ws


def sheet_630(wb, d630):
    ws = add_sheet(
        wb, "05_630調査R5", "精神保健福祉資料（630調査）令和5年度・福島県",
        "令和5年6月30日午前0時現在。別表第四のCは「令和五年における精神病床における入院期間が"
        "1年以上である入院患者数」であり、A1〜B2が「当該都道府県の区域に住所を有する者」を前提と"
        "していることから、住所地ベースの値を採るのが整合的。単位：人（実人数）。",
        [26, 18, 18, 18, 18, 16])

    r = 5
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_title(ws.cell(row=r, column=1),
                "1. 在院患者数（入院期間×住所地・施設所在地×年齢）　Ⅲ.2.(11)",
                fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["入院期間", "住所地\n65歳未満", "住所地\n65歳以上",
                             "所在地\n65歳未満", "所在地\n65歳以上", "住所地 計"])
    r += 1
    for i, (lb, a, b, c, d) in enumerate(d630.get("期間別", [])):
        total = (a or 0) + (b or 0)
        write_row(ws, r, [lb, a, b, c, d, total],
                  alt=(i % 2 == 1),
                  aligns=["left"] + ["right"] * 5,
                  numfmts=[None] + [INT] * 5,
                  fills=[COLORS["calc"] if lb == "1年以上" else None] * 6)
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_note(ws.cell(row=r, column=1),
               "※「1年以上」の住所地計が別表第四のCに相当する。"
               "福島県第7期障がい福祉計画が掲げる長期在院者数（令和4年度 65歳未満995人・"
               "65歳以上1,813人）と同じ系統であり、Cは630調査の住所地ベース実人数と解される。")
    ws.row_dimensions[r].height = 34
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_title(ws.cell(row=r, column=1),
                "2. 在院患者数（年齢階級別）　Ⅲ.2.(4)", fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["年齢階級", "男性", "女性", "不明", "計", "備考"])
    r += 1
    for i, (lb, m, f, u, t) in enumerate(d630.get("年齢階級", [])):
        note = "第8期で新設された目標区分。ただし在院期間との交差表がない" if lb == "75歳以上" else ""
        write_row(ws, r, [lb, m, f, u, t, note],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "left"],
                  numfmts=[None] + [INT] * 4 + [None])
        r += 1
    write_row(ws, r, ["総数", None, None, None, d630.get("総数"), "全在院期間の合計"],
              aligns=["left", "right", "right", "right", "right", "left"],
              numfmts=[None] + [INT] * 4 + [None],
              fills=[COLORS["band"]] * 6)
    r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    style_title(ws.cell(row=r, column=1),
                "3. 認知症治療病棟の在院患者数（在院期間別）　Ⅲ.4.(1)",
                fill=COLORS["subhead"], size=11)
    r += 1
    style_header_row(ws, r, ["在院期間", "患者数", "", "", "", "備考"])
    r += 1
    dem = d630.get("認知症病棟", [])
    for i, (lb, v) in enumerate(dem):
        write_row(ws, r, [lb, v, None, None, None, ""],
                  alt=(i % 2 == 1),
                  aligns=["left", "right", "right", "right", "right", "left"],
                  numfmts=[None, INT, None, None, None, None])
        r += 1
    long = sum(v for lb, v in dem
               if lb in ("1年以上5年未満", "5年以上10年未満",
                         "10年以上20年未満", "20年以上") and v is not None)
    write_row(ws, r, ["うち1年以上", long, None, None, None,
                      "認知症治療病棟に限った数であり、別表第四の「認知症である者」とは一致しない"
                      "（一般精神病棟にも認知症患者がいる）。認知症区分での按分にはNDB付表3.3を用いる"],
              aligns=["left", "right", "right", "right", "right", "left"],
              numfmts=[None, INT, None, None, None, None],
              fills=[COLORS["band"]] * 6)
    ws.row_dimensions[r].height = 34
    return ws


def sheet_beppyo4(wb, data):
    ws = add_sheet(
        wb, "04_別表第四", "基本指針 別表第四への当てはめ（福島県）",
        "別表第四のA1・A2・B1・B2は「令和11年における年齢階級別の推計患者数」であり、"
        "本シートに掲げるのは同じ区分の実績値である。推計への変換とX1・X2の適用は"
        "都道府県が行うため、本シートだけでは三の項（基盤整備量）は算定できない。単位：人（1日平均在院患者数）。",
        [10, 34, 18, 18, 60])
    r = 5
    for label, rows in data:
        y = latest_year(rows)
        if y is None:
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        style_title(ws.cell(row=r, column=1),
                    f"{label}　{y}年度（最新年度）実績", fill=COLORS["subhead"], size=11)
        r += 1
        style_header_row(ws, r, ["記号", "別表第四の定義", "実績値", "内訳件数", "備考"])
        r += 1
        agg = beppyo4(rows, y)
        notes = {
            "A1": "1年以上・65歳以上・認知症である者を除く。65〜74／75〜84／85歳以上の合計",
            "A2": "1年以上・65歳以上・認知症である者に限る。同上",
            "B1": "1年以上・65歳未満・認知症である者を除く",
            "B2": "1年以上・65歳未満・認知症である者に限る",
            "75": "第8期で新設された目標項目。A1・A2の内数",
            "C": "別表第四の備考のCは「令和5年における1年以上入院患者数」。"
                 "本値はNDBの1日平均在院患者数であり、630調査の実人数とは定義が異なる",
        }
        for code in ["A1", "A2", "B1", "B2", "75", "C"]:
            lbl, val, n = agg[code]
            write_row(ws, r, [code, lbl, val, n, notes[code]],
                      alt=(code in ("A2", "B2", "C")),
                      aligns=["center", "left", "right", "center", "left"],
                      numfmts=[None, None, DEC1, None, None])
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    style_title(ws.cell(row=r, column=1), "算定に不足している要素",
                fill=COLORS["障がい"], size=11)
    r += 1
    style_header_row(ws, r, ["区分", "不足している要素", "必要な理由", "入手先", "状態"])
    r += 1
    gaps = [
        ("推計値", "令和11年における年齢階級別の推計患者数（A1・A2・B1・B2）",
         "別表第四が求めるのは実績ではなく令和11年の推計患者数",
         "入院受療率（本ブック03の列）×福島県の令和11年年齢階級別推計人口（社人研）で算定可能",
         "算定可"),
        ("県知事裁量", "X1・X2",
         "令和11年の人口当たり慢性期推定入院患者数と「令和5年時点で人口当たり慢性期入院患者数が"
         "少ない県の水準」を比較して定める値。差分が2割未満なら差分の半分、2割以上なら0.1が標準",
         "「少ない県の水準」の値を国が示す必要がある。福島県への照会が必須",
         "県照会"),
        ("県算定値", "別表第四 三の項＝基盤整備量（利用者数）",
         "市町村はこの値を勘案して当該市町村区域の値を定める（別表第二 三の項・必須記載事項）",
         "福島県 第8期障がい福祉計画（令和8年度中策定）。第7期計画には三の項の記載がなく、"
         "一・二の項に相当する長期在院者数（65歳未満898人・65歳以上1,656人）のみ",
         "県照会"),
        ("市町村別", "北塩原村の区域における1年以上入院患者数",
         "当該市町村区域の基盤整備量を定めるための起点",
         "ReMHRADの市町村タブ（患者住所地別）、村・会津保健所の把握。"
         "現行計画に令和5年度5人の記載あり",
         "村・県照会"),
        ("定義確認", "Cを630調査の実人数とするかNDBの1日平均在院患者数とするか",
         "福島県第7期計画の長期在院者数は630調査ベースであり、NDBの1日平均在院患者数と定義が異なる",
         "令和5年度630調査 Ⅲ.2.(11) の住所地ベース1年以上入院患者数（65歳未満901人・"
         "65歳以上1,690人・計2,591人）が県計画と同系統。Cはこれと解する",
         "解消"),
        ("年度", "令和5年度630調査の福島県データ",
         "別表第四のCは「令和五年における」入院患者数",
         "受領済み（05_630調査R5シート）",
         "解消"),
        ("年齢区分", "40歳以上の認知症である者の1年以上長期入院患者数",
         "第8期で新設された目標項目",
         "付表3.3は366日以上を0〜64歳／65〜74歳／75〜84歳／85歳以上でしか公表しておらず、"
         "0〜64歳のうち40歳以上を切り出せない。県・国への照会が必要",
         "算定不可"),
    ]
    for i, rec in enumerate(gaps):
        write_row(ws, r, list(rec), alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 46
        r += 1
    return ws


def main():
    args = sys.argv[1:] or SRC_DEFAULT
    labels = ["通常集計", "公費含む"]
    sources = []
    loaded = []
    for i, path in enumerate(args):
        if not os.path.exists(path):
            print(f"  スキップ（未配置）: {path}")
            continue
        label = labels[i] if i < len(labels) else os.path.basename(path)
        print(f"  読込: {label} <- {os.path.basename(path)}")
        loaded.append((label, read_source(path)))
        sources.append((label, path))
    if not loaded:
        raise SystemExit("抽出元ファイルが1つも見つかりません。引数でパスを指定してください。")

    ensure_out_dir()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_overview(wb, sources)
    sheet_seikatsu(wb, [(lb, d["生活日数"]) for lb, d in loaded])
    sheet_sainyuin(wb, [(lb, d["再入院"]) for lb, d in loaded])
    sheet_zaiin(wb, [(lb, d["在院"]) for lb, d in loaded])
    if os.path.exists(SRC_630):
        print(f"  読込: 630調査R5 <- {os.path.basename(SRC_630)}")
        sheet_630(wb, read_630(SRC_630))
    else:
        print(f"  スキップ（未配置）: {SRC_630}")
    sheet_beppyo4(wb, [(lb, d["在院"]) for lb, d in loaded])
    wb.save(OUT_FILE)
    print(f"作成: {OUT_FILE}")

    for lb, d in loaded:
        y = latest_year(d["在院"])
        agg = beppyo4(d["在院"], y)
        parts = " ".join(f"{k}={agg[k][1]}" for k in ["A1", "A2", "B1", "B2", "C"])
        print(f"  {lb} {y}年度: {parts}")


if __name__ == "__main__":
    main()
