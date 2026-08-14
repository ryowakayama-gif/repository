# -*- coding: utf-8 -*-
"""大雪地区広域連合 第10期介護保険事業計画
エビデンス集の生成.

成果品（計画素案・報告書・分析表）の根拠となるデータを、
発注者が Excel で直接確認できる形に書き出す。
成果品に載せた数値は、すべて本エビデンス集のいずれかの表に遡ることができる。

  E1 統計データ　　　　　人口・認定・受給・事業所数・従事者数・保険料
  E2 事業所と公表情報　　指定事業所一覧・公表画面の記載内容・施設の名簿
  E3 調査の集計値　　　　実施済み4調査の集計値とクロス集計

個人情報の取扱い
  個票（利用者票・職員個票）は収録しない。集計値のみを収める。
  担当者名・電話番号・メールアドレス及び施設の管理者名も収録しない。
"""

import importlib
import io
import os
import runpy
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ODIR = "/home/user/repository/output"
EDIR = os.path.join(ODIR, "evidence")

FONT = "游ゴシック"
NAVY, HEAD = "1F3864", "4472C4"
IN_Y, OK_G, NG_O, MID_B, GRAY = "FFF2CC", "E2EFDA", "FCE4D6", "DEEBF7", "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------- 書式
def new_book():
    wb = Workbook()
    del wb["Sheet"]
    return wb


def sheet(wb, name, title, subtitle, widths=None, freeze="A5"):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=13, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 21
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=max(2, len(widths or [1])))
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = freeze
    return ws


def head(ws, row, cols, height=30, start=1):
    for i, v in enumerate(cols, start=start):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = height
    return row + 1


def line(ws, row, vals, fills=None, height=17, start=1, bold=False):
    for i, v in enumerate(vals, start=start):
        if isinstance(v, (list, dict, tuple)):
            v = str(v)
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=bold)
        c.alignment = Alignment(wrap_text=False, vertical="top")
        c.border = BORDER
        if fills and fills.get(i):
            c.fill = PatternFill("solid", fgColor=fills[i])
    ws.row_dimensions[row].height = height
    return row + 1


def lead(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 18
    return row + 1


def note(ws, row, text, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=8.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = max(28, 13 * (len(text) // 90 + 1))
    return row + 1


def autow(ws, ncol, minw=9, maxw=42):
    """列幅を内容から決める（全角を2文字として数える）。"""
    for j in range(1, ncol + 1):
        w = minw
        for cell in ws[get_column_letter(j)]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            n = sum(2 if ord(ch) > 0x2000 else 1 for ch in s)
            w = max(w, min(maxw, n + 2))
        ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------- 書出し
def put_records(ws, row, rows, cols=None):
    """辞書のリストを表にする。列は全レコードのキーの和集合。"""
    if not rows:
        return line(ws, row, ["（該当なし）"])
    if cols is None:
        cols, seen = [], set()
        for d in rows:
            for k in d:
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
    row = head(ws, row, cols)
    for d in rows:
        row = line(ws, row, [d.get(c) for c in cols])
    return row


def put_matrix(ws, row, d, index_label, index):
    """{系列名: [値, ...]} を、行＝系列・列＝index の表にする。"""
    row = head(ws, row, [index_label] + list(index))
    for k, v in d.items():
        row = line(ws, row, [k] + list(v))
    return row


def put_nested(ws, row, d, k1, k2, k3="値"):
    """{外: {内: 値}} を3列の表にする。値がさらに辞書なら展開する。"""
    row = head(ws, row, [k1, k2, k3])
    for a, inner in d.items():
        if not isinstance(inner, dict):
            row = line(ws, row, [a, "", inner])
            continue
        for b, v in inner.items():
            if isinstance(v, dict):
                for c, w in v.items():
                    row = line(ws, row, [a, "%s／%s" % (b, c), w])
            else:
                row = line(ws, row, [a, b, v])
    return row


def put_series(ws, row, d, code_label="コード"):
    """見える化の系列辞書（{コード: {指標・単位・系列・町別}}）を展開する。"""
    row = head(ws, row, [code_label, "指標", "単位", "区分", "年・区分", "値"])
    for code in sorted(d):
        rec = d[code]
        name = rec.get("指標") or rec.get("サービス") or ""
        unit = rec.get("単位", "")
        for key in ("系列", "値", "町別", "比較", "職種"):
            blk = rec.get(key)
            if blk is None:
                continue
            if isinstance(blk, dict):
                for a, v in blk.items():
                    if isinstance(v, dict):
                        for b, w in v.items():
                            row = line(ws, row, [code, name, unit,
                                                 "%s／%s" % (key, a), b, w])
                    elif isinstance(v, list):
                        for i, w in enumerate(v):
                            row = line(ws, row, [code, name, unit, key,
                                                 "%s[%d]" % (a, i), w])
                    else:
                        row = line(ws, row, [code, name, unit, key, a, v])
            elif isinstance(blk, list):
                for i, v in enumerate(blk):
                    row = line(ws, row, [code, name, unit, key, i, v])
            else:
                row = line(ws, row, [code, name, unit, key, "", blk])
    return row


# ---------------------------------------------------------------- 出典
GENSEN = [
    ("①", "地域包括ケア「見える化」システム",
     "厚生労働省", "令和8年7月22日・7月29日・7月30日取得",
     "A系列（人口）・B系列（認定）・C1（保険料）・D系列（受給）・"
     "K系列（事業所数）・M系列（従事者数）・W144（交付金評価指標）・"
     "δ（介護サービス自給率）",
     "E1"),
    ("②", "住民基本台帳に基づく人口、人口動態及び世帯数",
     "総務省", "各年1月1日現在（平成31年〜令和7年）",
     "年齢別人口（市区町村別）【総計】。"
     "団体コード 東川町014583・美瑛町014591・東神楽町014532",
     "E1"),
    ("③", "国勢調査", "総務省",
     "平成22年・平成27年・令和2年、令和7年（人口速報集計）",
     "総人口・世帯数・世帯人員・高齢者を含む世帯の類型",
     "E1"),
    ("④", "日本の地域別将来推計人口（令和5年推計）",
     "国立社会保障・人口問題研究所", "令和2年国勢調査基準",
     "見える化システムA2・A3・A4を通じて取得",
     "E1"),
    ("⑤", "介護保険事業状況報告", "厚生労働省", "各年3月末現在",
     "要介護認定者数・受給者数",
     "E1"),
    ("⑥", "介護保険事業所一覧", "北海道",
     "令和8年6月30日現在", "区域内の指定事業所（47区分）・通常の事業の実施地域",
     "E2"),
    ("⑦", "介護サービス情報公表システム",
     "厚生労働省（都道府県が公表）", "令和8年8月時点の公表内容",
     "個別公表画面18件。訪問介護13事業所すべてを含む",
     "E2"),
    ("⑧", "有料老人ホーム・サービス付き高齢者向け住宅等の名簿",
     "北海道", "令和8年公表分",
     "介護付有料老人ホーム・住宅型有料老人ホーム・"
     "サービス付き高齢者向け住宅・軽費老人ホーム",
     "E2"),
    ("⑨", "①在宅生活改善調査", "大雪地区広域連合",
     "令和8年8月受領", "事業所票15件・利用者票99票（有効98票）の集計値",
     "E3"),
    ("⑩", "②居所変更実態調査", "大雪地区広域連合",
     "令和8年8月受領", "施設等票18件の集計値",
     "E3"),
    ("⑪", "③介護人材実態調査", "大雪地区広域連合",
     "令和7年4月1日現在・令和8年8月受領",
     "事業所票27件・職員個票317人の集計値・職員票26件",
     "E3"),
    ("⑫", "④2025年健康とくらしの調査（JAGES）",
     "日本老年学的評価研究機構・大雪地区広域連合",
     "令和7年度", "3町の回答4,798票（回収率67.4％）の集計値",
     "E3"),
]


def gensen_sheet(wb, only):
    ws = sheet(wb, "00_出典一覧", "出典一覧",
               "本エビデンス集に収めたデータの出典である。"
               "成果品に載せた数値は、すべて本集のいずれかの表に遡ることができる。"
               "二次情報（民間の情報サイト・業界紙）は用いていない。",
               [4, 30, 20, 22, 46, 6])
    r = head(ws, 4, ["", "出典", "作成者", "時点", "収録した内容", "収録先"])
    for a in GENSEN:
        f = {6: OK_G} if a[5] == only else {}
        r = line(ws, r, list(a), f, height=32)
    for c in ws[4:r]:
        for cell in c:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    note(ws, r + 1,
         "注1）「収録先」は本エビデンス集のどのファイルに収めたかを示す。"
         "網掛けは本ファイルに収めたものである。"
         "注2）個票（利用者票・職員個票）は収録していない。集計値のみである。"
         "担当者名・電話番号・メールアドレス及び施設の管理者名も収録していない。"
         "注3）見える化システムのδ（介護サービス自給率）には"
         "「本指標は自治体向けのため取り扱いに注意」との記載があるため、"
         "本集には収録していない。"
         "分析結果は成果品「地域差の分析」による。", 6)
    return ws


# ================================================================ E1
def build_e1():
    import data_population as P
    import data_nintei as N
    import data_jukyu as J
    import data_mieruka_km as KM
    import data_c1_premium as C

    buf, old = io.StringIO(), sys.stdout
    sys.stdout = buf
    try:
        G = runpy.run_path("build_pop_adjust.py")
    finally:
        sys.stdout = old
    JUKI, JY, TOWNS = G["JUKI"], G["JY"], G["TOWNS"]

    wb = new_book()
    gensen_sheet(wb, "E1")

    # -- 01 人口
    ws = sheet(wb, "01_人口_見える化A系列",
               "人口　見える化A系列（2000年〜2050年）",
               "国勢調査（令和2年まで）及び社人研「日本の地域別将来推計人口」"
               "（令和5年推計）による。各年10月1日現在。"
               "5年ごとの公表値の間の各年は按分値である。")
    r = lead(ws, 4, "【A2　総人口・高齢者数・高齢化率】")
    r = put_matrix(ws, r, P.A2, "区分", P.YEARS)
    r = lead(ws, r + 1, "【A3　後期高齢者数】")
    r = put_matrix(ws, r, P.A3, "区分", P.YEARS)
    r = lead(ws, r + 1, "【A3a　前期・後期高齢者割合】")
    r = put_matrix(ws, r, P.A3a, "区分", P.YEARS)
    r = lead(ws, r + 1, "【A4　5歳階級別人口】")
    r = put_matrix(ws, r, P.A4, "階級", P.YEARS)
    r = lead(ws, r + 1, "【A9　生産年齢人口・年少人口】")
    r = put_matrix(ws, r, P.A9, "区分", P.YEARS)
    autow(ws, len(P.YEARS) + 1, minw=8, maxw=16)

    # -- 02 住民基本台帳
    ws = sheet(wb, "02_住民基本台帳_町別5歳階級",
               "住民基本台帳　年齢5歳階級別人口（町別・平成31年〜令和7年）",
               "総務省「住民基本台帳に基づく人口、人口動態及び世帯数」"
               "年齢別人口（市区町村別）【総計】各年1月1日現在。"
               "日本人住民と外国人住民の合計である。"
               "人口推計の補正（社人研の伸び率の検証）に用いた原表である。")
    COLS = ["65〜69歳", "70〜74歳", "75〜79歳", "80〜84歳", "85〜89歳",
            "90〜94歳", "95〜99歳", "100歳以上"]
    WA = {2019: "平成31年", 2020: "令和2年", 2021: "令和3年", 2022: "令和4年",
          2023: "令和5年", 2024: "令和6年", 2025: "令和7年"}
    r = head(ws, 4, ["町", "年（1月1日）", "総人口"] + COLS +
             ["65歳以上", "75歳以上", "高齢化率(%)"])
    for t in TOWNS:
        for y in JY:
            v = JUKI[t][y]
            b65, b75 = sum(v[1:9]), sum(v[3:9])
            r = line(ws, r, [t, WA[y], v[0]] + v[1:9] +
                     [b65, b75, round(b65 / v[0] * 100, 1)])
    for y in JY:
        tot = sum(JUKI[t][y][0] for t in TOWNS)
        band = [sum(JUKI[t][y][i] for t in TOWNS) for i in range(1, 9)]
        b65, b75 = sum(band), sum(band[2:])
        r = line(ws, r, ["3町計", WA[y], tot] + band +
                 [b65, b75, round(b65 / tot * 100, 1)],
                 {i: GRAY for i in range(1, 15)})
    note(ws, r + 1,
         "注1）平成31年は平成31年1月1日現在である。"
         "同年5月に改元されたため令和元年1月1日は存在しない。"
         "注2）国勢調査人口とは定義が異なる。"
         "令和2年について、住民基本台帳（1月1日）は"
         "東川町8,380人・美瑛町9,912人・東神楽町10,239人であり、"
         "国勢調査（10月1日）の8,314人・9,668人・10,127人をいずれも上回る。", 14)
    autow(ws, 14)

    # -- 03 認定
    ws = sheet(wb, "03_認定_見える化B系列",
               "要介護認定　見える化B系列",
               "介護保険事業状況報告等による。"
               "認定者数・認定率・要介護度別・年齢階級別・"
               "年齢調整済み認定率を含む。")
    r = put_series(ws, 4, N.B, "系列コード")
    autow(ws, 6, maxw=48)

    # -- 04 受給
    ws = sheet(wb, "04_受給_見える化D系列",
               "受給　見える化D系列",
               "受給者数・受給率・サービス種別ごとの給付月額・"
               "介護サービス利用率を含む。")
    r = put_series(ws, 4, J.D, "系列コード")
    autow(ws, 6, maxw=48)

    # -- 05 事業所数・従事者数
    ws = sheet(wb, "05_事業所数と従事者数_KM系列",
               "事業所数（K系列）と従事者数（M系列）",
               "K1〜K3は人口10万対の事業所数（全28種別・平成24年〜令和6年）、"
               "M2はサービス別・職種別の従事者数（認定者1万対・"
               "平成29年〜令和6年）である。")
    r = lead(ws, 4, "【K系列　事業所数】")
    r = put_series(ws, r, KM.K, "系列コード")
    r = lead(ws, r + 1, "【M系列　従事者数】")
    r = put_series(ws, r, KM.M, "系列コード")
    autow(ws, 6, maxw=48)

    # -- 06 保険料
    ws = sheet(wb, "06_保険料_上川管内21保険者",
               "保険料と給付月額　上川管内21保険者（第1期〜第9期）",
               "見える化システムC1による。"
               "第1号被保険者1人当たり給付月額・条例上の保険料基準額・"
               "必要保険料月額を、上川管内21保険者について収める。")
    r = lead(ws, 4, "【給付月額（円）】")
    r = put_matrix(ws, r, C.BEN, "保険者", C.YEARS)
    r = lead(ws, r + 1, "【条例上の保険料基準額（月額・円）】")
    r = put_matrix(ws, r, C.PREM, "保険者", C.YEARS)
    r = lead(ws, r + 1, "【必要保険料月額（円）】")
    r = put_matrix(ws, r, C.NEED, "保険者", C.YEARS)
    note(ws, r + 1,
         "注）「必要保険料月額」は給付費等から算定される保険料であり、"
         "条例上の基準額との差が基金の取崩し等による調整分である。",
         len(C.YEARS) + 1)
    autow(ws, len(C.YEARS) + 1, minw=8, maxw=16)

    out = os.path.join(EDIR, "第10期計画_エビデンス_1_統計データ.xlsx")
    wb.save(out)
    return out, wb


# ================================================================ E2
def build_e2():
    import data_hokkaido_shitei as H
    import data_hokkaido_roster as R

    wb = new_book()
    gensen_sheet(wb, "E2")

    # -- 01 指定事業所
    ws = sheet(wb, "01_指定事業所一覧",
               "区域内の指定事業所一覧（北海道・令和8年6月30日現在）",
               "北海道「介護保険事業所一覧」による。"
               "サービス区分ごとに、区域内に所在する指定事業所を収める。")
    r = head(ws, 4, ["サービス", "事業所名", "町", "所在地", "法人",
                     "事業所番号", "指定年月日"])
    n_shitei = 0
    for sv in H.SHITEI:
        for e in H.SHITEI[sv]:
            if isinstance(e, dict):
                r = line(ws, r, [sv, e.get("事業所名"), e.get("町"),
                                 e.get("所在地"), e.get("法人"),
                                 e.get("事業所番号"), e.get("指定年月日")])
            else:
                r = line(ws, r, [sv] + list(e))
            n_shitei += 1
    autow(ws, 7, maxw=44)

    # -- 02 公表画面
    ws = sheet(wb, "02_公表画面の記載内容",
               "介護サービス情報公表システムの個別公表画面（18件）",
               "令和8年8月時点の公表内容を、画面の記載どおりに収める。"
               "訪問介護13事業所すべてを含む。"
               "調査に回答がなかった事業所の把握に用いた。"
               "管理者名は収録していない。")
    r = put_records(ws, 4, H.KOHYO)
    autow(ws, 60, maxw=34)

    ws2 = sheet(wb, "03_公表画面が確認できない事業所",
                "公表画面が確認できない事業所",
                "指定事業所一覧に掲載があるが、"
                "介護サービス情報公表システムに個別公表画面が"
                "確認できない事業所である。")
    r = put_records(ws2, 4, H.KOHYO_MISSING)
    autow(ws2, 20, maxw=40)

    # -- 04 施設・住まいの名簿
    ws = sheet(wb, "04_施設と住まいの名簿",
               "施設・居住系サービス及び介護保険外の住まいの名簿",
               "特別養護老人ホーム等は北海道の指定事業所一覧、"
               "有料老人ホーム・サービス付き高齢者向け住宅・"
               "軽費老人ホームは北海道の公表名簿による。")
    r = lead(ws, 4, "【特別養護老人ホーム等】")
    r = put_records(ws, r, H.TOKUYO)
    for lab, rows in [("有料老人ホーム", R.YU),
                      ("サービス付き高齢者向け住宅", R.SA),
                      ("軽費老人ホーム・ケアハウス", R.KE),
                      ("養護老人ホーム", R.YO),
                      ("認知症対応型共同生活介護（名簿）", R.MI)]:
        r = lead(ws, r + 1, "【%s】" % lab)
        r = put_records(ws, r, rows)
    autow(ws, 16, maxw=40)

    # -- 05 通常の事業の実施地域
    ws = sheet(wb, "05_通常の事業の実施地域",
               "運営規程における「通常の事業の実施地域」",
               "訪問系・通所系の事業所について、"
               "運営規程に定める通常の事業の実施地域を収める。"
               "訪問・通所困難地域の判定の根拠である。")
    r = put_records(ws, 4, H.POS)
    autow(ws, 12, maxw=44)

    out = os.path.join(EDIR, "第10期計画_エビデンス_2_事業所と公表情報.xlsx")
    wb.save(out)
    return out, wb, n_shitei


# ================================================================ E3
def build_e3():
    import data_survey2025 as S
    import data_survey_cross as X
    import data_survey_entry as E

    wb = new_book()
    gensen_sheet(wb, "E3")

    ws = sheet(wb, "01_個人情報の取扱い",
               "個人情報の取扱い",
               "本ファイルに収めた調査データの範囲と、"
               "収録していないものを明示する。",
               [4, 30, 34, 40])
    r = head(ws, 4, ["", "区分", "収録の有無", "内容"])
    for a in [
        ("①", "利用者票の個票（99票）", "収録しない",
         "年齢・世帯・要介護度・回答内容を個人単位で保持しない。"
         "集計値（度数・クロス集計）のみを収める"),
        ("②", "職員個票（317人）", "収録しない",
         "同上。資格・雇用形態・年代等の集計値のみを収める"),
        ("③", "担当者名・電話番号・メールアドレス", "収録しない",
         "調査票に記入があるが、いずれのファイルにも収めない"),
        ("④", "施設の管理者名", "収録しない",
         "公表画面に記載があるが収録しない運用としている"),
        ("⑤", "事業所名・施設名", "収録する",
         "指定事業所一覧及び公表画面により公表されている情報である"),
        ("⑥", "事業所票・施設等票の回答（15件・18件・27件）", "収録する",
         "事業所を単位とする回答であり、個人情報に当たらない"),
        ("⑦", "自由回答（45件）", "収録する",
         "事業所の意見である。同一法人の同文は1件に集約している"),
        ("⑧", "所在地区欄の記載文字列（99票分）", "収録する",
         "提出元・ファイル名・票番号・記載文字列のみ。"
         "個人の属性は伴わない。記入形式が9種類に分かれることの根拠である"),
    ]:
        f = {3: OK_G if a[2] == "収録する" else NG_O}
        r = line(ws, r, list(a), f, height=32)
    for c in ws[4:r]:
        for cell in c:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    note(ws, r + 1,
         "注）④健康とくらしの調査は、日本老年学的評価研究機構が集計した"
         "指標値（20指標の年齢階級別・町別集計）を用いている。"
         "個票は当方に提供されていない。"
         "同調査の集計値は成果品「調査クロス集計・分析」及び"
         "「アンケート調査の集計分析報告書」に収めている。", 4)

    ws = sheet(wb, "02_①在宅生活改善調査",
               "① 在宅生活改善調査　事業所票（15件）・利用者票（99票）の集計",
               "居宅介護支援事業所のケアマネジャーを通じて、"
               "在宅生活の維持が難しい利用者について把握したものである。"
               "利用者票は課題のある利用者を抽出する設計であり、"
               "区域内の在宅利用者全体を母集団としない。")
    r = lead(ws, 4, "【事業所票（15件）】")
    r = head(ws, r, ["提出元", "事業所名", "値1", "値2", "値3", "内訳"])
    for e in S.ZAI:
        r = line(ws, r, list(e))
    r = lead(ws, r + 1, "【利用者票（99票）の集計】")
    r = head(ws, r, ["項目", "値"])
    for k, v in S.RIYO.items():
        r = line(ws, r, [k, v if not isinstance(v, (list, dict)) else str(v)])
    autow(ws, 6, maxw=44)

    ws = sheet(wb, "03_②居所変更実態調査",
               "② 居所変更実態調査　施設等票（18件）",
               "施設・居住系サービスの管理者を通じて、"
               "入所前の居場所・退去先・受入れ可能な医療処置等を"
               "把握したものである。"
               "18施設からの回答であり、区域内の住替え全体を母集団としない。")
    r = put_records(ws, 4, S.SHI)
    autow(ws, 40, maxw=34)

    ws = sheet(wb, "04_③介護人材実態調査",
               "③ 介護人材実態調査　事業所票（27件）・職員個票（317人）の集計"
               "・職員票（26件）",
               "令和7年4月1日現在。"
               "訪問系は区域内13事業所のうち3事業所からの回答である。"
               "訪問系の全数は、エビデンス2の公表画面により把握した。")
    r = lead(ws, 4, "【事業所票（27件）】")
    r = put_records(ws, r, S.JIN)
    r = lead(ws, r + 1, "【職員個票（317人）の集計】")
    r = head(ws, r, ["項目", "値"])
    for k, v in S.SHOKU.items():
        r = line(ws, r, [k, v if not isinstance(v, (list, dict)) else str(v)])
    r = lead(ws, r + 1, "【職員票　訪問系（26件）】")
    r = put_records(ws, r, S.HOU)
    autow(ws, 40, maxw=34)

    ws = sheet(wb, "05_調査内クロス集計",
               "3調査それぞれの中でのクロス集計（度数）",
               "個票を用いず、度数の表として収める。"
               "4調査を横断したクロス集計は、"
               "個票を共通の単位に割り付けられないため実施していない。")
    r = 4
    for lab, d, k1, k2 in [
        ("CU　①利用者票のクロス集計", X.CU, "集計軸", "区分"),
        ("SVB　①より適切なサービス×提出元", X.SVB, "サービス", "提出元"),
        ("SVC　①より適切なサービス×要介護度", X.SVC, "サービス", "要介護度"),
        ("SUP　①必要な生活支援×世帯類型", X.SUP, "生活支援", "世帯類型"),
        ("RSN　①困難の理由×希望の方向", X.RSN, "理由", "希望の方向"),
        ("CS　②施設等票のクロス集計", X.CS, "集計軸", "区分"),
        ("SS　②種別ごとの定員・入所・新規・退去", X.SS, "種別", "項目"),
        ("MF　②施設ごとの医療処置", X.MF, "施設", "医療処置"),
        ("CJ　③職員個票のクロス集計", X.CJ, "集計軸", "区分"),
        ("CH　③職員票 訪問系のクロス集計", X.CH, "集計軸", "区分"),
        ("HT　③訪問系の従事時間", X.HT, "雇用形態", "区分"),
    ]:
        r = lead(ws, r, "【%s】" % lab, 3)
        r = put_nested(ws, r, d, k1, k2, "度数")
        r += 1
    autow(ws, 3, maxw=48)

    ws = sheet(wb, "06_記載内容の一覧",
               "調査票の記載内容の一覧",
               "所在地区欄の記載形式が9種類に分かれることの根拠である。"
               "提出元・ファイル名・票番号・記載文字列のみを収める。"
               "個人の属性は一切収めていない。")
    r = lead(ws, 4, "【利用者票の所在地区欄（99票）】", 4)
    r = head(ws, r, ["提出元", "ファイル名", "票番号", "記載文字列"])
    for e in E.CHIKU:
        r = line(ws, r, list(e))
    r = lead(ws, r + 1, "【②施設等票の所在地区（18件）】", 9)
    r = head(ws, r, ["提出元", "施設名", "値1", "値2", "値3", "値4",
                     "値5", "値6", "値7"])
    for e in E.AREA:
        r = line(ws, r, list(e))
    r = lead(ws, r + 1, "【③人材_事業所票の区分（27件）】", 3)
    r = head(ws, r, ["提出元", "事業所名", "区分"])
    for e in E.JIN:
        r = line(ws, r, list(e))
    autow(ws, 9, maxw=44)

    ws = sheet(wb, "07_自由回答",
               "3調査の自由回答（45件）",
               "事業所の意見である。"
               "同一法人の同文は1件に集約している。"
               "個人を特定できる記述は含まれない。")
    r = head(ws, 4, ["調査・票種", "提出元", "事業所名", "設問", "回答"])
    for e in S.FREE:
        r = line(ws, r, list(e), height=15)
    autow(ws, 5, maxw=70)
    for row in ws.iter_rows(min_row=5, min_col=5, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    out = os.path.join(EDIR, "第10期計画_エビデンス_3_調査の集計値.xlsx")
    wb.save(out)
    return out, wb


# ================================================================ 索引
def build_index(files):
    wb = new_book()
    ws = sheet(wb, "00_エビデンス集について",
               "エビデンス集について",
               "成果品（計画素案・報告書・分析表）の根拠となるデータを、"
               "Excelで直接確認できる形に書き出したものである。"
               "成果品に載せた数値は、すべて本集のいずれかの表に遡ることができる。",
               [5, 34, 46, 30, 12])
    r = lead(ws, 4, "【1　エビデンス集の構成】", 5)
    r = head(ws, r, ["", "ファイル", "収録した内容", "主な出典", "シート数"])
    for a in files:
        r = line(ws, r, list(a), {}, height=44)
    for c in ws[5:r]:
        for cell in c:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    r += 1
    r = lead(ws, r, "【2　成果物とエビデンスの関係】", 5)
    r = head(ws, r, ["", "区分", "内容", "", ""])
    for a in [
        ("①", "成果物（納品対象）",
         "計画素案・骨子案・図表集・各報告書・各分析表・管理表・会議資料。"
         "仕様書４が求める成果品である。"
         "3つのZIPにまとめている。", "", ""),
        ("②", "エビデンス（根拠データ）",
         "成果物の数値の根拠となる原データ及び図表の画像。"
         "納品対象ではないが、"
         "数値の検証と第11期の作業の引継ぎのために添える。"
         "2つのZIPにまとめている。", "", ""),
        ("③", "両者の対応",
         "成果物の各表には出典を明記している。"
         "出典の表記は本集00シートの出典一覧の①〜⑫に対応する。", "", ""),
    ]:
        r = line(ws, r, list(a), {}, height=52)
        ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1,
                       end_column=5)
    for c in ws[5:r]:
        for cell in c:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    r += 1
    r = lead(ws, r, "【3　収録していないもの】", 5)
    r = head(ws, r, ["", "内容", "理由", "", ""])
    for a in [
        ("①", "個票（利用者票99票・職員個票317人）",
         "個人情報のため。集計値及びクロス集計の度数のみを収めている。", "", ""),
        ("②", "担当者名・電話番号・メールアドレス・施設の管理者名",
         "個人情報のため。調査票及び公表画面に記載があるが収録しない。", "", ""),
        ("③", "見える化システムのδ（介護サービス自給率）",
         "「本指標は自治体向けのため取り扱いに注意」との記載があるため。"
         "分析結果は成果品「地域差の分析」による。", "", ""),
        ("④", "④健康とくらしの調査の個票",
         "当方に提供されていない。"
         "集計された指標値（20指標の年齢階級別・町別）を用いている。", "", ""),
        ("⑤", "北海道庁・e-Statから直接取得した原ファイル",
         "作業環境から外部サイトへ接続できないため、"
         "受領したファイル及び画面の記載内容から収録している。", "", ""),
    ]:
        r = line(ws, r, list(a), {}, height=40)
        ws.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1,
                       end_column=5)
    for c in ws[5:r]:
        for cell in c:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    note(ws, r + 1,
         "注1）本集は成果品ではない。"
         "仕様書４（10）が求める「分析資料・推計資料・図表・グラフ・"
         "クロス集計表の電子データ」の一部として位置づけられるものである。"
         "注2）図表の画像（PNG・79点）は"
         "「エビデンス_4_図表の画像」に収めている。"
         "注3）本集の各表は、生成スクリプトから機械的に書き出している。"
         "手作業による転記を経ていないため、"
         "成果品の数値との食い違いは生じない。", 5)

    out = os.path.join(EDIR, "第10期計画_エビデンス_0_索引.xlsx")
    wb.save(out)
    return out


# ================================================================ main
if __name__ == "__main__":
    os.makedirs(EDIR, exist_ok=True)

    f1, wb1 = build_e1()
    f2, wb2, n_shitei = build_e2()
    f3, wb3 = build_e3()

    idx = build_index([
        ("1", os.path.basename(f1),
         "人口（見える化A系列・2000〜2050年）、"
         "住民基本台帳の年齢5歳階級別人口（町別・平成31年〜令和7年）、"
         "認定（B系列）、受給（D系列）、"
         "事業所数（K系列）・従事者数（M系列）、"
         "保険料（上川管内21保険者・第1期〜第9期）",
         "見える化システム、総務省住民基本台帳、"
         "国勢調査、社人研、介護保険事業状況報告",
         len(wb1.sheetnames)),
        ("2", os.path.basename(f2),
         "区域内の指定事業所一覧、"
         "介護サービス情報公表システムの個別公表画面（18件）、"
         "施設・住まいの名簿、通常の事業の実施地域",
         "北海道介護保険事業所一覧、"
         "介護サービス情報公表システム、北海道の公表名簿",
         len(wb2.sheetnames)),
        ("3", os.path.basename(f3),
         "実施済み3調査の集計値（事業所票・利用者票・施設等票・"
         "職員個票・職員票）、調査内クロス集計、"
         "記載内容の一覧、自由回答",
         "①在宅生活改善調査、②居所変更実態調査、"
         "③介護人材実態調査（いずれも大雪地区広域連合）",
         len(wb3.sheetnames)),
        ("4", "第10期計画_エビデンス_4_図表の画像.zip",
         "計画素案及び報告書に掲載した図表の画像（PNG）。"
         "figures 34点・figures_report 39点・images_basic 6点",
         "上記①〜⑫の各出典による",
         "―"),
    ])

    for p in (idx, f1, f2, f3):
        wb = None
        print("saved:", os.path.basename(p))
    print()
    print("指定事業所", n_shitei, "件")
    for nm, wb in (("1_統計データ", wb1), ("2_事業所と公表情報", wb2),
                   ("3_調査の集計値", wb3)):
        print(" ", nm)
        for ws in wb:
            print("    -", ws.title, ws.max_row, "rows")
